# -*- coding: utf-8 -*-
"""AUDITORIA VENTAS CORE - tanda 1 (SOLO LECTURA, julio 2026).

Ejecutar:  cd retailmind && python manage.py shell < <este archivo>
Invoca vistas reales via RequestFactory (patron _test_reportes_readonly) y
las cruza con oraculos independientes. Guarda anti-escritura: rollback.
"""
import json
import sys
import time
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from django.conf import settings
from django.db import connection, reset_queries, transaction
from django.db.models import Count, F, Q, Sum
from django.test import RequestFactory

from app.models import (Dte, Dte_Productos, EmpresaUser, Sucursal, Ticket,
                        Vendedor)
from app.utils_permisos import usuario_puede_ver_todas_sucursales
from django.contrib.auth import get_user_model

settings.DEBUG = True  # habilita connection.queries (solo memoria)

FI = date(2026, 7, 1)
FF = date(2026, 7, 31)
SEP = '=' * 78


def I(x):
    return int(x or 0)


def sec(t):
    print('\n' + SEP + '\n' + t + '\n' + SEP)


def _import_view(path):
    mod_name, fn = path.rsplit('.', 1)
    mod = __import__(mod_name, fromlist=[fn])
    return getattr(mod, fn)


def invocar(path, params, user, sucursal_id=None, empresa_id=None):
    """GET real a la vista. Devuelve dict con status/json/raw/queries/ms."""
    factory = RequestFactory()
    req = factory.get('/_aud', data=params)
    req.user = user
    req.session = {'idSucursalActual': sucursal_id, 'idEmpresaActual': empresa_id}
    out = {'status': None, 'json': None, 'raw': None, 'nq': None, 'ms': None, 'err': None}
    reset_queries()
    t0 = time.perf_counter()
    try:
        with transaction.atomic():
            resp = _import_view(path)(req)
            transaction.set_rollback(True)
    except Exception as e:
        out['err'] = f'{type(e).__name__}: {e}'
        out['ms'] = round((time.perf_counter() - t0) * 1000)
        out['nq'] = len(connection.queries)
        return out
    out['ms'] = round((time.perf_counter() - t0) * 1000)
    out['nq'] = len(connection.queries)
    malas = [q['sql'][:90] for q in connection.queries
             if q['sql'].lstrip().upper().startswith(
                 ('INSERT', 'UPDATE', 'DELETE', 'TRUNCATE', 'ALTER', 'DROP', 'CREATE'))]
    if malas:
        out['err'] = 'ESCRITURA DETECTADA: ' + '; '.join(malas[:2])
    out['status'] = resp.status_code
    out['raw'] = resp.content
    try:
        out['json'] = json.loads(resp.content)
    except Exception:
        pass
    return out


User = get_user_model()
ADMIN = (User.objects.filter(rol='administrador', is_active=True).first()
         or User.objects.filter(is_superuser=True, is_active=True).first())
print('admin =', getattr(ADMIN, 'username', None))

# usuario restringido (mismo criterio que la suite)
RESTR = None
RESTR_EMPRESAS = None
for u in User.objects.filter(is_active=True, is_superuser=False).exclude(rol='administrador')[:60]:
    eus = list(EmpresaUser.objects.filter(user=u, status=True)
               .values_list('empresa_id', 'sucursal_id'))
    if not eus:
        continue
    empresas = {e for e, _ in eus}
    try:
        if usuario_puede_ver_todas_sucursales(u):
            continue
    except Exception:
        continue
    suc = next((s for _, s in eus if s), None)
    if not suc:
        suc = Sucursal.objects.filter(empresa_id__in=empresas).values_list('id', flat=True).first()
    RESTR = (u, suc, sorted(empresas))
    RESTR_EMPRESAS = empresas
    break
print('restringido =', RESTR and (RESTR[0].username, 'suc', RESTR[1], 'empresas', RESTR[2]))

# ============================================================ S1 CENSO DTE JULIO
try:
    sec('S1 CENSO Dte JULIO 2026')
    base = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF)
    print('por tipo_documento (todos los DTE julio):')
    for r in base.values('tipo_documento').annotate(n=Count('id'), m=Sum('monto_con_iva')).order_by('-n'):
        print(f"  {r['tipo_documento']!s:24} n={r['n']:6}  monto={I(r['m']):>14,}")
    print('ventas (transacc VENTA/VP, no NC) por estado_dte / descartado:')
    v = base.filter(tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO']).exclude(tipo_documento='NOTA DE CREDITO')
    for r in v.values('estado_dte', 'descartado').annotate(n=Count('id'), m=Sum('monto_con_iva')).order_by('estado_dte'):
        print(f"  estado={r['estado_dte']!s:12} desc={r['descartado']!s:5} n={r['n']:6} monto={I(r['m']):>14,}")
    print('tipo_transaccion VENTA (no VP), no NC, estados vigentes, no descartado:')
    tv = v.filter(tipo_transaccion='VENTA', descartado=False).exclude(
        estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])
    a = tv.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"  n={a['n']} monto={I(a['m']):,}")
    print('NC julio por tipo_transaccion / estado / descartado:')
    nc = base.filter(tipo_documento='NOTA DE CREDITO')
    for r in nc.values('tipo_transaccion', 'estado_dte', 'descartado').annotate(
            n=Count('id'), m=Sum('monto_con_iva')).order_by('tipo_transaccion', 'estado_dte'):
        print(f"  tt={r['tipo_transaccion']!s:14} est={r['estado_dte']!s:10} desc={r['descartado']!s:5} "
              f"n={r['n']:4} monto={I(r['m']):>12,}")
    # ventas descartadas que ventas-global/comparativo INCLUYEN (sin filtro descartado)
    vd = v.filter(descartado=True).exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])
    a = vd.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"ventas descartado=True con estado vigente (entran a vglobal/comparativo): n={a['n']} monto={I(a['m']):,}")
    ncd = nc.filter(descartado=True).exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])
    a = ncd.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"NC OCULTAS (descartado=True, estado vigente) que vglobal resta: n={a['n']} monto={I(a['m']):,}")
    van = v.filter(estado_dte='ANULADO', descartado=False)
    a = van.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"ventas ANULADO no-descartadas (entran a vsucursal, NO a comisiones/vglobal): n={a['n']} monto={I(a['m']):,}")
    # monto_neto poblado (comisiones se calcula sobre monto_neto)
    vig = v.filter(descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'])
    nn = vig.filter(Q(monto_neto__isnull=True) | Q(monto_neto=0), monto_con_iva__gt=0)
    a = nn.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"ventas vigentes con monto_neto NULL/0 y monto_con_iva>0: n={a['n']} monto_iva={I(a['m']):,}")
    ncn = nc.filter(descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'])
    a2 = ncn.filter(Q(monto_neto__isnull=True) | Q(monto_neto=0), monto_con_iva__gt=0).aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"NC vigentes con monto_neto NULL/0: n={a2['n']} monto_iva={I(a2['m']):,}")
except Exception as e:
    print('S1 ERROR:', type(e).__name__, e)

# ============================================================ S2 TICKETS / BOLETA PAPEL / DOBLE CONTEO
try:
    sec('S2 TICKETS JULIO vs Dte (doble conteo / BOLETA PAPEL)')
    tk = Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF,
                               estado='PAGADO',
                               modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'])
    a = tk.aggregate(n=Count('id'), m=Sum('total'))
    print(f"tickets PAGADO julio (VP/POS/ECOM): n={a['n']} monto={I(a['m']):,}")
    tsd = tk.filter(dte_generado=False)
    a = tsd.aggregate(n=Count('id'), m=Sum('total'))
    print(f"  dte_generado=False (los que suman vglobal/comparativo): n={a['n']} monto={I(a['m']):,}")
    con_folio = tsd.exclude(folio_dte__isnull=True).exclude(folio_dte='')
    a = con_folio.aggregate(n=Count('id'), m=Sum('total'))
    print(f"  ...de esos, con folio_dte NO vacio (sospecha doble conteo): n={a['n']} monto={I(a['m']):,}")
    folios = set(con_folio.values_list('folio_dte', flat=True))
    if folios:
        dmatch = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                                    numero_documento__in=list(folios),
                                    tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'])\
            .exclude(tipo_documento='NOTA DE CREDITO')\
            .exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])
        a = dmatch.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
        print(f"  DTEs julio vigentes cuyo folio coincide (DOBLE CONTEO real vglobal): n={a['n']} monto={I(a['m']):,}")
        for r in dmatch.values('tipo_documento').annotate(n=Count('id'), m=Sum('monto_con_iva')):
            print(f"     match tipo={r['tipo_documento']}: n={r['n']} monto={I(r['m']):,}")
    bp = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                            tipo_documento='BOLETA PAPEL')
    a = bp.aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"Dte BOLETA PAPEL julio (todas): n={a['n']} monto={I(a['m']):,}")
    # drift Ticket.fecha (auto_now) vs created_at
    drift = 0
    tot = 0
    for t in tk.only('fecha', 'created_at')[:100000].iterator(chunk_size=2000):
        tot += 1
        from django.utils import timezone as _tz
        if t.fecha != _tz.localtime(t.created_at).date():
            drift += 1
    print(f"tickets julio con Ticket.fecha != date(created_at): {drift}/{tot}")
except Exception as e:
    print('S2 ERROR:', type(e).__name__, e)

# ============================================================ S3 VENTAS-SUCURSAL + VENDEDOR
try:
    sec('S3 VENTAS-SUCURSAL / VENDEDOR vs ORACULO')
    # oraculo (regla del reporte): estados EMITIDO/ACEPTADO/ANULADO, descartado=False
    ob = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
                            estado_dte__in=['EMITIDO', 'ACEPTADO', 'ANULADO'],
                            descartado=False, sucursal__isnull=False)
    ov = ob.exclude(tipo_documento='NOTA DE CREDITO').aggregate(m=Sum('monto_con_iva'), n=Count('id'))
    on = ob.filter(tipo_documento='NOTA DE CREDITO').aggregate(m=Sum('monto_con_iva'), n=Count('id'))
    ora_neto = I(ov['m']) - I(on['m'])
    print(f"ORACULO neto julio = {ora_neto:,}  (ventas {I(ov['m']):,}/{ov['n']} docs - NC {I(on['m']):,}/{on['n']})")

    r1 = invocar('app.views_modulo_reportes.obtener_ventas_por_sucursal_reporte',
                 {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    js = r1['json'] or {}
    filas = js.get('sucursales') or []
    tot_api = sum(I(f.get('ventas')) for f in filas)
    docs_api = sum(I(f.get('documentos')) for f in filas)
    dev_api = sum(I(f.get('devoluciones')) for f in filas)
    print(f"vista ventas-sucursal: status={r1['status']} nq={r1['nq']} ms={r1['ms']} err={r1['err']}")
    print(f"  suma filas.ventas = {tot_api:,}  delta vs oraculo = {tot_api - ora_neto:,}")
    print(f"  suma docs = {docs_api} (oraculo {ov['n']})  suma devoluciones = {dev_api:,} (oraculo {I(on['m']):,})")

    r2 = invocar('app.views_modulo_reportes.obtener_ventas_por_vendedor_reporte',
                 {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    js2 = r2['json'] or {}
    vend = js2.get('vendedores') or []
    kpis = js2.get('kpis') or {}
    suma_filas = sum(I(v.get('ventas')) for v in vend)
    fila_sv = next((v for v in vend if v.get('sin_vendedor')), None)
    print(f"vista ventas-vendedor: status={r2['status']} nq={r2['nq']} ms={r2['ms']} err={r2['err']}")
    print(f"  KPI total_ventas={I(kpis.get('total_ventas')):,}  KPI total_devoluciones={I(kpis.get('total_devoluciones')):,}")
    print(f"  suma filas (incl sin-vendedor)={suma_filas:,}  delta vs oraculo neto={suma_filas - ora_neto:,}")
    if fila_sv:
        print(f"  fila SIN VENDEDOR: brutas={I(fila_sv.get('ventas_brutas')):,} dev={I(fila_sv.get('devoluciones')):,} neto={I(fila_sv.get('ventas')):,}")
    print(f"  divergencia KPI-vendedor vs total-sucursal = {I(kpis.get('total_ventas')) - tot_api:,}")
    # composicion de la divergencia
    vsv = ob.exclude(tipo_documento='NOTA DE CREDITO').filter(vendedor__isnull=True).aggregate(m=Sum('monto_con_iva'), n=Count('id'))
    ncv = ob.filter(tipo_documento='NOTA DE CREDITO', vendedor__isnull=True).aggregate(m=Sum('monto_con_iva'), n=Count('id'))
    ncc = ob.filter(tipo_documento='NOTA DE CREDITO', vendedor__isnull=False).aggregate(m=Sum('monto_con_iva'), n=Count('id'))
    print(f"  ventas sin vendedor: {I(vsv['m']):,}/{vsv['n']}  NC sin vendedor: {I(ncv['m']):,}/{ncv['n']}  NC con vendedor: {I(ncc['m']):,}/{ncc['n']}")
    # NC con vendedor cuyo vendedor NO tiene ventas en julio (se pierden en el tab vendedor)
    vids_ventas = set(ob.exclude(tipo_documento='NOTA DE CREDITO').filter(vendedor__isnull=False)
                      .values_list('vendedor_id', flat=True).distinct())
    nc_sin_fila = ob.filter(tipo_documento='NOTA DE CREDITO', vendedor__isnull=False)\
        .exclude(vendedor_id__in=vids_ventas).aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"  NC con vendedor SIN fila de ventas en julio (se pierden del tab): n={nc_sin_fila['n']} monto={I(nc_sin_fila['m']):,}")
    # UPT oraculo del top vendedor
    top = next((v for v in vend if not v.get('sin_vendedor')), None)
    if top:
        u_ora = Dte_Productos.objects.filter(
            dte__fecha_emision__gte=FI, dte__fecha_emision__lte=FF,
            dte__tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
            dte__estado_dte__in=['EMITIDO', 'ACEPTADO', 'ANULADO'],
            dte__descartado=False, dte__vendedor_id=top['id'],
        ).exclude(dte__tipo_documento='NOTA DE CREDITO').aggregate(u=Sum('stock'))
        print(f"  top vendedor {top['nombre']!r}: unidades API={top.get('unidades')} oraculo={I(u_ora['u'])} "
              f"docs={top.get('documentos')} upt API={top.get('upt')} pct_dcto={top.get('pct_descuento')}")
except Exception as e:
    print('S3 ERROR:', type(e).__name__, e)

# ============================================================ S4 COMISIONES
try:
    sec('S4 COMISIONES-VENDEDOR (nunca auditado)')
    cens = Vendedor.objects.aggregate(
        n=Count('id'),
        con_pct=Count('id', filter=Q(comision__gt=0)),
    )
    print(f"vendedores: total={cens['n']}  con comision>0: {cens['con_pct']}")
    for r in Vendedor.objects.filter(comision__gt=0).values('comision').annotate(n=Count('id')).order_by('comision'):
        print(f"  comision={r['comision']}%: {r['n']} vendedores")

    r3 = invocar('app.views_modulo_reportes.obtener_comisiones_por_vendedor',
                 {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    js3 = r3['json'] or {}
    tot = js3.get('totales') or {}
    print(f"vista comisiones: status={r3['status']} nq={r3['nq']} ms={r3['ms']} err={r3['err']}")
    for k in ('total_ventas_brutas_con_iva', 'total_ventas_brutas_sin_iva', 'total_ventas_netas_sin_iva',
              'total_ventas_netas_con_iva', 'total_comisiones', 'total_documentos', 'total_devoluciones',
              'total_devoluciones_neto', 'cantidad_vendedores', 'devoluciones_sin_vendedor',
              'cantidad_ncs_sin_vendedor'):
        print(f"  {k} = {tot.get(k)}")

    # ORACULO independiente (a nivel de DTE, sin agrupar):
    qb = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
                            estado_dte__in=['EMITIDO', 'ACEPTADO'], descartado=False)
    qv = qb.exclude(tipo_documento='NOTA DE CREDITO').filter(vendedor__isnull=False, emisor__isnull=False)
    qn = qb.filter(tipo_documento='NOTA DE CREDITO', emisor__isnull=False)
    a = qv.aggregate(m=Sum('monto_con_iva'), neto=Sum('monto_neto'), n=Count('id'))
    print(f"ORACULO ventas (vendedor+emisor not null): brutas_iva={I(a['m']):,} brutas_neto={I(a['neto']):,} docs={a['n']}")
    print(f"  delta brutas_iva vista-oraculo = {I(tot.get('total_ventas_brutas_con_iva')) - I(a['m']):,}")
    # comision oraculo: sum(monto_neto*pct) ventas - sum NC imputada
    com_v = 0.0
    for r in qv.values('vendedor__comision').annotate(neto=Sum('monto_neto')):
        try:
            pct = float(r['vendedor__comision'] or 0)
        except Exception:
            pct = 0.0
        com_v += I(r['neto']) * pct / 100.0
    com_n = 0.0
    nc_sin_imputar = 0
    for r in qn.values('vendedor__comision', 'documento_afectado__vendedor__comision',
                       'vendedor_id', 'documento_afectado__vendedor_id').annotate(neto=Sum('monto_neto')):
        vid = r['vendedor_id'] or r['documento_afectado__vendedor_id']
        if not vid:
            nc_sin_imputar += I(r['neto'])
            continue
        pct_raw = r['vendedor__comision'] if r['vendedor_id'] else r['documento_afectado__vendedor__comision']
        try:
            pct = float(pct_raw or 0)
        except Exception:
            pct = 0.0
        com_n += I(r['neto']) * pct / 100.0
    ora_com = int(round(com_v - com_n))
    print(f"ORACULO comisiones = {ora_com:,}  vista = {I(tot.get('total_comisiones')):,}  delta = {I(tot.get('total_comisiones')) - ora_com:,}")
    print(f"  NC sin vendedor imputable (neto, no descuentan a nadie): {nc_sin_imputar:,}")
    an = qn.aggregate(m=Sum('monto_con_iva'), neto=Sum('monto_neto'), n=Count('id'))
    print(f"ORACULO NC (emisor not null): iva={I(an['m']):,} neto={I(an['neto']):,} n={an['n']}")
    # divergencia con tab vendedores (ANULADO)
    print(f"comisiones brutas_iva vs vendedor-tab brutas: comisiones excluye ANULADO por diseno")

    # Excel: mismos numeros que el JSON?
    r4 = invocar('app.views_modulo_reportes.exportar_comisiones_vendedor_excel',
                 {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    print(f"export excel: status={r4['status']} nq={r4['nq']} ms={r4['ms']} err={r4['err']}")
    if r4['status'] == 200 and r4['raw'][:2] == b'PK':
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r4['raw']), read_only=True, data_only=True)
        ws = wb.active
        fila_total = None
        for row in ws.iter_rows(min_col=1, max_col=8):
            if row[0].value == 'TOTAL GENERAL':
                fila_total = [c.value for c in row]
        if fila_total:
            print(f"  Excel TOTAL GENERAL: brutas_iva={fila_total[3]:,} dev_neto={fila_total[4]:,} "
                  f"netas_siniva={fila_total[5]:,} comision={fila_total[7]:,}")
            print(f"  JSON  equivalentes: brutas_iva={I(tot.get('total_ventas_brutas_con_iva')):,} "
                  f"dev_neto={I(tot.get('total_devoluciones_neto')):,} "
                  f"netas_siniva={I(tot.get('total_ventas_netas_sin_iva')):,} "
                  f"comision={I(tot.get('total_comisiones')):,}")
        else:
            print('  Excel sin fila TOTAL GENERAL (?)')
    elif r4['status'] == 403:
        print('  export devolvio 403 para admin (permiso puede_exportar?)', (r4['json'] or {}).get('error'))
except Exception as e:
    print('S4 ERROR:', type(e).__name__, e)

# ============================================================ S5 VENTAS-COMPARATIVO
try:
    sec('S5 VENTAS-COMPARATIVO (custom julio)')
    r5 = invocar('app.views_modulo_reportes.obtener_ventas_comparativo',
                 {'tipo_flujo': 'custom', 'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    js5 = r5['json'] or {}
    k = js5.get('kpis') or {}
    print(f"vista: status={r5['status']} nq={r5['nq']} ms={r5['ms']} err={r5['err']}")
    print(f"  ventas_actual(netas)={I(k.get('ventas_actual')):,} docs={k.get('documentos_actual')} "
          f"unidades={k.get('unidades_actual')} tasa_dev={k.get('tasa_devolucion_actual')}%")
    canales = {c['canal']: c for c in js5.get('canales') or []}
    suma_can = sum(I(c.get('ventas_actual')) for c in canales.values())
    print(f"  canales: ECOM={I(canales.get('ECOMMERCE', {}).get('ventas_actual')):,} "
          f"POS={I(canales.get('POS', {}).get('ventas_actual')):,} "
          f"DTE={I(canales.get('DTE', {}).get('ventas_actual')):,} suma={suma_can:,}")
    # oraculo brutas (replica): DTE ventas vigentes sin internas (SIN filtro descartado) + tickets sin dte
    dv = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'])\
        .exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])\
        .exclude(receptor__isnull=False, receptor_id=F('emisor_id'))\
        .exclude(tipo_documento='NOTA DE CREDITO')
    m_dte = I(dv.aggregate(m=Sum('monto_con_iva'))['m'])
    tk_sin = Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF,
                                   estado='PAGADO', dte_generado=False,
                                   modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'])
    m_tk = I(tk_sin.aggregate(m=Sum('total'))['m'])
    ncq = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                             tipo_documento='NOTA DE CREDITO',
                             tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
                             descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'])\
        .exclude(receptor__isnull=False, receptor_id=F('emisor_id'))
    m_nc = I(ncq.aggregate(m=Sum('monto_con_iva'))['m'])
    print(f"  ORACULO brutas = DTE {m_dte:,} + tickets-sin-dte {m_tk:,} = {m_dte + m_tk:,}; NC={m_nc:,} "
          f"-> netas oraculo={m_dte + m_tk - m_nc:,}")
    print(f"  delta netas vista-oraculo = {I(k.get('ventas_actual')) - (m_dte + m_tk - m_nc):,}")
    # mismatch resta canal DTE: ticket ecom total vs monto del DTE asociado
    ecom_cd = Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF,
                                    estado='PAGADO', modulo_origen='ECOMMERCE', dte_generado=True)
    a = ecom_cd.aggregate(n=Count('id'), m=Sum('total'))
    print(f"  tickets ECOMMERCE con DTE julio: n={a['n']} total_ticket={I(a['m']):,} (se restan del canal DTE)")
except Exception as e:
    print('S5 ERROR:', type(e).__name__, e)

# ============================================================ S6 COMPARATIVA MENSUAL
try:
    sec('S6 COMPARATIVA MENSUAL (fix F-16 anti doble conteo)')
    r6 = invocar('app.views_modulo_reportes.obtener_comparativa_mensual', {}, ADMIN)
    js6 = r6['json'] or {}
    cats = js6.get('categories') or []
    series = js6.get('series') or []
    print(f"vista: status={r6['status']} nq={r6['nq']} ms={r6['ms']} err={r6['err']} cats={cats}")
    if 'Jul 2026' in cats:
        idx = cats.index('Jul 2026')
        tot_jul = sum(I((s.get('data') or [0] * len(cats))[idx]) for s in series)
        # oraculo: tickets sin dte julio + DTE VP vigentes sin internas - NC (EMITIDO/ACEPTADO, sin internas)
        t = I(Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF, estado='PAGADO',
                                    dte_generado=False, modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
                                    sucursal__isnull=False).aggregate(m=Sum('total'))['m'])
        bd = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                                tipo_transaccion='VENTA_PUBLICO', sucursal__isnull=False)\
            .exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])\
            .exclude(receptor__isnull=False, receptor_id=F('emisor_id'))
        dvp = I(bd.exclude(tipo_documento='NOTA DE CREDITO').aggregate(m=Sum('monto_con_iva'))['m'])
        dnc = I(Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                                   tipo_documento='NOTA DE CREDITO',
                                   tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
                                   descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'],
                                   sucursal__isnull=False)
               .exclude(receptor__isnull=False, receptor_id=F('emisor_id'))
               .aggregate(m=Sum('monto_con_iva'))['m'])
        esperado = t + dvp - dnc
        dup = Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF, estado='PAGADO',
                                    dte_generado=True, modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'])\
            .aggregate(n=Count('id'), m=Sum('total'))
        print(f"  julio en grafico = {tot_jul:,}  oraculo = {esperado:,}  delta = {tot_jul - esperado:,}")
        print(f"  (si duplicara habria +{I(dup['m']):,} de {dup['n']} tickets con DTE)")
        print(f"  nota: comparativa NO filtra tipo VENTA (facturas concepto): el grafico es solo VP")
    else:
        print('  Jul 2026 no esta en categories:', cats)
except Exception as e:
    print('S6 ERROR:', type(e).__name__, e)

# ============================================================ S7 VENTAS-GLOBAL
try:
    sec('S7 VENTAS-GLOBAL (scoping F-01 + NC + doble conteo)')
    r7 = invocar('app.views_modulo_reportes.obtener_ventas_global_por_empresa',
                 {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, ADMIN)
    js7 = r7['json'] or {}
    k7 = js7.get('kpis') or {}
    print(f"vista(admin): status={r7['status']} nq={r7['nq']} ms={r7['ms']} err={r7['err']}")
    print(f"  KPI total_ventas={I(k7.get('total_ventas')):,} devoluciones={I(k7.get('total_devoluciones')):,} "
          f"docs={k7.get('total_documentos')} empresas={k7.get('total_empresas')} sucursales={k7.get('total_sucursales')}")
    for e_ in js7.get('empresas') or []:
        print(f"  empresa {e_['nombre']}: ventas={I(e_['ventas']):,} dev={I(e_['devoluciones']):,} docs={e_['documentos']}")
    # oraculo replica de _sumar_periodo (admin, sin scope):
    m_tk = I(Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF, estado='PAGADO',
                                   dte_generado=False, modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
                                   sucursal__isnull=False).aggregate(m=Sum('total'))['m'])
    dq = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'], sucursal__isnull=False)\
        .exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])\
        .exclude(tipo_documento='NOTA DE CREDITO')
    m_dte = I(dq.aggregate(m=Sum('monto_con_iva'))['m'])
    ncall = Dte.objects.filter(fecha_emision__gte=FI, fecha_emision__lte=FF,
                               tipo_documento='NOTA DE CREDITO',
                               tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'])\
        .exclude(estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO'])
    m_nc_all = I(ncall.aggregate(m=Sum('monto_con_iva'))['m'])
    # NC cuya sucursal NO tiene filas de venta (la vista las bota) o sucursal null
    sids_con_data = set(dq.values_list('sucursal_id', flat=True).distinct()) | set(
        Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF, estado='PAGADO',
                              dte_generado=False, modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
                              sucursal__isnull=False).values_list('sucursal_id', flat=True).distinct())
    nc_perdidas = ncall.exclude(sucursal_id__in=list(sids_con_data)).aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"  ORACULO-replica: tickets {m_tk:,} + DTE {m_dte:,} - NC {m_nc_all:,} = {m_tk + m_dte - m_nc_all:,}")
    print(f"  KPI vista - replica = {I(k7.get('total_ventas')) - (m_tk + m_dte - m_nc_all):,}")
    print(f"  NC botadas por 'sid not in data' o sucursal NULL: n={nc_perdidas['n']} monto={I(nc_perdidas['m']):,}")
    ncdesc = ncall.filter(descartado=True).aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    vdesc = dq.filter(descartado=True).aggregate(n=Count('id'), m=Sum('monto_con_iva'))
    print(f"  incluidas por NO filtrar descartado: ventas n={vdesc['n']}/{I(vdesc['m']):,}  NC n={ncdesc['n']}/{I(ncdesc['m']):,}")

    if RESTR:
        u, suc, emps = RESTR
        rr = invocar('app.views_modulo_reportes.obtener_ventas_global_por_empresa',
                     {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, u,
                     sucursal_id=suc, empresa_id=emps[0])
        jsr = rr['json'] or {}
        emp_ids = [e_.get('id') for e_ in jsr.get('empresas') or []]
        print(f"vista(RESTRINGIDO {u.username}, empresas {emps}): status={rr['status']} ve empresas={emp_ids} err={rr['err']}")
        fuga = [e for e in emp_ids if e not in emps]
        print(f"  FUGA multi-empresa: {fuga if fuga else 'NO'}")
        rc = invocar('app.views_modulo_reportes.obtener_comisiones_por_vendedor',
                     {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'}, u,
                     sucursal_id=suc, empresa_id=emps[0])
        print(f"comisiones(RESTRINGIDO): status={rc['status']} err={(rc['json'] or {}).get('error') or rc['err']}")
        if rc['status'] == 200 and (rc['json'] or {}).get('success'):
            emps_c = {v.get('empresa_id') for v in (rc['json'].get('vendedores') or [])}
            print(f"  empresas en filas comisiones: {sorted(x for x in emps_c if x)} (permitidas {emps})")
except Exception as e:
    print('S7 ERROR:', type(e).__name__, e)

# ============================================================ S8 DIAGNOSTICO CUADRATURA
try:
    sec('S8 DIAGNOSTICO CUADRATURA VS REPORTE (dia top julio)')
    tienda = (Ticket.objects.filter(created_at__date__gte=FI, created_at__date__lte=FF,
                                    estado='PAGADO', sucursal__es_centro_distribucion=False)
              .values('sucursal_id').annotate(n=Count('id')).order_by('-n').first())
    sid = tienda['sucursal_id']
    dia = (Dte.objects.filter(sucursal_id=sid, fecha_emision__gte=FI, fecha_emision__lte=FF)
           .values('fecha_emision').annotate(n=Count('id')).order_by('-n').first())['fecha_emision']
    print(f"sucursal={sid} dia={dia}")
    r8 = invocar('app.views_modulo_reportes.api_diagnostico_cuadratura_vs_reporte',
                 {'fecha': dia.strftime('%Y-%m-%d'), 'sucursal_id': sid}, ADMIN,
                 sucursal_id=sid)
    js8 = r8['json'] or {}
    print(f"vista: status={r8['status']} nq={r8['nq']} ms={r8['ms']} err={r8['err']}")
    if js8.get('success'):
        print(f"  cuadratura_total={I(js8.get('cuadratura_total')):,} reporte_total={I(js8.get('reporte_total')):,} "
              f"diferencia={I(js8.get('diferencia')):,}")
        sc = js8.get('solo_en_cuadratura') or []
        sr = js8.get('solo_en_reporte') or []
        print(f"  solo_en_cuadratura={len(sc)} (${sum(I(d['monto']) for d in sc):,}) "
              f"solo_en_reporte={len(sr)} (${sum(I(d['monto']) for d in sr):,}) en_ambos={js8.get('en_ambos')}")
        print(f"  tickets_sin_dte={len(js8.get('tickets_sin_dte') or [])} total={I(js8.get('tickets_sin_dte_total')):,}")
        sin_motivo = [d for d in sc if not d.get('motivos_exclusion_reporte')]
        print(f"  solo_en_cuadratura sin motivo: {len(sin_motivo)}")
        # factores que el diagnostico NO explica:
        dd = Dte.objects.filter(sucursal_id=sid, fecha_emision=dia)
        venta_tipo = dd.filter(tipo_transaccion='VENTA', descartado=False,
                               estado_dte__in=['EMITIDO', 'ACEPTADO'])\
            .exclude(tipo_documento='NOTA DE CREDITO').aggregate(n=Count('id'), m=Sum('monto_con_iva'))
        print(f"  DTE tipo VENTA vigentes ese dia (reporte los suma, cuadratura NO, diag los cree 'en ambos'): "
              f"n={venta_tipo['n']} monto={I(venta_tipo['m']):,}")
        anul = dd.filter(tipo_documento='NOTA DE CREDITO', tipo_transaccion='ANULACION',
                         descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'])\
            .aggregate(n=Count('id'), m=Sum('monto_con_iva'))
        print(f"  NC ANULACION ese dia (reporte resta, cuadratura NO resta, diag 'en ambos'): n={anul['n']} monto={I(anul['m']):,}")
        # NC DEVOLUCION con fecha_pago distinta (efecto cae otro dia en cuadratura)
        ncdev = dd.filter(tipo_documento='NOTA DE CREDITO', tipo_transaccion='DEVOLUCION',
                          descartado=False, estado_dte__in=['EMITIDO', 'ACEPTADO'])
        n_desplazadas = 0
        m_desplazadas = 0
        for x in ncdev.prefetch_related('dte_asociado'):
            fp = next((p.fecha_pago for p in x.dte_asociado.all() if p.fecha_pago), None)
            if fp and fp != dia:
                n_desplazadas += 1
                m_desplazadas += I(x.monto_con_iva)
        print(f"  NC DEVOLUCION emitidas ese dia con fecha_pago OTRO dia: n={n_desplazadas} monto={m_desplazadas:,}")
        resumen = js8.get('resumen_cuadratura') or {}
        print(f"  resumen_cuadratura: venta_total={I(resumen.get('venta_total')):,} tickets={I(resumen.get('total_tickets')):,} "
              f"bpapel={I(resumen.get('total_boletas_papel')):,} nc={I(resumen.get('total_notas_credito')):,}")
except Exception as e:
    print('S8 ERROR:', type(e).__name__, e)

print('\nFIN TANDA 1')
