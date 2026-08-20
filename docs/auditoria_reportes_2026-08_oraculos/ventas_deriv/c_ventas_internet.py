# -*- coding: utf-8 -*-
# Script C: ventas-internet a fondo (julio 2026)
import sys
from datetime import date
from io import BytesIO

sys.path.insert(0, r'C:\Users\cromo\AppData\Local\Temp\claude\c--Users-cromo-Documents-DjangoProyects-SistemaRetailMind\15dd0fbe-a29a-4c80-8ad2-c70c318e71b7\scratchpad\ventas_deriv')
from _boot import invocar, ADMIN  # noqa: E402

from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum  # noqa: E402
from app.models import (CambioDevolucionDetalle, PedidoEcommerce, Ticket,
                        TicketDetallePago)  # noqa: E402

FI, FF = date(2026, 7, 1), date(2026, 7, 31)

# ---------- 1. invocacion JSON ----------
r = invocar('app.views_modulo_reportes.obtener_reporte_ventas_internet',
            {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31',
             'page_size': 100})
js = r.get('json') or {}
res = js.get('resumen') or {}
print('JSON: status', r['status'], 'ms', r['ms'], 'nq', r['nq'],
      'escrituras', len(r['escrituras']))
print('resumen:', {x: res.get(x) for x in ('venta_internet', 'tickets', 'unidades',
      'ticket_promedio', 'ecommerce_count', 'pos_count', 'boletas',
      'monto_fallback_ecommerce', 'plataformas', 'empresas', 'sucursales',
      'pedidos_sin_vendedor')})
print('alertas:', js.get('alertas'))
print('paginacion:', js.get('paginacion'))
print('plataformas:', [(p['nombre'], p['tickets'], int(p['ventas']))
                       for p in (js.get('plataformas') or [])[:10]])
print('origenes:', [(o['nombre'], o['tickets'], int(o['ventas']))
                    for o in (js.get('origenes') or [])])

# ---------- 2. oraculo universo ----------
suc_ids = None  # admin ve todas
universo = Ticket.objects.filter(
    estado='PAGADO',
    created_at__date__gte=FI, created_at__date__lte=FF,
).filter(Q(modulo_origen='ECOMMERCE') | Q(pagos__metodo_pago='VENTA_INTERNET')).distinct()
n_uni = universo.count()
print('\noraculo universo tickets:', n_uni)

# monto: pagos VENTA_INTERNET de esos tickets + fallback total
ids_uni = list(universo.values_list('id', flat=True))
pagos = TicketDetallePago.objects.filter(
    ticket_id__in=ids_uni, metodo_pago='VENTA_INTERNET',
).values('ticket_id').annotate(m=Sum('monto'))
pago_por_ticket = {p['ticket_id']: int(p['m'] or 0) for p in pagos}
info = Ticket.objects.filter(id__in=ids_uni).values('id', 'total', 'modulo_origen')
ora_monto = 0
ora_fallback = 0
n_skip = 0
n_ecom = n_pos = 0
for t in info:
    p = pago_por_ticket.get(t['id'], 0)
    es_ecom = t['modulo_origen'] == 'ECOMMERCE'
    # nota: la vista tambien considera es_ecommerce si tiene pedido asociado
    if p > 0:
        ora_monto += p
    elif es_ecom:
        ora_monto += int(t['total'] or 0)
        ora_fallback += int(t['total'] or 0)
    else:
        n_skip += 1  # POS matcheado por join pero pago 0 -> la vista lo salta
        continue
    if es_ecom:
        n_ecom += 1
    else:
        n_pos += 1
print(f'oraculo: monto={ora_monto} fallback={ora_fallback} ecommerce={n_ecom} '
      f'pos={n_pos} saltados={n_skip}')
print('DELTA vista-oraculo: monto=', int(res.get('venta_internet') or 0) - ora_monto,
      'tickets=', int(res.get('tickets') or 0) - (n_ecom + n_pos))

# tickets con pedido pero modulo != ECOMMERCE (la vista los cuenta ecommerce)
con_pedido_no_ecom = Ticket.objects.filter(
    id__in=ids_uni, pedidos_ecommerce__isnull=False,
).exclude(modulo_origen='ECOMMERCE').distinct().count()
print('tickets con pedido y modulo!=ECOMMERCE:', con_pedido_no_ecom)

# ---------- 3. que NO resta: devoluciones/NC sobre esas ventas ----------
dev = CambioDevolucionDetalle.objects.filter(
    producto_original__isnull=False, cantidad_original__gt=0,
    producto_original__idTicket_id__in=ids_uni,
).aggregate(u=Sum('cantidad_original'),
            m=Sum(ExpressionWrapper(F('precio_original_unitario') * F('cantidad_original'),
                                    output_field=DecimalField())),
            n=Count('cambio_devolucion_id', distinct=True))
print('\ndevoluciones (cualquier fecha) sobre tickets internet julio:',
      dev['n'], 'ops,', int(dev['u'] or 0), 'u, $', int(dev['m'] or 0))
dev_jul = CambioDevolucionDetalle.objects.filter(
    producto_original__isnull=False, cantidad_original__gt=0,
    cambio_devolucion__fecha_ejecucion__date__gte=FI,
    cambio_devolucion__fecha_ejecucion__date__lte=FF,
    producto_original__idTicket__in=Ticket.objects.filter(
        Q(modulo_origen='ECOMMERCE') | Q(pagos__metodo_pago='VENTA_INTERNET')),
).aggregate(m=Sum(ExpressionWrapper(F('precio_original_unitario') * F('cantidad_original'),
                                    output_field=DecimalField())),
            n=Count('id'))
print('devoluciones EJECUTADAS en julio sobre ventas internet (cualquier fecha venta):',
      dev_jul['n'], 'lineas, $', int(dev_jul['m'] or 0))

# tickets internet ANULADOS julio (excluidos del reporte)
anul = Ticket.objects.filter(
    created_at__date__gte=FI, created_at__date__lte=FF,
).exclude(estado='PAGADO').filter(
    Q(modulo_origen='ECOMMERCE') | Q(pagos__metodo_pago='VENTA_INTERNET')
).distinct().values('estado').annotate(n=Count('id'), m=Sum('total'))
print('tickets internet julio NO pagados (fuera del reporte):',
      [(a['estado'], a['n'], int(a['m'] or 0)) for a in anul])

# ---------- 4. pedidos ecommerce julio vs universo ----------
peds = PedidoEcommerce.objects.filter(
    fecha_recepcion__date__gte=FI, fecha_recepcion__date__lte=FF)
print('\npedidos ecommerce recibidos julio por estado/sub_estado:')
for x in peds.values('estado', 'sub_estado').annotate(n=Count('id'), m=Sum('total')).order_by('-n'):
    print('  ', x['estado'], '/', x['sub_estado'], ':', x['n'], 'pedidos, $', int(x['m'] or 0))
sin_ticket = peds.filter(ticket__isnull=True)
print('pedidos julio SIN ticket:', sin_ticket.count(), '$',
      int(sin_ticket.aggregate(m=Sum('total'))['m'] or 0))
fe = peds.filter(sub_estado='FACTURADO_EXTERNO')
fe_sin = fe.filter(Q(ticket__isnull=True) | ~Q(ticket__estado='PAGADO'))
print('FACTURADO_EXTERNO julio:', fe.count(), '$', int(fe.aggregate(m=Sum('total'))['m'] or 0),
      '| de esos sin ticket PAGADO (fuera del reporte):', fe_sin.count(), '$',
      int(fe_sin.aggregate(m=Sum('total'))['m'] or 0))
# pedidos con ticket PAGADO fuera del rango julio (venta contada en otro mes)
fuera_rango = peds.filter(ticket__isnull=False, ticket__estado='PAGADO').exclude(
    ticket__created_at__date__gte=FI, ticket__created_at__date__lte=FF).count()
print('pedidos julio cuyo ticket PAGADO cae fuera de julio:', fuera_rango)

# canal_origen de los pedidos de tickets del universo
por_canal = PedidoEcommerce.objects.filter(ticket_id__in=ids_uni) \
    .values('canal_origen').annotate(n=Count('id')).order_by('-n')
print('canales de pedidos en el universo julio:', [(c['canal_origen'], c['n']) for c in por_canal])

# tickets ECOMMERCE julio SIN pedido asociado (AllConnected no linkeado)
ecom_sin_pedido = Ticket.objects.filter(
    estado='PAGADO', modulo_origen='ECOMMERCE',
    created_at__date__gte=FI, created_at__date__lte=FF,
    pedidos_ecommerce__isnull=True).count()
print('tickets ECOMMERCE julio sin pedido asociado:', ecom_sin_pedido)

# ---------- 5. export excel ----------
ex = invocar('app.views_modulo_reportes.exportar_reporte_ventas_internet',
             {'fecha_inicio': '2026-07-01', 'fecha_fin': '2026-07-31'})
print('\nexcel: status', ex['status'], 'ms', ex['ms'], 'nq', ex['nq'])
resp = ex.get('resp')
if resp is not None and ex['status'] == 200 and 'spreadsheet' in resp.get('Content-Type', ''):
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
    det = wb['Detalle pedidos']
    filas = sum(1 for _ in det.iter_rows(min_row=2))
    resw = wb['Resumen']
    vals = {row[0].value: row[1].value for row in resw.iter_rows(min_row=3)}
    print('excel detalle filas=', filas, 'vs paginacion.total=',
          (js.get('paginacion') or {}).get('total'))
    print('excel resumen venta internet=', vals.get('Venta internet'),
          'tickets=', vals.get('Tickets'),
          'vs JSON', res.get('venta_internet'), res.get('tickets'))
else:
    print('excel fallo:', getattr(resp, 'content', b'')[:200])
print('FIN C')
