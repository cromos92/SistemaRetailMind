# -*- coding: utf-8 -*-
# TANDA 3 — PAOLA bias, aging jul-2026, FIFO vs ficha, Excel, permisos, cruce existencias
import io, json, sys, time
from datetime import timedelta, date
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from django.test import RequestFactory
from django.db import connection, reset_queries
from django.utils import timezone
from django.db.models import Sum, Count, F, Q, BigIntegerField
from django.db.models.functions import Abs
from django.contrib.auth import get_user_model

from app.models import (AtributoOpcion, EmpresaUser, Movimientos_Producto,
                        Producto_Talla, Sucursal, LoteProducto, PermisoRol)
from app.constants_kardex import CONCEPTOS_VENTA
from app.views_inteligencia_compra import (
    obtener_plan_liquidacion, obtener_plan_liquidacion_por_anio,
    obtener_inteligencia_compra, exportar_plan_liquidacion_excel,
    _scope_plan, _detalle_query)

BI = BigIntegerField()
P = print
User = get_user_model()
factory = RequestFactory()
hoy = timezone.localdate()

admin = (User.objects.filter(rol='administrador', is_active=True).first()
         or User.objects.filter(is_superuser=True, is_active=True).first())
emp_ids = list(EmpresaUser.objects.filter(user=admin, status=True)
               .values_list('empresa_id', flat=True).distinct())
sucs = list(Sucursal.objects.filter(empresa_id__in=emp_ids)
            .values('id', 'alias', 'es_centro_distribucion'))
tiendas_ids = [s['id'] for s in sucs if not s['es_centro_distribucion']]


def req_as(user, params=None, ajax=False):
    kw = {'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'} if ajax else {}
    r = factory.get('/x', params or {}, **kw)
    r.user = user
    eu = EmpresaUser.objects.filter(user=user, status=True).first()
    suc = eu.sucursal_id if eu and eu.sucursal_id else tiendas_ids[0]
    emp = eu.empresa_id if eu else emp_ids[0]
    r.session = {'idSucursalActual': suc, 'idEmpresaActual': emp}
    return r


# ---- 1. PAOLA: fila del ranking vs ttm limpio ----
paola = AtributoOpcion.objects.filter(atributo__nombre__icontains='marca',
                                      valor__iexact='PAOLA').first()
P('PAOLA id=%s' % (paola.id if paola else None))
if paola:
    pt_p = Producto_Talla.objects.filter(stock__gt=0,
                                         producto__excluir_de_analitica=False,
                                         producto__sucursal_id__in=tiendas_ids,
                                         producto__atributo1_id=paola.id)
    inv = pt_p.aggregate(s_tot=Sum('stock', output_field=BI),
                         v_tot=Sum(F('stock') * F('producto__costo'), output_field=BI))
    mv = Movimientos_Producto.objects.filter(
        estado='COMPLETADO', concepto__in=CONCEPTOS_VENTA,
        ProductoTalla__producto__sucursal_id__in=tiendas_ids,
        ProductoTalla__producto__atributo1_id=paola.id,
        fecha__gte=hoy - timedelta(days=365))
    sucio = mv.aggregate(u=Sum(Abs('cantidad'), output_field=BI),
                         v=Sum(Abs(F('cantidad')) * F('precio'), output_field=BI),
                         c=Sum(Abs(F('cantidad')) * F('costo'), output_field=BI))
    limpio = mv.filter(ProductoTalla__producto__excluir_de_analitica=False).aggregate(
        u=Sum(Abs('cantidad'), output_field=BI),
        v=Sum(Abs(F('cantidad')) * F('precio'), output_field=BI),
        c=Sum(Abs(F('cantidad')) * F('costo'), output_field=BI))
    P('PAOLA inv=%s' % inv)
    P('PAOLA ttm VISTA(sin filtro analitica)=%s vs LIMPIO=%s' % (sucio, limpio))
    for tag, t in (('VISTA', sucio), ('LIMPIO', limpio)):
        u, v, c = t['u'] or 0, t['v'] or 0, t['c'] or 0
        s, val = inv['s_tot'] or 0, inv['v_tot'] or 0
        rot = round(u / s, 2) if s else None
        cob = round(s / (u / 12.0), 1) if u else None
        marg = (v - c) if (v > 0 and 0 < c < v) else None
        gmroi = round(marg / val, 2) if (marg and val) else None
        P('PAOLA %s: rotacion=%s cobertura=%s gmroi=%s' % (tag, rot, cob, gmroi))

# ---- 2. aging jul-2026: capital con lote recreado pero producto viejo ----
r = req_as(admin)
base_pt, mov_base, ctx = _scope_plan(r)
qs = _detalle_query(base_pt, mov_base, hoy)
jul = qs.filter(fecha_lote__gte=date(2026, 6, 1), fecha_lote__lt=date(2026, 9, 1))
o_jul = jul.aggregate(n=Count('id'), stock=Sum('stock_u', output_field=BI),
                      valor=Sum(F('costo') * F('stock_u'), output_field=BI))
o_jul_v = jul.filter(fecha_creacion__lt=date(2025, 10, 1)).aggregate(
    n=Count('id'), stock=Sum('stock_u', output_field=BI),
    valor=Sum(F('costo') * F('stock_u'), output_field=BI))
o_jul_v2 = jul.filter(fecha_creacion__lt=date(2025, 1, 1)).aggregate(
    n=Count('id'), stock=Sum('stock_u', output_field=BI),
    valor=Sum(F('costo') * F('stock_u'), output_field=BI))
P('AGING lote jun-ago2026: %s ; producto creado <oct2025: %s ; <ene2025: %s' % (
    o_jul, o_jul_v, o_jul_v2))
# total capital que la vista clasifica como anio 2026
o_2026 = qs.filter(fecha_lote__gte=date(2026, 1, 1)).aggregate(
    valor=Sum(F('costo') * F('stock_u'), output_field=BI),
    stock=Sum('stock_u', output_field=BI))
P('AGING capital anio-2026 por lote: %s' % o_2026)

# ---- 3. FIFO vs ficha (global tiendas) ----
lots = LoteProducto.objects.filter(
    activo=True, agotado=False, cantidad_disponible__gt=0,
    producto_talla__producto__sucursal_id__in=tiendas_ids,
    producto_talla__producto__excluir_de_analitica=False,
    producto_talla__stock__gt=0)
o_fifo = lots.aggregate(u=Sum('cantidad_disponible', output_field=BI),
                        v_fifo=Sum(F('cantidad_disponible') * F('costo_unitario'), output_field=BI),
                        v_ficha=Sum(F('cantidad_disponible') * F('producto_talla__producto__costo'), output_field=BI))
P('FIFO global: %s (plano=97032 / $1.529.634.125 ficha)' % o_fifo)

# ---- 4. Excel export vs JSON ----
t0 = time.perf_counter()
resp = exportar_plan_liquidacion_excel(req_as(admin))
ms = round((time.perf_counter() - t0) * 1000)
P('EXCEL status=%s ms=%s bytes=%s' % (resp.status_code, ms, len(resp.content)))
try:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    tot_pares = 0
    hojas = []
    res_costo = res_pares = None
    for ws in wb.worksheets:
        if ws.title in ('Resumen', 'Filtros'):
            continue
        pares_h = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[1] == 'TOTAL':
                pares_h = row[8] or 0
                break
        hojas.append((ws.title, pares_h))
        tot_pares += pares_h
    ws_r = wb['Resumen']
    res_pares = res_costo = 0
    seccion1 = True
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            seccion1 = False
            continue
        if seccion1 and isinstance(row[2], (int, float)):
            res_pares += row[2] or 0
            res_costo += row[3] or 0
    P('EXCEL hojas=%s' % hojas)
    P('EXCEL tot_pares=%s resumen{pares:%s costo:%s} vs ranking{97032 / 1529634125}' % (
        tot_pares, res_pares, res_costo))
except Exception as e:
    P('EXCEL parse error: %r' % e)

# ---- 5. permisos con usuario restringido ----
restr = None
for u in User.objects.filter(is_active=True, is_superuser=False).exclude(
        rol__in=('administrador', 'jefe_local', 'administracion'))[:40]:
    if EmpresaUser.objects.filter(user=u, status=True).exists():
        restr = u
        break
if restr:
    tiene = PermisoRol.tiene_permiso(restr, 'plan_liquidacion', 'puede_ver', None)
    P('RESTR user=%s rol=%s tiene_permiso(plan_liquidacion)=%s' % (
        restr.username, restr.rol, tiene))
    r1 = obtener_plan_liquidacion(req_as(restr, ajax=True))
    P('RESTR obtener_plan_liquidacion -> %s' % r1.status_code)
    r2 = obtener_plan_liquidacion_por_anio(req_as(restr, ajax=True))
    body2 = json.loads(r2.content) if r2.status_code == 200 else {}
    tot2 = body2.get('totales', {})
    P('RESTR obtener_plan_liquidacion_por_anio -> %s totales=%s' % (r2.status_code, tot2))
    marca_sk = AtributoOpcion.objects.filter(atributo__nombre__icontains='marca',
                                             valor__icontains='SKECHERS').first()
    r3 = obtener_inteligencia_compra(req_as(restr, {'marca_id': marca_sk.id,
                                                    'sucursal_id': 'todas'}, ajax=True))
    if r3.status_code == 200:
        b3 = json.loads(r3.content)
        if b3.get('success'):
            aliases = [x['alias'] for x in b3['data']['stock_suc']]
            P('RESTR inteligencia -> 200 stock_suc=%s inv_costo=%s' % (
                aliases, b3['data']['finanzas']['inv_costo']))
        else:
            P('RESTR inteligencia -> 200 error=%s' % b3.get('error'))
    else:
        P('RESTR inteligencia -> %s' % r3.status_code)
    emps_r = list(EmpresaUser.objects.filter(user=restr, status=True)
                  .values_list('empresa_id', flat=True).distinct())
    sucs_r = [s['alias'] for s in Sucursal.objects.filter(empresa_id__in=emps_r)
              .values('id', 'alias', 'es_centro_distribucion')]
    P('RESTR empresas=%s sucursales=%s' % (emps_r, sucs_r))
else:
    P('RESTR: no hay usuario restringido')

# ---- 6. cruce existencias-sucursal (NICK2 id=7) ----
SID = 7
pt_n = Producto_Talla.objects.filter(producto__sucursal_id=SID, stock__gt=0,
                                     producto__excluir_de_analitica=False)
stk_n = pt_n.aggregate(s=Sum('stock', output_field=BI))['s'] or 0
v30_origen = abs(Movimientos_Producto.objects.filter(
    concepto__in=CONCEPTOS_VENTA, estado='COMPLETADO',
    fecha__gte=hoy - timedelta(days=30), sucursal_origen_id=SID,
).aggregate(s=Sum('cantidad'))['s'] or 0)
v30_prod = Movimientos_Producto.objects.filter(
    concepto__in=CONCEPTOS_VENTA, estado='COMPLETADO',
    fecha__gte=hoy - timedelta(days=30),
    ProductoTalla__producto__sucursal_id=SID,
).aggregate(s=Sum(Abs('cantidad'), output_field=BI))['s'] or 0
u365_n = Movimientos_Producto.objects.filter(
    concepto__in=CONCEPTOS_VENTA, estado='COMPLETADO',
    fecha__gte=hoy - timedelta(days=365),
    ProductoTalla__producto__sucursal_id=SID,
).aggregate(s=Sum(Abs('cantidad'), output_field=BI))['s'] or 0
viejo_sinfiltro = LoteProducto.objects.filter(
    activo=True, cantidad_disponible__gt=0,
    producto_talla__producto__sucursal_id=SID,
    fecha_ingreso__date__lte=hoy - timedelta(days=181),
).aggregate(s=Sum('cantidad_disponible', output_field=BI))['s'] or 0
viejo_limpio = LoteProducto.objects.filter(
    activo=True, cantidad_disponible__gt=0,
    producto_talla__producto__sucursal_id=SID,
    producto_talla__producto__excluir_de_analitica=False,
    producto_talla__stock__gt=0,
    fecha_ingreso__date__lte=hoy - timedelta(days=181),
).aggregate(s=Sum('cantidad_disponible', output_field=BI))['s'] or 0
# dead-stock a la inteligencia (sin venta 180d) para la misma tienda
vend180_n = Movimientos_Producto.objects.filter(
    concepto__in=CONCEPTOS_VENTA, estado='COMPLETADO',
    fecha__gte=hoy - timedelta(days=180),
    ProductoTalla__producto__sucursal_id=SID).values('ProductoTalla')
dead_n = pt_n.exclude(id__in=vend180_n).aggregate(s=Sum('stock', output_field=BI))['s'] or 0
P('CRUCE NICK2: stock=%s v30_origen=%s v30_producto=%s u365=%s' % (
    stk_n, v30_origen, v30_prod, u365_n))
if v30_origen:
    P('CRUCE cobertura_dias(existencias-suc, v30 origen)=%s' % int(stk_n / (v30_origen / 30.0)))
if u365_n:
    P('CRUCE cobertura_meses(plan, u365)=%s' % round(stk_n / (u365_n / 12.0), 1))
P('CRUCE pct_stock_viejo(existencias: lotes>180d sin filtro)=%s%% (%s u)' % (
    round(100.0 * viejo_sinfiltro / stk_n, 1) if stk_n else 0, viejo_sinfiltro))
P('CRUCE pct_stock_viejo limpio=%s%% (%s u) ; dead180(sin venta)=%s%% (%s u)' % (
    round(100.0 * viejo_limpio / stk_n, 1) if stk_n else 0, viejo_limpio,
    round(100.0 * dead_n / stk_n, 1) if stk_n else 0, dead_n))

# ---- 7. delta kardex vs tickets 90d SKECHERS: composicion por concepto ----
sk = AtributoOpcion.objects.filter(atributo__nombre__icontains='marca',
                                   valor__icontains='SKECHERS').first()
comp = list(Movimientos_Producto.objects.filter(
    ProductoTalla__producto__atributo1_id=sk.id,
    ProductoTalla__producto__excluir_de_analitica=False,
    ProductoTalla__producto__sucursal__es_centro_distribucion=False,
    ticket__isnull=False,
    fecha__gte=hoy - timedelta(days=90), fecha__lt=hoy)
    .values('concepto', 'estado')
    .annotate(n=Count('id'), u=Sum(Abs('cantidad'), output_field=BI))
    .order_by('-u'))
P('DELTA mov con ticket 90d por concepto: %s' % comp)
P('FIN T3')
