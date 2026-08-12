"""
Módulo de Ventas - RetailMind
Contiene todas las vistas relacionadas con ventas, tickets, vendedores y POS
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, Http404, HttpResponseBadRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model

User = get_user_model()
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import (
    Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg,
    Case, When, IntegerField, Value, Exists, OuterRef, Prefetch, Subquery,
)
from django.db.models.functions import TruncDate, Coalesce
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
import json
import os
import re
import logging
from datetime import datetime, timedelta

# Importar funciones necesarias desde views.py
from .views import obtener_siguiente_correlativo, obtener_correlativo_existente, consumir_stock_fifo

# Helpers compartidos del módulo ventas (métodos de pago, agrupación, only() estándar)
from .utils_ventas import (
    obtener_nombre_metodo_pago,
    agrupar_metodos_pago,
    formatear_metodos_pago_str,
    get_sucursal_id,
    puede_editar_campo_dte,
    permisos_edicion_dte_context,
    persistir_costeo_fifo,
    puede_cambiar_tipo_dte,
    son_tipos_compatibles,
    tipos_compatibles_para,
    CODIGO_PERMISO_TIPO_DTE,
    ONLY_DTE_PRODUCTO,
    ONLY_DTE_PAGO,
    ONLY_TICKET_PRODUCTO_POS,
)
from .utils_permisos import obtener_configuracion_rango_arqueo, obtener_sucursales_usuario
from .decorators import requiere_rol

# Caching del módulo ventas (Redis / LocMem, ver app.cache_utils)
from .cache_utils import cache_ventas_json

# Importar servicios de Transbank
from .services.transbank_sdk_service import (
    run_transbank_operation, test_pos_connection, 
    execute_pos_sale, get_available_ports, cancel_pos_sale
)
from .services.inventario_service import (
    ingresar as ingresar_inventario,
    egresar as egresar_inventario,
)

from .models import (
    Ticket, Ticket_Productos, TicketDetallePago, TicketReferencia, Vendedor, Producto, Producto_Talla,
    Sucursal, EmpresaUser, Empresa, Movimientos_Producto, LoteProducto, Dte, Dte_Productos, Dte_Detalle_Pago,
    Correlativo, ESTADO_TICKET_CHOICES, METODO_PAGO_TICKET_CHOICES, ORIGEN_PAGO_CHOICES, TIPO_DOCUMENTO_CHOICES,
    ArqueoCaja, ESTADO_ARQUEO_CHOICES, RESULTADO_REVISION_CHOICES, GrupoDeposito, DepositoBancario,
    ObservacionArqueo, LogAccionCaja, log_accion_caja,
    ConfiguracionPOS, TransaccionPOS, LogPOS,
    TIPO_POS_CHOICES, ESTADO_TRANSACCION_POS_CHOICES, TIPO_TARJETA_CHOICES,
    # Modelos de Cambios y Devoluciones
    CambioDevolucion, CambioDevolucionDetalle, PagoCambioDevolucion, HistorialCambioDevolucion,
    TIPO_OPERACION_CAMBIO_CHOICES, ESTADO_CAMBIO_CHOICES, MOTIVO_CAMBIO_CHOICES, CONDICION_PRODUCTO_CHOICES,
    METODO_DEVOLUCION_NC_CHOICES,
    CodigoAutorizacionDinamico, RegistroAutorizacion, PermisoTemporalCambio,
    PermisoRol,
)


# ========== GESTIÓN DE VENDEDORES ==========

logger = logging.getLogger('app')

# Guard de cobro: exige que los pagos cubran el total antes de marcar un ticket
# como PAGADO. Se puede desactivar por entorno (VALIDAR_COBERTURA_PAGOS=0) si
# algún flujo legítimo quedara bloqueado, sin necesidad de tocar código.
VALIDAR_COBERTURA_PAGOS = os.environ.get('VALIDAR_COBERTURA_PAGOS', '1') != '0'
# Holgura en pesos para redondeos entre el front y el backend.
TOLERANCIA_COBERTURA_PAGOS = int(os.environ.get('TOLERANCIA_COBERTURA_PAGOS', '1'))


ACCIONES_TEMPORALES_CAMBIO = {
    PermisoTemporalCambio.ACCION_CANCELAR,
    PermisoTemporalCambio.ACCION_REVERTIR,
}
MINUTOS_PERMISO_TEMPORAL_VALIDOS = {15, 30, 60, 480}


class ConflictoInventarioCambio(Exception):
    """Conflicto de stock/sucursal que debe abortar toda la transaccion."""

    def __init__(self, mensaje, codigo='STOCK_CHANGED'):
        super().__init__(mensaje)
        self.codigo = codigo


def _vendedores_elegibles_para_sucursal(sucursal):
    """Vendedores activos que pueden operar en la empresa/sucursal indicada.

    Compatibilidad legacy:
    - empresa actual + sucursal exacta;
    - empresa actual sin asignaciones M2M (vendedor company-wide);
    - empresa nula + sucursal exacta.
    """
    return Vendedor.objects.filter(activo=True).filter(
        Q(empresa_id=sucursal.empresa_id, sucursales=sucursal)
        | Q(empresa_id=sucursal.empresa_id, sucursales__isnull=True)
        | Q(empresa__isnull=True, sucursales=sucursal)
    ).distinct()


def _vendedores_para_autorizacion_cambio(sucursal):
    """Vendedores que pueden autorizar/figurar en un cambio.

    A diferencia de _vendedores_elegibles_para_sucursal, NO restringe por
    sucursal NI por empresa: cualquier vendedor activo del sistema (de
    cualquier tienda o empresa) sirve para autorizar/figurar en el cambio. La
    sucursal del cambio sigue determinando el inventario; solo se relaja quién
    puede firmar. El parámetro `sucursal` se conserva por compatibilidad con
    los llamadores.
    """
    return Vendedor.objects.filter(activo=True)


def _bloquear_y_validar_inventario_cambio(detalles, sucursal_id, reversion=False):
    """Bloquea productos en orden estable y valida el stock final agregado."""
    entradas = {}
    salidas = {}

    def sumar(destino, producto_id, cantidad):
        if producto_id and cantidad and cantidad > 0:
            destino[producto_id] = destino.get(producto_id, 0) + cantidad

    for detalle in detalles:
        producto_original_id = (
            detalle.producto_original.ProductoTalla_id
            if detalle.producto_original_id else None
        )
        if detalle.apto_para_venta:
            sumar(entradas, producto_original_id, detalle.cantidad_original)
        sumar(salidas, detalle.producto_nuevo_id, detalle.cantidad_nueva)

    ids_producto = sorted(set(entradas) | set(salidas))
    bloqueados = list(
        Producto_Talla.objects.select_for_update()
        .select_related('producto')
        .filter(id__in=ids_producto)
        .order_by('id')
    )
    productos = {producto.id: producto for producto in bloqueados}

    if len(productos) != len(ids_producto):
        raise ConflictoInventarioCambio(
            'Uno de los productos del cambio ya no existe',
            codigo='PRODUCT_NOT_FOUND',
        )

    for producto_id in ids_producto:
        producto = productos[producto_id]
        if producto.producto.sucursal_id != int(sucursal_id):
            raise ConflictoInventarioCambio(
                f'El SKU {producto.sku} no pertenece a la sucursal del cambio',
                codigo='INVALID_PRODUCT_BRANCH',
            )

        if reversion:
            stock_final = (
                producto.stock
                + salidas.get(producto_id, 0)
                - entradas.get(producto_id, 0)
            )
        else:
            stock_final = (
                producto.stock
                + entradas.get(producto_id, 0)
                - salidas.get(producto_id, 0)
            )

        if stock_final < 0:
            operacion = 'revertir' if reversion else 'ejecutar'
            raise ConflictoInventarioCambio(
                f'Stock insuficiente para {operacion} el cambio: SKU {producto.sku}. '
                f'Disponible {producto.stock}, resultado proyectado {stock_final}.',
                codigo='STOCK_REVERSAL_CONFLICT' if reversion else 'STOCK_CHANGED',
            )

    return productos


def _usuario_es_administrador_activo(usuario):
    return bool(
        usuario
        and usuario.is_authenticated
        and usuario.is_active
        and getattr(usuario, 'es_activo', True)
        and getattr(usuario, 'rol', '') == 'administrador'
    )


def _autorizacion_fuera_plazo_previa(cambio):
    """Administrador que ya autorizó la excepción de plazo al crear la solicitud.

    El código dinámico es de un solo uso. Si al aprobar se vuelve a exigir un
    código de administrador, el mismo administrador tiene que emitir un segundo
    código para la misma operación: esa segunda firma no agrega control (la
    excepción ya quedó autorizada y registrada en el cambio) y en la práctica
    deja la solicitud creada pero imposible de aprobar hasta ubicar de nuevo al
    administrador.
    """
    if not cambio.es_fuera_de_plazo:
        return None
    autorizador = cambio.autorizado_por_usuario
    if _usuario_es_administrador_activo(autorizador):
        return autorizador
    return None


def _usuario_tiene_permiso_base_cambios(usuario, sucursal_id):
    if _usuario_es_administrador_activo(usuario):
        return True
    if (
        usuario
        and usuario.is_authenticated
        and usuario.is_active
        and getattr(usuario, 'es_activo', True)
        and getattr(usuario, 'rol', '') == 'jefe_local'
    ):
        return True
    return PermisoRol.tiene_permiso(
        usuario,
        'cambios_devoluciones',
        'puede_eliminar',
        sucursal_id=sucursal_id,
    )


def _permiso_temporal_vigente(usuario, cambio, accion):
    return PermisoTemporalCambio.vigente_para(
        usuario=usuario,
        empresa_id=cambio.sucursal.empresa_id,
        sucursal_id=cambio.sucursal_id,
        accion=accion,
    )


def _acciones_cambio_para_usuario(usuario, cambio):
    """Fuente de verdad para visibilidad y habilitación de acciones destructivas."""
    es_admin = _usuario_es_administrador_activo(usuario)
    tiene_base = _usuario_tiene_permiso_base_cambios(usuario, cambio.sucursal_id)
    permiso_cancelar = None if es_admin else _permiso_temporal_vigente(
        usuario, cambio, PermisoTemporalCambio.ACCION_CANCELAR
    )
    permiso_revertir = None if es_admin else _permiso_temporal_vigente(
        usuario, cambio, PermisoTemporalCambio.ACCION_REVERTIR
    )

    estado_cancelable = cambio.estado in ('SOLICITADO', 'APROBADO') and not cambio.ticket_nuevo_id
    ticket_pendiente = bool(cambio.ticket_nuevo_id and cambio.ticket_nuevo.estado == 'PENDIENTE')
    estado_revertible = cambio.estado in (
        'EJECUTADO', 'EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE'
    )
    integridad_reversion = not cambio.diferencia_condonada and not cambio.nc_generada

    cancelar_autorizado = estado_cancelable and tiene_base and (es_admin or bool(permiso_cancelar))
    revertir_autorizado = (
        estado_revertible and ticket_pendiente and integridad_reversion
        and tiene_base and (es_admin or bool(permiso_revertir))
    )

    vigencias = [
        permiso.vigente_hasta for permiso in (permiso_cancelar, permiso_revertir) if permiso
    ]
    return {
        'cancelar': cancelar_autorizado,
        'revertir': revertir_autorizado,
        'condonar': es_admin and cambio.estado == 'EJECUTADO_COBRO_PENDIENTE',
        'ajustar': es_admin and cambio.estado == 'EJECUTADO_COBRO_PENDIENTE',
        'puede_solicitar_cancelar': estado_cancelable and tiene_base,
        'puede_solicitar_revertir': (
            estado_revertible and ticket_pendiente and integridad_reversion and tiene_base
        ),
        'requiere_autorizacion_cancelar': (
            estado_cancelable and tiene_base and not cancelar_autorizado
        ),
        'requiere_autorizacion_revertir': (
            estado_revertible and ticket_pendiente and integridad_reversion
            and tiene_base and not revertir_autorizado
        ),
        'permiso_temporal_hasta': (
            timezone.localtime(max(vigencias)).strftime('%d/%m/%Y %H:%M') if vigencias else None
        ),
    }


def _registrar_intento_permiso_temporal(
    request, cambio, accion, exitoso, descripcion, autorizador=None,
    codigo_obj=None, sucursal_autorizador=None, minutos=None,
):
    return RegistroAutorizacion.objects.create(
        codigo_usado=codigo_obj if exitoso else None,
        usuario_solicitante=request.user,
        usuario_autorizador=autorizador if exitoso else None,
        tipo_operacion='OTRO',
        descripcion=descripcion,
        ip_origen=request.META.get('REMOTE_ADDR'),
        exitoso=exitoso,
        cambio_devolucion=cambio,
        sucursal_solicitante=cambio.sucursal,
        sucursal_autorizador=sucursal_autorizador if exitoso else None,
        es_cross_branch=bool(
            exitoso and sucursal_autorizador
            and sucursal_autorizador.id != cambio.sucursal_id
        ),
        requiere_revision=bool(
            exitoso and sucursal_autorizador
            and sucursal_autorizador.id != cambio.sucursal_id
        ),
        datos_adicionales={
            'accion': accion,
            'minutos': minutos,
            'cambio_id': cambio.id,
        },
    )


def _otorgar_permiso_temporal_desde_codigo(request, cambio, accion, codigo, motivo, minutos):
    if accion not in ACCIONES_TEMPORALES_CAMBIO:
        return None, JsonResponse({
            'success': False, 'code': 'INVALID_ACTION', 'error': 'Acción temporal no permitida'
        }, status=400)

    if not _usuario_tiene_permiso_base_cambios(request.user, cambio.sucursal_id):
        return None, JsonResponse({
            'success': False,
            'code': 'PERMISSION_DENIED',
            'error': 'Su perfil no tiene permiso para solicitar esta acción',
        }, status=403)

    try:
        minutos = int(minutos or 30)
    except (TypeError, ValueError):
        minutos = 30
    if minutos not in MINUTOS_PERMISO_TEMPORAL_VALIDOS:
        return None, JsonResponse({
            'success': False,
            'code': 'INVALID_DURATION',
            'error': 'Duración de autorización no permitida',
        }, status=400)

    hace_15_min = timezone.now() - timezone.timedelta(minutes=15)
    intentos_fallidos = RegistroAutorizacion.objects.filter(
        usuario_solicitante=request.user,
        tipo_operacion='OTRO',
        exitoso=False,
        fecha_hora__gte=hace_15_min,
        descripcion__icontains='permiso temporal de cambio',
    ).count()
    if intentos_fallidos >= 5:
        return None, JsonResponse({
            'success': False,
            'code': 'AUTH_LOCKED',
            'error': 'Demasiados intentos fallidos. Intente nuevamente en 15 minutos.',
        }, status=429)

    es_valido, mensaje, codigo_obj = CodigoAutorizacionDinamico.validar_codigo(codigo)
    if not es_valido or not codigo_obj:
        _registrar_intento_permiso_temporal(
            request, cambio, accion, False,
            f'Intento fallido de permiso temporal de cambio: {mensaje}',
            minutos=minutos,
        )
        return None, JsonResponse({
            'success': False,
            'code': 'INVALID_AUTH_CODE',
            'error': mensaje,
        }, status=403)

    # of=('self',): bloquea solo la fila del código. Sin esto,
    # select_related('generado_por') (FK nullable → LEFT OUTER JOIN) provoca
    # "FOR UPDATE cannot be applied to the nullable side of an outer join" en PostgreSQL.
    codigo_obj = CodigoAutorizacionDinamico.objects.select_for_update(of=('self',)).select_related(
        'generado_por'
    ).get(id=codigo_obj.id)
    if not codigo_obj.es_valido():
        _registrar_intento_permiso_temporal(
            request, cambio, accion, False,
            'Intento fallido de permiso temporal de cambio: código vencido o utilizado',
            minutos=minutos,
        )
        return None, JsonResponse({
            'success': False,
            'code': 'INVALID_AUTH_CODE',
            'error': 'Código de autorización vencido o ya utilizado',
        }, status=403)

    administrador = codigo_obj.generado_por
    if not _usuario_es_administrador_activo(administrador):
        _registrar_intento_permiso_temporal(
            request, cambio, accion, False,
            'Intento fallido de permiso temporal de cambio: autorizador no es administrador activo',
            minutos=minutos,
        )
        return None, JsonResponse({
            'success': False,
            'code': 'INVALID_AUTHORIZER',
            'error': 'El código no pertenece a un administrador activo',
        }, status=403)

    asignacion_admin = EmpresaUser.objects.filter(
        user=administrador,
        empresa_id=cambio.sucursal.empresa_id,
        status=True,
    ).select_related('sucursal').order_by('-active').first()
    if not asignacion_admin:
        _registrar_intento_permiso_temporal(
            request, cambio, accion, False,
            'Intento fallido de permiso temporal de cambio: administrador de otra empresa',
            minutos=minutos,
        )
        return None, JsonResponse({
            'success': False,
            'code': 'CROSS_COMPANY_AUTH',
            'error': 'El administrador debe pertenecer a la misma empresa',
        }, status=403)

    ahora = timezone.now()
    vigente_hasta = ahora + timezone.timedelta(minutes=minutos)
    codigo_obj.usado = True
    codigo_obj.save(update_fields=['usado'])
    registro = _registrar_intento_permiso_temporal(
        request,
        cambio,
        accion,
        True,
        f'Permiso temporal de cambio otorgado por {administrador.get_full_name() or administrador.username}',
        autorizador=administrador,
        codigo_obj=codigo_obj,
        sucursal_autorizador=asignacion_admin.sucursal,
        minutos=minutos,
    )
    permiso = PermisoTemporalCambio.objects.create(
        usuario=request.user,
        empresa_id=cambio.sucursal.empresa_id,
        sucursal=cambio.sucursal,
        accion=accion,
        otorgado_por=administrador,
        codigo_autorizacion=codigo_obj,
        motivo=motivo,
        vigente_desde=ahora,
        vigente_hasta=vigente_hasta,
    )
    logger.info(
        'Permiso temporal cambio otorgado usuario=%s accion=%s cambio=%s autorizador=%s registro=%s hasta=%s',
        request.user.username, accion, cambio.id, administrador.username, registro.id, vigente_hasta.isoformat(),
    )
    return permiso, None

@login_required
def gestion_vendedores(request):
    """Vista principal para gestión de vendedores"""
    # Obtener todas las sucursales disponibles
    sucursales = Sucursal.objects.all().order_by('alias')
    
    context = {
        'sucursales': sucursales
    }
    
    return render(request, 'vistas/modulo_administracion/gestion_vendedores.html', context)


@require_GET
@login_required
def obtener_vendedores(request):
    """API para obtener lista de vendedores con paginación y filtros"""
    try:
        # Parámetros de paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        # Parámetros de filtro
        search = request.GET.get('search', '').strip()
        estado = request.GET.get('estado', '')
        
        # Construir queryset base
        queryset = Vendedor.objects.all()
        
        # Aplicar filtros
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo_vendedor__icontains=search) |
                Q(correo__icontains=search) |
                Q(rut__icontains=search)
            )
        
        # Ordenar
        queryset = queryset.order_by('nombre')
        
        # Paginación
        paginator = Paginator(queryset, per_page)
        vendedores_page = paginator.get_page(page)
        
        # Serializar datos
        vendedores_data = []
        for vendedor in vendedores_page:
            sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
            vendedores_data.append({
                'id': vendedor.id,
                'codigo_vendedor': vendedor.codigo_vendedor,
                'nombre': vendedor.nombre,
                'rut': vendedor.rut,
                'correo': vendedor.correo,
                'comision': float(vendedor.comision) if vendedor.comision else 0,
                'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                'sucursales': sucursales_list,
                'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                'activo': True,  # Por defecto, ya que el modelo no tiene este campo
            })
        
        return JsonResponse({
            'success': True,
            'vendedores': vendedores_data,
            'pagination': {
                'current_page': vendedores_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': vendedores_page.has_next(),
                'has_previous': vendedores_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener vendedores: {str(e)}'
        }, status=500)


@require_GET
@login_required
def obtener_metricas_vendedores(request):
    """API para obtener métricas de vendedores"""
    try:
        total_vendedores = Vendedor.objects.count()
        vendedores_activos = Vendedor.objects.filter(activo=True).count()
        vendedores_inactivos = total_vendedores - vendedores_activos
        
        # Métricas de ventas (últimos 30 días)
        fecha_inicio = timezone.now() - timezone.timedelta(days=30)
        
        ventas_por_vendedor = Ticket.objects.filter(
            created_at__gte=fecha_inicio,
            estado='PAGADO'
        ).values('vendedor__nombre').annotate(
            total_ventas=Sum('total'),
            cantidad_tickets=Count('id')
        ).order_by('-total_ventas')[:5]
        
        return JsonResponse({
            'success': True,
            'metricas': {
                'total_vendedores': total_vendedores,
                'vendedores_activos': vendedores_activos,
                'vendedores_inactivos': vendedores_inactivos,
                'top_vendedores': list(ventas_por_vendedor)
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener métricas: {str(e)}'
        }, status=500)


def crear_vendedor(request):
    """Crear nuevo vendedor"""
    if request.method == 'GET':
        return render(request, 'vistas/modulo_administracion/crear_vendedor.html')
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar campos requeridos
            campos_requeridos = ['codigo_vendedor', 'nombre']
            for campo in campos_requeridos:
                if not data.get(campo):
                    return JsonResponse({
                        'success': False,
                        'error': f'El campo {campo} es requerido'
                    }, status=400)
            
            # Verificar que el código no exista
            if Vendedor.objects.filter(codigo_vendedor=data['codigo_vendedor']).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe un vendedor con ese código'
                }, status=400)
            
            # Crear vendedor
            vendedor = Vendedor.objects.create(
                codigo_vendedor=data['codigo_vendedor'],
                nombre=data['nombre'],
                rut=data.get('rut', ''),
                correo=data.get('correo', ''),
                comision=data.get('comision', 0),
                fecha_nacimiento=data.get('fecha_nacimiento') or None
            )
            
            # Asignar sucursales
            if 'sucursales' in data and data['sucursales']:
                sucursales_ids = data['sucursales'] if isinstance(data['sucursales'], list) else [data['sucursales']]
                vendedor.sucursales.set(sucursales_ids)
            
            sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
            
            return JsonResponse({
                'success': True,
                'message': 'Vendedor creado exitosamente',
                'vendedor': {
                    'id': vendedor.id,
                    'codigo_vendedor': vendedor.codigo_vendedor,
                    'nombre': vendedor.nombre,
                    'rut': vendedor.rut,
                    'correo': vendedor.correo,
                    'comision': float(vendedor.comision),
                    'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                    'sucursales': sucursales_list,
                    'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                    'activo': True
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear vendedor: {str(e)}'
            }, status=500)


@require_http_methods(["PUT"])
@login_required
@transaction.atomic
@csrf_exempt
def editar_vendedor(request):
    """Editar vendedor existente"""
    try:
        data = json.loads(request.body)
        logger.debug("editar_vendedor datos recibidos: %s", data)
        vendedor_id = data.get('id')
        
        if not vendedor_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de vendedor requerido'
            }, status=400)
        
        vendedor = get_object_or_404(Vendedor, id=vendedor_id)
        
        # Validar campos requeridos
        campos_requeridos = ['codigo_vendedor', 'nombre']
        for campo in campos_requeridos:
            if not data.get(campo):
                return JsonResponse({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }, status=400)
        
        # Verificar que el código no exista en otro vendedor
        if Vendedor.objects.filter(codigo_vendedor=data['codigo_vendedor']).exclude(id=vendedor_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'Ya existe otro vendedor con ese código'
            }, status=400)
        
        # Actualizar vendedor
        vendedor.codigo_vendedor = data['codigo_vendedor']
        vendedor.nombre = data['nombre']
        vendedor.rut = data.get('rut', '')
        vendedor.correo = data.get('correo', '')
        vendedor.comision = data.get('comision', 0)
        vendedor.fecha_nacimiento = data.get('fecha_nacimiento') or None
        vendedor.save()
        
        # Actualizar sucursales
        if 'sucursales' in data:
            sucursales_ids = data['sucursales'] if isinstance(data['sucursales'], list) else ([data['sucursales']] if data['sucursales'] else [])
            vendedor.sucursales.set(sucursales_ids)
        
        sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
        
        return JsonResponse({
            'success': True,
            'message': 'Vendedor actualizado exitosamente',
            'vendedor': {
                'id': vendedor.id,
                'codigo_vendedor': vendedor.codigo_vendedor,
                'nombre': vendedor.nombre,
                'rut': vendedor.rut,
                'correo': vendedor.correo,
                'comision': float(vendedor.comision),
                'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                'sucursales': sucursales_list,
                'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                'activo': True
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al editar vendedor: {str(e)}'
        }, status=500)


@require_http_methods(["DELETE"])
@login_required
@transaction.atomic
@csrf_exempt
def eliminar_vendedor(request, vendedor_id):
    """Eliminar vendedor (soft delete)"""
    try:
        vendedor = get_object_or_404(Vendedor, id=vendedor_id)
        
        # Verificar si tiene tickets asociados
        tickets_count = Ticket.objects.filter(vendedor=vendedor).count()
        
        if tickets_count > 0:
            # Soft delete - marcar como inactivo
            vendedor.activo = False
            vendedor.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Vendedor desactivado (tiene {tickets_count} tickets asociados)'
            })
        else:
            # Hard delete si no tiene tickets
            vendedor.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Vendedor eliminado exitosamente'
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar vendedor: {str(e)}'
        }, status=500)


@require_GET
@login_required
def exportar_vendedores(request):
    """Exportar lista de vendedores a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendedores"
        
        # Encabezados
        headers = [
            'Código', 'Nombre', 'Email', 'Teléfono', 
            'Comisión %', 'Estado', 'Fecha Creación'
        ]
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Obtener datos
        vendedores = Vendedor.objects.all().order_by('nombre')
        
        # Escribir datos
        for row, vendedor in enumerate(vendedores, 2):
            ws.cell(row=row, column=1, value=vendedor.codigo)
            ws.cell(row=row, column=2, value=vendedor.nombre)
            ws.cell(row=row, column=3, value=vendedor.email)
            ws.cell(row=row, column=4, value=vendedor.telefono)
            ws.cell(row=row, column=5, value=float(vendedor.comision_porcentaje) if vendedor.comision_porcentaje else 0)
            ws.cell(row=row, column=6, value='Activo' if vendedor.activo else 'Inactivo')
            ws.cell(row=row, column=7, value=vendedor.fecha_creacion.strftime('%d/%m/%Y') if vendedor.fecha_creacion else '')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="vendedores.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        }, status=500)


# ========== TICKET DE VENTA ==========

@login_required
def ticket_venta(request):
    """Vista principal para crear tickets de venta"""
    # Obtener sucursal actual del usuario (intentar ambas variables de sesión)
    sucursal_actual_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    sucursal_actual = None
    empresa_actual_nombre = request.session.get('nombreEmpresaActual', 'Sin empresa')
    
    # Si hay sucursal actual, obtenerla
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    # Si no hay sucursal actual, obtener las sucursales disponibles para el usuario
    sucursales_disponibles = []
    if not sucursal_actual:
        sucursales_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True,
            sucursal__isnull=False
        ).select_related('sucursal', 'empresa').distinct()
        
        sucursales_disponibles = [
            {
                'sucursal': eu.sucursal,
                'empresa': eu.empresa
            }
            for eu in sucursales_usuario
        ]
    
    # Validar que existe correlativo para tickets
    tiene_correlativo = False
    correlativo_info = None
    
    logger.debug(
        "ticket_venta session context: idSucursalActual=%s sucursalActual=%s sucursal_actual_id=%s sucursal_actual=%s",
        request.session.get('idSucursalActual'),
        request.session.get('sucursalActual'),
        sucursal_actual_id,
        sucursal_actual,
    )
    
    if sucursal_actual:
        logger.debug(
            "Buscando correlativo TICKET para sucursal=%s id=%s",
            sucursal_actual.alias,
            sucursal_actual.id,
        )
        try:
            correlativo = Correlativo.objects.get(
                sucursal=sucursal_actual,
                tipo_dte='TICKET'
            )
            logger.debug(
                "Correlativo TICKET encontrado: id=%s inicio=%s termino=%s puede_emitir=%s",
                correlativo.id,
                correlativo.inicio,
                correlativo.termino,
                correlativo.puede_emitir(),
            )
            
            tiene_correlativo = correlativo.puede_emitir()
            correlativo_info = {
                'disponibles': correlativo.disponibles,
                'inicio': correlativo.inicio,
                'termino': correlativo.termino,
                'estado': correlativo.estado
            }
        except Correlativo.DoesNotExist:
            logger.warning(
                "Correlativo TICKET no encontrado para sucursal_id=%s",
                sucursal_actual.id,
            )
            tiene_correlativo = False
            correlativo_info = None
    else:
        logger.warning(
            "ticket_venta sin sucursal actual: user_id=%s sucursal_actual_id=%s",
            request.user.id,
            sucursal_actual_id,
        )
    
    # Obtener vendedores de la sucursal actual (solo activos)
    if sucursal_actual:
        # Obtener vendedores activos asignados a esta sucursal
        vendedores = Vendedor.objects.filter(
            sucursales=sucursal_actual,
            activo=True
        ).order_by('nombre')
        
        # Si no hay vendedores asignados a la sucursal, buscar por empresa
        if not vendedores.exists() and sucursal_actual.empresa:
            vendedores = Vendedor.objects.filter(
                empresa=sucursal_actual.empresa,
                activo=True
            ).order_by('nombre')
    else:
        # Si no hay sucursal seleccionada, mostrar solo vendedores activos
        vendedores = Vendedor.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'sucursal_actual': sucursal_actual,
        'empresa_actual_nombre': empresa_actual_nombre,
        'vendedores': vendedores,
        'sucursales_disponibles': sucursales_disponibles,
        'necesita_seleccionar_sucursal': not sucursal_actual,
        'tiene_correlativo': tiene_correlativo,
        'correlativo_info': correlativo_info,
    }
    
    return render(request, 'vistas/modulo_ventas/ticket_venta.html', context)


@login_required
def buscar_vendedor_por_codigo(request):
    """Vista AJAX para buscar vendedor por código"""
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({
            'success': False,
            'error': 'Código de vendedor requerido'
        })
    
    try:
        vendedor = Vendedor.objects.get(codigo=codigo, activo=True)
        return JsonResponse({
            'success': True,
            'vendedor': {
                'id': vendedor.id,
                'codigo': vendedor.codigo,
                'nombre': vendedor.nombre,
                'comision_porcentaje': float(vendedor.comision_porcentaje) if vendedor.comision_porcentaje else 0
            }
        })
    except Vendedor.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vendedor no encontrado o inactivo'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar vendedor: {str(e)}'
        })


def buscar_producto_por_sku(request):
    """Buscar producto por SKU para ticket de venta"""
    sku = request.GET.get('sku', '').strip()
    sucursal_id = request.session.get('idSucursalActual')
    
    if not sku:
        return JsonResponse({
            'success': False,
            'error': 'SKU requerido'
        })
    
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal activa'
        })
    
    try:
        # Buscar producto por SKU (tolerante a SKUs duplicados en BD legacy)
        from .utils_producto_match import producto_talla_por_sku
        producto_talla = producto_talla_por_sku(
            sku, sucursal_id=sucursal_id,
            select_related=['producto', 'producto__categoria', 'producto__atributo1',
                            'producto__atributo2', 'producto__atributo3', 'producto__atributo4'])
        if not producto_talla:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado'})

        # Verificar stock en la sucursal
        stock_actual = producto_talla.stock_sucursal(sucursal_id)
        
        if stock_actual <= 0:
            return JsonResponse({
                'success': False,
                'error': 'Producto sin stock en esta sucursal'
            })
        
        return JsonResponse({
            'success': True,
            'producto': {
                'id': producto_talla.id,
                'sku': producto_talla.sku,
                'articulo': producto_talla.producto.articulo,
                'descripcion': producto_talla.producto.descripcion or '',
                'talla': producto_talla.talla if producto_talla.talla else 'Sin talla',
                'precio_venta': float(producto_talla.producto.precioventa),
                'stock': stock_actual,
                'marca': producto_talla.producto.atributo1.valor if producto_talla.producto.atributo1 else '',
                'color': producto_talla.producto.atributo2.valor if producto_talla.producto.atributo2 else '',
                'material': producto_talla.producto.atributo3.valor if producto_talla.producto.atributo3 else '',
                'categoria': producto_talla.producto.categoria.nombre if producto_talla.producto.categoria else ''
            }
        })
        
    except Producto_Talla.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar producto: {str(e)}'
        })


@require_GET
@login_required
def buscar_productos_pos_avanzado(request):
    """
    Búsqueda avanzada de productos para POS.

    Optimización:
    - Filtra `stock__gt=0` directamente en DB (antes: recorría hasta 200
      candidatos en memoria para encontrar 30 con stock).
    - Reduce columnas con `only()`.
    """
    sucursal_id = get_sucursal_id(request)

    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal activa'
        })

    search_term = request.GET.get('search', '').strip()
    if not search_term or len(search_term) < 2:
        return JsonResponse({
            'success': False,
            'error': 'Ingrese al menos 2 caracteres para buscar'
        })

    try:
        productos_query = (
            Producto_Talla.objects
            .filter(producto__sucursal_id=sucursal_id, stock__gt=0)
            .select_related(
                'producto',
                'producto__categoria',
                'producto__atributo1',
                'producto__atributo2',
                'producto__atributo3',
                'producto__atributo4',
            )
        )

        palabras = [p for p in search_term.split() if p.strip()]
        for palabra in palabras:
            # `sku` es BigIntegerField: un __icontains fuerza CAST a texto y un
            # seq scan de ~605.000 filas en CADA tecla. Mismo criterio que el
            # buscador de existencias: si el término es numérico, match exacto
            # contra el índice; si no, el SKU no participa del OR.
            filtros_palabra = (
                Q(producto__articulo__icontains=palabra) |
                Q(producto__atributo1__valor__icontains=palabra) |
                Q(producto__atributo2__valor__icontains=palabra) |
                Q(producto__atributo3__valor__icontains=palabra) |
                Q(producto__atributo4__valor__icontains=palabra) |
                Q(producto__categoria__nombre__icontains=palabra) |
                Q(talla__icontains=palabra)
            )
            if palabra.strip().isdigit():
                filtros_palabra = filtros_palabra | Q(sku=int(palabra.strip()))
            productos_query = productos_query.filter(filtros_palabra)

        productos_con_stock = []
        for pt in productos_query[:30]:
            prod = pt.producto
            productos_con_stock.append({
                'id': pt.id,
                'sku': pt.sku,
                'articulo': prod.articulo,
                'descripcion': prod.descripcion or '',
                'marca': prod.atributo1.valor if prod.atributo1 else '',
                'color': prod.atributo2.valor if prod.atributo2 else '',
                'material': prod.atributo3.valor if prod.atributo3 else '',
                'talla': pt.talla if pt.talla else 'Sin talla',
                'stock': max(0, pt.stock or 0),
                'precio_venta': float(prod.precioventa),
                'categoria': prod.categoria.nombre if prod.categoria else '',
            })

        return JsonResponse({
            'success': True,
            'productos': productos_con_stock,
            'total': len(productos_con_stock),
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        })


def buscar_productos_bodega(request):
    """Buscar productos en bodega para ticket de venta.

    Optimización:
    - Filtra `stock__gt=0` en DB (antes: recorría 20 resultados en memoria).
    - Arreglo de relaciones rotas (`producto__marca__nombre` → `atributo1`,
      `pt.precio_venta` → `pt.producto.precioventa`).
    - Restringe a la sucursal activa (antes devolvía productos de cualquier sucursal).
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})

    try:
        data = json.loads(request.body)
        termino = data.get('termino', '').strip()
        sucursal_id = get_sucursal_id(request)

        if not termino:
            return JsonResponse({
                'success': False,
                'error': 'Término de búsqueda requerido'
            })

        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa'
            })

        # Igual que arriba: nada de __icontains sobre el SKU numérico.
        filtros_termino = (
            Q(producto__articulo__icontains=termino) |
            Q(producto__atributo1__valor__icontains=termino)
        )
        if termino.strip().isdigit():
            filtros_termino = filtros_termino | Q(sku=int(termino.strip()))

        productos_query = (
            Producto_Talla.objects
            .filter(producto__sucursal_id=sucursal_id, stock__gt=0)
            .select_related(
                'producto',
                'producto__categoria',
                'producto__atributo1',
            )
            .filter(filtros_termino)
        )

        productos_data = []
        for pt in productos_query[:20]:
            prod = pt.producto
            productos_data.append({
                'id': pt.id,
                'sku': pt.sku,
                'nombre': prod.articulo,
                'talla': pt.talla if pt.talla else 'Sin talla',
                'precio_venta': float(prod.precioventa),
                'stock': max(0, pt.stock or 0),
                'marca': prod.atributo1.valor if prod.atributo1 else '',
                'categoria': prod.categoria.nombre if prod.categoria else '',
            })

        return JsonResponse({
            'success': True,
            'productos': productos_data
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en búsqueda: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def crear_ticket(request):
    """Crear nuevo ticket de venta.

    Antes estaba `@csrf_exempt` y sin autenticación: cualquiera podía crear
    tickets contra el sistema sin sesión.
    """
    try:
        data = json.loads(request.body)

        # ticket_venta.html envía { cliente: {...} } anidado; aplanar para que el
        # resto de la vista pueda usar data.get('cliente_nombre') etc.
        _cli_raw = data.get('cliente')
        if isinstance(_cli_raw, dict):
            data.setdefault('cliente_nombre', _cli_raw.get('nombre', ''))
            data.setdefault('cliente_rut', _cli_raw.get('rut', ''))
            data.setdefault('cliente_email', _cli_raw.get('email', ''))
            data.setdefault('cliente_telefono', _cli_raw.get('telefono', '') or _cli_raw.get('celular', ''))
            data.setdefault('cliente_celular', _cli_raw.get('celular', ''))
            data.setdefault('cliente_fecha_nacimiento', _cli_raw.get('fecha_nacimiento', ''))

        # Validaciones básicas
        vendedor_id = data.get('vendedor_id')
        productos = data.get('productos', [])
        metodo_pago = data.get('metodo_pago')
        
        if not vendedor_id:
            return JsonResponse({
                'success': False,
                'error': 'Vendedor requerido'
            })
        
        if not productos:
            return JsonResponse({
                'success': False,
                'error': 'Debe agregar al menos un producto'
            })
        
        if not metodo_pago:
            return JsonResponse({
                'success': False,
                'error': 'Método de pago requerido'
            })
        
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa'
            })
        
        with transaction.atomic():
            # Obtener vendedor y sucursal
            vendedor = get_object_or_404(Vendedor, id=vendedor_id)
            sucursal = get_object_or_404(Sucursal, id=sucursal_id)
            
            # Obtener límite de descuento del rol del usuario
            # Usamos Max para obtener el valor más alto guardado (todos deberían ser iguales)
            from .models import PermisoRol
            from django.db.models import Max
            limite_descuento_rol = 0
            if request.user.is_authenticated:
                rol_usuario = getattr(request.user, 'rol', None)
                if rol_usuario:
                    resultado = PermisoRol.objects.filter(rol=rol_usuario).aggregate(
                        max_limite=Max('limite_descuento_porcentaje')
                    )
                    if resultado['max_limite'] is not None:
                        limite_descuento_rol = float(resultado['max_limite'])
            
            # Validar promos NxM: las líneas gratis validadas quedan exentas del
            # límite de descuento por rol (un 100% legítimo, no un descuento manual).
            from .services.campanas_service import validar_promos_nxm_payload
            val_promo = validar_promos_nxm_payload(productos, sucursal)
            if not val_promo['ok']:
                return JsonResponse({
                    'success': False,
                    'error': 'Promoción NxM inválida: ' + '; '.join(
                        e['error'] for e in val_promo['errores']),
                    'error_tipo': 'PROMO_INVALIDA',
                }, status=400)
            idx_promo_exentos = set(val_promo['lineas_ok'].keys())

            # Validar descuentos por producto contra el límite del rol
            for idx, item in enumerate(productos):
                if idx in idx_promo_exentos:
                    continue  # línea gratis de promo NxM validada
                descuento_unitario = item.get('descuento_unitario', 0)
                precio_unitario = item.get('precio_unitario', 0)

                if descuento_unitario > 0 and precio_unitario > 0:
                    porcentaje_descuento = (descuento_unitario / precio_unitario) * 100

                    # Validar que no exceda el límite del rol
                    if porcentaje_descuento > limite_descuento_rol and limite_descuento_rol > 0:
                        return JsonResponse({
                            'success': False,
                            'error': f'El descuento aplicado ({porcentaje_descuento:.1f}%) excede el límite permitido para tu rol ({limite_descuento_rol}%). Producto: {item.get("articulo", "")}'
                        })

                    # Si el límite es 0, no permitir ningún descuento
                    if limite_descuento_rol == 0:
                        return JsonResponse({
                            'success': False,
                            'error': 'No tienes permisos para aplicar descuentos. Contacta al administrador.'
                        })
            
            # Obtener siguiente correlativo
            correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
            
            # Calcular totales
            subtotal = 0
            for item in productos:
                subtotal += item['cantidad'] * item['precio_unitario']
            
            descuento = data.get('descuento', 0)
            total = subtotal - descuento
            
            # Formatear RUT del cliente antes de guardar
            cliente_rut_raw = data.get('cliente_rut', '')
            cliente_rut_formateado = formatear_rut(cliente_rut_raw) if cliente_rut_raw else ''
            
            # Crear ticket
            ticket = Ticket.objects.create(
                correlativo=correlativo,
                vendedor=vendedor,
                sucursal=sucursal,
                subtotal=subtotal,
                descuento=descuento,
                total=total,
                estado='PENDIENTE',
                observaciones=data.get('observaciones', ''),
                cliente_nombre=data.get('cliente_nombre', ''),
                cliente_rut=cliente_rut_formateado,
                cliente_email=data.get('cliente_email', ''),
                cliente_telefono=data.get('cliente_telefono', '')
            )
            
            # Crear productos del ticket
            for item in productos:
                producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
                
                # Validar cantidad
                cantidad = item.get('cantidad', 0)
                
                # Validar que la cantidad sea un número entero positivo
                if not isinstance(cantidad, int) or cantidad < 1:
                    raise ValidationError(
                        f'Cantidad inválida para {producto_talla.sku}: debe ser un número entero positivo mayor a 0'
                    )
                
                # Verificar stock
                stock_actual = producto_talla.stock_sucursal(sucursal_id)
                if stock_actual < cantidad:
                    raise ValidationError(
                        f'Stock insuficiente para {producto_talla.sku}. Solicitado: {cantidad}, Disponible: {stock_actual}'
                    )
                
                Ticket_Productos.objects.create(
                    ticket=ticket,
                    productoTalla=producto_talla,
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio_unitario'],
                    descuento_unitario=item.get('descuento_unitario', 0)
                )
            
            # Crear detalle de pago
            TicketDetallePago.objects.create(
                ticket=ticket,
                metodo_pago=metodo_pago,
                monto=total,
                referencia=data.get('referencia_pago', ''),
                observaciones=data.get('observaciones_pago', '')
            )
            
            # Si el pago es efectivo o débito, marcar como pagado
            if metodo_pago in ['EFECTIVO', 'DEBITO']:
                ticket.estado = 'PAGADO'
                ticket.fecha_pago = timezone.now()
                ticket.save()
                
                # Consumir stock FIFO
                for item in productos:
                    producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
                    consumir_stock_fifo(
                        producto_talla=producto_talla,
                        cantidad_requerida=item['cantidad'],
                        responsable=request.user,
                        ticket=ticket,
                        observaciones=f'Venta ticket #{correlativo}'
                    )
            
            # Guardar cliente en la base de datos si tiene datos completos
            if ticket.cliente_rut and ticket.cliente_nombre:
                cliente_datos = {
                    'nombre': ticket.cliente_nombre,
                    'rut': ticket.cliente_rut,
                    'email': ticket.cliente_email,
                    'telefono': ticket.cliente_telefono,
                    'celular': data.get('cliente_celular', '') or '',
                    'fecha_nacimiento': data.get('cliente_fecha_nacimiento', '') or '',
                }
                guardar_o_actualizar_cliente(cliente_datos, request.user)

        return JsonResponse({
            'success': True,
            'message': 'Ticket creado exitosamente',
            'ticket_id': ticket.id,
            'correlativo': correlativo,
            'ticket_data': construir_ticket_data(ticket)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear ticket: {str(e)}'
        })


# ========== TICKETS DE VENTA - GESTIÓN ==========

@login_required
@require_POST
@transaction.atomic
def crear_ticket_venta(request):
    """Crear ticket de venta al público.

    OJO (deuda conocida, no corregida aquí): esta ruta descuenta el stock DOS
    veces — llama a `consumir_stock_fifo()`, que ya descuenta, y después a
    `registrar_movimiento_producto()`, que vuelve a descontar. Hay 3 tests que
    fallan por esto desde antes. No se toca en este cambio para no mezclar el
    arreglo de autenticación con un cambio de comportamiento de stock.
    """
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        vendedor_id = data.get('vendedor_id')
        productos = data.get('productos', [])
        
        if not vendedor_id:
            return JsonResponse({'success': False, 'error': 'Vendedor requerido'})
        
        if not productos:
            return JsonResponse({'success': False, 'error': 'Debe incluir al menos un producto'})
        
        # Obtener sucursal actual
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        vendedor = get_object_or_404(Vendedor, id=vendedor_id)
        
        # Validar que existe correlativo antes de crear el ticket
        try:
            correlativo_obj = Correlativo.objects.get(
                sucursal=sucursal,
                tipo_dte='TICKET'
            )
            if not correlativo_obj.puede_emitir():
                return JsonResponse({
                    'success': False, 
                    'error': f'No hay correlativos disponibles para TICKET en {sucursal.alias}. Por favor, configure un nuevo rango de correlativos.'
                })
        except Correlativo.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': f'No existe correlativo configurado para TICKET en {sucursal.alias}. Por favor, configure un correlativo antes de crear tickets.'
            })
        
        # Obtener siguiente correlativo
        correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
        
        # Calcular totales
        subtotal = sum(item['cantidad'] * item['precio_unitario'] for item in productos)
        descuento = data.get('descuento', 0)
        total = subtotal - descuento
        
        # Formatear RUT del cliente antes de guardar
        cliente_rut_raw = data.get('cliente_rut', '')
        cliente_rut_formateado = formatear_rut(cliente_rut_raw) if cliente_rut_raw else ''
        
        # Crear ticket
        ticket = Ticket.objects.create(
            correlativo=correlativo,
            vendedor=vendedor,
            sucursal=sucursal,
            subtotal=subtotal,
            descuento=descuento,
            total=total,
            estado='PAGADO',
            fecha_pago=timezone.now(),
            observaciones=data.get('observaciones', ''),
            cliente_nombre=data.get('cliente_nombre', ''),
            cliente_rut=cliente_rut_formateado,
            cliente_email=data.get('cliente_email', ''),
            cliente_telefono=data.get('cliente_telefono', '')
        )
        
        # Procesar productos
        for item in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
            
            # Validar cantidad
            cantidad = item.get('cantidad', 0)
            
            # Validar que la cantidad sea un número entero positivo
            if not isinstance(cantidad, int) or cantidad < 1:
                raise ValidationError(
                    f'Cantidad inválida para {producto_talla.sku}: debe ser un número entero positivo mayor a 0'
                )
            
            # Verificar stock disponible
            stock_actual = producto_talla.stock_sucursal(sucursal_id)
            if stock_actual < cantidad:
                raise ValidationError(
                    f'Stock insuficiente para {producto_talla.sku}. Solicitado: {cantidad}, Disponible: {stock_actual}'
                )
            
            # Crear detalle del ticket
            Ticket_Productos.objects.create(
                ticket=ticket,
                productoTalla=producto_talla,
                cantidad=item['cantidad'],
                precio_unitario=item['precio_unitario'],
                descuento_unitario=item.get('descuento_unitario', 0)
            )
            
            # Consumir stock FIFO
            consumir_stock_fifo(
                producto_talla=producto_talla,
                cantidad_requerida=item['cantidad'],
                responsable=request.user,
                ticket=ticket,
                observaciones=f'Venta ticket #{correlativo}'
            )
        
        # Crear detalle de pago
        metodo_pago = data.get('metodo_pago', 'EFECTIVO')
        TicketDetallePago.objects.create(
            ticket=ticket,
            metodo_pago=metodo_pago,
            monto=total,
            referencia=data.get('referencia_pago', ''),
            observaciones=data.get('observaciones_pago', '')
        )
        
        # Guardar cliente en la base de datos si tiene datos completos
        if ticket.cliente_rut and ticket.cliente_nombre:
            cliente_datos = {
                'nombre': ticket.cliente_nombre,
                'rut': ticket.cliente_rut,
                'email': ticket.cliente_email,
                'telefono': ticket.cliente_telefono,
                'celular': data.get('cliente_celular', '') or '',
                'fecha_nacimiento': data.get('cliente_fecha_nacimiento', '') or '',
            }
            guardar_o_actualizar_cliente(cliente_datos, request.user)

        return JsonResponse({
            'success': True,
            'message': 'Ticket creado exitosamente',
            'ticket_id': ticket.id,
            'correlativo': correlativo
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al crear ticket: {str(e)}'})


@require_GET
@login_required
def obtener_tickets_venta(request):
    """Obtener lista de tickets de venta con filtros"""
    try:
        # Parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        vendedor_id = request.GET.get('vendedor_id')
        estado = request.GET.get('estado')
        sucursal_id = request.session.get('idSucursalActual')
        
        # Construir queryset
        queryset = Ticket.objects.select_related('vendedor', 'sucursal').filter(
            sucursal_id=sucursal_id
        )
        
        # Aplicar filtros
        if fecha_inicio:
            queryset = queryset.filter(created_at__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(created_at__date__lte=fecha_fin)
        if vendedor_id:
            queryset = queryset.filter(vendedor_id=vendedor_id)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        # Ordenar por fecha descendente
        queryset = queryset.order_by('-created_at')
        
        # Paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        paginator = Paginator(queryset, per_page)
        tickets_page = paginator.get_page(page)
        
        # Serializar datos
        tickets_data = []
        for ticket in tickets_page:
            tickets_data.append({
                'id': ticket.id,
                'correlativo': ticket.correlativo,
                'vendedor': ticket.vendedor.nombre,
                'fecha_creacion': ticket.created_at.strftime('%d/%m/%Y %H:%M'),
                'subtotal': float(ticket.subtotal),
                'descuento': float(ticket.descuento),
                'total': float(ticket.total),
                'estado': ticket.estado,
                'cliente_nombre': ticket.cliente_nombre or '',
                'cliente_rut': ticket.cliente_rut or ''
            })
        
        return JsonResponse({
            'success': True,
            'tickets': tickets_data,
            'pagination': {
                'current_page': tickets_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': tickets_page.has_next(),
                'has_previous': tickets_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener tickets: {str(e)}'
        })


# ========== DASHBOARD POS ==========

def _get_qz_config(sucursal_id):
    """Retorna el dict qz_config para la sucursal dada (o defaults si no hay)."""
    if not sucursal_id:
        return {'habilitado': False, 'nombre_impresora': 'EPSON TM-T20II'}
    try:
        suc = Sucursal.objects.get(id=sucursal_id)
        return {
            'habilitado': getattr(suc, 'usar_qz_tray', False),
            'nombre_impresora': getattr(suc, 'nombre_impresora_termica', 'EPSON TM-T20II') or 'EPSON TM-T20II',
        }
    except Sucursal.DoesNotExist:
        return {'habilitado': False, 'nombre_impresora': 'EPSON TM-T20II'}


@login_required
def pos_dashboard(request):
    """Vista principal del dashboard POS"""
    # Obtener choices para los selects
    sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    
    # Obtener configuración POS guardada (para auto-conectar)
    config_pos = None
    if sucursal_id:
        config_pos = ConfiguracionPOS.objects.filter(
            sucursal_id=sucursal_id,
            tipo_pos='SDK_SERIAL',
            activo=True
        ).first()
    
    # Obtener límite de descuento del rol del usuario
    # Usamos Max para obtener el valor más alto guardado (todos deberían ser iguales)
    limite_descuento_rol = 0
    if request.user.is_authenticated:
        from .models import PermisoRol
        from django.db.models import Max
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario:
            resultado = PermisoRol.objects.filter(rol=rol_usuario).aggregate(
                max_limite=Max('limite_descuento_porcentaje')
            )
            if resultado['max_limite'] is not None:
                limite_descuento_rol = float(resultado['max_limite'])
    
    # Verificar si el usuario es administrador
    es_admin = getattr(request.user, 'rol', '') in ['administrador', 'administracion']

    context = {
        'metodo_pago_choices': METODO_PAGO_TICKET_CHOICES,
        'estado_ticket_choices': ESTADO_TICKET_CHOICES,
        'config_pos': config_pos,
        'limite_descuento_rol': limite_descuento_rol,
        'es_admin': es_admin,
        'qz_config': _get_qz_config(sucursal_id),
    }
    return render(request, 'vistas/modulo_ventas/generacionVentas.html', context)


@login_required
@require_GET
@cache_ventas_json('correlativos_disp', timeout=30)
def verificar_correlativos_disponibles(request):
    """API para verificar correlativos disponibles por tipo de documento"""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Tipos de documento para ventas al público (electrónicos y papel)
        tipos_documento = [
            'BOLETA_ELECTRONICA', 
            'BOLETA_PAPEL',
            'FACTURA_ELECTRONICA'
        ]
        
        correlativos_info = {}
        
        for tipo in tipos_documento:
            # Mapear nombres para la base de datos
            tipo_db = tipo
            if tipo == 'BOLETA_ELECTRONICA':
                tipo_db = 'BOLETA ELECTRONICA'
            elif tipo == 'BOLETA_PAPEL':
                tipo_db = 'BOLETA PAPEL'
            elif tipo == 'FACTURA_ELECTRONICA':
                tipo_db = 'FACTURA ELECTRONICA'
            
            try:
                correlativo = obtener_correlativo_existente(sucursal, tipo_db)
                if not correlativo:
                    raise Correlativo.DoesNotExist()
                
                correlativos_info[tipo] = {
                    'disponible': correlativo.puede_emitir(),
                    'numero_actual': correlativo.numero_actual,
                    'disponibles': correlativo.disponibles,
                    'estado': correlativo.estado,
                    'porcentaje_consumo': round(correlativo.porcentaje_consumo, 1),
                    'rango': f"{correlativo.inicio}-{correlativo.termino}"
                }
                
            except Correlativo.DoesNotExist:
                # Si no existe, crear uno automáticamente
                try:
                    # Esto creará el correlativo si no existe
                    numero = obtener_siguiente_correlativo(sucursal, tipo_db)
                    
                    # Obtener el correlativo recién creado
                    correlativo = obtener_correlativo_existente(sucursal, tipo_db)
                    if not correlativo:
                        raise Correlativo.DoesNotExist()
                    
                    correlativos_info[tipo] = {
                        'disponible': True,
                        'numero_actual': correlativo.numero_actual,
                        'disponibles': correlativo.disponibles,
                        'estado': correlativo.estado,
                        'porcentaje_consumo': round(correlativo.porcentaje_consumo, 1),
                        'rango': f"{correlativo.inicio}-{correlativo.termino}",
                        'recien_creado': True
                    }
                    
                except Exception as e:
                    correlativos_info[tipo] = {
                        'disponible': False,
                        'error': f'Error al crear correlativo: {str(e)}',
                        'numero_actual': 0,
                        'disponibles': 0,
                        'estado': 'error'
                    }

        return JsonResponse({
            'success': True,
            'correlativos': correlativos_info,
            'sucursal': {
                'id': sucursal.id,
                'alias': sucursal.alias
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar correlativos: {str(e)}'
        })


def _check_stock_ticket(ticket, sucursal_id):
    """
    Revisa si algún producto del ticket tiene stock insuficiente en la sucursal.
    Retorna dict con 'tiene_stock_insuf' y 'productos_stock_insuf'.

    IMPORTANTE: depende de que `ticket_productos__ProductoTalla__producto`
    estén prefetch_related. No gatilla queries adicionales a la DB.
    """
    sucursal_id_int = int(sucursal_id) if sucursal_id is not None else None
    problemas = []
    for tp in ticket.ticket_productos.all():
        pt = tp.ProductoTalla
        if not pt or not pt.producto:
            continue
        # Solo comparar stock si el producto pertenece a la sucursal actual
        if sucursal_id_int is None or pt.producto.sucursal_id != sucursal_id_int:
            continue
        stock_real = max(0, pt.stock or 0)
        cantidad_pedida = tp.stock  # campo 'stock' en Ticket_Productos = cantidad
        if stock_real < cantidad_pedida:
            problemas.append({
                'sku': str(pt.sku),
                'articulo': pt.producto.articulo or 'Sin nombre',
                'talla': str(pt.talla or ''),
                'stock_real': stock_real,
                'cantidad_pedida': int(cantidad_pedida),
            })
    return {
        'tiene_stock_insuf': len(problemas) > 0,
        'productos_stock_insuf': problemas,
    }


@login_required
def dashboard_stats(request):
    """API para obtener estadísticas del dashboard POS.

    Optimización clave:
    - Un solo `aggregate` condicional para ventas/cantidades (antes: 3 queries).
    - `_check_stock_ticket` usa únicamente datos prefetch (sin gatillar queries).
    - `ticket_productos.count()` evita la relación rota `ticket.productos`.
    - Nombre de sucursal leído con `get()` explícito, no con lambda opaca.
    """
    try:
        sucursal_id = get_sucursal_id(request)
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        from datetime import datetime, time
        hoy = timezone.localdate()
        inicio_dia = timezone.make_aware(datetime.combine(hoy, time.min))
        fin_dia = timezone.make_aware(datetime.combine(hoy, time.max))

        tickets_hoy = Ticket.objects.filter(
            sucursal_id=sucursal_id,
            created_at__range=[inicio_dia, fin_dia],
        )
        tickets_venta = tickets_hoy.exclude(modulo_origen='CAMBIO_DEVOLUCION')

        # Un solo aggregate para todos los contadores/sumas del día
        agg = tickets_venta.aggregate(
            ventas_hoy=Sum('total', filter=Q(estado='PAGADO')),
            tickets_pendientes=Count('id', filter=Q(estado='PENDIENTE')),
            tickets_pagados=Count('id', filter=Q(estado='PAGADO')),
        )
        ventas_hoy = agg['ventas_hoy'] or 0
        tickets_pendientes = agg['tickets_pendientes'] or 0
        tickets_pagados = agg['tickets_pagados'] or 0
        promedio_venta = (ventas_hoy / tickets_pagados) if tickets_pagados > 0 else 0

        # Tickets pendientes del día (últimos 20) - INCLUYENDO AMBOS TIPOS
        tickets_recientes = tickets_hoy.filter(estado='PENDIENTE').select_related(
            'vendedor', 'sucursal'
        ).prefetch_related(
            'ticket_productos__ProductoTalla__producto',
        ).order_by('-created_at')[:20]
        
        tickets_data = []
        for ticket in tickets_recientes:
            # Determinar tipo de ticket
            if ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
                tipo_ticket = 'Diferencia Cambio'
                tipo_ticket_class = 'warning'
                # Determinar si es cobro o devolución
                if 'A DEVOLVER AL CLIENTE' in (ticket.observaciones or ''):
                    tipo_detalle = 'Devolución'
                elif 'A COBRAR AL CLIENTE' in (ticket.observaciones or ''):
                    tipo_detalle = 'Cobro'
                else:
                    tipo_detalle = 'Cambio Directo'
            else:
                tipo_ticket = 'Venta'
                tipo_ticket_class = 'primary'
                tipo_detalle = 'Venta Normal'
            
            tickets_data.append({
                'correlativo': ticket.correlativo,
                'hora': ticket.created_at.strftime('%H:%M'),
                'cliente_nombre': ticket.cliente_nombre or 'Sin nombre',
                'cliente_rut': ticket.cliente_rut or '',
                'vendedor_nombre': f"{ticket.vendedor.codigo_vendedor} - {ticket.vendedor.nombre}" if ticket.vendedor else 'Sin vendedor',
                'total': int(ticket.total or 0),
                'estado': ticket.estado,
                'tipo_ticket': tipo_ticket,
                'tipo_ticket_class': tipo_ticket_class,
                'tipo_detalle': tipo_detalle,
                'modulo_origen': ticket.modulo_origen,
                **_check_stock_ticket(ticket, sucursal_id),
            })

        # Tickets pendientes para el wizard (con más detalles)
        tickets_pendientes_query = tickets_hoy.filter(estado='PENDIENTE').select_related(
            'vendedor'
        ).prefetch_related(
            'ticket_productos__ProductoTalla__producto',
        )[:10]
        tickets_pendientes_data = []
        for ticket in tickets_pendientes_query:
            # Determinar tipo de ticket
            if ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
                tipo_ticket = 'Diferencia Cambio'
                tipo_ticket_class = 'warning'
            else:
                tipo_ticket = 'Venta'
                tipo_ticket_class = 'primary'

            # productos prefetch → len() sobre la lista cacheada (0 queries)
            productos_count = len(ticket.ticket_productos.all())

            tickets_pendientes_data.append({
                'correlativo': ticket.correlativo,
                'cliente_nombre': ticket.cliente_nombre or 'Sin nombre',
                'cliente_rut': ticket.cliente_rut or '',
                'vendedor_nombre': f"{ticket.vendedor.codigo_vendedor} - {ticket.vendedor.nombre}" if ticket.vendedor else 'Sin vendedor',
                'total': int(ticket.total or 0),
                'hora': ticket.created_at.strftime('%H:%M'),
                'productos_count': productos_count,
                'tipo_ticket': tipo_ticket,
                'tipo_ticket_class': tipo_ticket_class,
                'modulo_origen': ticket.modulo_origen,
                **_check_stock_ticket(ticket, sucursal_id),
            })

        # TICKETS DE CAMBIOS/DEVOLUCIONES PENDIENTES (sin límite de fecha)
        tickets_cambio_pendientes = Ticket.objects.filter(
            sucursal_id=sucursal_id,
            estado='PENDIENTE',
            modulo_origen='CAMBIO_DEVOLUCION'
        ).select_related('vendedor').order_by('-created_at')[:20]
        
        tickets_cambio_data = []
        for ticket in tickets_cambio_pendientes:
            # Determinar tipo de operación
            if 'A DEVOLVER AL CLIENTE' in (ticket.observaciones or ''):
                tipo_op = 'DEVOLUCION'
                icono = '💵'
            elif 'A COBRAR AL CLIENTE' in (ticket.observaciones or ''):
                tipo_op = 'COBRO'
                icono = '💰'
            else:
                tipo_op = 'DIRECTO'
                icono = '🔄'
            
            tickets_cambio_data.append({
                'correlativo': ticket.correlativo,
                'cliente_nombre': ticket.cliente_nombre or 'Sin nombre',
                'total': int(ticket.total or 0),
                'tipo_operacion': tipo_op,
                'icono': icono,
                'metodo_pago': ticket.metodo_pago,
                'fecha': ticket.fecha.strftime('%d/%m/%Y'),
                'hora': ticket.created_at.strftime('%H:%M'),
            })

        sucursal_row = Sucursal.objects.filter(id=sucursal_id).values('alias', 'nombre').first()
        sucursal_nombre = ''
        if sucursal_row:
            sucursal_nombre = sucursal_row.get('alias') or sucursal_row.get('nombre') or ''

        return JsonResponse({
            'success': True,
            'stats': {
                'ventas_hoy': int(ventas_hoy),
                'tickets_pendientes': tickets_pendientes,
                'tickets_pagados': tickets_pagados,
                'promedio_venta': int(promedio_venta),
                'tickets_cambio_pendientes': len(tickets_cambio_data),
                'sucursal_nombre': sucursal_nombre,
            },
            'tickets': tickets_data,
            'tickets_pendientes': tickets_pendientes_data,
            'tickets_cambio': tickets_cambio_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estadísticas: {str(e)}'
        })


@login_required
def validar_rut_cliente(request):
    """API para validar RUT chileno"""
    rut = request.GET.get('rut', '').strip()
    
    if not rut:
        return JsonResponse({
            'success': False,
            'error': 'RUT requerido'
        })
    
    # Validar formato y dígito verificador
    def validar_rut_chileno(rut_completo):
        # Limpiar RUT
        rut_limpio = ''.join(c for c in rut_completo if c.isdigit() or c.lower() == 'k')
        
        if len(rut_limpio) < 2:
            return False
        
        cuerpo = rut_limpio[:-1]
        dv = rut_limpio[-1].lower()
        
        # Calcular dígito verificador
        suma = 0
        multiplicador = 2
        
        for i in range(len(cuerpo) - 1, -1, -1):
            suma += int(cuerpo[i]) * multiplicador
            multiplicador = 7 if multiplicador == 7 else multiplicador + 1
        
        resto = suma % 11
        dv_calculado = '0' if resto == 0 else 'k' if resto == 1 else str(11 - resto)
        
        return dv == dv_calculado
    
    es_valido = validar_rut_chileno(rut)
    
    return JsonResponse({
        'success': True,
        'valido': es_valido,
        'rut_formateado': formatear_rut(rut) if es_valido else rut
    })


def _liberar_cupon_de_venta(ticket, motivo):
    """
    Devuelve al cliente el cupón que consumió una venta que se anuló.

    Existe como helper porque hay CUATRO rutas que dejan una venta sin efecto y
    cada una llegó por su lado: `anular_ticket_pendiente` —la única que llamaba a
    `liberar_cupon`, y que exige estado PENDIENTE, en el cual jamás hay un cupón
    canjeado (el canje ocurre recién al pasar a PAGADO), o sea que la reversa era
    código inalcanzable—, `anular_documento_venta`, `eliminar_documento_venta`
    (la anulación desde cuadratura) y `anular_factura_dte` (la que emite la NC).

    Sin esto el cupón queda CANJEADO sobre una venta que ya no existe y, con una
    campaña de límite UNICO, el cliente no puede recibir otro nunca más.

    Best-effort e idempotente: nunca lanza (corre dentro de flujos de anulación
    que no debe tumbar) y re-anular no libera dos veces. Un cupón que venció
    mientras tanto queda EXPIRADO, no se le regala vigencia extra.
    """
    if ticket is None:
        return 0
    try:
        from .services import cupon_service
        return cupon_service.liberar_cupon(ticket, motivo=motivo)
    except Exception:
        logger.exception(
            "Error al liberar cupón ticket=%s", getattr(ticket, 'correlativo', None))
        return 0


@login_required
@require_POST
@csrf_exempt
def anular_ticket_pendiente(request):
    """Anular un ticket pendiente (solo si no ha sido pagado)"""
    try:
        data = json.loads(request.body)
        correlativo = data.get('correlativo')
        motivo = data.get('motivo', 'Sin motivo especificado')
        eliminar_diferencia = data.get('eliminar_diferencia', False)
        
        if not correlativo:
            return JsonResponse({
                'success': False,
                'error': 'Correlativo de ticket requerido'
            })
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        # Buscar el ticket
        try:
            ticket = Ticket.objects.get(
                correlativo=correlativo,
                sucursal_id=sucursal_id
            )
        except Ticket.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Ticket #{correlativo} no encontrado'
            })
        
        # ===== ELIMINACIÓN DE DIFERENCIA DE CAMBIO (Solo Admin) =====
        if eliminar_diferencia:
            # Verificar que sea administrador
            if getattr(request.user, 'rol', '') not in ['administrador', 'administracion']:
                return JsonResponse({
                    'success': False,
                    'error': '⛔ Solo los administradores pueden eliminar diferencias de cambio'
                })
            
            # Verificar que sea un ticket de cambio/devolución
            if ticket.modulo_origen != 'CAMBIO_DEVOLUCION':
                return JsonResponse({
                    'success': False,
                    'error': 'Este ticket no es de cambio/devolución'
                })
            
            # Verificar que esté pendiente
            if ticket.estado != 'PENDIENTE':
                return JsonResponse({
                    'success': False,
                    'error': f'Solo se pueden eliminar diferencias de tickets pendientes. Estado actual: {ticket.estado}'
                })
            
            with transaction.atomic():
                # Marcar el ticket como completado sin cobro/devolución
                ticket.estado = 'PAGADO'  # Marcamos como "pagado" para que ya no aparezca como pendiente
                ticket.metodo_pago = 'ELIMINADO_ADMIN'
                ticket.observaciones = (ticket.observaciones or '') + f'\n\n🔐 [DIFERENCIA ELIMINADA POR ADMIN] {timezone.now().strftime("%Y-%m-%d %H:%M")}\n' + \
                                       f'Usuario: {request.user.username}\nMotivo: {motivo}'
                ticket.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Diferencia del ticket #{correlativo} eliminada por administrador',
                'ticket_id': ticket.id
            })
        
        # ===== ANULACIÓN NORMAL DE TICKET =====
        # No permitir anulación directa de tickets de cambio/devolución
        if ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
            return JsonResponse({
                'success': False,
                'error': 'Los tickets de cambio/devolución no se pueden anular. Use la opción "Eliminar diferencia" desde la sección de Cambios (solo Admin).'
            })
        
        # Verificar que esté en estado PENDIENTE
        if ticket.estado != 'PENDIENTE':
            return JsonResponse({
                'success': False,
                'error': f'Solo se pueden anular tickets pendientes. Estado actual: {ticket.estado}'
            })
        
        with transaction.atomic():
            # Si es ticket de cambio/devolución, el stock ya fue ajustado
            # Solo revertir stock si es un ticket normal
            if ticket.modulo_origen != 'CAMBIO_DEVOLUCION':
                productos_ticket = Ticket_Productos.objects.filter(idTicket=ticket)
                
                for item in productos_ticket:
                    # Crear movimiento de devolución de stock
                    # ✅ Usar DTE si está disponible, si no usar correlativo del ticket
                    referencia = f'ANULACION_DTE_{ticket.folio_dte}' if ticket.folio_dte else f'ANULACION_TICKET_{ticket.correlativo}'
                    Movimientos_Producto.objects.create(
                        ProductoTalla=item.ProductoTalla,
                        cantidad=item.stock,  # Positivo para devolver al inventario
                        costo=item.ProductoTalla.producto.costo,
                        precio=int(item.precio),
                        concepto='ANULACION_TICKET',
                        tipo_movimiento='INGRESO',
                        responsable=request.user.username,
                        observaciones=f'Anulación de ticket #{ticket.correlativo} - Motivo: {motivo}',
                        referencia_externa=referencia
                    )
            
            # Cambiar estado del ticket a ANULADO
            ticket.estado = 'ANULADO'
            ticket.observaciones = (ticket.observaciones or '') + f'\n[ANULADO] {timezone.now().strftime("%Y-%m-%d %H:%M")} - {motivo}'
            ticket.save()

            # ===== FIDELIZACIÓN / GIFT CARD: reversa de la venta anulada =====
            # Devuelve los puntos acumulados por esta venta y recarga las gift
            # cards consumidas. Idempotente: re-anular no duplica la reversa.
            try:
                from .services import fidelizacion_service
                fidelizacion_service.reversar_venta(ticket, usuario=request.user)
            except Exception:
                logger.exception("Error al reversar puntos ticket=%s", ticket.correlativo)
            try:
                from .services import giftcard_service
                for pago_gc in ticket.pagos.filter(metodo_pago='GIFTCARD'):
                    codigo_gc = (pago_gc.voucher or '').strip()
                    if codigo_gc:
                        giftcard_service.reversar(
                            codigo_gc, pago_gc.monto,
                            ticket=ticket, usuario=request.user,
                        )
            except Exception:
                logger.exception("Error al reversar gift cards ticket=%s", ticket.correlativo)
            # El cupón vuelve a estar disponible para el cliente: si la venta se
            # anuló, el beneficio no se consumió.
            _liberar_cupon_de_venta(ticket, f'anulación: {motivo}')

        return JsonResponse({
            'success': True,
            'message': f'Ticket #{correlativo} anulado exitosamente',
            'ticket_id': ticket.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Formato JSON inválido'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al anular ticket: {str(e)}'
    })


@login_required
def buscar_cliente_rut(request):
    """API para buscar cliente por RUT.

    Busca en este orden y prioriza el match en `Empresa` para devolver
    `cliente_id` (porque el receptor de un DTE/NC es FK a `Empresa`,
    no a `Cliente` ni a `Ticket`):
      1. Tabla `Empresa` (clientes creados desde POS / gestion-DTE → NCN,
         o sincronizados desde Acepta — son receptor válido de DTE).
      2. Tabla `Cliente` (CRM clásico — solo enriquece datos, NO sirve
         como id de receptor de DTE).
      3. Tickets anteriores (histórico de ventas, igual: solo enriquece).

    Convención de respuesta:
      - `cliente_id`: SOLO se devuelve cuando el match es en `Empresa`.
        Es el id que se debe pasar a `asignar_receptor_dte` y a
        `anular_factura_dte` como `cliente_id`.
      - `cliente_origen`: 'EMPRESA' | 'CRM' | 'TICKET' — útil para que
        el frontend decida si mandar `cliente_id` o sólo RUT + datos.
      - Cuando es CRM/TICKET, `cliente_id` queda fuera de la respuesta
        para evitar la confusión de IDs cruzados que provocaba el error
        "El cliente seleccionado no existe" al intentar usar un Cliente.id
        (CRM) como Empresa.id.
    """
    from app.models import Cliente, Empresa

    rut = request.GET.get('rut', '').strip()

    if not rut:
        return JsonResponse({
            'success': False,
            'error': 'RUT requerido'
        })

    try:
        rut_limpio = rut.replace('.', '').replace('-', '').strip()
        rut_formateado = formatear_rut(rut_limpio)

        # Detectar RUT de persona jurídica (no fideliza)
        from .services.fidelizacion_service import es_rut_empresa as _es_empresa
        _cliente_es_empresa = _es_empresa(rut_formateado)

        # Tasas dinámicas desde el programa activo (fallback a defaults)
        try:
            from app.models import ProgramaFidelizacion as _Prog
            _prog_activo = _Prog.get_activo()
            _tasas_fid = {
                'PLATA': float(_prog_activo.tasa_plata),
                'ORO':   float(_prog_activo.tasa_oro),
                'PLATINO': float(_prog_activo.tasa_platino),
            }
        except Exception:
            _tasas_fid = {'PLATA': 3.0, 'ORO': 4.0, 'PLATINO': 5.0}

        # 1) Tabla Empresa — match aquí da cliente_id usable como receptor.
        empresa_cliente = Empresa.objects.filter(
            Q(rut__iexact=rut_formateado) |
            Q(rut__iexact=rut) |
            Q(rut__icontains=rut_limpio)
        ).filter(esProveedor=False).order_by('-id').first()

        if empresa_cliente:
            # contacto1 acepta emails o teléfonos — no usarlo como teléfono si tiene '@'.
            telefono_empresa = ''
            if empresa_cliente.contacto1 and '@' not in empresa_cliente.contacto1:
                telefono_empresa = empresa_cliente.contacto1
            telefono_empresa = telefono_empresa or empresa_cliente.telefono or ''

            cliente_data = {
                'nombre': empresa_cliente.nombre or empresa_cliente.razon_social or '',
                'rut': empresa_cliente.rut,
                'email': empresa_cliente.correoVendedor or empresa_cliente.email or '',
                'telefono': telefono_empresa,
                'giro': empresa_cliente.giro or '',
                'comuna': empresa_cliente.comuna or '',
                'ciudad': empresa_cliente.ciudad or '',
                'direccion': empresa_cliente.direccion or '',
                'telefono_secundario': empresa_cliente.contacto2 or '',
                'celular': '',
                'fecha_nacimiento': '',
                'email_facturacion': empresa_cliente.correoAdministrador or '',
            }

            # Enriquecer celular + fecha_nacimiento desde CRM si el mismo RUT tiene registro ahí.
            crm_cliente = Cliente.objects.filter(
                Q(rut__iexact=rut_formateado) | Q(rut__icontains=rut_limpio),
                activo=True,
            ).first()

            fidelizacion_data = None
            if crm_cliente and not _cliente_es_empresa:
                if crm_cliente.celular:
                    cliente_data['celular'] = crm_cliente.celular
                if crm_cliente.fecha_nacimiento:
                    cliente_data['fecha_nacimiento'] = crm_cliente.fecha_nacimiento.isoformat()
                try:
                    from .services import fidelizacion_service
                    saldo_info = fidelizacion_service.consultar_saldo(cliente=crm_cliente)
                    if saldo_info.get('nivel'):
                        fidelizacion_data = {
                            'saldo_puntos': saldo_info.get('saldo_puntos', 0),
                            'valor_pesos': saldo_info.get('valor_pesos', 0),
                            'nivel': saldo_info.get('nivel', 'PLATA'),
                            'tasa': _tasas_fid.get(saldo_info.get('nivel', 'PLATA'), 3.0),
                        }
                except Exception:
                    pass

            return JsonResponse({
                'success': True,
                'cliente': cliente_data,
                'mensaje': 'Cliente encontrado en empresas',
                'cliente_id': empresa_cliente.id,
                'cliente_origen': 'EMPRESA',
                'fidelizacion': fidelizacion_data,
                'es_empresa': _cliente_es_empresa,
            })

        # 2) Tabla Cliente (CRM): solo enriquece datos para auto-llenar el form.
        # NO se devuelve `cliente_id` porque ese id es de Cliente, no de Empresa,
        # y al usarlo como receptor explotaba con "El cliente seleccionado no existe".
        cliente = Cliente.objects.filter(
            Q(rut__iexact=rut_formateado) |
            Q(rut__icontains=rut_limpio)
        ).filter(activo=True).first()

        if cliente:
            cliente_data = {
                'nombre': cliente.nombre_completo,
                'rut': cliente.rut,
                'email': cliente.email or '',
                'telefono': cliente.telefono or cliente.celular or '',
                'giro': cliente.empresa.giro if cliente.empresa else '',
                'comuna': cliente.comuna or '',
                'ciudad': cliente.ciudad or '',
                'direccion': cliente.direccion or '',
                'telefono_secundario': cliente.celular if cliente.telefono else '',
                # Campos de fidelización para autocompletar en POS / ticket-venta
                'celular': cliente.celular or '',
                'fecha_nacimiento': cliente.fecha_nacimiento.isoformat() if cliente.fecha_nacimiento else '',
                'email_facturacion': cliente.email or '',
            }
            fidelizacion_data = None
            if not _cliente_es_empresa:
                try:
                    from .services import fidelizacion_service
                    saldo_info = fidelizacion_service.consultar_saldo(cliente=cliente)
                    if saldo_info.get('nivel'):
                        fidelizacion_data = {
                            'saldo_puntos': saldo_info.get('saldo_puntos', 0),
                            'valor_pesos': saldo_info.get('valor_pesos', 0),
                            'nivel': saldo_info.get('nivel', 'PLATA'),
                            'tasa': _tasas_fid.get(saldo_info.get('nivel', 'PLATA'), 3.0),
                        }
                except Exception:
                    pass
            return JsonResponse({
                'success': True,
                'cliente': cliente_data,
                'mensaje': 'Cliente encontrado en CRM (se creará Empresa al asignar)',
                'cliente_origen': 'CRM',
                'fidelizacion': fidelizacion_data,
                'es_empresa': _cliente_es_empresa,
                # No incluimos cliente_id: el backend resolverá por RUT y, si
                # es necesario, creará la Empresa con los datos del cliente.
            })

        # 3) Histórico de tickets: igual que CRM, solo enriquece datos.
        ticket_con_cliente = Ticket.objects.filter(
            Q(cliente_rut__iexact=rut_formateado) |
            Q(cliente_rut__icontains=rut_limpio)
        ).exclude(
            cliente_nombre__isnull=True
        ).exclude(
            cliente_nombre__exact=''
        ).order_by('-created_at').first()

        if ticket_con_cliente and ticket_con_cliente.cliente_nombre:
            cliente_data = {
                'nombre': ticket_con_cliente.cliente_nombre,
                'rut': ticket_con_cliente.cliente_rut,
                'email': ticket_con_cliente.cliente_email or '',
                'telefono': ticket_con_cliente.cliente_telefono or '',
                'giro': ticket_con_cliente.cliente_giro or '',
                'comuna': ticket_con_cliente.cliente_comuna or '',
                'ciudad': ticket_con_cliente.cliente_ciudad or '',
                'direccion': ticket_con_cliente.cliente_direccion or '',
                'telefono_secundario': ticket_con_cliente.cliente_telefono_secundario or '',
                # Celular: el ticket lo guarda en teléfono secundario
                'celular': ticket_con_cliente.cliente_telefono_secundario or '',
                'fecha_nacimiento': '',
                'email_facturacion': ticket_con_cliente.cliente_email_facturacion or '',
            }
            # Enriquecer fecha_nacimiento y celular desde CRM si el cliente ya existe ahí.
            # El ticket no almacena fecha_nacimiento, así que siempre hay que leerla del CRM.
            crm_tk = Cliente.objects.filter(
                Q(rut__iexact=rut_formateado) | Q(rut__icontains=rut_limpio)
            ).first()
            if crm_tk:
                if crm_tk.fecha_nacimiento:
                    cliente_data['fecha_nacimiento'] = crm_tk.fecha_nacimiento.isoformat()
                if crm_tk.celular and not cliente_data['celular']:
                    cliente_data['celular'] = crm_tk.celular
            # Intentar cargar fidelización (igual que paths EMPRESA/CRM)
            _fid_ticket = None
            if not _cliente_es_empresa:
                try:
                    from .services import fidelizacion_service as _fid_svc
                    _saldo_tk = _fid_svc.consultar_saldo(rut=rut_formateado)
                    if _saldo_tk and _saldo_tk.get('saldo_puntos', 0) > 0:
                        _nivel_tk = _saldo_tk.get('nivel', 'PLATA')
                        _fid_ticket = {
                            'saldo_puntos': _saldo_tk.get('saldo_puntos', 0),
                            'valor_pesos':  _saldo_tk.get('valor_pesos', 0),
                            'nivel': _nivel_tk,
                            'tasa': _tasas_fid.get(_nivel_tk, 3.0),
                        }
                except Exception:
                    pass

            return JsonResponse({
                'success': True,
                'cliente': cliente_data,
                'mensaje': 'Cliente encontrado en tickets anteriores',
                'cliente_origen': 'TICKET',
                'fidelizacion': _fid_ticket,
                'es_empresa': _cliente_es_empresa,
            })

        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado',
            'rut_formateado': rut_formateado,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar cliente: {str(e)}'
        })


def formatear_rut(rut):
    """Formatear RUT chileno SIN puntos, solo con guión"""
    # Limpiar RUT
    rut_limpio = ''.join(c for c in rut if c.isdigit() or c.lower() == 'k')
    
    if len(rut_limpio) < 2:
        return rut
    
    cuerpo = rut_limpio[:-1]
    dv = rut_limpio[-1]
    
    # Formatear SIN puntos (solo guión)
    return f"{cuerpo}-{dv}"


def guardar_o_actualizar_cliente(datos_cliente, usuario=None):
    """Guardar o actualizar cliente en la base de datos"""
    from app.models import Cliente

    try:
        if not datos_cliente.get('rut') or not datos_cliente.get('nombre'):
            return None  # No guardar si no hay datos mínimos
        
        rut = formatear_rut(datos_cliente['rut'])
        nombre_completo = datos_cliente['nombre']
        
        # Separar nombre y apellido
        partes_nombre = nombre_completo.split(' ', 1)
        nombre = partes_nombre[0] if len(partes_nombre) > 0 else nombre_completo
        apellido = partes_nombre[1] if len(partes_nombre) > 1 else ''
        
        # Buscar cliente existente
        cliente = Cliente.objects.filter(rut=rut).first()
        
        if cliente:
            # Política de actualización (cliente ya existente):
            #  - Si llega un valor NUEVO y no vacío -> se actualiza (permite corregir).
            #  - Si el campo llega vacío -> NO se pisa el dato existente.
            # Así, aunque el operador no confirme los datos precargados, nunca se
            # pierde un dato bueno; solo se sobrescribe cuando hay un valor nuevo.
            def _set_si_viene(attr, valor):
                v = (valor or '').strip() if isinstance(valor, str) else valor
                if v:
                    setattr(cliente, attr, v)

            _set_si_viene('email', datos_cliente.get('email'))
            _set_si_viene('telefono', datos_cliente.get('telefono'))
            _set_si_viene('direccion', datos_cliente.get('direccion'))
            _set_si_viene('comuna', datos_cliente.get('comuna'))
            _set_si_viene('ciudad', datos_cliente.get('ciudad'))
            # Celular: prioriza el campo propio; fallback al teléfono secundario.
            _set_si_viene('celular', datos_cliente.get('celular') or datos_cliente.get('telefono_secundario'))
            # Fecha de nacimiento (para fidelización: campañas de cumpleaños).
            _set_si_viene('fecha_nacimiento', datos_cliente.get('fecha_nacimiento'))

            # Nombre/apellido: actualizar solo si vienen (no borrar con vacío).
            if nombre and apellido:
                cliente.nombre = nombre
                cliente.apellido = apellido

            if usuario:
                cliente.updated_by = usuario
            cliente.save()
            
        else:
            # Determinar tipo de cliente basado en los datos
            tipo_cliente = 'INDIVIDUAL'
            if datos_cliente.get('giro') and datos_cliente.get('direccion'):
                tipo_cliente = 'EMPRESARIAL'
            
            # Crear nuevo cliente
            cliente = Cliente.objects.create(
                nombre=nombre,
                apellido=apellido,
                rut=rut,
                email=datos_cliente.get('email', ''),
                telefono=datos_cliente.get('telefono', ''),
                celular=datos_cliente.get('celular') or datos_cliente.get('telefono_secundario', ''),
                fecha_nacimiento=datos_cliente.get('fecha_nacimiento') or None,
                direccion=datos_cliente.get('direccion', ''),
                comuna=datos_cliente.get('comuna', ''),
                ciudad=datos_cliente.get('ciudad', ''),
                tipo_cliente=tipo_cliente,
                activo=True,
                created_by=usuario,
                observaciones=f'Creado automáticamente desde venta'
            )
        
        return cliente
        
    except Exception:
        # Si hay error, no detener el proceso de venta
        logger.exception("Error al guardar cliente desde venta")
        return None


# ========== FUNCIONES TICKET POS ==========

def construir_ticket_data(ticket):
    """Construir datos completos del ticket para POS"""
    from app.services.realsport_imagenes_service import resolver_foto_portada_url
    productos_procesados = []
    total_items = 0
    subtotal = 0

    empresa_id_ticket = ticket.sucursal.empresa_id if ticket.sucursal_id else None

    for tp in ticket.ticket_productos.select_related(
        'ProductoTalla',
        'ProductoTalla__producto',
        'ProductoTalla__producto__atributo1',
        'ProductoTalla__producto__atributo2',
        'ProductoTalla__producto__atributo3',
        'ProductoTalla__producto__atributo4',
        'promo_campana',
    ).all():
        producto_talla = tp.ProductoTalla
        producto = producto_talla.producto if producto_talla else None

        marca = ''
        if producto:
            atributo_marca = getattr(producto, 'atributo1', None)
            if atributo_marca:
                marca = getattr(atributo_marca, 'valor', '') or ''

        # Foto de portada (cacheada por el service — cero queries si ya estaba).
        foto_url = ''
        if producto and producto.articulo:
            foto_url = resolver_foto_portada_url(producto.articulo, empresa_id_ticket)

        subtotal += tp.subtotal
        total_items += tp.stock
        productos_procesados.append({
            'detalle_id': tp.id,
            'producto_talla_id': producto_talla.id if producto_talla else None,
            'producto_id': producto.id if producto else None,
            'sku': producto_talla.sku if producto_talla else '',
            'articulo': producto.articulo if producto else (tp.descripcion_linea or ''),
            'descripcion': producto.descripcion if producto else (tp.descripcion_linea or ''),
            'marca': marca,
            'talla': producto_talla.talla if producto_talla else '',
            'cantidad': tp.stock,
            'precio_unitario': tp.precio,
            'precio': tp.precio,  # Alias para compatibilidad con frontend
            'precio_original': tp.precio_original,
            'descuento_unitario': tp.descuento_unitario,
            'porcentaje_descuento': float(tp.porcentaje_descuento or 0),
            'subtotal': tp.subtotal,
            'costo_fifo': tp.costo_fifo,
            'lotes_utilizados': tp.lotes_utilizados,
            'stock_actual': producto_talla.stock if producto_talla else None,
            'stock': producto_talla.stock_sucursal(ticket.sucursal_id) if producto_talla else 0,  # Stock real de la sucursal del ticket
            'foto_portada_url': foto_url,
            # Flags de oferta NxM: la caja los usa para reabsorber la línea
            # gratis (idempotencia) en vez de duplicarla al cargar el ticket.
            'es_promo_nxm': tp.promo_campana_id is not None,
            'promo_campana_id': tp.promo_campana_id,
            'promo_label': (
                f"{tp.promo_campana.nxm_n}x{tp.promo_campana.nxm_m} · {tp.promo_campana.nombre}"
                if tp.promo_campana_id else None),
        })

    sucursal = ticket.sucursal
    empresa = sucursal.empresa if hasattr(sucursal, 'empresa') else None

    pagos_queryset = ticket.pagos.all().order_by('creado_en')
    pagos = [
        {
            'id': pago.id,
            'metodo_pago': pago.metodo_pago,
            'metodo_pago_display': pago.get_metodo_pago_display(),
            'monto': pago.monto,
            'voucher': pago.voucher or '',
            'tipo_tarjeta': pago.tipo_tarjeta or '',
            'numero_orden_compra': pago.numero_orden_compra or '',
            'notas': pago.notas or '',
            'creado_en': pago.creado_en.strftime('%Y-%m-%d %H:%M:%S'),
        }
        for pago in pagos_queryset
    ]

    total_pagado = sum(pago['monto'] for pago in pagos)
    saldo_por_pagar = (ticket.total or 0) - total_pagado
    if saldo_por_pagar < 0:
        saldo_por_pagar = 0

    return {
        'ticket_id': ticket.correlativo,
        'fecha': ticket.fecha.strftime('%Y-%m-%d'),
        'hora': ticket.hora.strftime('%H:%M:%S'),
        'tipo_documento': 'TICKET',
        'estado': ticket.estado,
        'modulo_origen': ticket.modulo_origen,  # ✅ Agregar módulo de origen para identificar tickets de cambio
        'metodo_pago_principal': ticket.metodo_pago,
        'total_pagado': total_pagado,
        'saldo_por_pagar': saldo_por_pagar,
        'responsable': ticket.responsable,
        'sucursal': {
            'alias': sucursal.alias,
            'nombre': getattr(sucursal, 'nombreSucursal', None) or sucursal.alias or '',
            'direccion': sucursal.direccion,
            'empresa': empresa.nombre if empresa else '',
            'rut_empresa': empresa.rut if empresa else ''
        },
        'vendedor': {
            'nombre': ticket.vendedor.nombre if ticket.vendedor else '',
            'codigo': ticket.vendedor.codigo_vendedor if ticket.vendedor else ''
        },
        'cliente': {
            'nombre': ticket.cliente_nombre or '',
            'rut': ticket.cliente_rut or '',
            'giro': ticket.cliente_giro or '',
            'comuna': ticket.cliente_comuna or '',
            'ciudad': ticket.cliente_ciudad or '',
            'direccion': ticket.cliente_direccion or '',
            'telefono': ticket.cliente_telefono or '',
            'telefono_secundario': ticket.cliente_telefono_secundario or '',
            # Celular: el ticket lo guarda en teléfono secundario; se expone como
            # 'celular' para precargar el campo nuevo del POS (fidelización).
            'celular': ticket.cliente_telefono_secundario or (ticket.cliente.celular if ticket.cliente_id and ticket.cliente else ''),
            # Fecha de nacimiento: vive en el Cliente CRM enlazado.
            'fecha_nacimiento': (ticket.cliente.fecha_nacimiento.isoformat()
                                 if ticket.cliente_id and ticket.cliente and ticket.cliente.fecha_nacimiento else ''),
            'email': ticket.cliente_email or '',
            'email_facturacion': ticket.cliente_email_facturacion or '',
        },
        'observaciones': ticket.observaciones or '',
        'observaciones_adicionales': ticket.observaciones_adicionales or '',
        'productos': productos_procesados,
        'pagos': pagos,
        'totales': {
            'items': total_items,
            'subtotal': subtotal,
            'descuento': ticket.descuento or 0,
            # Descuentos de cabecera: no están en ninguna línea, se aplican al
            # total al cobrar. Sin exponerlos, el comprobante térmico imprime un
            # TOTAL más bajo que la suma de sus propias líneas y sin ninguna
            # glosa que lo explique — el cliente reclama y el cajero no tiene qué
            # mostrarle.
            'descuento_cupon': ticket.descuento_cupon or 0,
            'descuento_fidelizacion': ticket.descuento_fidelizacion or 0,
            'total': ticket.total
        }
    }


def _obtener_ticket_para_pos(request, correlativo):
    """Función auxiliar para obtener ticket para POS"""
    sucursal_id = (
        request.session.get('idSucursalActual')
        or request.session.get('sucursalActual')
        or request.session.get('idSucursalActualPOS')
    )
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión'}, status=400)

    # Primero buscar en la sucursal activa
    ticket = (
        Ticket.objects
        .select_related('sucursal', 'vendedor', 'cliente')
        .prefetch_related('ticket_productos__ProductoTalla__producto', 'pagos')
        .filter(sucursal_id=sucursal_id, correlativo=correlativo)
        .first()
    )

    # Si no se encuentra, buscar en todas las sucursales del usuario
    # (para casos de cotizaciones facturadas desde otra sucursal)
    if not ticket:
        # Buscar en cualquier sucursal que el usuario tenga acceso
        ticket = (
            Ticket.objects
            .select_related('sucursal', 'vendedor', 'cliente')
            .prefetch_related('ticket_productos__ProductoTalla__producto', 'pagos')
            .filter(correlativo=correlativo)
            .order_by('-fecha', '-hora')  # El más reciente primero
            .first()
        )

    # Si sigue sin encontrarse, intentar buscar por folio_dte
    # (cuando se llama desde gestionVentasDocumentos pasando el número de DTE en lugar del correlativo de ticket)
    if not ticket:
        ticket = (
            Ticket.objects
            .select_related('sucursal', 'vendedor', 'cliente')
            .prefetch_related('ticket_productos__ProductoTalla__producto', 'pagos')
            .filter(folio_dte=correlativo)
            .order_by('-fecha', '-hora')
            .first()
        )

    if not ticket:
        return JsonResponse({'success': False, 'error': f'Ticket {correlativo} no encontrado'}, status=404)

    return JsonResponse({'success': True, 'ticket': construir_ticket_data(ticket)})


@login_required
@require_GET
def obtener_ticket_por_correlativo(request, correlativo):
    """Obtener ticket por correlativo para POS"""
    return _obtener_ticket_para_pos(request, correlativo)


@login_required
@require_POST
def buscar_ticket_pos(request):
    """Buscar ticket en POS"""
    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Formato JSON inválido'}, status=400)

    correlativo = payload.get('correlativo') or payload.get('ticket')
    if not correlativo:
        return JsonResponse({'success': False, 'error': 'Debe indicar el número de ticket'}, status=400)

    try:
        correlativo_int = int(correlativo)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Número de ticket inválido'}, status=400)

    return _obtener_ticket_para_pos(request, correlativo_int)


@login_required
@require_POST
def crear_ticket_pendiente_pos(request):
    """Crear un ticket PENDIENTE vacío desde el POS Dashboard para iniciar una nueva venta."""
    try:
        sucursal_id = (
            request.session.get('idSucursalActual')
            or request.session.get('sucursalActual')
        )
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión'}, status=400)

        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        vendedor = Vendedor.objects.filter(sucursales=sucursal, activo=True).first()
        if not vendedor:
            vendedor = Vendedor.objects.filter(
                empresa=sucursal.empresa, activo=True
            ).first() if hasattr(sucursal, 'empresa') and sucursal.empresa else None
        if not vendedor:
            return JsonResponse({
                'success': False,
                'error': 'No hay vendedores activos configurados para esta sucursal'
            }, status=400)

        with transaction.atomic():
            correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
            ticket = Ticket.objects.create(
                correlativo=correlativo,
                sucursal=sucursal,
                vendedor=vendedor,
                subTotal=0,
                descuento=0,
                total=0,
                estado='PENDIENTE',
                responsable=request.user.username,
                modulo_origen='POS',
            )

        return JsonResponse({
            'success': True,
            'ticket': construir_ticket_data(ticket),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al crear ticket: {str(e)}'}, status=500)


def cuadrar_detalle_neto(detalle, monto_neto, descuentos_recargos=None):
    """
    Ajusta el detalle de una FACTURA para que `sum(MontoItem) - descuentos`
    coincida exactamente con el `MntNeto` del header.

    Por qué hace falta: cada línea calcula su neto con `round(precio / 1.19)`
    mientras el header redondea el total completo una sola vez. Con varias
    líneas los redondeos no conmutan y queda una diferencia de $1..$N. Acepta
    rechaza el DTE cuando el detalle no cuadra con el header.

    El residuo se absorbe en la ÚLTIMA línea: el total cobrado no cambia, solo
    se reparte el peso del redondeo.

    Devuelve el residuo aplicado (0 si ya cuadraba). Muta `detalle` in-place.
    """
    if not detalle:
        return 0

    descuento_neto = sum(
        int(dr.get('valor_dr') or 0)
        for dr in (descuentos_recargos or [])
        if dr.get('tpo_mov') == 'D'
    )
    recargo_neto = sum(
        int(dr.get('valor_dr') or 0)
        for dr in (descuentos_recargos or [])
        if dr.get('tpo_mov') == 'R'
    )

    suma_items = sum(int(d['monto_item']) for d in detalle)
    residuo = int(monto_neto) - (suma_items - descuento_neto + recargo_neto)

    if residuo:
        detalle[-1]['monto_item'] = int(detalle[-1]['monto_item']) + residuo

    return residuo


def generar_dte_desde_ticket(ticket, tipo_documento, usuario, cotizacion=None):
    """
    Generar DTE (Boleta o Factura Electrónica) desde un Ticket
    Genera tanto el registro en BD como el archivo TXT para Acepta
    
    Args:
        cotizacion: Objeto Cotizacion_Empresa opcional para usar sus descripciones en el TXT
    """
    from decimal import Decimal
    
    # Mapear tipo de documento
    tipo_dte_map = {
        'BOLETA_ELECTRONICA': 'BOLETA ELECTRONICA',
        'BOLETA_PAPEL': 'BOLETA PAPEL',
        'FACTURA_ELECTRONICA': 'FACTURA ELECTRONICA',
    }
    
    tipo_dte = tipo_dte_map.get(tipo_documento, 'BOLETA ELECTRONICA')
    es_boleta = 'BOLETA' in tipo_dte
    
    # Obtener o crear receptor (cliente)
    receptor = None
    if ticket.cliente_rut and ticket.cliente_nombre:
        # Buscar si ya existe el cliente como Empresa.
        # OJO: `Empresa.rut` NO es único y hay ~35 RUTs con ficha duplicada;
        # sin `order_by` el `.first()` era NO determinístico y la factura podía
        # tomar la comuna/ciudad de cualquiera de las fichas. Además algunas
        # fichas legacy guardan el RUT CON puntos (12.345.678-9) mientras que
        # `ticket.cliente_rut` viene normalizado sin puntos, lo que provocaba
        # que no se encontrara y se creara una ficha duplicada nueva.
        rut_norm = formatear_rut(ticket.cliente_rut)  # sin puntos, con guión
        candidatos_rut = {ticket.cliente_rut, rut_norm}
        if '-' in rut_norm:
            _cuerpo, _dv = rut_norm.split('-', 1)
            if _cuerpo.isdigit():
                _cuerpo_puntos = f"{int(_cuerpo):,}".replace(',', '.')
                candidatos_rut.add(f"{_cuerpo_puntos}-{_dv}")
        receptor = (
            Empresa.objects
            .filter(rut__in=candidatos_rut)
            .order_by('id')
            .first()
        )

        if not receptor:
            # Crear empresa/cliente
            receptor = Empresa.objects.create(
                nombre=ticket.cliente_nombre,
                rut=ticket.cliente_rut,
                razon_social=ticket.cliente_nombre,
                nombre_fantasia=ticket.cliente_nombre,
                giro=ticket.cliente_giro or 'Consumidor Final',
                direccion=ticket.cliente_direccion or 'Sin dirección',
                comuna=ticket.cliente_comuna or 'Sin comuna',
                ciudad=ticket.cliente_ciudad or 'Sin ciudad',
                telefono=ticket.cliente_telefono or '',
                correoVendedor=ticket.cliente_email or '',
                correoAdministrador=ticket.cliente_email_facturacion or ticket.cliente_email or '',
                esProveedor=False,
            )
    
    # Obtener siguiente correlativo para el DTE
    correlativo_dte = obtener_siguiente_correlativo(ticket.sucursal, tipo_dte)

    # Calcular montos AUTORITATIVOS a partir de las líneas reales del ticket.
    # NO confiamos en ticket.total / ticket.descuento porque pueden quedar
    # stale si algún paso del flujo no los recalculó (ver bug histórico donde
    # al agregar un producto en el paso 3 el TXT del DTE mostraba una línea
    # "Descuento: $X" fantasma por desfase entre ticket.total y la suma real
    # de los items).
    descuento_real_lineas = sum(
        (tp.descuento_unitario or 0) * tp.stock
        for tp in ticket.ticket_productos.all()
    )
    suma_items_brutos = sum(
        (tp.precio or 0) * tp.stock
        for tp in ticket.ticket_productos.all()
    )
    total_real_lineas = suma_items_brutos - descuento_real_lineas

    # Descuentos de CABECERA: no viven en ninguna línea del ticket, se aplican
    # sobre el total al cobrar (ver registrar_pagos_ticket). Son excluyentes
    # entre sí por regla de negocio, pero se suman igual para que agregar un
    # tercero no vuelva a romper esta función.
    #   - descuento_fidelizacion → vale de puntos
    #   - descuento_cupon        → cupón nominativo
    # Omitir uno acá tiene tres consecuencias encadenadas, todas de plata: falsa
    # alarma de desfase, `ticket.total` reescrito al bruto en la BD, y el DTE
    # emitido por más de lo que el cliente pagó.
    descuento_vale = int(ticket.descuento_fidelizacion or 0)
    descuento_cupon = int(getattr(ticket, 'descuento_cupon', None) or 0)
    descuentos_cabecera = descuento_vale + descuento_cupon

    # Si ticket.total está sincronizado con las líneas, lo respetamos. Si no,
    # usamos el cálculo autoritativo basado en las líneas (y logeamos el desfase
    # para diagnosticar flujos mal cerrados). Esta comparación es en base a las
    # líneas de producto (bruto - descuento de línea), sumando de vuelta los
    # descuentos de cabecera — ticket.total ya los tiene restados, así que sin
    # eso habría una falsa alarma de desfase en cada venta con vale o cupón.
    ticket_total_guardado = int(ticket.total or 0) + descuentos_cabecera
    if ticket_total_guardado != total_real_lineas and total_real_lineas > 0:
        logger.warning(
            "Desfase ticket.total vs suma de lineas. ticket_id=%s, total_guardado=%s, "
            "total_lineas=%s. Usando suma de lineas como autoritativo.",
            ticket.id,
            ticket_total_guardado,
            total_real_lineas,
        )
        # Reconciliar también en DB para que el resto del flujo (resumen, cuadratura) use el valor correcto.
        ticket.total = total_real_lineas - descuentos_cabecera
        ticket.descuento = descuento_real_lineas
        if hasattr(ticket, 'subTotal'):
            ticket.subTotal = suma_items_brutos
        ticket.save(update_fields=['total', 'descuento', 'subTotal'] if hasattr(ticket, 'subTotal') else ['total', 'descuento'])

    # El DTE debe reflejar lo efectivamente cobrado: la parte pagada con vale de
    # puntos o rebajada por un cupón nunca llegó como pago en efectivo/tarjeta
    # (Dte_Detalle_Pago), así que se resta acá igual que el descuento por línea.
    # Si no, el DTE queda emitido por un monto mayor a la suma real de sus
    # propios Dte_Detalle_Pago — y se declara IVA de plata que no entró.
    total_dte = total_real_lineas - descuentos_cabecera

    total_con_iva = Decimal(total_dte)
    descuento = Decimal(descuento_real_lineas + descuentos_cabecera)

    # Descomponer el total para obtener neto e IVA
    # Total = Neto + IVA, donde IVA = Neto * 0.19
    # Total = Neto * 1.19
    # Neto = Total / 1.19
    neto = (total_con_iva / Decimal('1.19')).quantize(Decimal('0'))
    iva = total_con_iva - neto
    total = total_con_iva
    
    # Crear DTE con todos los campos requeridos
    from datetime import timedelta
    import json as _json

    # 1=Contado (pago inmediato), 2=Crédito (pago diferido)
    METODOS_CREDITO_DTE = {'CREDITO_TRABAJADOR', 'CREDITO_EXTERNO', 'CONVENIO', 'ORDEN_COMPRA'}
    forma_pago_dte = None
    try:
        notas_ticket = _json.loads(ticket.observaciones_adicionales or '{}')
        if isinstance(notas_ticket, dict) and notas_ticket.get('condicion_pago_dte') in (1, 2):
            forma_pago_dte = int(notas_ticket['condicion_pago_dte'])
    except (TypeError, ValueError):
        forma_pago_dte = None
    if forma_pago_dte is None:
        metodos_ticket = set(ticket.pagos.values_list('metodo_pago', flat=True))
        forma_pago_dte = 2 if metodos_ticket & METODOS_CREDITO_DTE else 1

    dias_credito_dte = 30 if (not es_boleta and forma_pago_dte == 2) else 0
    fecha_vencimiento_dte = ticket.fecha + timedelta(days=dias_credito_dte) if dias_credito_dte else ticket.fecha
    
    dte = Dte.objects.create(
        numero_documento=int(correlativo_dte),
        tipo_documento=tipo_dte,
        tipo_transaccion='VENTA_PUBLICO',
        fecha_emision=ticket.fecha,
        fecha_vencimiento=fecha_vencimiento_dte,
        diasCredito=dias_credito_dte,
        bultos=1,
        unidades_productos=sum(tp.stock for tp in ticket.ticket_productos.all()),
        emisor=ticket.sucursal.empresa,
        receptor=receptor,
        sucursal=ticket.sucursal,
        vendedor=ticket.vendedor,
        monto_neto=neto,
        monto_con_iva=total,
        descuento=descuento,
        estado_pago='PAGADO',
        estado_dte='EMITIDO',
        responsable=usuario.username if usuario else ticket.responsable,
        hora=ticket.hora,
        referencias=f'TICKET-{ticket.correlativo}'
    )
    
    # Copiar productos del ticket al DTE
    for tp in ticket.ticket_productos.all():
        if tp.ProductoTalla:
            costo_unitario = tp.ProductoTalla.producto.costo if tp.ProductoTalla.producto else 0
            sobreprecio_unitario = tp.ProductoTalla.producto.sobreprecio if tp.ProductoTalla.producto else 0
            descripcion_prod = (tp.ProductoTalla.producto.descripcion or tp.ProductoTalla.producto.articulo) if tp.ProductoTalla.producto else (tp.descripcion_linea or '')
        else:
            costo_unitario = 0
            sobreprecio_unitario = 0
            descripcion_prod = tp.descripcion_linea or 'Ítem pendiente de despacho'

        dcto_unit = tp.descuento_unitario or 0
        dcto_pct = float(tp.porcentaje_descuento or 0)
        dcto_monto_linea = dcto_unit * tp.stock if dcto_unit else 0

        # `precio_efectivo` = precio antes-de-descuento realmente cobrado.
        # Cubre el caso del "envío" con precio sistema fijo (ej. 500) que el
        # operador sube en el POS (a 800): si `tp.precio` se quedó con el
        # precio sistema pero `tp.subtotal` refleja lo cobrado, derivamos el
        # precio real desde subtotal/stock + descuento. Si `tp.precio` ya
        # cuadra con `tp.subtotal/tp.stock + dcto_unit`, no cambia nada.
        precio_efectivo = tp.precio
        if tp.stock and tp.subtotal:
            derivado = int(round(tp.subtotal / tp.stock)) + dcto_unit
            if derivado and derivado != tp.precio:
                precio_efectivo = derivado

        Dte_Productos.objects.create(
            dte=dte,
            productoTalla=tp.ProductoTalla,
            stock=tp.stock,
            costo=costo_unitario,
            sobreprecio=sobreprecio_unitario,
            precio=precio_efectivo,
            precio_unitario=precio_efectivo,
            descuento_pct=dcto_pct if dcto_pct > 0 else None,
            descuento_monto=dcto_monto_linea if dcto_monto_linea > 0 else None,
            monto_item=tp.subtotal,
            descripcion=descripcion_prod[:255],
            es_pendiente_despacho=tp.es_pendiente_despacho,
            # Espejo del ítem de cotización que originó la línea: sin esto no
            # hay forma de completar esta fila cuando el SKU se asigna después
            # (despacho diferido) y el costo del documento queda en $0.
            cotizacion_detalle_id=tp.cotizacion_detalle_id,
        )
    
    # Copiar métodos de pago
    for pago in ticket.pagos.all():
        Dte_Detalle_Pago.objects.create(
            dte=dte,
            metodo_pago=pago.metodo_pago,
            monto=pago.monto,
            tipo_tarjeta=pago.tipo_tarjeta or '',
            voucher=pago.voucher or '',
            notas=pago.notas or ''
        )
    
    # ✅ Actualizar movimientos del ticket para que también referencien el DTE
    from .models import Movimientos_Producto
    movimientos_ticket = Movimientos_Producto.objects.filter(ticket=ticket, dte__isnull=True)
    for mov in movimientos_ticket:
        mov.dte = dte
        mov.observaciones = f"{mov.observaciones or ''} - DTE {dte.tipo_documento} #{dte.numero_documento}".strip(' -')
        # ✅ Actualizar referencia_externa con el número de DTE
        mov.referencia_externa = f'DTE_{dte.numero_documento}'
        mov.save()
    
    # ✅ Actualizar el campo folio_dte del ticket
    ticket.folio_dte = dte.numero_documento
    ticket.dte_generado = True
    ticket.dte_fecha_generacion = timezone.now()
    ticket.save()
    
    logger.info(
        "DTE generado: ticket_id=%s, tipo=%s, numero=%s, movimientos_actualizados=%s",
        ticket.id,
        dte.tipo_documento,
        dte.numero_documento,
        movimientos_ticket.count(),
    )
    
    # ✅ Generar archivo TXT para Acepta (solo para documentos electrónicos, no para BOLETA PAPEL)
    archivo_txt_data = None
    
    # Solo generar TXT si NO es BOLETA PAPEL
    if tipo_dte != 'BOLETA PAPEL':
        try:
            # ✅ Importar AMBAS funciones al inicio del try. `limpiar_texto` se usa
            #    más abajo en metodos_pago_texto (línea ~2491); como Python trata
            #    cualquier nombre importado dentro de la función como local en TODO
            #    el ámbito, importarlo recién antes de los datos del emisor causaba
            #    UnboundLocalError ("referenced before assignment") y el TXT no se
            #    generaba (archivo_txt_data quedaba None → no se descargaba el TXT).
            from .views_modulo_documentos import generar_txt_dte_acepta, limpiar_texto

            # Preparar datos para TXT
            empresa = ticket.sucursal.empresa
            
            # Preparar información de métodos de pago (enriquecida para observaciones del TXT Acepta).
            # Incluye: método, monto, tipo de tarjeta, voucher/autorización Transbank y datos del POS
            # (terminal + número de operación guardados en el campo 'notas').
            metodos_pago_info = []
            for pago in ticket.pagos.all():
                metodo_nombre = dict(METODO_PAGO_TICKET_CHOICES).get(pago.metodo_pago, pago.metodo_pago)
                partes = [f"{metodo_nombre}: ${pago.monto:,}"]

                # Tipo de tarjeta (VISA, MASTERCARD, AMEX, etc.)
                if pago.tipo_tarjeta:
                    partes.append(f"Tarj: {pago.tipo_tarjeta}")

                # Código de autorización Transbank (voucher)
                if pago.voucher:
                    partes.append(f"Auth: {pago.voucher}")

                # Orden de compra si corresponde (ecommerce, convenios, etc.)
                if getattr(pago, 'numero_orden_compra', None):
                    partes.append(f"OC: {pago.numero_orden_compra}")

                # Notas: para pagos POS Transbank trae "Terminal: XXX | Op: YYY"
                if pago.notas:
                    # Compactar y truncar para que la observación global no explote
                    notas_compactas = ' '.join(str(pago.notas).split())[:80]
                    partes.append(notas_compactas)

                metodos_pago_info.append(' - '.join(partes))

            # Limpiar cada método por separado y unir con '|' DESPUÉS. Si se
            # limpiara la cadena ya unida, limpiar_texto() borraría los '|'
            # separadores (es el separador de campos Acepta) y la observación
            # del TXT solo mostraría el primer método cuando hay varios.
            metodos_pago_texto = (
                ' | '.join(limpiar_texto(m) for m in metodos_pago_info)
                if metodos_pago_info else 'EFECTIVO'
            )
            
            # ✅ DETECTAR SI ES TICKET DE CAMBIO/DEVOLUCIÓN
            es_ticket_cambio = (ticket.modulo_origen == 'CAMBIO_DEVOLUCION')
            
            # Preparar productos para el TXT
            productos_txt = []
            
            if es_ticket_cambio:
                logger.debug(
                    "Generando TXT para ticket de cambio: ticket_id=%s, productos con precio 0 y diferencia",
                    ticket.id,
                )
                
                # Para tickets de cambio: mostrar productos con precio $0
                for tp in ticket.ticket_productos.all():
                    if tp.ProductoTalla is None:
                        productos_txt.append({
                            # Marcador trazable, no el ID interno de tabla.
                            'sku': (f'PEND-{tp.cotizacion_detalle_id}'
                                    if tp.cotizacion_detalle_id else 'PEND'),
                            'nombre': (tp.descripcion_linea or 'Ítem pendiente')[:80],
                            'descripcion': '',
                            'cantidad': tp.stock,
                            'precio_unitario': 0,
                            'total': 0
                        })
                        continue
                    producto = tp.ProductoTalla.producto
                    # Usar descripción del producto si existe, sino artículo
                    nombre_producto = producto.descripcion if producto and producto.descripcion else producto.articulo
                    
                    productos_txt.append({
                        'sku': tp.ProductoTalla.sku,
                        'nombre': nombre_producto[:80],
                        'descripcion': '',  # Dejar vacío para evitar duplicados
                        'cantidad': tp.stock,
                        'precio_unitario': 0,  # ✅ PRECIO $0 para productos en cambio
                        'total': 0
                    })
                
                # Agregar ítem "DIFERENCIA DE CAMBIO" con el total (si es positivo)
                diferencia = int(ticket.total or 0)
                if diferencia > 0:
                    productos_txt.append({
                        'sku': 'DIF',
                        'nombre': 'DIFERENCIA DE CAMBIO',
                        'descripcion': '',
                        'cantidad': 1,
                        'precio_unitario': diferencia,
                        'total': diferencia
                    })
                else:
                    # Si es negativo o cero, agregar con $0
                    productos_txt.append({
                        'sku': 'DIF',
                        'nombre': 'DIFERENCIA DE CAMBIO',
                        'descripcion': '',
                        'cantidad': 1,
                        'precio_unitario': 0,
                        'total': 0
                    })
            else:
                # Ticket normal: productos con sus precios reales
                
                # ✅ Si viene de cotización, crear mapa de descripciones por SKU
                descripciones_cotizacion = {}
                if cotizacion:
                    try:
                        # Obtener descripciones de los items de la cotización
                        for item in cotizacion.items.all().prefetch_related('skus_asociados__producto_talla'):
                            for sku_rel in item.skus_asociados.all():
                                if sku_rel.producto_talla:
                                    # Usar la descripción del item de la cotización
                                    descripciones_cotizacion[sku_rel.producto_talla.sku] = item.descripcion
                        logger.debug(
                            "Descripciones de cotizacion cargadas para TXT: ticket_id=%s, productos=%s",
                            ticket.id,
                            len(descripciones_cotizacion),
                        )
                    except Exception:
                        logger.exception("Error al cargar descripciones de cotizacion para TXT ticket_id=%s", ticket.id)
                
                # SKU esperado de los ítems pendientes de la cotización, para no
                # emitir el ID interno de tabla como CdgItem del DTE.
                skus_pendientes_cot = {}
                if cotizacion:
                    for item in cotizacion.items.all():
                        if item.sku_producto_pendiente:
                            skus_pendientes_cot[item.id] = item.sku_producto_pendiente

                for tp in ticket.ticket_productos.all():
                    if tp.ProductoTalla is None:
                        # Ítem manual / pendiente de despacho — usar descripción de línea.
                        # El código del ítem debe ser algo trazable: el SKU esperado
                        # que cargó el vendedor, o un marcador PEND-<linea>. Antes se
                        # emitía `tp.cotizacion_detalle_id`, un ID de tabla interno.
                        sku = (
                            skus_pendientes_cot.get(tp.cotizacion_detalle_id)
                            or (f'PEND-{tp.cotizacion_detalle_id}'
                                if tp.cotizacion_detalle_id else 'PEND')
                        )
                        nombre_producto = tp.descripcion_linea or 'Ítem pendiente de despacho'
                    else:
                        producto = tp.ProductoTalla.producto
                        sku = tp.ProductoTalla.sku
                        
                        # ✅ PRIORIDAD: 1) Descripción de cotización, 2) Descripción del producto, 3) Artículo
                        if sku in descripciones_cotizacion and descripciones_cotizacion[sku]:
                            nombre_producto = descripciones_cotizacion[sku]
                            logger.debug(
                                "TXT ticket_id=%s sku=%s usando descripcion de cotizacion: %s",
                                ticket.id,
                                sku,
                                nombre_producto[:40],
                            )
                        elif producto and producto.descripcion:
                            nombre_producto = producto.descripcion
                        else:
                            nombre_producto = producto.articulo if producto else str(sku)
                    
                    if not es_boleta:
                        # Factura: precio completo neto por unidad; el descuento
                        # va como DscRcgGlobal (tabla 3) para que aparezca como
                        # línea visible en el documento. monto_item = precio_neto × qty.
                        precio_unitario_txt = int(round(Decimal(tp.precio) / Decimal('1.19')))
                        monto_descuento_txt = 0  # el descuento se muestra en DscRcgGlobal
                        monto_item_txt = precio_unitario_txt * tp.stock
                    else:
                        # Boleta: precio IVA-inclusive completo; el descuento va
                        # como DscRcgGlobal (tabla 4) después de las observaciones.
                        precio_unitario_txt = tp.precio
                        monto_descuento_txt = 0
                        monto_item_txt = tp.precio * tp.stock
                    
                    productos_txt.append({
                        'sku': sku,
                        'nombre': nombre_producto[:80],
                        'descripcion': '',
                        'cantidad': tp.stock,
                        'precio_unitario': precio_unitario_txt,
                        'descuento_pct': 0,          # el descuento va en DscRcgGlobal
                        'monto_descuento': monto_descuento_txt,
                        'total': monto_item_txt
                    })
            
            # (limpiar_texto ya fue importado al inicio del try, junto a generar_txt_dte_acepta)

            # Datos del documento - ✅ Aplicar limpiar_texto para eliminar acentos y Ñ
            datos_txt = {
                'documento': {
                    'tipo_documento': 39 if es_boleta else 33,  # 39=Boleta, 33=Factura
                    'folio': dte.numero_documento,
                    'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
                    'forma_pago': forma_pago_dte,
                    'fecha_vencimiento': fecha_vencimiento_dte.strftime('%Y-%m-%d'),
                    'ind_servicio': 3,  # Venta y servicios (para boleta)
                    'timestamp': timezone.now().strftime('%Y-%m-%dT%H:%M:%S')
                },
                'emisor': {
                    'rut': empresa.rut,
                    'razon_social': limpiar_texto(empresa.razon_social or empresa.nombre),
                    'giro': limpiar_texto(empresa.giro or 'Sin giro'),
                    'acteco': empresa.acteco or '',
                    # La dirección/comuna/ciudad del emisor en el DTE debe ser la de la SUCURSAL
                    # donde se hizo la venta (boletas físicas de cada tienda), cayendo a la
                    # casa matriz solo si la sucursal no tiene el dato cargado.
                    'direccion': limpiar_texto((ticket.sucursal.direccion if ticket.sucursal else '') or empresa.direccion or ''),
                    'comuna': limpiar_texto((ticket.sucursal.comuna if ticket.sucursal else '') or empresa.comuna or ''),
                    'ciudad': limpiar_texto((ticket.sucursal.ciudad if ticket.sucursal else '') or empresa.ciudad or ''),
                    'codigo_vendedor': limpiar_texto(ticket.vendedor.codigo_vendedor if ticket.vendedor else 'VENDEDOR'),
                    'nombre_vendedor': limpiar_texto(ticket.vendedor.nombre if ticket.vendedor else 'Sin vendedor'),
                    'vendedor_impresion': limpiar_texto(
                        ticket.vendedor.nombre if ticket.vendedor else (ticket.responsable or 'VENDEDOR')
                    ),
                    # Ya viene limpio método a método (preservando los '|').
                    'metodos_pago': metodos_pago_texto,
                    'correlativo_ticket': ticket.correlativo,
                    'telefono': empresa.contacto1 or '',
                    'nombre_impresora_boleta': getattr(ticket.sucursal, 'nombre_impresora_boleta', 'boleta') or 'boleta',
                    'nombre_impresora_factura': getattr(ticket.sucursal, 'nombre_impresora_factura', 'factura') or 'factura',
                    'sucursal': limpiar_texto(ticket.sucursal.alias if ticket.sucursal else ''),
                },
                # Receptor de factura: la fuente de verdad son los datos del TICKET
                # (lo que el cajero vio y corrigió en el POS), con fallback a la ficha
                # Empresa. Antes se leía SOLO `receptor.*`, así que la factura salía con
                # la comuna/ciudad vieja de la ficha (o de una ficha duplicada arbitraria)
                # e ignoraba lo tipeado en el POS. Para boleta el receptor es consumidor final.
                'receptor': {
                    'rut': (ticket.cliente_rut or (receptor.rut if receptor else '')) if not es_boleta else '66666666-6',
                    'razon_social': limpiar_texto((ticket.cliente_nombre or (receptor.razon_social if receptor else '')) if not es_boleta else 'CONSUMIDOR FINAL'),
                    'giro': limpiar_texto((ticket.cliente_giro or (receptor.giro if receptor else '')) if not es_boleta else ''),
                    'direccion': limpiar_texto((ticket.cliente_direccion or (receptor.direccion if receptor else '')) if not es_boleta else ''),
                    'comuna': limpiar_texto((ticket.cliente_comuna or (receptor.comuna if receptor else '')) if not es_boleta else ''),
                    'ciudad': limpiar_texto((ticket.cliente_ciudad or (receptor.ciudad if receptor else '')) if not es_boleta else '')
                },
                'totales': {
                    'monto_neto': int(neto),
                    'monto_exento': 0,
                    'tasa_iva': 19,
                    'iva': int(iva),
                    'monto_total': int(total),
                    'descuento_global': 0
                },
                'detalle': [],
                'observaciones': ticket.observaciones or '',
                'observaciones_adicionales': ticket.observaciones_adicionales or ''
            }

            if not es_boleta:
                referencias_txt = []
                referencias_modelo = ticket.referencias.all()
                if referencias_modelo.exists():
                    for ref in referencias_modelo:
                        referencias_txt.append({
                            'tipo_documento': ref.tipo_documento,
                            'folio': ref.folio,
                            'fecha': ref.fecha.strftime('%Y-%m-%d'),
                            'razon': '',
                        })
                elif ticket.referencia_tipo and ticket.referencia_folio:
                    referencias_txt.append({
                        'tipo_documento': ticket.referencia_tipo,
                        'folio': ticket.referencia_folio,
                        'fecha': ticket.referencia_fecha.strftime('%Y-%m-%d') if ticket.referencia_fecha else '',
                        'razon': '',
                    })
                datos_txt['referencias'] = referencias_txt
            
            for prod_txt in productos_txt:
                sku_str = str(prod_txt.get('sku', ''))
                datos_txt['detalle'].append({
                    'codigo': limpiar_texto(sku_str[:35]),
                    'sku': limpiar_texto(sku_str),
                    'nombre': limpiar_texto(prod_txt['nombre']),
                    'descripcion': limpiar_texto(prod_txt.get('descripcion', '')),
                    'cantidad': prod_txt['cantidad'],
                    'unidad': 'UN',
                    'precio_unitario': prod_txt['precio_unitario'],
                    'descuento_pct': prod_txt.get('descuento_pct', 0),
                    'monto_descuento': prod_txt.get('monto_descuento', 0),
                    'monto_item': prod_txt['total']
                })
            
            # Detect discounts (per-item or global) and add Tabla 4 block + fix total.
            #
            # IMPORTANTE: El único descuento "real" es el que está materializado
            # en las líneas (tp.descuento_unitario). NO usamos ticket.descuento
            # como fallback porque ese campo puede quedar stale al agregar
            # productos nuevos en el paso 3 (ver recalc en registrar_pagos_ticket).
            # Si ticket.descuento > 0 pero ninguna línea trae descuento, lo
            # ignoramos para evitar generar una línea "Descuento: $X" fantasma.
            descuento_items = sum(
                (tp.descuento_unitario or 0) * tp.stock
                for tp in ticket.ticket_productos.all()
            )
            descuento_efectivo = descuento_items  # solo líneas; no arrastrar stale

            # Boleta (39/41): DscRcgGlobal en monto IVA-inclusive — la sección
            # tabla 4 va DESPUÉS de las observaciones (formato Acepta oficial).
            # Factura (33/34): DscRcgGlobal en monto NETO — tabla 3, antes de
            # referencias. El monto_item de cada línea usa precio completo neto
            # para que sum(items) - dcto_neto = monto_neto del header.
            def _valor_dr(monto_bruto):
                if es_boleta:
                    return int(monto_bruto)  # IVA-inclusive
                return int(round(Decimal(monto_bruto) / Decimal('1.19')))  # neto

            # Los descuentos de CABECERA (vale de puntos, cupón nominativo) ya
            # están restados del header (`total`/`neto`), pero el detalle va a
            # precio completo: si no se declaran también como DscRcgGlobal, el
            # documento queda con sum(detalle) != MntTotal y Acepta lo timbra por
            # el monto bruto — se declara IVA de plata que nunca se cobró y la
            # boleta que se lleva el cliente dice más de lo que pagó.
            # En un ticket de cambio la línea sintética "DIFERENCIA DE CAMBIO"
            # ya vale `ticket.total`, o sea con los descuentos de cabecera YA
            # restados: declararlos otra vez acá los aplicaría dos veces.
            # (El cobro además rechaza el cupón en esos tickets; este guard es
            # la segunda barrera, y cubre los vales heredados.)
            if es_ticket_cambio:
                descuento_vale = 0
                descuento_cupon = 0

            descuentos_recargos_txt = []
            if descuento_efectivo > 0:
                descuentos_recargos_txt.append({
                    'tpo_mov': 'D',
                    'glosa_dr': 'Descuento',
                    'tpo_valor': '$',
                    'valor_dr': _valor_dr(descuento_efectivo),
                })
            if descuento_vale > 0:
                descuentos_recargos_txt.append({
                    'tpo_mov': 'D',
                    'glosa_dr': 'Descuento Puntos Fidelizacion',
                    'tpo_valor': '$',
                    'valor_dr': _valor_dr(descuento_vale),
                })
            if descuento_cupon > 0:
                descuentos_recargos_txt.append({
                    'tpo_mov': 'D',
                    'glosa_dr': 'Descuento Cupon',
                    'tpo_valor': '$',
                    'valor_dr': _valor_dr(descuento_cupon),
                })

            if descuentos_recargos_txt:
                datos_txt['descuentos_recargos'] = descuentos_recargos_txt
                logger.debug(
                    "TXT %s con descuento global: ticket_id=%s, lineas=%s, vale=%s, "
                    "cupon=%s, glosas=%s, tipo_valor=%s",
                    'Boleta' if es_boleta else 'Factura',
                    ticket.id,
                    descuento_efectivo,
                    descuento_vale,
                    descuento_cupon,
                    [d['glosa_dr'] for d in descuentos_recargos_txt],
                    'IVA-incl' if es_boleta else 'neto',
                )
                # monto_total ya es correcto (int(total) = ticket.total = monto
                # descontado IVA-inclusive). NO se sobreescribe aquí.

            # ── Cuadratura del detalle contra el header (solo facturas) ──────
            # En boleta el detalle va IVA-inclusive y ya cuadra por construcción.
            if not es_boleta:
                residuo = cuadrar_detalle_neto(
                    datos_txt['detalle'],
                    int(neto),
                    datos_txt.get('descuentos_recargos'),
                )
                if residuo:
                    logger.info(
                        "TXT factura: residuo de redondeo %s absorbido en la última "
                        "línea. ticket_id=%s dte=%s neto=%s",
                        residuo, ticket.id, dte.numero_documento, int(neto),
                    )

            # Generar TXT
            contenido_txt = generar_txt_dte_acepta(datos_txt)
            
            # Preparar datos del archivo para retornar
            nombre_archivo = f"{tipo_dte.replace(' ', '_')}_{dte.numero_documento}_{ticket.correlativo}.txt"
            archivo_txt_data = {
                'contenido': contenido_txt,
                'nombre_archivo': nombre_archivo
            }
            
            logger.info(
                "Archivo TXT generado: ticket_id=%s, dte_numero=%s, archivo=%s",
                ticket.id,
                dte.numero_documento,
                nombre_archivo,
            )
            
        except Exception as e:
            # Antes este error se tragaba en silencio (solo print): el TXT no se
            # generaba y la UI no descargaba nada SIN decir por qué (típico al
            # facturar un pedido de otra sucursal desde la sesión actual). Ahora
            # se loguea con stacktrace + contexto de sucursal y se expone el
            # motivo en `dte._txt_error` para mostrarlo en el resultado. La boleta
            # se sigue emitiendo (el TXT de Acepta es un export aparte).
            logger.error(
                'Error al generar TXT DTE %s | sucursal_facturacion=%s | tipo=%s: %s',
                getattr(dte, 'numero_documento', '?'),
                getattr(ticket, 'sucursal_id', None),
                tipo_dte, e, exc_info=True,
            )
            dte._txt_error = str(e)[:300]

    # Guardar datos del TXT en el DTE para referencia
    dte.archivo_txt_data = archivo_txt_data

    return dte


@login_required
@require_http_methods(["POST"])
def registrar_pagos_ticket(request, correlativo):
    """Registrar pagos para un ticket en POS"""
    logger.debug("Inicio registrar_pagos_ticket ticket=%s", correlativo)
    
    sucursal_id = (
        request.session.get('idSucursalActual')
        or request.session.get('sucursalActual')
        or request.session.get('idSucursalActualPOS')
    )
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión'}, status=400)

    # =========================================================================
    # NUEVO: Manejar cotizaciones cargadas como ticket
    # Si el correlativo empieza con "COT-", es una cotización que necesita
    # crear un ticket nuevo antes de procesar el pago
    # =========================================================================
    es_cotizacion = str(correlativo).startswith('COT-')
    cotizacion_obj = None
    productos_ya_creados_desde_cotizacion = False  # Bandera para evitar duplicar productos
    
    if es_cotizacion:
        logger.info("registrar_pagos_ticket detecto cotizacion correlativo=%s", correlativo)
        try:
            payload_check = json.loads(request.body or '{}')
            cotizacion_id = payload_check.get('cotizacion_id')
            
            if not cotizacion_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'Cotización detectada pero falta cotizacion_id en el payload'
                }, status=400)
            
            # Importar modelo de cotización
            from .models import Cotizacion_Empresa, Historial_Cotizacion
            
            # ✅ Buscar la cotización por ID (sin filtrar por sucursal activa)
            cotizacion_obj = Cotizacion_Empresa.objects.filter(id=cotizacion_id).first()
            
            if not cotizacion_obj:
                return JsonResponse({
                    'success': False, 
                    'error': f'Cotización {cotizacion_id} no encontrada'
                }, status=404)
            
            # ✅ IMPORTANTE: Usar la sucursal de la COTIZACIÓN, no la de la sesión
            # La cotización se debe facturar en su sucursal original
            sucursal_id = cotizacion_obj.sucursal_id
            
            # ✅ Verificar si la cotización ya fue facturada
            if cotizacion_obj.facturada:
                return JsonResponse({
                    'success': False, 
                    'error': f'La cotización {cotizacion_obj.numero_cotizacion} ya fue facturada con documento {cotizacion_obj.numero_factura}'
                }, status=400)
            
            if not cotizacion_obj.esta_vigente:
                return JsonResponse({
                    'success': False, 
                    'error': 'La cotización no está vigente o está vencida'
                }, status=400)
            
            sucursal = get_object_or_404(Sucursal, id=sucursal_id)
            
            # Obtener productos de la cotización
            productos_cotizacion = payload_check.get('productos', [])
            if not productos_cotizacion:
                return JsonResponse({
                    'success': False, 
                    'error': 'No hay productos en la cotización'
                }, status=400)
            
            # ✅ VALIDAR STOCK ANTES DE CREAR EL TICKET
            # IMPORTANTE: Usar producto_talla_id si está disponible, ya que el SKU puede existir
            # en múltiples sucursales y .first() retornaría el incorrecto
            productos_sin_stock = []
            
            for prod_data in productos_cotizacion:
                producto_talla = None
                
                # Skip pending items — they have no SKU and don't need stock validation
                if prod_data.get('es_pendiente_despacho'):
                    continue
                
                # ✅ PRIMERO: Intentar obtener por producto_talla_id (más preciso)
                producto_talla_id = prod_data.get('producto_talla_id')
                if producto_talla_id:
                    producto_talla = Producto_Talla.objects.filter(id=producto_talla_id).first()
                
                # FALLBACK: Si no hay producto_talla_id, buscar por SKU
                if not producto_talla:
                    sku = prod_data.get('sku')
                    if sku:
                        # Filtrar por SKU Y por sucursal de la cotización para evitar ambigüedad
                        producto_talla = Producto_Talla.objects.filter(
                            sku=sku,
                            producto__sucursal_id=sucursal_id
                        ).first()
                        # Si no existe en la sucursal, buscar global (compatibilidad)
                        if not producto_talla:
                            producto_talla = Producto_Talla.objects.filter(sku=sku).first()
                
                if producto_talla:
                    stock_disponible = producto_talla.stock_sucursal(sucursal_id)
                    cantidad_requerida = int(prod_data.get('cantidad', 1))
                    sku_display = prod_data.get('sku') or producto_talla.sku
                    
                    if stock_disponible < cantidad_requerida:
                        productos_sin_stock.append({
                            'sku': str(sku_display),
                            'nombre': producto_talla.producto.articulo if producto_talla.producto else 'Sin nombre',
                            'stock_disponible': stock_disponible,
                            'cantidad_requerida': cantidad_requerida
                        })
            
            if productos_sin_stock:
                detalle = ', '.join([f"SKU {p['sku']}: {p['stock_disponible']}/{p['cantidad_requerida']}" for p in productos_sin_stock])
                logger.warning(
                    "Stock insuficiente al facturar cotizacion=%s detalle=%s",
                    cotizacion_obj.numero_cotizacion,
                    detalle,
                )
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuficiente para facturar. {detalle}',
                    'error_tipo': 'STOCK_INSUFICIENTE',
                    'productos_sin_stock': productos_sin_stock
                }, status=400)
            
            logger.debug(
                "Stock validado para cotizacion=%s productos=%s",
                cotizacion_obj.numero_cotizacion,
                len(productos_cotizacion),
            )
            
            # Crear ticket desde la cotización
            logger.info(
                "Creando ticket desde cotizacion=%s sucursal_id=%s",
                cotizacion_obj.numero_cotizacion,
                sucursal_id,
            )
            
            # Obtener siguiente correlativo para ticket
            nuevo_correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
            
            # Calcular totales
            subtotal_calc = sum(p.get('subtotal', 0) for p in productos_cotizacion)
            total = int(cotizacion_obj.total)
            
            # Datos del cliente
            datos_cliente = payload_check.get('cliente', {})
            
            # Obtener vendedor (usar el de la cotización o buscar uno por defecto)
            vendedor = cotizacion_obj.vendedor
            if not vendedor:
                # Buscar vendedor activo asignado a esta sucursal (ManyToMany)
                vendedor = Vendedor.objects.filter(sucursales=sucursal, activo=True).first()
                if not vendedor:
                    # Buscar cualquier vendedor de la sucursal
                    vendedor = Vendedor.objects.filter(sucursales=sucursal).first()
                    if not vendedor:
                        return JsonResponse({
                            'success': False, 
                            'error': 'No hay vendedores configurados para esta sucursal'
                        }, status=400)
            
            # Crear el ticket
            ticket = Ticket.objects.create(
                correlativo=nuevo_correlativo,
                sucursal=sucursal,
                vendedor=vendedor,
                subTotal=int(subtotal_calc),  # Campo es subTotal con T mayúscula
                descuento=int(cotizacion_obj.descuento or 0),
                total=total,
                estado='PENDIENTE',
                responsable=request.user.username,  # Campo obligatorio
                observaciones=cotizacion_obj.observaciones or '',
                observaciones_adicionales=f'Facturación de cotización {cotizacion_obj.numero_cotizacion}. {cotizacion_obj.descripcion or ""}',
                cliente_nombre=datos_cliente.get('nombre', cotizacion_obj.cliente.nombre),
                cliente_rut=formatear_rut(datos_cliente.get('rut', cotizacion_obj.cliente.rut)),
                cliente_giro=datos_cliente.get('giro', cotizacion_obj.cliente.giro or ''),
                cliente_direccion=datos_cliente.get('direccion', cotizacion_obj.cliente.direccion or ''),
                cliente_comuna=datos_cliente.get('comuna', cotizacion_obj.cliente.comuna or ''),
                cliente_ciudad=datos_cliente.get('ciudad', cotizacion_obj.cliente.ciudad or ''),
                cliente_email=datos_cliente.get('email', cotizacion_obj.cliente.correoIntercambio or ''),
                cliente_email_facturacion=datos_cliente.get('email_facturacion', cotizacion_obj.cliente.correoAdministrador or ''),
                modulo_origen='POS'  # Usar POS ya que COTIZACION no existe en choices
            )
            
            # Crear productos del ticket
            # ✅ IMPORTANTE: Usar producto_talla_id si está disponible
            for prod_data in productos_cotizacion:
                producto_talla = None
                sku_display = prod_data.get('sku', 'N/A')

                cantidad = int(prod_data.get('cantidad', 1))
                precio = int(prod_data.get('precio_unitario', prod_data.get('precio', 0)))
                descuento = int(prod_data.get('descuento_unitario', 0))
                subtotal_prod = int(prod_data.get('subtotal', cantidad * precio))

                # Ítem pendiente de despacho (sin SKU) — crear línea manual
                if prod_data.get('es_pendiente_despacho'):
                    Ticket_Productos.objects.create(
                        idTicket=ticket,
                        ProductoTalla=None,
                        stock=cantidad,
                        precio=precio,
                        precio_original=precio,
                        descuento_unitario=descuento,
                        subtotal=subtotal_prod,
                        porcentaje_descuento=0,
                        descripcion_linea=prod_data.get('articulo') or prod_data.get('descripcion') or 'Ítem pendiente',
                        es_pendiente_despacho=True,
                        cotizacion_detalle_id=prod_data.get('cotizacion_item_id'),
                    )
                    logger.debug(
                        "Item pendiente de despacho agregado a ticket=%s articulo=%s cantidad=%s",
                        ticket.correlativo,
                        prod_data.get('articulo'),
                        cantidad,
                    )
                    continue
                
                # ✅ PRIMERO: Intentar obtener por producto_talla_id (más preciso)
                producto_talla_id = prod_data.get('producto_talla_id')
                if producto_talla_id:
                    producto_talla = Producto_Talla.objects.filter(id=producto_talla_id).first()
                
                # FALLBACK: Si no hay producto_talla_id, buscar por SKU
                if not producto_talla:
                    sku = prod_data.get('sku')
                    if sku:
                        # Filtrar por SKU Y por sucursal de la cotización
                        producto_talla = Producto_Talla.objects.filter(
                            sku=sku,
                            producto__sucursal_id=sucursal_id
                        ).first()
                        # Si no existe en la sucursal, buscar global (compatibilidad)
                        if not producto_talla:
                            producto_talla = Producto_Talla.objects.filter(sku=sku).first()
                        sku_display = sku
                
                if not producto_talla:
                    logger.warning(
                        "ProductoTalla no encontrado al crear ticket desde cotizacion=%s producto_talla_id=%s sku=%s",
                        cotizacion_obj.numero_cotizacion,
                        producto_talla_id,
                        sku_display,
                    )
                    continue
                
                cantidad = int(prod_data.get('cantidad', 1))
                precio = int(prod_data.get('precio_unitario', prod_data.get('precio', 0)))
                descuento = int(prod_data.get('descuento_unitario', 0))
                subtotal_prod = int(prod_data.get('subtotal', cantidad * precio))
                
                Ticket_Productos.objects.create(
                    idTicket=ticket,
                    ProductoTalla=producto_talla,
                    stock=cantidad,
                    precio=precio,
                    precio_original=precio,
                    descuento_unitario=descuento,
                    subtotal=subtotal_prod,
                    porcentaje_descuento=0
                )
                logger.debug(
                    "Producto agregado a ticket=%s sku=%s producto_talla_id=%s cantidad=%s",
                    ticket.correlativo,
                    producto_talla.sku,
                    producto_talla.id,
                    cantidad,
                )
            
            # ✅ NUEVO: Actualizar datos de la Empresa si el usuario completó campos faltantes
            empresa_cliente = cotizacion_obj.cliente
            empresa_actualizada = False
            campos_actualizados = []
            
            # Solo actualizar si hay datos nuevos que la empresa no tenía
            if datos_cliente.get('giro') and not empresa_cliente.giro:
                empresa_cliente.giro = datos_cliente.get('giro')
                empresa_actualizada = True
                campos_actualizados.append('giro')
            
            if datos_cliente.get('direccion') and not empresa_cliente.direccion:
                empresa_cliente.direccion = datos_cliente.get('direccion')
                empresa_actualizada = True
                campos_actualizados.append('direccion')
            
            if datos_cliente.get('comuna') and not empresa_cliente.comuna:
                empresa_cliente.comuna = datos_cliente.get('comuna')
                empresa_actualizada = True
                campos_actualizados.append('comuna')
            
            if datos_cliente.get('ciudad') and not empresa_cliente.ciudad:
                empresa_cliente.ciudad = datos_cliente.get('ciudad')
                empresa_actualizada = True
                campos_actualizados.append('ciudad')
            
            if datos_cliente.get('email') and not empresa_cliente.correoIntercambio:
                empresa_cliente.correoIntercambio = datos_cliente.get('email')
                empresa_actualizada = True
                campos_actualizados.append('email')
            
            if datos_cliente.get('telefono') and not empresa_cliente.contacto1:
                empresa_cliente.contacto1 = datos_cliente.get('telefono')
                empresa_actualizada = True
                campos_actualizados.append('telefono')
            
            if empresa_actualizada:
                empresa_cliente.save()
                logger.info(
                    "Empresa cliente actualizada desde cotizacion=%s rut=%s campos=%s",
                    cotizacion_obj.numero_cotizacion,
                    empresa_cliente.rut,
                    ', '.join(campos_actualizados),
                )
            
            # Actualizar correlativo para el resto del proceso
            correlativo = nuevo_correlativo
            # Marcar que los productos ya fueron creados (evitar duplicados)
            productos_ya_creados_desde_cotizacion = True
            logger.info(
                "Ticket creado desde cotizacion ticket=%s cotizacion=%s",
                correlativo,
                cotizacion_obj.numero_cotizacion,
            )
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Formato JSON inválido'}, status=400)
        except Exception as e:
            logger.exception(
                "Error creando ticket desde cotizacion correlativo=%s sucursal_id=%s",
                correlativo,
                sucursal_id,
            )
            return JsonResponse({'success': False, 'error': f'Error al procesar cotización: {str(e)}'}, status=500)
    
    # =========================================================================
    # FIN: Manejo de cotizaciones
    # =========================================================================

    ticket = (
        Ticket.objects
        .select_related('sucursal', 'vendedor')
        .prefetch_related('pagos', 'ticket_productos__ProductoTalla__producto')
        .filter(sucursal_id=sucursal_id, correlativo=correlativo)
        .first()
    )

    if not ticket:
        return JsonResponse({'success': False, 'error': f'Ticket {correlativo} no encontrado'}, status=404)

    # Bloquear reprocesamiento de tickets ya finalizados
    if ticket.estado in ('PAGADO', 'ANULADO', 'DEVUELTO'):
        estados_display = {
            'PAGADO': 'ya fue pagado',
            'ANULADO': 'está anulado',
            'DEVUELTO': 'está devuelto',
        }
        return JsonResponse({
            'success': False,
            'error': f'El ticket #{correlativo} {estados_display.get(ticket.estado, "ya fue procesado")} y no puede volver a procesarse.',
            'ticket_ya_procesado': True,
            'estado': ticket.estado,
        }, status=400)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Formato JSON inválido'}, status=400)

    datos_cliente = payload.get('cliente', {})
    ticket.cliente_nombre = datos_cliente.get('nombre') or ''
    
    # Formatear RUT sin puntos antes de guardar
    rut_cliente = datos_cliente.get('rut') or ''
    if rut_cliente:
        ticket.cliente_rut = formatear_rut(rut_cliente)
    else:
        ticket.cliente_rut = ''
    
    ticket.cliente_giro = datos_cliente.get('giro') or ''
    ticket.cliente_comuna = datos_cliente.get('comuna') or ''
    ticket.cliente_ciudad = datos_cliente.get('ciudad') or ''
    ticket.cliente_direccion = datos_cliente.get('direccion') or ''
    ticket.cliente_telefono = datos_cliente.get('telefono') or ''
    ticket.cliente_telefono_secundario = datos_cliente.get('telefono_secundario') or ''
    ticket.cliente_email = datos_cliente.get('email') or ''
    ticket.cliente_email_facturacion = datos_cliente.get('email_facturacion') or ''

    ticket.observaciones = payload.get('observaciones') or ''
    ticket.observaciones_adicionales = payload.get('observaciones_adicionales') or ''

    # ✅ Guardar datos de referencia ÚNICA (Retrocompatibilidad)
    ticket.referencia_tipo = payload.get('referencia_tipo') or None
    ticket.referencia_folio = payload.get('referencia_folio') or None
    if payload.get('referencia_fecha'):
        from datetime import datetime
        try:
            ticket.referencia_fecha = datetime.strptime(payload.get('referencia_fecha'), '%Y-%m-%d').date()
        except:
            ticket.referencia_fecha = None
    else:
        ticket.referencia_fecha = None
    
    # ✅ Procesar MÚLTIPLES REFERENCIAS (Nuevo sistema)
    referencias_payload = payload.get('referencias', [])
    if referencias_payload and isinstance(referencias_payload, list) and len(referencias_payload) > 0:
        # Eliminar referencias anteriores
        ticket.referencias.all().delete()
        
        # Crear nuevas referencias
        for ref_data in referencias_payload:
            try:
                from datetime import datetime
                fecha_ref = datetime.strptime(ref_data.get('fecha'), '%Y-%m-%d').date()
                
                TicketReferencia.objects.create(
                    ticket=ticket,
                    tipo_documento=ref_data.get('tipo_documento', ''),
                    folio=ref_data.get('folio', ''),
                    fecha=fecha_ref,
                    observaciones=ref_data.get('observaciones', '')
                )
            except Exception as e:
                logger.exception(
                    "Error al crear referencia para ticket=%s data=%s",
                    ticket.correlativo,
                    ref_data,
                )
                continue

    nuevo_estado = payload.get('estado')
    if nuevo_estado and nuevo_estado in dict(ESTADO_TICKET_CHOICES):
        ticket.estado = nuevo_estado
    
    # ✅ NUEVO: Si es cotización y tiene pagos válidos, marcar automáticamente como PAGADO
    if productos_ya_creados_desde_cotizacion and ticket.estado == 'PENDIENTE':
        pagos_payload = payload.get('pagos', [])
        total_pagado = sum(int(p.get('monto', 0)) for p in pagos_payload if p.get('monto'))
        if total_pagado >= ticket.total:
            logger.info(
                "Cotizacion con pagos completos ticket=%s total_pagado=%s total_ticket=%s",
                ticket.correlativo,
                total_pagado,
                ticket.total,
            )
            ticket.estado = 'PAGADO'

    metodo_principal = payload.get('metodo_pago_principal')
    if metodo_principal and metodo_principal in dict(METODO_PAGO_TICKET_CHOICES):
        ticket.metodo_pago = metodo_principal

    # Guardar condición de pago DTE elegida por el usuario (1=Contado, 2=Crédito)
    condicion_pago_dte = payload.get('condicion_pago_dte')
    if condicion_pago_dte in (1, 2):
        import json as _json
        try:
            notas = _json.loads(ticket.observaciones_adicionales or '{}')
        except (ValueError, TypeError):
            notas = {}
        notas['condicion_pago_dte'] = condicion_pago_dte
        ticket.observaciones_adicionales = _json.dumps(notas)

    correlativo_confirmacion = payload.get('correlativo')
    # Solo validar correlativo si NO es una cotización (las cotizaciones tienen formato COT-xxx)
    if correlativo_confirmacion and not str(correlativo_confirmacion).startswith('COT-'):
        try:
            if int(correlativo_confirmacion) != ticket.correlativo:
                return JsonResponse({'success': False, 'error': 'Correlativo no coincide con el ticket cargado'}, status=400)
        except ValueError:
            pass  # Si no se puede convertir a int, ignorar validación

    # ✅ NUEVO: Procesar productos actualizados (incluye productos agregados como bolsas)
    # IMPORTANTE: Si es cotización, los productos ya fueron creados al crear el ticket
    # IMPORTANTE: Si es cambio/devolución, NO se permiten cambios en los productos
    productos_payload = payload.get('productos', [])
    if ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
        logger.debug("Saltando productos para ticket de cambio/devolucion ticket=%s", ticket.correlativo)
    elif productos_ya_creados_desde_cotizacion:
        logger.debug("Saltando productos ya creados desde cotizacion ticket=%s", ticket.correlativo)
    elif productos_payload and isinstance(productos_payload, list):
        logger.debug(
            "Procesando productos payload ticket=%s lineas=%s",
            ticket.correlativo,
            len(productos_payload),
        )

        from collections import defaultdict

        # --- Agrupar Ticket_Productos existentes por SKU (lista, no dict) ---
        existentes_por_sku = defaultdict(list)
        pt_por_sku = {}
        for tp in ticket.ticket_productos.select_related('ProductoTalla', 'ProductoTalla__producto').all():
            if tp.ProductoTalla is not None:
                existentes_por_sku[tp.ProductoTalla.sku].append(tp)
                pt_por_sku[tp.ProductoTalla.sku] = tp.ProductoTalla

        # --- Validar stock total por SKU (solo cuando la cantidad total cambia) ---
        cantidades_payload_sku = defaultdict(int)
        for pd in productos_payload:
            s = pd.get('sku', '')
            if s:
                cantidades_payload_sku[s] += int(pd.get('cantidad', 1))

        for sku_val, cant_payload in cantidades_payload_sku.items():
            cant_existente = sum(tp.stock for tp in existentes_por_sku.get(sku_val, []))
            if cant_payload > cant_existente:
                pt = pt_por_sku.get(sku_val)
                if not pt:
                    pt = Producto_Talla.objects.filter(
                        sku=sku_val, producto__sucursal_id=ticket.sucursal_id
                    ).select_related('producto').first()
                if pt:
                    stock_real = pt.stock_sucursal(ticket.sucursal_id)
                    if cant_payload > stock_real:
                        logger.warning(
                            "Stock insuficiente total ticket=%s sku=%s solicitado=%s disponible=%s",
                            ticket.correlativo,
                            sku_val,
                            cant_payload,
                            stock_real,
                        )
                        return JsonResponse({
                            'success': False,
                            'error': f'Stock insuficiente para SKU {sku_val} ({pt.producto.articulo}). '
                                     f'Disponible: {stock_real}, Solicitado: {cant_payload}.',
                            'error_tipo': 'STOCK_INSUFICIENTE',
                            'sku': str(sku_val),
                            'stock_disponible': stock_real,
                            'stock_requerido': cant_payload,
                        }, status=400)

        # --- Validar promos NxM del payload contra las campañas vigentes ---
        # Función pura sin escritura: reejecutable en cada sync sin efectos
        # dobles (registrar_pagos_ticket no es atómico).
        from .services.campanas_service import validar_promos_nxm_payload
        val_promo = validar_promos_nxm_payload(productos_payload, ticket.sucursal)
        if not val_promo['ok']:
            return JsonResponse({
                'success': False,
                'error': 'Promoción NxM inválida: ' + '; '.join(
                    e['error'] for e in val_promo['errores']),
                'error_tipo': 'PROMO_INVALIDA',
                'detalles': val_promo['errores'],
            }, status=400)

        ids_existentes_usados = set()

        for prod_data in productos_payload:
            sku = prod_data.get('sku', '')
            if not sku:
                continue

            cantidad = int(prod_data.get('cantidad', 1))
            precio_unitario = int(prod_data.get('precio_unitario', 0))
            precio_original_payload = int(prod_data.get('precio_original', precio_unitario))
            descuento_unitario = int(prod_data.get('descuento_unitario', 0))
            subtotal = int(prod_data.get('subtotal', cantidad * precio_unitario))
            porcentaje_descuento = round((descuento_unitario / precio_unitario) * 100, 2) if precio_unitario > 0 and descuento_unitario > 0 else 0
            promo_campana_id = prod_data.get('promo_campana_id') if prod_data.get('es_promo_nxm') else None

            # Buscar un TP existente del mismo SKU que aún no haya sido emparejado
            tp_match = None
            candidatos = existentes_por_sku.get(sku, [])
            for tp_c in candidatos:
                if tp_c.id not in ids_existentes_usados:
                    tp_match = tp_c
                    break

            if tp_match:
                ids_existentes_usados.add(tp_match.id)
                algo_cambio = (
                    tp_match.stock != cantidad
                    or tp_match.precio != precio_unitario
                    or tp_match.descuento_unitario != descuento_unitario
                    or tp_match.subtotal != subtotal
                    or tp_match.precio_original != precio_original_payload
                )
                if algo_cambio:
                    if precio_unitario != precio_original_payload:
                        logger.debug(
                            "Actualizando linea ticket=%s linea_id=%s sku=%s cantidad=%s->%s precio=%s->%s precio_original=%s",
                            ticket.correlativo,
                            tp_match.id,
                            sku,
                            tp_match.stock,
                            cantidad,
                            tp_match.precio,
                            precio_unitario,
                            precio_original_payload,
                        )
                    else:
                        logger.debug(
                            "Actualizando linea ticket=%s linea_id=%s sku=%s cantidad=%s->%s descuento=%s->%s",
                            ticket.correlativo,
                            tp_match.id,
                            sku,
                            tp_match.stock,
                            cantidad,
                            tp_match.descuento_unitario,
                            descuento_unitario,
                        )
                    tp_match.stock = cantidad
                    tp_match.precio = precio_unitario
                    tp_match.precio_original = precio_original_payload
                    tp_match.descuento_unitario = descuento_unitario
                    tp_match.porcentaje_descuento = porcentaje_descuento
                    tp_match.subtotal = subtotal
                    tp_match.promo_campana_id = promo_campana_id
                    tp_match.save()
                elif tp_match.promo_campana_id != promo_campana_id:
                    # Sin otros cambios pero cambió el marcador de promo NxM.
                    tp_match.promo_campana_id = promo_campana_id
                    tp_match.save(update_fields=['promo_campana'])
            else:
                producto_talla_id = prod_data.get('producto_talla_id')
                producto_talla = None
                if producto_talla_id:
                    producto_talla = Producto_Talla.objects.filter(id=producto_talla_id).first()
                if not producto_talla:
                    producto_talla = Producto_Talla.objects.filter(sku=sku).first()

                if producto_talla:
                    logger.debug(
                        "Creando nueva linea ticket=%s sku=%s cantidad=%s precio=%s precio_original=%s",
                        ticket.correlativo,
                        sku,
                        cantidad,
                        precio_unitario,
                        precio_original_payload,
                    )
                    tp_nuevo = Ticket_Productos.objects.create(
                        idTicket=ticket,
                        ProductoTalla=producto_talla,
                        stock=cantidad,
                        precio=precio_unitario,
                        precio_original=precio_original_payload,
                        descuento_unitario=descuento_unitario,
                        subtotal=subtotal,
                        porcentaje_descuento=porcentaje_descuento,
                        promo_campana_id=promo_campana_id,
                    )
                    ids_existentes_usados.add(tp_nuevo.id)
                else:
                    logger.warning(
                        "ProductoTalla no encontrado al procesar payload ticket=%s sku=%s",
                        ticket.correlativo,
                        sku,
                    )

        # Eliminar líneas huérfanas (existían en DB pero ya no están en el payload)
        for sku_list in existentes_por_sku.values():
            for tp_orphan in sku_list:
                if tp_orphan.id not in ids_existentes_usados:
                    logger.debug(
                        "Eliminando linea huerfana ticket=%s linea_id=%s sku=%s",
                        ticket.correlativo,
                        tp_orphan.id,
                        tp_orphan.ProductoTalla.sku,
                    )
                    tp_orphan.delete()

        # Recalcular totales del ticket — authoritative server-side calc
        todas_lineas = list(ticket.ticket_productos.all())
        nuevo_descuento_prod = sum((tp.descuento_unitario or 0) * tp.stock for tp in todas_lineas)

        # Recalculate each line's subtotal so we never trust a faulty
        # frontend value (e.g. JS `0 || gross` gives gross instead of 0).
        for tp in todas_lineas:
            correcto = (tp.precio - (tp.descuento_unitario or 0)) * tp.stock
            if tp.subtotal != correcto:
                logger.debug(
                    "Corrigiendo subtotal ticket=%s sku=%s subtotal=%s correcto=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku if tp.ProductoTalla else '?',
                    tp.subtotal,
                    correcto,
                )
                tp.subtotal = correcto
                tp.save(update_fields=['subtotal'])

        # tp.subtotal ya tiene el descuento aplicado línea por línea,
        # así que nuevo_subtotal equivale al total final después del dcto.
        # El "subtotal bruto" del ticket es la suma sin descuentos.
        nuevo_subtotal_neto = sum(tp.subtotal for tp in todas_lineas)
        nuevo_subtotal_bruto = sum((tp.precio or 0) * tp.stock for tp in todas_lineas)

        # Server-authoritative: descuento y total se derivan SIEMPRE de las
        # líneas actuales del ticket. No arrastramos un `ticket.descuento`
        # previo (evita el bug donde agregar un producto nuevo en paso 3 hacía
        # que el TXT del DTE generara una línea "Descuento: $X" fantasma igual
        # al precio del producto agregado, por haber dejado stale el campo
        # `ticket.descuento` de una cotización / intento anterior).
        ticket.descuento = nuevo_descuento_prod
        ticket.total = nuevo_subtotal_neto
        if hasattr(ticket, 'subTotal'):
            ticket.subTotal = nuevo_subtotal_bruto
        logger.debug(
            "Nuevo total ticket=%s total=%s subtotal_bruto=%s descuento=%s",
            ticket.correlativo,
            ticket.total,
            nuevo_subtotal_bruto,
            ticket.descuento or 0,
        )

    # ── Cupón de descuento nominativo ────────────────────────────────────────
    # El POS envía el código como campo opcional `codigo_cupon_descuento`. Se
    # valida sin consumir aquí; el consumo real ocurre después de
    # `ticket_se_pago` (mismo patrón que el vale de puntos y las gift cards).
    #
    # Va ANTES del vale porque son EXCLUYENTES: en una venta hay un solo
    # beneficio (ni descuento manual, ni vale, ni cupón a la vez).
    _codigo_cupon = (payload.get('codigo_cupon_descuento') or '').strip().upper()
    _codigo_vale_pts = (payload.get('codigo_vale_canje') or '').strip().upper()
    _cupon_descuento = 0

    if _codigo_cupon and _codigo_vale_pts:
        # Estado que la UI no debería permitir jamás. No se resuelve eligiendo
        # uno en silencio: si llegó acá, algo se desincronizó y hay que verlo.
        logger.warning(
            "Cobro con cupón Y vale a la vez ticket=%s cupon=%s vale=%s usuario=%s",
            ticket.correlativo, _codigo_cupon, _codigo_vale_pts, request.user.username,
        )
        return JsonResponse({
            'success': False,
            'error': ('Esta venta tiene un cupón de descuento y un vale de puntos. '
                      'Es uno u otro: quite uno para continuar.'),
            'error_tipo': 'BENEFICIOS_EXCLUYENTES',
        }, status=400)

    if _codigo_cupon:
        from .services import cupon_service as _cup_svc
        from .services import fidelizacion_service as _fid_svc

        # Un ticket de cambio/devolución no admite cupón: su total ES la
        # diferencia a cobrar y el TXT la emite como una única línea sintética
        # "DIFERENCIA DE CAMBIO" tomada de `ticket.total`. Si además se
        # declarara el cupón como descuento global, el documento quedaría con
        # el descuento restado dos veces y Acepta lo rechazaría.
        if ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
            return JsonResponse({
                'success': False,
                'error': ('Los cupones de descuento no aplican a cambios ni '
                          'devoluciones, sólo a ventas nuevas.'),
                'error_tipo': 'CUPON_NO_APLICA',
            }, status=400)

        # Mismo gate que el vale: es beneficio de cliente particular.
        _fideliza_cup, _motivo_cup = _fid_svc.venta_fideliza(
            ticket,
            tipo_documento=payload.get('tipo_documento'),
            cotizacion=cotizacion_obj,
        )
        if not _fideliza_cup:
            logger.warning(
                "Cupón rechazado ticket=%s codigo=%s motivo=%s",
                ticket.correlativo, _codigo_cupon, _motivo_cup,
            )
            return JsonResponse({
                'success': False,
                'error': ('Los cupones de descuento solo aplican a clientes '
                          f'particulares ({_motivo_cup}).'),
                'error_tipo': 'CUPON_NO_APLICA',
            }, status=400)

        # El chequeo de "descuento manual" se rehace ACÁ, sobre las líneas ya
        # recalculadas: si sólo confiáramos en la UI, bastaba con validar el
        # cupón primero y aplicar el descuento después para acumular ambos.
        _tiene_dcto_manual = ticket.ticket_productos.filter(
            descuento_unitario__gt=0).exists()

        _info_cupon = _cup_svc.validar_cupon(
            _codigo_cupon,
            monto=int(ticket.total or 0),
            rut_cliente=ticket.cliente_rut,
            sucursal=ticket.sucursal,
            tiene_dcto_manual=_tiene_dcto_manual,
            tiene_vale_puntos=False,   # ya se cortó arriba si venían los dos
        )
        if not _info_cupon.get('valido'):
            logger.warning(
                "Cupón no válido ticket=%s codigo=%s motivo=%s",
                ticket.correlativo, _codigo_cupon,
                _info_cupon.get('motivo_codigo'),
            )
            return JsonResponse({
                'success': False,
                'error': _info_cupon.get('motivo') or 'Cupón no válido.',
                'error_tipo': f"CUPON_{_info_cupon.get('motivo_codigo', 'INVALIDO')}",
            }, status=400)

        _cupon_descuento = int(_info_cupon.get('descuento_pesos') or 0)
        _cupon_descuento = min(_cupon_descuento, int(ticket.total or 0))
        ticket.descuento_cupon = _cupon_descuento
        ticket.total = ticket.total - _cupon_descuento
        logger.info(
            "Cupón aplicado ticket=%s codigo=%s descuento=%s nuevo_total=%s",
            ticket.correlativo, _codigo_cupon, _cupon_descuento, ticket.total,
        )
    else:
        ticket.descuento_cupon = None
    # ─────────────────────────────────────────────────────────────────────────

    # ── Vale de puntos fidelización ──────────────────────────────────────────
    # El POS envía el código del vale (leído/escaneado en paso 3) como campo
    # opcional `codigo_vale_canje`. Se valida sin debitar aquí; el débito real
    # ocurre después de `ticket_se_pago` (mismo patrón que GIFTCARD).
    _vale_descuento_pts = 0
    if _codigo_vale_pts:
        from .services import fidelizacion_service as _fid_svc

        # Los vales de puntos son beneficio de cliente particular: no se pueden
        # aplicar a facturas ni a ventas originadas en una cotización. El POS
        # oculta la tarjeta en esos casos, pero el backend es el que manda.
        _fideliza, _motivo_no_fid = _fid_svc.venta_fideliza(
            ticket,
            tipo_documento=payload.get('tipo_documento'),
            cotizacion=cotizacion_obj,
        )
        if not _fideliza:
            logger.warning(
                "Vale de puntos rechazado ticket=%s codigo=%s motivo=%s",
                ticket.correlativo, _codigo_vale_pts, _motivo_no_fid,
            )
            return JsonResponse(
                {'success': False,
                 'error': ('Los vales de puntos solo aplican a clientes '
                           f'particulares ({_motivo_no_fid}).'),
                 'error_tipo': 'VALE_NO_APLICA'},
                status=400,
            )

        _info_vale = _fid_svc.validar_vale(_codigo_vale_pts)
        if not _info_vale.get('canjeable'):
            return JsonResponse(
                {'success': False,
                 'error': (f"Vale de puntos no válido o ya utilizado "
                           f"({_info_vale.get('estado', '?')}).")},
                status=400,
            )
        _vale_descuento_pts = int(_info_vale.get('valor_pesos') or 0)
        _vale_descuento_pts = min(_vale_descuento_pts, ticket.total)  # no negativos
        ticket.descuento_fidelizacion = _vale_descuento_pts
        ticket.total = ticket.total - _vale_descuento_pts
        logger.debug(
            "Vale pts aplicado ticket=%s codigo=%s descuento=%s nuevo_total=%s",
            ticket.correlativo, _codigo_vale_pts, _vale_descuento_pts, ticket.total,
        )
    else:
        ticket.descuento_fidelizacion = None
    # ─────────────────────────────────────────────────────────────────────────

    pagos = payload.get('pagos') or []
    ids_existentes = list(ticket.pagos.values_list('id', flat=True))

    # ─────────── GUARD: los pagos deben cubrir el total para marcar PAGADO ───────────
    # El estado venía del cliente sin ninguna comprobación: un POST con
    # estado='PAGADO' y pagos=[] dejaba la venta pagada, consumía el stock y
    # emitía el DTE. Aquí el total ya es definitivo (descuentos y vale aplicados)
    # y los pagos del payload están completos, así que es el punto correcto.
    #
    # Se excluye CAMBIO_DEVOLUCION: ahí el total es la diferencia y puede ser 0 o
    # negativo, con su propio flujo de cobro.
    if VALIDAR_COBERTURA_PAGOS and ticket.estado == 'PAGADO' and ticket.modulo_origen != 'CAMBIO_DEVOLUCION':
        total_a_cubrir = int(ticket.total or 0)
        if total_a_cubrir > 0:
            total_pagado = 0
            for _p in pagos:
                try:
                    _monto = int(_p.get('monto', 0))
                except (TypeError, ValueError):
                    continue
                if _monto > 0:
                    total_pagado += _monto
            # Los pagos ya registrados en el ticket también cuentan.
            total_pagado += int(
                ticket.pagos.exclude(id__in=[p.get('id') for p in pagos if p.get('id')])
                .aggregate(t=Sum('monto'))['t'] or 0
            )

            if total_pagado < total_a_cubrir - TOLERANCIA_COBERTURA_PAGOS:
                logger.warning(
                    "Cobro rechazado por pagos insuficientes ticket=%s total=%s pagado=%s usuario=%s",
                    ticket.correlativo, total_a_cubrir, total_pagado, request.user.username,
                )
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'Los pagos (${total_pagado:,.0f}) no cubren el total del '
                        f'ticket (${total_a_cubrir:,.0f}). Faltan '
                        f'${total_a_cubrir - total_pagado:,.0f}.'
                    ).replace(',', '.'),
                    'error_tipo': 'PAGOS_INSUFICIENTES',
                    'total_ticket': total_a_cubrir,
                    'total_pagado': total_pagado,
                }, status=400)
    # ─────────────────────────────────────────────────────────────────────────────────

    # ── Pre-validación de stock de TODO el ticket ────────────────────────────
    # Corre ANTES de escribir pagos, del candado y del canje de vale/gift card:
    # es la única operación que puede abortar el cobro y aquí todavía no se
    # escribió nada, así que el 400 sale con el ticket intacto (antes revertía
    # el estado a mano y dejaba pagos ya escritos).
    #
    # AGREGADA POR TALLA: el POS crea varias líneas del mismo SKU (NxM, precio
    # modificado) y validar línea a línea dejaba pasar un 2x1 con stock 1.
    # Consulta fresca a propósito: `ticket.ticket_productos.all()` devuelve el
    # caché del prefetch de la carga del ticket y no ve líneas escritas durante
    # este mismo request.
    if ticket.estado == 'PAGADO' and ticket.modulo_origen != 'CAMBIO_DEVOLUCION':
        _requerido_por_talla = {}
        _tallas_ticket = {}
        faltantes = []
        for tp in Ticket_Productos.objects.filter(idTicket=ticket).select_related('ProductoTalla'):
            if tp.ProductoTalla_id is None:
                continue
            _requerido_por_talla[tp.ProductoTalla_id] = (
                _requerido_por_talla.get(tp.ProductoTalla_id, 0) + tp.stock
            )
            _tallas_ticket[tp.ProductoTalla_id] = tp.ProductoTalla
        for _talla_id, _requerido in _requerido_por_talla.items():
            _disponible = _tallas_ticket[_talla_id].stock_sucursal(ticket.sucursal_id)
            if _disponible < _requerido:
                faltantes.append({
                    'sku': str(_tallas_ticket[_talla_id].sku),
                    'stock_disponible': _disponible,
                    'stock_requerido': _requerido,
                })
        if faltantes:
            logger.warning(
                "Cobro abortado antes de tocar pagos/stock ticket=%s faltantes=%s",
                ticket.correlativo, faltantes,
            )
            detalle = ', '.join(
                f"SKU {f['sku']} (disp. {f['stock_disponible']}, req. {f['stock_requerido']})"
                for f in faltantes
            )
            return JsonResponse({
                'success': False,
                'error': f'Stock insuficiente en {len(faltantes)} producto(s): {detalle}',
                'error_tipo': 'STOCK_INSUFICIENTE',
                'sku': faltantes[0]['sku'],
                'stock_disponible': faltantes[0]['stock_disponible'],
                'stock_requerido': faltantes[0]['stock_requerido'],
                'faltantes': faltantes,
            }, status=400)
    # ─────────────────────────────────────────────────────────────────────────

    # ── Candado anti doble cobro (compare-and-set atómico) ──────────────────
    # El guard de estado del inicio lee el ticket ANTES de que el otro POST
    # commitee PAGADO, así que dos cobros simultáneos (doble click / reintento
    # del front con timeout) pasaban ambos y el stock se descontaba dos veces:
    # 25 tickets medidos en prod jun→ago-2026, created_at idéntico al segundo.
    # El UPDATE condicional es atómico en PostgreSQL: exactamente UNO gana.
    # Va ANTES del bloque de pagos a propósito: más abajo el perdedor borraría
    # los TicketDetallePago del ganador (delete de ids_existentes) y con ellos
    # la TransaccionPOS por CASCADE — eso sí descuadra el arqueo de caja.
    ticket_se_pago = False
    if ticket.estado == 'PAGADO':
        ticket_se_pago = Ticket.objects.filter(
            id=ticket.id, estado='PENDIENTE'
        ).update(estado='PAGADO') == 1
        if not ticket_se_pago:
            _ticket_bd = Ticket.objects.get(id=ticket.id)
            logger.warning(
                "Cobro duplicado ignorado ticket=%s usuario=%s estado_real=%s",
                ticket.correlativo, request.user.username, _ticket_bd.estado,
            )
            if _ticket_bd.estado != 'PAGADO':
                # Cambió a ANULADO/DEVUELTO en el intertanto: no hay cobro válido.
                return JsonResponse({
                    'success': False,
                    'error': (f'El ticket #{correlativo} cambió de estado y no puede '
                              f'cobrarse (estado: {_ticket_bd.estado}).'),
                    'error_tipo': 'COBRO_NO_COMPLETADO',
                    'estado': _ticket_bd.estado,
                }, status=409)
            # El otro POST ya cobró este ticket: responder éxito SIN escribir
            # nada, con la misma forma que la respuesta normal, para que el POS
            # no muestre error ni reintente. Se adjunta el DTE si ya se emitió.
            respuesta_dup = {
                'success': True,
                'ticket': construir_ticket_data(_ticket_bd),
                'cobro_duplicado_ignorado': True,
            }
            _dte_dup = Dte.objects.filter(
                sucursal_id=ticket.sucursal_id,
                referencias__icontains=f'TICKET-{ticket.correlativo}',
            ).order_by('-id').first()
            if _dte_dup:
                respuesta_dup['dte_generado'] = {
                    'id': _dte_dup.id,
                    'numero': _dte_dup.numero_documento,
                    'tipo': _dte_dup.tipo_documento,
                }
            return JsonResponse(respuesta_dup)
    # ─────────────────────────────────────────────────────────────────────────

    # Procesar pagos (sin transaction.atomic anidado para evitar TransactionManagementError)
    for pago in pagos:
        pago_id = pago.get('id')
        try:
            monto = int(pago.get('monto', 0))
        except (TypeError, ValueError):
            continue
        if monto <= 0:
            continue

        metodo_pago = pago.get('metodo_pago', 'OTRO')
        if metodo_pago not in dict(METODO_PAGO_TICKET_CHOICES):
            metodo_pago = 'OTRO'

        # Origen del pago (MANUAL vs POS_INTEGRADO). El frontend lo envía
        # explícitamente según qué botón usó el cajero (F6/F7 manuales vs
        # POS TBK Automático con SDK). Si no viene y el método es de tarjeta
        # Transbank, asumimos MANUAL para que no quede ambiguo.
        METODOS_TBK_TARJETA = {
            'TBK_DEBITO_POS', 'TBK_CREDITO_POS', 'TBK_PREPAGO_POS',
            'TBK_POS_INTEGRADO', 'TBK_MANUAL',
            'TARJETA_DEBITO', 'TARJETA_CREDITO',
        }
        origen_pago_raw = (pago.get('origen_pago') or '').strip().upper()
        if origen_pago_raw not in dict(ORIGEN_PAGO_CHOICES):
            origen_pago_raw = 'MANUAL' if metodo_pago in METODOS_TBK_TARJETA else ''
        origen_pago_val = origen_pago_raw or None

        if pago_id and pago_id in ids_existentes:
            TicketDetallePago.objects.filter(id=pago_id, ticket=ticket).update(
                metodo_pago=metodo_pago,
                tipo_tarjeta=pago.get('tipo_tarjeta'),
                voucher=pago.get('voucher'),
                numero_orden_compra=pago.get('numero_orden_compra'),
                monto=monto,
                notas=pago.get('notas', ''),
                origen_pago=origen_pago_val,
            )
            ids_existentes.remove(pago_id)
        else:
            TicketDetallePago.objects.create(
                ticket=ticket,
                metodo_pago=metodo_pago,
                tipo_tarjeta=pago.get('tipo_tarjeta'),
                voucher=pago.get('voucher'),
                numero_orden_compra=pago.get('numero_orden_compra'),
                monto=monto,
                notas=pago.get('notas', ''),
                origen_pago=origen_pago_val,
            )

    if ids_existentes:
        TicketDetallePago.objects.filter(id__in=ids_existentes, ticket=ticket).delete()
    
    # `ticket_se_pago` viene del compare-and-set de más arriba (el read-then-
    # write que vivía aquí era la ventana de carrera del doble cobro). Este
    # save persiste totales/descuentos; el estado ya lo escribió el UPDATE.
    ticket.save()

    # (La pre-validación de stock del ticket completo se movió ANTES del bloque
    # de pagos y del candado anti doble cobro: así el 400 de stock insuficiente
    # sale sin haber escrito nada — ni pagos, ni estado, ni vale/gift card.)

    # ===== GIFT CARD: descontar saldo por cada pago con método GIFTCARD =====
    # Se procesa con su propia transacción + lock (el servicio lo maneja) y es
    # idempotente por pago, así que un reintento del POS no descuenta dos veces.
    # El código de la gift card viaja en el campo `voucher` del pago.
    if ticket_se_pago:
        from .services import giftcard_service
        for pago_gc in ticket.pagos.filter(metodo_pago='GIFTCARD'):
            codigo_gc = (pago_gc.voucher or '').strip()
            if not codigo_gc:
                continue
            try:
                giftcard_service.consumir(
                    codigo_gc, pago_gc.monto,
                    ticket=ticket, pago_ticket=pago_gc,
                    sucursal=ticket.sucursal, usuario=request.user,
                )
            except giftcard_service.GiftCardError:
                # No tumbar la venta (el cobro ya se confirmó); registrar para
                # revisión. La pre-validación AJAX debió evitar este caso.
                logger.exception(
                    "Error al consumir gift card en cobro ticket=%s codigo=%s",
                    ticket.correlativo, codigo_gc,
                )

    # ── Consumir cupón de descuento (después de confirmar el pago) ───────────
    if ticket_se_pago and _codigo_cupon:
        from .services import cupon_service as _cup_svc
        try:
            _cup_svc.canjear_cupon(
                _codigo_cupon,
                ticket=ticket,
                sucursal=ticket.sucursal,
                usuario=request.user,
                monto_descuento=_cupon_descuento,
            )
        except _cup_svc.CuponError:
            # Carrera extrema: la pre-validación debió impedirlo. No se tumba la
            # venta (ya se cobró); queda en el log para revisión manual. El
            # descuento SÍ se aplicó, así que el cupón quedaría reutilizable:
            # por eso el log es warning-con-datos y no un silencio.
            logger.exception(
                "Error al consumir cupón ticket=%s codigo=%s descuento=%s",
                ticket.correlativo, _codigo_cupon, _cupon_descuento,
            )
    # ─────────────────────────────────────────────────────────────────────────

    # ── Debitar vale de puntos (después de confirmar el pago) ────────────────
    if ticket_se_pago and _codigo_vale_pts:
        from .services import fidelizacion_service as _fid_svc
        try:
            _fid_svc.canjear_vale(
                _codigo_vale_pts,
                ticket=ticket,
                sucursal=ticket.sucursal,
                usuario=request.user,
            )
        except _fid_svc.FidelizacionError:
            # Race condition extrema: la pre-validación debió impedirlo.
            # No tumbar la venta; queda registrado para revisión manual.
            logger.exception(
                "Error al debitar vale de puntos ticket=%s codigo=%s",
                ticket.correlativo, _codigo_vale_pts,
            )
    # ─────────────────────────────────────────────────────────────────────────

    # Si el ticket se acaba de pagar, consumir stock FIFO y crear movimientos
    # ⚠️ IMPORTANTE: NO descontar stock si el ticket viene de CAMBIO_DEVOLUCION
    # porque el stock ya se ajustó al aprobar el cambio
    logger.debug(
        "Estado pago ticket=%s modulo_origen=%s ticket_se_pago=%s",
        ticket.correlativo,
        ticket.modulo_origen,
        ticket_se_pago,
    )
    
    if ticket_se_pago and ticket.modulo_origen != 'CAMBIO_DEVOLUCION':
        logger.debug("Iniciando descuento de stock ticket=%s", ticket.correlativo)

        # La pre-validación de stock del ticket completo ya corrió más arriba
        # (antes de pagos, candado y vale/gift card). Las verificaciones por
        # talla de este bucle se conservan como red de seguridad ante una
        # venta concurrente que se lleve el stock en el intertanto.

        # Si alguna talla ya consumió, el ticket NO puede volver a PENDIENTE:
        # reabrir el candado permitiría un re-cobro que descuenta de nuevo lo
        # ya consumido (la carrera original, pero con más daño).
        hubo_consumo = False

        for tp in ticket.ticket_productos.all():
            # Saltar ítems sin ProductoTalla (pendientes de despacho)
            if tp.ProductoTalla is None:
                logger.debug(
                    "Item manual sin stock ticket=%s descripcion=%s cantidad=%s",
                    ticket.correlativo,
                    tp.descripcion_linea,
                    tp.stock,
                )
                continue

            # Usar stock_sucursal para obtener el stock real de la sucursal
            stock_antes = tp.ProductoTalla.stock_sucursal(ticket.sucursal_id)
            logger.debug(
                "Descuento stock ticket=%s sku=%s stock_antes=%s cantidad=%s",
                ticket.correlativo,
                tp.ProductoTalla.sku,
                stock_antes,
                tp.stock,
            )
            
            # Verificar que hay stock disponible en la sucursal
            if stock_antes < tp.stock:
                error_msg = f'Stock insuficiente para SKU {tp.ProductoTalla.sku}. Disponible: {stock_antes}, Requerido: {tp.stock}'
                logger.warning(
                    "Stock insuficiente al pagar ticket=%s sku=%s disponible=%s requerido=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                    stock_antes,
                    tp.stock,
                )
                # Revertir a PENDIENTE SOLO si aún no se consumió nada: si una
                # talla anterior ya bajó stock, reabrir el candado permitiría
                # un re-cobro que la descuenta de nuevo. En ese caso el ticket
                # queda PAGADO a medias y marcado para revisión manual.
                if not hubo_consumo:
                    ticket.estado = 'PENDIENTE'
                    ticket.save(update_fields=['estado'])
                else:
                    logger.error(
                        "Consumo PARCIAL: ticket queda PAGADO para no reabrir el "
                        "candado ticket=%s sku_fallido=%s",
                        ticket.correlativo, tp.ProductoTalla.sku,
                    )
                return JsonResponse({
                    'success': False,
                    'error': error_msg,
                    'error_tipo': 'STOCK_INSUFICIENTE',
                    'sku': str(tp.ProductoTalla.sku),
                    'stock_disponible': stock_antes,
                    'stock_requerido': tp.stock,
                    'requiere_revision': hubo_consumo
                }, status=400)
            
            # Intentar consumir stock FIFO. Se marca ANTES de llamar: si la
            # llamada muere a medias no sabemos qué alcanzó a bajar, y el lado
            # seguro es NO reabrir el candado.
            hubo_consumo = True
            try:
                # Consumir stock FIFO (esto crea automáticamente el movimiento de EGRESO)
                # ✅ No pasar referencia_externa para que consumir_stock_fifo use DTE si está disponible
                costo_total_fifo, lotes_fifo = consumir_stock_fifo(
                    producto_talla=tp.ProductoTalla,
                    cantidad_requerida=tp.stock,
                    responsable=request.user.username,
                    ticket=ticket,
                    observaciones=f'Pago de ticket #{ticket.correlativo}',
                    referencia_externa=None  # Dejamos que consumir_stock_fifo determine la referencia correcta
                )

                # Trazabilidad: dejar registrado de qué lotes (y por tanto de qué
                # DTE de compra) salió esta línea. Sin esto el dato se pierde al
                # terminar el cobro y el margen real queda en 0.
                persistir_costeo_fifo(tp, costo_total_fifo, lotes_fifo)

                # Recargar para ver el stock actualizado
                tp.ProductoTalla.refresh_from_db()
                stock_despues = tp.ProductoTalla.stock
                logger.info(
                    "Stock consumido FIFO ticket=%s sku=%s stock_antes=%s stock_despues=%s diferencia=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                    stock_antes,
                    stock_despues,
                    stock_antes - stock_despues,
                )
                
            except Exception as e:
                logger.exception(
                    "Error FIFO ticket=%s sku=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                )
                
                # Recargar stock para verificar si FIFO ya lo descontó
                tp.ProductoTalla.refresh_from_db()
                stock_despues_error = tp.ProductoTalla.stock
                logger.debug(
                    "Stock despues de error FIFO ticket=%s sku=%s stock=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                    stock_despues_error,
                )
                
                # ⚠️ CRÍTICO: Solo descontar manualmente si FIFO NO descontó
                if stock_despues_error == stock_antes:
                    logger.warning(
                        "FIFO no desconto; procediendo con descuento manual ticket=%s sku=%s",
                        ticket.correlativo,
                        tp.ProductoTalla.sku,
                    )
                    # Antes de bajar el stock plano, consumir los lotes FIFO que SÍ
                    # existan (parcial). consumir_stock_fifo es todo-o-nada: si no
                    # hay lotes para cubrir TODA la cantidad no baja ninguno, y el
                    # descuento manual de stock dejaba los lotes intactos →
                    # desincronización (saldo_lotes > stock). Consumir lo que haya
                    # mantiene los lotes alineados; el resto cae al stock plano.
                    # (Si no hay lotes, devuelve consumido=0 y se comporta como antes.)
                    # _consumir_lotes_fifo_ajuste usa select_for_update(), que exige
                    # transacción: esta vista corre sin atomic (ver nota en la línea
                    # del comentario "sin transaction.atomic anidado"), así que la
                    # abrimos aquí, acotada al descuento de lotes de esta talla.
                    from .views_edicion_productos import _consumir_lotes_fifo_ajuste
                    try:
                        with transaction.atomic():
                            _lotes_bajados, _consumido_lotes = _consumir_lotes_fifo_ajuste(
                                tp.ProductoTalla, tp.stock
                            )
                        if _consumido_lotes:
                            logger.info(
                                "Lotes FIFO bajados parcialmente en fallback ticket=%s sku=%s consumido_lotes=%s/%s",
                                ticket.correlativo,
                                tp.ProductoTalla.sku,
                                _consumido_lotes,
                                tp.stock,
                            )
                    except Exception:
                        # No bloquear la venta si el ajuste de lotes falla; el
                        # descuento de stock plano (abajo) sigue siendo la verdad.
                        logger.exception(
                            "No se pudieron bajar lotes FIFO parciales en fallback ticket=%s sku=%s",
                            ticket.correlativo,
                            tp.ProductoTalla.sku,
                        )
                    # Crear movimiento manual si falla FIFO
                    # ✅ Usar DTE si está disponible, si no usar correlativo del ticket
                    referencia = f'DTE_{ticket.folio_dte}' if ticket.folio_dte else f'TICKET_{ticket.correlativo}'
                    Movimientos_Producto.objects.create(
                        ticket=ticket,
                        ProductoTalla=tp.ProductoTalla,
                        sucursal_origen=ticket.sucursal,
                        cantidad=-tp.stock,  # Negativo para egreso
                        costo=tp.ProductoTalla.producto.costo,
                        precio=tp.precio,
                        sobreprecio=tp.ProductoTalla.producto.sobreprecio if hasattr(tp.ProductoTalla.producto, 'sobreprecio') else 0,
                        concepto='VENTA_DIRECTA',
                        tipo_movimiento='EGRESO',
                        responsable=request.user.username,
                        observaciones=f'Venta ticket #{ticket.correlativo} - Consumo manual (FIFO no disponible)',
                        referencia_externa=referencia,
                        fecha=timezone.localdate(),
                        hora=timezone.localtime().time()
                    )
                    # Actualizar stock manualmente
                    tp.ProductoTalla.stock -= tp.stock
                    tp.ProductoTalla.save()
                    logger.warning(
                        "Stock consumido manualmente ticket=%s sku=%s cantidad=%s stock_antes=%s stock_despues=%s",
                        ticket.correlativo,
                        tp.ProductoTalla.sku,
                        tp.stock,
                        stock_antes,
                        tp.ProductoTalla.stock,
                    )
                else:
                    logger.warning(
                        "FIFO ya desconto parcialmente ticket=%s sku=%s stock_antes=%s stock_despues=%s",
                        ticket.correlativo,
                        tp.ProductoTalla.sku,
                        stock_antes,
                        stock_despues_error,
                    )
                    # No descontar de nuevo, solo crear movimiento de registro
                    # ✅ Usar DTE si está disponible, si no usar correlativo del ticket
                    referencia = f'DTE_{ticket.folio_dte}' if ticket.folio_dte else f'TICKET_{ticket.correlativo}'
                    Movimientos_Producto.objects.create(
                        ticket=ticket,
                        ProductoTalla=tp.ProductoTalla,
                        sucursal_origen=ticket.sucursal,
                        cantidad=-tp.stock,
                        costo=tp.ProductoTalla.producto.costo,
                        precio=tp.precio,
                        sobreprecio=tp.ProductoTalla.producto.sobreprecio if hasattr(tp.ProductoTalla.producto, 'sobreprecio') else 0,
                        concepto='VENTA_DIRECTA',
                        tipo_movimiento='EGRESO',
                        responsable=request.user.username,
                        observaciones=f'Venta ticket #{ticket.correlativo} - Movimiento de registro (FIFO parcial)',
                        referencia_externa=referencia,
                        fecha=timezone.localdate(),
                        hora=timezone.localtime().time()
                    )
                    logger.info(
                        "Movimiento de registro creado sin descuento adicional ticket=%s sku=%s",
                        ticket.correlativo,
                        tp.ProductoTalla.sku,
                    )
    elif ticket_se_pago and ticket.modulo_origen == 'CAMBIO_DEVOLUCION':
        logger.info(
            "Ticket cambio/devolucion pagado sin descuento adicional ticket=%s",
            ticket.correlativo,
        )
        
        # Auto-completar el CambioDevolucion asociado
        cambio_asociado = CambioDevolucion.objects.filter(
            ticket_nuevo=ticket,
            estado__in=['EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE']
        ).first()
        
        if not cambio_asociado:
            cambio_asociado = CambioDevolucion.objects.filter(
                ticket_diferencia=ticket,
                estado__in=['EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE']
            ).first()
        
        if cambio_asociado:
            estado_anterior_cd = cambio_asociado.estado
            cambio_asociado.estado = 'COMPLETADO'
            cambio_asociado.fecha_completado = timezone.now()
            if estado_anterior_cd == 'EJECUTADO_COBRO_PENDIENTE':
                cambio_asociado.fecha_pago_diferencia = timezone.now()
            cambio_asociado.save()
            
            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio_asociado,
                accion='COMPLETADO_AUTO',
                estado_anterior=estado_anterior_cd,
                estado_nuevo='COMPLETADO',
                usuario=request.user,
                descripcion=f'Cambio completado automáticamente al procesar pago del ticket #{ticket.correlativo}.',
                datos_adicionales={
                    'ticket_pagado': ticket.correlativo,
                    'monto_pagado': float(ticket.total),
                }
            )
            logger.info(
                "CambioDevolucion completado automaticamente numero=%s estado_anterior=%s",
                cambio_asociado.numero_operacion,
                estado_anterior_cd,
            )
    
    # Guardar o actualizar cliente en la base de datos si tiene datos
    if datos_cliente and datos_cliente.get('rut') and datos_cliente.get('nombre'):
        guardar_o_actualizar_cliente(datos_cliente, request.user)
    
    # Generar DTE si el tipo de documento lo requiere
    tipo_documento_seleccionado = payload.get('tipo_documento', '')
    dte_generado = None
    
    if tipo_documento_seleccionado in ['BOLETA_ELECTRONICA', 'BOLETA_PAPEL', 'FACTURA_ELECTRONICA'] and ticket.estado == 'PAGADO':
        try:
            logger.debug("Generando DTE para ticket=%s", ticket.correlativo)
            
            # ✅ CRÍTICO: Refrescar el ticket desde la BD para tener los pagos actualizados
            ticket.refresh_from_db()
            
            # Verificar pagos ANTES de generar DTE
            pagos_count = ticket.pagos.count()
            logger.debug(
                "Ticket pagos registrados ticket=%s pagos_count=%s",
                ticket.correlativo,
                pagos_count,
            )
            for pago in ticket.pagos.all():
                logger.debug(
                    "Pago ticket=%s metodo=%s monto=%s",
                    ticket.correlativo,
                    pago.metodo_pago,
                    pago.monto,
                )
            
            # Verificar stock ANTES de generar DTE
            for tp in ticket.ticket_productos.all():
                if tp.ProductoTalla is None:
                    continue
                logger.debug(
                    "Pre-DTE stock ticket=%s sku=%s stock=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                    tp.ProductoTalla.stock,
                )
            
            # ✅ Pasar cotizacion_obj para usar sus descripciones en el TXT
            dte_generado = generar_dte_desde_ticket(ticket, tipo_documento_seleccionado, request.user, cotizacion=cotizacion_obj)
            logger.info(
                "DTE generado desde ticket=%s tipo=%s numero=%s",
                ticket.correlativo,
                dte_generado.tipo_documento,
                dte_generado.numero_documento,
            )

            # Si un intento anterior había quedado marcado como fallido
            # (ver except más abajo / reintentar_generar_dte_ticket), limpiar
            # la bandera ahora que el DTE se generó con éxito.
            if ticket.dte_generacion_fallida:
                ticket.dte_generacion_fallida = False
                ticket.dte_error_detalle = None
                ticket.save(update_fields=['dte_generacion_fallida', 'dte_error_detalle'])

            # Verificar stock DESPUÉS de generar DTE
            for tp in ticket.ticket_productos.all():
                if tp.ProductoTalla is None:
                    continue
                tp.ProductoTalla.refresh_from_db()
                logger.debug(
                    "Post-DTE stock ticket=%s sku=%s stock=%s",
                    ticket.correlativo,
                    tp.ProductoTalla.sku,
                    tp.ProductoTalla.stock,
                )
            
            # =========================================================================
            # NUEVO: Marcar cotización como facturada si viene de una cotización
            # =========================================================================
            if cotizacion_obj and dte_generado:
                try:
                    from .models import Historial_Cotizacion
                    
                    # Usar solo el número de documento (el campo numero_factura tiene max_length=20)
                    numero_documento_corto = str(dte_generado.numero_documento)[:20]
                    numero_documento_completo = f"{dte_generado.tipo_documento} #{dte_generado.numero_documento}"

                    # ⚠️ Pasar el flag real de pendientes. Sin él,
                    # marcar_como_facturada() usa su default (tiene_pendientes=False)
                    # y deja estado_despacho=COMPLETADO aunque haya ítems sin SKU:
                    # el flujo de "Despacho Diferido" nunca se ofrecía y ese stock
                    # jamás se descontaba.
                    tiene_pendientes_cot = cotizacion_obj.items.filter(
                        es_producto_pendiente=True,
                        sku_asignado_post_factura=False,
                    ).exists()
                    cotizacion_obj.marcar_como_facturada(
                        numero_documento_corto,
                        tiene_pendientes=tiene_pendientes_cot,
                        dte=dte_generado,
                    )

                    # Registrar en historial
                    Historial_Cotizacion.objects.create(
                        cotizacion=cotizacion_obj,
                        usuario=request.user,
                        accion='FACTURADA',
                        descripcion=(
                            f'Cotización facturada desde POS. Documento: {numero_documento_completo}. '
                            f'Ticket: #{ticket.correlativo}'
                            + (' Quedan ítems con despacho diferido pendiente.'
                               if tiene_pendientes_cot else '')
                        ),
                        ip_address=request.META.get('REMOTE_ADDR', '')
                    )
                    logger.info(
                        "Cotizacion marcada como facturada numero=%s factura=%s",
                        cotizacion_obj.numero_cotizacion,
                        ticket.folio_dte,
                    )
                except Exception as cot_error:
                    logger.exception(
                        "Error al marcar cotizacion como facturada numero=%s",
                        cotizacion_obj.numero_cotizacion,
                    )
                
        except Exception as e:
            # No fallar el pago si hay error en DTE (la venta ya está cobrada
            # y el stock ya se descontó) — pero dejar rastro persistente para
            # que no quede un ticket PAGADO sin documento tributario en
            # silencio. El reintento manual es responsabilidad de
            # `reintentar_generar_dte_ticket`.
            logger.exception("Error al generar DTE desde ticket=%s", ticket.correlativo)
            ticket.dte_generacion_fallida = True
            ticket.dte_error_detalle = f"{type(e).__name__}: {e}"[:2000]
            ticket.save(update_fields=['dte_generacion_fallida', 'dte_error_detalle'])

    response_data = {
        'success': True,
        'ticket': construir_ticket_data(ticket)
    }

    if ticket.dte_generacion_fallida:
        response_data['dte_generacion_fallo'] = True
        response_data['dte_error_mensaje'] = (
            'La venta se cobró correctamente pero no se pudo generar el '
            'documento tributario. Reintente desde Consulta de Documentos.'
        )

    if dte_generado:
        response_data['dte_generado'] = {
            'id': dte_generado.id,
            'numero': dte_generado.numero_documento,
            'tipo': dte_generado.tipo_documento
        }

        # Incluir datos del archivo TXT si se generó
        if hasattr(dte_generado, 'archivo_txt_data') and dte_generado.archivo_txt_data:
            response_data['archivo_txt'] = dte_generado.archivo_txt_data

        # El DTE se emite aunque el TXT de Acepta falle (son cosas separadas),
        # pero el cajero tiene que enterarse: antes el motivo quedaba solo en
        # `dte._txt_error` y en el log, así que la venta se veía 100% OK y
        # nadie sabía que no había archivo para subir a Acepta.
        _txt_error = getattr(dte_generado, '_txt_error', None)
        if _txt_error:
            response_data['archivo_txt_error'] = _txt_error
            response_data['archivo_txt_error_mensaje'] = (
                f'El documento #{dte_generado.numero_documento} se emitió, pero no '
                'se pudo generar el archivo TXT para Acepta. Reintente desde '
                'Consulta de Documentos.'
            )
    
    # Incluir info de cotización si fue facturada desde una cotización
    if cotizacion_obj:
        response_data['cotizacion_facturada'] = {
            'id': cotizacion_obj.id,
            'numero_cotizacion': cotizacion_obj.numero_cotizacion,
            'numero_factura': cotizacion_obj.numero_factura
        }

    # ===== FIDELIZACIÓN: acumular puntos al cliente identificado por RUT =====
    # Solo si el ticket quedó PAGADO y la venta es a cliente particular: se
    # excluyen facturas y ventas originadas en cotización (ver venta_fideliza).
    # Venta anónima (sin cliente en CRM) tampoco acumula. Idempotente por
    # ticket. No debe tumbar la respuesta del cobro.
    if ticket.estado == 'PAGADO':
        try:
            from .services import fidelizacion_service
            resultado_pts = fidelizacion_service.acumular_puntos_por_venta(
                ticket,
                usuario=request.user,
                tipo_documento=tipo_documento_seleccionado,
                cotizacion=cotizacion_obj,
            )
            if resultado_pts:
                response_data['fidelizacion'] = resultado_pts
        except Exception:
            logger.exception(
                "Error al acumular puntos de fidelización ticket=%s",
                ticket.correlativo,
            )

    logger.debug("Fin registrar_pagos_ticket ticket=%s", correlativo)
    return JsonResponse(response_data)


@login_required
@require_http_methods(["POST"])
def reintentar_generar_dte_ticket(request, ticket_id):
    """Reintenta generar el DTE de un ticket que quedó PAGADO sin documento
    (ver `dte_generacion_fallida` en el modelo Ticket).

    No reusa `generar_dte_desde_ticket_api` (views_modulo_documentos.py) porque
    esa función solo genera el TXT descargable, consume un folio nuevo cada vez
    que se llama y no crea el registro `Dte` en BD — reutilizarla aquí
    duplicaría folios sin documento real. Este endpoint llama directamente a
    la función autoritativa `generar_dte_desde_ticket` de este mismo archivo,
    que recalcula el monto desde las líneas del ticket.
    """
    sucursal_id = (
        request.session.get('idSucursalActual')
        or request.session.get('sucursalActual')
        or request.session.get('idSucursalActualPOS')
    )
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión'}, status=400)

    ticket = (
        Ticket.objects
        .select_related('sucursal')
        .filter(id=ticket_id, sucursal_id=sucursal_id)
        .first()
    )
    if not ticket:
        return JsonResponse({'success': False, 'error': 'Ticket no encontrado'}, status=404)

    if ticket.estado != 'PAGADO':
        return JsonResponse({
            'success': False,
            'error': f'El ticket #{ticket.correlativo} no está en estado PAGADO (estado actual: {ticket.estado}).',
        }, status=400)

    if ticket.folio_dte:
        return JsonResponse({
            'success': False,
            'error': f'El ticket #{ticket.correlativo} ya tiene un DTE generado (folio {ticket.folio_dte}).',
        }, status=400)

    tipo_documento = ticket.tipo_dte
    if tipo_documento not in ('BOLETA_ELECTRONICA', 'BOLETA_PAPEL', 'FACTURA_ELECTRONICA'):
        return JsonResponse({
            'success': False,
            'error': f'Tipo de documento del ticket ("{tipo_documento}") no es facturable electrónicamente.',
        }, status=400)

    try:
        dte_generado = generar_dte_desde_ticket(ticket, tipo_documento, request.user)
        ticket.dte_generacion_fallida = False
        ticket.dte_error_detalle = None
        ticket.save(update_fields=['dte_generacion_fallida', 'dte_error_detalle'])
        logger.info(
            "DTE generado en reintento manual ticket=%s tipo=%s numero=%s",
            ticket.correlativo,
            dte_generado.tipo_documento,
            dte_generado.numero_documento,
        )
        return JsonResponse({
            'success': True,
            'dte_generado': {
                'id': dte_generado.id,
                'numero': dte_generado.numero_documento,
                'tipo': dte_generado.tipo_documento,
            },
        })
    except Exception as e:
        logger.exception("Error en reintento manual de DTE ticket=%s", ticket.correlativo)
        ticket.dte_generacion_fallida = True
        ticket.dte_error_detalle = f"{type(e).__name__}: {e}"[:2000]
        ticket.save(update_fields=['dte_generacion_fallida', 'dte_error_detalle'])
        return JsonResponse({'success': False, 'error': f'Error al generar DTE: {e}'}, status=500)


@login_required
@require_GET
def listar_tickets_dte_fallido(request):
    """Lista los tickets PAGADOS de la sucursal actual cuyo DTE no se pudo
    generar (`dte_generacion_fallida=True`), para el botón de reintento en
    Consulta de Documentos."""
    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal seleccionada'})

    tickets = (
        Ticket.objects
        .filter(sucursal_id=sucursal_id, dte_generacion_fallida=True)
        .order_by('-fecha', '-hora')
        .only('id', 'correlativo', 'fecha', 'hora', 'total', 'cliente_nombre', 'dte_error_detalle')
    )

    return JsonResponse({
        'success': True,
        'tickets': [
            {
                'id': t.id,
                'correlativo': t.correlativo,
                'fecha': t.fecha.strftime('%Y-%m-%d') if t.fecha else '',
                'hora': t.hora.strftime('%H:%M') if t.hora else '',
                'total': t.total,
                'cliente_nombre': t.cliente_nombre or '',
                'error_detalle': t.dte_error_detalle or '',
            }
            for t in tickets
        ],
    })


@login_required
@require_GET
def ticket_pago_pos(request):
    """Vista para página de pagos de tickets POS"""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    if sucursal_actual_id:
        sucursal_actual = Sucursal.objects.filter(id=sucursal_actual_id).first()

    context = {
        'metodo_pago_choices': METODO_PAGO_TICKET_CHOICES,
        'estado_ticket_choices': ESTADO_TICKET_CHOICES,
        'sucursal_actual': sucursal_actual,
    }
    return render(request, 'vistas/modulo_ventas/ticket_pago_pos.html', context)


# ========== GESTIÓN DE DOCUMENTOS DE VENTAS ==========

# Máximo de documentos que `listar_documentos_ventas` devuelve en una página.
# Acota el costo del endpoint sin importar lo que pida el cliente.
PER_PAGE_MAX_DOCUMENTOS_VENTAS = 1000

@login_required
def gestion_ventas_documentos(request):
    """Vista principal para gestión de ventas y documentos"""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None

    user_rol = getattr(request.user, 'rol', '') or ''
    es_admin = user_rol == 'administrador'

    # Permisos granulares de edición de DTE (3 campos x 4 tipos).
    # El template los usa para mostrar/habilitar cada control del modal.
    permisos_dte = permisos_edicion_dte_context(request.user, sucursal_actual_id)

    # Pares de tipos intercambiables habilitados para el usuario
    # (ej.: {"BOLETA ELECTRONICA": ["BOLETA ELECTRONICA", "BOLETA PAPEL"], ...}).
    # Sólo se incluyen entradas donde el usuario tiene permiso para ambos
    # extremos del grupo, para que el frontend pueda ofrecer el cambio.
    compatibles_por_tipo = permisos_dte.get('compatibles_por_tipo', {})

    # Vendedores para el selector del modal "DTE manual".
    # Se devuelven TODOS los vendedores activos sin filtrar por sucursal,
    # porque puede existir el caso en el que un vendedor opera en una
    # sucursal a la que no está asignado vía M2M (datos migrados,
    # cobertura ad-hoc, etc.) y el operador igual necesita poder
    # seleccionarlo. Se incluye la lista de alias de sucursales
    # asignadas para que el buscador del Select2 pueda matchear por
    # sucursal además del nombre/código del vendedor.
    vendedores_qs = (
        Vendedor.objects
        .filter(activo=True)
        .prefetch_related('sucursales')
        .order_by('nombre')
    )
    vendedores_sucursal = []
    for v in vendedores_qs:
        sucursales_alias = sorted(
            (s.alias or s.direccion or '') for s in v.sucursales.all()
            if (s.alias or s.direccion)
        )
        vendedores_sucursal.append({
            'id': v.id,
            'nombre': v.nombre or '',
            'codigo_vendedor': v.codigo_vendedor or '',
            'sucursales_str': ', '.join(sucursales_alias),
        })

    # Hoy en zona horaria Chile: se usa como `value` por defecto del
    # input de fecha en el modal de DTE manual (regla timezone-chile).
    fecha_hoy_str = timezone.localdate().strftime('%Y-%m-%d')

    # Tickets cobrados cuyo DTE no se pudo generar (ver `dte_generacion_fallida`
    # en el modelo Ticket) — requieren reintento manual desde este módulo.
    tickets_dte_fallido_count = 0
    if sucursal_actual_id:
        tickets_dte_fallido_count = Ticket.objects.filter(
            sucursal_id=sucursal_actual_id,
            dte_generacion_fallida=True,
        ).count()

    context = {
        'sucursal_actual': sucursal_actual,
        'metodo_pago_choices': METODO_PAGO_TICKET_CHOICES,
        'estado_ticket_choices': ESTADO_TICKET_CHOICES,
        'tipo_documento_choices': TIPO_DOCUMENTO_CHOICES,
        'qz_config': _get_qz_config(sucursal_actual_id),
        'user_rol': user_rol,
        'es_admin': es_admin,
        # Flags por campo
        'puede_editar_fecha_dte': permisos_dte['campo']['fecha'],
        'puede_editar_numero_dte': permisos_dte['campo']['numero_documento'],
        'puede_editar_pago_dte': permisos_dte['campo']['pago'],
        # Permiso para cambiar el vendedor asignado al DTE.
        # Bypass: el rol `administrador` siempre puede editar vendedor,
        # consistente con el resto de operaciones admin-only del módulo
        # (`crear_dte_manual`, `eliminar_documento_venta`). Esto evita
        # que el usuario quede bloqueado cuando la migración
        # `0151_permiso_dte_editar_vendedor` aún no se ha aplicado en
        # el servidor; otros roles siguen necesitando el permiso
        # granular `dte_editar_vendedor.puede_editar`.
        'puede_editar_vendedor_dte': (
            es_admin or permisos_dte['campo'].get('vendedor', False)
        ),
        # Flags por tipo de DTE (nombre amigable: sin espacios para usar en template)
        'puede_editar_tipo_boleta_electronica': permisos_dte['tipo']['BOLETA ELECTRONICA'],
        'puede_editar_tipo_boleta_papel': permisos_dte['tipo']['BOLETA PAPEL'],
        'puede_editar_tipo_factura_electronica': permisos_dte['tipo']['FACTURA ELECTRONICA'],
        'puede_editar_tipo_factura_exenta': permisos_dte['tipo']['FACTURA EXENTA'],
        # ¿Puede editar algo en algún tipo? → controla visibilidad del modal.
        # Admin siempre lo ve (consistente con el bypass aplicado al
        # resto de los flags por campo y por tipo). Los demás roles
        # se rigen por `permisos_dte['cualquiera']` (al menos un par
        # campo+tipo habilitado).
        'puede_editar_algun_dte': es_admin or permisos_dte['cualquiera'],
        # Mapa serializado {tipo_origen: [tipos_destino]} para el JS del modal
        # (usado para mostrar el selector "Tipo de Documento" en cambios
        # compatibles, p. ej. BOLETA ELECTRONICA ↔ BOLETA PAPEL).
        'tipos_dte_compatibles_json': json.dumps(compatibles_por_tipo),
        'puede_cambiar_tipo_dte_flag': bool(compatibles_por_tipo),
        # Creación de DTE manual: solo administradores. La pantalla
        # esconde el botón cuando el flag es False; el endpoint también
        # valida el rol del lado servidor.
        'puede_crear_dte_manual': es_admin,
        'vendedores_sucursal': vendedores_sucursal,
        'fecha_hoy_str': fecha_hoy_str,
        'tickets_dte_fallido_count': tickets_dte_fallido_count,
    }
    return render(request, 'vistas/modulo_ventas/gestionVentasDocumentos.html', context)


@login_required
@require_GET
def listar_documentos_ventas(request):
    """API para listar documentos de ventas (tickets, boletas, facturas).

    Optimizaciones clave:
    - Paginación y ordenamiento se realizan en la DB (Paginator + order_by).
    - Estadísticas globales se obtienen en un solo `aggregate` condicional.
    - Prefetch con `only()` para reducir columnas traídas.
    - Totales por DTE (total_pagos / subtotal_bruto) calculados via annotate con Subquery.
    """
    try:
        sucursal_id = get_sucursal_id(request)
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_documento = request.GET.get('tipo_documento')
        estado = request.GET.get('estado')
        metodo_pago = request.GET.get('metodo_pago')
        buscar = request.GET.get('buscar', '').strip()
        page = int(request.GET.get('page', 1))
        # Tope de `per_page`: sin él, "Copiar Tabla" pedía 99999 y el endpoint
        # armaba en memoria todos los DTE del rango con sus líneas y pagos.
        # Se devuelve `per_page_solicitado` para que el frontend pueda avisar
        # cuando el tope recorta — una tabla truncada en silencio se lee como
        # completa, que es peor que un error visible.
        per_page_solicitado = int(request.GET.get('per_page', 20))
        per_page = max(1, min(per_page_solicitado, PER_PAGE_MAX_DOCUMENTOS_VENTAS))
        monto_min_raw = request.GET.get('monto_min', '').strip()
        monto_max_raw = request.GET.get('monto_max', '').strip()
        monto_min = int(monto_min_raw) if monto_min_raw.isdigit() else None
        monto_max = int(monto_max_raw) if monto_max_raw.isdigit() else None

        # === SOLO DTEs (Facturas/Boletas Electrónicas) ===
        # Se excluyen los descartados: la eliminación lógica de un DTE
        # (gestionVentasDocumentos) los marca como `descartado=True` y
        # debe sacarlos del listado y de las KPIs derivadas.
        dtes_filtrados = (
            Dte.objects
            .filter(
                sucursal_id=sucursal_id,
                tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
                descartado=False,
            )
            .select_related('vendedor', 'receptor')
        )

        # Aplicar filtros de fecha
        if fecha_desde:
            dtes_filtrados = dtes_filtrados.filter(fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            dtes_filtrados = dtes_filtrados.filter(fecha_emision__lte=fecha_hasta)

        # Aplicar filtros por tipo de DTE
        if tipo_documento:
            tipo_db_map = {
                'BOLETA_ELECTRONICA': 'BOLETA ELECTRONICA',
                'BOLETA_PAPEL': 'BOLETA PAPEL',
                'FACTURA_ELECTRONICA': 'FACTURA ELECTRONICA',
                'FACTURA_EXENTA': 'FACTURA EXENTA',
            }
            tipo_db = tipo_db_map.get(tipo_documento)
            if tipo_db:
                dtes_filtrados = dtes_filtrados.filter(tipo_documento=tipo_db)

        # Filtrar por estado (independiente del tipo de documento)
        if estado:
            estado_dte_map = {
                'PENDIENTE': 'PENDIENTE',
                'PAGADO': 'EMITIDO',
                'ANULADO': 'ANULADO',
            }
            if estado in estado_dte_map:
                dtes_filtrados = dtes_filtrados.filter(estado_dte=estado_dte_map[estado])

        # Filtrar por método de pago (independiente del tipo de documento).
        # MULTIPLE es un valor sintético: DTEs con más de un método de pago distinto.
        # Algunas opciones agrupan variantes históricas y POS Transbank, porque el
        # mismo tipo "lógico" se almacena con distintos códigos según el origen
        # (migración Laravel, POS integrado, ingreso manual, etc.).
        if metodo_pago:
            if metodo_pago == 'MULTIPLE':
                dtes_filtrados = dtes_filtrados.annotate(
                    _n_metodos=Count('dte_asociado__metodo_pago', distinct=True)
                ).filter(_n_metodos__gt=1)
            else:
                metodo_pago_grupos = {
                    # "Tarjeta Débito" agrupa genérico (histórico) + POS Transbank
                    'TARJETA_DEBITO': ['TARJETA_DEBITO', 'TBK_DEBITO_POS'],
                    # "Tarjeta Crédito" agrupa genérico + POS Transbank + manual
                    'TARJETA_CREDITO': [
                        'TARJETA_CREDITO',
                        'TBK_CREDITO_POS',
                        'TBK_MANUAL',
                    ],
                    # "Transbank POS" cubre cualquier pago procesado por el SDK
                    'TBK_POS_INTEGRADO': [
                        'TBK_POS_INTEGRADO',
                        'TBK_DEBITO_POS',
                        'TBK_CREDITO_POS',
                        'TBK_PREPAGO_POS',
                    ],
                }
                valores = metodo_pago_grupos.get(metodo_pago, [metodo_pago])
                dtes_filtrados = dtes_filtrados.filter(
                    dte_asociado__metodo_pago__in=valores
                ).distinct()

        if buscar:
            dtes_filtrados = dtes_filtrados.filter(
                Q(numero_documento__icontains=buscar) |
                Q(receptor__nombre__icontains=buscar) |
                Q(receptor__rut__icontains=buscar) |
                Q(vendedor__nombre__icontains=buscar) |
                Q(dte_productos__productoTalla__sku__icontains=buscar) |
                Q(dte_productos__productoTalla__producto__articulo__icontains=buscar)
            ).distinct()

        if monto_min is not None:
            dtes_filtrados = dtes_filtrados.filter(monto_con_iva__gte=monto_min)
        if monto_max is not None:
            dtes_filtrados = dtes_filtrados.filter(monto_con_iva__lte=monto_max)

        # --- Orden en DB ---
        orden_campo = request.GET.get('orden_campo', 'fecha')
        orden_direccion = request.GET.get('orden_direccion', 'desc')
        orden_db_map = {
            'fecha': 'fecha_emision',
            'tipo_documento': 'tipo_documento',
            'numero_documento': 'numero_documento',
            'cliente_nombre': 'receptor__nombre',
            'vendedor_nombre': 'vendedor__nombre',
            'total': 'monto_con_iva',
            'estado': 'estado_dte',
        }
        campo_db = orden_db_map.get(orden_campo, 'fecha_emision')
        prefix = '-' if orden_direccion == 'desc' else ''
        dtes_filtrados = dtes_filtrados.order_by(f'{prefix}{campo_db}', '-id')

        # --- Estadísticas globales en UN solo aggregate ---
        # `total_ventas` debe coincidir con la columna "total" que se muestra
        # por fila: cuando hay pagos registrados, usamos la suma de pagos
        # (ya refleja descuentos aplicados en el cobro); si no, caemos a
        # `monto_con_iva`. Sumar directamente `monto_con_iva` daba un monto
        # mayor cuando el descuento quedaba sólo en el detalle de pago y no
        # se reflejaba en la cabecera del DTE.
        pagos_sum_sub = (
            Dte_Detalle_Pago.objects
            .filter(dte_id=OuterRef('pk'))
            .values('dte_id')
            .annotate(t=Sum('monto'))
            .values('t')
        )
        dtes_con_total = dtes_filtrados.annotate(
            _total_pagos=Coalesce(
                Subquery(pagos_sum_sub, output_field=DecimalField(max_digits=14, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)),
            ),
        ).annotate(
            _total_real=Case(
                When(_total_pagos__gt=0, then=F('_total_pagos')),
                default=F('monto_con_iva'),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
        )

        stats = dtes_con_total.aggregate(
            total_documentos=Count('id', distinct=True),
            total_ventas=Sum('_total_real'),
            total_pendientes=Count('id', filter=Q(estado_dte='PENDIENTE'), distinct=True),
            total_facturas=Count(
                'id',
                filter=Q(tipo_documento__in=['FACTURA ELECTRONICA', 'FACTURA EXENTA']),
                distinct=True,
            ),
            total_boletas=Count(
                'id',
                filter=Q(tipo_documento__in=['BOLETA ELECTRONICA', 'BOLETA PAPEL']),
                distinct=True,
            ),
            total_boletas_electronicas=Count(
                'id', filter=Q(tipo_documento='BOLETA ELECTRONICA'), distinct=True,
            ),
            total_boletas_papel=Count(
                'id', filter=Q(tipo_documento='BOLETA PAPEL'), distinct=True,
            ),
        )

        total_documentos = stats['total_documentos'] or 0

        # --- Paginación en DB, prefetch solo de la página visible ---
        paginator = Paginator(dtes_filtrados, per_page)
        page_obj = paginator.get_page(page)

        pk_pagina = [dte.pk for dte in page_obj.object_list]

        # Traemos sólo los DTE paginados con sus productos y pagos via Prefetch
        pagos_qs = Dte_Detalle_Pago.objects.only(*ONLY_DTE_PAGO)
        productos_qs = (
            Dte_Productos.objects
            .select_related('productoTalla__producto')
            .only(*ONLY_DTE_PRODUCTO)
        )
        dtes_pagina = (
            Dte.objects
            .filter(pk__in=pk_pagina)
            .select_related('vendedor', 'receptor')
            .prefetch_related(
                Prefetch('dte_asociado', queryset=pagos_qs),
                Prefetch('dte_productos', queryset=productos_qs),
            )
            .order_by(f'{prefix}{campo_db}', '-id')
        )

        from datetime import time as dt_time
        documentos_paginados = []

        # Batch-lookup: tickets vinculados a los DTEs de esta página
        # para poder mostrar el correlativo del ticket en el modal.
        ticket_map_by_folio = {}
        folios_pagina = [dte.numero_documento for dte in dtes_pagina if dte.numero_documento]
        if folios_pagina:
            tickets_vinculados = (
                Ticket.objects
                .filter(
                    sucursal_id=sucursal_id,
                    folio_dte__in=folios_pagina,
                    estado='PAGADO',
                )
                .only('correlativo', 'folio_dte', 'tipo_dte', 'hora', 'fecha')
            )
            for tk in tickets_vinculados:
                key = tk.folio_dte
                if key not in ticket_map_by_folio:
                    ticket_map_by_folio[key] = tk

        # Batch-lookup: DTEs de la página con nota de crédito vigente que los
        # referencia. El modal de edición necesita avisarlo antes de dejar
        # tocar el folio (el backend además lo bloquea).
        dtes_con_nc = set(
            Dte.objects
            .filter(documento_afectado_id__in=pk_pagina, es_nota_credito=True)
            .exclude(estado_dte='ANULADO')
            .values_list('documento_afectado_id', flat=True)
        ) if pk_pagina else set()

        for dte in dtes_pagina:
            productos = []
            subtotal_bruto = 0
            for dp in dte.dte_productos.all():
                linea_subtotal = (dp.precio or 0) * (dp.stock or 0)
                subtotal_bruto += linea_subtotal
                dcto_monto = int(dp.descuento_monto or 0)
                pt = dp.productoTalla
                productos.append({
                    'sku': pt.sku if pt else '',
                    'nombre': (
                        pt.producto.articulo if (pt and pt.producto) else dp.descripcion
                    ),
                    'talla': pt.talla if pt else '',
                    'cantidad': dp.stock,
                    'precio_unitario': dp.precio,
                    'subtotal': linea_subtotal,
                    'descuento_monto': dcto_monto,
                    'monto_item': dp.monto_item or (linea_subtotal - dcto_monto),
                    'costo': dp.costo,
                    'sobreprecio': dp.sobreprecio,
                })

            metodos_pago_raw = []
            total_pagos = 0
            # Los pagos heredan la fecha del DTE (Dte_Detalle_Pago no tiene
            # campo `fecha` propio). Se expone aquí para poder renderizarla
            # en la grilla y comprobar que sigue alineada al DTE tras una
            # edición de `fecha_emision`.
            fecha_pago_dte = (
                dte.fecha_emision.strftime('%Y-%m-%d')
                if dte.fecha_emision else None
            )
            for pago in dte.dte_asociado.all():
                total_pagos += pago.monto or 0
                metodos_pago_raw.append({
                    'id': pago.id,
                    'metodo': pago.metodo_pago,
                    'metodo_display': obtener_nombre_metodo_pago(pago.metodo_pago),
                    'monto': pago.monto,
                    'voucher': pago.voucher or '',
                    'tipo_tarjeta': pago.tipo_tarjeta or '',
                    'notas': pago.notas or '',
                    'fecha_pago': fecha_pago_dte,
                })
            metodos_pago = agrupar_metodos_pago(metodos_pago_raw)

            monto_lista = int(dte.monto_con_iva or 0)
            total_real = total_pagos if total_pagos > 0 else monto_lista

            descuento_guardado = int(dte.descuento or 0)
            if descuento_guardado > 0:
                descuento_efectivo = descuento_guardado
            elif total_pagos > 0 and total_pagos < monto_lista:
                descuento_efectivo = monto_lista - total_pagos
            else:
                descuento_efectivo = max(0, subtotal_bruto - monto_lista)

            estado_display = 'PAGADO' if dte.estado_dte == 'EMITIDO' else dte.estado_dte

            fecha_dt = timezone.datetime.combine(dte.fecha_emision, dt_time.min)
            created_at_dte = (
                timezone.make_aware(fecha_dt)
                if timezone.is_naive(fecha_dt) else fecha_dt
            )

            documentos_paginados.append({
                'id': dte.id,
                'tipo': dte.tipo_documento,
                'tipo_documento': dte.tipo_documento,
                'numero': dte.numero_documento,
                'numero_documento': dte.numero_documento,
                'fecha': dte.fecha_emision,
                'cliente_nombre': dte.receptor.nombre if dte.receptor else 'Sin nombre',
                'cliente_rut': dte.receptor.rut if dte.receptor else '',
                'cliente_giro': dte.receptor.giro if dte.receptor else '',
                'cliente_email': dte.receptor.correoVendedor if dte.receptor else '',
                'cliente_direccion': dte.receptor.direccion if dte.receptor else '',
                'cliente_comuna': dte.receptor.comuna if dte.receptor else '',
                'vendedor_id': dte.vendedor_id,
                'vendedor_nombre': (
                    f"{dte.vendedor.codigo_vendedor} - {dte.vendedor.nombre}"
                    if dte.vendedor else 'Sin vendedor'
                ),
                'total': total_real,
                'subtotal_bruto': subtotal_bruto,
                'monto_neto': int(dte.monto_neto or 0),
                'descuento': descuento_efectivo,
                'estado': estado_display,
                'estado_dte': dte.estado_dte,
                'hora': dte.hora.strftime('%H:%M') if dte.hora else '',
                'created_at': created_at_dte,
                'productos': productos,
                'metodos_pago': metodos_pago,
                'pagos_raw': metodos_pago_raw,
                'total_productos': len(productos),
                'metodos_pago_str': formatear_metodos_pago_str(metodos_pago),
                'ticket_correlativo': (
                    ticket_map_by_folio[dte.numero_documento].correlativo
                    if dte.numero_documento in ticket_map_by_folio else None
                ),
                'ticket_id': (
                    ticket_map_by_folio[dte.numero_documento].id
                    if dte.numero_documento in ticket_map_by_folio else None
                ),
                'tiene_nc': dte.id in dtes_con_nc,
                'observaciones': getattr(dte, 'referencias', '') or '',
                'es_manual': bool(getattr(dte, 'es_manual', False)),
            })

        return JsonResponse({
            'success': True,
            'documentos': documentos_paginados,
            'total': total_documentos,
            'pagination': {
                'current_page': page_obj.number,
                'per_page': per_page,
                'per_page_solicitado': per_page_solicitado,
                'total_pages': paginator.num_pages,
                'total_items': total_documentos,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            },
            'estadisticas': {
                'total_documentos': total_documentos,
                'total_ventas': int(stats['total_ventas'] or 0),
                'total_pendientes': stats['total_pendientes'] or 0,
                'total_facturas': stats['total_facturas'] or 0,
                'total_boletas': stats['total_boletas'] or 0,
                'total_boletas_electronicas': stats['total_boletas_electronicas'] or 0,
                'total_boletas_papel': stats['total_boletas_papel'] or 0,
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener documentos: {str(e)}'
        })


@login_required
@require_GET
def exportar_documentos_ventas_excel(request):
    """
    API para exportar documentos de ventas (DTEs) a Excel
    Utiliza los mismos filtros que listar_documentos_ventas
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Función helper para convertir códigos de método de pago
        def obtener_nombre_metodo_pago(codigo):
            nombres_metodos = {
                'EFECTIVO': 'Efectivo',
                'TARJETA_DEBITO': 'Tarjeta Débito',
                'TARJETA_CREDITO': 'Tarjeta Crédito',
                'TRANSFERENCIA': 'Transferencia',
                'CHEQUE': 'Cheque',
                'OTRO': 'Otro',
                'TBK_POS_INTEGRADO': 'Transbank POS',
                'TBK_MANUAL': 'Transbank Manual',
                'TBK_DEBITO_POS': 'TBK Débito POS',
                'TBK_CREDITO_POS': 'TBK Crédito POS',
                'TBK_PREPAGO_POS': 'TBK Prepago POS',
                'TARJETA_COMERCIAL': 'Tarjeta Comercial',
                'VENTA_INTERNET': 'Venta por Internet',
                'ORDEN_COMPRA': 'Orden de Compra',
                'CREDITO_TRABAJADOR': 'Crédito Trabajador',
                'CREDITO_EXTERNO': 'Crédito Externo',
            }
            return nombres_metodos.get(codigo, codigo)
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        # Obtener nombre de sucursal
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id)
            sucursal_nombre = sucursal.alias or sucursal.nombre
        except Sucursal.DoesNotExist:
            sucursal_nombre = 'Sucursal'
        
        # Parámetros de filtro (los mismos que listar_documentos_ventas)
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_documento = request.GET.get('tipo_documento')
        estado = request.GET.get('estado')
        metodo_pago = request.GET.get('metodo_pago')
        buscar = request.GET.get('buscar', '').strip()
        
        # Parámetros de ordenamiento
        orden_campo = request.GET.get('orden_campo', 'fecha')
        orden_direccion = request.GET.get('orden_direccion', 'desc')
        
        # Query base de DTEs
        dtes_query = Dte.objects.select_related(
            'vendedor', 
            'receptor'
        ).prefetch_related(
            # Solo los pagos: el Excel no exporta líneas de producto, así que
            # prefetchear `dte_productos` materializaba miles de filas para nada.
            'dte_asociado',
        ).filter(
            sucursal_id=sucursal_id,
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
            # Mismo criterio que `listar_documentos_ventas`: sin esto el Excel
            # incluía los DTE eliminados lógicamente y no cuadraba con la pantalla.
            descartado=False,
        )
        
        # Aplicar filtros de fecha
        if fecha_desde:
            dtes_query = dtes_query.filter(fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            dtes_query = dtes_query.filter(fecha_emision__lte=fecha_hasta)
        
        # Aplicar filtros por tipo de DTE
        if tipo_documento:
            if tipo_documento == 'BOLETA_ELECTRONICA':
                dtes_query = dtes_query.filter(tipo_documento='BOLETA ELECTRONICA')
            elif tipo_documento == 'BOLETA_PAPEL':
                dtes_query = dtes_query.filter(tipo_documento='BOLETA PAPEL')
            elif tipo_documento == 'FACTURA_ELECTRONICA':
                dtes_query = dtes_query.filter(tipo_documento='FACTURA ELECTRONICA')
            elif tipo_documento == 'FACTURA_EXENTA':
                dtes_query = dtes_query.filter(tipo_documento='FACTURA EXENTA')
        
        # Filtrar por estado
        if estado:
            estado_dte_map = {
                'PENDIENTE': 'PENDIENTE',
                'PAGADO': 'EMITIDO',
                'ANULADO': 'ANULADO'
            }
            if estado in estado_dte_map:
                dtes_query = dtes_query.filter(estado_dte=estado_dte_map[estado])
        
        # Filtrar por método de pago
        # MULTIPLE es un valor sintético: DTEs con más de un método de pago distinto
        if metodo_pago:
            if metodo_pago == 'MULTIPLE':
                dtes_query = dtes_query.annotate(
                    _n_metodos=Count('dte_asociado__metodo_pago', distinct=True)
                ).filter(_n_metodos__gt=1)
            else:
                dtes_query = dtes_query.filter(dte_asociado__metodo_pago=metodo_pago).distinct()
        
        # Filtro de búsqueda
        if buscar:
            dtes_query = dtes_query.filter(
                Q(numero_documento__icontains=buscar) |
                Q(receptor__nombre__icontains=buscar) |
                Q(receptor__rut__icontains=buscar) |
                Q(vendedor__nombre__icontains=buscar) |
                Q(dte_productos__productoTalla__sku__icontains=buscar) |
                Q(dte_productos__productoTalla__producto__articulo__icontains=buscar)
            ).distinct()
        
        # Recolectar datos de documentos
        documentos_data = []
        for dte in dtes_query:
            # Obtener métodos de pago
            metodos_pago_list = []
            for pago in dte.dte_asociado.all():
                nombre_metodo = obtener_nombre_metodo_pago(pago.metodo_pago)
                if pago.tipo_tarjeta:
                    nombre_metodo += f" ({pago.tipo_tarjeta})"
                metodos_pago_list.append(nombre_metodo)
            metodos_pago_str = ', '.join(metodos_pago_list) if metodos_pago_list else 'Sin pagos'
            
            # Mapear estado DTE
            estado_display = 'PAGADO' if dte.estado_dte == 'EMITIDO' else dte.estado_dte
            
            # Crear datetime para ordenamiento
            from datetime import time as dt_time
            fecha_dt = timezone.datetime.combine(dte.fecha_emision, dt_time.min)
            created_at_dte = timezone.make_aware(fecha_dt) if timezone.is_naive(fecha_dt) else fecha_dt
            
            documentos_data.append({
                'fecha': dte.fecha_emision,
                'created_at': created_at_dte,
                'tipo': dte.tipo_documento,
                'numero': dte.numero_documento,
                'cliente_nombre': dte.receptor.nombre if dte.receptor else 'Sin nombre',
                'cliente_rut': dte.receptor.rut if dte.receptor else '',
                'vendedor_nombre': f"{dte.vendedor.codigo_vendedor} - {dte.vendedor.nombre}" if dte.vendedor else 'Sin vendedor',
                # `Dte` no tiene `monto_total` ni `iva`: son `monto_neto` y
                # `monto_con_iva`. Leer los nombres inexistentes lanzaba
                # AttributeError en el primer documento, lo tragaba el
                # `except Exception` de más abajo y el usuario recibía un JSON
                # de error en vez del .xlsx.
                'neto': int(dte.monto_neto or 0),
                'iva': int((dte.monto_con_iva or 0) - (dte.monto_neto or 0)),
                'total': int(dte.monto_con_iva or 0),
                'metodos_pago': metodos_pago_str,
                'estado': estado_display,
            })
        
        # Ordenar documentos
        orden_map = {
            'fecha': 'created_at',
            'tipo_documento': 'tipo',
            'numero_documento': 'numero',
            'cliente_nombre': 'cliente_nombre',
            'vendedor_nombre': 'vendedor_nombre',
            'total': 'total',
            'estado': 'estado',
        }
        campo_ordenar = orden_map.get(orden_campo, 'created_at')
        reverse_order = (orden_direccion == 'desc')
        
        try:
            if campo_ordenar == 'total':
                documentos_data.sort(key=lambda x: x.get(campo_ordenar, 0) or 0, reverse=reverse_order)
            elif campo_ordenar == 'numero':
                documentos_data.sort(key=lambda x: int(x.get(campo_ordenar, 0) or 0), reverse=reverse_order)
            else:
                documentos_data.sort(key=lambda x: str(x.get(campo_ordenar, '') or '').lower(), reverse=reverse_order)
        except (TypeError, ValueError):
            documentos_data.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Documentos de Ventas"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = f"DOCUMENTOS DE VENTAS - {sucursal_nombre.upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:K1')
        
        # Periodo
        periodo_texto = "Todos los documentos"
        if fecha_desde and fecha_hasta:
            periodo_texto = f"Período: {fecha_desde} al {fecha_hasta}"
        elif fecha_desde:
            periodo_texto = f"Desde: {fecha_desde}"
        elif fecha_hasta:
            periodo_texto = f"Hasta: {fecha_hasta}"
        ws['A2'] = periodo_texto
        ws['A2'].font = Font(italic=True)
        
        # Encabezados
        headers = [
            "Fecha", "Tipo Documento", "Número", "Cliente", "RUT", 
            "Vendedor", "Neto", "IVA", "Total", "Método Pago", "Estado"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        total_neto = 0
        total_iva = 0
        total_general = 0
        
        for row_idx, doc in enumerate(documentos_data, 5):
            ws.cell(row=row_idx, column=1, value=doc['fecha'].strftime('%d/%m/%Y') if doc['fecha'] else '').border = border
            ws.cell(row=row_idx, column=2, value=doc['tipo']).border = border
            ws.cell(row=row_idx, column=3, value=doc['numero']).border = border
            ws.cell(row=row_idx, column=4, value=doc['cliente_nombre']).border = border
            ws.cell(row=row_idx, column=5, value=doc['cliente_rut']).border = border
            ws.cell(row=row_idx, column=6, value=doc['vendedor_nombre']).border = border
            
            cell_neto = ws.cell(row=row_idx, column=7, value=doc['neto'])
            cell_neto.number_format = '#,##0'
            cell_neto.border = border
            
            cell_iva = ws.cell(row=row_idx, column=8, value=doc['iva'])
            cell_iva.number_format = '#,##0'
            cell_iva.border = border
            
            cell_total = ws.cell(row=row_idx, column=9, value=doc['total'])
            cell_total.number_format = '#,##0'
            cell_total.border = border
            
            ws.cell(row=row_idx, column=10, value=doc['metodos_pago']).border = border
            ws.cell(row=row_idx, column=11, value=doc['estado']).border = border
            
            total_neto += doc['neto']
            total_iva += doc['iva']
            total_general += doc['total']
        
        # Fila de totales
        if documentos_data:
            row_totales = len(documentos_data) + 5
            ws.cell(row=row_totales, column=6, value="TOTALES:").font = Font(bold=True)
            
            cell_total_neto = ws.cell(row=row_totales, column=7, value=total_neto)
            cell_total_neto.number_format = '#,##0'
            cell_total_neto.font = Font(bold=True)
            
            cell_total_iva = ws.cell(row=row_totales, column=8, value=total_iva)
            cell_total_iva.number_format = '#,##0'
            cell_total_iva.font = Font(bold=True)
            
            cell_total_general = ws.cell(row=row_totales, column=9, value=total_general)
            cell_total_general.number_format = '#,##0'
            cell_total_general.font = Font(bold=True)
        
        # Ajustar anchos de columna
        column_widths = [12, 22, 12, 35, 15, 25, 12, 12, 12, 25, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Generar respuesta
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        fecha_actual = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"documentos_ventas_{fecha_actual}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.exception("Error al exportar documentos de ventas")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar documentos: {str(e)}'
        })


@login_required
@require_POST
def convertir_ticket_a_factura(request):
    """Convertir un ticket a factura electrónica"""
    try:
        data = json.loads(request.body)
        documento_id = data.get('documento_id')
        
        if not documento_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de documento requerido'
            })

        # Obtener el ticket original (prefetch para evitar N+1 al recorrer líneas)
        ticket = get_object_or_404(
            Ticket.objects.prefetch_related('ticket_productos__ProductoTalla__producto'),
            id=documento_id,
        )
        
        # Verificar que el ticket esté pagado
        if ticket.estado != 'PAGADO':
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden convertir tickets pagados'
            })

        # Verificar si ya existe una factura para este ticket
        factura_existente = Dte.objects.filter(
            referencias__icontains=f'TICKET-{ticket.correlativo}'
        ).first()
        
        if factura_existente:
            return JsonResponse({
                'success': False,
                'error': f'Ya existe la factura #{factura_existente.numero_documento} para este ticket'
            })

        with transaction.atomic():
            # Crear o obtener empresa receptora
            receptor_data = {
                'nombre': data.get('cliente_razon_social'),
                'rut': data.get('cliente_rut'),
                'giro': data.get('cliente_giro', ''),
                'direccion': data.get('cliente_direccion', ''),
                'comuna': data.get('cliente_comuna', ''),
                'correoVendedor': data.get('cliente_email', ''),
            }
            
            receptor, created = Empresa.objects.get_or_create(
                rut=receptor_data['rut'],
                defaults=receptor_data
            )
            
            if not created:
                # Actualizar datos si ya existe
                for key, value in receptor_data.items():
                    if value:  # Solo actualizar si hay valor
                        setattr(receptor, key, value)
                receptor.save()

            # Obtener siguiente correlativo para factura
            numero_factura = obtener_siguiente_correlativo(ticket.sucursal, 'FACTURA ELECTRONICA')

            # Crear la factura
            factura = Dte.objects.create(
                emisor=ticket.sucursal.empresa,
                receptor=receptor,
                numero_documento=numero_factura,
                tipo_documento='FACTURA ELECTRONICA',
                monto_con_iva=ticket.total,
                monto_neto=int(ticket.total / 1.19),  # Calcular neto (asumiendo IVA 19%)
                estado_pago='PAGADO',
                estado_dte='EMITIDO',
                responsable=request.user.username,
                fecha_emision=data.get('fecha_emision', timezone.localdate()),
                fecha_vencimiento=data.get('fecha_emision', timezone.localdate()),
                diasCredito=0,
                bultos=1,
                unidades_productos=ticket.ticket_productos.aggregate(
                    total=Sum('stock')
                )['total'] or 0,
                vendedor=ticket.vendedor,
                descuento=ticket.descuento or 0,
                sucursal=ticket.sucursal,
                tipo_transaccion='VENTA',
                referencias=f'TICKET-{ticket.correlativo}'
            )

            # Copiar productos del ticket a la factura (con descuentos)
            for ticket_producto in ticket.ticket_productos.all():
                costo_unitario = ticket_producto.ProductoTalla.producto.costo
                sobreprecio_unitario = ticket_producto.ProductoTalla.producto.sobreprecio

                dcto_u = ticket_producto.descuento_unitario or 0
                dcto_p = float(ticket_producto.porcentaje_descuento or 0)
                dcto_linea = dcto_u * ticket_producto.stock if dcto_u else 0

                # Igual que en la copia ticket→DTE de boleta: derivar el
                # precio efectivo desde subtotal/stock + descuento. Cubre
                # el caso del "envío" con precio sistema fijo cobrado a
                # otro precio. Sin esto, la línea del TXT NC sale con
                # precio sistema en vez del cobrado.
                precio_efectivo_f = ticket_producto.precio
                if ticket_producto.stock and ticket_producto.subtotal:
                    derivado = int(round(ticket_producto.subtotal / ticket_producto.stock)) + dcto_u
                    if derivado and derivado != ticket_producto.precio:
                        precio_efectivo_f = derivado

                Dte_Productos.objects.create(
                    dte=factura,
                    productoTalla=ticket_producto.ProductoTalla,
                    descripcion=f"{ticket_producto.ProductoTalla.producto.articulo} - {ticket_producto.ProductoTalla.talla}",
                    costo=costo_unitario,
                    sobreprecio=sobreprecio_unitario,
                    precio=precio_efectivo_f,
                    precio_unitario=precio_efectivo_f,
                    descuento_pct=dcto_p if dcto_p > 0 else None,
                    descuento_monto=dcto_linea if dcto_linea > 0 else None,
                    monto_item=ticket_producto.subtotal,
                    stock=ticket_producto.stock,
                    activo=True
                )

            # Copiar pagos del ticket a la factura
            for pago_ticket in ticket.pagos.all():
                Dte_Detalle_Pago.objects.create(
                    dte=factura,
                    metodo_pago=pago_ticket.get_metodo_pago_display(),
                    tipo_tarjeta=pago_ticket.tipo_tarjeta or '',
                    voucher=pago_ticket.voucher or '',
                    monto=pago_ticket.monto,
                    notas=pago_ticket.notas or ''
                )

        return JsonResponse({
            'success': True,
            'message': 'Factura creada exitosamente',
            'numero_factura': numero_factura,
            'factura_id': factura.id
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear factura: {str(e)}'
        })


@login_required
@require_GET
def detalle_documento_venta(request, documento_id):
    """Obtener detalle completo de un documento de venta"""
    try:
        tipo_documento = request.GET.get('tipo', 'TICKET')

        def obtener_nombre_metodo_pago(codigo):
            nombres_metodos = {
                'EFECTIVO': 'Efectivo',
                'TARJETA_DEBITO': 'Tarjeta Débito',
                'TARJETA_CREDITO': 'Tarjeta Crédito',
                'TRANSFERENCIA': 'Transferencia',
                'CHEQUE': 'Cheque',
                'OTRO': 'Otro',
                'TBK_POS_INTEGRADO': 'Transbank POS',
                'TBK_MANUAL': 'Transbank Manual',
                'TBK_DEBITO_POS': 'TBK Débito POS',
                'TBK_CREDITO_POS': 'TBK Crédito POS',
                'TBK_PREPAGO_POS': 'TBK Prepago POS',
                'TARJETA_COMERCIAL': 'Tarjeta Comercial',
                'VENTA_INTERNET': 'Venta por Internet',
                'ORDEN_COMPRA': 'Orden de Compra',
                'CREDITO_TRABAJADOR': 'Crédito Trabajador',
                'CREDITO_EXTERNO': 'Crédito Externo',
            }
            return nombres_metodos.get(codigo, codigo)
        
        if tipo_documento == 'TICKET':
            documento = get_object_or_404(Ticket, id=documento_id)
            
            # Obtener productos del ticket
            productos = []
            for tp in documento.ticket_productos.select_related('ProductoTalla__producto').all():
                if tp.ProductoTalla:
                    productos.append({
                        'sku': tp.ProductoTalla.sku,
                        'nombre': tp.ProductoTalla.producto.articulo if tp.ProductoTalla.producto else (tp.descripcion_linea or ''),
                        'talla': tp.ProductoTalla.talla,
                        'cantidad': tp.stock,
                        'precio_unitario': tp.precio,
                        'subtotal': tp.subtotal,
                    })
                else:
                    productos.append({
                        'sku': '',
                        'nombre': tp.descripcion_linea or 'Ítem pendiente de despacho',
                        'talla': '',
                        'cantidad': tp.stock,
                        'precio_unitario': tp.precio,
                        'subtotal': tp.subtotal,
                    })
            
            # Obtener pagos
            pagos = []
            for pago in documento.pagos.all():
                pagos.append({
                    'metodo': pago.get_metodo_pago_display(),
                    'monto': pago.monto,
                    'voucher': pago.voucher or '',
                    'notas': pago.notas or '',
                })
            
            detalle = {
                'tipo': 'TICKET',
                'numero': documento.correlativo,
                'fecha': documento.created_at.date(),
                'hora': documento.created_at.time(),
                'estado': documento.estado,
                'cliente': {
                    'nombre': documento.cliente_nombre or '',
                    'rut': documento.cliente_rut or '',
                    'email': documento.cliente_email or '',
                    'telefono': documento.cliente_telefono or '',
                },
                'vendedor': documento.vendedor.nombre if documento.vendedor else '',
                'productos': productos,
                'pagos': pagos,
                'totales': {
                    'subtotal': documento.subTotal,
                    'descuento': documento.descuento or 0,
                    'total': documento.total,
                },
                'observaciones': documento.observaciones or '',
            }
            
        else:  # DTE (Factura/Boleta)
            documento = get_object_or_404(Dte, id=documento_id)
            
            # Buscar ticket vinculado para mostrar el correlativo
            ticket_vinculado = (
                Ticket.objects
                .filter(
                    sucursal=documento.sucursal,
                    folio_dte=documento.numero_documento,
                    estado='PAGADO',
                )
                .only('correlativo', 'hora', 'fecha')
                .first()
            )
            
            # Obtener productos del DTE
            productos = []
            for dp in documento.dte_productos.select_related('productoTalla__producto').all():
                productos.append({
                    'sku': dp.productoTalla.sku if dp.productoTalla else '',
                    'nombre': dp.descripcion,
                    'cantidad': dp.stock,
                    'precio_unitario': dp.precio,
                    'subtotal': dp.precio * dp.stock,
                })
            
            # Obtener pagos.
            # Nota: `Dte_Detalle_Pago` no tiene un campo `fecha` propio: la
            # fecha del pago HEREDA de `dte.fecha_emision`. Por eso la
            # exponemos como `fecha_pago` para que el frontend pueda mostrarla
            # al usuario y verificar que siempre está alineada con el DTE
            # (útil cuando se edita la fecha del DTE desde /ventas/documentos).
            fecha_pago_dte = (
                documento.fecha_emision.strftime('%Y-%m-%d')
                if documento.fecha_emision else None
            )
            pagos_raw = []
            for pago in documento.dte_asociado.all():
                pagos_raw.append({
                    'id': pago.id,
                    'metodo': pago.metodo_pago,
                    'metodo_display': obtener_nombre_metodo_pago(pago.metodo_pago),
                    'monto': pago.monto,
                    'voucher': pago.voucher or '',
                    'tipo_tarjeta': pago.tipo_tarjeta or '',
                    'notas': pago.notas or '',
                    'fecha_pago': fecha_pago_dte,
                })

            # Agrupar pagos por método y sumar montos
            pagos = []
            agrupados = {}
            for pago in pagos_raw:
                metodo = pago.get('metodo') or ''
                metodo_display = pago.get('metodo_display') or metodo
                tipo_tarjeta = pago.get('tipo_tarjeta') or ''
                key = (metodo, metodo_display, tipo_tarjeta)
                if key not in agrupados:
                    agrupados[key] = {
                        'metodo': metodo,
                        'metodo_display': metodo_display,
                        'monto': 0,
                        'voucher': '',
                        'tipo_tarjeta': tipo_tarjeta,
                        'notas': '',
                        'fecha_pago': fecha_pago_dte,
                        '_vouchers': set(),
                        '_notas': set(),
                    }
                agrupados[key]['monto'] += pago.get('monto') or 0
                if pago.get('voucher'):
                    agrupados[key]['_vouchers'].add(str(pago['voucher']))
                if pago.get('notas'):
                    agrupados[key]['_notas'].add(str(pago['notas']))

            for item in agrupados.values():
                if item['_vouchers']:
                    item['voucher'] = ', '.join(sorted(item['_vouchers']))
                if item['_notas']:
                    item['notas'] = ' | '.join(sorted(item['_notas']))
                item.pop('_vouchers', None)
                item.pop('_notas', None)
                pagos.append(item)
            
            detalle = {
                'tipo': 'FACTURA' if 'FACTURA' in documento.tipo_documento else 'BOLETA',
                'tipo_documento': documento.tipo_documento,
                'numero': documento.numero_documento,
                'numero_documento': documento.numero_documento,
                'fecha': documento.fecha_emision,
                'hora': documento.hora.strftime('%H:%M') if documento.hora else '',
                'estado': documento.estado_dte,
                'ticket_correlativo': ticket_vinculado.correlativo if ticket_vinculado else None,
                'cliente': {
                    'nombre': documento.receptor.nombre if documento.receptor else '',
                    'rut': documento.receptor.rut if documento.receptor else '',
                    'giro': documento.receptor.giro if documento.receptor else '',
                    'direccion': documento.receptor.direccion if documento.receptor else '',
                },
                'vendedor': documento.vendedor.nombre if documento.vendedor else '',
                'productos': productos,
                'pagos': pagos,
                'pagos_raw': pagos_raw,
                'totales': {
                    'neto': documento.monto_neto,
                    'iva': documento.monto_con_iva - documento.monto_neto,
                    # Mismo criterio que `listar_documentos_ventas` (ver su
                    # comentario junto a `_total_pagos`): en DTEs históricos
                    # emitidos antes del fix de `descuento_fidelizacion` en
                    # `generar_dte_desde_ticket`, `monto_con_iva` puede quedar
                    # mayor que lo realmente cobrado si parte se pagó con vale
                    # de fidelización (el vale no genera un `Dte_Detalle_Pago`).
                    # Usamos la suma de pagos reales cuando existe; si no hay
                    # pagos registrados, caemos a `monto_con_iva`.
                    'total': (
                        sum(p.get('monto') or 0 for p in pagos_raw)
                        or documento.monto_con_iva
                    ),
                },
                'referencias': documento.referencias or '',
            }

        return JsonResponse({
            'success': True,
            'documento': detalle
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener detalle: {str(e)}'
        })


@login_required
@require_POST
def anular_documento_venta(request):
    """Anular un documento de venta"""
    try:
        data = json.loads(request.body)
        documento_id = data.get('documento_id')
        tipo_documento = data.get('tipo', 'TICKET')
        
        if not documento_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de documento requerido'
            })

        with transaction.atomic():
            if tipo_documento == 'TICKET':
                documento = get_object_or_404(Ticket, id=documento_id)
                
                if documento.estado == 'ANULADO':
                    return JsonResponse({
                        'success': False,
                        'error': 'El ticket ya está anulado'
                    })
                
                # Anular ticket
                documento.estado = 'ANULADO'
                documento.save()

                # El cupón vuelve al cliente: la venta que lo consumió ya no existe.
                _liberar_cupon_de_venta(documento, 'anulación de documento de venta')

                # Devolver stock si estaba pagado
                # ⚠️ RAMA MUERTA (bug pre-existente, NO corregido acá A PROPÓSITO):
                # el estado se pisa con 'ANULADO' tres líneas más arriba, así que
                # esta condición nunca es True y este endpoint jamás devolvió
                # stock. Habilitarla cambia el inventario de forma retroactiva y
                # exige verificar antes que no duplique con las otras rutas de
                # anulación (eliminar_documento_venta / anular_factura_dte).
                if documento.estado == 'PAGADO':
                    for tp in documento.ticket_productos.all():
                        if tp.ProductoTalla is None:
                            continue  # Sin stock que devolver para ítems manuales
                        # Crear movimiento de devolución de stock
                        # ✅ Usar DTE si está disponible, si no usar correlativo del ticket
                        referencia = f'ANULACION_DTE_{documento.folio_dte}' if documento.folio_dte else f'ANULACION_TICKET_{documento.correlativo}'
                        Movimientos_Producto.objects.create(
                            ticket=documento,
                            ProductoTalla=tp.ProductoTalla,
                            cantidad=tp.stock,  # Cantidad positiva para devolver
                            costo=tp.ProductoTalla.producto.costo if tp.ProductoTalla.producto else 0,
                            precio=tp.precio,
                            concepto='DEVOLUCION_CLIENTE',
                            tipo_movimiento='INGRESO',
                            responsable=request.user.username,
                            observaciones=f'Anulación ticket #{documento.correlativo}',
                            referencia_externa=referencia
                        )
                
            else:  # DTE
                documento = get_object_or_404(Dte, id=documento_id)
                
                if documento.estado_dte == 'ANULADO':
                    return JsonResponse({
                        'success': False,
                        'error': 'El documento ya está anulado'
                    })
                
                documento.estado_dte = 'ANULADO'
                documento.save()

                # En "Documentos de Ventas" una boleta del POS se lista como
                # DTE, así que anularla por acá es el camino natural del
                # supervisor. El ticket se resuelve ACOTADO POR SUCURSAL:
                # `folio_dte` no es único entre empresas del holding.
                ticket_del_dte = Ticket.objects.filter(
                    sucursal_id=documento.sucursal_id,
                    folio_dte=documento.numero_documento,
                ).first()
                _liberar_cupon_de_venta(
                    ticket_del_dte, 'anulación de DTE de venta')

        return JsonResponse({
            'success': True,
            'message': 'Documento anulado exitosamente'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al anular documento: {str(e)}'
        })


@login_required
@require_POST
def eliminar_documento_venta(request):
    """
    Elimina (soft delete) un DTE desde gestión de documentos de ventas:

    1. Marca el DTE como ``descartado=True`` y ``estado_dte='ANULADO'``.
       - El listado de documentos lo filtra (`descartado=False`).
       - La cuadratura de caja lo filtra (mismo criterio).

    2. Devuelve a bodega el stock vendido. La fuente del stock depende del
       vínculo:
         - Si el DTE tiene un Ticket vinculado (`Ticket.folio_dte ==
           Dte.numero_documento`), recorremos `Ticket_Productos` y
           devolvemos cada `ProductoTalla.stock`.
         - Si no hay ticket, recorremos `Dte_Productos` (sólo los que
           tengan `productoTalla` no nulo).
         - Si el DTE es una NOTA DE CREDITO, NO se toca stock: la NC ya
           devolvió el stock vendido a bodega al emitirse, y volver a
           sumarlo aquí lo duplicaría.
       Por cada línea se crea un `Movimientos_Producto` con
       ``concepto='DEVOLUCION_CLIENTE'`` y ``tipo_movimiento='INGRESO'`` con
       ``referencia_externa='ELIMINACION_DTE_<numero>'`` para mantener
       trazabilidad del stock devuelto.

    3. Si hay ticket vinculado, lo marca como ``estado='ANULADO'`` (queda
       fuera de tickets pagados de la cuadratura).

    Restricción: solo administradores (`request.user.rol == 'administrador'`).
    El soft delete deja el DTE en BD para auditoría / reversión manual con
    `restaurar_dte`.

    Body esperado::

        { "documento_id": 123, "motivo": "Texto opcional" }
    """
    try:
        data = json.loads(request.body or '{}')
        documento_id = data.get('documento_id')
        motivo = (data.get('motivo') or '').strip()[:200]

        if not documento_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de documento requerido'
            })

        # Solo administradores pueden eliminar.
        rol_usuario = getattr(request.user, 'rol', '') or ''
        if rol_usuario != 'administrador':
            return JsonResponse({
                'success': False,
                'error': 'Solo los administradores pueden eliminar documentos'
            }, status=403)

        # Trazabilidad de quién y cuándo descartó.
        responsable = (
            request.user.get_full_name()
            or request.user.username
            or 'Sistema'
        )

        with transaction.atomic():
            dte = (
                Dte.objects
                .select_for_update()
                .filter(id=documento_id)
                .first()
            )
            if not dte:
                return JsonResponse({
                    'success': False,
                    'error': 'Documento no encontrado'
                })

            if dte.descartado:
                return JsonResponse({
                    'success': False,
                    'error': 'El documento ya fue eliminado'
                })

            referencia = f'ELIMINACION_DTE_{dte.numero_documento}'
            obs_base = (
                f'Eliminación DTE #{dte.numero_documento} '
                f'({dte.tipo_documento}) por {responsable}'
                + (f'. Motivo: {motivo}' if motivo else '')
            )

            ticket_vinculado = None
            if dte.numero_documento:
                ticket_vinculado = (
                    Ticket.objects
                    .select_for_update()
                    .filter(
                        sucursal_id=dte.sucursal_id,
                        folio_dte=dte.numero_documento,
                    )
                    .first()
                )

            stock_devuelto = []  # [{sku, cantidad}, ...]
            movimientos_creados = 0

            if dte.tipo_documento == 'NOTA DE CREDITO':
                # Una NC ya devolvió el stock vendido a bodega al emitirse
                # (para eso existe el flujo de devolución). Si aquí
                # también sumáramos stock por sus propias `Dte_Productos`
                # duplicaríamos el ingreso a inventario, así que el soft
                # delete de una NC NO toca stock ni busca ticket vinculado
                # (una NC no tiene ticket propio: `ticket_vinculado` es
                # siempre None para este tipo).
                pass
            elif ticket_vinculado:
                # Caso 1: el stock fue descontado al pagar el ticket; lo
                # devolvemos por las líneas del ticket (mismas SKUs / cant.).
                productos_ticket = (
                    Ticket_Productos.objects
                    .filter(idTicket=ticket_vinculado)
                    .select_related('ProductoTalla', 'ProductoTalla__producto')
                )
                for tp in productos_ticket:
                    if tp.ProductoTalla is None or not tp.stock:
                        continue
                    pt = tp.ProductoTalla
                    pt.stock = (pt.stock or 0) + int(tp.stock)
                    pt.save(update_fields=['stock'])

                    Movimientos_Producto.objects.create(
                        ticket=ticket_vinculado,
                        ProductoTalla=pt,
                        sucursal_destino=ticket_vinculado.sucursal,
                        cantidad=int(tp.stock),
                        costo=(
                            pt.producto.costo if pt.producto else 0
                        ),
                        precio=int(tp.precio or 0),
                        concepto='DEVOLUCION_CLIENTE',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=responsable,
                        observaciones=obs_base,
                        referencia_externa=referencia,
                    )
                    movimientos_creados += 1
                    stock_devuelto.append({
                        'sku': getattr(pt, 'sku', '') or '',
                        'cantidad': int(tp.stock),
                    })

                # Anular ticket: lo saca de la cuadratura
                # (Ticket queries filtran por estado='PAGADO').
                if ticket_vinculado.estado != 'ANULADO':
                    ticket_vinculado.estado = 'ANULADO'
                    ticket_vinculado.save(update_fields=['estado'])
                    # El cupón vuelve al cliente. Va dentro del `if` para que
                    # eliminar dos veces el mismo documento no toque un cupón que
                    # el cliente ya volvió a usar en otra venta.
                    _liberar_cupon_de_venta(
                        ticket_vinculado, 'eliminación de documento desde cuadratura')
            else:
                # Caso 2: DTE sin ticket vinculado (factura emitida directa,
                # NC, etc.). Devolvemos por Dte_Productos.
                productos_dte = (
                    Dte_Productos.objects
                    .filter(dte=dte)
                    .select_related('productoTalla', 'productoTalla__producto')
                )
                for dp in productos_dte:
                    pt = dp.productoTalla
                    if pt is None or not dp.stock:
                        # Líneas manuales / sin SKU: nada que devolver.
                        continue
                    pt.stock = (pt.stock or 0) + int(dp.stock)
                    pt.save(update_fields=['stock'])

                    Movimientos_Producto.objects.create(
                        dte=dte,
                        ProductoTalla=pt,
                        sucursal_destino=dte.sucursal,
                        cantidad=int(dp.stock),
                        costo=int(dp.costo or 0),
                        sobreprecio=int(dp.sobreprecio or 0),
                        precio=int(dp.precio or 0),
                        concepto='DEVOLUCION_CLIENTE',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=responsable,
                        observaciones=obs_base,
                        referencia_externa=referencia,
                    )
                    movimientos_creados += 1
                    stock_devuelto.append({
                        'sku': getattr(pt, 'sku', '') or '',
                        'cantidad': int(dp.stock),
                    })

            # Soft delete del DTE: lo saca del listado y de la cuadratura.
            # `estado_dte='ANULADO'` reforzaría el filtro existente en otros
            # módulos que aún no respetan `descartado`.
            dte.descartado = True
            dte.fecha_descarte = timezone.now()
            dte.descartado_por = responsable[:100]
            dte.motivo_descarte = (
                motivo[:200]
                if motivo
                else f'Eliminado desde gestión de documentos por {responsable}'
            )[:200]
            dte.estado_dte = 'ANULADO'
            dte.save(update_fields=[
                'descartado',
                'fecha_descarte',
                'descartado_por',
                'motivo_descarte',
                'estado_dte',
            ])

        return JsonResponse({
            'success': True,
            'message': (
                f'DTE #{dte.numero_documento} eliminado. '
                f'Stock devuelto: {len(stock_devuelto)} línea(s).'
            ),
            'documento': {
                'id': dte.id,
                'tipo_documento': dte.tipo_documento,
                'numero_documento': dte.numero_documento,
            },
            'ticket_anulado': bool(ticket_vinculado),
            'ticket_id': ticket_vinculado.id if ticket_vinculado else None,
            'movimientos_creados': movimientos_creados,
            'stock_devuelto': stock_devuelto,
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar documento: {str(e)}'
        })


# Mapa `Dte.tipo_documento` -> `Ticket.tipo_dte` (choices del modelo Ticket).
# Se usa para (a) resolver el ticket vinculado sin ambigüedad y (b) propagar
# el cambio de tipo al ticket. Sin esto, la búsqueda del ticket sólo miraba
# sucursal + folio, y hay 6 pares (sucursal, folio_dte) con más de un ticket
# en producción: `.first()` sin `order_by` podía tomar cualquiera de ellos.
_TIPO_DTE_A_TIPO_TICKET = {
    'BOLETA ELECTRONICA': 'BOLETA_ELECTRONICA',
    'BOLETA PAPEL': 'BOLETA',
    'FACTURA ELECTRONICA': 'FACTURA_ELECTRONICA',
    'FACTURA EXENTA': 'FACTURA_EXENTA',
}


def _validar_folio_destino_dte(dte, tipo_destino, folio):
    """Valida que `folio` esté libre para el RUT del emisor + `tipo_destino`.

    Devuelve `(ok, mensaje_error, advertencias)`.

    El chequeo que había antes filtraba por **sucursal**, y ése es su punto
    ciego: dos sucursales que operan bajo el mismo RUT pueden emitir el mismo
    folio del mismo tipo de documento sin que la validación lo note. El SII
    asigna los folios por RUT + tipo de documento (CAF), no por sucursal, así
    que la unicidad hay que buscarla sobre todas las fichas de `Empresa` que
    comparten el RUT del emisor.
    """
    from .utils_folio_dte import validar_folio_dte, empresas_con_mismo_rut

    tipo_actual = (dte.tipo_documento or '').upper().strip()
    tipo_destino = (tipo_destino or '').upper().strip()

    if tipo_destino == tipo_actual:
        # Mismo tipo: el helper además compara contra el rango del
        # `Correlativo` de la sucursal y devuelve advertencias informativas.
        resultado = validar_folio_dte(dte, folio)
        return (
            resultado['ok'],
            ' '.join(resultado['errores']),
            resultado['advertencias'],
        )

    emisor_ids = empresas_con_mismo_rut(getattr(dte, 'emisor', None))
    qs = (
        Dte.objects
        .filter(
            tipo_documento=tipo_destino,
            numero_documento=folio,
            descartado=False,
        )
        .exclude(id=dte.id)
        .select_related('sucursal')
    )
    # Sin emisor conocido no se puede razonar por RUT: se conserva el
    # criterio antiguo (misma sucursal) para no dejar el folio sin chequeo.
    qs = (
        qs.filter(emisor_id__in=emisor_ids) if emisor_ids
        else qs.filter(sucursal_id=dte.sucursal_id)
    )
    choques = list(qs[:5])
    if not choques:
        return True, '', []

    detalle = '; '.join(
        'DTE #{id} ({suc}, {fec})'.format(
            id=o.id,
            suc=(getattr(o.sucursal, 'alias', '') or 'SIN SUCURSAL'),
            fec=o.fecha_emision.strftime('%Y-%m-%d') if o.fecha_emision else '?',
        )
        for o in choques
    )
    return False, (
        f'El folio {folio} ya está emitido para {tipo_destino} bajo el RUT '
        f'{getattr(dte.emisor, "rut", "?")}: {detalle}'
    ), []


@login_required
@require_POST
def editar_dte_boleta_papel(request):
    """
    Edita campos puntuales de un DTE aplicando permisos granulares.

    Cada campo editable está protegido por DOS permisos que deben ser
    `puede_editar=True` simultáneamente:

      1. Permiso del campo:
           - fecha            -> `dte_editar_fecha`
           - numero_documento -> `dte_editar_numero`
           - pago             -> `dte_editar_pago`
           - vendedor         -> `dte_editar_vendedor`

      2. Permiso del tipo de DTE:
           - BOLETA ELECTRONICA   -> `dte_editar_tipo_boleta_electronica`
           - BOLETA PAPEL         -> `dte_editar_tipo_boleta_papel`
           - FACTURA ELECTRONICA  -> `dte_editar_tipo_factura_electronica`
           - FACTURA EXENTA       -> `dte_editar_tipo_factura_exenta`

    Sólo se actualizan los campos presentes en el body Y para los que
    el usuario tenga ambos permisos. El resto se ignora en silencio.

    Body esperado::

        {
          "documento_id": 123,
          "numero_documento": 456,               // opcional
          "fecha_emision": "2026-04-17",         // opcional
          "pagos": [                             // opcional
             {"id": 10, "metodo_pago": "EFECTIVO", "monto": 50000},
             {"id": 11, "metodo_pago": "TRANSFERENCIA", "monto": 30000}
          ]
        }

    Por compatibilidad, el nombre `editar_dte_boleta_papel` se mantiene
    (la URL es usada por el template). Internamente ya es genérico.
    """
    try:
        data = json.loads(request.body)
        documento_id = data.get('documento_id')

        if not documento_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de documento requerido'
            })

        tiene_numero = 'numero_documento' in data and data.get('numero_documento') is not None
        tiene_fecha = 'fecha_emision' in data and data.get('fecha_emision') is not None
        tiene_pagos = 'pagos' in data and data.get('pagos') is not None
        tiene_tipo = 'tipo_documento' in data and data.get('tipo_documento') is not None
        # vendedor_id puede llegar como null/"" para "no tocar"; sólo se
        # considera "cambio solicitado" si trae un valor entero válido.
        tiene_vendedor = (
            'vendedor_id' in data
            and data.get('vendedor_id') is not None
            and str(data.get('vendedor_id')).strip() != ''
        )

        if not (tiene_numero or tiene_fecha or tiene_pagos or tiene_tipo or tiene_vendedor):
            return JsonResponse({
                'success': False,
                'error': 'No se enviaron campos para editar'
            })

        # Motivo declarado por el usuario. El modal ya lo exige (mínimo 5
        # caracteres, ver `confirmarEdicionDte` en gestionVentasDocumentos.html)
        # pero el endpoint nunca lo leía: cambiar folio, fecha o tipo de un
        # documento tributario no dejaba ningún rastro
        # (`HistorialCambioFolioDte` tiene 0 filas en producción).
        motivo = str(data.get('motivo') or '').strip()
        if (tiene_numero or tiene_fecha or tiene_tipo) and len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'error': (
                    'Debe indicar un motivo (mínimo 5 caracteres) para editar '
                    'el folio, la fecha o el tipo de un documento tributario'
                )
            }, status=400)

        # Parse / validaciones básicas antes de tocar DB.
        nuevo_numero = None
        if tiene_numero:
            try:
                nuevo_numero = int(data.get('numero_documento'))
            except (TypeError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Número de documento inválido'
                })
            if nuevo_numero <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'El número de documento debe ser un entero positivo'
                })

        fecha_parsed = None
        if tiene_fecha:
            from datetime import datetime as _dt
            try:
                fecha_parsed = _dt.strptime(
                    str(data.get('fecha_emision')).strip(), '%Y-%m-%d'
                ).date()
            except (TypeError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Fecha inválida. Formato esperado YYYY-MM-DD'
                })
            # Una fecha de emisión futura descuadra la caja del día (el
            # ticket se reimputa a `fecha_emision`) y no tiene sentido
            # tributario. Hoy hay 0 casos en producción: la guarda no
            # invalida nada existente.
            if fecha_parsed > timezone.localdate():
                return JsonResponse({
                    'success': False,
                    'error': 'La fecha de emisión no puede ser futura'
                })

        nuevo_tipo = None
        if tiene_tipo:
            nuevo_tipo = str(data.get('tipo_documento') or '').upper().strip()
            if nuevo_tipo not in CODIGO_PERMISO_TIPO_DTE:
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo de documento destino no soportado: {nuevo_tipo}'
                })

        nuevo_vendedor_id = None
        if tiene_vendedor:
            try:
                nuevo_vendedor_id = int(data.get('vendedor_id'))
            except (TypeError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'vendedor_id inválido'
                })
            if nuevo_vendedor_id <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'vendedor_id debe ser un entero positivo'
                })

        pagos_payload = []
        if tiene_pagos:
            pagos_raw = data.get('pagos') or []
            if not isinstance(pagos_raw, list):
                return JsonResponse({
                    'success': False,
                    'error': '`pagos` debe ser una lista'
                })
            metodos_validos = {c for c, _ in METODO_PAGO_TICKET_CHOICES}
            # Para VENTA_INTERNET la plataforma se persiste en
            # `Dte_Detalle_Pago.tipo_tarjeta` y el N° de pedido en `voucher`.
            # Se exigen ambos para mantener la trazabilidad operativa
            # (módulo de cuadratura clasifica por plataforma).
            for idx, item in enumerate(pagos_raw):
                if not isinstance(item, dict):
                    return JsonResponse({
                        'success': False,
                        'error': f'Pago #{idx + 1} inválido'
                    })
                try:
                    pago_id = int(item.get('id'))
                    monto = int(item.get('monto'))
                except (TypeError, ValueError):
                    return JsonResponse({
                        'success': False,
                        'error': f'Pago #{idx + 1}: id o monto inválido'
                    })
                metodo = str(item.get('metodo_pago') or '').strip().upper()
                if metodo not in metodos_validos:
                    return JsonResponse({
                        'success': False,
                        'error': f'Pago #{idx + 1}: método de pago inválido ({metodo})'
                    })
                if monto < 0:
                    return JsonResponse({
                        'success': False,
                        'error': f'Pago #{idx + 1}: monto no puede ser negativo'
                    })

                # `tipo_tarjeta` y `voucher` son opcionales para la mayoría
                # de los métodos pero obligatorios cuando es VENTA_INTERNET
                # (plataforma + N° de pedido). Si no vienen en el payload
                # los dejamos en None → "no tocar" al aplicar.
                tipo_tarjeta_in = item.get('tipo_tarjeta', None)
                voucher_in = item.get('voucher', None)
                tipo_tarjeta_val = (
                    str(tipo_tarjeta_in).strip()
                    if tipo_tarjeta_in is not None else None
                )
                voucher_val = (
                    str(voucher_in).strip()
                    if voucher_in is not None else None
                )

                if metodo == 'VENTA_INTERNET':
                    if not tipo_tarjeta_val:
                        return JsonResponse({
                            'success': False,
                            'error': (
                                f'Pago #{idx + 1}: Venta por Internet requiere '
                                'plataforma (Mercado Pago, Falabella, Paris…)'
                            )
                        })
                    if not voucher_val:
                        return JsonResponse({
                            'success': False,
                            'error': (
                                f'Pago #{idx + 1}: Venta por Internet requiere '
                                'N° de pedido / voucher'
                            )
                        })

                pagos_payload.append({
                    'id': pago_id,
                    'metodo_pago': metodo,
                    'monto': monto,
                    'tipo_tarjeta': tipo_tarjeta_val,
                    'voucher': voucher_val,
                })

        sucursal_id_sesion = get_sucursal_id(request)

        with transaction.atomic():
            # Scoping: el endpoint recibía sólo `documento_id` y cargaba
            # CUALQUIER Dte por id (IDOR). Se acota al mismo universo que
            # muestra el listado de esta pantalla (`listar_documentos_ventas`):
            # ventas no descartadas de la sucursal activa.
            # Ojo: NO agregar `select_related` de FKs nullables aquí —
            # PostgreSQL rechaza `FOR UPDATE` sobre el lado nullable de un
            # OUTER JOIN ("FOR UPDATE cannot be applied to the nullable side
            # of an outer join") y el endpoint completo devuelve error.
            dte = (
                Dte.objects
                .select_for_update()
                .filter(
                    id=documento_id,
                    tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
                    descartado=False,
                )
                .first()
            )
            if not dte:
                return JsonResponse({
                    'success': False,
                    'error': 'Documento no encontrado'
                })

            if (
                sucursal_id_sesion
                and dte.sucursal_id
                and str(dte.sucursal_id) != str(sucursal_id_sesion)
            ):
                return JsonResponse({
                    'success': False,
                    'error': 'El documento no pertenece a la sucursal activa'
                }, status=403)

            if dte.estado_dte == 'ANULADO':
                return JsonResponse({
                    'success': False,
                    'error': 'No se puede editar un documento anulado'
                })

            # Guarda tributaria: si el documento ya tiene una nota de crédito
            # vigente que lo referencia, cambiarle el folio o el tipo rompe esa
            # referencia (la NC guarda el folio en su JSON `referencias` y en
            # `motivo_nc`, y su TXT ya se envió al SII). En producción hay 452
            # documentos en esta situación. La fecha, el vendedor y los pagos
            # sí se pueden seguir corrigiendo.
            if tiene_numero or tiene_tipo:
                ncs_vigentes = list(
                    Dte.objects
                    .filter(documento_afectado_id=dte.id, es_nota_credito=True)
                    .exclude(estado_dte='ANULADO')
                    .values_list('numero_documento', flat=True)[:5]
                )
                if ncs_vigentes:
                    detalle_nc = ', '.join(str(n) for n in ncs_vigentes)
                    return JsonResponse({
                        'success': False,
                        'error': (
                            'No se puede cambiar el folio ni el tipo: el '
                            f'documento tiene nota(s) de crédito asociada(s) '
                            f'({detalle_nc}) que lo referencian. Anule primero '
                            'la NC o corrija por la vía tributaria.'
                        )
                    }, status=409)

            # El tipo de DTE debe ser uno de los reconocidos por la matriz
            # de permisos (si no, no hay forma de autorizar el cambio).
            if (dte.tipo_documento or '').upper() not in CODIGO_PERMISO_TIPO_DTE:
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo de documento no editable: {dte.tipo_documento}'
                }, status=403)

            # Validar permisos campo + tipo para cada cambio solicitado.
            #
            # El rol `administrador` salta la matriz de permisos granulares
            # SÓLO en los campos sin efecto tributario (`pago` y `vendedor`),
            # que es donde el bypass hacía falta cuando las migraciones de
            # permisos (0140 / 0151) no estaban aplicadas.
            #
            # Para `numero_documento` y `fecha` se exige el permiso real: son
            # los campos que alteran la identidad del documento ante el SII y
            # el frontend ya los habilita/deshabilita con ese mismo permiso
            # (`puede_editar_numero_dte` / `puede_editar_fecha_dte` en el
            # contexto de `gestion_ventas_documentos`), de modo que el bypass
            # sólo servía para saltarse la matriz por POST directo.
            # Verificado en producción: los 8 administradores activos tienen
            # `puede_editar=True` en `dte_editar_numero`, `dte_editar_fecha` y
            # los 4 `dte_editar_tipo_*`, sin PermisoSucursal ni PermisoUsuario
            # que los recorte, así que el cambio es transparente hoy.
            CAMPOS_CON_BYPASS_ADMIN = {'pago', 'vendedor'}
            errores_permiso = []
            es_admin_request = (
                getattr(request.user, 'rol', '') == 'administrador'
            )

            def _check(campo):
                if es_admin_request and campo in CAMPOS_CON_BYPASS_ADMIN:
                    return
                if not puede_editar_campo_dte(
                    request.user, campo, dte.tipo_documento,
                    sucursal_id=sucursal_id_sesion,
                ):
                    errores_permiso.append(campo)

            if tiene_numero:
                _check('numero_documento')
            if tiene_fecha:
                _check('fecha')
            if tiene_pagos:
                _check('pago')
            if tiene_vendedor:
                _check('vendedor')

            if errores_permiso:
                labels = {
                    'numero_documento': 'N° documento',
                    'fecha': 'fecha',
                    'pago': 'pagos',
                    'vendedor': 'vendedor',
                }
                campos_txt = ', '.join(labels.get(c, c) for c in errores_permiso)
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'No tiene permisos para editar {campos_txt} en '
                        f'{dte.tipo_documento}'
                    )
                }, status=403)

            # Validar que el vendedor solicitado exista y esté activo.
            # El cambio sólo tiene sentido si difiere del actual; si es
            # el mismo lo silenciamos (sin error) para que un guardado
            # repetido no pinche.
            cambiar_vendedor = False
            vendedor_obj = None
            if tiene_vendedor and nuevo_vendedor_id != (dte.vendedor_id or None):
                vendedor_obj = (
                    Vendedor.objects.filter(id=nuevo_vendedor_id).first()
                )
                if not vendedor_obj:
                    return JsonResponse({
                        'success': False,
                        'error': 'Vendedor no encontrado',
                    }, status=400)
                if not getattr(vendedor_obj, 'activo', True):
                    return JsonResponse({
                        'success': False,
                        'error': (
                            'No se puede asignar un vendedor inactivo '
                            f'({vendedor_obj.nombre or vendedor_obj.codigo_vendedor})'
                        ),
                    }, status=400)
                cambiar_vendedor = True

            # --- Validación de cambio de TIPO (solo pares compatibles) ---
            # Nota: se permite en cualquier estado excepto ANULADO (ya
            # bloqueado arriba). Cuando el DTE está EMITIDO/ACEPTADO el folio
            # origen queda "saltado" y se reserva uno nuevo del correlativo
            # destino; la responsabilidad de la trazabilidad contable es del
            # operador. La UI muestra un warning en estos casos.
            cambiar_tipo = False
            if tiene_tipo:
                tipo_actual = (dte.tipo_documento or '').upper().strip()
                if nuevo_tipo != tipo_actual:
                    if not son_tipos_compatibles(tipo_actual, nuevo_tipo):
                        return JsonResponse({
                            'success': False,
                            'error': (
                                f'No se puede cambiar de {tipo_actual} a '
                                f'{nuevo_tipo}: no son tipos compatibles '
                                '(se permite solo BOLETA ELECTRONICA ↔ '
                                'BOLETA PAPEL).'
                            )
                        }, status=400)
                    if not puede_cambiar_tipo_dte(
                        request.user, tipo_actual, nuevo_tipo,
                        sucursal_id=sucursal_id_sesion,
                    ):
                        return JsonResponse({
                            'success': False,
                            'error': (
                                'No tiene permisos para cambiar el tipo del '
                                f'DTE a {nuevo_tipo}.'
                            )
                        }, status=403)
                    cambiar_tipo = True

            # Validación de folio duplicado por RUT emisor + tipo (NO por
            # sucursal: ver `_validar_folio_destino_dte`). Si también cambia
            # el tipo, la validación se pospone a cuando ya se asignó el folio
            # del nuevo tipo (más abajo).
            advertencias_folio = []
            if tiene_numero and not cambiar_tipo and nuevo_numero != dte.numero_documento:
                ok_folio, error_folio, advertencias_folio = _validar_folio_destino_dte(
                    dte, dte.tipo_documento, nuevo_numero
                )
                if not ok_folio:
                    return JsonResponse({
                        'success': False,
                        'error': error_folio
                    }, status=400)

            # Validación específica de pagos:
            #  - Los ids enviados deben pertenecer al DTE.
            #  - El conjunto enviado debe cubrir TODOS los pagos del DTE
            #    (editamos, no agregamos/eliminamos).
            #  - La suma debe coincidir con `monto_con_iva` para mantener
            #    la cuadratura del documento.
            pagos_actualizar = []
            if tiene_pagos:
                ids_enviados = [p['id'] for p in pagos_payload]
                if len(ids_enviados) != len(set(ids_enviados)):
                    return JsonResponse({
                        'success': False,
                        'error': 'Hay ids de pago duplicados en el body'
                    })

                pagos_existentes = list(
                    Dte_Detalle_Pago.objects.select_for_update()
                    .filter(dte_id=dte.id)
                )
                ids_existentes = {p.id for p in pagos_existentes}
                ids_enviados_set = set(ids_enviados)

                if ids_enviados_set != ids_existentes:
                    return JsonResponse({
                        'success': False,
                        'error': (
                            'Los pagos enviados no coinciden con los del DTE. '
                            'Debe enviar todos los pagos existentes (solo se '
                            'permite editar, no agregar/eliminar).'
                        )
                    })

                suma_pagos = sum(p['monto'] for p in pagos_payload)
                monto_esperado = int(dte.monto_con_iva or 0)
                if monto_esperado > 0 and suma_pagos != monto_esperado:
                    return JsonResponse({
                        'success': False,
                        'error': (
                            f'La suma de los pagos ({suma_pagos:,}) no coincide '
                            f'con el total del DTE ({monto_esperado:,}).'
                        )
                    })

                pagos_por_id = {p.id: p for p in pagos_existentes}
                for item in pagos_payload:
                    obj = pagos_por_id[item['id']]
                    obj.metodo_pago = item['metodo_pago']
                    obj.monto = item['monto']

                    # Si el método final NO es VENTA_INTERNET y el cliente
                    # mandó valores nuevos, los aceptamos tal cual; si
                    # cambió a un método que no usa plataforma/voucher
                    # limpiamos para que la trazabilidad quede consistente
                    # (ej: pasar de VENTA_INTERNET → EFECTIVO no debería
                    # dejar la plataforma "Mercado Pago" colgando).
                    if 'tipo_tarjeta' in item:
                        if item['metodo_pago'] == 'VENTA_INTERNET':
                            obj.tipo_tarjeta = item['tipo_tarjeta'] or None
                        elif item['tipo_tarjeta'] is not None:
                            obj.tipo_tarjeta = item['tipo_tarjeta'] or None
                        else:
                            obj.tipo_tarjeta = None
                    if 'voucher' in item:
                        if item['metodo_pago'] == 'VENTA_INTERNET':
                            obj.voucher = item['voucher'] or None
                        elif item['voucher'] is not None:
                            obj.voucher = item['voucher'] or None
                        else:
                            obj.voucher = None
                    pagos_actualizar.append(obj)

            # Aplicar cambios ----------------------------------------------
            numero_anterior = dte.numero_documento
            fecha_anterior = dte.fecha_emision
            tipo_anterior = dte.tipo_documento

            # Localizar el Ticket vinculado ANTES de modificar el DTE, para
            # poder resolverlo aunque cambie numero_documento o tipo_documento.
            # El match es por sucursal + folio_dte (numero original). Si no
            # existe ticket asociado (ej: DTE emitido directo sin ticket
            # previo), ticket_vinculado queda en None y no se propaga nada.
            #
            # La búsqueda era `filter(...).first()` sin `order_by` y sin
            # tipo: en producción hay 6 pares (sucursal, folio_dte) con más de
            # un ticket, y el motor podía devolver cualquiera de ellos (mismo
            # patrón que provocó el bug de la comuna equivocada al facturar).
            # Ahora se prefiere el ticket del tipo equivalente y, a igualdad de
            # condiciones, el de menor id.
            _tipo_ticket_esperado = _TIPO_DTE_A_TIPO_TICKET.get(
                (dte.tipo_documento or '').upper().strip()
            )
            _tickets_qs = (
                Ticket.objects
                .select_for_update()
                .filter(
                    sucursal_id=dte.sucursal_id,
                    folio_dte=dte.numero_documento,
                )
                .order_by('id')
            )
            ticket_vinculado = None
            if _tipo_ticket_esperado:
                ticket_vinculado = _tickets_qs.filter(
                    tipo_dte=_tipo_ticket_esperado
                ).first()
            if ticket_vinculado is None:
                # Fallback al comportamiento histórico (tickets migrados sin
                # `tipo_dte` coherente), pero ya determinístico por id.
                ticket_vinculado = _tickets_qs.first()

            update_fields = []

            # Si cambia el tipo, tomamos un folio nuevo del correlativo del
            # tipo destino (el folio anterior del tipo origen queda como
            # "saltado"). Si además llegó `numero_documento` en el payload,
            # usamos ese (validando duplicado contra el nuevo tipo).
            if cambiar_tipo:
                if tiene_numero:
                    ok_folio, error_folio, advertencias_folio = (
                        _validar_folio_destino_dte(dte, nuevo_tipo, nuevo_numero)
                    )
                    if not ok_folio:
                        return JsonResponse({
                            'success': False,
                            'error': error_folio
                        }, status=400)
                    numero_asignado = nuevo_numero
                else:
                    try:
                        numero_asignado = obtener_siguiente_correlativo(
                            dte.sucursal, nuevo_tipo
                        )
                    except Exception as exc:
                        return JsonResponse({
                            'success': False,
                            'error': (
                                'No se pudo obtener un folio del tipo '
                                f'{nuevo_tipo}: {exc}'
                            )
                        })
                    # El correlativo puede venir desfasado (rangos redefinidos,
                    # documentos cargados a mano, folios reusados entre fichas
                    # del mismo RUT). Sin este chequeo el cambio de tipo podía
                    # generar un duplicado silencioso de folio, que es un
                    # problema tributario y no sólo de datos.
                    ok_folio, error_folio, advertencias_folio = (
                        _validar_folio_destino_dte(dte, nuevo_tipo, numero_asignado)
                    )
                    if not ok_folio:
                        return JsonResponse({
                            'success': False,
                            'error': (
                                f'El correlativo de {nuevo_tipo} entregó el folio '
                                f'{numero_asignado}, que ya está ocupado. '
                                f'{error_folio}'
                            )
                        }, status=409)

                dte.tipo_documento = nuevo_tipo
                dte.numero_documento = numero_asignado
                update_fields.extend(['tipo_documento', 'numero_documento'])
            elif tiene_numero:
                dte.numero_documento = nuevo_numero
                update_fields.append('numero_documento')

            if tiene_fecha:
                dte.fecha_emision = fecha_parsed
                update_fields.append('fecha_emision')
                # Para boletas papel el pago es al contado.
                if (dte.tipo_documento or '').upper() == 'BOLETA PAPEL':
                    dte.fecha_vencimiento = fecha_parsed
                    update_fields.append('fecha_vencimiento')

            # Snapshot del vendedor previo para la auditoría (cambiar el
            # vendedor mueve la comisión de una persona a otra).
            vendedor_anterior = dte.vendedor
            if cambiar_vendedor and vendedor_obj is not None:
                dte.vendedor = vendedor_obj
                update_fields.append('vendedor')

            if update_fields:
                dte.save(update_fields=update_fields)

            for obj in pagos_actualizar:
                obj.save(update_fields=[
                    'metodo_pago', 'monto', 'tipo_tarjeta', 'voucher',
                ])

            # Propagar al Ticket vinculado (si existe) para mantener la
            # consistencia con la cuadratura. Se ejecuta SIEMPRE al guardar,
            # aunque la fecha no haya cambiado, para que el resumen siempre
            # quede alineado con la fecha_emision actual del DTE.
            #
            # Se usa .update() (queryset) en lugar de .save() porque
            # `Ticket.fecha` tiene auto_now=True y un save() lo reescribiria
            # a "hoy" en vez de a la fecha del DTE.
            ticket_sincronizado = False
            ticket_pagos_sincronizados = 0
            ticket_pagos_resync_modo = None  # 'none' | 'update' | 'rebuild'
            if ticket_vinculado:
                ticket_fields = {'fecha': dte.fecha_emision}
                if cambiar_tipo or tiene_numero:
                    ticket_fields['folio_dte'] = dte.numero_documento
                if cambiar_tipo:
                    # El ticket también guarda el tipo de documento; si no se
                    # propaga, queda apuntando al tipo viejo (bug latente:
                    # hoy 0 casos porque el cambio de tipo aún no se ha usado).
                    ticket_fields['tipo_dte'] = _TIPO_DTE_A_TIPO_TICKET.get(
                        (dte.tipo_documento or '').upper().strip(),
                        ticket_vinculado.tipo_dte,
                    )
                Ticket.objects.filter(pk=ticket_vinculado.pk).update(
                    **ticket_fields
                )
                ticket_sincronizado = True

                # Reconciliación idempotente TicketDetallePago ↔ Dte_Detalle_Pago.
                #
                # `_calcular_cuadratura_data` lee los pagos desde el Ticket
                # cuando hay vínculo (Ticket.folio_dte == Dte.numero_documento),
                # así que cualquier cambio en los pagos del DTE tiene que
                # replicarse aquí o la caja queda descuadrada.
                #
                # Se ejecuta SIEMPRE (no sólo cuando llegó `pagos`) para que
                # un guardado con los mismos datos sirva además como
                # reparación de registros desalineados por flujos anteriores.
                dte_pagos_actuales = list(
                    Dte_Detalle_Pago.objects
                    .filter(dte_id=dte.id)
                    .order_by('id')
                )
                ticket_pagos_actuales = list(
                    TicketDetallePago.objects
                    .select_for_update()
                    .filter(ticket_id=ticket_vinculado.pk)
                    .order_by('id')
                )

                def _pago_difiere(dp, tp):
                    return (
                        (dp.metodo_pago or '') != (tp.metodo_pago or '')
                        or int(dp.monto or 0) != int(tp.monto or 0)
                        or (dp.tipo_tarjeta or '') != (tp.tipo_tarjeta or '')
                        or (dp.voucher or '') != (tp.voucher or '')
                    )

                mismo_largo = (
                    len(dte_pagos_actuales) == len(ticket_pagos_actuales)
                )
                hay_drift = (
                    not mismo_largo
                    or any(
                        _pago_difiere(dp, tp)
                        for dp, tp in zip(
                            dte_pagos_actuales, ticket_pagos_actuales
                        )
                    )
                )

                if hay_drift and mismo_largo:
                    # 1:1 por orden de id. Propagamos también
                    # tipo_tarjeta + voucher para que la cuadratura clasifique
                    # correctamente VENTA_INTERNET por plataforma.
                    for dp, tp in zip(
                        dte_pagos_actuales, ticket_pagos_actuales
                    ):
                        tp.metodo_pago = dp.metodo_pago
                        tp.monto = int(dp.monto or 0)
                        tp.tipo_tarjeta = dp.tipo_tarjeta
                        tp.voucher = dp.voucher
                        tp.save(update_fields=[
                            'metodo_pago', 'monto', 'tipo_tarjeta', 'voucher',
                        ])
                        ticket_pagos_sincronizados += 1
                    ticket_pagos_resync_modo = 'update'
                elif hay_drift and tiene_pagos:
                    # Cantidad de pagos distinta: reconstruimos los del
                    # ticket a partir del DTE preservando plataforma y
                    # voucher (clave para VENTA_INTERNET y tarjetas).
                    #
                    # Sólo se reconstruye cuando el usuario editó los pagos
                    # explícitamente. Antes corría en CUALQUIER guardado (por
                    # ejemplo, al corregir sólo el vendedor) y borraba filas de
                    # `TicketDetallePago` cuyo id es la clave de idempotencia
                    # del consumo de gift cards (`consumo:{id}`) y del resto de
                    # los hooks de cobro. En producción hay 34 documentos con
                    # distinta cantidad de pagos que caían aquí sin que nadie
                    # lo pidiera.
                    TicketDetallePago.objects.filter(
                        ticket_id=ticket_vinculado.pk
                    ).delete()
                    for dp in dte_pagos_actuales:
                        TicketDetallePago.objects.create(
                            ticket_id=ticket_vinculado.pk,
                            metodo_pago=dp.metodo_pago,
                            monto=int(dp.monto or 0),
                            tipo_tarjeta=dp.tipo_tarjeta,
                            voucher=dp.voucher,
                        )
                        ticket_pagos_sincronizados += 1
                    ticket_pagos_resync_modo = 'rebuild'
                elif hay_drift:
                    # Drift de cantidad detectado pero el usuario no editó los
                    # pagos: no se toca nada y se informa en la respuesta.
                    ticket_pagos_resync_modo = 'skipped_drift'
                    logger.warning(
                        '[DTE-EDIT] Drift de cantidad de pagos no reconciliado '
                        'DTE #%s (%s pagos) vs Ticket #%s (%s pagos): se omite '
                        'la reconstrucción porque no se editaron los pagos.',
                        dte.id, len(dte_pagos_actuales),
                        ticket_vinculado.pk, len(ticket_pagos_actuales),
                    )
                else:
                    ticket_pagos_resync_modo = 'none'

            # Dejar traza en la bitácora de los arqueos afectados cuando
            # cambia la fecha y el día origen o destino ya tiene arqueo
            # cerrado / con diferencias: sus snapshots de teóricos quedaron
            # desalineados respecto al recálculo en vivo del modal.
            arqueos_afectados_info = []
            cambio_fecha_real = (
                tiene_fecha
                and fecha_anterior
                and fecha_anterior != dte.fecha_emision
            )
            if cambio_fecha_real and dte.sucursal_id:
                fechas_a_revisar = {fecha_anterior, dte.fecha_emision}
                arqueos_afectados_qs = ArqueoCaja.objects.filter(
                    sucursal_id=dte.sucursal_id,
                    fecha_arqueo__in=fechas_a_revisar,
                ).exclude(estado='ABIERTO')

                for arq in arqueos_afectados_qs:
                    ObservacionArqueo.objects.create(
                        arqueo=arq,
                        usuario=request.user,
                        tipo='SISTEMA',
                        texto=(
                            f'DTE #{dte.numero_documento} '
                            f'({dte.tipo_documento}) editado desde gestión '
                            f'de documentos. Fecha: '
                            f'{fecha_anterior.strftime("%d/%m/%Y")} → '
                            f'{dte.fecha_emision.strftime("%d/%m/%Y")}. '
                            'Los teóricos guardados del arqueo pueden no '
                            'coincidir con el recálculo en vivo.'
                        ),
                        visible_para_cajera=True,
                    )
                    arqueos_afectados_info.append({
                        'id': arq.id,
                        'fecha': arq.fecha_arqueo.strftime('%Y-%m-%d'),
                        'estado': arq.estado,
                    })

            # ================= AUDITORÍA =================
            # Hasta ahora este endpoint cambiaba folio, fecha, tipo, vendedor
            # y pagos de documentos tributarios ya emitidos sin dejar ningún
            # registro: no se podía saber quién editó qué ni por qué.
            from .utils_folio_dte import (
                registrar_cambio_folio_dte,
                registrar_cambio_vendedor_dte,
                reemplazar_folio_en_texto,
            )

            folio_cambio_real = (
                numero_anterior is not None
                and numero_anterior != dte.numero_documento
            )

            movimientos_actualizados = 0
            if folio_cambio_real:
                # Las observaciones del kardex citan el folio del documento.
                # Se reemplaza SOLO el token completo: un `str.replace()` a
                # secas corrompe el texto cuando el folio es corto (el folio 1
                # convierte '10400190076' en un número inventado).
                for mov in Movimientos_Producto.objects.filter(dte=dte).exclude(
                    observaciones__isnull=True
                ).exclude(observaciones=''):
                    obs_nueva = reemplazar_folio_en_texto(
                        mov.observaciones, numero_anterior, dte.numero_documento
                    )
                    if obs_nueva != mov.observaciones:
                        mov.observaciones = obs_nueva
                        mov.save(update_fields=['observaciones'])
                        movimientos_actualizados += 1

                registrar_cambio_folio_dte(
                    dte, numero_anterior, dte.numero_documento,
                    motivo=(
                        f'{motivo} [editar_dte_boleta_papel'
                        + (f'; tipo {tipo_anterior} -> {dte.tipo_documento}'
                           if cambiar_tipo else '')
                        + (f'; estado SII {dte.estado_dte}'
                           if dte.estado_dte == 'ACEPTADO' else '')
                        + f'; movimientos actualizados: {movimientos_actualizados}]'
                    ),
                    request=request,
                )

            if cambiar_vendedor and vendedor_obj is not None:
                registrar_cambio_vendedor_dte(
                    dte, vendedor_anterior, vendedor_obj,
                    request=request, motivo=motivo,
                )

            # Nota: se usa '->' y no la flecha unicode porque el handler de
            # consola en Windows (cp1252) no puede codificarla y ensucia el
            # log con un UnicodeEncodeError por cada edición.
            cambios_log = []
            if folio_cambio_real:
                cambios_log.append(
                    f'folio {numero_anterior} -> {dte.numero_documento}'
                )
            if cambiar_tipo:
                cambios_log.append(f'tipo {tipo_anterior} -> {dte.tipo_documento}')
            if tiene_fecha and fecha_anterior != dte.fecha_emision:
                cambios_log.append(
                    f'fecha {fecha_anterior} -> {dte.fecha_emision}'
                )
            if pagos_actualizar:
                cambios_log.append(f'{len(pagos_actualizar)} pago(s)')
            if cambiar_vendedor:
                cambios_log.append('vendedor')
            if cambios_log:
                logger.info(
                    '[DTE-EDIT] %s #%s (id=%s, sucursal=%s, estado=%s) editado '
                    'por %s: %s | motivo: %s',
                    dte.tipo_documento, dte.numero_documento, dte.id,
                    dte.sucursal_id, dte.estado_dte,
                    getattr(request.user, 'username', '?'),
                    '; '.join(cambios_log), motivo or '(sin motivo)',
                )

            # El documento ya declarado al SII se puede seguir corrigiendo
            # (el modal lo advierte en rojo), pero queda avisado en la
            # respuesta y registrado en la auditoría de arriba.
            advertencias_respuesta = list(advertencias_folio or [])
            if dte.estado_dte == 'ACEPTADO' and (
                tiene_numero or tiene_fecha or tiene_tipo
            ):
                advertencias_respuesta.append(
                    'El documento está ACEPTADO por el SII: el cambio NO se '
                    'informa al organismo y el sistema queda distinto de lo '
                    'declarado.'
                )
            if ticket_pagos_resync_modo == 'skipped_drift':
                advertencias_respuesta.append(
                    'Los pagos del ticket vinculado tienen una cantidad '
                    'distinta a los del DTE. No se tocaron: edite los pagos '
                    'desde este modal si quiere reconciliarlos.'
                )

        return JsonResponse({
            'success': True,
            'message': 'Documento actualizado correctamente',
            'advertencias': advertencias_respuesta,
            'documento': {
                'id': dte.id,
                'tipo_documento': dte.tipo_documento,
                'numero_documento': dte.numero_documento,
                'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
                'tipo_anterior': tipo_anterior,
                'numero_anterior': numero_anterior,
                'fecha_anterior': (
                    fecha_anterior.strftime('%Y-%m-%d') if fecha_anterior else None
                ),
                'pagos_actualizados': len(pagos_actualizar),
                'tipo_cambiado': cambiar_tipo,
                'vendedor_cambiado': cambiar_vendedor,
                'vendedor_id': dte.vendedor_id,
                'vendedor_nombre': (
                    f"{dte.vendedor.codigo_vendedor} - {dte.vendedor.nombre}"
                    if dte.vendedor else None
                ),
                'ticket_sincronizado': ticket_sincronizado,
                'ticket_id': ticket_vinculado.pk if ticket_vinculado else None,
                'ticket_pagos_sincronizados': ticket_pagos_sincronizados,
                'ticket_pagos_resync_modo': ticket_pagos_resync_modo,
                'arqueos_afectados': arqueos_afectados_info,
                'movimientos_actualizados': movimientos_actualizados,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al editar documento: {str(e)}'
        })


# ========== DTE MANUAL (CUADRATURA INFORMATIVA) ==========

# Tipos de DTE permitidos para creación manual desde Gestión de
# Documentos. No incluye FACTURA EXENTA ni NOTA DE CREDITO porque
# estos requieren reglas de negocio adicionales (servicios sin IVA,
# documento afectado, motivo, etc.) que escapan al alcance del
# "DTE informativo para cuadratura".
TIPOS_DTE_MANUAL_PERMITIDOS = (
    'BOLETA ELECTRONICA',
    'BOLETA PAPEL',
    'FACTURA ELECTRONICA',
)


@login_required
@require_POST
def crear_dte_manual(request):
    """Crea un DTE manual (sin productos) para que figure en cuadratura
    y reportes de ventas.

    Pensado para registrar boletas/facturas que se emitieron fuera del
    sistema (boleta papel a mano, factura externa, etc.) y necesitan
    aparecer en `/app/ventas/documentos/`, en el resumen de caja y en
    `/app/reportes/ventas-sucursal/`. NO genera movimientos de stock,
    NO emite TXT a Acepta y NO consume correlativos del talonario
    electrónico: el operador digita el folio que corresponda.

    Body esperado::

        {
          "tipo_documento": "BOLETA ELECTRONICA" | "BOLETA PAPEL" | "FACTURA ELECTRONICA",
          "numero_documento": 12345,
          "fecha_emision": "2026-04-30",        // opcional, default = hoy Chile
          "monto_total": 10000,
          "vendedor_id": 7,
          "metodo_pago": "EFECTIVO",
          "tipo_tarjeta": "...",                 // opcional
          "voucher": "...",                      // opcional (obligatorio en VENTA_INTERNET)
          "referencias": "..."                   // opcional, texto libre
        }

    Sólo administradores (`request.user.rol == 'administrador'`).
    """
    from decimal import Decimal

    # Restricción de rol: solo administradores. La pantalla ya esconde el
    # botón cuando no corresponde, pero validamos también del lado servidor.
    if getattr(request.user, 'rol', '') != 'administrador':
        return JsonResponse({
            'success': False,
            'error': 'Solo los administradores pueden crear DTEs manuales'
        }, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        }, status=400)

    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal seleccionada en la sesión'
        }, status=400)

    tipo_documento = (data.get('tipo_documento') or '').strip().upper()
    if tipo_documento not in TIPOS_DTE_MANUAL_PERMITIDOS:
        return JsonResponse({
            'success': False,
            'error': (
                f'Tipo de documento no permitido para DTE manual: {tipo_documento}. '
                f'Permitidos: {", ".join(TIPOS_DTE_MANUAL_PERMITIDOS)}'
            )
        }, status=400)

    try:
        numero_documento = int(data.get('numero_documento'))
    except (TypeError, ValueError):
        return JsonResponse({
            'success': False,
            'error': 'Número de documento inválido'
        }, status=400)
    if numero_documento <= 0:
        return JsonResponse({
            'success': False,
            'error': 'El número de documento debe ser un entero positivo'
        }, status=400)

    fecha_raw = (data.get('fecha_emision') or '').strip()
    if fecha_raw:
        try:
            fecha_emision = datetime.strptime(fecha_raw, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Fecha inválida. Formato esperado YYYY-MM-DD'
            }, status=400)
    else:
        # Default: hoy en zona horaria Chile (regla timezone-chile).
        fecha_emision = timezone.localdate()

    try:
        monto_total = int(data.get('monto_total'))
    except (TypeError, ValueError):
        return JsonResponse({
            'success': False,
            'error': 'Monto total inválido'
        }, status=400)
    if monto_total <= 0:
        return JsonResponse({
            'success': False,
            'error': 'El monto total debe ser mayor a 0'
        }, status=400)

    vendedor_id_raw = data.get('vendedor_id')
    try:
        vendedor_id = int(vendedor_id_raw) if vendedor_id_raw is not None else None
    except (TypeError, ValueError):
        vendedor_id = None
    if not vendedor_id:
        return JsonResponse({
            'success': False,
            'error': 'Debe seleccionar un vendedor'
        }, status=400)

    metodo_pago = (data.get('metodo_pago') or '').strip().upper()
    metodos_validos = {c for c, _ in METODO_PAGO_TICKET_CHOICES}
    if metodo_pago not in metodos_validos:
        return JsonResponse({
            'success': False,
            'error': f'Método de pago inválido: {metodo_pago}'
        }, status=400)

    tipo_tarjeta = (data.get('tipo_tarjeta') or '').strip() or None
    voucher = (data.get('voucher') or '').strip() or None
    referencias = (data.get('referencias') or '').strip()

    # VENTA_INTERNET: misma regla de validación que `editar_dte_boleta_papel`
    # (se usa el campo `tipo_tarjeta` para la plataforma y `voucher` para el
    # N° de pedido). Sin estos datos la cuadratura no puede clasificar el
    # ingreso por plataforma (Falabella, Mercado Pago, etc.).
    if metodo_pago == 'VENTA_INTERNET':
        if not tipo_tarjeta:
            return JsonResponse({
                'success': False,
                'error': 'Venta por Internet requiere plataforma (Mercado Pago, Falabella, Paris…)'
            }, status=400)
        if not voucher:
            return JsonResponse({
                'success': False,
                'error': 'Venta por Internet requiere N° de pedido / voucher'
            }, status=400)

    try:
        sucursal = Sucursal.objects.select_related('empresa').get(id=sucursal_id)
    except Sucursal.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Sucursal de la sesión no encontrada'
        }, status=400)

    if not sucursal.empresa_id:
        return JsonResponse({
            'success': False,
            'error': 'La sucursal no tiene empresa asociada (emisor del DTE)'
        }, status=400)

    try:
        vendedor = Vendedor.objects.get(id=vendedor_id)
    except Vendedor.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vendedor no encontrado'
        }, status=400)

    # Folio único por (sucursal, tipo_documento). Mismo invariante que
    # `editar_dte_boleta_papel` para evitar pisar un DTE real ya emitido.
    if Dte.objects.filter(
        sucursal_id=sucursal_id,
        tipo_documento=tipo_documento,
        numero_documento=numero_documento,
    ).exists():
        return JsonResponse({
            'success': False,
            'error': (
                f'Ya existe un {tipo_documento} con el número '
                f'{numero_documento} en esta sucursal'
            )
        }, status=400)

    # Cálculo neto/IVA: las boletas guardan total IVA-inclusive y derivan
    # el neto dividiendo por 1.19. Las facturas guardan neto y agregan
    # 19% de IVA al total. Replica la misma lógica que
    # `_generar_dte_desde_ticket` para mantener consistencia con DTEs
    # generados desde tickets (cuadratura usa `monto_con_iva`).
    es_boleta = tipo_documento in ('BOLETA ELECTRONICA', 'BOLETA PAPEL')
    total_dec = Decimal(monto_total)
    if es_boleta:
        monto_con_iva = total_dec
        monto_neto = (total_dec / Decimal('1.19')).quantize(Decimal('1'))
    else:
        monto_neto = total_dec
        iva = (total_dec * Decimal('0.19')).quantize(Decimal('1'))
        monto_con_iva = total_dec + iva

    referencias_final = referencias or 'DTE MANUAL'

    try:
        with transaction.atomic():
            dte = Dte.objects.create(
                emisor=sucursal.empresa,
                receptor=None,
                numero_documento=numero_documento,
                tipo_documento=tipo_documento,
                monto_con_iva=monto_con_iva,
                monto_neto=monto_neto,
                estado_pago='PAGADO',
                estado_dte='EMITIDO',
                responsable=request.user.username or '',
                fecha_emision=fecha_emision,
                fecha_vencimiento=fecha_emision,
                diasCredito=0,
                bultos=0,
                unidades_productos=0,
                vendedor=vendedor,
                descuento=0,
                sucursal=sucursal,
                hora=timezone.localtime().time(),
                tipo_transaccion='VENTA_PUBLICO',
                referencias=referencias_final,
                descartado=False,
                es_manual=True,
            )

            Dte_Detalle_Pago.objects.create(
                dte=dte,
                metodo_pago=metodo_pago,
                tipo_tarjeta=tipo_tarjeta,
                voucher=voucher,
                monto=int(monto_total),
                notas='DTE manual (cuadratura informativa)',
            )

        return JsonResponse({
            'success': True,
            'message': 'DTE manual creado correctamente',
            'documento': {
                'id': dte.id,
                'tipo_documento': dte.tipo_documento,
                'numero_documento': dte.numero_documento,
                'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
                'monto_con_iva': int(dte.monto_con_iva or 0),
                'monto_neto': int(dte.monto_neto or 0),
                'metodo_pago': metodo_pago,
                'es_manual': True,
            },
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear DTE manual: {str(e)}'
        }, status=500)


# ========== CUADRATURA Y ARQUEO DE CAJA ==========

@login_required
def revision_arqueos(request):
    """Vista de supervisión: revisión de arqueos, comprobantes bancarios, depósitos."""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            pass
    if not sucursal_actual:
        return redirect('dashboard')

    rol_usuario = getattr(request.user, 'rol', None)
    es_supervisor = rol_usuario in ['administrador', 'administracion']
    if not es_supervisor:
        return redirect('cuadratura_caja')

    return render(request, 'vistas/modulo_ventas/revisionArqueos.html', {
        'sucursal_actual': sucursal_actual,
    })


@login_required
def cuadratura_caja(request):
    """Vista principal para cuadratura y arqueo de caja"""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    if not sucursal_actual:
        return redirect('dashboard')
    
    # Obtener rol del usuario
    rol_usuario = getattr(request.user, 'rol', None)
    
    # Verificar si el usuario es administrador (para permisos de reabrir arqueos)
    es_administrador = rol_usuario == 'administrador'

    # Verificar si tiene permisos de supervisión (administrador o administración)
    es_supervisor = rol_usuario in ['administrador', 'administracion']

    # Cajero/vendedor/jefe_local: puede declarar depósitos pero no confirmarlos
    es_cajero = not es_supervisor and rol_usuario in ['cajero', 'vendedor', 'jefe_local']

    # Permiso de reabrir: administrador (siempre) o jefe_local/administracion (con tolerancia)
    puede_reabrir = rol_usuario in ['administrador', 'jefe_local', 'administracion']

    # Rango configurable para crear arqueos históricos
    config_rango_arqueo = obtener_configuracion_rango_arqueo(rol_usuario)
    dias_tolerancia_arqueo = config_rango_arqueo['dias_equivalentes']

    # Permisos granulares de edición de DTE (3 campos x 4 tipos).
    # Se reutiliza el mismo helper que usa Gestión de Documentos para que
    # el modal "Detalle de Métodos de Pago" del Resumen de Caja muestre
    # u oculte los controles de edición de fecha por cada DTE asociado.
    permisos_dte = permisos_edicion_dte_context(request.user, sucursal_actual_id)

    context = {
        'sucursal_actual': sucursal_actual,
        'es_administrador': es_administrador,
        'es_supervisor': es_supervisor,
        'es_cajero': es_cajero,
        'puede_reabrir': puede_reabrir,
        'rol_usuario': rol_usuario or 'sin_rol',
        'dias_tolerancia_arqueo': dias_tolerancia_arqueo,
        'rango_arqueo_tipo': config_rango_arqueo['tipo'],
        'rango_arqueo_valor': config_rango_arqueo['valor'],
        'rango_arqueo_label': config_rango_arqueo['label'],
        'fecha_minima_arqueo': config_rango_arqueo['fecha_minima'].strftime('%Y-%m-%d'),
        'qz_config': _get_qz_config(sucursal_actual_id),
        # Flags por campo (se evalúan además por tipo de DTE en runtime).
        'puede_editar_fecha_dte': permisos_dte['campo']['fecha'],
        'puede_editar_numero_dte': permisos_dte['campo']['numero_documento'],
        'puede_editar_pago_dte': permisos_dte['campo']['pago'],
        # Flag global: ¿hay algún par (campo, tipo) habilitado? → controla
        # si se muestran los botones de "Ver detalle" / "Editar fecha" en
        # el modal de Resumen de Caja.
        'puede_editar_algun_dte': permisos_dte['cualquiera'],
    }
    return render(request, 'vistas/modulo_ventas/cuadraturaCaja.html', context)


# Métodos con los que se graba el Dte_Detalle_Pago de una NC de devolución
# cuando la venta original fue A CRÉDITO (ver `METODO_PAGO_NC_POR_DG` en
# devolucion_garantia_service). No sacan plata de la caja: rebajan la cuenta
# por cobrar, así que descuentan de `total_credito_externo` y no del efectivo.
_METODOS_PAGO_CREDITO_NC = {'CREDITO_EXTERNO', 'CREDITO_TRABAJADOR', 'CONVENIO', 'ORDEN_COMPRA'}


def _calcular_cuadratura_data(sucursal, fecha_str):
    """
    Función helper para calcular datos de cuadratura.
    Puede ser usada tanto por el endpoint POST como por el exportador Excel.
    
    Args:
        sucursal: Instancia de Sucursal
        fecha_str: Fecha en formato 'YYYY-MM-DD'
    
    Returns:
        dict: Datos de cuadratura calculados
    """
    from datetime import datetime
    from datetime import time as dt_time
    
    # Convertir fecha string a date object
    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    
    # Crear datetime para filtros con timezone aware
    inicio_dia = timezone.make_aware(datetime.combine(fecha_obj, dt_time.min))
    fin_dia = timezone.make_aware(datetime.combine(fecha_obj, dt_time.max))
    
    # Inicializar totales
    cuadratura_data = {
        'fecha_cuadratura': fecha_str,
        'total_efectivo': 0,
        'total_tarjeta_debito': 0,
        'total_tarjeta_credito': 0,
        'total_transbank': 0,
        # Tarjetas Comerciales (Hites, Presto)
        'total_hites': 0,
        'total_presto': 0,
        'total_tarjetas_comerciales': 0,
        # Venta Internet (Falabella, Paris, Ripley, MercadoPago, Klap)
        'total_falabella': 0,
        'total_paris': 0,
        'total_ripley': 0,
        'total_mercadopago': 0,
        'total_klap': 0,
        'total_venta_internet': 0,
        # Otros
        'total_transferencia': 0,
        'total_cheque': 0,
        'total_convenio': 0,
        'total_credito_trabajador': 0,
        'total_credito_externo': 0,
        'total_orden_compra': 0,
        'total_nota_credito': 0,
        'total_descuentos': 0,  # Descuentos aplicados
        'total_descuento_puntos': 0,  # Descuentos por canje de puntos de fidelización
        # Descuentos por cupón nominativo. Se cuenta aparte de los puntos porque
        # el pasivo es distinto: el vale consume saldo que el cliente ya ganó, el
        # cupón es margen que la empresa regala.
        # OJO: todavía NO se persiste en el modelo Cuadratura (no hay campo
        # `total_descuento_cupones_teorico`); se ve en la cuadratura del día.
        'total_descuento_cupones': 0,
        # Documentos
        'total_tickets': 0,
        'total_boletas': 0,
        'total_boletas_electronicas': 0,
        'total_boletas_papel': 0,
        'total_facturas': 0,
        'total_facturas_exentas': 0,
        'total_notas_credito': 0,
        'total_nc_efectivo': 0,
        'total_nc_transferencia': 0,
        # NC de devolución que rebajan la cuenta por cobrar en vez de sacar
        # plata de un medio de caja (ventas a crédito). Descuentan de
        # `total_credito_externo`, no del efectivo ni de las transferencias.
        'total_nc_credito': 0,
        # Totales "display" (bruto por tipo de documento) — suman TODOS los
        # DTEs del día, incluso los que tienen ticket asociado. Sirven para
        # mostrar en la sección "Documentos (referencia)" del Resumen de
        # Caja, sin afectar la lógica de cuadratura (que sigue usando los
        # `total_*` deduplicados para no sumar dos veces el mismo monto).
        'total_boletas_electronicas_display': 0,
        'total_boletas_papel_display': 0,
        'total_facturas_display': 0,
        'total_facturas_exentas_display': 0,
        'cantidad_notas_credito': 0,
        'cantidad_tickets': 0,
        'cantidad_boletas': 0,
        'cantidad_boletas_electronicas': 0,
        'cantidad_boletas_papel': 0,
        'cantidad_facturas': 0,
        'cantidad_facturas_exentas': 0,
        'venta_total': 0,
    }
    
    # ========== PROCESAR TICKETS ==========
    # Usar campo `fecha` (DateField, auto_now) en vez de `created_at` (DateTimeField)
    # porque `fecha` se actualiza al pagar el ticket, mientras que `created_at`
    # refleja cuando se creó (posiblemente como PENDIENTE en otro momento).
    tickets_del_dia = Ticket.objects.filter(
        sucursal=sucursal,
        fecha=fecha_obj,
        estado='PAGADO'
    ).prefetch_related('pagos')
    
    for ticket in tickets_del_dia:
        cuadratura_data['total_tickets'] += ticket.total or 0
        cuadratura_data['cantidad_tickets'] += 1
        cuadratura_data['total_descuento_puntos'] += ticket.descuento_fidelizacion or 0
        cuadratura_data['total_descuento_cupones'] += ticket.descuento_cupon or 0

        # Procesar pagos del ticket
        for pago in ticket.pagos.all():
            metodo = pago.metodo_pago
            tipo_tarjeta = (pago.tipo_tarjeta or '').upper()
            monto = pago.monto or 0
            
            if metodo == 'EFECTIVO':
                cuadratura_data['total_efectivo'] += monto
            elif metodo == 'TARJETA_DEBITO':
                # ✅ TARJETA_DEBITO se considera Transbank (datos migrados y genéricos)
                cuadratura_data['total_tarjeta_debito'] += monto
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TARJETA_CREDITO':
                # ✅ TARJETA_CREDITO se considera Transbank (datos migrados y genéricos)
                cuadratura_data['total_tarjeta_credito'] += monto
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TBK_DEBITO_POS':
                # ✅ Transbank POS Débito
                cuadratura_data['total_tarjeta_debito'] += monto
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TBK_CREDITO_POS':
                # ✅ Transbank POS Crédito
                cuadratura_data['total_tarjeta_credito'] += monto
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TBK_PREPAGO_POS':
                # ✅ Transbank POS Prepago (va a débito por convención)
                cuadratura_data['total_tarjeta_debito'] += monto
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TBK_POS_INTEGRADO' or metodo == 'TBK_MANUAL':
                # ✅ Transbank genérico (datos históricos)
                cuadratura_data['total_transbank'] += monto
            elif metodo == 'TRANSFERENCIA':
                cuadratura_data['total_transferencia'] += monto
            elif metodo == 'CHEQUE':
                cuadratura_data['total_cheque'] += monto
            elif metodo == 'CONVENIO':
                cuadratura_data['total_convenio'] += monto
            elif metodo == 'CREDITO_TRABAJADOR':
                cuadratura_data['total_credito_trabajador'] += monto
            elif metodo == 'CREDITO_EXTERNO':
                cuadratura_data['total_credito_externo'] += monto
            elif metodo == 'ORDEN_COMPRA':
                cuadratura_data['total_orden_compra'] += monto
            elif metodo == 'TARJETA_COMERCIAL':
                # Clasificar por tipo_tarjeta
                cuadratura_data['total_tarjetas_comerciales'] += monto
                if 'HITES' in tipo_tarjeta:
                    cuadratura_data['total_hites'] += monto
                elif 'PRESTO' in tipo_tarjeta:
                    cuadratura_data['total_presto'] += monto
            elif metodo == 'VENTA_INTERNET':
                cuadratura_data['total_venta_internet'] += monto
                # ✅ Clasificar por tipo_tarjeta (igual que con DTEs)
                if 'FALABELLA' in tipo_tarjeta or 'WALMART' in tipo_tarjeta:
                    cuadratura_data['total_falabella'] += monto
                elif 'PARIS' in tipo_tarjeta:
                    cuadratura_data['total_paris'] += monto
                elif 'RIPLEY' in tipo_tarjeta:
                    cuadratura_data['total_ripley'] += monto
                elif 'MERCADO' in tipo_tarjeta or 'MERCADOPAGO' in tipo_tarjeta or 'SHOPIFY' in tipo_tarjeta:
                    cuadratura_data['total_mercadopago'] += monto
                elif 'KLAP' in tipo_tarjeta:
                    cuadratura_data['total_klap'] += monto
                else:
                    cuadratura_data['total_mercadopago'] += monto
    
    # ========== PROCESAR DTEs (FACTURAS/BOLETAS ELECTRÓNICAS) ==========
    # Obtener folios de DTEs que ya tienen ticket asociado para evitar duplicar pagos
    folios_tickets = Ticket.objects.filter(
        sucursal=sucursal,
        fecha=fecha_obj,
        folio_dte__isnull=False
    ).values_list('folio_dte', flat=True)
    
    # Incluimos también NCs con tipo_transaccion='ANULACION': aunque la
    # política de negocio es que no descuenten del efectivo teórico, el
    # usuario necesita verlas en el Resumen de Caja como documento emitido
    # del día (si no la NC "desaparece" del día en que se solicitó).
    # La clasificación posterior (si descuenta o no efectivo) sigue dependiendo
    # de si la NC tiene un Dte_Detalle_Pago en EFECTIVO, no del tipo_transaccion.
    # Solo ventas al público (POS / módulo ventas) participan de la caja.
    # `tipo_transaccion='VENTA'` se reserva para DTEs emitidos FUERA del POS
    # —por concepto (sin mercadería), despacho externo, compensación de compra
    # y documentos— que NO son ventas de mesón y no deben aparecer ni sumar al
    # VENTA TOTAL del Resumen de Caja. Las conversiones ticket→factura también
    # quedan como 'VENTA', pero su dinero ya está contabilizado en el ticket
    # (`total_tickets`), así que excluirlas además evita doble conteo.
    # DEVOLUCION/ANULACION se mantienen porque son las NC del día.
    dtes_del_dia = Dte.objects.filter(
        sucursal=sucursal,
        fecha_emision=fecha_obj,
        estado_dte__in=['EMITIDO', 'ACEPTADO'],
        tipo_transaccion__in=['VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
        # Los DTEs marcados como descartados desde la gestión de
        # documentos no participan de la cuadratura (stock ya devuelto,
        # documento sacado del listado).
        descartado=False,
    ).prefetch_related('dte_asociado')
    
    folios_tickets_set = set(folios_tickets)

    for dte in dtes_del_dia:
        monto_dte = dte.monto_con_iva or 0
        tiene_ticket_asociado = dte.numero_documento in folios_tickets_set
        
        # Calcular suma de pagos para detectar descuentos
        suma_pagos_dte = sum((p.monto or 0) for p in dte.dte_asociado.all())
        descuento_dte = max(0, monto_dte - suma_pagos_dte)
        
        # Usar monto real pagado (con descuento aplicado) para cuadratura
        monto_real = suma_pagos_dte if suma_pagos_dte > 0 else monto_dte
        
        # NC: 3 modalidades coordinadas con el modal de gestion-DTE.
        #   - DEVOLUCION  → resta del efectivo/transferencia teorica y del
        #                   venta_total (aparece en cuadratura).
        #   - ANULACION   → modalidad "informativa": cuenta como documento
        #                   emitido del dia pero NO descuenta del venta_total
        #                   ni de los teoricos. Asi el operador ve la NC
        #                   en el resumen sin distorsionar la cuadratura.
        #   - OCULTA (descartado=True) → ya esta filtrada por el query (linea
        #                   `descartado=False`), no llega aqui.
        if dte.tipo_documento == 'NOTA DE CREDITO':
            if dte.tipo_transaccion == 'DEVOLUCION':
                # El efecto de las NC de devolución (conteo, monto y resta de
                # efectivo/transferencia teórico) se imputa por su FECHA DE
                # CUADRATURA = `Dte_Detalle_Pago.fecha_pago`, que puede diferir
                # de `fecha_emision` cuando el operador eligió otro día en el
                # modal de gestión-DTE. Se procesa en el bloque dedicado más
                # abajo; aquí se omite para no duplicar ni anclar el efecto al
                # día de emisión.
                continue
            # ANULACION (informativa): cuenta como documento emitido del día,
            # pero NO suma a total_notas_credito ni resta venta_total/teóricos.
            cuadratura_data['cantidad_notas_credito'] += 1
        elif not tiene_ticket_asociado:
            # Solo contar montos DTE si NO tienen ticket asociado (evita doble conteo)
            if dte.tipo_documento == 'BOLETA ELECTRONICA':
                cuadratura_data['total_boletas_electronicas'] += monto_real
                cuadratura_data['total_boletas_electronicas_display'] += monto_real
                cuadratura_data['cantidad_boletas_electronicas'] += 1
                cuadratura_data['total_descuentos'] += descuento_dte
            elif dte.tipo_documento == 'BOLETA PAPEL':
                cuadratura_data['total_boletas_papel'] += monto_real
                cuadratura_data['total_boletas_papel_display'] += monto_real
                cuadratura_data['cantidad_boletas_papel'] += 1
                cuadratura_data['total_descuentos'] += descuento_dte
            elif dte.tipo_documento == 'FACTURA ELECTRONICA':
                cuadratura_data['total_facturas'] += monto_real
                cuadratura_data['total_facturas_display'] += monto_real
                cuadratura_data['cantidad_facturas'] += 1
                cuadratura_data['total_descuentos'] += descuento_dte
            elif dte.tipo_documento == 'FACTURA EXENTA':
                cuadratura_data['total_facturas_exentas'] += monto_real
                cuadratura_data['total_facturas_exentas_display'] += monto_real
                cuadratura_data['cantidad_facturas_exentas'] += 1
                cuadratura_data['total_descuentos'] += descuento_dte
        else:
            # Ticket con DTE: solo registrar cantidades de documentos (no montos
            # en `total_*`, para evitar doble conteo con `total_tickets`).
            # Los `_display` SÍ los sumamos para que la sección "Documentos
            # (referencia)" del Resumen de Caja refleje el monto emitido por
            # tipo de documento, aunque el dinero ya esté contabilizado en el
            # ticket asociado.
            if dte.tipo_documento == 'BOLETA ELECTRONICA':
                cuadratura_data['cantidad_boletas_electronicas'] += 1
                cuadratura_data['total_boletas_electronicas_display'] += monto_real
            elif dte.tipo_documento == 'BOLETA PAPEL':
                cuadratura_data['cantidad_boletas_papel'] += 1
                cuadratura_data['total_boletas_papel_display'] += monto_real
            elif dte.tipo_documento == 'FACTURA ELECTRONICA':
                cuadratura_data['cantidad_facturas'] += 1
                cuadratura_data['total_facturas_display'] += monto_real
            elif dte.tipo_documento == 'FACTURA EXENTA':
                cuadratura_data['cantidad_facturas_exentas'] += 1
                cuadratura_data['total_facturas_exentas_display'] += monto_real
        
        # Procesar pagos del DTE SOLO si no tiene ticket asociado Y no es NC.
        # Las NC ya se contabilizaron arriba a través de
        # `total_nc_efectivo` / `total_nc_transferencia`; si volviéramos a
        # iterar sus `Dte_Detalle_Pago` aquí terminaríamos sumando el monto
        # de la NC como si fuera un ingreso de caja, y el descuento final
        # (`total_efectivo -= total_nc_efectivo`) se neutralizaría contra
        # esa suma dejando el teórico sin cambios — anulando el efecto
        # esperado de la NC sobre la cuadratura.
        if not tiene_ticket_asociado and dte.tipo_documento != 'NOTA DE CREDITO':
            for pago in dte.dte_asociado.all():
                metodo = pago.metodo_pago or ''
                tipo_tarjeta = pago.tipo_tarjeta or ''
                monto = pago.monto or 0
                
                metodo_upper = metodo.upper()
                tarjeta_upper = tipo_tarjeta.upper()
                
                # Efectivo
                if metodo_upper == 'EFECTIVO':
                    cuadratura_data['total_efectivo'] += monto
                
                # Transbank Débito (solo por método, tipo_tarjeta no importa)
                elif metodo_upper in ['TBK_DEBITO_POS', 'TARJETA_DEBITO']:
                    cuadratura_data['total_tarjeta_debito'] += monto
                    cuadratura_data['total_transbank'] += monto
                
                # Transbank Crédito (solo por método, tipo_tarjeta no importa)
                elif metodo_upper in ['TBK_CREDITO_POS', 'TARJETA_CREDITO']:
                    cuadratura_data['total_tarjeta_credito'] += monto
                    cuadratura_data['total_transbank'] += monto
                
                # Transbank Prepago
                elif metodo_upper == 'TBK_PREPAGO_POS':
                    cuadratura_data['total_tarjeta_debito'] += monto
                    cuadratura_data['total_transbank'] += monto
                
                # Transbank genérico
                elif metodo_upper in ['TBK_POS_INTEGRADO', 'TBK_MANUAL']:
                    cuadratura_data['total_transbank'] += monto
                
                # Transferencia
                elif 'TRANSFERENCIA' in metodo_upper:
                    cuadratura_data['total_transferencia'] += monto
                
                # Cheque
                elif 'CHEQUE' in metodo_upper:
                    cuadratura_data['total_cheque'] += monto
                
                # Convenio
                elif metodo_upper == 'CONVENIO':
                    cuadratura_data['total_convenio'] += monto
                
                # Crédito trabajador
                elif metodo_upper == 'CREDITO_TRABAJADOR':
                    cuadratura_data['total_credito_trabajador'] += monto
                
                # Crédito externo
                elif metodo_upper == 'CREDITO_EXTERNO':
                    cuadratura_data['total_credito_externo'] += monto
                
                # Orden de compra
                elif metodo_upper == 'ORDEN_COMPRA' or ('ORDEN' in metodo_upper and 'COMPRA' in metodo_upper):
                    cuadratura_data['total_orden_compra'] += monto
                
                # Tarjeta Comercial
                elif metodo_upper == 'TARJETA_COMERCIAL':
                    cuadratura_data['total_tarjetas_comerciales'] += monto
                    if 'HITES' in tarjeta_upper:
                        cuadratura_data['total_hites'] += monto
                    elif 'PRESTO' in tarjeta_upper:
                        cuadratura_data['total_presto'] += monto
                
                # Venta Internet - buscar en tipo_tarjeta para clasificar
                elif metodo_upper == 'VENTA_INTERNET':
                    cuadratura_data['total_venta_internet'] += monto
                    # Clasificar por tipo_tarjeta
                    if 'FALABELLA' in tarjeta_upper or 'WALMART' in tarjeta_upper:
                        cuadratura_data['total_falabella'] += monto
                    elif 'PARIS' in tarjeta_upper:
                        cuadratura_data['total_paris'] += monto
                    elif 'RIPLEY' in tarjeta_upper:
                        cuadratura_data['total_ripley'] += monto
                    elif 'MERCADO' in tarjeta_upper or 'SHOPIFY' in tarjeta_upper:
                        cuadratura_data['total_mercadopago'] += monto
                    elif 'KLAP' in tarjeta_upper:
                        cuadratura_data['total_klap'] += monto
    
    # ========== NC DE DEVOLUCIÓN: EFECTO POR FECHA DE CUADRATURA ==========
    # A diferencia del resto de DTEs (que se imputan por `fecha_emision`), las
    # NC de devolución afectan la caja del día elegido por el operador al
    # emitirlas, guardado en `Dte_Detalle_Pago.fecha_pago`. La `fecha_emision`
    # de la NC sigue siendo el día real de emisión (correcta para el SII y
    # para gestión-DTE), pero el egreso de caja se imputa a `fecha_pago`.
    #
    # Compatibilidad: para NC antiguas sin `fecha_pago` (null), la fecha de
    # efecto cae a `fecha_emision`, replicando exactamente el comportamiento
    # anterior (los tests de regresión de cuadratura siguen pasando).
    from django.db.models import Q
    ncs_devolucion = (
        Dte.objects.filter(
            sucursal=sucursal,
            tipo_documento='NOTA DE CREDITO',
            tipo_transaccion='DEVOLUCION',
            estado_dte__in=['EMITIDO', 'ACEPTADO'],
            descartado=False,
        )
        .filter(
            Q(dte_asociado__fecha_pago=fecha_obj)
            | Q(dte_asociado__fecha_pago__isnull=True, fecha_emision=fecha_obj)
        )
        .distinct()
        .prefetch_related('dte_asociado')
    )
    for nc in ncs_devolucion:
        pagos_nc = list(nc.dte_asociado.all())
        # Fecha de efecto: la primera fecha_pago no nula del reembolso; si
        # ninguna está fijada (NC antigua), se usa la fecha de emisión.
        fecha_efecto = next(
            (p.fecha_pago for p in pagos_nc if p.fecha_pago), None
        ) or nc.fecha_emision
        if fecha_efecto != fecha_obj:
            continue
        monto_nc = nc.monto_con_iva or 0
        cuadratura_data['cantidad_notas_credito'] += 1
        cuadratura_data['total_notas_credito'] += monto_nc
        if any((p.metodo_pago or '').upper() == 'EFECTIVO' for p in pagos_nc):
            cuadratura_data['total_nc_efectivo'] += monto_nc
        elif any((p.metodo_pago or '').upper() == 'TRANSFERENCIA' for p in pagos_nc):
            cuadratura_data['total_nc_transferencia'] += monto_nc
        elif any((p.metodo_pago or '').upper() in _METODOS_PAGO_CREDITO_NC for p in pagos_nc):
            # Devolución sobre una venta A CRÉDITO: no salió plata de ningún
            # medio de caja, se rebajó la cuenta por cobrar. Se descuenta del
            # bucket de crédito (que igual entra a "Efectivo y Otros" del
            # Resumen) para que el VENTA TOTAL cuadre con `total_notas_credito`.
            cuadratura_data['total_nc_credito'] += monto_nc

    # ========== CALCULAR TOTALES GENERALES ==========
    # Tarjetas comerciales: solo Hites
    cuadratura_data['total_tarjetas_comerciales'] = cuadratura_data['total_hites']
    
    # Alias para compatibilidad con frontend.
    # total_transbank = crédito + débito + genérico/migrado (TBK_POS_INTEGRADO/TBK_MANUAL);
    # total_tarjeta_debito = débito + prepago. La resta = crédito + genérico, así el bucket
    # VISA-MC-AMEX incluye el Transbank migrado (coherente con el data-metodos de su fila) y
    # ninguna venta con tarjeta queda fuera del VENTA TOTAL del Resumen de Caja.
    cuadratura_data['total_visa_mc_amex'] = (
        cuadratura_data['total_transbank'] - cuadratura_data['total_tarjeta_debito']
    )
    
    # Venta Internet ya se calcula en el loop, pero asegurar el total
    # (ya se suma en cada if de venta internet arriba)
    
    cuadratura_data['venta_total'] = (
        cuadratura_data['total_tickets'] +
        cuadratura_data['total_boletas_electronicas'] +
        cuadratura_data['total_boletas_papel'] +
        cuadratura_data['total_facturas'] +
        cuadratura_data['total_facturas_exentas'] -
        cuadratura_data['total_notas_credito']
    )

    # NC en efectivo resta del efectivo teórico de caja
    cuadratura_data['total_efectivo'] -= cuadratura_data['total_nc_efectivo']
    # NC por transferencia resta del teórico de transferencias (simétrico
    # al tratamiento del efectivo). Antes sólo se sumaba en total_nc_transferencia
    # pero nunca se descontaba → el arqueo cerraba con diferencia en transferencia
    # cada vez que se emitía una NC por devolución vía transferencia.
    cuadratura_data['total_transferencia'] -= cuadratura_data['total_nc_transferencia']
    # NC sobre ventas a crédito: rebajan la cuenta por cobrar, no la caja.
    # `total_credito_externo` es parte de "Efectivo y Otros" en el Resumen, así
    # que restar aquí mantiene esa suma coherente con el VENTA TOTAL (que ya
    # descuenta la NC vía `total_notas_credito`). Puede quedar negativo si la
    # venta a crédito original es de otro día — es correcto: representa plata
    # que se dejó de cobrar, no un ingreso.
    cuadratura_data['total_credito_externo'] -= cuadratura_data['total_nc_credito']

    return cuadratura_data


# Mapeo campo de ArqueoCaja ← key del dict devuelto por
# `_calcular_cuadratura_data`. Se declara a nivel de módulo para que
# `_recalcular_teoricos_arqueo` y `crear_arqueo` coincidan en el snapshot
# y no se desincronicen al agregar un nuevo método de pago.
_MAPEO_TEORICOS_ARQUEO = (
    ('total_efectivo_teorico', 'total_efectivo'),
    ('total_tarjeta_debito_teorico', 'total_tarjeta_debito'),
    ('total_tarjeta_credito_teorico', 'total_tarjeta_credito'),
    ('total_transbank_teorico', 'total_transbank'),
    ('total_hites_teorico', 'total_hites'),
    ('total_tarjetas_comerciales_teorico', 'total_tarjetas_comerciales'),
    ('total_falabella_teorico', 'total_falabella'),
    ('total_paris_teorico', 'total_paris'),
    ('total_ripley_teorico', 'total_ripley'),
    ('total_mercadopago_teorico', 'total_mercadopago'),
    ('total_klap_teorico', 'total_klap'),
    ('total_venta_internet_teorico', 'total_venta_internet'),
    ('total_transferencia_teorico', 'total_transferencia'),
    ('total_cheque_teorico', 'total_cheque'),
    ('total_convenio_teorico', 'total_convenio'),
    ('total_credito_trabajador_teorico', 'total_credito_trabajador'),
    ('total_tickets_teorico', 'total_tickets'),
    ('total_boletas_electronicas_teorico', 'total_boletas_electronicas'),
    ('total_facturas_teorico', 'total_facturas'),
    ('total_facturas_exentas_teorico', 'total_facturas_exentas'),
    ('total_notas_credito_teorico', 'total_notas_credito'),
    ('cantidad_tickets', 'cantidad_tickets'),
    ('cantidad_boletas_electronicas', 'cantidad_boletas_electronicas'),
    ('cantidad_facturas', 'cantidad_facturas'),
    ('cantidad_facturas_exentas', 'cantidad_facturas_exentas'),
    ('venta_total_teorica', 'venta_total'),
    ('total_descuento_puntos_teorico', 'total_descuento_puntos'),
)


def _to_int(value):
    """Cast defensivo a int (los Decimal/floats vienen serializados)."""
    if value is None:
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def _recalcular_teoricos_arqueo(
    arqueo, usuario=None, registrar_bitacora=False, razon=''
):
    """Re-snapshot de los `total_*_teorico` del arqueo desde la cuadratura actual.

    Vuelve a llamar a `_calcular_cuadratura_data` para la fecha y sucursal
    del arqueo, actualiza los campos denormalizados y recalcula las
    diferencias `físico - teórico`. Devuelve un dict con:

        {
          'cambios': { campo: {'antes': .., 'despues': ..}, ... },
          'cuadratura': dict completo (como lo entrega _calcular_cuadratura_data),
          'hay_cambios': bool,
        }

    Parámetros:
      - `usuario`: si `registrar_bitacora=True`, autor de la observación.
      - `registrar_bitacora`: si hay diferencias, crea una entrada en
        `ObservacionArqueo` tipo SISTEMA con los montos antes/después.
      - `razon`: texto libre que se anexa a la bitácora para contexto
        (ej: "solicitado por admin", "auto al abrir arqueo abierto").

    Este helper es la pieza central de la política C ("recalcular mientras
    está abierto, manual cuando está cerrado"): permite tener un solo
    punto de verdad y evita que `ArqueoCaja.total_*_teorico` queden
    fuera de sincronía con `_calcular_cuadratura_data`.
    """
    # `fecha_arqueo` suele llegar como `date`, pero si el ORM devolvió una
    # string (tests / fixtures) respetamos el formato esperado por
    # `_calcular_cuadratura_data` (YYYY-MM-DD).
    fecha_arqueo = arqueo.fecha_arqueo
    fecha_str = (
        fecha_arqueo.strftime('%Y-%m-%d')
        if hasattr(fecha_arqueo, 'strftime') else str(fecha_arqueo)
    )
    cuadratura = _calcular_cuadratura_data(arqueo.sucursal, fecha_str)

    update_fields = []
    cambios = {}
    for campo_arqueo, key_cuadratura in _MAPEO_TEORICOS_ARQUEO:
        nuevo = _to_int(cuadratura.get(key_cuadratura, 0))
        actual = _to_int(getattr(arqueo, campo_arqueo, 0))
        if nuevo != actual:
            cambios[campo_arqueo] = {'antes': actual, 'despues': nuevo}
            setattr(arqueo, campo_arqueo, nuevo)
            update_fields.append(campo_arqueo)

    # Diferencias físico-teórico se recalculan siempre con los valores
    # actualizados. Para `diferencia_efectivo` replicamos la misma fórmula
    # que usa `ArqueoCaja.save()` (`físico - (teorico + fondo_fijo)`) para
    # no divergir cuando el signal se dispare más adelante.
    nueva_dif_efectivo = (
        _to_int(arqueo.total_efectivo_fisico)
        - (_to_int(arqueo.total_efectivo_teorico) + _to_int(arqueo.fondo_fijo_snapshot))
    )
    nueva_dif_tbk = (
        _to_int(arqueo.cierre_pos_fisico) - _to_int(arqueo.total_transbank_teorico)
    )
    nueva_dif_debito = (
        _to_int(arqueo.cierre_debito_fisico) - _to_int(arqueo.total_tarjeta_debito_teorico)
    )
    nueva_dif_credito = (
        _to_int(arqueo.cierre_credito_fisico) - _to_int(arqueo.total_tarjeta_credito_teorico)
    )
    if arqueo.diferencia_efectivo != nueva_dif_efectivo:
        arqueo.diferencia_efectivo = nueva_dif_efectivo
        update_fields.append('diferencia_efectivo')
    if arqueo.diferencia_transbank != nueva_dif_tbk:
        arqueo.diferencia_transbank = nueva_dif_tbk
        update_fields.append('diferencia_transbank')
    if arqueo.diferencia_debito != nueva_dif_debito:
        arqueo.diferencia_debito = nueva_dif_debito
        update_fields.append('diferencia_debito')
    if arqueo.diferencia_credito != nueva_dif_credito:
        arqueo.diferencia_credito = nueva_dif_credito
        update_fields.append('diferencia_credito')

    # Usamos `QuerySet.update()` para NO disparar `ArqueoCaja.save()`, que
    # recomputa `total_efectivo_fisico` desde billetes/monedas y pisaría
    # los valores ingresados en modo Express (donde el total viene como
    # input libre y las denominaciones quedan en 0).
    if update_fields:
        ArqueoCaja.objects.filter(pk=arqueo.pk).update(
            **{f: getattr(arqueo, f) for f in update_fields}
        )

    if registrar_bitacora and cambios and usuario is not None:
        # Construir un resumen legible de los cambios más importantes para
        # que quede evidencia clara en la bitácora del arqueo.
        resumen = [
            f'{campo.replace("total_", "").replace("_teorico", "")}: '
            f'${c["antes"]:,} → ${c["despues"]:,}'
            for campo, c in cambios.items()
            if campo.startswith('total_') and campo.endswith('_teorico')
        ]
        texto = (
            f'Recálculo de teóricos ({razon or "manual"}). Cambios: '
            + ('; '.join(resumen) if resumen else 'solo cantidades/totales de documentos')
        )
        ObservacionArqueo.objects.create(
            arqueo=arqueo,
            usuario=usuario,
            tipo='SISTEMA',
            texto=texto,
            visible_para_cajera=True,
        )

    return {
        'cambios': cambios,
        'cuadratura': cuadratura,
        'hay_cambios': bool(update_fields),
    }


@login_required
@require_POST
def generar_cuadratura_caja(request):
    """Generar cuadratura de caja para una fecha específica"""
    try:
        fecha_cuadratura = request.POST.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha_cuadratura:
            return JsonResponse({
                'success': False,
                'error': 'Fecha requerida'
            })
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Usar la función helper para calcular los datos
        cuadratura_data = _calcular_cuadratura_data(sucursal, fecha_cuadratura)

        # Anti-fraude: mantener la bandera de conteo ciego para ocultar los
        # teóricos dentro del modal de Arqueo (cajeros/vendedores), pero dejar
        # visible el efectivo en el Resumen de Caja para evitar el mensaje
        # "Pendiente de conteo" que resulta confuso en ese contexto.
        rol_usuario = getattr(request.user, 'rol', None)
        cuadratura_data['modo_conteo_ciego'] = rol_usuario in ('cajero', 'vendedor')

        return JsonResponse({
            'success': True,
            'cuadratura': cuadratura_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al generar cuadratura: {str(e)}'
        })


# ========== DETALLE DE MÉTODOS DE PAGO PARA RESUMEN DE CAJA ==========

# Mapeo Ticket.tipo_dte (con guiones bajos, snake_case histórico) →
# Dte.tipo_documento (con espacios, formato "BOLETA ELECTRONICA"). Se
# usa para desambiguar el match folio_dte ↔ DTE cuando varios DTEs de
# distintos tipos comparten un mismo `numero_documento` en la misma
# sucursal (caso real visto: NOTA DE DEBITO #1 vs BOLETA #1, etc).
_TIPO_DTE_TICKET_A_DTE = {
    'BOLETA': 'BOLETA PAPEL',
    'BOLETA_PAPEL': 'BOLETA PAPEL',
    'BOLETA_ELECTRONICA': 'BOLETA ELECTRONICA',
    'FACTURA_ELECTRONICA': 'FACTURA ELECTRONICA',
    'FACTURA_EXENTA': 'FACTURA EXENTA',
}


def _tipo_dte_objetivo(ticket) -> str | None:
    """Tipo de DTE esperado según `ticket.tipo_dte`. None si es TICKET puro."""
    return _TIPO_DTE_TICKET_A_DTE.get((ticket.tipo_dte or '').upper().strip())


# Mapa código método de pago → categoría lógica que usa el modal
# "Resumen de Caja" (tarjetas, efectivo, internet). Se calcula una sola
# vez a nivel de módulo para no recomputarlo en cada request.
_CATEGORIAS_METODO_PAGO = {
    # Tarjetas Transbank y comerciales
    'TARJETA_DEBITO': 'tarjetas',
    'TARJETA_CREDITO': 'tarjetas',
    'TBK_DEBITO_POS': 'tarjetas',
    'TBK_CREDITO_POS': 'tarjetas',
    'TBK_PREPAGO_POS': 'tarjetas',
    'TBK_POS_INTEGRADO': 'tarjetas',
    'TBK_MANUAL': 'tarjetas',
    'TARJETA_COMERCIAL': 'tarjetas',
    # Efectivo y equivalentes (incluye transferencia, convenio, crédito,
    # cheque y orden de compra: son los que se agrupan en la tarjeta
    # "Efectivo y Otros" del Resumen).
    'EFECTIVO': 'efectivo',
    'TRANSFERENCIA': 'efectivo',
    'CHEQUE': 'efectivo',
    'CONVENIO': 'efectivo',
    'CREDITO_TRABAJADOR': 'efectivo',
    'CREDITO_EXTERNO': 'efectivo',
    'ORDEN_COMPRA': 'efectivo',
    'OTRO': 'efectivo',
    # Marketplaces / venta por internet
    'VENTA_INTERNET': 'internet',
}


def _categoria_metodo_pago(metodo: str | None) -> str:
    return _CATEGORIAS_METODO_PAGO.get((metodo or '').upper(), 'otros')


@login_required
@require_GET
def obtener_detalle_cuadratura_metodos_pago(request):
    """
    Devuelve el detalle desagregado de pagos del día para el modal
    "Resumen de Caja — <fecha>".

    Cada ítem corresponde a un `TicketDetallePago` o a un
    `Dte_Detalle_Pago` (cuando el DTE no tiene ticket vinculado) del día
    de la sucursal activa, más el DTE asociado resuelto por
    (sucursal, folio) cuando existe.

    La respuesta incluye por cada ítem un bloque `permisos` indicando
    qué puede editar el usuario sobre el DTE asociado (fecha, número,
    pago). El frontend usa estos flags para habilitar/ocultar el botón
    de edición de fecha por cada fila.

    Query params:
        fecha      (str, YYYY-MM-DD, requerido)
        categoria  (opcional: tarjetas | efectivo | internet | todo)
    """
    from datetime import datetime as _dt

    fecha_str = (request.GET.get('fecha') or '').strip()
    categoria = (request.GET.get('categoria') or 'todo').strip().lower()
    if not fecha_str:
        return JsonResponse({'success': False, 'error': 'Fecha requerida'})

    try:
        fecha_obj = _dt.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Fecha inválida. Formato esperado YYYY-MM-DD',
        })

    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal seleccionada',
        })

    sucursal = get_object_or_404(Sucursal, id=sucursal_id)

    # Permisos granulares del usuario (se aplican POR DTE según su tipo).
    permisos_dte = permisos_edicion_dte_context(request.user, sucursal_id)
    flags_campo = permisos_dte.get('campo', {})
    flags_tipo = permisos_dte.get('tipo', {})

    # Eliminar un ítem (boleta/factura/NC) desde este modal, y editar la
    # fecha de cuadratura de una NC, no tienen permiso granular por tipo
    # (`CODIGO_PERMISO_TIPO_DTE` no incluye 'NOTA DE CREDITO'), así que se
    # restringen al mismo gate de rol que usa `eliminar_documento_venta`.
    rol_usuario = getattr(request.user, 'rol', '') or ''
    es_admin = rol_usuario == 'administrador'

    # ---------------------------------------------------------------
    # 1) Pagos desde TICKETS PAGADOS del día (con o sin DTE asociado).
    # ---------------------------------------------------------------
    tickets_qs = (
        Ticket.objects
        .filter(sucursal=sucursal, fecha=fecha_obj, estado='PAGADO')
        .prefetch_related('pagos')
        .only(
            'id', 'correlativo', 'hora', 'created_at', 'fecha', 'total',
            'cliente_nombre', 'cliente_rut', 'folio_dte', 'tipo_dte',
            'sucursal_id',
        )
        .order_by('hora', 'id')
    )

    # Resolver DTEs vinculados a los tickets del día.
    # Se indexan por (folio, tipo_documento) para evitar el bug de cruzar
    # un Ticket BOLETA_ELECTRONICA con folio_dte=1 contra una NOTA DE
    # DEBITO #1 cualquiera de la misma sucursal. La preferencia para
    # match es:
    #   1) Coincidencia exacta (folio + tipo_dte_objetivo del ticket)
    #   2) Coincidencia sólo por folio si el ticket no declara tipo_dte
    #      o el tipo no está en el mapa.
    folios_tickets = [t.folio_dte for t in tickets_qs if t.folio_dte]
    dtes_por_folio_tipo: dict[tuple, object] = {}
    dtes_por_folio: dict[int, object] = {}
    if folios_tickets:
        dtes_vinc_qs = (
            Dte.objects
            .filter(
                sucursal_id=sucursal.id,
                numero_documento__in=folios_tickets,
            )
            .only(
                'id', 'numero_documento', 'tipo_documento', 'fecha_emision',
                'estado_dte', 'monto_con_iva', 'sucursal_id',
            )
        )
        for d in dtes_vinc_qs:
            tipo_up = (d.tipo_documento or '').upper().strip()
            key_compuesta = (d.numero_documento, tipo_up)
            existente = dtes_por_folio_tipo.get(key_compuesta)
            if (
                existente is None
                or (
                    existente.estado_dte not in ('EMITIDO', 'ACEPTADO')
                    and d.estado_dte in ('EMITIDO', 'ACEPTADO')
                )
            ):
                dtes_por_folio_tipo[key_compuesta] = d
            # Fallback por folio (preferir el EMITIDO/ACEPTADO, igual
            # que antes para no romper tickets sin tipo_dte declarado).
            existente2 = dtes_por_folio.get(d.numero_documento)
            if (
                existente2 is None
                or (
                    existente2.estado_dte not in ('EMITIDO', 'ACEPTADO')
                    and d.estado_dte in ('EMITIDO', 'ACEPTADO')
                )
            ):
                dtes_por_folio[d.numero_documento] = d

    items = []

    def _permisos_para_dte(dte_obj):
        """Retorna flags {editar_fecha, editar_numero, editar_pago} sobre
        un DTE, combinando permisos de campo con permisos del tipo DTE.
        Para tickets sin DTE, `editar_fecha_ticket` indica si el usuario
        puede mover el ticket a otra fecha."""
        if not dte_obj:
            return {
                'editar_fecha': False,
                'editar_numero': False,
                'editar_pago': False,
                'editar_fecha_ticket': bool(flags_campo.get('fecha')),
                'eliminar': False,
            }
        tipo_up = (dte_obj.tipo_documento or '').upper().strip()
        tipo_ok = flags_tipo.get(tipo_up, False)
        return {
            'editar_fecha': bool(flags_campo.get('fecha') and tipo_ok),
            'editar_numero': bool(
                flags_campo.get('numero_documento') and tipo_ok
            ),
            'editar_pago': bool(flags_campo.get('pago') and tipo_ok),
            'editar_fecha_ticket': False,
            'eliminar': es_admin,
        }

    for ticket in tickets_qs:
        dte = None
        if ticket.folio_dte:
            tipo_obj = _tipo_dte_objetivo(ticket)
            if tipo_obj:
                dte = dtes_por_folio_tipo.get((ticket.folio_dte, tipo_obj))
            if dte is None:
                candidato = dtes_por_folio.get(ticket.folio_dte)
                if candidato is not None and tipo_obj is None:
                    dte = candidato
                elif (
                    candidato is not None
                    and (candidato.tipo_documento or '').upper().strip() == tipo_obj
                ):
                    dte = candidato
        permisos_item = _permisos_para_dte(dte)
        hora_str = ticket.hora.strftime('%H:%M') if ticket.hora else ''
        ticket_fecha_str = ticket.fecha.strftime('%Y-%m-%d') if ticket.fecha else None
        drift_fecha = bool(
            dte and ticket.fecha and ticket.fecha != dte.fecha_emision
        )
        # Tipo de documento esperado según el ticket (para mostrar en
        # el frontend incluso cuando no hay DTE vinculado).
        tipo_dte_ticket = _tipo_dte_objetivo(ticket) or (ticket.tipo_dte or '').upper().strip()
        for pago in ticket.pagos.all():
            metodo = (pago.metodo_pago or '').upper()
            cat = _categoria_metodo_pago(metodo)
            items.append({
                'origen': 'TICKET',
                'pago_id': pago.id,
                'ticket_id': ticket.id,
                'ticket_correlativo': ticket.correlativo,
                'ticket_fecha': ticket_fecha_str,
                'tipo_dte_ticket': tipo_dte_ticket,
                'dte_id': dte.id if dte else None,
                'correlativo': ticket.correlativo,
                'hora': hora_str,
                'cliente': (ticket.cliente_nombre or '').strip() or 'Cliente General',
                'cliente_rut': (ticket.cliente_rut or '').strip(),
                'metodo_pago': metodo,
                'metodo_pago_display': obtener_nombre_metodo_pago(metodo),
                'tipo_tarjeta': (pago.tipo_tarjeta or '').strip(),
                'voucher': (pago.voucher or '').strip(),
                'monto': int(pago.monto or 0),
                'categoria': cat,
                'dte': {
                    'id': dte.id,
                    'folio': dte.numero_documento,
                    'tipo': dte.tipo_documento,
                    'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
                    'estado': dte.estado_dte,
                } if dte else None,
                'drift_fecha': drift_fecha,
                'permisos': permisos_item,
            })

    # ---------------------------------------------------------------
    # 2) Pagos desde DTEs del día que NO tienen ticket vinculado
    #    (ej: boletas/facturas emitidas directo). Se evita duplicar los
    #    pagos cuando el DTE ya aparece por el ticket.
    # ---------------------------------------------------------------
    folios_con_ticket = {
        t.folio_dte for t in tickets_qs if t.folio_dte is not None
    }
    dtes_directos_qs = (
        Dte.objects
        .filter(
            sucursal_id=sucursal.id,
            fecha_emision=fecha_obj,
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
            estado_dte__in=['EMITIDO', 'ACEPTADO'],
        )
        .prefetch_related('dte_asociado', 'receptor')
        .only(
            'id', 'numero_documento', 'tipo_documento', 'fecha_emision',
            'hora', 'estado_dte', 'monto_con_iva', 'sucursal_id',
            'receptor_id',
        )
        .order_by('hora', 'id')
    )

    for dte in dtes_directos_qs:
        if dte.numero_documento in folios_con_ticket:
            continue  # Ya contado vía ticket
        permisos_item = _permisos_para_dte(dte)
        hora_str = dte.hora.strftime('%H:%M') if dte.hora else ''
        receptor = getattr(dte, 'receptor', None)
        cliente_nombre = (
            (receptor.razon_social if receptor else '') or 'Cliente General'
        )
        cliente_rut = (receptor.rut if receptor else '') or ''
        for pago in dte.dte_asociado.all():
            metodo = (pago.metodo_pago or '').upper()
            cat = _categoria_metodo_pago(metodo)
            items.append({
                'origen': 'DTE',
                'pago_id': pago.id,
                'ticket_id': None,
                'ticket_correlativo': None,
                'ticket_fecha': None,
                'tipo_dte_ticket': (dte.tipo_documento or '').upper().strip(),
                'dte_id': dte.id,
                'correlativo': dte.numero_documento,
                'hora': hora_str,
                'cliente': cliente_nombre,
                'cliente_rut': cliente_rut,
                'metodo_pago': metodo,
                'metodo_pago_display': obtener_nombre_metodo_pago(metodo),
                'tipo_tarjeta': (pago.tipo_tarjeta or '').strip(),
                'voucher': (pago.voucher or '').strip(),
                'monto': int(pago.monto or 0),
                'categoria': cat,
                'dte': {
                    'id': dte.id,
                    'folio': dte.numero_documento,
                    'tipo': dte.tipo_documento,
                    'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
                    'estado': dte.estado_dte,
                },
                'drift_fecha': False,
                'permisos': permisos_item,
            })

    # ---------------------------------------------------------------
    # 3) NC de devolución que afectan la caja de este día. Se imputan por
    #    `Dte_Detalle_Pago.fecha_pago` (no por `fecha_emision`), mismo
    #    criterio que usa `_calcular_cuadratura_data`. Solo NC con
    #    tipo_transaccion='DEVOLUCION' (mueven dinero); las 'ANULACION'
    #    son informativas y no se listan aquí.
    # ---------------------------------------------------------------
    ncs_devolucion_qs = (
        Dte.objects
        .filter(
            sucursal_id=sucursal.id,
            tipo_documento='NOTA DE CREDITO',
            tipo_transaccion='DEVOLUCION',
            estado_dte__in=['EMITIDO', 'ACEPTADO'],
            descartado=False,
        )
        .filter(
            Q(dte_asociado__fecha_pago=fecha_obj)
            | Q(dte_asociado__fecha_pago__isnull=True, fecha_emision=fecha_obj)
        )
        .distinct()
        .prefetch_related('dte_asociado', 'receptor')
        .only(
            'id', 'numero_documento', 'tipo_documento', 'fecha_emision',
            'hora', 'estado_dte', 'monto_con_iva', 'sucursal_id',
            'receptor_id',
        )
    )

    for nc in ncs_devolucion_qs:
        pagos_nc = list(nc.dte_asociado.all())
        fecha_efecto = next(
            (p.fecha_pago for p in pagos_nc if p.fecha_pago), None
        ) or nc.fecha_emision
        if fecha_efecto != fecha_obj:
            continue
        hora_str = nc.hora.strftime('%H:%M') if nc.hora else ''
        receptor = getattr(nc, 'receptor', None)
        cliente_nombre = (
            (receptor.razon_social if receptor else '') or 'Cliente General'
        )
        cliente_rut = (receptor.rut if receptor else '') or ''
        for pago in pagos_nc:
            metodo = (pago.metodo_pago or '').upper()
            cat = _categoria_metodo_pago(metodo)
            fecha_pago_str = (
                pago.fecha_pago.strftime('%Y-%m-%d')
                if pago.fecha_pago else nc.fecha_emision.strftime('%Y-%m-%d')
            )
            items.append({
                'origen': 'NC',
                'pago_id': pago.id,
                'ticket_id': None,
                'ticket_correlativo': None,
                'ticket_fecha': None,
                'tipo_dte_ticket': 'NOTA DE CREDITO',
                'dte_id': nc.id,
                'correlativo': nc.numero_documento,
                'hora': hora_str,
                'cliente': cliente_nombre,
                'cliente_rut': cliente_rut,
                'metodo_pago': metodo,
                'metodo_pago_display': obtener_nombre_metodo_pago(metodo),
                'tipo_tarjeta': (pago.tipo_tarjeta or '').strip(),
                'voucher': (pago.voucher or '').strip(),
                # Negativo: la NC resta del método de pago (coherente con
                # el descuento que aplica `_calcular_cuadratura_data` al
                # efectivo/transferencia teórico del día).
                'monto': -int(pago.monto or 0),
                'categoria': cat,
                'es_nota_credito': True,
                'dte': {
                    'id': nc.id,
                    'folio': nc.numero_documento,
                    'tipo': nc.tipo_documento,
                    'fecha_emision': nc.fecha_emision.strftime('%Y-%m-%d'),
                    'fecha_pago': fecha_pago_str,
                    'estado': nc.estado_dte,
                },
                'drift_fecha': False,
                'permisos': {
                    'editar_fecha': False,
                    'editar_numero': False,
                    'editar_pago': False,
                    'editar_fecha_ticket': False,
                    # Edita `fecha_pago` (fecha de cuadratura), no
                    # `fecha_emision` — botón distinto en el frontend.
                    'editar_fecha_pago_nc': es_admin,
                    'eliminar': es_admin,
                },
            })

    # Filtro por categoría (opcional)
    if categoria and categoria != 'todo':
        items = [it for it in items if it['categoria'] == categoria]

    # Ordenar por hora y luego correlativo
    items.sort(key=lambda it: (it['hora'] or '', it['correlativo'] or 0))

    # Totales por método (para pintar subtotales en el modal de detalle)
    totales_por_metodo: dict[str, dict] = {}
    for it in items:
        key = it['metodo_pago']
        bucket = totales_por_metodo.setdefault(
            key,
            {
                'metodo_pago': key,
                'metodo_pago_display': it['metodo_pago_display'],
                'categoria': it['categoria'],
                'cantidad': 0,
                'total': 0,
            },
        )
        bucket['cantidad'] += 1
        bucket['total'] += it['monto']

    return JsonResponse({
        'success': True,
        'fecha': fecha_str,
        'categoria': categoria,
        'items': items,
        'totales_por_metodo': list(totales_por_metodo.values()),
        'permisos_usuario': {
            'cualquier_edicion': bool(permisos_dte.get('cualquiera')),
            'campo': flags_campo,
            'tipo': flags_tipo,
            'puede_eliminar': es_admin,
        },
    })


@login_required
@require_POST
def editar_fecha_pago_nc(request):
    """
    Edita la fecha de cuadratura (`Dte_Detalle_Pago.fecha_pago`) de una NC
    de devolución, usada por `_calcular_cuadratura_data` para imputar el
    efecto de la NC a un día de caja distinto de `fecha_emision`.

    A diferencia de `editar_dte_boleta_papel` (que edita `fecha_emision`
    y requiere el permiso granular `dte_editar_fecha` + `dte_editar_tipo_*`),
    aquí no existe permiso granular por tipo para NOTA DE CREDITO
    (`CODIGO_PERMISO_TIPO_DTE` no la incluye), así que se restringe al
    mismo gate de rol que usa `eliminar_documento_venta`: solo
    administrador.

    Body JSON: { "dte_id": 123, "fecha_pago": "YYYY-MM-DD" }
    """
    rol_usuario = getattr(request.user, 'rol', '') or ''
    if rol_usuario != 'administrador':
        return JsonResponse({
            'success': False,
            'error': 'Solo los administradores pueden editar la fecha de cuadratura de una NC',
        }, status=403)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})

    dte_id = data.get('dte_id')
    fecha_pago_str = (data.get('fecha_pago') or '').strip()
    if not dte_id or not fecha_pago_str:
        return JsonResponse({
            'success': False,
            'error': 'dte_id y fecha_pago son requeridos',
        })

    from datetime import datetime as _dt2
    try:
        nueva_fecha = _dt2.strptime(fecha_pago_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Formato de fecha inválido (YYYY-MM-DD)',
        })

    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal seleccionada',
        })

    with transaction.atomic():
        nc = (
            Dte.objects
            .select_for_update()
            .filter(
                pk=dte_id,
                sucursal_id=sucursal_id,
                tipo_documento='NOTA DE CREDITO',
            )
            .first()
        )
        if not nc:
            return JsonResponse({
                'success': False,
                'error': 'Nota de crédito no encontrada',
            })
        if nc.tipo_transaccion != 'DEVOLUCION':
            return JsonResponse({
                'success': False,
                'error': 'Solo aplica a NC de devolución',
            })

        pagos_actualizados = Dte_Detalle_Pago.objects.filter(
            dte=nc
        ).update(fecha_pago=nueva_fecha)

    return JsonResponse({
        'success': True,
        'message': (
            f'Fecha de cuadratura de NC #{nc.numero_documento} '
            f'actualizada a {fecha_pago_str}'
        ),
        'pagos_actualizados': pagos_actualizados,
    })


@login_required
@require_POST
def sincronizar_fecha_ticket_dte(request):
    """
    Sincroniza `Ticket.fecha` con `Dte.fecha_emision` cuando hay drift.

    Caso de uso: en un flujo histórico se editó la fecha del DTE desde
    Gestión de Documentos pero no se propagó al ticket vinculado, así
    que la cuadratura del día sigue mostrando esos pagos en la fecha
    "vieja" del ticket. Este endpoint usa `.update()` para evitar el
    `auto_now=True` del campo `Ticket.fecha`.

    Permisos: requiere `dte_editar_fecha` + permiso del tipo de DTE
    (mismo criterio que `editar_dte_boleta_papel`). Por seguridad sólo
    actúa cuando hay drift real (ticket.fecha != dte.fecha_emision).

    Body JSON:
        { "dte_id": 123 }   o   { "ticket_id": 456 }
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'})

    dte_id = data.get('dte_id')
    ticket_id = data.get('ticket_id')
    if not dte_id and not ticket_id:
        return JsonResponse({
            'success': False,
            'error': 'Se requiere dte_id o ticket_id',
        })

    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal seleccionada',
        })

    with transaction.atomic():
        # Cargar el DTE (vía ticket si no llegó dte_id)
        if not dte_id:
            ticket = (
                Ticket.objects
                .select_for_update()
                .filter(pk=ticket_id, sucursal_id=sucursal_id)
                .first()
            )
            if not ticket:
                return JsonResponse({
                    'success': False,
                    'error': 'Ticket no encontrado',
                })
            if not ticket.folio_dte:
                return JsonResponse({
                    'success': False,
                    'error': 'Este ticket no tiene DTE asociado',
                })
            tipo_obj = _TIPO_DTE_TICKET_A_DTE.get(
                (ticket.tipo_dte or '').upper().strip()
            )
            qs = Dte.objects.filter(
                sucursal_id=sucursal_id,
                numero_documento=ticket.folio_dte,
            )
            if tipo_obj:
                qs = qs.filter(tipo_documento=tipo_obj)
            dte = qs.first()
            if not dte:
                return JsonResponse({
                    'success': False,
                    'error': 'No se encontró el DTE asociado al ticket',
                })
        else:
            dte = (
                Dte.objects
                .select_for_update()
                .filter(pk=dte_id, sucursal_id=sucursal_id)
                .first()
            )
            if not dte:
                return JsonResponse({
                    'success': False,
                    'error': 'DTE no encontrado',
                })

        # Validar permiso (campo fecha + tipo del DTE)
        if not puede_editar_campo_dte(
            request.user, 'fecha', dte.tipo_documento,
            sucursal_id=sucursal_id,
        ):
            return JsonResponse({
                'success': False,
                'error': (
                    f'No tiene permisos para editar fecha en {dte.tipo_documento}'
                ),
            }, status=403)

        # Localizar ticket vinculado por (sucursal, folio, tipo) — preferimos
        # el match por tipo para evitar cruces con otros tipos.
        # Mapeo inverso: Dte.tipo_documento → Ticket.tipo_dte
        tipo_dte_inv = {v: k for k, v in _TIPO_DTE_TICKET_A_DTE.items()}
        tipo_ticket_obj = tipo_dte_inv.get(
            (dte.tipo_documento or '').upper().strip()
        )
        ticket_qs = Ticket.objects.select_for_update().filter(
            sucursal_id=dte.sucursal_id,
            folio_dte=dte.numero_documento,
        )
        ticket = None
        if tipo_ticket_obj:
            ticket = ticket_qs.filter(tipo_dte=tipo_ticket_obj).first()
        if ticket is None:
            ticket = ticket_qs.first()
        if not ticket:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró ticket vinculado al DTE',
            })

        if ticket.fecha == dte.fecha_emision:
            return JsonResponse({
                'success': True,
                'mensaje': 'Las fechas ya estaban sincronizadas',
                'cambio': False,
                'ticket_id': ticket.id,
                'dte_id': dte.id,
                'fecha': dte.fecha_emision.strftime('%Y-%m-%d'),
            })

        fecha_anterior = ticket.fecha
        # `.update()` para evitar el `auto_now=True` de `Ticket.fecha`.
        Ticket.objects.filter(pk=ticket.pk).update(
            fecha=dte.fecha_emision,
        )

        # Bitácora en arqueos afectados (origen y destino), si hay alguno
        # cerrado o con diferencias en esas fechas.
        try:
            fechas_a_revisar = {fecha_anterior, dte.fecha_emision}
            arqueos_qs = ArqueoCaja.objects.filter(
                sucursal_id=dte.sucursal_id,
                fecha_arqueo__in=fechas_a_revisar,
            ).exclude(estado='ABIERTO')
            for arq in arqueos_qs:
                ObservacionArqueo.objects.create(
                    arqueo=arq,
                    usuario=request.user,
                    tipo='SISTEMA',
                    texto=(
                        f'Sincronización de fecha del Ticket #{ticket.correlativo}'
                        f' con DTE #{dte.numero_documento} ({dte.tipo_documento}). '
                        f'Fecha ticket: '
                        f'{fecha_anterior.strftime("%d/%m/%Y") if fecha_anterior else "—"}'
                        f' → {dte.fecha_emision.strftime("%d/%m/%Y")}. '
                        'Los teóricos del arqueo pueden requerir recálculo.'
                    ),
                    visible_para_cajera=True,
                )
        except Exception:
            # La bitácora no es bloqueante.
            pass

    return JsonResponse({
        'success': True,
        'mensaje': (
            f'Ticket #{ticket.correlativo} movido al '
            f'{dte.fecha_emision.strftime("%d/%m/%Y")} '
            '(según fecha de emisión del DTE).'
        ),
        'cambio': True,
        'ticket_id': ticket.id,
        'dte_id': dte.id,
        'fecha_anterior': (
            fecha_anterior.strftime('%Y-%m-%d') if fecha_anterior else None
        ),
        'fecha_nueva': dte.fecha_emision.strftime('%Y-%m-%d'),
    })


@login_required
@require_POST
def editar_fecha_ticket_sin_dte(request):
    """
    Permite cambiar `Ticket.fecha` de un ticket que NO tiene DTE asociado
    (tipo_dte == 'TICKET' o sin folio_dte).

    Caso de uso: tickets con pagos por tarjeta que quedaron en un día
    incorrecto porque en un flujo anterior se modificó la fecha del DTE
    pero no la de los métodos de pago (ticket sin DTE). El usuario
    necesita moverlos al día correcto para que la cuadratura cuadre.

    Body JSON:
        { "ticket_id": 456, "nueva_fecha": "2026-04-15" }
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'})

    ticket_id = data.get('ticket_id')
    nueva_fecha_str = data.get('nueva_fecha')
    if not ticket_id or not nueva_fecha_str:
        return JsonResponse({
            'success': False,
            'error': 'Se requiere ticket_id y nueva_fecha',
        })

    try:
        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'error': 'Formato de fecha inválido (se espera YYYY-MM-DD)',
        })

    sucursal_id = get_sucursal_id(request)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'No hay sucursal seleccionada',
        })

    permisos = permisos_edicion_dte_context(request.user, sucursal_id)
    if not permisos.get('campo', {}).get('fecha'):
        return JsonResponse({
            'success': False,
            'error': 'No tiene permisos para editar fechas',
        }, status=403)

    with transaction.atomic():
        ticket = (
            Ticket.objects
            .select_for_update()
            .filter(pk=ticket_id, sucursal_id=sucursal_id, estado='PAGADO')
            .first()
        )
        if not ticket:
            return JsonResponse({
                'success': False,
                'error': 'Ticket no encontrado o no está pagado',
            })

        if ticket.fecha == nueva_fecha:
            return JsonResponse({
                'success': True,
                'mensaje': 'La fecha ya era la indicada',
                'cambio': False,
                'ticket_id': ticket.id,
                'fecha': nueva_fecha_str,
            })

        fecha_anterior = ticket.fecha
        Ticket.objects.filter(pk=ticket.pk).update(fecha=nueva_fecha)

        try:
            fechas_a_revisar = {fecha_anterior, nueva_fecha}
            arqueos_qs = ArqueoCaja.objects.filter(
                sucursal_id=sucursal_id,
                fecha_arqueo__in=fechas_a_revisar,
            ).exclude(estado='ABIERTO')
            for arq in arqueos_qs:
                ObservacionArqueo.objects.create(
                    arqueo=arq,
                    usuario=request.user,
                    tipo='SISTEMA',
                    texto=(
                        f'Cambio de fecha del Ticket #{ticket.correlativo}'
                        f' (sin DTE). '
                        f'Fecha: '
                        f'{fecha_anterior.strftime("%d/%m/%Y") if fecha_anterior else "—"}'
                        f' → {nueva_fecha.strftime("%d/%m/%Y")}. '
                        'Los teóricos del arqueo pueden requerir recálculo.'
                    ),
                    visible_para_cajera=True,
                )
        except Exception:
            pass

    return JsonResponse({
        'success': True,
        'mensaje': (
            f'Ticket #{ticket.correlativo} movido al '
            f'{nueva_fecha.strftime("%d/%m/%Y")}.'
        ),
        'cambio': True,
        'ticket_id': ticket.id,
        'fecha_anterior': (
            fecha_anterior.strftime('%Y-%m-%d') if fecha_anterior else None
        ),
        'fecha_nueva': nueva_fecha_str,
    })


@login_required
@require_POST
def guardar_cuadratura_completa(request):
    """
    Guardar cuadratura de caja completa con depósitos bancarios
    Sistema simplificado - prioriza sencillez
    """
    try:
        from datetime import datetime
        import json
        
        # Obtener datos del request
        fecha = request.POST.get('fecha')
        efectivo_teorico = int(request.POST.get('efectivo_teorico', 0))
        efectivo_real = int(request.POST.get('efectivo_real', 0))
        cierre_pos = int(request.POST.get('cierre_pos', 0))
        numero_lote = request.POST.get('numero_lote', '')
        observaciones = request.POST.get('observaciones', '')
        depositos_json = request.POST.get('depositos', '[]')
        cuadratura_completa_json = request.POST.get('cuadratura_completa', '{}')
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        # Validaciones
        if not fecha or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal son requeridos'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # Parsear JSON
        depositos_data = json.loads(depositos_json)
        cuadratura_completa = json.loads(cuadratura_completa_json)
        
        # Verificar si ya existe un arqueo para esta fecha
        arqueo_existente = ArqueoCaja.objects.filter(
            fecha_arqueo=fecha_obj,
            sucursal=sucursal
        ).first()
        
        if arqueo_existente:
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una cuadratura para el {fecha_obj}. Elimínala primero si deseas crear una nueva.'
            })
        
        # === CREAR ARQUEO DE CAJA ===
        # Primero crear con valores básicos
        arqueo = ArqueoCaja(
            fecha_arqueo=fecha_obj,
            sucursal=sucursal,
            usuario_responsable=request.user,
            
            # Totales teóricos del sistema
            total_efectivo_teorico=efectivo_teorico,
            total_tarjeta_debito_teorico=cuadratura_completa.get('total_tarjeta_debito', 0),
            total_tarjeta_credito_teorico=cuadratura_completa.get('total_tarjeta_credito', 0),
            total_transbank_teorico=cuadratura_completa.get('total_transbank', 0),
            # Tarjetas Comerciales (solo Hites)
            total_hites_teorico=cuadratura_completa.get('total_hites', 0),
            total_tarjetas_comerciales_teorico=cuadratura_completa.get('total_tarjetas_comerciales', 0),
            # Venta Internet (Falabella, Paris, Ripley, MercadoPago, Klap)
            total_falabella_teorico=cuadratura_completa.get('total_falabella', 0),
            total_paris_teorico=cuadratura_completa.get('total_paris', 0),
            total_ripley_teorico=cuadratura_completa.get('total_ripley', 0),
            total_mercadopago_teorico=cuadratura_completa.get('total_mercadopago', 0),
            total_klap_teorico=cuadratura_completa.get('total_klap', 0),
            total_venta_internet_teorico=cuadratura_completa.get('total_venta_internet', 0),
            # Otros
            total_transferencia_teorico=cuadratura_completa.get('total_transferencia', 0),
            total_credito_trabajador_teorico=cuadratura_completa.get('total_credito_trabajador', 0),
            total_descuento_puntos_teorico=cuadratura_completa.get('total_descuento_puntos', 0),

            # Cierre POS
            numero_lote_pos=numero_lote,
            
            # Observaciones
            observaciones=observaciones,

            fondo_fijo_snapshot=sucursal.fondo_fijo_caja,
            fecha_cierre=timezone.now()
        )
        
        # Guardar primero para obtener el ID
        arqueo.save()
        
        # Los CONTADOS (físico) sí vienen del cajero: son el resultado de contar
        # la caja. Se guardan con update() para que `save()` no los recalcule
        # desde las denominaciones (que aquí van vacías).
        ArqueoCaja.objects.filter(id=arqueo.id).update(
            total_efectivo_fisico=efectivo_real,
            cierre_pos_fisico=cierre_pos,
        )
        arqueo.refresh_from_db()

        # Los TEÓRICOS, en cambio, no pueden venir del navegador: enviando
        # teórico == contado se cerraba cualquier caja con diferencia $0. Se
        # re-snapshotean desde la cuadratura real del servidor, que es la misma
        # fuente que usan reabrir y recalcular.
        try:
            _recalcular_teoricos_arqueo(
                arqueo,
                usuario=request.user,
                registrar_bitacora=True,
                razon='cierre de caja (teóricos recalculados en el servidor)',
            )
            arqueo.refresh_from_db()
        except Exception:
            logger.exception(
                "No se pudieron recalcular los teoricos del arqueo id=%s; "
                "se conservan los enviados por el cliente", arqueo.id,
            )

        # El estado se decide con las diferencias ya recalculadas contra los
        # teóricos del servidor, no contra los del formulario.
        diferencia_efectivo = _to_int(arqueo.diferencia_efectivo)
        diferencia_transbank = _to_int(arqueo.diferencia_transbank)
        estado_final = 'CERRADO' if diferencia_efectivo == 0 and diferencia_transbank == 0 else 'CON_DIFERENCIAS'

        ArqueoCaja.objects.filter(id=arqueo.id).update(estado=estado_final)
        arqueo.refresh_from_db()

        logger.info(
            "Arqueo creado id=%s efectivo_teorico=%s efectivo_fisico=%s cierre_pos=%s transbank_teorico=%s",
            arqueo.id,
            arqueo.total_efectivo_teorico,
            arqueo.total_efectivo_fisico,
            arqueo.cierre_pos_fisico,
            arqueo.total_transbank_teorico,
        )
        
        # === CREAR DEPÓSITOS BANCARIOS ===
        deposito_confirmado_por_supervisor = getattr(request.user, 'rol', None) in [
            'administrador',
            'administracion',
        ]
        depositos_creados = []
        for dep in depositos_data:
            try:
                fecha_dep = datetime.strptime(dep['fecha'], '%Y-%m-%d').date()
                monto_dep = int(dep['monto'])
                deposito = DepositoBancario.objects.create(
                    arqueo=arqueo,
                    fecha_deposito=fecha_dep,
                    monto=monto_dep if deposito_confirmado_por_supervisor else 0,
                    monto_declarado=monto_dep,
                    monto_confirmado=monto_dep if deposito_confirmado_por_supervisor else 0,
                    banco=dep.get('banco', 'ESTADO'),
                    numero_comprobante=dep.get('comprobante', ''),
                    observaciones=dep.get('observaciones', ''),
                    declarado_por=request.user,
                    fecha_declaracion=timezone.now(),
                    registrado_por=request.user,
                    verificado=deposito_confirmado_por_supervisor,
                    verificado_por=request.user if deposito_confirmado_por_supervisor else None,
                    fecha_verificacion=timezone.now() if deposito_confirmado_por_supervisor else None,
                )
                depositos_creados.append({
                    'id': deposito.id,
                    'monto': deposito.monto,
                    'banco': deposito.get_banco_display()
                })
            except Exception as e:
                logger.exception(
                    "Error al crear deposito para arqueo_id=%s datos=%s",
                    arqueo.id,
                    dep,
                )
                continue
        
        log_accion_caja(request, 'GUARDAR_CONTEO', arqueo)

        return JsonResponse({
            'success': True,
            'message': '¡Cuadratura guardada exitosamente!',
            'arqueo_id': arqueo.id,
            'depositos_creados': len(depositos_creados),
            'diferencia': arqueo.diferencia_efectivo,
            'estado': arqueo.get_estado_display()
        })
        
    except Exception as e:
        logger.exception(
            "Error al guardar cuadratura user_id=%s sucursal_id=%s",
            request.user.id,
            request.session.get('idSucursalActual') or request.session.get('sucursalActual'),
        )
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar cuadratura: {str(e)}'
        })


@login_required
@require_GET
def verificar_cuadratura_existente(request):
    """Verificar si ya existe una cuadratura para la fecha dada"""
    try:
        from datetime import datetime
        
        fecha = request.GET.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha or not sucursal_id:
            return JsonResponse({
                'existe': False
            })
        
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        arqueo = ArqueoCaja.objects.filter(
            fecha_arqueo=fecha_obj,
            sucursal=sucursal
        ).first()
        
        if arqueo:
            return JsonResponse({
                'existe': True,
                'datos': {
                    'id': arqueo.id,
                    'fecha_arqueo': arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
                    'usuario': arqueo.usuario_responsable.get_full_name() or arqueo.usuario_responsable.username,
                    'estado': arqueo.get_estado_display(),
                    'diferencia_efectivo': arqueo.diferencia_efectivo,
                    'efectivo_teorico': arqueo.total_efectivo_teorico,
                    'efectivo_fisico': arqueo.total_efectivo_fisico
                }
            })
        else:
            return JsonResponse({
                'existe': False
            })
            
    except Exception:
        logger.exception("Error al verificar cuadratura existente")
        return JsonResponse({
            'existe': False
        })


@login_required
@require_POST
def eliminar_cuadratura(request, arqueo_id):
    """Eliminar una cuadratura existente"""
    try:
        # Verificar permisos
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ('administrador', 'administracion'):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para eliminar cuadraturas. Se requiere rol de Administración o Administrador.'
            }, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        arqueo = get_object_or_404(
            ArqueoCaja,
            id=arqueo_id,
            sucursal_id=sucursal_id
        )

        if arqueo.estado not in ('ABIERTO',):
            return JsonResponse({
                'success': False,
                'error': f'Solo se pueden eliminar arqueos en estado Abierto. Estado actual: {arqueo.get_estado_display()}'
            })

        # Guardar info antes de eliminar
        fecha_arqueo = arqueo.fecha_arqueo.strftime('%d/%m/%Y')

        # Registrar auditoría antes de eliminar
        log_accion_caja(request, 'ELIMINAR_ARQUEO', arqueo, fecha=fecha_arqueo)

        # Eliminar (los depósitos se eliminan automáticamente por CASCADE)
        arqueo.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Cuadratura del {fecha_arqueo} eliminada exitosamente'
        })
        
    except Exception as e:
        logger.exception("Error al eliminar cuadratura %s", arqueo_id)
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar: {str(e)}'
        })


@login_required
@require_GET
def obtener_sucursales(request):
    """Obtener listado de todas las sucursales"""
    try:
        sucursales = Sucursal.objects.all().order_by('alias')
        
        sucursales_data = []
        for suc in sucursales:
            sucursales_data.append({
                'id': suc.id,
                'alias': suc.alias,
                'nombre': suc.nombre,
                'direccion': suc.direccion
            })
        
        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener sucursales: {str(e)}'
        })


@login_required
@require_GET
def listar_cuadraturas(request):
    """Listar cuadraturas/arqueos con filtros.

    Optimización clave:
    - Paginación server-side (por defecto 20 items / página).
    - Precomputa pagos por (fecha, método) en 2 queries (tickets + DTEs)
      agrupando por `Ticket.fecha` / `Dte.fecha_emision` y mapeando en
      memoria. Antes se hacía un query por cada arqueo + uno por cada
      ticket (N×M).
    - `total_depositos` y `cantidad_depositos` se resuelven con `annotate`
      en lugar de properties que gatillan queries.
    - Los DTE que ya están representados por un ticket (Ticket.folio_dte)
      se excluyen del agregado de Dte_Detalle_Pago para no doble-contar.
    """
    try:
        from datetime import datetime

        fecha_filtro = request.GET.get('fecha')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        sucursal_actual_id = get_sucursal_id(request)
        if not sucursal_actual_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay una sucursal activa en la sesión'
            }, status=400)

        # Query base — siempre restringida a la sucursal de la sesión + annotate
        arqueos_qs = (
            ArqueoCaja.objects
            .filter(sucursal_id=sucursal_actual_id)
            .select_related('usuario_responsable', 'sucursal')
            .annotate(
                ann_total_depositos=Sum('depositos__monto'),
                ann_cantidad_depositos=Count('depositos'),
            )
            .order_by('-fecha_arqueo')
        )

        # Rango de fechas a evaluar para precomputar pagos
        if fecha_filtro:
            fecha_obj = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
            arqueos_qs = arqueos_qs.filter(fecha_arqueo=fecha_obj)
            rango_inicio = rango_fin = fecha_obj
        else:
            hoy = timezone.localdate()
            primer_dia_mes = hoy.replace(day=1)
            arqueos_qs = arqueos_qs.filter(
                fecha_arqueo__gte=primer_dia_mes,
                fecha_arqueo__lte=hoy,
            )
            rango_inicio = primer_dia_mes
            rango_fin = hoy

        paginator = Paginator(arqueos_qs, per_page)
        arqueos_page = paginator.get_page(page)

        # Fechas únicas de arqueos de la página actual (para precomputar pagos)
        fechas_arqueo = sorted({a.fecha_arqueo for a in arqueos_page.object_list})
        if not fechas_arqueo:
            return JsonResponse({
                'success': True,
                'arqueos': [],
                'pagination': {
                    'current_page': arqueos_page.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': arqueos_page.has_next(),
                    'has_previous': arqueos_page.has_previous(),
                }
            })

        # --- Pagos de tickets por fecha y método (1 query) ---
        #
        # Importante: agrupar por `ticket__fecha` (DateField) y NO por
        # `ticket__created_at`. `editar_dte_boleta_papel` sincroniza
        # `Ticket.fecha` al editar un DTE, pero `created_at` queda con el
        # valor original; si agrupáramos por `created_at` un cambio de
        # fecha desde gestión de documentos no se reflejaría en los
        # teóricos de la tabla.
        #
        # También se acota a `estado='PAGADO'`: el modal de Resumen
        # (`_calcular_cuadratura_data`) y el detalle de arqueo
        # (`obtener_detalle_arqueo`) usan el mismo filtro, así que incluir
        # PENDIENTE_PAGO / PARCIALMENTE_PAGADO mostraba efectivo teórico
        # aquí que no aparecía en las otras vistas.
        pagos_ticket_qs = (
            TicketDetallePago.objects
            .filter(
                ticket__sucursal_id=sucursal_actual_id,
                ticket__fecha__gte=fechas_arqueo[0],
                ticket__fecha__lte=fechas_arqueo[-1],
                ticket__estado='PAGADO',
            )
            .values('ticket__fecha', 'metodo_pago')
            .annotate(total=Sum('monto'))
        )

        # Folios de DTE que ya están contados vía ticket (Ticket.folio_dte):
        # sus pagos se leen desde TicketDetallePago, así que hay que
        # excluirlos del agregado de Dte_Detalle_Pago para evitar el doble
        # conteo. Mismo criterio que `_calcular_cuadratura_data`.
        folios_con_ticket = set(
            Ticket.objects
            .filter(
                sucursal_id=sucursal_actual_id,
                fecha__gte=fechas_arqueo[0],
                fecha__lte=fechas_arqueo[-1],
                folio_dte__isnull=False,
                estado='PAGADO',
            )
            .values_list('folio_dte', flat=True)
        )

        # --- Pagos de DTEs por fecha y método (1 query) ---
        pagos_dte_qs = (
            Dte_Detalle_Pago.objects
            .filter(
                dte__sucursal_id=sucursal_actual_id,
                dte__fecha_emision__gte=fechas_arqueo[0],
                dte__fecha_emision__lte=fechas_arqueo[-1],
            )
            .exclude(dte__estado_dte='ANULADO')
            .exclude(dte__numero_documento__in=folios_con_ticket)
            .values('dte__fecha_emision', 'metodo_pago')
            .annotate(total=Sum('monto'))
        )

        # Acumular en diccionarios: { fecha: { 'EFECTIVO': 123, 'TARJETA_DEBITO': 456, ... } }
        pagos_por_fecha: dict = {}
        for row in pagos_ticket_qs:
            f = row['ticket__fecha']
            pagos_por_fecha.setdefault(f, {})
            pagos_por_fecha[f][row['metodo_pago']] = (
                pagos_por_fecha[f].get(row['metodo_pago'], 0) + (row['total'] or 0)
            )
        for row in pagos_dte_qs:
            f = row['dte__fecha_emision']
            pagos_por_fecha.setdefault(f, {})
            pagos_por_fecha[f][row['metodo_pago']] = (
                pagos_por_fecha[f].get(row['metodo_pago'], 0) + (row['total'] or 0)
            )

        METODOS_TRANSBANK = {'TARJETA_DEBITO', 'TARJETA_CREDITO', 'TARJETA'}

        datos = []
        for arqueo in arqueos_page.object_list:
            por_metodo = pagos_por_fecha.get(arqueo.fecha_arqueo, {})

            efectivo_teo = por_metodo.get('EFECTIVO', 0)
            transbank_teo = sum(
                por_metodo.get(m, 0) for m in METODOS_TRANSBANK
            )
            convenio_teo = por_metodo.get('CONVENIO', 0)
            credito_trab_teo = por_metodo.get('CREDITO_TRABAJADOR', 0)
            credito_ext_teo = por_metodo.get('CREDITO_EXTERNO', 0)

            total_depositos = arqueo.ann_total_depositos or 0
            efectivo_fisico = arqueo.total_efectivo_fisico or 0
            total_efectivo_real = efectivo_fisico + total_depositos
            fondo_fijo = arqueo.fondo_fijo_snapshot or 0

            diferencia_efectivo = total_efectivo_real - (efectivo_teo + fondo_fijo)
            diferencia_cajero = efectivo_fisico - (efectivo_teo + fondo_fijo)
            diferencia_transbank = (arqueo.cierre_pos_fisico or 0) - transbank_teo
            diferencia_total = diferencia_efectivo + diferencia_transbank

            datos.append({
                'id': arqueo.id,
                'fecha_arqueo': arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
                'sucursal': arqueo.sucursal.alias if arqueo.sucursal else 'N/A',
                'sucursal_id': arqueo.sucursal.id if arqueo.sucursal else None,
                'usuario': (
                    arqueo.usuario_responsable.get_full_name()
                    or arqueo.usuario_responsable.username
                ),
                'efectivo_teorico': efectivo_teo,
                'efectivo_fisico': efectivo_fisico,
                'total_depositos': total_depositos,
                'total_efectivo_real': total_efectivo_real,
                'diferencia_efectivo': diferencia_efectivo,
                'diferencia_efectivo_real': diferencia_efectivo,
                'diferencia_cajero': diferencia_cajero,
                'total_transbank_teorico': transbank_teo,
                'total_convenio_teorico': convenio_teo,
                'total_credito_trabajador_teorico': credito_trab_teo,
                'total_credito_externo_teorico': credito_ext_teo,
                'cierre_pos_fisico': arqueo.cierre_pos_fisico,
                'numero_lote_pos': arqueo.numero_lote_pos or '',
                'diferencia_transbank': diferencia_transbank,
                'diferencia_total_real': diferencia_total,
                'estado': arqueo.get_estado_display(),
                'estado_codigo': arqueo.estado,
                'observaciones': arqueo.observaciones,
                'cantidad_depositos': arqueo.ann_cantidad_depositos or 0,
                'fondo_fijo': fondo_fijo,
            })

        return JsonResponse({
            'success': True,
            'arqueos': datos,
            'pagination': {
                'current_page': arqueos_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': arqueos_page.has_next(),
                'has_previous': arqueos_page.has_previous(),
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@require_GET
def obtener_detalle_arqueo(request, arqueo_id):
    """Obtener detalle completo de un arqueo específico - RECALCULA VALORES TEÓRICOS EN TIEMPO REAL"""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        
        qs = ArqueoCaja.objects.select_related('usuario_responsable', 'sucursal').prefetch_related('depositos')
        if es_supervisor:
            arqueo = get_object_or_404(qs, id=arqueo_id)
            sucursal_id = arqueo.sucursal_id
        else:
            arqueo = get_object_or_404(qs, id=arqueo_id, sucursal_id=sucursal_id)
        
        # ========== RECALCULAR VALORES TEÓRICOS EN TIEMPO REAL ==========
        #
        # Se delega en `_calcular_cuadratura_data` para que el detalle del
        # arqueo muestre EXACTAMENTE los mismos totales que el modal
        # "Resumen de Caja". Antes estaba duplicada la lógica y faltaban
        # casos (NC en efectivo, tipo_transaccion='DEVOLUCION'), por lo
        # que el detalle de un arqueo cerrado podía diferir del Resumen
        # del mismo día.
        fecha_obj = arqueo.fecha_arqueo
        sucursal = arqueo.sucursal

        cuadratura = _calcular_cuadratura_data(
            sucursal, fecha_obj.strftime('%Y-%m-%d')
        )

        total_efectivo_teorico = cuadratura['total_efectivo']
        total_tarjeta_debito_teorico = cuadratura['total_tarjeta_debito']
        total_tarjeta_credito_teorico = cuadratura['total_tarjeta_credito']
        total_transbank_teorico = cuadratura['total_transbank']
        total_hites_teorico = cuadratura['total_hites']
        total_falabella_teorico = cuadratura['total_falabella']
        total_paris_teorico = cuadratura['total_paris']
        total_ripley_teorico = cuadratura['total_ripley']
        total_mercadopago_teorico = cuadratura['total_mercadopago']
        total_klap_teorico = cuadratura['total_klap']
        total_venta_internet_teorico = cuadratura['total_venta_internet']
        total_transferencia_teorico = cuadratura['total_transferencia']
        total_credito_trabajador_teorico = cuadratura['total_credito_trabajador']
        total_convenio_teorico = cuadratura['total_convenio']
        total_credito_externo_teorico = cuadratura['total_credito_externo']
        total_tarjetas_comerciales_teorico = cuadratura['total_tarjetas_comerciales']

        # Calcular diferencias ACTUALIZADAS
        diferencia_efectivo = arqueo.total_efectivo_fisico - total_efectivo_teorico
        diferencia_transbank = arqueo.cierre_pos_fisico - total_transbank_teorico
        
        # Serializar depositos
        depositos_data = []
        for deposito in arqueo.depositos.all():
            depositos_data.append({
                'id': deposito.id,
                'fecha_deposito': deposito.fecha_deposito.strftime('%d/%m/%Y'),
                'fecha_deposito_iso': deposito.fecha_deposito.strftime('%Y-%m-%d'),
                'monto': deposito.monto,
                'monto_declarado': deposito.monto_declarado,
                'monto_confirmado': deposito.monto_confirmado,
                'diferencia_deposito': deposito.diferencia_deposito,
                'verificado': deposito.verificado,
                'banco': deposito.banco,
                'banco_display': deposito.get_banco_display(),
                'numero_comprobante': deposito.numero_comprobante,
                'observaciones': deposito.observaciones,
                'declarado_por': deposito.declarado_por.get_full_name() if deposito.declarado_por else '',
                'fecha_declaracion': deposito.fecha_declaracion.strftime('%d/%m/%Y %H:%M') if deposito.fecha_declaracion else '',
                'verificado_por': deposito.verificado_por.get_full_name() if deposito.verificado_por else '',
                'registrado_por': deposito.registrado_por.username if deposito.registrado_por else '',
                'fecha_registro': deposito.fecha_registro.strftime('%d/%m/%Y %H:%M') if deposito.fecha_registro else ''
            })
        
        # Venta total: usar el valor calculado por `_calcular_cuadratura_data`
        # (ventas netas = boletas + facturas - notas de crédito). Es el
        # mismo número que se muestra en el modal "Resumen de Caja".
        venta_total = cuadratura['venta_total']
        
        arqueo_data = {
            'id': arqueo.id,
            'fecha_arqueo': arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
            'usuario': arqueo.usuario_responsable.get_full_name() or arqueo.usuario_responsable.username,
            'sucursal': arqueo.sucursal.alias if arqueo.sucursal else 'N/A',
            'estado': arqueo.estado,
            'observaciones': arqueo.observaciones,
            
            # Totales RECALCULADOS
            'venta_total': venta_total,
            'total_efectivo_teorico': total_efectivo_teorico,
            'total_efectivo_fisico': arqueo.total_efectivo_fisico,
            'diferencia_efectivo': diferencia_efectivo,
            
            # Transbank POS RECALCULADO
            'total_transbank_teorico': total_transbank_teorico,
            'cierre_pos_fisico': arqueo.cierre_pos_fisico,
            'diferencia_transbank': diferencia_transbank,
            'numero_lote_pos': arqueo.numero_lote_pos or '',
            
            # Transbank Detalle RECALCULADO
            'total_tarjeta_debito_teorico': total_tarjeta_debito_teorico,
            'total_tarjeta_credito_teorico': total_tarjeta_credito_teorico,
            
            # Tarjetas Comerciales RECALCULADO (solo Hites)
            'total_hites_teorico': total_hites_teorico,
            'total_tarjetas_comerciales_teorico': total_tarjetas_comerciales_teorico,
            
            # Venta Internet RECALCULADO
            'total_falabella_teorico': total_falabella_teorico,
            'total_paris_teorico': total_paris_teorico,
            'total_ripley_teorico': total_ripley_teorico,
            'total_mercadopago_teorico': total_mercadopago_teorico,
            'total_klap_teorico': total_klap_teorico,
            'total_venta_internet_teorico': total_venta_internet_teorico,
            
            # Otros RECALCULADO
            'total_transferencia_teorico': total_transferencia_teorico,
            'total_credito_trabajador_teorico': total_credito_trabajador_teorico,
            'total_convenio_teorico': total_convenio_teorico,
            'total_credito_externo_teorico': total_credito_externo_teorico,
            
            # Depósitos
            'depositos': depositos_data,
            # Revisión
            'resultado_revision': getattr(arqueo, 'resultado_revision', 'PENDIENTE'),
            'observaciones_supervisor': arqueo.observaciones_supervisor or '',
            'supervisor': arqueo.supervisor_revision.get_full_name() if arqueo.supervisor_revision else '',
            'fecha_revision': arqueo.fecha_revision.strftime('%d/%m/%Y %H:%M') if arqueo.fecha_revision else '',
            # Bitácora
            'bitacora': [{
                'id': obs.id,
                'tipo': obs.tipo,
                'tipo_display': obs.get_tipo_display(),
                'texto': obs.texto,
                'usuario': obs.usuario.get_full_name() or obs.usuario.username,
                'fecha': obs.fecha.strftime('%d/%m/%Y %H:%M'),
                'visible_para_cajera': obs.visible_para_cajera,
            } for obs in arqueo.bitacora.all()[:20]],
        }
        
        logger.debug(
            "Detalle de arqueo %s recalculado: efectivo_teorico=%s, transbank_teorico=%s, venta_total=%s",
            arqueo_id,
            total_efectivo_teorico,
            total_transbank_teorico,
            venta_total,
        )
        
        return JsonResponse({
            'success': True,
            'arqueo': arqueo_data
        })
        
    except Exception as e:
        logger.exception("Error al obtener detalle de arqueo %s", arqueo_id)
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@require_POST
def agregar_deposito_arqueo(request):
    """Agregar un depósito bancario a un arqueo existente (solo supervisores)"""
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        if not es_supervisor:
            return JsonResponse({'success': False, 'error': 'No tiene permisos para registrar depósitos directamente. Solo supervisores.'}, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        arqueo_id = request.POST.get('arqueo_id')
        fecha_deposito = request.POST.get('fecha_deposito')
        monto = int(request.POST.get('monto', 0))
        banco = request.POST.get('banco')
        numero_comprobante = request.POST.get('numero_comprobante', '')
        observaciones = request.POST.get('observaciones', '')
        
        # Validaciones
        if not arqueo_id or not fecha_deposito or not monto or not banco:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos'
            })
        
        # Obtener el arqueo
        arqueo = get_object_or_404(
            ArqueoCaja,
            id=arqueo_id,
            sucursal_id=sucursal_id
        )
        
        # Crear el depósito
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha_deposito, '%Y-%m-%d').date()
        
        deposito = DepositoBancario.objects.create(
            arqueo=arqueo,
            fecha_deposito=fecha_obj,
            monto=monto,
            monto_declarado=monto,
            monto_confirmado=monto,
            banco=banco,
            numero_comprobante=numero_comprobante,
            observaciones=observaciones,
            declarado_por=request.user,
            fecha_declaracion=timezone.now(),
            registrado_por=request.user,
            verificado=True,
            verificado_por=request.user,
            fecha_verificacion=timezone.now(),
        )

        arqueo.refresh_from_db()
        
        # Usar properties del modelo para cálculos
        total_depositos = arqueo.total_depositos
        efectivo_en_caja = arqueo.efectivo_en_caja
        diferencia_efectivo_real = arqueo.diferencia_efectivo_real
        diferencia_total_real = arqueo.diferencia_total_real
        
        # Recalcular estado basado en la diferencia REAL (considerando depósitos)
        if abs(diferencia_efectivo_real) <= 1000 and abs(arqueo.diferencia_transbank) <= 1000:
            arqueo.estado = 'CERRADO'
        else:
            arqueo.estado = 'CON_DIFERENCIAS'
        
        ArqueoCaja.objects.filter(id=arqueo.id).update(estado=arqueo.estado)
        
        logger.info(
            "Deposito agregado al arqueo %s: monto=%s, total_depositos=%s, efectivo_fisico=%s, "
            "efectivo_en_caja=%s, efectivo_teorico=%s, diferencia_efectivo_real=%s, "
            "diferencia_total_real=%s, estado=%s",
            arqueo_id,
            monto,
            total_depositos,
            arqueo.total_efectivo_fisico,
            efectivo_en_caja,
            arqueo.total_efectivo_teorico,
            diferencia_efectivo_real,
            diferencia_total_real,
            arqueo.estado,
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Depósito agregado correctamente',
            'deposito': {
                'id': deposito.id,
                'fecha': deposito.fecha_deposito.strftime('%d/%m/%Y'),
                'monto': deposito.monto,
                'banco': deposito.get_banco_display()
            },
            'arqueo_actualizado': {
                'total_depositos': total_depositos,
                'efectivo_fisico': arqueo.total_efectivo_fisico,
                'efectivo_en_caja': efectivo_en_caja,
                'efectivo_teorico': arqueo.total_efectivo_teorico,
                'diferencia_efectivo_real': diferencia_efectivo_real,
                'diferencia_transbank': arqueo.diferencia_transbank,
                'diferencia_total_real': diferencia_total_real,
                'estado': arqueo.get_estado_display(),
                'estado_codigo': arqueo.estado
            }
        })
        
    except Exception as e:
        logger.exception("Error al agregar deposito al arqueo")
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@require_POST
def eliminar_deposito_bancario(request):
    """Eliminar un depósito bancario específico (solo supervisores o quien lo declaró si aún no está verificado)"""
    try:
        from .models import PermisoRol

        data = json.loads(request.body)
        deposito_id = data.get('deposito_id')
        
        if not deposito_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de depósito requerido'
            })
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        # Permiso de eliminación: flag "Eliminar" (puede_eliminar) de la opción
        # Revisión Arqueos y Depósitos (revision_arqueos), gestionable por rol
        # en la UI de permisos.
        puede_eliminar = PermisoRol.tiene_permiso(
            request.user, 'revision_arqueos', 'puede_eliminar',
            sucursal_id=int(sucursal_id) if sucursal_id else None,
        )

        # Obtener el depósito
        deposito = get_object_or_404(DepositoBancario, id=deposito_id)

        # Verificar que el arqueo pertenece a la sucursal actual
        if deposito.arqueo.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para eliminar este depósito'
            })

        # Sin permiso de eliminar: solo se permite borrar el PROPIO depósito aún
        # no verificado (caso cajero). Con el permiso, puede borrar cualquiera.
        if not puede_eliminar:
            es_propio = getattr(deposito, 'declarado_por_id', None) == request.user.id
            if not es_propio or deposito.verificado:
                return JsonResponse({
                    'success': False,
                    'error': 'Solo puede eliminar sus propios depósitos no verificados. Solicite a un supervisor.'
                }, status=403)

        arqueo = deposito.arqueo
        monto_eliminado = deposito.monto
        
        # Eliminar el depósito
        deposito.delete()
        
        # Usar properties del modelo para cálculos
        total_depositos = arqueo.total_depositos
        efectivo_en_caja = arqueo.efectivo_en_caja
        diferencia_efectivo_real = arqueo.diferencia_efectivo_real
        diferencia_total_real = arqueo.diferencia_total_real
        
        # Recalcular el estado según el veredicto canónico del conteo.
        #
        # Antes se usaba `diferencia_efectivo_real`, que resta los depósitos al
        # efectivo físico: como el conteo se hace ANTES de depositar, eso
        # descontaba la plata dos veces y un día perfectamente cuadrado pasaba a
        # CON_DIFERENCIAS por el solo hecho de haber depositado. Y se aplicaba
        # tolerancia ±$1.000 mientras `cerrar_arqueo` exigía $0 exacto, así que
        # un faltante real de $999 se blanqueaba al tocar un depósito.
        #
        # Un arqueo ya revisado no se degrada: la revisión del supervisor es un
        # hecho auditado y borrar un depósito mal cargado no debe deshacerla.
        if arqueo.estado != 'REVISADO':
            nuevo_estado_arqueo = (
                'CERRADO'
                if abs(arqueo.diferencia_efectivo or 0) <= TOLERANCIA_ARQUEO_EFECTIVO
                else 'CON_DIFERENCIAS'
            )
        else:
            nuevo_estado_arqueo = arqueo.estado

        # `.update()` y NO `save()`: `ArqueoCaja.save()` recalcula
        # `total_efectivo_fisico` sumando billetes y monedas, que en modo
        # EXPRESS están todos en 0. Este endpoint era el único del módulo que
        # se saltaba esa regla, así que borrar un depósito de un arqueo express
        # dejaba el efectivo contado en $0 y la diferencia en -(teórico+fondo).
        ArqueoCaja.objects.filter(id=arqueo.id).update(estado=nuevo_estado_arqueo)
        arqueo.estado = nuevo_estado_arqueo


        logger.info(
            "Deposito %s eliminado del arqueo %s: monto=%s, total_depositos=%s, "
            "efectivo_en_caja=%s, diferencia_efectivo_real=%s, diferencia_total_real=%s, estado=%s",
            deposito_id,
            arqueo.id,
            monto_eliminado,
            total_depositos,
            efectivo_en_caja,
            diferencia_efectivo_real,
            diferencia_total_real,
            arqueo.estado,
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Depósito eliminado correctamente',
            'arqueo_actualizado': {
                'total_depositos': total_depositos,
                'efectivo_fisico': arqueo.total_efectivo_fisico,
                'efectivo_en_caja': efectivo_en_caja,
                'efectivo_teorico': arqueo.total_efectivo_teorico,
                'diferencia_efectivo_real': diferencia_efectivo_real,
                'diferencia_transbank': arqueo.diferencia_transbank,
                'diferencia_total_real': diferencia_total_real,
                'estado': arqueo.get_estado_display(),
                'estado_codigo': arqueo.estado
            }
        })
        
    except Exception as e:
        logger.exception("Error al eliminar deposito %s", deposito_id)
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@require_POST
def cambiar_fecha_deposito(request):
    """
    Corrige la fecha de un depósito bancario existente. Gateado por el flag
    'Editar' (puede_editar) de la opción Revisión Arqueos y Depósitos
    (revision_arqueos), gestionable por rol en la UI de permisos.
    Solo cambia la fecha: no toca montos ni el cache del arqueo.
    """
    from .models import PermisoRol
    from datetime import datetime
    try:
        data = json.loads(request.body)
        deposito_id = data.get('deposito_id')
        nueva_fecha = data.get('fecha_deposito') or data.get('fecha')

        if not deposito_id or not nueva_fecha:
            return JsonResponse({'success': False, 'error': 'deposito_id y fecha_deposito son requeridos'}, status=400)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        if not PermisoRol.tiene_permiso(
            request.user, 'revision_arqueos', 'puede_editar',
            sucursal_id=int(sucursal_id) if sucursal_id else None,
        ):
            return JsonResponse({'success': False, 'error': 'No tiene permiso para cambiar la fecha de depósitos.'}, status=403)

        deposito = get_object_or_404(DepositoBancario, id=deposito_id)

        if sucursal_id and deposito.arqueo.sucursal_id != int(sucursal_id):
            return JsonResponse({'success': False, 'error': 'El depósito no pertenece a su sucursal.'}, status=403)

        try:
            fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Fecha inválida (formato AAAA-MM-DD).'}, status=400)

        fecha_anterior = deposito.fecha_deposito
        deposito.fecha_deposito = fecha_obj
        deposito.save(update_fields=['fecha_deposito'])

        try:
            log_accion_caja(request, 'EDITAR_FECHA_DEPOSITO', deposito.arqueo)
        except Exception:
            pass
        logger.info("Fecha de deposito %s cambiada: %s -> %s por %s",
                    deposito_id, fecha_anterior, fecha_obj, request.user)

        return JsonResponse({
            'success': True,
            'message': 'Fecha del depósito actualizada.',
            'fecha_deposito': fecha_obj.strftime('%Y-%m-%d'),
        })
    except Exception as e:
        logger.exception("Error al cambiar fecha de deposito")
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required
@require_POST
def declarar_deposito(request):
    """
    El cajero declara un depósito con comprobante bancario.
    Soporta FormData (multipart) para subir imagen del comprobante.
    Permite múltiples depósitos por arqueo (ej: efectivo + cheque).
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ('cajero', 'vendedor', 'jefe_local', 'administracion', 'administrador'):
            return JsonResponse({'success': False, 'error': 'No tiene permisos para declarar depósitos.'}, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        # Soportar tanto JSON como FormData
        if request.content_type and 'multipart' in request.content_type:
            arqueo_id = request.POST.get('arqueo_id')
            monto_declarado = int(request.POST.get('monto_declarado', 0))
            tipo_medio = request.POST.get('tipo_medio', 'EFECTIVO')
            banco = request.POST.get('banco', 'ESTADO')
            numero_comprobante = request.POST.get('numero_comprobante', '')
            observaciones = request.POST.get('observaciones', '')
            imagen_comprobante = request.FILES.get('imagen_comprobante')
        else:
            data = json.loads(request.body)
            arqueo_id = data.get('arqueo_id')
            monto_declarado = int(data.get('monto_declarado', 0))
            tipo_medio = data.get('tipo_medio', 'EFECTIVO')
            banco = data.get('banco', 'ESTADO')
            numero_comprobante = data.get('numero_comprobante', '')
            observaciones = data.get('observaciones', '')
            imagen_comprobante = None

        if not arqueo_id or monto_declarado <= 0:
            return JsonResponse({'success': False, 'error': 'Se requiere arqueo_id y monto_declarado > 0'})

        if tipo_medio not in ('EFECTIVO', 'CHEQUE'):
            return JsonResponse({'success': False, 'error': 'tipo_medio debe ser EFECTIVO o CHEQUE'})

        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id, sucursal_id=sucursal_id)

        # Validar que no se declare más de lo posible
        from django.db.models import Sum
        total_ya_declarado = arqueo.depositos.aggregate(total=Sum('monto_declarado'))['total'] or 0
        max_depositable = arqueo.total_efectivo_fisico + arqueo.total_cheque_teorico
        if max_depositable > 0 and (total_ya_declarado + monto_declarado) > max_depositable * 1.1:
            return JsonResponse({
                'success': False,
                'error': f'El total declarado (${total_ya_declarado + monto_declarado:,}) excede el máximo depositable (${max_depositable:,})'
            })

        from django.utils import timezone as tz
        deposito = DepositoBancario(
            arqueo=arqueo,
            fecha_deposito=arqueo.fecha_arqueo,
            monto=0,
            monto_declarado=monto_declarado,
            monto_confirmado=0,
            tipo_medio=tipo_medio,
            banco=banco,
            numero_comprobante=numero_comprobante,
            observaciones=observaciones,
            declarado_por=request.user,
            fecha_declaracion=tz.now(),
            verificado=False,
            registrado_por=request.user,
        )
        if imagen_comprobante:
            deposito.imagen_comprobante = imagen_comprobante
        deposito.save()

        log_accion_caja(request, 'DECLARAR_DEPOSITO', arqueo, monto=monto_declarado, tipo_medio=tipo_medio)

        return JsonResponse({
            'success': True,
            'message': 'Depósito declarado exitosamente. El supervisor deberá confirmarlo.',
            'deposito': {
                'id': deposito.id,
                'monto_declarado': deposito.monto_declarado,
                'tipo_medio': deposito.tipo_medio,
                'tipo_medio_display': deposito.get_tipo_medio_display(),
                'banco': deposito.banco,
                'banco_display': deposito.get_banco_display(),
                'numero_comprobante': deposito.numero_comprobante,
                'tiene_imagen': bool(deposito.imagen_comprobante),
                'declarado_por': request.user.get_full_name() or request.user.username,
                'fecha_declaracion': deposito.fecha_declaracion.strftime('%d/%m/%Y %H:%M'),
                'verificado': False,
            }
        })

    except Exception as e:
        logger.exception("Error al declarar deposito")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def finalizar_declaracion(request):
    """
    El cajero señala que terminó de declarar todos sus depósitos (efectivo, cheque, etc.)
    Transiciona el arqueo de CERRADO/CON_DIFERENCIAS → DEPOSITO_DECLARADO.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ('cajero', 'vendedor', 'jefe_local', 'administracion', 'administrador'):
            return JsonResponse({'success': False, 'error': 'No tiene permisos para finalizar declaración de depósitos.'}, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')

        if not arqueo_id:
            return JsonResponse({'success': False, 'error': 'Se requiere arqueo_id'})

        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id, sucursal_id=sucursal_id)

        # Verificar que el arqueo está en estado válido para finalizar declaración
        if arqueo.estado not in ('CERRADO', 'CON_DIFERENCIAS'):
            return JsonResponse({
                'success': False,
                'error': f'El arqueo debe estar Cerrado o Con Diferencias para finalizar declaración. Estado actual: {arqueo.get_estado_display()}'
            })

        # Verificar que exista al menos un depósito declarado
        depositos_declarados = arqueo.depositos.filter(monto_declarado__gt=0)
        if not depositos_declarados.exists():
            return JsonResponse({'success': False, 'error': 'Debe declarar al menos un depósito antes de finalizar'})

        # Transicionar estado
        ArqueoCaja.objects.filter(id=arqueo.id).update(estado='DEPOSITO_DECLARADO')
        arqueo.refresh_from_db()

        # Resumen de depósitos
        resumen = []
        for dep in depositos_declarados:
            resumen.append({
                'id': dep.id,
                'tipo_medio': dep.get_tipo_medio_display(),
                'monto_declarado': dep.monto_declarado,
                'banco': dep.get_banco_display(),
                'numero_comprobante': dep.numero_comprobante,
                'tiene_imagen': bool(dep.imagen_comprobante),
            })

        return JsonResponse({
            'success': True,
            'message': 'Declaración finalizada. Los depósitos serán revisados por el supervisor.',
            'estado': arqueo.estado,
            'estado_display': arqueo.get_estado_display(),
            'depositos': resumen,
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        logger.exception("Error al finalizar declaracion de depositos")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def confirmar_deposito(request, deposito_id):
    """
    El supervisor confirma un depósito declarado por el cajero.
    Los datos bancarios (banco, comprobante, imagen) ya vienen del cajero.
    El supervisor solo verifica visualmente, confirma monto y agrega observaciones si hay discrepancia.
    """
    try:
        if request.method not in ('POST', 'PATCH'):
            return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']

        if not es_supervisor:
            return JsonResponse({'success': False, 'error': 'Sin permisos. Solo supervisores pueden confirmar depósitos.'}, status=403)

        deposito = get_object_or_404(DepositoBancario, id=deposito_id)
        if deposito.arqueo.sucursal_id != int(sucursal_id):
            return JsonResponse({'success': False, 'error': 'Depósito no pertenece a su sucursal'}, status=403)

        # El supervisor confirma o ajusta el monto
        monto_confirmado = int(request.POST.get('monto_confirmado', request.POST.get('monto', 0)))
        observaciones_supervisor = request.POST.get('observaciones_supervisor', request.POST.get('observaciones', ''))

        # Fecha de depósito: usar la existente o una nueva si el supervisor la proporciona
        fecha_deposito = request.POST.get('fecha_deposito') or request.POST.get('fecha')

        if monto_confirmado <= 0:
            return JsonResponse({'success': False, 'error': 'Se requiere monto_confirmado > 0'})

        from datetime import datetime
        from django.utils import timezone as tz

        if fecha_deposito:
            fecha_obj = datetime.strptime(fecha_deposito, '%Y-%m-%d').date()
            deposito.fecha_deposito = fecha_obj

        deposito.monto = monto_confirmado
        deposito.monto_confirmado = monto_confirmado

        # Supervisor puede sobreescribir banco/comprobante si es necesario
        banco_override = request.POST.get('banco')
        if banco_override:
            deposito.banco = banco_override
        numero_comprobante_override = request.POST.get('numero_comprobante')
        if numero_comprobante_override:
            deposito.numero_comprobante = numero_comprobante_override

        if observaciones_supervisor:
            deposito.observaciones_supervisor = observaciones_supervisor

        if 'imagen_comprobante' in request.FILES:
            deposito.imagen_comprobante = request.FILES['imagen_comprobante']

        deposito.verificado = True
        deposito.verificado_por = request.user
        deposito.fecha_verificacion = tz.now()
        deposito.save()

        log_accion_caja(request, 'CONFIRMAR_DEPOSITO', deposito.arqueo, monto=monto_confirmado)

        arqueo = deposito.arqueo
        arqueo.refresh_from_db()
        total_depositos = arqueo.total_depositos
        efectivo_en_caja = arqueo.efectivo_en_caja
        diferencia_efectivo_real = arqueo.diferencia_efectivo_real

        # Verificar si todos los depósitos del arqueo están confirmados
        todos_confirmados = not arqueo.depositos.filter(verificado=False, monto_declarado__gt=0).exists()

        if todos_confirmados and arqueo.estado == 'DEPOSITO_DECLARADO':
            # Todos los depósitos confirmados → transicionar a DEPOSITO_CONFIRMADO
            ArqueoCaja.objects.filter(id=arqueo.id).update(estado='DEPOSITO_CONFIRMADO')
            arqueo.refresh_from_db()
        elif todos_confirmados:
            # Flujo legacy: si no pasó por DEPOSITO_DECLARADO, usar lógica original
            if abs(diferencia_efectivo_real) <= 1000 and abs(arqueo.diferencia_transbank) <= 1000:
                ArqueoCaja.objects.filter(id=arqueo.id).update(estado='CERRADO')
            else:
                ArqueoCaja.objects.filter(id=arqueo.id).update(estado='CON_DIFERENCIAS')
            arqueo.refresh_from_db()

        return JsonResponse({
            'success': True,
            'message': 'Depósito confirmado correctamente',
            'deposito': {
                'id': deposito.id,
                'monto_declarado': deposito.monto_declarado,
                'monto_confirmado': deposito.monto_confirmado,
                'diferencia': deposito.diferencia_deposito,
                'banco': deposito.get_banco_display(),
                'numero_comprobante': deposito.numero_comprobante,
                'verificado': True,
                'verificado_por': request.user.get_full_name() or request.user.username,
            },
            'arqueo_actualizado': {
                'total_depositos': total_depositos,
                'efectivo_en_caja': efectivo_en_caja,
                'diferencia_efectivo_real': diferencia_efectivo_real,
                'estado': arqueo.get_estado_display(),
                'estado_codigo': arqueo.estado,
            }
        })

    except Exception as e:
        logger.exception("Error al confirmar deposito %s", deposito_id)
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def obtener_depositos_pendientes(request):
    """Retorna depósitos declarados pero sin verificar para el panel del supervisor."""
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        if not es_supervisor:
            return JsonResponse({'success': True, 'depositos': [], 'total': 0})

        # Acepta `sucursal_id` por query param (validado contra las sucursales
        # del usuario). Antes leía SIEMPRE la sucursal de sesión, así que en la
        # pantalla de revisión el panel quedaba desincronizado con la sucursal
        # elegida en las pills.
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        pedida = request.GET.get('sucursal_id')
        if pedida and str(pedida).isdigit():
            permitidas, _ = _sucursales_permitidas(request)
            if not permitidas or int(pedida) in permitidas:
                sucursal_id = int(pedida)
        fecha_str = request.GET.get('fecha')

        qs = DepositoBancario.objects.filter(
            arqueo__sucursal_id=sucursal_id,
            verificado=False,
            monto_declarado__gt=0,
        ).select_related('arqueo', 'declarado_por').order_by('-fecha_declaracion')

        if fecha_str:
            qs = qs.filter(arqueo__fecha_arqueo=fecha_str)

        depositos = []
        for d in qs:
            depositos.append({
                'id': d.id,
                'arqueo_id': d.arqueo_id,
                'fecha_arqueo': d.arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
                'monto_declarado': d.monto_declarado,
                'tipo_medio': d.tipo_medio,
                'tipo_medio_display': d.get_tipo_medio_display(),
                'banco': d.banco,
                'banco_display': d.get_banco_display(),
                'numero_comprobante': d.numero_comprobante,
                'tiene_imagen': bool(d.imagen_comprobante),
                'imagen_url': d.imagen_comprobante.url if d.imagen_comprobante else '',
                'declarado_por': d.declarado_por.get_full_name() if d.declarado_por else '—',
                'fecha_declaracion': d.fecha_declaracion.strftime('%d/%m/%Y %H:%M') if d.fecha_declaracion else '—',
                'observaciones': d.observaciones,
            })

        return JsonResponse({'success': True, 'depositos': depositos, 'total': len(depositos)})

    except Exception as e:
        logger.exception("Error al obtener depositos pendientes")
        return JsonResponse({'success': False, 'error': str(e)})


# ========== DEPÓSITO MULTI-DÍA ==========

@login_required
@require_GET
def listar_arqueos_para_deposito(request):
    """
    Retorna arqueos de la sucursal que tienen efectivo pendiente de depositar.
    Se usa en el modal de depósito multi-día para elegir qué días incluir.
    Solo supervisores.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        if not es_supervisor:
            return JsonResponse({'success': False, 'error': 'Solo supervisores pueden acceder a depósitos multi-día.'}, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Sin sucursal'})

        arqueos = ArqueoCaja.objects.filter(
            sucursal_id=sucursal_id,
            estado__in=['CERRADO', 'CON_DIFERENCIAS', 'ABIERTO'],
        ).order_by('-fecha_arqueo')[:60]

        resultado = []
        for a in arqueos:
            efectivo_teorico = a.total_efectivo_teorico
            total_depositado = a.total_depositos
            pendiente = efectivo_teorico - total_depositado
            resultado.append({
                'id': a.id,
                'fecha': a.fecha_arqueo.strftime('%Y-%m-%d'),
                'fecha_display': a.fecha_arqueo.strftime('%d/%m/%Y'),
                'efectivo_teorico': efectivo_teorico,
                'efectivo_fisico': a.total_efectivo_fisico,
                'total_depositado': total_depositado,
                'pendiente_depositar': pendiente,
                'estado': a.get_estado_display(),
            })

        return JsonResponse({'success': True, 'arqueos': resultado})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
@csrf_exempt
def crear_deposito_multidia(request):
    """
    Crea un GrupoDeposito (1 comprobante bancario) con desglose por día.
    Valida que la suma del desglose coincida con el monto del comprobante.
    Solo supervisores.
    """
    try:
        import json
        from datetime import datetime

        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        if not es_supervisor:
            return JsonResponse({'success': False, 'error': 'Solo supervisores pueden crear depósitos multi-día.'}, status=403)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Sin sucursal'})

        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        fecha_deposito = request.POST.get('fecha_deposito')
        monto_total = int(request.POST.get('monto_total', 0))
        banco = request.POST.get('banco', 'ESTADO')
        numero_comprobante = request.POST.get('numero_comprobante', '')
        observaciones = request.POST.get('observaciones', '')
        desglose_json = request.POST.get('desglose', '[]')

        if not fecha_deposito or monto_total <= 0:
            return JsonResponse({'success': False, 'error': 'Fecha y monto son requeridos'})

        desglose = json.loads(desglose_json)
        if not desglose:
            return JsonResponse({'success': False, 'error': 'Debe incluir al menos un día en el desglose'})

        suma_desglose = sum(int(d.get('monto', 0)) for d in desglose)
        if abs(suma_desglose - monto_total) > 1:
            return JsonResponse({
                'success': False,
                'error': f'La suma del desglose (${suma_desglose:,}) no coincide con el monto del comprobante (${monto_total:,})'
            })
        # Si hay diferencia de exactamente $1, se ajusta el monto_total al valor real del desglose
        if suma_desglose != monto_total:
            monto_total = suma_desglose

        fecha_obj = datetime.strptime(fecha_deposito, '%Y-%m-%d').date()

        grupo = GrupoDeposito.objects.create(
            sucursal=sucursal,
            fecha_deposito=fecha_obj,
            monto_total=monto_total,
            banco=banco,
            numero_comprobante=numero_comprobante,
            observaciones=observaciones,
            registrado_por=request.user,
        )

        if 'imagen_comprobante' in request.FILES:
            grupo.imagen_comprobante = request.FILES['imagen_comprobante']
            grupo.save()

        depositos_creados = []
        for item in desglose:
            arqueo_id = int(item['arqueo_id'])
            monto_dia = int(item['monto'])

            arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id, sucursal=sucursal)

            dep = DepositoBancario.objects.create(
                arqueo=arqueo,
                grupo=grupo,
                fecha_deposito=fecha_obj,
                monto=monto_dia,
                monto_declarado=monto_dia,
                monto_confirmado=monto_dia,
                banco=banco,
                numero_comprobante=numero_comprobante,
                observaciones=f"Depósito multi-día (Grupo #{grupo.id})",
                declarado_por=request.user,
                fecha_declaracion=timezone.now(),
                registrado_por=request.user,
                verificado=True,
                verificado_por=request.user,
                fecha_verificacion=timezone.now(),
            )
            depositos_creados.append({
                'id': dep.id,
                'arqueo_id': arqueo.id,
                'fecha_arqueo': arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
                'monto': dep.monto,
            })

            arqueo.refresh_from_db()
            if abs(arqueo.diferencia_efectivo_real) <= 1000 and abs(arqueo.diferencia_transbank) <= 1000:
                estado_actualizado = 'CERRADO'
            else:
                estado_actualizado = 'CON_DIFERENCIAS'
            ArqueoCaja.objects.filter(id=arqueo.id).update(estado=estado_actualizado)

        return JsonResponse({
            'success': True,
            'message': f'Depósito multi-día registrado exitosamente ({len(depositos_creados)} días)',
            'grupo_id': grupo.id,
            'depositos': depositos_creados,
        })

    except Exception as e:
        logger.exception("Error al crear deposito multidia")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_GET
def detalle_grupo_deposito(request, grupo_id):
    """Retorna el detalle de un grupo de depósito con su desglose por día."""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        grupo = get_object_or_404(GrupoDeposito, id=grupo_id, sucursal_id=sucursal_id)

        depositos = []
        for d in grupo.depositos.select_related('arqueo').all():
            depositos.append({
                'id': d.id,
                'arqueo_id': d.arqueo_id,
                'fecha_arqueo': d.arqueo.fecha_arqueo.strftime('%d/%m/%Y'),
                'monto': d.monto,
                'efectivo_teorico': d.arqueo.total_efectivo_teorico,
            })

        return JsonResponse({
            'success': True,
            'grupo': {
                'id': grupo.id,
                'fecha_deposito': grupo.fecha_deposito.strftime('%d/%m/%Y'),
                'monto_total': grupo.monto_total,
                'banco': grupo.get_banco_display(),
                'numero_comprobante': grupo.numero_comprobante,
                'observaciones': grupo.observaciones,
                'esta_cuadrado': grupo.esta_cuadrado,
                'suma_desglose': grupo.suma_desglose,
                'diferencia': grupo.diferencia,
                'cantidad_dias': grupo.cantidad_dias,
                'registrado_por': grupo.registrado_por.get_full_name() or grupo.registrado_por.username,
                'fecha_registro': grupo.fecha_registro.strftime('%d/%m/%Y %H:%M'),
            },
            'depositos': depositos,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def editar_cuadratura(request, arqueo_id):
    """Editar una cuadratura existente"""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        # Obtener el arqueo
        arqueo = get_object_or_404(
            ArqueoCaja,
            id=arqueo_id,
            sucursal_id=sucursal_id
        )
        
        data = json.loads(request.body)
        logger.debug(
            "Editando arqueo %s. Datos recibidos=%s. Valores antes: efectivo_fisico=%s, cierre_pos=%s",
            arqueo_id,
            data,
            arqueo.total_efectivo_fisico,
            arqueo.cierre_pos_fisico,
        )
        
        # Actualizar efectivo físico
        if 'efectivo_real' in data:
            arqueo.total_efectivo_fisico = data['efectivo_real']
            arqueo.diferencia_efectivo = arqueo.total_efectivo_fisico - arqueo.total_efectivo_teorico
            logger.debug(
                "Efectivo actualizado en arqueo %s: fisico=%s, teorico=%s, diferencia=%s",
                arqueo_id,
                arqueo.total_efectivo_fisico,
                arqueo.total_efectivo_teorico,
                arqueo.diferencia_efectivo,
            )
        
        # Actualizar cierre POS
        if 'cierre_pos' in data:
            arqueo.cierre_pos_fisico = data['cierre_pos']
            arqueo.diferencia_transbank = arqueo.cierre_pos_fisico - arqueo.total_transbank_teorico
            logger.debug(
                "Cierre POS actualizado en arqueo %s: fisico=%s, teorico=%s, diferencia=%s",
                arqueo_id,
                arqueo.cierre_pos_fisico,
                arqueo.total_transbank_teorico,
                arqueo.diferencia_transbank,
            )
        
        if 'numero_lote' in data:
            arqueo.numero_lote_pos = data['numero_lote']
        
        # Actualizar observaciones
        if 'observaciones' in data:
            arqueo.observaciones = data['observaciones']
        
        # Actualizar depósitos (eliminar y recrear)
        if 'depositos' in data:
            # Eliminar depósitos anteriores
            arqueo.depositos.all().delete()
            
            # Crear nuevos depósitos
            deposito_confirmado_por_supervisor = getattr(request.user, 'rol', None) in [
                'administrador',
                'administracion',
            ]
            for dep in data['depositos']:
                monto_dep = int(dep['monto'])
                DepositoBancario.objects.create(
                    arqueo=arqueo,
                    fecha_deposito=dep['fecha'],
                    monto=monto_dep if deposito_confirmado_por_supervisor else 0,
                    monto_declarado=monto_dep,
                    monto_confirmado=monto_dep if deposito_confirmado_por_supervisor else 0,
                    banco=dep['banco'],
                    numero_comprobante=dep.get('comprobante', ''),
                    observaciones=dep.get('observaciones', ''),
                    declarado_por=request.user,
                    fecha_declaracion=timezone.now(),
                    registrado_por=request.user,
                    verificado=deposito_confirmado_por_supervisor,
                    verificado_por=request.user if deposito_confirmado_por_supervisor else None,
                    fecha_verificacion=timezone.now() if deposito_confirmado_por_supervisor else None,
                )
        
        # Recalcular estado
        if abs(arqueo.diferencia_efectivo) <= 1000 and abs(arqueo.diferencia_transbank) <= 1000:
            arqueo.estado = 'CERRADO'
        else:
            arqueo.estado = 'CON_DIFERENCIAS'
        
        # IMPORTANTE: Usar update() en lugar de save() para evitar que el método save() 
        # del modelo recalcule automáticamente el total_efectivo_fisico desde las denominaciones
        ArqueoCaja.objects.filter(id=arqueo.id).update(
            total_efectivo_fisico=arqueo.total_efectivo_fisico,
            diferencia_efectivo=arqueo.diferencia_efectivo,
            cierre_pos_fisico=arqueo.cierre_pos_fisico,
            diferencia_transbank=arqueo.diferencia_transbank,
            numero_lote_pos=arqueo.numero_lote_pos,
            observaciones=arqueo.observaciones,
            estado=arqueo.estado
        )
        
        # Recargar el objeto para verificar
        arqueo.refresh_from_db()
        
        logger.info(
            "Arqueo %s actualizado: efectivo_fisico=%s, cierre_pos=%s, diferencia_efectivo=%s, "
            "diferencia_transbank=%s, estado=%s",
            arqueo_id,
            arqueo.total_efectivo_fisico,
            arqueo.cierre_pos_fisico,
            arqueo.diferencia_efectivo,
            arqueo.diferencia_transbank,
            arqueo.estado,
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Cuadratura actualizada correctamente'
        })
        
    except Exception as e:
        logger.exception("Error al editar cuadratura %s", arqueo_id)
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@login_required
@require_GET
def exportar_cuadratura_excel(request):
    """Exportar cuadratura a Excel usando datos en tiempo real"""
    try:
        fecha = request.GET.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal requeridas'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Usar la función helper directamente para calcular la cuadratura
        cuadratura_data = _calcular_cuadratura_data(sucursal, fecha)
        
        # Convertir fecha string a date object
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # Crear Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Cuadratura {fecha}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = f"CUADRATURA DE CAJA - {sucursal.alias}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Datos de la cuadratura
        row = 4
        
        # Métodos de pago
        ws[f'A{row}'] = "MÉTODO DE PAGO"
        ws[f'B{row}'] = "MONTO"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        
        row += 1
        
        metodos_pago = [
            ('Efectivo', cuadratura_data.get('total_efectivo', 0)),
            ('Tarjeta Débito', cuadratura_data.get('total_tarjeta_debito', 0)),
            ('Tarjeta Crédito', cuadratura_data.get('total_tarjeta_credito', 0)),
            ('Transferencia', cuadratura_data.get('total_transferencia', 0)),
            ('Cheque', cuadratura_data.get('total_cheque', 0)),
            ('Convenio', cuadratura_data.get('total_convenio', 0)),
            ('VISA/MC/AMEX', cuadratura_data.get('total_visa_mc_amex', 0)),
            ('Presto', cuadratura_data.get('total_presto', 0)),
            ('AbcDin', cuadratura_data.get('total_abcdin', 0)),
            ('Tricot', cuadratura_data.get('total_tricot', 0)),
            ('Hites', cuadratura_data.get('total_hites', 0)),
            ('Ripley', cuadratura_data.get('total_ripley', 0)),
            ('Falabella', cuadratura_data.get('total_falabella', 0)),
            ('Paris', cuadratura_data.get('total_paris', 0)),
        ]
        
        for metodo, monto in metodos_pago:
            ws[f'A{row}'] = metodo
            ws[f'B{row}'] = monto
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

        # Documentos (referencia): montos por tipo de documento incluyendo
        # ventas cerradas con ticket (usamos los totales `_display`).
        row += 1
        ws[f'A{row}'] = "DOCUMENTO"
        ws[f'B{row}'] = "CANTIDAD"
        ws[f'C{row}'] = "MONTO"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
        row += 1
        documentos = [
            (
                'Boleta Electrónica',
                cuadratura_data.get('cantidad_boletas_electronicas', 0),
                cuadratura_data.get(
                    'total_boletas_electronicas_display',
                    cuadratura_data.get('total_boletas_electronicas', 0),
                ),
            ),
            (
                'Boleta Papel',
                cuadratura_data.get('cantidad_boletas_papel', 0),
                cuadratura_data.get(
                    'total_boletas_papel_display',
                    cuadratura_data.get('total_boletas_papel', 0),
                ),
            ),
            (
                'Factura Electrónica',
                cuadratura_data.get('cantidad_facturas', 0),
                cuadratura_data.get(
                    'total_facturas_display',
                    cuadratura_data.get('total_facturas', 0),
                ),
            ),
            (
                'Factura Exenta',
                cuadratura_data.get('cantidad_facturas_exentas', 0),
                cuadratura_data.get(
                    'total_facturas_exentas_display',
                    cuadratura_data.get('total_facturas_exentas', 0),
                ),
            ),
            (
                'Notas de Crédito',
                cuadratura_data.get('cantidad_notas_credito', 0),
                cuadratura_data.get('total_notas_credito', 0),
            ),
        ]
        for nombre, cantidad, monto in documentos:
            ws[f'A{row}'] = nombre
            ws[f'B{row}'] = cantidad
            ws[f'C{row}'] = monto
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1

        row += 1
        # Total
        ws[f'A{row}'] = "TOTAL VENTA"
        ws[f'B{row}'] = cuadratura_data.get('venta_total', 0)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        
        # Preparar respuesta
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cuadratura_{fecha}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


@login_required
@csrf_exempt
def obtener_transacciones_dia(request):
    """Obtener todas las transacciones del día (tickets, boletas, facturas)"""
    try:
        import json
        from datetime import datetime, time as dt_time
        
        data = json.loads(request.body)
        fecha = data.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal requeridas'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # Crear rango de fechas
        inicio_dia = timezone.make_aware(datetime.combine(fecha_obj, dt_time.min))
        fin_dia = timezone.make_aware(datetime.combine(fecha_obj, dt_time.max))
        
        # ========== OBTENER TICKETS ==========
        tickets_del_dia = Ticket.objects.filter(
            sucursal=sucursal,
            fecha=fecha_obj,
            estado='PAGADO'
        ).prefetch_related('pagos').order_by('created_at')
        
        tickets_data = []
        for ticket in tickets_del_dia:
            # Obtener método de pago predominante
            metodo_pago = 'N/A'
            if ticket.pagos.exists():
                primer_pago = ticket.pagos.first()
                metodo_pago = primer_pago.get_metodo_pago_display() if primer_pago else 'N/A'
            
            tickets_data.append({
                'id': ticket.id,
                'numero': ticket.correlativo,
                'hora': ticket.created_at.strftime('%H:%M:%S'),
                'cliente': ticket.cliente_nombre if ticket.cliente_nombre else 'Cliente General',
                'metodo_pago': metodo_pago,
                'total': ticket.total,
                'estado': ticket.estado
            })
        
        # ========== OBTENER BOLETAS ELECTRÓNICAS ==========
        boletas_del_dia = Dte.objects.filter(
            sucursal=sucursal,
            fecha_emision=fecha_obj,
            tipo_documento='BOLETA ELECTRONICA',
            tipo_transaccion='VENTA_PUBLICO',
            estado_dte__in=['EMITIDO', 'ACEPTADO']
        ).select_related('receptor').order_by('fecha_emision', 'hora')
        
        boletas_data = []
        for boleta in boletas_del_dia:
            # Calcular monto IVA
            monto_iva = boleta.monto_con_iva - boleta.monto_neto
            
            boletas_data.append({
                'id': boleta.id,
                'folio': boleta.numero_documento,
                'hora': boleta.hora.strftime('%H:%M:%S') if boleta.hora else 'N/A',
                'rut_cliente': boleta.receptor.rut if boleta.receptor else '66666666-6',
                'razon_social': boleta.receptor.razon_social if boleta.receptor else 'Cliente General',
                'monto_neto': float(boleta.monto_neto),
                'monto_iva': float(monto_iva),
                'monto_total': float(boleta.monto_con_iva),
                'estado': boleta.estado_dte
            })
        
        # ========== OBTENER FACTURAS ELECTRÓNICAS ==========
        facturas_del_dia = Dte.objects.filter(
            sucursal=sucursal,
            fecha_emision=fecha_obj,
            tipo_documento__in=['FACTURA ELECTRONICA', 'FACTURA EXENTA'],
            tipo_transaccion='VENTA_PUBLICO',
            estado_dte__in=['EMITIDO', 'ACEPTADO']
        ).select_related('receptor').order_by('fecha_emision', 'hora')
        
        facturas_data = []
        for factura in facturas_del_dia:
            # Calcular monto IVA
            monto_iva = factura.monto_con_iva - factura.monto_neto
            
            facturas_data.append({
                'id': factura.id,
                'folio': factura.numero_documento,
                'hora': factura.hora.strftime('%H:%M:%S') if factura.hora else 'N/A',
                'rut_cliente': factura.receptor.rut if factura.receptor else 'N/A',
                'razon_social': factura.receptor.razon_social if factura.receptor else 'Cliente',
                'monto_neto': float(factura.monto_neto),
                'monto_iva': float(monto_iva),
                'monto_total': float(factura.monto_con_iva),
                'estado': factura.estado_dte,
                'tipo': factura.tipo_documento
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'tickets': tickets_data,
                'boletas': boletas_data,
                'facturas': facturas_data,
                'totales': {
                    'total_tickets': len(tickets_data),
                    'total_boletas': len(boletas_data),
                    'total_facturas': len(facturas_data),
                    'total_documentos': len(tickets_data) + len(boletas_data) + len(facturas_data)
                }
            }
        })
        
    except Exception as e:
        logger.exception("Error al obtener transacciones del dia")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener transacciones: {str(e)}'
        })


# ========== NUEVAS FUNCIONALIDADES DE ARQUEO ==========

# ===== TOLERANCIAS DEL MÓDULO DE CAJA =====
# Antes convivían cuatro umbrales distintos sin relación entre sí: $0 decidía si
# el arqueo quedaba `CERRADO` o `CON_DIFERENCIAS`, $500 exigía observación al
# cerrar, $1.000 activaba `requiere_supervision` y daba el depósito por cuadrado,
# y $5.000 pedía categorizar. Con eso un mismo arqueo de $300 salía "cuadrado"
# para el cierre y "con diferencias" en el listado. Se unifican aquí para que
# todas las pantallas den el mismo veredicto.
TOLERANCIA_ARQUEO_EFECTIVO = 500     # bajo esto el día se considera cuadrado
TOLERANCIA_ARQUEO_DEPOSITO = 1000    # holgura al comparar depósito vs teórico
UMBRAL_ARQUEO_CATEGORIA = 5000       # sobre esto se exige categoría de diferencia


def _sucursales_permitidas(request):
    """IDs de sucursal que el usuario puede consultar en el módulo de caja.

    Un supervisor con 13 sucursales tenía que entrar sucursal por sucursal
    (13 recargas) porque el endpoint sólo aceptaba un id escalar. Devolvemos el
    universo permitido para poder soportar `sucursal_id=all` y listas separadas
    por coma sin abrir el acceso a sucursales ajenas.
    """
    rol_usuario = getattr(request.user, 'rol', None)
    es_supervisor = rol_usuario in ['administrador', 'administracion']
    try:
        permitidas = [
            int(s['sucursal_id'])
            for s in obtener_sucursales_usuario(request.user)
            if s.get('sucursal_id')
        ]
    except Exception:
        permitidas = []
    if not permitidas:
        actual = get_sucursal_id(request)
        permitidas = [int(actual)] if actual else []
    return permitidas, es_supervisor


def _resolver_sucursales_filtro(request):
    """Traduce el parámetro `sucursal_id` a una lista de ids ya autorizada.

    Acepta: vacío (sucursal activa), `all`/`todas`, o "1,4,7".
    """
    permitidas, es_supervisor = _sucursales_permitidas(request)
    raw = (request.GET.get('sucursal_id') or '').strip().lower()
    activa = get_sucursal_id(request)

    if not raw:
        return ([int(activa)] if activa else permitidas), es_supervisor

    if raw in ('all', 'todas', '*'):
        # Sin permisos de supervisión el "todas" se degrada a la sucursal activa
        # en vez de filtrar de más y mostrar datos de otra empresa.
        if not es_supervisor:
            return ([int(activa)] if activa else []), es_supervisor
        return permitidas, es_supervisor

    pedidas = []
    for parte in raw.split(','):
        parte = parte.strip()
        if parte.isdigit():
            pedidas.append(int(parte))
    if not pedidas:
        return ([int(activa)] if activa else permitidas), es_supervisor

    if es_supervisor:
        autorizadas = [s for s in pedidas if not permitidas or s in permitidas]
    else:
        autorizadas = [s for s in pedidas if s in permitidas]
    if not autorizadas:
        autorizadas = [int(activa)] if activa else []
    return autorizadas, es_supervisor


@login_required
@require_GET
def listar_arqueos(request):
    """API para listar arqueos con indicadores del período consultado.

    Los indicadores se calculan sobre EL MISMO rango y las MISMAS sucursales que
    la tabla. Antes se calculaban siempre sobre el mes en curso mientras la tabla
    respetaba el filtro, así que al consultar junio los KPIs seguían mostrando
    julio; el template lo había parcheado con un cartel ("no dependen del rango")
    en vez de arreglar el cálculo.

    Se mantiene `indicadores_mensuales` en la respuesta porque `cuadraturaCaja`
    también consume este endpoint.
    """
    try:
        from datetime import date, timedelta as _td
        from calendar import monthrange

        sucursal_ids, es_supervisor = _resolver_sucursales_filtro(request)
        rol_usuario = getattr(request.user, 'rol', None)

        if not sucursal_ids:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        # `sucursal_id` escalar se conserva para los cálculos que siguen siendo
        # mono-sucursal (indicadores mensuales de compatibilidad).
        sucursal_id = sucursal_ids[0]
        multi_sucursal = len(sucursal_ids) > 1

        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        estado = request.GET.get('estado')
        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 20)), 500)

        # --- Filtros nuevos (todos server-side) ---
        # Antes sólo existían fecha/estado; `resultado_revision` se filtraba en
        # JavaScript sobre las filas ya descargadas, así que "Requiere acción"
        # sólo encontraba los que hubieran entrado en la primera página.
        resultado = (request.GET.get('resultado') or '').strip().upper()
        tipo_dif = (request.GET.get('tipo_dif') or '').strip().upper()
        try:
            min_dif = int(request.GET.get('min_dif') or 0)
        except (TypeError, ValueError):
            min_dif = 0
        solo_express = request.GET.get('solo_express') in ('1', 'true', 'True')
        solo_dep_pend = request.GET.get('solo_dep_pendiente') in ('1', 'true', 'True')
        solo_sin_expl = request.GET.get('solo_sin_explicacion') in ('1', 'true', 'True')
        solo_accion = request.GET.get('solo_accion') in ('1', 'true', 'True')
        usuario_filtro = request.GET.get('usuario_id')
        try:
            dias_min = int(request.GET.get('dias_min') or 0)
        except (TypeError, ValueError):
            dias_min = 0
        orden = (request.GET.get('orden') or 'fecha').strip()

        # ========== INDICADORES MENSUALES (compatibilidad con cuadraturaCaja) ==========
        hoy = timezone.localdate()
        primer_dia_mes = date(hoy.year, hoy.month, 1)
        ultimo_dia_mes = date(hoy.year, hoy.month, monthrange(hoy.year, hoy.month)[1])

        # Días operativos del mes. Antes se usaba el calendario (lunes a sábado,
        # `weekday() < 6`), lo que dejaba dos huecos: los domingos con venta
        # nunca aparecían como día faltante, y a una sucursal cerrada los lunes
        # se le contaba un faltante inexistente. Ahora el día "cuenta" si hubo
        # venta pagada, que es la única razón por la que debe existir un arqueo.
        dias_calendario = []
        dia_actual = primer_dia_mes
        while dia_actual <= min(hoy, ultimo_dia_mes):
            dias_calendario.append(dia_actual)
            dia_actual += _td(days=1)

        dias_venta_mes = set(
            Ticket.objects.filter(
                sucursal_id__in=sucursal_ids,
                fecha__gte=primer_dia_mes,
                fecha__lte=hoy,
                estado='PAGADO',
            ).values_list('sucursal_id', 'fecha').distinct()
        )
        if dias_venta_mes:
            pares_esperados = dias_venta_mes
        else:
            # Sin datos de venta (sucursal nueva o Ticket vacío) se cae al
            # calendario lunes-sábado para no reportar 0 días operativos.
            pares_esperados = {
                (s, d) for s in sucursal_ids
                for d in dias_calendario if d.weekday() < 6
            }
        total_dias_habiles = len(pares_esperados)

        arqueos_mes_qs = ArqueoCaja.objects.filter(
            sucursal_id__in=sucursal_ids,
            fecha_arqueo__gte=primer_dia_mes,
            fecha_arqueo__lte=hoy,
        )

        # Único aggregate para todos los contadores y sumas del mes
        indic = arqueos_mes_qs.aggregate(
            arqueos_pendientes=Count('id', filter=Q(estado='ABIERTO')),
            # Se cuenta por el MONTO de la diferencia, no por el estado. Con
            # `Count(estado='CON_DIFERENCIAS')` el KPI se autoborraba: al aprobar
            # el arqueo `revisar_arqueo` lo pasa a `REVISADO`, así que un día con
            # $80.000 de faltante desaparecía del indicador apenas el supervisor
            # lo revisaba, y el mes cerraba en 0 con la caja igual de descuadrada.
            arqueos_con_diferencias=Count(
                'id',
                filter=(
                    Q(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO)
                    | Q(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO)
                ),
            ),
            arqueos_cerrados=Count('id', filter=Q(estado='CERRADO')),
            # "Revisado" es tener un veredicto del supervisor, no estar en el
            # estado REVISADO: `REQUIERE_ACCION` deja el estado intacto, así que
            # por estado un arqueo ya atendido figuraba como nunca revisado.
            arqueos_revisados=Count(
                'id', filter=~Q(resultado_revision='PENDIENTE')
            ),
            arqueos_sin_revision=Count(
                'id',
                filter=Q(resultado_revision='PENDIENTE') & ~Q(estado='ABIERTO'),
            ),
            arqueos_requieren_accion=Count(
                'id', filter=Q(resultado_revision='REQUIERE_ACCION')
            ),
            # Se lee la diferencia ya calculada en vez de recalcular
            # físico - teórico: `ArqueoCaja.save()` descuenta además el
            # `fondo_fijo_snapshot`, así que ambas fórmulas divergen apenas
            # una sucursal configure fondo fijo (hoy todas están en 0).
            total_diferencia_efectivo=Sum('diferencia_efectivo'),
            # Sólo sobre arqueos que informaron cierre POS. Sumar los que tienen
            # `cierre_pos_fisico = 0` restaba el Transbank teórico completo y
            # producía un "faltante" inventado del tamaño de la venta con tarjeta.
            total_diferencia_transbank=Sum(
                F('cierre_pos_fisico') - F('total_transbank_teorico'),
                filter=Q(cierre_pos_fisico__gt=0),
                output_field=IntegerField(),
            ),
            arqueos_sin_cierre_pos=Count(
                'id',
                filter=Q(cierre_pos_fisico=0, total_transbank_teorico__gt=0),
            ),
            total_teorico_efectivo_mes=Sum('total_efectivo_teorico'),
        )

        pares_con_arqueo = set(
            arqueos_mes_qs.values_list('sucursal_id', 'fecha_arqueo').distinct()
        )
        arqueos_realizados = len(pares_con_arqueo)
        pares_faltantes = sorted(
            pares_esperados - pares_con_arqueo, key=lambda p: (p[1], p[0])
        )
        arqueos_faltantes = len(pares_faltantes)

        dep_mes_agg = DepositoBancario.objects.filter(
            arqueo__sucursal_id__in=sucursal_ids,
            arqueo__fecha_arqueo__gte=primer_dia_mes,
            arqueo__fecha_arqueo__lte=hoy,
        ).aggregate(
            depositos_pendientes_conf=Count(
                'id', filter=Q(verificado=False, monto_declarado__gt=0)
            ),
            monto_pendiente_conf=Sum(
                'monto_declarado',
                filter=Q(verificado=False, monto_declarado__gt=0),
            ),
            total_depositado_mes=Sum(
                Case(
                    When(
                        verificado=True,
                        monto_confirmado__gt=0,
                        then=F('monto_confirmado'),
                    ),
                    When(verificado=True, then=F('monto')),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
        )

        total_depositado_mes = dep_mes_agg['total_depositado_mes'] or 0
        total_teorico_efectivo_mes = indic['total_teorico_efectivo_mes'] or 0
        arqueos_revisados = indic['arqueos_revisados'] or 0

        indicadores_mensuales = {
            'mes_actual': hoy.strftime('%B %Y'),
            'dias_habiles': total_dias_habiles,
            'arqueos_realizados': arqueos_realizados,
            'arqueos_faltantes': arqueos_faltantes,
            'arqueos_pendientes': indic['arqueos_pendientes'] or 0,
            'arqueos_con_diferencias': indic['arqueos_con_diferencias'] or 0,
            'arqueos_cerrados': indic['arqueos_cerrados'] or 0,
            # Se topa en 100: antes el numerador contaba los arqueos de domingo
            # y el denominador excluía los domingos, así que una sucursal que
            # arqueaba el fin de semana mostraba 103% de cumplimiento.
            'porcentaje_cumplimiento': round(
                min(100.0, arqueos_realizados / total_dias_habiles * 100)
                if total_dias_habiles > 0 else 0, 1
            ),
            'dias_faltantes': [d.strftime('%Y-%m-%d') for _s, d in pares_faltantes[:10]],
            'total_diferencia_efectivo': indic['total_diferencia_efectivo'] or 0,
            'total_diferencia_transbank': indic['total_diferencia_transbank'] or 0,
            'arqueos_sin_cierre_pos': indic['arqueos_sin_cierre_pos'] or 0,
            'arqueos_revisados': arqueos_revisados,
            'arqueos_sin_revision': indic['arqueos_sin_revision'] or 0,
            'arqueos_requieren_accion': indic['arqueos_requieren_accion'] or 0,
            'porcentaje_revisados': round(
                min(100.0, arqueos_revisados / arqueos_realizados * 100)
                if arqueos_realizados > 0 else 0, 1
            ),
            'depositos_pendientes_confirmacion': dep_mes_agg['depositos_pendientes_conf'] or 0,
            'monto_depositos_pendientes': dep_mes_agg['monto_pendiente_conf'] or 0,
            'total_depositado_mes': total_depositado_mes,
            'total_teorico_efectivo_mes': total_teorico_efectivo_mes,
            'diferencia_depositos_mes': total_depositado_mes - total_teorico_efectivo_mes,
        }

        # ========== QUERYSET PRINCIPAL ==========
        # Los campos `cache_*` de ArqueoCaja vienen denormalizados via signal
        # post_save/post_delete de DepositoBancario, evitando JOINs aquí.
        # Solo anotamos lo que no se puede denormalizar trivialmente:
        # existencia de comprobante y conteos de relaciones externas.
        # `base` sólo lleva filtros (sin anotaciones): sobre él se cuentan las
        # filas y se calculan los indicadores del período. Contar sobre el
        # queryset anotado obligaba a Postgres a resolver el GROUP BY completo
        # (dos `Count` sobre relaciones distintas + `Exists` + los dos JOIN del
        # `select_related`) dos veces por request, una para `count()` y otra
        # para la página.
        base = ArqueoCaja.objects.filter(sucursal_id__in=sucursal_ids)

        if fecha_desde:
            base = base.filter(fecha_arqueo__gte=fecha_desde)
        if fecha_hasta:
            base = base.filter(fecha_arqueo__lte=fecha_hasta)
        if estado:
            # Admite varios estados separados por coma para que el supervisor
            # pueda pedir "lo que aún no está cerrado" en una sola consulta.
            estados = [e.strip() for e in estado.split(',') if e.strip()]
            base = base.filter(estado__in=estados) if len(estados) > 1 \
                else base.filter(estado=estados[0])
        if resultado:
            resultados = [r.strip() for r in resultado.split(',') if r.strip()]
            base = base.filter(resultado_revision__in=resultados)
        if tipo_dif == 'FALTANTE':
            base = base.filter(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO)
        elif tipo_dif == 'SOBRANTE':
            base = base.filter(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO)
        elif tipo_dif == 'EXACTO':
            base = base.filter(
                diferencia_efectivo__gte=-TOLERANCIA_ARQUEO_EFECTIVO,
                diferencia_efectivo__lte=TOLERANCIA_ARQUEO_EFECTIVO,
            )
        if min_dif > 0:
            base = base.filter(
                Q(diferencia_efectivo__gte=min_dif)
                | Q(diferencia_efectivo__lte=-min_dif)
            )
        if solo_express:
            base = base.filter(modo_conteo='EXPRESS')
        if solo_dep_pend:
            base = base.filter(cache_depositos_pendientes__gt=0)
        if solo_sin_expl:
            # Diferencia relevante sin que nadie haya escrito por qué.
            base = base.filter(
                Q(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO)
                | Q(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO)
            ).filter(
                Q(observaciones_diferencia__isnull=True)
                | Q(observaciones_diferencia='')
            )
        if solo_accion:
            base = base.filter(resultado_revision='REQUIERE_ACCION')
        if usuario_filtro and str(usuario_filtro).isdigit():
            base = base.filter(usuario_responsable_id=int(usuario_filtro))
        if dias_min > 0:
            # "Sin revisar hace más de N días". `dias_sin_revision` es una
            # property de Python y no se puede filtrar en SQL, pero equivale a
            # acotar la fecha del arqueo entre los que aún no tienen veredicto.
            base = base.filter(
                fecha_arqueo__lte=hoy - _td(days=dias_min),
                resultado_revision='PENDIENTE',
            ).exclude(estado='ABIERTO')

        # ========== INDICADORES DEL PERÍODO CONSULTADO ==========
        # Estos sí responden al filtro que el usuario tiene puesto.
        ind_periodo = base.aggregate(
            total=Count('id'),
            faltante=Sum(
                'diferencia_efectivo',
                filter=Q(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO),
            ),
            sobrante=Sum(
                'diferencia_efectivo',
                filter=Q(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO),
            ),
            n_faltante=Count(
                'id', filter=Q(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO)
            ),
            n_sobrante=Count(
                'id', filter=Q(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO)
            ),
            n_exacto=Count('id', filter=Q(diferencia_efectivo=0)),
            n_cerrados=Count('id', filter=~Q(estado='ABIERTO')),
            n_express=Count('id', filter=Q(modo_conteo='EXPRESS')),
            n_sin_explicacion=Count(
                'id',
                filter=(
                    Q(diferencia_efectivo__gt=TOLERANCIA_ARQUEO_EFECTIVO)
                    | Q(diferencia_efectivo__lt=-TOLERANCIA_ARQUEO_EFECTIVO)
                ) & (
                    Q(observaciones_diferencia__isnull=True)
                    | Q(observaciones_diferencia='')
                ),
            ),
            n_sin_categoria=Count(
                'id',
                filter=(
                    Q(diferencia_efectivo__gt=UMBRAL_ARQUEO_CATEGORIA)
                    | Q(diferencia_efectivo__lt=-UMBRAL_ARQUEO_CATEGORIA)
                ) & (
                    Q(categoria_diferencia__isnull=True)
                    | Q(categoria_diferencia='')
                ),
            ),
            n_pend_revision=Count(
                'id',
                filter=Q(resultado_revision='PENDIENTE') & ~Q(estado='ABIERTO'),
            ),
            n_requiere_accion=Count(
                'id', filter=Q(resultado_revision='REQUIERE_ACCION')
            ),
            n_abiertos=Count('id', filter=Q(estado='ABIERTO')),
            n_dep_pendientes=Count('id', filter=Q(cache_depositos_pendientes__gt=0)),
            tbk_diferencia=Sum(
                F('cierre_pos_fisico') - F('total_transbank_teorico'),
                filter=Q(cierre_pos_fisico__gt=0),
                output_field=IntegerField(),
            ),
            n_tbk_sin_cierre=Count(
                'id', filter=Q(cierre_pos_fisico=0, total_transbank_teorico__gt=0)
            ),
        )

        _falt = ind_periodo['faltante'] or 0
        _sobr = ind_periodo['sobrante'] or 0
        _n_cerr = ind_periodo['n_cerrados'] or 0

        # Cobertura del período consultado: días-sucursal CON VENTA que tienen
        # su arqueo. Se calcula sólo cuando el rango es acotado y no hay filtros
        # de contenido activos (con un filtro puesto la tabla ya no representa
        # "todos los días del período" y el porcentaje no significaría nada).
        cobertura_pct = None
        cobertura_detalle = {}
        filtros_de_contenido = any([
            estado, resultado, tipo_dif, min_dif, solo_express,
            solo_dep_pend, solo_sin_expl, solo_accion, usuario_filtro, dias_min,
        ])
        if fecha_desde and fecha_hasta and not filtros_de_contenido:
            try:
                _d1 = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                _d2 = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                _d1 = _d2 = None
            if _d1 and _d2 and 0 <= (_d2 - _d1).days <= 400:
                dias_venta_periodo = set(
                    Ticket.objects.filter(
                        sucursal_id__in=sucursal_ids,
                        fecha__gte=_d1, fecha__lte=_d2,
                        estado='PAGADO',
                    ).values_list('sucursal_id', 'fecha').distinct()
                )
                dias_arqueo_periodo = set(
                    base.values_list('sucursal_id', 'fecha_arqueo').distinct()
                )
                if dias_venta_periodo:
                    cubiertos = len(dias_venta_periodo & dias_arqueo_periodo)
                    cobertura_pct = round(cubiertos / len(dias_venta_periodo) * 100, 1)
                    sin_arqueo = sorted(
                        dias_venta_periodo - dias_arqueo_periodo,
                        key=lambda p: (p[1], p[0]), reverse=True,
                    )
                    alias_suc = dict(
                        Sucursal.objects.filter(
                            id__in={s for s, _ in sin_arqueo[:20]}
                        ).values_list('id', 'alias')
                    )
                    cobertura_detalle = {
                        'dias_con_venta': len(dias_venta_periodo),
                        'dias_con_arqueo': cubiertos,
                        'dias_sin_arqueo': len(sin_arqueo),
                        # El backend siempre supo QUÉ días faltan; la UI sólo
                        # mostraba el número, así que el supervisor no tenía
                        # forma de saber cuáles perseguir.
                        'detalle_sin_arqueo': [
                            {'sucursal': alias_suc.get(s, f'#{s}'), 'fecha': d.strftime('%Y-%m-%d')}
                            for s, d in sin_arqueo[:20]
                        ],
                    }

        # Arqueos viejos sin veredicto: la cola real de trabajo del supervisor.
        n_atrasados = base.filter(
            resultado_revision='PENDIENTE',
            fecha_arqueo__lte=hoy - _td(days=3),
        ).exclude(estado='ABIERTO').count()

        indicadores = {
            'total_arqueos': ind_periodo['total'] or 0,
            # Exposición real: faltantes y sobrantes NO se compensan entre sí.
            # `Sum('diferencia_efectivo')` a secas dejaba que un sobrante de
            # $50.000 tapara un faltante de $50.000 y el KPI marcaba $0.
            'exposicion_efectivo': abs(_falt) + abs(_sobr),
            'faltante_total': _falt,
            'sobrante_total': _sobr,
            'neto_efectivo': _falt + _sobr,
            'n_faltante': ind_periodo['n_faltante'] or 0,
            'n_sobrante': ind_periodo['n_sobrante'] or 0,
            'n_exacto': ind_periodo['n_exacto'] or 0,
            'n_cerrados': _n_cerr,
            # Señal de riesgo: conteo idéntico al teórico al peso. El propio
            # `AnalisisFraudeCaja` marca sobre 80% como copia del teórico.
            'pct_exacto': round(
                (ind_periodo['n_exacto'] or 0) / _n_cerr * 100, 1
            ) if _n_cerr else 0,
            'n_express': ind_periodo['n_express'] or 0,
            'n_sin_explicacion': ind_periodo['n_sin_explicacion'] or 0,
            'n_sin_categoria': ind_periodo['n_sin_categoria'] or 0,
            'n_pend_revision': ind_periodo['n_pend_revision'] or 0,
            'n_requiere_accion': ind_periodo['n_requiere_accion'] or 0,
            'n_abiertos': ind_periodo['n_abiertos'] or 0,
            'n_atrasados': n_atrasados,
            'n_dep_pendientes': ind_periodo['n_dep_pendientes'] or 0,
            'tbk_diferencia': ind_periodo['tbk_diferencia'] or 0,
            'n_tbk_sin_cierre': ind_periodo['n_tbk_sin_cierre'] or 0,
            'tolerancia': TOLERANCIA_ARQUEO_EFECTIVO,
            'multi_sucursal': multi_sucursal,
            'sucursales_consultadas': len(sucursal_ids),
            'cumplimiento_pct': cobertura_pct,
            'cobertura': cobertura_detalle,
        }

        # Bandeja de alertas accionables: cada una trae el filtro que la aísla,
        # para que el supervisor pueda saltar directo a esa cola en un click.
        indicadores['alertas'] = [
            a for a in [
                {'codigo': 'ABIERTOS', 'label': 'Arqueos sin cerrar',
                 'count': indicadores['n_abiertos'], 'nivel': 'warning',
                 'filtro': {'estado': 'ABIERTO'}},
                {'codigo': 'ATRASADOS', 'label': 'Sin revisar hace +3 días',
                 'count': n_atrasados, 'nivel': 'error',
                 'filtro': {'dias_min': 3}},
                {'codigo': 'ACCION', 'label': 'Requieren acción correctiva',
                 'count': indicadores['n_requiere_accion'], 'nivel': 'error',
                 'filtro': {'solo_accion': 1}},
                {'codigo': 'DEP_PEND', 'label': 'Depósitos sin confirmar',
                 'count': indicadores['n_dep_pendientes'], 'nivel': 'warning',
                 'filtro': {'solo_dep_pendiente': 1}},
                {'codigo': 'SIN_EXPL', 'label': 'Diferencias sin explicación',
                 'count': indicadores['n_sin_explicacion'], 'nivel': 'error',
                 'filtro': {'solo_sin_explicacion': 1}},
                {'codigo': 'EXPRESS', 'label': 'Conteos express por validar',
                 'count': indicadores['n_express'], 'nivel': 'info',
                 'filtro': {'solo_express': 1}},
                {'codigo': 'TBK_SIN_CIERRE', 'label': 'Sin cierre POS informado',
                 'count': indicadores['n_tbk_sin_cierre'], 'nivel': 'info',
                 'filtro': {}},
                # Días con venta que nunca se arquearon: no aparecen en la
                # tabla (no existe la fila), así que sin esta alerta eran
                # invisibles.
                {'codigo': 'SIN_ARQUEO', 'label': 'Días con venta sin arqueo',
                 'count': cobertura_detalle.get('dias_sin_arqueo', 0),
                 'nivel': 'error', 'filtro': {}},
            ] if a['count'] > 0
        ]

        # ========== PÁGINA ==========
        ordenes = {
            'fecha': ('-fecha_arqueo', '-id'),
            'fecha_asc': ('fecha_arqueo', 'id'),
            'diferencia': ('diferencia_efectivo', '-id'),
            'sucursal': ('sucursal__alias', '-fecha_arqueo'),
        }
        order_by = ordenes.get(orden, ordenes['fecha'])

        total_items = base.count()
        per_page = max(1, per_page)
        total_pages = max(1, -(-total_items // per_page))
        page = min(max(1, page), total_pages)
        offset = (page - 1) * per_page

        ids_pagina = list(
            base.order_by(*order_by).values_list('id', flat=True)[offset:offset + per_page]
        )

        # Las anotaciones caras se aplican sólo a las filas de la página.
        arqueos_page = (
            ArqueoCaja.objects
            .filter(id__in=ids_pagina)
            .select_related('usuario_responsable', 'supervisor_revision', 'sucursal')
            .annotate(
                ann_tiene_comprobante=Exists(
                    DepositoBancario.objects.filter(
                        arqueo_id=OuterRef('pk'),
                        verificado=True,
                    ).exclude(numero_comprobante='')
                ),
                ann_reaperturas=Count('historial_reaperturas', distinct=True),
                ann_bitacora_count=Count('bitacora', distinct=True),
            )
            .order_by(*order_by)
        )

        resultado_revision_dict = dict(RESULTADO_REVISION_CHOICES)
        arqueos_lista = list(arqueos_page)

        # Fallback para arqueos con la cache de depósitos desincronizada
        # (legacy o signal no ejecutado). Antes se resolvía con un `aggregate`
        # DENTRO del loop: hasta 100 queries extra en una sola página. Ahora es
        # una única consulta agrupada para las filas que lo necesitan.
        ids_fallback = [
            a.id for a in arqueos_lista
            if (a.cache_total_dep_verificado or 0) == 0
            and (a.cache_depositos_confirmados or 0) > 0
            and (a.cache_total_depositos or 0) > 0
        ]
        fallback_por_arqueo = {}
        if ids_fallback:
            for fila in (
                DepositoBancario.objects
                .filter(arqueo_id__in=ids_fallback, verificado=True)
                .values('arqueo_id')
                .annotate(
                    total=Sum('monto'),
                    efectivo=Sum('monto', filter=Q(tipo_medio='EFECTIVO')),
                    cheque=Sum('monto', filter=Q(tipo_medio='CHEQUE')),
                )
            ):
                fallback_por_arqueo[fila['arqueo_id']] = fila

        arqueos_data = []
        for arqueo in arqueos_lista:
            # Leer desde campos denormalizados (actualizados por signal)
            depositos_declarados = arqueo.cache_depositos_declarados or 0
            total_dep_efectivo = arqueo.cache_total_dep_efectivo_verif or 0
            total_dep_cheque = arqueo.cache_total_dep_cheque_verif or 0
            total_dep_verif_all = arqueo.cache_total_dep_verificado or 0

            fallback_dep = fallback_por_arqueo.get(arqueo.id)
            if fallback_dep:
                total_dep_verif_all = fallback_dep['total'] or 0
                total_dep_efectivo = fallback_dep['efectivo'] or 0
                total_dep_cheque = fallback_dep['cheque'] or 0

            teorico_ef = arqueo.total_efectivo_teorico or 0
            teorico_ch = arqueo.total_cheque_teorico or 0
            dif_dep_vs_teorico = total_dep_efectivo - teorico_ef
            dif_cheques_vs_teorico = total_dep_cheque - teorico_ch

            esperado_total = teorico_ef + teorico_ch
            if esperado_total == 0 or total_dep_verif_all == 0:
                estado_deposito = 'SIN_DEPOSITO'
            elif abs(total_dep_verif_all - esperado_total) <= TOLERANCIA_ARQUEO_DEPOSITO:
                estado_deposito = 'COMPLETO'
            else:
                estado_deposito = 'PARCIAL'

            dif_ef = arqueo.diferencia_efectivo or 0
            # Veredicto único del día, con la misma tolerancia en todas las
            # pantallas. Antes cada consumidor aplicaba la suya ($0, $500,
            # $1.000) y el mismo arqueo salía "cuadrado" en una y "con
            # diferencias" en otra.
            if abs(dif_ef) <= TOLERANCIA_ARQUEO_EFECTIVO:
                veredicto = 'CUADRA'
            elif dif_ef < 0:
                veredicto = 'FALTANTE'
            else:
                veredicto = 'SOBRANTE'
            revisado = arqueo.resultado_revision != 'PENDIENTE'

            arqueos_data.append({
                'id': arqueo.id,
                'fecha_arqueo': arqueo.fecha_arqueo.strftime('%Y-%m-%d'),
                'fecha_creacion': arqueo.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'usuario_responsable': arqueo.usuario_responsable.username,
                'estado': arqueo.estado,
                'estado_display': arqueo.get_estado_display(),
                'efectivo_teorico': arqueo.total_efectivo_teorico,
                'efectivo_fisico': arqueo.total_efectivo_fisico,
                'diferencia_efectivo': arqueo.diferencia_efectivo,
                'diferencia_absoluta': arqueo.diferencia_absoluta,
                'tipo_diferencia': arqueo.tipo_diferencia,
                'porcentaje_diferencia': round(arqueo.porcentaje_diferencia, 2),
                'requiere_supervision': arqueo.requiere_supervision,
                'venta_total': arqueo.venta_total_teorica,
                'transbank_teorico': arqueo.total_transbank_teorico,
                'transbank_fisico': arqueo.cierre_pos_fisico,
                'diferencia_transbank': arqueo.diferencia_transbank,
                'debito_teorico': arqueo.total_tarjeta_debito_teorico,
                'debito_fisico': arqueo.cierre_debito_fisico,
                'diferencia_debito': arqueo.diferencia_debito,
                'credito_teorico': arqueo.total_tarjeta_credito_teorico,
                'credito_fisico': arqueo.cierre_credito_fisico,
                'diferencia_credito': arqueo.diferencia_credito,
                'numero_lote': arqueo.numero_lote_pos or '',
                'observaciones': arqueo.observaciones or '',
                'supervisor': arqueo.supervisor_revision.username if arqueo.supervisor_revision else '',
                'fecha_cierre': arqueo.fecha_cierre.strftime('%d/%m/%Y %H:%M') if arqueo.fecha_cierre else '',
                'tiene_comprobante': bool(arqueo.ann_tiene_comprobante),
                'total_depositado_verificado': total_dep_verif_all,
                'depositos_declarados': depositos_declarados,
                'depositos_confirmados': arqueo.cache_depositos_confirmados or 0,
                'depositos_pendientes': arqueo.cache_depositos_pendientes or 0,
                'tiene_depositos': depositos_declarados > 0,
                'reaperturas': arqueo.ann_reaperturas or 0,
                'total_deposito_efectivo': total_dep_efectivo,
                'total_deposito_cheque': total_dep_cheque,
                'diferencia_deposito_vs_teorico': dif_dep_vs_teorico,
                'diferencia_cheques_vs_teorico': dif_cheques_vs_teorico,
                'estado_deposito': estado_deposito,
                'dias_sin_revision': arqueo.dias_sin_revision,
                'requiere_revision_urgente': arqueo.requiere_revision_urgente,
                'modo_conteo': arqueo.modo_conteo,
                'requiere_revision_express': arqueo.requiere_revision_express,
                'fondo_fijo': arqueo.fondo_fijo_snapshot,
                'observaciones_diferencia': arqueo.observaciones_diferencia or '',
                'categoria_diferencia': arqueo.categoria_diferencia or '',
                'observaciones_supervisor': arqueo.observaciones_supervisor or '',
                'resultado_revision': getattr(arqueo, 'resultado_revision', 'PENDIENTE'),
                'resultado_revision_display': resultado_revision_dict.get(
                    getattr(arqueo, 'resultado_revision', 'PENDIENTE'), 'Pendiente'
                ),
                'cantidad_observaciones': arqueo.ann_bitacora_count or 0,
                'ultima_obs_supervisor': '',
                # --- Campos nuevos ---
                'sucursal_id': arqueo.sucursal_id,
                'sucursal_alias': arqueo.sucursal.alias if arqueo.sucursal else '',
                'veredicto': veredicto,
                'revisado': revisado,
                'sin_explicacion': (
                    veredicto != 'CUADRA' and not (arqueo.observaciones_diferencia or '').strip()
                ),
                'requiere_categoria': (
                    abs(dif_ef) > UMBRAL_ARQUEO_CATEGORIA
                    and not (arqueo.categoria_diferencia or '').strip()
                ),
                'fecha_revision': (
                    arqueo.fecha_revision.strftime('%d/%m/%Y %H:%M')
                    if arqueo.fecha_revision else ''
                ),
                # Aprobable en lote sólo si no hay nada que juzgar: cuadra
                # dentro de tolerancia, no está abierto y el depósito no quedó
                # a medias.
                'aprobable_en_lote': (
                    veredicto == 'CUADRA'
                    and arqueo.estado != 'ABIERTO'
                    and not revisado
                    and estado_deposito in ('COMPLETO', 'SIN_DEPOSITO')
                    and (arqueo.cache_depositos_pendientes or 0) == 0
                ),
            })

        return JsonResponse({
            'success': True,
            'arqueos': arqueos_data,
            # Indicadores del rango + sucursales realmente consultados.
            'indicadores': indicadores,
            # Se mantiene por compatibilidad: `cuadraturaCaja.html` consume
            # este mismo endpoint y lee `indicadores_mensuales`.
            'indicadores_mensuales': indicadores_mensuales,
            'sucursales_consultadas': sucursal_ids,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_items': total_items,
                'per_page': per_page,
                'has_next': page < total_pages,
                'has_previous': page > 1,
            }
        })

    except Exception as e:
        logger.exception("Error en listar_arqueos")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener arqueos: {str(e)}'
        })


@login_required
@require_POST
def corregir_arqueos_express(request):
    """Corrige UN arqueo express mal guardado con el monto realmente contado.

    Antes este endpoint recorría todos los arqueos descuadrados de la sucursal
    y les escribía `total_efectivo_fisico = total_efectivo_teorico` con
    `diferencia_efectivo = 0`: o sea, daba por cuadrada la caja sin que nadie
    hubiese contado el dinero. Sobre la base actual eso son 34 arqueos en 6
    sucursales que habrían quedado falseados de un clic, borrando faltantes y
    sobrantes reales.

    Ahora:

    * Con `arqueo_id` + `monto_real` corrige ese arqueo y sólo ese, con el
      monto que declara el operador, recalculando la diferencia igual que
      `ArqueoCaja.save()` y dejando una `ObservacionArqueo` con la traza.
    * Sin esos datos NO escribe nada: devuelve la lista de candidatos para que
      se corrijan uno por uno.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario != 'administrador':
            return JsonResponse({
                'success': False,
                'error': 'Solo el Administrador puede usar la corrección express.'
            }, status=403)

        # El front hace una primera llamada "sonda" (sin datos) para pedir la
        # lista de candidatos; toleramos body vacío para no responder un error
        # de JSON cuando lo único que falta son los datos de la corrección.
        data = json.loads(request.body or b'{}')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        # Arqueos candidatos: todas las denominaciones en 0, físico en 0 y un
        # teórico distinto de 0 (síntoma del conteo express que no se guardó).
        arqueos_problematicos = ArqueoCaja.objects.filter(
            sucursal_id=sucursal_id,
            billetes_20000=0,
            billetes_10000=0,
            billetes_5000=0,
            billetes_2000=0,
            billetes_1000=0,
            monedas_500=0,
            monedas_100=0,
            monedas_50=0,
            monedas_10=0,
            monedas_5=0,
            monedas_1=0,
            total_efectivo_fisico=0
        ).exclude(
            total_efectivo_teorico=0  # Excluir casos donde realmente no había ventas
        )

        arqueo_id = data.get('arqueo_id')
        monto_real_raw = data.get('monto_real')

        if not arqueo_id or monto_real_raw is None:
            candidatos = [
                {
                    'id': a.id,
                    'fecha': a.fecha_arqueo.strftime('%Y-%m-%d'),
                    'estado': a.estado,
                    'total_efectivo_teorico': a.total_efectivo_teorico,
                    # El fondo fijo entra en la diferencia (ver `ArqueoCaja.save`):
                    # el front lo necesita para mostrar cuánto se esperaba en caja.
                    'fondo_fijo': a.fondo_fijo_snapshot or 0,
                }
                for a in arqueos_problematicos.order_by('-fecha_arqueo')[:100]
            ]
            if candidatos:
                mensaje = (
                    f'Se encontraron {len(candidatos)} arqueo(s) sin conteo '
                    'guardado. La corrección masiva quedó desactivada porque '
                    'daba la caja por cuadrada sin contar el dinero: corrija '
                    'cada arqueo indicando el monto realmente contado '
                    '(arqueo_id + monto_real).'
                )
            else:
                mensaje = (
                    'No hay arqueos Express pendientes de regularizar en esta '
                    'sucursal.'
                )
            return JsonResponse({
                'success': False,
                'arqueos_corregidos': 0,
                'candidatos': candidatos,
                'error': mensaje
            }, status=400)

        try:
            monto_real = int(monto_real_raw)
        except (TypeError, ValueError):
            return JsonResponse({
                'success': False,
                'error': 'monto_real inválido: debe ser un entero'
            }, status=400)
        if monto_real < 0:
            return JsonResponse({
                'success': False,
                'error': 'monto_real no puede ser negativo'
            }, status=400)

        arqueo = arqueos_problematicos.filter(id=arqueo_id).first()
        if not arqueo:
            return JsonResponse({
                'success': False,
                'error': (
                    'El arqueo no existe en esta sucursal o no corresponde a '
                    'un conteo express sin guardar.'
                )
            }, status=404)

        diferencia = monto_real - (
            arqueo.total_efectivo_teorico + (arqueo.fondo_fijo_snapshot or 0)
        )
        with transaction.atomic():
            ArqueoCaja.objects.filter(id=arqueo.id).update(
                total_efectivo_fisico=monto_real,
                diferencia_efectivo=diferencia,
                modo_conteo='EXPRESS',
                requiere_revision_express=True,
                timestamp_conteo_fisico=timezone.now(),
            )
            ObservacionArqueo.objects.create(
                arqueo=arqueo,
                usuario=request.user,
                tipo='SISTEMA',
                texto=(
                    'Conteo express regularizado manualmente: efectivo físico '
                    f'${monto_real:,} (teórico ${arqueo.total_efectivo_teorico:,}, '
                    f'diferencia ${diferencia:,}).'
                ).replace(',', '.'),
                visible_para_cajera=True,
            )
        log_accion_caja(request, 'CORREGIR_EXPRESS', arqueo)
        logger.info(
            'Arqueo express regularizado: arqueo_id=%s, monto_real=%s, '
            'teorico=%s, diferencia=%s, usuario=%s',
            arqueo.id, monto_real, arqueo.total_efectivo_teorico, diferencia,
            getattr(request.user, 'username', '?'),
        )

        return JsonResponse({
            'success': True,
            'message': (
                f'Arqueo #{arqueo.id} corregido con el monto contado.'
            ),
            'arqueos_corregidos': 1,
            'arqueo': {
                'id': arqueo.id,
                'total_efectivo_fisico': monto_real,
                'diferencia_efectivo': diferencia,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al corregir arqueos: {str(e)}'
        })


@login_required
@require_POST
def crear_arqueo(request):
    """Crear nuevo arqueo basado en la cuadratura actual"""
    try:
        data = json.loads(request.body)
        fecha_arqueo = data.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha_arqueo or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal requeridas'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        # Validar que la fecha no sea futura y esté dentro del rango permitido por rol
        from datetime import datetime, date as dt_date
        fecha_obj = datetime.strptime(fecha_arqueo, '%Y-%m-%d').date()
        hoy = timezone.localdate()

        if fecha_obj > hoy:
            return JsonResponse({
                'success': False,
                'error': 'No puede crear arqueos para fechas futuras'
            })

        rol_usuario = getattr(request.user, 'rol', None)
        config_rango_arqueo = obtener_configuracion_rango_arqueo(
            rol_usuario,
            fecha_referencia=hoy,
        )

        if fecha_obj < config_rango_arqueo['fecha_minima']:
            return JsonResponse({
                'success': False,
                'error': (
                    'Solo puede crear arqueos dentro del rango configurado '
                    f'para su rol ({config_rango_arqueo["label"]}). '
                    'La fecha mínima permitida es '
                    f'{config_rango_arqueo["fecha_minima"].strftime("%d/%m/%Y")}.'
                )
            })

        # Verificar si ya existe un arqueo para esta fecha
        arqueo_existente = ArqueoCaja.objects.filter(
            fecha_arqueo=fecha_arqueo,
            sucursal=sucursal
        ).first()

        if arqueo_existente:
            return JsonResponse({
                'success': False,
                'error': f'Ya existe un arqueo para el {fecha_arqueo}',
                'arqueo_id': arqueo_existente.id,
                'estado': arqueo_existente.estado,
                'estado_display': arqueo_existente.get_estado_display(),
            })

        # Generar cuadratura en tiempo real para obtener los datos
        from django.test import RequestFactory
        factory = RequestFactory()
        fake_request = factory.post('/fake/', {'fecha': fecha_arqueo})
        fake_request.session = request.session
        fake_request.user = request.user
        
        response_data = generar_cuadratura_caja(fake_request)
        cuadratura_json = json.loads(response_data.content)
        
        if not cuadratura_json.get('success'):
            return JsonResponse({
                'success': False,
                'error': 'Error al generar datos de cuadratura'
            })
        
        cuadratura_data = cuadratura_json['cuadratura']
        
        # Crear el arqueo con los datos teóricos
        # Función auxiliar para convertir valores a int (los Decimal vienen como strings del JSON)
        def to_int(value):
            if value is None:
                return 0
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return 0
        
        arqueo = ArqueoCaja.objects.create(
            fecha_arqueo=fecha_arqueo,
            sucursal=sucursal,
            usuario_responsable=request.user,
            
            # Totales teóricos de la cuadratura
            # Tarjetas Comerciales (solo Hites)
            total_hites_teorico=to_int(cuadratura_data.get('total_hites', 0)),
            total_tarjetas_comerciales_teorico=to_int(cuadratura_data.get('total_tarjetas_comerciales', 0)),
            
            total_efectivo_teorico=to_int(cuadratura_data.get('total_efectivo', 0)),
            
            # Venta Internet (Falabella, Paris, Ripley, MercadoPago, Klap)
            total_falabella_teorico=to_int(cuadratura_data.get('total_falabella', 0)),
            total_paris_teorico=to_int(cuadratura_data.get('total_paris', 0)),
            total_ripley_teorico=to_int(cuadratura_data.get('total_ripley', 0)),
            total_mercadopago_teorico=to_int(cuadratura_data.get('total_mercadopago', 0)),
            total_klap_teorico=to_int(cuadratura_data.get('total_klap', 0)),
            total_venta_internet_teorico=to_int(cuadratura_data.get('total_venta_internet', 0)),
            
            total_tarjeta_debito_teorico=to_int(cuadratura_data.get('total_tarjeta_debito', 0)),
            total_tarjeta_credito_teorico=to_int(cuadratura_data.get('total_tarjeta_credito', 0)),
            total_transbank_teorico=to_int(cuadratura_data.get('total_transbank', 0)),
            total_transferencia_teorico=to_int(cuadratura_data.get('total_transferencia', 0)),
            total_cheque_teorico=to_int(cuadratura_data.get('total_cheque', 0)),
            total_convenio_teorico=to_int(cuadratura_data.get('total_convenio', 0)),
            total_credito_trabajador_teorico=to_int(cuadratura_data.get('total_credito_trabajador', 0)),
            
            total_tickets_teorico=to_int(cuadratura_data.get('total_tickets', 0)),
            total_boletas_electronicas_teorico=to_int(cuadratura_data.get('total_boletas_electronicas', 0)),
            total_facturas_teorico=to_int(cuadratura_data.get('total_facturas', 0)),
            total_facturas_exentas_teorico=to_int(cuadratura_data.get('total_facturas_exentas', 0)),
            total_notas_credito_teorico=to_int(cuadratura_data.get('total_notas_credito', 0)),
            
            cantidad_tickets=to_int(cuadratura_data.get('cantidad_tickets', 0)),
            cantidad_boletas_electronicas=to_int(cuadratura_data.get('cantidad_boletas_electronicas', 0)),
            cantidad_facturas=to_int(cuadratura_data.get('cantidad_facturas', 0)),
            cantidad_facturas_exentas=to_int(cuadratura_data.get('cantidad_facturas_exentas', 0)),
            
            venta_total_teorica=to_int(cuadratura_data.get('venta_total', 0)),

            fondo_fijo_snapshot=sucursal.fondo_fijo_caja,
            estado='ABIERTO'
        )
        
        log_accion_caja(request, 'GENERAR_CUADRATURA', arqueo)

        return JsonResponse({
            'success': True,
            'message': 'Arqueo creado exitosamente',
            'arqueo_id': arqueo.id,
            'cuadratura': cuadratura_data
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear arqueo: {str(e)}'
        })


@login_required
@require_POST
def guardar_conteo_fisico(request):
    """Guardar conteo físico del arqueo"""
    try:
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')
        
        if not arqueo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de arqueo requerido'
            })
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        arqueos_qs = ArqueoCaja.objects.all()
        if sucursal_id:
            arqueos_qs = arqueos_qs.filter(sucursal_id=sucursal_id)
        arqueo = get_object_or_404(arqueos_qs, id=arqueo_id)
        
        # Verificar que el usuario puede modificar este arqueo
        if arqueo.estado not in ['ABIERTO', 'CON_DIFERENCIAS']:
            return JsonResponse({
                'success': False,
                'error': 'Este arqueo ya está cerrado'
            })
        
        # Verificar modo de arqueo
        modo_express = data.get('modo_express', False)
        
        if modo_express:
            # Modo Express: usar monto total directamente
            monto_total = data.get('monto_total_express', 0)
            
            # Validar monto
            try:
                monto_total = int(monto_total)
                if monto_total < 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'El monto debe ser mayor o igual a 0'
                    })
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': 'Monto inválido. Debe ser un número válido'
                })
            
            # Limpiar denominaciones (ya que no se usan en modo express)
            arqueo.billetes_20000 = 0
            arqueo.billetes_10000 = 0
            arqueo.billetes_5000 = 0
            arqueo.billetes_2000 = 0
            arqueo.billetes_1000 = 0
            arqueo.monedas_500 = 0
            arqueo.monedas_100 = 0
            arqueo.monedas_50 = 0
            arqueo.monedas_10 = 0
            arqueo.monedas_5 = 0
            arqueo.monedas_1 = 0
            
            # Establecer el total físico directamente
            arqueo.total_efectivo_fisico = monto_total
            # Misma fórmula que `ArqueoCaja.save()` (models/caja.py): el fondo
            # fijo de caja chica es parte del efectivo que debería estar en el
            # cajón. Sin sumarlo, el modo express reportaba una diferencia
            # distinta a la del modo detallado para el mismo conteo. Hoy los
            # 624 arqueos tienen `fondo_fijo_snapshot = 0`, así que el cambio
            # no altera ningún dato existente.
            arqueo.diferencia_efectivo = monto_total - (
                arqueo.total_efectivo_teorico + (arqueo.fondo_fijo_snapshot or 0)
            )
            
            logger.debug(
                "Conteo de arqueo %s en modo express: total_fisico=%s, diferencia=%s",
                arqueo.id,
                monto_total,
                arqueo.diferencia_efectivo,
            )
            
        else:
            # Modo Detallado: usar denominaciones
            arqueo.billetes_20000 = int(data.get('billetes_20000', 0))
            arqueo.billetes_10000 = int(data.get('billetes_10000', 0))
            arqueo.billetes_5000 = int(data.get('billetes_5000', 0))
            arqueo.billetes_2000 = int(data.get('billetes_2000', 0))
            arqueo.billetes_1000 = int(data.get('billetes_1000', 0))
            
            arqueo.monedas_500 = int(data.get('monedas_500', 0))
            arqueo.monedas_100 = int(data.get('monedas_100', 0))
            arqueo.monedas_50 = int(data.get('monedas_50', 0))
            arqueo.monedas_10 = int(data.get('monedas_10', 0))
            arqueo.monedas_5 = int(data.get('monedas_5', 0))
            arqueo.monedas_1 = int(data.get('monedas_1', 0))
            
            logger.debug("Conteo de arqueo %s en modo detallado desde denominaciones", arqueo.id)

        arqueo.timestamp_conteo_fisico = timezone.now()
        arqueo.modo_conteo = 'EXPRESS' if modo_express else 'DETALLADO'
        if modo_express:
            arqueo.requiere_revision_express = True

        # Observaciones
        arqueo.observaciones = data.get('observaciones', '')
        arqueo.observaciones_diferencia = data.get('observaciones_diferencia', '')
        
        # Datos de Transbank (cierre POS)
        cierre_debito = int(data.get('cierre_debito', 0))
        cierre_credito = int(data.get('cierre_credito', 0))
        numero_lote = data.get('numero_lote', '')
        
        # Calcular total y diferencias de Transbank
        cierre_pos_total = cierre_debito + cierre_credito
        diferencia_debito = cierre_debito - arqueo.total_tarjeta_debito_teorico
        diferencia_credito = cierre_credito - arqueo.total_tarjeta_credito_teorico
        diferencia_transbank = cierre_pos_total - arqueo.total_transbank_teorico
        
        logger.debug(
            "Conteo Transbank arqueo %s: debito=%s teorico_debito=%s, credito=%s teorico_credito=%s, "
            "total=%s teorico_total=%s",
            arqueo.id,
            cierre_debito,
            arqueo.total_tarjeta_debito_teorico,
            cierre_credito,
            arqueo.total_tarjeta_credito_teorico,
            cierre_pos_total,
            arqueo.total_transbank_teorico,
        )
        
        # Guardar según el modo
        if modo_express:
            # En modo express, guardamos directamente sin recalcular
            # Usar update() para evitar que el método save() recalcule automáticamente
            ArqueoCaja.objects.filter(id=arqueo.id).update(
                total_efectivo_fisico=arqueo.total_efectivo_fisico,
                diferencia_efectivo=arqueo.diferencia_efectivo,
                observaciones=arqueo.observaciones,
                observaciones_diferencia=arqueo.observaciones_diferencia,
                billetes_20000=arqueo.billetes_20000,
                billetes_10000=arqueo.billetes_10000,
                billetes_5000=arqueo.billetes_5000,
                billetes_2000=arqueo.billetes_2000,
                billetes_1000=arqueo.billetes_1000,
                monedas_500=arqueo.monedas_500,
                monedas_100=arqueo.monedas_100,
                monedas_50=arqueo.monedas_50,
                monedas_10=arqueo.monedas_10,
                monedas_5=arqueo.monedas_5,
                monedas_1=arqueo.monedas_1,
                # Transbank
                cierre_debito_fisico=cierre_debito,
                cierre_credito_fisico=cierre_credito,
                cierre_pos_fisico=cierre_pos_total,
                numero_lote_pos=numero_lote,
                diferencia_debito=diferencia_debito,
                diferencia_credito=diferencia_credito,
                diferencia_transbank=diferencia_transbank,
                # Estos tres se asignaban en memoria más arriba y se perdían:
                # el `.update()` no los incluía. Por eso en producción los 624
                # arqueos quedaron con `modo_conteo='DETALLADO'` y sólo 3 con
                # `timestamp_conteo_fisico`, dejando muertos los indicadores
                # `uso_express_pct` y `anomalias_timing` de AnalisisFraudeCaja.
                timestamp_conteo_fisico=arqueo.timestamp_conteo_fisico,
                modo_conteo=arqueo.modo_conteo,
                requiere_revision_express=arqueo.requiere_revision_express,
            )
            logger.info(
                "Conteo guardado en modo express: arqueo_id=%s, total_fisico=%s, diferencia=%s",
                arqueo.id,
                arqueo.total_efectivo_fisico,
                arqueo.diferencia_efectivo,
            )
        else:
            # En modo detallado, save() calculará automáticamente el total físico y diferencia
            # Pero primero guardamos los valores de Transbank
            arqueo.cierre_debito_fisico = cierre_debito
            arqueo.cierre_credito_fisico = cierre_credito
            arqueo.cierre_pos_fisico = cierre_pos_total
            arqueo.numero_lote_pos = numero_lote
            arqueo.diferencia_debito = diferencia_debito
            arqueo.diferencia_credito = diferencia_credito
            arqueo.diferencia_transbank = diferencia_transbank
            arqueo.save()
            logger.info(
                "Conteo guardado en modo detallado: arqueo_id=%s, total_fisico=%s",
                arqueo.id,
                arqueo.total_efectivo_fisico,
            )
        
        # Recargar el objeto para obtener los valores actualizados
        arqueo.refresh_from_db()

        log_accion_caja(request, 'GUARDAR_CONTEO', arqueo)

        return JsonResponse({
            'success': True,
            'message': 'Conteo guardado exitosamente',
            'modo_usado': 'express' if modo_express else 'detallado',
            'arqueo': {
                'id': arqueo.id,
                'efectivo_fisico': arqueo.total_efectivo_fisico,
                'efectivo_teorico': arqueo.total_efectivo_teorico,
                'diferencia': arqueo.diferencia_efectivo,
                'diferencia_absoluta': arqueo.diferencia_absoluta,
                'tipo_diferencia': arqueo.tipo_diferencia,
                'porcentaje_diferencia': round(arqueo.porcentaje_diferencia, 2),
                'requiere_supervision': arqueo.requiere_supervision,
                'estado': arqueo.estado
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar conteo: {str(e)}'
        })


@login_required
@require_POST
def cerrar_arqueo(request):
    """Cerrar arqueo definitivamente"""
    try:
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')

        if not arqueo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de arqueo requerido'
            })

        # Scoping por sucursal activa (mismo patrón que `guardar_conteo_fisico`).
        # Antes se cargaba el arqueo sólo por id: cualquier usuario autenticado
        # podía cerrar el arqueo de otra sucursal mandando su id por POST.
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        arqueos_qs = ArqueoCaja.objects.all()
        if sucursal_id:
            arqueos_qs = arqueos_qs.filter(sucursal_id=sucursal_id)
        arqueo = get_object_or_404(arqueos_qs, id=arqueo_id)
        
        # Verificar que el usuario puede cerrar este arqueo
        if arqueo.estado not in ['ABIERTO', 'CON_DIFERENCIAS']:
            return JsonResponse({
                'success': False,
                'error': 'Este arqueo ya está cerrado'
            })
        
        logger.debug(
            "Intentando cerrar arqueo %s: estado=%s, diferencia=%s, observaciones_diferencia=%s, "
            "requiere_supervision=%s",
            arqueo_id,
            arqueo.estado,
            arqueo.diferencia_efectivo,
            arqueo.observaciones_diferencia,
            arqueo.requiere_supervision,
        )
        
        # Validar observaciones SOLO si hay diferencias significativas (mayor a $500)
        diferencia_absoluta = abs(arqueo.diferencia_efectivo)
        
        if diferencia_absoluta == 0:
            logger.debug("Arqueo %s sin diferencias", arqueo_id)
        elif diferencia_absoluta > 500:
            obs = (arqueo.observaciones_diferencia or '').strip()
            if len(obs) < 20 or len(set(obs.split())) < 3:
                return JsonResponse({
                    'success': False,
                    'error': f'Debe agregar observaciones detalladas (mínimo 20 caracteres y 3 palabras distintas) para diferencias mayores a $500 (actual: ${diferencia_absoluta:,})'
                })
        else:
            logger.debug(
                "Arqueo %s con diferencia menor sin observacion obligatoria: diferencia_absoluta=%s",
                arqueo_id,
                diferencia_absoluta,
            )
        
        # Cerrar arqueo
        fecha_cierre = timezone.now()
        
        # Determinar estado final
        if arqueo.diferencia_efectivo == 0:
            estado_final = 'CERRADO'
        else:
            estado_final = 'CON_DIFERENCIAS'
        
        # Usar update() en lugar de save() para NO recalcular el total_efectivo_fisico
        # Esto es crítico para mantener el valor correcto en modo Express
        ArqueoCaja.objects.filter(id=arqueo.id).update(
            fecha_cierre=fecha_cierre,
            estado=estado_final
        )
        
        # Recargar para obtener valores actualizados
        arqueo.refresh_from_db()
        
        logger.info("Arqueo %s cerrado exitosamente: estado_final=%s", arqueo_id, arqueo.estado)

        log_accion_caja(request, 'CERRAR_ARQUEO', arqueo)

        return JsonResponse({
            'success': True,
            'message': 'Arqueo cerrado exitosamente',
            'estado_final': arqueo.get_estado_display(),
            'arqueo': {
                'id': arqueo.id,
                'estado': arqueo.estado,
                'fecha_cierre': arqueo.fecha_cierre.strftime('%d/%m/%Y %H:%M'),
                'diferencia': arqueo.diferencia_efectivo,
                'requiere_supervision': arqueo.requiere_supervision
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cerrar arqueo: {str(e)}'
        })


# ========== FUNCIONES DE SUPERVISIÓN (ADMINISTRACIÓN/ADMINISTRADOR) ==========

@login_required
@require_POST
def revisar_arqueo(request):
    """
    Revisar y aprobar un arqueo (solo supervisores: administración/administrador)
    Soporta resultado_revision: OK, OK_CON_OBS, REQUIERE_ACCION
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        
        if not es_supervisor:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para revisar arqueos. Se requiere rol de Administración o Administrador.'
            })
        
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')
        observaciones_supervisor = data.get('observaciones', '')
        aprobar = data.get('aprobar', True)
        resultado = data.get('resultado_revision', '')
        # Permite aprobar OK dejando constancia de que el depósito todavía no
        # está confirmado (ver más abajo por qué la puerta dura no servía).
        forzar_sin_deposito = bool(data.get('forzar_sin_deposito'))

        if not arqueo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de arqueo requerido'
            })

        # Scoping por sucursal: antes se cargaba sólo por id, así que un
        # supervisor de una empresa podía aprobar arqueos de otra mandando el
        # id por POST.
        sucursales_ok, _ = _sucursales_permitidas(request)
        arqueos_qs = ArqueoCaja.objects.all()
        if sucursales_ok:
            arqueos_qs = arqueos_qs.filter(sucursal_id__in=sucursales_ok)
        arqueo = get_object_or_404(arqueos_qs, id=arqueo_id)

        es_aprobacion_ok = resultado == 'OK' or (
            aprobar and resultado not in ('OK_CON_OBS', 'REQUIERE_ACCION')
        )
        nota_deposito_forzado = ''
        if es_aprobacion_ok:
            esperado_deposito = (
                (arqueo.total_efectivo_teorico or 0)
                + (arqueo.total_cheque_teorico or 0)
            )
            if esperado_deposito > 0:
                dep_agg = arqueo.depositos.aggregate(
                    total_verificado=Sum(
                        Case(
                            When(
                                verificado=True,
                                monto_confirmado__gt=0,
                                then=F('monto_confirmado'),
                            ),
                            When(verificado=True, then=F('monto')),
                            default=Value(0),
                            output_field=IntegerField(),
                        )
                    )
                )
                total_verificado = dep_agg['total_verificado'] or 0
                if total_verificado == 0:
                    estado_deposito = 'SIN_DEPOSITO'
                elif abs(total_verificado - esperado_deposito) <= TOLERANCIA_ARQUEO_DEPOSITO:
                    estado_deposito = 'COMPLETO'
                else:
                    estado_deposito = 'PARCIAL'

                # La puerta exigía que el depósito verificado igualara el
                # efectivo teórico del día (±$1.000) para poder aprobar OK. En
                # la operación real eso casi nunca se cumple: el depósito se
                # hace con rezago, un mismo comprobante cubre varios días
                # (GrupoDeposito existe justo para eso) y parte del efectivo se
                # queda como fondo fijo. Resultado: el supervisor no podía
                # aprobar prácticamente ningún día con efectivo y quedaba
                # obligado a escribir una justificación a mano cada vez —
                # trabajo repetido sin control adicional.
                #
                # Ahora sigue bloqueando por defecto (nada se aprueba en
                # silencio), pero es franqueable de forma explícita, con motivo
                # obligatorio y traza en bitácora + log de auditoría.
                if estado_deposito != 'COMPLETO' and not forzar_sin_deposito:
                    return JsonResponse({
                        'success': False,
                        'error': (
                            'El depósito bancario de este día todavía no está '
                            'confirmado. Confírmelo, o apruebe indicando el '
                            'motivo por el que se aprueba sin depósito.'
                        ),
                        'estado_deposito': estado_deposito,
                        'total_depositado_verificado': total_verificado,
                        'total_esperado_deposito': esperado_deposito,
                        'diferencia_deposito': total_verificado - esperado_deposito,
                        # El frontend usa esto para ofrecer "aprobar igual"
                        # pidiendo el motivo, en vez de dejar al supervisor sin
                        # salida.
                        'puede_forzar': True,
                    }, status=400)

                if estado_deposito != 'COMPLETO' and forzar_sin_deposito:
                    if len((observaciones_supervisor or '').strip()) < 10:
                        return JsonResponse({
                            'success': False,
                            'error': (
                                'Para aprobar sin depósito confirmado debe '
                                'indicar el motivo (mínimo 10 caracteres).'
                            ),
                        }, status=400)
                    nota_deposito_forzado = (
                        f'Aprobado sin depósito confirmado '
                        f'(estado depósito: {estado_deposito}, '
                        f'verificado ${total_verificado:,} de ${esperado_deposito:,}).'
                    )

        # Segregación de funciones: nadie aprueba su propio arqueo. `crear_arqueo`
        # no tiene gate de rol, así que un administrador que además opera caja
        # podía crear, contar, cerrar y aprobar el mismo día de punta a punta —
        # el control de cuatro ojos existía sólo de palabra.
        if arqueo.usuario_responsable_id == request.user.id:
            return JsonResponse({
                'success': False,
                'error': (
                    'No puede revisar un arqueo del que usted es responsable. '
                    'Debe revisarlo otro supervisor.'
                ),
            }, status=403)

        if resultado == 'REQUIERE_ACCION':
            if not observaciones_supervisor or len(observaciones_supervisor.strip()) < 10:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe explicar qué acción se requiere (mínimo 10 caracteres).'
                })
            nuevo_estado = arqueo.estado
            resultado_rev = 'REQUIERE_ACCION'
            accion_texto = 'marcado como requiere acción'
        elif resultado == 'OK_CON_OBS':
            if not observaciones_supervisor:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe incluir observaciones al aprobar con notas.'
                })
            nuevo_estado = 'REVISADO'
            resultado_rev = 'OK_CON_OBS'
            accion_texto = 'aprobado con observaciones'
        elif resultado == 'OK' or aprobar:
            nuevo_estado = 'REVISADO'
            resultado_rev = 'OK'
            accion_texto = 'aprobado'
        else:
            nuevo_estado = arqueo.estado
            resultado_rev = 'PENDIENTE'
            accion_texto = 'marcado como pendiente de revisión'
        
        if nota_deposito_forzado:
            observaciones_supervisor = (
                f'{observaciones_supervisor.strip()}\n[{nota_deposito_forzado}]'
            )

        ArqueoCaja.objects.filter(id=arqueo.id).update(
            estado=nuevo_estado,
            supervisor_revision=request.user,
            fecha_revision=timezone.now(),
            observaciones_supervisor=observaciones_supervisor,
            resultado_revision=resultado_rev
        )

        if observaciones_supervisor:
            ObservacionArqueo.objects.create(
                arqueo=arqueo,
                usuario=request.user,
                tipo='SUPERVISOR',
                texto=observaciones_supervisor,
                visible_para_cajera=True
            )
        
        arqueo.refresh_from_db()
        
        log_accion_caja(request, 'REVISAR_ARQUEO', arqueo)

        return JsonResponse({
            'success': True,
            'message': f'Arqueo {accion_texto} exitosamente',
            'arqueo': {
                'id': arqueo.id,
                'estado': arqueo.estado,
                'estado_display': arqueo.get_estado_display(),
                'resultado_revision': arqueo.resultado_revision,
                'supervisor': request.user.username,
                'fecha_revision': arqueo.fecha_revision.strftime('%d/%m/%Y %H:%M') if arqueo.fecha_revision else ''
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        logger.exception("Error al revisar arqueo")
        return JsonResponse({
            'success': False,
            'error': f'Error al revisar arqueo: {str(e)}'
        })


@login_required
@require_POST
def revisar_arqueos_lote(request):
    """Aprueba en lote los arqueos que no tienen nada que juzgar.

    Revisar un mes de 13 sucursales son ~340 arqueos y, uno por uno, entre 700 y
    1.700 clicks. La inmensa mayoría de esos días cuadra al peso y no requiere
    criterio humano: el supervisor sólo estaba firmando lo obvio, y ese trabajo
    mecánico es el que hace que la revisión se sienta repetitiva.

    Aquí sólo entran arqueos que cumplen TODAS las condiciones:
      * diferencia de efectivo dentro de la tolerancia,
      * no están ABIERTO,
      * todavía sin veredicto,
      * sin depósitos declarados pendientes de confirmar,
      * el revisor no es el responsable del arqueo.

    Cualquier arqueo con diferencia, depósito a medias o acción pendiente queda
    fuera y se sigue revisando a mano, con su motivo. Cada aprobación deja su
    `LogAccionCaja`, igual que la individual.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'administracion']:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para revisar arqueos.'
            }, status=403)

        data = json.loads(request.body or '{}')
        ids = data.get('arqueo_ids') or []
        if not isinstance(ids, list) or not ids:
            return JsonResponse({
                'success': False, 'error': 'Debe indicar al menos un arqueo.'
            }, status=400)
        ids = [int(i) for i in ids if str(i).isdigit()][:500]

        sucursales_ok, _ = _sucursales_permitidas(request)
        candidatos = ArqueoCaja.objects.filter(id__in=ids)
        if sucursales_ok:
            candidatos = candidatos.filter(sucursal_id__in=sucursales_ok)

        aprobados, omitidos = [], []
        ahora = timezone.now()

        for arqueo in candidatos.select_related('sucursal'):
            etiqueta = f'{arqueo.sucursal.alias} {arqueo.fecha_arqueo:%d-%m-%Y}'
            motivo = None
            if arqueo.estado == 'ABIERTO':
                motivo = 'sigue abierto'
            elif arqueo.resultado_revision != 'PENDIENTE':
                motivo = 'ya tiene veredicto'
            elif abs(arqueo.diferencia_efectivo or 0) > TOLERANCIA_ARQUEO_EFECTIVO:
                motivo = 'tiene diferencia de efectivo'
            elif (arqueo.cache_depositos_pendientes or 0) > 0:
                motivo = 'tiene depósitos sin confirmar'
            elif arqueo.usuario_responsable_id == request.user.id:
                motivo = 'usted es el responsable'

            if motivo:
                omitidos.append({'id': arqueo.id, 'etiqueta': etiqueta, 'motivo': motivo})
                continue

            with transaction.atomic():
                ArqueoCaja.objects.filter(id=arqueo.id).update(
                    estado='REVISADO',
                    supervisor_revision=request.user,
                    fecha_revision=ahora,
                    resultado_revision='OK',
                )
                ObservacionArqueo.objects.create(
                    arqueo=arqueo,
                    usuario=request.user,
                    tipo='SISTEMA',
                    texto=(
                        'Aprobado en lote: efectivo dentro de tolerancia '
                        f'(±${TOLERANCIA_ARQUEO_EFECTIVO:,}) y sin depósitos pendientes.'
                    ),
                    visible_para_cajera=True,
                )
                log_accion_caja(
                    request, 'REVISAR_ARQUEO', arqueo, modo='lote'
                )
            aprobados.append({'id': arqueo.id, 'etiqueta': etiqueta})

        return JsonResponse({
            'success': True,
            'aprobados': len(aprobados),
            'omitidos': len(omitidos),
            'detalle_aprobados': aprobados,
            'detalle_omitidos': omitidos,
            'message': (
                f'{len(aprobados)} arqueo(s) aprobados. '
                f'{len(omitidos)} quedaron para revisión manual.'
            ),
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        logger.exception("Error en revision en lote de arqueos")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def crear_observacion_arqueo(request):
    """Agregar observación a la bitácora de un arqueo (cajera o supervisor)."""
    try:
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')
        texto = (data.get('texto') or '').strip()

        if not arqueo_id or not texto:
            return JsonResponse({'success': False, 'error': 'Arqueo y texto son requeridos.'})
        if len(texto) < 5:
            return JsonResponse({'success': False, 'error': 'La observación debe tener al menos 5 caracteres.'})

        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)

        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        tipo = 'SUPERVISOR' if es_supervisor else 'CAJERA'
        visible = data.get('visible_para_cajera', True)

        obs = ObservacionArqueo.objects.create(
            arqueo=arqueo,
            usuario=request.user,
            tipo=tipo,
            texto=texto,
            visible_para_cajera=visible,
        )

        return JsonResponse({
            'success': True,
            'message': 'Observación registrada.',
            'observacion': {
                'id': obs.id,
                'tipo': obs.tipo,
                'tipo_display': obs.get_tipo_display(),
                'texto': obs.texto,
                'usuario': request.user.get_full_name() or request.user.username,
                'fecha': obs.fecha.strftime('%d/%m/%Y %H:%M'),
                'visible_para_cajera': obs.visible_para_cajera,
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def obtener_bitacora_arqueo(request, arqueo_id):
    """Obtener la bitácora completa de observaciones de un arqueo."""
    try:
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)

        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']

        qs = arqueo.bitacora.select_related('usuario').all()
        if not es_supervisor:
            qs = qs.filter(visible_para_cajera=True)

        observaciones = [{
            'id': obs.id,
            'tipo': obs.tipo,
            'tipo_display': obs.get_tipo_display(),
            'texto': obs.texto,
            'usuario': obs.usuario.get_full_name() or obs.usuario.username,
            'fecha': obs.fecha.strftime('%d/%m/%Y %H:%M'),
            'visible_para_cajera': obs.visible_para_cajera,
        } for obs in qs[:50]]

        return JsonResponse({
            'success': True,
            'observaciones': observaciones,
            'resultado_revision': getattr(arqueo, 'resultado_revision', 'PENDIENTE'),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def obtener_bloqueos_arqueo(request, fecha):
    """Retornar lista de bloqueos activos para cerrar un día."""
    try:
        from datetime import date as dt_date
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Sin sucursal.'})

        fecha_obj = dt_date.fromisoformat(fecha)
        bloqueos = []
        completados = []

        try:
            arqueo = ArqueoCaja.objects.get(fecha_arqueo=fecha_obj, sucursal_id=sucursal_id)
        except ArqueoCaja.DoesNotExist:
            bloqueos.append({
                'codigo': 'SIN_ARQUEO',
                'titulo': 'Arqueo no iniciado',
                'descripcion': 'No se ha creado el arqueo para este día.',
                'bloqueante': True,
                'icono': 'ri-calendar-close-line',
            })
            return JsonResponse({'success': True, 'bloqueos': bloqueos, 'completados': completados})

        tiene_conteo = arqueo.total_efectivo_fisico > 0 or arqueo.modo_conteo == 'EXPRESS'
        if arqueo.estado == 'ABIERTO' and not tiene_conteo:
            bloqueos.append({
                'codigo': 'SIN_CONTEO',
                'titulo': 'Falta conteo de efectivo',
                'descripcion': 'Debe contar el efectivo físico en caja.',
                'bloqueante': True,
                'icono': 'ri-money-dollar-circle-line',
            })
        else:
            completados.append({
                'codigo': 'CONTEO_OK',
                'titulo': 'Conteo de efectivo realizado',
                'icono': 'ri-check-line',
            })

        if abs(arqueo.diferencia_efectivo) > 500 and not arqueo.observaciones_diferencia:
            bloqueos.append({
                'codigo': 'SIN_EXPLICACION',
                'titulo': 'Diferencia > $500 sin explicar',
                'descripcion': f'Diferencia de ${abs(arqueo.diferencia_efectivo):,}. Debe agregar observaciones (min 20 chars, 3 palabras).',
                'bloqueante': True,
                'icono': 'ri-error-warning-line',
            })
        elif abs(arqueo.diferencia_efectivo) > 500:
            completados.append({
                'codigo': 'EXPLICACION_OK',
                'titulo': 'Diferencia explicada',
                'icono': 'ri-check-line',
            })

        if arqueo.estado in ['CERRADO', 'CON_DIFERENCIAS', 'DEPOSITO_DECLARADO', 'DEPOSITO_CONFIRMADO', 'REVISADO']:
            completados.append({
                'codigo': 'CIERRE_OK',
                'titulo': 'Arqueo cerrado',
                'icono': 'ri-check-double-line',
            })
        elif arqueo.estado == 'ABIERTO':
            bloqueos.append({
                'codigo': 'SIN_CIERRE',
                'titulo': 'Arqueo aún abierto',
                'descripcion': 'Complete el conteo y cierre el arqueo.',
                'bloqueante': False,
                'icono': 'ri-lock-line',
            })

        tiene_deposito = arqueo.depositos.filter(monto_declarado__gt=0).exists() or arqueo.depositos.filter(verificado=True).exists()
        deposito_confirmado = arqueo.depositos.filter(verificado=True).exists()
        if deposito_confirmado:
            completados.append({
                'codigo': 'DEPOSITO_OK',
                'titulo': 'Depósito confirmado',
                'icono': 'ri-bank-line',
            })
        elif tiene_deposito:
            completados.append({
                'codigo': 'DEPOSITO_DECLARADO',
                'titulo': 'Depósito declarado (pendiente confirmación)',
                'icono': 'ri-time-line',
            })
        elif arqueo.estado not in ['ABIERTO']:
            bloqueos.append({
                'codigo': 'SIN_DEPOSITO',
                'titulo': 'Depósito pendiente',
                'descripcion': 'Declare el depósito bancario del efectivo.',
                'bloqueante': False,
                'icono': 'ri-bank-line',
            })

        resultado_rev = getattr(arqueo, 'resultado_revision', 'PENDIENTE')
        if resultado_rev == 'REQUIERE_ACCION':
            ultima_obs = arqueo.bitacora.filter(tipo='SUPERVISOR').first()
            bloqueos.append({
                'codigo': 'REQUIERE_ACCION',
                'titulo': 'El supervisor requiere acción',
                'descripcion': ultima_obs.texto[:120] if ultima_obs else 'Revise las observaciones del supervisor.',
                'bloqueante': False,
                'icono': 'ri-alarm-warning-line',
            })
        elif resultado_rev in ['OK', 'OK_CON_OBS']:
            completados.append({
                'codigo': 'REVISION_OK',
                'titulo': 'Revisado por supervisor',
                'icono': 'ri-shield-check-line',
            })

        return JsonResponse({
            'success': True,
            'bloqueos': bloqueos,
            'completados': completados,
            'estado': arqueo.estado,
            'resultado_revision': resultado_rev,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def registrar_comprobante_supervisor(request):
    """
    Registrar comprobante de pago bancario (solo supervisores)
    Permite adjuntar imagen del comprobante
    """
    try:
        # Verificar permisos de supervisor
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        
        if not es_supervisor:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para registrar comprobantes. Se requiere rol de Administración o Administrador.'
            })
        
        arqueo_id = request.POST.get('arqueo_id')
        monto = int(request.POST.get('monto', 0))
        banco = request.POST.get('banco', 'ESTADO')
        numero_comprobante = request.POST.get('numero_comprobante', '')
        observaciones = request.POST.get('observaciones', '')
        fecha_deposito = request.POST.get('fecha_deposito')
        imagen = request.FILES.get('imagen_comprobante')
        
        if not arqueo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de arqueo requerido'
            })
        
        if monto <= 0:
            return JsonResponse({
                'success': False,
                'error': 'El monto debe ser mayor a 0'
            })
        
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
        
        # Validar que no tenga ya un comprobante bancario verificado
        comprobante_existente = DepositoBancario.objects.filter(
            arqueo=arqueo,
            verificado=True,
            numero_comprobante__gt=''
        ).exists()
        if comprobante_existente:
            return JsonResponse({
                'success': False,
                'error': f'Este arqueo ({arqueo.fecha_arqueo.strftime("%d/%m/%Y")}) ya tiene un comprobante bancario registrado. '
                         'Si necesita corregirlo, primero elimine el existente.'
            })
        
        # Convertir fecha
        from datetime import datetime
        if fecha_deposito:
            fecha_dep = datetime.strptime(fecha_deposito, '%Y-%m-%d').date()
        else:
            fecha_dep = arqueo.fecha_arqueo
        
        # Crear depósito bancario
        deposito = DepositoBancario.objects.create(
            arqueo=arqueo,
            fecha_deposito=fecha_dep,
            monto=monto,
            monto_declarado=monto,
            monto_confirmado=monto,
            banco=banco,
            numero_comprobante=numero_comprobante,
            observaciones=observaciones,
            imagen_comprobante=imagen,
            declarado_por=request.user,
            fecha_declaracion=timezone.now(),
            registrado_por=request.user,
            verificado=True,  # Registrado por supervisor = verificado automáticamente
            verificado_por=request.user,
            fecha_verificacion=timezone.now()
        )
        
        logger.info(
            "Comprobante bancario registrado: arqueo_id=%s, monto=%s, banco=%s",
            arqueo_id,
            monto,
            banco,
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Comprobante de ${monto:,} registrado exitosamente',
            'deposito': {
                'id': deposito.id,
                'monto': deposito.monto,
                'banco': deposito.get_banco_display(),
                'numero_comprobante': deposito.numero_comprobante,
                'tiene_imagen': bool(deposito.imagen_comprobante)
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al registrar comprobante: {str(e)}'
        })


@login_required
@require_POST
def verificar_deposito(request):
    """
    Verificar un depósito bancario (solo supervisores)
    """
    try:
        # Verificar permisos de supervisor
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ['administrador', 'administracion']
        
        if not es_supervisor:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para verificar depósitos.'
            })
        
        data = json.loads(request.body)
        deposito_id = data.get('deposito_id')
        
        if not deposito_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de depósito requerido'
            })
        
        deposito = get_object_or_404(DepositoBancario, id=deposito_id)
        
        DepositoBancario.objects.filter(id=deposito_id).update(
            verificado=True,
            verificado_por=request.user,
            fecha_verificacion=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Depósito verificado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar depósito: {str(e)}'
        })


@login_required
@require_GET
def obtener_depositos_arqueo(request, arqueo_id):
    """Obtener depósitos bancarios de un arqueo"""
    try:
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
        
        depositos = DepositoBancario.objects.filter(arqueo=arqueo).order_by('-fecha_declaracion', '-fecha_deposito')
        
        depositos_data = []
        for dep in depositos:
            depositos_data.append({
                'id': dep.id,
                'fecha_deposito': dep.fecha_deposito.strftime('%d/%m/%Y'),
                'monto': dep.monto,
                'monto_declarado': dep.monto_declarado,
                'monto_confirmado': dep.monto_confirmado,
                'tipo_medio': dep.tipo_medio,
                'tipo_medio_display': dep.get_tipo_medio_display(),
                'banco': dep.banco,
                'banco_display': dep.get_banco_display(),
                'numero_comprobante': dep.numero_comprobante,
                'observaciones': dep.observaciones,
                'tiene_imagen': bool(dep.imagen_comprobante),
                'imagen_url': dep.imagen_comprobante.url if dep.imagen_comprobante else None,
                'verificado': dep.verificado,
                'verificado_por': dep.verificado_por.username if dep.verificado_por else None,
                'declarado_por': (dep.declarado_por.get_full_name() or dep.declarado_por.username) if dep.declarado_por else None,
                'fecha_declaracion': dep.fecha_declaracion.strftime('%d/%m/%Y %H:%M') if dep.fecha_declaracion else None,
                'registrado_por': dep.registrado_por.username if dep.registrado_por else None,
                'fecha_registro': dep.fecha_registro.strftime('%d/%m/%Y %H:%M') if dep.fecha_registro else None
            })
        
        return JsonResponse({
            'success': True,
            'depositos': depositos_data,
            'total': sum((d['monto_confirmado'] or d['monto'] or d['monto_declarado'] or 0) for d in depositos_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener depósitos: {str(e)}'
        })


@login_required
@require_GET
def verificar_ventas_post_cierre(request):
    """
    Verificar si hay ventas registradas después de cerrar el arqueo del día.
    Este es un caso común donde el cajero cierra el arqueo pero sigue vendiendo.
    """
    try:
        fecha_str = request.GET.get('fecha')
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not fecha_str or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal requeridas'
            })
        
        from datetime import datetime, time as dt_time
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Buscar arqueo cerrado para la fecha
        arqueo = ArqueoCaja.objects.filter(
            sucursal=sucursal,
            fecha_arqueo=fecha_obj,
            estado__in=['CERRADO', 'CON_DIFERENCIAS', 'REVISADO']
        ).first()
        
        if not arqueo or not arqueo.fecha_cierre:
            return JsonResponse({
                'success': True,
                'tiene_ventas_post_cierre': False,
                'message': 'No hay arqueo cerrado para verificar'
            })
        
        hora_cierre = arqueo.fecha_cierre
        
        # Buscar tickets creados después del cierre del arqueo (mismo día)
        fin_del_dia = timezone.make_aware(datetime.combine(fecha_obj, dt_time.max))
        
        tickets_post_cierre = Ticket.objects.filter(
            sucursal=sucursal,
            created_at__gt=hora_cierre,
            created_at__lte=fin_del_dia,
            estado='PAGADO'
        ).select_related().order_by('created_at')
        
        # También buscar DTEs post-cierre
        dtes_post_cierre = Dte.objects.filter(
            sucursal=sucursal,
            fecha_emision=fecha_obj,  # fecha_emision es DateField, no necesita __date
            created_at__gt=hora_cierre,
            created_at__lte=fin_del_dia
        ).exclude(
            tipo_documento__in=['NOTA DE CREDITO', 'GUIA']
        ).order_by('created_at')
        
        cantidad_tickets = tickets_post_cierre.count()
        cantidad_dtes = dtes_post_cierre.count()
        cantidad_total = cantidad_tickets + cantidad_dtes
        
        if cantidad_total == 0:
            return JsonResponse({
                'success': True,
                'tiene_ventas_post_cierre': False,
                'arqueo_id': arqueo.id,
                'hora_cierre': hora_cierre.strftime('%H:%M:%S')
            })
        
        # Calcular monto total de ventas post-cierre
        monto_tickets = sum(t.total or 0 for t in tickets_post_cierre)
        monto_dtes = sum(d.monto_con_iva or 0 for d in dtes_post_cierre)
        monto_total = monto_tickets + monto_dtes
        
        # Generar detalle
        detalle = []
        
        for ticket in tickets_post_cierre[:10]:  # Limitar a 10 para el modal
            metodos_pago = ', '.join([p.metodo_pago for p in ticket.pagos.all()]) if ticket.pagos.exists() else 'N/A'
            detalle.append({
                'hora': ticket.created_at.strftime('%H:%M:%S'),
                'tipo_documento': 'Ticket',
                'numero': ticket.numero_documento or ticket.id,
                'metodo_pago': metodos_pago,
                'monto': ticket.total or 0
            })
        
        for dte in dtes_post_cierre[:10]:
            metodos_pago = ', '.join([p.metodo_pago for p in dte.dte_asociado.all()]) if dte.dte_asociado.exists() else 'N/A'
            detalle.append({
                'hora': dte.created_at.strftime('%H:%M:%S') if dte.created_at else '-',
                'tipo_documento': dte.tipo_documento,
                'numero': dte.numero_documento,
                'metodo_pago': metodos_pago,
                'monto': dte.monto_con_iva or 0
            })
        
        # Ordenar detalle por hora
        detalle.sort(key=lambda x: x['hora'])
        
        return JsonResponse({
            'success': True,
            'tiene_ventas_post_cierre': True,
            'arqueo_id': arqueo.id,
            'hora_cierre': hora_cierre.strftime('%H:%M:%S'),
            'cantidad_ventas': cantidad_total,
            'cantidad_tickets': cantidad_tickets,
            'cantidad_dtes': cantidad_dtes,
            'monto_total': monto_total,
            'monto_tickets': monto_tickets,
            'monto_dtes': monto_dtes,
            'detalle': detalle,
            'mensaje': f'Se encontraron {cantidad_total} ventas por ${monto_total:,} después del cierre a las {hora_cierre.strftime("%H:%M")}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar ventas post-cierre: {str(e)}'
        })


@login_required
@require_POST
def reabrir_arqueo(request):
    """
    Reabrir un arqueo cerrado para incluir ventas post-cierre.
    Recalcula automáticamente los totales teóricos.
    Permisos con tolerancia de días:
    - administrador: sin límite (configurable via ParametroGlobal)
    - jefe_local / administracion: dentro de N días (default 2)
    Requiere justificación obligatoria. Crea registro de auditoría.
    """
    try:
        from datetime import date as dt_date
        from app.models.caja import HistorialReaperturaArqueo
        from app.models.precios import ParametroGlobal

        rol_usuario = getattr(request.user, 'rol', None)

        data = json.loads(request.body)
        fecha_str = data.get('fecha')
        justificacion = (data.get('justificacion') or '').strip()
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        if not fecha_str or not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Fecha y sucursal requeridas'
            })

        if not justificacion or len(justificacion) < 10:
            return JsonResponse({
                'success': False,
                'error': 'Debe ingresar una justificación de al menos 10 caracteres'
            })

        from datetime import datetime
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        # Verificar permisos y tolerancia de días
        dias_desde_arqueo = (timezone.localdate() - fecha_obj).days

        if rol_usuario == 'administrador':
            param = ParametroGlobal.objects.filter(nombre='DIAS_TOLERANCIA_REAPERTURA_ADMIN').first()
            max_dias = param.valor_entero if param else 0  # 0 = ilimitado
            if max_dias > 0 and dias_desde_arqueo > max_dias:
                return JsonResponse({
                    'success': False,
                    'error': f'Solo puede reabrir arqueos de los últimos {max_dias} días'
                })
        elif rol_usuario in ['jefe_local', 'administracion']:
            param = ParametroGlobal.objects.filter(nombre='DIAS_TOLERANCIA_REAPERTURA_JEFE_LOCAL').first()
            max_dias = param.valor_entero if param else 2
            if dias_desde_arqueo > max_dias:
                return JsonResponse({
                    'success': False,
                    'error': f'Solo puede reabrir arqueos de los últimos {max_dias} días. Han pasado {dias_desde_arqueo} días.'
                })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para reabrir arqueos'
            })

        # Buscar arqueo cerrado para la fecha
        arqueo = ArqueoCaja.objects.filter(
            sucursal=sucursal,
            fecha_arqueo=fecha_obj,
            estado__in=['CERRADO', 'CON_DIFERENCIAS', 'DEPOSITO_DECLARADO', 'DEPOSITO_CONFIRMADO', 'REVISADO']
        ).first()

        if not arqueo:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró un arqueo cerrado para reabrir'
            })

        # Guardar el estado anterior para el log
        estado_anterior = arqueo.estado

        # Crear registro de auditoría
        HistorialReaperturaArqueo.objects.create(
            arqueo=arqueo,
            usuario=request.user,
            estado_anterior=estado_anterior,
            justificacion=justificacion,
        )

        log_accion_caja(request, 'REABRIR_ARQUEO', arqueo, justificacion=justificacion)

        # Reabrir el arqueo
        arqueo.estado = 'ABIERTO'
        arqueo.fecha_cierre = None
        nombre_usuario = request.user.get_full_name() or request.user.username
        arqueo.observaciones = (arqueo.observaciones or '') + f'\n[REABIERTO {timezone.now().strftime("%d/%m/%Y %H:%M")} por {nombre_usuario}] {justificacion}. Estado anterior: {estado_anterior}'
        # Persistir el cambio de estado/observaciones antes de delegar el
        # recálculo de los teóricos al helper unificado (así
        # `_recalcular_teoricos_arqueo` trabaja sobre el arqueo ya "abierto").
        arqueo.save(update_fields=['estado', 'fecha_cierre', 'observaciones'])

        # Recalcular totales teóricos delegando en la misma lógica que usa
        # la cuadratura de caja (`_calcular_cuadratura_data` a través de
        # `_recalcular_teoricos_arqueo`). Esto evita tener dos
        # implementaciones divergentes del cálculo — el bloque manual que
        # había acá omitía, por ejemplo, `total_nc_transferencia` y el
        # descuento de transferencia cuando la NC se pagaba por transferencia.
        _recalcular_teoricos_arqueo(
            arqueo,
            usuario=request.user,
            registrar_bitacora=True,
            razon=f'reapertura — {justificacion[:80]}',
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Arqueo reabierto exitosamente',
            'arqueo': {
                'id': arqueo.id,
                'estado': arqueo.estado,
                'total_efectivo_teorico': arqueo.total_efectivo_teorico,
                'total_tickets_teorico': arqueo.total_tickets_teorico,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al reabrir arqueo: {str(e)}'
        })


@login_required
@require_POST
def cancelar_arqueo(request):
    """Cancelar un arqueo abierto (eliminar)"""
    try:
        data = json.loads(request.body)
        arqueo_id = data.get('arqueo_id')
        
        if not arqueo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de arqueo requerido'
            })
        
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
        
        # Solo se puede cancelar arqueos abiertos
        if arqueo.estado != 'ABIERTO':
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden cancelar arqueos en estado ABIERTO'
            })
        
        # Verificar que el usuario tiene permiso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if arqueo.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permiso para cancelar este arqueo'
            })
        
        # Eliminar el arqueo
        arqueo_fecha = arqueo.fecha_arqueo
        arqueo.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Arqueo del {arqueo_fecha.strftime("%d/%m/%Y")} cancelado exitosamente'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cancelar arqueo: {str(e)}'
        })


@login_required
@require_GET
def analisis_fraude_caja(request):
    """
    Análisis de patrones sospechosos en arqueos de caja.
    Solo accesible para administrador/administración.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ('administrador', 'administracion'):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para ver el análisis de fraude.'
            }, status=403)

        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        usuario_id = request.GET.get('usuario_id')
        meses = int(request.GET.get('meses', 3))

        from app.services.analisis_caja import AnalisisFraudeCaja
        servicio = AnalisisFraudeCaja()

        if usuario_id:
            resultado = servicio.analizar_cajero(int(usuario_id), sucursal_id, meses)
            return JsonResponse({'success': True, 'tipo': 'cajero', 'analisis': resultado})
        elif sucursal_id:
            resultados = servicio.analizar_sucursal(int(sucursal_id), meses)
            return JsonResponse({'success': True, 'tipo': 'sucursal', 'analisis': resultados})
        else:
            return JsonResponse({'success': False, 'error': 'Se requiere sucursal_id o usuario_id'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error en análisis: {str(e)}'})


@login_required
@require_GET
def obtener_arqueo_detalle(request, arqueo_id):
    """Obtener detalle completo de un arqueo.

    Política de recálculo (opción C, "recalcular mientras abierto"):
      - Si el arqueo está `ABIERTO` o `CON_DIFERENCIAS`, re-snapshoteamos
        los teóricos en cada lectura para reflejar ediciones de fecha de
        DTE u otros cambios del día en curso, sin que el operador tenga
        que tocar un botón.
      - Para estados ya finales (`CERRADO`, `REVISADO`,
        `DEPOSITO_DECLARADO`, `DEPOSITO_CONFIRMADO`) se respeta el
        snapshot histórico para auditoría. Si un admin necesita
        actualizarlo, debe llamar explícitamente al endpoint
        `recalcular_teoricos_arqueo`, que deja traza en la bitácora.
    """
    try:
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
        
        # Verificar que el usuario tiene acceso a esta sucursal
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if arqueo.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a este arqueo'
            })

        # Auto-recálculo silencioso para arqueos todavía en curso. No se
        # registra en bitácora porque en un arqueo abierto es el
        # comportamiento esperado (no hay evento que auditar).
        recalculo_auto_aplicado = False
        if arqueo.estado in ('ABIERTO', 'CON_DIFERENCIAS'):
            resultado_recalc = _recalcular_teoricos_arqueo(
                arqueo,
                usuario=None,
                registrar_bitacora=False,
            )
            recalculo_auto_aplicado = resultado_recalc['hay_cambios']

        arqueo_data = {
            'id': arqueo.id,
            'fecha_arqueo': arqueo.fecha_arqueo.strftime('%Y-%m-%d'),
            'fecha_creacion': arqueo.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'usuario_responsable': arqueo.usuario_responsable.username,
            'estado': arqueo.estado,
            'estado_display': arqueo.get_estado_display(),
            # Metadatos del recálculo para que el frontend pueda mostrar un
            # aviso "Teóricos actualizados desde DTE" cuando corresponda.
            'recalculo_auto_aplicado': recalculo_auto_aplicado,
            'puede_recalcular_manual': (
                getattr(request.user, 'rol', None) in ('administrador', 'administracion')
            ),
            
            # Totales teóricos
            'totales_teoricos': {
                'efectivo': arqueo.total_efectivo_teorico,
                'tarjetas_comerciales': {
                    'hites': arqueo.total_hites_teorico,
                    'total': arqueo.total_tarjetas_comerciales_teorico,
                },
                'venta_internet': {
                    'falabella': arqueo.total_falabella_teorico,
                    'paris': arqueo.total_paris_teorico,
                    'ripley': arqueo.total_ripley_teorico,
                    'mercadopago': arqueo.total_mercadopago_teorico,
                    'klap': arqueo.total_klap_teorico,
                    'total': arqueo.total_venta_internet_teorico,
                },
                'transbank': {
                    'debito': arqueo.total_tarjeta_debito_teorico,
                    'credito': arqueo.total_tarjeta_credito_teorico,
                    'total': arqueo.total_transbank_teorico,
                },
                'otros': {
                    'tarjeta_debito': arqueo.total_tarjeta_debito_teorico,
                    'tarjeta_credito': arqueo.total_tarjeta_credito_teorico,
                    'transbank': arqueo.total_transbank_teorico,
                    'transferencia': arqueo.total_transferencia_teorico,
                    'cheque': arqueo.total_cheque_teorico,
                    'convenio': arqueo.total_convenio_teorico,
                },
                'documentos': {
                    'tickets': arqueo.total_tickets_teorico,
                    'boletas_electronicas': arqueo.total_boletas_electronicas_teorico,
                    'facturas': arqueo.total_facturas_teorico,
                    'facturas_exentas': arqueo.total_facturas_exentas_teorico,
                    'notas_credito': arqueo.total_notas_credito_teorico,
                },
                'venta_total': arqueo.venta_total_teorica,
            },
            
            # Conteo físico
            'conteo_fisico': {
                'billetes': {
                    '20000': arqueo.billetes_20000,
                    '10000': arqueo.billetes_10000,
                    '5000': arqueo.billetes_5000,
                    '2000': arqueo.billetes_2000,
                    '1000': arqueo.billetes_1000,
                },
                'monedas': {
                    '500': arqueo.monedas_500,
                    '100': arqueo.monedas_100,
                    '50': arqueo.monedas_50,
                    '10': arqueo.monedas_10,
                    '5': arqueo.monedas_5,
                    '1': arqueo.monedas_1,
                },
                'total_fisico': arqueo.total_efectivo_fisico,
            },
            
            # Cierre Transbank
            'cierre_transbank': {
                'debito_fisico': arqueo.cierre_debito_fisico,
                'credito_fisico': arqueo.cierre_credito_fisico,
                'total_fisico': arqueo.cierre_pos_fisico,
                'numero_lote': arqueo.numero_lote_pos,
                'diferencia_debito': arqueo.diferencia_debito,
                'diferencia_credito': arqueo.diferencia_credito,
                'diferencia_total': arqueo.diferencia_transbank,
            },
            
            # Diferencias
            'diferencias': {
                'efectivo': arqueo.diferencia_efectivo,
                'transbank': arqueo.diferencia_transbank,
                'absoluta': arqueo.diferencia_absoluta,
                'tipo': arqueo.tipo_diferencia,
                'porcentaje': round(arqueo.porcentaje_diferencia, 2),
                'requiere_supervision': arqueo.requiere_supervision,
            },
            
            # Observaciones
            'observaciones': arqueo.observaciones or '',
            'observaciones_diferencia': arqueo.observaciones_diferencia or '',
            
            # Supervisión
            'supervisor': arqueo.supervisor_revision.username if arqueo.supervisor_revision else '',
            'fecha_revision': arqueo.fecha_revision.strftime('%d/%m/%Y %H:%M') if arqueo.fecha_revision else '',
            'observaciones_supervisor': arqueo.observaciones_supervisor or '',
            'resultado_revision': getattr(arqueo, 'resultado_revision', 'PENDIENTE'),
            
            # Fechas
            'fecha_cierre': arqueo.fecha_cierre.strftime('%d/%m/%Y %H:%M') if arqueo.fecha_cierre else '',
            
            # Bitácora visible para cajera
            'bitacora': [{
                'id': obs.id,
                'tipo': obs.tipo,
                'tipo_display': obs.get_tipo_display(),
                'texto': obs.texto,
                'usuario': obs.usuario.get_full_name() or obs.usuario.username,
                'fecha': obs.fecha.strftime('%d/%m/%Y %H:%M'),
            } for obs in arqueo.bitacora.filter(visible_para_cajera=True)[:20]],
        }
        
        return JsonResponse({
            'success': True,
            'arqueo': arqueo_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener arqueo: {str(e)}'
        })


@login_required
@require_POST
def recalcular_teoricos_arqueo(request, arqueo_id):
    """Fuerza el recálculo de los `total_*_teorico` de un arqueo desde los DTEs.

    Pensado para cuando el admin corrige la fecha de un DTE emitido en un
    día cuyo arqueo ya está cerrado/revisado, y necesita que los teóricos
    vuelvan a cuadrar con la nueva realidad.

    Permisos:
      - `ABIERTO` / `CON_DIFERENCIAS`: cualquier usuario con acceso a la
        sucursal (es equivalente al recálculo automático de
        `obtener_arqueo_detalle`, pero explícito).
      - Estados finales (`CERRADO`, `REVISADO`, `DEPOSITO_DECLARADO`,
        `DEPOSITO_CONFIRMADO`): sólo administrador/administración, y
        siempre se registra una observación SISTEMA en la bitácora.
    """
    try:
        arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        rol_usuario = getattr(request.user, 'rol', None)
        es_supervisor = rol_usuario in ('administrador', 'administracion')

        if not es_supervisor and arqueo.sucursal_id != int(sucursal_id or 0):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a este arqueo'
            }, status=403)

        estado_abierto = arqueo.estado in ('ABIERTO', 'CON_DIFERENCIAS')
        if not estado_abierto and not es_supervisor:
            return JsonResponse({
                'success': False,
                'error': (
                    'Este arqueo está cerrado. Sólo administración '
                    'puede recalcular sus teóricos.'
                )
            }, status=403)

        # Intentar extraer razón del body (opcional, JSON libre).
        razon = ''
        if request.body:
            try:
                body = json.loads(request.body)
                razon = str(body.get('razon', '') or '').strip()[:500]
            except (ValueError, TypeError):
                razon = ''

        # Sólo registrar bitácora cuando el usuario disparó el recálculo
        # sobre un arqueo cerrado o con diferencias — en los ABIERTOS es
        # el estado de trabajo normal y saturaría la bitácora.
        registrar = not estado_abierto or bool(razon)

        resultado = _recalcular_teoricos_arqueo(
            arqueo,
            usuario=request.user,
            registrar_bitacora=registrar,
            razon=razon or ('recálculo manual' if not estado_abierto else 'recálculo desde modal'),
        )

        try:
            log_accion_caja(
                request,
                'RECALCULAR_TEORICOS',
                arqueo=arqueo,
                cambios=resultado['cambios'],
                razon=razon,
            )
        except Exception:
            # Nunca romper la operación si el log auxiliar falla.
            pass

        cuadratura = resultado['cuadratura']

        return JsonResponse({
            'success': True,
            'hay_cambios': resultado['hay_cambios'],
            'cambios': resultado['cambios'],
            'arqueo': {
                'id': arqueo.id,
                'estado': arqueo.estado,
                'total_efectivo_teorico': arqueo.total_efectivo_teorico,
                'total_tarjeta_debito_teorico': arqueo.total_tarjeta_debito_teorico,
                'total_tarjeta_credito_teorico': arqueo.total_tarjeta_credito_teorico,
                'total_transbank_teorico': arqueo.total_transbank_teorico,
                'total_hites_teorico': arqueo.total_hites_teorico,
                'total_venta_internet_teorico': arqueo.total_venta_internet_teorico,
                'venta_total_teorica': arqueo.venta_total_teorica,
                'diferencia_efectivo': arqueo.diferencia_efectivo,
                'diferencia_transbank': arqueo.diferencia_transbank,
                'diferencia_debito': arqueo.diferencia_debito,
                'diferencia_credito': arqueo.diferencia_credito,
            },
            'cuadratura': cuadratura,
        })

    except Exception as e:
        logger.exception("Error al recalcular arqueo desde dashboard")
        return JsonResponse({
            'success': False,
            'error': f'Error al recalcular: {str(e)}'
        }, status=500)


# ========== GESTIÓN POS TRANSBANK ==========

@login_required
def gestion_pos_transbank(request):
    """Vista principal para gestión de POS Transbank"""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    if not sucursal_actual:
        return redirect('dashboard')
    
    # Obtener configuraciones POS de la sucursal
    configuraciones_pos = ConfiguracionPOS.objects.filter(
        sucursal=sucursal_actual
    ).order_by('-es_principal', 'nombre')
    
    context = {
        'sucursal_actual': sucursal_actual,
        'configuraciones_pos': configuraciones_pos,
        'tipo_pos_choices': TIPO_POS_CHOICES,
        'metodo_pago_choices': METODO_PAGO_TICKET_CHOICES,
    }
    return render(request, 'vistas/modulo_ventas/gestion_pos_transbank_simple.html', context)


@login_required
@require_POST
@csrf_exempt
def detectar_terminales_pos(request):
    """Detectar y guardar terminales POS automáticamente"""
    try:
        data = json.loads(request.body)
        puertos_detectados = data.get('puertos_detectados', [])
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        terminales_creados = []
        terminales_existentes = []
        
        for puerto in puertos_detectados:
            # Verificar si ya existe una configuración para este puerto
            config_existente = ConfiguracionPOS.objects.filter(
                sucursal=sucursal,
                puerto_conexion=puerto
            ).first()
            
            if config_existente:
                # Actualizar estado de conexión
                config_existente.estado_conexion = 'DETECTADO'
                config_existente.ultima_conexion = timezone.now()
                config_existente.save()
                terminales_existentes.append({
                    'id': config_existente.id,
                    'nombre': config_existente.nombre,
                    'puerto': puerto,
                    'estado': 'existente'
                })
            else:
                # Crear nueva configuración automática
                nombre_auto = f"Terminal Auto {puerto}"
                tipo_pos = 'VERIFONE_520'  # Tipo por defecto, se puede detectar después
                
                nueva_config = ConfiguracionPOS.objects.create(
                    sucursal=sucursal,
                    nombre=nombre_auto,
                    tipo_pos=tipo_pos,
                    puerto_conexion=puerto,
                    velocidad_conexion=115200,  # Velocidad estándar
                    timeout_conexion=30,
                    estado_conexion='DETECTADO',
                    ultima_conexion=timezone.now(),
                    activo=True,
                    es_principal=len(terminales_creados) == 0,  # El primero es principal
                    observaciones=f'Terminal detectado automáticamente en puerto {puerto}'
                )
                
                terminales_creados.append({
                    'id': nueva_config.id,
                    'nombre': nueva_config.nombre,
                    'puerto': puerto,
                    'estado': 'nuevo'
                })
        
        # Marcar como desconectados los terminales que no fueron detectados
        ConfiguracionPOS.objects.filter(
            sucursal=sucursal,
            activo=True
        ).exclude(
            puerto_conexion__in=puertos_detectados
        ).update(
            estado_conexion='DESCONECTADO'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Detección completada: {len(terminales_creados)} nuevos, {len(terminales_existentes)} existentes',
            'terminales_creados': terminales_creados,
            'terminales_existentes': terminales_existentes,
            'total_detectados': len(puertos_detectados)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en detección automática: {str(e)}'
        })


@login_required
@require_GET
def obtener_configuraciones_pos(request):
    """API para obtener configuraciones POS de la sucursal actual"""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        configuraciones = ConfiguracionPOS.objects.filter(
            sucursal_id=sucursal_id
        ).order_by('-es_principal', 'nombre')
        
        configuraciones_data = []
        for config in configuraciones:
            configuraciones_data.append({
                'id': config.id,
                'nombre': config.nombre,
                'tipo_pos': config.tipo_pos,
                'tipo_pos_display': config.get_tipo_pos_display(),
                'puerto_conexion': config.puerto_conexion,
                'velocidad_conexion': config.velocidad_conexion,
                'activo': config.activo,
                'es_principal': config.es_principal,
                'estado_conexion': config.estado_conexion,
                'estado_conexion_display': config.get_estado_conexion_display(),
                'ultima_conexion': config.ultima_conexion.strftime('%d/%m/%Y %H:%M') if config.ultima_conexion else '',
                'numero_serie': config.numero_serie or '',
                'version_firmware': config.version_firmware or '',
                'timeout_conexion': config.timeout_conexion,
                'observaciones': config.observaciones or '',
            })
        
        return JsonResponse({
            'success': True,
            'configuraciones': configuraciones_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener configuraciones: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def crear_configuracion_pos(request):
    """Crear nueva configuración POS"""
    try:
        data = json.loads(request.body)
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Validar campos requeridos
        campos_requeridos = ['nombre', 'tipo_pos', 'puerto_conexion']
        for campo in campos_requeridos:
            if not data.get(campo):
                return JsonResponse({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                })
        
        # Verificar que el nombre no exista en la sucursal
        if ConfiguracionPOS.objects.filter(
            sucursal=sucursal, 
            nombre=data['nombre']
        ).exists():
            return JsonResponse({
                'success': False,
                'error': 'Ya existe una configuración con ese nombre'
            })
        
        # Crear configuración
        configuracion = ConfiguracionPOS.objects.create(
            sucursal=sucursal,
            nombre=data['nombre'],
            tipo_pos=data['tipo_pos'],
            puerto_conexion=data['puerto_conexion'],
            velocidad_conexion=data.get('velocidad_conexion', 115200),
            activo=data.get('activo', True),
            es_principal=data.get('es_principal', False),
            timeout_conexion=data.get('timeout_conexion', 30),
            numero_serie=data.get('numero_serie', ''),
            version_firmware=data.get('version_firmware', ''),
            observaciones=data.get('observaciones', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Configuración POS creada exitosamente',
            'configuracion_id': configuracion.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear configuración: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def probar_conexion_pos(request):
    """Probar conexión con terminal POS"""
    try:
        data = json.loads(request.body)
        configuracion_id = data.get('configuracion_id')
        
        if not configuracion_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de configuración requerido'
            })
        
        configuracion = get_object_or_404(ConfiguracionPOS, id=configuracion_id)
        
        # Verificar que el usuario tiene acceso a esta sucursal
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if configuracion.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a esta configuración'
            })
        
        # NOTA: La conexión real se prueba desde el frontend usando el SDK JavaScript
        # El agente Transbank usa Socket.IO, no WebSocket directo
        # Por lo tanto, el backend solo valida la configuración y retorna datos
        
        # Validar que la configuración es correcta
        result = {
            'success': True,
            'message': 'Configuración validada. La conexión real se probará desde el navegador.',
            'puerto': configuracion.puerto_conexion,
            'velocidad': configuracion.velocidad_conexion,
            'tipo_pos': configuracion.get_tipo_pos_display(),
            'note': 'Use el SDK de JavaScript en el navegador para conectarse al agente Transbank'
        }
        
        # Actualizar estado de conexión
        if result['success']:
            configuracion.ultima_conexion = timezone.now()
            configuracion.estado_conexion = 'VALIDADO'
            configuracion.save()
            
            # Crear log exitoso
            LogPOS.objects.create(
                configuracion_pos=configuracion,
                tipo_evento='VALIDACION',
                mensaje=f'Configuración validada - {result["message"]}',
                datos_tecnicos={
                    'puerto': configuracion.puerto_conexion,
                    'velocidad': configuracion.velocidad_conexion,
                    'tipo_pos': configuracion.tipo_pos,
                    'resultado': 'VALIDADO',
                    'nota': result.get('note', '')
                }
            )
            
            return JsonResponse({
                'success': True,
                'message': result['message'],
                'estado_conexion': configuracion.get_estado_conexion_display(),
                'ultima_conexion': configuracion.ultima_conexion.strftime('%d/%m/%Y %H:%M'),
                'puerto': result.get('puerto', ''),
                'velocidad': result.get('velocidad', 0),
                'tipo_pos': result.get('tipo_pos', ''),
                'note': result.get('note', '')
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        # En caso de error, actualizar estado
        if 'configuracion' in locals():
            configuracion.estado_conexion = 'ERROR'
            configuracion.save()
            
            LogPOS.objects.create(
                configuracion_pos=configuracion,
                tipo_evento='ERROR',
                mensaje=f'Error en prueba de conexión: {str(e)}',
                datos_tecnicos={
                    'puerto': configuracion.puerto_conexion,
                    'error': str(e)
                }
            )
        
        return JsonResponse({
            'success': False,
            'error': f'Error al probar conexión: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def iniciar_venta_pos(request):
    """Iniciar venta en terminal POS"""
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        monto = data.get('monto')
        configuracion_id = data.get('configuracion_id')
        ticket_id = data.get('ticket_id')  # ID del ticket de venta
        
        if not all([monto, configuracion_id]):
            return JsonResponse({
                'success': False,
                'error': 'Monto y configuración POS requeridos'
            })
        
        try:
            monto = float(monto)
            if monto <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'El monto debe ser mayor a 0'
                })
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Monto inválido'
            })
        
        configuracion = get_object_or_404(ConfiguracionPOS, id=configuracion_id)
        
        # Verificar que el usuario tiene acceso a esta sucursal
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if configuracion.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a esta configuración'
            })
        
        # Verificar que la configuración esté activa
        if not configuracion.activo:
            return JsonResponse({
                'success': False,
                'error': 'La configuración POS no está activa'
            })
        
        # Obtener ticket si se proporcionó
        # NOTA: ticket_id puede ser:
        # - Un ID numérico de ticket de venta existente (para asociar pago a ticket)
        # - Un string generado (TXNxxxxxx) para identificar la transacción POS
        ticket = None
        ticket_referencia = ticket_id  # Guardar para usar como referencia
        
        if ticket_id:
            # Intentar convertir a número (si es ID de ticket real)
            try:
                ticket_id_num = int(ticket_id)
                # Es un número, buscar ticket en BD
                try:
                    ticket = Ticket.objects.get(id=ticket_id_num, sucursal_id=sucursal_id)
                except Ticket.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Ticket {ticket_id_num} no encontrado'
                    })
            except (ValueError, TypeError):
                # No es un número, es un string generado (TXNxxxxxx)
                # Esto es válido - se usa como referencia de transacción
                # No se asocia a un ticket de venta
                ticket = None
        
        # Crear transacción POS
        observaciones = data.get('observaciones', '')
        if ticket_referencia and not ticket:
            # Si hay referencia de ticket pero no se asoció a un Ticket de BD
            observaciones = f"Ref: {ticket_referencia}. {observaciones}".strip()
        
        transaccion = TransaccionPOS.objects.create(
            configuracion_pos=configuracion,
            ticket=ticket,
            monto=monto,
            tipo_transaccion='VENTA',
            estado='INICIADA',
            usuario_operador=request.user,
            ip_origen=request.META.get('REMOTE_ADDR'),
            observaciones=observaciones
        )
        
        # Crear log de inicio
        LogPOS.objects.create(
            configuracion_pos=configuracion,
            transaccion_pos=transaccion,
            tipo_evento='COMANDO_ENVIADO',
            mensaje=f'Iniciando venta por ${monto:,}',
            datos_tecnicos={
                'monto': monto,
                'ticket_pos': transaccion.ticket_pos,
                'puerto': configuracion.puerto_conexion
            }
        )
        
        # Ejecutar venta real con SDK Transbank
        from decimal import Decimal
        result = run_transbank_operation(
            execute_pos_sale,
            Decimal(str(monto)),
            transaccion.ticket_pos,
            configuracion.puerto_conexion,
            configuracion.velocidad_conexion
        )
        
        # Actualizar transacción con resultado
        if result['success']:
            transaccion.estado = result['status']
            transaccion.codigo_respuesta = result.get('response_code', '')
            transaccion.mensaje_respuesta = result.get('message', '')
            transaccion.codigo_autorizacion = result.get('authorization_code', '')
            transaccion.tipo_tarjeta = result.get('card_type', 'DESCONOCIDO')
            transaccion.ultimos_4_digitos = result.get('card_number', '')[-4:] if result.get('card_number') else ''
            transaccion.nombre_tarjeta = result.get('card_brand', '')
            transaccion.numero_operacion = result.get('operation_number', '')
            transaccion.numero_cuotas = result.get('installments', 1)
            transaccion.codigo_comercio = result.get('commerce_code', '')
            transaccion.terminal_id = result.get('terminal_id', '')
            transaccion.save()
            
            # Crear log de respuesta
            LogPOS.objects.create(
                configuracion_pos=configuracion,
                transaccion_pos=transaccion,
                tipo_evento='RESPUESTA_RECIBIDA',
                mensaje=f'Venta {result["status"].lower()}: {result["message"]}',
                datos_tecnicos=result.get('raw_response', result)
            )
            
            # Si la transacción fue exitosa y hay un ticket asociado, procesar pago
            if transaccion.es_exitosa and ticket:
                # Determinar método de pago según el tipo de tarjeta
                metodo_pago = 'TBK_POS_INTEGRADO'
                if transaccion.tipo_tarjeta == 'DEBITO':
                    metodo_pago = 'TBK_DEBITO_POS'
                elif transaccion.tipo_tarjeta == 'CREDITO':
                    metodo_pago = 'TBK_CREDITO_POS'
                elif transaccion.tipo_tarjeta == 'PREPAGO':
                    metodo_pago = 'TBK_PREPAGO_POS'
                
                # Crear detalle de pago
                detalle_pago = TicketDetallePago.objects.create(
                    ticket=ticket,
                    metodo_pago=metodo_pago,
                    tipo_tarjeta=transaccion.nombre_tarjeta,
                    voucher=transaccion.codigo_autorizacion,
                    monto=int(transaccion.monto),
                    notas=f'POS {configuracion.nombre} - Oper: {transaccion.numero_operacion}',
                    origen_pago='POS_INTEGRADO',
                )
                
                # Asociar el detalle de pago con la transacción
                transaccion.detalle_pago = detalle_pago
                transaccion.save()
                
                # Actualizar estado del ticket si está completamente pagado
                if ticket.saldo_por_pagar <= 0:
                    ticket.estado = 'PAGADO'
                    ticket.save()
            
            return JsonResponse({
                'success': True,
                'transaccion': {
                    'id': transaccion.id,
                    'ticket_pos': transaccion.ticket_pos,
                    'monto': float(transaccion.monto),
                    'estado': transaccion.estado,
                    'codigo_autorizacion': transaccion.codigo_autorizacion,
                    'tipo_tarjeta': transaccion.tipo_tarjeta,
                    'mensaje': result['message'],
                    'puerto_conexion': configuracion.puerto_conexion
                }
            })
        else:
            # Error en la venta
            transaccion.estado = 'ERROR'
            transaccion.error_detalle = result.get('error', 'Error desconocido')
            transaccion.save()
            
            # Crear log de error
            LogPOS.objects.create(
                configuracion_pos=configuracion,
                transaccion_pos=transaccion,
                tipo_evento='ERROR',
                mensaje=f'Error en venta: {result.get("error", "Error desconocido")}',
                datos_tecnicos=result
            )
            
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Error en la venta POS'),
                'suggestion': result.get('suggestion', 'Verifique el terminal y intente nuevamente'),
                'transaccion_id': transaccion.id
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al iniciar venta POS: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def guardar_venta_pos(request):
    """Guardar venta POS procesada desde el frontend"""
    try:
        data = json.loads(request.body)
        
        sale_response = data.get('sale_response', {})
        ticket_id = data.get('ticket_id')
        monto = data.get('monto')
        
        if not sale_response:
            return JsonResponse({
                'success': False,
                'error': 'Respuesta de venta requerida'
            })
        
        # Obtener sucursal y configuración
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa'
            })
        
        # Buscar configuración POS activa (si hay una detectada recientemente)
        configuracion = ConfiguracionPOS.objects.filter(
            sucursal_id=sucursal_id,
            activo=True
        ).order_by('-ultima_conexion').first()
        
        if not configuracion:
            # Crear configuración temporal si no existe
            from .models import Sucursal
            sucursal = get_object_or_404(Sucursal, id=sucursal_id)
            configuracion = ConfiguracionPOS.objects.create(
                sucursal=sucursal,
                nombre=f"POS Auto",
                tipo_pos='VERIFONE_520',
                puerto_conexion=sale_response.get('activePort', 'AUTO'),
                velocidad_conexion=115200,
                activo=True,
                es_principal=True
            )
        
        # Obtener ticket si se proporcionó ID numérico
        ticket = None
        if ticket_id:
            try:
                ticket_id_num = int(ticket_id)
                ticket = Ticket.objects.get(id=ticket_id_num, sucursal_id=sucursal_id)
            except (ValueError, TypeError, Ticket.DoesNotExist):
                ticket = None
        
        # Crear transacción POS
        transaccion = TransaccionPOS.objects.create(
            configuracion_pos=configuracion,
            ticket=ticket,
            monto=monto or sale_response.get('amount', 0),
            tipo_transaccion='VENTA',
            estado='APROBADA' if sale_response.get('responseCode') == 0 else 'RECHAZADA',
            codigo_respuesta=str(sale_response.get('responseCode', '')),
            mensaje_respuesta=sale_response.get('responseMessage', ''),
            codigo_autorizacion=sale_response.get('authorizationCode', ''),
            tipo_tarjeta='DEBITO' if sale_response.get('cardType') == 'DB' else 'CREDITO',
            ultimos_4_digitos=sale_response.get('last4Digits', ''),
            nombre_tarjeta=sale_response.get('cardBrand', ''),
            numero_operacion=sale_response.get('operationNumber', ''),
            numero_cuotas=1,
            codigo_comercio=sale_response.get('commerceCode', ''),
            terminal_id=sale_response.get('terminalId', ''),
            usuario_operador=request.user,
            ip_origen=request.META.get('REMOTE_ADDR'),
            observaciones=f"Ref: {ticket_id}" if ticket_id and not ticket else ''
        )
        
        # Si hay ticket asociado, crear pago
        if transaccion.es_exitosa and ticket:
            metodo_pago = 'TBK_DEBITO_POS' if sale_response.get('cardType') == 'DB' else 'TBK_CREDITO_POS'
            
            detalle_pago = TicketDetallePago.objects.create(
                ticket=ticket,
                metodo_pago=metodo_pago,
                tipo_tarjeta=sale_response.get('cardBrand', ''),
                voucher=sale_response.get('authorizationCode', ''),
                monto=int(transaccion.monto),
                notas=f'POS - Oper: {sale_response.get("operationNumber", "")}',
                origen_pago='POS_INTEGRADO',
            )
            
            transaccion.detalle_pago = detalle_pago
            transaccion.save()
            
            if ticket.saldo_por_pagar <= 0:
                ticket.estado = 'PAGADO'
                ticket.save()
        
        return JsonResponse({
            'success': True,
            'transaccion_id': transaccion.id,
            'message': 'Transacción guardada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error guardando transacción: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def validar_password_usuario(request):
    """Validar contraseña del usuario actual para autorizaciones"""
    try:
        data = json.loads(request.body)
        password = data.get('password')
        
        if not password:
            return JsonResponse({
                'success': False,
                'error': 'Contraseña requerida'
            })
        
        # Validar contraseña del usuario actual
        usuario = request.user
        
        if usuario.check_password(password):
            return JsonResponse({
                'success': True,
                'usuario': usuario.username,
                'mensaje': 'Contraseña correcta'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Contraseña incorrecta'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en validación: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def completar_transaccion_pos(request):
    """Completar transacción POS con respuesta del terminal"""
    try:
        data = json.loads(request.body)
        
        ticket_pos = data.get('ticket_pos')
        respuesta_pos = data.get('respuesta_pos', {})
        
        if not ticket_pos:
            return JsonResponse({
                'success': False,
                'error': 'Ticket POS requerido'
            })
        
        transaccion = get_object_or_404(TransaccionPOS, ticket_pos=ticket_pos)
        
        # Verificar que el usuario tiene acceso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if transaccion.configuracion_pos.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a esta transacción'
            })
        
        # Actualizar transacción con respuesta del POS
        transaccion.codigo_respuesta = respuesta_pos.get('response_code', '')
        transaccion.mensaje_respuesta = respuesta_pos.get('response_message', '')
        transaccion.codigo_autorizacion = respuesta_pos.get('authorization_code', '')
        transaccion.tipo_tarjeta = respuesta_pos.get('card_type', 'DESCONOCIDO')
        transaccion.ultimos_4_digitos = respuesta_pos.get('card_number', '')[-4:] if respuesta_pos.get('card_number') else ''
        transaccion.nombre_tarjeta = respuesta_pos.get('card_brand', '')
        transaccion.numero_operacion = respuesta_pos.get('operation_number', '')
        transaccion.numero_cuotas = respuesta_pos.get('installments', 1)
        transaccion.codigo_comercio = respuesta_pos.get('commerce_code', '')
        transaccion.terminal_id = respuesta_pos.get('terminal_id', '')
        
        # Determinar estado final
        if respuesta_pos.get('success', False) and transaccion.codigo_autorizacion:
            transaccion.estado = 'APROBADA'
        else:
            transaccion.estado = 'RECHAZADA'
            transaccion.error_detalle = respuesta_pos.get('error_message', 'Transacción rechazada')
        
        transaccion.save()
        
        # Crear log de respuesta
        LogPOS.objects.create(
            configuracion_pos=transaccion.configuracion_pos,
            transaccion_pos=transaccion,
            tipo_evento='RESPUESTA_RECIBIDA',
            mensaje=f'Transacción {transaccion.get_estado_display().lower()}',
            datos_tecnicos=respuesta_pos
        )
        
        # Si la transacción fue exitosa y hay un ticket asociado, crear el detalle de pago
        if transaccion.es_exitosa and transaccion.ticket:
            # Determinar método de pago según el tipo de tarjeta
            metodo_pago = 'TBK_POS_INTEGRADO'
            if transaccion.tipo_tarjeta == 'DEBITO':
                metodo_pago = 'TBK_DEBITO_POS'
            elif transaccion.tipo_tarjeta == 'CREDITO':
                metodo_pago = 'TBK_CREDITO_POS'
            elif transaccion.tipo_tarjeta == 'PREPAGO':
                metodo_pago = 'TBK_PREPAGO_POS'
            
            # Crear detalle de pago
            detalle_pago = TicketDetallePago.objects.create(
                ticket=transaccion.ticket,
                metodo_pago=metodo_pago,
                tipo_tarjeta=transaccion.nombre_tarjeta,
                voucher=transaccion.codigo_autorizacion,
                monto=int(transaccion.monto),
                notas=f'POS {transaccion.configuracion_pos.nombre} - Oper: {transaccion.numero_operacion}',
                origen_pago='POS_INTEGRADO',
            )
            
            # Asociar el detalle de pago con la transacción
            transaccion.detalle_pago = detalle_pago
            transaccion.save()
            
            # Actualizar estado del ticket si está completamente pagado
            if transaccion.ticket.saldo_por_pagar <= 0:
                transaccion.ticket.estado = 'PAGADO'
                transaccion.ticket.save()
        
        return JsonResponse({
            'success': True,
            'transaccion': {
                'id': transaccion.id,
                'ticket_pos': transaccion.ticket_pos,
                'estado': transaccion.estado,
                'estado_display': transaccion.get_estado_display(),
                'es_exitosa': transaccion.es_exitosa,
                'codigo_autorizacion': transaccion.codigo_autorizacion,
                'tipo_tarjeta': transaccion.get_tipo_tarjeta_display() if transaccion.tipo_tarjeta else '',
                'ultimos_4_digitos': transaccion.ultimos_4_digitos,
                'nombre_tarjeta': transaccion.nombre_tarjeta,
                'duracion': transaccion.duracion_transaccion,
                'puede_anular': transaccion.puede_anular
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al completar transacción: {str(e)}'
        })


@login_required
@require_GET
def obtener_transacciones_pos(request):
    """Obtener historial de transacciones POS"""
    try:
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        estado = request.GET.get('estado')
        configuracion_id = request.GET.get('configuracion_id')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Construir queryset
        queryset = TransaccionPOS.objects.select_related(
            'configuracion_pos', 'ticket', 'detalle_pago', 'usuario_operador'
        ).filter(
            configuracion_pos__sucursal_id=sucursal_id
        )
        
        # Aplicar filtros
        if fecha_desde:
            queryset = queryset.filter(fecha_inicio__date__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_inicio__date__lte=fecha_hasta)
        if estado:
            queryset = queryset.filter(estado=estado)
        if configuracion_id:
            queryset = queryset.filter(configuracion_pos_id=configuracion_id)
        
        # Paginación
        from django.core.paginator import Paginator
        paginator = Paginator(queryset, per_page)
        transacciones_page = paginator.get_page(page)
        
        # Serializar datos
        transacciones_data = []
        for transaccion in transacciones_page:
            transacciones_data.append({
                'id': transaccion.id,
                'ticket_pos': transaccion.ticket_pos,
                'fecha_inicio': transaccion.fecha_inicio.strftime('%d/%m/%Y %H:%M:%S'),
                'fecha_completada': transaccion.fecha_completada.strftime('%d/%m/%Y %H:%M:%S') if transaccion.fecha_completada else '',
                'monto': float(transaccion.monto),
                'estado': transaccion.estado,
                'estado_display': transaccion.get_estado_display(),
                'tipo_transaccion': transaccion.get_tipo_transaccion_display(),
                'configuracion_pos': transaccion.configuracion_pos.nombre,
                'codigo_autorizacion': transaccion.codigo_autorizacion or '',
                'tipo_tarjeta': transaccion.get_tipo_tarjeta_display() if transaccion.tipo_tarjeta else '',
                'ultimos_4_digitos': transaccion.ultimos_4_digitos or '',
                'nombre_tarjeta': transaccion.nombre_tarjeta or '',
                'ticket_id': transaccion.ticket.id if transaccion.ticket else None,
                'ticket_correlativo': transaccion.ticket.correlativo if transaccion.ticket else '',
                'usuario_operador': transaccion.usuario_operador.username if transaccion.usuario_operador else '',
                'duracion': transaccion.duracion_transaccion,
                'es_exitosa': transaccion.es_exitosa,
                'puede_anular': transaccion.puede_anular,
                'error_detalle': transaccion.error_detalle or '',
                # `numero_operacion` es el dato que aparece en el voucher y en
                # el cierre de lote del POS; `observaciones` es hoy el único
                # lugar donde queda el correlativo del ticket
                # ('Ticket POS: TKT184117') cuando la FK `ticket` viene nula.
                # Sin ellos la conciliación manual obliga a abrir otra pantalla.
                'numero_operacion': transaccion.numero_operacion or '',
                'observaciones': transaccion.observaciones or '',
            })
        
        return JsonResponse({
            'success': True,
            'transacciones': transacciones_data,
            'pagination': {
                'current_page': transacciones_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': transacciones_page.has_next(),
                'has_previous': transacciones_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener transacciones: {str(e)}'
        })


@login_required
@require_POST
@csrf_exempt
def anular_transaccion_pos(request):
    """Anular transacción POS"""
    try:
        data = json.loads(request.body)
        transaccion_id = data.get('transaccion_id')
        
        if not transaccion_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de transacción requerido'
            })
        
        transaccion = get_object_or_404(TransaccionPOS, id=transaccion_id)
        
        # Verificar que el usuario tiene acceso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if transaccion.configuracion_pos.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a esta transacción'
            })
        
        # Verificar que se puede anular
        if not transaccion.puede_anular:
            return JsonResponse({
                'success': False,
                'error': 'Esta transacción no puede ser anulada'
            })
        
        with transaction.atomic():
            # Crear nueva transacción de anulación
            anulacion = TransaccionPOS.objects.create(
                configuracion_pos=transaccion.configuracion_pos,
                ticket=transaccion.ticket,
                monto=transaccion.monto,
                tipo_transaccion='ANULACION',
                estado='INICIADA',
                usuario_operador=request.user,
                ip_origen=request.META.get('REMOTE_ADDR'),
                observaciones=f'Anulación de transacción {transaccion.ticket_pos}'
            )
            
            # Aquí iría la lógica real de anulación con el SDK de Transbank
            # Por ahora simulamos una anulación exitosa
            
            anulacion.estado = 'APROBADA'
            anulacion.codigo_respuesta = '00'
            anulacion.mensaje_respuesta = 'Anulación aprobada'
            anulacion.codigo_autorizacion = f'ANU-{transaccion.codigo_autorizacion}'
            anulacion.save()
            
            # Marcar transacción original como anulada
            transaccion.estado = 'ANULADA'
            transaccion.save()
            
            # Si había un detalle de pago asociado, eliminarlo o marcarlo como anulado
            if transaccion.detalle_pago:
                transaccion.detalle_pago.notas += f' - ANULADO {timezone.now().strftime("%d/%m/%Y %H:%M")}'
                transaccion.detalle_pago.save()
            
            # Crear logs
            LogPOS.objects.create(
                configuracion_pos=transaccion.configuracion_pos,
                transaccion_pos=anulacion,
                tipo_evento='COMANDO_ENVIADO',
                mensaje=f'Anulación exitosa de {transaccion.ticket_pos}',
                datos_tecnicos={
                    'transaccion_original': transaccion.ticket_pos,
                    'codigo_autorizacion_original': transaccion.codigo_autorizacion,
                    'monto': float(transaccion.monto)
                }
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Transacción anulada exitosamente',
            'anulacion': {
                'id': anulacion.id,
                'ticket_pos': anulacion.ticket_pos,
                'codigo_autorizacion': anulacion.codigo_autorizacion
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al anular transacción: {str(e)}'
        })


@login_required
@require_GET
def obtener_logs_pos(request, configuracion_id):
    """Obtener logs de una configuración POS específica"""
    try:
        configuracion = get_object_or_404(ConfiguracionPOS, id=configuracion_id)
        
        # Verificar que el usuario tiene acceso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if configuracion.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a esta configuración'
            })
        
        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_evento = request.GET.get('tipo_evento')
        limit = int(request.GET.get('limit', 100))
        
        # Construir queryset
        queryset = LogPOS.objects.filter(
            configuracion_pos=configuracion
        ).select_related('transaccion_pos')
        
        # Aplicar filtros
        if fecha_desde:
            queryset = queryset.filter(timestamp__date__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(timestamp__date__lte=fecha_hasta)
        if tipo_evento:
            queryset = queryset.filter(tipo_evento=tipo_evento)
        
        # Limitar resultados
        logs = queryset[:limit]
        
        # Serializar datos
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'timestamp': log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                'tipo_evento': log.tipo_evento,
                'tipo_evento_display': log.get_tipo_evento_display(),
                'mensaje': log.mensaje,
                'transaccion_pos': log.transaccion_pos.ticket_pos if log.transaccion_pos else '',
                'datos_tecnicos': log.datos_tecnicos,
            })
        
        return JsonResponse({
            'success': True,
            'logs': logs_data,
            'configuracion': {
                'id': configuracion.id,
                'nombre': configuracion.nombre,
                'tipo_pos': configuracion.get_tipo_pos_display()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener logs: {str(e)}'
        })


# ========== MÓDULO DE CAMBIOS Y DEVOLUCIONES ==========

@login_required
def gestion_cambios_devoluciones(request):
    """Vista principal para gestión de cambios y devoluciones"""
    sucursal_actual_id = request.session.get('sucursalActual') or request.session.get('idSucursalActual')
    sucursal_actual = None
    
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    if not sucursal_actual:
        return redirect('verHome')
    
    # ✅ Obtener tickets de cambio PENDIENTES de cobro/devolución
    tickets_cambio_pendientes = Ticket.objects.filter(
        sucursal=sucursal_actual,
        modulo_origen='CAMBIO_DEVOLUCION',
        estado='PENDIENTE'
    ).select_related('vendedor').order_by('-created_at')[:20]
    
    tickets_cambio_data = []
    for ticket in tickets_cambio_pendientes:
        # Determinar tipo de operación
        if 'A DEVOLVER AL CLIENTE' in (ticket.observaciones or ''):
            tipo_op = 'DEVOLUCION'
            icono = '💵'
            texto = 'Devolver'
            clase = 'success'
        elif 'A COBRAR AL CLIENTE' in (ticket.observaciones or ''):
            tipo_op = 'COBRO'
            icono = '💰'
            texto = 'Cobrar'
            clase = 'danger'
        else:
            tipo_op = 'DIRECTO'
            icono = '🔄'
            texto = 'Cambio'
            clase = 'info'
        
        tickets_cambio_data.append({
            'correlativo': ticket.correlativo,
            'cliente_nombre': ticket.cliente_nombre or 'Cliente General',
            'cliente_rut': ticket.cliente_rut or '',
            'total': int(ticket.total or 0),
            'tipo_operacion': tipo_op,
            'icono': icono,
            'texto': texto,
            'clase': clase,
            'vendedor': ticket.vendedor.nombre if ticket.vendedor else '-',
            'fecha': ticket.fecha.strftime('%d/%m/%Y') if ticket.fecha else '-',
            'hora': ticket.created_at.strftime('%H:%M') if ticket.created_at else '-',
            'observaciones': ticket.observaciones or '',
        })
    
    # Contar pendientes de revisión gerencial
    revision_pendiente_count = CambioDevolucion.objects.filter(
        sucursal=sucursal_actual,
        requiere_revision_gerencial=True,
        revisado_por_gerencia__isnull=True,
    ).count()

    context = {
        'sucursal_actual': sucursal_actual,
        'tipo_operacion_choices': TIPO_OPERACION_CAMBIO_CHOICES,
        'estado_choices': ESTADO_CAMBIO_CHOICES,
        'motivo_choices': MOTIVO_CAMBIO_CHOICES,
        'condicion_producto_choices': CONDICION_PRODUCTO_CHOICES,
        'metodo_pago_choices': METODO_PAGO_TICKET_CHOICES,
        'tickets_cambio_pendientes': tickets_cambio_data,
        'total_tickets_pendientes': len(tickets_cambio_data),
        'qz_config': _get_qz_config(sucursal_actual_id),
        'user_rol': getattr(request.user, 'rol', ''),
        'revision_pendiente_count': revision_pendiente_count,
    }
    return render(request, 'vistas/modulo_ventas/gestion_cambios_devoluciones.html', context)


@login_required
@require_GET
def listar_cambios_devoluciones(request):
    """API para listar cambios y devoluciones con filtros"""
    try:
        sucursal_id = get_sucursal_id(request)
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_operacion = request.GET.get('tipo_operacion')
        estado = request.GET.get('estado')
        buscar = request.GET.get('buscar', '').strip()
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        # Base filter (sin estado) — reutilizado para el aggregate y el filtrado
        base_qs = CambioDevolucion.objects.filter(sucursal_id=sucursal_id)

        # Aplicar filtros (fecha, tipo, búsqueda — sin estado todavía)
        if fecha_desde:
            base_qs = base_qs.filter(fecha_solicitud__date__gte=fecha_desde)
        if fecha_hasta:
            base_qs = base_qs.filter(fecha_solicitud__date__lte=fecha_hasta)
        if tipo_operacion:
            base_qs = base_qs.filter(tipo_operacion=tipo_operacion)
        if buscar:
            base_qs = base_qs.filter(
                Q(numero_operacion__icontains=buscar) |
                Q(ticket_original__correlativo__icontains=buscar) |
                Q(ticket_original__cliente_nombre__icontains=buscar) |
                Q(ticket_original__cliente_rut__icontains=buscar) |
                Q(observaciones_cliente__icontains=buscar) |
                Q(observaciones_vendedor__icontains=buscar)
            )

        # --- 1 sola query: todos los conteos de tabs + estadísticas adicionales ---
        agg = base_qs.aggregate(
            todos=Count('id', distinct=True),
            solicitados=Count('id', filter=Q(estado='SOLICITADO'), distinct=True),
            aprobados=Count('id', filter=Q(estado='APROBADO'), distinct=True),
            por_cobrar=Count(
                'id',
                filter=Q(estado__in=[
                    'EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE',
                ]),
                distinct=True,
            ),
            completados=Count('id', filter=Q(estado='COMPLETADO'), distinct=True),
            cancelados=Count(
                'id',
                filter=Q(estado__in=['CANCELADO', 'RECHAZADO', 'REVERTIDO']),
                distinct=True,
            ),
            total_diferencia=Sum('diferencia_monto'),
            cambios_fuera_plazo=Count('id', filter=Q(es_fuera_de_plazo=True), distinct=True),
            cambios_cross_branch=Count(
                'id', filter=Q(es_autorizacion_cross_branch=True), distinct=True
            ),
            cambios_revision_pendiente=Count(
                'id',
                filter=Q(requiere_revision_gerencial=True, revisado_por_gerencia__isnull=True),
                distinct=True,
            ),
        )

        conteos_tab = {
            'todos': agg['todos'] or 0,
            'solicitados': agg['solicitados'] or 0,
            'aprobados': agg['aprobados'] or 0,
            'por_cobrar': agg['por_cobrar'] or 0,
            'completados': agg['completados'] or 0,
            'cancelados': agg['cancelados'] or 0,
        }

        # --- Queryset filtrado por estado (tab activo) con prefetch ---
        queryset = base_qs.select_related(
            'ticket_original', 'ticket_nuevo', 'sucursal', 'solicitado_por', 'aprobado_por',
            'autorizado_por_usuario', 'revisado_por_gerencia', 'nota_credito',
        ).prefetch_related(
            'detalles__producto_original__ProductoTalla__producto',
            'detalles__producto_nuevo__producto',
            'pagos',
        )

        if estado:
            if estado == 'CANCELADO':
                queryset = queryset.filter(estado__in=['CANCELADO', 'RECHAZADO', 'REVERTIDO'])
            elif estado == 'EJECUTADO_COBRO_PENDIENTE':
                queryset = queryset.filter(
                    estado__in=['EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE']
                )
            else:
                queryset = queryset.filter(estado=estado)

        paginator = Paginator(queryset, per_page)
        cambios_page = paginator.get_page(page)

        cambios_data = []
        for cambio in cambios_page:
            acciones_permitidas = _acciones_cambio_para_usuario(request.user, cambio)
            detalles = list(cambio.detalles.all())
            total_productos_devueltos = sum(1 for d in detalles if d.producto_original_id)
            total_productos_nuevos = sum(1 for d in detalles if d.producto_nuevo_id)
            cant_devueltos = sum(d.cantidad_original for d in detalles if d.producto_original_id)
            cant_nuevos = sum(d.cantidad_nueva for d in detalles if d.producto_nuevo_id)

            productos_resumen = []
            for d in detalles[:3]:
                if d.producto_original_id:
                    try:
                        nombre = d.producto_original.ProductoTalla.producto.articulo
                    except Exception:
                        nombre = 'Producto'
                    productos_resumen.append({'nombre': nombre[:30], 'tipo': 'devuelto', 'cantidad': d.cantidad_original})
                if d.producto_nuevo_id:
                    try:
                        nombre = d.producto_nuevo.producto.articulo
                    except Exception:
                        nombre = 'Producto'
                    productos_resumen.append({'nombre': nombre[:30], 'tipo': 'nuevo', 'cantidad': d.cantidad_nueva})

            solicitante_nombre = ''
            if cambio.solicitado_por:
                solicitante_nombre = cambio.solicitado_por.get_full_name() or cambio.solicitado_por.username

            ticket_pendiente_pago = False
            if cambio.ticket_nuevo and cambio.ticket_nuevo.estado == 'PENDIENTE':
                ticket_pendiente_pago = True

            cambios_data.append({
                'id': cambio.id,
                'numero_operacion': cambio.numero_operacion,
                'fecha_solicitud': cambio.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
                'fecha_solicitud_iso': cambio.fecha_solicitud.isoformat(),
                'tipo_operacion': cambio.tipo_operacion,
                'tipo_operacion_display': cambio.get_tipo_operacion_display(),
                'estado': cambio.estado,
                'estado_display': cambio.get_estado_display(),
                'ticket_original': cambio.ticket_original.correlativo,
                'ticket_nuevo': {
                    'id': cambio.ticket_nuevo.id,
                    'correlativo': cambio.ticket_nuevo.correlativo,
                    'estado': cambio.ticket_nuevo.estado,
                    'estado_display': cambio.ticket_nuevo.get_estado_display(),
                    'metodo_pago': cambio.ticket_nuevo.metodo_pago,
                    'tipo_dte': cambio.ticket_nuevo.tipo_dte or ''
                } if cambio.ticket_nuevo else None,
                'ticket_pendiente_pago': ticket_pendiente_pago,
                'cliente_nombre': cambio.ticket_original.cliente_nombre or 'Sin nombre',
                'cliente_rut': cambio.ticket_original.cliente_rut or '',
                'monto_original': float(cambio.monto_original),
                'monto_nuevo': float(cambio.monto_nuevo),
                'diferencia_monto': float(cambio.diferencia_monto),
                'motivo_principal': cambio.get_motivo_principal_display(),
                'motivo_principal_codigo': cambio.motivo_principal,
                'solicitado_por': cambio.solicitado_por.username,
                'solicitado_por_nombre': solicitante_nombre,
                'aprobado_por': cambio.aprobado_por.username if cambio.aprobado_por else '',
                'fecha_aprobacion': cambio.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if cambio.fecha_aprobacion else '',
                'fecha_completado': cambio.fecha_completado.strftime('%d/%m/%Y %H:%M') if cambio.fecha_completado else '',
                'fecha_limite': cambio.fecha_limite_cambio.strftime('%d/%m/%Y'),
                'dias_desde_venta': cambio.dias_desde_venta,
                'dentro_del_plazo': cambio.dentro_del_plazo,
                'puede_completar': cambio.puede_completar,
                'requiere_pago_adicional': cambio.requiere_pago_adicional,
                'genera_devolucion': cambio.genera_devolucion,
                'total_productos_devueltos': total_productos_devueltos,
                'total_productos_nuevos': total_productos_nuevos,
                'cant_devueltos': cant_devueltos,
                'cant_nuevos': cant_nuevos,
                'productos_resumen': productos_resumen,
                'requiere_autorizacion': cambio.requiere_autorizacion,
                'autorizado_excepcion': cambio.autorizado_excepcion,
                'cobro_pendiente': cambio.cobro_pendiente,
                'devolucion_pendiente': cambio.devolucion_pendiente,
                'tiene_obligacion_pendiente': cambio.tiene_obligacion_pendiente,
                # Nuevos campos de trazabilidad
                'es_fuera_de_plazo': cambio.es_fuera_de_plazo,
                'dias_fuera_de_plazo': cambio.dias_fuera_de_plazo,
                'tipo_cambio_especial': cambio.tipo_cambio_especial,
                'es_autorizacion_cross_branch': cambio.es_autorizacion_cross_branch,
                'es_cambio_concepto': cambio.es_cambio_concepto,
                'autorizado_por_usuario': cambio.autorizado_por_usuario.get_full_name() if cambio.autorizado_por_usuario else None,
                'excepcion_plazo_ya_autorizada': _autorizacion_fuera_plazo_previa(cambio) is not None,
                'score_riesgo': cambio.score_riesgo,
                'requiere_revision_gerencial': cambio.requiere_revision_gerencial,
                'revisado_por_gerencia': cambio.revisado_por_gerencia.get_full_name() if cambio.revisado_por_gerencia else None,
                # Nota de Crédito
                'nc_generada': cambio.nc_generada,
                'metodo_devolucion': cambio.metodo_devolucion,
                'metodo_devolucion_display': cambio.get_metodo_devolucion_display() if cambio.metodo_devolucion != 'SIN_NC' else '',
                'nota_credito_numero': cambio.nota_credito.numero_documento if cambio.nota_credito_id else None,
                'acciones_permitidas': acciones_permitidas,
            })

        return JsonResponse({
            'success': True,
            'cambios': cambios_data,
            'conteos_tab': conteos_tab,
            'pagination': {
                'current_page': cambios_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': cambios_page.has_next(),
                'has_previous': cambios_page.has_previous(),
            },
            'estadisticas': {
                'total_cambios': conteos_tab['todos'],
                'cambios_pendientes': conteos_tab['solicitados'],
                'cambios_aprobados': conteos_tab['aprobados'],
                'cambios_por_cobrar': conteos_tab['por_cobrar'],
                'cambios_completados': conteos_tab['completados'],
                'total_diferencia': float(agg['total_diferencia'] or 0),
                'cambios_fuera_plazo': agg['cambios_fuera_plazo'] or 0,
                'cambios_cross_branch': agg['cambios_cross_branch'] or 0,
                'cambios_revision_pendiente': agg['cambios_revision_pendiente'] or 0,
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener cambios: {str(e)}'
        })


# Estados en los que un cambio ya "consumió" unidades de la venta original.
ESTADOS_CAMBIO_VIGENTES = [
    'SOLICITADO', 'APROBADO', 'EJECUTADO',
    'EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE', 'COMPLETADO',
]

# Estados en los que el cambio ya generó su ticket de reemplazo.
ESTADOS_CAMBIO_CON_TICKET_NUEVO = [
    'EJECUTADO', 'EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE', 'COMPLETADO',
]


def _ticket_raiz_cambio(ticket):
    """Sube por la cadena de cambios hasta el ticket de la venta original.

    El `ticket_nuevo` de un cambio es un comprobante del delta (línea negativa por
    lo devuelto + positiva por lo entregado), no un reemplazo de la venta.
    """
    actual = ticket
    visitados = set()

    while actual and actual.id not in visitados:
        visitados.add(actual.id)
        cambio = CambioDevolucion.objects.filter(
            ticket_nuevo=actual
        ).select_related('ticket_original').order_by('fecha_solicitud').first()

        if cambio and cambio.ticket_original and cambio.ticket_original_id not in visitados:
            actual = cambio.ticket_original
        else:
            break

    return actual


def _cadena_cambios_ticket(ticket_raiz):
    """Devuelve (tickets, cambios) de toda la cadena que cuelga del ticket raíz.

    `tickets` trae la venta original más los tickets de reemplazo ya generados;
    `cambios` trae todos los cambios de la cadena, en cualquier estado.
    """
    tickets = {ticket_raiz.id: ticket_raiz}
    cambios = {}
    pendientes = [ticket_raiz.id]

    while pendientes:
        cambios_qs = CambioDevolucion.objects.filter(
            ticket_original_id__in=pendientes
        ).select_related('ticket_nuevo').prefetch_related(
            'detalles__producto_original__ProductoTalla__producto',
            'detalles__producto_nuevo__producto',
        )
        pendientes = []

        for cambio in cambios_qs:
            if cambio.id in cambios:
                continue
            cambios[cambio.id] = cambio

            nuevo = cambio.ticket_nuevo
            if (nuevo and nuevo.id not in tickets
                    and cambio.estado in ESTADOS_CAMBIO_CON_TICKET_NUEVO):
                tickets[nuevo.id] = nuevo
                pendientes.append(nuevo.id)

    cambios_ordenados = sorted(cambios.values(), key=lambda c: c.fecha_solicitud)
    return list(tickets.values()), cambios_ordenados


def _describir_producto_talla(producto_talla):
    """Texto corto de un Producto_Talla para mostrarlo en el historial."""
    if not producto_talla:
        return ''
    articulo = producto_talla.producto.articulo if producto_talla.producto else ''
    talla = producto_talla.talla or ''
    return f'{articulo} T{talla}'.strip() if talla else articulo


def _productos_cambio_data(tickets, cambios):
    """Productos cambiables de toda la cadena, con su procedencia.

    Incluye las líneas de la venta original y también los artículos que entraron
    como reemplazo en cambios anteriores, para que se puedan volver a cambiar.
    """
    cambio_por_ticket_nuevo = {c.ticket_nuevo_id: c for c in cambios if c.ticket_nuevo_id}

    lineas_por_ticket = {}
    ids_lineas = []
    for ticket in tickets:
        lineas = list(ticket.ticket_productos.filter(precio__gt=0, stock__gt=0)
                      .select_related('ProductoTalla__producto'))
        lineas_por_ticket[ticket.id] = lineas
        ids_lineas.extend(linea.id for linea in lineas)

    # Una sola consulta para saber qué se cambió de cada línea (antes era N+1).
    cantidad_cambiada = {}
    cambios_por_linea = {}
    if ids_lineas:
        detalles = CambioDevolucionDetalle.objects.filter(
            producto_original_id__in=ids_lineas,
            cambio_devolucion__estado__in=ESTADOS_CAMBIO_VIGENTES,
        ).select_related('cambio_devolucion', 'producto_nuevo__producto')

        for detalle in detalles:
            linea_id = detalle.producto_original_id
            cantidad_cambiada[linea_id] = cantidad_cambiada.get(linea_id, 0) + (detalle.cantidad_original or 0)
            cambio = detalle.cambio_devolucion
            cambios_por_linea.setdefault(linea_id, []).append({
                'numero_operacion': cambio.numero_operacion,
                'estado': cambio.get_estado_display(),
                'fecha': timezone.localtime(cambio.fecha_solicitud).strftime('%d/%m/%Y'),
                'cantidad': detalle.cantidad_original or 0,
                'reemplazo': _describir_producto_talla(detalle.producto_nuevo),
            })

    productos_data = []
    disponibles_count = 0

    for ticket in tickets:
        cambio_origen = cambio_por_ticket_nuevo.get(ticket.id)

        for linea in lineas_por_ticket[ticket.id]:
            if linea.ProductoTalla is None:
                continue  # ítems sin SKU (pendientes de despacho) no se cambian

            ya_cambiada = cantidad_cambiada.get(linea.id, 0)
            disponible = max(0, linea.stock - ya_cambiada)
            if disponible > 0:
                disponibles_count += 1

            descuento = linea.descuento_unitario or 0
            precio_pagado = linea.precio - descuento
            producto = linea.ProductoTalla.producto

            productos_data.append({
                'id': linea.id,
                'sku': linea.ProductoTalla.sku,
                'articulo': producto.articulo if producto else (linea.descripcion_linea or ''),
                'descripcion': producto.descripcion if producto else (linea.descripcion_linea or ''),
                'talla': linea.ProductoTalla.talla,
                'cantidad_original': linea.stock,
                'cantidad_ya_cambiada': ya_cambiada,
                'cantidad_disponible': disponible,
                'precio_unitario': float(precio_pagado),
                'precio_lista': float(linea.precio),
                'descuento_unitario': float(descuento),
                'tiene_descuento': descuento > 0,
                'subtotal': float(precio_pagado * linea.stock),
                'ya_cambiado': disponible == 0,
                # Procedencia: distingue lo vendido de lo que entró por un cambio previo
                'origen': 'CAMBIO' if cambio_origen else 'VENTA',
                'es_reemplazo': bool(cambio_origen),
                'origen_cambio': cambio_origen.numero_operacion if cambio_origen else None,
                'origen_cambio_fecha': (
                    timezone.localtime(cambio_origen.fecha_solicitud).strftime('%d/%m/%Y')
                    if cambio_origen else None
                ),
                'ticket_id': ticket.id,
                'ticket_correlativo': ticket.correlativo,
                'cambios_linea': cambios_por_linea.get(linea.id, []),
            })

    return productos_data, disponibles_count


def _historial_cambios_data(cambios):
    """Historial de la cadena, con el detalle de qué se cambió por qué."""
    historial = []

    for cambio in cambios:
        detalles = []
        for detalle in cambio.detalles.all():
            origen = detalle.producto_original.ProductoTalla if detalle.producto_original_id else None
            detalles.append({
                'de': _describir_producto_talla(origen),
                'de_sku': origen.sku if origen else '',
                'a': _describir_producto_talla(detalle.producto_nuevo),
                'a_sku': detalle.producto_nuevo.sku if detalle.producto_nuevo_id else '',
                'cantidad': detalle.cantidad_original or 0,
            })

        historial.append({
            'numero_operacion': cambio.numero_operacion,
            'tipo_operacion': cambio.get_tipo_operacion_display(),
            'estado': cambio.get_estado_display(),
            'estado_codigo': cambio.estado,
            'fecha_solicitud': timezone.localtime(cambio.fecha_solicitud).strftime('%d/%m/%Y'),
            'diferencia_monto': float(cambio.diferencia_monto or 0),
            'ticket_nuevo': cambio.ticket_nuevo.correlativo if cambio.ticket_nuevo_id else None,
            'detalles': detalles,
        })

    return historial


@login_required
@require_POST
def crear_cambio_devolucion(request):
    """Crear nueva solicitud de cambio o devolución"""
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        documento_id = data.get('documento_id')
        documento_tipo = data.get('documento_tipo', 'TICKET')
        documento_numero = data.get('documento_numero')
        tipo_operacion = data.get('tipo_operacion')
        motivo_principal = data.get('motivo_principal')
        productos_cambio = data.get('productos', [])
        
        # Retrocompatibilidad con ticket_correlativo
        if not documento_numero and data.get('ticket_correlativo'):
            documento_numero = data.get('ticket_correlativo')
            documento_tipo = 'TICKET'
        
        if not all([documento_numero, tipo_operacion, motivo_principal]):
            return JsonResponse({
                'success': False,
                'error': f'Faltan datos obligatorios. documento_numero: {documento_numero}, tipo_operacion: {tipo_operacion}, motivo: {motivo_principal}'
            })
        
        if not productos_cambio:
            return JsonResponse({
                'success': False,
                'error': 'Debe incluir al menos un producto'
            })

        # En este módulo NO se devuelve dinero: toda operación tiene que entregar
        # algo a cambio. Sin producto de salida la operación nace debiéndole plata
        # al cliente, queda en "Devolución Pendiente" y además bloquea cualquier
        # cambio posterior sobre esa venta (así quedó atascada CD-7-202608-0014).
        # Las devoluciones de dinero van por Devolución Garantía.
        if tipo_operacion not in ('CAMBIO_CONCEPTO', 'DEVOLUCION_CONCEPTO'):
            if not any(item.get('producto_nuevo_id') for item in productos_cambio):
                return JsonResponse({
                    'success': False,
                    'code': 'SIN_PRODUCTO_SALIDA',
                    'error': 'Falta registrar el producto que se lleva el cliente. '
                             'En este módulo no se devuelve dinero: si el cliente no '
                             'se lleva nada, la operación va por Devolución Garantía.',
                }, status=400)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Buscar documento original (Ticket o DTE)
        ticket_original = None
        dte_original = None
        
        if documento_tipo == 'DTE':
            # Buscar DTE y crear ticket asociado si no existe
            try:
                # El folio NO es único entre empresas del holding, así que un
                # `.get()` por número revienta con MultipleObjectsReturned.
                dte_original = Dte.objects.select_related('receptor', 'vendedor', 'sucursal').filter(
                    numero_documento=documento_numero,
                    sucursal=sucursal,
                    tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
                ).order_by('-fecha_emision', '-id').first()

                if not dte_original:
                    raise Dte.DoesNotExist


                # Primero: intentar encontrar el ticket ORIGINAL del POS (tiene descuentos correctos)
                if dte_original.referencias and 'TICKET-' in dte_original.referencias:
                    try:
                        corr_orig = dte_original.referencias.split('TICKET-')[1].strip().split()[0]
                        ticket_original = Ticket.objects.filter(
                            correlativo=corr_orig,
                            sucursal_id=sucursal_id,
                            estado='PAGADO'
                        ).first()
                    except Exception:
                        pass
                
                # Si no se encontró el original, buscar ticket de referencia existente
                if not ticket_original:
                    ticket_original = Ticket.objects.filter(
                        observaciones__icontains=f'para DTE #{documento_numero} -'
                    ).first()
                
                if not ticket_original:
                    ticket_original = Ticket.objects.filter(
                        observaciones__icontains=f'DTE #{documento_numero} -'
                    ).first()
                
                if not ticket_original:
                    # Crear ticket de referencia desde el DTE
                    
                    correlativo_ticket = obtener_siguiente_correlativo(sucursal, 'TICKET')
                    
                    # Crear ticket de referencia con solo los campos que existen en el modelo
                    ticket_original = Ticket.objects.create(
                        correlativo=correlativo_ticket,
                        vendedor=dte_original.vendedor,
                        sucursal=sucursal,
                        subTotal=int(dte_original.monto_neto),
                        descuento=int(dte_original.descuento) if dte_original.descuento else 0,
                        total=int(dte_original.monto_con_iva),
                        estado='PAGADO',
                        responsable=dte_original.responsable,
                        cliente_nombre=dte_original.receptor.razon_social if dte_original.receptor else '',
                        cliente_rut=dte_original.receptor.rut if dte_original.receptor else '',
                        cliente_email=dte_original.receptor.correoVendedor if dte_original.receptor else '',
                        cliente_telefono='',
                        cliente_giro=dte_original.receptor.giro if dte_original.receptor else '',
                        cliente_direccion=dte_original.receptor.direccion if dte_original.receptor else '',
                        cliente_comuna=dte_original.receptor.comuna if dte_original.receptor else '',
                        cliente_ciudad=dte_original.receptor.ciudad if dte_original.receptor else '',
                        observaciones=f'Ticket de referencia para DTE #{documento_numero} - {dte_original.tipo_documento}'
                    )
                    
                    # Copiar productos del DTE al ticket y crear mapeo (con descuentos)
                    mapeo_productos = {}  # dte_producto_id → ticket_producto_id
                    es_boleta_ref = dte_original.tipo_documento in ['39', '41', 'BOLETA ELECTRONICA', 'BOLETA EXENTA']
                    
                    for dp in dte_original.dte_productos.all():
                        dcto_u = 0
                        if dp.descuento_monto and dp.stock and dp.stock > 0:
                            dcto_u = int(dp.descuento_monto / dp.stock)
                        elif es_boleta_ref and dp.monto_item and dp.stock and dp.stock > 0:
                            precio_ef = int(dp.monto_item / dp.stock)
                            if precio_ef < dp.precio:
                                dcto_u = dp.precio - precio_ef
                        sub = (dp.precio - dcto_u) * dp.stock

                        tp = Ticket_Productos.objects.create(
                            idTicket=ticket_original,
                            ProductoTalla=dp.productoTalla,
                            stock=dp.stock,
                            precio=dp.precio,
                            precio_original=dp.precio,
                            descuento_unitario=dcto_u,
                            subtotal=sub,
                            porcentaje_descuento=dp.descuento_pct or 0,
                            descripcion_linea=dp.descripcion if not dp.productoTalla else None,
                            es_pendiente_despacho=dp.es_pendiente_despacho,
                        )
                        mapeo_productos[dp.id] = tp.id
                    
                    # Guardar mapeo en la sesión o en el ticket para referencia
                    ticket_original.observaciones += f"\n[MAPEO_PRODUCTOS: {mapeo_productos}]"
                    ticket_original.save()

                    
            except Dte.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'DTE #{documento_numero} no encontrado'
                })
        else:
            # Buscar Ticket
            try:
                ticket_original = Ticket.objects.get(
                    correlativo=documento_numero,
                    sucursal=sucursal
                )
            except Ticket.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'Ticket #{documento_numero} no encontrado'
                })
            
            # Verificar que el ticket esté pagado
            if ticket_original.estado != 'PAGADO':
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se pueden procesar cambios de tickets pagados'
                })
        
        # Un cambio anterior NO reemplaza la venta: se trabaja siempre sobre el
        # ticket original y se aceptan además las líneas de reemplazo que dejaron
        # los cambios ya ejecutados (para poder cambiar de nuevo ese artículo).
        ticket_original = _ticket_raiz_cambio(ticket_original)
        tickets_cadena, _cambios_cadena = _cadena_cambios_ticket(ticket_original)
        ids_tickets_cadena = [t.id for t in tickets_cadena]

        # Verificar plazo (30 días por defecto)
        from datetime import timedelta
        fecha_base_plazo = None

        # Caso 1: el backend recibió el DTE directamente
        if dte_original:
            fecha_base_plazo = dte_original.fecha_emision

        # Caso 2: el frontend indica que el documento original era un DTE
        if not fecha_base_plazo:
            dte_numero_original = data.get('dte_numero_original')
            if dte_numero_original:
                try:
                    dte_query = Dte.objects.filter(numero_documento=dte_numero_original)
                    if sucursal:
                        dte_query = dte_query.filter(sucursal=sucursal)
                    dte_ref = dte_query.order_by('-fecha_emision').first()
                    if dte_ref:
                        fecha_base_plazo = dte_ref.fecha_emision
                except Exception:
                    pass

        # Caso 3: ticket puro → usar fecha del ticket
        if not fecha_base_plazo:
            fecha_base_plazo = ticket_original.fecha

        fecha_limite = fecha_base_plazo + timedelta(days=30)
        fuera_de_plazo = timezone.localdate() > fecha_limite
        
        # Los cambios fuera de plazo usan exclusivamente el código dinámico de la navbar.
        codigo_autorizacion = str(
            data.get('codigo_autorizacion') or data.get('supervisor_pin') or ''
        ).strip()
        supervisor_autorizo = False
        supervisor = None
        dias_fuera = 0
        codigo_dinamico_obj = None
        sucursal_supervisor = None

        if fuera_de_plazo:
            dias_fuera = (timezone.localdate() - fecha_limite).days

            if not codigo_autorizacion:
                return JsonResponse({
                    'success': False,
                    'code': 'AUTH_CODE_REQUIRED',
                    'error': f'El plazo para cambios venció el {fecha_limite.strftime("%d/%m/%Y")}',
                    'requiere_autorizacion': True,
                    'fecha_limite': fecha_limite.strftime('%d/%m/%Y'),
                    'fecha_compra': fecha_base_plazo.strftime('%d/%m/%Y'),
                    'dias_transcurridos': (timezone.localdate() - fecha_base_plazo).days,
                    'dias_fuera_de_plazo': dias_fuera,
                })

            es_valido_cod, mensaje_codigo, codigo_dinamico_obj = \
                CodigoAutorizacionDinamico.validar_codigo(codigo_autorizacion)
            supervisor = codigo_dinamico_obj.generado_por if codigo_dinamico_obj else None
            if not es_valido_cod or not _usuario_es_administrador_activo(supervisor):
                return JsonResponse({
                    'success': False,
                    'code': 'INVALID_AUTH_CODE',
                    'error': mensaje_codigo if not es_valido_cod else 'El código no pertenece a un administrador activo',
                    'requiere_autorizacion': True,
                }, status=403)

            asignacion_supervisor = EmpresaUser.objects.filter(
                user=supervisor,
                empresa_id=sucursal.empresa_id,
                status=True,
            ).select_related('sucursal').order_by('-active').first()
            if not asignacion_supervisor:
                return JsonResponse({
                    'success': False,
                    'code': 'CROSS_COMPANY_AUTH',
                    'error': 'El administrador debe pertenecer a la misma empresa',
                    'requiere_autorizacion': True,
                }, status=403)
            sucursal_supervisor = asignacion_supervisor.sucursal

            supervisor_autorizo = True
        
        # Validar que no existan cambios con obligaciones financieras pendientes para este ticket
        cambios_con_pago_pendiente = CambioDevolucion.objects.filter(
            ticket_original_id__in=ids_tickets_cadena,
            estado__in=['EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE']
        ).first()

        if not cambios_con_pago_pendiente:
            # También verificar el patrón legacy: COMPLETADO pero ticket_nuevo PENDIENTE
            cambios_con_pago_pendiente = CambioDevolucion.objects.filter(
                ticket_original_id__in=ids_tickets_cadena,
                estado='COMPLETADO',
                ticket_nuevo__estado='PENDIENTE'
            ).first()
        
        if cambios_con_pago_pendiente:
            monto_pend = abs(int(cambios_con_pago_pendiente.diferencia_monto or 0))
            return JsonResponse({
                'success': False,
                'error': f'El cambio anterior {cambios_con_pago_pendiente.numero_operacion} tiene una diferencia de '
                         f'${monto_pend:,} que NO se cobró. Debe cobrarla o condonarla (administrador) '
                         f'antes de crear un nuevo cambio sobre este documento.'
            })

        # Aviso (permitir + confirmar): el ticket tuvo un cambio previo con la diferencia
        # condonada. El admin ya perdonó ese cobro, así que dejamos re-cambiar pero avisando.
        if not data.get('aceptar_recambio_condonado'):
            cambio_condonado = CambioDevolucion.objects.filter(
                ticket_original_id__in=ids_tickets_cadena,
                diferencia_condonada=True,
                estado='COMPLETADO'
            ).order_by('-fecha_condonacion').first()
            if cambio_condonado:
                monto_cond = abs(int(cambio_condonado.diferencia_monto or 0))
                admin_nombre = getattr(cambio_condonado.condonada_por, 'username', None) or 'un administrador'
                fecha_cond = (timezone.localtime(cambio_condonado.fecha_condonacion).strftime('%d/%m/%Y')
                              if cambio_condonado.fecha_condonacion else '')
                mensaje = (f'El cambio anterior {cambio_condonado.numero_operacion} tuvo una diferencia de '
                           f'${monto_cond:,} que NO se cobró (condonada por {admin_nombre}'
                           + (f' el {fecha_cond}' if fecha_cond else '') + '). ¿Desea continuar igual?')
                return JsonResponse({
                    'success': False,
                    'requiere_confirmacion_condonacion': True,
                    'mensaje_condonacion': mensaje,
                })

        with transaction.atomic():
            if codigo_dinamico_obj:
                codigo_dinamico_obj = CodigoAutorizacionDinamico.objects.select_for_update().get(
                    id=codigo_dinamico_obj.id
                )
                if not codigo_dinamico_obj.es_valido():
                    return JsonResponse({
                        'success': False,
                        'code': 'AUTH_CODE_ALREADY_USED',
                        'error': 'El código fue utilizado o venció antes de completar la solicitud',
                        'requiere_autorizacion': True,
                    }, status=409)

            # Cambios por concepto: monto viene directamente del frontend
            es_concepto = tipo_operacion in ('CAMBIO_CONCEPTO', 'DEVOLUCION_CONCEPTO')
            if es_concepto:
                monto_original_calculado = int(data.get('concepto_monto_original', 0))
            else:
                # Calcular monto_original basado en el precio efectivo (con descuento aplicado)
                monto_original_calculado = 0
            for item in productos_cambio:
                if es_concepto:
                    break  # No iterar productos para cambios por concepto
                try:
                    ticket_producto = Ticket_Productos.objects.get(
                        idTicket_id__in=ids_tickets_cadena,
                        id=item['ticket_producto_id']
                    )
                    cantidad_cambio = item.get('cantidad', 0)
                    precio_efectivo = ticket_producto.precio - (ticket_producto.descuento_unitario or 0)
                    monto_original_calculado += precio_efectivo * cantidad_cambio
                except Ticket_Productos.DoesNotExist:
                    pass
            
            # Crear cambio/devolución con el monto correcto
            obs_vendedor = data.get('observaciones_vendedor', '')
            if supervisor_autorizo:
                obs_vendedor = f'[AUTORIZADO FUERA DE PLAZO por {supervisor.get_full_name() or supervisor.username}] {obs_vendedor}'.strip()

            # Determinar tipo especial y cross-branch
            tipo_especial = 'NORMAL'
            es_cross_branch = False
            if fuera_de_plazo:
                tipo_especial = 'FUERA_PLAZO'

            if tipo_operacion in ('CAMBIO_CONCEPTO', 'DEVOLUCION_CONCEPTO'):
                tipo_especial = 'CONCEPTO'

            if supervisor:
                es_cross_branch = bool(
                    sucursal_supervisor and sucursal_supervisor.id != sucursal.id
                )

            # Crear registro de autorización con trazabilidad completa
            registro_auth = None
            if supervisor_autorizo:
                from .models import RegistroAutorizacion
                metodo_auth = 'código de autorización del navbar'
                registro_auth = RegistroAutorizacion.objects.create(
                    codigo_usado=codigo_dinamico_obj,
                    usuario_solicitante=request.user,
                    usuario_autorizador=supervisor,
                    tipo_operacion='APROBACION_CAMBIO',
                    descripcion=f'Autorización fuera de plazo ({dias_fuera} días) vía {metodo_auth} por {supervisor.get_full_name() or supervisor.username}',
                    ip_origen=request.META.get('REMOTE_ADDR'),
                    exitoso=True,
                    sucursal_solicitante=sucursal,
                    sucursal_autorizador=sucursal_supervisor,
                    es_cross_branch=es_cross_branch,
                    requiere_revision=es_cross_branch or dias_fuera > 15,
                    datos_adicionales={
                        'dias_fuera_de_plazo': dias_fuera,
                        'fecha_limite': fecha_limite.strftime('%Y-%m-%d'),
                        'fecha_compra': fecha_base_plazo.strftime('%Y-%m-%d'),
                        'supervisor_username': supervisor.username,
                        'supervisor_sucursal': str(sucursal_supervisor) if sucursal_supervisor else None,
                        'metodo_autorizacion': metodo_auth,
                    }
                )
                # Marcar el código dinámico como usado (único uso) una vez registrada la autorización
                if codigo_dinamico_obj:
                    codigo_dinamico_obj.marcar_como_usado()

            # Determinar si requiere revisión gerencial (auto-escalamiento)
            requiere_revision = (
                fuera_de_plazo or
                es_cross_branch or
                monto_original_calculado > 200000  # Umbral configurable
            )

            cambio = CambioDevolucion.objects.create(
                ticket_original=ticket_original,
                sucursal=sucursal,
                tipo_operacion=tipo_operacion,
                monto_original=monto_original_calculado,
                motivo_principal=motivo_principal,
                observaciones_cliente=data.get('observaciones_cliente', ''),
                observaciones_vendedor=obs_vendedor,
                solicitado_por=request.user,
                requiere_autorizacion=True if supervisor_autorizo else data.get('requiere_autorizacion', False),
                fecha_limite_cambio=fecha_limite,
                # Nuevos campos de trazabilidad
                autorizado_por_usuario=supervisor if supervisor_autorizo else None,
                sucursal_autorizador=sucursal_supervisor if supervisor_autorizo else None,
                es_autorizacion_cross_branch=es_cross_branch,
                es_fuera_de_plazo=fuera_de_plazo,
                dias_fuera_de_plazo=dias_fuera if fuera_de_plazo else 0,
                tipo_cambio_especial=tipo_especial,
                registro_autorizacion=registro_auth,
                es_cambio_concepto=tipo_operacion in ('CAMBIO_CONCEPTO', 'DEVOLUCION_CONCEPTO'),
                concepto_descripcion=data.get('concepto_descripcion', ''),
                concepto_monto_original=data.get('concepto_monto_original'),
                documento_referencia_legacy=data.get('documento_referencia_legacy', ''),
                requiere_revision_gerencial=requiere_revision,
            )

            # Vincular registro de autorización al cambio
            if registro_auth:
                registro_auth.cambio_devolucion = cambio
                registro_auth.save(update_fields=['cambio_devolucion'])
            
            # Procesar productos
            monto_nuevo_total = 0
            monto_original_real = 0  # Recalcular para asegurar consistencia
            productos_procesados = set()  # Para evitar duplicar devoluciones
            
            for item in productos_cambio:
                # Buscar producto original en el ticket
                try:
                    ticket_producto = Ticket_Productos.objects.get(
                        idTicket_id__in=ids_tickets_cadena,
                        id=item['ticket_producto_id']
                    )
                except Ticket_Productos.DoesNotExist:
                    raise ValidationError(f'Producto no encontrado en el ticket')
                
                # Validar cantidad
                cantidad_cambio = item.get('cantidad', 0)
                
                # ✅ CORREGIDO: Detectar si es un producto nuevo ADICIONAL
                es_producto_adicional = item.get('es_producto_adicional', False) or (cantidad_cambio == 0 and item.get('producto_nuevo_id'))
                
                # Solo validar y contar devolución si NO es producto adicional
                if not es_producto_adicional:
                    if cantidad_cambio <= 0 or cantidad_cambio > ticket_producto.stock:
                        raise ValidationError(f'Cantidad inválida para {ticket_producto.ProductoTalla.producto.articulo}')
                    
                    # Solo sumar al monto original si no hemos procesado este producto ya
                    producto_key = f"{ticket_producto.id}_{cantidad_cambio}"
                    if producto_key not in productos_procesados:
                        precio_efectivo = ticket_producto.precio - (ticket_producto.descuento_unitario or 0)
                        monto_original_real += precio_efectivo * cantidad_cambio
                        productos_procesados.add(producto_key)
                
                # Producto nuevo (si es cambio)
                producto_nuevo = None
                precio_nuevo = 0
                cantidad_nueva = 0
                
                if tipo_operacion in ['CAMBIO_SIMPLE', 'CAMBIO_CON_DIFERENCIA']:
                    producto_nuevo_id = item.get('producto_nuevo_id')
                    if producto_nuevo_id:
                        try:
                            producto_nuevo = Producto_Talla.objects.get(id=producto_nuevo_id)
                            precio_catalogo = producto_nuevo.producto.precioventa
                            
                            # ✅ Usar el precio enviado desde el frontend si es mayor o igual al precio catálogo
                            precio_enviado = item.get('precio_nuevo', 0)
                            if precio_enviado and precio_enviado >= precio_catalogo:
                                precio_nuevo = precio_enviado
                            else:
                                precio_nuevo = precio_catalogo
                            
                            # Usar la cantidad enviada o 1 por defecto
                            cantidad_nueva = item.get('cantidad_nueva', 1) or 1
                            monto_nuevo_total += precio_nuevo * cantidad_nueva
                        except Producto_Talla.DoesNotExist:
                            raise ValidationError(f'Producto nuevo no encontrado')
                
                # ✅ Crear detalle del cambio
                # Separar la creación según sea devolución, cambio o producto adicional
                
                if es_producto_adicional:
                    # Producto ADICIONAL: solo producto nuevo, sin devolución asociada
                    if producto_nuevo:
                        detalle = CambioDevolucionDetalle.objects.create(
                            cambio_devolucion=cambio,
                            producto_original=None,  # No hay producto original
                            cantidad_original=0,
                            producto_nuevo=producto_nuevo,
                            cantidad_nueva=cantidad_nueva,
                            precio_nuevo=precio_nuevo,
                            precio_original_unitario=0,
                            condicion_producto='PERFECTO',
                            apto_para_venta=True,
                            observaciones=item.get('observaciones', '') + ' [PRODUCTO ADICIONAL]'
                        )
                else:
                    # Producto con DEVOLUCIÓN (puede o no tener producto nuevo asociado)
                    if cantidad_cambio > 0:
                        precio_efectivo_unitario = ticket_producto.precio - (ticket_producto.descuento_unitario or 0)
                        detalle = CambioDevolucionDetalle.objects.create(
                            cambio_devolucion=cambio,
                            producto_original=ticket_producto,
                            cantidad_original=cantidad_cambio,
                            producto_nuevo=producto_nuevo,
                            cantidad_nueva=cantidad_nueva,
                            precio_nuevo=precio_nuevo,
                            precio_original_unitario=precio_efectivo_unitario,
                            condicion_producto=item.get('condicion_producto', 'PERFECTO'),
                            apto_para_venta=item.get('apto_para_venta', True),
                            observaciones=item.get('observaciones', '')
                        )
            
            # ✅ Usar el monto original recalculado para mayor precisión
            if monto_original_real > 0:
                cambio.monto_original = monto_original_real
            
            cambio.monto_nuevo = monto_nuevo_total
            cambio.diferencia_monto = monto_nuevo_total - float(cambio.monto_original)
            
            # VALIDACIÓN: En CAMBIOS no se permite diferencia a favor del cliente
            if tipo_operacion in ['CAMBIO_SIMPLE', 'CAMBIO_CON_DIFERENCIA']:
                if cambio.diferencia_monto < 0:
                    raise ValidationError(
                        f'En un CAMBIO no se permite que los productos nuevos tengan menor valor que los devueltos. '
                        f'Diferencia: ${abs(cambio.diferencia_monto):,.0f} a favor del cliente. '
                        f'Para devolver dinero al cliente, use DEVOLUCIÓN en lugar de CAMBIO.'
                    )
            
            cambio.save()
            
            # Crear historial
            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion='CREADO',
                estado_nuevo='SOLICITADO',
                usuario=request.user,
                descripcion=f'Solicitud de {cambio.get_tipo_operacion_display().lower()} creada',
                datos_adicionales={
                    'motivo': motivo_principal,
                    'productos_count': len(productos_cambio),
                    'monto_diferencia': float(cambio.diferencia_monto)
                }
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Solicitud creada exitosamente',
            'cambio_id': cambio.id,
            'numero_operacion': cambio.numero_operacion,
            'diferencia_monto': float(cambio.diferencia_monto)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear solicitud: {str(e)}'
        })


@login_required
@require_GET
def obtener_detalle_cambio(request, cambio_id):
    """Obtener detalle completo de un cambio/devolución"""
    try:
        cambio = get_object_or_404(
            CambioDevolucion.objects.select_related(
                'ticket_original', 'ticket_nuevo', 'sucursal', 
                'solicitado_por', 'aprobado_por', 'nota_credito'
            ).prefetch_related(
                'detalles__producto_original__ProductoTalla__producto',
                'detalles__producto_nuevo__producto',
                'pagos',
                'historial__usuario'
            ),
            id=cambio_id
        )
        
        # Verificar acceso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'code': 'BRANCH_REQUIRED',
                'error': 'No hay una sucursal seleccionada'
            }, status=400)
        if cambio.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a este cambio'
            })
        
        # Datos del cambio
        acciones_permitidas = _acciones_cambio_para_usuario(request.user, cambio)
        cambio_data = {
            'id': cambio.id,
            'numero_operacion': cambio.numero_operacion,
            'tipo_operacion': cambio.tipo_operacion,
            'tipo_operacion_display': cambio.get_tipo_operacion_display(),
            'estado': cambio.estado,
            'estado_display': cambio.get_estado_display(),
            'fecha_solicitud': cambio.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'fecha_aprobacion': cambio.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if cambio.fecha_aprobacion else '',
            'fecha_completado': cambio.fecha_completado.strftime('%d/%m/%Y %H:%M') if cambio.fecha_completado else '',
            'fecha_limite_cambio': cambio.fecha_limite_cambio.strftime('%d/%m/%Y'),
            'dias_desde_venta': cambio.dias_desde_venta,
            'dentro_del_plazo': cambio.dentro_del_plazo,
            'puede_completar': cambio.puede_completar,
            
            # Montos
            'monto_original': float(cambio.monto_original),
            'monto_nuevo': float(cambio.monto_nuevo),
            'diferencia_monto': float(cambio.diferencia_monto),
            'requiere_pago_adicional': cambio.requiere_pago_adicional,
            'genera_devolucion': cambio.genera_devolucion,
            'cobro_pendiente': cambio.cobro_pendiente,
            'devolucion_pendiente': cambio.devolucion_pendiente,
            'tiene_obligacion_pendiente': cambio.tiene_obligacion_pendiente,
            
            # Responsables
            'solicitado_por': cambio.solicitado_por.username,
            'aprobado_por': cambio.aprobado_por.username if cambio.aprobado_por else '',
            
            # Observaciones
            'motivo_principal': cambio.motivo_principal,
            'motivo_principal_display': cambio.get_motivo_principal_display(),
            'observaciones_cliente': cambio.observaciones_cliente or '',
            'observaciones_vendedor': cambio.observaciones_vendedor or '',
            'observaciones_aprobacion': cambio.observaciones_aprobacion or '',
            
            # Políticas
            'requiere_autorizacion': cambio.requiere_autorizacion,
            'autorizado_excepcion': cambio.autorizado_excepcion,

            # Trazabilidad y control
            'es_fuera_de_plazo': cambio.es_fuera_de_plazo,
            'dias_fuera_de_plazo': cambio.dias_fuera_de_plazo,
            'tipo_cambio_especial': cambio.tipo_cambio_especial,
            'es_autorizacion_cross_branch': cambio.es_autorizacion_cross_branch,
            'es_cambio_concepto': cambio.es_cambio_concepto,
            'concepto_descripcion': cambio.concepto_descripcion or '',
            'autorizado_por_usuario': cambio.autorizado_por_usuario.get_full_name() if cambio.autorizado_por_usuario else None,
            'excepcion_plazo_ya_autorizada': _autorizacion_fuera_plazo_previa(cambio) is not None,
            'sucursal_autorizador': cambio.sucursal_autorizador.alias if cambio.sucursal_autorizador else None,
            'score_riesgo': cambio.score_riesgo,
            'requiere_revision_gerencial': cambio.requiere_revision_gerencial,
            'revisado_por_gerencia': cambio.revisado_por_gerencia.get_full_name() if cambio.revisado_por_gerencia else None,
            'fecha_revision_gerencia': cambio.fecha_revision_gerencia.strftime('%d/%m/%Y %H:%M') if cambio.fecha_revision_gerencia else None,
            'notas_revision_gerencia': cambio.notas_revision_gerencia or '',
            'acciones_permitidas': acciones_permitidas,

            # Tickets
            'ticket_original': {
                'correlativo': cambio.ticket_original.correlativo,
                'fecha': cambio.ticket_original.fecha.strftime('%d/%m/%Y'),
                'total': float(cambio.ticket_original.total),
                'cliente_nombre': cambio.ticket_original.cliente_nombre or '',
                'cliente_rut': cambio.ticket_original.cliente_rut or '',
                'vendedor': cambio.ticket_original.vendedor.nombre if cambio.ticket_original.vendedor else '',
                'vendedor_id': cambio.ticket_original.vendedor.id if cambio.ticket_original.vendedor else None,
                'vendedor_codigo': cambio.ticket_original.vendedor.codigo_vendedor if cambio.ticket_original.vendedor else '',
            },
            'ticket_nuevo': {
                'correlativo': cambio.ticket_nuevo.correlativo if cambio.ticket_nuevo else '',
                'total': float(cambio.ticket_nuevo.total) if cambio.ticket_nuevo else 0,
            } if cambio.ticket_nuevo else None,

            # Nota de Crédito
            'nc_generada': cambio.nc_generada,
            'metodo_devolucion': cambio.metodo_devolucion,
            'metodo_devolucion_display': cambio.get_metodo_devolucion_display() if cambio.metodo_devolucion != 'SIN_NC' else '',
            'fecha_nc': cambio.fecha_nc.strftime('%d/%m/%Y %H:%M') if cambio.fecha_nc else None,
            'nota_credito': {
                'id': cambio.nota_credito.id,
                'numero': cambio.nota_credito.numero_documento,
                'monto': float(cambio.nota_credito.monto_con_iva),
                'fecha': cambio.nota_credito.fecha_emision.strftime('%d/%m/%Y'),
                'estado': cambio.nota_credito.estado_dte,
            } if cambio.nota_credito else None,
        }
        
        # Detalles de productos
        productos_detalle = []
        for detalle in cambio.detalles.all():
            producto_original = detalle.producto_original
            
            # Manejar producto original (puede ser NULL para productos adicionales)
            producto_original_data = None
            if producto_original:
                producto_original_data = {
                    'sku': producto_original.ProductoTalla.sku,
                    'articulo': producto_original.ProductoTalla.producto.articulo,
                    'descripcion': producto_original.ProductoTalla.producto.descripcion,
                    'talla': producto_original.ProductoTalla.talla,
                    'cantidad_original': producto_original.stock,
                    'precio_unitario': float(producto_original.precio),
                }
            
            productos_detalle.append({
                'id': detalle.id,
                'producto_original': producto_original_data,
                'cantidad_cambio': detalle.cantidad_original,
                'precio_original_unitario': float(detalle.precio_original_unitario),
                'valor_original_total': float(detalle.valor_original_total),
                
                'producto_nuevo': {
                    'sku': detalle.producto_nuevo.sku if detalle.producto_nuevo else '',
                    'articulo': detalle.producto_nuevo.producto.articulo if detalle.producto_nuevo else '',
                    'descripcion': detalle.producto_nuevo.producto.descripcion if detalle.producto_nuevo else '',
                    'talla': detalle.producto_nuevo.talla if detalle.producto_nuevo else '',
                    'precio_unitario': float(detalle.precio_nuevo),
                } if detalle.producto_nuevo else None,
                
                'cantidad_nueva': detalle.cantidad_nueva,
                'valor_nuevo_total': float(detalle.valor_nuevo_total),
                'diferencia_unitaria': float(detalle.diferencia_unitaria),
                'diferencia_total': float(detalle.diferencia_total),
                
                'condicion_producto': detalle.condicion_producto,
                'condicion_producto_display': detalle.get_condicion_producto_display(),
                'apto_para_venta': detalle.apto_para_venta,
                'observaciones': detalle.observaciones or '',
                
                'es_cambio': detalle.es_cambio,
                'es_devolucion': detalle.es_devolucion,
            })
        
        # Pagos asociados
        pagos_data = []
        for pago in cambio.pagos.all():
            pagos_data.append({
                'id': pago.id,
                'tipo_pago': pago.tipo_pago,
                'tipo_pago_display': pago.get_tipo_pago_display(),
                'metodo_pago': pago.metodo_pago,
                'metodo_pago_display': pago.get_metodo_pago_display(),
                'monto': float(pago.monto),
                'referencia_pago': pago.referencia_pago or '',
                'numero_autorizacion': pago.numero_autorizacion or '',
                'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y %H:%M'),
                'procesado_por': pago.procesado_por.username,
                'observaciones': pago.observaciones or '',
            })
        
        # Historial
        historial_data = []
        for hist in cambio.historial.order_by('timestamp'):
            historial_data.append({
                'id': hist.id,
                'accion': hist.accion,
                'accion_display': hist.get_accion_display(),
                'estado_anterior': hist.estado_anterior or '',
                'estado_nuevo': hist.estado_nuevo or '',
                'usuario': hist.usuario.username if hist.usuario else '-',
                'descripcion': hist.descripcion,
                'timestamp': hist.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                'datos_adicionales': hist.datos_adicionales or {},
            })
        
        # Datos del ticket nuevo para impresión
        ticket_data = None
        if cambio.ticket_nuevo:
            ticket_data = construir_ticket_data(cambio.ticket_nuevo)
            ticket_data['es_ticket_cambio'] = True
            ticket_data['numero_operacion']            = cambio.numero_operacion
            ticket_data['ticket_original_correlativo'] = (
                cambio.ticket_original.correlativo if cambio.ticket_original else None
            )
            ticket_data['tipo_operacion']         = cambio.tipo_operacion
            ticket_data['tipo_operacion_display']  = cambio.get_tipo_operacion_display()
            ticket_data['monto_original']  = int(cambio.monto_original)
            ticket_data['monto_nuevo']     = int(cambio.monto_nuevo)
            ticket_data['diferencia_monto'] = int(cambio.diferencia_monto)

        return JsonResponse({
            'success': True,
            'cambio': cambio_data,
            'productos': productos_detalle,
            'pagos': pagos_data,
            'historial': historial_data,
            'ticket_data': ticket_data  # Datos completos del ticket para impresión
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener detalle: {str(e)}'
        })


@login_required
@require_POST
def cancelar_cambio_devolucion(request):
    """Cancelar una solicitud antes de que genere tickets o movimientos."""
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        motivo = str(data.get('motivo') or '').strip()
        codigo_autorizacion = str(data.get('codigo_autorizacion') or '').strip()
        minutos_autorizacion = data.get('minutos_autorizacion', 30)
        
        if not cambio_id:
            return JsonResponse({'success': False, 'error': 'ID de cambio requerido'}, status=400)
        if len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'code': 'REASON_REQUIRED',
                'error': 'Debe indicar un motivo de al menos 5 caracteres',
            }, status=400)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal seleccionada'}, status=400)

        with transaction.atomic():
            # of=('self',): bloquea solo la fila de CambioDevolucion. Sin esto,
            # select_related('ticket_nuevo') (FK nullable → LEFT OUTER JOIN) provoca
            # "FOR UPDATE cannot be applied to the nullable side of an outer join" en PostgreSQL.
            cambio = get_object_or_404(
                CambioDevolucion.objects.select_for_update(of=('self',)).select_related(
                    'sucursal__empresa', 'ticket_nuevo'
                ),
                id=cambio_id,
            )
            if cambio.sucursal_id != int(sucursal_id):
                return JsonResponse({
                    'success': False,
                    'code': 'BRANCH_DENIED',
                    'error': 'No tiene acceso a este cambio',
                }, status=403)
            if cambio.estado not in ('SOLICITADO', 'APROBADO') or cambio.ticket_nuevo_id:
                return JsonResponse({
                    'success': False,
                    'code': 'INVALID_STATE',
                    'error': 'Solo se pueden cancelar solicitudes no ejecutadas',
                }, status=409)

            acciones = _acciones_cambio_para_usuario(request.user, cambio)
            permiso_temporal = None
            if not acciones['cancelar']:
                if not acciones['puede_solicitar_cancelar']:
                    return JsonResponse({
                        'success': False,
                        'code': 'PERMISSION_DENIED',
                        'error': 'Su perfil no tiene permiso para cancelar cambios',
                    }, status=403)
                if not codigo_autorizacion:
                    return JsonResponse({
                        'success': False,
                        'code': 'TEMP_AUTH_REQUIRED',
                        'requiere_autorizacion_temporal': True,
                        'error': 'Ingrese el código de autorización de un administrador',
                    }, status=403)
                permiso_temporal, error_response = _otorgar_permiso_temporal_desde_codigo(
                    request,
                    cambio,
                    PermisoTemporalCambio.ACCION_CANCELAR,
                    codigo_autorizacion,
                    motivo,
                    minutos_autorizacion,
                )
                if error_response:
                    return error_response

            estado_anterior = cambio.estado
            cambio.estado = 'CANCELADO'
            cambio.observaciones_aprobacion = (
                f'Cancelado por {request.user.username} el '
                f'{timezone.localtime().strftime("%d/%m/%Y %H:%M")}. Motivo: {motivo}'
            )
            cambio.save(update_fields=['estado', 'observaciones_aprobacion'])
            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion='CANCELADO',
                estado_anterior=estado_anterior,
                estado_nuevo='CANCELADO',
                usuario=request.user,
                descripcion=f'Solicitud cancelada. Motivo: {motivo}',
                datos_adicionales={
                    'motivo': motivo,
                    'permiso_temporal_id': permiso_temporal.id if permiso_temporal else None,
                    'fecha_cancelacion': timezone.now().isoformat(),
                },
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Cambio {cambio.numero_operacion} cancelado correctamente',
            'nuevo_estado': 'CANCELADO',
            'permiso_temporal_hasta': (
                timezone.localtime(permiso_temporal.vigente_hasta).strftime('%H:%M')
                if permiso_temporal else None
            ),
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        logger.exception('Error al cancelar cambio/devolución')
        return JsonResponse({
            'success': False,
            'error': f'Error al cancelar cambio: {str(e)}'
        }, status=500)


@login_required
@require_POST
def revertir_cambio_devolucion(request):
    """Revertir un cambio ejecutado cuyo ticket financiero aún está pendiente.
    Deshace todos los movimientos de stock y cancela el ticket pendiente."""
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        motivo = str(data.get('motivo') or '').strip()
        codigo_autorizacion = str(data.get('codigo_autorizacion') or '').strip()
        minutos_autorizacion = data.get('minutos_autorizacion', 30)

        if not cambio_id:
            return JsonResponse({'success': False, 'error': 'ID de cambio requerido'}, status=400)
        if len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'code': 'REASON_REQUIRED',
                'error': 'Debe indicar un motivo de al menos 5 caracteres',
            }, status=400)

        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal seleccionada'}, status=400)

        with transaction.atomic():
            cambio = get_object_or_404(
                CambioDevolucion.objects.select_for_update().select_related(
                    'sucursal__empresa'
                ),
                id=cambio_id,
            )
            if cambio.sucursal_id != int(sucursal_id):
                return JsonResponse({
                    'success': False,
                    'code': 'BRANCH_DENIED',
                    'error': 'No tiene acceso a este cambio',
                }, status=403)

            estados_revertibles = (
                'EJECUTADO', 'EJECUTADO_COBRO_PENDIENTE', 'EJECUTADO_DEVOL_PENDIENTE'
            )
            if cambio.estado not in estados_revertibles:
                return JsonResponse({
                    'success': False,
                    'code': 'INVALID_STATE',
                    'error': 'Solo se pueden revertir cambios ejecutados con cobro o devolución pendiente',
                }, status=409)
            if cambio.diferencia_condonada or cambio.nc_generada:
                return JsonResponse({
                    'success': False,
                    'code': 'FINANCIAL_OPERATION_CLOSED',
                    'error': 'Este cambio ya tiene una operación financiera cerrada y no se puede revertir',
                }, status=409)
            ticket_ids = sorted(filter(None, [
                cambio.ticket_nuevo_id,
                cambio.ticket_diferencia_id,
            ]))
            tickets_bloqueados = {
                ticket.id: ticket
                for ticket in Ticket.objects.select_for_update()
                .filter(id__in=ticket_ids)
                .order_by('id')
            }
            ticket_nuevo = tickets_bloqueados.get(cambio.ticket_nuevo_id)
            ticket_diferencia = tickets_bloqueados.get(cambio.ticket_diferencia_id)
            if not ticket_nuevo or ticket_nuevo.estado != 'PENDIENTE':
                return JsonResponse({
                    'success': False,
                    'code': 'TICKET_NOT_PENDING',
                    'error': 'El ticket ya fue pagado, anulado o completado; use el flujo formal de anulacion',
                }, status=409)

            pagos_ticket = list(
                TicketDetallePago.objects.select_for_update()
                .filter(ticket_id__in=ticket_ids)
                .values_list('id', flat=True)
            )
            pagos_cambio = list(
                PagoCambioDevolucion.objects.select_for_update()
                .filter(cambio_devolucion=cambio)
                .values_list('id', flat=True)
            )
            if pagos_ticket or pagos_cambio:
                return JsonResponse({
                    'success': False,
                    'code': 'FINANCIAL_OPERATION_STARTED',
                    'error': 'El cambio ya registra pagos y no se puede revertir desde este flujo',
                }, status=409)

            # Poblar las relaciones con los objetos que ya quedaron bloqueados.
            cambio.ticket_nuevo = ticket_nuevo
            cambio.ticket_diferencia = ticket_diferencia

            detalles_bloqueados = list(
                cambio.detalles.select_for_update().order_by('id')
            )
            productos_bloqueados = _bloquear_y_validar_inventario_cambio(
                detalles_bloqueados,
                cambio.sucursal_id,
                reversion=True,
            )

            acciones = _acciones_cambio_para_usuario(request.user, cambio)
            permiso_temporal = None
            if not acciones['revertir']:
                if not acciones['puede_solicitar_revertir']:
                    return JsonResponse({
                        'success': False,
                        'code': 'PERMISSION_DENIED',
                        'error': 'Su perfil no tiene permiso para revertir cambios',
                    }, status=403)
                if not codigo_autorizacion:
                    return JsonResponse({
                        'success': False,
                        'code': 'TEMP_AUTH_REQUIRED',
                        'requiere_autorizacion_temporal': True,
                        'error': 'Ingrese el código de autorización de un administrador',
                    }, status=403)
                permiso_temporal, error_response = _otorgar_permiso_temporal_desde_codigo(
                    request,
                    cambio,
                    PermisoTemporalCambio.ACCION_REVERTIR,
                    codigo_autorizacion,
                    motivo,
                    minutos_autorizacion,
                )
                if error_response:
                    return error_response

            sucursal = cambio.sucursal
            # 1) Los productos nuevos entregados regresan a stock y FIFO.
            for item in detalles_bloqueados:
                if not item.producto_nuevo_id or item.cantidad_nueva <= 0:
                    continue
                producto_talla = productos_bloqueados[item.producto_nuevo_id]
                ingresar_inventario(
                    producto_talla=producto_talla,
                    cantidad=item.cantidad_nueva,
                    concepto='REVERSION_CAMBIO',
                    responsable=request.user.username,
                    sucursal_destino=cambio.sucursal,
                    ticket=ticket_nuevo,
                    costo_unitario=producto_talla.producto.costo or 0,
                    precio_unitario=int(item.precio_nuevo),
                    observaciones=(
                        f'Reversion - Cambio #{cambio.numero_operacion}. '
                        'Producto entregado devuelto al stock.'
                    ),
                    referencia_externa=cambio.numero_operacion,
                )

            # 2) Los productos originales aptos salen nuevamente. El servicio
            # rechaza la reversion si esas unidades ya no estan disponibles.
            for item in detalles_bloqueados:
                if (
                    not item.producto_original_id
                    or item.cantidad_original <= 0
                    or not item.apto_para_venta
                ):
                    continue
                producto_id = item.producto_original.ProductoTalla_id
                egresar_inventario(
                    producto_talla=productos_bloqueados[producto_id],
                    cantidad=item.cantidad_original,
                    concepto='REVERSION_CAMBIO',
                    responsable=request.user.username,
                    sucursal_origen=cambio.sucursal,
                    ticket=cambio.ticket_original,
                    precio_unitario=int(item.precio_original_unitario),
                    observaciones=(
                        f'Reversion - Cambio #{cambio.numero_operacion}. '
                        'Se revierte el ingreso de devolucion.'
                    ),
                    referencia_externa=cambio.numero_operacion,
                )

            # 3) Cancelar ticket pendiente
            ticket = cambio.ticket_nuevo
            ticket.estado = 'ANULADO'
            ticket.observaciones = (ticket.observaciones or '') + f'\n\nANULADO por reversión de cambio #{cambio.numero_operacion}'
            ticket.save()

            # 4) Cancelar ticket de diferencia si existe
            if cambio.ticket_diferencia and cambio.ticket_diferencia.estado == 'PENDIENTE':
                cambio.ticket_diferencia.estado = 'ANULADO'
                cambio.ticket_diferencia.observaciones = (cambio.ticket_diferencia.observaciones or '') + f'\n\nANULADO por reversión de cambio #{cambio.numero_operacion}'
                cambio.ticket_diferencia.save()

            # 5) Cambiar estado del cambio
            estado_anterior = cambio.estado
            cambio.estado = 'REVERTIDO'
            cambio.save()

            # 6) Registrar historial
            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion='REVERTIDO',
                estado_anterior=estado_anterior,
                estado_nuevo='REVERTIDO',
                usuario=request.user,
                descripcion=f'Cambio revertido por {request.user.get_full_name() or request.user.username}. Stock restaurado, ticket #{ticket.correlativo} anulado.' + (f' Motivo: {motivo}' if motivo else ''),
                datos_adicionales={
                    'motivo': motivo,
                    'ticket_anulado': ticket.correlativo,
                    'permiso_temporal_id': permiso_temporal.id if permiso_temporal else None,
                    'fecha_reversion': timezone.now().isoformat()
                }
            )

        return JsonResponse({
            'success': True,
            'message': f'Cambio {cambio.numero_operacion} revertido exitosamente. Stock restaurado.',
            'nuevo_estado': 'REVERTIDO',
            'permiso_temporal_hasta': (
                timezone.localtime(permiso_temporal.vigente_hasta).strftime('%H:%M')
                if permiso_temporal else None
            ),
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except ConflictoInventarioCambio as e:
        return JsonResponse({
            'success': False,
            'code': e.codigo,
            'error': str(e),
        }, status=409)
    except Exception as e:
        logger.exception("Error al revertir cambio/devolucion")
        return JsonResponse({'success': False, 'error': f'Error al revertir cambio: {str(e)}'}, status=500)


@login_required
@require_POST
def aprobar_cambio_devolucion(request):
    """Ruta legacy deshabilitada: la aprobacion se ejecuta en el flujo unificado."""
    return JsonResponse({
        'success': False,
        'code': 'LEGACY_FLOW_DISABLED',
        'error': (
            'El flujo separado de aprobacion fue deshabilitado. '
            'Use la autorizacion y ejecucion unificada con codigo de Administrador.'
        ),
    }, status=410)


@login_required
@require_POST
def ejecutar_cambio_devolucion(request):
    """Ruta legacy deshabilitada: la ejecucion requiere el flujo unificado."""
    return JsonResponse({
        'success': False,
        'code': 'LEGACY_FLOW_DISABLED',
        'error': (
            'La ejecucion separada fue deshabilitada. '
            'Use la autorizacion y ejecucion unificada con codigo de Administrador.'
        ),
    }, status=410)


@login_required
@require_POST
def registrar_pago_diferencia(request):
    """Ruta legacy deshabilitada: el POS es el cierre financiero canonico."""
    return JsonResponse({
        'success': False,
        'code': 'LEGACY_FLOW_DISABLED',
        'error': (
            'El cierre manual de diferencias fue deshabilitado. '
            'El cobro o devolucion debe procesarse desde el POS.'
        ),
    }, status=410)


@require_POST
@requiere_rol('administrador')
def condonar_diferencia_cobro(request):
    """
    Condonar (perdonar) la diferencia de cobro pendiente de un cambio, con justificación.

    Solo administradores. NO anula el cambio: no se revierte stock ni se deshacen
    movimientos (eso lo hace revertir_cambio_devolucion). Solo se perdona el cobro,
    el cambio pasa a COMPLETADO y deja de contar en "Pend. Cobro".
    """
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        motivo = str(data.get('motivo', '') or '').strip()

        if not cambio_id:
            return JsonResponse({'success': False, 'error': 'Falta el identificador del cambio'})

        if len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'error': 'Debe indicar una justificación (mínimo 5 caracteres) para condonar la diferencia'
            })

        cambio = get_object_or_404(CambioDevolucion, id=cambio_id)

        # Verificar acceso por sucursal (fail-closed: exigir sucursal en sesión)
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal seleccionada'})
        if cambio.sucursal_id != int(sucursal_id):
            return JsonResponse({'success': False, 'error': 'No tiene acceso a este cambio'})

        with transaction.atomic():
            # Bloquear la fila para idempotencia frente a doble click / operaciones
            # concurrentes (cobro o reversión simultáneos): se re-lee y re-valida el estado.
            cambio = CambioDevolucion.objects.select_for_update().get(id=cambio_id)

            # Solo cambios con cobro de diferencia pendiente (re-verificado bajo el lock)
            if cambio.estado != 'EJECUTADO_COBRO_PENDIENTE':
                return JsonResponse({
                    'success': False,
                    'error': f'Solo se puede condonar un cambio con cobro pendiente. Estado actual: {cambio.get_estado_display()}'
                })

            monto = cambio.diferencia_monto or 0
            estado_anterior = cambio.estado

            # Anular el ticket pendiente del cobro (NO se revierte stock ni movimientos).
            ticket_pendiente = None
            if cambio.ticket_diferencia and cambio.ticket_diferencia.estado == 'PENDIENTE':
                ticket_pendiente = cambio.ticket_diferencia
            elif cambio.ticket_nuevo and cambio.ticket_nuevo.estado == 'PENDIENTE':
                ticket_pendiente = cambio.ticket_nuevo
            if ticket_pendiente:
                ticket_pendiente.estado = 'ANULADO'
                nota = f'[CONDONACIÓN] Diferencia condonada por {request.user.username}: {motivo}'
                ticket_pendiente.observaciones = ((ticket_pendiente.observaciones or '') + f'\n{nota}').strip()
                ticket_pendiente.save()

            # Marcar la condonación en el cambio sin anularlo
            cambio.estado = 'COMPLETADO'
            cambio.diferencia_condonada = True
            cambio.motivo_condonacion = motivo
            cambio.condonada_por = request.user
            cambio.fecha_condonacion = timezone.now()
            cambio.fecha_completado = timezone.now()
            cambio.save()

            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion='CONDONACION_DIFERENCIA',
                estado_anterior=estado_anterior,
                estado_nuevo='COMPLETADO',
                usuario=request.user,
                descripcion=f'Diferencia de ${int(monto):,} condonada por {request.user.username}. Motivo: {motivo}',
                datos_adicionales={
                    'monto_condonado': float(monto),
                    'motivo': motivo,
                    'condonada_por': request.user.username,
                    'fecha': timezone.now().isoformat(),
                }
            )

        return JsonResponse({
            'success': True,
            'message': f'Diferencia de ${int(monto):,} condonada. El cambio quedó completado.',
            'cambio_id': cambio.id,
            'estado_final': cambio.get_estado_display(),
            'estado_final_codigo': cambio.estado,
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al condonar la diferencia: {str(e)}'})


@require_POST
@requiere_rol('administrador')
def ajustar_diferencia_cobro(request):
    """
    Ajustar (rebajar) la diferencia de cobro pendiente de un cambio a un monto
    menor, con justificación. Solo administradores.

    A diferencia de condonar (que perdona el 100% y cierra el cambio), acá el
    cobro sigue vivo: el ticket pendiente queda por el monto ajustado y se cobra
    normal en el POS. La rebaja se materializa como una línea manual negativa
    en el ticket, porque la generación del DTE recalcula el total desde la suma
    de líneas (ver construir DTE: total autoritativo = suma de ticket_productos)
    y un total editado sin línea de respaldo sería revertido en la emisión.
    """
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        motivo = str(data.get('motivo', '') or '').strip()

        if not cambio_id:
            return JsonResponse({'success': False, 'error': 'Falta el identificador del cambio'})

        if len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'error': 'Debe indicar una justificación (mínimo 5 caracteres) para ajustar la diferencia'
            })

        try:
            nuevo_monto = int(data.get('nuevo_monto'))
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'El nuevo monto debe ser un número entero'})

        if nuevo_monto <= 0:
            return JsonResponse({
                'success': False,
                'error': 'El nuevo monto debe ser mayor que $0. Para perdonar todo el cobro use "Condonar".'
            })

        get_object_or_404(CambioDevolucion, id=cambio_id)

        # Verificar acceso por sucursal (fail-closed: exigir sucursal en sesión)
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal seleccionada'})

        with transaction.atomic():
            # Bloquear la fila para idempotencia frente a doble click / operaciones
            # concurrentes (cobro, condonación o reversión simultáneos).
            cambio = CambioDevolucion.objects.select_for_update().get(id=cambio_id)

            if cambio.sucursal_id != int(sucursal_id):
                return JsonResponse({'success': False, 'error': 'No tiene acceso a este cambio'})

            # Solo cambios con cobro de diferencia pendiente (re-verificado bajo el lock)
            if cambio.estado != 'EJECUTADO_COBRO_PENDIENTE':
                return JsonResponse({
                    'success': False,
                    'error': f'Solo se puede ajustar un cambio con cobro pendiente. Estado actual: {cambio.get_estado_display()}'
                })

            diferencia_actual = int(cambio.diferencia_monto or 0)
            if nuevo_monto >= diferencia_actual:
                return JsonResponse({
                    'success': False,
                    'error': f'El nuevo monto (${nuevo_monto:,}) debe ser menor que la diferencia actual (${diferencia_actual:,})'
                })

            # Ubicar y bloquear el ticket pendiente del cobro (misma prioridad que
            # condonar: ticket_diferencia primero, luego ticket_nuevo). El lock
            # evita la carrera con un cobro simultáneo en el POS.
            ticket_ids = [tid for tid in (cambio.ticket_diferencia_id, cambio.ticket_nuevo_id) if tid]
            tickets_bloqueados = {
                t.id: t
                for t in Ticket.objects.select_for_update().filter(id__in=sorted(ticket_ids))
            }
            ticket_pendiente = None
            for tid in ticket_ids:
                candidato = tickets_bloqueados.get(tid)
                if candidato and candidato.estado == 'PENDIENTE':
                    ticket_pendiente = candidato
                    break
            if not ticket_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': 'El ticket del cobro ya no está pendiente (fue pagado o anulado)'
                })

            rebaja = int(ticket_pendiente.total or 0) - nuevo_monto
            if rebaja <= 0:
                return JsonResponse({
                    'success': False,
                    'error': f'El nuevo monto (${nuevo_monto:,}) debe ser menor que el total del ticket pendiente (${int(ticket_pendiente.total or 0):,})'
                })

            # Línea manual negativa que respalda la rebaja: así la suma de líneas
            # del ticket cuadra con el nuevo total y el DTE se emite por el monto
            # realmente cobrado.
            Ticket_Productos.objects.create(
                idTicket=ticket_pendiente,
                ProductoTalla=None,
                stock=1,
                precio=-rebaja,
                precio_original=-rebaja,
                descuento_unitario=0,
                subtotal=-rebaja,
                descripcion_linea=f'AJUSTE DE DIFERENCIA (admin {request.user.username})'[:255],
            )

            nota = (f'[AJUSTE DIFERENCIA] ${diferencia_actual:,} → ${nuevo_monto:,} '
                    f'por {request.user.username}: {motivo}')
            ticket_pendiente.total = nuevo_monto
            ticket_pendiente.subTotal = nuevo_monto
            ticket_pendiente.observaciones = ((ticket_pendiente.observaciones or '') + f'\n{nota}').strip()
            ticket_pendiente.save(update_fields=['total', 'subTotal', 'observaciones'])

            # Registrar el ajuste en el cambio SIN cerrar el cobro: el estado se
            # mantiene EJECUTADO_COBRO_PENDIENTE y se completa al pagar en el POS.
            if cambio.monto_diferencia_original is None:
                cambio.monto_diferencia_original = cambio.diferencia_monto
            cambio.diferencia_monto = nuevo_monto
            cambio.diferencia_ajustada = True
            cambio.motivo_ajuste = motivo
            cambio.ajustada_por = request.user
            cambio.fecha_ajuste = timezone.now()
            cambio.save(update_fields=[
                'monto_diferencia_original', 'diferencia_monto', 'diferencia_ajustada',
                'motivo_ajuste', 'ajustada_por', 'fecha_ajuste',
            ])

            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion='AJUSTE_DIFERENCIA',
                estado_anterior=cambio.estado,
                estado_nuevo=cambio.estado,
                usuario=request.user,
                descripcion=(
                    f'Diferencia ajustada de ${diferencia_actual:,} a ${nuevo_monto:,} '
                    f'(rebaja de ${rebaja:,}) por {request.user.username}. Motivo: {motivo}'
                ),
                datos_adicionales={
                    'monto_anterior': float(diferencia_actual),
                    'monto_nuevo': float(nuevo_monto),
                    'monto_rebajado': float(rebaja),
                    'motivo': motivo,
                    'ajustada_por': request.user.username,
                    'ticket_correlativo': ticket_pendiente.correlativo,
                    'fecha': timezone.now().isoformat(),
                }
            )

        return JsonResponse({
            'success': True,
            'message': (
                f'Diferencia ajustada de ${diferencia_actual:,} a ${nuevo_monto:,}. '
                f'El cobro sigue pendiente en el POS por el monto nuevo.'
            ),
            'cambio_id': cambio.id,
            'nuevo_monto': nuevo_monto,
            'monto_rebajado': rebaja,
            'ticket_correlativo': ticket_pendiente.correlativo,
            'estado_final': cambio.get_estado_display(),
            'estado_final_codigo': cambio.estado,
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al ajustar la diferencia: {str(e)}'})


@login_required
@require_POST
def aprobar_cambio_generar_ticket(request):
    """Autorizar, aprobar y ejecutar el cambio en una sola transacción."""
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        vendedor_id = data.get('vendedor_id')
        observaciones = data.get('observaciones', '')
        # Código dinámico de autorización de la barra superior (6 dígitos).
        # Fuera de plazo debe ser de un ADMINISTRADOR; dentro de plazo basta admin o jefe de local.
        credencial = str(data.get('codigo_autorizacion') or '').strip()

        if not all([cambio_id, vendedor_id, credencial]):
            return JsonResponse({
                'success': False,
                'code': 'AUTH_CODE_REQUIRED',
                'error': 'ID de cambio, vendedor y código de autorización requeridos'
            }, status=400)

        # Obtener cambio
        cambio = get_object_or_404(CambioDevolucion, id=cambio_id)

        # Verificar acceso
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'code': 'BRANCH_REQUIRED',
                'error': 'No hay una sucursal seleccionada'
            }, status=400)
        if cambio.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'error': 'No tiene acceso a este cambio'
            }, status=403)

        # ¿Requiere autorización especial de administrador?
        # La excepción de plazo se autoriza UNA sola vez, al crear la solicitud.
        # Si el cambio ya trae esa firma, aprobar vuelve a ser un paso normal
        # (código de administrador o de jefe de local).
        #
        # Se evalúa sobre `es_fuera_de_plazo` (marcado al crear) y no sobre
        # `dentro_del_plazo` (que se recalcula contra la fecha de hoy): un cambio
        # creado en plazo que queda días esperando aprobación no es una excepción
        # y no debe empezar a exigir un administrador por el solo paso del tiempo.
        autorizacion_previa = _autorizacion_fuera_plazo_previa(cambio)
        requiere_admin = cambio.es_fuera_de_plazo and autorizacion_previa is None

        # Toda autorización se hace con el código dinámico de la barra superior.
        es_valido_codigo, mensaje_codigo, codigo_obj = \
            CodigoAutorizacionDinamico.validar_codigo(credencial)
        usuario_autorizador = codigo_obj.generado_por if (es_valido_codigo and codigo_obj) else None

        if not es_valido_codigo or not usuario_autorizador:
            return JsonResponse({
                'success': False,
                'code': 'INVALID_AUTH_CODE',
                'error': mensaje_codigo or 'Código de autorización inválido',
            }, status=403)

        if not (usuario_autorizador.is_active and getattr(usuario_autorizador, 'es_activo', True)):
            codigo_obj = None
            return JsonResponse({
                'success': False,
                'code': 'INVALID_AUTH_CODE',
                'error': 'El código no pertenece a un usuario activo',
            }, status=403)

        # Casos especiales (fuera de plazo) → SOLO el código de un ADMINISTRADOR.
        # Cambios normales (dentro de plazo) → código de admin o de jefe de local.
        if requiere_admin and not _usuario_es_administrador_activo(usuario_autorizador):
            return JsonResponse({
                'success': False,
                'code': 'ADMIN_REQUIRED',
                'error': (
                    'Este cambio fuera de plazo no tiene una autorización de administrador '
                    'vigente, así que requiere el código de un ADMINISTRADOR para aprobarse.'
                ),
            }, status=403)

        # Cambios normales: el código debe ser de un administrador o jefe de
        # local (mismo criterio que para poder generar el código del navbar).
        if not requiere_admin and not (
            _usuario_es_administrador_activo(usuario_autorizador)
            or getattr(usuario_autorizador, 'rol', '') == 'jefe_local'
        ):
            return JsonResponse({
                'success': False,
                'code': 'INVALID_AUTHORIZER',
                'error': 'El código debe ser de un administrador o de un jefe de local',
            }, status=403)

        asignacion_autorizador = EmpresaUser.objects.filter(
            user=usuario_autorizador,
            empresa_id=cambio.sucursal.empresa_id,
            status=True,
        ).select_related('sucursal').order_by('-active').first()
        es_cross_company = asignacion_autorizador is None
        if es_cross_company:
            # La excepción de plazo sigue amarrada a la empresa: solo un
            # administrador de la MISMA empresa puede firmarla.
            if requiere_admin:
                return JsonResponse({
                    'success': False,
                    'code': 'CROSS_COMPANY_AUTH',
                    'error': 'El administrador debe pertenecer a la misma empresa',
                }, status=403)
            # Cambio normal: el código de admin/jefe de local vale aunque sea de
            # otra empresa. Su asignación de origen se conserva solo para
            # auditoría y la aprobación queda marcada para revisión gerencial.
            asignacion_autorizador = EmpresaUser.objects.filter(
                user=usuario_autorizador,
                status=True,
            ).select_related('sucursal').order_by('-active').first()

        # Verificar estado
        if cambio.estado not in ('SOLICITADO', 'APROBADO'):
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden aprobar/ejecutar cambios en estado Solicitado o Aprobado'
            })

        # Validar vendedor (vendedor_id es el ID del modelo Vendedor, no User).
        # Se acepta cualquier vendedor activo de la empresa (de cualquier tienda).
        vendedor_obj = _vendedores_para_autorizacion_cambio(cambio.sucursal).filter(
            id=vendedor_id
        ).first()
        if not vendedor_obj:
            return JsonResponse({
                'success': False,
                'code': 'SELLER_NOT_AVAILABLE',
                'error': 'El vendedor no esta activo',
            }, status=403)
        
        logger.info("Iniciando transaccion atomica para aprobar cambio %s", cambio.id)
        
        with transaction.atomic():
            cambio = CambioDevolucion.objects.select_for_update().select_related(
                'sucursal__empresa', 'ticket_original'
            ).get(id=cambio_id)
            if cambio.estado not in ('SOLICITADO', 'APROBADO'):
                return JsonResponse({
                    'success': False,
                    'code': 'OPERATION_CHANGED',
                    'error': 'El cambio fue modificado por otro usuario. Actualice la pantalla.',
                }, status=409)

            vendedor_obj = _vendedores_para_autorizacion_cambio(cambio.sucursal).filter(
                id=vendedor_id
            ).first()
            if not vendedor_obj:
                return JsonResponse({
                    'success': False,
                    'code': 'SELLER_NOT_AVAILABLE',
                    'error': 'El vendedor no esta activo',
                }, status=403)

            detalles_bloqueados = list(
                cambio.detalles.select_for_update().order_by('id')
            )
            productos_bloqueados = _bloquear_y_validar_inventario_cambio(
                detalles_bloqueados,
                cambio.sucursal_id,
                reversion=False,
            )

            # Todo código dinámico es de un solo uso: se consume aquí, sea de
            # administrador o de jefe de local.
            if codigo_obj is not None:
                codigo_obj = CodigoAutorizacionDinamico.objects.select_for_update().get(id=codigo_obj.id)
                if not codigo_obj.es_valido():
                    return JsonResponse({
                        'success': False,
                        'code': 'AUTH_CODE_ALREADY_USED',
                        'error': 'El código fue utilizado o venció antes de ejecutar el cambio',
                    }, status=409)
                codigo_obj.usado = True
                codigo_obj.save(update_fields=['usado'])

            _autorizador_nombre = (
                usuario_autorizador.get_full_name() or usuario_autorizador.username
            )
            registro_autorizacion = RegistroAutorizacion.objects.create(
                codigo_usado=codigo_obj,
                usuario_solicitante=request.user,
                usuario_autorizador=usuario_autorizador,
                tipo_operacion='APROBACION_CAMBIO',
                descripcion=(
                    (f'Aprobación y ejecución (fuera de plazo) autorizada por {_autorizador_nombre}'
                     if requiere_admin else
                     (f'Aprobación y ejecución (fuera de plazo ya autorizada por '
                      f'{autorizacion_previa.get_full_name() or autorizacion_previa.username}) '
                      f'ejecutada por {_autorizador_nombre}'
                      if autorizacion_previa else
                      f'Aprobación y ejecución (cambio normal) autorizada por {_autorizador_nombre}'))
                ),
                ip_origen=request.META.get('REMOTE_ADDR'),
                exitoso=True,
                cambio_devolucion=cambio,
                sucursal_solicitante=cambio.sucursal,
                sucursal_autorizador=(
                    asignacion_autorizador.sucursal if asignacion_autorizador else None
                ),
                es_cross_branch=bool(
                    es_cross_company
                    or (asignacion_autorizador
                        and asignacion_autorizador.sucursal_id
                        and asignacion_autorizador.sucursal_id != cambio.sucursal_id)
                ),
                requiere_revision=bool(
                    es_cross_company
                    or (asignacion_autorizador
                        and asignacion_autorizador.sucursal_id
                        and asignacion_autorizador.sucursal_id != cambio.sucursal_id)
                ),
                datos_adicionales={
                    'cambio_id': cambio.id,
                    'vendedor_id': vendedor_id,
                },
            )

            # Aprobar el cambio (solo si no está ya aprobado)
            if cambio.estado == 'SOLICITADO':
                logger.debug("Aprobando cambio %s", cambio.id)
                cambio.aprobar_cambio(request.user, observaciones)
                logger.info("Cambio %s aprobado: estado=%s", cambio.id, cambio.estado)
            else:
                logger.debug("Cambio %s ya estaba en estado %s; continuando ejecucion", cambio.id, cambio.estado)

            # No pisar al administrador que autorizó la excepción de plazo al crear:
            # es el dato que justifica el cambio ante una auditoría, y es el que
            # evita volver a exigir un código de administrador si esta aprobación
            # se reintenta (el flujo es re-entrante y admite revertir + re-aprobar).
            campos_autorizacion = ['registro_autorizacion', 'requiere_autorizacion']
            if autorizacion_previa is None:
                cambio.autorizado_por_usuario = usuario_autorizador
                cambio.sucursal_autorizador = (
                    asignacion_autorizador.sucursal if asignacion_autorizador else None
                )
                campos_autorizacion += ['autorizado_por_usuario', 'sucursal_autorizador']
            cambio.registro_autorizacion = registro_autorizacion
            cambio.requiere_autorizacion = bool(requiere_admin or autorizacion_previa)
            cambio.save(update_fields=campos_autorizacion)
            
            # GENERAR TICKET DE VENTA
            # Obtener correlativo usando la función centralizada
            sucursal = get_object_or_404(Sucursal, id=sucursal_id)
            
            logger.debug("Obteniendo correlativo para sucursal %s", sucursal.alias)
            try:
                nuevo_correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
            except Exception as e:
                logger.exception("Error al obtener correlativo para sucursal %s", sucursal.id)
                raise  # Re-lanzar para hacer rollback
            logger.debug("Correlativo obtenido para cambio %s: ticket=%s", cambio.id, nuevo_correlativo)
            
            # Calcular totales
            total_devuelto = cambio.monto_original
            total_nuevo = cambio.monto_nuevo
            diferencia = total_nuevo - total_devuelto
            
            logger.debug(
                "Creando ticket para cambio %s: correlativo=%s, sucursal=%s, vendedor=%s, "
                "total=%s, diferencia=%s, tipo_diferencia=%s",
                cambio.id,
                nuevo_correlativo,
                sucursal.alias,
                vendedor_obj.nombre,
                abs(diferencia),
                diferencia,
                'a cobrar' if diferencia > 0 else 'a devolver',
            )
            
            try:
                # Determinar estado según la diferencia
                if diferencia > 0:
                    # Cliente debe pagar
                    estado_ticket = 'PENDIENTE'
                    metodo_pago_ticket = 'PENDIENTE_COBRO'
                    tipo_documento = 'TICKET_COBRO_CAMBIO'
                elif diferencia < 0:
                    # Se devuelve dinero al cliente
                    estado_ticket = 'PENDIENTE'
                    metodo_pago_ticket = 'PENDIENTE_DEVOLUCION'
                    tipo_documento = 'TICKET_DEVOLUCION'
                else:
                    # Sin diferencia - cambio directo
                    estado_ticket = 'PAGADO'
                    metodo_pago_ticket = 'SIN_DIFERENCIA'
                    tipo_documento = 'TICKET_CAMBIO_DIRECTO'
                
                ticket = Ticket.objects.create(
                    correlativo=nuevo_correlativo,
                    sucursal=sucursal,
                    vendedor=vendedor_obj,
                    responsable=request.user.get_full_name() or request.user.username,
                    cliente_nombre=cambio.ticket_original.cliente_nombre if cambio.ticket_original else 'Cliente General',
                    cliente_rut=cambio.ticket_original.cliente_rut if cambio.ticket_original else '',
                    subTotal=int(abs(diferencia)),
                    total=int(abs(diferencia)),
                    descuento=0,
                    estado=estado_ticket,
                    metodo_pago=metodo_pago_ticket,
                    modulo_origen='CAMBIO_DEVOLUCION',
                    tipo_dte=tipo_documento,  # Usar tipo_dte para identificar
                    observaciones=f'🔄 CAMBIO/DEVOLUCIÓN #{cambio.numero_operacion}\n' +
                                 f'📋 Ticket Original: #{cambio.ticket_original.correlativo}\n\n' +
                                 f'📦 Productos devueltos: ${int(total_devuelto):,}\n' +
                                 f'✨ Productos nuevos: ${int(total_nuevo):,}\n' +
                                 f'💰 Diferencia: ${int(diferencia):,}\n\n' +
                                 (f'💵 A DEVOLVER AL CLIENTE: ${abs(int(diferencia)):,}\n\n' if diferencia < 0 else 
                                  f'💰 A COBRAR AL CLIENTE: ${int(diferencia):,}\n\n' if diferencia > 0 else 
                                  f'✅ SIN DIFERENCIA - CAMBIO DIRECTO\n\n') +
                                 (observaciones if observaciones else '')
                )
                logger.info(
                    "Ticket creado para cambio %s: ticket_id=%s, correlativo=%s",
                    cambio.id,
                    ticket.id,
                    ticket.correlativo,
                )
            except Exception as e:
                logger.exception("Error al crear ticket para cambio %s", cambio.id)
                raise  # Re-lanzar para hacer rollback
            
            # Agregar productos al ticket usando el modelo Ticket_Productos
            logger.debug(
                "Agregando productos al ticket de cambio %s: detalles=%s",
                cambio.id,
                cambio.detalles.count(),
            )
            
            try:
                # ✅ CORREGIDO: Agrupar productos para evitar duplicados
                # Estructura: { producto_talla_id: { 'producto': obj, 'cantidad': n, 'precio': p, 'subtotal': s } }
                productos_devueltos_agrupados = {}
                productos_nuevos_agrupados = {}
                
                # AGRUPAR PRODUCTOS DEVUELTOS (con precio negativo)
                # Filtrar solo detalles con producto_original (excluye productos adicionales)
                for item in cambio.detalles.filter(producto_original__isnull=False, cantidad_original__gt=0):
                    producto_talla = item.producto_original.ProductoTalla
                    pt_id = producto_talla.id
                    precio = abs(int(item.precio_original_unitario or 0))
                    
                    if pt_id in productos_devueltos_agrupados:
                        # Sumar cantidad al existente
                        productos_devueltos_agrupados[pt_id]['cantidad'] += item.cantidad_original
                        productos_devueltos_agrupados[pt_id]['subtotal'] += precio * item.cantidad_original
                    else:
                        productos_devueltos_agrupados[pt_id] = {
                            'producto': producto_talla,
                            'cantidad': item.cantidad_original,
                            'precio': precio,
                            'subtotal': precio * item.cantidad_original
                        }
                
                # AGRUPAR PRODUCTOS NUEVOS (con precio positivo)
                for item in cambio.detalles.all():
                    if item.producto_nuevo and item.cantidad_nueva and item.cantidad_nueva > 0:
                        producto_talla = item.producto_nuevo
                        pt_id = producto_talla.id
                        precio = int(item.precio_nuevo or producto_talla.producto.precioventa)
                        
                        if pt_id in productos_nuevos_agrupados:
                            # Sumar cantidad al existente
                            productos_nuevos_agrupados[pt_id]['cantidad'] += item.cantidad_nueva
                            productos_nuevos_agrupados[pt_id]['subtotal'] += precio * item.cantidad_nueva
                        else:
                            productos_nuevos_agrupados[pt_id] = {
                                'producto': producto_talla,
                                'cantidad': item.cantidad_nueva,
                                'precio': precio,
                                'subtotal': precio * item.cantidad_nueva
                            }
                
                # CREAR REGISTROS DE PRODUCTOS DEVUELTOS
                for pt_id, data in productos_devueltos_agrupados.items():
                    logger.debug(
                        "Producto devuelto en cambio %s: producto=%s, cantidad=%s",
                        cambio.id,
                        data['producto'].producto.articulo,
                        data['cantidad'],
                    )
                    Ticket_Productos.objects.create(
                        idTicket=ticket,
                        ProductoTalla=data['producto'],
                        stock=data['cantidad'],
                        precio=-data['precio'],  # Negativo
                        precio_original=-data['precio'],
                        descuento_unitario=0,
                        subtotal=-data['subtotal']
                    )
                
                # CREAR REGISTROS DE PRODUCTOS NUEVOS
                for pt_id, data in productos_nuevos_agrupados.items():
                    logger.debug(
                        "Producto nuevo en cambio %s: producto=%s, cantidad=%s",
                        cambio.id,
                        data['producto'].producto.articulo,
                        data['cantidad'],
                    )
                    Ticket_Productos.objects.create(
                        idTicket=ticket,
                        ProductoTalla=data['producto'],
                        stock=data['cantidad'],
                        precio=data['precio'],
                        precio_original=data['precio'],
                        descuento_unitario=0,
                        subtotal=data['subtotal']
                    )
                
                total_productos = len(productos_devueltos_agrupados) + len(productos_nuevos_agrupados)
                logger.info(
                    "Productos agregados al ticket de cambio %s: total_productos=%s, detalles=%s",
                    cambio.id,
                    total_productos,
                    cambio.detalles.count(),
                )
            except Exception as e:
                logger.exception("Error al agregar productos al ticket de cambio %s", cambio.id)
                raise
            
            # EJECUTAR MOVIMIENTOS DE INVENTARIO AUTOMÁTICAMENTE
            logger.debug("Ejecutando movimientos de inventario para cambio %s", cambio.id)
            
            try:
                # Primero ingresan las devoluciones aptas. Las no aptas quedan
                # solo en kardex y nunca se consideran stock disponible.
                for item in detalles_bloqueados:
                    if not item.producto_original_id or item.cantidad_original <= 0:
                        continue
                    producto_id = item.producto_original.ProductoTalla_id
                    producto_talla = productos_bloqueados[producto_id]
                    if item.apto_para_venta:
                        ingresar_inventario(
                            producto_talla=producto_talla,
                            cantidad=item.cantidad_original,
                            concepto='CAMBIO_PRODUCTO_ENTRADA',
                            responsable=request.user.username,
                            sucursal_destino=sucursal,
                            ticket=cambio.ticket_original,
                            costo_unitario=producto_talla.producto.costo or 0,
                            precio_unitario=int(item.precio_original_unitario),
                            observaciones=(
                                f'Devolucion apta - Cambio #{cambio.numero_operacion}. '
                                f'Condicion: {item.get_condicion_producto_display()}.'
                            ),
                            referencia_externa=cambio.numero_operacion,
                        )
                    else:
                        Movimientos_Producto.objects.create(
                            ProductoTalla=producto_talla,
                            tipo_movimiento='AJUSTE',
                            concepto='DEVOLUCION_NO_APTA',
                            cantidad=0,
                            responsable=request.user.username,
                            sucursal_destino=sucursal,
                            ticket=cambio.ticket_original,
                            precio=int(item.precio_original_unitario),
                            costo=0,
                            estado='COMPLETADO',
                            referencia_externa=cambio.numero_operacion,
                            observaciones=(
                                f'Devolucion NO APTA - Cambio #{cambio.numero_operacion}. '
                                'No se suma al inventario.'
                            ),
                        )

                # Luego salen los productos entregados. egresar_inventario
                # mantiene stock plano, FIFO y kardex en la misma transaccion.
                for item in detalles_bloqueados:
                    if not item.producto_nuevo_id or item.cantidad_nueva <= 0:
                        continue
                    egresar_inventario(
                        producto_talla=productos_bloqueados[item.producto_nuevo_id],
                        cantidad=item.cantidad_nueva,
                        concepto='CAMBIO_PRODUCTO_SALIDA',
                        responsable=request.user.username,
                        sucursal_origen=sucursal,
                        ticket=ticket,
                        precio_unitario=int(item.precio_nuevo),
                        observaciones=f'Entrega - Cambio #{cambio.numero_operacion}',
                        referencia_externa=cambio.numero_operacion,
                    )

                logger.info("Movimientos de inventario ejecutados para cambio %s", cambio.id)
            except Exception:
                logger.exception("Error en movimientos de inventario para cambio %s", cambio.id)
                raise
            # Vincular ticket nuevo al cambio
            logger.debug("Vinculando ticket %s al cambio %s", ticket.id, cambio.id)
            cambio.ticket_nuevo = ticket
            estado_anterior_cambio = cambio.estado

            # Determinar estado final según diferencia y estado del ticket
            diferencia = float(cambio.diferencia_monto)
            if ticket.estado == 'PENDIENTE' and diferencia > 0:
                cambio.estado = 'EJECUTADO_COBRO_PENDIENTE'
                cambio.fecha_ejecucion = timezone.now()
            elif ticket.estado == 'PENDIENTE' and diferencia < 0:
                cambio.estado = 'EJECUTADO_DEVOL_PENDIENTE'
                cambio.fecha_ejecucion = timezone.now()
            else:
                cambio.estado = 'COMPLETADO'
                cambio.fecha_ejecucion = timezone.now()
                cambio.fecha_completado = timezone.now()

            cambio.save()
            logger.info("Cambio %s actualizado: estado=%s, ticket_nuevo_id=%s", cambio.id, cambio.estado, ticket.id)
            
            # Crear historial
            logger.debug("Creando historial para cambio %s", cambio.id)
            accion_historial = 'APROBADO_Y_EJECUTADO'
            if cambio.estado == 'EJECUTADO_COBRO_PENDIENTE':
                accion_historial = 'EJECUTADO_COBRO_PENDIENTE'
            elif cambio.estado == 'EJECUTADO_DEVOL_PENDIENTE':
                accion_historial = 'EJECUTADO_DEVOL_PENDIENTE'

            HistorialCambioDevolucion.objects.create(
                cambio_devolucion=cambio,
                accion=accion_historial,
                estado_anterior=estado_anterior_cambio,
                estado_nuevo=cambio.estado,
                usuario=request.user,
                descripcion=f'Cambio aprobado y ejecutado por {request.user.get_full_name() or request.user.username}. Ticket #{nuevo_correlativo} generado. Movimientos de inventario realizados.',
                datos_adicionales={
                    'observaciones': observaciones,
                    'vendedor_id': vendedor_id,
                    'vendedor_nombre': vendedor_obj.nombre,
                    'vendedor_codigo': vendedor_obj.codigo_vendedor,
                    'ticket_generado': nuevo_correlativo,
                    'fecha_aprobacion': timezone.now().isoformat()
                }
            )
            logger.info(
                "Cambio ejecutado exitosamente: cambio=%s, numero_operacion=%s, estado=%s, "
                "ticket_correlativo=%s, ticket_id=%s, pendiente_tipo=%s, pendiente_monto=%s",
                cambio.id,
                cambio.numero_operacion,
                cambio.estado,
                nuevo_correlativo,
                ticket.id,
                ('Cobro' if diferencia > 0 else 'Devolucion') if cambio.estado != 'COMPLETADO' else '',
                abs(int(diferencia)) if cambio.estado != 'COMPLETADO' else 0,
            )
        
        # Construir datos del ticket para impresión
        ticket_data = construir_ticket_data(ticket)

        # ── Contexto adicional del cambio para el ticket visual ──────────────
        ticket_data['es_ticket_cambio'] = True
        ticket_data['numero_operacion']          = cambio.numero_operacion
        ticket_data['ticket_original_correlativo'] = (
            cambio.ticket_original.correlativo if cambio.ticket_original else None
        )
        ticket_data['tipo_operacion']  = cambio.tipo_operacion
        ticket_data['tipo_operacion_display'] = cambio.get_tipo_operacion_display()
        ticket_data['monto_original']  = int(cambio.monto_original)
        ticket_data['monto_nuevo']     = int(cambio.monto_nuevo)
        ticket_data['diferencia_monto'] = int(cambio.diferencia_monto)

        return JsonResponse({
            'success': True,
            'message': 'Cambio aprobado, ticket generado e inventario actualizado',
            'ticket_id': ticket.id,
            'ticket_correlativo': nuevo_correlativo,
            'diferencia_cobrar': cambio.diferencia_monto,
            'nuevo_estado': cambio.estado,
            'nuevo_estado_display': cambio.get_estado_display(),
            'cobro_pendiente': cambio.cobro_pendiente,
            'devolucion_pendiente': cambio.devolucion_pendiente,
            'ticket_data': ticket_data
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except ConflictoInventarioCambio as e:
        return JsonResponse({
            'success': False,
            'code': e.codigo,
            'error': str(e),
        }, status=409)
    except Exception as e:
        logger.exception("Error al aprobar cambio y generar ticket")
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@login_required
@require_POST
def validar_codigo_vendedor(request):
    """Valida un vendedor dentro del alcance del cambio solicitado."""
    try:
        data = json.loads(request.body)
        codigo = str(data.get('codigo') or '').strip()
        cambio_id = data.get('cambio_id')

        if not codigo or not cambio_id:
            return JsonResponse({
                'success': False,
                'code': 'SELLER_CONTEXT_REQUIRED',
                'error': 'Codigo de vendedor e identificador del cambio requeridos',
            }, status=400)

        cambio = CambioDevolucion.objects.select_related('sucursal__empresa').filter(
            id=cambio_id
        ).first()
        if not cambio:
            return JsonResponse({
                'success': False,
                'code': 'CHANGE_NOT_FOUND',
                'error': 'Cambio no encontrado',
            }, status=404)

        sucursal_id = get_sucursal_id(request)
        if not sucursal_id or cambio.sucursal_id != int(sucursal_id):
            return JsonResponse({
                'success': False,
                'code': 'BRANCH_DENIED',
                'error': 'No tiene acceso a este cambio',
            }, status=403)

        candidatos = list(
            _vendedores_para_autorizacion_cambio(cambio.sucursal)
            .filter(codigo_vendedor=codigo)
            .order_by('id')[:2]
        )
        if not candidatos:
            return JsonResponse({
                'success': False,
                'code': 'SELLER_NOT_AVAILABLE',
                'error': 'El vendedor no esta activo',
            }, status=404)
        if len(candidatos) > 1:
            return JsonResponse({
                'success': False,
                'code': 'SELLER_CODE_AMBIGUOUS',
                'error': 'El codigo corresponde a mas de un vendedor. Corrija la configuracion.',
            }, status=409)

        vendedor_obj = candidatos[0]
        return JsonResponse({
            'success': True,
            'vendedor': {
                'id': vendedor_obj.id,
                'nombre_completo': vendedor_obj.nombre or f'Vendedor {vendedor_obj.codigo_vendedor}',
                'codigo': vendedor_obj.codigo_vendedor,
            },
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON invalidos',
        }, status=400)
    except Exception:
        logger.exception("Error al validar vendedor para cambio/devolucion")
        return JsonResponse({
            'success': False,
            'error': 'No fue posible validar el vendedor',
        }, status=500)

# ========== CÓDIGOS DE AUTORIZACIÓN DINÁMICOS ==========

@login_required
def obtener_codigo_autorizacion_actual(request):
    """
    Obtiene el código de autorización dinámico actual.
    Solo accesible para usuarios con rol 'administrador' o 'jefe_local'
    """
    try:
        from .models import CodigoAutorizacionDinamico
        
        # Verificar que el usuario tenga el rol apropiado
        rol_usuario = getattr(request.user, 'rol', None)
        
        if rol_usuario not in ['administrador', 'jefe_local']:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para acceder a los códigos de autorización',
                'requiere_rol': 'Administrador o Jefe Local'
            }, status=403)
        
        # Obtener o generar el código actual para este supervisor
        codigo_obj = CodigoAutorizacionDinamico.obtener_codigo_actual(request.user)
        
        if not codigo_obj:
            return JsonResponse({
                'success': False,
                'error': 'No se pudo generar el código de autorización'
            }, status=500)
        
        # Calcular tiempo restante usando hora de Chile
        import pytz
        ahora_utc = timezone.now()
        chile_tz = pytz.timezone('America/Santiago')
        ahora = ahora_utc.astimezone(chile_tz)
        tiempo_restante = codigo_obj.fecha_hora_fin - ahora
        minutos_restantes = int(tiempo_restante.total_seconds() / 60)
        
        return JsonResponse({
            'success': True,
            'codigo': {
                'codigo': codigo_obj.codigo,
                'valido_desde': codigo_obj.fecha_hora_inicio.strftime('%H:%M'),
                'valido_hasta': codigo_obj.fecha_hora_fin.strftime('%H:%M'),
                'minutos_restantes': minutos_restantes,
                'fecha_actual': ahora.strftime('%d/%m/%Y %H:%M:%S')
            }
        })
        
    except Exception as e:
        logger.exception("Error al obtener codigo de autorizacion")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener código: {str(e)}'
        }, status=500)


@login_required
@require_POST
def validar_codigo_autorizacion(request):
    """
    Valida un código de autorización dinámico ingresado por el usuario.
    Incluye protección anti brute-force: máximo 5 intentos fallidos en 15 minutos.
    """
    try:
        from .models import CodigoAutorizacionDinamico, RegistroAutorizacion
        
        # Protección anti brute-force: max 5 intentos fallidos en 15 minutos
        hace_15_min = timezone.now() - timezone.timedelta(minutes=15)
        intentos_fallidos = RegistroAutorizacion.objects.filter(
            usuario_solicitante=request.user,
            exitoso=False,
            fecha_hora__gte=hace_15_min
        ).count()
        
        if intentos_fallidos >= 5:
            return JsonResponse({
                'success': False,
                'error': 'Demasiados intentos fallidos. Intente nuevamente en 15 minutos.'
            }, status=429)
        
        data = json.loads(request.body)
        codigo_ingresado = data.get('codigo', '').strip()
        tipo_operacion = data.get('tipo_operacion', 'APROBACION_CAMBIO')
        cambio_id = data.get('cambio_id', None)
        
        if not codigo_ingresado:
            return JsonResponse({
                'success': False,
                'error': 'Debe ingresar un código de autorización'
            })
        
        # Validar el código
        es_valido, mensaje, codigo_obj = CodigoAutorizacionDinamico.validar_codigo(codigo_ingresado)
        
        # Registrar el intento de autorización con trazabilidad del supervisor
        try:
            cambio_obj = None
            if cambio_id:
                cambio_obj = CambioDevolucion.objects.get(id=cambio_id)
            
            supervisor = codigo_obj.generado_por if (es_valido and codigo_obj) else None
            
            registro = RegistroAutorizacion.objects.create(
                codigo_usado=codigo_obj if es_valido else None,
                usuario_solicitante=request.user,
                usuario_autorizador=supervisor,
                tipo_operacion=tipo_operacion,
                descripcion=f"{'Autorización exitosa' if es_valido else 'Intento fallido'}: {mensaje}",
                ip_origen=request.META.get('REMOTE_ADDR'),
                exitoso=es_valido,
                cambio_devolucion=cambio_obj,
                datos_adicionales={
                    'codigo_ingresado': codigo_ingresado,
                    'mensaje': mensaje,
                    'supervisor_id': supervisor.id if supervisor else None,
                    'supervisor_nombre': supervisor.get_full_name() if supervisor else None
                }
            )
        except Exception:
            logger.exception("Error al registrar intento de autorizacion")
        
        if not es_valido:
            return JsonResponse({
                'success': False,
                'error': mensaje
            })
        
        # Marcar el código como usado (un solo uso por código)
        codigo_obj.marcar_como_usado()
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Código de autorización validado correctamente',
            'codigo': {
                'codigo': codigo_obj.codigo,
                'valido_hasta': codigo_obj.fecha_hora_fin.strftime('%H:%M'),
                'supervisor': codigo_obj.generado_por.get_full_name() if codigo_obj.generado_por else None
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        logger.exception("Error al validar codigo de autorizacion")
        return JsonResponse({
            'success': False,
            'error': f'Error al validar código: {str(e)}'
        })


@login_required
@require_POST
def completar_cambio_devolucion(request):
    """Ruta legacy deshabilitada: completar requiere el flujo unificado."""
    return JsonResponse({
        'success': False,
        'code': 'LEGACY_FLOW_DISABLED',
        'error': (
            'El flujo separado de completado fue deshabilitado. '
            'Use la autorizacion y ejecucion unificada con codigo de Administrador.'
        ),
    }, status=410)


@login_required
@require_GET
def buscar_documento_cambio(request):
    """Buscar documento (Ticket o DTE) para iniciar proceso de cambio/devolución"""
    try:
        numero = request.GET.get('numero', '').strip()
        tipo_documento = request.GET.get('tipo_documento', 'dte')
        fecha_compra = request.GET.get('fecha_compra', '').strip()
        tipo_dte = request.GET.get('tipo_dte', '').strip()  # 33, 39, 34, etc.
        
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not numero:
            return JsonResponse({
                'success': False,
                'error': 'Número de documento requerido'
            })
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        # Buscar según el tipo
        if tipo_documento == 'dte':
            # Buscar DTE - check count for informative errors
            query_debug = Dte.objects.filter(numero_documento=numero)
            
            if not query_debug.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'DTE #{numero} no encontrado en el sistema.'
                })
            
            # Buscar DTE con filtros para ventas
            query = Dte.objects.select_related('vendedor', 'receptor', 'sucursal').prefetch_related(
                'dte_productos__productoTalla__producto'
            ).filter(
                numero_documento=numero,
                sucursal_id=sucursal_id
            )
            
            # Filtrar por tipo de transacción (solo ventas)
            query = query.filter(tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'])
            
            # Filtrar por tipo de DTE específico si se proporcionó
            if tipo_dte:
                query = query.filter(tipo_documento=tipo_dte)
            
            if fecha_compra:
                query = query.filter(fecha_emision=fecha_compra)
            
            dte = query.first()
            
            if not dte:
                # Verificar primero si existe el DTE con el número
                dte_existe = Dte.objects.filter(numero_documento=numero).first()
                
                if not dte_existe:
                    return JsonResponse({
                        'success': False,
                        'error': f'DTE #{numero} no encontrado en el sistema.'
                    })
                
                # Verificar si existe en otra sucursal
                if dte_existe.sucursal_id != sucursal_id:
                    return JsonResponse({
                        'success': False,
                        'error': f'DTE #{numero} pertenece a otra sucursal ({dte_existe.sucursal.alias}). Solo puede procesar documentos de la sucursal actual.'
                    })
                
                # El DTE existe, verificar por qué no pasó los filtros
                # Verificar tipo de transacción
                if dte_existe.tipo_transaccion not in ['VENTA', 'VENTA_PUBLICO']:
                    return JsonResponse({
                        'success': False,
                        'error': f'DTE #{numero} es tipo "{dte_existe.tipo_transaccion}". Solo se permiten cambios de documentos de VENTA o VENTA_PUBLICO.'
                    })
                
                # Si llegó aquí, el tipo es correcto pero otros filtros no coinciden
                # Verificar si hay múltiples DTEs con ese número
                query_disponibles = Dte.objects.filter(
                    numero_documento=numero,
                    sucursal_id=sucursal_id,
                    tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO']
                )
                
                if fecha_compra and tipo_dte:
                    # Usuario especificó fecha y tipo, pero no se encontró
                    return JsonResponse({
                        'success': False,
                        'error': f'No se encontró DTE #{numero} tipo {tipo_dte} con fecha {fecha_compra}. Verifique los datos.'
                    })
                
                count_dtes = query_disponibles.count()
                
                if count_dtes > 1:
                    # Hay múltiples DTEs, crear tabla informativa
                    dtes_info = []
                    tipo_dte_nombres = {
                        '33': 'Factura Electrónica',
                        '34': 'Factura Exenta',
                        '39': 'Boleta Electrónica',
                        '41': 'Boleta Exenta',
                        '61': 'Nota de Crédito',
                        '56': 'Nota de Débito'
                    }
                    
                    for d in query_disponibles.all():
                        tipo_nombre = tipo_dte_nombres.get(d.tipo_documento, f'Tipo {d.tipo_documento}')
                        dtes_info.append({
                            'tipo': d.tipo_documento,
                            'tipo_nombre': tipo_nombre,
                            'fecha': d.fecha_emision.strftime('%d/%m/%Y'),
                            'monto': f'${int(d.monto_con_iva):,}'
                        })
                    
                    # Crear mensaje con tabla HTML
                    tabla_html = '<div class="table-responsive"><table class="table table-sm table-bordered">'
                    tabla_html += '<thead><tr><th>Tipo DTE</th><th>Código</th><th>Fecha</th><th>Monto</th></tr></thead><tbody>'
                    
                    for info in dtes_info:
                        tabla_html += f'<tr><td><strong>{info["tipo_nombre"]}</strong></td><td>{info["tipo"]}</td><td>{info["fecha"]}</td><td>{info["monto"]}</td></tr>'
                    
                    tabla_html += '</tbody></table></div>'
                    
                    return JsonResponse({
                        'success': False,
                        'error': f'Se encontraron {count_dtes} documentos con el número #{numero}',
                        'multiple_documents': True,
                        'documents_html': tabla_html,
                        'message': 'Por favor, seleccione el tipo de DTE específico (Boleta o Factura) en el formulario de búsqueda.'
                    })
                else:
                    # Un solo DTE pero no coincide con la fecha
                    return JsonResponse({
                        'success': False,
                        'error': f'DTE #{numero} encontrado con fecha {dte_existe.fecha_emision.strftime("%d/%m/%Y")}, pero usted buscó con fecha {fecha_compra}. Corrija la fecha de compra.'
                    })
            
            # Verificar que esté emitido/pagado
            if dte.estado_dte not in ['EMITIDO', 'ACEPTADO']:
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se pueden procesar cambios de documentos emitidos'
                })
            
            # Crear o buscar ticket de referencia para el DTE
            from datetime import timedelta
            fecha_limite = dte.fecha_emision + timedelta(days=30)
            dentro_del_plazo = timezone.localdate() <= fecha_limite
            
            # Intentar encontrar el ticket ORIGINAL del POS (fuente con descuentos correctos)
            ticket_original_pos = None
            if dte.referencias and 'TICKET-' in dte.referencias:
                try:
                    corr_original = dte.referencias.split('TICKET-')[1].strip().split()[0]
                    ticket_original_pos = Ticket.objects.filter(
                        correlativo=corr_original,
                        sucursal_id=sucursal_id,
                        estado='PAGADO'
                    ).first()
                except Exception:
                    pass

            # Si encontramos el ticket original del POS, usarlo directamente
            if ticket_original_pos:
                ticket_referencia = ticket_original_pos
            else:
                # Buscar si ya existe un ticket de referencia asociado a este DTE
                ticket_referencia = Ticket.objects.filter(
                    observaciones__icontains=f'DTE #{dte.numero_documento}'
                ).first()
                
                if ticket_referencia:
                    # Repair: sync discounts from DTE or original ticket
                    for dp in dte.dte_productos.all():
                        if not dp.productoTalla:
                            continue
                        dcto_unit = 0
                        if dp.descuento_monto and dp.stock and dp.stock > 0:
                            dcto_unit = int(dp.descuento_monto / dp.stock)
                        if dcto_unit > 0:
                            tp_ref = ticket_referencia.ticket_productos.filter(
                                ProductoTalla=dp.productoTalla,
                                stock=dp.stock,
                                descuento_unitario=0
                            ).first()
                            if tp_ref:
                                tp_ref.descuento_unitario = dcto_unit
                                tp_ref.subtotal = (dp.precio - dcto_unit) * dp.stock
                                tp_ref.porcentaje_descuento = dp.descuento_pct or 0
                                tp_ref.save()
            
            if not ticket_referencia:
                # Crear ticket de referencia
                from .views import obtener_siguiente_correlativo
                correlativo_ticket = obtener_siguiente_correlativo(dte.sucursal, 'TICKET')
                
                ticket_referencia = Ticket.objects.create(
                    correlativo=correlativo_ticket,
                    vendedor=dte.vendedor,
                    sucursal=dte.sucursal,
                    subTotal=int(dte.monto_neto),
                    descuento=int(dte.descuento) if dte.descuento else 0,
                    total=int(dte.monto_con_iva),
                    estado='PAGADO',
                    responsable=dte.responsable,
                    cliente_nombre=dte.receptor.razon_social if dte.receptor else '',
                    cliente_rut=dte.receptor.rut if dte.receptor else '',
                    cliente_email=dte.receptor.correoVendedor if dte.receptor else '',
                    cliente_telefono='',
                    cliente_giro=dte.receptor.giro if dte.receptor else '',
                    cliente_direccion=dte.receptor.direccion if dte.receptor else '',
                    cliente_comuna=dte.receptor.comuna if dte.receptor else '',
                    cliente_ciudad=dte.receptor.ciudad if dte.receptor else '',
                    observaciones=f'Ticket de referencia para DTE #{dte.numero_documento} - {dte.tipo_documento}'
                )
                
                # Copiar productos del DTE al ticket (con descuentos)
                es_boleta = dte.tipo_documento in ['39', '41', 'BOLETA ELECTRONICA', 'BOLETA EXENTA']
                for dp in dte.dte_productos.all():
                    precio_lista = dp.precio
                    dcto_unitario = 0

                    if dp.descuento_monto and dp.stock and dp.stock > 0:
                        dcto_unitario = int(dp.descuento_monto / dp.stock)
                    elif es_boleta and dp.monto_item and dp.stock and dp.stock > 0:
                        precio_efectivo_por_unidad = int(dp.monto_item / dp.stock)
                        if precio_efectivo_por_unidad < dp.precio:
                            dcto_unitario = dp.precio - precio_efectivo_por_unidad

                    subtotal = (precio_lista - dcto_unitario) * dp.stock

                    Ticket_Productos.objects.create(
                        idTicket=ticket_referencia,
                        ProductoTalla=dp.productoTalla,
                        stock=dp.stock,
                        precio=precio_lista,
                        precio_original=precio_lista,
                        descuento_unitario=dcto_unitario,
                        subtotal=subtotal,
                        porcentaje_descuento=dp.descuento_pct or 0,
                        descripcion_linea=dp.descripcion if not dp.productoTalla else None,
                        es_pendiente_despacho=dp.es_pendiente_despacho,
                    )
                
            # Un cambio anterior NO reemplaza la venta: su `ticket_nuevo` solo trae
            # el delta (lo devuelto en negativo + lo entregado). Se muestra la venta
            # completa MÁS los artículos de reemplazo, así lo que no se cambió sigue
            # disponible y lo ya cambiado se puede volver a cambiar.
            ticket_referencia = _ticket_raiz_cambio(ticket_referencia)
            ticket_original_correlativo = ticket_referencia.correlativo

            tickets_cadena, cambios_cadena = _cadena_cambios_ticket(ticket_referencia)
            productos_data, productos_disponibles_count = _productos_cambio_data(
                tickets_cadena, cambios_cadena
            )
            cambios_anteriores = _historial_cambios_data(cambios_cadena)

            return JsonResponse({
                'success': True,
                'documento': {
                    'id': ticket_referencia.id,
                    'tipo': 'DTE',
                    'numero_documento': dte.numero_documento,
                    'tipo_documento': dte.tipo_documento,
                    'correlativo': ticket_referencia.correlativo,
                    'correlativo_original': ticket_original_correlativo,
                    'fue_redirigido': False,
                    'fecha': dte.fecha_emision.strftime('%d/%m/%Y'),
                    'total': float(dte.monto_con_iva),
                    'vendedor': dte.vendedor.nombre if dte.vendedor else 'Sin vendedor',
                    'cliente_nombre': dte.receptor.razon_social if dte.receptor else 'Sin nombre',
                    'cliente_rut': dte.receptor.rut if dte.receptor else '',
                    'fecha_limite_cambio': fecha_limite.strftime('%d/%m/%Y'),
                    'dentro_del_plazo': dentro_del_plazo,
                    'dias_transcurridos': (timezone.localdate() - dte.fecha_emision).days,
                    'productos': productos_data,
                    'productos_disponibles': productos_disponibles_count,
                    'cambios_anteriores': cambios_anteriores,
                    'tiene_cambios_previos': bool(cambios_anteriores),
                    'puede_cambiar': productos_disponibles_count > 0,
                }
            })
        elif tipo_documento == 'ticket':
            # Buscar directamente por correlativo de ticket
            ticket = Ticket.objects.filter(
                correlativo=numero,
                sucursal_id=sucursal_id
            ).first()
            
            if not ticket:
                return JsonResponse({
                    'success': False,
                    'error': f'Ticket #{numero} no encontrado en esta sucursal'
                })
            
            return buscar_ticket_para_cambio_response(ticket, request)
        
        elif tipo_documento == 'ticket_cambio':
            # Buscar por Ticket de Cambio (número del ticket original)
            
            # El ticket de cambio puede tener varios formatos:
            # 1. Nuevo formato: TC-{SUCURSAL}-{TICKET}-{FECHA} (ej: TC-SUC1-123-250120)
            # 2. Formato anterior: TC-{TICKET} (ej: TC-123)
            # 3. Solo número: 123
            numero_limpio = numero.upper().strip()
            numero_ticket = None
            fecha_extraida = None
            
            if numero_limpio.startswith('TC-'):
                partes = numero_limpio.split('-')
                if len(partes) >= 4:
                    # Nuevo formato: TC-SUCURSAL-TICKET-FECHA
                    # TC-SUC1-123-250120 → ticket=123, fecha=2025-01-20
                    numero_ticket = partes[2]  # El tercer elemento es el número de ticket
                    # Extraer fecha del formato YYMMDD
                    if len(partes[3]) == 6:
                        try:
                            fecha_str = partes[3]  # YYMMDD
                            year = 2000 + int(fecha_str[0:2])
                            month = int(fecha_str[2:4])
                            day = int(fecha_str[4:6])
                            fecha_extraida = f"{year}-{month:02d}-{day:02d}"
                            logger.debug("Fecha extraida del codigo de cambio: %s", fecha_extraida)
                        except:
                            pass
                elif len(partes) == 2:
                    # Formato anterior: TC-123
                    numero_ticket = partes[1]
                else:
                    numero_ticket = numero_limpio.replace('TC-', '')
            else:
                # Solo número
                numero_ticket = numero_limpio
            
            # Si se extrajo fecha del código y no se proporcionó fecha_compra, usarla
            if fecha_extraida and not fecha_compra:
                fecha_compra = fecha_extraida
                logger.debug("Usando fecha extraida del codigo de cambio: %s", fecha_compra)
            
            # Buscar el ticket original
            ticket_query = Ticket.objects.select_related(
                'vendedor', 'sucursal'
            ).prefetch_related(
                'ticket_productos__ProductoTalla__producto',
                'cambios_devoluciones'
            ).filter(
                correlativo=numero_ticket,
                sucursal_id=sucursal_id
            )
            
            if fecha_compra:
                ticket_query = ticket_query.filter(fecha=fecha_compra)
            
            ticket = ticket_query.first()
            
            if not ticket:
                # Buscar sin filtro de fecha para dar mejor mensaje
                ticket_sin_fecha = Ticket.objects.filter(
                    correlativo=numero_ticket,
                    sucursal_id=sucursal_id
                ).first()
                
                if ticket_sin_fecha:
                    return JsonResponse({
                        'success': False,
                        'error': f'Ticket de Cambio #{numero_ticket} encontrado con fecha {ticket_sin_fecha.fecha.strftime("%d/%m/%Y")}, pero usted buscó con fecha {fecha_compra}. Corrija la fecha.'
                    })
                
                return JsonResponse({
                    'success': False,
                    'error': f'Ticket de Cambio #{numero_ticket} no encontrado. Verifique el número del ticket y la sucursal.'
                })
            
            # Subir a la venta original: el ticket de cambio solo trae el delta
            ticket_original_correlativo = ticket.correlativo
            ticket_raiz = _ticket_raiz_cambio(ticket)
            fue_redirigido = (ticket_raiz.id != ticket.id)
            ticket = ticket_raiz

            # Verificar estado y plazo - permitir PAGADO y PENDIENTE (tickets de cambio pueden estar pendientes)
            if ticket.estado not in ('PAGADO', 'PENDIENTE'):
                return JsonResponse({
                    'success': False,
                    'error': f'El ticket referenciado está en estado "{ticket.get_estado_display()}". No se puede procesar.'
                })

            from datetime import timedelta
            fecha_limite = ticket.fecha + timedelta(days=30) if ticket.fecha else timezone.localdate() + timedelta(days=30)
            dentro_del_plazo = timezone.localdate() <= fecha_limite

            # Productos de toda la cadena (venta original + reemplazos de cambios previos)
            tickets_cadena, cambios_cadena = _cadena_cambios_ticket(ticket)
            productos_data, productos_disponibles_count = _productos_cambio_data(
                tickets_cadena, cambios_cadena
            )
            cambios_anteriores = _historial_cambios_data(cambios_cadena)

            return JsonResponse({
                'success': True,
                'documento': {
                    'id': ticket.id,
                    'tipo': 'TICKET_CAMBIO',
                    'correlativo': ticket.correlativo,
                    'correlativo_original': ticket_original_correlativo,
                    'fue_redirigido': fue_redirigido,
                    'fecha': ticket.fecha.strftime('%d/%m/%Y'),
                    'hora': ticket.hora.strftime('%H:%M') if ticket.hora else '',
                    'total': float(ticket.total),
                    'estado': ticket.estado,
                    'vendedor': ticket.vendedor.nombre if ticket.vendedor else '',
                    'cliente_nombre': ticket.cliente_nombre or '',
                    'cliente_rut': ticket.cliente_rut or '',
                    'fecha_limite_cambio': fecha_limite.strftime('%d/%m/%Y'),
                    'dentro_del_plazo': dentro_del_plazo,
                    'dias_transcurridos': (timezone.localdate() - ticket.fecha).days,
                    'productos': productos_data,
                    'productos_disponibles': productos_disponibles_count,
                    'cambios_anteriores': cambios_anteriores,
                    'tiene_cambios_previos': bool(cambios_anteriores),
                    'puede_cambiar': productos_disponibles_count > 0,
                }
            })

        else:
            # Buscar Ticket (lógica existente)
            return buscar_ticket_para_cambio_original(request, numero, fecha_compra, sucursal_id)
        
    except Exception as e:
        logger.exception("Error en buscar_documento_cambio")
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar documento: {str(e)}'
        })


def buscar_ticket_para_cambio_original(request, correlativo, fecha_compra, sucursal_id):
    """Función para buscar tickets (llamada desde buscar_documento_cambio)"""
    query = Ticket.objects.select_related(
        'vendedor', 'sucursal'
    ).prefetch_related(
        'ticket_productos__ProductoTalla__producto',
        'cambios_devoluciones'
    ).filter(
        correlativo=correlativo,
        sucursal_id=sucursal_id
    )
    
    if fecha_compra:
        query = query.filter(fecha=fecha_compra)
    
    ticket = query.first()
    
    if not ticket:
        return JsonResponse({
            'success': False,
            'error': f'Ticket #{correlativo} no encontrado' + (f' para la fecha {fecha_compra}' if fecha_compra else '')
        })
    
    # La respuesta resuelve sola la cadena de cambios (ver
    # buscar_ticket_para_cambio_response): siempre se trabaja sobre la venta original.
    return buscar_ticket_para_cambio_response(ticket, request)


def buscar_ticket_para_cambio_response(ticket, request):
    """Genera la respuesta con datos del ticket para cambio"""
    # Si lo buscado es el comprobante de un cambio, subir a la venta original: ese
    # ticket solo trae el delta y por sí solo esconde el resto de la compra.
    ticket_buscado = ticket
    ticket = _ticket_raiz_cambio(ticket)

    # Verificar que esté pagado
    if ticket.estado != 'PAGADO':
        return JsonResponse({
            'success': False,
            'error': 'Solo se pueden procesar cambios de tickets pagados'
        })

    # Verificar plazo
    from datetime import timedelta
    fecha_limite = ticket.fecha + timedelta(days=30)
    dentro_del_plazo = timezone.localdate() <= fecha_limite

    # Productos de toda la cadena: la venta original + lo que entró por cambios previos
    tickets_cadena, cambios_cadena = _cadena_cambios_ticket(ticket)
    productos_data, productos_disponibles_count = _productos_cambio_data(
        tickets_cadena, cambios_cadena
    )
    cambios_anteriores = _historial_cambios_data(cambios_cadena)

    # Se avisa solo si hubo que subir desde el comprobante de un cambio
    fue_redirigido = ticket_buscado.id != ticket.id

    ticket_data = {
        'id': ticket.id,
        'tipo': 'TICKET',
        'correlativo': ticket.correlativo,
        'correlativo_original': ticket_buscado.correlativo,
        'fue_redirigido': fue_redirigido,
        'fecha': ticket.fecha.strftime('%d/%m/%Y'),
        'hora': ticket.hora.strftime('%H:%M') if ticket.hora else '',
        'total': float(ticket.total),
        'estado': ticket.estado,
        'vendedor': ticket.vendedor.nombre if ticket.vendedor else '',
        'cliente_nombre': ticket.cliente_nombre or '',
        'cliente_rut': ticket.cliente_rut or '',
        'cliente_email': ticket.cliente_email or '',
        'cliente_telefono': ticket.cliente_telefono or '',
        'observaciones': ticket.observaciones or '',
        'fecha_limite_cambio': fecha_limite.strftime('%d/%m/%Y'),
        'dentro_del_plazo': dentro_del_plazo,
        'dias_transcurridos': (timezone.localdate() - ticket.fecha).days,
        'productos': productos_data,
        'productos_disponibles': productos_disponibles_count,
        'cambios_anteriores': cambios_anteriores,
        'tiene_cambios_previos': bool(cambios_anteriores),
        'puede_cambiar': productos_disponibles_count > 0,
    }

    return JsonResponse({
        'success': True,
        'documento': ticket_data
        })


@login_required
@require_GET
def buscar_ticket_para_cambio(request):
    """Buscar ticket para iniciar proceso de cambio/devolución.

    Busca por el FOLIO del DTE impreso en la boleta (Ticket.folio_dte) y, como
    respaldo tolerante, por el correlativo interno. Acotado a la sucursal actual
    (dentro de una sucursal el folio es único, así que no hace falta fecha ni
    desambiguar entre sucursales)."""
    try:
        correlativo = request.GET.get('correlativo', '').strip()
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        if not correlativo:
            return JsonResponse({
                'success': False,
                'error': 'Número de boleta requerido'
            })

        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })

        try:
            num = int(correlativo)
        except (TypeError, ValueError):
            return JsonResponse({
                'success': False,
                'error': f'Número de boleta inválido: {correlativo}'
            })

        query = Ticket.objects.select_related(
                'vendedor', 'sucursal'
            ).prefetch_related(
                'ticket_productos__ProductoTalla__producto',
                'cambios_devoluciones'
        ).filter(
                Q(folio_dte=num) | Q(correlativo=num),
                sucursal_id=sucursal_id
            ).order_by('-id')

        ticket = query.first()

        if not ticket:
            return JsonResponse({
                'success': False,
                'error': f'Boleta N° {correlativo} no encontrada en esta sucursal'
            })

        return buscar_ticket_para_cambio_response(ticket, request)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar ticket: {str(e)}'
        })


@login_required
@require_GET
def buscar_productos_para_cambio(request):
    """Buscar productos disponibles para cambio"""
    try:
        termino = request.GET.get('termino', '').strip()
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not termino:
            return JsonResponse({
                'success': False,
                'error': 'Término de búsqueda requerido'
            })
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal seleccionada'
            })
        
        # Buscar productos con stock disponible
        productos_query = Producto_Talla.objects.select_related(
            'producto', 'producto__atributo1', 'producto__atributo2'
        ).filter(
            Q(sku__icontains=termino) |
            Q(producto__articulo__icontains=termino) |
            Q(producto__descripcion__icontains=termino) |
            Q(producto__atributo1__valor__icontains=termino) |
            Q(producto__atributo2__valor__icontains=termino),
            producto__sucursal_id=sucursal_id,
            stock__gt=0
        )[:100]  # Límite ampliado para cubrir marcas con muchos modelos/tallas
        
        productos_data = []
        for pt in productos_query:
            productos_data.append({
                'id': pt.id,
                'sku': pt.sku,
                'articulo': pt.producto.articulo,
                'descripcion': pt.producto.descripcion,
                'talla': pt.talla,
                'precio_venta': float(pt.producto.precioventa),
                'stock_disponible': pt.stock,
                'marca': pt.producto.atributo1.valor if pt.producto.atributo1 else '',
                'color': pt.producto.atributo2.valor if pt.producto.atributo2 else '',
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        })


# ========== GESTIÓN DE CLIENTES POS ==========

@require_POST
@login_required
def guardar_cliente_pos(request):
    """Guardar datos del cliente desde el POS / flujo NCN de gestion-DTE.

    Usa la tabla `Empresa` (no `Cliente` CRM). Si encuentra un registro con el
    mismo RUT lo actualiza; si no, crea uno nuevo siempre que llegue RUT.
    """
    import logging as _logging
    _log = _logging.getLogger(__name__)

    try:
        try:
            data = json.loads(request.body or b'{}')
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            }, status=400)

        nombre = (data.get('nombre') or '').strip()
        rut_raw = (data.get('rut') or '').strip()
        email = (data.get('email') or '').strip()
        telefono = (data.get('telefono') or '').strip()
        tipo_documento = data.get('tipo_documento', 'BOLETA_ELECTRONICA')

        giro = (data.get('giro') or '').strip()
        direccion = (data.get('direccion') or '').strip()
        comuna = (data.get('comuna') or '').strip()
        ciudad = (data.get('ciudad') or '').strip()
        telefono_secundario = (data.get('telefono_secundario') or '').strip()
        email_facturacion = (data.get('email_facturacion') or '').strip()
        fecha_nacimiento_str = (data.get('fecha_nacimiento') or '').strip()
        celular_fid = (data.get('celular') or telefono_secundario or '').strip()

        if not nombre and not rut_raw:
            return JsonResponse({
                'success': False,
                'error': 'Debe proporcionar al menos nombre o RUT'
            }, status=400)

        # Normalizar RUT: si viene sin guión lo formateamos para que matchee
        # con registros existentes y respete el regex validator de Empresa.rut
        rut = formatear_rut(rut_raw) if rut_raw else ''

        cliente = None
        if rut:
            cliente = Empresa.objects.filter(
                Q(rut__iexact=rut) | Q(rut__iexact=rut_raw)
            ).order_by('-id').first()

        if cliente:
            if nombre:
                cliente.nombre = nombre
                cliente.razon_social = nombre

            if email or email_facturacion:
                cliente.correoVendedor = email or cliente.correoVendedor or ''
                cliente.correoAdministrador = email_facturacion or cliente.correoAdministrador or ''

            if telefono or telefono_secundario:
                cliente.contacto1 = telefono or cliente.contacto1 or ''
                cliente.contacto2 = telefono_secundario or cliente.contacto2 or ''

            if tipo_documento == 'FACTURA_ELECTRONICA' or giro or direccion or comuna or ciudad:
                if giro:
                    cliente.giro = giro
                if direccion:
                    cliente.direccion = direccion
                if comuna:
                    cliente.comuna = comuna
                if ciudad:
                    cliente.ciudad = ciudad

            cliente.save()
            _log.info("Cliente actualizado (Empresa ID %s) - giro: %s", cliente.id, cliente.giro)
        else:
            if not rut:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe proporcionar un RUT para crear un nuevo cliente'
                }, status=400)

            cliente = Empresa.objects.create(
                nombre=nombre or f'Cliente {rut}',
                rut=rut,
                nombre_fantasia=nombre or '',
                razon_social=nombre or '',
                giro=giro,
                direccion=direccion,
                comuna=comuna,
                ciudad=ciudad,
                esProveedor=False,
                correoVendedor=email or '',
                correoAdministrador=email_facturacion or '',
                correoIntercambio='',
                contacto1=telefono or '',
                contacto2=telefono_secundario or '',
            )
            _log.info("Cliente creado (Empresa ID %s) - giro: %s", cliente.id, cliente.giro)

        # Sync fecha_nacimiento y celular al CRM (tabla Cliente) para personas naturales.
        # La tabla Empresa no tiene esos campos; los guardamos en el CRM para fidelización.
        if (fecha_nacimiento_str or celular_fid) and rut:
            try:
                from .services.fidelizacion_service import es_rut_empresa as _is_emp
                if not _is_emp(rut):
                    from app.models import Cliente as _Cliente
                    from datetime import date as _date
                    crm_cli, _ = _Cliente.objects.get_or_create(
                        rut=rut,
                        defaults={'nombre': nombre or rut, 'activo': True}
                    )
                    _changed = []
                    if fecha_nacimiento_str and not crm_cli.fecha_nacimiento:
                        try:
                            crm_cli.fecha_nacimiento = _date.fromisoformat(fecha_nacimiento_str)
                            _changed.append('fecha_nacimiento')
                        except ValueError:
                            pass
                    if celular_fid and not crm_cli.celular:
                        crm_cli.celular = celular_fid
                        _changed.append('celular')
                    if _changed:
                        crm_cli.save(update_fields=_changed)
            except Exception:
                _log.exception("Error sync CRM en guardar_cliente_pos rut=%s", rut)

        return JsonResponse({
            'success': True,
            'cliente_id': cliente.id,
            'mensaje': 'Cliente guardado exitosamente'
        })

    except Exception as e:
        _log.exception("Error inesperado en guardar_cliente_pos")
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar cliente: {str(e)}'
        }, status=500)


@require_POST
@login_required
def enviar_ticket_email(request):
    """Enviar ticket por email al cliente"""
    try:
        data = json.loads(request.body)
        ticket_id = data.get('ticket_id')
        email = data.get('email', '').strip()
        
        if not ticket_id or not email:
            return JsonResponse({
                'success': False,
                'error': 'Debe proporcionar ticket_id y email'
            })
        
        # Buscar el ticket
        ticket = Ticket.objects.filter(correlativo=ticket_id).first()
        
        if not ticket:
            return JsonResponse({
                'success': False,
                'error': f'No se encontró el ticket #{ticket_id}'
            })
        
        # TODO: Implementar envío de email
        # Por ahora, solo simular el envío
        # En producción, usar Django's send_mail o un servicio de email
        
        # from django.core.mail import send_mail
        # from django.template.loader import render_to_string
        
        # asunto = f'Ticket de Venta #{ticket.correlativo}'
        # mensaje = render_to_string('emails/ticket_venta.html', {'ticket': ticket})
        # send_mail(asunto, mensaje, 'noreply@retailmind.cl', [email], html_message=mensaje)
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Ticket enviado exitosamente a {email}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al enviar email: {str(e)}'
        })


# ========== DASHBOARD DE VENTAS ==========

@login_required
def dashboard_ventas(request):
    """Vista principal del dashboard de ventas - NEXO Design System"""
    return render(request, 'vistas/modulo_dashboards/dashboard_ventas_nexo.html')


@login_required
def dashboard_ventas_mejorado(request):
    """Vista del dashboard de ventas mejorado - NEXO Design System"""
    from app.models import Categoria, Productos_Atributos, AtributoOpcion
    # Categorías v1.2 para el filtro (árbol Padre › Hijo, ocultando deprecadas).
    cats_raw = list(Categoria.objects.exclude(nombre__startswith='_ZZ_')
                    .values('id', 'nombre', 'padre_id'))
    hijos = {}
    for c in cats_raw:
        hijos.setdefault(c['padre_id'], []).append(c)
    categorias = []
    for raiz in sorted(hijos.get(None, []), key=lambda c: c['nombre']):
        categorias.append({'id': raiz['id'], 'nombre': raiz['nombre']})
        for h in sorted(hijos.get(raiz['id'], []), key=lambda c: c['nombre']):
            categorias.append({'id': h['id'], 'nombre': '› ' + h['nombre']})
    esp_attr = Productos_Atributos.objects.filter(nombre__iexact='Especialidad').first()
    especialidades = (list(AtributoOpcion.objects.filter(atributo=esp_attr)
                           .order_by('valor').values('id', 'valor')) if esp_attr else [])
    return render(request, 'vistas/modulo_dashboards/dashboard_ventas_nexo.html', {
        'categorias_filtro': categorias,
        'especialidades_filtro': especialidades,
    })


@require_GET
@login_required
@cache_ventas_json('ind_globales_v2', timeout=60)
def obtener_indicadores_globales_ventas(request):
    """
    API para obtener indicadores globales de ventas
    Incluye: ventas totales, ticket promedio, cantidad ventas, crecimiento

    Base de cálculo (compartida con el resto del dashboard vía
    `_tickets_venta_periodo`): created_at, estado PAGADO por defecto y sin
    tickets de cambio/devolución. La evolución diaria agrupa por la MISMA fecha
    que los totales, así el gráfico cuadra con la tarjeta.
    """
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal_id = request.GET.get('sucursal_id')
        vendedor_id = request.GET.get('vendedor_id')
        metodo_pago = request.GET.get('metodo_pago')
        estado = request.GET.get('estado', '')  # Vacío por defecto para mostrar todos
        periodo_comparacion = request.GET.get('periodo_comparacion', 'mes_anterior')
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            fecha_fin = timezone.localdate()
            fecha_inicio = fecha_fin - timedelta(days=30)
        else:
            try:
                fecha_inicio = timezone.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin = timezone.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }, status=400)
        
        def build_queryset(f_inicio, f_fin):
            # created_at = fecha real de venta (Ticket.fecha es auto_now) y sin
            # tickets de cambio/devolución: helper único del dashboard.
            qs = _tickets_venta_periodo(request, f_inicio, f_fin, sucursal_id=sucursal_id)
            if vendedor_id:
                qs = qs.filter(vendedor_id=vendedor_id)
            if metodo_pago:
                qs = qs.filter(pagos__metodo_pago=metodo_pago).distinct()
            return qs

        queryset = build_queryset(fecha_inicio, fecha_fin)

        # --- Un solo aggregate para todas las métricas del período actual ---
        agg = queryset.aggregate(
            ventas_totales=Sum('total'),
            cantidad_ventas=Count('id', distinct=True),
            ventas_con_factura=Count(
                'id',
                filter=Q(tipo_dte__in=['FACTURA_ELECTRONICA', 'FACTURA_EXENTA']),
                distinct=True,
            ),
            ventas_con_boleta=Count(
                'id', filter=Q(tipo_dte='BOLETA_ELECTRONICA'), distinct=True,
            ),
            tickets_offline=Count('id', filter=Q(created_offline=True), distinct=True),
        )
        ventas_totales = agg['ventas_totales'] or 0
        cantidad_ventas = agg['cantidad_ventas'] or 0
        ticket_promedio = ventas_totales / cantidad_ventas if cantidad_ventas > 0 else 0

        # --- Período de comparación ---
        if periodo_comparacion == 'mes_anterior':
            dias_diferencia = (fecha_fin - fecha_inicio).days
            fecha_comp_fin = fecha_inicio - timedelta(days=1)
            fecha_comp_inicio = fecha_comp_fin - timedelta(days=dias_diferencia)
        elif periodo_comparacion == 'mes_mismo_anio_anterior':
            fecha_comp_inicio = fecha_inicio.replace(year=fecha_inicio.year - 1)
            fecha_comp_fin = fecha_fin.replace(year=fecha_fin.year - 1)
        else:  # semana_anterior
            fecha_comp_fin = fecha_inicio - timedelta(days=1)
            fecha_comp_inicio = fecha_comp_fin - timedelta(days=6)

        agg_comp = build_queryset(fecha_comp_inicio, fecha_comp_fin).aggregate(
            ventas_comp=Sum('total'),
            cantidad_comp=Count('id', distinct=True),
        )
        ventas_comp = agg_comp['ventas_comp'] or 0
        cantidad_comp = agg_comp['cantidad_comp'] or 0
        ticket_comp = ventas_comp / cantidad_comp if cantidad_comp > 0 else 0

        crecimiento_ventas = ((ventas_totales - ventas_comp) / ventas_comp * 100) if ventas_comp > 0 else 0
        crecimiento_cantidad = ((cantidad_ventas - cantidad_comp) / cantidad_comp * 100) if cantidad_comp > 0 else 0
        crecimiento_ticket = ((ticket_promedio - ticket_comp) / ticket_comp * 100) if ticket_comp > 0 else 0

        # Cambios y devoluciones
        cambios_qs = CambioDevolucion.objects.filter(
            fecha_solicitud__date__gte=fecha_inicio,
            fecha_solicitud__date__lte=fecha_fin,
        )
        cambios_qs = _scope_suc_emp(cambios_qs, request, sucursal_id)
        cantidad_cambios = cambios_qs.count()
        ratio_cambios = (cantidad_cambios / cantidad_ventas * 100) if cantidad_ventas > 0 else 0

        # Descuentos aplicados en el periodo
        desc_agg = Ticket_Productos.objects.filter(idTicket__in=queryset).aggregate(
            descuento_total=Sum(
                ExpressionWrapper(
                    F('stock') * F('descuento_unitario'),
                    output_field=DecimalField(),
                )
            ),
            descuento_prom_pct=Avg('porcentaje_descuento'),
        )
        descuento_total = float(desc_agg['descuento_total'] or 0)
        descuento_prom_pct = float(desc_agg['descuento_prom_pct'] or 0)

        ventas_con_factura = agg['ventas_con_factura'] or 0
        ventas_con_boleta = agg['ventas_con_boleta'] or 0
        tickets_offline = agg['tickets_offline'] or 0

        # Evolución diaria de ventas — un solo values+annotate y fill en memoria.
        # Agrupa por TruncDate('created_at'), la MISMA fecha con la que se calculó
        # "Ventas Totales": antes agrupaba por `fecha` (auto_now), así que el
        # gráfico no sumaba lo mismo que la tarjeta y además inyectaba días fuera
        # del rango pedido (jun-2026: 2 días fantasma, incluido 2020-01-01).
        evolucion_diaria = queryset.annotate(dia=TruncDate('created_at')).values('dia').annotate(
            total=Sum('total'),
            cantidad=Count('id'),
        ).order_by('dia')

        todas_fechas = {}
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            todas_fechas[fecha_actual] = {'total': 0, 'cantidad': 0}
            fecha_actual += timedelta(days=1)
        fuera_de_rango = 0
        for item in evolucion_diaria:
            dia = item['dia']
            if dia is None or dia not in todas_fechas:
                # No se inventa un punto en el gráfico: se descarta y se deja
                # traza (sólo puede ocurrir con datos corruptos de fecha).
                fuera_de_rango += 1
                continue
            todas_fechas[dia] = {
                'total': float(item['total'] or 0),
                'cantidad': item['cantidad'],
            }
        if fuera_de_rango:
            logger.warning(
                'Dashboard ventas: %s días fuera del rango %s..%s descartados de la evolución diaria',
                fuera_de_rango, fecha_inicio, fecha_fin,
            )
        evolucion_data = [
            {
                'fecha': fecha.strftime('%d/%m'),
                'total': datos['total'],
                'cantidad': datos['cantidad'],
            }
            for fecha, datos in sorted(todas_fechas.items())
        ]
        
        return JsonResponse({
            'success': True,
            'ventas_totales': float(ventas_totales),
            'cantidad_ventas': cantidad_ventas,
            'ticket_promedio': float(ticket_promedio),
            'cantidad_cambios': cantidad_cambios,
            'ratio_cambios': float(ratio_cambios),
            'crecimiento_ventas': float(crecimiento_ventas),
            'crecimiento_cantidad': float(crecimiento_cantidad),
            'crecimiento_ticket': float(crecimiento_ticket),
            'evolucion_diaria': evolucion_data,
            'descuento_total': descuento_total,
            'descuento_prom_pct': round(descuento_prom_pct, 2),
            'ventas_con_factura': ventas_con_factura,
            'ventas_con_boleta': ventas_con_boleta,
            'tickets_offline': tickets_offline,
            'periodo': {
                'inicio': fecha_inicio.strftime('%d/%m/%Y'),
                'fin': fecha_fin.strftime('%d/%m/%Y')
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener indicadores globales: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('ventas_vendedor_v2', timeout=60)
def obtener_ventas_por_vendedor(request):
    """
    API para obtener ventas por vendedor con métricas individuales
    Incluye: ranking, comisiones, participación

    Usa la MISMA base que la tarjeta "Ventas Totales" (`_tickets_venta_periodo`):
    antes filtraba por `Ticket.fecha` (auto_now) y sin default de estado, así que
    incluía ANULADO/PENDIENTE y la tabla nunca cuadraba con el KPI.
    """
    try:
        # Obtener parámetros de filtro
        sucursal_id = request.GET.get('sucursal_id')

        # Validar fechas
        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        queryset = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        # Calcular total general para participación
        total_general = queryset.aggregate(total=Sum('total'))['total'] or 0
        
        # Agrupar por vendedor
        ventas_vendedor = queryset.values(
            'vendedor__id',
            'vendedor__codigo_vendedor',
            'vendedor__nombre',
            'vendedor__comision'
        ).annotate(
            total_vendido=Sum('total'),
            cantidad_ventas=Count('id'),
            ticket_promedio=Avg('total')
        ).order_by('-total_vendido')
        
        vendedores_data = []
        top_vendedores = []
        
        for idx, venta in enumerate(ventas_vendedor):
            total_vendido = float(venta['total_vendido'] or 0)
            cantidad_ventas = venta['cantidad_ventas']
            ticket_promedio = float(venta['ticket_promedio'] or 0)
            comision_porcentaje = float(venta['vendedor__comision'] or 0)
            comision_total = total_vendido * (comision_porcentaje / 100)
            participacion = (total_vendido / total_general * 100) if total_general > 0 else 0
            
            # Calcular rendimiento (basado en participación relativa)
            if idx == 0 and total_vendido > 0:
                rendimiento = 100
            elif total_vendido > 0 and ventas_vendedor[0]['total_vendido']:
                rendimiento = (total_vendido / float(ventas_vendedor[0]['total_vendido']) * 100)
            else:
                rendimiento = 0
            
            vendedor_info = {
                'id': venta['vendedor__id'],
                'codigo': venta['vendedor__codigo_vendedor'] or 'S/C',
                'nombre': venta['vendedor__nombre'] or 'Sin nombre',
                'cantidad_ventas': cantidad_ventas,
                'total_vendido': total_vendido,
                'ticket_promedio': ticket_promedio,
                'comision_porcentaje': comision_porcentaje,
                'comision_total': comision_total,
                'participacion': float(participacion),
                'rendimiento': float(rendimiento)
            }
            
            vendedores_data.append(vendedor_info)
            
            # Top 10 vendedores para gráfico
            if idx < 10:
                top_vendedores.append({
                    'nombre': venta['vendedor__nombre'] or 'Sin nombre',
                    'total': total_vendido
                })
        
        return JsonResponse({
            'success': True,
            'vendedores': vendedores_data,
            'top_vendedores': top_vendedores,
            'total_vendedores': len(vendedores_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener ventas por vendedor: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('suc_dashboard_v2', timeout=300, vary_on_session=True)
def obtener_sucursales_dashboard(request):
    """
    API para obtener empresas y sucursales para los filtros del dashboard.
    Cada sucursal incluye `empresa_id` para la cascada Empresa → Sucursal.
    Usa la utilidad centralizada de permisos para determinar visibilidad.
    """
    try:
        from .utils_permisos import (
            obtener_sucursales_usuario, usuario_puede_ver_todas_sucursales,
        )

        sucursales = obtener_sucursales_usuario(request.user).select_related('empresa')

        sucursales_data = []
        for sucursal in sucursales:
            sucursales_data.append({
                'id': sucursal.id,
                'nombre': sucursal.alias,
                'alias': sucursal.alias,
                'direccion': sucursal.direccion or '',
                'empresa_id': sucursal.empresa_id,
                'empresa': sucursal.empresa.nombre if sucursal.empresa else '',
                'es_cd': sucursal.es_centro_distribucion,
            })

        # Empresas para el dropdown: SOLO las que operan sucursales visibles (las
        # operadoras del holding), derivadas de las propias sucursales. OJO: NO
        # usar la tabla Empresa completa ni obtener_empresas_usuario: `Empresa`
        # también almacena clientes/proveedores (miles) y ensuciaría el filtro.
        empresas_dict = {}
        for s in sucursales_data:
            eid = s['empresa_id']
            if not eid:
                continue
            e = empresas_dict.setdefault(eid, {'id': eid, 'nombre': s['empresa'], 'tiene_tiendas': False})
            if not s['es_cd']:
                e['tiene_tiendas'] = True
        empresas_data = sorted(empresas_dict.values(), key=lambda e: e['nombre'])

        empresa_user = EmpresaUser.objects.filter(
            user=request.user,
            active=True
        ).first()

        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data,
            'empresas': empresas_data,
            'es_admin': usuario_puede_ver_todas_sucursales(request.user),
            'sucursal_actual': empresa_user.sucursal_id if empresa_user else None
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener sucursales: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('ventas_sucursal_v2', timeout=60)
def obtener_ventas_por_sucursal(request):
    """
    API para obtener análisis comparativo de ventas por sucursal.

    Misma base de venta que el resto del dashboard (created_at, PAGADO por
    defecto, sin cambios/devoluciones). NO aplica el filtro de sucursal a
    propósito: esta sección compara tiendas entre sí (igual que mix-por-sucursal).
    """
    try:
        # Validar fechas
        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        queryset = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, aplicar_scope=False)

        # Consultar ventas por sucursal
        ventas_sucursal = queryset.values(
            'sucursal__id',
            'sucursal__alias'
        ).annotate(
            total_ventas=Sum('total'),
            cantidad=Count('id')
        ).order_by('-total_ventas')
        
        sucursales_data = []
        for venta in ventas_sucursal:
            total = float(venta['total_ventas'] or 0)
            cantidad = venta['cantidad']
            ticket_promedio = total / cantidad if cantidad > 0 else 0
            
            sucursales_data.append({
                'id': venta['sucursal__id'],
                'sucursal': venta['sucursal__alias'] or 'Sin nombre',
                'total': total,
                'cantidad': cantidad,
                'ticket_promedio': ticket_promedio
            })
        
        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener ventas por sucursal: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('ventas_metodo_pago_v2', timeout=60)
def obtener_ventas_por_metodo_pago(request):
    """
    API para obtener distribución de ventas por método de pago.
    Misma base de venta que la tarjeta "Ventas Totales" (created_at, PAGADO por
    defecto, sin cambios/devoluciones).
    """
    try:
        # Obtener parámetros de filtro
        sucursal_id = request.GET.get('sucursal_id')

        # Validar fechas
        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        queryset = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        # Obtener IDs de tickets que cumplen con los filtros
        ticket_ids = queryset.values_list('id', flat=True)
        
        # Agrupar por método de pago desde TicketDetallePago
        ventas_metodo = TicketDetallePago.objects.filter(
            ticket_id__in=ticket_ids
        ).values('metodo_pago').annotate(
            total=Sum('monto'),
            cantidad=Count('id')
        ).order_by('-total')
        
        metodos_data = []
        total_general = 0
        
        for metodo in ventas_metodo:
            total = float(metodo['total'] or 0)
            total_general += total
            
            # Obtener nombre legible del método
            metodo_nombre = dict(METODO_PAGO_TICKET_CHOICES).get(
                metodo['metodo_pago'], 
                metodo['metodo_pago']
            )
            
            metodos_data.append({
                'metodo': metodo_nombre,
                'codigo': metodo['metodo_pago'],
                'total': total,
                'cantidad': metodo['cantidad']
            })
        
        # Calcular porcentajes
        for metodo in metodos_data:
            metodo['porcentaje'] = (metodo['total'] / total_general * 100) if total_general > 0 else 0
        
        return JsonResponse({
            'success': True,
            'metodos_pago': metodos_data,
            'total': total_general
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener ventas por método de pago: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('analisis_cambios_v2', timeout=120)
def obtener_analisis_cambios_devoluciones(request):
    """
    API para obtener análisis de cambios y devoluciones
    Incluye: ratio, motivos, impacto financiero
    """
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal_id = request.GET.get('sucursal_id')
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            fecha_fin = timezone.localdate()
            fecha_inicio = fecha_fin - timedelta(days=30)
        else:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Consultar cambios y devoluciones
        queryset = CambioDevolucion.objects.filter(
            fecha_solicitud__date__gte=fecha_inicio,
            fecha_solicitud__date__lte=fecha_fin
        )
        
        queryset = _scope_suc_emp(queryset, request, sucursal_id)
        
        # Métricas generales
        total_cambios = queryset.count()
        monto_total = queryset.aggregate(
            total=Sum('monto_original')
        )['total'] or 0
        
        # Total de ventas para calcular ratio — mismo denominador que la tarjeta
        # "Transacciones" (created_at, PAGADO, sin tickets de cambio/devolución;
        # contarlos inflaba el denominador y bajaba artificialmente el ratio).
        ventas_total = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        cantidad_ventas = ventas_total.count()
        ratio = (total_cambios / cantidad_ventas * 100) if cantidad_ventas > 0 else 0
        
        # Análisis por motivo (desde CambioDevolucion)
        motivos_cambio = queryset.filter(
            motivo_principal__isnull=False
        ).values('motivo_principal').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')
        
        motivos_data = []
        for item in motivos_cambio:
            if item['motivo_principal']:
                motivo_nombre = dict(MOTIVO_CAMBIO_CHOICES).get(
                    item['motivo_principal'],
                    item['motivo_principal']
                )
                motivos_data.append({
                    'motivo': motivo_nombre,
                    'cantidad': item['cantidad']
                })
        
        # Análisis por tipo de operación
        por_tipo = queryset.values('tipo_operacion').annotate(
            cantidad=Count('id'),
            monto=Sum('monto_original')
        )
        
        tipos_data = []
        for tipo in por_tipo:
            tipo_nombre = dict(TIPO_OPERACION_CAMBIO_CHOICES).get(
                tipo['tipo_operacion'],
                tipo['tipo_operacion']
            )
            tipos_data.append({
                'tipo': tipo_nombre,
                'cantidad': tipo['cantidad'],
                'monto': float(tipo['monto'] or 0)
            })
        
        # Análisis por estado
        por_estado = queryset.values('estado').annotate(
            cantidad=Count('id')
        )
        
        estados_data = []
        for estado in por_estado:
            estado_nombre = dict(ESTADO_CAMBIO_CHOICES).get(
                estado['estado'],
                estado['estado']
            )
            estados_data.append({
                'estado': estado_nombre,
                'cantidad': estado['cantidad']
            })
        
        return JsonResponse({
            'success': True,
            'total_cambios': total_cambios,
            'monto_total': float(monto_total),
            'ratio': float(ratio),
            'por_motivo': motivos_data,
            'por_tipo': tipos_data,
            'por_estado': estados_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener análisis de cambios: {str(e)}'
        }, status=500)


@require_GET
@login_required
def obtener_analisis_fraude_cambios(request):
    """
    API para obtener análisis de detección de fraude en cambios y devoluciones.
    Solo accesible para administradores, jefes locales y administración.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos para acceder a esta información'}, status=403)

        from .services.fraud_detection import (
            detectar_vendedores_alto_retorno,
            detectar_productos_multiples_cambios,
            detectar_perdidas_no_apto,
            detectar_cambios_fuera_plazo,
            detectar_patrones_cross_branch,
        )

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')

        if fecha_inicio and fecha_fin:
            from datetime import datetime
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        else:
            fecha_fin = timezone.localdate()
            fecha_inicio = fecha_fin - timedelta(days=30)

        vendedores = detectar_vendedores_alto_retorno(sucursal_id, fecha_inicio, fecha_fin)
        productos = detectar_productos_multiples_cambios(sucursal_id, fecha_inicio, fecha_fin)
        perdidas = detectar_perdidas_no_apto(sucursal_id, fecha_inicio, fecha_fin)
        fuera_plazo = detectar_cambios_fuera_plazo(sucursal_id, fecha_inicio, fecha_fin)
        cross_branch = detectar_patrones_cross_branch(fecha_inicio, fecha_fin)

        alertas_vendedores = len([v for v in vendedores if v['alerta']])
        alertas_productos = len(productos)
        alertas_total = alertas_vendedores + alertas_productos + (1 if perdidas['total_items'] > 0 else 0) + (1 if fuera_plazo['total'] > 0 else 0) + (1 if cross_branch['pendientes_revision'] > 0 else 0)

        return JsonResponse({
            'success': True,
            'alertas_total': alertas_total,
            'vendedores_alto_retorno': vendedores[:10],
            'productos_multiples_cambios': productos[:10],
            'perdidas_no_apto': perdidas,
            'cambios_fuera_plazo': fuera_plazo,
            'patrones_cross_branch': cross_branch,
            'periodo': {
                'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            },
        })
    except Exception as e:
        logger.exception("Error al obtener analisis de fraude")
        return JsonResponse({'success': False, 'error': f'Error al obtener análisis de fraude: {str(e)}'}, status=500)


@require_GET
@login_required
def obtener_analisis_cambios_avanzado(request):
    """
    API para obtener análisis avanzado completo de cambios y devoluciones.
    Solo accesible para administradores, jefes locales y administración.
    """
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos para acceder a esta información'}, status=403)

        from .services.fraud_detection import obtener_analisis_avanzado

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')

        if fecha_inicio and fecha_fin:
            from datetime import datetime
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        else:
            fecha_fin = timezone.localdate()
            fecha_inicio = fecha_fin - timedelta(days=30)

        analisis = obtener_analisis_avanzado(sucursal_id, fecha_inicio, fecha_fin)
        analisis['success'] = True
        return JsonResponse(analisis)
    except Exception as e:
        logger.exception("Error al obtener analisis avanzado de cambios")
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@require_GET
@login_required
def listar_autorizaciones_cross_branch(request):
    """
    Lista autorizaciones cross-branch para revisión gerencial.
    """
    try:
        from .models import RegistroAutorizacion

        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos'}, status=403)

        estado = request.GET.get('estado', 'todos')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')

        qs = RegistroAutorizacion.objects.filter(
            es_cross_branch=True, exitoso=True,
        ).select_related(
            'usuario_solicitante', 'usuario_autorizador',
            'sucursal_solicitante', 'sucursal_autorizador',
            'cambio_devolucion', 'revisado_por',
        ).order_by('-fecha_hora')

        if estado == 'pendientes':
            qs = qs.filter(requiere_revision=True, revisado_por__isnull=True)
        elif estado == 'revisados':
            qs = qs.filter(revisado_por__isnull=False)

        if fecha_desde:
            qs = qs.filter(fecha_hora__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_hora__date__lte=fecha_hasta)

        registros = []
        for r in qs[:50]:
            registros.append({
                'id': r.id,
                'fecha': r.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                'usuario_solicitante': r.usuario_solicitante.get_full_name() or r.usuario_solicitante.username if r.usuario_solicitante else 'N/A',
                'usuario_autorizador': r.usuario_autorizador.get_full_name() or r.usuario_autorizador.username if r.usuario_autorizador else 'N/A',
                'sucursal_solicitante': r.sucursal_solicitante.alias if r.sucursal_solicitante else 'N/A',
                'sucursal_autorizador': r.sucursal_autorizador.alias if r.sucursal_autorizador else 'N/A',
                'tipo_operacion': r.get_tipo_operacion_display(),
                'descripcion': r.descripcion,
                'cambio_id': r.cambio_devolucion_id,
                'cambio_numero': r.cambio_devolucion.numero_operacion if r.cambio_devolucion else None,
                'requiere_revision': r.requiere_revision,
                'revisado': r.revisado_por is not None,
                'revisado_por': r.revisado_por.get_full_name() if r.revisado_por else None,
                'fecha_revision': r.fecha_revision.strftime('%d/%m/%Y %H:%M') if r.fecha_revision else None,
                'notas_revision': r.notas_revision,
            })

        pendientes = RegistroAutorizacion.objects.filter(
            es_cross_branch=True, exitoso=True,
            requiere_revision=True, revisado_por__isnull=True,
        ).count()

        return JsonResponse({
            'success': True,
            'registros': registros,
            'pendientes_revision': pendientes,
        })
    except Exception as e:
        logger.exception("Error al listar autorizaciones cross-branch")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
@csrf_exempt
def revisar_autorizacion(request, registro_id):
    """Marca una autorización cross-branch como revisada."""
    try:
        from .models import RegistroAutorizacion

        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos'}, status=403)

        registro = RegistroAutorizacion.objects.get(id=registro_id)
        data = json.loads(request.body)

        registro.revisado_por = request.user
        registro.fecha_revision = timezone.now()
        registro.notas_revision = data.get('notas', '')
        registro.save(update_fields=['revisado_por', 'fecha_revision', 'notas_revision'])

        return JsonResponse({'success': True, 'mensaje': 'Autorización marcada como revisada'})
    except RegistroAutorizacion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Registro no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
@login_required
def obtener_cola_revision_gerencial(request):
    """Obtiene la cola de cambios que requieren revisión gerencial."""
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos'}, status=403)

        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')

        qs = CambioDevolucion.objects.filter(
            requiere_revision_gerencial=True,
            revisado_por_gerencia__isnull=True,
        ).select_related(
            'ticket_original', 'sucursal', 'solicitado_por', 'autorizado_por_usuario',
        ).order_by('-fecha_solicitud')

        qs = _scope_suc_emp(qs, request, sucursal_id)

        items = []
        for c in qs[:50]:
            items.append({
                'id': c.id,
                'numero_operacion': c.numero_operacion,
                'tipo_operacion': c.get_tipo_operacion_display(),
                'estado': c.get_estado_display(),
                'monto_original': float(c.monto_original),
                'diferencia_monto': float(c.diferencia_monto),
                'es_fuera_de_plazo': c.es_fuera_de_plazo,
                'dias_fuera_de_plazo': c.dias_fuera_de_plazo,
                'es_cross_branch': c.es_autorizacion_cross_branch,
                'tipo_especial': c.tipo_cambio_especial,
                'score_riesgo': c.score_riesgo,
                'solicitado_por': c.solicitado_por.get_full_name() or c.solicitado_por.username,
                'autorizado_por': c.autorizado_por_usuario.get_full_name() if c.autorizado_por_usuario else None,
                'sucursal': c.sucursal.alias if c.sucursal else '',
                'fecha': c.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
                'ticket_original': c.ticket_original.correlativo if c.ticket_original else '',
                'motivo': c.get_motivo_principal_display(),
            })

        return JsonResponse({'success': True, 'items': items, 'total_pendientes': qs.count()})
    except Exception as e:
        logger.exception("Error al obtener cola de revision gerencial")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
@csrf_exempt
def revisar_cambio_gerencial(request):
    """Marca un cambio como revisado por gerencia."""
    try:
        rol_usuario = getattr(request.user, 'rol', None)
        if rol_usuario not in ['administrador', 'jefe_local', 'administracion']:
            return JsonResponse({'success': False, 'error': 'No tiene permisos'}, status=403)

        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        notas = data.get('notas', '')

        cambio = CambioDevolucion.objects.get(id=cambio_id)
        cambio.revisado_por_gerencia = request.user
        cambio.fecha_revision_gerencia = timezone.now()
        cambio.notas_revision_gerencia = notas
        cambio.save(update_fields=['revisado_por_gerencia', 'fecha_revision_gerencia', 'notas_revision_gerencia'])

        HistorialCambioDevolucion.objects.create(
            cambio_devolucion=cambio,
            usuario=request.user,
            accion='MODIFICADO',
            estado_anterior=cambio.estado,
            estado_nuevo=cambio.estado,
            descripcion=f'Revisión gerencial completada. Notas: {notas[:200]}',
        )

        return JsonResponse({'success': True, 'mensaje': 'Cambio marcado como revisado por gerencia'})
    except CambioDevolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cambio no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
@login_required
def exportar_cambios_devoluciones(request):
    """Exporta listado de cambios y devoluciones a Excel."""
    try:
        import io
        from django.http import HttpResponse

        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')

        qs = CambioDevolucion.objects.select_related(
            'ticket_original', 'sucursal', 'solicitado_por',
            'aprobado_por', 'autorizado_por_usuario',
        ).order_by('-fecha_solicitud')

        if fecha_desde:
            qs = qs.filter(fecha_solicitud__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_solicitud__date__lte=fecha_hasta)
        qs = _scope_suc_emp(qs, request, sucursal_id)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cambios y Devoluciones"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )

            headers = [
                'N Operacion', 'Fecha', 'Tipo', 'Estado', 'Tipo Especial',
                'Sucursal', 'Cliente', 'RUT', 'Ticket Original',
                'Monto Original', 'Monto Nuevo', 'Diferencia',
                'Motivo', 'Solicitado Por', 'Aprobado Por',
                'Fuera Plazo', 'Dias Fuera', 'Cross-Branch',
                'Autorizado Por', 'Score Riesgo', 'Observaciones',
            ]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            alert_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            for row_num, cambio in enumerate(qs[:5000], 2):
                row_data = [
                    cambio.numero_operacion,
                    cambio.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if cambio.fecha_solicitud else '',
                    cambio.get_tipo_operacion_display(),
                    cambio.get_estado_display(),
                    cambio.tipo_cambio_especial,
                    cambio.sucursal.alias if cambio.sucursal else '',
                    cambio.ticket_original.cliente_nombre if cambio.ticket_original else '',
                    cambio.ticket_original.cliente_rut if cambio.ticket_original else '',
                    cambio.ticket_original.correlativo if cambio.ticket_original else '',
                    float(cambio.monto_original),
                    float(cambio.monto_nuevo),
                    float(cambio.diferencia_monto),
                    cambio.get_motivo_principal_display(),
                    cambio.solicitado_por.get_full_name() if cambio.solicitado_por else '',
                    cambio.aprobado_por.get_full_name() if cambio.aprobado_por else '',
                    'Si' if cambio.es_fuera_de_plazo else 'No',
                    cambio.dias_fuera_de_plazo,
                    'Si' if cambio.es_autorizacion_cross_branch else 'No',
                    cambio.autorizado_por_usuario.get_full_name() if cambio.autorizado_por_usuario else '',
                    cambio.score_riesgo,
                    cambio.observaciones_vendedor or '',
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    if cambio.es_fuera_de_plazo or cambio.score_riesgo >= 50:
                        cell.fill = alert_fill

            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="cambios_devoluciones_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            return response

        except ImportError:
            import csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="cambios_devoluciones_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response)
            writer.writerow(['N Operacion', 'Fecha', 'Tipo', 'Estado', 'Sucursal', 'Monto Original', 'Diferencia', 'Motivo', 'Fuera Plazo', 'Score Riesgo'])
            for cambio in qs[:5000]:
                writer.writerow([
                    cambio.numero_operacion,
                    cambio.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if cambio.fecha_solicitud else '',
                    cambio.get_tipo_operacion_display(),
                    cambio.get_estado_display(),
                    cambio.sucursal.alias if cambio.sucursal else '',
                    float(cambio.monto_original),
                    float(cambio.diferencia_monto),
                    cambio.get_motivo_principal_display(),
                    'Si' if cambio.es_fuera_de_plazo else 'No',
                    cambio.score_riesgo,
                ])
            return response

    except Exception as e:
        logger.exception("Error al exportar cambios/devoluciones")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
@login_required
@cache_ventas_json('estado_cuadraturas', timeout=120)
def obtener_estado_cuadraturas(request):
    """
    API para obtener estado de cuadraturas de caja
    Incluye: exitosas, con diferencias, pendientes
    """
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal_id = request.GET.get('sucursal_id')
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            fecha_fin = timezone.localdate()
            fecha_inicio = fecha_fin - timedelta(days=30)
        else:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Consultar arqueos/cuadraturas
        queryset = ArqueoCaja.objects.filter(
            fecha_arqueo__gte=fecha_inicio,
            fecha_arqueo__lte=fecha_fin
        )
        
        queryset = _scope_suc_emp(queryset, request, sucursal_id)
        
        total_cuadraturas = queryset.count()
        
        # Calcular diferencias
        cuadraturas_con_datos = []
        exitosas = 0
        con_diferencias = 0
        
        for arqueo in queryset:
            # `total_efectivo_fisico` es el único valor correcto: lo calcula
            # `ArqueoCaja.save()` a partir de las denominaciones en el modo
            # detallado y lo escribe directo en el modo express (donde todas
            # las denominaciones quedan en 0). Rehacer la suma aquí, además,
            # se olvidaba de las monedas de $5 y de $1.
            total_conteo = arqueo.total_efectivo_fisico or 0

            # Misma fórmula que el modelo: el fondo fijo de caja chica es
            # parte del efectivo esperado en el cajón.
            diferencia = total_conteo - (
                arqueo.total_efectivo_teorico + (arqueo.fondo_fijo_snapshot or 0)
            )

            cuadraturas_con_datos.append({
                'id': arqueo.id,
                'fecha': arqueo.fecha_arqueo,
                'diferencia': abs(diferencia)
            })
            
            if abs(diferencia) <= 1000:  # Tolerancia de $1000
                exitosas += 1
            else:
                con_diferencias += 1
        
        # Cuadraturas pendientes (días sin cuadratura)
        dias_periodo = (fecha_fin - fecha_inicio).days + 1
        pendientes = max(0, dias_periodo - total_cuadraturas)
        
        # Calcular diferencia total y promedio
        diferencia_total = sum(c['diferencia'] for c in cuadraturas_con_datos)
        promedio_diferencia = diferencia_total / len(cuadraturas_con_datos) if cuadraturas_con_datos else 0
        
        return JsonResponse({
            'success': True,
            'exitosas': exitosas,
            'con_diferencias': con_diferencias,
            'pendientes': pendientes,
            'total': total_cuadraturas,
            'diferencia_total': float(diferencia_total),
            'promedio_diferencia': float(promedio_diferencia)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estado de cuadraturas: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('prod_mas_vendidos_v3', timeout=120)
def obtener_productos_mas_vendidos(request):
    """Top de productos vendidos agrupado por MODELO (articulo), enriquecido con
    marca / categoría Padre › Hija / género / especialidad(es).

    Cambios vs versión anterior (deliberados, para alinear con los charts v1.2):
      · Agrupa por `articulo` (modelo), no por SKU (talla): Producto es por
        sucursal, así el ranking se consolida entre tiendas y por curva de tallas.
      · Usa `created_at` (fecha real de venta), no `fecha` (auto_now).
      · Usa Sum('subtotal') (ingreso real tras descuentos), no stock*precio.
      · `participacion` es sobre el total del período, no sobre el top-N.
      · Excluye productos con excluir_de_analitica.
    Las especialidades (multi-etiqueta) se resuelven en query aparte con
    .distinct() para no duplicar filas ni inflar las sumas."""
    try:
        from app.models import Producto, ProductoAtributoValor, Categoria
        limite = int(request.GET.get('limite', 20))

        # 1) Tickets del período (created_at real, estado PAGADO por defecto,
        #    sucursal_id del GET) — mismo helper que los charts v1.2.
        ticket_ids = _tickets_pagados_periodo(request).values_list('id', flat=True)
        lineas = (Ticket_Productos.objects
                  .filter(idTicket_id__in=ticket_ids, ProductoTalla__isnull=False)
                  .exclude(ProductoTalla__producto__excluir_de_analitica=True))

        # 2) Filtros cruzados que el dashboard ya envía
        categoria_id = request.GET.get('categoria_id')
        if categoria_id:  # el padre incluye a sus hijas
            cat_ids = [int(categoria_id)]
            cat_ids += list(Categoria.objects.filter(padre_id=categoria_id)
                            .values_list('id', flat=True))
            lineas = lineas.filter(ProductoTalla__producto__categoria_id__in=cat_ids)
        especialidad_id = request.GET.get('especialidad_id')
        if especialidad_id:  # Exists → no duplica filas aunque el producto tenga varias etiquetas
            lineas = lineas.filter(Exists(
                ProductoAtributoValor.objects.filter(
                    producto_id=OuterRef('ProductoTalla__producto_id'),
                    opcion_id=especialidad_id)))

        # 3) Métricas por MODELO — sin join a atributos (cero doble conteo)
        agg = list(lineas
                   .values(articulo=F('ProductoTalla__producto__articulo'))
                   .annotate(cantidad_vendida=Sum('stock'),
                             total_ventas=Sum('subtotal'))
                   .order_by('-cantidad_vendida'))
        total_general = sum(float(a['total_ventas'] or 0) for a in agg)  # sobre TODO el período
        top = agg[:limite]
        arts = [a['articulo'] for a in top if a['articulo']]

        # 4) Atributos por articulo (Producto es por-sucursal → preferir no-nulo al fusionar)
        attrs = {}
        for p in (Producto.objects.filter(articulo__in=arts)
                  .values('articulo', 'descripcion',
                          'atributo1__valor',   # Marca
                          'atributo3__valor',   # Género
                          'categoria__nombre', 'categoria__padre__nombre')):
            cur = attrs.setdefault(p['articulo'], dict(p))
            for k, v in p.items():
                if not cur.get(k) and v:
                    cur[k] = v

        # 5) Especialidades por articulo (distinct → una fila por (articulo, slug))
        esp_map = {}
        for art, slug in (ProductoAtributoValor.objects
                          .filter(producto__articulo__in=arts,
                                  atributo__nombre__iexact='Especialidad')
                          .values_list('producto__articulo', 'opcion__valor')
                          .distinct()):
            if slug:
                esp_map.setdefault(art, set()).add(slug)

        productos_data = []
        for a in top:
            art = a['articulo']
            at = attrs.get(art, {})
            cantidad = a['cantidad_vendida'] or 0
            total_ventas = float(a['total_ventas'] or 0)
            hija = at.get('categoria__nombre') or 'Sin categoría'
            padre = at.get('categoria__padre__nombre') or ''
            productos_data.append({
                'articulo': art or 'Sin código',
                'nombre': art or 'Sin código',            # compat con el JS actual (p.nombre)
                'descripcion': at.get('descripcion') or '',
                'marca': at.get('atributo1__valor') or '—',
                'categoria': hija,
                'categoria_label': (padre + ' › ' + hija) if padre else hija,
                'genero': at.get('atributo3__valor') or '—',
                'especialidades': sorted(esp_map.get(art, [])),
                'cantidad': cantidad,
                'total_ventas': total_ventas,
                'precio_promedio': total_ventas / cantidad if cantidad > 0 else 0,
                'participacion': round(total_ventas / total_general * 100, 1) if total_general > 0 else 0,
            })

        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'total_productos': len(productos_data),
            # Universo COMPLETO del período (no el top-N): "Unidades Vendidas" y
            # el UPT se pintaban sumando sólo el top-20 (jun-2026: 438 de 5.638
            # unidades reales, 92% por debajo). Se exponen aquí para que ninguna
            # pantalla vuelva a derivarlos de la tabla.
            'unidades_periodo': int(sum(a['cantidad_vendida'] or 0 for a in agg)),
            'total_periodo': total_general,
            'modelos_periodo': len(agg),
        })

    except Exception as e:
        logger.error(f'Error productos más vendidos: {e}')
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos más vendidos: {str(e)}'
        }, status=500)


def _rango_periodo(request):
    """Helper: (fecha_inicio, fecha_fin) del período pedido; default últimos 30 días."""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    if not fecha_inicio or not fecha_fin:
        fecha_fin = timezone.localdate()
        fecha_inicio = fecha_fin - timedelta(days=30)
    else:
        fecha_inicio = timezone.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = timezone.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    return fecha_inicio, fecha_fin


def _scope_suc_emp(qs, request, sucursal_id=None, campo_empresa='sucursal__empresa_id'):
    """Filtra un queryset (con FK a Sucursal) por sucursal específica o, si no hay
    sucursal, por empresa (empresa_id). Así el filtro Empresa del dashboard acota
    todos los endpoints sin romper el filtro de sucursal existente.

    · Con sucursal_id  → una tienda (empresa se ignora, la tienda ya la implica).
    · Sin sucursal pero con empresa_id → todas las tiendas de esa empresa.
    · Sin ninguno → sin filtro (todas las visibles).

    `campo_empresa` permite el caso de líneas (ej. 'idTicket__sucursal__empresa_id'
    o 'producto__sucursal__empresa_id')."""
    if sucursal_id is None:
        sucursal_id = request.GET.get('sucursal_id')
    if sucursal_id:
        # el campo de sucursal directa depende del queryset; asumimos 'sucursal_id'
        return qs.filter(sucursal_id=sucursal_id)
    empresa_id = request.GET.get('empresa_id')
    if empresa_id:
        return qs.filter(**{campo_empresa: empresa_id})
    return qs


# Tickets que NO son venta nueva: la "venta" de un CAMBIO_DEVOLUCION es sólo la
# diferencia cobrada en un cambio. El POS, el reporte de ventas y el dashboard
# home ya los excluyen; los endpoints de este dashboard eran los únicos que los
# sumaban (jun-2026: $551.300 en 154 tickets, +$1.286 de ticket promedio falso).
MODULO_ORIGEN_NO_VENTA = 'CAMBIO_DEVOLUCION'

# Los montos de Ticket/Ticket_Productos son BRUTOS (con IVA); costo_fifo es NETO.
# Para comparar peras con peras el margen se calcula sobre ingreso neto.
IVA_FACTOR = 1.19


def _tickets_venta_periodo(request, fecha_inicio=None, fecha_fin=None,
                           aplicar_estado=True, excluir_cambios=True,
                           aplicar_scope=True, sucursal_id=None):
    """Fuente ÚNICA de verdad de "qué tickets son la venta del período" para todo
    el dashboard de ventas. Antes cada endpoint la construía a mano y ninguno
    coincidía con el de al lado.

    Reglas (todas deliberadas, ver docs/PLAN_DASHBOARDS_2026-07-25.md):
      · Fecha  → `created_at` (fecha real de la venta). `Ticket.fecha` es
        `auto_now`: se reescribe en cada save (reimpresión, cambio de estado,
        generación de DTE, sync desktop), así que filtrar por él mide "tickets
        tocados en el rango", no vendidos.
      · Estado → el del filtro; si viene vacío, PAGADO (mismo default que ya
        usaban indicadores-globales y tendencias, ahora en todos).
      · Excluye `modulo_origen='CAMBIO_DEVOLUCION'` (no es venta nueva).
      · Scope de sucursal/empresa vía `_scope_suc_emp`.

    `aplicar_estado=False` (panel operacional) mantiene todos los estados;
    `excluir_cambios=False` conserva los cambios cuando el desglose los necesita;
    `aplicar_scope=False` para las vistas que comparan tiendas entre sí.
    """
    if fecha_inicio is None or fecha_fin is None:
        fecha_inicio, fecha_fin = _rango_periodo(request)
    tickets = Ticket.objects.filter(
        created_at__date__gte=fecha_inicio, created_at__date__lte=fecha_fin)
    if aplicar_estado:
        estado = request.GET.get('estado', '')
        tickets = tickets.filter(estado=estado) if estado else tickets.filter(estado='PAGADO')
    if excluir_cambios:
        tickets = tickets.exclude(modulo_origen=MODULO_ORIGEN_NO_VENTA)
    if aplicar_scope:
        tickets = _scope_suc_emp(tickets, request, sucursal_id)
    return tickets


def _tickets_pagados_periodo(request):
    """Alias histórico (charts v1.2 / indicador de compra). Delega en
    `_tickets_venta_periodo` para que hereden el mismo criterio de fecha,
    estado y exclusión de cambios/devoluciones."""
    return _tickets_venta_periodo(request)


@require_GET
@login_required
def obtener_ventas_por_categoria(request):
    """Ventas ($ y unidades) agrupadas por categoría v1.2 (Padre › Hijo).
    Recién con la recategorización esto es útil (antes RAMA CASUAL = todo).
    Opcional: ?especialidad_id= para ver una especialidad dentro."""
    try:
        especialidad_id = request.GET.get('especialidad_id')
        ticket_ids = _tickets_pagados_periodo(request).values_list('id', flat=True)
        lineas = (Ticket_Productos.objects
                  .filter(idTicket_id__in=ticket_ids, ProductoTalla__isnull=False)
                  .exclude(ProductoTalla__producto__excluir_de_analitica=True))
        if especialidad_id:
            lineas = lineas.filter(ProductoTalla__producto__atributos__opcion_id=especialidad_id)
        agg = (lineas.values(
                   'ProductoTalla__producto__categoria__nombre',
                   'ProductoTalla__producto__categoria__padre__nombre')
               .annotate(cantidad=Sum('stock'), total=Sum('subtotal'))
               .order_by('-total'))
        agg = list(agg)
        total_general = sum(float(a['total'] or 0) for a in agg)
        categorias = []
        for a in agg[:20]:
            nombre = a['ProductoTalla__producto__categoria__nombre'] or 'Sin categoría'
            padre = a['ProductoTalla__producto__categoria__padre__nombre'] or ''
            total = float(a['total'] or 0)
            categorias.append({
                'categoria': nombre,
                'padre': padre,
                'label': (padre + ' › ' + nombre) if padre else nombre,
                'cantidad': a['cantidad'] or 0,
                'total': total,
                'participacion': round(total / total_general * 100, 1) if total_general else 0,
            })
        return JsonResponse({'success': True, 'categorias': categorias})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error ventas por categoría: {str(e)}'}, status=500)


@require_GET
@login_required
def obtener_ventas_por_especialidad(request):
    """Ventas ($ y unidades) por especialidad v1.2 (deporte/uso). Multi-etiqueta:
    un producto con [running, urbano] suma a AMBAS (es una vista de atribución,
    no una partición; la suma total puede exceder las ventas). Opcional:
    ?categoria_id= (padre incluye hijas) para acotar el deporte a una categoría."""
    try:
        categoria_id = request.GET.get('categoria_id')
        ticket_ids = _tickets_pagados_periodo(request).values_list('id', flat=True)
        lineas = (Ticket_Productos.objects
                  .filter(idTicket_id__in=ticket_ids, ProductoTalla__isnull=False)
                  .exclude(ProductoTalla__producto__excluir_de_analitica=True))
        if categoria_id:
            from app.models import Categoria
            cat_ids = [int(categoria_id)]
            cat_ids += list(Categoria.objects.filter(padre_id=categoria_id).values_list('id', flat=True))
            lineas = lineas.filter(ProductoTalla__producto__categoria_id__in=cat_ids)
        lineas = lineas.filter(
            ProductoTalla__producto__atributos__atributo__nombre__iexact='Especialidad')
        agg = (lineas.values('ProductoTalla__producto__atributos__opcion__valor')
               .annotate(cantidad=Sum('stock'), total=Sum('subtotal'))
               .order_by('-total'))
        agg = list(agg)
        especialidades = [{
            'slug': a['ProductoTalla__producto__atributos__opcion__valor'] or '',
            'cantidad': a['cantidad'] or 0,
            'total': float(a['total'] or 0),
        } for a in agg[:25] if a['ProductoTalla__producto__atributos__opcion__valor']]
        return JsonResponse({'success': True, 'especialidades': especialidades})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error ventas por especialidad: {str(e)}'}, status=500)


@require_GET
@login_required
def obtener_indicador_compra_categoria(request):
    """Indicador de compra por categoría v1.2: cruza ventas (unidades del período)
    contra stock actual → cobertura en días → veredicto (¿qué comprar / qué sobra?).

    cobertura_dias = stock_actual / (unidades_vendidas / dias_periodo)
        · muy baja + hay ventas  → QUIEBRE  (reponer / comprar)
        · alta o sin ventas       → SOBRE-STOCK / SIN ROTACIÓN (no comprar, liquidar)

    Respeta los mismos filtros del dashboard (fecha, sucursal, estado) y excluye
    excluir_de_analitica. Solo categorías hijas v1.2 (con padre)."""
    try:
        from app.models import Producto_Talla
        fecha_inicio, fecha_fin = _rango_periodo(request)
        dias = max(1, (fecha_fin - fecha_inicio).days + 1)
        sucursal_id = request.GET.get('sucursal_id')

        # 1) Ventas del período (unidades y monto) por categoría hija
        ticket_ids = _tickets_pagados_periodo(request).values_list('id', flat=True)
        ventas_agg = (Ticket_Productos.objects
                      .filter(idTicket_id__in=ticket_ids, ProductoTalla__isnull=False)
                      .exclude(ProductoTalla__producto__excluir_de_analitica=True)
                      .filter(ProductoTalla__producto__categoria__padre__isnull=False)
                      .values('ProductoTalla__producto__categoria_id')
                      .annotate(unidades=Sum('stock'), monto=Sum('subtotal')))
        ventas = {v['ProductoTalla__producto__categoria_id']:
                  (int(v['unidades'] or 0), float(v['monto'] or 0)) for v in ventas_agg}

        # 2) Stock actual por categoría hija (mismo scope de sucursal)
        #    Sin sucursal específica excluimos los centros de distribución: su
        #    stock es mayorista y no vende a público, así la cobertura refleja
        #    la rotación real del punto de venta (no el bodegón).
        stock_qs = (Producto_Talla.objects
                    .filter(producto__categoria__padre__isnull=False)
                    .exclude(producto__excluir_de_analitica=True))
        if sucursal_id:
            stock_qs = stock_qs.filter(producto__sucursal_id=sucursal_id)
        else:
            empresa_id = request.GET.get('empresa_id')
            if empresa_id:
                stock_qs = stock_qs.filter(producto__sucursal__empresa_id=empresa_id)
            stock_qs = stock_qs.exclude(producto__sucursal__es_centro_distribucion=True)
        stock_agg = (stock_qs.values(
                         'producto__categoria_id',
                         'producto__categoria__nombre',
                         'producto__categoria__padre__nombre')
                     # OJO: el alias no puede llamarse 'stock' — sombrearía al
                     # campo dentro de la expresión de 'dinero'.
                     .annotate(stock_total=Sum('stock'),
                               # $ a costo del stock parado (solo filas positivas)
                               dinero=Sum(F('stock') * F('producto__costo'),
                                          filter=Q(stock__gt=0))))

        indicadores = []
        for s in stock_agg:
            cid = s['producto__categoria_id']
            stock = max(0, int(s['stock_total'] or 0))
            dinero = float(s['dinero'] or 0)
            unidades, monto = ventas.get(cid, (0, 0.0))
            if stock == 0 and unidades == 0:
                continue  # categoría sin stock ni ventas: irrelevante
            venta_diaria = unidades / dias if unidades else 0
            if venta_diaria > 0:
                cobertura = round(stock / venta_diaria, 1)
            else:
                cobertura = None  # sin rotación
            # Sell-through del período: vendido / (vendido + stock). En un
            # catálogo sobre-stockeado la cobertura satura (todo >120d) y es el
            # ST% el que discrimina qué rota de verdad (rango real 5%-28%).
            st_pct = round(unidades / (unidades + stock) * 100, 1) if (unidades + stock) else 0.0
            # Veredicto accionable
            if unidades == 0:
                veredicto, color, orden = ('SIN ROTACIÓN', 'muerto', 4)
            elif cobertura is not None and cobertura < 15:
                veredicto, color, orden = ('QUIEBRE', 'quiebre', 0)
            elif cobertura is not None and cobertura < 30:
                veredicto, color, orden = ('BAJO', 'bajo', 1)
            elif cobertura is not None and cobertura <= 120:
                veredicto, color, orden = ('SANO', 'sano', 2)
            else:
                veredicto, color, orden = ('SOBRE-STOCK', 'sobre', 3)
            nombre = s['producto__categoria__nombre'] or 'Sin categoría'
            padre = s['producto__categoria__padre__nombre'] or ''
            indicadores.append({
                'categoria_id': cid,
                'categoria': nombre,
                'padre': padre,
                'label': (padre + ' › ' + nombre) if padre else nombre,
                'unidades': unidades,
                'monto': monto,
                'stock': stock,
                'dinero_inmovilizado': dinero,
                'sell_through': st_pct,
                'cobertura_dias': cobertura,
                'veredicto': veredicto,
                'color': color,
                '_orden': orden,
            })
        # Comprar-primero: quiebre/bajo arriba (menor cobertura primero); entre
        # iguales, mayor sell-through primero (lo que mejor rota encabeza).
        indicadores.sort(key=lambda x: (x['_orden'],
                                        x['cobertura_dias'] if x['cobertura_dias'] is not None else 1e9,
                                        -x['sell_through']))
        for x in indicadores:
            x.pop('_orden', None)
        return JsonResponse({'success': True, 'dias': dias, 'indicadores': indicadores})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error indicador de compra: {str(e)}'}, status=500)


@require_GET
@login_required
@cache_ventas_json('mix_sucursal_v2', timeout=120, vary_on_session=True)
def obtener_mix_por_sucursal(request):
    """Comparativo ENTRE tiendas (siempre todas las visibles del usuario, sin
    centros de distribución). IGNORA a propósito sucursal_id y vendedor_id: la
    sección compara tiendas, no filtra a una. SÍ respeta fecha/estado/categoría.

    GET:
      · dimension = especialidad | categoria   (default especialidad)
      · top_n     = nº de segmentos globales    (default 6, máx 12)
      · categoria_id                            (acota el mix a una categoría)
      · fecha_inicio_comp / fecha_fin_comp      (opcionales → variación % por tienda)

    El % del mix se normaliza sobre la suma de segmentos de cada tienda (vista de
    atribución: un producto multi-etiqueta suma a todas sus especialidades), de
    modo que cada barra suma 100% y es comparable entre tiendas. `cobertura_pct`
    expone qué proporción de las unidades de la tienda tiene etiqueta (control de
    calidad; es una cota superior porque las multi-etiqueta se cuentan N veces)."""
    try:
        from app.models import Categoria
        dimension = request.GET.get('dimension', 'especialidad')
        if dimension not in ('especialidad', 'categoria'):
            return JsonResponse({'success': False, 'error': 'dimension inválida'}, status=400)
        top_n = max(1, min(int(request.GET.get('top_n', 6)), 12))

        # Universo de tiendas: permisos server-side (no se puede burlar por GET),
        # sin centros de distribución (stock mayorista, no venden a público).
        # Si viene empresa_id, se acota a las tiendas de esa empresa (el filtro
        # Empresa del dashboard); sucursal_id se ignora aquí (comparamos tiendas).
        suc_qs = obtener_sucursales_usuario(request.user).filter(es_centro_distribucion=False)
        empresa_id = request.GET.get('empresa_id')
        if empresa_id:
            suc_qs = suc_qs.filter(empresa_id=empresa_id)
        sucursales = list(suc_qs.values('id', 'alias'))
        suc_ids = [s['id'] for s in sucursales]
        if not suc_ids:
            return JsonResponse({'success': True, 'dimension': dimension,
                                 'segmentos': [], 'sucursales': []})

        fecha_inicio, fecha_fin = _rango_periodo(request)
        estado = request.GET.get('estado') or 'PAGADO'  # mezclar ANULADO distorsiona el mix

        def _tickets(fi, ff):
            # Sin tickets de CAMBIO_DEVOLUCION: la diferencia cobrada en un
            # cambio no es venta de la tienda (mismo criterio que el resto del
            # dashboard, el POS y el reporte de ventas).
            return Ticket.objects.filter(
                created_at__date__gte=fi, created_at__date__lte=ff,
                estado=estado, sucursal_id__in=suc_ids,
            ).exclude(modulo_origen=MODULO_ORIGEN_NO_VENTA)

        # ── Totales por tienda (período actual) ──
        tot_map = {r['sucursal_id']: r for r in
                   _tickets(fecha_inicio, fecha_fin).values('sucursal_id')
                   .annotate(total=Sum('total'), cantidad=Count('id'))}

        # ── Comparativo opcional (para variación %) ──
        tot_comp = {}
        fic, ffc = request.GET.get('fecha_inicio_comp'), request.GET.get('fecha_fin_comp')
        if fic and ffc:
            try:
                fic = timezone.datetime.strptime(fic, '%Y-%m-%d').date()
                ffc = timezone.datetime.strptime(ffc, '%Y-%m-%d').date()
                tot_comp = {r['sucursal_id']: float(r['total'] or 0) for r in
                            _tickets(fic, ffc).values('sucursal_id').annotate(total=Sum('total'))}
            except ValueError:
                pass  # comparativo malformado → sin variación

        # ── Líneas del período para el mix ──
        ticket_ids = _tickets(fecha_inicio, fecha_fin).values_list('id', flat=True)
        lineas = (Ticket_Productos.objects
                  .filter(idTicket_id__in=ticket_ids, ProductoTalla__isnull=False)
                  .exclude(ProductoTalla__producto__excluir_de_analitica=True))
        categoria_id = request.GET.get('categoria_id')
        if categoria_id:
            cat_ids = [int(categoria_id)]
            cat_ids += list(Categoria.objects.filter(padre_id=categoria_id)
                            .values_list('id', flat=True))
            lineas = lineas.filter(ProductoTalla__producto__categoria_id__in=cat_ids)

        # Unidades totales por tienda (ANTES de unir atributos → sin fanout)
        uds_totales = {r['idTicket__sucursal_id']: int(r['u'] or 0) for r in
                       lineas.values('idTicket__sucursal_id').annotate(u=Sum('stock'))}

        if dimension == 'especialidad':
            lineas_dim = lineas.filter(
                ProductoTalla__producto__atributos__atributo__nombre__iexact='Especialidad')
            clave_expr = F('ProductoTalla__producto__atributos__opcion__valor')
        else:
            lineas_dim = lineas
            clave_expr = Coalesce(F('ProductoTalla__producto__categoria__padre__nombre'),
                                  F('ProductoTalla__producto__categoria__nombre'))

        agg = list(lineas_dim.values('idTicket__sucursal_id', clave=clave_expr)
                   .annotate(total=Sum('subtotal'), cantidad=Sum('stock')))

        # ── Top-N GLOBAL por total $ del período (mismos segmentos/colores en todas las barras) ──
        tot_por_clave = {}
        for r in agg:
            if r['clave']:
                tot_por_clave[r['clave']] = tot_por_clave.get(r['clave'], 0) + float(r['total'] or 0)
        segmentos = [k for k, _ in sorted(tot_por_clave.items(), key=lambda kv: -kv[1])[:top_n]]
        seg_set = set(segmentos)
        hay_otras = len(tot_por_clave) > len(segmentos)

        # ── Mix por tienda: bucket 'otras' + % sobre la suma de segmentos de la tienda ──
        mix_suc = {sid: {} for sid in suc_ids}
        uds_tag = {sid: 0 for sid in suc_ids}
        for r in agg:
            sid, clave = r['idTicket__sucursal_id'], r['clave']
            if not clave:
                continue
            seg = clave if clave in seg_set else 'otras'
            d = mix_suc[sid].setdefault(seg, {'total': 0.0, 'cantidad': 0})
            d['total'] += float(r['total'] or 0)
            d['cantidad'] += int(r['cantidad'] or 0)
            uds_tag[sid] += int(r['cantidad'] or 0)  # atribución (multi-tag cuenta N veces)

        data = []
        for s in sucursales:
            sid = s['id']
            t = tot_map.get(sid, {})
            total = float(t.get('total') or 0)
            cant = int(t.get('cantidad') or 0)
            base_mix = sum(v['total'] for v in mix_suc[sid].values())
            mix = {seg: {'total': v['total'], 'cantidad': v['cantidad'],
                         'pct': round(v['total'] / base_mix * 100, 1) if base_mix else 0}
                   for seg, v in mix_suc[sid].items()}
            comp = tot_comp.get(sid)
            if comp is None:
                variacion = None
            elif comp > 0:
                variacion = round((total - comp) / comp * 100, 1)
            else:
                variacion = 100.0 if total > 0 else None
            ut = uds_totales.get(sid, 0)
            data.append({
                'id': sid, 'alias': s['alias'], 'total': total, 'cantidad': cant,
                'ticket_promedio': total / cant if cant else 0,
                'variacion_pct': variacion, 'mix': mix,
                'cobertura_pct': round(min(uds_tag[sid], ut) / ut * 100, 1) if ut else None,
            })
        data.sort(key=lambda x: -x['total'])
        if hay_otras:
            segmentos = segmentos + ['otras']
        return JsonResponse({'success': True, 'dimension': dimension,
                             'segmentos': segmentos, 'sucursales': data})
    except Exception as e:
        logger.error(f'Error mix por sucursal: {e}')
        return JsonResponse({'success': False, 'error': f'Error mix por sucursal: {str(e)}'}, status=500)


@require_GET
@login_required
@cache_ventas_json('tendencias_ventas_v2', timeout=120)
def obtener_tendencias_ventas(request):
    """
    API para obtener tendencias de ventas
    Incluye: ventas por hora y por día de la semana.

    Dos correcciones respecto de la versión anterior:
      · Usaba `Ticket.hora` y `Ticket.fecha`, ambos `auto_now`: la "hora pico"
        era la hora del último guardado, no la de la venta. Ahora sale de
        `created_at` (ExtractHour/ExtractIsoWeekDay, convertido por la BD a
        America/Santiago).
      · Iteraba en Python TODOS los tickets del período (dos veces). Ahora son
        dos agregados en la base de datos.
    """
    from django.db.models.functions import ExtractHour, ExtractIsoWeekDay

    try:
        sucursal_id = request.GET.get('sucursal_id')

        # Validar fechas
        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        queryset = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        # Ventas por hora del día (0-23) — un solo GROUP BY
        ventas_por_hora = [0.0] * 24
        cantidad_por_hora = [0] * 24
        for fila in (queryset.annotate(h=ExtractHour('created_at'))
                     .values('h').annotate(total=Sum('total'), cantidad=Count('id'))
                     .order_by('h')):
            h = fila['h']
            if h is None or not (0 <= h <= 23):
                continue
            ventas_por_hora[h] = float(fila['total'] or 0)
            cantidad_por_hora[h] = fila['cantidad']

        por_hora_data = [
            {'hora': i, 'total': ventas_por_hora[i], 'cantidad': cantidad_por_hora[i]}
            for i in range(24)
        ]

        # Ventas por día de la semana (0=Lunes … 6=Domingo, igual que antes).
        # ExtractIsoWeekDay devuelve 1=Lunes … 7=Domingo.
        ventas_por_dia = [0.0] * 7
        for fila in (queryset.annotate(d=ExtractIsoWeekDay('created_at'))
                     .values('d').annotate(total=Sum('total'))
                     .order_by('d')):
            d = fila['d']
            if d is None or not (1 <= d <= 7):
                continue
            ventas_por_dia[d - 1] = float(fila['total'] or 0)

        return JsonResponse({
            'success': True,
            'por_hora': por_hora_data,
            'por_dia_semana': ventas_por_dia
        })

    except Exception as e:
        logger.error('Error al obtener tendencias de ventas: %s', e)
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener tendencias de ventas: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('ind_avanzados_v2', timeout=120)
def obtener_indicadores_avanzados_ventas(request):
    """
    API para obtener indicadores avanzados de retail con datos reales.
    Calcula: Margen Bruto (FIFO), Sell-Through Rate, Rotacion, Dias de Stock,
    GMROI, Descuento Promedio, Costo de Ventas real.

    BASE DEL MARGEN (elegida y documentada, antes estaba mezclada):
      · Ingreso = Sum('subtotal') / 1.19 → NETO de IVA y YA con los descuentos
        aplicados. Antes era Σ(stock × precio): precio BRUTO (con IVA) y sin
        restar `descuento_unitario`, contra un costo FIFO que es NETO. El margen
        salía inflado ~19 puntos por el IVA más lo que sobraba por los descuentos.
      · Costo = Σ(stock × costo_fifo) SOLO sobre las líneas con costo_fifo > 0.
        Una línea sin costo FIFO no aporta "100% de margen": queda fuera del
        cálculo y se refleja en `cobertura_costeo_pct`.
      · Si NO hay ninguna línea costeada, el margen NO se inventa: `margen_bruto`,
        `margen_pct` y `gmroi` vuelven en null con `margen_calculable=false` y
        `margen_nota` explicando por qué.
    Universo de líneas: excluye productos con `excluir_de_analitica` (mismo
    criterio que los charts v1.2, el top de productos y el indicador de compra),
    por lo que `ingresos` puede quedar levemente por debajo de "Ventas Totales".
    """
    try:
        sucursal_id = request.GET.get('sucursal_id')

        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        tickets_qs = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        ticket_ids = tickets_qs.values_list('id', flat=True)
        # Denominador del UPT: transacciones del MISMO universo que las unidades.
        cantidad_tickets = tickets_qs.count()

        lineas = (Ticket_Productos.objects
                  .filter(idTicket_id__in=ticket_ids)
                  .exclude(ProductoTalla__producto__excluir_de_analitica=True))

        costeada = Q(costo_fifo__gt=0)
        agg = lineas.aggregate(
            ingresos_brutos=Sum('subtotal'),
            ingresos_brutos_costeados=Sum('subtotal', filter=costeada),
            costo_ventas=Sum(
                ExpressionWrapper(F('stock') * F('costo_fifo'), output_field=DecimalField()),
                filter=costeada),
            unidades_vendidas=Sum('stock'),
            unidades_costeadas=Sum('stock', filter=costeada),
            descuento_prom=Avg('porcentaje_descuento'),
            descuento_total_monto=Sum(ExpressionWrapper(F('stock') * F('descuento_unitario'), output_field=DecimalField())),
        )

        ingresos = float(agg['ingresos_brutos'] or 0) / IVA_FACTOR
        ingresos_costeados = float(agg['ingresos_brutos_costeados'] or 0) / IVA_FACTOR
        costo_ventas = float(agg['costo_ventas'] or 0)
        unidades_vendidas = int(agg['unidades_vendidas'] or 0)
        unidades_costeadas = int(agg['unidades_costeadas'] or 0)
        descuento_promedio = float(agg['descuento_prom'] or 0)
        descuento_total_monto = float(agg['descuento_total_monto'] or 0)

        cobertura_costeo = (unidades_costeadas / unidades_vendidas * 100) if unidades_vendidas > 0 else 0.0
        margen_calculable = unidades_costeadas > 0 and ingresos_costeados > 0
        if margen_calculable:
            margen_bruto = ingresos_costeados - costo_ventas
            margen_pct = margen_bruto / ingresos_costeados * 100
            margen_nota = (
                'Margen neto de IVA sobre las líneas con costo FIFO '
                f'({cobertura_costeo:.1f}% de las unidades del período).'
            )
        else:
            margen_bruto = None
            margen_pct = None
            margen_nota = (
                'No calculable: ninguna línea vendida del período tiene costo FIFO '
                '(costo_fifo = 0). Antes esto se mostraba como 100% de margen.'
            )
            logger.warning(
                'Dashboard ventas: margen no calculable %s..%s (0 de %s unidades con costo FIFO)',
                fecha_inicio, fecha_fin, unidades_vendidas,
            )

        stock_filter = {}
        if sucursal_id:
            stock_filter['producto__sucursal_id'] = sucursal_id
        elif request.GET.get('empresa_id'):
            stock_filter['producto__sucursal__empresa_id'] = request.GET.get('empresa_id')
        stock_actual = Producto_Talla.objects.filter(
            stock__gt=0, **stock_filter
        ).aggregate(
            total_unidades=Sum('stock'),
        )
        stock_total_unidades = int(stock_actual['total_unidades'] or 0)

        sell_through = 0
        if (unidades_vendidas + stock_total_unidades) > 0:
            sell_through = (unidades_vendidas / (unidades_vendidas + stock_total_unidades)) * 100

        dias_periodo = max(1, (fecha_fin - fecha_inicio).days + 1)

        if stock_total_unidades > 0 and unidades_vendidas > 0:
            venta_diaria = unidades_vendidas / dias_periodo
            dias_stock = stock_total_unidades / venta_diaria
            rotacion_periodo = unidades_vendidas / stock_total_unidades
            rotacion_anualizada = rotacion_periodo * (365 / dias_periodo)
        else:
            dias_stock = 0
            rotacion_periodo = 0
            rotacion_anualizada = 0

        # GMROI = margen bruto / inversión en inventario, valorizado con el costo
        # unitario de lo efectivamente costeado. Si el margen no es calculable el
        # GMROI tampoco lo es (antes devolvía 0, que se leía como "malísimo").
        if margen_calculable and unidades_costeadas > 0:
            inventario_costo_est = stock_total_unidades * (costo_ventas / unidades_costeadas)
            gmroi = (margen_bruto / inventario_costo_est) if inventario_costo_est > 0 else None
        else:
            inventario_costo_est = 0
            gmroi = None

        return JsonResponse({
            'success': True,
            'ingresos': ingresos,                       # neto de IVA, con descuentos
            'ingresos_costeados': ingresos_costeados,   # base real del margen
            'costo_ventas': costo_ventas,
            'margen_bruto': margen_bruto,
            'margen_pct': round(margen_pct, 2) if margen_pct is not None else None,
            'margen_calculable': margen_calculable,
            'margen_nota': margen_nota,
            'cobertura_costeo_pct': round(cobertura_costeo, 1),
            'unidades_vendidas': unidades_vendidas,
            'unidades_costeadas': unidades_costeadas,
            'unidades_sin_costo': unidades_vendidas - unidades_costeadas,
            'upt': round(unidades_vendidas / cantidad_tickets, 2) if cantidad_tickets else 0,
            'cantidad_tickets': cantidad_tickets,
            'stock_actual': stock_total_unidades,
            'sell_through': round(sell_through, 2),
            'rotacion_periodo': round(rotacion_periodo, 2),
            'rotacion_anualizada': round(rotacion_anualizada, 2),
            'dias_stock': round(dias_stock, 1),
            'gmroi': round(gmroi, 2) if gmroi is not None else None,
            'descuento_promedio': round(descuento_promedio, 2),
            'descuento_total_monto': descuento_total_monto,
            'base_calculo': (
                'created_at · estado PAGADO por defecto · sin CAMBIO_DEVOLUCION · '
                'sin productos excluidos de analítica · ingreso neto de IVA'
            ),
        })

    except Exception as e:
        logger.error('Error al obtener indicadores avanzados: %s', e)
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener indicadores avanzados: {str(e)}'
        }, status=500)


@require_GET
@login_required
@cache_ventas_json('estado_operacional_v2', timeout=60)
def obtener_estado_operacional_ventas(request):
    """
    API para obtener el estado operacional completo del modulo de ventas.
    Cubre: tickets por estado, ventas por modulo, POS, cambios/devoluciones,
    depositos, DTEs pendientes, regularizaciones.

    Panel de OPERACIÓN, no de venta: conserva a propósito todos los estados y
    todos los `modulo_origen` (incluido CAMBIO_DEVOLUCION, que es justamente una
    de las filas del desglose). Lo único que cambia es la fecha: `created_at` en
    vez de `Ticket.fecha` (auto_now).
    """
    try:
        sucursal_id = request.GET.get('sucursal_id')

        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido'
            }, status=400)

        tickets_qs = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin,
            aplicar_estado=False, excluir_cambios=False, sucursal_id=sucursal_id)

        # --- Tickets por estado ---
        tickets_por_estado = list(
            tickets_qs.values('estado').annotate(
                cantidad=Count('id'),
                monto=Sum('total')
            ).order_by('estado')
        )
        total_tickets = tickets_qs.count()
        anulados = tickets_qs.filter(estado='ANULADO').count()
        pct_anulados = (anulados / total_tickets * 100) if total_tickets > 0 else 0

        pendientes_pago = tickets_qs.filter(estado='PENDIENTE')
        pendientes_count = pendientes_pago.count()
        pendientes_monto = float(pendientes_pago.aggregate(t=Sum('total'))['t'] or 0)

        # --- Ventas por modulo de origen ---
        por_modulo = list(
            tickets_qs.values('modulo_origen').annotate(
                cantidad=Count('id'),
                monto=Sum('total')
            ).order_by('-monto')
        )
        for item in por_modulo:
            item['monto'] = float(item['monto'] or 0)

        # --- Ventas por tipo DTE ---
        por_tipo_dte = list(
            tickets_qs.exclude(tipo_dte__isnull=True).values('tipo_dte').annotate(
                cantidad=Count('id')
            ).order_by('-cantidad')
        )

        # --- Tickets offline ---
        tickets_offline = tickets_qs.filter(created_offline=True).count()

        # --- Transacciones POS ---
        pos_qs = TransaccionPOS.objects.filter(
            fecha_inicio__date__gte=fecha_inicio,
            fecha_inicio__date__lte=fecha_fin
        )
        if sucursal_id:
            pos_qs = pos_qs.filter(configuracion_pos__sucursal_id=sucursal_id)
        elif request.GET.get('empresa_id'):
            pos_qs = pos_qs.filter(configuracion_pos__sucursal__empresa_id=request.GET.get('empresa_id'))

        pos_por_estado = list(
            pos_qs.values('estado').annotate(
                cantidad=Count('id'),
                monto=Sum('monto')
            ).order_by('estado')
        )
        for item in pos_por_estado:
            item['monto'] = float(item['monto'] or 0)

        pos_total = pos_qs.count()
        pos_completadas = pos_qs.filter(estado='COMPLETADA').count()
        pos_tasa_exito = (pos_completadas / pos_total * 100) if pos_total > 0 else 0

        # --- Cambios y Devoluciones ---
        cambios_qs = CambioDevolucion.objects.filter(
            fecha_solicitud__date__gte=fecha_inicio,
            fecha_solicitud__date__lte=fecha_fin
        )
        cambios_qs = _scope_suc_emp(cambios_qs, request, sucursal_id)

        cambios_por_estado = list(
            cambios_qs.values('estado').annotate(
                cantidad=Count('id'),
                monto=Sum('monto_original')
            ).order_by('estado')
        )
        for item in cambios_por_estado:
            item['monto'] = float(item['monto'] or 0)
            item['estado_display'] = dict(ESTADO_CAMBIO_CHOICES).get(item['estado'], item['estado'])

        cambios_por_tipo = list(
            cambios_qs.values('tipo_operacion').annotate(
                cantidad=Count('id'),
                monto=Sum('monto_original')
            ).order_by('-cantidad')
        )
        for item in cambios_por_tipo:
            item['monto'] = float(item['monto'] or 0)
            item['tipo_display'] = dict(TIPO_OPERACION_CAMBIO_CHOICES).get(item['tipo_operacion'], item['tipo_operacion'])

        motivos = list(
            cambios_qs.filter(motivo_principal__isnull=False)
            .values('motivo_principal')
            .annotate(cantidad=Count('id'))
            .order_by('-cantidad')[:5]
        )
        for item in motivos:
            item['motivo_display'] = dict(MOTIVO_CAMBIO_CHOICES).get(item['motivo_principal'], item['motivo_principal'])

        cambios_monto_total = float(cambios_qs.aggregate(t=Sum('monto_original'))['t'] or 0)
        cambios_pendientes_aprobacion = cambios_qs.filter(estado__in=['SOLICITADO', 'EN_PROCESO']).count()

        # --- Depositos Bancarios ---
        depositos_qs = DepositoBancario.objects.filter(
            fecha_deposito__gte=fecha_inicio,
            fecha_deposito__lte=fecha_fin
        )
        if sucursal_id:
            depositos_qs = depositos_qs.filter(arqueo__sucursal_id=sucursal_id)

        depositos_verificados = depositos_qs.filter(verificado=True).count()
        depositos_pendientes = depositos_qs.filter(verificado=False).count()
        depositos_monto_verificado = float(
            depositos_qs.filter(verificado=True).aggregate(t=Sum('monto'))['t'] or 0
        )
        depositos_monto_pendiente = float(
            depositos_qs.filter(verificado=False).aggregate(t=Sum('monto'))['t'] or 0
        )

        # --- DTEs pendientes ---
        dtes_pendientes = Dte.objects.filter(
            estado_dte='EMITIDO',
            tipo_transaccion='TRASPASO'
        ).count()

        # --- Regularizaciones pendientes ---
        from .models import Solicitud_Regularizacion
        regularizaciones_pendientes = Solicitud_Regularizacion.objects.filter(
            estado__in=['PENDIENTE', 'EN_REVISION']
        ).count()

        return JsonResponse({
            'success': True,
            'tickets': {
                'por_estado': [{
                    'estado': t['estado'],
                    'cantidad': t['cantidad'],
                    'monto': float(t['monto'] or 0)
                } for t in tickets_por_estado],
                'total': total_tickets,
                'anulados': anulados,
                'pct_anulados': round(pct_anulados, 2),
                'pendientes_pago': pendientes_count,
                'pendientes_monto': pendientes_monto,
            },
            'modulo_origen': por_modulo,
            'tipo_dte': por_tipo_dte,
            'tickets_offline': tickets_offline,
            'pos': {
                'por_estado': pos_por_estado,
                'total': pos_total,
                'completadas': pos_completadas,
                'tasa_exito': round(pos_tasa_exito, 2),
            },
            'cambios_devoluciones': {
                'por_estado': cambios_por_estado,
                'por_tipo': cambios_por_tipo,
                'motivos_principales': motivos,
                'monto_total': cambios_monto_total,
                'pendientes_aprobacion': cambios_pendientes_aprobacion,
            },
            'depositos': {
                'verificados': depositos_verificados,
                'pendientes': depositos_pendientes,
                'monto_verificado': depositos_monto_verificado,
                'monto_pendiente': depositos_monto_pendiente,
            },
            'dtes_pendientes': dtes_pendientes,
            'regularizaciones_pendientes': regularizaciones_pendientes,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estado operacional: {str(e)}'
        }, status=500)


@require_GET
@login_required
def exportar_dashboard_ventas_excel(request):
    """
    API para exportar dashboard de ventas a Excel
    Incluye todas las métricas e indicadores.

    Consume la MISMA base que la pantalla (`_tickets_venta_periodo`): antes
    filtraba por `Ticket.fecha` (auto_now), forzaba estado PAGADO ignorando el
    filtro, no aplicaba empresa_id, contaba los tickets de cambio/devolución y
    valorizaba los productos con stock × precio mientras la tabla en pantalla usa
    Sum('subtotal'). El archivo que se mandaba a gerencia contradecía al
    dashboard del que salía.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Obtener parámetros de filtro
        sucursal_id = request.GET.get('sucursal_id')

        # Validar fechas
        try:
            fecha_inicio, fecha_fin = _rango_periodo(request)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }, status=400)

        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        
        # ===== HOJA 1: RESUMEN EJECUTIVO =====
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        
        ws1['A1'] = "DASHBOARD DE VENTAS - RESUMEN EJECUTIVO"
        ws1['A1'].font = title_font
        ws1['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        # Obtener datos de indicadores globales — misma base que la pantalla
        queryset = _tickets_venta_periodo(
            request, fecha_inicio, fecha_fin, sucursal_id=sucursal_id)

        ventas_totales = queryset.aggregate(total=Sum('total'))['total'] or 0
        cantidad_ventas = queryset.count()
        ticket_promedio = ventas_totales / cantidad_ventas if cantidad_ventas > 0 else 0

        ws1['A4'] = "INDICADORES PRINCIPALES"
        ws1['A4'].font = header_font
        ws1['A4'].fill = header_fill

        ws1['A5'] = "Ventas Totales"
        ws1['B5'] = f"${ventas_totales:,.0f}"
        ws1['A6'] = "Cantidad de Ventas"
        ws1['B6'] = cantidad_ventas
        ws1['A7'] = "Ticket Promedio"
        ws1['B7'] = f"${ticket_promedio:,.0f}"
        ws1['A8'] = "Base de cálculo"
        ws1['B8'] = (
            f"Fecha de venta (created_at) · estado "
            f"{request.GET.get('estado') or 'PAGADO'} · sin cambios/devoluciones"
        )

        # ===== HOJA 2: VENTAS POR VENDEDOR =====
        ws2 = wb.create_sheet("Ventas por Vendedor")
        
        headers_vendedor = ["Código", "Vendedor", "Cant. Ventas", "Total Vendido", 
                           "Ticket Promedio", "Comisión %", "Comisión Total", "% Participación"]
        
        for col, header in enumerate(headers_vendedor, 1):
            cell = ws2.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        ventas_vendedor = queryset.values(
            'vendedor__codigo_vendedor',
            'vendedor__nombre',
            'vendedor__comision'
        ).annotate(
            total_vendido=Sum('total'),
            cantidad_ventas=Count('id'),
            ticket_promedio=Avg('total')
        ).order_by('-total_vendido')
        
        row = 2
        for venta in ventas_vendedor:
            total_vendido = float(venta['total_vendido'] or 0)
            comision_porcentaje = float(venta['vendedor__comision'] or 0)
            comision_total = total_vendido * (comision_porcentaje / 100)
            participacion = (total_vendido / ventas_totales * 100) if ventas_totales > 0 else 0
            
            ws2.cell(row, 1, venta['vendedor__codigo_vendedor'])
            ws2.cell(row, 2, venta['vendedor__nombre'])
            ws2.cell(row, 3, venta['cantidad_ventas'])
            ws2.cell(row, 4, f"${total_vendido:,.0f}")
            ws2.cell(row, 5, f"${float(venta['ticket_promedio']):,.0f}")
            ws2.cell(row, 6, f"{comision_porcentaje:.2f}%")
            ws2.cell(row, 7, f"${comision_total:,.0f}")
            ws2.cell(row, 8, f"{participacion:.2f}%")
            row += 1
        
        # ===== HOJA 3: PRODUCTOS MÁS VENDIDOS =====
        ws3 = wb.create_sheet("Productos Más Vendidos")
        
        headers_productos = ["#", "SKU", "Producto", "Categoría", "Cantidad", "Total Ventas", 
                            "Precio Promedio", "% Participación"]
        
        for col, header in enumerate(headers_productos, 1):
            cell = ws3.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        ticket_ids = queryset.values_list('id', flat=True)

        # Sum('subtotal') = ingreso real tras descuentos, igual que la tabla en
        # pantalla (antes stock × precio, que ignora `descuento_unitario`), y sin
        # los productos marcados como excluidos de analítica.
        productos_vendidos = list(Ticket_Productos.objects.filter(
            idTicket_id__in=ticket_ids
        ).exclude(
            ProductoTalla__producto__excluir_de_analitica=True
        ).values(
            'ProductoTalla__sku',
            'ProductoTalla__producto__articulo',
            'ProductoTalla__producto__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('stock'),
            total_ventas=Sum('subtotal')
        ).order_by('-cantidad_vendida')[:50])

        total_productos = sum(float(p['total_ventas'] or 0) for p in productos_vendidos)
        
        row = 2
        for idx, producto in enumerate(productos_vendidos, 1):
            total_ventas_prod = float(producto['total_ventas'] or 0)
            cantidad = producto['cantidad_vendida'] or 0
            precio_prom = total_ventas_prod / cantidad if cantidad > 0 else 0
            participacion = (total_ventas_prod / total_productos * 100) if total_productos > 0 else 0
            
            ws3.cell(row, 1, idx)
            ws3.cell(row, 2, producto['ProductoTalla__sku'])
            ws3.cell(row, 3, producto['ProductoTalla__producto__articulo'])
            ws3.cell(row, 4, producto['ProductoTalla__producto__categoria__nombre'] or 'Sin categoría')
            ws3.cell(row, 5, cantidad)
            ws3.cell(row, 6, f"${total_ventas_prod:,.0f}")
            ws3.cell(row, 7, f"${precio_prom:,.0f}")
            ws3.cell(row, 8, f"{participacion:.2f}%")
            row += 1
        
        # Ajustar ancho de columnas
        for ws in [ws1, ws2, ws3]:
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Dashboard_Ventas_{fecha_inicio}_{fecha_fin}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar dashboard: {str(e)}'
        }, status=500)


# ========== NOTA DE CRÉDITO DESDE DEVOLUCIONES ==========

@login_required
@require_POST
@transaction.atomic
def generar_nc_devolucion(request):
    """
    Genera una Nota de Crédito (NC) a partir de un CambioDevolucion completado.
    La NC se vincula al DTE original del ticket y afecta la cuadratura de caja
    según el método de devolución elegido (efectivo caja o transferencia bancaria).
    """
    from datetime import date
    from decimal import Decimal
    from collections import defaultdict
    from .views_modulo_documentos import generar_txt_nota_credito_acepta, limpiar_texto, normalizar_detalle_para_tipo
    import logging
    logger = logging.getLogger(__name__)

    try:
        body = json.loads(request.body)
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    cambio_id = body.get('cambio_devolucion_id')
    metodo_devolucion = body.get('metodo_devolucion')

    if not cambio_id:
        return JsonResponse({'success': False, 'error': 'ID de cambio/devolución requerido'}, status=400)
    if metodo_devolucion not in ('EFECTIVO_CAJA', 'TRANSFERENCIA_BANCARIA'):
        return JsonResponse({'success': False, 'error': 'Método de devolución inválido. Use EFECTIVO_CAJA o TRANSFERENCIA_BANCARIA'}, status=400)

    # Validar permisos: solo admin, administracion, jefe_local
    try:
        empresa_user = EmpresaUser.objects.get(user=request.user, active=True)
    except EmpresaUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no tiene empresa asignada'}, status=403)

    rol = getattr(empresa_user, 'rol', None) or getattr(request.user, 'rol', '')
    if rol not in ('administrador', 'administracion', 'jefe_local'):
        return JsonResponse({'success': False, 'error': 'No tiene permisos para generar Notas de Crédito'}, status=403)

    # Obtener CambioDevolucion
    try:
        cambio = CambioDevolucion.objects.select_related(
            'ticket_original', 'sucursal'
        ).prefetch_related('detalles__producto_original__ProductoTalla__producto').get(id=cambio_id)
    except CambioDevolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cambio/Devolución no encontrado'}, status=404)

    # Validar estado
    estados_validos = ('COMPLETADO', 'EJECUTADO_DEVOL_PENDIENTE', 'EJECUTADO', 'EJECUTADO_COBRO_PENDIENTE')
    if cambio.estado not in estados_validos:
        return JsonResponse({
            'success': False,
            'error': f'El cambio/devolución debe estar en estado completado o ejecutado. Estado actual: {cambio.get_estado_display()}'
        }, status=400)

    # Validar tipo de operación (solo devoluciones)
    tipos_devolucion = ('DEVOLUCION_TOTAL', 'DEVOLUCION_PARCIAL')
    if cambio.tipo_operacion not in tipos_devolucion:
        return JsonResponse({
            'success': False,
            'error': 'Solo se puede generar NC para devoluciones (total o parcial)'
        }, status=400)

    # Validar que no tenga NC generada
    if cambio.nc_generada:
        return JsonResponse({
            'success': False,
            'error': 'Ya se generó una Nota de Crédito para esta devolución',
            'nota_credito_id': cambio.nota_credito_id
        }, status=400)

    # Validar sucursal del usuario
    sucursal_id = request.session.get('idSucursalActual')
    if cambio.sucursal_id != int(sucursal_id):
        return JsonResponse({'success': False, 'error': 'El cambio/devolución no pertenece a su sucursal actual'}, status=403)

    # Buscar DTE original del ticket
    ticket_original = cambio.ticket_original
    dte_original = None

    if ticket_original.folio_dte:
        dte_original = Dte.objects.filter(
            numero_documento=ticket_original.folio_dte,
            sucursal=cambio.sucursal,
            tipo_documento__in=['BOLETA ELECTRONICA', 'FACTURA ELECTRONICA', 'BOLETA PAPEL'],
            estado_dte__in=['EMITIDO', 'ACEPTADO']
        ).first()

    if not dte_original:
        dte_original = Dte.objects.filter(
            sucursal=cambio.sucursal,
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
            tipo_documento__in=['BOLETA ELECTRONICA', 'FACTURA ELECTRONICA', 'BOLETA PAPEL'],
            estado_dte__in=['EMITIDO', 'ACEPTADO'],
            fecha_emision=ticket_original.fecha
        ).order_by('-id').first()

    empresa_id = request.session.get('idEmpresaActual')
    empresa = Empresa.objects.get(id=empresa_id)

    # Calcular monto de la NC según los detalles de la devolución
    detalles = cambio.detalles.all()
    monto_devolucion = abs(cambio.diferencia_monto) if cambio.diferencia_monto < 0 else cambio.monto_original

    if cambio.tipo_operacion == 'DEVOLUCION_PARCIAL':
        monto_devolucion = sum(
            abs(d.precio_original_unitario * d.cantidad_original)
            for d in detalles if d.cantidad_original > 0
        )

    monto_neto = int(round(monto_devolucion / Decimal('1.19')))
    iva = int(monto_devolucion) - monto_neto
    monto_con_iva = int(monto_devolucion)

    # Obtener correlativo para NC
    numero_nc = obtener_siguiente_correlativo(cambio.sucursal, 'NOTA DE CREDITO')

    # Determinar tipo SII del documento original para la referencia
    tipo_sii_original = 39  # boleta por defecto
    folio_original = ''
    fecha_original = timezone.localdate().strftime('%Y-%m-%d')

    if dte_original:
        if 'FACTURA' in dte_original.tipo_documento:
            tipo_sii_original = 33
        elif 'BOLETA' in dte_original.tipo_documento:
            tipo_sii_original = 39
        folio_original = str(dte_original.numero_documento)
        fecha_original = dte_original.fecha_emision.strftime('%Y-%m-%d')
    elif ticket_original.folio_dte:
        folio_original = str(ticket_original.folio_dte)
        fecha_original = ticket_original.fecha.strftime('%Y-%m-%d')

    referencias_json = json.dumps([{
        'tipo_documento': tipo_sii_original,
        'folio': folio_original,
        'fecha': fecha_original,
        'razon': '1'
    }])

    # Crear el DTE tipo NC
    nc = Dte.objects.create(
        emisor=empresa,
        receptor=dte_original.receptor if dte_original and dte_original.receptor else None,
        tipo_documento='NOTA DE CREDITO',
        numero_documento=numero_nc,
        monto_neto=monto_neto,
        monto_con_iva=monto_con_iva,
        descuento=0,
        fecha_emision=timezone.localdate(),
        fecha_vencimiento=timezone.localdate(),
        diasCredito=0,
        bultos=0,
        unidades_productos=sum(d.cantidad_original for d in detalles if d.cantidad_original > 0),
        estado_dte='EMITIDO',
        estado_pago='PAGADO',
        tipo_transaccion='DEVOLUCION',
        responsable=request.user.username,
        sucursal=cambio.sucursal,
        hora=timezone.localtime().time(),
        es_nota_credito=True,
        documento_afectado=dte_original,
        motivo_nc=f"Devolución {cambio.get_tipo_operacion_display()} - {cambio.numero_operacion}. Motivo: {cambio.get_motivo_principal_display()}",
        referencias=referencias_json,
    )

    # Crear líneas de productos en la NC
    for detalle in detalles:
        if detalle.cantidad_original <= 0:
            continue
        if detalle.producto_original and detalle.producto_original.ProductoTalla:
            pt = detalle.producto_original.ProductoTalla
            Dte_Productos.objects.create(
                dte=nc,
                productoTalla=pt,
                descripcion=pt.producto.articulo if pt.producto else '',
                costo=0,
                sobreprecio=0,
                precio=detalle.precio_original_unitario,
                stock=detalle.cantidad_original,
                activo=True
            )

    # Crear detalle de pago de la NC según método de devolución
    metodo_pago_nc = 'EFECTIVO' if metodo_devolucion == 'EFECTIVO_CAJA' else 'TRANSFERENCIA'
    Dte_Detalle_Pago.objects.create(
        dte=nc,
        metodo_pago=metodo_pago_nc,
        monto=monto_con_iva,
    )

    # Actualizar CambioDevolucion
    cambio.nota_credito = nc
    cambio.nc_generada = True
    cambio.metodo_devolucion = metodo_devolucion
    cambio.fecha_nc = timezone.now()
    cambio.save(update_fields=['nota_credito', 'nc_generada', 'metodo_devolucion', 'fecha_nc'])

    # Registrar en historial
    HistorialCambioDevolucion.objects.create(
        cambio_devolucion=cambio,
        accion='NC_GENERADA',
        estado_anterior=cambio.estado,
        estado_nuevo=cambio.estado,
        usuario=request.user,
        descripcion=f"Nota de Crédito #{numero_nc} generada. Método: {cambio.get_metodo_devolucion_display()}. Monto: ${monto_con_iva:,}",
        datos_adicionales={
            'nc_id': nc.id,
            'nc_numero': numero_nc,
            'metodo_devolucion': metodo_devolucion,
            'monto_nc': monto_con_iva,
            'dte_original_id': dte_original.id if dte_original else None,
        }
    )

    # Generar TXT Acepta
    contenido_txt = None
    nombre_archivo = f"NC_61_{numero_nc}_{nc.fecha_emision.strftime('%Y%m%d')}.txt"
    try:
        productos_agrupados = defaultdict(lambda: {
            'tallas': [], 'cantidad_total': 0, 'precio': 0,
            'monto_total': 0, 'articulo': '', 'marca': '', 'color': ''
        })
        for dp in nc.dte_productos.select_related('productoTalla__producto'):
            producto = dp.productoTalla.producto
            key = producto.articulo
            g = productos_agrupados[key]
            talla_nombre = str(dp.productoTalla.talla) if hasattr(dp.productoTalla, 'talla') and dp.productoTalla.talla else 'U'
            g['tallas'].append(f"{dp.stock}:{talla_nombre}")
            g['cantidad_total'] += dp.stock
            g['precio'] = dp.precio
            g['monto_total'] += dp.stock * dp.precio
            g['articulo'] = producto.articulo
            if not g['marca'] and producto.atributo1:
                g['marca'] = producto.atributo1.valor
            if not g['color'] and producto.atributo2:
                g['color'] = producto.atributo2.valor

        detalle_txt = []
        for articulo, g in productos_agrupados.items():
            tallas_str = ' '.join(g['tallas'])
            marca_limpia = limpiar_texto(g['marca'] or '')
            color_limpio = limpiar_texto(g['color'] or '')
            marca_color = f"{marca_limpia} {color_limpio}".strip()
            nombre_final = f"{marca_color} {tallas_str}".strip() if marca_color else tallas_str
            detalle_txt.append({
                'nombre': limpiar_texto(nombre_final),
                'descripcion': '',
                'cantidad': g['cantidad_total'],
                'unidad': 'UN',
                'precio_unitario': g['precio'],
                'descuento_pct': 0,
                'monto_descuento': 0,
                'monto_item': g['monto_total'],
                'codigo': limpiar_texto(g['articulo'])
            })

        referencias_nc = json.loads(nc.referencias) if isinstance(nc.referencias, str) else []

        datos_txt = {
            'documento': {
                'tipo_documento': 61,
                'folio': nc.numero_documento,
                'fecha_emision': nc.fecha_emision.strftime('%Y-%m-%d'),
                'fecha_vencimiento': nc.fecha_vencimiento.strftime('%Y-%m-%d'),
                'forma_pago': 1,
                'timestamp': timezone.now().strftime('%Y-%m-%dT%H:%M:%S')
            },
            'emisor': {
                'rut': empresa.rut,
                'razon_social': limpiar_texto(empresa.razon_social or ''),
                'giro': limpiar_texto(empresa.giro or ''),
                'acteco': empresa.acteco or '',
                # Dirección/comuna/ciudad de la SUCURSAL del cambio/devolución,
                # con fallback a la casa matriz si la sucursal no tiene el dato.
                'direccion': limpiar_texto((cambio.sucursal.direccion if cambio.sucursal else '') or empresa.direccion or ''),
                'comuna': limpiar_texto((cambio.sucursal.comuna if cambio.sucursal else '') or empresa.comuna or ''),
                'ciudad': limpiar_texto((cambio.sucursal.ciudad if cambio.sucursal else '') or empresa.ciudad or ''),
                'codigo_vendedor': limpiar_texto(request.user.username or 'USUARIO'),
                'sucursal': limpiar_texto(cambio.sucursal.alias if cambio.sucursal else ''),
                'telefono': empresa.contacto1 or '',
            },
            'receptor': {
                'rut': nc.receptor.rut if nc.receptor else '66666666-6',
                'razon_social': limpiar_texto(nc.receptor.razon_social if nc.receptor else 'CONSUMIDOR FINAL'),
                'giro': limpiar_texto(nc.receptor.giro if nc.receptor else ''),
                'direccion': limpiar_texto(nc.receptor.direccion if nc.receptor else ''),
                'comuna': limpiar_texto(nc.receptor.comuna if nc.receptor else ''),
                'ciudad': limpiar_texto(nc.receptor.ciudad if nc.receptor else ''),
            },
            'totales': {
                'monto_neto': monto_neto,
                'monto_exento': 0,
                'tasa_iva': 19,
                'iva': iva,
                'monto_total': monto_con_iva,
                'descuento_global': 0
            },
            'detalle': detalle_txt,
            'referencias': referencias_nc
        }

        # Las líneas vienen CON IVA (precio público); para el TXT NC tipo 61
        # deben ir NETAS y sumar el MntNeto. Normaliza contra el cabezal.
        normalizar_detalle_para_tipo(datos_txt['detalle'], datos_txt['totales'], 61)

        contenido_txt = generar_txt_nota_credito_acepta(datos_txt)

        import os
        ruta_nc = os.path.join('MEDIA', 'documentos_electronicos', 'nc')
        os.makedirs(ruta_nc, exist_ok=True)
        with open(os.path.join(ruta_nc, nombre_archivo), 'w', encoding='utf-8') as f:
            f.write(contenido_txt)

    except Exception as e:
        logger.error(f"Error generando TXT Acepta para NC #{numero_nc}: {e}")
        contenido_txt = None

    return JsonResponse({
        'success': True,
        'message': f'Nota de Crédito #{numero_nc} generada exitosamente',
        'data': {
            'nc_id': nc.id,
            'nc_numero': numero_nc,
            'nc_monto': monto_con_iva,
            'metodo_devolucion': metodo_devolucion,
            'metodo_devolucion_display': dict(METODO_DEVOLUCION_NC_CHOICES).get(metodo_devolucion, ''),
            'cambio_id': cambio.id,
            'numero_operacion': cambio.numero_operacion,
            'txt_generado': contenido_txt is not None,
            'nombre_archivo_txt': nombre_archivo if contenido_txt else None,
        }
    })


@login_required
@require_GET
def detalle_nc_devolucion(request, cambio_id):
    """
    Retorna el detalle de la Nota de Crédito asociada a un CambioDevolucion.
    """
    try:
        cambio = CambioDevolucion.objects.select_related(
            'nota_credito', 'nota_credito__emisor', 'nota_credito__receptor',
            'ticket_original', 'sucursal'
        ).get(id=cambio_id)
    except CambioDevolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cambio/Devolución no encontrado'}, status=404)

    sucursal_id = request.session.get('idSucursalActual')
    if cambio.sucursal_id != int(sucursal_id):
        return JsonResponse({'success': False, 'error': 'Sin acceso a esta operación'}, status=403)

    if not cambio.nc_generada or not cambio.nota_credito:
        return JsonResponse({
            'success': True,
            'nc_generada': False,
            'data': None
        })

    nc = cambio.nota_credito
    productos_nc = []
    for dp in nc.dte_productos.select_related('productoTalla__producto'):
        producto = dp.productoTalla.producto if dp.productoTalla else None
        productos_nc.append({
            'articulo': producto.articulo if producto else dp.descripcion,
            'talla': str(dp.productoTalla.talla) if dp.productoTalla and hasattr(dp.productoTalla, 'talla') else '',
            'cantidad': dp.stock,
            'precio_unitario': float(dp.precio),
            'subtotal': float(dp.precio * dp.stock),
        })

    return JsonResponse({
        'success': True,
        'nc_generada': True,
        'data': {
            'nc_id': nc.id,
            'nc_numero': nc.numero_documento,
            'nc_fecha': nc.fecha_emision.strftime('%d/%m/%Y'),
            'nc_monto_neto': float(nc.monto_neto),
            'nc_iva': float(nc.monto_con_iva - nc.monto_neto),
            'nc_monto_total': float(nc.monto_con_iva),
            'nc_estado': nc.estado_dte,
            'metodo_devolucion': cambio.metodo_devolucion,
            'metodo_devolucion_display': cambio.get_metodo_devolucion_display(),
            'fecha_nc': cambio.fecha_nc.strftime('%d/%m/%Y %H:%M') if cambio.fecha_nc else None,
            'numero_operacion': cambio.numero_operacion,
            'motivo': nc.motivo_nc,
            'productos': productos_nc,
            'documento_afectado': nc.documento_afectado.numero_documento if nc.documento_afectado else None,
        }
    })
