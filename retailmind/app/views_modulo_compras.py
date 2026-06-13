"""
Módulo de Compras - RetailMind
Contiene todas las vistas relacionadas con compras, DTEs de compras, recepciones y proveedores
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, Http404, HttpResponseBadRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg
from django.db.models.functions import Abs
from django.db import models
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
import json
import re
import csv
import logging
from decimal import Decimal

from .models import (
    Compras, Compras_Producto, Compras_Producto_Talla, Dte, Dte_Detalle_Pago,
    Dte_Productos, Empresa, Producto, Producto_Talla, Productos_Recepcionados,
    Sucursal, EmpresaUser, Movimientos_Producto, LoteProducto, Dte_Incidencia
)

# Método de pago usado para registrar una compensación factura-contra-factura
# (neteo de tesorería / cuentas por pagar, NO una relación tributaria SII: ambas
# facturas siguen siendo documentos tributarios válidos). Se guarda en el libro de
# pagos Dte_Detalle_Pago igual que una Nota de Crédito. Ver asociar_factura_compensacion.
# IMPORTANTE: este string debe coincidir EXACTO en todos los lugares que lo filtran
# (cargarDteCompra, pagosDTE, obtener_resumen_pendientes_anio).
METODO_COMPENSACION = 'Compensación con Factura'

logger = logging.getLogger('app')


# ========== GESTIÓN DE COMPRAS ==========

@login_required
def verGestionCompras(request):
    """Vista principal para gestión de compras"""
    empresas = Empresa.objects.filter(esProveedor=True).order_by('nombre')
    return render(request, 'vistas/modulo_compras/gestionCompras.html', {'empresas': empresas})


def crear_compra(request):
    """Crear nueva compra"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            empresa_id = request.POST.get('empresa_id')
            fecha_compra = request.POST.get('fecha_compra')
            numero_factura = request.POST.get('numero_factura')
            total = request.POST.get('total')
            observaciones = request.POST.get('observaciones', '')
            
            # Validaciones
            if not all([empresa_id, fecha_compra, numero_factura, total]):
                return JsonResponse({
                    'success': False,
                    'error': 'Todos los campos son requeridos'
                })
            
            # Crear la compra
            compra = Compras.objects.create(
                empresa_id=empresa_id,
                fecha_compra=fecha_compra,
                numero_factura=numero_factura,
                total=Decimal(total),
                observaciones=observaciones,
                usuario_creacion=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Compra creada exitosamente',
                'compra_id': compra.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear compra: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@require_GET
def obtener_compras_por_anio(request):
    """Obtener compras filtradas por año"""
    try:
        anio = request.GET.get('anio', timezone.now().year)
        empresa_id = request.GET.get('empresa_id')
        
        queryset = Compras.objects.filter(
            fecha__year=anio
        ).select_related('empresa')
        
        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        
        compras = queryset.order_by('-fecha')
        
        compras_data = []
        for compra in compras:
            compras_data.append({
                'id': compra.id,
                'nombre': compra.nombre,
                'empresa': compra.empresa.nombre,
                'fecha': compra.fecha.strftime('%d/%m/%Y'),
                'temporada': compra.temporada,
                'responsable': compra.responsable,
                'correlativo': compra.correlativo
            })
        
        return JsonResponse({
            'success': True,
            'compras': compras_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener compras: {str(e)}'
        })


def importar_csv_compra(request):
    """Importar compras desde archivo CSV"""
    if request.method == 'POST':
        try:
            csv_file = request.FILES.get('csv_file')
            if not csv_file:
                return JsonResponse({
                    'success': False,
                    'error': 'Archivo CSV requerido'
                })
            
            # Validar formato
            if not csv_file.name.endswith('.csv'):
                return JsonResponse({
                    'success': False,
                    'error': 'El archivo debe ser formato CSV'
                })
            
            # Leer archivo CSV
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            compras_creadas = 0
            errores = []
            
            with transaction.atomic():
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Validar campos requeridos
                        campos_requeridos = ['empresa_id', 'fecha_compra', 'numero_factura', 'total']
                        for campo in campos_requeridos:
                            if not row.get(campo):
                                errores.append(f'Fila {row_num}: Campo {campo} requerido')
                                continue
                        
                        # Crear compra
                        Compras.objects.create(
                            empresa_id=row['empresa_id'],
                            fecha_compra=row['fecha_compra'],
                            numero_factura=row['numero_factura'],
                            total=Decimal(row['total']),
                            observaciones=row.get('observaciones', ''),
                            usuario_creacion=request.user
                        )
                        compras_creadas += 1
                        
                    except Exception as e:
                        errores.append(f'Fila {row_num}: {str(e)}')
            
            return JsonResponse({
                'success': True,
                'message': f'{compras_creadas} compras importadas exitosamente',
                'errores': errores
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al importar CSV: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def recepcionar_compra(request):
    """
    Recepcionar productos de una compra
    
    IMPORTANTE: Esta función usa registrar_movimiento_producto que:
    - Crea el movimiento de ingreso
    - Actualiza el stock en Producto_Talla
    - Crea automáticamente el lote FIFO (cuando crear_lote_fifo=True y concepto es de ingreso)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            compra_id = data.get('compra_id')
            productos = data.get('productos', [])
            
            if not compra_id:
                return JsonResponse({
                    'success': False,
                    'error': 'ID de compra requerido'
                })
            
            if not productos:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe incluir al menos un producto'
                })
            
            compra = get_object_or_404(Compras, id=compra_id)
            sucursal_id = request.session.get('idSucursalActual')
            sucursal = Sucursal.objects.filter(id=sucursal_id).first() if sucursal_id else None
            responsable = request.user.username if hasattr(request.user, 'username') else 'Sistema'
            
            with transaction.atomic():
                productos_procesados = 0
                for item in productos:
                    producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
                    cantidad = int(item['cantidad'])
                    costo_unitario = Decimal(str(item['costo_unitario']))
                    precio_venta = Decimal(str(item.get('precio_venta', producto_talla.producto.precioventa)))
                    sobreprecio = Decimal(str(item.get('sobreprecio', producto_talla.producto.sobreprecio)))
                    
                    # Crear registro de recepción
                    Productos_Recepcionados.objects.create(
                        compra=compra,
                        producto_talla=producto_talla,
                        cantidad_recepcionada=cantidad,
                        costo_unitario=costo_unitario,
                        precio_venta_sugerido=precio_venta,
                        usuario_recepcion=request.user,
                        fecha_recepcion=timezone.now()
                    )
                    
                    # ✅ CORREGIDO: Usar registrar_movimiento_producto que:
                    # - Crea el movimiento con el concepto correcto
                    # - Actualiza el stock de Producto_Talla
                    # - Crea el lote FIFO automáticamente (cuando crear_lote_fifo=True)
                    # ⚠️ NO crear lote FIFO manualmente para evitar duplicación
                    from .views import registrar_movimiento_producto
                    registrar_movimiento_producto(
                        producto_talla=producto_talla,
                        concepto='RECEPCION_COMPRA',  # ✅ CORREGIDO: Usar concepto válido
                        cantidad=cantidad,
                        responsable=responsable,
                        sucursal_origen=sucursal,
                        sucursal_destino=sucursal,
                        observaciones=f'Recepción compra #{compra.numero_factura}',
                        referencia_externa=f'COMPRA_{compra.id}',
                        crear_lote_fifo=True  # ✅ El lote se crea automáticamente
                    )
                    
                    # Actualizar costo del producto si es diferente (costo más reciente)
                    if producto_talla.producto.costo != int(costo_unitario):
                        producto_talla.producto.costo = int(costo_unitario)
                        producto_talla.producto.sobreprecio = int(sobreprecio)
                        producto_talla.producto.precioventa = int(precio_venta)
                        producto_talla.producto.save()
                    
                    productos_procesados += 1
                
                # Actualizar estado de la compra
                compra.estado = 'RECEPCIONADA'
                compra.fecha_recepcion = timezone.now()
                compra.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Compra recepcionada exitosamente. {productos_procesados} producto(s) procesado(s)'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            })
        except Exception as e:
            logger.exception("Error en recepcionar_compra compra_id=%s", compra_id)
            return JsonResponse({
                'success': False,
                'error': f'Error al recepcionar compra: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ========== GESTIÓN DE DTEs DE COMPRAS ==========

@login_required
def verGestionDteCompras(request):
    """Vista principal para gestión de DTEs de compras"""
    return render(request, 'vistas/modulo_compras/gestionDteCompras.html')


def obtener_dte(request, dte_id):
    """Obtener detalles de un DTE específico"""
    try:
        dte = get_object_or_404(Dte, id=dte_id)
        
        # Obtener productos del DTE
        productos = []
        for dte_producto in dte.dte_productos.all():
            productos.append({
                'id': dte_producto.id,
                'producto_nombre': dte_producto.productoTalla.producto.articulo,
                'sku': dte_producto.productoTalla.sku,
                'talla': dte_producto.productoTalla.talla.nombre if dte_producto.productoTalla.talla else '',
                'cantidad': dte_producto.cantidad,
                'precio_unitario': float(dte_producto.precio_unitario),
                'total_linea': float(dte_producto.cantidad * dte_producto.precio_unitario)
            })
        
        # Obtener pagos
        pagos = []
        for pago in dte.dte_detalle_pago.all():
            pagos.append({
                'id': pago.id,
                'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                'monto': float(pago.monto),
                'metodo_pago': pago.metodo_pago,
                'referencia': pago.referencia or '',
                'observaciones': pago.observaciones or ''
            })
        
        dte_data = {
            'id': dte.id,
            'numero_dte': dte.numero_dte,
            'tipo_documento': dte.tipo_documento,
            'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y'),
            'emisor': dte.emisor.nombre,
            'receptor': dte.receptor.nombre if dte.receptor else '',
            'empresa_receptora_id': dte.receptor.id if dte.receptor else None,
            'empresa_receptora_nombre': dte.receptor.nombre if dte.receptor else '',
            'sucursal_receptora_id': dte.sucursal.id if dte.sucursal else None,
            'sucursal_receptora_nombre': dte.sucursal.alias if dte.sucursal else '',
            'subtotal': float(dte.subtotal),
            'iva': float(dte.iva),
            'total': float(dte.total),
            'estado_dte': dte.estado_dte,
            'productos': productos,
            'pagos': pagos
        }
        
        return JsonResponse({
            'success': True,
            'dte': dte_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener DTE: {str(e)}'
        })


def obtener_dte_compras(request):
    """Obtener lista de DTEs de compras con filtros"""
    try:
        # Parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        proveedor_id = request.GET.get('proveedor_id')
        estado = request.GET.get('estado')
        tipo_documento = request.GET.get('tipo_documento')
        
        # Construir queryset
        queryset = Dte.objects.filter(
            tipo_transaccion='COMPRA'
        ).select_related('emisor', 'receptor')
        
        # Aplicar filtros
        if fecha_inicio:
            queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_emision__lte=fecha_fin)
        if proveedor_id:
            queryset = queryset.filter(emisor_id=proveedor_id)
        if estado:
            queryset = queryset.filter(estado_dte=estado)
        if tipo_documento:
            queryset = queryset.filter(tipo_documento=tipo_documento)
        
        # Ordenar por fecha descendente
        queryset = queryset.order_by('-fecha_emision')
        
        # Paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        paginator = Paginator(queryset, per_page)
        dtes_page = paginator.get_page(page)
        
        # Serializar datos
        dtes_data = []
        for dte in dtes_page:
            # Calcular IVA
            iva = dte.monto_con_iva - dte.monto_neto
            
            dtes_data.append({
                'id': dte.id,
                'numero_dte': dte.numero_documento,  # Corregido: numero_documento
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y'),
                'emisor': dte.emisor.nombre,
                'emisor_rut': dte.emisor.rut if dte.emisor else '',
                'subtotal': float(dte.monto_neto),  # Corregido: monto_neto
                'iva': float(iva),  # Calculado
                'total': float(dte.monto_con_iva),  # Corregido: monto_con_iva
                'estado_dte': dte.estado_dte,
                'estado_pago': dte.estado_pago,
                'fecha_recepcion': dte.fecha_recepcion.strftime('%d/%m/%Y') if dte.fecha_recepcion else None
            })
        
        return JsonResponse({
            'success': True,
            'dtes': dtes_data,
            'pagination': {
                'current_page': dtes_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': dtes_page.has_next(),
                'has_previous': dtes_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener DTEs: {str(e)}'
        })


def crearDteCompras(request):
    """Crear nuevo DTE de compras"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar datos requeridos
            emisor_id = data.get('emisor_id')
            tipo_documento = data.get('tipo_documento')
            numero_dte = data.get('numero_dte')
            productos = data.get('productos', [])
            
            if not all([emisor_id, tipo_documento, numero_dte]):
                return JsonResponse({
                    'success': False,
                    'error': 'Emisor, tipo de documento y número DTE son requeridos'
                })
            
            if not productos:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe incluir al menos un producto'
                })
            
            with transaction.atomic():
                # Calcular totales
                subtotal = sum(
                    Decimal(item['cantidad']) * Decimal(item['precio_unitario']) 
                    for item in productos
                )
                iva = subtotal * Decimal('0.19')  # 19% IVA
                total = subtotal + iva
                
                # Crear DTE
                dte = Dte.objects.create(
                    numero_dte=numero_dte,
                    tipo_documento=tipo_documento,
                    tipo_transaccion='COMPRA',
                    fecha_emision=data.get('fecha_emision', timezone.localdate()),
                    emisor_id=emisor_id,
                    receptor_id=data.get('receptor_id'),
                    subtotal=subtotal,
                    iva=iva,
                    total=total,
                    estado_dte='EMITIDO',
                    observaciones=data.get('observaciones', '')
                )
                
                # Crear productos del DTE
                for item in productos:
                    producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
                    cantidad = int(item['cantidad'])
                    precio_unitario = int(item['precio_unitario'])
                    costo = producto_talla.producto.costo
                    
                    # Guardar sobreprecio tal cual está en el producto (es un DELTA/MARGEN)
                    sobreprecio_unitario = producto_talla.producto.sobreprecio
                    
                    Dte_Productos.objects.create(
                        dte=dte,
                        productoTalla=producto_talla,
                        descripcion=f"{producto_talla.producto.articulo} - Talla {producto_talla.talla}",
                        costo=costo,
                        sobreprecio=sobreprecio_unitario,
                        precio=precio_unitario,
                        stock=cantidad,
                        activo=True
                    )
            
            return JsonResponse({
                'success': True,
                'message': 'DTE creado exitosamente',
                'dte_id': dte.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear DTE: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def empresas_proveedoras(request):
    """Obtener lista de empresas proveedoras"""
    try:
        empresas = Empresa.objects.filter(
            esProveedor=True
        ).order_by('nombre')
        
        empresas_data = []
        for empresa in empresas:
            empresas_data.append({
                'id': empresa.id,
                'nombre': empresa.nombre,
                'rut': empresa.rut
            })
        
        return JsonResponse(empresas_data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener empresas: {str(e)}'
        })


def cargarDteCompra(request):
    """Cargar DTE de compra desde archivo XML"""
    if request.method == 'POST':
        try:
            xml_file = request.FILES.get('xml_file')
            if not xml_file:
                return JsonResponse({
                    'success': False,
                    'error': 'Archivo XML requerido'
                })
            
            # Validar formato
            if not xml_file.name.endswith('.xml'):
                return JsonResponse({
                    'success': False,
                    'error': 'El archivo debe ser formato XML'
                })
            
            # El parser XML de DTEs aún no está implementado. Devolver éxito
            # aquí engañaría al usuario (el archivo no se guarda ni procesa),
            # así que respondemos un error honesto en lugar de un falso OK.
            logger.warning(
                "cargarDteCompra: intento de carga XML '%s' — parser no implementado",
                xml_file.name,
            )
            return JsonResponse({
                'success': False,
                'error': 'La carga de DTE por XML aún no está disponible. '
                         'Ingresa el documento de compra de forma manual.'
            }, status=501)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al cargar DTE: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ========== GESTIÓN DE PAGOS DTE ==========

def registrarPagoDTE(request):
    """Registrar pago para un DTE"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            dte_id = data.get('dte_id')
            monto = data.get('monto')
            metodo_pago = data.get('metodo_pago')
            fecha_pago = data.get('fecha_pago')
            
            if not all([dte_id, monto, metodo_pago, fecha_pago]):
                return JsonResponse({
                    'success': False,
                    'error': 'Todos los campos son requeridos'
                })
            
            dte = get_object_or_404(Dte, id=dte_id)
            
            # Crear detalle de pago
            pago = Dte_Detalle_Pago.objects.create(
                dte=dte,
                fecha_pago=fecha_pago,
                monto=Decimal(monto),
                metodo_pago=metodo_pago,
                referencia=data.get('referencia', ''),
                observaciones=data.get('observaciones', '')
            )
            
            # Verificar si el DTE está completamente pagado
            total_pagado = dte.dte_detalle_pago.aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            if total_pagado >= dte.total:
                dte.estado_dte = 'PAGADO'
                dte.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Pago registrado exitosamente',
                'pago_id': pago.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al registrar pago: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def obtenerDetallePago(request, dte_id):
    """Obtener detalles de pagos de un DTE"""
    try:
        dte = get_object_or_404(Dte, id=dte_id)
        
        pagos = []
        for pago in dte.dte_detalle_pago.all().order_by('-fecha_pago'):
            pagos.append({
                'id': pago.id,
                'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                'monto': float(pago.monto),
                'metodo_pago': pago.metodo_pago,
                'referencia': pago.referencia or '',
                'observaciones': pago.observaciones or ''
            })
        
        total_pagado = sum(pago['monto'] for pago in pagos)
        saldo_pendiente = float(dte.total) - total_pagado
        
        return JsonResponse({
            'success': True,
            'dte': {
                'id': dte.id,
                'numero_dte': dte.numero_dte,
                'total': float(dte.total),
                'total_pagado': total_pagado,
                'saldo_pendiente': saldo_pendiente
            },
            'pagos': pagos
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener pagos: {str(e)}'
        })


def pagosDTE(request, dte_id):
    """Vista para gestionar pagos de un DTE"""
    try:
        dte = get_object_or_404(Dte, id=dte_id)
        context = {
            'dte': dte
        }
        return render(request, 'vistas/modulo_compras/pagos_dte.html', context)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


def eliminarPago(request, pago_id):
    """Eliminar un pago de DTE"""
    if request.method == 'DELETE':
        try:
            pago = get_object_or_404(Dte_Detalle_Pago, id=pago_id)
            dte = pago.dte
            
            pago.delete()
            
            # Recalcular estado del DTE
            total_pagado = dte.dte_detalle_pago.aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            if total_pagado < dte.total:
                dte.estado_dte = 'EMITIDO'
                dte.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Pago eliminado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar pago: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def detallePago(request, pago_id):
    """Obtener detalles de un pago específico"""
    try:
        pago = get_object_or_404(Dte_Detalle_Pago, id=pago_id)
        
        pago_data = {
            'id': pago.id,
            'dte_id': pago.dte.id,
            'fecha_pago': pago.fecha_pago.strftime('%Y-%m-%d'),
            'monto': float(pago.monto),
            'metodo_pago': pago.metodo_pago,
            'referencia': pago.referencia or '',
            'observaciones': pago.observaciones or ''
        }
        
        return JsonResponse({
            'success': True,
            'pago': pago_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener pago: {str(e)}'
        })


def editarPago(request, pago_id):
    """Editar un pago existente"""
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            pago = get_object_or_404(Dte_Detalle_Pago, id=pago_id)
            
            # Actualizar campos
            pago.fecha_pago = data.get('fecha_pago', pago.fecha_pago)
            pago.monto = Decimal(data.get('monto', pago.monto))
            pago.metodo_pago = data.get('metodo_pago', pago.metodo_pago)
            pago.referencia = data.get('referencia', pago.referencia)
            pago.observaciones = data.get('observaciones', pago.observaciones)
            pago.save()
            
            # Recalcular estado del DTE
            dte = pago.dte
            total_pagado = dte.dte_detalle_pago.aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            if total_pagado >= dte.total:
                dte.estado_dte = 'PAGADO'
            else:
                dte.estado_dte = 'EMITIDO'
            dte.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Pago actualizado exitosamente'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al editar pago: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ========== NOTAS DE CRÉDITO ==========

def notasCredito(request, dte_id):
    """Vista para gestionar notas de crédito de un DTE"""
    try:
        dte = get_object_or_404(Dte, id=dte_id)
        context = {
            'dte': dte
        }
        return render(request, 'vistas/modulo_compras/notas_credito.html', context)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


def agregarNotaCredito(request):
    """Agregar nota de crédito a un DTE"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # TODO: Implementar lógica de notas de crédito
            
            return JsonResponse({
                'success': True,
                'message': 'Nota de crédito agregada (funcionalidad en desarrollo)'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al agregar nota de crédito: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def eliminarNotaCredito(request, nc_id):
    """Eliminar nota de crédito"""
    if request.method == 'DELETE':
        try:
            # TODO: Implementar eliminación de nota de crédito
            
            return JsonResponse({
                'success': True,
                'message': 'Nota de crédito eliminada (funcionalidad en desarrollo)'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar nota de crédito: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def eliminar_dte(request, dte_id):
    """Eliminar DTE de compras"""
    if request.method == 'DELETE':
        try:
            import json
            body = json.loads(request.body) if request.body else {}
            forzar = body.get('forzar', False)  # Parámetro para forzar eliminación
            
            dte = get_object_or_404(Dte, id=dte_id)
            
            # Si no es forzado, hacer validaciones normales
            if not forzar:
                # Verificar si se puede eliminar
                if dte.estado_pago == 'Pagado':
                    return JsonResponse({
                        'success': False,
                        'error': 'No se puede eliminar un DTE pagado',
                        'puede_forzar': True
                    }, status=400)
                
                # Verificar si tiene recepciones asociadas
                if dte.fecha_recepcion:
                    return JsonResponse({
                        'success': False,
                        'error': 'No se puede eliminar un DTE que ya fue recepcionado',
                        'puede_forzar': True
                    }, status=400)
                
                # Verificar si tiene productos asociados
                productos_dte = Dte_Productos.objects.filter(dte=dte).count()
                if productos_dte > 0:
                    return JsonResponse({
                        'success': False,
                        'error': f'No se puede eliminar un DTE con {productos_dte} producto(s) asociado(s)',
                        'puede_forzar': True,
                        'productos_count': productos_dte
                    }, status=400)
                
                # Verificar si tiene Notas de Crédito asociadas
                if dte.tipo_documento in ['FACTURA ELECTRONICA', '33', 'FACTURA']:
                    nc_asociadas = Dte_Detalle_Pago.objects.filter(
                        dte=dte,
                        metodo_pago='Nota de Crédito'
                    ).count()
                    
                    if nc_asociadas > 0:
                        return JsonResponse({
                            'success': False,
                            'error': f'No se puede eliminar una factura con {nc_asociadas} Nota(s) de Crédito asociada(s)',
                            'puede_forzar': True
                        }, status=400)
                
                # Verificar si es una NC enlazada a una factura
                if dte.tipo_documento in ['NOTA DE CREDITO', '61', 'NC']:
                    pago_nc = Dte_Detalle_Pago.objects.filter(
                        voucher=str(dte.numero_documento),
                        metodo_pago='Nota de Crédito'
                    ).first()
                    
                    if pago_nc:
                        return JsonResponse({
                            'success': False,
                            'error': f'No se puede eliminar una NC enlazada a la Factura #{pago_nc.dte.numero_documento}',
                            'puede_forzar': True
                        }, status=400)
            
            # Eliminación forzada o sin restricciones
            if forzar:
                with transaction.atomic():
                    # Eliminar productos asociados
                    Dte_Productos.objects.filter(dte=dte).delete()
                    
                    # Eliminar pagos/NCs asociadas
                    Dte_Detalle_Pago.objects.filter(dte=dte).delete()
                    
                    # Eliminar el DTE
                    dte.delete()
                    
                logger.warning("DTE de compra eliminado forzadamente: dte_id=%s numero=%s", dte_id, dte.numero_documento)
            else:
                dte.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'DTE eliminado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar DTE: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# ========== DASHBOARDS DE COMPRAS ==========

@login_required
@require_GET
def dashboard_compras_estrategico(request):
    """Dashboard estratégico de compras con métricas avanzadas basado en datos reales"""
    try:
        from datetime import datetime, timedelta
        from django.db.models.functions import TruncMonth
        
        # Parámetros de filtro
        anio = int(request.GET.get('anio', timezone.localdate().year))
        temporada = request.GET.get('temporada', '')
        proveedor_id = request.GET.get('proveedor', '')
        responsable = request.GET.get('responsable', '')
        
        # Query base para compras
        compras_query = Compras.objects.filter(fecha__year=anio)

        # Filtro de temporada: acepta código de familia normalizada
        # (VERANO/OTONO/INVIERNO/PRIMAVERA) o texto libre (fallback legacy).
        if temporada:
            _t_norm = temporada.strip().upper()
            if _t_norm in {'VERANO', 'OTONO', 'INVIERNO', 'PRIMAVERA'}:
                compras_query = compras_query.filter(temporada_familia=_t_norm)
            else:
                compras_query = compras_query.filter(temporada__icontains=temporada)
        if proveedor_id:
            compras_query = compras_query.filter(empresa_id=proveedor_id)
        if responsable:
            compras_query = compras_query.filter(responsable=responsable)

        # Obtener IDs de compras para filtrar
        compras_ids = list(compras_query.values_list('id', flat=True))
        
        # ===== CÁLCULO DE MÉTRICAS REALES =====
        
        # 1. Total de compras y unidades
        total_compras = compras_query.count()
        
        # 2. Productos y tallas de las compras
        productos_compras = Compras_Producto.objects.filter(compras__in=compras_ids)
        total_productos = productos_compras.count()
        
        tallas_compras = Compras_Producto_Talla.objects.filter(
            compra_producto__compras__in=compras_ids
        )
        
        # 3. Total de unidades esperadas
        total_unidades_esperadas = tallas_compras.aggregate(
            total=Sum('stock')
        )['total'] or 0
        
        # 4. Total de unidades recepcionadas
        recepciones = Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras__in=compras_ids
        )
        total_unidades_recepcionadas = recepciones.aggregate(
            total=Sum('stockArribado')
        )['total'] or 0
        
        # 5. Cálculo de cumplimiento (% recepcionado vs esperado)
        cumplimiento_general = 0
        if total_unidades_esperadas > 0:
            cumplimiento_general = round((total_unidades_recepcionadas / total_unidades_esperadas) * 100, 1)
        
        # 6. Cálculo de inversión total (costo)
        inversion_total = productos_compras.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # 7. Valor esperado de venta (precio sugerido)
        valor_venta_esperado = productos_compras.aggregate(
            total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # 8. Cálculo de ROI promedio (basado en precio sugerido vs costo)
        roi_promedio = 0
        if inversion_total > 0:
            ganancia_esperada = valor_venta_esperado - inversion_total
            roi_promedio = round((ganancia_esperada / inversion_total) * 100, 1)
        
        # 9. Rotación de inventario (estimada - productos con recepciones)
        productos_con_recepcion = recepciones.values('producto_talla').distinct().count()
        rotacion_inventario = round(productos_con_recepcion / max(total_productos, 1), 2)
        
        # 10. Precisión de pronóstico (% de cumplimiento de recepciones)
        precision_pronostico = cumplimiento_general  # Similar al cumplimiento
        
        # ===== CUMPLIMIENTO POR PROVEEDOR =====
        cumplimiento_proveedores = []
        proveedores = compras_query.values('empresa__id', 'empresa__nombre').distinct()
        
        for proveedor in proveedores:
            compras_proveedor = compras_query.filter(empresa_id=proveedor['empresa__id'])
            compras_proveedor_ids = list(compras_proveedor.values_list('id', flat=True))
            
            tallas_proveedor = Compras_Producto_Talla.objects.filter(
                compra_producto__compras__in=compras_proveedor_ids
            )
            esperadas_proveedor = tallas_proveedor.aggregate(total=Sum('stock'))['total'] or 0
            
            recepciones_proveedor = Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto__compras__in=compras_proveedor_ids
            )
            recepcionadas_proveedor = recepciones_proveedor.aggregate(total=Sum('stockArribado'))['total'] or 0
            
            cumplimiento = 0
            if esperadas_proveedor > 0:
                cumplimiento = round((recepcionadas_proveedor / esperadas_proveedor) * 100, 1)
            
            cumplimiento_proveedores.append({
                'proveedor': proveedor['empresa__nombre'],
                'cumplimiento': cumplimiento
            })
        
        # ===== ROI POR TEMPORADA (usa agrupamiento normalizado con YoY) =====
        roi_temporadas = calcular_roi_temporadas_mejorado(compras_query, compras_ids)
        
        # ===== RENDIMIENTO DETALLADO POR COMPRA =====
        rendimiento_detallado = []
        
        for compra in compras_query[:10]:  # Limitar a las primeras 10
            # Unidades de esta compra
            tallas_compra = Compras_Producto_Talla.objects.filter(
                compra_producto__compras=compra
            )
            unidades_esperadas = tallas_compra.aggregate(total=Sum('stock'))['total'] or 0
            
            # Recepciones de esta compra
            recepciones_compra = Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto__compras=compra
            )
            unidades_recibidas = recepciones_compra.aggregate(total=Sum('stockArribado'))['total'] or 0
            
            # Cumplimiento
            cumplimiento_compra = 0
            if unidades_esperadas > 0:
                cumplimiento_compra = round((unidades_recibidas / unidades_esperadas) * 100, 1)
            
            # ROI de la compra
            productos_compra = Compras_Producto.objects.filter(compras=compra)
            inversion_compra = productos_compra.aggregate(
                total=Sum(F('costo') * F('compras_producto_talla__stock'))
            )['total'] or 0
            
            valor_compra = productos_compra.aggregate(
                total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
            )['total'] or 0
            
            roi_compra = 0
            if inversion_compra > 0:
                ganancia_compra = valor_compra - inversion_compra
                roi_compra = round((ganancia_compra / inversion_compra) * 100, 1)
            
            # Determinar estado
            estado = 'Pendiente'
            if cumplimiento_compra >= 100:
                estado = 'Completado'
            elif cumplimiento_compra >= 80:
                estado = 'Pendiente'
            else:
                estado = 'Retrasado'
            
            rendimiento_detallado.append({
                'nombre': compra.nombre if hasattr(compra, 'nombre') and compra.nombre else f'Compra #{compra.id}',
                'proveedor': compra.empresa.nombre if compra.empresa else 'Sin proveedor',
                'temporada': compra.temporada if hasattr(compra, 'temporada') and compra.temporada else 'N/A',
                'cumplimiento': cumplimiento_compra,
                'roi': roi_compra,
                'rotacion': round(unidades_recibidas / max(unidades_esperadas, 1), 2),
                'precision': cumplimiento_compra,
                'estado': estado
            })
        
        # ===== ALERTAS =====
        alertas = []
        
        if cumplimiento_general < 80:
            alertas.append({
                'mensaje': f'Cumplimiento general bajo ({cumplimiento_general}%). Revisar procesos de recepción.'
            })
        
        compras_sin_recepcion = compras_query.count() - Compras.objects.filter(
            id__in=compras_ids,
            compras_producto__compras_producto_talla__productos_recepcionados__isnull=False
        ).distinct().count()
        
        if compras_sin_recepcion > 0:
            alertas.append({
                'mensaje': f'{compras_sin_recepcion} compra(s) sin recepción registrada.'
            })
        
        if roi_promedio < 15:
            alertas.append({
                'mensaje': f'ROI promedio bajo ({roi_promedio}%). Revisar precios y costos.'
            })
        
        # ===== RECOMENDACIONES =====
        recomendaciones = []
        
        if cumplimiento_general < 90:
            recomendaciones.append({
                'mensaje': 'Implementar seguimiento más estricto de recepciones para mejorar cumplimiento.'
            })
        
        if rotacion_inventario < 0.5:
            recomendaciones.append({
                'mensaje': 'Optimizar gestión de inventario para aumentar rotación de productos.'
            })
        
        if len(cumplimiento_proveedores) > 0:
            proveedores_bajo_cumplimiento = [p for p in cumplimiento_proveedores if p['cumplimiento'] < 80]
            if proveedores_bajo_cumplimiento:
                recomendaciones.append({
                    'mensaje': f'Revisar desempeño de {len(proveedores_bajo_cumplimiento)} proveedor(es) con bajo cumplimiento.'
                })
        
        # ===== COMPRAS POR MES (para gráfico de tendencias) =====
        compras_por_mes = []
        for mes in range(1, 13):
            compras_mes = compras_query.filter(fecha__month=mes)
            if compras_mes.exists():
                productos_mes = Compras_Producto.objects.filter(compras__in=compras_mes.values_list('id', flat=True))
                inversion_mes = productos_mes.aggregate(
                    total=Sum(F('costo') * F('compras_producto_talla__stock'))
                )['total'] or 0
                
                compras_por_mes.append({
                    'mes': mes,
                    'nombre_mes': datetime(anio, mes, 1).strftime('%B'),
                    'total_compras': compras_mes.count(),
                    'inversion': float(inversion_mes) if inversion_mes else 0
                })
        
        # ===== TENDENCIAS (calculadas) =====
        # Calcular tendencias comparando con período anterior
        compras_anterior = Compras.objects.filter(fecha__year=anio-1)
        if temporada:
            compras_anterior = compras_anterior.filter(temporada__icontains=temporada)
        if proveedor_id:
            compras_anterior = compras_anterior.filter(empresa_id=proveedor_id)
        if responsable:
            compras_anterior = compras_anterior.filter(responsable=responsable)
            
        compras_anterior_ids = list(compras_anterior.values_list('id', flat=True))
        
        # Cumplimiento anterior
        tallas_anterior = Compras_Producto_Talla.objects.filter(
            compra_producto__compras__in=compras_anterior_ids
        )
        esperadas_anterior = tallas_anterior.aggregate(total=Sum('stock'))['total'] or 1
        
        recepciones_anterior = Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras__in=compras_anterior_ids
        )
        recepcionadas_anterior = recepciones_anterior.aggregate(total=Sum('stockArribado'))['total'] or 0
        cumplimiento_anterior = (recepcionadas_anterior / esperadas_anterior * 100) if esperadas_anterior > 0 else 0
        
        # ROI anterior
        productos_anterior = Compras_Producto.objects.filter(compras__in=compras_anterior_ids)
        inversion_anterior = productos_anterior.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 1
        valor_anterior = productos_anterior.aggregate(
            total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
        )['total'] or 0
        roi_anterior = ((valor_anterior - inversion_anterior) / inversion_anterior * 100) if inversion_anterior > 0 else 0
        
        tendencias = {
            'trend_cumplimiento': round(cumplimiento_general - cumplimiento_anterior, 1),
            'trend_roi': round(roi_promedio - roi_anterior, 1),
            'trend_rotacion': 0,  # Se puede calcular con datos de ventas si están disponibles
            'trend_precision': round(precision_pronostico - cumplimiento_anterior, 1)
        }
        
        # ===== TOP 5 PROVEEDORES POR INVERSIÓN =====
        top_proveedores = []
        proveedores_inversiones = {}
        for proveedor in proveedores:
            compras_proveedor = compras_query.filter(empresa_id=proveedor['empresa__id'])
            compras_proveedor_ids = list(compras_proveedor.values_list('id', flat=True))
            
            productos_prov = Compras_Producto.objects.filter(compras__in=compras_proveedor_ids)
            inversion_prov = productos_prov.aggregate(
                total=Sum(F('costo') * F('compras_producto_talla__stock'))
            )['total'] or 0
            
            if inversion_prov > 0:
                proveedores_inversiones[proveedor['empresa__nombre']] = float(inversion_prov)
        
        # Ordenar y tomar top 5
        top_proveedores = sorted(
            [{'proveedor': k, 'inversion': v} for k, v in proveedores_inversiones.items()],
            key=lambda x: x['inversion'],
            reverse=True
        )[:5]
        
        # ===== PRODUCTOS CON MAYOR INVERSIÓN =====
        top_productos = []
        productos_con_inversion = Compras_Producto.objects.filter(
            compras__in=compras_ids
        ).annotate(
            inversion_total=Sum(F('costo') * F('compras_producto_talla__stock'))
        ).order_by('-inversion_total')[:10]
        
        for prod in productos_con_inversion:
            top_productos.append({
                'nombre': prod.nombre,
                'marca': prod.atributo1 or '-',
                'inversion': float(prod.inversion_total) if prod.inversion_total else 0,
                'unidades': Compras_Producto_Talla.objects.filter(compra_producto=prod).aggregate(
                    total=Sum('stock')
                )['total'] or 0
            })
        
        # ===== RESPUESTA FINAL =====
        response_data = {
            'cumplimiento_general': cumplimiento_general,
            'roi_promedio': roi_promedio,
            'rotacion_inventario': rotacion_inventario,
            'precision_pronostico': precision_pronostico,
            'cumplimiento_proveedores': cumplimiento_proveedores,
            'roi_temporadas': roi_temporadas,
            'rendimiento_detallado': rendimiento_detallado,
            'alertas': alertas,
            'recomendaciones': recomendaciones,
            'compras_por_mes': compras_por_mes,
            'top_proveedores': top_proveedores,
            'top_productos': top_productos,
            **tendencias,
            # Métricas adicionales
            'metricas_adicionales': {
                'total_compras': total_compras,
                'total_productos': total_productos,
                'total_unidades_esperadas': total_unidades_esperadas,
                'total_unidades_recepcionadas': total_unidades_recepcionadas,
                'inversion_total': float(inversion_total) if inversion_total else 0,
                'valor_venta_esperado': float(valor_venta_esperado) if valor_venta_esperado else 0,
                'ganancia_esperada': float(valor_venta_esperado - inversion_total) if (valor_venta_esperado and inversion_total) else 0
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'Error al generar dashboard: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@require_GET
def exportar_dashboard_compras(request):
    """Exportar datos del dashboard de compras a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Obtener datos del dashboard
        dashboard_response = dashboard_compras_estrategico(request)
        dashboard_payload = json.loads(dashboard_response.content)
        if isinstance(dashboard_payload, dict) and dashboard_payload.get('success') is False:
            return JsonResponse({
                'success': False,
                'error': dashboard_payload.get('error', 'Error al generar dashboard')
            })

        dashboard_data = dashboard_payload.get('dashboard', dashboard_payload)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja de métricas generales
        ws_metricas = wb.active
        ws_metricas.title = "Métricas Generales"
        
        # Escribir métricas
        metricas = dashboard_data.get('metricas_generales', {})
        metricas_adicionales = dashboard_data.get('metricas_adicionales', {})
        total_compras = metricas.get('total_compras', metricas_adicionales.get('total_compras', 0))
        monto_total = metricas.get('monto_total', metricas_adicionales.get('inversion_total', 0))
        promedio_compra = metricas.get(
            'promedio_compra',
            round(monto_total / total_compras, 2) if total_compras else 0
        )
        dtes_pendientes = metricas.get('dtes_pendientes', 0)
        monto_pendiente = metricas.get('monto_pendiente', 0)

        ws_metricas.append(['Métrica', 'Valor'])
        ws_metricas.append(['Total Compras', total_compras])
        ws_metricas.append(['Monto Total', monto_total])
        ws_metricas.append(['Promedio por Compra', promedio_compra])
        ws_metricas.append(['DTEs Pendientes', dtes_pendientes])
        ws_metricas.append(['Monto Pendiente', monto_pendiente])
        
        # Hoja de compras por proveedor
        ws_proveedores = wb.create_sheet("Compras por Proveedor")
        ws_proveedores.append(['Proveedor', 'Total Compras', 'Monto Total'])
        
        compras_por_proveedor = dashboard_data.get('compras_por_proveedor')
        if not compras_por_proveedor:
            compras_por_proveedor = [
                {
                    'proveedor': item.get('proveedor', '-'),
                    'total_compras': item.get('total_compras', ''),
                    'monto_total': item.get('inversion', 0)
                }
                for item in dashboard_data.get('top_proveedores', [])
            ]

        for item in compras_por_proveedor:
            ws_proveedores.append([
                item.get('proveedor', '-'),
                item.get('total_compras', ''),
                item.get('monto_total', 0)
            ])
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dashboard_compras.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


@login_required
def verDashboardCompras(request):
    """Vista principal del dashboard de compras - Redirige al dashboard mejorado"""
    return render(request, 'vistas/modulo_dashboards/dashboard_compras_mejorado.html')


@login_required
def verDiagnosticoCompras(request):
    """Vista para diagnóstico de datos de compras"""
    return render(request, 'vistas/modulo_compras/diagnostico_compras.html')


@login_required
def diagnostico_datos_compras(request):
    """API para diagnóstico de calidad de datos de compras"""
    try:
        # Análisis de calidad de datos
        total_compras = Compras.objects.count()
        compras_sin_proveedor = Compras.objects.filter(empresa__isnull=True).count()
        compras_sin_fecha = Compras.objects.filter(fecha_compra__isnull=True).count()
        compras_sin_total = Compras.objects.filter(total__isnull=True).count()
        
        # DTEs con problemas
        dtes_sin_productos = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            dte_productos__isnull=True
        ).count()
        
        dtes_sin_emisor = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            emisor__isnull=True
        ).count()
        
        diagnostico = {
            'compras': {
                'total': total_compras,
                'sin_proveedor': compras_sin_proveedor,
                'sin_fecha': compras_sin_fecha,
                'sin_total': compras_sin_total,
                'calidad_score': max(0, 100 - (
                    (compras_sin_proveedor + compras_sin_fecha + compras_sin_total) * 100 / max(total_compras, 1)
                ))
            },
            'dtes': {
                'sin_productos': dtes_sin_productos,
                'sin_emisor': dtes_sin_emisor
            }
        }
        
        return JsonResponse({
            'success': True,
            'diagnostico': diagnostico
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en diagnóstico: {str(e)}'
        })


def obtenerDetalleComprasPorParametros(request):
    """Obtener detalle de compras por parámetros específicos"""
    # Placeholder para funcionalidad futura
    return JsonResponse({
        'success': True,
        'message': 'Funcionalidad en desarrollo'
    })


@login_required
@require_GET
def obtener_resumen_pendientes_anio(request):
    """Obtener resumen de DTEs pendientes del año actual con KPIs de vencimiento"""
    try:
        # Obtener año actual
        ahora = timezone.localtime()
        anio_actual = ahora.year
        hoy = timezone.localdate()
        empresa_actual_id = request.session.get('idEmpresaActual')
        
        # Query para DTEs de compra del año actual
        queryset = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            fecha_emision__year=anio_actual
        ).filter(
            Q(receptor_id=empresa_actual_id) | Q(receptor__isnull=True)
        )
        
        # Contar pendientes
        pendientes = queryset.filter(
            Q(estado_pago='Pendiente') | Q(estado_pago='Parcial')
        )
        
        cantidad_pendientes = pendientes.count()
        
        # Calcular monto total pendiente y clasificar por vencimiento
        monto_total_pendiente = 0
        vencidos = 0  # Ya pasó la fecha de vencimiento
        por_vencer_pronto = 0  # Vencen en los próximos 7 días
        al_dia = 0  # Más de 7 días para vencer
        
        monto_vencidos = 0
        monto_por_vencer = 0
        monto_al_dia = 0
        
        for dte in pendientes:
            monto_dte = dte.monto_con_iva
            
            # Restar notas de crédito y compensaciones con factura asociadas
            notas_credito = Dte_Detalle_Pago.objects.filter(
                dte=dte,
                metodo_pago__in=['Nota de Crédito', METODO_COMPENSACION]
            ).aggregate(total=Sum('monto'))['total'] or 0

            saldo_pendiente = float(monto_dte - notas_credito)
            monto_total_pendiente += saldo_pendiente
            
            # Clasificar por vencimiento
            dias_hasta_vencimiento = (dte.fecha_vencimiento - hoy).days
            
            if dias_hasta_vencimiento < 0:
                # Ya venció
                vencidos += 1
                monto_vencidos += saldo_pendiente
            elif dias_hasta_vencimiento <= 7:
                # Por vencer en 7 días o menos
                por_vencer_pronto += 1
                monto_por_vencer += saldo_pendiente
            else:
                # Al día (más de 7 días)
                al_dia += 1
                monto_al_dia += saldo_pendiente
        
        # Estadísticas adicionales
        total_dtes = queryset.count()
        pagados = queryset.filter(estado_pago='Pagado').count()
        
        return JsonResponse({
            'success': True,
            'anio': anio_actual,
            'cantidad_pendientes': cantidad_pendientes,
            'monto_pendiente': monto_total_pendiente,
            'total_dtes': total_dtes,
            'pagados': pagados,
            # KPIs de vencimiento
            'vencidos': vencidos,
            'monto_vencidos': monto_vencidos,
            'por_vencer_pronto': por_vencer_pronto,
            'monto_por_vencer': monto_por_vencer,
            'al_dia': al_dia,
            'monto_al_dia': monto_al_dia
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener resumen: {str(e)}'
        })


# ========== EXPORTACIÓN DE COMPRAS ACTUALES ==========

@require_GET
@login_required
def exportar_compras_excel(request):
    """Exportar compras actuales a Excel con productos y detalles"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        anio = request.GET.get('anio', timezone.now().year)
        
        # Obtener compras del año
        compras = Compras.objects.filter(fecha__year=anio).select_related('empresa').order_by('-fecha')
        
        if not compras.exists():
            return JsonResponse({
                'success': False,
                'error': f'No hay compras para el año {anio}'
            })
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen de Compras
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Compras"
        
        # Estilos
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers_resumen = [
            'ID', 'Proveedor', 'RUT Proveedor', 'Nombre Compra', 'Temporada',
            'Fecha Inicio', 'Fecha Término', 'Fecha Registro', 'Responsable',
            'Total Productos', 'Total Unidades', 'Costo Total', 'Venta Esperada', 'Recepcionado'
        ]
        
        for col, header in enumerate(headers_resumen, start=1):
            cell = ws_resumen.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row_idx, compra in enumerate(compras, start=2):
            # Calcular totales
            productos = Compras_Producto.objects.filter(compras=compra)
            total_productos = productos.count()
            
            total_unidades = Compras_Producto_Talla.objects.filter(
                compra_producto__compras=compra
            ).aggregate(total=Sum('stock'))['total'] or 0
            
            costo_total = productos.aggregate(
                total=Sum(F('costo') * F('compras_producto_talla__stock'))
            )['total'] or 0
            
            venta_esperada = productos.aggregate(
                total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
            )['total'] or 0
            
            recepcionado = Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto__compras=compra
            ).aggregate(total=Sum('stockArribado'))['total'] or 0
            
            # Escribir datos
            rut_proveedor = ''
            if compra.empresa and compra.empresa.rut:
                rut_proveedor = compra.empresa.rut.replace('.', '')
            
            ws_resumen.cell(row=row_idx, column=1, value=compra.id)
            ws_resumen.cell(row=row_idx, column=2, value=compra.empresa.nombre if compra.empresa else '')
            ws_resumen.cell(row=row_idx, column=3, value=rut_proveedor)
            ws_resumen.cell(row=row_idx, column=4, value=compra.nombre)
            ws_resumen.cell(row=row_idx, column=5, value=compra.temporada)
            ws_resumen.cell(row=row_idx, column=6, value=compra.fechaInicioTemporada)
            ws_resumen.cell(row=row_idx, column=7, value=compra.fechaTerminoTemporada)
            ws_resumen.cell(row=row_idx, column=8, value=compra.fecha)
            ws_resumen.cell(row=row_idx, column=9, value=compra.responsable)
            ws_resumen.cell(row=row_idx, column=10, value=total_productos)
            ws_resumen.cell(row=row_idx, column=11, value=total_unidades)
            ws_resumen.cell(row=row_idx, column=12, value=float(costo_total) if costo_total else 0)
            ws_resumen.cell(row=row_idx, column=13, value=float(venta_esperada) if venta_esperada else 0)
            ws_resumen.cell(row=row_idx, column=14, value=recepcionado)
            
            # Aplicar bordes
            for col in range(1, 15):
                ws_resumen.cell(row=row_idx, column=col).border = border
        
        # Ajustar ancho de columnas
        for col in ws_resumen.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_resumen.column_dimensions[column].width = adjusted_width
        
        # Hoja 2: Detalle de Productos
        ws_detalle = wb.create_sheet("Detalle Productos")
        
        headers_detalle = [
            'ID Compra', 'Nombre Compra', 'Proveedor', 'Nombre Producto',
            'Descripción', 'Marca', 'Color', 'Género', 'Costo', 'Precio Sugerido',
            'Talla', 'Stock', 'Recepcionado', 'Factura DTE'
        ]
        
        for col, header in enumerate(headers_detalle, start=1):
            cell = ws_detalle.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos de productos
        row_idx = 2
        for compra in compras:
            productos = Compras_Producto.objects.filter(compras=compra).prefetch_related('compras_producto_talla_set')
            
            for producto in productos:
                tallas = producto.compras_producto_talla_set.all()
                
                for talla in tallas:
                    # Obtener recepción si existe
                    recepcion = Productos_Recepcionados.objects.filter(
                        compra_producto_talla=talla
                    ).first()
                    
                    recepcionado = recepcion.stockArribado if recepcion else 0
                    
                    # Obtener factura asociada si existe
                    factura = ''
                    if recepcion and recepcion.dte:
                        factura = f"{recepcion.dte.numero_documento}"
                    
                    ws_detalle.cell(row=row_idx, column=1, value=compra.id)
                    ws_detalle.cell(row=row_idx, column=2, value=compra.nombre)
                    ws_detalle.cell(row=row_idx, column=3, value=compra.empresa.nombre if compra.empresa else '')
                    ws_detalle.cell(row=row_idx, column=4, value=producto.nombre)
                    ws_detalle.cell(row=row_idx, column=5, value=producto.descripcion or '')
                    ws_detalle.cell(row=row_idx, column=6, value=producto.atributo1)
                    ws_detalle.cell(row=row_idx, column=7, value=producto.atributo2)
                    ws_detalle.cell(row=row_idx, column=8, value=producto.atributo3)
                    ws_detalle.cell(row=row_idx, column=9, value=producto.costo)
                    ws_detalle.cell(row=row_idx, column=10, value=producto.precioSugerido)
                    ws_detalle.cell(row=row_idx, column=11, value=talla.talla)
                    ws_detalle.cell(row=row_idx, column=12, value=talla.stock)
                    ws_detalle.cell(row=row_idx, column=13, value=recepcionado)
                    ws_detalle.cell(row=row_idx, column=14, value=factura)
                    
                    # Aplicar bordes
                    for col in range(1, 15):
                        ws_detalle.cell(row=row_idx, column=col).border = border
                    
                    row_idx += 1
        
        # Ajustar ancho de columnas en detalle
        for col in ws_detalle.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detalle.column_dimensions[column].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="compras_{anio}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar compras: {str(e)}'
        })


@require_GET
@login_required
def exportar_compras_csv(request):
    """Exportar compras actuales a CSV"""
    try:
        anio = request.GET.get('anio', timezone.now().year)
        
        # Obtener compras del año
        compras = Compras.objects.filter(fecha__year=anio).select_related('empresa').order_by('-fecha')
        
        if not compras.exists():
            return JsonResponse({
                'success': False,
                'error': f'No hay compras para el año {anio}'
            })
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="compras_{anio}.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'ID Compra', 'Proveedor', 'RUT Proveedor', 'Nombre Compra', 'Temporada',
            'Fecha Inicio', 'Fecha Término', 'Fecha Registro', 'Responsable',
            'Nombre Producto', 'Descripción', 'Marca', 'Color', 'Género',
            'Costo', 'Precio Sugerido', 'Talla', 'Stock', 'Recepcionado', 'Factura DTE'
        ])
        
        # Datos
        for compra in compras:
            productos = Compras_Producto.objects.filter(compras=compra).prefetch_related('compras_producto_talla_set')
            
            for producto in productos:
                tallas = producto.compras_producto_talla_set.all()
                
                for talla in tallas:
                    # Obtener recepción si existe
                    recepcion = Productos_Recepcionados.objects.filter(
                        compra_producto_talla=talla
                    ).first()
                    
                    recepcionado = recepcion.stockArribado if recepcion else 0
                    factura = ''
                    if recepcion and recepcion.dte:
                        factura = str(recepcion.dte.numero_documento)
                    
                    # Limpiar RUT (sin puntos)
                    rut_proveedor = ''
                    if compra.empresa and compra.empresa.rut:
                        rut_proveedor = compra.empresa.rut.replace('.', '')
                    
                    writer.writerow([
                        compra.id,
                        compra.empresa.nombre if compra.empresa else '',
                        rut_proveedor,
                        compra.nombre,
                        compra.temporada,
                        compra.fechaInicioTemporada.strftime('%Y-%m-%d') if compra.fechaInicioTemporada else '',
                        compra.fechaTerminoTemporada.strftime('%Y-%m-%d') if compra.fechaTerminoTemporada else '',
                        compra.fecha.strftime('%Y-%m-%d'),
                        compra.responsable,
                        producto.nombre,
                        producto.descripcion or '',
                        producto.atributo1,
                        producto.atributo2,
                        producto.atributo3,
                        producto.costo,
                        producto.precioSugerido,
                        talla.talla,
                        talla.stock,
                        recepcionado,
                        factura
                    ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar compras: {str(e)}'
        })


# ========== GESTIÓN DE PROVEEDORES - IMPORTACIÓN ==========

@login_required
def ver_importacion_proveedores(request):
    """Vista para importar proveedores desde CSV/Excel"""
    return render(request, 'vistas/modulo_compras/importacion_proveedores.html')


@require_POST
def importar_proveedores_csv(request):
    """Importar proveedores desde archivo CSV/Excel"""
    try:
        archivo = request.FILES.get('archivo_proveedores')
        if not archivo:
            return JsonResponse({
                'success': False,
                'error': 'No se proporcionó ningún archivo'
            })
        
        # Obtener modo de importación
        modo_actualizacion = request.POST.get('modo_actualizacion', 'crear_y_actualizar')
        
        # Validar extensión
        nombre_archivo = archivo.name.lower()
        if not (nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls')):
            return JsonResponse({
                'success': False,
                'error': 'Formato de archivo no válido. Use CSV o Excel (.xlsx, .xls)'
            })
        
        # Leer datos según formato
        if nombre_archivo.endswith('.csv'):
            # Leer CSV
            decoded_file = archivo.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            datos = list(reader)
        else:
            # Leer Excel
            import openpyxl
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Obtener encabezados (primera fila) y limpiarlos
            headers = []
            for cell in ws[1]:
                header = str(cell.value or '').strip().lower()
                headers.append(header)
            
            # Leer datos
            datos = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Saltar filas vacías
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            # Convertir None a string vacío
                            row_dict[headers[i]] = str(value) if value is not None else ''
                    datos.append(row_dict)
        
        # Debug de importación
        if datos and len(datos) > 0:
            logger.debug(
                "Importacion proveedores: headers=%s primera_fila_campos=%s",
                list(datos[0].keys()),
                {k: bool(v) for k, v in datos[0].items()},
            )
        
        # Procesar datos
        proveedores_creados = 0
        proveedores_actualizados = 0
        proveedores_omitidos = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    # Validar campos requeridos
                    rut = str(fila.get('rut', '') or '').strip()
                    nombre = str(fila.get('nombre', '') or '').strip()
                    
                    # Verificar si la fila está completamente vacía
                    valores_fila = [str(v or '').strip() for v in fila.values()]
                    if not any(valores_fila):
                        # Fila vacía, ignorar silenciosamente
                        continue
                    
                    # Si hay algún dato pero falta RUT o nombre, reportar error
                    if not rut or not nombre:
                        logger.debug(
                            "Importacion proveedores fila incompleta: fila=%s rut=%s nombre=%s campos=%s",
                            idx,
                            rut,
                            nombre,
                            {k: bool(v) for k, v in fila.items()},
                        )
                        # Solo reportar error si hay otros datos en la fila
                        if any(valores_fila):
                            errores.append(f'Fila {idx}: RUT y Nombre son requeridos (RUT="{rut}", Nombre="{nombre}")')
                        continue
                    
                    # Validar formato RUT (básico)
                    if not validar_rut_basico(rut):
                        errores.append(f'Fila {idx}: RUT "{rut}" no válido')
                        continue
                    
                    # Verificar si el proveedor ya existe
                    proveedor_existente = Empresa.objects.filter(rut=rut).first()
                    
                    # Preparar datos del proveedor (usar minúsculas para keys)
                    datos_proveedor = {
                        'rut': rut,
                        'nombre': nombre,
                        'nombre_fantasia': str(fila.get('nombre_fantasia', fila.get('nombre', nombre))).strip(),
                        'razon_social': str(fila.get('razon_social', fila.get('nombre', nombre))).strip(),
                        'giro': str(fila.get('giro', '')).strip() or 'Sin especificar',
                        'direccion': str(fila.get('direccion', '')).strip() or 'Sin especificar',
                        'comuna': str(fila.get('comuna', '')).strip() or 'Sin especificar',
                        'ciudad': str(fila.get('ciudad', '')).strip() or 'Sin especificar',
                        'esProveedor': True,
                        'correoVendedor': str(fila.get('correo_vendedor', '') or fila.get('email', '')).strip() or 'sin@correo.com',
                        'correoIntercambio': str(fila.get('correo_intercambio', '') or fila.get('email', '')).strip() or 'sin@correo.com',
                        'correoAdministrador': str(fila.get('correo_administrador', '') or fila.get('email', '')).strip() or 'sin@correo.com',
                        'acteco': str(fila.get('acteco', '')).strip() or None,
                        'contacto1': str(fila.get('telefono', '') or fila.get('contacto1', '')).strip() or None,
                        'contacto2': str(fila.get('contacto2', '')).strip() or None,
                    }
                    
                    if proveedor_existente:
                        if modo_actualizacion == 'crear_y_actualizar':
                            # Actualizar proveedor existente
                            for key, value in datos_proveedor.items():
                                setattr(proveedor_existente, key, value)
                            proveedor_existente.save()
                            proveedores_actualizados += 1
                        elif modo_actualizacion == 'solo_crear':
                            # Omitir proveedores existentes
                            proveedores_omitidos += 1
                        elif modo_actualizacion == 'solo_actualizar':
                            # Actualizar solo proveedores existentes
                            for key, value in datos_proveedor.items():
                                setattr(proveedor_existente, key, value)
                            proveedor_existente.save()
                            proveedores_actualizados += 1
                    else:
                        if modo_actualizacion in ['crear_y_actualizar', 'solo_crear']:
                            # Crear nuevo proveedor
                            Empresa.objects.create(**datos_proveedor)
                            proveedores_creados += 1
                        else:
                            # Modo solo_actualizar: omitir nuevos
                            proveedores_omitidos += 1
                        
                except Exception as e:
                    errores.append(f'Fila {idx}: {str(e)}')
                    continue
        
        # Preparar mensaje de respuesta
        mensaje = []
        if proveedores_creados > 0:
            mensaje.append(f'{proveedores_creados} proveedores creados')
        if proveedores_actualizados > 0:
            mensaje.append(f'{proveedores_actualizados} proveedores actualizados')
        if proveedores_omitidos > 0:
            mensaje.append(f'{proveedores_omitidos} proveedores omitidos')
        
        return JsonResponse({
            'success': True,
            'message': ', '.join(mensaje) if mensaje else 'No se procesaron proveedores',
            'proveedores_creados': proveedores_creados,
            'proveedores_actualizados': proveedores_actualizados,
            'proveedores_omitidos': proveedores_omitidos,
            'errores': errores
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al importar proveedores: {str(e)}'
        })


def validar_rut_basico(rut):
    """
    Validación de formato RUT chileno
    Acepta formatos: 76123456-7, 76.123.456-7
    """
    # Eliminar puntos y espacios
    rut_limpio = rut.replace('.', '').replace(' ', '').upper()
    
    # Validar que tenga guión
    if '-' not in rut_limpio:
        return False
    
    # Separar número y dígito verificador por guión
    partes = rut_limpio.split('-')
    if len(partes) != 2:
        return False
    
    numero = partes[0]
    dv = partes[1]
    
    # Validar longitud
    if len(numero) < 7 or len(numero) > 8:
        return False
    
    # Validar que el número sea numérico
    if not numero.isdigit():
        return False
    
    # Validar DV (debe ser dígito o K)
    if not (dv.isdigit() or dv == 'K'):
        return False
    
    # Calcular dígito verificador
    suma = 0
    multiplicador = 2
    
    for digit in reversed(numero):
        suma += int(digit) * multiplicador
        multiplicador += 1
        if multiplicador > 7:
            multiplicador = 2
    
    resto = suma % 11
    dv_calculado = 11 - resto
    
    if dv_calculado == 11:
        dv_calculado = '0'
    elif dv_calculado == 10:
        dv_calculado = 'K'
    else:
        dv_calculado = str(dv_calculado)
    
    return dv == dv_calculado


@require_GET
def descargar_formato_proveedores(request):
    """Descargar formato CSV de ejemplo para importar proveedores"""
    try:
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="formato_proveedores.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'rut', 'nombre', 'nombre_fantasia', 'razon_social', 
            'giro', 'direccion', 'comuna', 'ciudad',
            'email', 'telefono', 'acteco'
        ])
        
        # Filas de ejemplo (RUT sin puntos)
        writer.writerow([
            '76123456-7', 
            'Empresa Ejemplo SPA', 
            'Ejemplo', 
            'Empresa Ejemplo Sociedad por Acciones',
            'Comercio al por mayor',
            'Av. Principal 123',
            'Santiago',
            'Santiago',
            'contacto@ejemplo.cl',
            '+56912345678',
            '471010'
        ])
        writer.writerow([
            '77234567-8', 
            'Distribuidora ABC Ltda', 
            'ABC Distribuidora', 
            'Distribuidora ABC Limitada',
            'Distribución de productos',
            'Calle Comercio 456',
            'Providencia',
            'Santiago',
            'ventas@abc.cl',
            '+56987654321',
            '471020'
        ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al generar formato: {str(e)}'
        })


@require_GET
def exportar_proveedores_actuales(request):
    """Exportar todos los proveedores actuales a CSV"""
    try:
        # Obtener proveedores
        proveedores = Empresa.objects.filter(esProveedor=True).order_by('nombre')
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="proveedores_actuales.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'id', 'rut', 'nombre', 'nombre_fantasia', 'razon_social', 
            'giro', 'direccion', 'comuna', 'ciudad',
            'email', 'telefono', 'acteco'
        ])
        
        # Datos
        for proveedor in proveedores:
            # Limpiar RUT (sin puntos)
            rut_limpio = proveedor.rut.replace('.', '') if proveedor.rut else ''
            
            writer.writerow([
                proveedor.id,
                rut_limpio,
                proveedor.nombre,
                proveedor.nombre_fantasia,
                proveedor.razon_social,
                proveedor.giro,
                proveedor.direccion,
                proveedor.comuna,
                proveedor.ciudad,
                proveedor.correoVendedor or proveedor.correoIntercambio or proveedor.correoAdministrador,
                proveedor.contacto1 or '',
                proveedor.acteco or ''
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar proveedores: {str(e)}'
        })


@require_GET
def exportar_proveedores_excel(request):
    """Exportar proveedores actuales a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Obtener proveedores
        proveedores = Empresa.objects.filter(esProveedor=True).order_by('nombre')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        
        # Estilos
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Encabezados
        headers = [
            'ID', 'RUT', 'Nombre', 'Nombre Fantasía', 'Razón Social',
            'Giro', 'Dirección', 'Comuna', 'Ciudad',
            'Email', 'Teléfono', 'Acteco'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, proveedor in enumerate(proveedores, start=2):
            # Limpiar RUT (sin puntos)
            rut_limpio = proveedor.rut.replace('.', '') if proveedor.rut else ''
            
            ws.cell(row=row_idx, column=1, value=proveedor.id)
            ws.cell(row=row_idx, column=2, value=rut_limpio)
            ws.cell(row=row_idx, column=3, value=proveedor.nombre)
            ws.cell(row=row_idx, column=4, value=proveedor.nombre_fantasia)
            ws.cell(row=row_idx, column=5, value=proveedor.razon_social)
            ws.cell(row=row_idx, column=6, value=proveedor.giro)
            ws.cell(row=row_idx, column=7, value=proveedor.direccion)
            ws.cell(row=row_idx, column=8, value=proveedor.comuna)
            ws.cell(row=row_idx, column=9, value=proveedor.ciudad)
            ws.cell(row=row_idx, column=10, value=proveedor.correoVendedor or proveedor.correoIntercambio or proveedor.correoAdministrador)
            ws.cell(row=row_idx, column=11, value=proveedor.contacto1 or '')
            ws.cell(row=row_idx, column=12, value=proveedor.acteco or '')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="proveedores_actuales.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar a Excel: {str(e)}'
        })


# ========== GESTIÓN DE DTEs - IMPORTACIÓN ==========

@login_required
def ver_importacion_dtes(request):
    """Vista para importar DTEs desde CSV/Excel"""
    # Obtener proveedores para el selector
    proveedores = Empresa.objects.filter(esProveedor=True).order_by('nombre')
    return render(request, 'vistas/modulo_compras/importacion_dtes.html', {
        'proveedores': proveedores
    })


@require_POST
def importar_dtes_csv(request):
    """Importar DTEs desde archivo CSV/Excel"""
    try:
        archivo = request.FILES.get('archivo_dtes')
        if not archivo:
            return JsonResponse({
                'success': False,
                'error': 'No se proporcionó ningún archivo'
            })
        
        # Obtener parámetros adicionales
        tipo_busqueda = request.POST.get('tipo_busqueda', 'rut')  # 'rut' o 'id'
        modo_actualizacion = request.POST.get('modo_actualizacion', 'solo_crear')  # 'solo_crear' o 'crear_y_actualizar'
        
        # Validar extensión
        nombre_archivo = archivo.name.lower()
        if not (nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls')):
            return JsonResponse({
                'success': False,
                'error': 'Formato de archivo no válido. Use CSV o Excel (.xlsx, .xls)'
            })
        
        # Leer datos según formato
        if nombre_archivo.endswith('.csv'):
            decoded_file = archivo.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            datos = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Obtener encabezados (primera fila) y limpiarlos
            headers = []
            for cell in ws[1]:
                header = str(cell.value or '').strip().lower()
                headers.append(header)
            
            # Leer datos
            datos = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            # Convertir None a string vacío
                            row_dict[headers[i]] = str(value) if value is not None else ''
                    datos.append(row_dict)
        
        # Debug de importación
        if datos and len(datos) > 0:
            logger.debug(
                "Importacion DTE compras: headers=%s primera_fila_campos=%s",
                list(datos[0].keys()),
                {k: bool(v) for k, v in datos[0].items()},
            )
        
        # Procesar DTEs
        dtes_creados = 0
        dtes_actualizados = 0
        dtes_omitidos = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    # Verificar si la fila está completamente vacía
                    valores_fila = [str(v or '').strip() for v in fila.values()]
                    if not any(valores_fila):
                        # Fila vacía, ignorar silenciosamente
                        continue
                    
                    # Buscar proveedor (emisor) - usar minúsculas para compatibilidad
                    if tipo_busqueda == 'rut':
                        rut_proveedor = str(fila.get('rut_proveedor', '') or fila.get('rut_emisor', '') or '').strip()
                        if not rut_proveedor:
                            errores.append(f'Fila {idx}: RUT de proveedor requerido')
                            continue
                        
                        # Buscar con RUT limpio (sin puntos)
                        rut_busqueda = rut_proveedor.replace('.', '').replace(' ', '')
                        emisor = Empresa.objects.filter(rut=rut_busqueda, esProveedor=True).first()
                        if not emisor:
                            # Intentar buscar con el RUT original
                            emisor = Empresa.objects.filter(rut=rut_proveedor, esProveedor=True).first()
                        if not emisor:
                            errores.append(f'Fila {idx}: Proveedor con RUT "{rut_proveedor}" no encontrado')
                            continue
                    else:  # id
                        id_proveedor = str(fila.get('id_proveedor', '') or fila.get('proveedor_id', '') or '').strip()
                        if not id_proveedor:
                            errores.append(f'Fila {idx}: ID de proveedor requerido')
                            continue
                        emisor = Empresa.objects.filter(id=int(id_proveedor), esProveedor=True).first()
                        if not emisor:
                            errores.append(f'Fila {idx}: Proveedor con ID "{id_proveedor}" no encontrado')
                            continue
                    
                    # Validar campos requeridos del DTE
                    numero_documento = str(fila.get('numero_documento', fila.get('numero_dte', '')) or '').strip()
                    tipo_documento = str(fila.get('tipo_documento', '33') or '33').strip()
                    
                    if not numero_documento:
                        errores.append(f'Fila {idx}: Número de documento requerido')
                        continue
                    
                    # Calcular montos: Acepta monto_neto O monto_con_iva
                    monto_neto_str = str(fila.get('monto_neto', fila.get('subtotal', '')) or '').strip().replace(',', '.')
                    monto_con_iva_str = str(fila.get('monto_con_iva', fila.get('total', fila.get('monto_total', ''))) or '').strip().replace(',', '.')
                    
                    if monto_neto_str and monto_neto_str != '0':
                        # Si viene monto_neto, calcular IVA y total
                        monto_neto = Decimal(monto_neto_str)
                        iva = monto_neto * Decimal('0.19')
                        total = monto_neto + iva
                    elif monto_con_iva_str and monto_con_iva_str != '0':
                        # Si viene monto_con_iva (total), calcular monto_neto
                        total = Decimal(monto_con_iva_str)
                        monto_neto = total / Decimal('1.19')
                        iva = total - monto_neto
                    else:
                        errores.append(f'Fila {idx}: Debe proporcionar monto_neto O monto_con_iva')
                        continue
                    
                    # Fechas
                    from datetime import datetime, timedelta
                    fecha_emision_str = str(fila.get('fecha_emision', '') or '').strip()
                    logger.debug("Importacion DTE fila=%s fecha_emision_csv=%s", idx, fecha_emision_str)
                    
                    if fecha_emision_str and fecha_emision_str != 'None' and fecha_emision_str != '':
                        # Si viene con hora (formato Excel datetime), extraer solo la fecha
                        if ' ' in fecha_emision_str:
                            fecha_emision_str = fecha_emision_str.split(' ')[0]
                            logger.debug("Importacion DTE fila=%s fecha limpiada=%s", idx, fecha_emision_str)
                        
                        try:
                            # Intentar YYYY-MM-DD
                            fecha_emision = datetime.strptime(fecha_emision_str, '%Y-%m-%d').date()
                            logger.debug("Importacion DTE fila=%s fecha parseada YYYY-MM-DD=%s", idx, fecha_emision)
                        except:
                            try:
                                # Intentar DD/MM/YYYY
                                fecha_emision = datetime.strptime(fecha_emision_str, '%d/%m/%Y').date()
                                logger.debug("Importacion DTE fila=%s fecha parseada DD/MM/YYYY=%s", idx, fecha_emision)
                            except:
                                try:
                                    # Intentar DD-MM-YYYY
                                    fecha_emision = datetime.strptime(fecha_emision_str, '%d-%m-%Y').date()
                                    logger.debug("Importacion DTE fila=%s fecha parseada DD-MM-YYYY=%s", idx, fecha_emision)
                                except:
                                    fecha_emision = timezone.localdate()
                                    logger.warning("Importacion DTE fila=%s fecha invalida; usando hoy=%s", idx, fecha_emision)
                    else:
                        fecha_emision = timezone.localdate()
                        logger.warning("Importacion DTE fila=%s sin fecha; usando hoy=%s", idx, fecha_emision)
                    
                    dias_credito_str = str(fila.get('dias_credito', '30') or '30').strip()
                    dias_credito = int(dias_credito_str) if dias_credito_str and dias_credito_str.isdigit() else 30
                    fecha_vencimiento = fecha_emision + timedelta(days=dias_credito)
                    
                    # Preparar campos numéricos ANTES de verificar duplicados
                    bultos_str = str(fila.get('bultos', '') or '').strip()
                    bultos = int(bultos_str) if bultos_str and bultos_str.isdigit() else 0
                    
                    unidades_str = str(fila.get('unidades', '') or '').strip()
                    unidades = int(unidades_str) if unidades_str and unidades_str.isdigit() else 0
                    
                    descuento_str = str(fila.get('descuento', '') or '').strip().replace(',', '.')
                    descuento = Decimal(descuento_str) if descuento_str and descuento_str != '' else Decimal('0')
                    
                    referencias = str(fila.get('referencias', '') or '').strip()
                    
                    # Verificar si el DTE ya existe
                    dte_existente = Dte.objects.filter(
                        emisor=emisor,
                        numero_documento=int(numero_documento),
                        tipo_documento=tipo_documento
                    ).first()
                    
                    if dte_existente:
                        if modo_actualizacion == 'crear_y_actualizar':
                            # Actualizar DTE existente
                            dte_existente.monto_neto = monto_neto
                            dte_existente.monto_con_iva = total
                            dte_existente.fecha_emision = fecha_emision
                            dte_existente.fecha_vencimiento = fecha_vencimiento
                            dte_existente.diasCredito = dias_credito
                            dte_existente.bultos = bultos
                            dte_existente.unidades_productos = unidades
                            dte_existente.descuento = descuento
                            dte_existente.referencias = referencias
                            dte_existente.save()
                            dtes_actualizados += 1
                            logger.info(
                                "DTE compra actualizado por importacion: dte_id=%s numero=%s fecha=%s",
                                dte_existente.id,
                                dte_existente.numero_documento,
                                fecha_emision,
                            )
                            continue
                        else:
                            # Modo solo_crear: omitir duplicados
                            dtes_omitidos += 1
                            continue
                    
                    # Obtener empresa actual del usuario (receptor)
                    empresa_actual_id = request.session.get('idEmpresaActual')
                    receptor_id = fila.get('receptor_id') if fila.get('receptor_id') else empresa_actual_id
                    
                    # Crear DTE (responsable y receptor automáticos)
                    dte = Dte.objects.create(
                        emisor=emisor,
                        receptor_id=receptor_id,  # Empresa actual o especificada
                        numero_documento=int(numero_documento),
                        tipo_documento=tipo_documento,
                        monto_neto=monto_neto,
                        monto_con_iva=total,
                        estado_pago='Pendiente',  # Siempre pendiente para que el usuario registre el pago
                        estado_dte='EMITIDO',  # Estado por defecto
                        responsable=request.user.username,  # Usuario que importa
                        fecha_emision=fecha_emision,
                        fecha_vencimiento=fecha_vencimiento,
                        diasCredito=dias_credito,
                        bultos=bultos,
                        unidades_productos=unidades,
                        descuento=descuento,
                        tipo_transaccion='COMPRA',
                        referencias=referencias
                    )
                    
                    dtes_creados += 1
                    logger.info(
                        "DTE compra creado por importacion: dte_id=%s numero=%s emisor_id=%s receptor_id=%s",
                        dte.id,
                        dte.numero_documento,
                        emisor.id,
                        receptor_id,
                    )
                    
                except Exception as e:
                    errores.append(f'Fila {idx}: {str(e)}')
                    logger.warning("Error en fila de importacion DTE compras: fila=%s error=%s", idx, e)
                    continue
        
        logger.info(
            "Resumen importacion DTE compras: creados=%s actualizados=%s omitidos=%s errores=%s",
            dtes_creados,
            dtes_actualizados,
            dtes_omitidos,
            len(errores),
        )
        
        # Preparar mensaje
        mensaje = []
        if dtes_creados > 0:
            mensaje.append(f'{dtes_creados} DTEs creados')
        if dtes_actualizados > 0:
            mensaje.append(f'{dtes_actualizados} DTEs actualizados')
        if dtes_omitidos > 0:
            mensaje.append(f'{dtes_omitidos} DTEs omitidos (duplicados)')
        
        return JsonResponse({
            'success': True,
            'message': ', '.join(mensaje) if mensaje else 'No se procesaron DTEs',
            'dtes_creados': dtes_creados,
            'dtes_actualizados': dtes_actualizados,
            'dtes_omitidos': dtes_omitidos,
            'errores': errores
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al importar DTEs: {str(e)}'
        })


@require_GET
def descargar_formato_dtes(request):
    """Descargar formato CSV de ejemplo para importar DTEs"""
    try:
        tipo_busqueda = request.GET.get('tipo', 'rut')  # 'rut' o 'id'
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="formato_dtes.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezados según tipo de búsqueda (sin responsable)
        # IMPORTANTE: Los nombres deben coincidir EXACTAMENTE con lo que busca el código
        if tipo_busqueda == 'rut':
            writer.writerow([
                'rut_proveedor', 
                'numero_documento', 
                'tipo_documento', 
                'fecha_emision', 
                'monto_con_iva', 
                'dias_credito', 
                'bultos', 
                'unidades', 
                'referencias'
            ])
            # Ejemplo 1 (RUT sin puntos, monto con IVA)
            writer.writerow([
                '76123456-7', '12345', '33', 
                '2024-12-11', '119000', '30',
                '2', '50', 'Orden de Compra 001'
            ])
            # Ejemplo 2
            writer.writerow([
                '77234567-8', '12346', '33', 
                '2024-12-10', '297500', '45',
                '5', '100', 'Orden de Compra 002'
            ])
        else:  # id
            writer.writerow([
                'id_proveedor', 
                'numero_documento', 
                'tipo_documento', 
                'fecha_emision', 
                'monto_con_iva', 
                'dias_credito',
                'bultos', 
                'unidades', 
                'referencias'
            ])
            writer.writerow([
                '1', '12345', '33',
                '2024-12-11', '119000', '30',
                '2', '50', 'Orden de Compra 001'
            ])
            writer.writerow([
                '2', '12346', '33',
                '2024-12-10', '297500', '45',
                '5', '100', 'Orden de Compra 002'
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al generar formato: {str(e)}'
        })


@require_GET
def exportar_dtes_actuales(request):
    """Exportar DTEs de compras actuales a CSV"""
    try:
        tipo_exportacion = request.GET.get('tipo', 'rut')  # 'rut' o 'id'
        
        # Obtener DTEs de compras
        dtes = Dte.objects.filter(tipo_transaccion='COMPRA').select_related('emisor').order_by('-fecha_emision')
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="dtes_compras_actuales.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Encabezados según tipo
        if tipo_exportacion == 'rut':
            writer.writerow([
                'id_dte', 'rut_proveedor', 'nombre_proveedor', 'numero_documento', 'tipo_documento',
                'fecha_emision', 'monto_neto', 'monto_iva', 'total', 'dias_credito',
                'bultos', 'unidades', 'referencias', 'estado_dte', 'estado_pago'
            ])
            
            for dte in dtes:
                # Limpiar RUT (sin puntos)
                rut_limpio = dte.emisor.rut.replace('.', '') if dte.emisor.rut else ''
                
                writer.writerow([
                    dte.id,
                    rut_limpio,
                    dte.emisor.nombre,
                    dte.numero_documento,
                    dte.tipo_documento,
                    dte.fecha_emision.strftime('%Y-%m-%d'),
                    float(dte.monto_neto),
                    float(dte.monto_con_iva - dte.monto_neto),
                    float(dte.monto_con_iva),
                    dte.diasCredito,
                    dte.bultos,
                    dte.unidades_productos,
                    dte.referencias or '',
                    dte.estado_dte,
                    dte.estado_pago
                ])
        else:  # id
            writer.writerow([
                'id_dte', 'id_proveedor', 'nombre_proveedor', 'numero_documento', 'tipo_documento',
                'fecha_emision', 'monto_neto', 'monto_iva', 'total', 'dias_credito',
                'bultos', 'unidades', 'referencias', 'estado_dte', 'estado_pago'
            ])
            
            for dte in dtes:
                writer.writerow([
                    dte.id,
                    dte.emisor.id,
                    dte.emisor.nombre,
                    dte.numero_documento,
                    dte.tipo_documento,
                    dte.fecha_emision.strftime('%Y-%m-%d'),
                    float(dte.monto_neto),
                    float(dte.monto_con_iva - dte.monto_neto),
                    float(dte.monto_con_iva),
                    dte.diasCredito,
                    dte.bultos,
                    dte.unidades_productos,
                    dte.referencias or '',
                    dte.estado_dte,
                    dte.estado_pago
                ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar DTEs: {str(e)}'
        })


@require_GET
def exportar_dtes_excel(request):
    """Exportar DTEs de compras actuales a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        tipo_exportacion = request.GET.get('tipo', 'rut')
        
        # Obtener DTEs
        dtes = Dte.objects.filter(tipo_transaccion='COMPRA').select_related('emisor').order_by('-fecha_emision')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DTEs Compras"
        
        # Estilos
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Encabezados
        if tipo_exportacion == 'rut':
            headers = [
                'ID DTE', 'RUT Proveedor', 'Nombre Proveedor', 'Nº Documento', 'Tipo',
                'Fecha Emisión', 'Monto Neto', 'IVA', 'Total', 'Días Crédito',
                'Bultos', 'Unidades', 'Referencias', 'Estado DTE', 'Estado Pago'
            ]
        else:
            headers = [
                'ID DTE', 'ID Proveedor', 'Nombre Proveedor', 'Nº Documento', 'Tipo',
                'Fecha Emisión', 'Monto Neto', 'IVA', 'Total', 'Días Crédito',
                'Bultos', 'Unidades', 'Referencias', 'Estado DTE', 'Estado Pago'
            ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, dte in enumerate(dtes, start=2):
            if tipo_exportacion == 'rut':
                # Limpiar RUT (sin puntos)
                rut_limpio = dte.emisor.rut.replace('.', '') if dte.emisor.rut else ''
                ws.cell(row=row_idx, column=1, value=dte.id)
                ws.cell(row=row_idx, column=2, value=rut_limpio)
            else:
                ws.cell(row=row_idx, column=1, value=dte.id)
                ws.cell(row=row_idx, column=2, value=dte.emisor.id)
            
            ws.cell(row=row_idx, column=3, value=dte.emisor.nombre)
            ws.cell(row=row_idx, column=4, value=dte.numero_documento)
            ws.cell(row=row_idx, column=5, value=dte.tipo_documento)
            ws.cell(row=row_idx, column=6, value=dte.fecha_emision)
            ws.cell(row=row_idx, column=7, value=float(dte.monto_neto))
            ws.cell(row=row_idx, column=8, value=float(dte.monto_con_iva - dte.monto_neto))
            ws.cell(row=row_idx, column=9, value=float(dte.monto_con_iva))
            ws.cell(row=row_idx, column=10, value=dte.diasCredito)
            ws.cell(row=row_idx, column=11, value=dte.bultos)
            ws.cell(row=row_idx, column=12, value=dte.unidades_productos)
            ws.cell(row=row_idx, column=13, value=dte.referencias or '')
            ws.cell(row=row_idx, column=14, value=dte.estado_dte)
            ws.cell(row=row_idx, column=15, value=dte.estado_pago)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dtes_compras_actuales.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar a Excel: {str(e)}'
        })


# ========== DASHBOARD COMPRAS MEJORADO ==========

@login_required
def verDashboardComprasMejorado(request):
    """Vista principal del dashboard de compras mejorado"""
    return render(request, 'vistas/modulo_dashboards/dashboard_compras_mejorado.html')


@login_required
@require_GET
def dashboard_compras_mejorado_api(request):
    """
    API completa para el dashboard de compras mejorado.
    Proporciona métricas estratégicas, gráficos y análisis para toma de decisiones.
    """
    try:
        from datetime import datetime, timedelta
        from django.db.models.functions import TruncMonth, ExtractMonth
        
        # Parámetros de filtro
        anio = int(request.GET.get('anio', timezone.localdate().year))
        periodo = request.GET.get('periodo', 'anual')
        temporada = request.GET.get('temporada', '')
        proveedor_id = request.GET.get('proveedor', '')
        
        # Calcular rango de fechas según período
        hoy = timezone.localtime()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
        else:  # anual
            fecha_inicio = datetime(anio, 1, 1)
        
        fecha_fin = hoy
        
        # Query base para compras
        compras_query = Compras.objects.filter(fecha__year=anio)
        
        # Filtro de temporada: acepta código de familia normalizada
        # (VERANO/OTONO/INVIERNO/PRIMAVERA) o texto libre (fallback legacy).
        if temporada:
            _t_norm = temporada.strip().upper()
            if _t_norm in {'VERANO', 'OTONO', 'INVIERNO', 'PRIMAVERA'}:
                compras_query = compras_query.filter(temporada_familia=_t_norm)
            else:
                compras_query = compras_query.filter(temporada__icontains=temporada)
        if proveedor_id:
            compras_query = compras_query.filter(empresa_id=proveedor_id)

        # IDs de compras para filtrar relaciones
        compras_ids = list(compras_query.values_list('id', flat=True))
        
        # ===== MÉTRICAS PRINCIPALES =====
        metricas = calcular_metricas_principales_mejorado(compras_query, compras_ids, anio)
        
        # ===== EVOLUCIÓN MENSUAL =====
        evolucion_mensual = calcular_evolucion_mensual_mejorado(anio, temporada, proveedor_id)
        
        # ===== PARETO PROVEEDORES =====
        pareto_proveedores = calcular_pareto_proveedores_mejorado(compras_query, compras_ids)
        
        # ===== COMPARATIVA ANUAL =====
        comparativa_anual = calcular_comparativa_anual_mejorado(anio, temporada, proveedor_id)
        
        # ===== ROI POR TEMPORADA =====
        roi_temporadas = calcular_roi_temporadas_mejorado(compras_query, compras_ids)
        
        # ===== TOP PROVEEDORES =====
        top_proveedores = calcular_top_proveedores_mejorado(compras_query, compras_ids)
        
        # ===== TOP PRODUCTOS =====
        top_productos = calcular_top_productos_mejorado(compras_ids)
        
        # ===== CUMPLIMIENTO PROVEEDORES =====
        cumplimiento_proveedores = calcular_cumplimiento_proveedores_mejorado(compras_query, compras_ids)
        
        # ===== RENDIMIENTO DETALLADO =====
        rendimiento_detallado = calcular_rendimiento_detallado_mejorado(compras_query, compras_ids)
        
        # ===== ALERTAS INTELIGENTES =====
        alertas = generar_alertas_compras_mejorado(metricas, cumplimiento_proveedores, roi_temporadas)
        
        # ===== INSIGHTS ESTRATÉGICOS =====
        insights = generar_insights_compras_mejorado(metricas, pareto_proveedores, comparativa_anual)
        
        # ===== MÉTRICAS DE DISTRIBUCIÓN (CENTRO DE COMPRAS) =====
        import logging
        _logger = logging.getLogger(__name__)

        try:
            distribucion = calcular_metricas_distribucion(anio, compras_ids)
        except Exception as e:
            _logger.warning(f'Error en metricas distribucion: {e}')
            distribucion = {'unidades_compradas': 0, 'unidades_despachadas': 0, 'stock_centro_distribucion': 0, 'eficiencia_distribucion': 0}

        try:
            despachos_sucursal = calcular_despachos_por_sucursal(anio)
        except Exception as e:
            _logger.warning(f'Error en despachos sucursal: {e}')
            despachos_sucursal = []

        try:
            sucursales_destino = calcular_rendimiento_sucursales_destino(anio)
        except Exception as e:
            _logger.warning(f'Error en rendimiento sucursales destino: {e}')
            sucursales_destino = []

        try:
            flujo_distribucion = calcular_flujo_distribucion_mensual(anio)
        except Exception as e:
            _logger.warning(f'Error en flujo distribucion: {e}')
            flujo_distribucion = []

        # ===== MÁRGENES CENTRO DE DISTRIBUCIÓN =====
        try:
            margenes_cd = calcular_margenes_centro_distribucion(anio)
        except Exception as e:
            _logger.warning(f'Error en margenes CD: {e}')
            margenes_cd = {'margen_total_cd': 0, 'costo_proveedor_total': 0, 'costo_destino_total': 0, 'margen_promedio_pct': 0, 'unidades_despachadas': 0, 'detalle_por_sucursal': [], 'centros_distribucion': []}

        try:
            comparativa_costos = calcular_comparativa_costos_cd_vs_sucursales(anio)
        except Exception as e:
            _logger.warning(f'Error en comparativa costos: {e}')
            comparativa_costos = []

        try:
            rentabilidad_tipo = calcular_rentabilidad_por_tipo_sucursal(anio)
        except Exception as e:
            _logger.warning(f'Error en rentabilidad tipo: {e}')
            rentabilidad_tipo = {'centros_distribucion': [], 'sucursales_vendedoras': []}
        
        return JsonResponse({
            'success': True,
            'metricas': metricas,
            'evolucion_mensual': evolucion_mensual,
            'pareto_proveedores': pareto_proveedores,
            'comparativa_anual': comparativa_anual,
            'roi_temporadas': roi_temporadas,
            'top_proveedores': top_proveedores,
            'top_productos': top_productos,
            'cumplimiento_proveedores': cumplimiento_proveedores,
            'rendimiento_detallado': rendimiento_detallado,
            'alertas': alertas,
            'insights': insights,
            # Datos de distribución (Centro de Compras)
            'distribucion': distribucion,
            'despachos_sucursal': despachos_sucursal,
            'sucursales_destino': sucursales_destino,
            'flujo_distribucion': flujo_distribucion,
            # Datos de márgenes Centro de Distribución
            'margenes_cd': margenes_cd,
            'comparativa_costos': comparativa_costos,
            'rentabilidad_tipo_sucursal': rentabilidad_tipo,
            'filtros_aplicados': {
                'anio': anio,
                'periodo': periodo,
                'temporada': temporada,
                'proveedor_id': proveedor_id
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'Error al generar dashboard: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


def calcular_metricas_principales_mejorado(compras_query, compras_ids, anio):
    """Calcula las métricas principales del dashboard"""
    
    # Total compras
    total_compras = compras_query.count()
    
    # Productos y tallas
    productos_compras = Compras_Producto.objects.filter(compras__in=compras_ids)
    tallas_compras = Compras_Producto_Talla.objects.filter(
        compra_producto__compras__in=compras_ids
    )
    
    # Unidades esperadas y recepcionadas
    unidades_esperadas = tallas_compras.aggregate(
        total=Sum('stock')
    )['total'] or 0
    
    recepciones = Productos_Recepcionados.objects.filter(
        compra_producto_talla__compra_producto__compras__in=compras_ids
    )
    unidades_recepcionadas = recepciones.aggregate(
        total=Sum('stockArribado')
    )['total'] or 0
    
    # Cumplimiento general
    cumplimiento_general = 0
    if unidades_esperadas > 0:
        cumplimiento_general = round((unidades_recepcionadas / unidades_esperadas) * 100, 1)
    
    # Inversión total
    inversion_total = productos_compras.aggregate(
        total=Sum(F('costo') * F('compras_producto_talla__stock'))
    )['total'] or 0
    
    # Valor de venta esperado
    valor_venta = productos_compras.aggregate(
        total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
    )['total'] or 0
    
    # ROI promedio
    roi_promedio = 0
    if inversion_total > 0:
        roi_promedio = round(((valor_venta - inversion_total) / inversion_total) * 100, 1)
    
    # ===== TENDENCIAS vs AÑO ANTERIOR =====
    compras_anterior = Compras.objects.filter(fecha__year=anio-1)
    total_anterior = compras_anterior.count()
    
    compras_anterior_ids = list(compras_anterior.values_list('id', flat=True))
    productos_anterior = Compras_Producto.objects.filter(compras__in=compras_anterior_ids)
    inversion_anterior = productos_anterior.aggregate(
        total=Sum(F('costo') * F('compras_producto_talla__stock'))
    )['total'] or 1  # Evitar división por cero
    
    trend_compras = 0
    if total_anterior > 0:
        trend_compras = round(((total_compras - total_anterior) / total_anterior) * 100, 1)
    
    trend_inversion = 0
    if inversion_anterior > 0:
        trend_inversion = round(((float(inversion_total) - float(inversion_anterior)) / float(inversion_anterior)) * 100, 1)
    
    # ROI anterior para tendencia
    valor_anterior = productos_anterior.aggregate(
        total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
    )['total'] or 0
    roi_anterior = 0
    if inversion_anterior > 0:
        roi_anterior = ((valor_anterior - inversion_anterior) / inversion_anterior) * 100
    
    trend_roi = round(roi_promedio - roi_anterior, 1)
    
    return {
        'total_compras': total_compras,
        'inversion_total': float(inversion_total) if inversion_total else 0,
        'valor_venta': float(valor_venta) if valor_venta else 0,
        'unidades_esperadas': int(unidades_esperadas),
        'unidades_recepcionadas': int(unidades_recepcionadas),
        'cumplimiento_general': cumplimiento_general,
        'roi_promedio': roi_promedio,
        'productos_distintos': productos_compras.values('nombre').distinct().count(),
        'proveedores_activos': compras_query.values('empresa').distinct().count(),
        'trend_compras': trend_compras,
        'trend_inversion': trend_inversion,
        'trend_roi': trend_roi
    }


def calcular_evolucion_mensual_mejorado(anio, temporada='', proveedor_id=''):
    """Calcula la evolución mensual de compras vs ventas"""
    from .models import Ticket
    
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    resultado = []
    
    for mes in range(1, 13):
        # Query de compras del mes
        compras_mes = Compras.objects.filter(fecha__year=anio, fecha__month=mes)

        if temporada:
            _t_norm = temporada.strip().upper()
            if _t_norm in {'VERANO', 'OTONO', 'INVIERNO', 'PRIMAVERA'}:
                compras_mes = compras_mes.filter(temporada_familia=_t_norm)
            else:
                compras_mes = compras_mes.filter(temporada__icontains=temporada)
        if proveedor_id:
            compras_mes = compras_mes.filter(empresa_id=proveedor_id)
        
        compras_mes_ids = list(compras_mes.values_list('id', flat=True))
        
        # Inversión del mes
        productos_mes = Compras_Producto.objects.filter(compras__in=compras_mes_ids)
        inversion_mes = productos_mes.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # Ventas relacionadas del mes (simplificado - ventas totales del mes)
        try:
            ventas_mes = Ticket.objects.filter(
                created_at__year=anio,
                created_at__month=mes,
                estado='PAGADO'
            ).aggregate(total=Sum('total'))['total'] or 0
        except:
            ventas_mes = 0
        
        resultado.append({
            'mes': mes,
            'mes_nombre': meses_nombres[mes - 1],
            'inversion': float(inversion_mes),
            'ventas': float(ventas_mes),
            'total_compras': compras_mes.count()
        })
    
    return resultado


def calcular_pareto_proveedores_mejorado(compras_query, compras_ids):
    """Calcula el análisis Pareto (80/20) de proveedores"""
    
    proveedores = compras_query.values('empresa__id', 'empresa__nombre').distinct()
    
    inversiones = []
    for prov in proveedores:
        if not prov['empresa__id']:
            continue
            
        compras_prov = compras_query.filter(empresa_id=prov['empresa__id'])
        compras_prov_ids = list(compras_prov.values_list('id', flat=True))
        
        productos_prov = Compras_Producto.objects.filter(compras__in=compras_prov_ids)
        inversion = productos_prov.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        if inversion > 0:
            inversiones.append({
                'proveedor': prov['empresa__nombre'] or 'Sin nombre',
                'inversion': float(inversion)
            })
    
    # Ordenar por inversión descendente
    inversiones.sort(key=lambda x: x['inversion'], reverse=True)
    
    return inversiones[:15]  # Top 15 para el gráfico


def calcular_comparativa_anual_mejorado(anio, temporada='', proveedor_id=''):
    """Calcula comparativa de inversión mes a mes: año actual vs anterior.

    El filtro ``temporada`` acepta:
      - Código de familia normalizada (VERANO / OTONO / INVIERNO / PRIMAVERA)
        → filtra por ``temporada_familia``.
      - Texto libre → fallback por ``temporada__icontains`` (compatibilidad).
    """

    FAMILIAS_NORMALIZADAS = {'VERANO', 'OTONO', 'INVIERNO', 'PRIMAVERA'}

    def _aplicar_filtros(qs):
        if temporada:
            t_norm = temporada.strip().upper()
            if t_norm in FAMILIAS_NORMALIZADAS:
                qs = qs.filter(temporada_familia=t_norm)
            else:
                qs = qs.filter(temporada__icontains=temporada)
        if proveedor_id:
            qs = qs.filter(empresa_id=proveedor_id)
        return qs

    resultado = {
        'actual': [],
        'anterior': []
    }

    for mes in range(1, 13):
        compras_actual = _aplicar_filtros(
            Compras.objects.filter(fecha__year=anio, fecha__month=mes)
        )
        compras_actual_ids = list(compras_actual.values_list('id', flat=True))
        productos_actual = Compras_Producto.objects.filter(compras__in=compras_actual_ids)
        inversion_actual = productos_actual.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0

        compras_anterior = _aplicar_filtros(
            Compras.objects.filter(fecha__year=anio - 1, fecha__month=mes)
        )
        compras_anterior_ids = list(compras_anterior.values_list('id', flat=True))
        productos_anterior = Compras_Producto.objects.filter(compras__in=compras_anterior_ids)
        inversion_anterior = productos_anterior.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0

        resultado['actual'].append(float(inversion_actual))
        resultado['anterior'].append(float(inversion_anterior))

    return resultado


def calcular_roi_temporadas_mejorado(compras_query, compras_ids):
    """Calcula ROI por temporada, preferentemente agrupando por
    `temporada_familia` + `temporada_anio` (normalizados) para permitir
    comparativas YoY reales (ej. Invierno 2025 vs Invierno 2026).

    Fallback: si la compra aún no tiene los campos normalizados (data legacy),
    se agrupa por el texto libre `temporada` como antes.

    Devuelve una lista de objetos:
      {
        'temporada': 'Invierno 2025',
        'temporada_familia': 'INVIERNO',
        'temporada_anio': 2025,
        'roi': 45.2,
        'inversion': 1234.0,
        'valor_venta': 1800.0,
        # Comparativa YoY (mismo rubro, año anterior)
        'roi_anterior': 38.1,
        'inversion_anterior': 1000.0,
        'delta_roi': 7.1,
        'delta_inversion_pct': 23.4,
      }
    """

    FAMILIA_LEGIBLE = {
        'VERANO': 'Verano',
        'OTONO': 'Otoño',
        'INVIERNO': 'Invierno',
        'PRIMAVERA': 'Primavera',
    }

    def _metricas(qs_compras):
        ids = list(qs_compras.values_list('id', flat=True))
        prods = Compras_Producto.objects.filter(compras__in=ids)
        inv = prods.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        val = prods.aggregate(
            total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
        )['total'] or 0
        roi = round(((float(val) - float(inv)) / float(inv)) * 100, 1) if inv else 0
        return float(inv), float(val), roi

    resultado = []

    # 1) Rubros normalizados (familia + año) — base de la comparativa YoY
    rubros_normalizados = compras_query.exclude(
        temporada_familia__isnull=True
    ).exclude(
        temporada_anio__isnull=True
    ).values('temporada_familia', 'temporada_anio').distinct()

    familias_anios_vistos = set()

    for r in rubros_normalizados:
        familia = r['temporada_familia']
        anio = r['temporada_anio']
        familias_anios_vistos.add((familia, anio))

        compras_temp = compras_query.filter(
            temporada_familia=familia, temporada_anio=anio
        )
        inv, val, roi = _metricas(compras_temp)

        # Año anterior (mismo rubro) — NO filtra por compras_query porque el
        # query original ya está acotado al año actual; buscamos el anterior
        # en toda la tabla de compras (respetando eliminadas no incluidas aquí).
        compras_temp_ant = Compras.objects.filter(
            temporada_familia=familia,
            temporada_anio=anio - 1,
        ).exclude(estado='ELIMINADA')
        inv_ant, _val_ant, roi_ant = _metricas(compras_temp_ant)

        delta_roi = round(roi - roi_ant, 1)
        delta_inv_pct = 0.0
        if inv_ant > 0:
            delta_inv_pct = round(((inv - inv_ant) / inv_ant) * 100, 1)

        resultado.append({
            'temporada': f"{FAMILIA_LEGIBLE.get(familia, familia)} {anio}",
            'temporada_familia': familia,
            'temporada_anio': anio,
            'roi': roi,
            'inversion': inv,
            'valor_venta': val,
            'roi_anterior': roi_ant,
            'inversion_anterior': inv_ant,
            'delta_roi': delta_roi,
            'delta_inversion_pct': delta_inv_pct,
        })

    # 2) Fallback: compras viejas sin campos normalizados (texto libre)
    legacy = compras_query.filter(
        Q(temporada_familia__isnull=True) | Q(temporada_anio__isnull=True)
    ).values('temporada').distinct()

    for temp in legacy:
        temporada_nombre = temp['temporada']
        if not temporada_nombre:
            continue
        compras_temp = compras_query.filter(
            temporada=temporada_nombre,
            temporada_familia__isnull=True,
        )
        inv, val, roi = _metricas(compras_temp)
        resultado.append({
            'temporada': temporada_nombre,
            'temporada_familia': None,
            'temporada_anio': None,
            'roi': roi,
            'inversion': inv,
            'valor_venta': val,
            # Sin comparativo YoY confiable cuando no hay familia/año
            'roi_anterior': 0,
            'inversion_anterior': 0,
            'delta_roi': 0,
            'delta_inversion_pct': 0,
        })

    return resultado


def calcular_top_proveedores_mejorado(compras_query, compras_ids):
    """Calcula top 10 proveedores por inversión con cumplimiento"""
    
    proveedores = compras_query.values('empresa__id', 'empresa__nombre').distinct()
    
    resultado = []
    for prov in proveedores:
        if not prov['empresa__id']:
            continue
        
        compras_prov = compras_query.filter(empresa_id=prov['empresa__id'])
        compras_prov_ids = list(compras_prov.values_list('id', flat=True))
        
        productos_prov = Compras_Producto.objects.filter(compras__in=compras_prov_ids)
        inversion = productos_prov.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # Cumplimiento
        tallas_prov = Compras_Producto_Talla.objects.filter(
            compra_producto__compras__in=compras_prov_ids
        )
        esperadas = tallas_prov.aggregate(total=Sum('stock'))['total'] or 0
        
        recepciones = Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras__in=compras_prov_ids
        )
        recibidas = recepciones.aggregate(total=Sum('stockArribado'))['total'] or 0
        
        cumplimiento = 0
        if esperadas > 0:
            cumplimiento = round((recibidas / esperadas) * 100, 1)
        
        if inversion > 0:
            resultado.append({
                'proveedor': prov['empresa__nombre'] or 'Sin nombre',
                'inversion': float(inversion),
                'total_compras': compras_prov.count(),
                'cumplimiento': cumplimiento
            })
    
    # Ordenar y limitar
    resultado.sort(key=lambda x: x['inversion'], reverse=True)
    return resultado[:10]


def calcular_top_productos_mejorado(compras_ids):
    """Calcula top 10 productos por inversión"""
    
    productos = Compras_Producto.objects.filter(
        compras__in=compras_ids
    ).values(
        'nombre', 'atributo1'  # atributo1 = marca
    ).annotate(
        inversion_total=Sum(F('costo') * F('compras_producto_talla__stock')),
        unidades_total=Sum('compras_producto_talla__stock')
    ).order_by('-inversion_total')[:10]
    
    resultado = []
    for prod in productos:
        resultado.append({
            'nombre': prod['nombre'] or 'Sin nombre',
            'marca': prod['atributo1'] or '-',
            'inversion': float(prod['inversion_total'] or 0),
            'unidades': int(prod['unidades_total'] or 0)
        })
    
    return resultado


def calcular_cumplimiento_proveedores_mejorado(compras_query, compras_ids):
    """Calcula cumplimiento detallado por proveedor"""
    
    proveedores = compras_query.values('empresa__id', 'empresa__nombre').distinct()
    
    resultado = []
    for prov in proveedores:
        if not prov['empresa__id']:
            continue
        
        compras_prov = compras_query.filter(empresa_id=prov['empresa__id'])
        compras_prov_ids = list(compras_prov.values_list('id', flat=True))
        
        tallas_prov = Compras_Producto_Talla.objects.filter(
            compra_producto__compras__in=compras_prov_ids
        )
        esperadas = tallas_prov.aggregate(total=Sum('stock'))['total'] or 0
        
        recepciones = Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras__in=compras_prov_ids
        )
        recibidas = recepciones.aggregate(total=Sum('stockArribado'))['total'] or 0
        
        cumplimiento = 0
        if esperadas > 0:
            cumplimiento = round((recibidas / esperadas) * 100, 1)
        
        if esperadas > 0:
            resultado.append({
                'proveedor': prov['empresa__nombre'] or 'Sin nombre',
                'cumplimiento': cumplimiento,
                'esperadas': int(esperadas),
                'recibidas': int(recibidas)
            })
    
    resultado.sort(key=lambda x: x['cumplimiento'], reverse=True)
    return resultado[:12]


def calcular_rendimiento_detallado_mejorado(compras_query, compras_ids):
    """Calcula rendimiento detallado por compra para la tabla"""
    
    resultado = []
    
    for compra in compras_query.select_related('empresa')[:20]:
        productos_compra = Compras_Producto.objects.filter(compras=compra)
        
        inversion = productos_compra.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        valor_venta = productos_compra.aggregate(
            total=Sum(F('precioSugerido') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        tallas_compra = Compras_Producto_Talla.objects.filter(
            compra_producto__compras=compra
        )
        unidades_esperadas = tallas_compra.aggregate(total=Sum('stock'))['total'] or 0
        
        recepciones = Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras=compra
        )
        unidades_recibidas = recepciones.aggregate(total=Sum('stockArribado'))['total'] or 0
        
        cumplimiento = 0
        if unidades_esperadas > 0:
            cumplimiento = round((unidades_recibidas / unidades_esperadas) * 100, 1)
        
        roi = 0
        if inversion > 0:
            roi = round(((valor_venta - inversion) / inversion) * 100, 1)
        
        if cumplimiento >= 100:
            estado = 'Completado'
        elif cumplimiento >= 80:
            estado = 'Pendiente'
        else:
            estado = 'Retrasado'
        
        resultado.append({
            'nombre': compra.nombre if hasattr(compra, 'nombre') and compra.nombre else f'Compra #{compra.id}',
            'proveedor': compra.empresa.nombre if compra.empresa else 'Sin proveedor',
            'temporada': compra.temporada if hasattr(compra, 'temporada') and compra.temporada else 'N/A',
            'inversion': float(inversion),
            'cumplimiento': cumplimiento,
            'roi': roi,
            'unidades_esperadas': int(unidades_esperadas),
            'unidades_recibidas': int(unidades_recibidas),
            'estado': estado
        })
    
    return resultado


def generar_alertas_compras_mejorado(metricas, cumplimiento_proveedores, roi_temporadas):
    """Genera alertas inteligentes basadas en métricas"""
    
    alertas = []
    
    if metricas['cumplimiento_general'] < 80:
        alertas.append({
            'tipo': 'warning',
            'titulo': 'Cumplimiento Bajo',
            'mensaje': f"El cumplimiento general ({metricas['cumplimiento_general']}%) está por debajo del objetivo del 80%."
        })
    
    proveedores_criticos = [p for p in cumplimiento_proveedores if p['cumplimiento'] < 70]
    if proveedores_criticos:
        alertas.append({
            'tipo': 'danger',
            'titulo': 'Proveedores Críticos',
            'mensaje': f"{len(proveedores_criticos)} proveedor(es) tienen cumplimiento inferior al 70%."
        })
    
    temporadas_bajo_roi = [t for t in roi_temporadas if t['roi'] < 15]
    if temporadas_bajo_roi:
        temp_nombres = ', '.join([t['temporada'] for t in temporadas_bajo_roi])
        alertas.append({
            'tipo': 'warning',
            'titulo': 'ROI Bajo por Temporada',
            'mensaje': f"Las temporadas {temp_nombres} tienen ROI inferior al 15%."
        })
    
    diferencia = metricas['unidades_esperadas'] - metricas['unidades_recepcionadas']
    if diferencia > 0 and metricas['unidades_esperadas'] > 0:
        porcentaje_pendiente = (diferencia / metricas['unidades_esperadas']) * 100
        if porcentaje_pendiente > 20:
            alertas.append({
                'tipo': 'info',
                'titulo': 'Recepciones Pendientes',
                'mensaje': f"Faltan {diferencia:,} unidades por recepcionar ({porcentaje_pendiente:.1f}%)."
            })
    
    if len(alertas) == 0 and metricas['total_compras'] > 0:
        alertas.append({
            'tipo': 'success',
            'titulo': 'Excelente Desempeño',
            'mensaje': 'Todos los indicadores están dentro de los parámetros esperados.'
        })
    
    return alertas


def generar_insights_compras_mejorado(metricas, pareto_proveedores, comparativa_anual):
    """Genera insights estratégicos para toma de decisiones"""
    
    insights = []
    
    if len(pareto_proveedores) >= 2:
        total_inversion = sum(p['inversion'] for p in pareto_proveedores)
        if total_inversion > 0:
            top_2 = pareto_proveedores[:2]
            concentracion = sum(p['inversion'] for p in top_2) / total_inversion * 100
            
            if concentracion > 60:
                insights.append({
                    'titulo': 'Alta Concentración',
                    'descripcion': f'Los 2 principales proveedores representan {concentracion:.1f}% de la inversión.',
                    'valor': f'{concentracion:.0f}%',
                    'icono': 'bi-building',
                    'color': 'warning'
                })
    
    total_actual = sum(comparativa_anual.get('actual', []))
    total_anterior = sum(comparativa_anual.get('anterior', []))
    
    if total_anterior > 0:
        crecimiento = ((total_actual - total_anterior) / total_anterior) * 100
        
        if crecimiento > 10:
            insights.append({
                'titulo': 'Crecimiento Positivo',
                'descripcion': 'La inversión en compras ha aumentado respecto al año anterior.',
                'valor': f'+{crecimiento:.1f}%',
                'icono': 'bi-graph-up-arrow',
                'color': 'success'
            })
        elif crecimiento < -10:
            insights.append({
                'titulo': 'Reducción de Inversión',
                'descripcion': 'La inversión en compras ha disminuido respecto al año anterior.',
                'valor': f'{crecimiento:.1f}%',
                'icono': 'bi-graph-down-arrow',
                'color': 'danger'
            })
    
    if metricas['roi_promedio'] >= 25:
        insights.append({
            'titulo': 'ROI Excelente',
            'descripcion': 'El retorno sobre inversión supera el 25%.',
            'valor': f"{metricas['roi_promedio']}%",
            'icono': 'bi-trophy',
            'color': 'success'
        })
    
    return insights


# ========== FUNCIONES DE DISTRIBUCIÓN (CENTRO DE COMPRAS) ==========

def calcular_metricas_distribucion(anio, compras_ids):
    """
    Calcula métricas de distribución desde el centro de compras hacia sucursales vendedoras.
    Analiza el flujo: Compras → Recepciones → Despachos → Ventas
    """
    from .models import Movimientos_Producto, Traspaso, Traspaso_Detalle, Ticket, Ticket_Productos
    
    # Unidades compradas (recepcionadas de proveedores)
    unidades_compradas = Productos_Recepcionados.objects.filter(
        compra_producto_talla__compra_producto__compras__in=compras_ids
    ).aggregate(total=Sum('stockArribado'))['total'] or 0
    
    # Unidades despachadas a otras sucursales (traspasos salida)
    unidades_despachadas = Movimientos_Producto.objects.filter(
        fecha__year=anio,
        concepto='TRASPASO_SALIDA',
        estado='COMPLETADO'
    ).aggregate(total=Sum('cantidad'))['total'] or 0
    
    # También contar desde Traspasos completados
    traspasos_despachados = Traspaso_Detalle.objects.filter(
        traspaso__fecha_solicitud__year=anio,
        traspaso__estado__in=['EN_TRANSITO', 'RECIBIDO']
    ).aggregate(total=Sum('cantidad_enviada'))['total'] or 0
    
    unidades_despachadas = max(unidades_despachadas, traspasos_despachados)
    
    # Stock en centro de distribución (comprado - despachado)
    stock_centro = max(0, unidades_compradas - unidades_despachadas)
    
    # Eficiencia de distribución
    eficiencia = 0
    if unidades_compradas > 0:
        eficiencia = round((unidades_despachadas / unidades_compradas) * 100, 1)
    
    return {
        'unidades_compradas': int(unidades_compradas),
        'unidades_despachadas': int(unidades_despachadas),
        'stock_centro_distribucion': int(stock_centro),
        'eficiencia_distribucion': eficiencia
    }


def calcular_despachos_por_sucursal(anio):
    """Calcula los despachos realizados a cada sucursal destino"""
    from .models import Movimientos_Producto, Traspaso, Traspaso_Detalle
    
    # Desde movimientos de producto
    despachos = Movimientos_Producto.objects.filter(
        fecha__year=anio,
        concepto='TRASPASO_SALIDA',
        estado='COMPLETADO',
        sucursal_destino__isnull=False
    ).values(
        'sucursal_destino__id',
        'sucursal_destino__alias',
        'sucursal_destino__empresa__nombre'
    ).annotate(
        unidades=Sum('cantidad')
    ).order_by('-unidades')
    
    resultado = []
    for d in despachos:
        resultado.append({
            'sucursal_id': d['sucursal_destino__id'],
            'sucursal': d['sucursal_destino__alias'] or 'Sin nombre',
            'empresa': d['sucursal_destino__empresa__nombre'] or '-',
            'unidades': int(abs(d['unidades'] or 0))  # abs porque los egresos son negativos
        })
    
    # Si no hay datos en movimientos, intentar con Traspasos
    if not resultado:
        traspasos = Traspaso.objects.filter(
            fecha_solicitud__year=anio,
            estado__in=['EN_TRANSITO', 'RECIBIDO']
        ).values(
            'sucursal_destino__id',
            'sucursal_destino__alias',
            'sucursal_destino__empresa__nombre'
        ).annotate(
            unidades=Sum('detalles__cantidad_enviada')
        ).order_by('-unidades')
        
        for t in traspasos:
            if t['unidades']:
                resultado.append({
                    'sucursal_id': t['sucursal_destino__id'],
                    'sucursal': t['sucursal_destino__alias'] or 'Sin nombre',
                    'empresa': t['sucursal_destino__empresa__nombre'] or '-',
                    'unidades': int(t['unidades'] or 0)
                })
    
    return resultado[:10]


def calcular_rendimiento_sucursales_destino(anio):
    """
    Calcula el rendimiento de cada sucursal destino:
    - Unidades despachadas recibidas
    - Unidades vendidas
    - Ventas en dinero
    - Margen estimado
    """
    from .models import Movimientos_Producto, Traspaso, Ticket, Ticket_Productos, Sucursal
    
    resultado = []
    
    try:
        # Obtener sucursales que han recibido mercadería
        sucursales_destino = Movimientos_Producto.objects.filter(
            fecha__year=anio,
            concepto='TRASPASO_ENTRADA',
            estado='COMPLETADO'
        ).values('sucursal_destino__id').distinct()
        
        sucursal_ids = [s['sucursal_destino__id'] for s in sucursales_destino if s['sucursal_destino__id']]
    except:
        return resultado
    
    for suc_id in sucursal_ids:
        try:
            # Solo obtener campos básicos para evitar error si hay campos nuevos no migrados
            sucursal = Sucursal.objects.only('id', 'alias', 'direccion', 'empresa_id').select_related('empresa').get(id=suc_id)
            
            # Unidades despachadas (recibidas en esta sucursal)
            despachado = Movimientos_Producto.objects.filter(
                fecha__year=anio,
                sucursal_destino_id=suc_id,
                concepto='TRASPASO_ENTRADA',
                estado='COMPLETADO'
            ).aggregate(total=Sum('cantidad'))['total'] or 0
            
            # Ventas de esta sucursal
            tickets = Ticket.objects.filter(
                sucursal_id=suc_id,
                created_at__year=anio,
                estado='PAGADO'
            )
            
            ventas_monto = tickets.aggregate(total=Sum('total'))['total'] or 0
            
            # Unidades vendidas
            vendido = Ticket_Productos.objects.filter(
                ticket__in=tickets
            ).aggregate(total=Sum('cantidad'))['total'] or 0
            
            # Costo estimado (basado en promedio de compras)
            costo_estimado = despachado * 10000  # Placeholder - se debería calcular con FIFO
            
            resultado.append({
                'sucursal_id': suc_id,
                'sucursal': sucursal.alias or sucursal.nombre or 'Sin nombre',
                'empresa': sucursal.empresa.nombre if sucursal.empresa else '-',
                'despachado': int(despachado),
                'vendido': int(vendido),
                'ventas_monto': float(ventas_monto),
                'costo': float(costo_estimado)
            })
        except Sucursal.DoesNotExist:
            continue
    
    # Ordenar por ventas
    resultado.sort(key=lambda x: x['ventas_monto'], reverse=True)
    
    return resultado[:15]


def calcular_flujo_distribucion_mensual(anio):
    """
    Calcula el flujo mensual: Compras → Despachos → Ventas
    Para visualizar la cadena de suministro
    """
    from .models import Movimientos_Producto, Ticket
    
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    resultado = []
    
    for mes in range(1, 13):
        # Compras del mes (inversión)
        compras_mes = Compras.objects.filter(fecha__year=anio, fecha__month=mes)
        compras_ids = list(compras_mes.values_list('id', flat=True))
        
        productos_mes = Compras_Producto.objects.filter(compras__in=compras_ids)
        inversion = productos_mes.aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # Despachos del mes
        despachos_mes = Movimientos_Producto.objects.filter(
            fecha__year=anio,
            fecha__month=mes,
            concepto='TRASPASO_SALIDA',
            estado='COMPLETADO'
        ).aggregate(total=Sum(F('cantidad') * F('costo')))['total'] or 0
        
        # Ventas del mes
        try:
            ventas_mes = Ticket.objects.filter(
                created_at__year=anio,
                created_at__month=mes,
                estado='PAGADO'
            ).aggregate(total=Sum('total'))['total'] or 0
        except:
            ventas_mes = 0
        
        resultado.append({
            'mes': mes,
            'mes_nombre': meses_nombres[mes - 1],
            'inversion': float(inversion),
            'despachos': abs(float(despachos_mes)),  # abs porque egresos son negativos
            'ventas': float(ventas_mes)
        })
    
    return resultado


# ========== FUNCIONES DE ANÁLISIS DE MÁRGENES CENTRO DE DISTRIBUCIÓN ==========

def calcular_margenes_centro_distribucion(anio, sucursal_cd_id=None):
    """
    Calcula los márgenes que aplica el Centro de Distribución (EDEL, GILD) 
    al despachar productos a sucursales vendedoras.
    
    El costo para las sucursales vendedoras = Costo proveedor + Sobreprecio CD
    
    Retorna:
    - Margen bruto del CD (sobreprecio total)
    - Margen % promedio aplicado
    - Desglose por sucursal destino
    - Comparativa costo proveedor vs costo sucursal
    """
    from .models import Movimientos_Producto, Traspaso, Traspaso_Detalle, Sucursal
    
    # Identificar sucursales que son Centros de Distribución
    if sucursal_cd_id:
        sucursales_cd = Sucursal.objects.filter(id=sucursal_cd_id)
    else:
        # Intentar buscar por los nuevos campos (pueden no existir si no se aplicó migración)
        try:
            sucursales_cd = Sucursal.objects.filter(
                models.Q(es_centro_distribucion=True) | 
                models.Q(tipo_sucursal='CENTRO_DISTRIBUCION')
            )
            if not sucursales_cd.exists():
                raise Exception("No hay CD marcados")
        except:
            # Fallback: identificar por empresa proveedora
            sucursales_cd = Sucursal.objects.filter(empresa__esProveedor=True)
    
    sucursales_cd_ids = list(sucursales_cd.values_list('id', flat=True))
    
    # 1. Desde Traspasos (tienen sobreprecio, costo, costo_destino)
    traspasos_detalles = Traspaso_Detalle.objects.filter(
        traspaso__fecha_solicitud__year=anio,
        traspaso__sucursal_origen__in=sucursales_cd_ids,
        traspaso__estado__in=['EN_TRANSITO', 'RECIBIDO']
    ).select_related('traspaso__sucursal_destino', 'traspaso__sucursal_origen')
    
    margen_total_cd = 0
    costo_proveedor_total = 0
    unidades_total = 0
    detalle_por_sucursal = {}
    
    for detalle in traspasos_detalles:
        cantidad = detalle.cantidad_enviada or 0
        costo_unitario = detalle.costo or 0
        sobreprecio = detalle.sobreprecio if hasattr(detalle, 'sobreprecio') and detalle.sobreprecio else 0
        costo_destino = detalle.costo_destino if hasattr(detalle, 'costo_destino') and detalle.costo_destino else (costo_unitario + sobreprecio)
        
        # Calcular margen
        margen_cd = sobreprecio * cantidad
        costo_prov = costo_unitario * cantidad
        
        margen_total_cd += margen_cd
        costo_proveedor_total += costo_prov
        unidades_total += cantidad
        
        # Agrupar por sucursal destino
        suc_destino = detalle.traspaso.sucursal_destino
        suc_key = suc_destino.id
        
        if suc_key not in detalle_por_sucursal:
            detalle_por_sucursal[suc_key] = {
                'sucursal_id': suc_destino.id,
                'sucursal': suc_destino.alias,
                'empresa': suc_destino.empresa.nombre if suc_destino.empresa else '-',
                'unidades': 0,
                'costo_proveedor_total': 0,
                'sobreprecio_total': 0,
                'costo_destino_total': 0
            }
        
        detalle_por_sucursal[suc_key]['unidades'] += cantidad
        detalle_por_sucursal[suc_key]['costo_proveedor_total'] += costo_prov
        detalle_por_sucursal[suc_key]['sobreprecio_total'] += margen_cd
        detalle_por_sucursal[suc_key]['costo_destino_total'] += costo_destino * cantidad
    
    # 2. Si no hay datos en Traspasos, intentar con Movimientos_Producto
    if unidades_total == 0:
        movimientos = Movimientos_Producto.objects.filter(
            fecha__year=anio,
            concepto='TRASPASO_SALIDA',
            sucursal_origen__in=sucursales_cd_ids,
            estado='COMPLETADO'
        ).select_related('sucursal_destino', 'sucursal_origen')
        
        for mov in movimientos:
            cantidad = abs(mov.cantidad or 0)
            costo_unitario = mov.costo or 0
            sobreprecio = mov.sobreprecio or 0
            precio = mov.precio or 0
            
            margen_cd = sobreprecio * cantidad
            costo_prov = costo_unitario * cantidad
            
            margen_total_cd += margen_cd
            costo_proveedor_total += costo_prov
            unidades_total += cantidad
            
            if mov.sucursal_destino:
                suc_key = mov.sucursal_destino.id
                
                if suc_key not in detalle_por_sucursal:
                    detalle_por_sucursal[suc_key] = {
                        'sucursal_id': mov.sucursal_destino.id,
                        'sucursal': mov.sucursal_destino.alias,
                        'empresa': mov.sucursal_destino.empresa.nombre if mov.sucursal_destino.empresa else '-',
                        'unidades': 0,
                        'costo_proveedor_total': 0,
                        'sobreprecio_total': 0,
                        'costo_destino_total': 0
                    }
                
                detalle_por_sucursal[suc_key]['unidades'] += cantidad
                detalle_por_sucursal[suc_key]['costo_proveedor_total'] += costo_prov
                detalle_por_sucursal[suc_key]['sobreprecio_total'] += margen_cd
                detalle_por_sucursal[suc_key]['costo_destino_total'] += (costo_unitario + sobreprecio) * cantidad
    
    # Calcular margen promedio %
    margen_promedio_pct = 0
    if costo_proveedor_total > 0:
        margen_promedio_pct = round((margen_total_cd / costo_proveedor_total) * 100, 2)
    
    # Ordenar detalle por unidades
    detalle_lista = sorted(detalle_por_sucursal.values(), key=lambda x: x['unidades'], reverse=True)
    
    # Calcular margen % para cada sucursal
    for item in detalle_lista:
        if item['costo_proveedor_total'] > 0:
            item['margen_pct'] = round((item['sobreprecio_total'] / item['costo_proveedor_total']) * 100, 2)
        else:
            item['margen_pct'] = 0
    
    return {
        'centros_distribucion': [{'id': s.id, 'alias': s.alias} for s in sucursales_cd],
        'margen_total_cd': float(margen_total_cd),
        'costo_proveedor_total': float(costo_proveedor_total),
        'costo_destino_total': float(costo_proveedor_total + margen_total_cd),
        'margen_promedio_pct': margen_promedio_pct,
        'unidades_despachadas': int(unidades_total),
        'detalle_por_sucursal': detalle_lista[:10]
    }


def calcular_comparativa_costos_cd_vs_sucursales(anio):
    """
    Compara el costo de productos según origen:
    - Costo proveedor externo (lo que paga EDEL/GILD)
    - Costo interno (lo que pagan las sucursales vendedoras a EDEL/GILD)
    
    Muestra el incremento de costo por pasar por el CD.
    """
    from .models import Movimientos_Producto, Sucursal, LoteProducto
    
    # Sucursales CD (con fallback si no existe el campo nuevo)
    try:
        sucursales_cd = Sucursal.objects.filter(
            models.Q(es_centro_distribucion=True) | 
            models.Q(tipo_sucursal='CENTRO_DISTRIBUCION') |
            models.Q(empresa__esProveedor=True)
        )
    except:
        sucursales_cd = Sucursal.objects.filter(empresa__esProveedor=True)
    sucursales_cd_ids = list(sucursales_cd.values_list('id', flat=True))
    
    # Sucursales vendedoras (no son CD)
    sucursales_vendedoras = Sucursal.objects.exclude(id__in=sucursales_cd_ids)
    
    comparativa = []
    
    for suc_vendedora in sucursales_vendedoras[:10]:
        # Lotes recibidos en la sucursal vendedora
        lotes = LoteProducto.objects.filter(
            sucursal=suc_vendedora,
            fecha_ingreso__year=anio
        )
        
        total_costo = 0
        total_sobreprecio = 0
        total_unidades = 0
        
        for lote in lotes:
            total_costo += lote.costo_unitario * lote.cantidad_inicial
            total_sobreprecio += (lote.sobreprecio_unitario or 0) * lote.cantidad_inicial
            total_unidades += lote.cantidad_inicial
        
        costo_promedio = 0
        sobreprecio_promedio = 0
        
        if total_unidades > 0:
            costo_promedio = round(total_costo / total_unidades)
            sobreprecio_promedio = round(total_sobreprecio / total_unidades)
        
        comparativa.append({
            'sucursal_id': suc_vendedora.id,
            'sucursal': suc_vendedora.alias,
            'empresa': suc_vendedora.empresa.nombre if suc_vendedora.empresa else '-',
            'unidades_recibidas': int(total_unidades),
            'costo_promedio': costo_promedio,
            'sobreprecio_promedio': sobreprecio_promedio,
            'costo_total_promedio': costo_promedio + sobreprecio_promedio,
            'incremento_pct': round((sobreprecio_promedio / costo_promedio * 100), 2) if costo_promedio > 0 else 0
        })
    
    return sorted(comparativa, key=lambda x: x['unidades_recibidas'], reverse=True)


def calcular_rentabilidad_por_tipo_sucursal(anio):
    """
    Calcula la rentabilidad diferenciada:
    - CD: Compra a proveedor → Vende a sucursales con sobreprecio
    - Sucursales vendedoras: Compra al CD → Vende a cliente final
    
    Analiza márgenes en cada etapa de la cadena.
    """
    from .models import Sucursal, Ticket, Ticket_Productos, Movimientos_Producto
    
    resultado = {
        'centros_distribucion': [],
        'sucursales_vendedoras': []
    }
    
    # Sucursales CD (con fallback si no existe el campo nuevo)
    try:
        sucursales_cd = Sucursal.objects.filter(
            models.Q(es_centro_distribucion=True) | 
            models.Q(tipo_sucursal='CENTRO_DISTRIBUCION') |
            models.Q(empresa__esProveedor=True)
        )
    except:
        sucursales_cd = Sucursal.objects.filter(empresa__esProveedor=True)
    
    for suc_cd in sucursales_cd:
        # Inversión en compras a proveedores
        compras_cd = Compras.objects.filter(
            responsable__sucursales=suc_cd,
            fecha__year=anio
        )
        compras_ids = list(compras_cd.values_list('id', flat=True))
        
        inversion = Compras_Producto.objects.filter(
            compras__in=compras_ids
        ).aggregate(
            total=Sum(F('costo') * F('compras_producto_talla__stock'))
        )['total'] or 0
        
        # Ingresos por despachos (sobreprecio)
        despachos = Movimientos_Producto.objects.filter(
            fecha__year=anio,
            sucursal_origen=suc_cd,
            concepto='TRASPASO_SALIDA',
            estado='COMPLETADO'
        ).aggregate(
            total_sobreprecio=Sum(F('sobreprecio') * Abs(F('cantidad'))),
            total_costo=Sum(F('costo') * Abs(F('cantidad')))
        )
        
        sobreprecio_generado = despachos['total_sobreprecio'] or 0
        costo_despachado = despachos['total_costo'] or 0
        
        rentabilidad_cd = 0
        if costo_despachado > 0:
            rentabilidad_cd = round((sobreprecio_generado / costo_despachado) * 100, 2)
        
        resultado['centros_distribucion'].append({
            'sucursal_id': suc_cd.id,
            'sucursal': suc_cd.alias,
            'empresa': suc_cd.empresa.nombre if suc_cd.empresa else '-',
            'inversion_proveedores': float(inversion),
            'costo_despachado': float(costo_despachado),
            'sobreprecio_generado': float(sobreprecio_generado),
            'rentabilidad_pct': rentabilidad_cd
        })
    
    # Sucursales vendedoras
    sucursales_cd_ids = list(sucursales_cd.values_list('id', flat=True))
    sucursales_vendedoras = Sucursal.objects.exclude(id__in=sucursales_cd_ids)[:10]
    
    for suc_vend in sucursales_vendedoras:
        # Ventas realizadas
        try:
            ventas = Ticket.objects.filter(
                sucursal=suc_vend,
                created_at__year=anio,
                estado='PAGADO'
            )
            
            total_ventas = ventas.aggregate(total=Sum('total'))['total'] or 0
            
            # Costo de lo vendido (incluyendo sobreprecio del CD)
            costo_ventas = Ticket_Productos.objects.filter(
                ticket__in=ventas
            ).aggregate(
                total=Sum(F('costo') * F('cantidad'))
            )['total'] or 0
            
            margen_bruto = total_ventas - costo_ventas
            rentabilidad_vend = 0
            if total_ventas > 0:
                rentabilidad_vend = round((margen_bruto / total_ventas) * 100, 2)
            
            resultado['sucursales_vendedoras'].append({
                'sucursal_id': suc_vend.id,
                'sucursal': suc_vend.alias,
                'empresa': suc_vend.empresa.nombre if suc_vend.empresa else '-',
                'ventas_total': float(total_ventas),
                'costo_ventas': float(costo_ventas),
                'margen_bruto': float(margen_bruto),
                'rentabilidad_pct': rentabilidad_vend
            })
        except Exception as e:
            continue

    return resultado


# ========== COMPENSACIÓN FACTURA-CONTRA-FACTURA ("Pagar con Factura") ==========
#
# Cuando ya no se puede emitir/asociar una Nota de Crédito a una factura de compra
# antigua, el proveedor puede pedir saldarla asociándole OTRA factura del mismo
# proveedor como instrumento de pago. Esto es una compensación de tesorería / neteo
# de cuentas por pagar, NO una relación tributaria SII: ambas facturas siguen siendo
# documentos tributarios válidos. Se registra en el libro de pagos (Dte_Detalle_Pago)
# con metodo_pago=METODO_COMPENSACION, igual que una Nota de Crédito.
# Espeja el patrón de obtener_ncs_disponibles / asociar_nc_existente / desasociar_nc.


@login_required
@require_GET
def obtener_facturas_compensar_disponibles(request):
    """
    Lista facturas de compra del mismo proveedor que pueden usarse como instrumento
    de compensación contra una factura objetivo. Excluye la propia factura objetivo,
    las ya usadas como instrumento de compensación y las que no tienen saldo propio
    disponible. Espeja obtener_ncs_disponibles (views.py).
    """
    try:
        empresa_id = request.session.get('idEmpresaActual')
        if not empresa_id:
            return JsonResponse({'success': False, 'error': 'Empresa no identificada en sesión'}, status=403)

        dte_id = request.GET.get('dte_id')
        if not dte_id:
            return JsonResponse({'success': False, 'error': 'ID de factura objetivo requerido'}, status=400)

        try:
            objetivo = Dte.objects.get(id=dte_id)
        except Dte.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Factura objetivo no encontrada'}, status=404)

        busqueda = (request.GET.get('busqueda', '') or '').strip()
        try:
            limit = int(request.GET.get('limit', 200))
        except (TypeError, ValueError):
            limit = 200

        candidatos = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            receptor_id=empresa_id,
            tipo_documento='FACTURA ELECTRONICA',
            emisor=objetivo.emisor,           # mismo proveedor
        ).exclude(id=objetivo.id)

        if busqueda:
            candidatos = candidatos.filter(numero_documento__icontains=busqueda)

        # Excluir facturas ya usadas como instrumento de compensación (mismo idiom que NC).
        numeros_usados = set(
            Dte_Detalle_Pago.objects.filter(
                metodo_pago=METODO_COMPENSACION,
                dte__receptor_id=empresa_id,
            ).exclude(voucher__isnull=True)
             .values_list('voucher', flat=True)
        )
        if numeros_usados:
            candidatos = candidatos.exclude(numero_documento__in=numeros_usados)

        candidatos = list(
            candidatos.select_related('emisor').order_by('-fecha_emision')[:limit]
        )

        # Saldo propio disponible de cada candidato (monto - sus propios pagos),
        # en una sola consulta. Sólo se ofrecen facturas con saldo > 0.
        ids = [c.id for c in candidatos]
        pagos_map = dict(
            Dte_Detalle_Pago.objects.filter(dte_id__in=ids)
                .values_list('dte_id')
                .annotate(total=Sum('monto'))
                .values_list('dte_id', 'total')
        )

        resultado = []
        for c in candidatos:
            pagado = float(pagos_map.get(c.id, 0) or 0)
            saldo_disponible = float(c.monto_con_iva or 0) - pagado
            if saldo_disponible <= 0:
                continue
            resultado.append({
                'id': c.id,
                'numero_documento': c.numero_documento,
                'proveedor': c.emisor.nombre if c.emisor else 'N/A',
                'fecha_emision': c.fecha_emision.strftime('%Y-%m-%d') if c.fecha_emision else '',
                'monto_con_iva': float(c.monto_con_iva or 0),
                'saldo_disponible': saldo_disponible,
                'estado': c.estado_dte,
            })

        return JsonResponse({'success': True, 'facturas': resultado})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def asociar_factura_compensacion(request):
    """
    Asocia una factura existente como instrumento de compensación (pago no en
    efectivo) sobre una factura objetivo. Espeja asociar_nc_existente, pero el
    instrumento es una factura del mismo proveedor y el monto es editable/parcial.

    NOTA TRIBUTARIA: compensación de tesorería / neteo de cuentas por pagar, NO una
    relación tributaria SII. Ambas facturas siguen siendo documentos tributarios
    válidos; sólo se registra en el libro de pagos (Dte_Detalle_Pago).
    """
    try:
        data = json.loads(request.body)
        dte_id = data.get('dte_id')
        instrumento_id = data.get('factura_compensadora_id')
        monto_pedido = data.get('monto')

        if not dte_id or not instrumento_id:
            return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)

        try:
            objetivo = Dte.objects.get(id=dte_id)
            instrumento = Dte.objects.get(
                id=instrumento_id,
                tipo_documento='FACTURA ELECTRONICA',
                tipo_transaccion='COMPRA',
            )
        except Dte.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Documento no encontrado'}, status=404)

        # Guard: no auto-compensar
        if instrumento.id == objetivo.id:
            return JsonResponse({'success': False, 'error': 'Una factura no puede compensarse a sí misma'}, status=400)

        # Guard: mismo proveedor
        if instrumento.emisor_id != objetivo.emisor_id:
            return JsonResponse({'success': False, 'error': 'La factura de compensación debe ser del mismo proveedor'}, status=400)

        # Guard: incidencias pendientes en la factura objetivo (espeja registrarPagoDTE)
        if Dte_Incidencia.objects.filter(dte=objetivo, estado__in=['PENDIENTE', 'EN_GESTION']).exists():
            return JsonResponse({'success': False, 'error': 'No se puede compensar mientras existan incidencias pendientes o en gestión para esta factura.'}, status=400)

        # Guard: el instrumento no debe estar ya usado como compensación (espeja NC ya_asociada)
        if Dte_Detalle_Pago.objects.filter(
            metodo_pago=METODO_COMPENSACION,
            voucher=str(instrumento.numero_documento),
        ).exists():
            return JsonResponse({'success': False, 'error': 'Esta factura ya fue usada como compensación en otro documento'}, status=400)

        with transaction.atomic():
            # Saldos reales (monto - pagos) de ambas facturas
            pagos_objetivo = Dte_Detalle_Pago.objects.filter(dte=objetivo).aggregate(total=Sum('monto'))['total'] or 0
            saldo_objetivo = float(objetivo.monto_con_iva or 0) - float(pagos_objetivo)

            pagos_instrumento = Dte_Detalle_Pago.objects.filter(dte=instrumento).aggregate(total=Sum('monto'))['total'] or 0
            saldo_instrumento = float(instrumento.monto_con_iva or 0) - float(pagos_instrumento)

            if saldo_objetivo <= 0:
                return JsonResponse({'success': False, 'error': 'La factura objetivo no tiene saldo pendiente'}, status=400)
            if saldo_instrumento <= 0:
                return JsonResponse({'success': False, 'error': 'La factura de compensación no tiene saldo disponible'}, status=400)

            # Monto a aplicar: editable, acotado por el saldo del objetivo y del instrumento.
            try:
                monto_default = min(saldo_objetivo, saldo_instrumento)
                monto_aplicado = float(monto_pedido) if monto_pedido not in (None, '') else monto_default
            except (TypeError, ValueError):
                return JsonResponse({'success': False, 'error': 'Monto inválido'}, status=400)

            # monto es IntegerField -> redondear; nunca exceder ninguno de los dos saldos.
            monto_aplicado = int(round(min(monto_aplicado, saldo_objetivo, saldo_instrumento)))
            if monto_aplicado <= 0:
                return JsonResponse({'success': False, 'error': 'El monto a compensar debe ser mayor a cero'}, status=400)

            Dte_Detalle_Pago.objects.create(
                dte=objetivo,
                metodo_pago=METODO_COMPENSACION,
                voucher=str(instrumento.numero_documento),
                monto=monto_aplicado,
                notas=f'Compensación con Factura #{instrumento.numero_documento} - {instrumento.emisor.nombre if instrumento.emisor else "N/A"}',
                fecha_pago=timezone.localdate(),
            )

            _recalcular_estado_pago(objetivo)

        return JsonResponse({
            'success': True,
            'message': 'Factura asociada como compensación correctamente',
            'monto_aplicado': monto_aplicado,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def desasociar_factura_compensacion(request, pago_id):
    """
    Revierte una compensación: elimina la fila de Dte_Detalle_Pago y recalcula el
    estado de pago de la factura objetivo. Espeja desasociar_nc, pero por id de fila
    (más preciso que el lookup por voucher).
    """
    try:
        try:
            pago = Dte_Detalle_Pago.objects.select_related('dte').get(
                id=pago_id, metodo_pago=METODO_COMPENSACION
            )
        except Dte_Detalle_Pago.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Compensación no encontrada'}, status=404)

        objetivo = pago.dte

        if Dte_Incidencia.objects.filter(dte=objetivo, estado__in=['PENDIENTE', 'EN_GESTION']).exists():
            return JsonResponse({'success': False, 'error': 'No se puede modificar la compensación mientras existan incidencias pendientes o en gestión.'}, status=400)

        with transaction.atomic():
            pago.delete()
            _recalcular_estado_pago(objetivo)

        return JsonResponse({'success': True, 'message': 'Compensación revertida correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_GET
def obtener_info_compensacion(request, dte_id):
    """
    Devuelve las compensaciones (facturas usadas como pago) registradas sobre una
    factura, para el panel 'Ver compensaciones'.
    """
    try:
        compensaciones = list(
            Dte_Detalle_Pago.objects.filter(
                dte_id=dte_id, metodo_pago=METODO_COMPENSACION
            ).values('id', 'voucher', 'monto', 'notas', 'fecha_pago').order_by('id')
        )
        for c in compensaciones:
            if c.get('fecha_pago'):
                c['fecha_pago'] = c['fecha_pago'].isoformat()
        return JsonResponse({'success': True, 'compensaciones': compensaciones})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _recalcular_estado_pago(dte):
    """
    Recalcula y guarda estado_pago de una factura según sus pagos (efectivo + NC +
    compensaciones). Usa los mismos valores que el flujo de pagos existente:
    'PAGADO' / 'Parcial' / 'Pendiente' (pagoBadge() en el front es case-insensitive).
    Tolerancia de 1 peso por el redondeo de Dte_Detalle_Pago.monto (IntegerField).
    """
    total_pagos = Dte_Detalle_Pago.objects.filter(dte=dte).aggregate(total=Sum('monto'))['total'] or 0
    monto_total = float(dte.monto_con_iva or 0)
    if total_pagos >= monto_total - 1:
        dte.estado_pago = 'PAGADO'
    elif total_pagos > 0:
        dte.estado_pago = 'Parcial'
    else:
        dte.estado_pago = 'Pendiente'
    dte.save(update_fields=['estado_pago'])
