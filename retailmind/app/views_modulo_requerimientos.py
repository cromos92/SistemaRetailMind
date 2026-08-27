"""
Módulo de Requerimientos - RetailMind
Gestión completa de requerimientos de garantías, devoluciones y reclamos
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, Http404, HttpResponseBadRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import (
    Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg,
    Case, When, Value, IntegerField,
)
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
from django.core.mail import send_mail, EmailMessage, EmailMultiAlternatives, get_connection
from django.core.validators import validate_email
from django.conf import settings
from django.template.loader import render_to_string
import json
import os
import re
import logging
from decimal import Decimal
from datetime import datetime, timedelta

from .models import (
    Producto, Producto_Talla, Sucursal, EmpresaUser, Empresa,
    Requerimiento, FotoRequerimiento, HistorialRequerimiento,
    TipoFotoRequerimiento, MAX_FOTOS_POR_TIPO, EnvioCorreo,
    ESTADO_REQUERIMIENTO_CHOICES, TIPO_REQUERIMIENTO_CHOICES,
    ORIGEN_REQUERIMIENTO_CHOICES, ETAPA_POR_ESTADO, ESTADOS_CERRADOS,
    Ticket, Dte, Dte_Productos, Movimientos_Producto, LoteProducto,
    DocumentoCompraLegacy,
)
from .services.pdf_requerimiento_proveedor import (
    generar_pdf_requerimiento, nombre_archivo_pdf,
)
from .services.correo_service import enviar_correo_trazado, CorreoError

logger = logging.getLogger('app')

# Tope de adjuntos del correo al proveedor. El PDF del formato ya lleva las
# fotos incrustadas y reescaladas; las originales van ADEMÁS solo si caben.
# Sin esto, 8 fotos de celular (~35MB) hacen fallar el envío entero: Gmail y
# MailerSend cortan cerca de los 25MB.
MAX_ADJUNTOS_MB = int(os.environ.get('REQUERIMIENTOS_MAX_ADJUNTOS_MB', '12'))
PLAZO_RESPUESTA_DIAS = int(os.environ.get('REQUERIMIENTOS_PLAZO_RESPUESTA_DIAS', '7'))
# Timeout (segundos) de cada operación SMTP del envío al proveedor. Sin él,
# un servidor de correo colgado deja la request esperando sin límite y el
# usuario mirando el spinner.
EMAIL_TIMEOUT_SEGUNDOS = int(os.environ.get('REQUERIMIENTOS_EMAIL_TIMEOUT', '30'))


# ========== SISTEMA DE PERMISOS ==========

def obtener_rol_usuario(user):
    """Obtiene el rol del usuario"""
    if hasattr(user, 'rol'):
        return user.rol

    return 'vendedor'  # Por defecto


# Acciones que un jefe de local puede ejercer sobre los requerimientos de SU
# sucursal. 'editar' está incluido porque completar los datos que faltan
# (proveedor, factura, RUT del cliente) es justamente el trabajo de quien
# revisa, no del que creó el ticket. 'validar'/'rechazar_interno' son la
# primera aprobación del circuito: decidir si el reclamo procede.
ACCIONES_JEFE_LOCAL = frozenset({
    'ver', 'revisar', 'validar', 'rechazar_interno',
    'aprobar_simple', 'rechazar_simple', 'comentar', 'escalar', 'asignar',
    'editar', 'completar', 'cancelar',
})

# Reservadas al administrador: son las que hablan con el proveedor o cierran
# el caso con su respuesta.
ACCIONES_SOLO_ADMIN = frozenset({
    'enviar_proveedor', 'registrar_respuesta_proveedor',
})


def usuario_puede_realizar_accion(user, requerimiento, accion):
    """
    Valida si el usuario puede realizar una acción sobre el requerimiento

    Roles:
    - administrador: Puede hacer TODO
    - jefe_local (Supervisor): valida/rechaza y completa datos de SU sucursal,
      pero no le escribe al proveedor
    - cajero/vendedor: Solo puede ver y crear
    """
    rol = obtener_rol_usuario(user)

    # Administrador puede todo
    if rol == 'administrador':
        return True

    # Jefe Local (Supervisor)
    if rol == 'jefe_local':
        # Obtener sucursal del usuario
        empresa_user = EmpresaUser.objects.filter(user=user).first()
        if not empresa_user or not empresa_user.sucursal:
            return False

        # Solo puede gestionar requerimientos de su sucursal
        if requerimiento.sucursal != empresa_user.sucursal:
            return False

        if accion in ACCIONES_SOLO_ADMIN:
            return False
        return accion in ACCIONES_JEFE_LOCAL

    # Cajero/Vendedor
    if rol in ['cajero', 'vendedor']:
        # Solo puede ver sus propios requerimientos y crear nuevos
        if accion == 'crear':
            return True
        if accion == 'ver':
            return requerimiento.usuario_creador == user or requerimiento.sucursal in obtener_sucursales_usuario(user)
        if accion == 'editar' or accion == 'cancelar':
            return requerimiento.usuario_creador == user and requerimiento.estado == 'PENDIENTE'

        return False

    return False


def obtener_sucursales_usuario(user):
    """Obtiene las sucursales a las que el usuario tiene acceso"""
    return Sucursal.objects.filter(
        empresa__empresauser__user=user
    )


# El circuito real, en orden:
#   la tienda crea → alguien lo revisa → la empresa decide si procede →
#   si procede se le manda al proveedor → el proveedor decide → se cierra.
#
# APROBADO/RECHAZADO son la decisión DEL PROVEEDOR y por eso solo se llega a
# ellos desde ESPERANDO_RESPUESTA: antes se podía saltar desde EN_REVISION y
# quedaba un caso marcado "Aprobado por Proveedor" que el proveedor nunca vio.
TRANSICIONES_PERMITIDAS = {
    'PENDIENTE': ['EN_REVISION', 'VALIDADO', 'RECHAZADO_INTERNO', 'CANCELADO'],
    'EN_REVISION': ['VALIDADO', 'RECHAZADO_INTERNO', 'ESPERANDO_RESPUESTA', 'CANCELADO'],
    'VALIDADO': ['ESPERANDO_RESPUESTA', 'RECHAZADO_INTERNO', 'EN_PROCESO',
                 'COMPLETADO', 'CANCELADO'],
    # Vuelve a revisión cuando el proveedor pide más antecedentes en vez de
    # resolver: el caso sigue vivo y hay que completarlo y reenviarlo.
    'ESPERANDO_RESPUESTA': ['APROBADO', 'RECHAZADO', 'EN_REVISION'],
    'APROBADO': ['EN_PROCESO', 'COMPLETADO'],
    # Que el proveedor rechace no cierra el caso hacia el cliente: la tienda
    # todavía tiene que resolverlo (asumirlo, devolver el dinero, etc.).
    'RECHAZADO': ['EN_PROCESO', 'COMPLETADO'],
    'EN_PROCESO': ['COMPLETADO'],
    'RECHAZADO_INTERNO': ['EN_REVISION'],  # reapertura si aparecen antecedentes
    'COMPLETADO': [],
    'CANCELADO': [],
}

# Estados a los que NO se llega escribiendo el estado a mano: exigen su propia
# acción porque cada una guarda datos que el cambio de estado genérico no pide
# (motivo de la decisión, respuesta del proveedor, correo enviado…).
ESTADOS_CON_ACCION_PROPIA = {
    'ESPERANDO_RESPUESTA': 'Use la acción "Enviar a proveedor"',
    'APROBADO': 'Use la acción "Registrar respuesta del proveedor"',
    'RECHAZADO': 'Use la acción "Registrar respuesta del proveedor"',
    'VALIDADO': 'Use la acción "Validar" para dejar constancia del motivo',
    'RECHAZADO_INTERNO': 'Use la acción "Rechazar" para dejar constancia del motivo',
    'COMPLETADO': 'Use la acción "Completar" para registrar la resolución',
}


def puede_cambiar_estado(estado_actual, estado_nuevo):
    """Valida si la transición de estado es permitida"""
    if estado_actual == estado_nuevo:
        return True
    return estado_nuevo in TRANSICIONES_PERMITIDAS.get(estado_actual, [])


# ========== HELPERS DE CORREO ==========

def _correo_proveedor(empresa):
    """Primer correo configurado de la ficha del proveedor.

    La ficha de Empresa tiene 4 campos de correo; el envío histórico solo
    miraba correoVendedor y fallaba con proveedores que solo tienen `email`.
    """
    if not empresa:
        return None
    for campo in ('correoVendedor', 'email', 'correoIntercambio'):
        valor = (getattr(empresa, campo, '') or '').strip()
        if valor:
            return valor
    return None


def _correo_copia_default(user):
    """Correo de control que recibe el resumen (sin fotos) de cada envío.

    Configurable con la env var REQUERIMIENTOS_CORREO_COPIA; si no está
    definida se usa el correo del usuario que envía.
    """
    return (
        os.environ.get('REQUERIMIENTOS_CORREO_COPIA', '').strip()
        or (user.email or '').strip()
    )


# Cómo se pinta cada estado de entrega. `indicativo` marca los estados que NO
# prueban nada: una apertura puede ser Apple Mail precargando imágenes, no el
# proveedor leyendo. La interfaz tiene que decirlo, no dar por leído algo que
# no lo está.
ESTILO_ESTADO_CORREO = {
    'ENVIADO':    ('Enviado',            'secondary', 'ri-send-plane-line',   False),
    'ENTREGADO':  ('Entregado',          'success',   'ri-mail-check-line',   False),
    'ABIERTO':    ('Abierto',            'info',      'ri-eye-line',          True),
    'CLICK':      ('Abrió el enlace',    'primary',   'ri-cursor-line',       False),
    'RESPONDIDO': ('Respondió',          'success',   'ri-reply-line',        False),
    'REBOTADO':   ('REBOTÓ (no llegó)',  'danger',    'ri-mail-close-line',   False),
    'SPAM':       ('Marcado como spam',  'danger',    'ri-spam-2-line',       False),
    'FALLIDO':    ('Falló el envío',     'danger',    'ri-error-warning-line', False),
}


def _fmt(fecha):
    return timezone.localtime(fecha).strftime('%d/%m/%Y %H:%M') if fecha else ''


def _badge_correo(envio):
    """Versión compacta del estado de entrega, para el listado."""
    if not envio:
        return None
    etiqueta, clase, icono, indicativo = ESTILO_ESTADO_CORREO.get(
        envio.estado, (envio.estado, 'secondary', 'ri-mail-line', False))
    return {
        'estado': envio.estado,
        'etiqueta': etiqueta,
        'clase': clase,
        'icono': icono,
        'es_indicativo': indicativo,
        'hay_problema': envio.estado in ('REBOTADO', 'SPAM', 'FALLIDO'),
        'detalle': envio.estado_detalle or '',
        'destinatario': envio.destinatario,
    }


def _seguimiento_correo(requerimiento_id):
    """Estado de entrega del último correo enviado al proveedor.

    Devuelve None si el requerimiento es anterior a la bitácora (los casos
    viejos no tienen envío registrado y la ficha simplemente no muestra la
    sección, en vez de inventar un estado).
    """
    envio = (EnvioCorreo.objects
             .filter(modulo='REQUERIMIENTO', objeto_id=requerimiento_id,
                     es_copia_control=False)
             .order_by('-creado_en')
             .first())
    if not envio:
        return None

    etiqueta, clase, icono, indicativo = ESTILO_ESTADO_CORREO.get(
        envio.estado, (envio.estado, 'secondary', 'ri-mail-line', False))

    # La línea de tiempo solo incluye hitos que REALMENTE ocurrieron: un paso
    # sin fecha no se dibuja "en gris", se omite. Prometer un dato que no
    # tenemos es peor que no mostrarlo.
    linea_tiempo = []
    if envio.enviado_en:
        linea_tiempo.append({'titulo': 'Enviado', 'fecha': _fmt(envio.enviado_en),
                             'clase': 'secondary', 'icono': 'ri-send-plane-line',
                             'nota': envio.destinatario})
    if envio.entregado_en:
        linea_tiempo.append({'titulo': 'Entregado en el buzón',
                             'fecha': _fmt(envio.entregado_en),
                             'clase': 'success', 'icono': 'ri-mail-check-line',
                             'nota': 'Confirmado por el servidor del proveedor'})
    if envio.abierto_en:
        veces = f' ({envio.aperturas} veces)' if envio.aperturas > 1 else ''
        linea_tiempo.append({'titulo': f'Abierto{veces}', 'fecha': _fmt(envio.abierto_en),
                             'clase': 'info', 'icono': 'ri-eye-line',
                             'nota': 'Indicativo: no prueba que lo hayan leído'})
    if envio.click_en:
        linea_tiempo.append({'titulo': 'Abrió el enlace del correo',
                             'fecha': _fmt(envio.click_en),
                             'clase': 'primary', 'icono': 'ri-cursor-line',
                             'nota': 'Evidencia fuerte: alguien hizo clic'})
    if envio.estado in ('REBOTADO', 'SPAM', 'FALLIDO'):
        linea_tiempo.append({'titulo': etiqueta, 'fecha': _fmt(envio.estado_en),
                             'clase': 'danger', 'icono': icono,
                             'nota': envio.estado_detalle or envio.error[:255] or ''})

    return {
        'envio_id': envio.id,
        'estado': envio.estado,
        'etiqueta': etiqueta,
        'clase': clase,
        'icono': icono,
        'es_indicativo': indicativo,
        'hay_problema': envio.hubo_problema,
        'llego': envio.llego,
        'destinatario': envio.destinatario,
        'enviado_en': _fmt(envio.enviado_en),
        'aperturas': envio.aperturas,
        'clicks': envio.clicks,
        'detalle': envio.estado_detalle or (envio.error[:255] if envio.error else ''),
        'linea_tiempo': linea_tiempo,
        # Si nunca hubo confirmación de entrega puede ser que el proveedor no
        # reporte eventos (webhook sin configurar) o que el correo se haya
        # perdido. No es lo mismo y no hay que afirmar ninguna de las dos.
        'sin_confirmacion': envio.estado == 'ENVIADO',
    }


# ========== VISTAS PRINCIPALES ==========

@login_required
def modulo_requerimientos(request):
    """Vista principal del módulo de requerimientos"""
    # Obtener rol del usuario
    rol_usuario = obtener_rol_usuario(request.user)
    sucursales = Sucursal.objects.filter(empresa__empresauser__user=request.user)
    proveedores = Empresa.objects.filter(esProveedor=True).order_by('nombre')
    
    context = {
        'rol_usuario': rol_usuario,
        'sucursales': sucursales,
        'proveedores': proveedores,
    }
    
    return render(request, 'vistas/modulo_requerimientos/gestion_requerimientos.html', context)


@login_required
def crear_requerimiento_vista(request):
    """Vista para crear nuevo requerimiento"""
    url_base = reverse('modulo_requerimientos')
    return redirect(f"{url_base}?panel=crear")


@login_required
def detalle_requerimiento_vista(request, requerimiento_id):
    """Vista de detalle de un requerimiento"""
    requerimiento = get_object_or_404(
        Requerimiento.objects.select_related('sucursal', 'usuario_creador'),
        id=requerimiento_id,
    )

    # La página no validaba nada: cualquier usuario logueado podía abrir el
    # detalle de otra empresa. La API sí filtraba, así que la pantalla cargaba
    # vacía sin decir por qué.
    if not usuario_puede_realizar_accion(request.user, requerimiento, 'ver'):
        logger.warning('Acceso denegado al requerimiento %s por usuario %s',
                       requerimiento_id, request.user)
        return redirect(f"{reverse('modulo_requerimientos')}?sin_acceso=1")

    context = {
        'requerimiento': requerimiento,
        # Para el modal de completar datos: quien revisa asigna el proveedor
        # que la tienda no tiene cómo saber.
        'proveedores': Empresa.objects.filter(esProveedor=True).order_by('nombre'),
    }
    return render(request, 'vistas/modulo_requerimientos/detalle_requerimiento.html', context)


@login_required
def gestionar_requerimientos_vista(request):
    """Vista para gestionar requerimientos (administrador)"""
    return redirect('modulo_requerimientos')


# ========== APIs DE CREACIÓN Y GESTIÓN ==========

@login_required
@require_POST
def crear_requerimiento(request):
    """API para crear nuevo requerimiento"""
    try:
        # Obtener datos del formulario o JSON
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        # Validar datos requeridos
        campos_requeridos = ['tipo', 'sku', 'nombre_producto', 'motivo']
        for campo in campos_requeridos:
            if not data.get(campo):
                return JsonResponse({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }, status=400)

        if data.get('tipo') not in dict(TIPO_REQUERIMIENTO_CHOICES):
            return JsonResponse({
                'success': False,
                'error': f"Tipo de requerimiento no válido: {data.get('tipo')}"
            }, status=400)

        # Origen: sin cliente solo se acepta cuando la falla se detectó en el
        # stock de la tienda (merma de bodega). En un reclamo de cliente el
        # nombre sigue siendo obligatorio: es el dato con el que el proveedor
        # y la propia tienda rastrean el caso.
        origen = data.get('origen') or 'CLIENTE'
        if origen not in dict(ORIGEN_REQUERIMIENTO_CHOICES):
            origen = 'CLIENTE'
        if origen == 'CLIENTE' and not data.get('cliente_nombre'):
            return JsonResponse({
                'success': False,
                'error': 'El nombre del cliente es requerido cuando el requerimiento '
                         'nace de un reclamo. Si el producto se detectó en bodega, '
                         'marque el origen como "Detectado en stock".'
            }, status=400)

        try:
            cantidad = int(data.get('cantidad') or 1)
        except (TypeError, ValueError):
            cantidad = 1
        if cantidad < 1:
            cantidad = 1

        # Obtener sucursal actual
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No se ha seleccionado una sucursal'
            }, status=400)

        # La sucursal viene de la sesión: hay que verificar que sea del alcance
        # del usuario o un requerimiento puede quedar cargado a otra empresa.
        sucursal = Sucursal.objects.filter(
            id=sucursal_id, empresa__empresauser__user=request.user
        ).first()
        if not sucursal:
            return JsonResponse({
                'success': False,
                'error': 'La sucursal activa no pertenece a su usuario. '
                         'Vuelva a seleccionar sucursal.'
            }, status=403)

        dte_compra = None
        if data.get('dte_compra_id'):
            dte_compra = Dte.objects.filter(
                id=data.get('dte_compra_id'), tipo_transaccion='COMPRA'
            ).first()

        # Buscar producto_talla por SKU
        producto_talla = None
        try:
            # Usar filter().first() para evitar error si hay duplicados
            producto_talla = Producto_Talla.objects.filter(sku=data.get('sku')).first()
        except Exception:
            pass  # El producto puede no existir en el sistema
        
        # Crear requerimiento
        tipo_req = data.get('tipo')
        with transaction.atomic():
            requerimiento = Requerimiento.objects.create(
                tipo=tipo_req,
                subtipo=data.get('subtipo', '') or None,
                origen=origen,
                sucursal=sucursal,
                usuario_creador=request.user,
                producto_talla=producto_talla,
                sku=data.get('sku'),
                nombre_producto=data.get('nombre_producto'),
                cantidad=cantidad,
                dte_compra=dte_compra,
                numero_factura_compra=data.get('numero_factura_compra', '') or None,
                fecha_factura_compra=data.get('fecha_factura_compra') or None,
                numero_boleta=data.get('numero_boleta', ''),
                tipo_documento=data.get('tipo_documento', ''),
                fecha_compra=data.get('fecha_compra') if data.get('fecha_compra') else None,
                cliente_rut=data.get('cliente_rut', ''),
                cliente_nombre=data.get('cliente_nombre', '') or '',
                cliente_telefono=data.get('cliente_telefono', ''),
                cliente_email=data.get('cliente_email', ''),
                motivo=data.get('motivo'),
                descripcion_problema=data.get('descripcion_problema', ''),
                prioridad=data.get('prioridad', 'MEDIA'),
                proveedor_id=data.get('proveedor_id') if data.get('proveedor_id') else None,
                severidad_defecto=data.get('severidad_defecto', '') or None,
                condicion_producto=data.get('condicion_producto', '') or None,
                producto_esperado=data.get('producto_esperado', '') or None,
            )

            # Registrar en historial
            HistorialRequerimiento.objects.create(
                requerimiento=requerimiento,
                accion='CREADO',
                estado_nuevo='PENDIENTE',
                comentario='Requerimiento creado',
                usuario=request.user
            )

            # Procesar fotos por tipo (foto_FOTO_GENERAL, foto_FOTO_DEFECTO, etc.)
            max_fotos = MAX_FOTOS_POR_TIPO.get(tipo_req, 5)
            orden_counter = 1
            if request.FILES:
                # Fotos con tipo definido
                tipos_foto_db = {
                    tf.codigo: tf for tf in TipoFotoRequerimiento.objects.filter(activo=True)
                }
                # Las guiadas van PRIMERO: `request.FILES` no garantiza orden y
                # el corte por `max_fotos` estaba descartando fotos obligatorias
                # cuando el usuario adjuntaba muchas adicionales.
                claves = sorted(
                    request.FILES.keys(),
                    key=lambda k: (0 if k.startswith('foto_FOTO_') else 1, k),
                )
                for key in claves:
                    if orden_counter > max_fotos:
                        logger.info(
                            'Requerimiento %s: se omitió la foto %s por superar el '
                            'máximo de %s para el tipo %s',
                            requerimiento.numero_requerimiento, key, max_fotos, tipo_req,
                        )
                        continue
                    tipo_foto_obj = None
                    if key.startswith('foto_FOTO_'):
                        codigo = key.replace('foto_', '', 1)
                        tipo_foto_obj = tipos_foto_db.get(codigo)
                    elif key.startswith('foto_adicional_'):
                        tipo_foto_obj = tipos_foto_db.get('FOTO_ADICIONAL')
                    elif key.startswith('foto_'):
                        # Retrocompatibilidad: foto_1, foto_2, etc.
                        pass
                    else:
                        continue

                    desc_key = f'descripcion_{key}'
                    FotoRequerimiento.objects.create(
                        requerimiento=requerimiento,
                        imagen=request.FILES[key],
                        tipo_foto=tipo_foto_obj,
                        descripcion=data.get(desc_key, '') or '',
                        orden=orden_counter,
                        usuario=request.user
                    )
                    orden_counter += 1

            # Verificar completitud de fotos obligatorias
            requerimiento.fotos_completas = requerimiento.verificar_fotos_completas()
            requerimiento.save(update_fields=['fotos_completas'])
        
        return JsonResponse({
            'success': True,
            'message': 'Requerimiento creado exitosamente',
            'requerimiento_id': requerimiento.id,
            'numero_requerimiento': requerimiento.numero_requerimiento
        })

    except Exception as e:
        logger.exception('Error al crear requerimiento (usuario %s)', request.user)
        return JsonResponse({
            'success': False,
            'error': f'Error al crear requerimiento: {str(e)}'
        }, status=500)


@login_required
def listar_requerimientos(request):
    """Listar requerimientos con filtros según rol del usuario"""
    try:
        # Parámetros de filtro
        estado = request.GET.get('estado')
        tipo = request.GET.get('tipo')
        prioridad = request.GET.get('prioridad')
        sucursal_id = request.GET.get('sucursal_id')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        busqueda = request.GET.get('busqueda', '')
        urgencia = request.GET.get('urgencia')  # Nueva: filtro por urgencia
        sin_respuesta = request.GET.get('sin_respuesta')  # Nueva: > 7 días sin respuesta
        
        # Parámetros de paginación
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        etapa = request.GET.get('etapa')  # EMPRESA / PROVEEDOR / RESOLUCION / CERRADO
        incompletos = request.GET.get('incompletos')  # sin proveedor o sin factura

        # Query base
        requerimientos = Requerimiento.objects.select_related(
            'sucursal', 'usuario_creador', 'proveedor', 'producto_talla', 'asignado_a'
        ).prefetch_related('fotos')
        
        # Filtrar según rol del usuario
        rol_usuario = obtener_rol_usuario(request.user)
        
        if rol_usuario == 'administrador':
            # Administrador ve TODO
            pass
        elif rol_usuario == 'jefe_local':
            # Supervisor solo ve su sucursal
            empresa_user = EmpresaUser.objects.filter(user=request.user).first()
            if empresa_user and empresa_user.sucursal:
                requerimientos = requerimientos.filter(sucursal=empresa_user.sucursal)
            else:
                requerimientos = requerimientos.none()
        else:
            # Cajero/Vendedor solo ve sus requerimientos y los de su sucursal
            sucursales_usuario = obtener_sucursales_usuario(request.user)
            requerimientos = requerimientos.filter(
                Q(usuario_creador=request.user) | Q(sucursal__in=sucursales_usuario)
            )
        
        # Aplicar filtros
        if estado:
            requerimientos = requerimientos.filter(estado=estado)
        if tipo:
            requerimientos = requerimientos.filter(tipo=tipo)
        if prioridad:
            requerimientos = requerimientos.filter(prioridad=prioridad)
        if sucursal_id:
            requerimientos = requerimientos.filter(sucursal_id=sucursal_id)
        if fecha_inicio:
            try:
                dt_inicio = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
                requerimientos = requerimientos.filter(fecha_creacion__gte=dt_inicio)
            except (ValueError, TypeError):
                pass
        if fecha_fin:
            try:
                dt_fin = timezone.make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
                requerimientos = requerimientos.filter(fecha_creacion__lte=dt_fin)
            except (ValueError, TypeError):
                pass
        if busqueda:
            requerimientos = requerimientos.filter(
                Q(numero_requerimiento__icontains=busqueda) |
                Q(sku__icontains=busqueda) |
                Q(cliente_nombre__icontains=busqueda) |
                Q(cliente_rut__icontains=busqueda) |
                Q(numero_boleta__icontains=busqueda)
            )
        
        # Correo que NO llegó. Es el filtro más urgente del módulo: un caso
        # "esperando respuesta" cuyo correo rebotó no está esperando nada, y
        # hasta ahora se veía idéntico a uno que sí llegó.
        correo_filtro = (request.GET.get('correo') or '').strip()
        if correo_filtro == 'problema':
            ids_problema = (EnvioCorreo.objects
                            .filter(modulo='REQUERIMIENTO', es_copia_control=False,
                                    estado__in=('REBOTADO', 'SPAM', 'FALLIDO'))
                            .values_list('objeto_id', flat=True))
            requerimientos = requerimientos.filter(id__in=list(ids_problema))
        elif correo_filtro == 'sin_confirmar':
            # Salió, pero nadie confirmó que llegara.
            ids_confirmados = (EnvioCorreo.objects
                               .filter(modulo='REQUERIMIENTO', es_copia_control=False,
                                       estado__in=('ENTREGADO', 'ABIERTO', 'CLICK',
                                                   'RESPONDIDO'))
                               .values_list('objeto_id', flat=True))
            requerimientos = requerimientos.filter(
                correo_enviado_proveedor=True
            ).exclude(id__in=list(ids_confirmados))

        # Etapa del circuito: "quién tiene la pelota" es la pregunta real del
        # analista, y no se contestaba sin conocer los 10 estados de memoria.
        if etapa:
            estados_etapa = [e for e, et in ETAPA_POR_ESTADO.items() if et == etapa]
            if estados_etapa:
                requerimientos = requerimientos.filter(estado__in=estados_etapa)

        # Bandeja de "les falta algo para poder salir": sin proveedor asignado
        # o sin la factura de compra que el proveedor exige.
        if incompletos == 'true':
            requerimientos = requerimientos.filter(
                Q(proveedor__isnull=True) |
                (Q(numero_factura_compra__isnull=True) | Q(numero_factura_compra=''))
                & Q(dte_compra__isnull=True)
            ).exclude(estado__in=ESTADOS_CERRADOS)

        # Filtros especiales de seguimiento
        if sin_respuesta == 'true':
            # Requerimientos esperando proveedor sin respuesta > 7 días
            # Mismo plazo que el aviso y las estadisticas: con la env var en
            # otro valor, el boton "Ver cuales" traia un conjunto distinto al
            # que el propio aviso acababa de contar.
            fecha_limite = timezone.now() - timedelta(days=PLAZO_RESPUESTA_DIAS)
            requerimientos = requerimientos.filter(
                estado='ESPERANDO_RESPUESTA',
                fecha_envio_proveedor__lt=fecha_limite,
                fecha_respuesta_proveedor__isnull=True
            )
        
        # Paginación
        paginator = Paginator(requerimientos, page_size)
        page_obj = paginator.get_page(page)

        # Estado de entrega de toda la página en UNA consulta: pedirlo por
        # fila sería un N+1 de 25 queries por listado.
        ids_pagina = [r.id for r in page_obj]
        envio_por_req = {}
        if ids_pagina:
            for envio in (EnvioCorreo.objects
                          .filter(modulo='REQUERIMIENTO', objeto_id__in=ids_pagina,
                                  es_copia_control=False)
                          .only('objeto_id', 'estado', 'estado_detalle',
                                'destinatario', 'creado_en')
                          .order_by('objeto_id', '-creado_en')):
                # El primero de cada grupo es el más reciente: los siguientes
                # del mismo requerimiento son reenvíos anteriores.
                envio_por_req.setdefault(envio.objeto_id, envio)

        # Serializar resultados
        requerimientos_data = []
        for req in page_obj:
            requerimientos_data.append({
                'id': req.id,
                'numero_requerimiento': req.numero_requerimiento,
                'tipo': req.get_tipo_display(),
                'tipo_codigo': req.tipo,
                'subtipo': req.subtipo or '',
                'estado': req.get_estado_display(),
                'estado_codigo': req.estado,
                'prioridad': req.get_prioridad_display(),
                'prioridad_codigo': req.prioridad,
                'sucursal': req.sucursal.alias,
                'sku': req.sku,
                'nombre_producto': req.nombre_producto,
                'cantidad': req.cantidad,
                'origen_codigo': req.origen,
                # Bandera de triage para el analista: sin factura de compra la
                # mayoría de los proveedores no cursa la garantía.
                'tiene_factura_compra': req.tiene_respaldo_compra,
                'etapa': req.etapa,
                'decision_interna': req.decision_interna or '',
                'decision_proveedor': req.decision_proveedor or '',
                # Lo que le impide salir al proveedor, calculado en el modelo
                # para que listado, detalle y modal digan exactamente lo mismo.
                'faltantes': req.faltantes_para_enviar,
                'cliente_nombre': req.cliente_nombre,
                'fecha_creacion': req.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'dias_transcurridos': req.dias_transcurridos,
                'cantidad_fotos': req.cantidad_fotos,
                'fotos_completas': req.fotos_completas,
                'max_fotos': req.max_fotos,
                'usuario_creador': req.usuario_creador.get_full_name() if req.usuario_creador else '',
                'proveedor': req.proveedor.nombre if req.proveedor else '',
                'asignado_a': req.asignado_a.get_full_name() if req.asignado_a else '',
                'correo_enviado_proveedor': req.correo_enviado_proveedor,
                # Para que la lista pueda decir A QUIÉN y CUÁNDO salió el
                # correo, no solo que salió.
                'correo_proveedor_destino': req.correo_proveedor_destino or '',
                'fecha_envio_proveedor': (
                    req.fecha_envio_proveedor.strftime('%d/%m/%Y %H:%M')
                    if req.fecha_envio_proveedor else ''),
                'dias_sin_respuesta': req.dias_sin_respuesta,
                'requiere_recordatorio': req.requiere_recordatorio,
                'nivel_urgencia': req.nivel_urgencia,
                # Badge de entrega: lo urgente de la lista es distinguir "sin
                # respuesta" de "nunca le llegó", que antes se veían igual.
                'correo_estado': _badge_correo(envio_por_req.get(req.id)),
            })
        
        return JsonResponse({
            'success': True,
            'requerimientos': requerimientos_data,
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener requerimientos: {str(e)}'
        }, status=500)


def _siguiente_paso(requerimiento, permisos):
    """Qué corresponde hacer ahora con este requerimiento, en una frase.

    Devuelve ``{titulo, detalle, accion, tono}``. La pantalla lo muestra
    arriba de todo: sin esto el usuario ve ocho botones y ninguno le dice
    cuál es el que toca.
    """
    estado = requerimiento.estado
    faltantes = requerimiento.faltantes_para_enviar

    if estado in ESTADOS_CERRADOS:
        return {
            'titulo': f'Caso cerrado — {requerimiento.get_estado_display()}',
            'detalle': requerimiento.motivo_resolucion or requerimiento.resolucion or '',
            'accion': None, 'tono': 'secondary',
        }

    if estado in ('PENDIENTE', 'EN_REVISION'):
        if faltantes:
            return {
                'titulo': f'Falta {" y ".join(faltantes)}',
                'detalle': 'Complete los datos que la tienda no puede saber antes de decidir.',
                'accion': 'editar' if permisos.get('puede_editar') else None,
                'tono': 'warning',
            }
        if permisos.get('puede_validar'):
            return {
                'titulo': 'Listo para decidir',
                'detalle': 'Están el proveedor, la factura y las fotos: valide si procede reclamarlo.',
                'accion': 'decidir', 'tono': 'primary',
            }
        return {
            'titulo': 'Esperando revisión',
            'detalle': 'Un administrador o el jefe de local debe validar el caso.',
            'accion': None, 'tono': 'info',
        }

    if estado == 'VALIDADO':
        if permisos.get('puede_enviar_proveedor'):
            return {
                'titulo': 'Validado: falta enviarlo al proveedor',
                'detalle': f'Se le enviará a {requerimiento.proveedor.nombre}.' if requerimiento.proveedor else '',
                'accion': 'enviar', 'tono': 'primary',
            }
        return {
            'titulo': 'Validado, a la espera del envío',
            'detalle': 'Solo un administrador puede enviarlo al proveedor.',
            'accion': None, 'tono': 'info',
        }

    if estado == 'ESPERANDO_RESPUESTA':
        dias = requerimiento.dias_sin_respuesta
        vencido = dias > PLAZO_RESPUESTA_DIAS
        return {
            'titulo': (f'Sin respuesta hace {dias} día(s)' if vencido
                       else f'Enviado al proveedor hace {dias} día(s)'),
            'detalle': ('Ya pasó el plazo: corresponde recordatorio o registrar lo que respondió.'
                        if vencido else 'Registre la respuesta apenas conteste.'),
            'accion': 'respuesta' if permisos.get('puede_registrar_respuesta') else None,
            'tono': 'danger' if vencido else 'info',
        }

    if estado in ('APROBADO', 'RECHAZADO', 'EN_PROCESO'):
        aprobado = estado == 'APROBADO'
        return {
            'titulo': ('El proveedor aprobó: falta resolverlo con el cliente' if aprobado
                       else 'El proveedor rechazó: falta cerrar el caso'
                       if estado == 'RECHAZADO' else 'En proceso de resolución'),
            'detalle': requerimiento.motivo_resolucion or '',
            'accion': 'completar' if permisos.get('puede_completar') else None,
            'tono': 'success' if aprobado else 'warning',
        }

    return {'titulo': requerimiento.get_estado_display(), 'detalle': '',
            'accion': None, 'tono': 'secondary'}


@login_required
def detalle_requerimiento(request, requerimiento_id):
    """Obtener detalles completos de un requerimiento"""
    try:
        requerimiento = get_object_or_404(
            Requerimiento.objects.select_related(
                'sucursal', 'usuario_creador', 'usuario_gestor', 'proveedor', 
                'producto_talla', 'asignado_a'
            ).prefetch_related('fotos', 'historial'),
            id=requerimiento_id
        )
        
        # Validar permisos de visualización
        if not usuario_puede_realizar_accion(request.user, requerimiento, 'ver'):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para ver este requerimiento'
            }, status=403)
        
        # Obtener rol del usuario actual
        rol_usuario = obtener_rol_usuario(request.user)
        
        # Serializar fotos con tipo
        fotos = []
        for foto in requerimiento.fotos.select_related('tipo_foto').all():
            fotos.append({
                'id': foto.id,
                'url': foto.imagen.url if foto.imagen else '',
                'descripcion': foto.descripcion or '',
                'orden': foto.orden,
                'fecha': foto.fecha_subida.strftime('%d/%m/%Y %H:%M'),
                'tipo_foto_codigo': foto.tipo_foto.codigo if foto.tipo_foto else None,
                'tipo_foto_nombre': foto.tipo_foto.nombre if foto.tipo_foto else 'Sin clasificar',
                'tipo_foto_icono': foto.tipo_foto.icono if foto.tipo_foto else 'ri-image-line',
            })
        
        # Serializar historial
        historial = []
        for hist in requerimiento.historial.all():
            historial.append({
                'id': hist.id,
                'accion': hist.accion,
                'estado_anterior': hist.estado_anterior,
                'estado_nuevo': hist.estado_nuevo,
                'comentario': hist.comentario or '',
                'usuario': hist.usuario.get_full_name() if hist.usuario else '',
                'fecha': hist.fecha.strftime('%d/%m/%Y %H:%M')
            })
        
        requerimiento_data = {
            'id': requerimiento.id,
            'numero_requerimiento': requerimiento.numero_requerimiento,
            'tipo': requerimiento.get_tipo_display(),
            'tipo_codigo': requerimiento.tipo,
            'estado': requerimiento.get_estado_display(),
            'estado_codigo': requerimiento.estado,
            'prioridad': requerimiento.get_prioridad_display(),
            'prioridad_codigo': requerimiento.prioridad,
            
            # Sucursal y usuarios
            'sucursal': {
                'id': requerimiento.sucursal.id,
                'nombre': requerimiento.sucursal.alias
            },
            'usuario_creador': requerimiento.usuario_creador.get_full_name() if requerimiento.usuario_creador else '',
            'usuario_gestor': requerimiento.usuario_gestor.get_full_name() if requerimiento.usuario_gestor else '',
            'asignado_a': requerimiento.asignado_a.get_full_name() if requerimiento.asignado_a else '',
            'asignado_a_id': requerimiento.asignado_a.id if requerimiento.asignado_a else None,
            
            # Producto
            'sku': requerimiento.sku,
            'nombre_producto': requerimiento.nombre_producto,
            'cantidad': requerimiento.cantidad,

            # Origen y respaldo de compra
            'origen': requerimiento.get_origen_display(),
            'origen_codigo': requerimiento.origen,
            'numero_factura_compra': requerimiento.numero_factura_compra or (
                str(requerimiento.dte_compra.numero_documento)
                if requerimiento.dte_compra_id else ''),
            'fecha_factura_compra': (
                requerimiento.fecha_factura_compra.strftime('%d/%m/%Y')
                if requerimiento.fecha_factura_compra else (
                    requerimiento.dte_compra.fecha_emision.strftime('%d/%m/%Y')
                    if requerimiento.dte_compra_id and requerimiento.dte_compra.fecha_emision
                    else '')),
            # ISO para el <input type="date"> del modal de edición
            'fecha_factura_compra_iso': (
                requerimiento.fecha_factura_compra.strftime('%Y-%m-%d')
                if requerimiento.fecha_factura_compra else ''),

            # Documento
            'tipo_documento': requerimiento.tipo_documento or '',
            'numero_boleta': requerimiento.numero_boleta or '',
            'fecha_compra': requerimiento.fecha_compra.strftime('%d/%m/%Y') if requerimiento.fecha_compra else '',
            
            # Cliente
            'cliente_nombre': requerimiento.cliente_nombre,
            'cliente_rut': requerimiento.cliente_rut or '',
            'cliente_telefono': requerimiento.cliente_telefono or '',
            'cliente_email': requerimiento.cliente_email or '',
            
            # Descripcion
            'motivo': requerimiento.motivo,
            'descripcion_problema': requerimiento.descripcion_problema or '',

            # Clasificacion de defecto
            'subtipo': requerimiento.subtipo_display,
            'subtipo_codigo': requerimiento.subtipo or '',
            'severidad_defecto': requerimiento.get_severidad_defecto_display() if requerimiento.severidad_defecto else '',
            'severidad_defecto_codigo': requerimiento.severidad_defecto or '',
            'condicion_producto': requerimiento.get_condicion_producto_display() if requerimiento.condicion_producto else '',
            'condicion_producto_codigo': requerimiento.condicion_producto or '',
            'producto_esperado': requerimiento.producto_esperado or '',
            # Las ve quien puede completar el requerimiento: si no, el modal de
            # edición las mandaría vacías y borraría lo que escribió otro.
            'notas_internas': (requerimiento.notas_internas or ''
                               if rol_usuario in ('administrador', 'jefe_local') else ''),
            'fotos_completas': requerimiento.fotos_completas,
            'max_fotos': requerimiento.max_fotos,

            # Proveedor
            'proveedor': {
                'id': requerimiento.proveedor.id if requerimiento.proveedor else None,
                'nombre': requerimiento.proveedor.nombre if requerimiento.proveedor else '',
                'correo': _correo_proveedor(requerimiento.proveedor) or '',
                'correo_administrador': (requerimiento.proveedor.correoAdministrador or '') if requerimiento.proveedor else '',
            },
            'correo_copia_default': _correo_copia_default(request.user),
            'correo_enviado_proveedor': requerimiento.correo_enviado_proveedor,
            'fecha_envio_proveedor': requerimiento.fecha_envio_proveedor.strftime('%d/%m/%Y %H:%M') if requerimiento.fecha_envio_proveedor else '',
            'correo_proveedor_destino': requerimiento.correo_proveedor_destino or '',
            'intentos_envio': requerimiento.intentos_envio,
            # Estado de entrega del último correo: si llegó, si rebotó, si lo
            # abrieron. None en los casos anteriores a la bitácora.
            'seguimiento_correo': _seguimiento_correo(requerimiento.id),
            'dias_sin_respuesta': requerimiento.dias_sin_respuesta,
            'requiere_recordatorio': requerimiento.requiere_recordatorio,
            'respuesta_proveedor': requerimiento.respuesta_proveedor or '',
            'fecha_respuesta_proveedor': requerimiento.fecha_respuesta_proveedor.strftime('%d/%m/%Y %H:%M') if requerimiento.fecha_respuesta_proveedor else '',
            'decision_proveedor': requerimiento.decision_proveedor or '',

            # Decisión interna (la de la empresa, previa al proveedor)
            'decision_interna': requerimiento.decision_interna or '',
            'motivo_decision_interna': requerimiento.motivo_decision_interna or '',
            'fecha_decision_interna': (
                requerimiento.fecha_decision_interna.strftime('%d/%m/%Y %H:%M')
                if requerimiento.fecha_decision_interna else ''),
            'usuario_decision_interna': (
                requerimiento.usuario_decision_interna.get_full_name()
                if requerimiento.usuario_decision_interna_id else ''),

            # Etapa y bloqueos: lo que la pantalla necesita para decir qué
            # falta y quién tiene la pelota, sin recalcularlo en el navegador.
            'etapa': requerimiento.etapa,
            'esta_cerrado': requerimiento.esta_cerrado,
            'faltantes': requerimiento.faltantes_para_enviar,
            'listo_para_proveedor': requerimiento.listo_para_proveedor,


            # Resolución
            'resolucion': requerimiento.resolucion or '',
            'motivo_resolucion': requerimiento.motivo_resolucion or '',
            'fecha_resolucion': requerimiento.fecha_resolucion.strftime('%d/%m/%Y %H:%M') if requerimiento.fecha_resolucion else '',
            
            # Fechas
            'fecha_creacion': requerimiento.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'fecha_actualizacion': requerimiento.fecha_actualizacion.strftime('%d/%m/%Y %H:%M'),
            'dias_transcurridos': requerimiento.dias_transcurridos,
            'nivel_urgencia': requerimiento.nivel_urgencia,
            
            # Tipos de foto requeridos para este tipo de requerimiento
            'tipos_foto_requeridos': [
                {
                    'codigo': tf.codigo,
                    'nombre': tf.nombre,
                    'descripcion_guia': tf.descripcion_guia,
                    'icono': tf.icono,
                    'es_obligatorio': tf.es_obligatorio,
                }
                for tf in TipoFotoRequerimiento.tipos_para(requerimiento.tipo)
            ],

            # Relacionados
            'fotos': fotos,
            'historial': historial,
            
            # Transiciones de estado válidas desde el estado actual (para el
            # modal Cambiar Estado — antes ofrecía los 8 estados y el backend
            # rechazaba la mayoría)
            'transiciones_permitidas': TRANSICIONES_PERMITIDAS.get(requerimiento.estado, []),
            'estados_labels': dict(ESTADO_REQUERIMIENTO_CHOICES),

            # Permisos del usuario actual
            'permisos': {
                'puede_editar': usuario_puede_realizar_accion(request.user, requerimiento, 'editar'),
                'puede_revisar': usuario_puede_realizar_accion(request.user, requerimiento, 'revisar'),
                'puede_validar': usuario_puede_realizar_accion(request.user, requerimiento, 'validar'),
                'puede_rechazar_interno': usuario_puede_realizar_accion(request.user, requerimiento, 'rechazar_interno'),
                'puede_aprobar': usuario_puede_realizar_accion(request.user, requerimiento, 'aprobar_simple'),
                'puede_rechazar': usuario_puede_realizar_accion(request.user, requerimiento, 'rechazar_simple'),
                'puede_enviar_proveedor': usuario_puede_realizar_accion(request.user, requerimiento, 'enviar_proveedor'),
                'puede_registrar_respuesta': usuario_puede_realizar_accion(request.user, requerimiento, 'registrar_respuesta_proveedor'),
                'puede_completar': usuario_puede_realizar_accion(request.user, requerimiento, 'completar'),
                'puede_cancelar': usuario_puede_realizar_accion(request.user, requerimiento, 'cancelar'),
                'puede_ver_notas': rol_usuario in ('administrador', 'jefe_local'),
            },
            'rol_usuario': rol_usuario,
        }
        # La pantalla no tiene que deducir el siguiente paso a punta de ifs:
        # el backend, que es quien conoce las transiciones y los permisos,
        # dice cuál es la acción que corresponde ahora.
        requerimiento_data['siguiente_paso'] = _siguiente_paso(
            requerimiento, requerimiento_data['permisos'])
        
        return JsonResponse({
            'success': True,
            'requerimiento': requerimiento_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener requerimiento: {str(e)}'
        }, status=500)


@login_required
@require_POST
def actualizar_estado_requerimiento(request, requerimiento_id):
    """Actualizar estado de un requerimiento con validación de permisos"""
    try:
        data = json.loads(request.body)
        
        nuevo_estado = data.get('estado')
        comentario = data.get('comentario', '')
        
        if not nuevo_estado:
            return JsonResponse({
                'success': False,
                'error': 'Nuevo estado es requerido'
            }, status=400)
        
        requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)
        estado_anterior = requerimiento.estado

        # Validar permisos según estado y rol
        rol_usuario = obtener_rol_usuario(request.user)

        # Validar que la transición sea permitida
        if not puede_cambiar_estado(estado_anterior, nuevo_estado):
            etiquetas = dict(ESTADO_REQUERIMIENTO_CHOICES)
            return JsonResponse({
                'success': False,
                'error': (f'No se puede pasar de "{etiquetas.get(estado_anterior, estado_anterior)}" '
                          f'a "{etiquetas.get(nuevo_estado, nuevo_estado)}"')
            }, status=400)

        # Hay estados que exigen su propia acción porque guardan datos que este
        # endpoint no pide (el motivo de la decisión, la respuesta del
        # proveedor, el correo enviado). Permitirlos acá dejaba casos marcados
        # "Aprobado por el proveedor" sin una sola línea de respuesta.
        if nuevo_estado in ESTADOS_CON_ACCION_PROPIA:
            return JsonResponse({
                'success': False,
                'error': ESTADOS_CON_ACCION_PROPIA[nuevo_estado],
            }, status=400)

        # Validar permisos por rol
        if rol_usuario == 'jefe_local':
            # Supervisor solo puede gestionar casos de su sucursal
            empresa_user = EmpresaUser.objects.filter(user=request.user).first()
            if not empresa_user or empresa_user.sucursal != requerimiento.sucursal:
                return JsonResponse({
                    'success': False,
                    'error': 'Solo puede gestionar requerimientos de su sucursal'
                }, status=403)

        elif rol_usuario in ['cajero', 'vendedor']:
            # Vendedores solo pueden cancelar sus propios req PENDIENTES. Sin
            # el chequeo de estado podían cancelar uno ya validado por el jefe
            # de local y tirar abajo esa decisión.
            if not (requerimiento.usuario_creador == request.user
                    and nuevo_estado == 'CANCELADO'
                    and estado_anterior == 'PENDIENTE'):
                return JsonResponse({
                    'success': False,
                    'error': 'Solo puede cancelar un requerimiento propio que siga pendiente'
                }, status=403)

        elif rol_usuario != 'administrador':
            # Cierre por defecto. Antes cualquier rol que no fuera jefe_local,
            # cajero o vendedor —incluido un rol nuevo o mal escrito— caía por
            # el hueco de los elif y cambiaba el estado de cualquier
            # requerimiento de cualquier empresa.
            return JsonResponse({
                'success': False,
                'error': f'El rol "{rol_usuario}" no tiene permisos sobre requerimientos'
            }, status=403)


        with transaction.atomic():
            requerimiento.estado = nuevo_estado
            
            # Si cambia a EN_REVISION, asignar al usuario actual
            if nuevo_estado == 'EN_REVISION' and not requerimiento.asignado_a:
                requerimiento.asignado_a = request.user
            
            requerimiento.save()
            
            # Registrar en historial
            HistorialRequerimiento.objects.create(
                requerimiento=requerimiento,
                accion='CAMBIO_ESTADO',
                estado_anterior=estado_anterior,
                estado_nuevo=nuevo_estado,
                comentario=comentario,
                usuario=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Estado actualizado exitosamente',
            'nuevo_estado': requerimiento.get_estado_display()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar estado: {str(e)}'
        }, status=500)


@login_required
@require_POST
def decidir_requerimiento(request, requerimiento_id):
    """Primera aprobación del circuito: la que toma la empresa, no el proveedor.

    La tienda levanta el caso; quien revisa decide si procede reclamárselo al
    proveedor (VALIDADO) o si se cierra acá (RECHAZADO_INTERNO). Antes esta
    decisión no existía como acción: se marcaba APROBADO —la etiqueta que dice
    "Aprobado por el proveedor"— y el caso quedaba mintiendo, además de
    perderse quién decidió y por qué.
    """
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Cuerpo inválido'}, status=400)

    decision = (data.get('decision') or '').upper()
    motivo = (data.get('motivo') or '').strip()

    if decision not in ('APROBADO', 'RECHAZADO'):
        return JsonResponse({
            'success': False,
            'error': 'La decisión debe ser APROBADO o RECHAZADO'
        }, status=400)
    if not motivo:
        return JsonResponse({
            'success': False,
            'error': 'Indique el motivo: queda en el historial y es lo que se '
                     'le explica a la tienda que levantó el caso.'
        }, status=400)

    requerimiento = get_object_or_404(
        Requerimiento.objects.select_related('proveedor', 'sucursal'),
        id=requerimiento_id,
    )

    accion_permiso = 'validar' if decision == 'APROBADO' else 'rechazar_interno'
    if not usuario_puede_realizar_accion(request.user, requerimiento, accion_permiso):
        return JsonResponse({
            'success': False,
            'error': 'No tiene permisos para decidir sobre este requerimiento'
        }, status=403)

    nuevo_estado = 'VALIDADO' if decision == 'APROBADO' else 'RECHAZADO_INTERNO'
    estado_anterior = requerimiento.estado

    if not puede_cambiar_estado(estado_anterior, nuevo_estado):
        return JsonResponse({
            'success': False,
            'error': f'El requerimiento está "{requerimiento.get_estado_display()}" '
                     f'y ya pasó la etapa de validación interna'
        }, status=400)

    # Validar sin proveedor asignado deja el caso en un callejón: el estado
    # dice "listo para el proveedor" y no hay a quién mandárselo.
    if decision == 'APROBADO' and not requerimiento.proveedor_id:
        return JsonResponse({
            'success': False,
            'error': 'Asigne el proveedor antes de validar: sin proveedor el '
                     'requerimiento no se le puede enviar a nadie.',
            'falta': 'proveedor',
        }, status=400)

    with transaction.atomic():
        requerimiento.estado = nuevo_estado
        requerimiento.decision_interna = decision
        requerimiento.motivo_decision_interna = motivo
        requerimiento.fecha_decision_interna = timezone.now()
        requerimiento.usuario_decision_interna = request.user
        if not requerimiento.asignado_a:
            requerimiento.asignado_a = request.user
        if decision == 'RECHAZADO':
            # Se cierra acá: la resolución visible es la propia decisión.
            requerimiento.fecha_resolucion = timezone.now()
            requerimiento.motivo_resolucion = motivo
        requerimiento.save()

        HistorialRequerimiento.objects.create(
            requerimiento=requerimiento,
            accion='VALIDADO_INTERNAMENTE' if decision == 'APROBADO' else 'RECHAZADO_INTERNAMENTE',
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            comentario=motivo[:2000],
            usuario=request.user,
        )

    faltantes = requerimiento.faltantes_para_enviar
    return JsonResponse({
        'success': True,
        'message': ('Requerimiento validado: ya se puede enviar al proveedor'
                    if decision == 'APROBADO'
                    else 'Requerimiento rechazado internamente'),
        'nuevo_estado': requerimiento.get_estado_display(),
        'nuevo_estado_codigo': nuevo_estado,
        # El validador se entera acá mismo de lo que le falta al caso para
        # poder salir, en vez de descubrirlo al intentar enviarlo.
        'faltantes': faltantes,
    })


CAMPOS_EDITABLES_ANALISTA = {
    # Lo que la tienda no tiene cómo saber y completa quien revisa
    'proveedor_id': 'Proveedor',
    'numero_factura_compra': 'N° factura de compra',
    'fecha_factura_compra': 'Fecha factura de compra',
    # Correcciones de lo cargado en tienda
    'cantidad': 'Cantidad',
    'prioridad': 'Prioridad',
    'subtipo': 'Subtipo',
    'severidad_defecto': 'Severidad',
    'condicion_producto': 'Condición del producto',
    'producto_esperado': 'Producto esperado',
    'motivo': 'Motivo',
    'descripcion_problema': 'Detalle del problema',
    'cliente_nombre': 'Nombre del cliente',
    'cliente_rut': 'RUT del cliente',
    'cliente_telefono': 'Teléfono del cliente',
    'cliente_email': 'Email del cliente',
    'notas_internas': 'Notas internas',
}


@login_required
@require_POST
def editar_requerimiento(request, requerimiento_id):
    """Completar/corregir un requerimiento ya creado.

    Existe porque el reclamo nace incompleto por diseño: la tienda sabe qué
    falló y tiene el producto en la mano, pero NO sabe a qué proveedor se le
    compró ni con qué factura — eso lo averigua quien revisa. Sin esta vista,
    un requerimiento al que le falta la factura queda muerto: no había forma
    de completarlo desde ninguna pantalla.

    Deja rastro en el historial de qué campos cambiaron y con qué valores.
    """
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Cuerpo inválido'}, status=400)

    requerimiento = get_object_or_404(
        Requerimiento.objects.select_related('proveedor', 'sucursal'),
        id=requerimiento_id,
    )

    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'administrador':
        pass  # puede completar cualquier requerimiento
    elif rol_usuario == 'jefe_local':
        empresa_user = EmpresaUser.objects.filter(user=request.user).first()
        if not empresa_user or empresa_user.sucursal_id != requerimiento.sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'Solo puede editar requerimientos de su sucursal'
            }, status=403)
    else:
        # El creador puede corregir lo suyo mientras nadie lo haya tomado
        if not (requerimiento.usuario_creador_id == request.user.id
                and requerimiento.estado == 'PENDIENTE'):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para editar este requerimiento'
            }, status=403)

    if requerimiento.estado in ('COMPLETADO', 'CANCELADO'):
        return JsonResponse({
            'success': False,
            'error': f'El requerimiento está {requerimiento.get_estado_display()} y ya no se edita'
        }, status=400)

    cambios = []
    with transaction.atomic():
        for campo, etiqueta in CAMPOS_EDITABLES_ANALISTA.items():
            if campo not in data:
                continue
            valor = data.get(campo)

            if campo == 'proveedor_id':
                nuevo = Empresa.objects.filter(id=valor).first() if valor else None
                if nuevo != requerimiento.proveedor:
                    anterior = requerimiento.proveedor.nombre if requerimiento.proveedor else '—'
                    requerimiento.proveedor = nuevo
                    cambios.append(f'{etiqueta}: {anterior} → {nuevo.nombre if nuevo else "—"}')
                continue

            if campo == 'cantidad':
                try:
                    valor = max(1, int(valor))
                except (TypeError, ValueError):
                    continue
            elif campo == 'fecha_factura_compra':
                valor = valor or None
            elif isinstance(valor, str):
                valor = valor.strip() or None

            anterior = getattr(requerimiento, campo)
            if str(anterior or '') == str(valor or ''):
                continue
            setattr(requerimiento, campo, valor)
            # Las notas internas no se copian al historial: pueden traer
            # información que no corresponde repetir en la bitácora visible.
            if campo == 'notas_internas':
                cambios.append(f'{etiqueta}: actualizada')
            else:
                cambios.append(f'{etiqueta}: {anterior or "—"} → {valor or "—"}')

        if not cambios:
            return JsonResponse({
                'success': True,
                'message': 'No hubo cambios que guardar',
                'cambios': [],
            })

        requerimiento.save()
        HistorialRequerimiento.objects.create(
            requerimiento=requerimiento,
            accion='DATOS_ACTUALIZADOS',
            comentario=' · '.join(cambios)[:2000],
            usuario=request.user,
        )

    return JsonResponse({
        'success': True,
        'message': f'{len(cambios)} dato(s) actualizado(s)',
        'cambios': cambios,
    })


@login_required
@require_POST
def enviar_a_proveedor(request, requerimiento_id):
    """Enviar requerimiento al proveedor por correo (con fotos adjuntas).

    Además despacha una copia-resumen SIN fotos a un correo de control para
    certificar que el envío al proveedor ocurrió (env REQUERIMIENTOS_CORREO_COPIA,
    campo correo_copia del POST, o el correo del usuario que envía).
    """
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}

    requerimiento = get_object_or_404(
        Requerimiento.objects.select_related('proveedor', 'sucursal', 'sucursal__empresa'),
        id=requerimiento_id
    )

    # Validar permisos (solo administrador)
    if not usuario_puede_realizar_accion(request.user, requerimiento, 'enviar_proveedor'):
        return JsonResponse({
            'success': False,
            'error': 'No tiene permisos para enviar a proveedor'
        }, status=403)

    if not requerimiento.proveedor:
        return JsonResponse({
            'success': False,
            'error': 'El requerimiento no tiene proveedor asignado'
        }, status=400)

    # Un caso cerrado no se le manda a nadie. Sin esta guarda se podía
    # "enviar" un requerimiento CANCELADO y quedaba resucitado en
    # ESPERANDO_RESPUESTA.
    ESTADOS_ENVIABLES = ('PENDIENTE', 'EN_REVISION', 'VALIDADO', 'ESPERANDO_RESPUESTA')
    if requerimiento.estado not in ESTADOS_ENVIABLES:
        return JsonResponse({
            'success': False,
            'error': f'El requerimiento está "{requerimiento.get_estado_display()}" '
                     f'y ya no corresponde enviarlo al proveedor'
        }, status=400)

    # Correo destino: manual > último envío > correoVendedor > email > correoIntercambio.
    # El "último envío" importa para el recordatorio rápido de la lista (que no
    # manda correo en el POST): sin él, el reenvío se iba al correo de la ficha
    # aunque el envío original se hubiera hecho a un correo tipeado a mano.
    correo_destino = (
        (data.get('correo_destino') or '').strip()
        or (requerimiento.correo_proveedor_destino or '').strip()
        or _correo_proveedor(requerimiento.proveedor)
    )
    if not correo_destino:
        return JsonResponse({
            'success': False,
            'error': 'El proveedor no tiene ningún correo configurado en su ficha. Ingrese uno manualmente.'
        }, status=400)
    try:
        validate_email(correo_destino)
    except ValidationError:
        return JsonResponse({
            'success': False,
            'error': f'El correo destino no es válido: {correo_destino}'
        }, status=400)

    # Correo de copia (resumen sin fotos)
    correo_copia = (data.get('correo_copia') or '').strip() or _correo_copia_default(request.user)
    if correo_copia:
        try:
            validate_email(correo_copia)
        except ValidationError:
            return JsonResponse({
                'success': False,
                'error': f'El correo de copia no es válido: {correo_copia}'
            }, status=400)

    mensaje_adicional = (data.get('mensaje') or '').strip()
    es_reenvio = bool(data.get('es_reenvio')) or requerimiento.correo_enviado_proveedor

    # Cada foto se descarga UNA sola vez del storage DEL CAMPO (puede ser
    # Spaces y no el default: `default_storage` miraría el disco local) y los
    # bytes se reusan para el PDF y para los adjuntos. Antes cada foto viajaba
    # 3 veces por la red (exists + lectura del PDF + lectura del adjunto) y el
    # envío demoraba proporcionalmente.
    fotos = list(requerimiento.fotos.select_related('tipo_foto').all())
    fotos_bytes = {}
    for foto in fotos:
        if not foto.imagen:
            continue
        try:
            with foto.imagen.storage.open(foto.imagen.name, 'rb') as fh:
                fotos_bytes[foto.id] = fh.read()
        except Exception as e:
            logger.warning("Error al leer foto de requerimiento %s: %s", requerimiento.id, e)
    fotos_adjuntables = [foto for foto in fotos if foto.id in fotos_bytes]
    fotos_registradas = len(fotos)

    # Formato propio de RetailMind: el documento formal que el proveedor
    # archiva y responde. Lleva las fotos incrustadas y reescaladas, así que
    # va SIEMPRE aunque las originales no quepan como adjunto.
    pdf_bytes = None
    try:
        pdf_bytes = generar_pdf_requerimiento(
            requerimiento, usuario=request.user, plazo_dias=PLAZO_RESPUESTA_DIAS,
            fotos_bytes=fotos_bytes,
        )
    except Exception:
        # Si el formato falla, el correo igual sale: perder el envío por un
        # problema de maquetación sería peor que mandarlo sin el PDF.
        logger.exception('No se pudo generar el formato PDF del requerimiento %s',
                         requerimiento.id)

    # Las fotos se leen y se presupuestan ANTES de armar el correo: el texto
    # decía "se adjuntan N fotos" con las fotos *adjuntables*, pero el tope de
    # MB podía dejar varias afuera y el proveedor recibía la promesa de una
    # evidencia que no venía.
    presupuesto = MAX_ADJUNTOS_MB * 1024 * 1024 - len(pdf_bytes or b'')
    adjuntos_fotos = []      # [(nombre, contenido)] que realmente se envían
    fotos_omitidas = []      # no cupieron: van igual dentro del PDF
    for foto in fotos_adjuntables:
        contenido = fotos_bytes[foto.id]
        if len(contenido) > presupuesto:
            fotos_omitidas.append(foto.imagen.name)
            continue
        presupuesto -= len(contenido)
        adjuntos_fotos.append((os.path.basename(foto.imagen.name), contenido))

    fotos_enviadas = len(adjuntos_fotos)

    context = {
        'requerimiento': requerimiento,
        'fotos': fotos,
        'cantidad_fotos_adjuntas': fotos_enviadas,
        # Solo en el PDF: no cupieron como adjunto pero sí van incrustadas.
        'fotos_solo_en_pdf': len(fotos_omitidas) if pdf_bytes else 0,
        # Su archivo ya no está en el servidor: no van en ninguna parte.
        'fotos_no_disponibles': fotos_registradas - len(fotos_adjuntables),
        'usuario': request.user,
        'empresa': requerimiento.sucursal.empresa,
        'mensaje_adicional': mensaje_adicional,
        'es_reenvio': es_reenvio,
        'lleva_formato_pdf': pdf_bytes is not None,
        'plazo_respuesta_dias': PLAZO_RESPUESTA_DIAS,
        'fecha_limite_respuesta': timezone.localdate() + timedelta(days=PLAZO_RESPUESTA_DIAS),
    }

    # Asunto filtrable por el proveedor: sin SKU ni factura no puede buscarlo
    # en su bandeja ni cruzarlo con su sistema.
    prefijo = 'RECORDATORIO · ' if es_reenvio else ''
    referencia_compra = requerimiento.numero_factura_compra or (
        str(requerimiento.dte_compra.numero_documento) if requerimiento.dte_compra_id else '')
    partes_asunto = [
        f'[{requerimiento.get_tipo_display().upper()}]',
        requerimiento.numero_requerimiento,
        f'SKU {requerimiento.sku}',
    ]
    if referencia_compra:
        partes_asunto.append(f'FAC {referencia_compra}')
    partes_asunto.append(requerimiento.sucursal.empresa.nombre)
    asunto = prefijo + ' · '.join(partes_asunto)

    html_message = render_to_string('emails/requerimiento_proveedor.html', context)
    texto_plano = (
        f'Requerimiento {requerimiento.numero_requerimiento} - {requerimiento.get_tipo_display()}\n'
        f'Proveedor: {requerimiento.proveedor.nombre}\n'
        f'Producto: {requerimiento.sku} - {requerimiento.nombre_producto} '
        f'(cantidad: {requerimiento.cantidad})\n'
        f'{("Factura de compra: " + referencia_compra) if referencia_compra else "Sin factura de compra registrada"}\n'
        f'Motivo: {requerimiento.motivo}\n'
        f'{("Mensaje: " + mensaje_adicional) if mensaje_adicional else ""}\n'
        f'{"Se adjunta el formato del requerimiento en PDF con la evidencia fotográfica. " if pdf_bytes else ""}'
        f'Se adjuntan {fotos_enviadas} foto(s) en su resolución original.\n'
        f'Por favor responda indicando si procede.\n'
        f'Contacto: {request.user.get_full_name()} - {requerimiento.sucursal.empresa.nombre}'
    )

    # CC al administrador del proveedor (si existe y no es el mismo destino)
    cc = []
    correo_admin_proveedor = (requerimiento.proveedor.correoAdministrador or '').strip()
    if correo_admin_proveedor and correo_admin_proveedor.lower() != correo_destino.lower():
        cc.append(correo_admin_proveedor)

    # Una sola conexión SMTP para el correo al proveedor Y la copia de control:
    # cada send() suelto abre, negocia TLS y autentica de nuevo, o sea el
    # usuario pagaba dos handshakes completos por envío.
    connection = get_connection(timeout=EMAIL_TIMEOUT_SEGUNDOS)

    # Reply-To adicional: el usuario que envía y la casilla de control. La
    # casilla genérica con el token la antepone `enviar_correo_trazado` cuando
    # está configurada (CORREO_BUZON_RESPUESTAS), de modo que la respuesta del
    # proveedor pueda pegarse sola en esta ficha en vez de morir en el correo
    # personal de quien lo mandó.
    reply_to = []
    for direccion in ((request.user.email or '').strip(), correo_copia):
        if direccion and direccion.lower() not in (d.lower() for d in reply_to):
            reply_to.append(direccion)

    # 1) El formato propio va primero: es el documento del reclamo.
    # 2) Después las fotos originales que entraron en el presupuesto, ya leídas
    #    más arriba (por el storage del campo y no por `.path`: con
    #    almacenamiento remoto `.path` lanza NotImplementedError).
    adjuntos = []
    if pdf_bytes:
        adjuntos.append((nombre_archivo_pdf(requerimiento), pdf_bytes, 'application/pdf'))
    adjuntos += [(nombre, contenido) for nombre, contenido in adjuntos_fotos]

    if fotos_omitidas:
        logger.info(
            'Requerimiento %s: %s foto(s) no se adjuntaron en original por el tope '
            'de %sMB (van igual dentro del PDF)',
            requerimiento.id, len(fotos_omitidas), MAX_ADJUNTOS_MB,
        )

    try:
        # Abrir a mano: si la abre send(), Django la cierra al terminar ese
        # send() y la copia vuelve a pagar la conexión completa.
        connection.open()
        envio = enviar_correo_trazado(
            modulo='REQUERIMIENTO',
            objeto_id=requerimiento.id,
            asunto=asunto,
            texto=texto_plano,
            html=html_message,
            destinatario=correo_destino,
            cc=cc,
            reply_to=reply_to,
            adjuntos=adjuntos,
            from_email=(getattr(settings, 'REQUERIMIENTOS_FROM_EMAIL', '')
                        or settings.DEFAULT_FROM_EMAIL),
            usuario=request.user,
            connection=connection,
            tags=['requerimiento', requerimiento.tipo.lower()],
        )
    except CorreoError as e:
        try:
            connection.close()
        except Exception:
            pass
        # El fallo queda en el historial, no solo en el log: sin esto, mañana
        # nadie sabe que este requerimiento nunca le llegó al proveedor.
        HistorialRequerimiento.objects.create(
            requerimiento=requerimiento,
            accion='ENVIO_FALLIDO',
            comentario=(f'NO se pudo enviar a {requerimiento.proveedor.nombre} '
                        f'({correo_destino}): {e}'),
            usuario=request.user,
        )
        return JsonResponse({
            'success': False,
            'error': f'Error al enviar correo al proveedor: {e}'
        }, status=500)

    # Actualizar requerimiento + historial
    estado_anterior = requerimiento.estado
    with transaction.atomic():
        requerimiento.correo_enviado_proveedor = True
        requerimiento.fecha_envio_proveedor = timezone.now()
        requerimiento.correo_proveedor_destino = correo_destino
        requerimiento.intentos_envio = (requerimiento.intentos_envio or 0) + 1
        if es_reenvio:
            requerimiento.ultimo_recordatorio = timezone.now()
        requerimiento.estado = 'ESPERANDO_RESPUESTA'
        requerimiento.save()

        HistorialRequerimiento.objects.create(
            requerimiento=requerimiento,
            accion='RECORDATORIO_ENVIADO' if es_reenvio else 'ENVIADO_A_PROVEEDOR',
            estado_anterior=estado_anterior,
            estado_nuevo='ESPERANDO_RESPUESTA',
            comentario=(
                f'Correo enviado a {requerimiento.proveedor.nombre} ({correo_destino}) '
                f'- Intento #{requerimiento.intentos_envio} '
                f'- {"con" if pdf_bytes else "SIN"} formato PDF '
                f'- {fotos_enviadas} foto(s) adjuntas'
                + (f' [envio #{envio.id}]' if envio else '')
            ),
            usuario=request.user
        )

    # Copia-resumen de control: SIN fotos adjuntas (solo el conteo)
    copia_enviada = False
    if correo_copia:
        try:
            context_copia = dict(context)
            context_copia.update({
                'correo_destino': correo_destino,
                'correo_cc': ', '.join(cc),
                'reply_to': ', '.join(reply_to),
                'fecha_envio': requerimiento.fecha_envio_proveedor,
                'intento': requerimiento.intentos_envio,
                'url_detalle': request.build_absolute_uri(f'/app/requerimientos/{requerimiento.id}/'),
            })
            html_copia = render_to_string('emails/requerimiento_copia_resumen.html', context_copia)
            texto_copia = (
                f'COPIA DE CONTROL - Requerimiento {requerimiento.numero_requerimiento}\n'
                f'Enviado al proveedor {requerimiento.proveedor.nombre} ({correo_destino}) '
                f'el {requerimiento.fecha_envio_proveedor.strftime("%d/%m/%Y %H:%M")} '
                f'por {request.user.get_full_name()}.\n'
                f'Producto: {requerimiento.sku} - {requerimiento.nombre_producto}\n'
                f'Motivo: {requerimiento.motivo}\n'
                f'Formato PDF adjuntado: {"si" if pdf_bytes else "NO"}\n'
                f'Fotos adjuntadas al proveedor: {fotos_enviadas} (esta copia no incluye adjuntos).'
            )
            # La copia va sin píxel ni token de respuesta: es correo interno,
            # medir su apertura no aporta nada y el token debe apuntar al
            # correo del proveedor, no a este.
            enviar_correo_trazado(
                modulo='REQUERIMIENTO',
                objeto_id=requerimiento.id,
                asunto=f'[COPIA] {asunto} → {correo_destino}',
                texto=texto_copia,
                html=html_copia,
                destinatario=correo_copia,
                from_email=(getattr(settings, 'REQUERIMIENTOS_FROM_EMAIL', '')
                            or settings.DEFAULT_FROM_EMAIL),
                usuario=request.user,
                connection=connection,
                tags=['requerimiento', 'copia-control'],
                con_pixel=False,
                con_token_respuesta=False,
                es_copia_control=True,
            )
            copia_enviada = True

            HistorialRequerimiento.objects.create(
                requerimiento=requerimiento,
                accion='COPIA_RESUMEN_ENVIADA',
                comentario=f'Copia-resumen (sin fotos) enviada a {correo_copia}',
                usuario=request.user
            )
        except Exception as e:
            # La copia es de control: su falla no revierte el envío al proveedor
            logger.exception("Error al enviar copia-resumen del requerimiento %s a %s", requerimiento.id, correo_copia)

    try:
        connection.close()
    except Exception:
        pass

    mensaje_out = (
        f'Requerimiento enviado a {requerimiento.proveedor.nombre} ({correo_destino})'
        f'{" con el formato PDF" if pdf_bytes else " SIN el formato PDF (revisar logs)"}'
        f' y {fotos_enviadas} foto(s) adjuntas'
    )
    if fotos_omitidas:
        mensaje_out += (f'. {len(fotos_omitidas)} foto(s) superaron el tope de '
                        f'{MAX_ADJUNTOS_MB}MB y van solo dentro del PDF')
    if copia_enviada:
        mensaje_out += f'. Copia-resumen enviada a {correo_copia}'
    elif correo_copia:
        mensaje_out += f'. ATENCIÓN: falló la copia-resumen a {correo_copia} (revisar logs)'

    return JsonResponse({
        'success': True,
        'message': mensaje_out,
        'fecha_envio': requerimiento.fecha_envio_proveedor.strftime('%d/%m/%Y %H:%M'),
        'correo_destino': correo_destino,
        'copia_enviada': copia_enviada,
        'correo_copia': correo_copia or '',
        'fotos_adjuntas': fotos_enviadas,
        'fotos_omitidas': len(fotos_omitidas),
        'formato_pdf_adjunto': bool(pdf_bytes),
        # Para que la ficha pueda mostrar después si llegó, si rebotó o si lo
        # abrieron, sin tener que volver a buscar el envío.
        'envio_id': envio.id,
        'relay_id': envio.proveedor_message_id or '',
    })


@login_required
@require_POST
def registrar_respuesta_proveedor(request, requerimiento_id):
    """Registrar respuesta del proveedor (solo administrador)"""
    try:
        data = json.loads(request.body)
        
        requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)
        
        # Validar permisos (solo administrador)
        if not usuario_puede_realizar_accion(request.user, requerimiento, 'registrar_respuesta_proveedor'):
            return JsonResponse({
                'success': False,
                'error': 'Solo administradores pueden registrar respuestas de proveedores'
            }, status=403)
        
        respuesta = data.get('respuesta')
        decision = data.get('decision')  # 'APROBADO' o 'RECHAZADO'
        motivo = data.get('motivo', '')  # Motivo visible al usuario
        
        if not respuesta or not decision:
            return JsonResponse({
                'success': False,
                'error': 'La respuesta y decisión son requeridas'
            }, status=400)
        
        if decision not in ['APROBADO', 'RECHAZADO', 'PARCIAL']:
            return JsonResponse({
                'success': False,
                'error': 'Decisión debe ser APROBADO o RECHAZADO'
            }, status=400)

        # Solo se registra la respuesta de un requerimiento que efectivamente
        # se envió. Sin esta guarda se podía cerrar como "Aprobado por el
        # proveedor" un caso que nunca salió de la tienda.
        if requerimiento.estado != 'ESPERANDO_RESPUESTA':
            return JsonResponse({
                'success': False,
                'error': (f'El requerimiento está "{requerimiento.get_estado_display()}". '
                          f'Solo se registra la respuesta de un caso enviado al proveedor.')
            }, status=400)

        # Fecha real de la respuesta: el modal la pide y hasta ahora se
        # descartaba (siempre se guardaba el instante del registro, que puede
        # ser días después de que el proveedor contestó).
        fecha_respuesta = timezone.now()
        if data.get('fecha_respuesta'):
            try:
                parseada = datetime.strptime(data['fecha_respuesta'][:16], '%Y-%m-%dT%H:%M')
                fecha_respuesta = timezone.make_aware(parseada)
            except (ValueError, TypeError):
                pass  # formato inesperado: se deja el instante actual

        with transaction.atomic():
            requerimiento.respuesta_proveedor = respuesta
            requerimiento.fecha_respuesta_proveedor = fecha_respuesta
            requerimiento.decision_proveedor = decision
            # PARCIAL es una aprobación acotada (el proveedor acepta parte del
            # reclamo): dejarlo como RECHAZADO cerraba el caso sin poder
            # tramitar lo que sí aprobó.
            requerimiento.estado = 'RECHAZADO' if decision == 'RECHAZADO' else 'APROBADO'
            requerimiento.fecha_resolucion = timezone.now()
            
            # Motivo visible al usuario
            if motivo:
                requerimiento.motivo_resolucion = motivo
            else:
                requerimiento.motivo_resolucion = f"{decision}: {respuesta[:200]}"
            
            requerimiento.save()
            
            # Registrar en historial
            HistorialRequerimiento.objects.create(
                requerimiento=requerimiento,
                accion='RESPUESTA_PROVEEDOR_REGISTRADA',
                estado_anterior='ESPERANDO_RESPUESTA',
                estado_nuevo=requerimiento.estado,
                comentario=(f'Proveedor {requerimiento.proveedor.nombre if requerimiento.proveedor_id else "(sin proveedor)"} '
                            f'respondió: {decision} - {respuesta[:100]}...'),
                usuario=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Respuesta del proveedor registrada exitosamente',
            'decision': decision,
            'nuevo_estado': requerimiento.get_estado_display()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al registrar respuesta: {str(e)}'
        }, status=500)


@login_required
@require_POST
def completar_requerimiento(request, requerimiento_id):
    """Cerrar el requerimiento con su resolución final.

    Esta vista no validaba NADA: ni permisos ni estado. Cualquier usuario
    logueado podía cerrar como COMPLETADO el requerimiento de otra empresa,
    incluso uno recién creado que nunca se revisó.
    """
    try:
        data = json.loads(request.body)

        resolucion = data.get('resolucion')

        if not resolucion:
            return JsonResponse({
                'success': False,
                'error': 'La resolución es requerida'
            }, status=400)

        requerimiento = get_object_or_404(
            Requerimiento.objects.select_related('sucursal'), id=requerimiento_id)
        estado_anterior = requerimiento.estado

        if not usuario_puede_realizar_accion(request.user, requerimiento, 'completar'):
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para completar este requerimiento'
            }, status=403)

        if not puede_cambiar_estado(estado_anterior, 'COMPLETADO'):
            return JsonResponse({
                'success': False,
                'error': (f'El requerimiento está "{requerimiento.get_estado_display()}" '
                          f'y desde ahí no se puede completar')
            }, status=400)

        with transaction.atomic():
            requerimiento.resolucion = resolucion
            requerimiento.fecha_resolucion = timezone.now()
            requerimiento.estado = 'COMPLETADO'
            requerimiento.save()
            
            # Registrar en historial
            HistorialRequerimiento.objects.create(
                requerimiento=requerimiento,
                accion='COMPLETADO',
                estado_anterior=estado_anterior,
                estado_nuevo='COMPLETADO',
                comentario=resolucion,
                usuario=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Requerimiento completado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al completar requerimiento: {str(e)}'
        }, status=500)


# ========== BÚSQUEDA Y UTILIDADES ==========

@login_required
def buscar_producto_sku(request):
    """Buscar producto por SKU"""
    try:
        sku = request.GET.get('sku', '')
        
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU es requerido'
            }, status=400)
        
        try:
            from .utils_producto_match import producto_talla_por_sku
            producto_talla = producto_talla_por_sku(
                sku, sucursal_id=request.session.get('idSucursalActual'),
                select_related=['producto'])
            if not producto_talla:
                return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)

            return JsonResponse({
                'success': True,
                'producto': {
                    'id': producto_talla.id,
                    'sku': producto_talla.sku,
                    'nombre': producto_talla.producto.articulo,
                    'descripcion': producto_talla.producto.descripcion,
                    'talla': producto_talla.talla,
                    'precio': producto_talla.producto.precioventa,
                }
            })
        except Producto_Talla.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Producto no encontrado'
            }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar producto: {str(e)}'
        }, status=500)


@login_required
def buscar_ticket_por_folio(request):
    """Buscar ticket/documento por folio o correlativo en TODAS las sucursales de la DB"""
    try:
        folio = request.GET.get('folio', '').strip()

        if not folio:
            return JsonResponse({
                'success': False,
                'error': 'Folio o correlativo es requerido'
            }, status=400)

        resultados = []

        if folio.isdigit():
            folio_num = int(folio)

            # 1. Buscar en Tickets por folio_dte (TODAS las sucursales)
            for ticket in Ticket.objects.filter(
                folio_dte=folio_num
            ).select_related('vendedor', 'sucursal')[:10]:
                resultados.append(_serializar_ticket(ticket))

            # 2. Buscar en Tickets por correlativo (si no hay resultados por folio_dte)
            if not resultados:
                for ticket in Ticket.objects.filter(
                    correlativo=folio_num
                ).select_related('vendedor', 'sucursal')[:10]:
                    resultados.append(_serializar_ticket(ticket))

            # 3. Buscar en DTEs
            if not resultados:
                for dte in Dte.objects.filter(
                    numero_documento=folio_num,
                    tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO']
                ).select_related('vendedor', 'emisor', 'receptor', 'sucursal')[:10]:
                    resultados.append(_serializar_dte(dte))

        if not resultados:
            return JsonResponse({
                'success': False,
                'error': 'Documento no encontrado'
            }, status=404)

        # Si hay un solo resultado, devolver directo (retrocompatible)
        if len(resultados) == 1:
            return JsonResponse({
                'success': True,
                'documento': resultados[0],
            })

        # Multiples resultados: devolver lista para que el usuario elija
        return JsonResponse({
            'success': True,
            'multiple': True,
            'documentos': resultados,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar documento: {str(e)}'
        }, status=500)


def _serializar_ticket(ticket):
    """Serializa un Ticket para la respuesta de busqueda"""
    productos = []
    for tp in ticket.ticket_productos.select_related('ProductoTalla__producto').all():
        if tp.ProductoTalla:
            productos.append({
                'sku': tp.ProductoTalla.sku,
                'nombre': tp.ProductoTalla.producto.articulo,
                'talla': tp.ProductoTalla.talla,
                'cantidad': tp.stock,
                'precio': tp.precio,
            })

    return {
        'tipo_fuente': 'ticket',
        'sucursal': ticket.sucursal.alias if ticket.sucursal else '',
        'sucursal_id': ticket.sucursal.id if ticket.sucursal else None,
        'correlativo': ticket.correlativo,
        'folio_dte': ticket.folio_dte,
        'tipo_dte': ticket.get_tipo_dte_display() if ticket.tipo_dte else 'Ticket',
        'tipo_dte_codigo': ticket.tipo_dte or 'TICKET',
        'fecha': ticket.fecha.strftime('%Y-%m-%d'),
        'total': ticket.total,
        'vendedor': ticket.vendedor.nombre if ticket.vendedor else '',
        'cliente_nombre': ticket.cliente_nombre or '',
        'cliente_rut': ticket.cliente_rut or '',
        'cliente_email': ticket.cliente_email or '',
        'cliente_telefono': ticket.cliente_telefono or '',
        'cliente_direccion': ticket.cliente_direccion or '',
        'cliente_comuna': ticket.cliente_comuna or '',
        'productos': productos,
    }


def _serializar_dte(dte):
    """Serializa un DTE para la respuesta de busqueda"""
    productos = []
    for dp in Dte_Productos.objects.filter(dte=dte).select_related('productoTalla__producto'):
        productos.append({
            'sku': dp.productoTalla.sku if dp.productoTalla else '',
            'nombre': dp.productoTalla.producto.articulo if dp.productoTalla else dp.descripcion,
            'talla': dp.productoTalla.talla if dp.productoTalla else '',
            'cantidad': dp.stock,
            'precio': int(dp.precio),
        })

    cliente_nombre = ''
    cliente_rut = ''
    cliente_email = ''
    cliente_direccion = ''
    cliente_comuna = ''
    if dte.receptor:
        cliente_nombre = dte.receptor.nombre
        cliente_rut = dte.receptor.rut
        cliente_email = dte.receptor.correoAdministrador or ''
        cliente_direccion = dte.receptor.direccion or ''
        cliente_comuna = dte.receptor.comuna or ''

    return {
        'tipo_fuente': 'dte',
        'sucursal': dte.sucursal.alias if dte.sucursal else 'N/A',
        'sucursal_id': dte.sucursal.id if dte.sucursal else None,
        'correlativo': dte.numero_documento,
        'folio_dte': dte.numero_documento,
        'tipo_dte': dte.get_tipo_documento_display(),
        'tipo_dte_codigo': dte.tipo_documento,
        'fecha': dte.fecha_emision.strftime('%Y-%m-%d'),
        'total': int(dte.monto_con_iva),
        'vendedor': dte.vendedor.nombre if dte.vendedor else '',
        'cliente_nombre': cliente_nombre,
        'cliente_rut': cliente_rut,
        'cliente_email': cliente_email,
        'cliente_telefono': '',
        'cliente_direccion': cliente_direccion,
        'cliente_comuna': cliente_comuna,
        'productos': productos,
    }


@login_required
def validar_rut_chileno(request):
    """Validar formato y dígito verificador de RUT chileno"""
    try:
        rut = request.GET.get('rut', '').strip()
        
        if not rut:
            return JsonResponse({'success': False, 'error': 'RUT es requerido'}, status=400)
        
        # Limpiar RUT
        rut = rut.replace('.', '').replace('-', '').upper()
        
        if len(rut) < 2:
            return JsonResponse({'success': False, 'error': 'RUT inválido'}, status=400)
        
        # Separar número y dígito verificador
        numero = rut[:-1]
        dv = rut[-1]
        
        # Validar que el número sea numérico
        if not numero.isdigit():
            return JsonResponse({'success': False, 'error': 'RUT inválido'}, status=400)
        
        # Calcular dígito verificador
        suma = 0
        multiplicador = 2
        
        for digito in reversed(numero):
            suma += int(digito) * multiplicador
            multiplicador = multiplicador + 1 if multiplicador < 7 else 2
        
        resto = suma % 11
        dv_calculado = 11 - resto
        
        if dv_calculado == 11:
            dv_calculado = '0'
        elif dv_calculado == 10:
            dv_calculado = 'K'
        else:
            dv_calculado = str(dv_calculado)
        
        # Formatear RUT
        rut_formateado = f"{numero[:-6]}.{numero[-6:-3]}.{numero[-3:]}-{dv}" if len(numero) > 6 else f"{numero}-{dv}"
        
        if dv == dv_calculado:
            return JsonResponse({
                'success': True,
                'valido': True,
                'rut_formateado': rut_formateado,
                'message': 'RUT válido'
            })
        else:
            return JsonResponse({
                'success': True,
                'valido': False,
                'message': f'RUT inválido. DV correcto debería ser: {dv_calculado}'
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al validar RUT: {str(e)}'
        }, status=500)


@login_required
def buscar_cliente_por_rut(request):
    """Buscar cliente por RUT en la base de datos"""
    try:
        from app.models import Cliente

        rut = request.GET.get('rut', '').strip()
        
        if not rut:
            return JsonResponse({
                'success': False,
                'error': 'RUT es requerido'
            }, status=400)
        
        # Limpiar RUT (quitar puntos y guiones)
        rut_limpio = rut.replace('.', '').replace('-', '')
        
        # Buscar cliente por RUT (con o sin formato)
        cliente = Cliente.objects.filter(
            Q(rut__icontains=rut_limpio) | Q(rut__icontains=rut)
        ).first()
        
        if not cliente:
            return JsonResponse({
                'success': False,
                'error': 'Cliente no encontrado'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'cliente': {
                'id': cliente.id,
                'nombre': cliente.nombre_completo,
                'rut': cliente.rut or '',
                'email': cliente.email or '',
                'telefono': cliente.telefono or cliente.celular or '',
                'direccion': cliente.direccion or '',
                'comuna': cliente.comuna or '',
                'ciudad': cliente.ciudad or '',
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar cliente: {str(e)}'
        }, status=500)


@login_required
@require_POST
def crear_cliente_rapido(request):
    """Crear cliente rápido desde formulario de requerimientos"""
    try:
        from app.models import Cliente

        data = json.loads(request.body)
        
        # Validar campos requeridos
        if not data.get('nombre') or not data.get('apellido'):
            return JsonResponse({
                'success': False,
                'error': 'Nombre y apellido son requeridos'
            }, status=400)
        
        rut = data.get('rut', '').strip()
        
        # Validar RUT si se proporciona
        if rut:
            # Verificar que no exista
            if Cliente.objects.filter(rut=rut).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe un cliente con este RUT'
                }, status=400)
        
        # Crear cliente
        with transaction.atomic():
            cliente = Cliente.objects.create(
                nombre=data.get('nombre'),
                apellido=data.get('apellido', ''),
                rut=rut if rut else None,
                email=data.get('email', ''),
                telefono=data.get('telefono', ''),
                direccion=data.get('direccion', ''),
                comuna=data.get('comuna', ''),
                ciudad=data.get('ciudad', ''),
                tipo_cliente='INDIVIDUAL',
                created_by=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Cliente creado exitosamente',
            'cliente': {
                'id': cliente.id,
                'nombre': cliente.nombre_completo,
                'rut': cliente.rut or '',
                'email': cliente.email or '',
                'telefono': cliente.telefono or '',
                'direccion': cliente.direccion or '',
                'comuna': cliente.comuna or '',
                'ciudad': cliente.ciudad or '',
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear cliente: {str(e)}'
        }, status=500)


@login_required
def obtener_estadisticas_requerimientos(request):
    """Obtener estadísticas del módulo de requerimientos"""
    try:
        # Filtrar por sucursal si no es admin
        requerimientos = Requerimiento.objects.all()
        
        # MISMO alcance que `listar_requerimientos`: el jefe de local ve su
        # SUCURSAL en la tabla, pero los KPI y los contadores de las pestañas
        # sumaban toda la EMPRESA, así que mostraban números que al hacer clic
        # no se podían abrir.
        rol_usuario = obtener_rol_usuario(request.user)
        if rol_usuario == 'jefe_local':
            empresa_user = EmpresaUser.objects.filter(user=request.user).first()
            if empresa_user and empresa_user.sucursal:
                requerimientos = requerimientos.filter(sucursal=empresa_user.sucursal)
            else:
                requerimientos = requerimientos.none()
        elif rol_usuario != 'administrador':
            requerimientos = requerimientos.filter(
                Q(usuario_creador=request.user) |
                Q(sucursal__in=obtener_sucursales_usuario(request.user))
            )
        
        # Estadísticas generales
        total = requerimientos.count()
        
        # Obtener choices del modelo
        from .models import ESTADO_REQUERIMIENTO_CHOICES, TIPO_REQUERIMIENTO_CHOICES
        
        por_estado = {}
        for estado_code, estado_name in ESTADO_REQUERIMIENTO_CHOICES:
            count = requerimientos.filter(estado=estado_code).count()
            por_estado[estado_code] = {
                'nombre': estado_name,
                'cantidad': count
            }
        
        por_tipo = {}
        for tipo_code, tipo_name in TIPO_REQUERIMIENTO_CHOICES:
            count = requerimientos.filter(tipo=tipo_code).count()
            por_tipo[tipo_code] = {
                'nombre': tipo_name,
                'cantidad': count
            }
        
        # Requerimientos recientes
        recientes = requerimientos.order_by('-fecha_creacion')[:5]
        recientes_data = []
        for req in recientes:
            recientes_data.append({
                'id': req.id,
                'numero': req.numero_requerimiento,
                'tipo': req.get_tipo_display(),
                'estado': req.get_estado_display(),
                'dias': req.dias_transcurridos
            })
        
        # Contadores especiales de seguimiento
        fecha_limite = timezone.now() - timedelta(days=PLAZO_RESPUESTA_DIAS)
        sin_respuesta_7dias = requerimientos.filter(
            estado='ESPERANDO_RESPUESTA',
            fecha_envio_proveedor__lt=fecha_limite,
            fecha_respuesta_proveedor__isnull=True
        ).count()

        # Agregados por etapa: es lo que el KPI debe mostrar. "Pendientes" a
        # secas no distinguía un caso recién creado de uno esperando al
        # proveedor hace tres semanas.
        por_etapa = {}
        for nombre_etapa in ('EMPRESA', 'PROVEEDOR', 'RESOLUCION', 'CERRADO'):
            estados_etapa = [e for e, et in ETAPA_POR_ESTADO.items() if et == nombre_etapa]
            por_etapa[nombre_etapa] = requerimientos.filter(
                estado__in=estados_etapa).count()

        # Bandeja "les falta algo": no pueden salir al proveedor tal como están.
        incompletos = requerimientos.exclude(estado__in=ESTADOS_CERRADOS).filter(
            Q(proveedor__isnull=True) |
            (Q(numero_factura_compra__isnull=True) | Q(numero_factura_compra=''))
            & Q(dte_compra__isnull=True)
        ).count()

        return JsonResponse({
            'success': True,
            'estadisticas': {
                'total': total,
                'por_estado': por_estado,
                'por_tipo': por_tipo,
                'por_etapa': por_etapa,
                'recientes': recientes_data,
                'sin_respuesta_7dias': sin_respuesta_7dias,
                'incompletos': incompletos,
                'plazo_respuesta_dias': PLAZO_RESPUESTA_DIAS,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estadísticas: {str(e)}'
        }, status=500)


@login_required
def exportar_requerimientos(request):
    """Exportar requerimientos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Filtros. Deben ser LOS MISMOS que los del listado: el botón de Excel
        # manda todo lo que hay puesto en pantalla, y al ignorar la mitad el
        # archivo bajaba más filas de las que el usuario estaba viendo.
        estado = request.GET.get('estado')
        tipo = request.GET.get('tipo')
        prioridad = request.GET.get('prioridad')
        busqueda = request.GET.get('busqueda', '')
        etapa = request.GET.get('etapa')
        incompletos = request.GET.get('incompletos')
        sin_respuesta = request.GET.get('sin_respuesta')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Query
        requerimientos = Requerimiento.objects.select_related(
            'sucursal', 'usuario_creador', 'proveedor'
        ).all()

        # MISMO alcance que el listado: sin esto cualquier usuario logueado
        # descargaba los requerimientos de todas las empresas del holding.
        rol_usuario = obtener_rol_usuario(request.user)
        if rol_usuario == 'jefe_local':
            empresa_user = EmpresaUser.objects.filter(user=request.user).first()
            if empresa_user and empresa_user.sucursal:
                requerimientos = requerimientos.filter(sucursal=empresa_user.sucursal)
            else:
                requerimientos = requerimientos.none()
        elif rol_usuario != 'administrador':
            requerimientos = requerimientos.filter(
                Q(usuario_creador=request.user) |
                Q(sucursal__in=obtener_sucursales_usuario(request.user))
            )

        # Aplicar filtros
        if estado:
            requerimientos = requerimientos.filter(estado=estado)
        if tipo:
            requerimientos = requerimientos.filter(tipo=tipo)
        if prioridad:
            requerimientos = requerimientos.filter(prioridad=prioridad)
        if busqueda:
            requerimientos = requerimientos.filter(
                Q(numero_requerimiento__icontains=busqueda) |
                Q(sku__icontains=busqueda) |
                Q(cliente_nombre__icontains=busqueda) |
                Q(cliente_rut__icontains=busqueda) |
                Q(numero_boleta__icontains=busqueda)
            )
        if etapa:
            estados_etapa = [e for e, et in ETAPA_POR_ESTADO.items() if et == etapa]
            if estados_etapa:
                requerimientos = requerimientos.filter(estado__in=estados_etapa)
        if incompletos == 'true':
            requerimientos = requerimientos.exclude(estado__in=ESTADOS_CERRADOS).filter(
                Q(proveedor__isnull=True) |
                (Q(dte_compra__isnull=True)
                 & (Q(numero_factura_compra__isnull=True) | Q(numero_factura_compra='')))
            )
        if sin_respuesta == 'true':
            requerimientos = requerimientos.filter(
                estado='ESPERANDO_RESPUESTA',
                fecha_envio_proveedor__lt=timezone.now() - timedelta(days=PLAZO_RESPUESTA_DIAS),
                fecha_respuesta_proveedor__isnull=True)
        if fecha_inicio:
            try:
                dt_inicio = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
                requerimientos = requerimientos.filter(fecha_creacion__gte=dt_inicio)
            except (ValueError, TypeError):
                pass
        if fecha_fin:
            try:
                dt_fin = timezone.make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
                requerimientos = requerimientos.filter(fecha_creacion__lte=dt_fin)
            except (ValueError, TypeError):
                pass

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requerimientos"
        
        # Encabezados
        headers = [
            'Número', 'Tipo', 'Estado', 'Etapa', 'Prioridad', 'Origen', 'Sucursal', 'SKU',
            'Producto', 'Cantidad', 'Cliente', 'Boleta', 'Factura Compra', 'Fecha Creación',
            'Días', 'Decisión Interna', 'Motivo Decisión Interna', 'Proveedor',
            'Estado Proveedor', 'Decisión Proveedor', 'Días sin Respuesta'
        ]
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Escribir datos
        for row, req in enumerate(requerimientos, 2):
            factura = req.numero_factura_compra or (
                str(req.dte_compra.numero_documento) if req.dte_compra_id else '')
            ws.cell(row=row, column=1, value=req.numero_requerimiento)
            ws.cell(row=row, column=2, value=req.get_tipo_display())
            ws.cell(row=row, column=3, value=req.get_estado_display())
            ws.cell(row=row, column=4, value=req.etapa)
            ws.cell(row=row, column=5, value=req.get_prioridad_display())
            ws.cell(row=row, column=6, value=req.get_origen_display())
            ws.cell(row=row, column=7, value=req.sucursal.alias)
            ws.cell(row=row, column=8, value=req.sku)
            ws.cell(row=row, column=9, value=req.nombre_producto)
            ws.cell(row=row, column=10, value=req.cantidad)
            ws.cell(row=row, column=11, value=req.cliente_nombre)
            ws.cell(row=row, column=12, value=req.numero_boleta or '')
            ws.cell(row=row, column=13, value=factura)
            ws.cell(row=row, column=14, value=req.fecha_creacion.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=15, value=req.dias_transcurridos)
            ws.cell(row=row, column=16, value=req.decision_interna or '')
            ws.cell(row=row, column=17, value=req.motivo_decision_interna or '')
            ws.cell(row=row, column=18, value=req.proveedor.nombre if req.proveedor else '')
            ws.cell(row=row, column=19, value='Enviado' if req.correo_enviado_proveedor else 'No enviado')
            ws.cell(row=row, column=20, value=req.decision_proveedor or '')
            ws.cell(row=row, column=21, value=req.dias_sin_respuesta)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="requerimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        }, status=500)


# ========== FORMATO PROPIO (PDF) ==========

@login_required
@require_GET
def descargar_formato_requerimiento(request, requerimiento_id):
    """Formato RetailMind del requerimiento en PDF.

    Es el MISMO documento que se adjunta al correo del proveedor, así que el
    analista puede revisarlo antes de enviar (y la tienda imprimirlo para el
    archivador). `?descargar=1` fuerza la descarga; por defecto abre en el
    navegador.
    """
    requerimiento = get_object_or_404(
        Requerimiento.objects.select_related(
            'sucursal', 'sucursal__empresa', 'proveedor', 'producto_talla',
            'producto_talla__producto', 'usuario_creador', 'dte_compra',
        ),
        id=requerimiento_id,
    )

    if not usuario_puede_realizar_accion(request.user, requerimiento, 'ver'):
        return JsonResponse({
            'success': False,
            'error': 'No tiene permisos para ver este requerimiento'
        }, status=403)

    try:
        pdf = generar_pdf_requerimiento(
            requerimiento, usuario=request.user, plazo_dias=PLAZO_RESPUESTA_DIAS,
        )
    except Exception as e:
        logger.exception('Error al generar el formato PDF del requerimiento %s',
                         requerimiento_id)
        return JsonResponse({
            'success': False,
            'error': f'No se pudo generar el formato: {e}'
        }, status=500)

    disposicion = 'attachment' if request.GET.get('descargar') else 'inline'
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'{disposicion}; filename="{nombre_archivo_pdf(requerimiento)}"')
    return response


# ========== BÚSQUEDA DEL RESPALDO DE COMPRA ==========
#
# El proveedor exige la factura con la que se le compró el producto, y ese
# dato es el que más se pierde. Buscarlo solo en las LÍNEAS de los DTE de
# compra deja ciego al analista en dos escenarios muy frecuentes:
#
#   1. DTE de compra sin líneas — buena parte de las compras cargadas solo
#      como cabecera (monto y proveedor, sin detalle de productos).
#   2. Datos migrados desde Laravel — el kardex trae el ingreso del producto,
#      pero la migración NO copió el N° de documento: `referencia_externa`
#      quedó ocupada con "MIG:<id de MySQL>" y el FK al DTE quedó vacío.
#
# Por eso la búsqueda cruza CUATRO fuentes y marca de cuál salió cada
# candidato, en vez de decir "no hay compras registradas" y dejar al usuario
# tecleando el número de memoria.

# Conceptos de kardex que significan "esta unidad entró comprada" (no por
# traspaso, ni por devolución de un cliente, ni por ajuste de inventario).
CONCEPTOS_INGRESO_COMPRA = (
    'RECEPCION_COMPRA', 'INGRESO_INICIAL', 'INGRESO_MANUAL', 'REPOSICION_STOCK',
)

# Cuánta confianza da cada fuente para elegir la factura correcta.
CONFIANZA_FUENTE = {
    'LINEA_DTE': 'ALTA',        # el documento dice explícitamente este SKU
    'LOTE': 'ALTA',             # el lote FIFO quedó apuntando al documento
    # El sistema anterior anotaba la factura EN el movimiento del producto al
    # crearlo, así que el número es tan bueno como el de una línea de DTE.
    'LEGACY': 'ALTA',
    'KARDEX_DTE': 'MEDIA',      # el movimiento apunta al DTE, sin línea de detalle
    'KARDEX_SIN_DOC': 'BAJA',   # hay ingreso pero el documento se perdió
    'CABECERA_PROVEEDOR': 'BAJA',  # compra del proveedor en fecha cercana
}


def _tallas_de_consulta(consulta):
    """Producto_Talla que matchean un SKU exacto o un texto del artículo.

    Devuelve una LISTA y no un queryset: la versión anterior retornaba un
    queryset ya rebanado (`[:400]`) y cualquier `.first()` posterior reventaba
    con "Cannot reorder a query once a slice has been taken" — un 500 en la
    búsqueda por texto, que es el camino más usado con productos legacy.
    """
    campos = ('producto', 'producto__atributo1')
    if consulta.isdigit():
        try:
            exactas = list(Producto_Talla.objects.filter(sku=int(consulta))
                           .select_related(*campos)[:400])
        except (ValueError, OverflowError):
            exactas = []
        if exactas:
            return exactas
    return list(
        Producto_Talla.objects
        .filter(Q(producto__articulo__icontains=consulta) |
                Q(producto__descripcion__icontains=consulta))
        .select_related(*campos)[:400]
    )


def _filtro_alcance_dte(queryset, user, prefijo='dte__'):
    """Acota a las empresas/sucursales del usuario.

    La compra histórica viene con `sucursal` NULL, así que anclar solo en
    sucursal la haría desaparecer: se acepta también por empresa receptora.
    """
    from .utils_permisos import ids_empresas_alcance, ids_sucursales_alcance
    empresas_ids = ids_empresas_alcance(user)
    if empresas_ids is None:
        return queryset
    sucursales_ids = ids_sucursales_alcance(user) or []
    return queryset.filter(
        Q(**{f'{prefijo}receptor_id__in': empresas_ids}) |
        Q(**{f'{prefijo}sucursal_id__in': sucursales_ids})
    )


def _candidato(fuente, *, dte=None, fecha=None, proveedor_id=None, proveedor='',
               sucursal='—', sku='', articulo='', talla='', cantidad=None,
               costo=0, nota=''):
    """Normaliza un hallazgo de cualquiera de las cuatro fuentes."""
    return {
        'fuente': fuente,
        'confianza': CONFIANZA_FUENTE.get(fuente, 'BAJA'),
        'dte_id': dte.id if dte else None,
        'numero_documento': dte.numero_documento if dte else '',
        'tipo_documento': dte.get_tipo_documento_display() if dte else '',
        'fecha': fecha.strftime('%Y-%m-%d') if fecha else '',
        'fecha_texto': fecha.strftime('%d/%m/%Y') if fecha else '',
        'proveedor_id': proveedor_id,
        'proveedor': proveedor,
        'sucursal': sucursal,
        'sku': sku,
        'articulo': articulo,
        'talla': talla,
        'cantidad': cantidad,
        'costo_unitario': costo or 0,
        'nota': nota,
    }


@login_required
@require_GET
def sugerir_proveedor_por_sku(request):
    """Proveedor y factura de compra más probables para un SKU.

    La tienda no sabe a qué proveedor reclamarle y por eso el campo llega
    vacío al analista. Se busca en el mismo orden de confianza que el buscador
    completo, así que ahora también responde con datos migrados desde Laravel
    (antes devolvía "no hay compras registradas" para todo el catálogo
    histórico) y, en última instancia, propone el proveedor por la MARCA.
    """
    sku = (request.GET.get('sku') or '').strip()
    if not sku:
        return JsonResponse({'success': False, 'error': 'SKU es requerido'}, status=400)
    # `Requerimiento.sku` es texto libre y `Producto_Talla.sku` es numerico:
    # un SKU alfanumerico llegaba al ORM y devolvia 500 en vez de "no existe".
    if not sku.isdigit():
        return JsonResponse({
            'success': False,
            'error': f'"{sku}" no es un SKU del catálogo (debe ser numérico)'
        }, status=404)

    from .utils_producto_match import producto_talla_por_sku
    producto_talla = producto_talla_por_sku(
        sku, sucursal_id=request.session.get('idSucursalActual'),
        select_related=['producto', 'producto__atributo1'])
    if not producto_talla:
        return JsonResponse({
            'success': False,
            'error': 'El SKU no existe en el sistema'
        }, status=404)

    # 1) Línea de DTE de compra: la fuente que nombra el SKU.
    # Mismo alcance que el buscador completo: sin esto la sugerencia filtraba
    # proveedor, N° de factura y costo de compras de OTRA empresa del holding.
    linea = (
        _filtro_alcance_dte(
            Dte_Productos.objects
            .filter(productoTalla=producto_talla, dte__tipo_transaccion='COMPRA')
            .select_related('dte', 'dte__emisor'),
            request.user)
        .order_by('-dte__fecha_emision', '-dte__id')
        .first()
    )
    if linea and linea.dte.emisor_id:
        dte = linea.dte
        return JsonResponse({
            'success': True,
            'fuente': 'LINEA_DTE',
            'confianza': 'ALTA',
            'proveedor': {'id': dte.emisor_id, 'nombre': dte.emisor.nombre,
                          'rut': dte.emisor.rut},
            'compra': {
                'dte_id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.get_tipo_documento_display(),
                'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d') if dte.fecha_emision else '',
                'costo_unitario': linea.costo or 0,
            },
        })

    # 2) Lote FIFO con documento: cuando el DTE existe pero sin líneas.
    lote = (
        _filtro_alcance_dte(
            LoteProducto.objects
            .filter(producto_talla=producto_talla, dte__isnull=False,
                    dte__tipo_transaccion='COMPRA')
            .select_related('dte', 'dte__emisor'),
            request.user)
        .order_by('-fecha_ingreso')
        .first()
    )
    if lote and lote.dte.emisor_id:
        dte = lote.dte
        return JsonResponse({
            'success': True,
            'fuente': 'LOTE',
            'confianza': 'ALTA',
            'proveedor': {'id': dte.emisor_id, 'nombre': dte.emisor.nombre,
                          'rut': dte.emisor.rut},
            'compra': {
                'dte_id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.get_tipo_documento_display(),
                'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d') if dte.fecha_emision else '',
                'costo_unitario': lote.costo_unitario or 0,
            },
        })

    # 3) Kardex con DTE asociado (ingreso por compra sin línea de detalle).
    mov = (
        _filtro_alcance_dte(
            Movimientos_Producto.objects
            .filter(ProductoTalla=producto_talla, dte__isnull=False,
                    dte__tipo_transaccion='COMPRA')
            .select_related('dte', 'dte__emisor'),
            request.user)
        .order_by('-fecha', '-hora')
        .first()
    )
    if mov and mov.dte.emisor_id:
        dte = mov.dte
        return JsonResponse({
            'success': True,
            'fuente': 'KARDEX_DTE',
            'confianza': 'MEDIA',
            'proveedor': {'id': dte.emisor_id, 'nombre': dte.emisor.nombre,
                          'rut': dte.emisor.rut},
            'compra': {
                'dte_id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.get_tipo_documento_display(),
                'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d') if dte.fecha_emision else '',
                'costo_unitario': mov.costo or 0,
            },
        })

    # 4) Último recurso: el proveedor de la MARCA. No da la factura, pero al
    #    menos evita que el analista tenga que adivinar a quién reclamarle.
    proveedor_marca = _proveedor_por_marca(producto_talla)
    if proveedor_marca:
        return JsonResponse({
            'success': True,
            'fuente': 'MARCA',
            'confianza': 'BAJA',
            'proveedor': {'id': proveedor_marca.id, 'nombre': proveedor_marca.nombre,
                          'rut': proveedor_marca.rut},
            'compra': None,
            'aviso': ('Sugerido por la marca del producto, no por una compra. '
                      'Confirme la factura en el buscador.'),
        })

    return JsonResponse({
        'success': False,
        'error': 'No hay compras registradas de este SKU en ninguna fuente'
    }, status=404)


def _proveedor_por_marca(producto_talla):
    """Proveedor cuyo nombre coincide con la marca del producto.

    En este catálogo la marca ES el proveedor casi siempre (NIKE → NIKE DE
    CHILE S.A.). No es prueba de nada, pero cuando no hay ni una compra
    rastreable es mejor que dejar el campo vacío.
    """
    producto = getattr(producto_talla, 'producto', None)
    marca = getattr(getattr(producto, 'atributo1', None), 'valor', None) if producto else None
    if not marca or len(marca.strip()) < 3:
        return None
    marca = marca.strip()
    return (
        Empresa.objects.filter(esProveedor=True)
        .filter(Q(nombre__istartswith=marca) | Q(nombre__icontains=marca))
        .order_by('nombre')
        .first()
    )


@login_required
@require_GET
def buscar_compras_producto(request):
    """Dónde entró este producto, cruzando las cuatro fuentes disponibles.

    Un mismo modelo se compra varias veces, a veces a proveedores distintos, y
    el par que falló pertenece a una partida concreta. Acá se listan los
    candidatos —fecha, proveedor, documento, cantidad y costo— con la fuente y
    la confianza de cada uno, para que quien revisa elija en vez de adivinar.

    Busca por SKU exacto o por texto del artículo/descripción.
    """
    consulta = (request.GET.get('q') or '').strip()
    if len(consulta) < 2:
        return JsonResponse({
            'success': False,
            'error': 'Ingrese al menos 2 caracteres (SKU o nombre del artículo)'
        }, status=400)

    try:
        # `max(1, ...)`: un `limite` negativo del query string llegaba al
        # slice del queryset y Django respondía 500.
        limite = max(1, min(int(request.GET.get('limite', 40)), 100))
    except (TypeError, ValueError):
        limite = 40

    tallas = _tallas_de_consulta(consulta)
    ids_tallas = [t.id for t in tallas]
    if not ids_tallas:
        return JsonResponse({
            'success': True, 'compras': [], 'total': 0,
            'diagnostico': {
                'mensaje': f'Ningún producto del catálogo coincide con "{consulta}".',
                'sugerencia': 'Pruebe con el SKU exacto o con parte del nombre del artículo.',
            },
        })

    candidatos = []
    vistos_dte = set()          # dte_id ya incorporado (la fuente más confiable gana)
    fechas_ingreso = []         # para acotar la búsqueda por cabecera
    proveedores_detectados = {}

    # ── Fuente 1: líneas de DTE de compra ────────────────────────────────
    lineas = _filtro_alcance_dte(
        Dte_Productos.objects
        .filter(productoTalla_id__in=ids_tallas, dte__tipo_transaccion='COMPRA')
        .select_related('dte', 'dte__emisor', 'dte__sucursal',
                        'productoTalla', 'productoTalla__producto'),
        request.user,
    ).order_by('-dte__fecha_emision', '-dte__id')[:limite]

    for linea in lineas:
        dte, talla = linea.dte, linea.productoTalla
        vistos_dte.add(dte.id)
        if dte.emisor_id:
            proveedores_detectados[dte.emisor_id] = dte.emisor.nombre
        if dte.fecha_emision:
            fechas_ingreso.append(dte.fecha_emision)
        candidatos.append(_candidato(
            'LINEA_DTE', dte=dte, fecha=dte.fecha_emision,
            proveedor_id=dte.emisor_id,
            proveedor=dte.emisor.nombre if dte.emisor_id else '',
            sucursal=dte.sucursal.alias if dte.sucursal_id else '—',
            sku=talla.sku if talla else '',
            articulo=(talla.producto.articulo if talla and talla.producto_id
                      else linea.descripcion),
            talla=talla.talla if talla else '',
            cantidad=linea.stock, costo=linea.costo or 0,
            nota='El documento detalla este SKU',
        ))

    # ── Fuente 2: lotes FIFO con documento ───────────────────────────────
    lotes = _filtro_alcance_dte(
        LoteProducto.objects
        .filter(producto_talla_id__in=ids_tallas, dte__isnull=False,
                dte__tipo_transaccion='COMPRA')
        .select_related('dte', 'dte__emisor', 'dte__sucursal',
                        'producto_talla', 'producto_talla__producto'),
        request.user,
    ).order_by('-fecha_ingreso')[:limite]

    for lote in lotes:
        dte, talla = lote.dte, lote.producto_talla
        if dte.id in vistos_dte:
            continue
        vistos_dte.add(dte.id)
        if dte.emisor_id:
            proveedores_detectados[dte.emisor_id] = dte.emisor.nombre
        if dte.fecha_emision:
            fechas_ingreso.append(dte.fecha_emision)
        candidatos.append(_candidato(
            'LOTE', dte=dte, fecha=dte.fecha_emision,
            proveedor_id=dte.emisor_id,
            proveedor=dte.emisor.nombre if dte.emisor_id else '',
            sucursal=dte.sucursal.alias if dte.sucursal_id else '—',
            sku=talla.sku if talla else '',
            articulo=talla.producto.articulo if talla and talla.producto_id else '',
            talla=talla.talla if talla else '',
            cantidad=lote.cantidad_inicial, costo=lote.costo_unitario or 0,
            nota='Lote FIFO ligado a este documento',
        ))

    # ── Fuente 3: N° recuperado del sistema anterior ─────────────────────
    # La migración desde Laravel leía `movimiento_productos.N_documento` y no
    # lo guardaba. `backfill_documento_movimientos_laravel` lo rescata en su
    # propia tabla —sin tocar el kardex ni los DTE, para no mover ningún
    # reporte— y acá se ofrece como candidato. Para el proveedor, el número
    # ES la respuesta, exista o no la cabecera en el sistema.
    #
    # Va ANTES del kardex a propósito: los movimientos ya rescatados no deben
    # repetirse como "ingreso sin documento" dos filas más abajo.
    legacy = (
        DocumentoCompraLegacy.objects
        .filter(producto_talla_id__in=ids_tallas)
        .select_related('dte', 'proveedor', 'sucursal')
        .order_by('-fecha_movimiento')[:limite]
    )
    movimientos_rescatados = set()
    for doc in legacy:
        movimientos_rescatados.add(doc.movimiento_origen_id)
        if doc.dte_id and doc.dte_id in vistos_dte:
            continue
        if doc.dte_id:
            vistos_dte.add(doc.dte_id)
        if doc.proveedor_id:
            proveedores_detectados[doc.proveedor_id] = doc.proveedor.nombre
        if doc.fecha_movimiento:
            fechas_ingreso.append(doc.fecha_movimiento)

        candidato = _candidato(
            'LEGACY', dte=doc.dte, fecha=doc.fecha_movimiento,
            proveedor_id=doc.proveedor_id,
            proveedor=doc.proveedor.nombre if doc.proveedor_id else (doc.marca_legacy or ''),
            sucursal=doc.sucursal.alias if doc.sucursal_id else '—',
            sku=doc.sku, cantidad=doc.cantidad, costo=doc.costo,
            nota=('N° recuperado del sistema anterior' if not doc.dte_id
                  else 'N° del sistema anterior, con su documento identificado'),
        )
        # El número se muestra SIEMPRE, haya o no cabecera en la BD: es
        # exactamente el dato que el proveedor pide.
        candidato['numero_documento'] = doc.numero_documento
        if not doc.dte_id:
            candidato['tipo_documento'] = 'Factura de compra'
        candidatos.append(candidato)

    # ── Fuente 4: kardex (ingresos que no quedaron rescatados) ───────────
    movimientos = (
        Movimientos_Producto.objects
        .filter(ProductoTalla_id__in=ids_tallas,
                concepto__in=CONCEPTOS_INGRESO_COMPRA,
                cantidad__gt=0)
        .select_related('dte', 'dte__emisor', 'sucursal_destino', 'sucursal_origen',
                        'ProductoTalla', 'ProductoTalla__producto')
    )
    # Mismo alcance que las otras fuentes. El kardex no cuelga de un DTE, así
    # que se acota por la sucursal del movimiento: sin esto un usuario de una
    # empresa veía los ingresos de otra.
    from .utils_permisos import ids_sucursales_alcance
    sucursales_alcance = ids_sucursales_alcance(request.user)
    if sucursales_alcance is not None:
        movimientos = movimientos.filter(
            Q(sucursal_destino_id__in=sucursales_alcance) |
            Q(sucursal_origen_id__in=sucursales_alcance)
        )
    # Los que SÍ apuntan a un DTE van primero: ordenando solo por fecha, el
    # corte por `limite` dejaba fuera justamente los movimientos rastreables
    # cuando había cientos de ingresos sin documento más nuevos.
    movimientos = movimientos.annotate(
        _sin_dte=Case(When(dte__isnull=True, then=Value(1)), default=Value(0),
                      output_field=IntegerField())
    ).order_by('_sin_dte', '-fecha', '-hora')[:limite]

    movimientos_sin_documento = 0
    for mov in movimientos:
        # Si su N° ya se rescató arriba, repetirlo acá como "sin documento"
        # sería mostrar dos veces la misma unidad y contradecirse.
        if mov.id in movimientos_rescatados:
            continue
        talla = mov.ProductoTalla
        # La sucursal del ingreso: destino si existe, si no origen (el mismo
        # criterio que usa el reporte de movimientos).
        sucursal = mov.sucursal_destino or mov.sucursal_origen
        if mov.dte_id and mov.dte.tipo_transaccion == 'COMPRA':
            if mov.dte_id in vistos_dte:
                continue
            vistos_dte.add(mov.dte_id)
            if mov.dte.emisor_id:
                proveedores_detectados[mov.dte.emisor_id] = mov.dte.emisor.nombre
            if mov.dte.fecha_emision:
                fechas_ingreso.append(mov.dte.fecha_emision)
            candidatos.append(_candidato(
                'KARDEX_DTE', dte=mov.dte, fecha=mov.dte.fecha_emision or mov.fecha,
                proveedor_id=mov.dte.emisor_id,
                proveedor=mov.dte.emisor.nombre if mov.dte.emisor_id else '',
                sucursal=sucursal.alias if sucursal else '—',
                sku=talla.sku if talla else '',
                articulo=talla.producto.articulo if talla and talla.producto_id else '',
                talla=talla.talla if talla else '',
                cantidad=mov.cantidad, costo=mov.costo or 0,
                nota='El movimiento apunta al documento (sin línea de detalle)',
            ))
            continue

        # Ingreso sin FK al DTE: no da la factura, pero sí la FECHA y la
        # SUCURSAL en que entró, que es lo que permite ubicar la compra del
        # proveedor más abajo. El N° recuperado del sistema anterior se busca
        # aparte (fuente 4), en su propia tabla.
        if mov.fecha:
            fechas_ingreso.append(mov.fecha)
        movimientos_sin_documento += 1
        migrado = (mov.referencia_externa or '').startswith('MIG:')
        candidatos.append(_candidato(
            'KARDEX_SIN_DOC', fecha=mov.fecha,
            sucursal=sucursal.alias if sucursal else '—',
            sku=talla.sku if talla else '',
            articulo=talla.producto.articulo if talla and talla.producto_id else '',
            talla=talla.talla if talla else '',
            cantidad=mov.cantidad, costo=mov.costo or 0,
            nota=('Ingreso migrado del sistema anterior: la migración no trajo '
                  'el N° de documento' if migrado
                  else f'Ingreso por {mov.get_concepto_display()} sin documento asociado'),
        ))

    # ── Fuente 5: compras del proveedor en fechas cercanas ───────────────
    # Cuando el ingreso existe pero perdió el documento, la factura suele ser
    # una de las compras que ese proveedor emitió alrededor de esa fecha.
    if movimientos_sin_documento and fechas_ingreso:
        proveedor_id = request.GET.get('proveedor_id')
        ids_proveedor = ([int(proveedor_id)] if proveedor_id and proveedor_id.isdigit()
                         else list(proveedores_detectados.keys()))
        if not ids_proveedor:
            proveedor_marca = _proveedor_por_marca(tallas[0] if tallas else None)
            if proveedor_marca:
                ids_proveedor = [proveedor_marca.id]

        if ids_proveedor:
            # Ventana alrededor del ingreso MÁS RECIENTE. Usar min()/max()
            # global abría un rango de años y devolvía las compras más nuevas
            # del proveedor, que no tienen nada que ver con este producto.
            referencia = max(fechas_ingreso)
            desde = referencia - timedelta(days=45)
            hasta = referencia + timedelta(days=45)
            cabeceras = _filtro_alcance_dte(
                Dte.objects.filter(
                    tipo_transaccion='COMPRA',
                    emisor_id__in=ids_proveedor,
                    fecha_emision__range=(desde, hasta),
                ).exclude(id__in=vistos_dte)
                .select_related('emisor', 'sucursal'),
                request.user, prefijo='',
            ).order_by('-fecha_emision')[:15]

            for dte in cabeceras:
                candidatos.append(_candidato(
                    'CABECERA_PROVEEDOR', dte=dte, fecha=dte.fecha_emision,
                    proveedor_id=dte.emisor_id,
                    proveedor=dte.emisor.nombre if dte.emisor_id else '',
                    sucursal=dte.sucursal.alias if dte.sucursal_id else '—',
                    cantidad=dte.unidades_productos, costo=0,
                    nota='Compra del proveedor en fecha cercana al ingreso — verifique',
                ))

    # Más confiable primero y, dentro de cada nivel, lo más reciente.
    orden_confianza = {'ALTA': 0, 'MEDIA': 1, 'BAJA': 2}
    candidatos.sort(
        key=lambda c: (orden_confianza.get(c['confianza'], 3), -_orden_fecha(c['fecha'])))
    candidatos = candidatos[:limite]

    diagnostico = _diagnostico_busqueda(candidatos, movimientos_sin_documento, consulta)

    return JsonResponse({
        'success': True,
        'compras': candidatos,
        'total': len(candidatos),
        'diagnostico': diagnostico,
    })


def _orden_fecha(fecha_iso):
    """Fecha ISO como entero para ordenar descendente sin parsear."""
    return int((fecha_iso or '0000-00-00').replace('-', ''))


def _diagnostico_busqueda(candidatos, sin_documento, consulta):
    """Explica QUÉ se encontró y qué hacer, en vez de una lista muda."""
    if not candidatos:
        return {
            'mensaje': f'No hay ningún ingreso registrado para "{consulta}".',
            'sugerencia': ('Escriba el N° de factura a mano si lo tiene en el '
                           'documento físico, o valide el caso sin respaldo '
                           'dejando la razón en las notas internas.'),
            'nivel': 'warning',
        }
    # Lo que cuenta es tener el NÚMERO, no que exista la cabecera en la BD:
    # al proveedor se le manda el número. Mirar solo `dte_id` hacía que el
    # buscador dijera "ninguno conserva el N°" con el número en pantalla.
    con_documento = [c for c in candidatos if c['numero_documento']]
    if not con_documento:
        return {
            'mensaje': (f'Hay {sin_documento} ingreso(s) de este producto, pero '
                        f'ninguno conserva el N° de documento.'),
            'sugerencia': ('Son datos traídos del sistema anterior: use la fecha '
                           'del ingreso para ubicar la factura del proveedor.'),
            'nivel': 'warning',
        }
    altas = [c for c in con_documento if c['confianza'] == 'ALTA']
    if altas:
        return {
            'mensaje': f'{len(altas)} factura(s) de compra identificadas para este producto.',
            'sugerencia': 'Elija la partida que corresponde a la unidad que falló.',
            'nivel': 'success',
        }
    return {
        'mensaje': 'No hay un documento que detalle el SKU; estos son los más probables.',
        'sugerencia': 'Verifique contra el documento físico antes de reclamarlo.',
        'nivel': 'info',
    }


@login_required
@require_GET
def buscar_dte_compra_por_numero(request):
    """Valida un N° de factura de compra tipeado a mano.

    El analista suele tener el número en el documento físico. Antes lo
    escribía a ciegas y nadie comprobaba que existiera ni de qué proveedor
    era; con esto, si el DTE está en el sistema, quedan enlazados el
    proveedor, la fecha y el documento de una sola vez.
    """
    numero = (request.GET.get('numero') or '').strip()
    if not numero.isdigit():
        return JsonResponse({
            'success': False,
            'error': 'Ingrese el número de documento (solo dígitos)'
        }, status=400)

    dtes = _filtro_alcance_dte(
        Dte.objects.filter(numero_documento=int(numero), tipo_transaccion='COMPRA')
        .select_related('emisor', 'sucursal'),
        request.user, prefijo='',
    ).order_by('-fecha_emision')[:10]

    if not dtes:
        return JsonResponse({
            'success': True, 'encontrados': [],
            'aviso': (f'No hay ninguna factura de compra N° {numero} en el sistema. '
                      f'Puede usarla igual: se guardará como dato tipeado.'),
        })

    return JsonResponse({
        'success': True,
        'encontrados': [{
            'dte_id': d.id,
            'numero_documento': d.numero_documento,
            'tipo_documento': d.get_tipo_documento_display(),
            'fecha': d.fecha_emision.strftime('%Y-%m-%d') if d.fecha_emision else '',
            'fecha_texto': d.fecha_emision.strftime('%d/%m/%Y') if d.fecha_emision else '',
            'proveedor_id': d.emisor_id,
            'proveedor': d.emisor.nombre if d.emisor_id else '',
            'sucursal': d.sucursal.alias if d.sucursal_id else '—',
            'monto': int(d.monto_con_iva or 0),
        } for d in dtes],
    })


# ========== API TIPOS DE FOTO ==========

@login_required
@require_GET
def obtener_tipos_foto(request):
    """Retorna los tipos de foto requeridos/opcionales para un tipo de requerimiento"""
    tipo = request.GET.get('tipo', '')
    if not tipo:
        return JsonResponse({'success': False, 'error': 'Tipo es requerido'}, status=400)

    tipos_foto = TipoFotoRequerimiento.tipos_para(tipo)

    return JsonResponse({
        'success': True,
        'tipos_foto': [
            {
                'codigo': tf.codigo,
                'nombre': tf.nombre,
                'descripcion_guia': tf.descripcion_guia,
                'icono': tf.icono,
                'es_obligatorio': tf.es_obligatorio,
                'orden': tf.orden,
            }
            for tf in tipos_foto
        ],
        'max_fotos': MAX_FOTOS_POR_TIPO.get(tipo, 5),
    })
