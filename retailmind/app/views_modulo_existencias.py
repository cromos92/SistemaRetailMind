"""
Módulo de Existencias - RetailMind
Contiene todas las vistas relacionadas con productos, inventario, FIFO, lotes, movimientos y gestión de stock
"""

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET
from django.db.models import Sum, F, Q
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
import json
from decimal import Decimal

from .models import (
    Producto, Producto_Talla, LoteProducto, Movimientos_Producto, Sucursal,
    AjusteInventario, AjusteInventario_Detalle, Productos_Recepcionados
)
# LIMPIEZA DE IMPORTS 2026-08-22 (Fase D, F-06). Se fueron 23 nombres que el
# módulo importaba y no referenciaba en ninguna parte — verificado por AST
# (ningún `ast.Name` los usa) y confirmado línea a línea: salvo su propia línea
# de import, no aparecen en el archivo.
#   - por el borrado de las 3 FBV de traspaso: Traspaso, Traspaso_Detalle
#   - ya muertos de antes: redirect, Http404, HttpResponseBadRequest,
#     require_http_methods, csrf_exempt, ExpressionWrapper, DecimalField,
#     Count, Avg, re, Productos_Atributos, AtributoOpcion, Categoria,
#     GuiaTalla, GuiaTallaItem, GuiaTallaProducto, EmpresaUser, Empresa,
#     Ticket, Dte, Font, PatternFill, Alignment (estos 3 últimos, en el
#     import local de openpyxl más abajo).
# `Ticket_Productos` NO estaba en este import: se importa localmente dentro de
# la función que lo usa.


# ========== GESTIÓN DE PRODUCTOS ==========

# NOTA 2026-07: aquí vivía una copia muerta de `verGestionProducto` que
# renderizaba verGestionProductos.html SIN el contexto de atributos
# (id_atributo_marca / color / genero / especialidad), de modo que quien
# editara esta copia por error dejaba el modal "Crear Producto Manual" sin sus
# select2. La versión viva y ruteada es `views.verGestionProducto`
# (app/urls.py línea 516); este módulo ni siquiera está importado por urls.py.


@require_GET
@login_required
def obtener_productos_base(request):
    """Obtener lista base de productos para reportes"""
    try:
        # Parámetros de filtro
        search = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria_id')
        marca_id = request.GET.get('marca_id')
        
        # Construir queryset
        queryset = Producto.objects.select_related('categoria', 'marca')
        
        # Aplicar filtros
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo__icontains=search) |
                Q(marca__nombre__icontains=search)
            )
        
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)
        
        if marca_id:
            queryset = queryset.filter(marca_id=marca_id)
        
        # Ordenar
        queryset = queryset.order_by('nombre')
        
        # Limitar resultados
        queryset = queryset[:100]
        
        productos_data = []
        for producto in queryset:
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.marca.nombre if producto.marca else '',
                'activo': producto.activo
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos: {str(e)}'
        })


@require_GET
@login_required
def obtener_productos(request):
    """Obtener productos con filtros avanzados y paginación"""
    try:
        # Parámetros de paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Parámetros de filtro
        search = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria_id')
        marca_id = request.GET.get('marca_id')
        estado = request.GET.get('estado')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Función para limpiar prefijos
        def limpiar_prefijo(valor):
            if valor and valor.startswith('categoria_'):
                return valor.replace('categoria_', '')
            elif valor and valor.startswith('marca_'):
                return valor.replace('marca_', '')
            return valor
        
        categoria_id = limpiar_prefijo(categoria_id)
        marca_id = limpiar_prefijo(marca_id)
        
        # Construir queryset
        queryset = Producto.objects.select_related('categoria', 'marca').prefetch_related('producto_talla_set')
        
        # Aplicar filtros
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo__icontains=search) |
                Q(marca__nombre__icontains=search) |
                Q(categoria__nombre__icontains=search)
            )
        
        if categoria_id and categoria_id.isdigit():
            queryset = queryset.filter(categoria_id=int(categoria_id))
        
        if marca_id and marca_id.isdigit():
            queryset = queryset.filter(marca_id=int(marca_id))
        
        if estado:
            queryset = queryset.filter(activo=(estado == 'activo'))
        
        if fecha_inicio:
            queryset = queryset.filter(fecha_creacion__date__gte=fecha_inicio)
        
        if fecha_fin:
            queryset = queryset.filter(fecha_creacion__date__lte=fecha_fin)
        
        # Ordenar
        queryset = queryset.order_by('-fecha_creacion')
        
        # Paginación
        paginator = Paginator(queryset, per_page)
        productos_page = paginator.get_page(page)
        
        # Serializar datos
        productos_data = []
        for producto in productos_page:
            # Calcular stock total
            stock_total = sum(
                pt.stock_total() for pt in producto.producto_talla_set.all()
            )
            
            # Calcular tallas disponibles
            tallas_count = producto.producto_talla_set.count()
            
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.marca.nombre if producto.marca else '',
                'stock_total': stock_total,
                'tallas_count': tallas_count,
                'activo': producto.activo,
                'fecha_creacion': producto.fecha_creacion.strftime('%d/%m/%Y') if producto.fecha_creacion else ''
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'pagination': {
                'current_page': productos_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': productos_page.has_next(),
                'has_previous': productos_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos: {str(e)}'
        })


def obtener_productos_para_crear(request):
    """Obtener productos disponibles para crear desde recepciones"""
    try:
        # Obtener productos recepcionados que no tienen producto creado
        productos_pendientes = Productos_Recepcionados.objects.filter(
            producto_creado=False
        ).select_related('compra', 'compra__empresa').order_by('-fecha_recepcion')
        
        productos_data = []
        for item in productos_pendientes:
            productos_data.append({
                'id': item.id,
                'nombre_producto': item.nombre_producto,
                'marca': item.marca or '',
                'categoria': item.categoria or '',
                'proveedor': item.compra.empresa.nombre,
                'cantidad_recepcionada': item.cantidad_recepcionada,
                'costo_unitario': float(item.costo_unitario),
                'precio_venta_sugerido': float(item.precio_venta_sugerido),
                'fecha_recepcion': item.fecha_recepcion.strftime('%d/%m/%Y')
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos: {str(e)}'
        })


def detalle_producto_para_crear(request, producto_id):
    """Obtener detalles de un producto para crear"""
    try:
        producto_recepcionado = get_object_or_404(Productos_Recepcionados, id=producto_id)
        
        producto_data = {
            'id': producto_recepcionado.id,
            'nombre_producto': producto_recepcionado.nombre_producto,
            'marca': producto_recepcionado.marca or '',
            'categoria': producto_recepcionado.categoria or '',
            'proveedor': producto_recepcionado.compra.empresa.nombre,
            'cantidad_recepcionada': producto_recepcionado.cantidad_recepcionada,
            'costo_unitario': float(producto_recepcionado.costo_unitario),
            'precio_venta_sugerido': float(producto_recepcionado.precio_venta_sugerido),
            'observaciones': producto_recepcionado.observaciones or '',
            'atributos': producto_recepcionado.atributos_json or {}
        }
        
        return JsonResponse({
            'success': True,
            'producto': producto_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener producto: {str(e)}'
        })


# NOTA 2026-08-22 (F-06, Fase D de la auditoría de Reportes): aquí vivían dos
# copias muertas — `crear_producto_desde_recepcion` y `crear_producto_manual` —
# que llamaban directo a `views.crear_producto`. Las vivas y ruteadas son
# `views.crear_producto_desde_recepcion` (app/urls.py:706 → app/views.py:20880) y
# `views.crear_producto_manual` (app/urls.py:783 → app/views.py:22841); el
# template verGestionProductos.html pega a /app/crear_producto_desde_recepcion/
# y /app/crear_producto_manual/, que resuelven a esas dos. Este módulo no lo
# importa NADIE (0 `from .views_modulo_existencias import` en todo el repo), así
# que editar estas copias no tenía efecto alguno en producción.


@require_GET
@login_required
def buscar_productos_existentes(request):
    """Buscar productos existentes para copiar o referenciar"""
    try:
        termino = request.GET.get('termino', '').strip()
        
        if not termino:
            return JsonResponse({
                'success': False,
                'error': 'Término de búsqueda requerido'
            })
        
        # Buscar productos
        productos = Producto.objects.filter(
            Q(nombre__icontains=termino) |
            Q(codigo__icontains=termino) |
            Q(marca__nombre__icontains=termino)
        ).select_related('categoria', 'marca')[:20]
        
        productos_data = []
        for producto in productos:
            # Calcular stock total
            stock_total = sum(
                pt.stock_total() for pt in producto.producto_talla_set.all()
            )
            
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.marca.nombre if producto.marca else '',
                'stock_total': stock_total,
                'activo': producto.activo
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en búsqueda: {str(e)}'
        })


@require_GET
@login_required
def detalle_producto_para_copiar(request, producto_id):
    """Obtener detalles de un producto para copiar"""
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        
        # Obtener atributos
        atributos = []
        for attr in producto.productos_atributos_set.all():
            atributos.append({
                'atributo_id': attr.atributo.id,
                'atributo_nombre': attr.atributo.nombre,
                'opcion_id': attr.opcion.id if attr.opcion else None,
                'opcion_nombre': attr.opcion.nombre if attr.opcion else '',
                'valor_personalizado': attr.valor_personalizado or ''
            })
        
        # Obtener tallas
        tallas = []
        for pt in producto.producto_talla_set.all():
            tallas.append({
                'id': pt.id,
                'talla_id': pt.talla.id if pt.talla else None,
                'talla_nombre': pt.talla.nombre if pt.talla else 'Sin talla',
                'sku': pt.sku,
                'precio_venta': float(pt.precio_venta),
                'stock_total': pt.stock_total()
            })
        
        producto_data = {
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'descripcion': producto.descripcion or '',
            'categoria_id': producto.categoria.id if producto.categoria else None,
            'categoria_nombre': producto.categoria.nombre if producto.categoria else '',
            'marca_id': producto.marca.id if producto.marca else None,
            'marca_nombre': producto.marca.nombre if producto.marca else '',
            'atributos': atributos,
            'tallas': tallas
        }
        
        return JsonResponse({
            'success': True,
            'producto': producto_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener producto: {str(e)}'
        })


@require_GET
@login_required
def tallas_producto(request, producto_id):
    """Obtener tallas de un producto específico"""
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        sucursal_id = request.session.get('idSucursalActual')
        
        tallas_data = []
        for pt in producto.producto_talla_set.all():
            stock_sucursal = pt.stock_sucursal(sucursal_id) if sucursal_id else 0
            stock_total = pt.stock_total()
            
            tallas_data.append({
                'id': pt.id,
                'talla_id': pt.talla.id if pt.talla else None,
                'talla_nombre': pt.talla.nombre if pt.talla else 'Sin talla',
                'sku': pt.sku,
                'precio_venta': float(pt.precio_venta),
                'stock_sucursal': stock_sucursal,
                'stock_total': stock_total,
                'activo': pt.activo
            })
        
        return JsonResponse({
            'success': True,
            'tallas': tallas_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener tallas: {str(e)}'
        })


# ========== MOVIMIENTOS DE PRODUCTOS ==========

@login_required
def verMovimientosProducto(request):
    """Vista principal para movimientos de productos"""
    return render(request, 'vistas/modulo_existencias/gestionMovimientos.html')


@require_GET
@login_required
def obtener_movimientos_producto(request):
    """Obtener movimientos de productos con filtros"""
    try:
        # Parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        producto_id = request.GET.get('producto_id')
        concepto = request.GET.get('concepto')
        sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')
        
        # Función para parsear fecha
        def parse_fecha_ddmmyyyy(fecha_str):
            if not fecha_str:
                return None
            try:
                from datetime import datetime
                return datetime.strptime(fecha_str, '%d/%m/%Y').date()
            except ValueError:
                try:
                    return datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except ValueError:
                    return None
        
        # Construir queryset
        queryset = Movimientos_Producto.objects.select_related(
            'ProductoTalla__producto', 'sucursal_origen', 'sucursal_destino'
        )
        
        # Aplicar filtros
        if fecha_inicio:
            fecha_inicio_parsed = parse_fecha_ddmmyyyy(fecha_inicio)
            if fecha_inicio_parsed:
                queryset = queryset.filter(fecha__gte=fecha_inicio_parsed)
        
        if fecha_fin:
            fecha_fin_parsed = parse_fecha_ddmmyyyy(fecha_fin)
            if fecha_fin_parsed:
                queryset = queryset.filter(fecha__lte=fecha_fin_parsed)
        
        if producto_id:
            queryset = queryset.filter(ProductoTalla__producto_id=producto_id)
        
        if concepto:
            queryset = queryset.filter(concepto=concepto)
        
        if sucursal_id:
            queryset = queryset.filter(
                Q(sucursal_origen_id=sucursal_id) | Q(sucursal_destino_id=sucursal_id)
            )
        
        # ✅ ORDEN: Más recientes primero (fecha + hora descendente)
        queryset = queryset.order_by('-fecha', '-hora')
        
        # Paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        paginator = Paginator(queryset, per_page)
        movimientos_page = paginator.get_page(page)
        
        # Serializar datos
        movimientos_data = []
        for mov in movimientos_page:
            # Combinar fecha y hora para mostrar
            fecha_hora = f"{mov.fecha.strftime('%d/%m/%Y')} {mov.hora.strftime('%H:%M')}" if mov.hora else mov.fecha.strftime('%d/%m/%Y')
            movimientos_data.append({
                'id': mov.id,
                'fecha_creacion': fecha_hora,
                'concepto': mov.concepto,
                'tipo_movimiento': mov.tipo_movimiento,
                'producto_nombre': mov.ProductoTalla.producto.articulo,
                'sku': mov.ProductoTalla.sku,
                'cantidad': mov.cantidad,
                'responsable': mov.responsable,
                'sucursal_origen': mov.sucursal_origen.alias if mov.sucursal_origen else '',
                'sucursal_destino': mov.sucursal_destino.alias if mov.sucursal_destino else '',
                'observaciones': mov.observaciones or '',
                'referencia_externa': mov.referencia_externa or ''
            })
        
        return JsonResponse({
            'success': True,
            'movimientos': movimientos_data,
            'pagination': {
                'current_page': movimientos_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': movimientos_page.has_next(),
                'has_previous': movimientos_page.has_previous(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener movimientos: {str(e)}'
        })


# NOTA 2026-08-22 (Fase D de la auditoría de Reportes): aquí vivían
# `crear_traspaso`, `aprobar_traspaso` y `recibir_traspaso`. Eran las ÚNICAS
# definiciones de esos tres nombres en el repo y sus rutas se eliminaron el
# 2026-07-28 (ver el comentario en app/urls.py:758): movían stock sin
# `@login_required` (POST anónimo) y `aprobar_traspaso` descontaba de nuevo el
# stock que `api_crear_despacho_masivo` ya había descontado. Quedaron sin ruta
# pero con el cuerpo intacto, listas para que alguien las volviera a colgar de
# una URL. Borradas. El circuito real de traspasos es por DTE (recepcion-dte/).


# ========== AJUSTES DE INVENTARIO ==========

@require_POST
@transaction.atomic
def crear_ajuste_inventario(request):
    """Crear ajuste de inventario"""
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        sucursal_id = data.get('sucursal_id') or request.session.get('idSucursalActual')
        tipo_ajuste = data.get('tipo_ajuste')  # 'ENTRADA' o 'SALIDA'
        productos = data.get('productos', [])
        motivo = data.get('motivo', '')
        
        if not all([sucursal_id, tipo_ajuste, productos]):
            return JsonResponse({
                'success': False,
                'error': 'Sucursal, tipo de ajuste y productos son requeridos'
            })
        
        if tipo_ajuste not in ['ENTRADA', 'SALIDA']:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de ajuste debe ser ENTRADA o SALIDA'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Crear ajuste
        ajuste = AjusteInventario.objects.create(
            sucursal=sucursal,
            tipo_ajuste=tipo_ajuste,
            motivo=motivo,
            usuario_responsable=request.user,
            observaciones=data.get('observaciones', '')
        )
        
        # Procesar productos
        from .views import registrar_movimiento_producto, crear_lote_producto, consumir_stock_fifo
        
        for item in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=item['producto_talla_id'])
            cantidad = int(item['cantidad'])
            
            # Crear detalle del ajuste
            AjusteInventario_Detalle.objects.create(
                ajuste=ajuste,
                producto_talla=producto_talla,
                cantidad=cantidad,
                observaciones=item.get('observaciones', '')
            )
            
            if tipo_ajuste == 'ENTRADA':
                # Registrar movimiento de entrada
                registrar_movimiento_producto(
                    producto_talla=producto_talla,
                    concepto='AJUSTE_ENTRADA',
                    cantidad=cantidad,
                    responsable=request.user,
                    observaciones=f'Ajuste inventario #{ajuste.id}: {motivo}',
                    referencia_externa=str(ajuste.id)
                )
                
                # Crear lote
                costo_unitario = Decimal(item.get('costo_unitario', 0)) or producto_talla.precio_venta * Decimal('0.7')
                
                crear_lote_producto(
                    producto_talla=producto_talla,
                    cantidad=cantidad,
                    costo_unitario=costo_unitario,
                    sobreprecio_unitario=0,
                    precio_venta_unitario=producto_talla.precio_venta,
                    observaciones=f'Ajuste inventario #{ajuste.id}: {motivo}'
                )
                
            else:  # SALIDA
                # Verificar stock disponible
                stock_disponible = producto_talla.stock_sucursal(sucursal_id)
                if stock_disponible < cantidad:
                    raise ValidationError(f'Stock insuficiente para {producto_talla.sku}')
                
                # Registrar movimiento de salida
                registrar_movimiento_producto(
                    producto_talla=producto_talla,
                    concepto='AJUSTE_SALIDA',
                    cantidad=-cantidad,  # Negativo para salida
                    responsable=request.user,
                    observaciones=f'Ajuste inventario #{ajuste.id}: {motivo}',
                    referencia_externa=str(ajuste.id)
                )
                
                # Consumir stock FIFO
                consumir_stock_fifo(
                    producto_talla=producto_talla,
                    cantidad_requerida=cantidad,
                    responsable=request.user,
                    observaciones=f'Ajuste inventario #{ajuste.id}: {motivo}',
                    referencia_externa=str(ajuste.id)
                )
        
        return JsonResponse({
            'success': True,
            'message': 'Ajuste de inventario creado exitosamente',
            'ajuste_id': ajuste.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear ajuste: {str(e)}'
        })


# ========== GESTIÓN DE LOTES FIFO ==========

@require_GET
@login_required
def ver_lotes_producto(request, producto_talla_id):
    """Vista para ver lotes de un producto específico"""
    try:
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        
        context = {
            'producto_talla': producto_talla,
            'producto': producto_talla.producto,
            'talla': producto_talla.talla
        }
        
        return render(request, 'vistas/modulo_existencias/lotes_producto.html', context)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        })


@require_GET
@login_required
def obtener_lotes_producto(request, producto_talla_id):
    """API para obtener lotes de un producto específico"""
    try:
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        
        # Parámetros de filtro
        estado = request.GET.get('estado', 'todos')
        
        # Construir queryset
        queryset = LoteProducto.objects.filter(
            producto_talla=producto_talla
        ).select_related('dte_origen', 'movimiento_origen')
        
        # Aplicar filtros
        if estado == 'activos':
            queryset = queryset.filter(cantidad_disponible__gt=0, activo=True)
        elif estado == 'agotados':
            queryset = queryset.filter(cantidad_disponible=0)
        elif estado == 'inactivos':
            queryset = queryset.filter(activo=False)
        
        # Ordenar por fecha de creación (FIFO)
        queryset = queryset.order_by('fecha_creacion')
        
        lotes_data = []
        for lote in queryset:
            lotes_data.append({
                'id': lote.id,
                'numero_lote': lote.numero_lote,
                'cantidad_inicial': lote.cantidad_inicial,
                'cantidad_disponible': lote.cantidad_disponible,
                'costo_unitario': float(lote.costo_unitario),
                'precio_venta_unitario': float(lote.precio_venta_unitario),
                'fecha_creacion': lote.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'fecha_vencimiento': lote.fecha_vencimiento.strftime('%d/%m/%Y') if lote.fecha_vencimiento else None,
                'dte_origen': lote.dte_origen.numero_dte if lote.dte_origen else None,
                'valor_inventario': float(lote.cantidad_disponible * lote.costo_unitario),
                'activo': lote.activo,
                'observaciones': lote.observaciones or ''
            })
        
        # Calcular totales
        total_cantidad = sum(lote['cantidad_disponible'] for lote in lotes_data)
        total_valor = sum(lote['valor_inventario'] for lote in lotes_data)
        costo_promedio = total_valor / total_cantidad if total_cantidad > 0 else 0
        
        return JsonResponse({
            'success': True,
            'lotes': lotes_data,
            'resumen': {
                'total_lotes': len(lotes_data),
                'total_cantidad': total_cantidad,
                'total_valor': total_valor,
                'costo_promedio': costo_promedio
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener lotes: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def crear_lote_manual(request):
    """Crear lote manualmente"""
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        producto_talla_id = data.get('producto_talla_id')
        cantidad = data.get('cantidad')
        costo_unitario = data.get('costo_unitario')
        precio_venta_unitario = data.get('precio_venta_unitario')
        
        if not all([producto_talla_id, cantidad, costo_unitario, precio_venta_unitario]):
            return JsonResponse({
                'success': False,
                'error': 'Todos los campos son requeridos'
            })
        
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        
        # Crear lote
        from .views import crear_lote_producto
        lote = crear_lote_producto(
            producto_talla=producto_talla,
            cantidad=int(cantidad),
            costo_unitario=Decimal(costo_unitario),
            sobreprecio_unitario=Decimal(data.get('sobreprecio_unitario', 0)),
            precio_venta_unitario=Decimal(precio_venta_unitario),
            numero_lote=data.get('numero_lote'),
            fecha_vencimiento=data.get('fecha_vencimiento'),
            observaciones=data.get('observaciones')
        )
        
        # Registrar movimiento de entrada
        from .views import registrar_movimiento_producto
        registrar_movimiento_producto(
            producto_talla=producto_talla,
            concepto='AJUSTE_ENTRADA',
            cantidad=int(cantidad),
            responsable=request.user,
            observaciones=f'Lote manual: {lote.numero_lote}'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Lote creado exitosamente',
            'lote_id': lote.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear lote: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def ajustar_lote(request, lote_id):
    """Ajustar cantidad de un lote existente"""
    try:
        data = json.loads(request.body)
        lote = get_object_or_404(LoteProducto, id=lote_id)
        
        nueva_cantidad = int(data.get('nueva_cantidad', 0))
        motivo = data.get('motivo', '')
        
        if nueva_cantidad < 0:
            return JsonResponse({
                'success': False,
                'error': 'La cantidad no puede ser negativa'
            })
        
        # Calcular diferencia
        cantidad_anterior = lote.cantidad_disponible
        diferencia = nueva_cantidad - cantidad_anterior
        
        # Actualizar lote
        lote.cantidad_disponible = nueva_cantidad
        lote.save()
        
        # Registrar movimiento
        if diferencia != 0:
            from .views import registrar_movimiento_producto
            concepto = 'AJUSTE_ENTRADA' if diferencia > 0 else 'AJUSTE_SALIDA'
            
            registrar_movimiento_producto(
                producto_talla=lote.producto_talla,
                concepto=concepto,
                cantidad=abs(diferencia),
                responsable=request.user,
                observaciones=f'Ajuste lote {lote.numero_lote}: {motivo}'
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Lote ajustado exitosamente',
            'diferencia': diferencia
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al ajustar lote: {str(e)}'
        })


# ========== DASHBOARD DE PRODUCTOS ==========

@login_required
def dashboard_productos(request):
    """Vista principal del dashboard de productos"""
    return render(request, 'vistas/modulo_dashboards/dashboard_productos.html')


@require_GET
@login_required
def obtener_datos_dashboard_productos(request):
    """API para obtener datos del dashboard de productos"""
    try:
        sucursal_id = request.session.get('idSucursalActual')
        
        # Métricas generales
        total_productos = Producto.objects.filter(activo=True).count()
        total_tallas = Producto_Talla.objects.filter(activo=True).count()
        
        # Stock por sucursal
        if sucursal_id:
            productos_con_stock = Producto_Talla.objects.filter(
                activo=True,
                loteproducto__cantidad_disponible__gt=0
            ).distinct().count()
            
            productos_sin_stock = total_tallas - productos_con_stock
        else:
            productos_con_stock = 0
            productos_sin_stock = 0
        
        # Valor del inventario
        valor_inventario = 0
        for pt in Producto_Talla.objects.filter(activo=True):
            from .views import obtener_valor_inventario_fifo
            valor_inventario += obtener_valor_inventario_fifo(pt)
        
        # Productos más vendidos (últimos 30 días)
        fecha_inicio = timezone.now() - timezone.timedelta(days=30)
        
        from .models import Ticket_Productos
        productos_vendidos = Ticket_Productos.objects.filter(
            ticket__created_at__gte=fecha_inicio,
            ticket__estado='PAGADO'
        ).values(
            'productoTalla__producto__nombre',
            'productoTalla__sku'
        ).annotate(
            total_vendido=Sum('cantidad'),
            ingresos=Sum(F('cantidad') * F('precio_unitario'))
        ).order_by('-total_vendido')[:10]
        
        # Productos con bajo stock
        productos_bajo_stock = []
        for pt in Producto_Talla.objects.filter(activo=True)[:50]:  # Limitar para performance
            stock_actual = sum(
                lote.cantidad_disponible 
                for lote in pt.loteproducto_set.filter(activo=True)
            )
            
            if stock_actual <= 5:  # Umbral de bajo stock
                productos_bajo_stock.append({
                    'sku': pt.sku,
                    'nombre': pt.producto.nombre,
                    'talla': pt.talla.nombre if pt.talla else 'Sin talla',
                    'stock': stock_actual
                })
        
        dashboard_data = {
            'metricas_generales': {
                'total_productos': total_productos,
                'total_tallas': total_tallas,
                'productos_con_stock': productos_con_stock,
                'productos_sin_stock': productos_sin_stock,
                'valor_inventario': float(valor_inventario)
            },
            'productos_vendidos': [
                {
                    'nombre': item['productoTalla__producto__nombre'],
                    'sku': item['productoTalla__sku'],
                    'total_vendido': item['total_vendido'],
                    'ingresos': float(item['ingresos'])
                }
                for item in productos_vendidos
            ],
            'productos_bajo_stock': productos_bajo_stock[:10]
        }
        
        return JsonResponse({
            'success': True,
            'dashboard': dashboard_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener datos del dashboard: {str(e)}'
        })


@require_GET
@login_required
def filtrar_productos_dashboard(request):
    """Filtrar productos para el dashboard"""
    try:
        # Parámetros de filtro
        categoria_id = request.GET.get('categoria_id')
        marca_id = request.GET.get('marca_id')
        estado_stock = request.GET.get('estado_stock')
        
        # Construir queryset
        queryset = Producto_Talla.objects.select_related(
            'producto', 'producto__categoria', 'producto__marca', 'talla'
        ).filter(activo=True)
        
        # Aplicar filtros
        if categoria_id:
            queryset = queryset.filter(producto__categoria_id=categoria_id)
        
        if marca_id:
            queryset = queryset.filter(producto__marca_id=marca_id)
        
        productos_data = []
        for pt in queryset[:100]:  # Limitar para performance
            stock_total = sum(
                lote.cantidad_disponible 
                for lote in pt.loteproducto_set.filter(activo=True)
            )
            
            # Filtrar por estado de stock
            if estado_stock == 'con_stock' and stock_total <= 0:
                continue
            elif estado_stock == 'sin_stock' and stock_total > 0:
                continue
            elif estado_stock == 'bajo_stock' and stock_total > 5:
                continue
            
            from .views import obtener_valor_inventario_fifo
            valor_inventario = obtener_valor_inventario_fifo(pt)
            
            productos_data.append({
                'sku': pt.sku,
                'nombre': pt.producto.nombre,
                'categoria': pt.producto.categoria.nombre if pt.producto.categoria else '',
                'marca': pt.producto.marca.nombre if pt.producto.marca else '',
                'talla': pt.talla.nombre if pt.talla else 'Sin talla',
                'precio_venta': float(pt.precio_venta),
                'stock_total': stock_total,
                'valor_inventario': float(valor_inventario)
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al filtrar productos: {str(e)}'
        })


@require_GET
@login_required
def exportar_dashboard_productos(request):
    """Exportar datos del dashboard de productos a Excel"""
    try:
        import openpyxl
        # `Font, PatternFill, Alignment` se importaban acá y no se usaban en
        # ninguna celda: fuera el 2026-08-22 (ver nota de imports arriba).

        # Obtener datos del dashboard
        dashboard_response = obtener_datos_dashboard_productos(request)
        dashboard_data = json.loads(dashboard_response.content)['dashboard']
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja de métricas generales
        ws_metricas = wb.active
        ws_metricas.title = "Métricas Generales"
        
        metricas = dashboard_data['metricas_generales']
        ws_metricas.append(['Métrica', 'Valor'])
        ws_metricas.append(['Total Productos', metricas['total_productos']])
        ws_metricas.append(['Total Tallas', metricas['total_tallas']])
        ws_metricas.append(['Productos con Stock', metricas['productos_con_stock']])
        ws_metricas.append(['Productos sin Stock', metricas['productos_sin_stock']])
        ws_metricas.append(['Valor Inventario', metricas['valor_inventario']])
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dashboard_productos.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


@require_GET
@login_required
def exportar_productos_filtrado(request):
    """Exportar productos filtrados a Excel"""
    try:
        import openpyxl
        
        # Obtener productos filtrados
        productos_response = filtrar_productos_dashboard(request)
        productos_data = json.loads(productos_response.content)['productos']
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos Filtrados"
        
        # Encabezados
        headers = [
            'SKU', 'Nombre', 'Categoría', 'Marca', 'Talla',
            'Precio Venta', 'Stock Total', 'Valor Inventario'
        ]
        
        ws.append(headers)
        
        # Datos
        for producto in productos_data:
            ws.append([
                producto['sku'],
                producto['nombre'],
                producto['categoria'],
                producto['marca'],
                producto['talla'],
                producto['precio_venta'],
                producto['stock_total'],
                producto['valor_inventario']
            ])
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="productos_filtrados.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })



# NOTA 2026-07: aqui vivia un set FIFO duplicado y muerto (dashboard_fifo,
# obtener_datos_dashboard_fifo, obtener_metricas_fifo, exportar_dashboard_fifo,
# reporte_fifo_general, obtener_analisis_fifo_detallado). Las versiones vivas
# y ruteadas estan en app/views.py.
