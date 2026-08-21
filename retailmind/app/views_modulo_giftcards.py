"""
Módulo Gift Cards - RetailMind

Emisión, consulta, recarga y anulación de tarjetas de regalo con saldo.
Vistas HTML (render) + APIs JSON, 100% function-based (patrón del proyecto).
La lógica de saldos vive en `app/services/giftcard_service.py`.
"""
import json
import logging
import os

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Count, F, Value
from django.db.models.functions import Coalesce
from django.utils import timezone

from .decorators import requiere_permiso
from .models import (
    GiftCard, MovimientoGiftCard, PermisoRol, Sucursal, Cliente, Vendedor, Empresa,
    ESTADO_GIFTCARD_CHOICES,
)
from .services import giftcard_service, fidelizacion_service

logger = logging.getLogger('app')

# Tope por operación de emisión/recarga. Una gift card es dinero canjeable en
# caja y hasta ahora se emitía sin ningún límite: un error de tipeo (un cero de
# más) o un abuso creaban saldo real sin contrapartida. Configurable por entorno.
MONTO_MAXIMO_GIFTCARD = int(os.environ.get('GIFTCARD_MONTO_MAXIMO', '2000000'))


def _validar_monto_giftcard(monto):
    """Devuelve un mensaje de error si el monto no es válido, o None si lo es."""
    if monto <= 0:
        return 'Monto inválido.'
    if monto > MONTO_MAXIMO_GIFTCARD:
        return (
            f'El monto (${monto:,.0f}) supera el máximo permitido por operación '
            f'(${MONTO_MAXIMO_GIFTCARD:,.0f}). Si es correcto, pide que suban el '
            f'límite (GIFTCARD_MONTO_MAXIMO).'
        ).replace(',', '.')
    return None


def _sucursal_actual(request):
    sid = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    if not sid:
        return None
    return Sucursal.objects.filter(id=sid).first()


def _empresa_actual(request):
    eid = request.session.get('idEmpresaActual') or request.session.get('empresaActual')
    if not eid:
        return None
    return Empresa.objects.filter(id=eid).first()


# ---------------------------------------------------------------------------
# Vigencia REAL de una gift card.
#
# `GiftCard.estado` solo pasa a VENCIDA cuando corre el comando
# `expirar_giftcards` (hoy NO está enganchado a `run_scheduler`), así que una
# tarjeta caducada sigue figurando como ACTIVA en la base. Nunca clasifiques
# por `estado` a secas: el vencimiento se evalúa siempre contra la fecha.
# ---------------------------------------------------------------------------

def _qs_con_saldo(qs):
    """Tarjetas que todavía representan dinero (excluye anuladas y saldo 0)."""
    return qs.filter(saldo_actual__gt=0).exclude(estado='ANULADA')


def _qs_bloqueadas(qs):
    return _qs_con_saldo(qs).filter(estado='BLOQUEADA')


def _qs_vencidas(qs, hoy):
    """Con saldo, no bloqueadas y con fecha de vencimiento ya pasada."""
    return _qs_con_saldo(qs).exclude(estado='BLOQUEADA').filter(
        fecha_vencimiento__isnull=False, fecha_vencimiento__lt=hoy,
    )


def _qs_vigentes(qs, hoy):
    """Deuda REAL con clientes: canjeable hoy en cualquier sucursal."""
    return _qs_con_saldo(qs).exclude(estado='BLOQUEADA').filter(
        Q(fecha_vencimiento__isnull=True) | Q(fecha_vencimiento__gte=hoy)
    )


def _estado_efectivo(gc, hoy):
    """Estado que se le debe mostrar al usuario (corrige el cron que no corre)."""
    if gc.estado in ('ANULADA', 'BLOQUEADA', 'AGOTADA'):
        return gc.estado
    if gc.fecha_vencimiento and gc.fecha_vencimiento < hoy:
        return 'VENCIDA'
    if gc.saldo_actual <= 0:
        return 'AGOTADA'
    return gc.estado


def _aplicar_filtros_giftcards(qs, request):
    """Filtros compartidos por el listado y la exportación de gift cards."""
    hoy = timezone.localdate()
    vigencia = (request.GET.get('vigencia') or '').strip()
    if vigencia == 'vigentes':
        qs = _qs_vigentes(qs, hoy)
    elif vigencia == 'vencidas':
        qs = _qs_vencidas(qs, hoy)
    elif vigencia == 'por_vencer':
        qs = _qs_vigentes(qs, hoy).filter(
            fecha_vencimiento__isnull=False,
            fecha_vencimiento__lte=hoy + timezone.timedelta(days=30),
        )
    elif vigencia == 'bloqueadas':
        qs = _qs_bloqueadas(qs)
    elif vigencia == 'con_saldo':
        qs = _qs_con_saldo(qs)

    # Estado de entrega del código: para saber a quién falta enviarle su
    # gift card (o revisar las ya enviadas) sin exportar todo el listado.
    envio = (request.GET.get('envio') or '').strip()
    if envio == 'enviadas':
        qs = qs.filter(correo_enviado_en__isnull=False)
    elif envio == 'sin_enviar':
        qs = qs.filter(correo_enviado_en__isnull=True)

    busqueda = (request.GET.get('q') or '').strip()
    if busqueda:
        qs = qs.filter(
            Q(codigo__icontains=busqueda) |
            Q(codigo_fisico__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(cliente__apellido__icontains=busqueda) |
            Q(cliente__rut__icontains=busqueda) |
            Q(correo_enviado_a__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    estado = (request.GET.get('estado') or '').strip()
    if estado:
        qs = qs.filter(estado=estado)
    motivo = (request.GET.get('motivo') or '').strip()
    if motivo:
        qs = qs.filter(motivo=motivo)
    sucursal_id = (request.GET.get('sucursal_id') or '').strip()
    if sucursal_id.isdigit():
        qs = qs.filter(sucursal_emision_id=int(sucursal_id))
    desde = _fecha_valida(request.GET.get('desde'))
    if desde:
        qs = qs.filter(fecha_emision__date__gte=desde)
    hasta = _fecha_valida(request.GET.get('hasta'))
    if hasta:
        qs = qs.filter(fecha_emision__date__lte=hasta)
    return qs


def _fecha_valida(valor):
    """'YYYY-MM-DD' -> str si es una fecha real; None si no (evita el 500)."""
    from datetime import datetime as _dt
    valor = (valor or '').strip()
    if not valor:
        return None
    try:
        _dt.strptime(valor, '%Y-%m-%d')
    except ValueError:
        return None
    return valor


def _aplicar_filtros_trazabilidad(qs, request):
    """Filtros compartidos por la trazabilidad global y su exportación."""
    busqueda = (request.GET.get('q') or '').strip()
    if busqueda:
        qs = qs.filter(
            Q(giftcard__codigo__icontains=busqueda) |
            Q(giftcard__codigo_fisico__icontains=busqueda) |
            Q(giftcard__cliente__nombre__icontains=busqueda) |
            Q(giftcard__cliente__rut__icontains=busqueda)
        )
    tipo = (request.GET.get('tipo') or '').strip()
    if tipo:
        qs = qs.filter(tipo=tipo)
    # `isdigit` obligatorio: un sucursal_id no numérico en la URL reventaba con
    # ValueError -> 500 en la trazabilidad y en su export.
    sucursal_id = (request.GET.get('sucursal_id') or '').strip()
    if sucursal_id.isdigit():
        qs = qs.filter(sucursal_id=int(sucursal_id))
    usuario = (request.GET.get('usuario') or '').strip()
    if usuario:
        qs = qs.filter(usuario__username__icontains=usuario)
    desde = _fecha_valida(request.GET.get('desde'))
    if desde:
        qs = qs.filter(fecha__date__gte=desde)
    hasta = _fecha_valida(request.GET.get('hasta'))
    if hasta:
        qs = qs.filter(fecha__date__lte=hasta)
    return qs


# ========== VISTAS HTML ==========

@requiere_permiso('giftcards_listado', 'puede_ver')
def modulo_giftcards(request):
    """Listado / gestión de gift cards. La emisión se hace en un modal aquí."""
    context = {
        'sucursal_actual': _sucursal_actual(request),
        'empresa_actual': _empresa_actual(request),
        'estado_choices': GiftCard._meta.get_field('estado').choices,
        'tipo_tarjeta_choices': GiftCard._meta.get_field('tipo_tarjeta').choices,
        'motivo_choices': GiftCard._meta.get_field('motivo').choices,
        'sucursales': Sucursal.objects.all().order_by('nombre'),
        'hoy': timezone.localdate(),
    }
    return render(request, 'vistas/modulo_giftcards/lista.html', context)


@requiere_permiso('giftcards_emitir', 'puede_crear')
def emitir_giftcard_vista(request):
    """
    Emisión movida a un modal dentro del listado. Esta ruta se conserva (la usa
    el menú/permiso `giftcards_emitir`) y redirige al listado abriendo el modal.
    """
    return redirect('/app/giftcards/?nueva=1')


@requiere_permiso('giftcards_listado', 'puede_ver')
def detalle_giftcard_vista(request, giftcard_id):
    """Detalle de una gift card con su historial de movimientos."""
    giftcard = get_object_or_404(
        GiftCard.objects.select_related('cliente', 'sucursal_emision', 'created_by', 'empresa'),
        id=giftcard_id,
    )
    hoy = timezone.localdate()
    movs = giftcard.movimientos.select_related('usuario', 'sucursal', 'ticket')

    def _sum(**f):
        return movs.filter(**f).aggregate(s=Sum('monto'))['s'] or 0

    ledger = movs.aggregate(s=Sum('monto'))['s'] or 0
    context = {
        'giftcard': giftcard,
        'movimientos': movs.order_by('-fecha'),
        'estado_efectivo': _estado_efectivo(giftcard, hoy),
        'estado_efectivo_display': dict(ESTADO_GIFTCARD_CHOICES).get(
            _estado_efectivo(giftcard, hoy), giftcard.estado),
        'cargado': _sum(tipo__in=['EMISION', 'CARGA']),
        'consumido': -_sum(tipo='CONSUMO'),
        'reversado': _sum(tipo='REVERSA'),
        'anulado': -_sum(tipo='ANULACION'),
        'saldo_ledger': ledger,
        'descuadrada': ledger != giftcard.saldo_actual,
        'dias_vencimiento': (giftcard.fecha_vencimiento - hoy).days if giftcard.fecha_vencimiento else None,
    }
    return render(request, 'vistas/modulo_giftcards/detalle.html', context)


# ========== APIs JSON ==========

@require_POST
@requiere_permiso('giftcards_emitir', 'puede_crear')
def api_emitir_giftcard(request):
    """
    Emite una nueva gift card desde el modal del listado.

    Tipos:
    - DIGITAL: requiere un cliente (se resuelve por RUT o cliente_id).
    - FISICA:  requiere `codigo_fisico` (impreso en la tarjeta); el RUT/cliente
      es opcional (se puede vincular al canjear).
    """
    try:
        data = json.loads(request.body or '{}')
        monto = int(data.get('monto') or 0)
        error_monto = _validar_monto_giftcard(monto)
        if error_monto:
            return JsonResponse({'success': False, 'error': error_monto}, status=400)

        tipo_tarjeta = (data.get('tipo_tarjeta') or 'DIGITAL').upper()

        # Resolver titular: por cliente_id explícito o por RUT.
        cliente = None
        if data.get('cliente_id'):
            cliente = Cliente.objects.filter(id=data['cliente_id']).first()
        rut = (data.get('rut') or '').strip()
        nombre_nuevo = (data.get('nombre') or '').strip()
        if cliente is None and rut:
            cliente = fidelizacion_service.resolver_cliente_por_rut(rut)
            if cliente is None:
                if nombre_nuevo:
                    # El RUT no existe: crear el cliente al vuelo (mínimo nombre + RUT).
                    try:
                        cliente, _, _ = fidelizacion_service.registrar_cliente_manual(
                            nombre=nombre_nuevo,
                            rut=rut,
                            usuario=request.user,
                            empresa=_empresa_actual(request),
                        )
                    except fidelizacion_service.FidelizacionError as e:
                        return JsonResponse({'success': False, 'error': str(e)}, status=400)
                elif tipo_tarjeta == 'DIGITAL':
                    # Pedir el nombre para poder crearlo (lo maneja el modal).
                    return JsonResponse({
                        'success': False,
                        'need_nombre': True,
                        'error': 'No existe un cliente con ese RUT. Ingresa el nombre para crearlo.',
                    }, status=400)
                # FÍSICA con RUT no encontrado y sin nombre → se emite sin titular.

        if tipo_tarjeta == 'DIGITAL' and cliente is None:
            return JsonResponse({
                'success': False,
                'error': 'Para una gift card digital debes indicar el RUT del cliente.',
            }, status=400)

        vendedor = None
        if data.get('vendedor_id'):
            vendedor = Vendedor.objects.filter(id=data['vendedor_id']).first()

        vencimiento = data.get('fecha_vencimiento') or None

        # Ámbito de canje: 'EMPRESA' (default) = solo sucursales de la empresa
        # activa de la sesión; 'TODAS' = toda la cadena (comportamiento viejo).
        ambito = (data.get('ambito') or 'EMPRESA').upper()
        if ambito not in ('EMPRESA', 'TODAS'):
            return JsonResponse({'success': False, 'error': 'Ámbito inválido.'}, status=400)
        empresa_gc = None
        if ambito == 'EMPRESA':
            empresa_gc = _empresa_actual(request)
            # Sin empresa en sesión NO se emite en silencio como global: el
            # usuario pidió acotarla y la tarjeta saldría canjeable en toda la
            # cadena sin aviso (mismo criterio que api_cambiar_ambito).
            if empresa_gc is None:
                return JsonResponse({
                    'success': False,
                    'error': 'No se pudo determinar la empresa activa para acotar la '
                             'gift card. Selecciona una empresa o emite con ámbito '
                             '"Todas las empresas".',
                }, status=400)

        giftcard = giftcard_service.emitir(
            monto,
            sucursal=_sucursal_actual(request),
            cliente=cliente,
            vendedor=vendedor,
            vencimiento=vencimiento,
            pin=(data.get('pin') or None),
            usuario=request.user,
            observaciones=data.get('observaciones', ''),
            tipo_tarjeta=tipo_tarjeta,
            codigo_fisico=(data.get('codigo_fisico') or None),
            motivo=(data.get('motivo') or 'OTRO'),
            descripcion=(data.get('descripcion') or ''),
            empresa=empresa_gc,
        )
        return JsonResponse({
            'success': True,
            'giftcard': {
                'id': giftcard.id,
                'codigo': giftcard.codigo,
                'codigo_fisico': giftcard.codigo_fisico,
                'tipo_tarjeta': giftcard.tipo_tarjeta,
                'cliente': giftcard.cliente.nombre_completo if giftcard.cliente else '',
                'cliente_email': (giftcard.cliente.email or '') if giftcard.cliente else '',
                'saldo_actual': giftcard.saldo_actual,
                'fecha_vencimiento': giftcard.fecha_vencimiento.isoformat() if giftcard.fecha_vencimiento else None,
                'ambito': 'EMPRESA' if giftcard.empresa_id else 'TODAS',
                'empresa': giftcard.empresa.nombre if giftcard.empresa_id else '',
            },
        })
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception("Error al emitir gift card")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_listar_giftcards(request):
    """Listado paginado de gift cards con filtros."""
    hoy = timezone.localdate()
    qs = GiftCard.objects.select_related('cliente', 'sucursal_emision', 'empresa').all()
    qs = _aplicar_filtros_giftcards(qs, request)

    # Totales del universo filtrado (no solo de la página visible): sin esto el
    # usuario no puede saber cuánta plata representa el filtro que aplicó.
    tot = qs.aggregate(n=Count('id'), saldo=Sum('saldo_actual'))
    saldo_vig = _qs_vigentes(qs, hoy).aggregate(s=Sum('saldo_actual'))['s'] or 0

    qs = qs.annotate(
        _cargado=Coalesce(
            Sum('movimientos__monto',
                filter=Q(movimientos__tipo__in=['EMISION', 'CARGA'])),
            Value(0),
        ),
        _consumido=Coalesce(
            Sum('movimientos__monto', filter=Q(movimientos__tipo='CONSUMO')),
            Value(0),
        ),
    # `annotate()` con GROUP BY descarta el ordering por defecto del Meta: sin
    # este order_by explícito la paginación repite/saltea filas entre páginas.
    ).order_by('-fecha_emision', '-id')

    per_page = min(max(int(request.GET.get('per_page', 20) or 20), 5), 200)
    paginator = Paginator(qs, per_page)
    page = paginator.get_page(request.GET.get('page', 1))

    items = []
    for gc in page:
        efectivo = _estado_efectivo(gc, hoy)
        dias = (gc.fecha_vencimiento - hoy).days if gc.fecha_vencimiento else None
        items.append({
            'id': gc.id,
            'codigo': gc.codigo,
            'codigo_fisico': gc.codigo_fisico,
            'tipo_tarjeta': gc.tipo_tarjeta,
            'tipo_display': gc.get_tipo_tarjeta_display(),
            'motivo': gc.motivo,
            'motivo_display': gc.get_motivo_display(),
            'descripcion': gc.descripcion or '',
            'saldo_inicial': gc.saldo_inicial,
            'saldo_actual': gc.saldo_actual,
            'cargado': gc._cargado,
            'consumido': -gc._consumido,
            'estado': gc.estado,
            'estado_display': gc.get_estado_display(),
            # Estado corregido por fecha: lo que el usuario debe creer.
            'estado_efectivo': efectivo,
            'estado_efectivo_display': dict(ESTADO_GIFTCARD_CHOICES).get(efectivo, efectivo),
            'cliente': gc.cliente.nombre_completo if gc.cliente else '',
            'cliente_email': (gc.cliente.email or '') if gc.cliente else '',
            'sucursal_emision': gc.sucursal_emision.nombre if gc.sucursal_emision else '',
            'fecha_emision': gc.fecha_emision.strftime('%Y-%m-%d %H:%M'),
            'fecha_vencimiento': gc.fecha_vencimiento.isoformat() if gc.fecha_vencimiento else None,
            'dias_vencimiento': dias,
            'vencida': gc.esta_vencida,
            'ambito': 'EMPRESA' if gc.empresa_id else 'TODAS',
            'empresa': gc.empresa.nombre if gc.empresa_id else '',
            # Entrega del código: a quién y cuándo se envió (vacío = nunca).
            'correo_enviado_a': gc.correo_enviado_a or '',
            'correo_enviado_en': (
                timezone.localtime(gc.correo_enviado_en).strftime('%Y-%m-%d %H:%M')
                if gc.correo_enviado_en else ''
            ),
            'correo_envios': gc.correo_envios or 0,
            # Plata que el cliente TODAVÍA puede gastar (0 si ya no es canjeable).
            'pendiente_por_usar': (
                gc.saldo_actual
                if (gc.estado == 'ACTIVA' and not gc.esta_vencida and gc.saldo_actual > 0)
                else 0
            ),
        })

    return JsonResponse({
        'success': True,
        'items': items,
        'page': page.number,
        'num_pages': paginator.num_pages,
        'total': paginator.count,
        'totales': {
            'tarjetas': tot['n'] or 0,
            'saldo': tot['saldo'] or 0,
            'saldo_vigente': saldo_vig,
        },
    })


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_detalle_giftcard(request, giftcard_id):
    """Detalle + historial de una gift card."""
    gc = get_object_or_404(GiftCard, id=giftcard_id)
    movimientos = [{
        'tipo': m.tipo,
        'tipo_display': m.get_tipo_display(),
        'monto': m.monto,
        'saldo_resultante': m.saldo_resultante,
        'fecha': m.fecha.strftime('%Y-%m-%d %H:%M'),
        'ticket': m.ticket.correlativo if m.ticket else None,
        'usuario': m.usuario.get_username() if m.usuario else '',
        'sucursal': m.sucursal.nombre if m.sucursal else '',
        'observaciones': m.observaciones or '',
    } for m in gc.movimientos.select_related('usuario', 'sucursal', 'ticket').order_by('-fecha')]

    return JsonResponse({
        'success': True,
        'giftcard': {
            'id': gc.id,
            'codigo': gc.codigo,
            'codigo_fisico': gc.codigo_fisico,
            'tipo_tarjeta': gc.tipo_tarjeta,
            'tipo_display': gc.get_tipo_tarjeta_display(),
            'motivo': gc.motivo,
            'motivo_display': gc.get_motivo_display(),
            'descripcion': gc.descripcion or '',
            'observaciones': gc.observaciones or '',
            'saldo_inicial': gc.saldo_inicial,
            'saldo_actual': gc.saldo_actual,
            'estado': gc.estado,
            'estado_display': gc.get_estado_display(),
            'cliente': gc.cliente.nombre_completo if gc.cliente else '',
            'cliente_email': (gc.cliente.email or '') if gc.cliente else '',
            'fecha_emision': gc.fecha_emision.strftime('%Y-%m-%d %H:%M'),
            'fecha_vencimiento': gc.fecha_vencimiento.isoformat() if gc.fecha_vencimiento else None,
            'ambito': 'EMPRESA' if gc.empresa_id else 'TODAS',
            'empresa': gc.empresa.nombre if gc.empresa_id else '',
        },
        'movimientos': movimientos,
    })


@require_GET
@login_required
def api_consultar_saldo_giftcard(request):
    """
    Consulta saldo por código (usada por el POS web y la pantalla de gestión).

    Gate en `login_required` (no permiso de módulo) a propósito: lo consume la
    CAJA al cobrar y el rol cajero/vendedor no tiene `giftcards_listado`.
    Mismo criterio que `obtener_promos_activas` / `obtener_ofertas_activas`.
    """
    codigo = (request.GET.get('codigo') or '').strip()
    if not codigo:
        return JsonResponse({'success': False, 'error': 'Código requerido.'}, status=400)
    try:
        info = giftcard_service.consultar_saldo(codigo)
        # Datos extra que la pantalla de gestión necesita para no obligar al
        # usuario a abrir el detalle en otra pestaña.
        gc = GiftCard.objects.select_related('cliente', 'sucursal_emision', 'empresa').filter(
            Q(codigo=codigo.upper()) | Q(codigo_fisico=codigo.upper())
        ).first()
        if gc is not None:
            hoy = timezone.localdate()
            info.update({
                'id': gc.id,
                'saldo_inicial': gc.saldo_inicial,
                'cliente': gc.cliente.nombre_completo if gc.cliente else '',
                'sucursal_emision': gc.sucursal_emision.nombre if gc.sucursal_emision else '',
                'fecha_emision': gc.fecha_emision.strftime('%Y-%m-%d %H:%M'),
                'estado_efectivo': _estado_efectivo(gc, hoy),
                'movimientos': gc.movimientos.count(),
            })
            # El correo del titular SOLO para quien administra gift cards: este
            # endpoint está en `login_required` porque lo consume la caja, y
            # cualquier usuario logueado que pruebe códigos físicos (impresos,
            # potencialmente correlativos) no debe cosechar emails de clientes.
            if PermisoRol.tiene_permiso(request.user, 'giftcards_listado', 'puede_ver'):
                info['cliente_email'] = (gc.cliente.email or '') if gc.cliente else ''
                info['correo_enviado_a'] = gc.correo_enviado_a or ''
                info['correo_enviado_en'] = (
                    timezone.localtime(gc.correo_enviado_en).strftime('%Y-%m-%d %H:%M')
                    if gc.correo_enviado_en else ''
                )
                info['correo_envios'] = gc.correo_envios or 0
        return JsonResponse({'success': True, **info})
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)


@require_POST
@login_required
def api_validar_giftcard(request):
    """
    Pre-validación de gift card antes de cobrar (no descuenta).
    La usa el POS al agregar un pago con gift card.

    Gate en `login_required` (no permiso de módulo): lo consume la CAJA y el
    rol cajero/vendedor no tiene `giftcards_listado` (criterio de promos POS).
    Valida también el ámbito por empresa contra la sucursal activa.
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        data = request.POST
    codigo = (data.get('codigo') or '').strip()
    monto = int(data.get('monto') or 0)
    pin = data.get('pin')
    resultado = giftcard_service.validar(
        codigo, monto, pin=pin, sucursal=_sucursal_actual(request),
    )
    return JsonResponse({'success': True, **resultado})


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_crear')
def api_recargar_giftcard(request):
    """
    Recarga saldo a una gift card existente.

    Guarda contra un agujero real: `giftcard_service.recargar` acepta tarjetas
    VENCIDAS, pero `consumir` las rechaza — la plata recargada quedaba muerta
    sin ningún aviso. Aquí se bloquea y se le dice al usuario que primero
    extienda el vencimiento (se hace desde el mismo modal).
    """
    try:
        data = json.loads(request.body or '{}')
        codigo = (data.get('codigo') or '').strip()
        monto = int(data.get('monto') or 0)
        error_monto = _validar_monto_giftcard(monto)
        if error_monto:
            return JsonResponse({'success': False, 'error': error_monto}, status=400)

        gc_prev = GiftCard.objects.filter(
            Q(codigo=codigo.upper()) | Q(codigo_fisico=codigo.upper())
        ).first()
        if gc_prev is not None and (gc_prev.esta_vencida or gc_prev.estado == 'VENCIDA'):
            return JsonResponse({
                'success': False,
                'vencida': True,
                'error': (
                    f'La gift card está vencida ({gc_prev.fecha_vencimiento}). '
                    'Extiende primero el vencimiento: si la recargas así, el '
                    'saldo no se puede canjear.'
                ),
            }, status=400)

        mov = giftcard_service.recargar(
            codigo, monto,
            sucursal=_sucursal_actual(request), usuario=request.user,
            observaciones=data.get('observaciones', ''),
        )
        return JsonResponse({'success': True, 'saldo_actual': mov.saldo_resultante})
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_editar')
def api_anular_giftcard(request):
    """
    Anula una gift card (lleva saldo a 0, estado ANULADA).

    Es dar de baja un pasivo, así que el motivo es OBLIGATORIO: queda en el
    ledger junto al usuario que la anuló.
    """
    try:
        data = json.loads(request.body or '{}')
        codigo = (data.get('codigo') or '').strip()
        motivo = (data.get('motivo') or '').strip()
        if len(motivo) < 5:
            return JsonResponse({
                'success': False,
                'error': 'Indica el motivo de la anulación (mínimo 5 caracteres).',
            }, status=400)
        gc = giftcard_service.anular(
            codigo, usuario=request.user,
            observaciones=f'Anulación: {motivo}',
        )
        return JsonResponse({'success': True, 'estado': gc.estado})
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_reporte_giftcards(request):
    """
    KPIs del programa de gift cards, en clave de PASIVO (dinero emitido que el
    negocio le debe a sus clientes).

    Respeta los mismos filtros del listado para que la tabla y los KPIs cuadren.

    Notas de veracidad (por qué no se usa `estado` a secas):
    - `estado='VENCIDA'` solo lo escribe el comando `expirar_giftcards`, que no
      está enganchado al scheduler: el vencimiento se evalúa por fecha.
    - "Total cargado" suma EMISION + CARGA del ledger; `saldo_inicial` ignora
      las recargas posteriores y subestima el dinero puesto en tarjetas.
    """
    hoy = timezone.localdate()
    qs = _aplicar_filtros_giftcards(GiftCard.objects.all(), request)

    def _agg(sub):
        r = sub.aggregate(n=Count('id'), s=Sum('saldo_actual'))
        return r['n'] or 0, r['s'] or 0

    n_vig, m_vig = _agg(_qs_vigentes(qs, hoy))
    n_ven, m_ven = _agg(_qs_vencidas(qs, hoy))
    n_blo, m_blo = _agg(_qs_bloqueadas(qs))
    n_pv, m_pv = _agg(_qs_vigentes(qs, hoy).filter(
        fecha_vencimiento__isnull=False,
        fecha_vencimiento__lte=hoy + timezone.timedelta(days=30),
    ))

    # Flujo desde el ledger (fuente de verdad auditable).
    movs = MovimientoGiftCard.objects.filter(giftcard__in=qs)

    def _sum(**f):
        return movs.filter(**f).aggregate(s=Sum('monto'))['s'] or 0

    cargado = _sum(tipo__in=['EMISION', 'CARGA'])
    consumido = -_sum(tipo='CONSUMO')
    reversado = _sum(tipo='REVERSA')
    anulado = -_sum(tipo='ANULACION')
    ajustes = _sum(tipo='AJUSTE')

    # Cuadratura ledger vs saldo denormalizado (una sola query).
    descuadres = qs.annotate(
        _led=Coalesce(Sum('movimientos__monto'), Value(0))
    ).exclude(_led=F('saldo_actual')).count()

    return JsonResponse({
        'success': True,
        # === PASIVO ===
        'vigentes': n_vig,
        'saldo_vigente': m_vig,          # deuda real con clientes
        'vencidas_con_saldo': n_ven,
        'saldo_vencido': m_ven,          # dinero que el cliente ya no puede usar
        'bloqueadas_con_saldo': n_blo,
        'saldo_bloqueado': m_blo,
        'por_vencer_30': n_pv,
        'saldo_por_vencer_30': m_pv,
        # === FLUJO (ledger) ===
        'total_emitidas': qs.count(),
        'total_cargado': cargado,
        'total_consumido': consumido,
        'total_reversado': reversado,
        'total_anulado': anulado,
        'total_ajustes': ajustes,
        'saldo_total_tarjetas': qs.aggregate(s=Sum('saldo_actual'))['s'] or 0,
        'descuadres_ledger': descuadres,
        'por_estado': list(
            qs.values('estado').annotate(n=Count('id')).order_by('-n')
        ),
        # === ENTREGA DEL CÓDIGO ===
        # Cuántas tarjetas del filtro ya salieron por correo y cuántas no:
        # responde "¿le envié su gift card a todos?" sin abrir el listado.
        'con_correo_enviado': qs.filter(correo_enviado_en__isnull=False).count(),
        'sin_correo_enviado': qs.filter(correo_enviado_en__isnull=True).count(),
        # Compatibilidad con consumidores antiguos del endpoint.
        'activas': n_vig,
        'saldo_circulante': m_vig,
        'monto_total_emitido': cargado,
    })


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_editar')
def api_bloquear_giftcard(request):
    """Bloquea temporalmente una gift card (estado BLOQUEADA, reversible)."""
    try:
        data = json.loads(request.body or '{}')
        codigo = (data.get('codigo') or '').strip()
        gc = giftcard_service.bloquear(
            codigo, usuario=request.user,
            observaciones=data.get('motivo', ''),
        )
        return JsonResponse({'success': True, 'estado': gc.estado})
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_editar')
def api_desbloquear_giftcard(request):
    """Levanta el bloqueo de una gift card."""
    try:
        data = json.loads(request.body or '{}')
        codigo = (data.get('codigo') or '').strip()
        gc = giftcard_service.desbloquear(
            codigo, usuario=request.user,
            observaciones=data.get('motivo', ''),
        )
        return JsonResponse({'success': True, 'estado': gc.estado})
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def _extender_vencimiento(codigo, nueva_fecha, usuario):
    """
    Cambia la fecha de vencimiento de una gift card dejando registro en el
    ledger (fila AJUSTE monto=0). Si la tarjeta estaba VENCIDA y la nueva fecha
    la revive, recalcula el estado por saldo.

    Va aquí y no en el servicio porque `giftcard_service.editar` no maneja
    fechas (ver `cambios_backend_pendientes`). Usa lock para no pisarse con un
    consumo simultáneo.
    """
    from datetime import date, datetime

    if isinstance(nueva_fecha, str):
        try:
            nueva_fecha = datetime.strptime(nueva_fecha.strip(), '%Y-%m-%d').date()
        except ValueError:
            raise giftcard_service.GiftCardError('Fecha de vencimiento inválida.')
    if nueva_fecha is not None and not isinstance(nueva_fecha, date):
        raise giftcard_service.GiftCardError('Fecha de vencimiento inválida.')

    codigo = (codigo or '').strip().upper()
    hoy = timezone.localdate()
    with transaction.atomic():
        gc = GiftCard.objects.select_for_update().filter(
            Q(codigo=codigo) | Q(codigo_fisico=codigo)
        ).first()
        if gc is None:
            raise giftcard_service.GiftCardError('Gift card no encontrada.')
        if gc.estado == 'ANULADA':
            raise giftcard_service.GiftCardError(
                'La gift card está anulada: no se puede cambiar su vencimiento.'
            )
        anterior = gc.fecha_vencimiento
        if anterior == nueva_fecha:
            return gc

        gc.fecha_vencimiento = nueva_fecha
        campos = ['fecha_vencimiento', 'updated_by', 'updated_at']
        vigente_ahora = nueva_fecha is None or nueva_fecha >= hoy
        if gc.estado == 'VENCIDA' and vigente_ahora:
            gc.estado = 'AGOTADA' if gc.saldo_actual <= 0 else 'ACTIVA'
            campos.append('estado')
        gc.updated_by = usuario
        gc.save(update_fields=campos)

        MovimientoGiftCard.objects.create(
            giftcard=gc,
            tipo='AJUSTE',
            monto=0,
            saldo_resultante=gc.saldo_actual,
            usuario=usuario,
            observaciones=(
                f"Vencimiento: {anterior or 'sin vencimiento'} -> "
                f"{nueva_fecha or 'sin vencimiento'}"
            ),
        )
    logger.info("GiftCard vencimiento actualizado codigo=%s %s -> %s",
                gc.codigo, anterior, nueva_fecha)
    return gc


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_editar')
def api_editar_giftcard(request):
    """
    Edita la descripción / motivo / observaciones de una gift card emitida.
    Acepta además `fecha_vencimiento` (opcional) para extender la vigencia sin
    tener que emitir una tarjeta nueva. Todo queda en el ledger.
    """
    try:
        data = json.loads(request.body or '{}')
        codigo = (data.get('codigo') or '').strip()
        gc = giftcard_service.editar(
            codigo,
            descripcion=data.get('descripcion'),
            motivo=data.get('motivo'),
            observaciones=data.get('observaciones'),
            usuario=request.user,
        )
        if 'fecha_vencimiento' in data:
            gc = _extender_vencimiento(
                codigo, (data.get('fecha_vencimiento') or None), request.user,
            )
        return JsonResponse({
            'success': True,
            'motivo': gc.motivo,
            'motivo_display': gc.get_motivo_display(),
            'descripcion': gc.descripcion or '',
            'estado': gc.estado,
            'fecha_vencimiento': gc.fecha_vencimiento.isoformat() if gc.fecha_vencimiento else None,
        })
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_editar')
def api_cambiar_ambito_giftcard(request):
    """
    Cambia el ámbito de canje de una gift card:
    - ambito='EMPRESA': solo sucursales de la empresa activa de la sesión
      (o de `empresa_id` si viene explícito).
    - ambito='TODAS':   válida en todas las empresas de la cadena.
    Queda una fila AJUSTE monto=0 en el ledger.
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'}, status=400)

    codigo = (data.get('codigo') or '').strip()
    ambito = (data.get('ambito') or '').upper()
    if ambito == 'TODAS':
        empresa = None
    elif ambito == 'EMPRESA':
        empresa_id = data.get('empresa_id')
        empresa = (Empresa.objects.filter(id=empresa_id).first()
                   if empresa_id else _empresa_actual(request))
        if empresa is None:
            return JsonResponse({
                'success': False,
                'error': 'No se pudo determinar la empresa para acotar el ámbito.',
            }, status=400)
    else:
        return JsonResponse({'success': False, 'error': 'Ámbito inválido (EMPRESA o TODAS).'}, status=400)

    try:
        gc = giftcard_service.cambiar_ambito(
            codigo, empresa, usuario=request.user,
            observaciones=data.get('observaciones', ''),
        )
        return JsonResponse({
            'success': True,
            'ambito': 'EMPRESA' if gc.empresa_id else 'TODAS',
            'empresa': gc.empresa.nombre if gc.empresa_id else '',
        })
    except giftcard_service.GiftCardError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# Meses en español para la fecha "humana" del correo (evita depender del
# locale del sistema operativo del servidor).
_MESES_ES = ['', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
             'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']


class CorreoGiftCardError(Exception):
    """El correo no se pudo entregar al servidor (nada quedó registrado)."""


def enviar_codigos_por_correo(giftcards, email_destino, *, nombre_destino='',
                              usuario=None, sucursal=None, beneficiarios=None,
                              connection=None):
    """
    Envía UN correo con los códigos de una o varias gift cards.

    Lo usan el botón de la pantalla y el comando de lote: un trabajador con
    varios hijos recibe UN correo con todas sus tarjetas (cada una rotulada con
    su beneficiario) en vez de N correos sueltos.

    - `beneficiarios`: dict {giftcard_id: 'NOMBRE'} para rotular cada tarjeta.
    - `connection`: conexión SMTP ya abierta (para envíos masivos; si no se
      entrega, se abre y cierra una propia con timeout).

    Registra el envío SOLO si el servidor lo aceptó: fila `ENVIO_CORREO` en el
    ledger de cada tarjeta + campos denormalizados. Lanza `CorreoGiftCardError`
    si el envío falla, sin tocar nada.
    """
    from django.conf import settings as dj_settings
    from django.core.mail import EmailMultiAlternatives, get_connection
    from django.template.loader import render_to_string

    giftcards = list(giftcards)
    if not giftcards:
        raise CorreoGiftCardError('No hay gift cards que enviar.')
    beneficiarios = beneficiarios or {}
    hoy = timezone.localdate()
    marca = os.environ.get('GIFTCARD_CORREO_MARCA', 'REALSPORT')

    # Vigencia y ámbito se toman de la primera tarjeta: en un lote son iguales
    # (mismo comando, misma empresa y vencimiento).
    ref = giftcards[0]
    vigencia_dias = (ref.fecha_vencimiento - hoy).days if ref.fecha_vencimiento else None
    fecha_venc_humana = ''
    if ref.fecha_vencimiento:
        f = ref.fecha_vencimiento
        fecha_venc_humana = f'{f.day} de {_MESES_ES[f.month]} de {f.year}'

    tiendas = []
    if ref.empresa_id:
        tiendas = list(
            Sucursal.objects.filter(empresa_id=ref.empresa_id)
            .exclude(tipo_sucursal='CENTRO_DISTRIBUCION')
            .order_by('alias').values_list('alias', flat=True)
        )

    tarjetas = [{
        'codigo': (gc.codigo_fisico if gc.tipo_tarjeta == 'FISICA' and gc.codigo_fisico
                   else gc.codigo),
        'monto': gc.saldo_actual,
        'beneficiario': (beneficiarios.get(gc.id) or '').strip(),
    } for gc in giftcards]
    monto_total = sum(t['monto'] for t in tarjetas)

    html = render_to_string('emails/giftcard_codigo.html', {
        'marca': marca,
        'nombre_destinatario': nombre_destino,
        'tarjetas': tarjetas,
        'fecha_vencimiento_humana': fecha_venc_humana,
        'vigencia_dias': vigencia_dias,
        'empresa_nombre': ref.empresa.nombre if ref.empresa_id else '',
        'empresa_rut': ref.empresa.rut if ref.empresa_id else '',
        'tiendas': tiendas,
    })

    saludo = f'Hola {nombre_destino}, ' if nombre_destino else 'Hola, '
    donde = (f'en cualquier tienda de {ref.empresa.nombre} ({", ".join(tiendas)})'
             if ref.empresa_id and tiendas else 'en cualquiera de nuestras tiendas')
    detalle_txt = '\n'.join(
        (f"{t['beneficiario']}: " if t['beneficiario'] else '')
        + f"{t['codigo']}  (${t['monto']:,})"
        for t in tarjetas
    )
    plural = len(tarjetas) > 1
    texto = (
        f'{marca} - GIFT CARD\n\n'
        f'{saludo}has recibido {len(tarjetas)} gift card{"s" if plural else ""} '
        f'por un total de ${monto_total:,}.\n\n'
        f'{"TUS CODIGOS" if plural else "TU CODIGO"}:\n{detalle_txt}\n\n'
        + (f'Validas hasta el {fecha_venc_humana}.\n\n' if fecha_venc_humana else '')
        + f'Preséntala{"s" if plural else ""} en caja {donde}. Si no usas todo el '
        'monto, el saldo queda disponible para una próxima compra dentro de la '
        'vigencia.\n\n'
        'IMPORTANTE - cuidado del código: esta gift card funciona al portador; '
        'cualquier persona que presente el código puede canjearla. La custodia del '
        'código es responsabilidad exclusiva del titular. No se repone el saldo '
        'canjeado por terceros en caso de pérdida, robo o divulgación. No es '
        'canjeable por dinero en efectivo.\n'
    ).replace(',', '.')

    asunto = (
        f'\U0001F381 Tus {len(tarjetas)} Gift Cards {marca} — ${monto_total:,}'
        if plural else f'\U0001F381 Tu Gift Card {marca} — ${monto_total:,}'
    ).replace(',', '.')

    propia = connection is None
    if propia:
        connection = get_connection(
            timeout=int(os.environ.get('GIFTCARD_EMAIL_TIMEOUT', '30'))
        )
    correo = EmailMultiAlternatives(
        subject=asunto,
        body=texto,
        from_email=dj_settings.DEFAULT_FROM_EMAIL,
        to=[email_destino],
        connection=connection,
    )
    correo.attach_alternative(html, 'text/html')
    try:
        enviados = correo.send(fail_silently=False)
    except Exception as e:
        logger.exception("Error al enviar correo de gift cards destino=%s", email_destino)
        raise CorreoGiftCardError(f'No se pudo enviar el correo: {e}.')
    finally:
        if propia:
            try:
                connection.close()
            except Exception:
                pass

    if not enviados:
        # El servidor NO aceptó el mensaje: no se registra como enviado (si no,
        # el sistema mostraría un envío que nunca ocurrió).
        logger.error("Correo de gift cards NO aceptado destino=%s", email_destino)
        raise CorreoGiftCardError(
            'El servidor de correo rechazó el mensaje (0 enviados). '
            'Revisa la dirección del destinatario.'
        )

    # Message-ID: permite rastrear ESTE envío en el panel del proveedor
    # (MailerSend/Gmail) si el destinatario dice no haberlo recibido.
    message_id = (correo.extra_headers or {}).get('Message-ID', '')
    if not message_id:
        try:
            message_id = correo.message()['Message-ID'] or ''
        except Exception:
            message_id = ''

    enviado_en = timezone.now()
    for gc in giftcards:
        # Sin idempotency_key: reenviar el código es legítimo y cada envío debe
        # quedar como fila propia del ledger.
        MovimientoGiftCard.objects.create(
            giftcard=gc,
            tipo='ENVIO_CORREO',
            monto=0,
            saldo_resultante=gc.saldo_actual,
            usuario=usuario,
            sucursal=sucursal,
            observaciones=(
                f'Código enviado a {email_destino}'
                + (f' ({nombre_destino})' if nombre_destino else '')
                + (f' — para {beneficiarios[gc.id]}'
                   if beneficiarios.get(gc.id) else '')
                + (f' [msg {message_id}]' if message_id else '')
            ),
        )
    GiftCard.objects.filter(pk__in=[g.pk for g in giftcards]).update(
        correo_enviado_a=email_destino,
        correo_enviado_en=enviado_en,
        correo_envios=F('correo_envios') + 1,
        correo_message_id=message_id[:255],
    )
    logger.info("Gift cards enviadas por correo n=%s destino=%s msg=%s",
                len(giftcards), email_destino, message_id)
    return {
        'enviado_en': enviado_en,
        'message_id': message_id,
        'tarjetas': len(giftcards),
        'monto_total': monto_total,
    }


@require_POST
@requiere_permiso('giftcards_emitir', 'puede_crear')
def api_enviar_correo_giftcard(request):
    """
    Envía por correo el código de una gift card al destinatario indicado.

    El correo es el CANAL DE ENTREGA del código (instrumento al portador):
    incluye monto, código, vigencia y el descargo de responsabilidad de
    custodia. Cada envío queda en el ledger (fila ENVIO_CORREO monto=0) con
    destinatario, usuario y sucursal — trazable en la trazabilidad global.

    Patrón de envío del proyecto (requerimientos): EmailMultiAlternatives con
    texto plano + alternativa HTML y get_connection(timeout=...) para que un
    SMTP caído no deje la request colgada.
    """
    from django.conf import settings as dj_settings
    from django.core.exceptions import ValidationError as DjangoValidationError
    from django.core.mail import EmailMultiAlternatives, get_connection
    from django.core.validators import validate_email
    from django.template.loader import render_to_string

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'}, status=400)

    codigo = (data.get('codigo') or '').strip().upper()
    email_destino = (data.get('email') or '').strip()
    nombre_destino = (data.get('nombre') or '').strip()

    if not codigo:
        return JsonResponse({'success': False, 'error': 'Código requerido.'}, status=400)
    try:
        validate_email(email_destino)
    except DjangoValidationError:
        return JsonResponse({'success': False, 'error': 'Correo de destino inválido.'}, status=400)

    gc = GiftCard.objects.select_related('cliente', 'empresa').filter(
        Q(codigo=codigo) | Q(codigo_fisico=codigo)
    ).first()
    if gc is None:
        return JsonResponse({'success': False, 'error': 'Gift card no encontrada.'}, status=404)
    # Solo se envía lo canjeable: mandar el diseño de regalo con una tarjeta
    # vencida/agotada/bloqueada garantiza un reclamo en caja.
    if gc.estado in ('ANULADA', 'BLOQUEADA'):
        return JsonResponse({
            'success': False,
            'error': f'La gift card está {gc.get_estado_display().lower()}: '
                     'no corresponde enviar su código.',
        }, status=400)
    if gc.esta_vencida:
        return JsonResponse({
            'success': False,
            'error': f'La gift card venció el {gc.fecha_vencimiento}. '
                     'Extiende primero el vencimiento desde Gestionar.',
        }, status=400)
    if gc.saldo_actual <= 0:
        return JsonResponse({
            'success': False,
            'error': 'La gift card no tiene saldo: no corresponde enviar su código.',
        }, status=400)

    hoy = timezone.localdate()
    codigo_mostrar = (gc.codigo_fisico
                      if gc.tipo_tarjeta == 'FISICA' and gc.codigo_fisico else gc.codigo)
    vigencia_dias = ((gc.fecha_vencimiento - hoy).days
                     if gc.fecha_vencimiento else None)
    fecha_venc_humana = ''
    if gc.fecha_vencimiento:
        f = gc.fecha_vencimiento
        fecha_venc_humana = f'{f.day} de {_MESES_ES[f.month]} de {f.year}'

    # Tiendas donde vale: las de la empresa del ámbito (sin el CD); global = texto genérico.
    tiendas = []
    if gc.empresa_id:
        tiendas = list(
            Sucursal.objects.filter(empresa_id=gc.empresa_id)
            .exclude(tipo_sucursal='CENTRO_DISTRIBUCION')
            .order_by('alias').values_list('alias', flat=True)
        )

    try:
        resultado = enviar_codigos_por_correo(
            [gc], email_destino, nombre_destino=nombre_destino,
            usuario=request.user, sucursal=_sucursal_actual(request),
            beneficiarios={gc.id: (data.get('beneficiario') or '').strip()},
        )
    except CorreoGiftCardError as e:
        return JsonResponse({
            'success': False,
            'error': f'{e} La gift card no se modificó; corrige el problema e intenta de nuevo.',
        }, status=502)

    gc.refresh_from_db()
    return JsonResponse({
        'success': True,
        'destinatario': email_destino,
        'enviado_en': timezone.localtime(resultado['enviado_en']).strftime('%Y-%m-%d %H:%M'),
        'message_id': resultado['message_id'],
        'envios': gc.correo_envios,
        'saldo_enviado': gc.saldo_actual,
    })


# ========== TRAZABILIDAD GLOBAL ==========

@requiere_permiso('giftcards_listado', 'puede_ver')
def trazabilidad_giftcards_vista(request):
    """Log global de uso de gift cards (todas las tarjetas), con filtros."""
    context = {
        'sucursales': Sucursal.objects.all().order_by('nombre'),
        'tipo_mov_choices': MovimientoGiftCard._meta.get_field('tipo').choices,
    }
    return render(request, 'vistas/modulo_giftcards/trazabilidad.html', context)


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_trazabilidad_giftcards(request):
    """Listado paginado del ledger de TODAS las gift cards, con filtros."""
    qs = MovimientoGiftCard.objects.select_related(
        'giftcard', 'giftcard__cliente', 'ticket', 'usuario', 'sucursal'
    ).all()
    qs = _aplicar_filtros_trazabilidad(qs, request).order_by('-fecha')

    try:
        per_page = min(max(int(request.GET.get('per_page', 30) or 30), 5), 200)
    except (TypeError, ValueError):
        per_page = 30

    # Resumen por sucursal de los CANJES del filtro (nº y $ canjeado). Se
    # calcula sobre el universo filtrado COMPLETO, no la página, cuando el
    # filtro pide tipo=CONSUMO — es el reporte "gift cards usadas por sucursal".
    resumen_sucursal = None
    if (request.GET.get('tipo') or '').strip() == 'CONSUMO':
        resumen_sucursal = [{
            'sucursal': r['sucursal__nombre'] or 'Sin sucursal',
            'canjes': r['n'],
            'monto': -(r['s'] or 0),
        } for r in (
            qs.values('sucursal__nombre')
            .annotate(n=Count('id'), s=Sum('monto'))
            .order_by('s')   # monto más negativo = más canjeado, primero
        )]

    paginator = Paginator(qs, per_page)
    page = paginator.get_page(request.GET.get('page', 1))

    items = [{
        'fecha': m.fecha.strftime('%Y-%m-%d %H:%M'),
        'codigo': m.giftcard.codigo,
        'giftcard_id': m.giftcard_id,
        'cliente': m.giftcard.cliente.nombre_completo if m.giftcard.cliente else '',
        'tipo': m.tipo,
        'tipo_display': m.get_tipo_display(),
        'monto': m.monto,
        'saldo_resultante': m.saldo_resultante,
        'ticket': m.ticket.correlativo if m.ticket else None,
        # Boleta/factura de la venta que canjeó la tarjeta (si el ticket ya
        # tiene DTE emitido): lo que el usuario cruza contra el SII.
        'folio_dte': m.ticket.folio_dte if m.ticket else None,
        'tipo_dte': m.ticket.tipo_dte if m.ticket else None,
        'total_venta': m.ticket.total if m.ticket else None,
        'usuario': str(m.usuario) if m.usuario else '',
        'sucursal': m.sucursal.nombre if m.sucursal else '',
        'observaciones': m.observaciones or '',
    } for m in page]

    return JsonResponse({
        'success': True,
        'items': items,
        'page': page.number,
        'num_pages': paginator.num_pages,
        'total': paginator.count,
        'resumen_sucursal': resumen_sucursal,
    })


# ========== EXPORTES A EXCEL ==========

def _excel_response(wb, filename):
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _estilar_encabezados(ws, headers):
    import openpyxl  # noqa: F401  (asegura dependencia disponible)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="405189", end_color="405189", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_exportar_giftcards(request):
    """Exporta el listado de gift cards (respeta filtros) a Excel."""
    import openpyxl
    qs = GiftCard.objects.select_related('cliente', 'sucursal_emision', 'empresa').all()
    qs = _aplicar_filtros_giftcards(qs, request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gift Cards"
    headers = [
        'Código', 'Código físico', 'Tipo', 'Motivo', 'Descripción', 'Cliente',
        'Saldo inicial', 'Saldo actual', 'Estado', 'Emisión', 'Vencimiento', 'Sucursal',
        'Vigencia real', 'Ámbito', 'Pendiente por usar',
        'Correo enviado a', 'Fecha envío', 'N° envíos',
    ]
    _estilar_encabezados(ws, headers)
    hoy = timezone.localdate()
    estados_dict = dict(ESTADO_GIFTCARD_CHOICES)
    for row, gc in enumerate(qs.iterator(), 2):
        ws.cell(row=row, column=1, value=gc.codigo)
        ws.cell(row=row, column=2, value=gc.codigo_fisico or '')
        ws.cell(row=row, column=3, value=gc.get_tipo_tarjeta_display())
        ws.cell(row=row, column=4, value=gc.get_motivo_display())
        ws.cell(row=row, column=5, value=gc.descripcion or '')
        ws.cell(row=row, column=6, value=gc.cliente.nombre_completo if gc.cliente else '')
        ws.cell(row=row, column=7, value=gc.saldo_inicial)
        ws.cell(row=row, column=8, value=gc.saldo_actual)
        ws.cell(row=row, column=9, value=gc.get_estado_display())
        ws.cell(row=row, column=10, value=gc.fecha_emision.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=11, value=gc.fecha_vencimiento.strftime('%Y-%m-%d') if gc.fecha_vencimiento else '')
        ws.cell(row=row, column=12, value=gc.sucursal_emision.nombre if gc.sucursal_emision else '')
        _ef = _estado_efectivo(gc, hoy)
        ws.cell(row=row, column=13, value=estados_dict.get(_ef, _ef))
        ws.cell(row=row, column=14,
                value=gc.empresa.nombre if gc.empresa_id else 'Todas las empresas')
        ws.cell(row=row, column=15, value=(
            gc.saldo_actual if _ef == 'ACTIVA' and gc.saldo_actual > 0 else 0
        ))
        ws.cell(row=row, column=16, value=gc.correo_enviado_a or '')
        ws.cell(row=row, column=17, value=(
            timezone.localtime(gc.correo_enviado_en).strftime('%Y-%m-%d %H:%M')
            if gc.correo_enviado_en else ''
        ))
        ws.cell(row=row, column=18, value=gc.correo_envios or 0)
    return _excel_response(wb, 'giftcards.xlsx')


@require_GET
@requiere_permiso('giftcards_listado', 'puede_ver')
def api_exportar_trazabilidad(request):
    """Exporta la trazabilidad global de uso (respeta filtros) a Excel."""
    import openpyxl
    qs = MovimientoGiftCard.objects.select_related(
        'giftcard', 'giftcard__cliente', 'ticket', 'usuario', 'sucursal'
    ).all()
    qs = _aplicar_filtros_trazabilidad(qs, request).order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trazabilidad Gift Cards"
    headers = [
        'Fecha', 'Código', 'Cliente', 'Tipo', 'Monto', 'Saldo', 'Ticket',
        'Boleta/Factura', 'Tipo doc', 'Total venta',
        'Usuario', 'Sucursal', 'Observaciones',
    ]
    _estilar_encabezados(ws, headers)
    for row, m in enumerate(qs.iterator(), 2):
        ws.cell(row=row, column=1, value=m.fecha.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=m.giftcard.codigo)
        ws.cell(row=row, column=3, value=m.giftcard.cliente.nombre_completo if m.giftcard.cliente else '')
        ws.cell(row=row, column=4, value=m.get_tipo_display())
        ws.cell(row=row, column=5, value=m.monto)
        ws.cell(row=row, column=6, value=m.saldo_resultante)
        ws.cell(row=row, column=7, value=m.ticket.correlativo if m.ticket else '')
        ws.cell(row=row, column=8, value=(m.ticket.folio_dte or '') if m.ticket else '')
        ws.cell(row=row, column=9, value=(m.ticket.tipo_dte or '') if m.ticket else '')
        ws.cell(row=row, column=10, value=m.ticket.total if m.ticket else '')
        ws.cell(row=row, column=11, value=str(m.usuario) if m.usuario else '')
        ws.cell(row=row, column=12, value=m.sucursal.nombre if m.sucursal else '')
        ws.cell(row=row, column=13, value=m.observaciones or '')
    return _excel_response(wb, 'trazabilidad_giftcards.xlsx')
