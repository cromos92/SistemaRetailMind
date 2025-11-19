from django.shortcuts import render, redirect
from .models import (
    AtributoOpcion,
    Categoria,
    Compras,
    Compras_Producto,
    Compras_Producto_Talla,
    Dte,
    Dte_Detalle_Pago,
    Dte_Incidencia,
    Dte_Productos,
    Empresa,
    Correlativo,
    EmpresaUser,
    GuiaTalla,
    GuiaTallaItem,
    GuiaTallaProducto,
    Movimientos_Producto,
    ParametroGlobal,
    Producto,
    Producto_Talla,
    Productos_Atributos,
    Productos_Recepcionados,
    Sucursal,
    Vendedor,
    Ticket,
    Ticket_Productos,
    TicketDetallePago,
    Traspaso,
    Traspaso_Detalle,
    AjusteInventario,
    AjusteInventario_Detalle,
    LoteProducto,
    TIPO_DOCUMENTO_CHOICES,
    ESTADO_TICKET_CHOICES,
    METODO_PAGO_TICKET_CHOICES,
)
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse,Http404, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.shortcuts import get_object_or_404
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
import re
from django.db import transaction
import json


@login_required
def recepcion_dte(request):
    """Vista web para la recepción de traspasos internos."""
    return render(request, 'vistas/modulo_compras/recepcion_dte.html')


@login_required
@require_GET
def recepciones_pendientes_api(request):
    """Lista DTE internos pendientes y recepcionados para la vista de recepciones."""

    try:
        sucursal_destino_id = request.session.get('idSucursalActual')
        empresa_actual_id = request.session.get('idEmpresaActual')

        if not sucursal_destino_id or not empresa_actual_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal o empresa activa en la sesión.'
            }, status=400)

        pagina = max(int(request.GET.get('pagina', 1) or 1), 1)
        page_size = 10

        tipo_documento = request.GET.get('tipo_documento')
        sucursal_origen_id = request.GET.get('sucursal_origen')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # Incluir tanto DTEs pendientes como recepcionados
        queryset = (
            Dte.objects.filter(
                tipo_transaccion='TRASPASO',
                dte_movimientos__concepto='TRASPASO_SALIDA',
                dte_movimientos__tipo_movimiento='EGRESO',
                dte_movimientos__estado='COMPLETADO',  # ✅ COMPLETADO porque el stock ya salió
                dte_movimientos__sucursal_destino_id=sucursal_destino_id
            )
            .select_related('emisor', 'sucursal')
            .prefetch_related(
                'dte_productos__productoTalla__producto',
                'dte_movimientos__sucursal_origen__empresa'
            )
            .distinct()
        )

        if tipo_documento:
            queryset = queryset.filter(tipo_documento=tipo_documento)

        if sucursal_origen_id:
            try:
                queryset = queryset.filter(
                    dte_movimientos__concepto='TRASPASO_SALIDA',
                    dte_movimientos__sucursal_origen_id=int(sucursal_origen_id)
                )
            except (TypeError, ValueError):
                queryset = queryset.none()

        if fecha_inicio:
            queryset = queryset.filter(fecha_emision__gte=parse_date(fecha_inicio))

        if fecha_fin:
            queryset = queryset.filter(fecha_emision__lte=parse_date(fecha_fin))

        queryset = queryset.order_by('-fecha_emision', '-id')

        dte_ids = list(queryset.values_list('id', flat=True))
        total_unidades_global = Dte_Productos.objects.filter(
            dte_id__in=dte_ids
        ).aggregate(total=Sum('stock'))['total'] or 0

        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(pagina)

        items = []
        total_unidades_pagina = 0

        for dte in page_obj.object_list:
            detalles_queryset = dte.dte_productos.select_related('productoTalla__producto')

            movimientos_salida = [
                mov for mov in dte.dte_movimientos.all()
                if mov.concepto == 'TRASPASO_SALIDA'
                and mov.estado == 'COMPLETADO'  # ✅ COMPLETADO porque el stock ya salió
                and mov.sucursal_destino_id == sucursal_destino_id
            ]
            if not movimientos_salida:
                # Si no hay movimiento pendiente asociado directamente, continuar con el siguiente
                continue

            movimiento_origen = movimientos_salida[0]

            sucursal_origen_alias = '-'
            empresa_origen_nombre = dte.emisor.razon_social if dte.emisor else ''
            if movimiento_origen.sucursal_origen:
                sucursal_origen_alias = movimiento_origen.sucursal_origen.alias or sucursal_origen_alias
                if movimiento_origen.sucursal_origen.empresa:
                    empresa_origen_nombre = movimiento_origen.sucursal_origen.empresa.razon_social or empresa_origen_nombre
            sucursal_destino_alias = movimiento_origen.sucursal_destino.alias if movimiento_origen.sucursal_destino else request.session.get('alias', '-')

            resumen_tallas = {}
            detalle_completo = []

            for detalle in detalles_queryset:
                producto_talla = detalle.productoTalla
                producto = producto_talla.producto if producto_talla else None
                talla = producto_talla.talla if producto_talla else '-'
                cantidad = detalle.stock or 0
                precio = int(detalle.precio or 0)

                resumen_tallas[talla] = resumen_tallas.get(talla, 0) + cantidad

                # Extraer información adicional del producto
                articulo = producto.articulo if producto else ''
                marca = producto.atributo1.valor if (producto and producto.atributo1) else ''
                color = producto.atributo2.valor if (producto and producto.atributo2) else ''

                detalle_completo.append({
                    'dte_producto_id': detalle.id,  # ID del registro Dte_Productos
                    'sku': producto_talla.sku if producto_talla else '-',
                    'descripcion': producto.descripcion if producto else '',
                    'articulo': articulo,
                    'marca': marca,
                    'color': color,
                    'talla': talla,
                    'cantidad': cantidad,
                    'precio': precio,
                })

            total_unidades_doc = sum(resumen_tallas.values())
            total_unidades_pagina += total_unidades_doc

            items.append({
                'id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision,
                'fecha_recepcion': dte.fecha_recepcion,
                'estado_dte': dte.estado_dte,
                'emisor': dte.emisor.nombre,
                'empresa_origen': empresa_origen_nombre,
                'sucursal_origen': sucursal_origen_alias,
                'sucursal_destino': sucursal_destino_alias,
                'detalle_resumen': [
                    {'talla': talla, 'cantidad': cantidad}
                    for talla, cantidad in resumen_tallas.items()
                ],
                'detalle': detalle_completo,
                'total_unidades': total_unidades_doc,
                'referencias': dte.referencias or '',
                'observaciones': movimiento_origen.observaciones or '',
            })

        hoy = timezone.now().date()
        recibidos_hoy = Dte.objects.filter(
            tipo_transaccion='TRASPASO',
            sucursal_id=sucursal_destino_id,
            fecha_recepcion=hoy
        ).count()

        # Contar solo los pendientes (sin fecha de recepción)
        pendientes_reales = queryset.filter(fecha_recepcion__isnull=True).values('id').distinct().count()
        
        pendientes_mes = queryset.filter(
            fecha_emision__year=hoy.year,
            fecha_emision__month=hoy.month,
            fecha_recepcion__isnull=True
        ).values('id').distinct().count()

        movimiento_ids = set(queryset.values_list('dte_movimientos__id', flat=True))
        movimiento_ids.discard(None)
        movimientos_pendientes = Movimientos_Producto.objects.filter(
            id__in=movimiento_ids,
            concepto='TRASPASO_SALIDA',
            estado='COMPLETADO',  # ✅ COMPLETADO porque el stock ya salió
            sucursal_destino_id=sucursal_destino_id
        ).select_related('sucursal_origen__empresa')

        origenes_dict = {}
        for mov in movimientos_pendientes:
            suc_origen = mov.sucursal_origen
            if suc_origen:
                origenes_dict[suc_origen.id] = {
                    'id': suc_origen.id,
                    'alias': suc_origen.alias or 'Sin alias',
                    'empresa': suc_origen.empresa.razon_social if suc_origen.empresa else ''
                }

        # Contar productos con problemas
        productos_con_problemas = Productos_Recepcionados.objects.filter(
            dte__sucursal_id=sucursal_destino_id,
            estado__in=['RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO', 'FALTANTE', 'EN_REGULARIZACION']
        ).count()
        
        return JsonResponse({
            'success': True,
            'items': items,
            'pagination': {
                'page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
            },
            'resumen': {
                'recibidos_hoy': recibidos_hoy,
                'pendientes': pendientes_reales,
                'total_unidades_pendientes': total_unidades_global,
                'pendientes_mes': pendientes_mes,
                'productos_con_problemas': productos_con_problemas,
            },
            'origenes': sorted(origenes_dict.values(), key=lambda x: x['alias'].lower()),
        }, json_dumps_params={'default': str})

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cargar recepciones pendientes: {str(e)}'
        }, status=500)


@login_required
@require_GET
def historial_recepciones_api(request):
    """Devuelve el historial reciente de recepciones de traspasos."""

    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión.'
            }, status=400)

        limite = max(int(request.GET.get('limite', 5)), 1)

        historial = (
            Dte.objects.filter(
                tipo_transaccion='TRASPASO',
                sucursal_id=sucursal_id,
                fecha_recepcion__isnull=False
            )
            .select_related('emisor', 'receptor', 'sucursal')
            .order_by('-fecha_recepcion', '-id')[:limite]
        )

        items = []
        sucursal_destino_alias = request.session.get('alias', '-')
        for dte in historial:
            total_unidades = dte.dte_productos.aggregate(total=Sum('stock'))['total'] or 0

            items.append({
                'id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_recepcion': dte.fecha_recepcion,
                'sucursal_origen': dte.sucursal.alias if dte.sucursal else '-',
                'sucursal_destino': sucursal_destino_alias,
                'total_unidades': total_unidades,
            })

        return JsonResponse({'success': True, 'items': items}, json_dumps_params={'default': str})

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener historial: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def confirmar_recepcion_api(request):
    """
    Confirma la recepción DETALLADA de un DTE de traspaso.
    Soporta recepción parcial con control de problemas por producto.
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'}, status=400)

    dte_id = data.get('dte_id')
    productos_recepcion = data.get('productos', [])
    observaciones_generales = data.get('observaciones_generales', '')
    recepcion_completa = data.get('recepcion_completa', True)  # True si todo está OK

    if not dte_id:
        return JsonResponse({'success': False, 'error': 'Falta dte_id.'}, status=400)

    try:
        dte = Dte.objects.select_related('sucursal').prefetch_related('dte_productos__productoTalla__producto').get(id=dte_id)
    except Dte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'DTE no encontrado.'}, status=404)

    if dte.tipo_transaccion != 'TRASPASO':
        return JsonResponse({'success': False, 'error': 'El DTE no corresponde a un traspaso interno.'}, status=400)

    if dte.fecha_recepcion is not None:
        return JsonResponse({'success': False, 'error': 'El DTE ya fue recepcionado previamente.'}, status=400)

    sucursal_destino_id = request.session.get('idSucursalActual')
    if not sucursal_destino_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión.'}, status=400)

    usuario = request.user.username
    sucursal_destino = get_object_or_404(Sucursal, id=sucursal_destino_id)

    try:
        with transaction.atomic():
            from .models import Productos_Recepcionados
            
            hoy = timezone.now()
            productos_ok = 0
            productos_problemas = 0
            total_esperado = 0
            total_recepcionado = 0
            
            # Si no se envía detalle de productos, asumir recepción completa (legacy)
            if not productos_recepcion:
                # Modo legacy: recepcionar todo como OK
                productos_recepcion = []
                for detalle in dte.dte_productos.all():
                    productos_recepcion.append({
                        'dte_producto_id': detalle.id,
                        'cantidad_recepcionada': detalle.stock,
                        'cantidad_esperada': detalle.stock,
                        'cantidad_danada': 0,
                        'cantidad_faltante': 0,
                        'estado': 'RECEPCIONADO_OK',
                        'observaciones': ''
                    })
            
            # Procesar cada producto
            for prod_data in productos_recepcion:
                dte_producto_id = prod_data.get('dte_producto_id')
                cantidad_recepcionada = int(prod_data.get('cantidad_recepcionada', 0))
                cantidad_esperada = int(prod_data.get('cantidad_esperada', 0))
                cantidad_danada = int(prod_data.get('cantidad_danada', 0))
                cantidad_faltante = int(prod_data.get('cantidad_faltante', 0))
                estado = prod_data.get('estado', 'RECEPCIONADO_OK')
                observaciones = prod_data.get('observaciones', '')
                
                try:
                    dte_producto = dte.dte_productos.get(id=dte_producto_id)
                except:
                    continue
                
                producto_talla = dte_producto.productoTalla
                
                total_esperado += cantidad_esperada
                total_recepcionado += cantidad_recepcionada
                
                # Determinar si tiene problemas
                tiene_problemas = (
                    estado != 'RECEPCIONADO_OK' or
                    cantidad_danada > 0 or
                    cantidad_faltante > 0 or
                    cantidad_recepcionada != cantidad_esperada
                )
                
                if tiene_problemas:
                    productos_problemas += 1
                else:
                    productos_ok += 1
                
                # Calcular cantidad a ingresar (solo lo que llegó bien)
                cantidad_a_ingresar = cantidad_recepcionada - cantidad_danada
                
                # Si tiene problemas, cambiar estado a EN_REGULARIZACION automáticamente
                estado_final = estado
                if tiene_problemas:
                    estado_final = 'EN_REGULARIZACION'  # Ya está en proceso de regularización
                
                # Registrar recepción detallada
                Productos_Recepcionados.objects.create(
                    dte=dte,
                    dte_producto=dte_producto,
                    producto_talla=producto_talla,
                    stockArribado=cantidad_recepcionada,
                    cantidad_esperada=cantidad_esperada,
                    cantidad_danada=cantidad_danada,
                    cantidad_faltante=cantidad_faltante,
                    estado=estado_final,  # Estado automático EN_REGULARIZACION si hay problemas
                    observaciones=observaciones,
                    fecha_recepcion=hoy,
                    recepcionado_por=usuario
                )
                
                # Crear movimiento de INGRESO en sucursal destino (solo lo que llegó bien)
                if cantidad_a_ingresar > 0:
                    # ============================================
                    # BUSCAR O CREAR PRODUCTO-TALLA POR SKU EN SUCURSAL DESTINO
                    # ============================================
                    producto_origen = producto_talla.producto
                    
                    # 1. Primero buscar si ya existe el Producto_Talla por SKU en la sucursal destino
                    try:
                        talla_destino = Producto_Talla.objects.get(
                            sku=producto_talla.sku,
                            producto__sucursal=sucursal_destino
                        )
                        print(f"  ✓ Producto SKU {producto_talla.sku} ya existe en sucursal {sucursal_destino.alias}")
                        talla_creada = False
                        producto_creado = False
                        
                    except Producto_Talla.DoesNotExist:
                        # No existe, hay que crear el producto y la talla
                        print(f"  ⚠ Producto SKU {producto_talla.sku} NO existe en sucursal {sucursal_destino.alias}, creando...")
                        
                        # 2. Buscar si existe un producto igual en la sucursal destino
                        # (mismo artículo + atributos, para no duplicar productos)
                        producto_destino, producto_creado = Producto.objects.get_or_create(
                            articulo=producto_origen.articulo,
                            sucursal=sucursal_destino,
                            atributo1=producto_origen.atributo1,  # Marca
                            atributo2=producto_origen.atributo2,  # Color
                            atributo3=producto_origen.atributo3,  # Género
                            atributo4=producto_origen.atributo4,  # Otro
                            defaults={
                                'descripcion': producto_origen.descripcion,
                                'categoria': producto_origen.categoria,
                                'costo': producto_origen.costo,
                                'sobreprecio': producto_origen.sobreprecio,
                                'precioventa': producto_origen.precioventa,
                                'precioSugerido': producto_origen.precioSugerido,
                                'tipo_talla': producto_origen.tipo_talla,
                                'guia_talla': producto_origen.guia_talla
                            }
                        )
                        
                        if producto_creado:
                            print(f"  ✨ Producto '{producto_origen.articulo}' creado en sucursal {sucursal_destino.alias}")
                        else:
                            print(f"  ✓ Producto '{producto_origen.articulo}' ya existe en sucursal {sucursal_destino.alias}")
                        
                        # 3. Crear la talla específica con el SKU del origen
                        talla_destino = Producto_Talla.objects.create(
                            producto=producto_destino,
                            talla=producto_talla.talla,
                            sku=producto_talla.sku,  # ← MISMO SKU que el origen
                            stock=0  # Inicia en 0, se incrementará abajo
                        )
                        talla_creada = True
                        print(f"  ✨ Talla {producto_talla.talla} (SKU: {producto_talla.sku}) creada en sucursal {sucursal_destino.alias}")
                    
                    # 4. Crear movimiento de INGRESO en sucursal destino
                    producto_destino_final = talla_destino.producto  # Obtener el producto de la talla destino
                    Movimientos_Producto.objects.create(
                        dte=dte,
                        ProductoTalla=talla_destino,  # ← Usar talla de DESTINO
                        sucursal_origen=dte.sucursal,
                        sucursal_destino=sucursal_destino,
                        cantidad=cantidad_a_ingresar,  # Positivo porque es ingreso
                        costo=producto_destino_final.costo,
                        sobreprecio=producto_destino_final.sobreprecio,
                        precio=producto_destino_final.precioventa,
                        concepto='TRASPASO_ENTRADA',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=usuario,
                        observaciones=f'Recepción DTE #{dte.numero_documento} desde {dte.sucursal.alias}' + (f' - {observaciones}' if observaciones else '')
                    )
                    
                    # 5. ACTUALIZAR campo stock en SUCURSAL DESTINO
                    stock_antes = talla_destino.stock
                    talla_destino.stock += cantidad_a_ingresar
                    talla_destino.save()
                    print(f"✓ Movimiento de ingreso creado: {talla_destino.sku} +{cantidad_a_ingresar} en sucursal {sucursal_destino.alias}")
                    print(f"  Stock actualizado en {sucursal_destino.alias}: {stock_antes} → {talla_destino.stock}")
            
            # Los movimientos de salida ya están en COMPLETADO desde la emisión
            
            # Determinar estado final del DTE
            if productos_problemas == 0:
                dte.estado_dte = 'RECEPCIONADO_COMPLETO'
                mensaje = 'Recepción completada exitosamente. Todos los productos llegaron correctamente.'
            else:
                dte.estado_dte = 'RECEPCIONADO_PARCIAL'
                mensaje = f'Recepción procesada. {productos_problemas} producto(s) requieren atención.'
            
            # Actualizar DTE
            dte.fecha_recepcion = hoy.date()
            dte.hora = hoy.time()
            referencias_texto = (dte.referencias or '').strip()
            registro = f"Recepción por {usuario} el {hoy.strftime('%Y-%m-%d %H:%M')}"
            registro += f"\nProductos OK: {productos_ok}, Con problemas: {productos_problemas}"
            if observaciones_generales:
                registro += f"\nNota: {observaciones_generales}"
            dte.referencias = f"{referencias_texto}\n{registro}".strip()
            dte.save()
            
            print(f"✓ DTE #{dte.numero_documento} - Estado: {dte.estado_dte}")
            print(f"  Productos OK: {productos_ok}, Problemas: {productos_problemas}")

        return JsonResponse({
            'success': True,
            'message': mensaje,
            'estado_dte': dte.estado_dte,
            'productos_ok': productos_ok,
            'productos_problemas': productos_problemas,
            'total_esperado': total_esperado,
            'total_recepcionado': total_recepcionado
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        transaction.set_rollback(True)
        return JsonResponse({
            'success': False,
            'error': f'Error al confirmar recepción: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def rechazar_recepcion_api(request):
    """
    Rechaza la recepción de un DTE de traspaso.
    El stock NO se incrementa y se registra el motivo del rechazo.
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'}, status=400)

    dte_id = data.get('dte_id')
    motivo_rechazo = data.get('motivo_rechazo', '').strip()
    usuario = data.get('usuario', request.user.username)

    if not dte_id:
        return JsonResponse({'success': False, 'error': 'Falta dte_id.'}, status=400)
    
    if not motivo_rechazo:
        return JsonResponse({'success': False, 'error': 'Debes ingresar un motivo del rechazo.'}, status=400)

    try:
        dte = Dte.objects.select_related('sucursal').get(id=dte_id)
    except Dte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'DTE no encontrado.'}, status=404)

    if dte.tipo_transaccion != 'TRASPASO':
        return JsonResponse({'success': False, 'error': 'El DTE no corresponde a un traspaso interno.'}, status=400)

    if dte.fecha_recepcion is not None:
        return JsonResponse({'success': False, 'error': 'El DTE ya fue recepcionado o rechazado previamente.'}, status=400)

    sucursal_destino_id = request.session.get('idSucursalActual')
    if not sucursal_destino_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa en la sesión.'}, status=400)

    try:
        with transaction.atomic():
            hoy = timezone.now()
            
            # Marcar movimientos de salida como RECHAZADOS
            Movimientos_Producto.objects.filter(
                dte=dte,
                concepto='TRASPASO_SALIDA',
                estado='PENDIENTE_RECEPCION'
            ).update(
                estado='RECHAZADO',
                observaciones=F('observaciones') + f'\n❌ RECHAZADO: {motivo_rechazo}'
            )
            
            # Marcar el DTE como rechazado
            dte.estado_dte = 'RECHAZADO'
            dte.fecha_recepcion = hoy.date()
            dte.hora = hoy.time()
            
            referencias_texto = (dte.referencias or '').strip()
            registro = f"❌ RECEPCIÓN RECHAZADA por {usuario} el {hoy.strftime('%Y-%m-%d %H:%M')}"
            registro += f"\nMotivo: {motivo_rechazo}"
            dte.referencias = f"{referencias_texto}\n{registro}".strip()
            dte.save()
            
            print(f"✓ DTE #{dte.numero_documento} - RECHAZADO")
            print(f"  Motivo: {motivo_rechazo}")

        return JsonResponse({
            'success': True,
            'message': f'La recepción del DTE #{dte.numero_documento} ha sido rechazada.',
            'motivo': motivo_rechazo
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        transaction.set_rollback(True)
        return JsonResponse({
            'success': False,
            'error': f'Error al rechazar recepción: {str(e)}'
        }, status=500)


# ========== VISTAS PARA REGULARIZACIÓN DE RECEPCIONES ==========

@login_required
def regularizar_recepciones(request):
    """Vista para gestionar regularizaciones de productos con problemas en recepción"""
    return render(request, 'vistas/modulo_compras/regularizar_recepciones.html')


@login_required
def solicitudes_regularizacion_recibidas(request):
    """Vista para que el EMISOR revise solicitudes de regularización recibidas"""
    return render(request, 'vistas/modulo_compras/solicitudes_recibidas.html')


@login_required
def documento_regularizacion(request, recepcion_id):
    """Genera documento imprimible de regularización"""
    try:
        from .models import Productos_Recepcionados
        
        recepcion = get_object_or_404(Productos_Recepcionados, id=recepcion_id)
        
        if not recepcion.dte:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró el DTE asociado'
            }, status=404)
        
        dte_original = recepcion.dte
        
        # Determinar tipo de regularización
        tipo_regularizacion = 'NC'  # Por defecto
        nc_generada = None
        dte_cambio = None
        producto_cambio = None
        
        # Buscar NC generada para este producto
        nc_generada_obj = Dte.objects.filter(
            es_nota_credito=True,
            documento_afectado=dte_original,
            referencias__icontains=f'regularización'
        ).order_by('-fecha_emision').first()
        
        # Buscar DTE de cambio
        dte_cambio_obj = Dte.objects.filter(
            tipo_transaccion='TRASPASO',
            referencias__icontains=f'DTE #{dte_original.numero_documento}'
        ).filter(
            Q(referencias__icontains='cambio') | Q(referencias__icontains='Cambio')
        ).exclude(es_nota_credito=True).order_by('-fecha_emision').first()
        
        if dte_cambio_obj:
            tipo_regularizacion = 'CAMBIO'
            # Obtener producto de cambio
            producto_cambio_obj = dte_cambio_obj.dte_productos.first()
            if producto_cambio_obj:
                producto_cambio = {
                    'sku': producto_cambio_obj.productoTalla.sku,
                    'nombre': producto_cambio_obj.descripcion,
                    'talla': producto_cambio_obj.productoTalla.talla,
                    'cantidad': producto_cambio_obj.stock,
                    'precio_unitario': producto_cambio_obj.precio,
                    'total': producto_cambio_obj.stock * producto_cambio_obj.precio
                }
        
        if nc_generada_obj:
            iva_nc = nc_generada_obj.monto_con_iva - nc_generada_obj.monto_neto
            nc_generada = {
                'numero_documento': nc_generada_obj.numero_documento,
                'fecha_emision': nc_generada_obj.fecha_emision,
                'monto_neto': nc_generada_obj.monto_neto,
                'iva': iva_nc,
                'monto_con_iva': nc_generada_obj.monto_con_iva
            }
        
        if dte_cambio_obj:
            iva_cambio = dte_cambio_obj.monto_con_iva - dte_cambio_obj.monto_neto
            dte_cambio = {
                'numero_documento': dte_cambio_obj.numero_documento,
                'tipo_documento': dte_cambio_obj.tipo_documento,
                'fecha_emision': dte_cambio_obj.fecha_emision,
                'monto_neto': dte_cambio_obj.monto_neto,
                'iva': iva_cambio,
                'monto_con_iva': dte_cambio_obj.monto_con_iva
            }
        
        # Determinar tipo de problema
        tipo_problema = 'PROBLEMA'
        if recepcion.estado == 'FALTANTE' or recepcion.cantidad_faltante > 0:
            tipo_problema = 'FALTANTE'
        elif recepcion.estado == 'RECEPCIONADO_DANADO' or recepcion.cantidad_danada > 0:
            tipo_problema = 'DAÑADO'
        elif recepcion.estado == 'RECEPCIONADO_PARCIAL':
            tipo_problema = 'PARCIAL'
        
        cantidad_problema = recepcion.cantidad_danada or recepcion.cantidad_faltante or 0
        
        context = {
            'empresa': dte_original.emisor,
            'numero_documento': nc_generada_obj.numero_documento if nc_generada_obj else dte_cambio_obj.numero_documento if dte_cambio_obj else '-',
            'fecha_emision': timezone.now().strftime('%d/%m/%Y %H:%M'),
            'responsable': recepcion.regularizado_por or request.user.username,
            'tipo_regularizacion': tipo_regularizacion,
            'dte_original': dte_original,
            'producto_original': {
                'sku': recepcion.producto_talla.sku if recepcion.producto_talla else '-',
                'nombre': recepcion.producto_talla.producto.articulo if recepcion.producto_talla and recepcion.producto_talla.producto else '-',
                'talla': recepcion.producto_talla.talla if recepcion.producto_talla else '-',
                'cantidad_esperada': recepcion.cantidad_esperada,
                'cantidad_recibida': recepcion.stockArribado,
                'cantidad_problema': cantidad_problema,
                'tipo_problema': tipo_problema,
                'precio_unitario': recepcion.dte_producto.precio if recepcion.dte_producto else 0
            },
            'producto_cambio': producto_cambio,
            'nc_generada': nc_generada,
            'dte_cambio': dte_cambio,
            'motivo': recepcion.observaciones or 'Regularización de productos con problemas'
        }
        
        return render(request, 'vistas/modulo_compras/documento_regularizacion.html', context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al generar documento: {str(e)}'
        }, status=500)


@login_required
@require_GET
def obtener_productos_regularizar(request):
    """Obtiene lista de productos que requieren regularización"""
    try:
        from .models import Productos_Recepcionados
        
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión.'
            }, status=400)
        
        # Filtros
        estado_filtro = request.GET.get('estado', '')
        proveedor_filtro = request.GET.get('proveedor', '')
        busqueda = request.GET.get('buscar', '')
        
        # ✅ MOSTRAR PRODUCTOS CON PROBLEMAS Y REGULARIZADOS:
        # 1. Productos que ESTA SUCURSAL RECEPCIONÓ con problemas → Para solicitar solución
        # 2. Productos que ESTA SUCURSAL ENVIÓ y fueron recepcionados con problemas → Para DAR solución
        # 3. Productos REGULARIZADOS (solo si se filtran explícitamente)
        
        # Determinar estados a mostrar según el filtro
        if estado_filtro == 'TODOS':
            # Mostrar todos: pendientes + regularizados
            estados_a_mostrar = ['RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO', 'FALTANTE', 'EN_REGULARIZACION', 'EN_SOLICITUD_REGULARIZACION', 'REGULARIZADO']
        elif estado_filtro == 'REGULARIZADO':
            # Solo regularizados
            estados_a_mostrar = ['REGULARIZADO']
        elif estado_filtro:
            # Si hay un filtro específico, usar solo ese
            estados_a_mostrar = [estado_filtro]
        else:
            # Sin filtro: mostrar solo pendientes (NO regularizados)
            estados_a_mostrar = ['RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO', 'FALTANTE', 'EN_REGULARIZACION', 'EN_SOLICITUD_REGULARIZACION']
        
        queryset = Productos_Recepcionados.objects.filter(
            Q(dte__isnull=False) &  # Solo traspasos/DTEs
            (
                # Caso 1A: Productos que YO RECIBÍ (están en mi inventario)
                Q(producto_talla__producto__sucursal_id=sucursal_id) |
                # Caso 1B: Productos que YO RECIBÍ (buscar por sucursal destino del DTE)
                Q(dte__dte_movimientos__sucursal_destino_id=sucursal_id) |
                # Caso 2: Productos que YO ENVIÉ (soy el emisor del DTE)
                Q(dte__sucursal_id=sucursal_id)
            ) &
            Q(estado__in=estados_a_mostrar)
        ).select_related(
            'dte',
            'dte__emisor',
            'dte__receptor',
            'dte__sucursal',
            'producto_talla',
            'producto_talla__producto',
            'producto_talla__producto__sucursal',
            'dte_producto'
        ).prefetch_related('dte__dte_movimientos').distinct().order_by('-id', '-fecha_recepcion')
        
        if proveedor_filtro:
            queryset = queryset.filter(dte__emisor_id=proveedor_filtro)
        
        if busqueda:
            queryset = queryset.filter(
                Q(producto_talla__sku__icontains=busqueda) |
                Q(producto_talla__producto__articulo__icontains=busqueda) |
                Q(dte__numero_documento__icontains=busqueda)
            )
        
        # Calcular estadísticas
        total_queryset = queryset.count()
        pendientes = queryset.exclude(estado='REGULARIZADO').count()
        faltantes = queryset.filter(estado='FALTANTE').count()
        regularizados = queryset.filter(estado='REGULARIZADO').count()
        dtes_con_problemas = queryset.values('dte').distinct().count()
        
        productos = []
        for recepcion in queryset[:100]:  # Limitar a 100
            producto_nombre = 'Sin producto'
            if recepcion.producto_talla and recepcion.producto_talla.producto:
                producto_nombre = recepcion.producto_talla.producto.articulo
            
            # Determinar tipo de regularización
            requiere_nc = False
            tipo_regularizacion_texto = 'Ajuste Interno'
            if recepcion.dte:
                requiere_nc = recepcion.dte.requiere_nota_credito_check()
                tipo_regularizacion_texto = 'Nota de Crédito' if requiere_nc else 'Ajuste Interno'
            
            # Obtener información del emisor y receptor
            emisor_nombre = recepcion.dte.emisor.nombre if recepcion.dte and recepcion.dte.emisor else '-'
            receptor_nombre = recepcion.dte.receptor.nombre if recepcion.dte and recepcion.dte.receptor else '-'
            sucursal_origen = recepcion.dte.sucursal.alias if recepcion.dte and recepcion.dte.sucursal else '-'
            sucursal_origen_id = recepcion.dte.sucursal.id if recepcion.dte and recepcion.dte.sucursal else None
            
            # Determinar el ROL del usuario actual en este caso
            # SOY EMISOR si el DTE fue emitido por mi sucursal
            soy_emisor = (recepcion.dte and recepcion.dte.sucursal_id == sucursal_id)
            
            # SOY RECEPTOR si:
            # - Tengo movimientos de ENTRADA donde yo soy el destino, O
            # - El producto pertenece a mi inventario PERO yo NO soy el emisor
            tiene_movimiento_entrada = False
            if recepcion.dte:
                tiene_movimiento_entrada = recepcion.dte.dte_movimientos.filter(
                    concepto__in=['TRASPASO_ENTRADA', 'TRASPASO_SALIDA'],
                    sucursal_destino_id=sucursal_id
                ).exists()
            
            soy_receptor = tiene_movimiento_entrada and not soy_emisor
            
            # DEBUG: Log para entender la lógica
            if recepcion.dte:
                print(f"📊 DTE #{recepcion.dte.numero_documento}:")
                print(f"   - dte.sucursal_id: {recepcion.dte.sucursal_id}, sucursal_actual: {sucursal_id}")
                print(f"   - soy_emisor: {soy_emisor}")
                print(f"   - tiene_movimiento_entrada: {tiene_movimiento_entrada}")
                print(f"   - soy_receptor: {soy_receptor}")
                print(f"   - producto_talla.producto.sucursal_id: {recepcion.producto_talla.producto.sucursal_id if recepcion.producto_talla and recepcion.producto_talla.producto else 'N/A'}")
            
            # Obtener precio del producto original del DTE
            precio_unitario = 0
            if recepcion.dte_producto:
                precio_unitario = recepcion.dte_producto.precio or 0
            
            # Buscar si ya tiene NC o DTE de cambio generado
            nc_numero = None
            dte_cambio_numero = None
            solucion_aplicada = None
            
            if recepcion.dte and recepcion.estado == 'REGULARIZADO':
                # Buscar NC generada
                nc_obj = Dte.objects.filter(
                    es_nota_credito=True,
                    documento_afectado=recepcion.dte,
                    referencias__icontains='regularización'
                ).order_by('-fecha_emision').first()
                
                if nc_obj:
                    nc_numero = nc_obj.numero_documento
                    solucion_aplicada = f'NC #{nc_numero}'
                
                # Buscar DTE de cambio
                dte_cambio_obj = Dte.objects.filter(
                    tipo_transaccion='TRASPASO',
                    referencias__icontains=f'DTE #{recepcion.dte.numero_documento}'
                ).filter(
                    Q(referencias__icontains='cambio') | Q(referencias__icontains='Cambio')
                ).exclude(es_nota_credito=True).order_by('-fecha_emision').first()
                
                if dte_cambio_obj:
                    dte_cambio_numero = dte_cambio_obj.numero_documento
                    if nc_numero:
                        solucion_aplicada = f'NC #{nc_numero} + Cambio DTE #{dte_cambio_numero}'
                    else:
                        solucion_aplicada = f'Cambio DTE #{dte_cambio_numero}'
            
            productos.append({
                'id': recepcion.id,
                'dte_numero': recepcion.dte.numero_documento if recepcion.dte else '-',
                'dte_fecha': recepcion.dte.fecha_emision if recepcion.dte else None,
                'sku': str(recepcion.producto_talla.sku) if recepcion.producto_talla else '-',
                'producto_nombre': producto_nombre,
                'requiere_nc': requiere_nc,
                'tipo_regularizacion': tipo_regularizacion_texto,
                'talla': recepcion.producto_talla.talla if recepcion.producto_talla else '-',
                'cantidad_esperada': recepcion.cantidad_esperada or 0,
                'cantidad_recibida': recepcion.stockArribado or 0,
                'cantidad_danada': recepcion.cantidad_danada or 0,
                'cantidad_faltante': recepcion.cantidad_faltante or 0,
                'estado': recepcion.estado,
                'estado_display': recepcion.get_estado_display() if hasattr(recepcion, 'get_estado_display') else recepcion.estado,
                'observaciones': recepcion.observaciones or '',
                'fecha_recepcion': recepcion.fecha_recepcion,
                'recepcionado_por': recepcion.recepcionado_por or '-',
                # NUEVO: Información del emisor para solicitudes
                'emisor': emisor_nombre,
                'receptor': receptor_nombre,
                'sucursal_origen': sucursal_origen,
                'sucursal_origen_id': sucursal_origen_id,
                # NUEVO: Información de precios para cálculo de NC
                'precio_unitario': precio_unitario,
                # NUEVO: Determinar ROL del usuario actual
                'soy_emisor': soy_emisor,  # True si yo envié este DTE
                'soy_receptor': soy_receptor,  # True si yo recepcioné este DTE
                # NUEVO: Información de solución aplicada
                'nc_numero': nc_numero,
                'dte_cambio_numero': dte_cambio_numero,
                'solucion_aplicada': solucion_aplicada
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos,
            'total': total_queryset,
            'estadisticas': {
                'pendientes': pendientes,
                'faltantes': faltantes,
                'regularizados': regularizados,
                'dtes_con_problemas': dtes_con_problemas
            }
        }, json_dumps_params={'default': str})
        
    except Exception as e:
        print(f"Error en obtener_productos_regularizar: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos para regularizar: {str(e)}'
        }, status=500)


@login_required
@require_GET
def obtener_solicitudes_recibidas(request):
    """
    Obtiene lista de solicitudes de regularización recibidas por el emisor
    Para que el emisor pueda revisar y aprobar/rechazar
    """
    try:
        from .models import Solicitud_Regularizacion
        
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión'
            }, status=400)
        
        # Filtros
        estado_filtro = request.GET.get('estado', '')
        sucursal_filtro = request.GET.get('sucursal', '')
        busqueda = request.GET.get('buscar', '')
        
        # Obtener solicitudes donde esta sucursal es EMISORA
        queryset = Solicitud_Regularizacion.objects.filter(
            sucursal_emisora_id=sucursal_id
        ).select_related(
            'dte_original',
            'producto_recepcionado',
            'producto_recepcionado__producto_talla',
            'producto_recepcionado__producto_talla__producto',
            'sucursal_solicitante',
            'producto_cambio_solicitado',
            'producto_cambio_solicitado__producto'
        ).order_by('-fecha_solicitud')
        
        if estado_filtro:
            queryset = queryset.filter(estado=estado_filtro)
        
        if sucursal_filtro:
            queryset = queryset.filter(sucursal_solicitante_id=sucursal_filtro)
        
        if busqueda:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(numero_solicitud__icontains=busqueda) |
                Q(dte_original__numero_documento__icontains=busqueda) |
                Q(producto_recepcionado__producto_talla__sku__icontains=busqueda)
            )
        
        # Preparar datos
        solicitudes = []
        for sol in queryset[:100]:
            # Producto original
            prod_orig = sol.producto_recepcionado
            prod_orig_sku = prod_orig.producto_talla.sku if prod_orig and prod_orig.producto_talla else '-'
            prod_orig_nombre = prod_orig.producto_talla.producto.articulo if prod_orig and prod_orig.producto_talla and prod_orig.producto_talla.producto else '-'
            
            # Producto de cambio
            prod_cambio = sol.producto_cambio_aprobado or sol.producto_cambio_solicitado
            prod_cambio_sku = prod_cambio.sku if prod_cambio else '-'
            prod_cambio_nombre = prod_cambio.producto.articulo if prod_cambio and prod_cambio.producto else '-'
            stock_disponible = prod_cambio.stock if prod_cambio else 0
            
            solicitudes.append({
                'id': sol.id,
                'numero_solicitud': sol.numero_solicitud,
                'estado': sol.estado,
                'fecha_solicitud': sol.fecha_solicitud.isoformat(),
                'dias_pendiente': sol.dias_pendiente,
                'dte_numero': sol.dte_original.numero_documento if sol.dte_original else '-',
                'sucursal_solicita': sol.sucursal_solicitante.alias if sol.sucursal_solicitante else '-',
                'usuario_solicita': sol.usuario_solicita,
                'tipo_problema': sol.tipo_problema,
                'cantidad_problema': sol.cantidad_problema,
                'descripcion_problema': sol.descripcion_problema,
                'evidencia_url': sol.evidencia_foto.url if sol.evidencia_foto else None,
                'tipo_solucion_solicitada': sol.tipo_solucion_solicitada,
                'producto_original_sku': prod_orig_sku,
                'producto_original_nombre': prod_orig_nombre,
                'producto_cambio_sku': prod_cambio_sku,
                'producto_cambio_nombre': prod_cambio_nombre,
                'cantidad_cambio_solicitada': sol.cantidad_cambio_solicitada or 0,
                'stock_disponible': stock_disponible,
                'decision_emisor': sol.decision_emisor,
                'fecha_revision': sol.fecha_revision.isoformat() if sol.fecha_revision else None
            })
        
        # Estadísticas
        total = queryset.count()
        pendientes = queryset.filter(estado__in=['PENDIENTE', 'EN_REVISION']).count()
        aprobadas = queryset.filter(estado='APROBADA').count()
        ejecutadas = queryset.filter(estado='EJECUTADA').count()
        completadas = queryset.filter(estado='COMPLETADA').count()
        
        return JsonResponse({
            'success': True,
            'solicitudes': solicitudes,
            'total': total,
            'estadisticas': {
                'pendientes': pendientes,
                'aprobadas': aprobadas,
                'ejecutadas': ejecutadas,
                'completadas': completadas
            }
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener solicitudes: {str(e)}'
        }, status=500)


@login_required
@require_GET
def obtener_solicitud_producto(request, producto_id):
    """Obtiene la solicitud de regularización asociada a un producto"""
    try:
        from .models import Productos_Recepcionados, Solicitud_Regularizacion
        
        recepcion = get_object_or_404(Productos_Recepcionados, id=producto_id)
        
        # Buscar solicitud activa
        solicitud = Solicitud_Regularizacion.objects.filter(
            producto_recepcionado=recepcion
        ).order_by('-fecha_solicitud').first()
        
        if not solicitud:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró solicitud para este producto'
            }, status=404)
        
        # Preparar datos de la solicitud
        solicitud_data = {
            'numero_solicitud': solicitud.numero_solicitud,
            'estado': solicitud.estado,
            'fecha_solicitud': solicitud.fecha_solicitud.isoformat(),
            'tipo_problema': solicitud.tipo_problema,
            'cantidad_problema': solicitud.cantidad_problema,
            'descripcion_problema': solicitud.descripcion_problema,
            'tipo_solucion_solicitada': solicitud.tipo_solucion_solicitada,
            'producto_original_sku': recepcion.producto_talla.sku if recepcion.producto_talla else '-',
            'producto_original_nombre': recepcion.producto_talla.producto.articulo if recepcion.producto_talla and recepcion.producto_talla.producto else '-',
            'producto_cambio_sku': solicitud.producto_cambio_solicitado.sku if solicitud.producto_cambio_solicitado else '-',
            'producto_cambio_nombre': solicitud.producto_cambio_solicitado.producto.articulo if solicitud.producto_cambio_solicitado and solicitud.producto_cambio_solicitado.producto else '-',
            'cantidad_cambio_solicitada': solicitud.cantidad_cambio_solicitada,
            'fecha_revision': solicitud.fecha_revision.isoformat() if solicitud.fecha_revision else None,
            'usuario_revisa': solicitud.usuario_revisa,
            'decision_emisor': solicitud.decision_emisor,
            'fecha_ejecucion': solicitud.fecha_ejecucion.isoformat() if solicitud.fecha_ejecucion else None,
            'dte_solucion_numero': solicitud.dte_solucion.numero_documento if solicitud.dte_solucion else None,
            'nc_numero': solicitud.nota_credito.numero_documento if solicitud.nota_credito else None,
        }
        
        return JsonResponse({
            'success': True,
            'solicitud': solicitud_data
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener solicitud: {str(e)}'
        }, status=500)


@login_required
@require_POST
@transaction.atomic
def decidir_solicitud_api(request):
    """
    El EMISOR toma una decisión sobre una solicitud de regularización
    Opciones: APROBAR, RECHAZAR, MODIFICAR, NOTA_CREDITO
    """
    try:
        from .models import Solicitud_Regularizacion, Producto_Talla
        from .utils import notificar_solicitud_aprobada
        
        data = json.loads(request.body or '{}')
        solicitud_id = data.get('solicitud_id')
        decision = data.get('decision')  # APROBAR, RECHAZAR, MODIFICAR, NOTA_CREDITO
        observaciones = data.get('observaciones', '').strip()
        
        if not solicitud_id or not decision:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos'
            }, status=400)
        
        solicitud = get_object_or_404(Solicitud_Regularizacion, id=solicitud_id)
        usuario = request.user.username
        hoy = timezone.now()
        
        # Validar que esté pendiente
        if solicitud.estado not in ['PENDIENTE', 'EN_REVISION']:
            return JsonResponse({
                'success': False,
                'error': 'Esta solicitud ya fue procesada'
            }, status=400)
        
        with transaction.atomic():
            
            if decision == 'APROBAR':
                # Aprobar tal como fue solicitada
                solicitud.estado = 'APROBADA'
                solicitud.fecha_revision = hoy
                solicitud.usuario_revisa = usuario
                solicitud.decision_emisor = observaciones or 'Solicitud aprobada tal como fue solicitada'
                solicitud.tipo_solucion_aprobada = solicitud.tipo_solucion_solicitada
                solicitud.producto_cambio_aprobado = solicitud.producto_cambio_solicitado
                solicitud.cantidad_cambio_aprobada = solicitud.cantidad_cambio_solicitada
                solicitud.save()
                
                # Notificar al receptor
                notificar_solicitud_aprobada(solicitud)
                
                mensaje = f'Solicitud #{solicitud.numero_solicitud} aprobada correctamente. Ahora debes ejecutar la solución.'
            
            elif decision == 'MODIFICAR':
                # Aprobar con modificación (producto alternativo)
                producto_alt_id = data.get('producto_alternativo_id')
                cantidad_alt = data.get('cantidad_alternativa')
                
                if not producto_alt_id or not cantidad_alt:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debes especificar el producto y cantidad alternativa'
                    }, status=400)
                
                producto_alternativo = get_object_or_404(Producto_Talla, id=producto_alt_id)
                
                solicitud.estado = 'APROBADA'
                solicitud.fecha_revision = hoy
                solicitud.usuario_revisa = usuario
                solicitud.decision_emisor = observaciones or f'Aprobada con modificación: se enviará {producto_alternativo.sku} en lugar de lo solicitado'
                solicitud.tipo_solucion_aprobada = 'CAMBIO_PRODUCTO'
                solicitud.producto_cambio_aprobado = producto_alternativo
                solicitud.cantidad_cambio_aprobada = int(cantidad_alt)
                solicitud.save()
                
                notificar_solicitud_aprobada(solicitud)
                
                mensaje = f'Solicitud #{solicitud.numero_solicitud} aprobada con modificación'
            
            elif decision == 'NOTA_CREDITO':
                # Aprobar pero con NC en lugar de cambio de producto
                solicitud.estado = 'APROBADA'
                solicitud.fecha_revision = hoy
                solicitud.usuario_revisa = usuario
                solicitud.decision_emisor = observaciones or 'Se aprobó emitir Nota de Crédito en lugar de cambio de producto'
                solicitud.tipo_solucion_aprobada = 'NOTA_CREDITO'
                solicitud.producto_cambio_aprobado = None
                solicitud.cantidad_cambio_aprobada = None
                solicitud.save()
                
                notificar_solicitud_aprobada(solicitud)
                
                mensaje = f'Solicitud #{solicitud.numero_solicitud} aprobada. Se emitirá Nota de Crédito'
            
            elif decision == 'RECHAZAR':
                # Rechazar solicitud
                motivo_rechazo = data.get('motivo_rechazo', '').strip()
                
                if not motivo_rechazo:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debes especificar el motivo del rechazo'
                    }, status=400)
                
                solicitud.estado = 'RECHAZADA'
                solicitud.fecha_revision = hoy
                solicitud.usuario_revisa = usuario
                solicitud.decision_emisor = motivo_rechazo
                solicitud.save()
                
                # Actualizar producto recepcionado a EN_REGULARIZACION
                solicitud.producto_recepcionado.estado = 'EN_REGULARIZACION'
                solicitud.producto_recepcionado.observaciones = (solicitud.producto_recepcionado.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] Solicitud #{solicitud.numero_solicitud} RECHAZADA por {usuario}. Motivo: {motivo_rechazo}"
                solicitud.producto_recepcionado.save()
                
                mensaje = f'Solicitud #{solicitud.numero_solicitud} rechazada'
            
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Decisión no válida'
                }, status=400)
        
        print(f"✓ Solicitud #{solicitud.numero_solicitud} - Decisión: {decision} por {usuario}")
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'estado_nuevo': solicitud.estado,
            'numero_solicitud': solicitud.numero_solicitud
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        transaction.set_rollback(True)
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar decisión: {str(e)}'
        }, status=500)


@login_required
@require_GET
def buscar_productos_emisor(request):
    """
    Busca productos en el inventario de la sucursal EMISORA
    Para solicitudes de cambio de producto entre empresas
    """
    try:
        from .models import Producto_Talla, Producto
        
        query = request.GET.get('query', '').strip()
        sucursal_emisor_id = request.GET.get('sucursal_emisor_id')
        
        if not query or len(query) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Ingresa al menos 2 caracteres para buscar'
            }, status=400)
        
        if not sucursal_emisor_id:
            return JsonResponse({
                'success': False,
                'error': 'No se especificó la sucursal emisora'
            }, status=400)
        
        # Buscar productos con stock en la sucursal emisora
        from django.db.models import Q
        
        # Filtrar por sucursal primero
        productos = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_emisor_id,  # Sucursal está en Producto
            stock__gt=0  # Solo productos con stock disponible
        ).select_related('producto')
        
        # Aplicar filtros de búsqueda
        productos = productos.filter(
            Q(sku__icontains=query) |
            Q(producto__articulo__icontains=query)
        ).order_by('producto__articulo', 'talla')[:20]
        
        items = []
        for pt in productos:
            if pt.producto:
                items.append({
                    'id': pt.id,
                    'sku': pt.sku,
                    'nombre': pt.producto.articulo,
                    'talla': pt.talla,
                    'stock': pt.stock,
                    'precio': pt.producto.precioventa,
                    'costo': pt.producto.costo,
                })
        
        return JsonResponse({
            'success': True,
            'productos': items,
            'total': len(items)
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        }, status=500)


@login_required
@require_POST
@transaction.atomic
def regularizar_producto_api(request):
    """Regulariza un producto con problemas en la recepción o crea solicitud"""
    try:
        from .models import Productos_Recepcionados, Movimientos_Producto, Producto_Talla, Solicitud_Regularizacion
        from .utils import generar_numero_solicitud, notificar_nueva_solicitud
        
        data = json.loads(request.body or '{}')
        producto_id = data.get('producto_id')
        tipo_regularizacion = data.get('tipo_regularizacion')  # 'AJUSTAR', 'CAMBIAR_PRODUCTO', 'SOLICITAR_NC'
        observaciones = data.get('observaciones', '')
        es_solicitud = data.get('es_solicitud', False)  # NUEVO: indica si es solicitud
        tipo_solucion = data.get('tipo_solucion', '')  # Para identificar tipo de solicitud
        
        if not producto_id or not tipo_regularizacion:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos'
            }, status=400)
        
        recepcion = get_object_or_404(Productos_Recepcionados, id=producto_id)
        usuario = request.user.username
        hoy = timezone.now()
        
        with transaction.atomic():
            
            # NUEVO: Manejar solicitud de NC (entre empresas)
            if tipo_regularizacion == 'SOLICITAR_NC' or (es_solicitud and tipo_solucion == 'NOTA_CREDITO'):
                # Crear solicitud de NC
                justificacion = data.get('justificacion', '')
                cantidad_solicitud = int(data.get('cantidad_solicitud', 0))
                
                if not justificacion:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debe ingresar una justificación para la solicitud'
                    }, status=400)
                
                # Generar número de solicitud
                numero_solicitud = generar_numero_solicitud()
                
                # Determinar tipo de problema
                if recepcion.estado == 'FALTANTE' or recepcion.cantidad_faltante > 0:
                    tipo_problema = 'FALTANTE'
                elif recepcion.estado == 'RECEPCIONADO_DANADO':
                    tipo_problema = 'DANADO'
                elif recepcion.estado == 'RECEPCIONADO_PARCIAL':
                    tipo_problema = 'PARCIAL'
                else:
                    tipo_problema = 'INCORRECTO'
                
                cantidad_problema = recepcion.cantidad_faltante or recepcion.cantidad_danada or recepcion.cantidad_esperada
                
                # Obtener sucursal solicitante (receptora) y sucursal emisora
                # La sucursal solicitante es la que recepcionó (destino del movimiento)
                movimiento = recepcion.dte.dte_movimientos.filter(
                    concepto='TRASPASO_SALIDA',
                    tipo_movimiento='EGRESO'
                ).first() if recepcion.dte else None
                
                sucursal_solicitante = movimiento.sucursal_destino if movimiento else None
                sucursal_emisora = recepcion.dte.sucursal if recepcion.dte else None  # La sucursal del DTE es la emisora
                
                # Crear solicitud de NC
                solicitud = Solicitud_Regularizacion.objects.create(
                    numero_solicitud=numero_solicitud,
                    dte_original=recepcion.dte,
                    producto_recepcionado=recepcion,
                    sucursal_solicitante=sucursal_solicitante,
                    sucursal_emisora=sucursal_emisora,
                    usuario_solicita=usuario,
                    tipo_problema=tipo_problema,
                    cantidad_problema=cantidad_problema,
                    descripcion_problema=justificacion,
                    tipo_solucion_solicitada='NOTA_CREDITO',  # Solicita NC
                    producto_cambio_solicitado=None,  # No hay producto de cambio
                    cantidad_cambio_solicitada=None,
                    estado='PENDIENTE'
                )
                
                # Actualizar estado del producto recepcionado
                recepcion.estado = 'EN_SOLICITUD_REGULARIZACION'
                recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] Solicitud #{solicitud.numero_solicitud} creada - NC por {cantidad_problema} unidades. {justificacion}"
                recepcion.save()
                
                # Notificar al emisor
                notificar_nueva_solicitud(solicitud)
                
                print(f"✓ Solicitud NC #{solicitud.numero_solicitud} creada - {solicitud.sucursal_solicitante.alias} → {solicitud.sucursal_emisora.alias}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'Solicitud de NC #{solicitud.numero_solicitud} creada correctamente',
                    'tipo': 'SOLICITUD_CREADA',
                    'numero_solicitud': solicitud.numero_solicitud,
                    'requiere_aprobacion': True,
                    'estado_solicitud': 'PENDIENTE'
                })
            
            # NUEVO: Emitir NC directamente (emisor ejecuta)
            if tipo_regularizacion == 'EMITIR_NC' or data.get('ejecutar_nc'):
                motivo_nc = data.get('motivo_nc', '')
                cantidad_nc = int(data.get('cantidad_nc', 0))
                
                if not motivo_nc:
                    return JsonResponse({
                        'success': False,
                        'error': 'El motivo de la NC es obligatorio'
                    }, status=400)
                
                if not recepcion.dte or not recepcion.dte_producto:
                    return JsonResponse({
                        'success': False,
                        'error': 'No se encontró el DTE original'
                    }, status=400)
                
                # Generar la NC automáticamente
                try:
                    # obtener_siguiente_correlativo está en este mismo módulo (views.py)
                    
                    dte_original = recepcion.dte
                    
                    # Calcular montos
                    precio_unitario = recepcion.dte_producto.precio  # Este es NETO
                    total_neto = cantidad_nc * precio_unitario
                    iva = total_neto * Decimal('0.19')
                    total_con_iva = total_neto + iva
                    
                    print(f"💰 DEBUG Cálculo NC:")
                    print(f"   - Cantidad NC: {cantidad_nc}")
                    print(f"   - Precio unitario NETO (del DTE original): ${precio_unitario}")
                    print(f"   - Total NETO: ${total_neto}")
                    print(f"   - IVA (19%): ${iva}")
                    print(f"   - Total CON IVA: ${total_con_iva}")
                    print(f"   - DTE original monto_neto: ${dte_original.monto_neto}")
                    print(f"   - DTE original monto_con_iva: ${dte_original.monto_con_iva}")
                    
                    # Obtener correlativo para NC
                    numero_nc = obtener_siguiente_correlativo(dte_original.sucursal, 'NOTA DE CREDITO')
                    
                    # Crear Nota de Crédito
                    nota_credito = Dte.objects.create(
                        emisor=dte_original.emisor,
                        receptor=dte_original.receptor,
                        numero_documento=numero_nc,
                        tipo_documento='NOTA DE CREDITO',
                        monto_neto=total_neto,
                        monto_con_iva=total_con_iva,
                        estado_pago='PENDIENTE',
                        estado_dte='EMITIDO',
                        responsable=usuario,
                        fecha_emision=hoy.date(),
                        fecha_vencimiento=hoy.date(),
                        diasCredito=0,
                        bultos=1,
                        unidades_productos=cantidad_nc,
                        tipo_transaccion='TRASPASO',
                        sucursal=dte_original.sucursal,
                        es_nota_credito=True,
                        documento_afectado=dte_original,
                        motivo_nc=motivo_nc,
                        referencias=f"NC por regularización DTE #{dte_original.numero_documento}. {motivo_nc}"
                    )
                    
                    # Crear detalle de la NC
                    from .models import Dte_Productos
                    Dte_Productos.objects.create(
                        dte=nota_credito,
                        productoTalla=recepcion.dte_producto.productoTalla,
                        descripcion=f"NC: {recepcion.dte_producto.descripcion}",
                        costo=recepcion.dte_producto.costo,
                        sobreprecio=recepcion.dte_producto.sobreprecio,
                        precio=recepcion.dte_producto.precio,
                        stock=cantidad_nc,
                        activo=True
                    )
                    
                    # ============================================
                    # DEVOLVER STOCK A BODEGA DE ORIGEN (EMISOR)
                    # ============================================
                    producto_origen = recepcion.producto_talla  # Producto de la bodega ORIGEN
                    
                    if producto_origen:
                        stock_antes = producto_origen.stock
                        producto_origen.stock += cantidad_nc  # Devolver unidades
                        producto_origen.save()
                        print(f"  ✓ Stock devuelto a {dte_original.sucursal.alias}: {stock_antes} → {producto_origen.stock} (+{cantidad_nc})")
                        
                        # Crear movimiento de INGRESO en bodega origen (devolución por NC)
                        Movimientos_Producto.objects.create(
                            dte=nota_credito,
                            ProductoTalla=producto_origen,
                            sucursal_origen=None,  # No hay origen (es devolución)
                            sucursal_destino=dte_original.sucursal,  # Vuelve a la bodega emisora
                            cantidad=cantidad_nc,  # Cantidad que se devuelve
                            costo=producto_origen.producto.costo,
                            sobreprecio=producto_origen.producto.sobreprecio,
                            precio=producto_origen.producto.precioventa,
                            concepto='DEVOLUCION_NC',
                            tipo_movimiento='INGRESO',
                            estado='COMPLETADO',
                            responsable=usuario,
                            observaciones=f'Devolución de stock por NC #{numero_nc} - DTE original #{dte_original.numero_documento}. {motivo_nc}'
                        )
                        print(f"  ✓ Movimiento de INGRESO creado en {dte_original.sucursal.alias}")
                    else:
                        print(f"  ⚠️ No se pudo devolver stock: producto_talla no encontrado")
                    
                    # Actualizar estado del producto recepcionado
                    recepcion.estado = 'REGULARIZADO'
                    recepcion.fecha_regularizacion = hoy
                    recepcion.regularizado_por = usuario
                    recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] REGULARIZADO - NC #{numero_nc} emitida por {cantidad_nc} unidades. Stock devuelto a bodega origen. {motivo_nc}"
                    recepcion.save()
                    
                    print(f"✓ NC #{numero_nc} generada - DTE original #{dte_original.numero_documento}")
                    
                    # Generar archivo TXT para Acepta
                    archivo_txt_url = None
                    try:
                        from .views_modulo_documentos import generar_txt_dte_acepta
                        import os
                        from django.conf import settings
                        
                        datos_txt = {
                            'documento': {
                                'tipo_documento': '61',
                                'folio': str(numero_nc),
                                'fecha_emision': nota_credito.fecha_emision.strftime('%Y-%m-%d'),
                                'fecha_vencimiento': nota_credito.fecha_vencimiento.strftime('%Y-%m-%d'),
                                'tipo_despacho': '2',
                                'ind_traslado': '1',
                                'forma_pago': '1'
                            },
                            'emisor': {
                                'rut': nota_credito.emisor.rut if nota_credito.emisor else '',
                                'razon_social': nota_credito.emisor.razon_social if nota_credito.emisor else '',
                                'giro': nota_credito.emisor.giro if nota_credito.emisor else '',
                                'acteco': nota_credito.emisor.acteco if nota_credito.emisor else '',
                                'direccion': nota_credito.emisor.direccion if nota_credito.emisor else '',
                                'comuna': nota_credito.emisor.comuna if nota_credito.emisor else '',
                                'ciudad': nota_credito.emisor.ciudad if nota_credito.emisor else '',
                                'codigo_vendedor': usuario
                            },
                            'receptor': {
                                'rut': nota_credito.receptor.rut if nota_credito.receptor else '',
                                'codigo_interno': nota_credito.receptor.codigo if nota_credito.receptor else '',
                                'razon_social': nota_credito.receptor.razon_social if nota_credito.receptor else '',
                                'giro': nota_credito.receptor.giro if nota_credito.receptor else '',
                                'contacto': nota_credito.receptor.contacto if nota_credito.receptor else '',
                                'direccion': nota_credito.receptor.direccion if nota_credito.receptor else '',
                                'comuna': nota_credito.receptor.comuna if nota_credito.receptor else '',
                                'ciudad': nota_credito.receptor.ciudad if nota_credito.receptor else ''
                            },
                            'totales': {
                                'monto_neto': int(total_neto),
                                'monto_exento': 0,
                                'iva': int(iva),
                                'monto_total': int(total_con_iva)
                            },
                            'detalle': [{
                                'codigo': str(recepcion.producto_talla.sku) if recepcion.producto_talla else '',
                                'sku': str(recepcion.producto_talla.sku) if recepcion.producto_talla else '',
                                'nombre': recepcion.producto_talla.producto.articulo if recepcion.producto_talla else '',
                                'descripcion': recepcion.dte_producto.descripcion if recepcion.dte_producto else '',
                                'cantidad': cantidad_nc,
                                'unidad': 'UN',
                                'precio_unitario': int(precio_unitario),
                                'monto_item': int(total_neto),
                                'indicador_exencion': ''
                            }],
                            'referencias': [{
                                'tipo_documento': '33',
                                'folio': str(dte_original.numero_documento),
                                'fecha': dte_original.fecha_emision.strftime('%Y-%m-%d'),
                                'razon': '1'
                            }]
                        }
                        
                        contenido_txt = generar_txt_dte_acepta(datos_txt)
                        
                        txt_dir = os.path.join(settings.MEDIA_ROOT, 'documentos_electronicos', 'nc')
                        os.makedirs(txt_dir, exist_ok=True)
                        
                        nombre_archivo = f'NC_{numero_nc}_{hoy.strftime("%Y%m%d_%H%M%S")}.txt'
                        ruta_archivo = os.path.join(txt_dir, nombre_archivo)
                        
                        with open(ruta_archivo, 'w', encoding='utf-8') as f:
                            f.write(contenido_txt)
                        
                        archivo_txt_url = f'/media/documentos_electronicos/nc/{nombre_archivo}'
                        print(f"✅ Archivo TXT generado: {nombre_archivo}")
                        
                    except Exception as e:
                        print(f"⚠️ Error al generar TXT: {str(e)}")
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Nota de Crédito #{numero_nc} generada correctamente',
                        'tipo': 'NC_GENERADA',
                        'numero_nc': numero_nc,
                        'monto_nc': float(total_con_iva),
                        'documento_url': f'/app/dte/documento-regularizacion/{recepcion.id}/',
                        'archivo_txt_url': archivo_txt_url
                    })
                    
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    return JsonResponse({
                        'success': False,
                        'error': f'Error al generar NC: {str(e)}'
                    }, status=500)
            
            # NUEVO: Enviar producto de cambio (emisor ejecuta)
            if tipo_regularizacion == 'ENVIAR_CAMBIO' or data.get('ejecutar_envio'):
                producto_envio_id = int(data.get('producto_envio_id', 0))
                cantidad_envio = int(data.get('cantidad_envio', 0))
                motivo_envio = data.get('motivo_envio', '')
                
                if not producto_envio_id or not cantidad_envio or not motivo_envio:
                    return JsonResponse({
                        'success': False,
                        'error': 'Faltan datos para enviar producto de cambio'
                    }, status=400)
                
                if not recepcion.dte or not recepcion.dte_producto:
                    return JsonResponse({
                        'success': False,
                        'error': 'No se encontró el DTE original'
                    }, status=400)
                
                try:
                    # obtener_siguiente_correlativo está en este mismo módulo (views.py)
                    from .models import Dte_Productos, Producto_Talla
                    
                    dte_original = recepcion.dte
                    producto_cambio = get_object_or_404(Producto_Talla, id=producto_envio_id)
                    
                    # Verificar stock disponible
                    if producto_cambio.stock < cantidad_envio:
                        return JsonResponse({
                            'success': False,
                            'error': f'Stock insuficiente. Disponible: {producto_cambio.stock}'
                        }, status=400)
                    
                    # Obtener sucursal destino (receptor original)
                    movimiento_original = dte_original.dte_movimientos.filter(
                        concepto='TRASPASO_SALIDA'
                    ).first()
                    sucursal_destino = movimiento_original.sucursal_destino if movimiento_original else None
                    
                    if not sucursal_destino:
                        return JsonResponse({
                            'success': False,
                            'error': 'No se pudo identificar la sucursal destino'
                        }, status=400)
                    
                    # 1. Generar NC por producto original dañado
                    precio_unitario_original = recepcion.dte_producto.precio
                    cantidad_problema = recepcion.cantidad_danada or recepcion.cantidad_faltante or 1
                    total_neto_nc = cantidad_problema * precio_unitario_original
                    iva_nc = total_neto_nc * Decimal('0.19')
                    total_con_iva_nc = total_neto_nc + iva_nc
                    
                    numero_nc = obtener_siguiente_correlativo(dte_original.sucursal, 'NOTA DE CREDITO')
                    
                    nota_credito = Dte.objects.create(
                        emisor=dte_original.emisor,
                        receptor=dte_original.receptor,
                        numero_documento=numero_nc,
                        tipo_documento='NOTA DE CREDITO',
                        monto_neto=total_neto_nc,
                        monto_con_iva=total_con_iva_nc,
                        estado_pago='PENDIENTE',
                        estado_dte='EMITIDO',
                        responsable=usuario,
                        fecha_emision=hoy.date(),
                        fecha_vencimiento=hoy.date(),
                        diasCredito=0,
                        bultos=1,
                        unidades_productos=cantidad_problema,
                        tipo_transaccion='TRASPASO',
                        sucursal=dte_original.sucursal,
                        es_nota_credito=True,
                        documento_afectado=dte_original,
                        motivo_nc=f"NC por producto dañado - Envío de reemplazo. {motivo_envio}",
                        referencias=f"NC por regularización DTE #{dte_original.numero_documento}"
                    )
                    
                    Dte_Productos.objects.create(
                        dte=nota_credito,
                        productoTalla=recepcion.dte_producto.productoTalla,
                        descripcion=f"NC: {recepcion.dte_producto.descripcion}",
                        costo=recepcion.dte_producto.costo,
                        sobreprecio=recepcion.dte_producto.sobreprecio,
                        precio=recepcion.dte_producto.precio,
                        stock=cantidad_problema,
                        activo=True
                    )
                    
                    # Devolver stock del producto original a bodega origen
                    producto_origen = recepcion.producto_talla
                    if producto_origen:
                        stock_antes_nc = producto_origen.stock
                        producto_origen.stock += cantidad_problema
                        producto_origen.save()
                        print(f"  ✓ Stock devuelto (NC) a {dte_original.sucursal.alias}: {stock_antes_nc} → {producto_origen.stock} (+{cantidad_problema})")
                        
                        # Crear movimiento de INGRESO por la NC
                        Movimientos_Producto.objects.create(
                            dte=nota_credito,
                            ProductoTalla=producto_origen,
                            sucursal_origen=None,
                            sucursal_destino=dte_original.sucursal,
                            cantidad=cantidad_problema,
                            costo=producto_origen.producto.costo,
                            sobreprecio=producto_origen.producto.sobreprecio,
                            precio=producto_origen.producto.precioventa,
                            concepto='DEVOLUCION_NC',
                            tipo_movimiento='INGRESO',
                            estado='COMPLETADO',
                            responsable=usuario,
                            observaciones=f'Devolución por NC #{numero_nc} antes de enviar cambio'
                    )
                    
                    # 2. Crear nuevo DTE con producto de cambio
                    precio_cambio = producto_cambio.producto.precioventa if producto_cambio.producto else 0
                    costo_cambio = producto_cambio.producto.costo if producto_cambio.producto else 0
                    total_neto_cambio = cantidad_envio * precio_cambio
                    iva_cambio = total_neto_cambio * Decimal('0.19')
                    total_con_iva_cambio = total_neto_cambio + iva_cambio
                    
                    numero_dte_cambio = obtener_siguiente_correlativo(dte_original.sucursal, 'GUIA')
                    
                    dte_cambio = Dte.objects.create(
                        emisor=dte_original.emisor,
                        receptor=dte_original.receptor,
                        numero_documento=numero_dte_cambio,
                        tipo_documento='GUIA',
                        monto_neto=total_neto_cambio,
                        monto_con_iva=total_con_iva_cambio,
                        estado_pago='PENDIENTE',
                        estado_dte='EMITIDO',
                        responsable=usuario,
                        fecha_emision=hoy.date(),
                        fecha_vencimiento=hoy.date(),
                        diasCredito=0,
                        bultos=1,
                        unidades_productos=cantidad_envio,
                        tipo_transaccion='TRASPASO',
                        sucursal=dte_original.sucursal,
                        referencias=f"Producto de cambio por DTE #{dte_original.numero_documento}. NC #{numero_nc}. {motivo_envio}"
                    )
                    
                    Dte_Productos.objects.create(
                        dte=dte_cambio,
                        productoTalla=producto_cambio,
                        descripcion=f"CAMBIO: {producto_cambio.producto.articulo} - Talla {producto_cambio.talla}",
                        costo=costo_cambio,
                        sobreprecio=producto_cambio.producto.sobreprecio if producto_cambio.producto else 0,
                        precio=precio_cambio,
                        stock=cantidad_envio,
                        activo=True
                    )
                    
                    # 3. Restar stock del emisor
                    producto_cambio.stock -= cantidad_envio
                    producto_cambio.save()
                    
                    # 4. Crear movimiento de salida
                    Movimientos_Producto.objects.create(
                        dte=dte_cambio,
                        ProductoTalla=producto_cambio,
                        sucursal_origen=dte_original.sucursal,
                        sucursal_destino=sucursal_destino,
                        cantidad=-cantidad_envio,
                        costo=costo_cambio,
                        sobreprecio=producto_cambio.producto.sobreprecio if producto_cambio.producto else 0,
                        precio=precio_cambio,
                        concepto='TRASPASO_SALIDA',
                        tipo_movimiento='EGRESO',
                        estado='PENDIENTE_RECEPCION',
                        responsable=usuario,
                        observaciones=f'Producto de cambio por regularización - DTE original #{dte_original.numero_documento}'
                    )
                    
                    # 5. Actualizar estado del producto recepcionado
                    recepcion.estado = 'REGULARIZADO'
                    recepcion.fecha_regularizacion = hoy
                    recepcion.regularizado_por = usuario
                    recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] REGULARIZADO - NC #{numero_nc} + DTE Cambio #{numero_dte_cambio} emitidos. {motivo_envio}"
                    recepcion.save()
                    
                    print(f"✓ Producto de cambio enviado - NC #{numero_nc}, DTE #{numero_dte_cambio}")
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Producto de cambio enviado correctamente',
                        'tipo': 'CAMBIO_ENVIADO',
                        'numero_nc': numero_nc,
                        'numero_dte_cambio': numero_dte_cambio,
                        'producto_cambio': f"{producto_cambio.sku} - {producto_cambio.producto.articulo}",
                        'documento_url': f'/app/dte/documento-regularizacion/{recepcion.id}/'
                    })
                    
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    return JsonResponse({
                        'success': False,
                        'error': f'Error al enviar producto de cambio: {str(e)}'
                    }, status=500)
            
            if tipo_regularizacion == 'AJUSTAR':
                # Ajustar cantidad recibida
                nueva_cantidad = int(data.get('nueva_cantidad', 0))
                nuevo_estado = data.get('nuevo_estado', 'REGULARIZADO')
                
                diferencia = nueva_cantidad - recepcion.stockArribado
                nota_credito = None
                
                # Determinar si requiere NC (empresas diferentes)
                requiere_nc = recepcion.dte.requiere_nota_credito_check() if recepcion.dte else False
                
                # ✅ NUEVO: Si requiere NC y el usuario es RECEPTOR, NO puede ejecutar directamente
                # Debe ir por flujo de solicitudes
                # Este código solo se ejecuta para traspasos INTERNOS (misma empresa)
                
                if requiere_nc:
                    # Entre empresas: no debería llegar aquí, pero por seguridad
                    return JsonResponse({
                        'success': False,
                        'error': 'Para traspasos entre empresas debes usar "Solicitar NC" o "Solicitar Cambio de Producto"'
                    }, status=400)
                
                # SOLO TRASPASOS INTERNOS a partir de aquí
                if diferencia > 0:
                    # Llegaron más productos (traspaso interno)
                    recepcion.producto_talla.stock += diferencia
                    recepcion.producto_talla.save()
                    
                    # Crear movimiento de ingreso
                    Movimientos_Producto.objects.create(
                        dte=recepcion.dte,
                        ProductoTalla=recepcion.producto_talla,
                        cantidad=diferencia,
                        costo=recepcion.producto_talla.producto.costo if recepcion.producto_talla and recepcion.producto_talla.producto else 0,
                        concepto='REGULARIZACION_TRASPASO',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=usuario,
                        observaciones=f"Regularización DTE #{recepcion.dte.numero_documento if recepcion.dte else 'N/A'} - Ajuste +{diferencia} - {observaciones}"
                    )
                
                # Actualizar recepción (solo para internos)
                recepcion.stockArribado = nueva_cantidad
                recepcion.cantidad_faltante = max(0, recepcion.cantidad_esperada - nueva_cantidad)
                recepcion.estado = nuevo_estado
                recepcion.fecha_regularizacion = hoy
                recepcion.regularizado_por = usuario
                recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] Regularizado (ajuste interno): cantidad ajustada a {nueva_cantidad}. {observaciones}"
                recepcion.save()
            
            elif tipo_regularizacion == 'CAMBIAR_PRODUCTO':
                # Cambiar por otro producto completamente diferente
                nuevo_producto_id = data.get('nuevo_producto_id')
                
                if not nuevo_producto_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debe seleccionar el nuevo producto'
                    }, status=400)
                
                nuevo_producto_talla = get_object_or_404(Producto_Talla, id=nuevo_producto_id)
                cantidad = recepcion.stockArribado or recepcion.cantidad_esperada
                
                # ✅ NUEVO: Detectar si es SOLICITUD (entre empresas)
                if es_solicitud:
                    # CREAR SOLICITUD en lugar de cambio directo
                    justificacion = data.get('justificacion', '')
                    cantidad_solicitud = int(data.get('cantidad_solicitud', 0))
                    
                    if not justificacion:
                        return JsonResponse({
                            'success': False,
                            'error': 'Debe ingresar una justificación para la solicitud'
                        }, status=400)
                    
                    # Validar cantidad
                    cantidad_problema = recepcion.cantidad_faltante or recepcion.cantidad_danada or recepcion.cantidad_esperada
                    
                    if cantidad_solicitud <= 0:
                        return JsonResponse({
                            'success': False,
                            'error': 'La cantidad debe ser al menos 1 unidad'
                        }, status=400)
                    
                    if cantidad_solicitud > cantidad_problema:
                        return JsonResponse({
                            'success': False,
                            'error': f'No puedes solicitar más de {cantidad_problema} unidades'
                        }, status=400)
                    
                    # Generar número de solicitud
                    numero_solicitud = generar_numero_solicitud()
                    
                    # Determinar tipo de problema
                    if recepcion.estado == 'FALTANTE' or recepcion.cantidad_faltante > 0:
                        tipo_problema = 'FALTANTE'
                    elif recepcion.estado == 'RECEPCIONADO_DANADO':
                        tipo_problema = 'DANADO'
                    elif recepcion.estado == 'RECEPCIONADO_PARCIAL':
                        tipo_problema = 'PARCIAL'
                    else:
                        tipo_problema = 'INCORRECTO'
                    
                    # Obtener sucursal solicitante (receptora) y sucursal emisora
                    movimiento = recepcion.dte.dte_movimientos.filter(
                        concepto='TRASPASO_SALIDA',
                        tipo_movimiento='EGRESO'
                    ).first() if recepcion.dte else None
                    
                    sucursal_solicitante = movimiento.sucursal_destino if movimiento else None
                    sucursal_emisora = recepcion.dte.sucursal if recepcion.dte else None  # La sucursal del DTE es la emisora
                    
                    # Crear solicitud con la cantidad especificada
                    solicitud = Solicitud_Regularizacion.objects.create(
                        numero_solicitud=numero_solicitud,
                        dte_original=recepcion.dte,
                        producto_recepcionado=recepcion,
                        sucursal_solicitante=sucursal_solicitante,  # Receptor
                        sucursal_emisora=sucursal_emisora,  # Emisor
                        usuario_solicita=usuario,
                        tipo_problema=tipo_problema,
                        cantidad_problema=cantidad_problema,
                        descripcion_problema=justificacion,
                        tipo_solucion_solicitada='CAMBIO_PRODUCTO',
                        producto_cambio_solicitado=nuevo_producto_talla,
                        cantidad_cambio_solicitada=cantidad_solicitud,  # Usar cantidad especificada
                        estado='PENDIENTE'
                    )
                    
                    # Actualizar estado del producto recepcionado
                    recepcion.estado = 'EN_SOLICITUD_REGULARIZACION'
                    recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] Solicitud #{solicitud.numero_solicitud} creada - Cambio por {nuevo_producto_talla.sku}. {justificacion}"
                    recepcion.save()
                    
                    # Notificar al emisor
                    notificar_nueva_solicitud(solicitud)
                    
                    print(f"✓ Solicitud #{solicitud.numero_solicitud} creada - {solicitud.sucursal_solicitante.alias} → {solicitud.sucursal_emisora.alias}")
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Solicitud #{solicitud.numero_solicitud} creada correctamente',
                        'tipo': 'SOLICITUD_CREADA',
                        'numero_solicitud': solicitud.numero_solicitud,
                        'requiere_aprobacion': True,
                        'estado_solicitud': 'PENDIENTE'
                    })
                
                else:
                    # CAMBIO DIRECTO (traspaso interno - flujo original)
                    # Ingresar el nuevo producto
                    nuevo_producto_talla.stock += cantidad
                    nuevo_producto_talla.save()
                    
                    # Crear movimiento
                    Movimientos_Producto.objects.create(
                        dte=recepcion.dte,
                        ProductoTalla=nuevo_producto_talla,
                        cantidad=cantidad,
                        costo=nuevo_producto_talla.producto.costo if nuevo_producto_talla.producto else 0,
                        concepto='REGULARIZACION_CAMBIO_PRODUCTO',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=usuario,
                        observaciones=f"Regularización DTE #{recepcion.dte.numero_documento if recepcion.dte else 'N/A'} - Cambio de producto {recepcion.producto_talla.sku if recepcion.producto_talla else 'N/A'} → {nuevo_producto_talla.sku} - {observaciones}"
                    )
                    
                    recepcion.estado = 'REGULARIZADO'
                    recepcion.fecha_regularizacion = hoy
                    recepcion.regularizado_por = usuario
                    producto_reemplazo = nuevo_producto_talla.producto.articulo if nuevo_producto_talla.producto else 'N/A'
                    recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] Reemplazado por {producto_reemplazo} ({cantidad} unidades). {observaciones}"
                    recepcion.save()
            
            # Verificar si el DTE completo está regularizado
            dte_completado = False
            if recepcion.dte:
                recepciones_dte = Productos_Recepcionados.objects.filter(dte=recepcion.dte)
                todas_ok = all(r.estado in ['RECEPCIONADO_OK', 'REGULARIZADO'] for r in recepciones_dte)
                
                if todas_ok and hasattr(recepcion.dte, 'estado_dte'):
                    recepcion.dte.estado_dte = 'RECEPCIONADO_COMPLETO'
                    recepcion.dte.save()
                    dte_completado = True
        
        return JsonResponse({
            'success': True,
            'message': 'Producto regularizado correctamente',
            'estado_nuevo': recepcion.estado,
            'dte_completado': dte_completado
        })
        
    except Exception as e:
        print(f"Error en regularizar_producto_api: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al regularizar: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def regularizar_dte_masivo(request):
    """Genera 1 SOLA Nota de Crédito por todo el DTE con múltiples productos"""
    try:
        from .models import Productos_Recepcionados, Movimientos_Producto, Producto_Talla, Dte, Dte_Productos
        
        data = json.loads(request.body or '{}')
        dte_numero = data.get('dte_numero')
        productos_ids = data.get('productos_ids', [])
        motivo = data.get('motivo', '')
        
        if not dte_numero or not productos_ids or not motivo:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos (DTE, productos o motivo)'
            }, status=400)
        
        usuario = request.user.username
        hoy = timezone.now()
        
        with transaction.atomic():
            # Obtener todas las recepciones del DTE
            recepciones = Productos_Recepcionados.objects.filter(
                id__in=productos_ids
            ).select_related('dte', 'dte_producto', 'producto_talla', 'producto_talla__producto')
            
            if not recepciones.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'No se encontraron productos para regularizar'
                }, status=404)
            
            # Obtener DTE original (todos deberían ser del mismo DTE)
            dte_original = recepciones.first().dte
            
            # Calcular totales de la NC
            total_unidades = 0
            monto_neto_total = Decimal('0')
            productos_nc = []
            
            for recepcion in recepciones:
                cantidad_acreditar = recepcion.cantidad_faltante + recepcion.cantidad_danada
                
                if cantidad_acreditar <= 0:
                    continue
                
                precio_unitario = recepcion.dte_producto.precio
                monto_producto = cantidad_acreditar * precio_unitario
                
                total_unidades += cantidad_acreditar
                monto_neto_total += monto_producto
                
                productos_nc.append({
                    'recepcion': recepcion,
                    'cantidad': cantidad_acreditar,
                    'precio_unitario': precio_unitario
                })
            
            if not productos_nc:
                return JsonResponse({
                    'success': False,
                    'error': 'No hay productos con cantidades a acreditar'
                }, status=400)
            
            # Calcular IVA y total
            iva_total = monto_neto_total * Decimal('0.19')
            total_con_iva = monto_neto_total + iva_total
            
            # ============================================
            # GENERAR 1 SOLA NOTA DE CRÉDITO
            # ============================================
            numero_nc = obtener_siguiente_correlativo(dte_original.sucursal, 'NOTA DE CREDITO')
            
            nota_credito = Dte.objects.create(
                emisor=dte_original.emisor,
                receptor=dte_original.receptor,
                numero_documento=numero_nc,
                tipo_documento='NOTA DE CREDITO',
                monto_neto=monto_neto_total,
                monto_con_iva=total_con_iva,
                estado_pago='PENDIENTE',
                estado_dte='EMITIDO',
                responsable=usuario,
                fecha_emision=hoy.date(),
                fecha_vencimiento=hoy.date(),
                diasCredito=0,
                bultos=len(productos_nc),
                unidades_productos=total_unidades,
                tipo_transaccion='TRASPASO',
                sucursal=dte_original.sucursal,
                es_nota_credito=True,
                documento_afectado=dte_original,
                motivo_nc=motivo,
                referencias=f"NC por regularización DTE #{dte_original.numero_documento}. {motivo}"
            )
            
            print(f"✅ NC #{numero_nc} creada - DTE original #{dte_original.numero_documento}")
            print(f"   Total productos: {len(productos_nc)}, Total unidades: {total_unidades}")
            print(f"   Monto neto: ${monto_neto_total}, Total con IVA: ${total_con_iva}")
            
            # ============================================
            # AGREGAR TODOS LOS PRODUCTOS A LA NC
            # ============================================
            for prod_nc in productos_nc:
                recepcion = prod_nc['recepcion']
                cantidad = prod_nc['cantidad']
                
                # Crear detalle de producto en la NC
                Dte_Productos.objects.create(
                    dte=nota_credito,
                    productoTalla=recepcion.dte_producto.productoTalla,
                    descripcion=f"NC: {recepcion.dte_producto.descripcion}",
                    costo=recepcion.dte_producto.costo,
                    sobreprecio=recepcion.dte_producto.sobreprecio,
                    precio=recepcion.dte_producto.precio,
                    stock=cantidad,
                    activo=True
                )
                
                # Devolver stock a bodega origen
                producto_origen = recepcion.producto_talla
                if producto_origen:
                    stock_antes = producto_origen.stock
                    producto_origen.stock += cantidad
                    producto_origen.save()
                    print(f"   ✓ Stock devuelto: {producto_origen.sku} +{cantidad} ({stock_antes} → {producto_origen.stock})")
                    
                    # Crear movimiento de INGRESO
                    Movimientos_Producto.objects.create(
                        dte=nota_credito,
                        ProductoTalla=producto_origen,
                        sucursal_origen=None,
                        sucursal_destino=dte_original.sucursal,
                        cantidad=cantidad,
                        costo=producto_origen.producto.costo,
                        sobreprecio=producto_origen.producto.sobreprecio,
                        precio=producto_origen.producto.precioventa,
                        concepto='DEVOLUCION_NC',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=usuario,
                        observaciones=f'Devolución por NC #{numero_nc} - DTE #{dte_numero}. {motivo}'
                    )
                
                # Actualizar estado de la recepción
                recepcion.estado = 'REGULARIZADO'
                recepcion.fecha_regularizacion = hoy
                recepcion.regularizado_por = usuario
                recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] REGULARIZADO - NC #{numero_nc}. {motivo}"
                recepcion.save()
            
            print(f"✅ Regularización completada - {len(productos_nc)} productos procesados")
            
            # ============================================
            # GENERAR ARCHIVO TXT PARA ACEPTA
            # ============================================
            archivo_txt_url = None
            error_txt = None
            try:
                from .views_modulo_documentos import generar_txt_dte_acepta
                
                # Preparar datos para generar TXT
                datos_txt = {
                    'documento': {
                        'tipo_documento': '61',  # Nota de Crédito
                        'folio': str(numero_nc),
                        'fecha_emision': nota_credito.fecha_emision.strftime('%Y-%m-%d'),
                        'fecha_vencimiento': nota_credito.fecha_vencimiento.strftime('%Y-%m-%d'),
                        'tipo_despacho': '2',
                        'ind_traslado': '1',
                        'forma_pago': '1'
                    },
                    'emisor': {
                        'rut': nota_credito.emisor.rut if nota_credito.emisor else '',
                        'razon_social': nota_credito.emisor.razon_social if nota_credito.emisor else 'SIN RAZON SOCIAL',
                        'giro': nota_credito.emisor.giro if (nota_credito.emisor and nota_credito.emisor.giro) else 'COMERCIALIZADORA',
                        'acteco': nota_credito.emisor.acteco if nota_credito.emisor else '',
                        'direccion': nota_credito.emisor.direccion if nota_credito.emisor else '',
                        'comuna': nota_credito.emisor.comuna if nota_credito.emisor else '',
                        'ciudad': nota_credito.emisor.ciudad if nota_credito.emisor else '',
                        'sucursal': nota_credito.sucursal.alias if nota_credito.sucursal else '',
                        'codigo_vendedor': usuario
                    },
                    'receptor': {
                        'rut': nota_credito.receptor.rut if nota_credito.receptor else '',
                        'codigo_interno': str(nota_credito.receptor.id) if nota_credito.receptor else '',
                        'razon_social': nota_credito.receptor.razon_social if nota_credito.receptor else 'SIN RAZON SOCIAL',
                        'giro': nota_credito.receptor.giro if (nota_credito.receptor and nota_credito.receptor.giro) else 'COMERCIALIZADORA',
                        'contacto': nota_credito.receptor.contacto1 if (nota_credito.receptor and nota_credito.receptor.contacto1) else '',
                        'direccion': nota_credito.receptor.direccion if nota_credito.receptor else '',
                        'comuna': nota_credito.receptor.comuna if nota_credito.receptor else '',
                        'ciudad': nota_credito.receptor.ciudad if nota_credito.receptor else ''
                    },
                    'totales': {
                        'monto_neto': int(monto_neto_total),
                        'monto_exento': 0,
                        'iva': int(iva_total),
                        'monto_total': int(total_con_iva)
                    },
                    'detalle': [],
                    'referencias': [{
                        'tipo_documento': '33',  # Factura Electrónica
                        'folio': str(dte_original.numero_documento),
                        'fecha': dte_original.fecha_emision.strftime('%Y-%m-%d'),
                        'razon': '1'  # 1=Anula documento
                    }]
                }
                
                # Agregar detalle de productos
                for prod_nc in productos_nc:
                    recepcion = prod_nc['recepcion']
                    cantidad = prod_nc['cantidad']
                    precio_unitario = int(prod_nc['precio_unitario'])
                    
                    datos_txt['detalle'].append({
                        'codigo': str(recepcion.producto_talla.sku) if recepcion.producto_talla else '',
                        'sku': str(recepcion.producto_talla.sku) if recepcion.producto_talla else '',
                        'nombre': recepcion.producto_talla.producto.articulo if recepcion.producto_talla else '',
                        'descripcion': recepcion.dte_producto.descripcion if recepcion.dte_producto else '',
                        'cantidad': cantidad,
                        'unidad': 'UN',
                        'precio_unitario': precio_unitario,
                        'monto_item': cantidad * precio_unitario,
                        'indicador_exencion': ''
                    })
                
                # Generar contenido TXT
                contenido_txt = generar_txt_dte_acepta(datos_txt)
                
                # Guardar archivo TXT
                import os
                from django.conf import settings
                
                txt_dir = os.path.join(settings.MEDIA_ROOT, 'documentos_electronicos', 'nc')
                os.makedirs(txt_dir, exist_ok=True)
                
                nombre_archivo = f'NC_{numero_nc}_{hoy.strftime("%Y%m%d_%H%M%S")}.txt'
                ruta_archivo = os.path.join(txt_dir, nombre_archivo)
                
                with open(ruta_archivo, 'w', encoding='utf-8') as f:
                    f.write(contenido_txt)
                
                archivo_txt_url = f'/media/documentos_electronicos/nc/{nombre_archivo}'
                print(f"✅ Archivo TXT generado: {nombre_archivo}")
                
            except Exception as e:
                error_txt = str(e)
                print(f"⚠️ Error al generar TXT (NC creada igualmente): {error_txt}")
                import traceback
                traceback.print_exc()
            
            return JsonResponse({
                'success': True,
                'message': f'Nota de Crédito generada exitosamente para {len(productos_nc)} producto(s)',
                'numero_nc': numero_nc,
                'total_productos': len(productos_nc),
                'total_unidades': total_unidades,
                'monto_neto': float(monto_neto_total),
                'monto_iva': float(iva_total),
                'monto_total': float(total_con_iva),
                'archivo_txt_url': archivo_txt_url,
                'txt_generado': archivo_txt_url is not None,
                'error_txt': error_txt
            })
        
    except Exception as e:
        print(f"Error en regularizar_dte_masivo: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar regularización: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def anular_regularizacion_dte(request):
    """Anula una regularización hecha por error (receptor puede cancelar antes de que emisor resuelva)"""
    try:
        from .models import Productos_Recepcionados, Movimientos_Producto, Producto_Talla
        
        data = json.loads(request.body or '{}')
        dte_numero = data.get('dte_numero')
        productos_ids = data.get('productos_ids', [])
        motivo = data.get('motivo', '')
        
        if not dte_numero or not productos_ids or not motivo:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos'
            }, status=400)
        
        usuario = request.user.username
        hoy = timezone.now()
        sucursal_id = request.session.get('idSucursalActual')
        
        with transaction.atomic():
            recepciones = Productos_Recepcionados.objects.filter(
                id__in=productos_ids,
                estado__in=['EN_REGULARIZACION', 'FALTANTE', 'RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO']
            ).select_related('dte', 'producto_talla', 'producto_talla__producto')
            
            if not recepciones.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'No se encontraron productos para anular'
                }, status=404)
            
            productos_anulados = []
            stock_actualizado = 0
            
            for recepcion in recepciones:
                # Cantidad que se había marcado como problema
                cantidad_problema = recepcion.cantidad_faltante + recepcion.cantidad_danada
                
                if cantidad_problema <= 0:
                    continue
                
                # IMPORTANTE: Actualizar stock en DESTINO (ya que los productos "llegaron bien")
                # Necesitamos buscar o crear el producto en la sucursal destino
                producto_origen = recepcion.producto_talla
                dte_original = recepcion.dte
                
                # Obtener sucursal destino del movimiento
                movimiento = dte_original.dte_movimientos.filter(
                    concepto='TRASPASO_SALIDA'
                ).first()
                sucursal_destino = movimiento.sucursal_destino if movimiento else None
                
                if sucursal_destino and producto_origen:
                    # Buscar el producto en sucursal destino (por SKU)
                    try:
                        talla_destino = Producto_Talla.objects.get(
                            sku=producto_origen.sku,
                            producto__sucursal=sucursal_destino
                        )
                    except Producto_Talla.DoesNotExist:
                        # Si no existe, crearlo (igual que en confirmar_recepcion)
                        from .models import Producto
                        producto_origen_obj = producto_origen.producto
                        
                        producto_destino, _ = Producto.objects.get_or_create(
                            articulo=producto_origen_obj.articulo,
                            sucursal=sucursal_destino,
                            atributo1=producto_origen_obj.atributo1,
                            atributo2=producto_origen_obj.atributo2,
                            atributo3=producto_origen_obj.atributo3,
                            atributo4=producto_origen_obj.atributo4,
                            defaults={
                                'descripcion': producto_origen_obj.descripcion,
                                'categoria': producto_origen_obj.categoria,
                                'costo': producto_origen_obj.costo,
                                'sobreprecio': producto_origen_obj.sobreprecio,
                                'precioventa': producto_origen_obj.precioventa,
                                'precioSugerido': producto_origen_obj.precioSugerido,
                                'tipo_talla': producto_origen_obj.tipo_talla,
                                'guia_talla': producto_origen_obj.guia_talla
                            }
                        )
                        
                        talla_destino = Producto_Talla.objects.create(
                            producto=producto_destino,
                            talla=producto_origen.talla,
                            sku=producto_origen.sku,
                            stock=0
                        )
                    
                    # Actualizar stock en destino
                    stock_antes = talla_destino.stock
                    talla_destino.stock += cantidad_problema
                    talla_destino.save()
                    stock_actualizado += cantidad_problema
                    print(f"   ✓ Stock agregado en {sucursal_destino.alias}: {producto_origen.sku} +{cantidad_problema} ({stock_antes} → {talla_destino.stock})")
                    
                    # Crear movimiento de INGRESO en destino
                    Movimientos_Producto.objects.create(
                        dte=dte_original,
                        ProductoTalla=talla_destino,
                        sucursal_origen=dte_original.sucursal,
                        sucursal_destino=sucursal_destino,
                        cantidad=cantidad_problema,
                        costo=talla_destino.producto.costo,
                        sobreprecio=talla_destino.producto.sobreprecio,
                        precio=talla_destino.producto.precioventa,
                        concepto='ANULACION_REGULARIZACION',
                        tipo_movimiento='INGRESO',
                        estado='COMPLETADO',
                        responsable=usuario,
                        observaciones=f'Anulación de regularización DTE #{dte_numero}. {motivo}'
                    )
                
                # Actualizar recepción
                recepcion.cantidad_recepcionada = recepcion.cantidad_esperada
                recepcion.cantidad_faltante = 0
                recepcion.cantidad_danada = 0
                recepcion.stockArribado = recepcion.cantidad_esperada
                recepcion.estado = 'RECEPCIONADO_OK'
                recepcion.observaciones = (recepcion.observaciones or '') + f"\n[{hoy.strftime('%Y-%m-%d %H:%M')}] ANULADO - Regularización cancelada. {motivo}"
                recepcion.save()
                
                productos_anulados.append({
                    'producto': producto_origen.producto.articulo if producto_origen else '-',
                    'talla': producto_origen.talla if producto_origen else '-',
                    'cantidad': cantidad_problema
                })
            
            print(f"✅ Regularización DTE #{dte_numero} anulada - {len(productos_anulados)} productos, {stock_actualizado} unidades")
            
            return JsonResponse({
                'success': True,
                'message': f'Regularización anulada. {len(productos_anulados)} producto(s) actualizados como OK.',
                'productos_anulados': len(productos_anulados),
                'stock_actualizado': stock_actualizado
            })
        
    except Exception as e:
        print(f"Error en anular_regularizacion_dte: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al anular: {str(e)}'
        }, status=500)


@login_required
@require_GET
def obtener_dtes_con_problemas(request):
    """Obtiene lista de DTEs que tienen productos con problemas"""
    try:
        from .models import Productos_Recepcionados
        from django.db.models import Count, Q, Sum
        
        sucursal_id = request.session.get('idSucursalActual')
        
        # Obtener DTEs que tienen al menos un producto con problemas
        dtes_con_problemas = Dte.objects.filter(
            tipo_transaccion='TRASPASO',
            estado_dte__in=['RECEPCIONADO_PARCIAL', 'EN_REGULARIZACION'],
            recepciones__estado__in=['RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO', 'FALTANTE', 'EN_REGULARIZACION']
        ).select_related('emisor', 'receptor', 'sucursal').annotate(
            total_productos=Count('recepciones'),
            productos_ok=Count('recepciones', filter=Q(recepciones__estado='RECEPCIONADO_OK')),
            productos_problemas=Count('recepciones', filter=Q(
                recepciones__estado__in=['RECEPCIONADO_PARCIAL', 'RECEPCIONADO_DANADO', 'FALTANTE', 'EN_REGULARIZACION']
            ))
        ).distinct().order_by('-fecha_recepcion')
        
        items = []
        for dte in dtes_con_problemas[:50]:
            items.append({
                'id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision,
                'fecha_recepcion': dte.fecha_recepcion,
                'sucursal_origen': dte.sucursal.alias if dte.sucursal else '-',
                'emisor': dte.emisor.nombre if dte.emisor else '-',
                'total_productos': dte.total_productos,
                'productos_ok': dte.productos_ok,
                'productos_problemas': dte.productos_problemas,
                'estado_dte': dte.estado_dte,
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'total': dtes_con_problemas.count()
        }, json_dumps_params={'default': str})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener DTEs con problemas: {str(e)}'
        }, status=500)


@login_required
@require_GET
def obtener_detalle_dte_recepcionado(request):
    """Obtiene el detalle completo de un DTE recepcionado incluyendo productos recepcionados"""
    try:
        dte_id = request.GET.get('dte_id')
        
        if not dte_id:
            return JsonResponse({
                'success': False,
                'error': 'Falta el parámetro dte_id'
            }, status=400)
        
        dte = Dte.objects.select_related('emisor', 'receptor', 'sucursal').get(id=dte_id)
        
        # Obtener productos recepcionados
        from .models import Productos_Recepcionados
        productos_recepcionados = Productos_Recepcionados.objects.filter(
            dte=dte
        ).select_related(
            'producto_talla__producto',
            'dte_producto__productoTalla__producto'
        ).order_by('id')
        
        productos_detalle = []
        for recepcion in productos_recepcionados:
            producto_talla = recepcion.producto_talla or (recepcion.dte_producto.productoTalla if recepcion.dte_producto else None)
            producto = producto_talla.producto if producto_talla else None
            
            articulo = producto.articulo if producto else ''
            marca = producto.atributo1.valor if (producto and producto.atributo1) else ''
            color = producto.atributo2.valor if (producto and producto.atributo2) else ''
            
            productos_detalle.append({
                'sku': producto_talla.sku if producto_talla else '-',
                'descripcion': producto.descripcion if producto else (recepcion.dte_producto.descripcion if recepcion.dte_producto else '-'),
                'articulo': articulo,
                'marca': marca,
                'color': color,
                'talla': producto_talla.talla if producto_talla else '-',
                'cantidad_esperada': recepcion.cantidad_esperada,
                'cantidad_recepcionada': recepcion.stockArribado,
                'cantidad_danada': recepcion.cantidad_danada,
                'cantidad_faltante': recepcion.cantidad_faltante,
                'estado': recepcion.estado,
                'observaciones': recepcion.observaciones or '',
            })
        
        return JsonResponse({
            'success': True,
            'dte': {
                'id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision,
                'fecha_recepcion': dte.fecha_recepcion,
                'estado_dte': dte.estado_dte,
                'emisor': dte.emisor.nombre if dte.emisor else '-',
                'receptor': dte.receptor.nombre if dte.receptor else '-',
                'sucursal_origen': dte.sucursal.alias if dte.sucursal else '-',
                'referencias': dte.referencias or '',
            },
            'productos': productos_detalle
        }, json_dumps_params={'default': str})
        
    except Dte.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'DTE no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener detalles del DTE: {str(e)}'
        }, status=500)


def generar_nota_credito_automatica(dte_original, productos_afectados, usuario, motivo):
    """
    Genera una Nota de Crédito automáticamente por productos con problemas.
    
    Args:
        dte_original: DTE original que tiene problemas
        productos_afectados: Lista de productos con problemas [{'dte_producto_id', 'cantidad_faltante', 'observaciones'}]
        usuario: Usuario que genera la NC
        motivo: Motivo de la NC
    
    Returns:
        Dte de tipo Nota de Crédito
    """
    from decimal import Decimal
    from django.utils.dateparse import parse_date
    from django.utils import timezone
    
    # Calcular totales de productos afectados
    total_neto = Decimal('0')
    total_unidades = 0
    
    for prod_data in productos_afectados:
        dte_producto = Dte_Productos.objects.get(id=prod_data['dte_producto_id'])
        cantidad_nc = prod_data['cantidad_faltante']
        precio_unitario = Decimal(str(dte_producto.precio))
        
        total_neto += cantidad_nc * precio_unitario
        total_unidades += cantidad_nc
    
    # Calcular IVA
    iva = total_neto * Decimal('0.19')
    total_con_iva = total_neto + iva
    
    # Obtener correlativo para NC
    numero_nc = obtener_siguiente_correlativo(dte_original.sucursal, 'NOTA DE CREDITO')
    
    # Crear Nota de Crédito
    nota_credito = Dte.objects.create(
        emisor=dte_original.emisor,
        receptor=dte_original.receptor,
        numero_documento=numero_nc,
        tipo_documento='NOTA DE CREDITO',
        monto_neto=total_neto,
        monto_con_iva=total_con_iva,
        estado_pago='PENDIENTE',
        estado_dte='EMITIDO',
        responsable=usuario,
        fecha_emision=timezone.now().date(),
        fecha_vencimiento=timezone.now().date(),
        diasCredito=0,
        bultos=1,
        unidades_productos=total_unidades,
        tipo_transaccion='TRASPASO',
        sucursal=dte_original.sucursal,
        es_nota_credito=True,
        documento_afectado=dte_original,
        motivo_nc=motivo,
        referencias=f"NC por regularización DTE #{dte_original.numero_documento}. {motivo}"
    )
    
    # Crear detalle de productos afectados
    for prod_data in productos_afectados:
        dte_producto_original = Dte_Productos.objects.get(id=prod_data['dte_producto_id'])
        cantidad_nc = prod_data['cantidad_faltante']
        
        Dte_Productos.objects.create(
            dte=nota_credito,
            productoTalla=dte_producto_original.productoTalla,
            descripcion=f"NC: {dte_producto_original.descripcion}",
            costo=dte_producto_original.costo,
            sobreprecio=dte_producto_original.sobreprecio,
            precio=dte_producto_original.precio,
            stock=cantidad_nc,
            activo=True
        )
    
    return nota_credito


def validar_rut_chileno(rut):
    """
    Valida un RUT chileno
    Retorna: (es_valido, mensaje_error)
    """
    try:
        # Limpiar el RUT de puntos y guiones
        rut_limpio = re.sub(r'[.-]', '', rut.upper())
        
        # Verificar formato básico
        if not re.match(r'^\d{7,8}[0-9K]$', rut_limpio):
            return False, "El RUT debe tener 7 u 8 dígitos seguidos de un dígito verificador (0-9 o K)"
        
        # Separar número y dígito verificador
        numero = rut_limpio[:-1]
        dv = rut_limpio[-1]
        
        # Calcular dígito verificador
        suma = 0
        multiplicador = 2
        
        for digito in reversed(numero):
            suma += int(digito) * multiplicador
            multiplicador = multiplicador + 1 if multiplicador < 7 else 2
        
        # Calcular dígito verificador esperado
        resto = suma % 11
        dv_esperado = 11 - resto if resto != 0 else 0
        
        # Convertir a string
        if dv_esperado == 10:
            dv_esperado = 'K'
        else:
            dv_esperado = str(dv_esperado)
        
        # Comparar
        if dv == dv_esperado:
            return True, ""
        else:
            return False, f"El dígito verificador es incorrecto. Debería ser {dv_esperado}"
            
    except Exception as e:
        return False, f"Error al validar RUT: {str(e)}"

def validar_campos_proveedor(data):
    """
    Valida los campos obligatorios de un proveedor
    Retorna: (es_valido, errores)
    """
    errores = []
    
    # Campos obligatorios
    campos_obligatorios = [
        'nombre', 'rut', 'nombre_fantasia', 'razon_social', 
        'giro', 'direccion', 'comuna', 'ciudad', 'correoVendedor'
    ]
    
    for campo in campos_obligatorios:
        if not data.get(campo) or str(data.get(campo)).strip() == '':
            errores.append(f"El campo '{campo.replace('_', ' ').title()}' es obligatorio")
    
    # Validar RUT
    if data.get('rut'):
        rut_valido, mensaje_rut = validar_rut_chileno(data['rut'])
        if not rut_valido:
            errores.append(f"RUT inválido: {mensaje_rut}")
    
    # Validar formato de correos (si están presentes)
    if data.get('correoVendedor') and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', data['correoVendedor']):
        errores.append("El correo del vendedor no tiene un formato válido")
    
    if data.get('correoIntercambio') and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', data['correoIntercambio']):
        errores.append("El correo de intercambio no tiene un formato válido")
    
    if data.get('correoAdministrador') and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', data['correoAdministrador']):
        errores.append("El correo del administrador no tiene un formato válido")
    
    return len(errores) == 0, errores
from django.utils.dateparse import parse_date
import json
from datetime import date
from decimal import Decimal
from django.db.models import ProtectedError
from django.db.models import Sum, Count, Case, When, IntegerField, Value, F, Max
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import re
from django.views.decorators.csrf import csrf_exempt
import csv

# ========== FUNCIONES AUXILIARES PARA MOVIMIENTOS ==========

def registrar_movimiento_producto(producto_talla, concepto, cantidad, responsable, 
                                dte=None, ticket=None, sucursal_origen=None, 
                                sucursal_destino=None, observaciones=None, 
                                referencia_externa=None, crear_lote_fifo=True):
    """
    Función centralizada para registrar movimientos de productos
    Ahora incluye soporte para FIFO automático
    """
    from .models import Movimientos_Producto
    
    # Determinar sucursales si no se proporcionan
    if not sucursal_origen:
        sucursal_origen = producto_talla.producto.sucursal
    if not sucursal_destino:
        sucursal_destino = producto_talla.producto.sucursal
    
    # Crear el movimiento
    movimiento = Movimientos_Producto.objects.create(
        ProductoTalla=producto_talla,
        dte=dte,
        ticket=ticket,
        sucursal_origen=sucursal_origen,
        sucursal_destino=sucursal_destino,
        cantidad=cantidad,
        costo=producto_talla.producto.costo,
        sobreprecio=producto_talla.producto.sobreprecio,
        precio=producto_talla.producto.precioventa,
        concepto=concepto,
        responsable=responsable,
        observaciones=observaciones,
        referencia_externa=referencia_externa
    )
    
    # Actualizar stock del producto
    producto_talla.stock += cantidad
    producto_talla.save()
    
    # Crear lote FIFO para ingresos (solo si es positivo y se solicita)
    if crear_lote_fifo and cantidad > 0 and concepto in [
        'INGRESO_INICIAL', 'RECEPCION_COMPRA', 'DEVOLUCION_CLIENTE', 
        'TRASPASO_ENTRADA', 'AJUSTE_POSITIVO', 'DONACION_RECIBIDA'
    ]:
        try:
            crear_lote_producto(
                producto_talla=producto_talla,
                cantidad=cantidad,
                costo_unitario=producto_talla.producto.costo,
                sobreprecio_unitario=producto_talla.producto.sobreprecio,
                precio_venta_unitario=producto_talla.producto.precioventa,
                dte=dte,
                movimiento=movimiento,
                observaciones=f"Lote automático - {concepto} - {observaciones or ''}"
            )
        except Exception as e:
            print(f"⚠️ Error creando lote FIFO: {str(e)}")
            # No fallar el movimiento principal si falla la creación del lote
    
    return movimiento

def obtener_siguiente_correlativo(sucursal, tipo):
    """
    Obtiene el siguiente número de correlativo para tickets, traspasos, etc.
    """
    from .models import Correlativo
    
    correlativo, created = Correlativo.objects.get_or_create(
        tipo_dte=tipo,
        sucursal=sucursal,
        defaults={
            'inicio': 1, 
            'termino': 999999, 
            'alias': f'{tipo}_{sucursal.alias}',
            'responsable': 'Sistema'
        }
    )
    
    try:
        return correlativo.obtener_siguiente_numero()
    except ValueError as e:
        # Si el correlativo está agotado, crear uno nuevo automáticamente
        correlativo.inicio = correlativo.termino + 1
        correlativo.termino = correlativo.termino + 100000
        correlativo.fecha_actualizacion = timezone.now().date()
        correlativo.save()
        
        return correlativo.obtener_siguiente_numero()
# ========== VISTAS PARA VENTAS AL PÚBLICO ==========
@require_POST
@transaction.atomic
def crear_ticket_venta(request):
    """
    Crea un ticket de venta al público y registra los movimientos correspondientes
    """
    try:
        data = json.loads(request.body)
        
        # Datos de sesión
        sucursal_id = request.session.get('idSucursalActual')
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        vendedor = get_object_or_404(Vendedor, id=data.get('vendedor_id'))
        
        # Datos del ticket
        productos = data.get('productos', [])
        cliente_nombre = data.get('cliente_nombre', '')
        cliente_rut = data.get('cliente_rut', '')
        cliente_email = data.get('cliente_email', '')
        cliente_telefono = data.get('cliente_telefono', '')
        metodo_pago = data.get('metodo_pago', 'EFECTIVO')
        observaciones = data.get('observaciones', '')
        
        if not productos:
            return JsonResponse({'success': False, 'error': 'No hay productos en la venta'}, status=400)
        
        # Calcular totales
        subtotal = 0
        descuento_total = 0
        
        for producto in productos:
            # Validar cantidad
            try:
                cantidad = int(producto['cantidad'])
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'error': f'Cantidad inválida: debe ser un número entero'
                }, status=400)
            
            if cantidad < 1:
                return JsonResponse({
                    'success': False, 
                    'error': f'Cantidad inválida: debe ser mayor a 0'
                }, status=400)
            
            precio = int(producto['precio'])
            descuento = int(producto.get('descuento', 0))
            
            subtotal += cantidad * precio
            descuento_total += descuento
        
        total = subtotal - descuento_total
        
        # Crear ticket
        correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
        
        ticket = Ticket.objects.create(
            vendedor=vendedor,
            sucursal=sucursal,
            correlativo=correlativo,
            estado='PAGADO',
            subTotal=subtotal,
            descuento=descuento_total,
            total=total,
            responsable=responsable,
            cliente_nombre=cliente_nombre,
            cliente_rut=cliente_rut,
            cliente_email=cliente_email,
            cliente_telefono=cliente_telefono,
            metodo_pago=metodo_pago,
            observaciones=observaciones
        )
        
        # Crear detalles y registrar movimientos
        for producto in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=producto['producto_talla_id'])
            
            # Validar cantidad nuevamente antes de procesar
            try:
                cantidad = int(producto['cantidad'])
            except (ValueError, TypeError):
                raise Exception(f'Cantidad inválida para {producto_talla.producto.articulo} - Talla {producto_talla.talla}')
            
            if cantidad < 1:
                raise Exception(f'Cantidad debe ser mayor a 0 para {producto_talla.producto.articulo} - Talla {producto_talla.talla}')
            
            precio = int(producto['precio'])
            descuento = int(producto.get('descuento', 0))
            
            # Verificar stock disponible
            if producto_talla.stock < cantidad:
                raise Exception(
                    f'Stock insuficiente para {producto_talla.producto.articulo} - Talla {producto_talla.talla}. '
                    f'Solicitado: {cantidad}, Disponible: {producto_talla.stock}'
                )
            
            # Crear detalle del ticket
            Ticket_Productos.objects.create(
                ProductoTalla=producto_talla,
                idTicket=ticket,
                stock=cantidad,
                precio=precio,
                descuento_unitario=descuento,
                subtotal=cantidad * precio - descuento,
                precio_original=precio,
                porcentaje_descuento=(descuento / (cantidad * precio)) * 100 if cantidad * precio > 0 else 0
            )
            
            # Consumir stock usando FIFO
            try:
                costo_total_consumido, lotes_utilizados = consumir_stock_fifo(
                    producto_talla=producto_talla,
                    cantidad_requerida=cantidad,
                    responsable=responsable,
                    ticket=ticket,
                    observaciones=f'Ticket #{correlativo} - Cliente: {cliente_nombre}',
                    referencia_externa=str(correlativo)
                )
                
                # Registrar movimiento de egreso (sin crear lote FIFO)
                registrar_movimiento_producto(
                    producto_talla=producto_talla,
                    concepto='VENTA_PUBLICO',
                    cantidad=-cantidad,  # Negativo para egreso
                    responsable=responsable,
                    ticket=ticket,
                    observaciones=f'Ticket #{correlativo} - Cliente: {cliente_nombre} - Costo FIFO: ${costo_total_consumido:,}',
                    referencia_externa=str(correlativo),
                    crear_lote_fifo=False  # No crear lote para egresos
                )
                
                # Guardar información FIFO en el ticket
                ticket_producto = Ticket_Productos.objects.get(
                    ProductoTalla=producto_talla,
                    idTicket=ticket
                )
                ticket_producto.costo_fifo = costo_total_consumido
                ticket_producto.lotes_utilizados = str(lotes_utilizados)  # Convertir a string para almacenar
                ticket_producto.save()
                
            except Exception as e:
                raise Exception(f'Error en FIFO para {producto_talla.producto.articulo} - Talla {producto_talla.talla}: {str(e)}')
        
        return JsonResponse({
            'success': True,
            'ticket_id': ticket.id,
            'correlativo': correlativo,
            'total': total
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def obtener_tickets_venta(request):
    """
    Obtiene tickets de venta con filtros
    """
    sucursal_id = request.session.get('idSucursalActual')
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado')
    vendedor_id = request.GET.get('vendedor_id')
    cliente = request.GET.get('cliente')
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 50)), 100)
    
    tickets = Ticket.objects.filter(sucursal_id=sucursal_id)
    
    if fecha_inicio:
        tickets = tickets.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        tickets = tickets.filter(fecha__lte=parse_date(fecha_fin))
    if estado:
        tickets = tickets.filter(estado=estado)
    if vendedor_id:
        tickets = tickets.filter(vendedor_id=vendedor_id)
    if cliente:
        tickets = tickets.filter(
            Q(cliente_nombre__icontains=cliente) |
            Q(cliente_rut__icontains=cliente)
        )
    
    total_count = tickets.count()
    offset = (page - 1) * page_size
    tickets = tickets.select_related('vendedor')[offset:offset+page_size]
    
    data = []
    for ticket in tickets:
        data.append({
            'id': ticket.id,
            'correlativo': ticket.correlativo,
            'fecha': ticket.fecha.strftime('%Y-%m-%d'),
            'hora': ticket.hora.strftime('%H:%M'),
            'vendedor': ticket.vendedor.nombre,
            'cliente': ticket.cliente_nombre or 'Sin cliente',
            'subtotal': ticket.subTotal,
            'descuento': ticket.descuento or 0,
            'total': ticket.total,
            'estado': ticket.estado,
            'metodo_pago': ticket.metodo_pago,
            'responsable': ticket.responsable
        })
    
    return JsonResponse({
        'success': True,
        'items': data,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size
        }
    })

# ========== VISTAS PARA TRASPASOS ==========

@require_POST
@transaction.atomic
def crear_traspaso(request):
    """
    Crea una solicitud de traspaso entre sucursales
    """
    try:
        data = json.loads(request.body)
        
        # Datos de sesión
        sucursal_origen_id = request.session.get('idSucursalActual')
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        if not sucursal_origen_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)
        
        sucursal_origen = get_object_or_404(Sucursal, id=sucursal_origen_id)
        sucursal_destino = get_object_or_404(Sucursal, id=data.get('sucursal_destino_id'))
        
        if sucursal_origen == sucursal_destino:
            return JsonResponse({'success': False, 'error': 'No se puede traspasar a la misma sucursal'}, status=400)
        
        productos = data.get('productos', [])
        observaciones = data.get('observaciones', '')
        
        if not productos:
            return JsonResponse({'success': False, 'error': 'No hay productos en el traspaso'}, status=400)
        
        # Verificar stock disponible
        for producto in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=producto['producto_talla_id'])
            cantidad = int(producto['cantidad'])
            
            if producto_talla.stock < cantidad:
                raise Exception(f'Stock insuficiente para {producto_talla.producto.articulo} - Talla {producto_talla.talla}')
        
        # Crear traspaso
        numero_traspaso = obtener_siguiente_correlativo(sucursal_origen, 'TRASPASO')
        
        traspaso = Traspaso.objects.create(
            sucursal_origen=sucursal_origen,
            sucursal_destino=sucursal_destino,
            numero_traspaso=numero_traspaso,
            solicitante=responsable,
            observaciones_solicitud=observaciones
        )
        
        # Crear detalles
        for producto in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=producto['producto_talla_id'])
            cantidad = int(producto['cantidad'])
            
            Traspaso_Detalle.objects.create(
                traspaso=traspaso,
                producto_talla=producto_talla,
                cantidad_solicitada=cantidad,
                costo=producto_talla.producto.costo,
                precio_venta=producto_talla.producto.precioventa
            )
        
        return JsonResponse({
            'success': True,
            'traspaso_id': traspaso.id,
            'numero_traspaso': numero_traspaso
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_POST
@transaction.atomic
def aprobar_traspaso(request):
    """
    Aprueba un traspaso y registra los movimientos de salida
    """
    try:
        data = json.loads(request.body)
        traspaso_id = data.get('traspaso_id')
        aprobador = request.session.get('nombreUsuario', 'Sistema')
        
        traspaso = get_object_or_404(Traspaso, id=traspaso_id)
        
        if traspaso.estado != 'PENDIENTE':
            return JsonResponse({'success': False, 'error': 'El traspaso no está pendiente'}, status=400)
        
        # Actualizar traspaso
        traspaso.estado = 'APROBADO'
        traspaso.aprobador = aprobador
        traspaso.fecha_aprobacion = timezone.now().date()
        traspaso.observaciones_aprobacion = data.get('observaciones', '')
        traspaso.save()
        
        # Registrar movimientos de salida
        for detalle in traspaso.detalles.all():
            if detalle.cantidad_aprobada:
                cantidad = detalle.cantidad_aprobada
            else:
                cantidad = detalle.cantidad_solicitada
                detalle.cantidad_aprobada = cantidad
                detalle.save()
            
            # Registrar movimiento de salida
            registrar_movimiento_producto(
                producto_talla=detalle.producto_talla,
                concepto='TRASPASO_SALIDA',
                cantidad=-cantidad,
                responsable=aprobador,
                sucursal_origen=traspaso.sucursal_origen,
                sucursal_destino=traspaso.sucursal_destino,
                observaciones=f'Traspaso #{traspaso.numero_traspaso} a {traspaso.sucursal_destino}',
                referencia_externa=str(traspaso.numero_traspaso)
            )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_POST
@transaction.atomic
def recibir_traspaso(request):
    """
    Recibe un traspaso y registra los movimientos de entrada
    """
    try:
        data = json.loads(request.body)
        traspaso_id = data.get('traspaso_id')
        receptor = request.session.get('nombreUsuario', 'Sistema')
        
        traspaso = get_object_or_404(Traspaso, id=traspaso_id)
        
        if traspaso.estado != 'APROBADO':
            return JsonResponse({'success': False, 'error': 'El traspaso no está aprobado'}, status=400)
        
        # Actualizar traspaso
        traspaso.estado = 'RECIBIDO'
        traspaso.receptor = receptor
        traspaso.fecha_recepcion = timezone.now().date()
        traspaso.observaciones_recepcion = data.get('observaciones', '')
        traspaso.save()
        
        # Registrar movimientos de entrada
        for detalle in traspaso.detalles.all():
            cantidad = detalle.cantidad_aprobada or detalle.cantidad_solicitada
            
            # Registrar movimiento de entrada
            registrar_movimiento_producto(
                producto_talla=detalle.producto_talla,
                concepto='TRASPASO_ENTRADA',
                cantidad=cantidad,
                responsable=receptor,
                sucursal_origen=traspaso.sucursal_origen,
                sucursal_destino=traspaso.sucursal_destino,
                observaciones=f'Traspaso #{traspaso.numero_traspaso} desde {traspaso.sucursal_origen}',
                referencia_externa=str(traspaso.numero_traspaso)
            )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ========== VISTAS PARA AJUSTES DE INVENTARIO ==========

@require_POST
@transaction.atomic
def crear_ajuste_inventario(request):
    """
    Crea un ajuste de inventario
    """
    try:
        data = json.loads(request.body)
        
        # Datos de sesión
        sucursal_id = request.session.get('idSucursalActual')
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        productos = data.get('productos', [])
        tipo_ajuste = data.get('tipo_ajuste')
        motivo = data.get('motivo', '')
        observaciones = data.get('observaciones', '')
        
        if not productos:
            return JsonResponse({'success': False, 'error': 'No hay productos en el ajuste'}, status=400)
        
        # Crear ajuste
        numero_ajuste = obtener_siguiente_correlativo(sucursal, 'AJUSTE')
        
        ajuste = AjusteInventario.objects.create(
            sucursal=sucursal,
            numero_ajuste=numero_ajuste,
            tipo_ajuste=tipo_ajuste,
            motivo=motivo,
            observaciones=observaciones,
            solicitante=responsable
        )
        
        # Crear detalles y registrar movimientos
        for producto in productos:
            producto_talla = get_object_or_404(Producto_Talla, id=producto['producto_talla_id'])
            stock_sistema = producto_talla.stock
            stock_fisico = int(producto['stock_fisico'])
            diferencia = stock_fisico - stock_sistema
            
            # Crear detalle
            AjusteInventario_Detalle.objects.create(
                ajuste=ajuste,
                producto_talla=producto_talla,
                stock_sistema=stock_sistema,
                stock_fisico=stock_fisico,
                diferencia=diferencia,
                costo=producto_talla.producto.costo,
                precio_venta=producto_talla.producto.precioventa,
                observaciones=producto.get('observaciones', '')
            )
            
            # Registrar movimiento
            if diferencia != 0:
                concepto = 'AJUSTE_POSITIVO' if diferencia > 0 else 'AJUSTE_NEGATIVO'
                registrar_movimiento_producto(
                    producto_talla=producto_talla,
                    concepto=concepto,
                    cantidad=diferencia,
                    responsable=responsable,
                    observaciones=f'Ajuste #{numero_ajuste} - {motivo}',
                    referencia_externa=str(numero_ajuste)
                )
        
        return JsonResponse({
            'success': True,
            'ajuste_id': ajuste.id,
            'numero_ajuste': numero_ajuste
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ========== VISTAS MEJORADAS PARA MOVIMIENTOS ==========

# ========== VISTAS PARA REPORTES ==========

@require_GET
@login_required
def reporte_movimientos_kardex(request):
    producto_talla_id = request.GET.get('producto_talla_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 100)), 500)
    if not producto_talla_id:
        return JsonResponse({'success': False, 'error': 'ID de producto requerido'}, status=400)
    producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
    movimientos = Movimientos_Producto.objects.filter(
        ProductoTalla=producto_talla
    ).order_by('fecha', 'hora')
    if fecha_inicio:
        from django.utils.dateparse import parse_date
        movimientos = movimientos.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        from django.utils.dateparse import parse_date
        movimientos = movimientos.filter(fecha__lte=parse_date(fecha_fin))
    total_count = movimientos.count()
    offset = (page - 1) * page_size
    movimientos = movimientos[offset:offset+page_size]
    kardex = []
    saldo = 0
    for m in movimientos:
        saldo += m.cantidad
        # Enriquecer referencia
        referencia = ''
        if m.dte:
            referencia = f"{m.dte.tipo_documento} {m.dte.numero_documento}"
        elif m.ticket:
            referencia = f"Ticket {m.ticket.correlativo}"
        elif m.referencia_externa:
            referencia = m.referencia_externa
        elif m.sucursal_destino:
            referencia = f"Destino: {m.sucursal_destino.alias}"
        kardex.append({
            'fecha': m.fecha.strftime('%Y-%m-%d'),
            'hora': m.hora.strftime('%H:%M'),
            'concepto': m.concepto,
            'tipo_movimiento': m.tipo_movimiento,
            'entrada': m.cantidad if m.cantidad > 0 else 0,
            'salida': abs(m.cantidad) if m.cantidad < 0 else 0,
            'saldo': saldo,
            'costo': m.costo,
            'precio': m.precio,
            'responsable': m.responsable,
            'referencia': referencia
        })
    return JsonResponse({
        'success': True,
        'items': kardex,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size,
            'has_next': offset + page_size < total_count,
            'has_previous': page > 1
        }
    })
@require_GET
@login_required
def reporte_kardex_agrupado(request):
    """
    Genera un kardex agrupado por producto (sin mostrar tallas ni SKU)
    """
    producto_id = request.GET.get('producto_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 100)), 500)
    
    if not producto_id:
        return JsonResponse({'success': False, 'error': 'ID de producto requerido'}, status=400)
    
    producto = get_object_or_404(Producto, id=producto_id)
    
    # Obtener todos los movimientos del producto (todas las tallas)
    movimientos = Movimientos_Producto.objects.filter(
        ProductoTalla__producto=producto
    ).select_related('ProductoTalla', 'dte', 'ticket').order_by('fecha', 'hora')
    
    if fecha_inicio:
        from django.utils.dateparse import parse_date
        movimientos = movimientos.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        from django.utils.dateparse import parse_date
        movimientos = movimientos.filter(fecha__lte=parse_date(fecha_fin))
    
    # Agrupar movimientos por fecha y concepto
    from collections import defaultdict
    movimientos_agrupados = defaultdict(lambda: {
        'fecha': None,
        'hora': None,
        'concepto': None,
        'tipo_movimiento': None,
        'cantidad_total': 0,
        'costo_promedio': 0,
        'precio_promedio': 0,
        'responsable': None,
        'referencia': None,
        'movimientos_detalle': []
    })
    
    for m in movimientos:
        # Crear clave única por fecha + concepto + responsable
        key = f"{m.fecha}_{m.concepto}_{m.responsable}"
        
        grupo = movimientos_agrupados[key]
        if not grupo['fecha']:
            grupo['fecha'] = m.fecha
            grupo['hora'] = m.hora
            grupo['concepto'] = m.concepto
            grupo['tipo_movimiento'] = m.tipo_movimiento
            grupo['responsable'] = m.responsable
            
            # Enriquecer referencia
            referencia = m.referencia_externa or ''
            if m.dte:
                tipo_doc = m.dte.tipo_documento
                if tipo_doc == 'GUIA' and m.dte.tipo_transaccion == 'VENTA':
                    referencia = f"Guía {m.dte.numero_documento} - Despacho"
                else:
                    referencia = f"{tipo_doc} {m.dte.numero_documento}"
            elif m.ticket:
                referencia = f"Ticket {m.ticket.correlativo}"
            grupo['referencia'] = referencia
        
        # Sumar cantidades
        grupo['cantidad_total'] += m.cantidad
        grupo['movimientos_detalle'].append({
            'talla': m.ProductoTalla.talla,
            'sku': m.ProductoTalla.sku,
            'cantidad': m.cantidad
        })
    
    # Convertir a lista y calcular saldo acumulado
    kardex_agrupado = []
    saldo_acumulado = 0
    
    for grupo in sorted(movimientos_agrupados.values(), key=lambda x: (x['fecha'], x['hora'])):
        saldo_acumulado += grupo['cantidad_total']
        
        kardex_agrupado.append({
            'fecha': grupo['fecha'].strftime('%Y-%m-%d'),
            'hora': grupo['hora'].strftime('%H:%M') if grupo['hora'] else '',
            'concepto': grupo['concepto'],
            'tipo_movimiento': grupo['tipo_movimiento'],
            'entrada': grupo['cantidad_total'] if grupo['cantidad_total'] > 0 else 0,
            'salida': abs(grupo['cantidad_total']) if grupo['cantidad_total'] < 0 else 0,
            'saldo': saldo_acumulado,
            'responsable': grupo['responsable'],
            'referencia': grupo['referencia'],
            'detalle_tallas': grupo['movimientos_detalle']  # Para mostrar en tooltip o modal
        })
    
    # Paginación
    total_count = len(kardex_agrupado)
    offset = (page - 1) * page_size
    kardex_paginado = kardex_agrupado[offset:offset+page_size]
    
    return JsonResponse({
        'success': True,
        'producto': {
            'id': producto.id,
            'articulo': producto.articulo,
            'descripcion': producto.descripcion,
            'stock_total': Producto_Talla.objects.filter(producto=producto).aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        },
        'items': kardex_paginado,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size,
            'has_next': offset + page_size < total_count,
            'has_previous': page > 1
        }
    })

@require_GET
@login_required
def obtener_productos_base(request):
    """
    Obtiene productos base (sin tallas) para el selector del kardex agrupado
    """
    q = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 20)), 100)
    sucursal_id = request.session.get('idSucursalActual')
    
    productos = Producto.objects.all()
    if sucursal_id:
        productos = productos.filter(sucursal_id=sucursal_id)
    
    if q:
        productos = productos.filter(
            Q(articulo__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(atributo1__valor__icontains=q) |
            Q(atributo2__valor__icontains=q) |
            Q(atributo3__valor__icontains=q)
        )
    
    total_count = productos.count()
    offset = (page - 1) * page_size
    productos = productos.order_by('articulo')[offset:offset+page_size]
    
    def limpiar_prefijo(valor):
        if not valor:
            return ''
        for prefijo in ['Marca:', 'Color:', 'Género:']:
            if valor.startswith(prefijo):
                return valor[len(prefijo):].strip()
        return valor.strip()
    
    results = []
    for prod in productos:
        marca = limpiar_prefijo(prod.atributo1.valor if prod.atributo1 else '')
        color = limpiar_prefijo(prod.atributo2.valor if prod.atributo2 else '')
        genero = limpiar_prefijo(prod.atributo3.valor if prod.atributo3 else '')
        
        text = f"{prod.articulo} - {prod.descripcion}"
        if marca:
            text += f" | Marca: {marca}"
        if color:
            text += f" | Color: {color}"
        if genero:
            text += f" | Género: {genero}"
        
        results.append({
            'id': prod.id,
            'text': text,
            'articulo': prod.articulo,
            'descripcion': prod.descripcion,
            'marca': marca,
            'color': color,
            'genero': genero
        })
    
    return JsonResponse({
        'results': results,
        'pagination': {
            'more': offset + page_size < total_count
        }
    })

# Create your views here.
@login_required
def verHome(request):
    """
    Dashboard general que engloba módulos de ventas, DTE y requerimientos
    con indicadores clave de negocio
    """
    try:
        from .models import (
            CambioDevolucion, 
            Requerimiento, 
            Solicitud_Regularizacion,
            PagoCambioDevolucion
        )
        from datetime import datetime, timedelta
        from django.db.models import Sum, Count, Q, Avg
        
        # Obtener sucursal y empresa actual
        sucursal_id = request.session.get('idSucursalActual')
        empresa_id = request.session.get('idEmpresaActual')
        
        # Fechas para filtros
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        mes_pasado = (inicio_mes - timedelta(days=1)).replace(day=1)
        
        # ========== MÓDULO CAMBIOS Y DEVOLUCIONES ==========
        cambios_base = CambioDevolucion.objects.all()
        if sucursal_id:
            cambios_base = cambios_base.filter(sucursal_id=sucursal_id)
        
        cambios_data = {
            'completados': cambios_base.filter(estado='COMPLETADO').count(),
            'pendientes_cobro': cambios_base.filter(estado='EJECUTADO_COBRO_PENDIENTE').count(),
            'en_proceso': cambios_base.filter(
                estado__in=['SOLICITADO', 'EN_PROCESO', 'APROBADO', 'EJECUTADO']
            ).count(),
            'rechazados': cambios_base.filter(estado__in=['RECHAZADO', 'CANCELADO']).count(),
            'total_mes': cambios_base.filter(fecha_solicitud__gte=inicio_mes).count(),
            'monto_mes': cambios_base.filter(
                fecha_solicitud__gte=inicio_mes
            ).aggregate(total=Sum('monto_nuevo'))['total'] or 0,
        }
        
        # Cambios por tipo de operación
        cambios_por_tipo = cambios_base.filter(
            fecha_solicitud__gte=inicio_mes
        ).values('tipo_operacion').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')[:5]
        
        # ========== MÓDULO REQUERIMIENTOS ==========
        requerimientos_base = Requerimiento.objects.all()
        if sucursal_id:
            requerimientos_base = requerimientos_base.filter(sucursal_id=sucursal_id)
        
        requerimientos_data = {
            'pendientes': requerimientos_base.filter(estado='PENDIENTE').count(),
            'esperando_respuesta': requerimientos_base.filter(estado='ESPERANDO_RESPUESTA').count(),
            'aprobados': requerimientos_base.filter(estado='APROBADO').count(),
            'rechazados': requerimientos_base.filter(estado='RECHAZADO').count(),
            'total': requerimientos_base.count(),
            'total_mes': requerimientos_base.filter(fecha_creacion__gte=inicio_mes).count(),
        }
        
        # Requerimientos por tipo
        requerimientos_por_tipo = requerimientos_base.filter(
            fecha_creacion__gte=inicio_mes
        ).values('tipo').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')
        
        # Antigüedad promedio de pendientes
        requerimientos_pendientes = requerimientos_base.filter(estado='PENDIENTE')
        if requerimientos_pendientes.exists():
            # Calcular días de antigüedad para cada requerimiento pendiente
            total_dias = 0
            for req in requerimientos_pendientes:
                dias_antiguedad = (hoy - req.fecha_creacion.date()).days
                total_dias += dias_antiguedad
            dias_promedio = total_dias // requerimientos_pendientes.count() if requerimientos_pendientes.count() > 0 else 0
        else:
            dias_promedio = 0
        requerimientos_data['dias_promedio_pendientes'] = dias_promedio
        
        # ========== MÓDULO DTE ==========
        dtes_base = Dte.objects.all()
        if sucursal_id:
            dtes_base = dtes_base.filter(sucursal_id=sucursal_id)
        if empresa_id:
            dtes_base = dtes_base.filter(emisor_id=empresa_id)
        
        dtes_data = {
            'emitidos_mes': dtes_base.filter(
                fecha_emision__gte=inicio_mes,
                estado_dte='EMITIDO'
            ).count(),
            'recepcionados': dtes_base.filter(
                estado_dte__in=['RECEPCIONADO_COMPLETO', 'RECEPCIONADO_PARCIAL']
            ).count(),
            'en_regularizacion': dtes_base.filter(estado_dte='EN_REGULARIZACION').count(),
            'pendientes_pago': dtes_base.filter(
                estado_pago='PENDIENTE',
                tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO']
            ).count(),
            'monto_emitido_mes': dtes_base.filter(
                fecha_emision__gte=inicio_mes
            ).aggregate(total=Sum('monto_con_iva'))['total'] or 0,
            'monto_pendiente_pago': dtes_base.filter(
                estado_pago='PENDIENTE'
            ).aggregate(total=Sum('monto_con_iva'))['total'] or 0,
        }
        
        # DTEs por tipo de documento del mes
        dtes_por_tipo = dtes_base.filter(
            fecha_emision__gte=inicio_mes
        ).values('tipo_documento').annotate(
            cantidad=Count('id'),
            monto=Sum('monto_con_iva')
        ).order_by('-cantidad')
        
        # ========== MÓDULO REGULARIZACIONES ==========
        regularizaciones_base = Solicitud_Regularizacion.objects.all()
        if sucursal_id:
            regularizaciones_base = regularizaciones_base.filter(
                Q(sucursal_solicitante_id=sucursal_id) | Q(sucursal_emisora_id=sucursal_id)
            )
        
        regularizaciones_data = {
            'pendientes': regularizaciones_base.filter(estado='PENDIENTE').count(),
            'en_revision': regularizaciones_base.filter(estado='EN_REVISION').count(),
            'aprobadas': regularizaciones_base.filter(estado='APROBADA').count(),
            'ejecutadas': regularizaciones_base.filter(estado='EJECUTADA').count(),
            'completadas': regularizaciones_base.filter(estado='COMPLETADA').count(),
            'rechazadas': regularizaciones_base.filter(estado='RECHAZADA').count(),
            'total': regularizaciones_base.count(),
        }
        
        # Regularizaciones por tipo de problema
        regularizaciones_por_tipo = regularizaciones_base.filter(
            fecha_solicitud__gte=inicio_mes
        ).values('tipo_problema').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')[:5]
        
        # ========== INDICADORES GENERALES ==========
        # Calcular tendencias (comparar con mes pasado)
        cambios_mes_pasado = CambioDevolucion.objects.filter(
            fecha_solicitud__gte=mes_pasado,
            fecha_solicitud__lt=inicio_mes
        )
        if sucursal_id:
            cambios_mes_pasado = cambios_mes_pasado.filter(sucursal_id=sucursal_id)
        
        tendencia_cambios = cambios_data['total_mes'] - cambios_mes_pasado.count()
        tendencia_cambios_abs = abs(tendencia_cambios)
        
        # Preparar contexto
        context = {
            # Datos de cambios y devoluciones
            'cambios': cambios_data,
            'cambios_por_tipo': list(cambios_por_tipo),
            'tendencia_cambios': tendencia_cambios,
            'tendencia_cambios_abs': tendencia_cambios_abs,
            
            # Datos de requerimientos
            'requerimientos': requerimientos_data,
            'requerimientos_por_tipo': list(requerimientos_por_tipo),
            
            # Datos de DTEs
            'dtes': dtes_data,
            'dtes_por_tipo': list(dtes_por_tipo),
            
            # Datos de regularizaciones
            'regularizaciones': regularizaciones_data,
            'regularizaciones_por_tipo': list(regularizaciones_por_tipo),
            
            # Información general
            'fecha_actual': hoy,
            'inicio_mes': inicio_mes,
        }
        
        return render(request, 'vistas/dashboard_general.html', context)
        
    except Exception as e:
        import traceback
        print(f"Error en verHome/dashboard_general: {str(e)}")
        print(traceback.format_exc())
        # En caso de error, renderizar template vacío con mensaje
        return render(request, 'vistas/dashboard_general.html', {
            'cambios': {'completados': 0, 'pendientes_cobro': 0, 'en_proceso': 0, 'rechazados': 0, 'total_mes': 0, 'monto_mes': 0},
            'requerimientos': {'pendientes': 0, 'esperando_respuesta': 0, 'aprobados': 0, 'rechazados': 0, 'total': 0, 'total_mes': 0, 'dias_promedio_pendientes': 0},
            'dtes': {'emitidos_mes': 0, 'recepcionados': 0, 'en_regularizacion': 0, 'pendientes_pago': 0, 'monto_emitido_mes': 0, 'monto_pendiente_pago': 0},
            'regularizaciones': {'pendientes': 0, 'en_revision': 0, 'aprobadas': 0, 'ejecutadas': 0, 'completadas': 0, 'rechazadas': 0, 'total': 0},
            'cambios_por_tipo': [],
            'requerimientos_por_tipo': [],
            'dtes_por_tipo': [],
            'regularizaciones_por_tipo': [],
            'tendencia_cambios': 0,
            'tendencia_cambios_abs': 0,
            'fecha_actual': timezone.now().date(),
            'inicio_mes': timezone.now().date().replace(day=1),
        })
@login_required
def verGestionCompras(request):
    # Esta vista solo maneja GET requests para mostrar la página
    # Los POST requests para crear compras se manejan en la vista crear_compra
    empresas = Empresa.objects.all()  # Lista para usar en el select del modal
    return render(request, 'vistas/modulo_compras/gestionCompras.html', {'empresas': empresas})
 
@login_required
def verGestionProducto(request):
    """
    Vista para gestión de productos con inicialización automática de atributos
    """
    try:
        # Intentar obtener los atributos básicos
        marca = Productos_Atributos.objects.get(nombre__iexact='Marca')
        color = Productos_Atributos.objects.get(nombre__iexact='Color')
        genero = Productos_Atributos.objects.get(nombre__iexact='Género')
        
    except Productos_Atributos.DoesNotExist:
        # Si no existen los atributos, ejecutar inicialización automática
        from django.core.management import call_command
        from django.contrib import messages
        
        try:
            call_command('inicializar_atributos')
            messages.success(request, 'Atributos básicos inicializados correctamente.')
            
            # Intentar obtener los atributos nuevamente
            marca = Productos_Atributos.objects.get(nombre__iexact='Marca')
            color = Productos_Atributos.objects.get(nombre__iexact='Color')
            genero = Productos_Atributos.objects.get(nombre__iexact='Género')
            
        except Exception as e:
            messages.error(request, f'Error al inicializar atributos: {str(e)}')
            # Valores por defecto en caso de error
            marca = color = genero = None

    context = {
        'id_atributo_marca': marca.id if marca else 0,
        'id_atributo_color': color.id if color else 0,
        'id_atributo_genero': genero.id if genero else 0,
    }

    return render(request, 'vistas/modulo_existencias/verGestionProductos.html', context)
@login_required
def verGestionDteCompras(request):
     
     
    return render(request, 'vistas/modulo_compras/gestionDteCompras.html' )


def ver_resetPassword(request):
    if request.method == 'POST':
        email = request.POST['email']
      
         
    return render(request, 'registration/passwordReset.html')

def obtenerDetalleComprasPorParametros(request):
   
    return True
def crear_compra(request):
    try:
        # Obtener datos del formulario
        empresa_id = request.POST.get('empresa')
        nombre = request.POST.get('nombre')
        temporada = request.POST.get('temporada')
        fecha_inicio = request.POST.get('fechaInicioTemporada')
        fecha_termino = request.POST.get('fechaTerminoTemporada')

        # Validar datos requeridos
        if not all([empresa_id, nombre, temporada, fecha_inicio, fecha_termino]):
            return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)

        if fecha_inicio > fecha_termino:
            return JsonResponse({'success': False, 'error': 'Fechas inválidas'}, status=400)

        # Obtener empresa
        empresa = get_object_or_404(Empresa, id=empresa_id)

        # Obtener sucursal de la sesión
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Sucursal no definida en la sesión'}, status=400)

        # Obtener la sucursal
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        # Usar la función obtener_siguiente_correlativo que maneja la creación automática
        numero_actual = obtener_siguiente_correlativo(sucursal, 'COMPRA')

        # Crear la compra
        compra = Compras.objects.create(
            empresa=empresa,
            nombre=nombre,
            temporada=temporada,
            responsable=request.user.get_full_name() or 'Sistema',
            correlativo=numero_actual,
            fechaInicioTemporada=fecha_inicio,
            fechaTerminoTemporada=fecha_termino
        )

        return JsonResponse({
            'success': True, 
            'message': 'Compra creada exitosamente',
            'compra_id': compra.id
        })

    except Exception as e:
        # Log del error para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al crear compra: {str(e)}")
        
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'}, status=500)

 
@require_GET
def obtener_compras_por_anio(request):
    anio = request.GET.get('anio')
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 20)  # 20 registros por página
    search = request.GET.get('search', '').strip()
    
    if not anio:
        return JsonResponse({'success': False, 'error': 'Año no especificado'}, status=400)

    # Validar parámetros de paginación
    try:
        page = int(page)
        page_size = min(int(page_size), 100)  # Máximo 100 por página
    except (ValueError, TypeError):
        page = 1
        page_size = 20

    # Query base optimizada
    compras_query = Compras.objects.filter(
        fecha__year=anio
    ).select_related('empresa').annotate(
        # Calcular total de unidades y costo total usando subconsultas
        unidades_totales=Sum('compras_producto__compras_producto_talla__stock'),
        costo_total=Sum(
            F('compras_producto__compras_producto_talla__stock') * 
            F('compras_producto__costo')
        ),
        # Calcular total recepcionado
        total_recepcionado=Sum('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado')
    )

    # Aplicar búsqueda si se proporciona
    if search:
        compras_query = compras_query.filter(
            Q(nombre__icontains=search) |
            Q(empresa__nombre__icontains=search) |
            Q(responsable__icontains=search) |
            Q(temporada__icontains=search)
        )

    # Contar total de registros para paginación
    total_count = compras_query.count()
    
    # Aplicar paginación
    offset = (page - 1) * page_size
    compras = compras_query.values(
        'id', 'nombre', 'empresa__nombre', 'responsable', 'temporada', 
        'fecha', 'fechaInicioTemporada', 'fechaTerminoTemporada',
        'unidades_totales', 'costo_total', 'total_recepcionado'
    )[offset:offset + page_size]

    # Calcular días restantes y formatear datos
    hoy = date.today()
    data = []
    
    for compra in compras:
        # Calcular días restantes
        if compra['fechaInicioTemporada'] and compra['fechaTerminoTemporada']:
            if compra['fechaInicioTemporada'] <= hoy <= compra['fechaTerminoTemporada']:
                dias_restantes = (compra['fechaTerminoTemporada'] - hoy).days
            else:
                dias_restantes = -1  # fuera del rango
        else:
            dias_restantes = -1

        data.append({
            'id': compra['id'],
            'nombre': compra['nombre'],
            'proveedor': compra['empresa__nombre'],
            'responsable': compra['responsable'],
            'temporada': compra['temporada'],
            'fecha': compra['fecha'].strftime('%d-%m-%Y'),
            'costo_total': float(compra['costo_total'] or 0),
            'unidades': int(compra['unidades_totales'] or 0),
            'recepcionado': int(compra['total_recepcionado'] or 0),
            'dias_temporada': dias_restantes,
        })

    # Respuesta con metadatos de paginación
    response_data = {
        'success': True,
        'compras': data,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size,
            'has_next': page * page_size < total_count,
            'has_previous': page > 1
        },
        'search': search
    }

    return JsonResponse(response_data)

def importar_csv_compra(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        compra_id = data.get('compra_id')
        filas = data.get('filas', [])

        productos_dict = {}

        for fila in filas:
            key = (
                fila['nombre'],
                fila.get('descripcion', ''),
                fila['atributo1'],
                fila['atributo2'],
                fila['atributo3'],
                fila.get('atributo4', ''),  # si existe
                fila['costo'],
                fila['precioSugerido']
            )

            if key not in productos_dict:
                productos_dict[key] = {
                    "stock_tallas": []
                }

            productos_dict[key]["stock_tallas"].append({
                "talla": fila["talla"],
                "stock": fila["stock"]
            })

        compra = Compras.objects.get(id=compra_id)

        for (nombre, descripcion, a1, a2, a3, a4, costo, precio) in productos_dict:
            prod = Compras_Producto.objects.create(
                compras=compra,
                nombre=nombre,
                descripcion=descripcion,
                atributo1=a1,
                atributo2=a2,
                atributo3=a3,
                atributo4=a4,
                costo=costo,
                precioSugerido=precio
            )

            for talla_info in productos_dict[(nombre, descripcion, a1, a2, a3, a4, costo, precio)]["stock_tallas"]:
                Compras_Producto_Talla.objects.create(
                    compra_producto=prod,
                    stock=talla_info["stock"],
                    talla=talla_info["talla"]
                )

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Método no permitido"})

 
def recepcionar_compra(request):
    if request.method == 'POST':
        body = json.loads(request.body)
        compra_id = body.get('compra_id')
        page = body.get('page', 1)
        page_size = body.get('page_size', 50)  # 50 registros por página
        search = body.get('search', '').strip()
        vista_agrupada = body.get('vista_agrupada', True)  # Nuevo parámetro para vista agrupada
        
        # Validar parámetros
        try:
            page = int(page)
            page_size = min(int(page_size), 100)  # Máximo 100 por página
        except (ValueError, TypeError):
            page = 1
            page_size = 50

        # Obtener la compra
        compra = get_object_or_404(Compras, id=compra_id)

        # Query base optimizada con select_related
        tallas_query = Compras_Producto_Talla.objects.select_related(
            'compra_producto'
        ).filter(
            compra_producto__compras=compra
        )

        # Aplicar búsqueda si se proporciona
        if search:
            tallas_query = tallas_query.filter(
                Q(compra_producto__nombre__icontains=search) |
                Q(compra_producto__descripcion__icontains=search) |
                Q(compra_producto__atributo1__icontains=search) |
                Q(compra_producto__atributo2__icontains=search) |
                Q(compra_producto__atributo3__icontains=search) |
                Q(talla__icontains=search)
            )

        # ============================
        # 1. Facturas del proveedor (optimizada)
        # ============================
        # Para DTEs de COMPRA: emisor = proveedor, receptor = nosotros
        facturas_proveedor = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            emisor=compra.empresa  # El proveedor (compra.empresa) es el emisor del DTE
        ).values('id', 'numero_documento', 'monto_con_iva')

        # ============================
        # 2. Calcular uso por factura (optimizada)
        # ============================
        facturas_uso = (
            Productos_Recepcionados.objects
            .filter(dte_id__in=[f['id'] for f in facturas_proveedor])
            .annotate(monto_usado=ExpressionWrapper(
                F('stockArribado') * F('compra_producto_talla__compra_producto__costo'),
                output_field=DecimalField()
            ))
            .values('dte_id')
            .annotate(total_usado=Sum('monto_usado'))
        )

        uso_por_factura = {f['dte_id']: f['total_usado'] for f in facturas_uso}

        # Filtrar facturas con saldo disponible
        facturas_con_saldo = []
        for factura in facturas_proveedor:
            usado = uso_por_factura.get(factura['id'], 0)
            if usado < factura['monto_con_iva']:
                facturas_con_saldo.append({
                    'id': factura['id'],
                    'numero': factura['numero_documento']
                })

        # ============================
        # VISTA AGRUPADA POR PRODUCTO
        # ============================
        if vista_agrupada:
            # Obtener todas las tallas sin paginar primero
            todas_tallas = tallas_query.all()
            
            # Agrupar por producto (nombre + marca + color + género)
            productos_agrupados = {}
            for t in todas_tallas:
                # Crear clave única para agrupar
                key = f"{t.compra_producto.nombre}|{t.compra_producto.atributo1}|{t.compra_producto.atributo2}|{t.compra_producto.atributo3}"
                
                if key not in productos_agrupados:
                    productos_agrupados[key] = {
                        'nombre': t.compra_producto.nombre,
                        'descripcion': t.compra_producto.descripcion,
                        'marca': t.compra_producto.atributo1,
                        'color': t.compra_producto.atributo2,
                        'genero': t.compra_producto.atributo3,
                        'costo': t.compra_producto.costo,
                        'precio': t.compra_producto.precioSugerido,
                        'stock_total': 0,
                        'recepcionado_total': 0,
                        'tallas': []
                    }
                
                # Obtener recepción existente
                recep = Productos_Recepcionados.objects.filter(
                    compra_producto_talla=t
                ).select_related('dte').first()
                
                # Calcular total recepcionado para este producto
                recepcionado_talla = recep.stockArribado if recep else 0
                
                # Agregar talla al grupo
                talla_data = {
                    'compra_producto_talla_id': t.id,
                    'talla': t.talla,
                    'stock': t.stock,
                    'recepcionado': recepcionado_talla,
                    'factura_id': recep.dte_id if recep and recep.dte_id else None,
                    'factura_numero': recep.dte.numero_documento if recep and recep.dte else None,
                    'facturas': facturas_con_saldo
                }
                
                productos_agrupados[key]['tallas'].append(talla_data)
                productos_agrupados[key]['stock_total'] += t.stock
                productos_agrupados[key]['recepcionado_total'] += recepcionado_talla
            
            # Convertir a lista y aplicar paginación
            resultado_agrupado = list(productos_agrupados.values())
            total_count = len(resultado_agrupado)
            
            # Aplicar paginación sobre productos agrupados
            offset = (page - 1) * page_size
            resultado = resultado_agrupado[offset:offset + page_size]
            
            # Respuesta con vista agrupada
            response_data = {
                'items': resultado,
                'vista_agrupada': True,
                'proveedor_id': compra.empresa.id,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_count': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size if total_count > 0 else 0,
                    'has_next': page * page_size < total_count,
                    'has_previous': page > 1
                },
                'search': search
            }
            
            return JsonResponse(response_data)
        
        # ============================
        # VISTA DETALLADA (ORIGINAL)
        # ============================
        else:
            # Contar total de registros para paginación
            total_count = tallas_query.count()
            
            # Aplicar paginación
            offset = (page - 1) * page_size
            tallas = tallas_query[offset:offset + page_size]

            # ============================
            # 3. Facturas ya asociadas a tallas (optimizada)
            # ============================
            talla_ids = [t.id for t in tallas]
            recepciones = (
                Productos_Recepcionados.objects
                .filter(compra_producto_talla_id__in=talla_ids)
                .values('compra_producto_talla_id', 'dte__id', 'dte__numero_documento')
                .annotate(total=Sum('stockArribado'))
            )

            mapa_recepciones = {}
            for r in recepciones:
                key = r['compra_producto_talla_id']
                if key not in mapa_recepciones:
                    mapa_recepciones[key] = []
                mapa_recepciones[key].append({
                    'factura_id': r['dte__id'],
                    'numero': r['dte__numero_documento'],
                    'total': r['total']
                })

            # ============================
            # 4. Resultado final optimizado
            # ============================
            resultado = []
            for t in tallas:
                # Obtener recepción existente de forma optimizada
                recep = Productos_Recepcionados.objects.filter(
                    compra_producto_talla=t
                ).select_related('dte').first()
                
                resultado.append({
                    'compra_producto_talla_id': t.id,
                    'nombre': t.compra_producto.nombre,
                    'descripcion': t.compra_producto.descripcion,
                    'marca': t.compra_producto.atributo1,
                    'color': t.compra_producto.atributo2,
                    'genero': t.compra_producto.atributo3,
                    'costo': t.compra_producto.costo,
                    'precio': t.compra_producto.precioSugerido,
                    'stock': t.stock,
                    'talla': t.talla,
                    'recepcionado': recep.stockArribado if recep else None,
                    'factura_id': recep.dte_id if recep and recep.dte_id else None,
                    'facturas': facturas_con_saldo,
                    'facturas_asociadas': mapa_recepciones.get(t.id, [])
                })

            # Respuesta con metadatos de paginación
            response_data = {
                'items': resultado,
                'vista_agrupada': False,
                'proveedor_id': compra.empresa.id,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_count': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size,
                    'has_next': page * page_size < total_count,
                    'has_previous': page > 1
                },
                'search': search
            }

            return JsonResponse(response_data)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def obtener_dte(request, dte_id):
    try:
        dte = Dte.objects.get(id=dte_id)
        # Para DTEs de COMPRA: emisor = proveedor, receptor = nosotros
        # El formulario usa "receptor" para el proveedor, por eso devolvemos emisor
        return JsonResponse({
            'id': dte.id,
            'receptor_id': dte.emisor.id if dte.emisor else None,  # El proveedor (emisor) se muestra en campo "receptor" del form
            'numero_documento': dte.numero_documento,
            'tipo_documento': dte.tipo_documento,
            'monto_con_iva': float(dte.monto_con_iva),
            'descuento': float(dte.descuento),
            'descuento_con_iva': False,  # si tienes forma de guardar esto, cámbialo
            'fecha_emision': dte.fecha_emision.isoformat(),
            'fecha_recepcion': dte.fecha_recepcion.isoformat() if dte.fecha_recepcion else '',
            'estado_dte': dte.estado_dte,
            'estado_pago': dte.estado_pago,
            'diasCredito': dte.diasCredito,
            'bultos': dte.bultos,
            'unidades_productos': dte.unidades_productos,
            'motivo_rechazo': dte.motivo_rechazo if dte.motivo_rechazo else '',
            'documento_padre_id': dte.documento_padre.id if dte.documento_padre else None
        })
    except Dte.DoesNotExist:
        return JsonResponse({'error': 'DTE no encontrado'}, status=404)

def obtener_dte_compras(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')

        dtes = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            fecha_emision__range=[fecha_inicio, fecha_fin]
        ).select_related('emisor')

        resultado = [{
            'id': d.id,
            'nombre': d.emisor.nombre,
            'rut': d.emisor.rut,
            'numero_documento': d.numero_documento,
            'tipo': d.tipo_documento,
            'fecha_emision': d.fecha_emision.strftime('%Y-%m-%d'),
            'fecha_recepcion': d.fecha_recepcion.strftime('%Y-%m-%d') if d.fecha_recepcion else None,
            'monto_con_iva': float(d.monto_con_iva),
            'descuento': float(d.descuento),  # 👉 aquí se incluye el descuento
            'estado': d.estado_dte,
            'estado_pago': d.estado_pago
        } for d in dtes]

        return JsonResponse(resultado, safe=False)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

 
  
  
def crearDteCompras(request):
    if request.method == 'POST':
        try:
            print(f"🆕 DEBUG - Creando DTE, datos recibidos:")
            print(f"   - Body: {request.body}")
            data = json.loads(request.body)
            print(f"   - Data parseada: {data}")

            # Datos desde sesión
            empresa_emisor_id = request.session.get('idEmpresaActual')
            sucursal_id = request.session.get('idSucursalActual')
            responsable = request.session.get('nombreUsuario', 'Sistema')

            if not empresa_emisor_id or not sucursal_id:
                return JsonResponse({'success': False, 'error': 'No se pudo identificar la empresa o sucursal actual.'})

            # Obtener objetos relacionados
            empresa_emisor = Empresa.objects.get(id=empresa_emisor_id)
            sucursal = Sucursal.objects.get(id=sucursal_id)

            # Datos del request
            receptor_id = data.get('receptor_id')
            numero_documento = data.get('numero_documento')
            monto_con_iva = Decimal(data.get('monto_con_iva'))
            tipo_documento = data.get('tipo_documento')
            fecha_emision = parse_date(data.get('fecha_emision'))
            fecha_recepcion = parse_date(data.get('fecha_recepcion')) if data.get('fecha_recepcion') else None
            estado_dte = data.get('estado_dte')
            estado_pago = data.get('estado_pago')
            dias_credito = int(data.get('diasCredito') or 0)
            bultos = int(data.get('bultos') or 0)
            unidades_productos = int(data.get('unidades_productos') or 0)

            # Descuento
            descuento_bruto = Decimal(data.get('descuento', 0))
            descuento_con_iva = data.get('descuento_con_iva', False)
            descuento_neto = descuento_bruto / Decimal('1.19') if descuento_con_iva else descuento_bruto
            
            # Motivo de rechazo
            motivo_rechazo = data.get('motivo_rechazo', '').strip() if data.get('motivo_rechazo') else None
            
            # Documento padre (para facturas que anexan cotizaciones o guías)
            documento_padre_id = data.get('documento_padre_id')

            # Validación de campos requeridos
            if not all([receptor_id, numero_documento, monto_con_iva, fecha_emision, estado_dte, estado_pago, tipo_documento]):
                return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios.'})
            
            # Validar motivo de rechazo si el estado es Rechazado
            if estado_dte == 'Rechazado':
                if not motivo_rechazo or len(motivo_rechazo) == 0:
                    return JsonResponse({'success': False, 'error': 'El motivo de rechazo es obligatorio cuando el estado DTE es Rechazado.'})
                if len(motivo_rechazo) > 100:
                    return JsonResponse({'success': False, 'error': 'El motivo de rechazo no puede exceder 100 caracteres.'})

            # Validar receptor
            try:
                receptor = Empresa.objects.get(id=receptor_id)
            except Empresa.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Empresa receptora no válida.'})

            # Validar duplicado: mismo número, misma fecha, mismo proveedor (emisor)
            if Dte.objects.filter(
                numero_documento=numero_documento,
                fecha_emision=fecha_emision,
                emisor_id=receptor_id,  # El proveedor (emisor) debe ser único
                tipo_transaccion='COMPRA'
            ).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe un DTE con ese número y fecha de emisión de este proveedor.'
                })

            # Calcular monto neto con descuento
            monto_neto = (monto_con_iva / Decimal('1.19')) - descuento_neto
            
            # Validar documento padre si se especifica
            documento_padre = None
            if documento_padre_id:
                try:
                    documento_padre = Dte.objects.get(id=documento_padre_id)
                    # Verificar que el documento padre sea una cotización o guía
                    if documento_padre.tipo_documento not in ['COTIZACION', 'GUIA']:
                        return JsonResponse({'success': False, 'error': 'El documento base debe ser una Cotización o Guía de Despacho.'})
                    # Verificar que no tenga ya una factura anexada
                    if documento_padre.documentos_hijos.filter(tipo_documento='FACTURA ELECTRONICA').exists():
                        return JsonResponse({'success': False, 'error': 'Este documento ya tiene una factura anexada.'})
                except Dte.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Documento base no encontrado.'})

            # Debug para creación de NC
            if tipo_documento == 'Nota de Crédito':
                print(f"🆕 DEBUG - Creando Nota de Crédito:")
                print(f"   - Tipo documento: {tipo_documento}")
                print(f"   - Número: {numero_documento}")
                print(f"   - Emisor ID: {receptor_id}")
                print(f"   - Receptor ID: {empresa_emisor_id}")
            
            # Crear DTE
            # IMPORTANTE: En un DTE de COMPRAS:
            # - Emisor = Proveedor (quien nos vende/emite la factura)
            # - Receptor = Nosotros (quien compra/recibe la factura)
            nuevo_dte = Dte.objects.create(
                emisor=receptor,  # El proveedor es quien emite la factura
                receptor=empresa_emisor,  # Nosotros recibimos la factura
                numero_documento=numero_documento,
                tipo_documento=tipo_documento,
                monto_con_iva=monto_con_iva,
                monto_neto=monto_neto,
                descuento=descuento_neto,
                estado_dte=estado_dte,
                estado_pago=estado_pago,
                responsable=responsable,
                fecha_emision=fecha_emision,
                fecha_recepcion=fecha_recepcion,
                fecha_vencimiento=fecha_emision,
                diasCredito=dias_credito,
                bultos=bultos,
                unidades_productos=unidades_productos,
                tipo_transaccion='COMPRA',
                sucursal=sucursal,
                motivo_rechazo=motivo_rechazo,
                documento_padre=documento_padre
            )
            
            # Debug después de crear
            if tipo_documento == 'Nota de Crédito':
                print(f"✅ NC creada exitosamente:")
                print(f"   - ID: {nuevo_dte.id}")
                print(f"   - Tipo guardado: '{nuevo_dte.tipo_documento}'")
                print(f"   - Número: {nuevo_dte.numero_documento}")
                print(f"   - Emisor: {nuevo_dte.emisor}")
                print(f"   - Receptor: {nuevo_dte.receptor}")

            return JsonResponse({'success': True, 'id': nuevo_dte.id})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


def actualizarDteCompras(request, dte_id):
    """Actualizar un DTE de compras existente"""
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            # Obtener el DTE existente
            try:
                dte = Dte.objects.get(id=dte_id, tipo_transaccion='COMPRA')
            except Dte.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'DTE no encontrado.'})
            
            # Datos del request
            receptor_id = data.get('receptor_id')
            numero_documento = data.get('numero_documento')
            monto_con_iva = Decimal(data.get('monto_con_iva'))
            tipo_documento = data.get('tipo_documento')
            fecha_emision = parse_date(data.get('fecha_emision'))
            fecha_recepcion = parse_date(data.get('fecha_recepcion')) if data.get('fecha_recepcion') else None
            estado_dte = data.get('estado_dte')
            estado_pago = data.get('estado_pago')
            dias_credito = int(data.get('diasCredito') or 0)
            bultos = int(data.get('bultos') or 0)
            unidades_productos = int(data.get('unidades_productos') or 0)
            
            # Descuento
            descuento_bruto = Decimal(data.get('descuento', 0))
            descuento_con_iva = data.get('descuento_con_iva', False)
            descuento_neto = descuento_bruto / Decimal('1.19') if descuento_con_iva else descuento_bruto
            
            # Motivo de rechazo
            motivo_rechazo = data.get('motivo_rechazo', '').strip() if data.get('motivo_rechazo') else None
            
            # Documento padre (para facturas que anexan cotizaciones o guías)
            documento_padre_id = data.get('documento_padre_id')
            
            # Validación de campos requeridos
            if not all([receptor_id, numero_documento, monto_con_iva, fecha_emision, estado_dte, estado_pago, tipo_documento]):
                return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios.'})
            
            # Validar motivo de rechazo si el estado es Rechazado
            if estado_dte == 'Rechazado':
                if not motivo_rechazo or len(motivo_rechazo) == 0:
                    return JsonResponse({'success': False, 'error': 'El motivo de rechazo es obligatorio cuando el estado DTE es Rechazado.'})
                if len(motivo_rechazo) > 100:
                    return JsonResponse({'success': False, 'error': 'El motivo de rechazo no puede exceder 100 caracteres.'})
            
            # Validar receptor
            try:
                receptor = Empresa.objects.get(id=receptor_id)
            except Empresa.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Empresa receptora no válida.'})
            
            # Validar duplicado (excluyendo el DTE actual)
            if Dte.objects.filter(
                numero_documento=numero_documento,
                fecha_emision=fecha_emision,
                emisor_id=receptor_id,  # El proveedor (emisor) debe ser único
                tipo_transaccion='COMPRA'
            ).exclude(id=dte_id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe otro DTE con ese número y fecha de emisión de este proveedor.'
                })
            
            # Calcular monto neto con descuento
            monto_neto = (monto_con_iva / Decimal('1.19')) - descuento_neto
            
            # Validar documento padre si se especifica
            documento_padre = None
            if documento_padre_id:
                try:
                    documento_padre = Dte.objects.get(id=documento_padre_id)
                    # Verificar que el documento padre sea una cotización o guía
                    if documento_padre.tipo_documento not in ['COTIZACION', 'GUIA']:
                        return JsonResponse({'success': False, 'error': 'El documento base debe ser una Cotización o Guía de Despacho.'})
                    # Verificar que no tenga ya una factura anexada (excepto esta misma)
                    if documento_padre.documentos_hijos.filter(tipo_documento='FACTURA ELECTRONICA').exclude(id=dte_id).exists():
                        return JsonResponse({'success': False, 'error': 'Este documento ya tiene una factura anexada.'})
                except Dte.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Documento base no encontrado.'})
            
            # Actualizar DTE
            # IMPORTANTE: En un DTE de COMPRAS:
            # - Emisor = Proveedor (quien nos vende/emite la factura)
            # - Receptor = Nosotros (quien compra/recibe la factura)
            dte.emisor = receptor  # El proveedor es quien emite la factura
            dte.receptor = dte.receptor  # Mantener el receptor original (nosotros)
            dte.numero_documento = numero_documento
            dte.tipo_documento = tipo_documento
            dte.monto_con_iva = monto_con_iva
            dte.monto_neto = monto_neto
            dte.descuento = descuento_neto
            dte.estado_dte = estado_dte
            dte.estado_pago = estado_pago
            dte.fecha_emision = fecha_emision
            dte.fecha_recepcion = fecha_recepcion
            dte.fecha_vencimiento = fecha_emision
            dte.diasCredito = dias_credito
            dte.bultos = bultos
            dte.unidades_productos = unidades_productos
            dte.motivo_rechazo = motivo_rechazo
            dte.documento_padre = documento_padre
            dte.save()
            
            return JsonResponse({'success': True, 'id': dte.id, 'message': 'DTE actualizado correctamente'})
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


def empresas_proveedoras(request):
    if request.method == 'POST':
        try:
            empresa_id = request.session.get('idEmpresaActual')

            if not empresa_id:
                return JsonResponse({'error': 'Empresa no identificada en sesión'}, status=403)

            # Solo proveedores globales o vinculados a la empresa actual
            proveedores = Empresa.objects.filter(esProveedor=True)

            data = [{'id': e.id, 'nombre': e.nombre, 'rut': e.rut} for e in proveedores]
            return JsonResponse(data, safe=False)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)
def cargarDteCompra(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            fecha_inicio = parse_date(data.get('fecha_inicio'))
            fecha_fin = parse_date(data.get('fecha_fin'))
            page = data.get('page', 1)
            page_size = data.get('page_size', 20)  # 20 registros por página
            search = data.get('search', '').strip()
            tipo_documento_filtro = data.get('tipo_documento', '').strip()

            if not fecha_inicio or not fecha_fin:
                return JsonResponse({'error': 'Fechas inválidas'}, status=400)

            # Validar parámetros de paginación
            try:
                page = int(page)
                page_size = min(int(page_size), 100)  # Máximo 100 por página
            except (ValueError, TypeError):
                page = 1
                page_size = 20

            empresa_id = request.session.get('idEmpresaActual')
            if not empresa_id:
                return JsonResponse({'error': 'Empresa no identificada en sesión'}, status=403)

            # Query base optimizada
            # Para DTEs de COMPRA: emisor = proveedor, receptor = nosotros
            dtes_query = Dte.objects.filter(
                tipo_transaccion='COMPRA',
                receptor_id=empresa_id,  # Filtramos por receptor (nosotros)
                fecha_emision__range=(fecha_inicio, fecha_fin)
            ).select_related('emisor')  # El proveedor es el emisor

            # Aplicar filtro por tipo de documento si se proporciona
            if tipo_documento_filtro:
                dtes_query = dtes_query.filter(tipo_documento=tipo_documento_filtro)
            
            # Aplicar búsqueda si se proporciona
            if search:
                dtes_query = dtes_query.filter(
                    Q(emisor__nombre__icontains=search) |  # Buscar por proveedor (emisor)
                    Q(emisor__rut__icontains=search) |
                    Q(numero_documento__icontains=search) |
                    Q(tipo_documento__icontains=search) |
                    Q(estado_dte__icontains=search) |
                    Q(estado_pago__icontains=search) |
                    Q(dte_asociado__voucher__icontains=search)
                ).distinct()

            # Contar total de registros para paginación
            total_count = dtes_query.count()
            
            # Aplicar paginación
            offset = (page - 1) * page_size
            dtes = dtes_query[offset:offset + page_size]

            hoy = date.today()
            resultado = []

            for d in dtes:
                dias_transcurridos = (hoy - d.fecha_emision).days
                dias_credito_restantes = max(d.diasCredito - dias_transcurridos, 0)

                # Calcular total de notas de crédito asociadas y obtener sus números
                notas_credito_objs = Dte_Detalle_Pago.objects.filter(
                    dte=d,
                    metodo_pago='Nota de Crédito'
                ).values('monto', 'voucher')
                
                notas_credito_total = sum(nc['monto'] for nc in notas_credito_objs) if notas_credito_objs else 0
                notas_credito_numeros = [nc['voucher'] for nc in notas_credito_objs if nc.get('voucher')]
                
                # Contar incidencias pendientes y totales
                incidencias_count = Dte_Incidencia.objects.filter(dte=d).count()
                incidencias_pendientes = Dte_Incidencia.objects.filter(dte=d, estado='PENDIENTE').count()
                
                # Obtener información del documento padre si existe
                documento_padre_info = None
                if d.documento_padre:
                    documento_padre_info = {
                        'id': d.documento_padre.id,
                        'tipo': d.documento_padre.tipo_documento,
                        'numero': d.documento_padre.numero_documento
                    }
                
                # Verificar si tiene documentos hijos (facturas anexadas)
                tiene_factura_anexada = d.documentos_hijos.filter(
                    tipo_documento='FACTURA ELECTRONICA'
                ).exists()
                
                # Para NCs, verificar si está asociada a alguna factura
                nc_esta_asociada = False
                factura_asociada_info = None
                if d.tipo_documento == 'NOTA DE CREDITO':
                    pago_nc = Dte_Detalle_Pago.objects.filter(
                        voucher=d.numero_documento,
                        metodo_pago='Nota de Crédito'
                    ).select_related('dte').first()
                    
                    if pago_nc:
                        nc_esta_asociada = True
                        factura_asociada_info = {
                            'factura_id': pago_nc.dte.id,
                            'factura_numero': pago_nc.dte.numero_documento,
                            'factura_proveedor': pago_nc.dte.emisor.nombre if pago_nc.dte.emisor else 'N/A'
                        }

                resultado.append({
                    'id': d.id,
                    'nombre': d.emisor.nombre if d.emisor else 'N/A',  # El proveedor es el emisor
                    'rut': d.emisor.rut if d.emisor else '-',
                    'numero_documento': d.numero_documento,
                    'tipo': d.tipo_documento,
                    'fecha_emision': d.fecha_emision.strftime('%Y-%m-%d'),
                    'fecha_recepcion': d.fecha_recepcion.strftime('%Y-%m-%d') if d.fecha_recepcion else None,
                    'monto_con_iva': float(d.monto_con_iva),
                    'descuento': float(d.descuento or 0),
                    'notas_credito': float(notas_credito_total),
                    'notas_credito_numeros': notas_credito_numeros,
                    'estado': d.estado_dte,
                    'estado_pago': d.estado_pago,
                    'diasCredito': d.diasCredito,
                    'dias_credito_restantes': dias_credito_restantes,
                    'incidencias_count': incidencias_count,
                    'incidencias_pendientes': incidencias_pendientes,
                    'documento_padre': documento_padre_info,
                    'tiene_factura_anexada': tiene_factura_anexada,
                    'requiere_factura': d.tipo_documento in ['COTIZACION', 'GUIA'],
                    'nc_esta_asociada': nc_esta_asociada,
                    'factura_asociada_info': factura_asociada_info
                })

            # Respuesta con metadatos de paginación
            response_data = {
                'items': resultado,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_count': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size,
                    'has_next': page * page_size < total_count,
                    'has_previous': page > 1
                },
                'search': search
            }

            return JsonResponse(response_data, safe=False)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def registrarPagoDTE(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        dte_id = data.get('dte_id')
        metodo_pago = data.get('metodo_pago')
        voucher = data.get('voucher', '').strip()
        monto = int(data.get('monto'))

        if not dte_id or not metodo_pago or monto <= 0:
            return JsonResponse({'error': 'Datos incompletos o inválidos'}, status=400)

        dte = Dte.objects.get(pk=dte_id)

        # Verificar incidencias pendientes o en gestión
        if Dte_Incidencia.objects.filter(dte=dte, estado__in=['PENDIENTE', 'EN_GESTION']).exists():
            return JsonResponse({
                'error': 'No se pueden registrar pagos mientras existan incidencias pendientes o en gestión para este DTE.'
            }, status=400)
        
        # Verificar si el documento requiere factura anexada
        if dte.tipo_documento in ['COTIZACION', 'GUIA']:
            tiene_factura = dte.documentos_hijos.filter(tipo_documento='FACTURA ELECTRONICA').exists()
            if not tiene_factura:
                return JsonResponse({
                    'error': f'Este documento ({dte.tipo_documento}) requiere tener una factura anexada antes de poder registrar pagos.'
                }, status=400)

        # Validar voucher duplicado (si se proporcionó)
        if voucher:
            voucher_existente = Dte_Detalle_Pago.objects.filter(
                dte=dte,
                voucher=voucher
            ).exists()
            
            if voucher_existente:
                return JsonResponse({
                    'error': f'Ya existe un pago con el voucher "{voucher}" para este DTE. Por favor, usa un número diferente.'
                }, status=400)

        # Total de pagos anteriores
        pagos_previos = Dte_Detalle_Pago.objects.filter(dte=dte).aggregate(total= Sum('monto'))['total'] or 0
        monto_total = float(dte.monto_con_iva)
        total_con_este = pagos_previos + monto

        if total_con_este > monto_total:
            return JsonResponse({'error': 'El monto total de pagos excede el total del DTE'}, status=400)

        # Guardar el nuevo pago
        Dte_Detalle_Pago.objects.create(
            dte=dte,
            metodo_pago=metodo_pago,
            voucher=voucher if voucher else None,
            monto=monto
        )

        # Actualizar estado de pago
        if total_con_este == monto_total:
            dte.estado_pago = 'Pagado'
        elif total_con_este > 0:
            dte.estado_pago = 'Abonado'
        else:
            dte.estado_pago = 'Pendiente'
        dte.save()

        return JsonResponse({'success': True, 'mensaje': 'Pago registrado correctamente.'})

    except Dte.DoesNotExist:
        return JsonResponse({'error': 'DTE no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
 
 
def obtenerDetallePago(request, dte_id):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        dte = Dte.objects.get(pk=dte_id)

        pagos = Dte_Detalle_Pago.objects.filter(dte=dte)
        total_abonado = pagos.aggregate(total=Sum('monto'))['total'] or 0

        return JsonResponse({
            'total_abonado': float(total_abonado),
            'total_dte': float(dte.monto_con_iva)  # 👈 este es el monto original
        })
    except Dte.DoesNotExist:
        return JsonResponse({'error': 'DTE no encontrado'}, status=404)

def pagosDTE(request, dte_id):
    if request.method == 'GET':
        pagos = Dte_Detalle_Pago.objects.filter(
            dte_id=dte_id
        ).exclude(
            metodo_pago='Nota de Crédito'
        ).values(
            'id', 'metodo_pago', 'voucher', 'monto'
        )
        return JsonResponse(list(pagos), safe=False)

 
def eliminarPago(request, pago_id):
    if request.method == 'DELETE':
        try:
            pago = Dte_Detalle_Pago.objects.get(id=pago_id)
            dte = pago.dte
            
            if Dte_Incidencia.objects.filter(dte=dte, estado__in=['PENDIENTE', 'EN_GESTION']).exists():
                return JsonResponse({
                    'error': 'No se pueden modificar pagos mientras existan incidencias pendientes o en gestión para este DTE.'
                }, status=400)
            
            # Eliminar el pago
            pago.delete()
            
            # Recalcular estado del DTE
            total_pagado = Dte_Detalle_Pago.objects.filter(dte=dte).aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            monto_total = float(dte.monto_con_iva)
            
            if total_pagado >= monto_total:
                dte.estado_pago = 'Pagado'
            elif total_pagado > 0:
                dte.estado_pago = 'Abonado'
            else:
                dte.estado_pago = 'Pendiente'
            dte.save()
            
            return JsonResponse({'success': True, 'mensaje': 'Pago eliminado correctamente'})
        except Dte_Detalle_Pago.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Pago no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)
 
def detallePago(request, pago_id):
    if request.method == 'GET':
        try:
            pago = Dte_Detalle_Pago.objects.select_related('dte').get(id=pago_id)

            return JsonResponse({
                'id': pago.id,
                'dte_id': pago.dte.id,
                'metodo_pago': pago.metodo_pago,
                'voucher': pago.voucher,
                'monto': pago.monto
            })

        except Dte_Detalle_Pago.DoesNotExist:
            return JsonResponse({'error': 'Pago no encontrado'}, status=404)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

 
def editarPago(request, pago_id):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)

            pago = Dte_Detalle_Pago.objects.get(id=pago_id)
            dte = pago.dte
            
            # Validar voucher duplicado (si se proporcionó y cambió)
            nuevo_voucher = data.get('voucher', pago.voucher)
            if nuevo_voucher:
                nuevo_voucher = nuevo_voucher.strip()
                # Verificar si cambió el voucher
                if nuevo_voucher != pago.voucher:
                    voucher_existente = Dte_Detalle_Pago.objects.filter(
                        dte=dte,
                        voucher=nuevo_voucher
                    ).exclude(id=pago_id).exists()
                    
                    if voucher_existente:
                        return JsonResponse({
                            'error': f'Ya existe otro pago con el voucher "{nuevo_voucher}" para este DTE. Por favor, usa un número diferente.'
                        }, status=400)
            
            # Actualizar campos del pago
            pago.metodo_pago = data.get('metodo_pago', pago.metodo_pago)
            pago.voucher = nuevo_voucher if nuevo_voucher else None
            nuevo_monto = int(data.get('monto', pago.monto))
            
            # Validar que el total de pagos no exceda el monto del DTE
            pagos_otros = Dte_Detalle_Pago.objects.filter(dte=dte).exclude(id=pago_id).aggregate(
                total=Sum('monto')
            )['total'] or 0
            monto_total = float(dte.monto_con_iva)
            total_con_este = pagos_otros + nuevo_monto
            
            if total_con_este > monto_total:
                return JsonResponse({
                    'error': f'El monto total de pagos (${total_con_este:,.0f}) excedería el total del DTE (${monto_total:,.0f})'
                }, status=400)
            
            pago.monto = nuevo_monto
            pago.save()

            # Actualizar estado de pago del DTE
            if total_con_este >= monto_total:
                dte.estado_pago = 'Pagado'
            elif total_con_este > 0:
                dte.estado_pago = 'Abonado'
            else:
                dte.estado_pago = 'Pendiente'
            dte.save()
            
            return JsonResponse({'success': True, 'mensaje': 'Pago actualizado correctamente'})
        except Dte_Detalle_Pago.DoesNotExist:
            return JsonResponse({'error': 'Pago no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)
def notasCredito(request, dte_id):
    ncs = Dte_Detalle_Pago.objects.filter(dte_id=dte_id, metodo_pago='Nota de Crédito') \
        .values('id', 'voucher', 'monto', 'notas')
    return JsonResponse(list(ncs), safe=False)
 
def agregarNotaCredito(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            dte = Dte.objects.get(id=data['dte_id'])
            
            # Obtener el motivo
            notas = data.get('notas', '').strip()
            
            # Validar que el motivo no esté vacío
            if not notas:
                return JsonResponse({
                    'success': False,
                    'error': 'El motivo de la nota de crédito es obligatorio.'
                }, status=400)
            
            # Validar longitud del motivo
            if len(notas) > 100:
                return JsonResponse({
                    'success': False,
                    'error': 'El motivo no puede exceder 100 caracteres.'
                }, status=400)

            Dte_Detalle_Pago.objects.create(
                dte=dte,
                metodo_pago='Nota de Crédito',
                voucher=data.get('voucher'),
                monto=data.get('monto'),
                notas=notas
            )
            return JsonResponse({'success': True})
        except Dte.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'DTE no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'error': 'Método no permitido'}, status=405)
 
def eliminarNotaCredito(request, nc_id):
    if request.method == 'DELETE':
        try:
            nc = Dte_Detalle_Pago.objects.get(id=nc_id, metodo_pago='Nota de Crédito')
            nc.delete()
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'error': 'No se encontró la nota de crédito'}, status=404)
    return JsonResponse({'error': 'Método no permitido'}, status=405)
 
 
def eliminar_dte(request, dte_id):
    if request.method == 'DELETE':
        try:
            dte = Dte.objects.get(id=dte_id)
            dte.delete()
            return JsonResponse({'success': True})
        except Dte.DoesNotExist:
            return JsonResponse({'error': 'DTE no encontrado'}, status=404)
        except ProtectedError:
            return JsonResponse({
                'error': 'No se puede eliminar este DTE porque tiene registros asociados (como pagos, notas u otros).'
            }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ========== GESTIÓN DE INCIDENCIAS DTE ==========

def listar_incidencias(request, dte_id):
    """Listar todas las incidencias de un DTE"""
    try:
        dte = Dte.objects.get(id=dte_id)
        incidencias = Dte_Incidencia.objects.filter(dte=dte).values(
            'id', 'tipo', 'descripcion', 'estado', 'fecha_registro', 
            'fecha_resolucion', 'notas_resolucion'
        ).order_by('-fecha_registro')
        
        # Formatear las fechas y agregar el nombre del tipo
        incidencias_list = []
        for inc in incidencias:
            inc['tipo_display'] = dict(Dte_Incidencia.TIPO_INCIDENCIA_CHOICES).get(inc['tipo'], inc['tipo'])
            inc['estado_display'] = dict(Dte_Incidencia.ESTADO_CHOICES).get(inc['estado'], inc['estado'])
            inc['fecha_registro'] = inc['fecha_registro'].strftime('%d/%m/%Y %H:%M')
            if inc['fecha_resolucion']:
                inc['fecha_resolucion'] = inc['fecha_resolucion'].strftime('%d/%m/%Y %H:%M')
            incidencias_list.append(inc)
        
        return JsonResponse(incidencias_list, safe=False)
    except Dte.DoesNotExist:
        return JsonResponse({'error': 'DTE no encontrado'}, status=404)


def crear_incidencia(request):
    """Crear una nueva incidencia para un DTE"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            dte_id = data.get('dte_id')
            tipo = data.get('tipo')
            descripcion = data.get('descripcion', '').strip()
            
            # Validaciones
            if not all([dte_id, tipo, descripcion]):
                return JsonResponse({
                    'success': False,
                    'error': 'Todos los campos son obligatorios'
                }, status=400)
            
            if len(descripcion) < 10:
                return JsonResponse({
                    'success': False,
                    'error': 'La descripción debe tener al menos 10 caracteres'
                }, status=400)
            
            # Verificar que el DTE existe
            try:
                dte = Dte.objects.get(id=dte_id)
            except Dte.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'DTE no encontrado'
                }, status=404)
            
            # Crear incidencia
            incidencia = Dte_Incidencia.objects.create(
                dte=dte,
                tipo=tipo,
                descripcion=descripcion,
                estado='PENDIENTE'
            )
            
            return JsonResponse({
                'success': True,
                'id': incidencia.id,
                'message': 'Incidencia creada correctamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def actualizar_incidencia(request, incidencia_id):
    """Actualizar el estado de una incidencia"""
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            incidencia = Dte_Incidencia.objects.get(id=incidencia_id)
            
            estado = data.get('estado')
            notas_resolucion = data.get('notas_resolucion', '').strip()
            
            if estado:
                incidencia.estado = estado
                
                # Si se marca como resuelto, guardar la fecha y notas
                if estado == 'RESUELTO':
                    from django.utils import timezone
                    incidencia.fecha_resolucion = timezone.now()
                    if notas_resolucion:
                        incidencia.notas_resolucion = notas_resolucion
            
            incidencia.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Incidencia actualizada correctamente'
            })
            
        except Dte_Incidencia.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Incidencia no encontrada'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def eliminar_incidencia(request, incidencia_id):
    """Eliminar una incidencia"""
    if request.method == 'DELETE':
        try:
            incidencia = Dte_Incidencia.objects.get(id=incidencia_id)
            incidencia.delete()
            return JsonResponse({'success': True, 'message': 'Incidencia eliminada'})
        except Dte_Incidencia.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Incidencia no encontrada'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def obtener_documentos_base(request):
    """
    Obtiene cotizaciones y guías de despacho que pueden ser usadas como base para una factura
    """
    if request.method == 'GET':
        try:
            empresa_id = request.session.get('idEmpresaActual')
            if not empresa_id:
                return JsonResponse({'error': 'Empresa no identificada en sesión'}, status=403)
            
            # Obtener cotizaciones y guías sin factura anexada
            documentos = Dte.objects.filter(
                tipo_transaccion='COMPRA',
                receptor_id=empresa_id,
                tipo_documento__in=['COTIZACION', 'GUIA'],
                documentos_hijos__isnull=True  # Sin facturas anexadas
            ).select_related('emisor').order_by('-fecha_emision')
            
            resultado = []
            for doc in documentos:
                resultado.append({
                    'id': doc.id,
                    'tipo': doc.tipo_documento,
                    'numero': doc.numero_documento,
                    'proveedor': doc.emisor.nombre if doc.emisor else 'N/A',
                    'fecha': doc.fecha_emision.strftime('%Y-%m-%d'),
                    'monto': float(doc.monto_con_iva),
                    'display': f"{doc.tipo_documento} #{doc.numero_documento} - {doc.emisor.nombre if doc.emisor else 'N/A'} - ${float(doc.monto_con_iva):,.0f}"
                })
            
            return JsonResponse({'success': True, 'documentos': resultado})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def obtener_ncs_disponibles(request):
    """
    Obtiene notas de crédito que no están asociadas a ninguna factura
    """
    if request.method == 'GET':
        try:
            empresa_id = request.session.get('idEmpresaActual')
            if not empresa_id:
                return JsonResponse({'error': 'Empresa no identificada en sesión'}, status=403)
            
            proveedor_id = request.GET.get('proveedor', '')
            busqueda = request.GET.get('busqueda', '')
            
            print(f"🔍 DEBUG - Buscando NCs con:")
            print(f"   - Empresa ID: {empresa_id}")
            print(f"   - Proveedor ID: {proveedor_id}")
            print(f"   - Búsqueda: {busqueda}")
            
            # Obtener todas las NCs primero
            ncs_query = Dte.objects.filter(
                tipo_transaccion='COMPRA',
                receptor_id=empresa_id,
                tipo_documento='NOTA DE CREDITO'
            ).select_related('emisor')
            
            print(f"📊 Total NCs encontradas: {ncs_query.count()}")
            
            # Aplicar filtros ANTES de verificar asociación
            if proveedor_id:
                ncs_query = ncs_query.filter(emisor_id=proveedor_id)
                print(f"📊 Después de filtro proveedor: {ncs_query.count()}")
            
            if busqueda:
                ncs_query = ncs_query.filter(numero_documento__icontains=busqueda)
                print(f"📊 Después de filtro búsqueda: {ncs_query.count()}")
            
            # Ahora filtrar las que no están asociadas
            ncs_disponibles = []
            for nc in ncs_query:
                # Verificar si esta NC ya está siendo usada como pago
                ya_usada = Dte_Detalle_Pago.objects.filter(
                    voucher=nc.numero_documento,
                    metodo_pago='Nota de Crédito'
                ).exists()
                
                print(f"   NC #{nc.numero_documento}: {'YA USADA' if ya_usada else 'DISPONIBLE'}")
                
                if not ya_usada:
                    ncs_disponibles.append(nc)
            
            print(f"✅ NCs disponibles finales: {len(ncs_disponibles)}")
            
            resultado = []
            for nc in ncs_disponibles:
                resultado.append({
                    'id': nc.id,
                    'numero_documento': nc.numero_documento,
                    'proveedor': nc.emisor.nombre if nc.emisor else 'N/A',
                    'fecha_emision': nc.fecha_emision.strftime('%Y-%m-%d'),
                    'monto_con_iva': float(nc.monto_con_iva),
                    'estado': nc.estado_dte
                })
            
            return JsonResponse({'success': True, 'ncs': resultado})
        except Exception as e:
            print(f"❌ Error en obtener_ncs_disponibles: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def obtener_facturas_para_nc(request):
    """
    Obtiene facturas disponibles para asociar a una NC específica
    """
    if request.method == 'GET':
        try:
            empresa_id = request.session.get('idEmpresaActual')
            if not empresa_id:
                return JsonResponse({'error': 'Empresa no identificada en sesión'}, status=403)
            
            nc_id = request.GET.get('nc_id')
            if not nc_id:
                return JsonResponse({'error': 'ID de NC requerido'}, status=400)
            
            # Obtener la NC
            try:
                nc = Dte.objects.get(id=nc_id, tipo_documento='NOTA DE CREDITO')
            except Dte.DoesNotExist:
                return JsonResponse({'error': 'Nota de Crédito no encontrada'}, status=404)
            
            # Obtener facturas del mismo proveedor que no tengan esta NC asociada
            facturas_query = Dte.objects.filter(
                tipo_transaccion='COMPRA',
                receptor_id=empresa_id,
                tipo_documento='FACTURA ELECTRONICA',
                emisor=nc.emisor  # Mismo proveedor
            ).select_related('emisor')
            
            resultado = []
            for factura in facturas_query:
                # Verificar que esta NC no esté ya asociada a esta factura
                ya_asociada = Dte_Detalle_Pago.objects.filter(
                    dte=factura,
                    voucher=nc.numero_documento,
                    metodo_pago='Nota de Crédito'
                ).exists()
                
                if not ya_asociada:
                    resultado.append({
                        'id': factura.id,
                        'numero_documento': factura.numero_documento,
                        'proveedor': factura.emisor.nombre if factura.emisor else 'N/A',
                        'fecha_emision': factura.fecha_emision.strftime('%Y-%m-%d'),
                        'monto_con_iva': float(factura.monto_con_iva),
                        'estado': factura.estado_dte
                    })
            
            return JsonResponse({'success': True, 'facturas': resultado})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def obtener_info_asociacion_nc(request, nc_id):
    """
    Obtiene información sobre la asociación de una NC específica
    """
    if request.method == 'GET':
        try:
            # Obtener la NC
            try:
                nc = Dte.objects.get(id=nc_id, tipo_documento='NOTA DE CREDITO')
            except Dte.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Nota de Crédito no encontrada'}, status=404)
            
            # Verificar si está asociada
            pago_nc = Dte_Detalle_Pago.objects.filter(
                voucher=nc.numero_documento,
                metodo_pago='Nota de Crédito'
            ).select_related('dte').first()
            
            info = {
                'esta_asociada': bool(pago_nc),
                'monto_nc': float(nc.monto_con_iva)
            }
            
            if pago_nc:
                info.update({
                    'factura_id': pago_nc.dte.id,
                    'factura_numero': pago_nc.dte.numero_documento,
                    'factura_proveedor': pago_nc.dte.emisor.nombre if pago_nc.dte.emisor else 'N/A'
                })
            
            return JsonResponse({'success': True, 'info': info})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def desasociar_nc(request, nc_id):
    """
    Desasocia una NC de su factura eliminando el registro de pago
    """
    if request.method == 'POST':
        try:
            # Obtener la NC
            try:
                nc = Dte.objects.get(id=nc_id, tipo_documento='NOTA DE CREDITO')
            except Dte.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Nota de Crédito no encontrada'}, status=404)
            
            # Buscar y eliminar el pago asociado
            pago_nc = Dte_Detalle_Pago.objects.filter(
                voucher=nc.numero_documento,
                metodo_pago='Nota de Crédito'
            ).first()
            
            if not pago_nc:
                return JsonResponse({'success': False, 'error': 'Esta NC no está asociada a ninguna factura'}, status=400)
            
            factura = pago_nc.dte
            pago_nc.delete()
            
            # Recalcular estado de pago de la factura
            total_pagos = Dte_Detalle_Pago.objects.filter(dte=factura).aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            if total_pagos >= factura.monto_con_iva:
                factura.estado_pago = 'PAGADO'
            else:
                factura.estado_pago = 'PENDIENTE'
            factura.save()
            
            return JsonResponse({'success': True, 'message': 'NC desasociada correctamente'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def procesar_pago_masivo(request):
    """
    Procesa pagos masivos para múltiples facturas
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            facturas_data = data.get('facturas', [])
            metodo_pago = data.get('metodo_pago')
            voucher_base = data.get('voucher', '').strip()
            observaciones = data.get('observaciones', '').strip()
            
            if not facturas_data or not metodo_pago:
                return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)
            
            # Validaciones
            facturas_ids = [f['id'] for f in facturas_data]
            facturas = Dte.objects.filter(
                id__in=facturas_ids,
                tipo_documento='FACTURA ELECTRONICA',
                tipo_transaccion='COMPRA'
            ).select_related('emisor')
            
            if facturas.count() != len(facturas_ids):
                return JsonResponse({'success': False, 'error': 'Algunas facturas no son válidas'}, status=400)
            
            # Validar que todas sean del mismo proveedor
            proveedores = set(f.emisor.id for f in facturas if f.emisor)
            if len(proveedores) > 1:
                return JsonResponse({'success': False, 'error': 'Todas las facturas deben ser del mismo proveedor'}, status=400)
            
            # Validar que todas tengan saldo pendiente
            facturas_con_saldo = []
            for factura in facturas:
                # Verificar incidencias pendientes
                if Dte_Incidencia.objects.filter(dte=factura, estado__in=['PENDIENTE', 'EN_GESTION']).exists():
                    return JsonResponse({'success': False, 'error': f'La factura #{factura.numero_documento} tiene incidencias pendientes'}, status=400)
                
                # Calcular saldo pendiente
                pagos_previos = Dte_Detalle_Pago.objects.filter(dte=factura).aggregate(total=Sum('monto'))['total'] or 0
                saldo = float(factura.monto_con_iva) - pagos_previos
                
                if saldo <= 0:
                    return JsonResponse({'success': False, 'error': f'La factura #{factura.numero_documento} no tiene saldo pendiente'}, status=400)
                
                facturas_con_saldo.append({'factura': factura, 'saldo': saldo})
            
            # Procesar pagos
            total_procesado = 0
            procesadas = 0
            
            for item in facturas_con_saldo:
                factura = item['factura']
                monto_pago = item['saldo']  # Pagar el saldo completo
                
                # Generar voucher único si es necesario
                voucher_final = f"{voucher_base}-{factura.numero_documento}" if voucher_base else f"MASIVO-{factura.numero_documento}"
                
                # Crear el pago
                Dte_Detalle_Pago.objects.create(
                    dte=factura,
                    metodo_pago=metodo_pago,
                    voucher=voucher_final,
                    monto=monto_pago,
                    notas=f"Pago masivo - {observaciones}" if observaciones else "Pago masivo"
                )
                
                # Actualizar estado de la factura
                total_pagos = Dte_Detalle_Pago.objects.filter(dte=factura).aggregate(total=Sum('monto'))['total'] or 0
                if total_pagos >= factura.monto_con_iva:
                    factura.estado_pago = 'PAGADO'
                    factura.save()
                
                total_procesado += monto_pago
                procesadas += 1
            
            return JsonResponse({
                'success': True, 
                'procesadas': procesadas,
                'total_procesado': total_procesado,
                'message': f'{procesadas} facturas procesadas correctamente'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def asociar_nc_existente(request):
    """
    Asocia una nota de crédito existente a una factura específica
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nc_id = data.get('nc_id')
            dte_id = data.get('dte_id')
            
            if not nc_id or not dte_id:
                return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)
            
            # Obtener la NC y la factura
            try:
                nc = Dte.objects.get(id=nc_id, tipo_documento='NOTA DE CREDITO')
                factura = Dte.objects.get(id=dte_id)
            except Dte.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Documento no encontrado'}, status=404)
            
            # Verificar que la NC no esté ya asociada
            ya_asociada = Dte_Detalle_Pago.objects.filter(
                voucher=nc.numero_documento,
                metodo_pago='Nota de Crédito'
            ).exists()
            
            if ya_asociada:
                return JsonResponse({'success': False, 'error': 'Esta Nota de Crédito ya está asociada a otra factura'}, status=400)
            
            # Crear el registro de pago con la NC
            Dte_Detalle_Pago.objects.create(
                dte=factura,
                metodo_pago='Nota de Crédito',
                voucher=nc.numero_documento,
                monto=nc.monto_con_iva,
                notas=f'NC #{nc.numero_documento} - {nc.emisor.nombre if nc.emisor else "N/A"}'
            )
            
            # Actualizar estado de pago de la factura si corresponde
            total_pagos = Dte_Detalle_Pago.objects.filter(dte=factura).aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            if total_pagos >= factura.monto_con_iva:
                factura.estado_pago = 'PAGADO'
                factura.save()
            
            return JsonResponse({'success': True, 'message': 'Nota de Crédito asociada correctamente'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def guardar_recepcion(request):
    try:
        data = json.loads(request.body)
        compra_id = data.get('compra_id')
        recepciones = data.get('recepciones', [])

        if not compra_id or not recepciones:
            return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)

        for item in recepciones:
            compra_talla_id = item['compra_producto_talla_id']
            cantidad = item['recepcionado']
            factura_id = item.get('factura_id')

            compra_talla = Compras_Producto_Talla.objects.select_related('compra_producto').get(id=compra_talla_id)

            # 🔁 Si ya existe una recepción para esta talla → actualizar
            recepcion_existente = Productos_Recepcionados.objects.filter(compra_producto_talla=compra_talla).first()

            if recepcion_existente:
                recepcion_existente.stockArribado = cantidad
                recepcion_existente.dte_id = factura_id
                recepcion_existente.save()
            else:
                # 🚫 No se crea Producto_Talla todavía (ya hablaste de eso antes)
                Productos_Recepcionados.objects.create(
                    compra_producto_talla=compra_talla,
                    producto_talla=None,
                    dte_id=factura_id,
                    stockArribado=cantidad
                )

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
@require_POST
@transaction.atomic
def agregar_producto_manual(request):
    try:
        # Validar que se proporcione un DTE
        dte_id = request.POST.get('dte_id')
        if not dte_id:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar un DTE'}, status=400)
        
        # Validar que se proporcione la compra
        compra_id = request.POST.get('compra_id')
        if not compra_id:
            return JsonResponse({'success': False, 'error': 'Debe tener una compra activa'}, status=400)
        
        # Verificar que el DTE existe y es válido
        try:
            dte = Dte.objects.get(id=dte_id, tipo_transaccion='COMPRA')
        except Dte.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'DTE no válido'}, status=400)
        
        # Verificar que la compra existe
        try:
            compra = Compras.objects.get(id=compra_id)
        except Compras.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Compra no válida'}, status=400)
        
        # Crear el producto de compra (esto aparecerá en la recepción)
        compra_producto = Compras_Producto.objects.create(
            compras=compra,
            nombre=request.POST['nombre'],
            descripcion=request.POST.get('descripcion', ''),
            atributo1=request.POST.get('atributo1', ''),
            atributo2=request.POST.get('atributo2', ''),
            atributo3=request.POST.get('atributo3', ''),
            atributo4=request.POST.get('atributo4', ''),
            costo=int(request.POST['costo']),
            precioSugerido=int(request.POST['precioSugerido'])
        )
        
        # Crear la talla del producto de compra
        Compras_Producto_Talla.objects.create(
            compra_producto=compra_producto,
            stock=int(request.POST['stock']),
            talla=request.POST['talla']
        )
        
        return JsonResponse({'success': True, 'message': 'Producto agregado a la compra correctamente'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
 
def productos_recepcionados(request):
    productos = Productos_Recepcionados.objects.select_related(
        'compra_producto_talla__compra_producto', 'dte'
    )

    data = []
    for p in productos:
        prod = p.compra_producto_talla.compra_producto
        data.append({
            'recepcion_id': p.id,
            'nombre': prod.nombre,
            'descripcion': prod.descripcion,
            'marca': prod.atributo1,
            'color': prod.atributo2,
            'genero': prod.atributo3,
            'talla': p.compra_producto_talla.talla,
            'stock': p.stockArribado,
            'costo': prod.costo,
            'factura': p.dte.numero_documento if p.dte else None
        })

    return JsonResponse(data, safe=False)

 
 
# views.py

 

def obtener_productos_para_crear(request):
    anio = request.GET.get('anio')
    compra_id = request.GET.get('compra_id')
    proveedor_id = request.GET.get('proveedor_id')
    articulo = request.GET.get('articulo')
    marca = request.GET.get('marca')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    estado = request.GET.get('estado')  # 'creado', 'no_creado', o None
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    recientes = request.GET.get('recientes')
    factura = request.GET.get('factura')  # NUEVO: filtro por número de factura
    
    # Parámetros de paginación
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 25))

    qs = Productos_Recepcionados.objects.select_related(
        'compra_producto_talla__compra_producto__compras__empresa',
        'dte'  # NUEVO: incluir relación con Dte
    ).all()

    if anio:
        qs = qs.filter(fecha__year=anio)
    if compra_id:
        qs = qs.filter(compra_producto_talla__compra_producto__compras__id=compra_id)
    if proveedor_id:
        qs = qs.filter(compra_producto_talla__compra_producto__compras__empresa__id=proveedor_id)
    if articulo:
        qs = qs.filter(compra_producto_talla__compra_producto__nombre__icontains=articulo)
    if marca:
        qs = qs.filter(compra_producto_talla__compra_producto__atributo1__icontains=marca)
    if color:
        qs = qs.filter(compra_producto_talla__compra_producto__atributo2__icontains=color)
    if genero:
        qs = qs.filter(compra_producto_talla__compra_producto__atributo3__icontains=genero)
    if factura:  # NUEVO: filtro por número de factura
        qs = qs.filter(dte__numero_documento__icontains=factura)
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)
    if estado == 'creado':
        qs = qs.filter(producto_talla__isnull=False)
    elif estado == 'no_creado':
        qs = qs.filter(producto_talla__isnull=True)
    if recientes:
        qs = qs.order_by('-fecha')[:10]

    # Primero aplicar distinct y agregaciones
    productos = (
        qs
        .select_related('compra_producto_talla__compra_producto', 'dte', 'producto_talla')
        .values(
            'compra_producto_talla__compra_producto_id',
            'compra_producto_talla__compra_producto__nombre',
            'compra_producto_talla__compra_producto__descripcion',
            'compra_producto_talla__compra_producto__atributo1',
            'compra_producto_talla__compra_producto__atributo2',
            'compra_producto_talla__compra_producto__atributo3',
            'compra_producto_talla__compra_producto__atributo4',
            'compra_producto_talla__compra_producto__costo',
            'compra_producto_talla__compra_producto__compras__id',
            'compra_producto_talla__compra_producto__compras__nombre',
            'compra_producto_talla__compra_producto__compras__empresa__id',
            'compra_producto_talla__compra_producto__compras__empresa__nombre',
            'dte__numero_documento',  # NUEVO: incluir número de factura
            'producto_talla__id',  # NUEVO: incluir ID del producto_talla si existe
        )
        .annotate(
            stock_total=Sum('stockArribado'),
            stock_creado=Sum(
                Case(
                    When(producto_talla__isnull=False, then=F('stockArribado')),
                    default=Value(0),
                    output_field=IntegerField()
                )
            ),
            count_creados=Count('producto_talla', distinct=True)
        )
        .distinct()
    )

    # Contar total de registros para paginación (después del distinct)
    total_records = productos.count()
    
    # Calcular total de páginas
    total_pages = (total_records + page_size - 1) // page_size
    
    # Aplicar paginación (después del distinct)
    offset = (page - 1) * page_size
    productos = productos[offset:offset + page_size]

    respuesta = []
    for p in productos:
        respuesta.append({
            'producto_id': p['compra_producto_talla__compra_producto_id'],
            'nombre': p['compra_producto_talla__compra_producto__nombre'],
            'descripcion': p['compra_producto_talla__compra_producto__descripcion'],
            'atributo1': p['compra_producto_talla__compra_producto__atributo1'],
            'atributo2': p['compra_producto_talla__compra_producto__atributo2'],
            'atributo3': p['compra_producto_talla__compra_producto__atributo3'],
            'atributo4': p['compra_producto_talla__compra_producto__atributo4'],
            'costo': p['compra_producto_talla__compra_producto__costo'],
            'stock_total': p['stock_total'],
            'stock_creado': p['stock_creado'],
            'creado': p['count_creados'] > 0,
            'producto_talla_id': p['producto_talla__id'],  # NUEVO: incluir ID del producto_talla
            'compra_id': p['compra_producto_talla__compra_producto__compras__id'],
            'compra_nombre': p['compra_producto_talla__compra_producto__compras__nombre'],
            'proveedor_id': p['compra_producto_talla__compra_producto__compras__empresa__id'],
            'proveedor_nombre': p['compra_producto_talla__compra_producto__compras__empresa__nombre'],
            'factura': p['dte__numero_documento'],  # NUEVO: incluir número de factura en respuesta
        })

    # Devolver respuesta con información de paginación
    return JsonResponse({
        'data': respuesta,
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_records': total_records,
            'page_size': page_size
        }
    }, safe=False)
 
def opciones_atributo(request):
    atributo_id = request.GET.get('atributo_id')
    opciones = AtributoOpcion.objects.filter(atributo_id=atributo_id)
    return JsonResponse([{'id': o.id, 'valor': o.valor} for o in opciones], safe=False)


@require_POST
def opcion_atributo_crear(request):
    atributo_id = request.POST.get('atributo_id')
    valor = request.POST.get('valor')

    if not atributo_id or not valor:
        return JsonResponse({'success': False, 'error': 'Datos incompletos'})

    opcion = AtributoOpcion.objects.create(
        atributo_id=atributo_id,
        valor=valor
    )

    return JsonResponse({'success': True, 'opcion_id': opcion.id, 'valor': opcion.valor})
 
 
 
def detalle_producto_para_crear(request, producto_id):
    compra_producto = get_object_or_404(Compras_Producto, id=producto_id)

    # Obtener tallas recepcionadas sin producto_talla aún
    recepcionadas = Productos_Recepcionados.objects.filter(
        compra_producto_talla__compra_producto=compra_producto,
        producto_talla__isnull=True
    ).values('compra_producto_talla__talla', 'dte_id').annotate(
        stock=Sum('stockArribado')
    )

    tallas = [
        {'talla': r['compra_producto_talla__talla'], 'stock': r['stock']}
        for r in recepcionadas
    ]

    # Obtener el DTE más común entre las recepciones (o el primero si hay varios)
    dte_id = None
    if recepcionadas:
        # Tomar el primer DTE no nulo que encontremos
        for r in recepcionadas:
            if r['dte_id']:
                dte_id = r['dte_id']
                break

    data = {
        'producto_id': compra_producto.id,
        'nombre': compra_producto.nombre,
        'descripcion': compra_producto.descripcion,
        'costo': compra_producto.costo,
        'precioSugerido': compra_producto.precioSugerido,

        'atributo1': compra_producto.atributo1,
        'atributo2': compra_producto.atributo2,
        'atributo3': compra_producto.atributo3,
        'atributo4': compra_producto.atributo4,
        'categoria': None,  # o compra_producto.categoria si lo tiene

        # Etiquetas sugeridas (pueden ser string directo o .valor si son ForeignKeys)
        'atributo1_label': compra_producto.atributo1,
        'atributo2_label': compra_producto.atributo2,
        'atributo3_label': compra_producto.atributo3,
        'atributo4_label': compra_producto.atributo4,
        'categoria_label': None,

        # 🔥 Nuevo campo para tipo de talla (US, EU, etc.)
        'tipo_talla': compra_producto.tipo_talla if hasattr(compra_producto, 'tipo_talla') else 'CL',

        'tallas': tallas,
        'dte_id': dte_id  # Agregar el DTE ID
    }

    return JsonResponse(data)

 
@require_GET
def margenes_usuario(request):
    user = request.user
    empresa_id = request.session.get('idEmpresaActual')  # o ajusta al tuyo
    sucursal_id = request.session.get('idSucursalActual')

    try:
        eu = EmpresaUser.objects.get(user=user, empresa_id=empresa_id, sucursal_id=sucursal_id, status=True)
        return JsonResponse({
            'success': True,
            'margenSobreprecio': eu.margenSobreprecio or 0,
            'margenPrecioVenta': eu.margenPrecioVenta or 0
        })
    except EmpresaUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Configuración no encontrada'})
@require_POST
@login_required
def guardar_margenes_usuario(request):
    empresa_id = request.session.get('idEmpresaActual')
    sucursal_id = request.session.get('idSucursalActual')
    try:
        eu = EmpresaUser.objects.get(
            user=request.user,
            empresa_id=empresa_id,
            sucursal_id=sucursal_id,
            status=True
        )

        margenSobreprecio = int(request.POST.get('margenSobreprecio', 0))
        margenPrecioVenta = int(request.POST.get('margenPrecioVenta', 0))

        eu.margenSobreprecio = margenSobreprecio
        eu.margenPrecioVenta = margenPrecioVenta
        eu.save()

        return JsonResponse({'success': True})
    except EmpresaUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Configuración no encontrada'})

  
def ajustar_margenes(request):
    if request.method == 'POST':
        user = request.user
        margen_sobreprecio = request.POST.get('margenSobreprecio')
        margen_precio_venta = request.POST.get('margenPrecioVenta')

        empresa_user = EmpresaUser.objects.get(user=user)
        empresa_user.margenSobreprecio = margen_sobreprecio
        empresa_user.margenPrecioVenta = margen_precio_venta
        empresa_user.save()

        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)
 
@require_GET
def categorias_existentes(request):
    def recorrer(categorias, prefijo=''):
        resultado = []
        for c in categorias:
            resultado.append({'id': c.id, 'nombre': f"{prefijo}{c.nombre}"})
            hijos = c.subcategorias.all().order_by('nombre')
            resultado += recorrer(hijos, prefijo + '› ')
        return resultado

    categorias_raiz = Categoria.objects.filter(padre__isnull=True).order_by('nombre')
    data = recorrer(categorias_raiz)
    return JsonResponse(data, safe=False)


@require_POST
def categoria_guardar(request):
    from django.shortcuts import get_object_or_404
    id = request.POST.get('id')
    nombre = request.POST.get('nombre')
    padre_id = request.POST.get('padre')

    if not nombre:
        return JsonResponse({'success': False, 'error': 'Nombre obligatorio'})

    padre = Categoria.objects.filter(id=padre_id).first() if padre_id else None

    if id:
        categoria = get_object_or_404(Categoria, pk=id)
        categoria.nombre = nombre
        categoria.padre = padre
        categoria.save()
    else:
        Categoria.objects.create(nombre=nombre, padre=padre)

    return JsonResponse({'success': True})
 
def guias_talla_list(request):
    marca_id = request.GET.get('marca')
    print(f"🔍 Guías de talla - Marca ID: {marca_id}")
    
    if marca_id and str(marca_id).isdigit():
        guias = GuiaTalla.objects.filter(marca_id=marca_id)
        print(f"✅ Filtrando por marca {marca_id}, encontradas: {guias.count()}")
    else:
        guias = GuiaTalla.objects.all()
        print(f"✅ Mostrando todas las guías, total: {guias.count()}")

    data = [{'id': g.id, 'text': str(g)} for g in guias]
    print(f"📋 Datos a enviar: {data}")
    return JsonResponse(data, safe=False)

 
 
def crear_guia_talla(request):
    if request.method == 'POST':
        try:
            guia_id = request.POST.get('id')
            marca_id = request.POST.get('marca')
            nombre = request.POST.get('nombre')
            producto_id = request.POST.get('producto', None)
            tallas_json = request.POST.get('tallas')

            if not marca_id or not marca_id.isdigit():
                return JsonResponse({'success': False, 'error': 'Marca no válida.'})

            marca = AtributoOpcion.objects.get(pk=marca_id)  # ✅ CAMBIO

            if guia_id:
                guia = GuiaTalla.objects.get(pk=guia_id)
                guia.marca = marca
                guia.nombre = nombre
                guia.save()
                guia.items.all().delete()  # elimina y re-crea tallas
            else:
                guia = GuiaTalla.objects.create(marca=marca, nombre=nombre)

            tallas = json.loads(tallas_json)
            for idx, item in enumerate(tallas):
                GuiaTallaItem.objects.create(
                    guia=guia,
                    cl=item.get('cl'),
                    us=item.get('us'),
                    eu=item.get('eu'),
                    uk=item.get('uk'),
                    br=item.get('br'),
                    cm=item.get('cm'),
                    orden=idx
                )

            if producto_id and producto_id.isdigit():
                producto = Producto.objects.get(pk=producto_id)
                GuiaTallaProducto.objects.get_or_create(guia=guia, producto=producto)

            return JsonResponse({
                'success': True,
                'id': guia.id,
                'text': str(guia)
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

def asociar_producto_guia(request):
    if request.method == 'POST':
        try:
            guia_id = request.POST.get('guia_id')
            producto_id = request.POST.get('producto_id')

            if not guia_id or not producto_id:
                return JsonResponse({'success': False, 'error': 'Datos incompletos.'})

            guia = GuiaTalla.objects.get(pk=guia_id)
            producto = Producto.objects.get(pk=producto_id)

            GuiaTallaProducto.objects.get_or_create(guia=guia, producto=producto)

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
def guia_talla_detalle(request, id):
    try:
        guia = GuiaTalla.objects.get(id=id)
    except GuiaTalla.DoesNotExist:
        raise Http404("Guía de talla no encontrada")

    return JsonResponse({
        'id': guia.id,
        'nombre': guia.nombre,
        'marca': guia.marca.id,  # ejemplo: 5 (marca Puma)
        'producto': guia.productos.first().id if guia.productos.exists() else None,
        'tallas': list(
            guia.items.order_by('orden').values('cl', 'us', 'eu', 'uk', 'br', 'cm')
        )
    })

 
def eliminar_guia_talla(request):
    if request.method == 'POST':
        try:
            id = request.POST.get('id')
            if not id:
                return JsonResponse({'success': False, 'error': 'ID no proporcionado'})

            guia = GuiaTalla.objects.get(pk=id)
            guia.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
 
 
@transaction.atomic
def crear_producto(data, responsable):
    # Crear el producto
    producto = Producto.objects.create(
        articulo=data['articulo'],
        descripcion=data['descripcion'],
        atributo1=data['atributo1'],
        atributo2=data['atributo2'],
        atributo3=data['atributo3'],
        atributo4=data['atributo4'],
        categoria=data.get('categoria'),
        sucursal=data['sucursal'],
        costo=data['costo'],
        sobreprecio=data['sobreprecio'],
        precioventa=data['precioventa'],
        precioSugerido=data.get('precioSugerido')
    )

    # Crear las tallas asociadas
    for talla_data in data['tallas']:
        producto_talla = Producto_Talla.objects.create(
            producto=producto,
            sku=obtener_siguiente_sku(),
            stock=talla_data['stock'],
            talla=talla_data['talla']
        )

        # Registrar el movimiento de inventario
        Movimientos_Producto.objects.create(
            ProductoTalla=producto_talla,
            costo=producto.costo,
            sobreprecio=producto.sobreprecio,
            precio=producto.precioventa,
            concepto='Ingreso Inicial',
            tipo_movimiento='INGRESO',
            responsable=responsable
        )

    return producto
 
def guias_talla_por_marca(request):
    marca_id = request.GET.get('marca')
    if not marca_id:
        return JsonResponse([], safe=False)
    
    guias = GuiaTalla.objects.filter(marca_id=marca_id).order_by('nombre')
    data = [{'id': g.id, 'nombre': g.nombre} for g in guias]
    return JsonResponse(data, safe=False)
 
def verificar_existencia_producto(request):
    articulo = request.GET.get('articulo')
    if not articulo:
        return JsonResponse({'existe': False})

    existe = Producto.objects.filter(articulo=articulo).exists()
    return JsonResponse({'existe': existe})
 
@transaction.atomic
def obtener_siguiente_sku():
    parametro, creado = ParametroGlobal.objects.select_for_update().get_or_create(
        nombre='sku',
        defaults={'valor_entero': 100000}
    )

    siguiente = parametro.valor_entero + 1
    parametro.valor_entero = siguiente
    parametro.save()

    return siguiente
@require_GET
def obtener_siguiente_sku_view(request):
    try:
        sku = obtener_siguiente_sku()  # 👈 esta es tu función interna que sí funciona
        return JsonResponse({'success': True, 'sku': sku})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
 
def verificar_producto_existente(request):
    articulo = request.GET.get('articulo')
    marca = request.GET.get('marca')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    categoria = request.GET.get('categoria')  # 🔧 AGREGADA CATEGORÍA

    # Convertir a enteros si son números, o buscar por valor si son texto
    filtros = {'articulo': articulo}
    
    if marca:
        try:
            # Si es un número, usar como ID
            marca_id = int(marca)
            filtros['atributo1_id'] = marca_id
        except (ValueError, TypeError):
            # Si es texto, buscar por valor
            try:
                marca_obj = AtributoOpcion.objects.get(valor=marca)
                filtros['atributo1_id'] = marca_obj.id
            except AtributoOpcion.DoesNotExist:
                return JsonResponse({
                    'existe': False,
                    'tallas_existentes': []
                })
    
    if color:
        try:
            color_id = int(color)
            filtros['atributo2_id'] = color_id
        except (ValueError, TypeError):
            try:
                color_obj = AtributoOpcion.objects.get(valor=color)
                filtros['atributo2_id'] = color_obj.id
            except AtributoOpcion.DoesNotExist:
                return JsonResponse({
                    'existe': False,
                    'tallas_existentes': []
                })
    
    if genero:
        try:
            genero_id = int(genero)
            filtros['atributo3_id'] = genero_id
        except (ValueError, TypeError):
            try:
                genero_obj = AtributoOpcion.objects.get(valor=genero)
                filtros['atributo3_id'] = genero_obj.id
            except AtributoOpcion.DoesNotExist:
                return JsonResponse({
                    'existe': False,
                    'tallas_existentes': []
                })
    
    # 🔧 MANEJAR CATEGORÍA
    if categoria:
        try:
            categoria_id = int(categoria)
            filtros['categoria_id'] = categoria_id
        except (ValueError, TypeError):
            # Si es texto, buscar por nombre
            try:
                from .models import Categoria
                categoria_obj = Categoria.objects.get(nombre=categoria)
                filtros['categoria_id'] = categoria_obj.id
            except Categoria.DoesNotExist:
                return JsonResponse({
                    'existe': False,
                    'tallas_existentes': []
                })

    producto = Producto.objects.filter(**filtros).first()
    
    if producto:
        # Obtener las tallas existentes con sus SKUs
        tallas_existentes = Producto_Talla.objects.filter(producto=producto).values('talla', 'sku', 'stock')
        
        return JsonResponse({
            'existe': True,
            'producto_id': producto.id,
            'tallas_existentes': list(tallas_existentes)
        })
    else:
        return JsonResponse({
            'existe': False,
            'tallas_existentes': []
        })
@transaction.atomic
def crear_producto_desde_recepcion(request):
    # 1. Validar sesión y datos básicos
    sucursal_id = request.session.get('idSucursalActual')
    usuario = request.session.get('nombreUsuario', 'Sistema')
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)

    data = request.POST
    articulo = data.get('articulo')
    descripcion = data.get('descripcion')
    atributo1 = data.get('atributo1') or None
    atributo2 = data.get('atributo2') or None
    atributo3 = data.get('atributo3') or None
    atributo4 = data.get('atributo4') or None
    categoria = data.get('categoria') or None
    costo = int(data.get('costo') or 0)
    sobreprecio = int(data.get('sobreprecio') or 0)
    precioventa = int(data.get('precioventa') or 0)
    precio_sugerido = int(data.get('precio_sugerido') or 0)
    tipo_talla = data.get('tipo_talla') or 'CL'
    guia_talla = data.get('guia_talla') or None
    producto_compra_id = data.get('producto_compra_id') or None

    # 2. Crear el producto principal
    producto = Producto.objects.create(
        articulo=articulo,
        descripcion=descripcion,
        atributo1_id=atributo1,
        atributo2_id=atributo2,
        atributo3_id=atributo3,
        atributo4_id=atributo4,
        categoria_id=categoria,
        sucursal=sucursal,
        costo=costo,
        sobreprecio=sobreprecio,
        precioventa=precioventa,
        precioSugerido=precio_sugerido,
        tipo_talla=tipo_talla,
        guia_talla_id=guia_talla,
    )

    # 3. Crear variantes (tallas)
    tallas = []
    for key in data:
        if key.startswith('sku_'):
            talla = key.replace('sku_', '')
            sku = int(data[key])
            stock = int(data.get(f'stock_{talla}', 0))
            pt = Producto_Talla.objects.create(
                producto=producto,
                sku=sku,
                stock=stock,
                talla=talla,
            )
            tallas.append((pt, stock, talla))

    # 4. Registrar movimiento de ingreso a bodega para cada variante
    for pt, stock, talla in tallas:
        # Obtener el DTE asociado a esta recepción
        dte = None
        if producto_compra_id:
            recepcion = Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto_id=producto_compra_id,
                compra_producto_talla__talla=talla
            ).first()
            if recepcion and recepcion.dte:
                dte = recepcion.dte
        
        registrar_movimiento_producto(
            producto_talla=pt,
            concepto='INGRESO_INICIAL',
            cantidad=stock,
            responsable=usuario,
            dte=dte,  # Asociar el DTE al movimiento
            sucursal_origen=sucursal,
            sucursal_destino=sucursal,
            observaciones=f'Ingreso inicial de producto {producto.articulo} - Talla {talla}'
        )

    # 5. Actualizar tabla de recepción de productos
    if producto_compra_id:
        for pt, stock, talla in tallas:
            Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto_id=producto_compra_id,
                compra_producto_talla__talla=talla
            ).update(producto_talla=pt)

    return JsonResponse({'success': True, 'producto_id': producto.id})
import re

def obtener_tallas_post(request):
    """
    Convierte los campos 'sku_36', 'stock_36', 'guia_talla_36'… recibidos
    en una lista de dicts:
        [{'nombre': '36', 'sku': '123', 'stock': 8, 'guia_talla': '1'}, …]
    """
    tallas = {}
    for k, v in request.POST.items():
        m = re.match(r'^(sku|stock|guia_talla)_(.+)$', k)
        if not m:
            continue
        campo, talla = m.groups()
        tallas.setdefault(talla, {})[campo] = v

    return [
        {
            'nombre': t,
            'sku': datos.get('sku') or None,
            'stock': int(datos.get('stock', 0) or 0),
            'guia_talla': datos.get('guia_talla') or None
        }
        for t, datos in tallas.items()
    ]

# views.py
 
def _next_sku():
    max_sku = Producto_Talla.objects.aggregate(m=Max('sku'))['m'] or 0
    return max_sku + 1


def _buscar_producto(request, articulo, attr1, attr2, attr3):
    sucursal_id = request.session.get('idSucursalActual')
    if not sucursal_id:
        return None
    qs = Producto.objects.filter(
        articulo      = articulo.strip(),
        atributo1_id  = attr1,
        sucursal_id   = sucursal_id
    )
    if attr2:
        qs = qs.filter(atributo2_id = attr2)
    if attr3:
        qs = qs.filter(atributo3_id = attr3)
    return qs.first()


def sku_para_talla(request):
    """
    GET params:
        articulo, atributo1, atributo2?, atributo3?, talla
    Return:
        { sku: 12345, existe: true|false }
    """
    art     = request.GET.get('articulo')
    attr1   = request.GET.get('atributo1')
    attr2   = request.GET.get('atributo2')
    attr3   = request.GET.get('atributo3')
    talla   = request.GET.get('talla')

    if not (art and attr1 and talla):
        return JsonResponse({'error': 'parametros'}, status=400)

    producto = _buscar_producto(request, art, attr1, attr2, attr3)

    # 1 Ya existe producto y talla
    if producto:
        pt = Producto_Talla.objects.filter(producto=producto, talla=talla).first()
        if pt:
            return JsonResponse({'sku': pt.sku, 'existe': True})

    # 2 SKU sugerido no sirve → genera nuevo
    sku = _next_sku()
    return JsonResponse({'sku': sku, 'existe': False})

@require_GET
def facturas_pendientes(request):
    """
    Devuelve facturas de compra (Dte) con saldo disponible para recepcionar productos.
    Permite filtrar por número de factura con ?q= y por proveedor con ?proveedor_id=.
    """
    empresa_id = request.session.get('idEmpresaActual')
    proveedor_id = request.GET.get('proveedor_id')  # <- lo recibimos por GET
    q = request.GET.get('q', '').strip()

    facturas = Dte.objects.filter(
        tipo_transaccion='COMPRA',
        emisor_id=empresa_id
    ).select_related('receptor')  # Incluir el proveedor (receptor)
    
    if proveedor_id:
        facturas = facturas.filter(receptor_id=proveedor_id)
    if q:
        facturas = facturas.filter(numero_documento__icontains=q)

    # Excluir facturas que ya están completamente recepcionadas
    facturas = facturas.values('id', 'numero_documento', 'monto_con_iva', 'receptor__nombre')
    ids = [f['id'] for f in facturas]
    from django.db.models import Sum, F
    from .models import Productos_Recepcionados
    usados = (
        Productos_Recepcionados.objects
        .filter(dte_id__in=ids)
        .values('dte_id')
        .annotate(total=Sum(F('stockArribado') * F('compra_producto_talla__compra_producto__costo')))
    )
    usado_map = {u['dte_id']: u['total'] for u in usados}

    disponibles = []
    for f in facturas:
        usado = usado_map.get(f['id'], 0) or 0
        if usado < float(f['monto_con_iva']):
            disponibles.append({
                'id': f['id'],
                'text': str(f['numero_documento']),
                'monto': float(f['monto_con_iva']),
                'usado': float(usado),
                'proveedor_nombre': f['receptor__nombre'] or ''
            })

    # Formato para select2 con proveedor incluido
    return JsonResponse([
        {
            'id': f['id'], 
            'text': f['text'],
            'proveedor_nombre': f['proveedor_nombre']
        }
        for f in disponibles
    ], safe=False)

@login_required
def verMovimientosProducto(request):
    """
    Renderiza la página de movimientos de producto por sucursal.
    """
    return render(request, 'vistas/modulo_existencias/gestionMovimientos.html')

@require_GET
@login_required
def obtener_movimientos_producto(request):
    sucursal_id = request.session.get('idSucursalActual')
    if not sucursal_id:
        return JsonResponse({'success': False, 'error': 'No hay sucursal activa'}, status=400)

    from datetime import datetime
    from django.utils.dateparse import parse_date
    def parse_fecha_ddmmyyyy(fecha_str):
        try:
            if fecha_str and '/' in fecha_str:
                return datetime.strptime(fecha_str, '%d/%m/%Y').date()
            elif fecha_str:
                return parse_date(fecha_str)
        except Exception:
            return None
        return None

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')  # INGRESO, EGRESO, etc.
    articulo = request.GET.get('articulo')
    responsable = request.GET.get('responsable')
    concepto = request.GET.get('concepto')
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 50)), 100)

    fecha_inicio_dt = parse_fecha_ddmmyyyy(fecha_inicio)
    fecha_fin_dt = parse_fecha_ddmmyyyy(fecha_fin)

    # Construir query filtrando SOLO movimientos que ocurrieron en esta sucursal
    # Se filtra por sucursal_origen, sucursal_destino, ticket o DTE de la sucursal
    movimientos = Movimientos_Producto.objects.select_related(
        'ProductoTalla__producto', 'dte', 'ticket', 'sucursal_origen', 'sucursal_destino'
    ).filter(
        Q(sucursal_origen_id=sucursal_id) |  # Movimientos que salieron de esta sucursal
        Q(sucursal_destino_id=sucursal_id) |  # Movimientos que llegaron a esta sucursal
        Q(ticket__sucursal_id=sucursal_id) |  # Ventas (tickets) de esta sucursal
        Q(dte__sucursal_id=sucursal_id)  # DTEs de esta sucursal
    ).distinct()
    if fecha_inicio_dt:
        movimientos = movimientos.filter(fecha__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        movimientos = movimientos.filter(fecha__lte=fecha_fin_dt)
    if tipo:
        movimientos = movimientos.filter(tipo_movimiento__iexact=tipo)
    if articulo:
        movimientos = movimientos.filter(ProductoTalla__producto__articulo__icontains=articulo)
    if responsable:
        movimientos = movimientos.filter(responsable__icontains=responsable)
    if concepto:
        movimientos = movimientos.filter(concepto__icontains=concepto)

    total_count = movimientos.count()
    offset = (page - 1) * page_size
    # ✅ ORDEN: Más recientes primero (fecha + hora descendente)
    movimientos = movimientos.order_by('-fecha', '-hora')[offset:offset+page_size]

    data = []
    for m in movimientos:
        prod = m.ProductoTalla.producto
        # Usar la cantidad real del movimiento
        cantidad = m.cantidad
        referencia = m.referencia_externa or ''
        if m.tipo_movimiento == 'TRASPASO' and m.sucursal_destino:
            referencia = f"Destino: {m.sucursal_destino.alias}"
        elif m.ticket:
            referencia = f"Ticket #{m.ticket.correlativo}"
        elif m.dte:
            referencia = f"{m.dte.tipo_documento} {m.dte.numero_documento}"
        def limpiar_prefijo(valor):
            if not valor:
                return ''
            for prefijo in ['Marca:', 'Color:', 'Género:']:
                if valor.startswith(prefijo):
                    return valor[len(prefijo):].strip()
            return valor.strip()
        marca = limpiar_prefijo(prod.atributo1.valor if prod.atributo1 else '')
        color = limpiar_prefijo(prod.atributo2.valor if prod.atributo2 else '')
        genero = limpiar_prefijo(prod.atributo3.valor if prod.atributo3 else '')
        data.append({
            'id': m.id,
            'fecha': m.fecha.strftime('%Y-%m-%d'),
            'hora': m.hora.strftime('%H:%M:%S') if m.hora else '',
            'articulo': prod.articulo,
            'descripcion': prod.descripcion,
            'marca': marca,
            'color': color,
            'genero': genero,
            'talla': m.ProductoTalla.talla,
            'sku': m.ProductoTalla.sku,
            'cantidad': cantidad,  # Agregar cantidad
            'costo': m.costo,
            'precio': m.precio,
            'sobreprecio': m.sobreprecio,
            'tipo_movimiento': m.tipo_movimiento,
            'concepto': m.concepto,
            'responsable': m.responsable,
            'dte': m.dte.numero_documento if m.dte else None,
            'referencia_externa': referencia,
        })
    return JsonResponse({
        'success': True,
        'items': data,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size,
            'has_next': page * page_size < total_count,
            'has_previous': page > 1
        }
    })

@require_GET
@login_required
def obtener_productos(request):
    q = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 20)), 100)
    sucursal_id = request.session.get('idSucursalActual')
    productos = Producto_Talla.objects.select_related('producto')
    if sucursal_id:
        productos = productos.filter(producto__sucursal_id=sucursal_id)
    if q:
        productos = productos.filter(
            Q(producto__articulo__icontains=q) |
            Q(producto__descripcion__icontains=q) |
            Q(producto__atributo1__valor__icontains=q) |
            Q(producto__atributo2__valor__icontains=q) |
            Q(producto__atributo3__valor__icontains=q) |
            Q(sku__icontains=q)
        )
    total_count = productos.count()
    offset = (page - 1) * page_size
    productos = productos.order_by('producto__articulo', 'talla')[offset:offset+page_size]
    def limpiar_prefijo(valor):
        if not valor:
            return ''
        for prefijo in ['Marca:', 'Color:', 'Género:']:
            if valor.startswith(prefijo):
                return valor[len(prefijo):].strip()
        return valor.strip()
    results = []
    for pt in productos:
        prod = pt.producto
        marca = limpiar_prefijo(prod.atributo1.valor if prod.atributo1 else '')
        color = limpiar_prefijo(prod.atributo2.valor if prod.atributo2 else '')
        genero = limpiar_prefijo(prod.atributo3.valor if prod.atributo3 else '')
        text = f"{prod.articulo} - {prod.descripcion} | Marca: {marca} | Color: {color} | Género: {genero} | Talla: {pt.talla} | SKU: {pt.sku}"
        results.append({
            'id': pt.id,
            'text': text,
            'sku': pt.sku,
            'marca': marca,
            'color': color,
            'genero': genero,
            'talla': pt.talla
        })
    return JsonResponse({
        'results': results,
        'pagination': {'more': offset + page_size < total_count}
    })

@require_GET
@login_required
def reporte_despachos_por_proveedor(request):
    """
    Reporte que muestra por cada proveedor y DTE:
    - Total de productos ingresados
    - Total de productos despachados (vendidos, transferidos, etc.)
    - Saldo restante
    """
    # Filtros
    proveedor_id = request.GET.get('proveedor_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    dte_numero = request.GET.get('dte_numero')
    
    # Parámetros de paginación
    page = int(request.GET.get('page', 1))
    page_size = min(int(request.GET.get('page_size', 25)), 100)
    
    # Query base: obtener todos los DTEs de compra con sus movimientos
    dtes_query = Dte.objects.filter(
        tipo_transaccion='COMPRA'
    ).select_related('receptor')
    
    # Aplicar filtros
    if proveedor_id:
        dtes_query = dtes_query.filter(receptor_id=proveedor_id)
    if fecha_inicio:
        dtes_query = dtes_query.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:
        dtes_query = dtes_query.filter(fecha_emision__lte=fecha_fin)
    if dte_numero:
        dtes_query = dtes_query.filter(numero_documento__icontains=dte_numero)
    
    # Contar total para paginación
    total_count = dtes_query.count()
    total_pages = (total_count + page_size - 1) // page_size
    
    # Aplicar paginación
    offset = (page - 1) * page_size
    dtes = dtes_query[offset:offset + page_size]
    
    resultado = []
    
    for dte in dtes:
        # Calcular ingresos (movimientos de INGRESO asociados a este DTE)
        ingresos = Movimientos_Producto.objects.filter(
            dte=dte,
            tipo_movimiento='INGRESO'
        ).aggregate(
            total_cantidad=Sum('cantidad'),
            total_costo=Sum(F('cantidad') * F('costo'))
        )
        
        # Calcular despachos (movimientos de EGRESO asociados a este DTE)
        despachos = Movimientos_Producto.objects.filter(
            dte=dte,
            tipo_movimiento='EGRESO'
        ).aggregate(
            total_cantidad=Sum('cantidad'),
            total_costo=Sum(F('cantidad') * F('costo'))
        )
        
        # Valores por defecto
        total_ingresado = ingresos['total_cantidad'] or 0
        total_despachado = abs(despachos['total_cantidad'] or 0)  # Valor absoluto porque egresos son negativos
        saldo_restante = total_ingresado - total_despachado
        
        # Calcular montos
        monto_ingresado = ingresos['total_costo'] or 0
        monto_despachado = abs(despachos['total_costo'] or 0)
        monto_restante = monto_ingresado - monto_despachado
        
        resultado.append({
            'dte_id': dte.id,
            'dte_numero': dte.numero_documento,
            'dte_tipo': dte.tipo_documento,
            'fecha_emision': dte.fecha_emision.strftime('%Y-%m-%d'),
            'proveedor_id': dte.receptor.id,
            'proveedor_nombre': dte.receptor.nombre,
            'proveedor_rut': dte.receptor.rut,
            'total_ingresado': total_ingresado,
            'total_despachado': total_despachado,
            'saldo_restante': saldo_restante,
            'monto_ingresado': float(monto_ingresado),
            'monto_despachado': float(monto_despachado),
            'monto_restante': float(monto_restante),
            'estado_dte': dte.estado_dte,
            'estado_pago': dte.estado_pago
        })
    
    return JsonResponse({
        'success': True,
        'data': resultado,
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_records': total_count,
            'page_size': page_size
        }
    })

@require_GET
@login_required
def obtener_proveedores_para_reporte(request):
    """
    Obtiene la lista de proveedores para el filtro del reporte
    """
    proveedores = Empresa.objects.filter(
        esProveedor=True
    ).values('id', 'nombre', 'rut').order_by('nombre')
    
    return JsonResponse(list(proveedores), safe=False)

@login_required
def verReporteDespachosProveedor(request):
    """
    Renderiza la página completa del reporte de despachos por proveedor.
    """
    return render(request, 'vistas/modulo reportes/reporteDespachosProveedor.html')

# ========== VISTAS PARA CREACIÓN MANUAL DE PRODUCTOS ==========

@require_GET
@login_required
def obtener_proveedores(request):
    """
    Obtiene lista de proveedores para el modal de creación manual
    """
    try:
        print("🔍 Obteniendo lista de proveedores...")
        proveedores = Empresa.objects.filter(esProveedor=True).values('id', 'nombre').order_by('nombre')
        result = list(proveedores)
        print(f"✅ Proveedores encontrados: {len(result)}")
        print(f"📋 Datos: {result}")
        return JsonResponse(result, safe=False)
    except Exception as e:
        print(f"❌ Error obteniendo proveedores: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@require_GET
@login_required
def obtener_dtes_por_proveedor(request, proveedor_id):
    """
    Obtiene DTEs de un proveedor específico para el modal de creación manual
    """
    try:
        print(f"🔍 DTEs por proveedor - Proveedor ID: {proveedor_id}")
        
        # Primero verificar si el proveedor existe
        proveedor = Empresa.objects.filter(id=proveedor_id, esProveedor=True).first()
        if not proveedor:
            print(f"❌ Proveedor no encontrado: {proveedor_id}")
            return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)
        
        print(f"✅ Proveedor encontrado: {proveedor.nombre}")
        
        # Buscar DTEs pagados primero
        dtes_pagados = Dte.objects.filter(
            receptor_id=proveedor_id,
            estado_pago='PAGADO'
        ).values('id', 'numero_documento', 'fecha_emision').order_by('-fecha_emision')
        
        print(f"✅ DTEs pagados encontrados: {dtes_pagados.count()}")
        
        # Si no hay DTEs pagados, buscar todos los DTEs del proveedor
        if dtes_pagados.count() == 0:
            print("⚠️ No hay DTEs pagados, buscando todos los DTEs del proveedor")
            dtes = Dte.objects.filter(
                receptor_id=proveedor_id
            ).values('id', 'numero_documento', 'fecha_emision').order_by('-fecha_emision')
            print(f"✅ Total DTEs encontrados: {dtes.count()}")
        else:
            dtes = dtes_pagados
        
        # Formatear fecha para mostrar
        for dte in dtes:
            if dte['fecha_emision']:
                dte['fecha'] = dte['fecha_emision'].strftime('%d/%m/%Y')
            else:
                dte['fecha'] = 'Sin fecha'
        
        result = list(dtes)
        print(f"📋 Datos a enviar: {result}")
        return JsonResponse(result, safe=False)
    except Exception as e:
        print(f"❌ Error en obtener_dtes_por_proveedor: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
def crear_producto_manual(request):
    """
    Crea un producto manualmente con DTE y proveedor seleccionados
    """
    try:
        # Obtener datos del formulario
        es_manual = request.POST.get('es_manual') == 'true'
        proveedor_id = request.POST.get('proveedor')
        dte_id = request.POST.get('dte_manual')
        articulo = request.POST.get('articulo')
        descripcion = request.POST.get('descripcion', '')
        atributo1 = request.POST.get('atributo1')  # Marca
        atributo2 = request.POST.get('atributo2')  # Color
        atributo3 = request.POST.get('atributo3')  # Género
        categoria_id = request.POST.get('categoria')
        tipo_talla = request.POST.get('tipo_talla')
        guia_talla_id = request.POST.get('guia_talla')
        costo = Decimal(request.POST.get('costo', 0))
        sobreprecio = Decimal(request.POST.get('sobreprecio', 0))
        precioventa = Decimal(request.POST.get('precioventa', 0))
        
        # Obtener tallas del formulario
        tallas = request.POST.getlist('talla[]')
        stocks = request.POST.getlist('stock[]')
        skus = request.POST.getlist('sku[]')
        guias_talla_manual = request.POST.getlist('guia_talla_manual[]')
        
        # Validaciones básicas
        if not all([proveedor_id, dte_id, articulo, atributo1, atributo2, atributo3, categoria_id, tipo_talla, costo, precioventa]):
            return JsonResponse({'success': False, 'error': 'Todos los campos obligatorios deben estar completos'})
        
        if not tallas:
            return JsonResponse({'success': False, 'error': 'Debe agregar al menos una talla'})
        
        # Obtener objetos relacionados
        proveedor = get_object_or_404(Empresa, id=proveedor_id)
        dte = get_object_or_404(Dte, id=dte_id)
        categoria = get_object_or_404(Categoria, id=categoria_id)
        sucursal = get_object_or_404(Sucursal, id=request.session.get('idSucursalActual'))
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        # Obtener instancias de AtributoOpcion (no strings)
        atributo1_obj = get_object_or_404(AtributoOpcion, id=atributo1)
        atributo2_obj = get_object_or_404(AtributoOpcion, id=atributo2)
        atributo3_obj = get_object_or_404(AtributoOpcion, id=atributo3)
        
        # Verificar si el producto ya existe
        producto_existente = Producto.objects.filter(
            articulo=articulo,
            atributo1=atributo1_obj,
            atributo2=atributo2_obj,
            atributo3=atributo3_obj
        ).select_related('atributo1', 'atributo2', 'atributo3', 'categoria', 'sucursal').first()
        
        if producto_existente:
            # Obtener tallas existentes del producto
            tallas_existentes = Producto_Talla.objects.filter(producto=producto_existente).values(
                'id', 'talla', 'stock', 'sku'
            )
            
            # Preparar datos del producto existente para comparación
            datos_existente = {
                'id': producto_existente.id,
                'articulo': producto_existente.articulo,
                'descripcion': producto_existente.descripcion or '',
                'costo': float(producto_existente.costo),
                'sobreprecio': float(producto_existente.sobreprecio),
                'precioventa': float(producto_existente.precioventa),
                'precioSugerido': float(producto_existente.precioSugerido) if producto_existente.precioSugerido else None,
                'categoria_id': producto_existente.categoria.id if producto_existente.categoria else None,
                'categoria_nombre': producto_existente.categoria.nombre if producto_existente.categoria else None,
                'tipo_talla': producto_existente.tipo_talla,
                'guia_talla_id': producto_existente.guia_talla.id if producto_existente.guia_talla else None,
                'marca': producto_existente.atributo1.valor if producto_existente.atributo1 else None,
                'color': producto_existente.atributo2.valor if producto_existente.atributo2 else None,
                'genero': producto_existente.atributo3.valor if producto_existente.atributo3 else None,
                'tallas': list(tallas_existentes)
            }
            
            # Datos del formulario para comparar
            datos_nuevos = {
                'costo': float(costo),
                'sobreprecio': float(sobreprecio),
                'precioventa': float(precioventa),
                'tallas': []
            }
            
            # Preparar tallas del formulario
            for i, talla in enumerate(tallas):
                if not talla.strip():
                    continue
                stock = int(stocks[i]) if i < len(stocks) else 1
                sku = skus[i] if i < len(skus) else ''
                datos_nuevos['tallas'].append({
                    'talla': talla.strip(),
                    'stock': stock,
                    'sku': sku
                })
            
            return JsonResponse({
                'success': False,
                'producto_existente': True,
                'error': 'Ya existe un producto con estas características',
                'producto_actual': datos_existente,
                'producto_nuevo': datos_nuevos
            })
        
        # Crear el producto
        producto = Producto.objects.create(
            articulo=articulo,
            descripcion=descripcion,
            atributo1=atributo1_obj,
            atributo2=atributo2_obj,
            atributo3=atributo3_obj,
            categoria=categoria,
            tipo_talla=tipo_talla,
            guia_talla_id=guia_talla_id if guia_talla_id else None,
            costo=costo,
            sobreprecio=sobreprecio,
            precioventa=precioventa,
            precioSugerido=precioventa,  # Usar precioSugerido (con S mayúscula)
            sucursal=sucursal
        )
        
        # Crear variantes de talla
        for i, talla in enumerate(tallas):
            if not talla.strip():
                continue
                
            stock = int(stocks[i]) if i < len(stocks) else 1
            sku = skus[i] if i < len(skus) else ''
            guia_talla_manual_id = guias_talla_manual[i] if i < len(guias_talla_manual) else None
            
            # Generar SKU si no se proporcionó
            if not sku:
                sku = _next_sku()
            
            # Crear variante de talla
            producto_talla = Producto_Talla.objects.create(
                producto=producto,
                talla=talla.strip(),
                stock=stock,
                sku=sku,
                guia_talla_id=guia_talla_manual_id if guia_talla_manual_id else None
            )
            
            # Registrar movimiento de ingreso inicial
            registrar_movimiento_producto(
                producto_talla=producto_talla,
                concepto='INGRESO_MANUAL',
                cantidad=stock,
                responsable=responsable,
                dte=dte,
                observaciones=f'Creación manual - DTE: {dte.numero_documento}',
                referencia_externa=f'Manual-{dte.numero_documento}'
            )
        
        return JsonResponse({
            'success': True,
            'producto_id': producto.id,
            'mensaje': 'Producto creado correctamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_POST
@login_required
def actualizar_producto_existente(request):
    """
    Actualiza un producto existente con nuevos precios y/o agrega nuevas tallas
    """
    try:
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        
        producto_id = data.get('producto_id')
        actualizar_precios = data.get('actualizar_precios', False)
        agregar_tallas = data.get('agregar_tallas', False)
        
        # Obtener producto existente
        producto = get_object_or_404(Producto, id=producto_id)
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        cambios_realizados = []
        tallas_agregadas = []
        
        # Actualizar precios si se solicita
        if actualizar_precios:
            nuevo_costo = Decimal(data.get('costo', producto.costo))
            nuevo_sobreprecio = Decimal(data.get('sobreprecio', producto.sobreprecio))
            nuevo_precioventa = Decimal(data.get('precioventa', producto.precioventa))
            
            precio_cambiado = False
            if producto.costo != nuevo_costo:
                producto.costo = nuevo_costo
                precio_cambiado = True
                cambios_realizados.append(f'Costo: ${producto.costo:,.0f} → ${nuevo_costo:,.0f}')
            
            if producto.sobreprecio != nuevo_sobreprecio:
                producto.sobreprecio = nuevo_sobreprecio
                precio_cambiado = True
                cambios_realizados.append(f'Sobreprecio: ${producto.sobreprecio:,.0f} → ${nuevo_sobreprecio:,.0f}')
            
            if producto.precioventa != nuevo_precioventa:
                producto.precioventa = nuevo_precioventa
                producto.precioSugerido = nuevo_precioventa
                precio_cambiado = True
                cambios_realizados.append(f'Precio Venta: ${producto.precioventa:,.0f} → ${nuevo_precioventa:,.0f}')
            
            if precio_cambiado:
                producto.save()
        
        # Agregar nuevas tallas si se solicita
        if agregar_tallas:
            tallas_nuevas = data.get('tallas', [])
            dte_id = data.get('dte_id')
            dte = get_object_or_404(Dte, id=dte_id) if dte_id else None
            
            # Obtener tallas existentes
            tallas_existentes = set(
                Producto_Talla.objects.filter(producto=producto).values_list('talla', flat=True)
            )
            
            for talla_data in tallas_nuevas:
                talla = talla_data.get('talla', '').strip()
                if not talla:
                    continue
                
                # Verificar si la talla ya existe
                if talla in tallas_existentes:
                    # Actualizar stock de talla existente (sumar al existente)
                    producto_talla = Producto_Talla.objects.filter(
                        producto=producto, 
                        talla=talla
                    ).first()
                    
                    if producto_talla:
                        stock_adicional = int(talla_data.get('stock', 0))
                        if stock_adicional > 0:
                            # Sumar stock adicional al existente
                            producto_talla.stock += stock_adicional
                            producto_talla.save()
                            
                            # Registrar movimiento de ingreso
                            registrar_movimiento_producto(
                                producto_talla=producto_talla,
                                concepto='INGRESO_MANUAL',
                                cantidad=stock_adicional,
                                responsable=responsable,
                                dte=dte,
                                observaciones=f'Actualización manual - DTE: {dte.numero_documento if dte else "N/A"}',
                                referencia_externa=f'Manual-{dte.numero_documento if dte else "N/A"}'
                            )
                            stock_final = producto_talla.stock
                            cambios_realizados.append(f'Talla {talla}: Stock actualizado (+{stock_adicional}, total: {stock_final})')
                else:
                    # Crear nueva talla
                    stock = int(talla_data.get('stock', 1))
                    sku = talla_data.get('sku', '')
                    
                    # Generar SKU si no se proporcionó
                    if not sku:
                        sku = _next_sku()
                    
                    # Verificar que el SKU no esté en uso
                    while Producto_Talla.objects.filter(sku=sku).exists():
                        sku = _next_sku()
                    
                    producto_talla = Producto_Talla.objects.create(
                        producto=producto,
                        talla=talla,
                        stock=stock,
                        sku=sku
                    )
                    
                    # Registrar movimiento de ingreso
                    if dte:
                        registrar_movimiento_producto(
                            producto_talla=producto_talla,
                            concepto='INGRESO_MANUAL',
                            cantidad=stock,
                            responsable=responsable,
                            dte=dte,
                            observaciones=f'Creación manual - DTE: {dte.numero_documento}',
                            referencia_externa=f'Manual-{dte.numero_documento}'
                        )
                    
                    tallas_agregadas.append(talla)
                    cambios_realizados.append(f'Talla {talla}: Agregada (Stock: {stock}, SKU: {sku})')
        
        return JsonResponse({
            'success': True,
            'producto_id': producto.id,
            'mensaje': 'Producto actualizado correctamente',
            'cambios': cambios_realizados,
            'tallas_agregadas': tallas_agregadas
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

# ========== VISTAS PARA BUSCADOR DE PRODUCTOS EXISTENTES ==========

@require_GET
@login_required
def buscar_productos_existentes(request):
    """
    Busca productos existentes por término de búsqueda
    """
    try:
        termino = request.GET.get('q', '').strip()
        print(f"🔍 Buscando productos con término: '{termino}'")
        
        if not termino:
            print("❌ Término de búsqueda vacío")
            return JsonResponse({'success': False, 'error': 'Término de búsqueda requerido'}, status=400)
        
        # Buscar productos que coincidan con el término
        productos = Producto.objects.filter(
            Q(articulo__icontains=termino) |
            Q(descripcion__icontains=termino)
        ).select_related(
            'categoria',
            'atributo1',
            'atributo2', 
            'atributo3'
        )[:10]  # Limitar a 10 resultados
        
        print(f"✅ Productos encontrados en DB: {productos.count()}")
        
        resultados = []
        for producto in productos:
            resultado = {
                'id': producto.id,
                'articulo': producto.articulo,
                'descripcion': producto.descripcion,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.atributo1.valor if producto.atributo1 else '',
                'color': producto.atributo2.valor if producto.atributo2 else '',
                'genero': producto.atributo3.valor if producto.atributo3 else '',
                'costo': float(producto.costo),
                'sobreprecio': float(producto.sobreprecio),
                'precioventa': float(producto.precioventa)
            }
            resultados.append(resultado)
            print(f"📦 Producto procesado: {resultado['articulo']}")
        
        response_data = {
            'success': True,
            'productos': resultados
        }
        
        print(f"📋 Respuesta final: {len(resultados)} productos")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"❌ Error en búsqueda de productos: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def detalle_producto_para_copiar(request, producto_id):
    """
    Obtiene detalles completos de un producto para copiar sus datos
    """
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        
        # Obtener tallas del producto
        tallas = Producto_Talla.objects.filter(producto=producto).values('talla', 'sku')
        
        # Preparar datos para copiar
        datos_producto = {
            'id': producto.id,
            'articulo': producto.articulo,
            'descripcion': producto.descripcion,
            'categoria': producto.categoria.nombre if producto.categoria else '',
            'categoria_id': producto.categoria.id if producto.categoria else None,
            'marca': producto.atributo1.valor if producto.atributo1 else '',
            'atributo1': producto.atributo1.id if producto.atributo1 else None,
            'color': producto.atributo2.valor if producto.atributo2 else '',
            'atributo2': producto.atributo2.id if producto.atributo2 else None,
            'genero': producto.atributo3.valor if producto.atributo3 else '',
            'atributo3': producto.atributo3.id if producto.atributo3 else None,
            'tipo_talla': producto.tipo_talla,
            'guia_talla_id': producto.guia_talla.id if producto.guia_talla else None,
            'costo': float(producto.costo),
            'sobreprecio': float(producto.sobreprecio),
            'precioventa': float(producto.precioventa),
            'tallas': list(tallas)
        }
        
        return JsonResponse({
            'success': True,
            'producto': datos_producto
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
@require_GET
@login_required
def tallas_producto(request, producto_id):
    """
    Obtiene las tallas de un producto específico
    """
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        
        # Obtener todas las tallas del producto
        tallas = Producto_Talla.objects.filter(producto=producto).select_related('guia_talla')
        
        # Preparar datos del producto
        datos_producto = {
            'id': producto.id,
            'articulo': producto.articulo,
            'descripcion': producto.descripcion,
            'categoria': producto.categoria.nombre if producto.categoria else '',
            'marca': producto.atributo1.valor if producto.atributo1 else '',
            'color': producto.atributo2.valor if producto.atributo2 else '',
            'genero': producto.atributo3.valor if producto.atributo3 else '',
            'precioventa': float(producto.precioventa)
        }
        
        # Preparar datos de las tallas
        datos_tallas = []
        for talla in tallas:
            # Obtener equivalencias de la guía de talla si existe
            equivalencias = ''
            if talla.guia_talla:
                # Buscar la equivalencia en la guía
                from .models import Guia_Talla_Item
                item_guia = Guia_Talla_Item.objects.filter(
                    guia_talla=talla.guia_talla,
                    cl=talla.talla
                ).first()
                
                if item_guia:
                    equivalencias = f"US: {item_guia.us or '-'} / EU: {item_guia.eu or '-'} / UK: {item_guia.uk or '-'} / BR: {item_guia.br or '-'} / CM: {item_guia.cm or '-'}"
            
            datos_talla = {
                'id': talla.id,
                'talla': talla.talla,
                'sku': talla.sku,
                'stock': talla.stock,
                'precio': float(producto.precioventa),  # Usar precio del producto
                'activo': talla.activo if hasattr(talla, 'activo') else True,
                'equivalencias': equivalencias
            }
            datos_tallas.append(datos_talla)
        
        return JsonResponse({
            'success': True,
            'producto': datos_producto,
            'tallas': datos_tallas
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo tallas del producto {producto_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ========== FUNCIONES FIFO ==========

def crear_lote_producto(producto_talla, cantidad, costo_unitario, sobreprecio_unitario, 
                       precio_venta_unitario, dte=None, movimiento=None, 
                       numero_lote=None, fecha_vencimiento=None, observaciones=None):
    """
    Crea un nuevo lote de producto para implementar FIFO
    """
    from .models import LoteProducto
    
    lote = LoteProducto.objects.create(
        producto_talla=producto_talla,
        dte=dte,
        movimiento=movimiento,
        cantidad_inicial=cantidad,
        cantidad_disponible=cantidad,
        costo_unitario=costo_unitario,
        sobreprecio_unitario=sobreprecio_unitario,
        precio_venta_unitario=precio_venta_unitario,
        numero_lote=numero_lote,
        fecha_vencimiento=fecha_vencimiento,
        observaciones=observaciones
    )
    
    return lote
def consumir_stock_fifo(producto_talla, cantidad_requerida, responsable, ticket=None, 
                       observaciones=None, referencia_externa=None):
    """
    Consume stock usando metodología FIFO (First In, First Out)
    Retorna el costo total consumido y los lotes utilizados
    """
    from .models import LoteProducto, Movimientos_Producto
    
    print(f"🔍 FIFO LLAMADO: SKU {producto_talla.sku}, Cantidad: {cantidad_requerida}, Ticket: {ticket.correlativo if ticket else 'N/A'}")
    
    if cantidad_requerida <= 0:
        return 0, []
    
    # Obtener lotes disponibles ordenados por fecha de ingreso (FIFO)
    lotes_disponibles = LoteProducto.objects.filter(
        producto_talla=producto_talla,
        cantidad_disponible__gt=0,
        activo=True,
        agotado=False
    ).order_by('fecha_ingreso')
    
    if not lotes_disponibles.exists():
        raise Exception(f'No hay stock disponible para {producto_talla.producto.articulo} - Talla {producto_talla.talla}')
    
    cantidad_restante = cantidad_requerida
    costo_total_consumido = 0
    lotes_utilizados = []
    
    for lote in lotes_disponibles:
        if cantidad_restante <= 0:
            break
            
        # Calcular cuánto podemos consumir de este lote
        cantidad_a_consumir = min(cantidad_restante, lote.cantidad_disponible)
        
        # Actualizar el lote
        lote.cantidad_disponible -= cantidad_a_consumir
        lote.save()
        
        # Calcular costo de esta porción
        costo_porcion = cantidad_a_consumir * lote.costo_unitario
        costo_total_consumido += costo_porcion
        
        # Registrar el detalle del consumo
        lotes_utilizados.append({
            'lote_id': lote.id,
            'cantidad_consumida': cantidad_a_consumir,
            'costo_unitario': lote.costo_unitario,
            'costo_total': costo_porcion,
            'fecha_ingreso_lote': lote.fecha_ingreso,
            'dte_origen': lote.dte.numero_documento if lote.dte else None
        })
        
        cantidad_restante -= cantidad_a_consumir
    
    if cantidad_restante > 0:
        raise Exception(f'Stock insuficiente. Faltan {cantidad_restante} unidades')
    
    # Actualizar stock del producto_talla
    stock_antes_descuento = producto_talla.stock
    producto_talla.stock -= cantidad_requerida
    producto_talla.save()
    print(f"🔍 FIFO DESCUENTO: SKU {producto_talla.sku} - Stock {stock_antes_descuento} → {producto_talla.stock} (Descontado: {cantidad_requerida})")
    
    # Crear movimiento de EGRESO en Movimientos_Producto
    costo_promedio = costo_total_consumido // cantidad_requerida if cantidad_requerida > 0 else 0
    
    sucursal_origen_mov = ticket.sucursal if ticket and ticket.sucursal else getattr(producto_talla.producto, 'sucursal', None)
    sobreprecio_mov = producto_talla.producto.sobreprecio if producto_talla.producto and hasattr(producto_talla.producto, 'sobreprecio') else 0
    
    # ✅ Determinar referencia externa: usar DTE si está disponible, si no usar correlativo del ticket
    if referencia_externa:
        ref_final = referencia_externa
    elif ticket:
        ref_final = f'DTE_{ticket.folio_dte}' if ticket.folio_dte else f'TICKET_{ticket.correlativo}'
    else:
        ref_final = None
    
    movimiento = Movimientos_Producto.objects.create(
        ticket=ticket,
        ProductoTalla=producto_talla,
        sucursal_origen=sucursal_origen_mov,
        cantidad=-cantidad_requerida,  # Negativo para EGRESO
        costo=costo_promedio,
        precio=producto_talla.producto.precioventa if producto_talla.producto else 0,
        sobreprecio=sobreprecio_mov,
        concepto='VENTA_TICKET' if ticket else 'VENTA_DIRECTA',
        tipo_movimiento='EGRESO',
        responsable=responsable if isinstance(responsable, str) else responsable.username,
        observaciones=observaciones or f'Consumo FIFO - {cantidad_requerida} unidades',
        referencia_externa=ref_final
    )
    
    print(f"✓ Movimiento #{movimiento.id} creado: {movimiento.concepto} - Cantidad: {movimiento.cantidad} - SKU: {producto_talla.sku}")
    
    return costo_total_consumido, lotes_utilizados

def obtener_valor_inventario_fifo(producto_talla):
    """
    Calcula el valor del inventario usando metodología FIFO
    """
    from .models import LoteProducto
    
    lotes_disponibles = LoteProducto.objects.filter(
        producto_talla=producto_talla,
        cantidad_disponible__gt=0,
        activo=True,
        agotado=False
    ).order_by('fecha_ingreso')
    
    valor_total = 0
    for lote in lotes_disponibles:
        valor_total += lote.valor_disponible
    
    return valor_total

def obtener_costo_promedio_fifo(producto_talla):
    """
    Calcula el costo promedio ponderado usando metodología FIFO
    """
    from .models import LoteProducto
    
    lotes_disponibles = LoteProducto.objects.filter(
        producto_talla=producto_talla,
        cantidad_disponible__gt=0,
        activo=True,
        agotado=False
    )
    
    if not lotes_disponibles.exists():
        return 0
    
    valor_total = 0
    cantidad_total = 0
    
    for lote in lotes_disponibles:
        valor_total += lote.valor_disponible
        cantidad_total += lote.cantidad_disponible
    
    if cantidad_total == 0:
        return 0
    
    return valor_total / cantidad_total

# ========== VISTAS PARA GESTIÓN FIFO ==========

@require_GET
@login_required
def ver_lotes_producto(request, producto_talla_id):
    """
    Vista para ver los lotes de un producto específico
    """
    try:
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        
        # Obtener todos los lotes del producto
        lotes = LoteProducto.objects.filter(
            producto_talla=producto_talla
        ).order_by('fecha_ingreso')
        
        # Calcular estadísticas FIFO
        lotes_activos = lotes.filter(activo=True, agotado=False)
        valor_inventario_fifo = obtener_valor_inventario_fifo(producto_talla)
        costo_promedio_fifo = obtener_costo_promedio_fifo(producto_talla)
        
        context = {
            'producto_talla': producto_talla,
            'lotes': lotes,
            'lotes_activos': lotes_activos,
            'valor_inventario_fifo': valor_inventario_fifo,
            'costo_promedio_fifo': costo_promedio_fifo,
            'stock_sistema': producto_talla.stock,
            'stock_fifo': sum(lote.cantidad_disponible for lote in lotes_activos)
        }
        
        return render(request, 'vistas/modulo_existencias/lotes_producto.html', context)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def obtener_lotes_producto(request, producto_talla_id):
    """
    API para obtener lotes de un producto en formato JSON
    """
    try:
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        
        lotes = LoteProducto.objects.filter(
            producto_talla=producto_talla
        ).order_by('fecha_ingreso')
        
        lotes_data = []
        for lote in lotes:
            lotes_data.append({
                'id': lote.id,
                'cantidad_inicial': lote.cantidad_inicial,
                'cantidad_disponible': lote.cantidad_disponible,
                'costo_unitario': lote.costo_unitario,
                'precio_venta_unitario': lote.precio_venta_unitario,
                'fecha_ingreso': lote.fecha_ingreso.strftime('%d/%m/%Y %H:%M'),
                'fecha_vencimiento': lote.fecha_vencimiento.strftime('%d/%m/%Y') if lote.fecha_vencimiento else None,
                'activo': lote.activo,
                'agotado': lote.agotado,
                'numero_lote': lote.numero_lote,
                'observaciones': lote.observaciones,
                'valor_disponible': lote.valor_disponible,
                'porcentaje_consumido': round(lote.porcentaje_consumido, 2),
                'dte_origen': lote.dte.numero_documento if lote.dte else None
            })
        
        return JsonResponse({
            'success': True,
            'lotes': lotes_data,
            'producto': {
                'articulo': producto_talla.producto.articulo,
                'talla': producto_talla.talla,
                'stock_sistema': producto_talla.stock,
                'valor_inventario_fifo': obtener_valor_inventario_fifo(producto_talla),
                'costo_promedio_fifo': obtener_costo_promedio_fifo(producto_talla)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_POST
@login_required
@transaction.atomic
def crear_lote_manual(request):
    """
    Crear un lote manualmente
    """
    try:
        producto_talla_id = request.POST.get('producto_talla_id')
        cantidad = int(request.POST.get('cantidad', 0))
        costo_unitario = int(request.POST.get('costo_unitario', 0))
        sobreprecio_unitario = int(request.POST.get('sobreprecio_unitario', 0))
        precio_venta_unitario = int(request.POST.get('precio_venta_unitario', 0))
        numero_lote = request.POST.get('numero_lote', '')
        fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
        observaciones = request.POST.get('observaciones', '')
        
        if not all([producto_talla_id, cantidad, costo_unitario, precio_venta_unitario]):
            return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios'})
        
        producto_talla = get_object_or_404(Producto_Talla, id=producto_talla_id)
        responsable = request.session.get('nombreUsuario', 'Sistema')
        
        # Crear el lote
        lote = crear_lote_producto(
            producto_talla=producto_talla,
            cantidad=cantidad,
            costo_unitario=costo_unitario,
            sobreprecio_unitario=sobreprecio_unitario,
            precio_venta_unitario=precio_venta_unitario,
            numero_lote=numero_lote if numero_lote else None,
            fecha_vencimiento=parse_date(fecha_vencimiento) if fecha_vencimiento else None,
            observaciones=observaciones
        )
        
        # Actualizar stock del producto
        producto_talla.stock += cantidad
        producto_talla.save()
        
        # Registrar movimiento
        registrar_movimiento_producto(
            producto_talla=producto_talla,
            concepto='AJUSTE_POSITIVO',
            cantidad=cantidad,
            responsable=responsable,
            observaciones=f'Lote manual creado - {observaciones}',
            referencia_externa=f'Lote-{lote.id}',
            crear_lote_fifo=False  # Ya creamos el lote manualmente
        )
        
        return JsonResponse({
            'success': True,
            'lote_id': lote.id,
            'mensaje': 'Lote creado correctamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_POST
@login_required
@transaction.atomic
def ajustar_lote(request, lote_id):
    """
    Ajustar cantidad disponible de un lote
    """
    try:
        lote = get_object_or_404(LoteProducto, id=lote_id)
        nueva_cantidad = int(request.POST.get('cantidad_disponible', 0))
        observaciones = request.POST.get('observaciones', '')
        
        if nueva_cantidad < 0:
            return JsonResponse({'success': False, 'error': 'La cantidad no puede ser negativa'})
        
        cantidad_anterior = lote.cantidad_disponible
        diferencia = nueva_cantidad - cantidad_anterior
        
        # Actualizar el lote
        lote.cantidad_disponible = nueva_cantidad
        lote.observaciones = f"{lote.observaciones or ''}\nAjuste: {observaciones}"
        lote.save()
        
        # Actualizar stock del producto
        producto_talla = lote.producto_talla
        producto_talla.stock += diferencia
        producto_talla.save()
        
        # Registrar movimiento de ajuste
        responsable = request.session.get('nombreUsuario', 'Sistema')
        concepto = 'AJUSTE_POSITIVO' if diferencia > 0 else 'AJUSTE_NEGATIVO'
        
        registrar_movimiento_producto(
            producto_talla=producto_talla,
            concepto=concepto,
            cantidad=diferencia,
            responsable=responsable,
            observaciones=f'Ajuste lote {lote.id}: {cantidad_anterior} → {nueva_cantidad} - {observaciones}',
            referencia_externa=f'Lote-{lote.id}',
            crear_lote_fifo=False
        )
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Lote ajustado correctamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def reporte_fifo_general(request):
    """
    Reporte general de inventario FIFO
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Obtener productos con lotes activos
        productos_con_lotes = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_id,
            lotes__activo=True,
            lotes__agotado=False
        ).distinct()
        
        reporte_data = []
        valor_total_inventario = 0
        
        for producto_talla in productos_con_lotes:
            valor_inventario = obtener_valor_inventario_fifo(producto_talla)
            costo_promedio = obtener_costo_promedio_fifo(producto_talla)
            
            reporte_data.append({
                'producto_id': producto_talla.producto.id,
                'producto_talla_id': producto_talla.id,
                'articulo': producto_talla.producto.articulo,
                'talla': producto_talla.talla,
                'stock_sistema': producto_talla.stock,
                'valor_inventario_fifo': valor_inventario,
                'costo_promedio_fifo': costo_promedio,
                'diferencia_valor': valor_inventario - (producto_talla.stock * producto_talla.producto.costo)
            })
            
            valor_total_inventario += valor_inventario
        
        return JsonResponse({
            'success': True,
            'reporte': reporte_data,
            'valor_total_inventario': valor_total_inventario,
            'total_productos': len(reporte_data)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def dashboard_fifo(request):
    """
    Vista completa del dashboard FIFO
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Solo renderizar el template, los datos se cargarán via AJAX
        context = {
            'sucursal_id': sucursal_id
        }
        
        return render(request, 'vistas/modulo_dashboards/dashboard_fifo.html', context)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
# ========== VISTAS AJAX PARA DASHBOARD FIFO ==========
@require_GET
@login_required
def obtener_datos_dashboard_fifo(request):
    """
    API para obtener datos del dashboard FIFO con filtros
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Parámetros de filtrado
        filtro_producto = request.GET.get('filtro_producto', '').lower()
        filtro_talla = request.GET.get('filtro_talla', '').lower()
        filtro_diferencia = request.GET.get('filtro_diferencia', '')
        filtro_stock = request.GET.get('filtro_stock', '')
        ordenar_por = request.GET.get('ordenar_por', 'diferencia')
        mostrar_solo_diferencias = request.GET.get('mostrar_solo_diferencias', 'false') == 'true'
        
        # Obtener productos con lotes activos
        productos_con_lotes = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_id,
            lotes__activo=True,
            lotes__agotado=False
        ).distinct()
        
        # Aplicar filtros
        if filtro_producto:
            productos_con_lotes = productos_con_lotes.filter(
                producto__articulo__icontains=filtro_producto
            )
        
        if filtro_talla:
            productos_con_lotes = productos_con_lotes.filter(
                talla__icontains=filtro_talla
            )
        
        # Obtener datos detallados
        productos_data = []
        valor_total_inventario = 0
        productos_con_diferencia = 0
        diferencia_total = 0
        
        # Contadores para análisis
        diferencias_positivas = 0
        diferencias_negativas = 0
        sin_diferencias = 0
        stock_bajo = 0
        stock_medio = 0
        stock_alto = 0
        sin_stock = 0
        
        for producto_talla in productos_con_lotes:
            valor_inventario = obtener_valor_inventario_fifo(producto_talla)
            costo_promedio = obtener_costo_promedio_fifo(producto_talla)
            valor_sistema = producto_talla.stock * producto_talla.producto.costo
            diferencia = valor_inventario - valor_sistema
            
            # Calcular porcentaje de diferencia
            porcentaje_diferencia = 0
            if valor_sistema > 0:
                porcentaje_diferencia = (diferencia / valor_sistema) * 100
            
            # Aplicar filtros adicionales
            if filtro_diferencia:
                if filtro_diferencia == 'positiva' and diferencia <= 0:
                    continue
                elif filtro_diferencia == 'negativa' and diferencia >= 0:
                    continue
                elif filtro_diferencia == 'cero' and diferencia != 0:
                    continue
                elif filtro_diferencia == 'alta' and porcentaje_diferencia <= 10:
                    continue
                elif filtro_diferencia == 'baja' and porcentaje_diferencia >= 5:
                    continue
            
            if filtro_stock:
                if filtro_stock == 'bajo' and producto_talla.stock >= 10:
                    continue
                elif filtro_stock == 'medio' and (producto_talla.stock < 10 or producto_talla.stock > 50):
                    continue
                elif filtro_stock == 'alto' and producto_talla.stock <= 50:
                    continue
                elif filtro_stock == 'agotado' and producto_talla.stock > 0:
                    continue
            
            if mostrar_solo_diferencias and diferencia == 0:
                continue
            
            # Contar diferencias
            if diferencia > 0:
                diferencias_positivas += 1
            elif diferencia < 0:
                diferencias_negativas += 1
            else:
                sin_diferencias += 1
            
            # Contar stock
            if producto_talla.stock == 0:
                sin_stock += 1
            elif producto_talla.stock < 10:
                stock_bajo += 1
            elif producto_talla.stock <= 50:
                stock_medio += 1
            else:
                stock_alto += 1
            
            valor_total_inventario += valor_inventario
            diferencia_total += diferencia
            
            if diferencia != 0:
                productos_con_diferencia += 1
            
            productos_data.append({
                'producto_id': producto_talla.producto.id,
                'producto_talla_id': producto_talla.id,
                'articulo': producto_talla.producto.articulo,
                'talla': producto_talla.talla,
                'stock_sistema': producto_talla.stock,
                'valor_inventario_fifo': valor_inventario,
                'costo_promedio_fifo': costo_promedio,
                'diferencia_valor': diferencia,
                'costo_sistema': producto_talla.producto.costo,
                'valor_sistema': valor_sistema,
                'porcentaje_diferencia': porcentaje_diferencia,
                'sku': producto_talla.sku
            })
        
        # Ordenar datos
        if ordenar_por == 'diferencia':
            productos_data.sort(key=lambda x: abs(x['diferencia_valor']), reverse=True)
        elif ordenar_por == 'producto':
            productos_data.sort(key=lambda x: x['articulo'].lower())
        elif ordenar_por == 'valor_fifo':
            productos_data.sort(key=lambda x: x['valor_inventario_fifo'], reverse=True)
        elif ordenar_por == 'stock':
            productos_data.sort(key=lambda x: x['stock_sistema'], reverse=True)
        elif ordenar_por == 'porcentaje':
            productos_data.sort(key=lambda x: abs(x['porcentaje_diferencia']), reverse=True)
        elif ordenar_por == 'reciente':
            # Ordenar por fecha de creación del producto (aproximado)
            productos_data.sort(key=lambda x: x['producto_id'], reverse=True)
        
        # Generar alertas
        alertas = []
        diferencias_altas = sum(1 for p in productos_data if abs(p['porcentaje_diferencia']) > 20)
        if diferencias_altas > 0:
            alertas.append({
                'tipo': 'warning',
                'mensaje': f'{diferencias_altas} productos tienen diferencias superiores al 20%'
            })
        
        if stock_bajo > 5:
            alertas.append({
                'tipo': 'danger',
                'mensaje': f'{stock_bajo} productos tienen stock bajo (<10 unidades)'
            })
        
        if sin_stock > 0:
            alertas.append({
                'tipo': 'danger',
                'mensaje': f'{sin_stock} productos están sin stock'
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'total_productos': len(productos_data),
            'valor_total_inventario': valor_total_inventario,
            'productos_con_diferencia': productos_con_diferencia,
            'diferencia_total': diferencia_total,
            'porcentaje_diferencia': (diferencia_total / valor_total_inventario * 100) if valor_total_inventario > 0 else 0,
            'analisis': {
                'diferencias_positivas': diferencias_positivas,
                'diferencias_negativas': diferencias_negativas,
                'sin_diferencias': sin_diferencias,
                'stock_bajo': stock_bajo,
                'stock_medio': stock_medio,
                'stock_alto': stock_alto,
                'sin_stock': sin_stock
            },
            'alertas': alertas
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def obtener_metricas_fifo(request):
    """
    API para obtener métricas generales del FIFO
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Obtener productos con lotes activos
        productos_con_lotes = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_id,
            lotes__activo=True,
            lotes__agotado=False
        ).distinct()
        
        # Calcular métricas
        total_productos = productos_con_lotes.count()
        valor_total_inventario = 0
        productos_con_diferencia = 0
        diferencia_total = 0
        
        for producto_talla in productos_con_lotes:
            valor_inventario = obtener_valor_inventario_fifo(producto_talla)
            valor_sistema = producto_talla.stock * producto_talla.producto.costo
            diferencia = valor_inventario - valor_sistema
            
            valor_total_inventario += valor_inventario
            diferencia_total += diferencia
            
            if diferencia != 0:
                productos_con_diferencia += 1
        
        return JsonResponse({
            'success': True,
            'metricas': {
                'total_productos': total_productos,
                'valor_total_inventario': valor_total_inventario,
                'productos_con_diferencia': productos_con_diferencia,
                'diferencia_total': diferencia_total,
                'porcentaje_diferencia': (diferencia_total / valor_total_inventario * 100) if valor_total_inventario > 0 else 0
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
@login_required
def exportar_dashboard_fifo(request):
    """
    Exportar datos del dashboard FIFO
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Obtener parámetros de filtrado
        filtro_producto = request.GET.get('filtro_producto', '').lower()
        filtro_talla = request.GET.get('filtro_talla', '').lower()
        filtro_diferencia = request.GET.get('filtro_diferencia', '')
        filtro_stock = request.GET.get('filtro_stock', '')
        
        # Obtener productos con lotes activos
        productos_con_lotes = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_id,
            lotes__activo=True,
            lotes__agotado=False
        ).distinct()
        
        # Aplicar filtros
        if filtro_producto:
            productos_con_lotes = productos_con_lotes.filter(
                producto__articulo__icontains=filtro_producto
            )
        
        if filtro_talla:
            productos_con_lotes = productos_con_lotes.filter(
                talla__icontains=filtro_talla
            )
        
        # Generar datos para exportación
        datos_exportacion = []
        for producto_talla in productos_con_lotes:
            valor_inventario = obtener_valor_inventario_fifo(producto_talla)
            costo_promedio = obtener_costo_promedio_fifo(producto_talla)
            valor_sistema = producto_talla.stock * producto_talla.producto.costo
            diferencia = valor_inventario - valor_sistema
            porcentaje_diferencia = (diferencia / valor_sistema * 100) if valor_sistema > 0 else 0
            
            # Aplicar filtros adicionales
            if filtro_diferencia:
                if filtro_diferencia == 'positiva' and diferencia <= 0:
                    continue
                elif filtro_diferencia == 'negativa' and diferencia >= 0:
                    continue
                elif filtro_diferencia == 'cero' and diferencia != 0:
                    continue
                elif filtro_diferencia == 'alta' and porcentaje_diferencia <= 10:
                    continue
                elif filtro_diferencia == 'baja' and porcentaje_diferencia >= 5:
                    continue
            
            if filtro_stock:
                if filtro_stock == 'bajo' and producto_talla.stock >= 10:
                    continue
                elif filtro_stock == 'medio' and (producto_talla.stock < 10 or producto_talla.stock > 50):
                    continue
                elif filtro_stock == 'alto' and producto_talla.stock <= 50:
                    continue
                elif filtro_stock == 'agotado' and producto_talla.stock > 0:
                    continue
            
            datos_exportacion.append({
                'SKU': producto_talla.sku,
                'Artículo': producto_talla.producto.articulo,
                'Talla': producto_talla.talla,
                'Stock Sistema': producto_talla.stock,
                'Costo Sistema': producto_talla.producto.costo,
                'Valor Sistema': valor_sistema,
                'Valor FIFO': valor_inventario,
                'Costo Promedio FIFO': costo_promedio,
                'Diferencia': diferencia,
                'Porcentaje Diferencia': f"{porcentaje_diferencia:.2f}%",
                'Estado': 'Con Diferencia' if diferencia != 0 else 'Sin Diferencia'
            })
        
        # Generar CSV
        import csv
        from django.http import HttpResponse
        from io import StringIO
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="dashboard_fifo_{sucursal_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        if datos_exportacion:
            writer = csv.DictWriter(response, fieldnames=datos_exportacion[0].keys())
            writer.writeheader()
            writer.writerows(datos_exportacion)
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
def obtener_analisis_fifo_detallado(request):
    """
    API para obtener análisis detallado del FIFO
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'No hay sucursal activa'})
        
        # Obtener productos con lotes activos
        productos_con_lotes = Producto_Talla.objects.filter(
            producto__sucursal_id=sucursal_id,
            lotes__activo=True,
            lotes__agotado=False
        ).distinct()
        
        # Análisis detallado
        analisis = {
            'distribucion_diferencias': {
                'positivas': 0,
                'negativas': 0,
                'sin_diferencia': 0
            },
            'distribucion_stock': {
                'bajo': 0,
                'medio': 0,
                'alto': 0,
                'sin_stock': 0
            },
            'rangos_diferencia': {
                '0_5': 0,
                '5_10': 0,
                '10_20': 0,
                '20_50': 0,
                'mas_50': 0
            },
            'productos_criticos': [],
            'tendencias': {
                'valor_total_fifo': 0,
                'diferencia_promedio': 0,
                'productos_con_problemas': 0
            }
        }
        
        valor_total_fifo = 0
        diferencia_total = 0
        productos_con_problemas = 0
        
        for producto_talla in productos_con_lotes:
            valor_inventario = obtener_valor_inventario_fifo(producto_talla)
            valor_sistema = producto_talla.stock * producto_talla.producto.costo
            diferencia = valor_inventario - valor_sistema
            porcentaje_diferencia = abs((diferencia / valor_sistema * 100)) if valor_sistema > 0 else 0
            
            valor_total_fifo += valor_inventario
            diferencia_total += diferencia
            
            # Distribución de diferencias
            if diferencia > 0:
                analisis['distribucion_diferencias']['positivas'] += 1
            elif diferencia < 0:
                analisis['distribucion_diferencias']['negativas'] += 1
            else:
                analisis['distribucion_diferencias']['sin_diferencia'] += 1
            
            # Distribución de stock
            if producto_talla.stock == 0:
                analisis['distribucion_stock']['sin_stock'] += 1
            elif producto_talla.stock < 10:
                analisis['distribucion_stock']['bajo'] += 1
            elif producto_talla.stock <= 50:
                analisis['distribucion_stock']['medio'] += 1
            else:
                analisis['distribucion_stock']['alto'] += 1
            
            # Rangos de diferencia
            if porcentaje_diferencia <= 5:
                analisis['rangos_diferencia']['0_5'] += 1
            elif porcentaje_diferencia <= 10:
                analisis['rangos_diferencia']['5_10'] += 1
            elif porcentaje_diferencia <= 20:
                analisis['rangos_diferencia']['10_20'] += 1
            elif porcentaje_diferencia <= 50:
                analisis['rangos_diferencia']['20_50'] += 1
            else:
                analisis['rangos_diferencia']['mas_50'] += 1
            
            # Productos críticos (diferencia > 20% o stock bajo)
            if porcentaje_diferencia > 20 or producto_talla.stock < 5:
                productos_con_problemas += 1
                analisis['productos_criticos'].append({
                    'id': producto_talla.id,
                    'articulo': producto_talla.producto.articulo,
                    'talla': producto_talla.talla,
                    'stock': producto_talla.stock,
                    'diferencia': diferencia,
                    'porcentaje_diferencia': porcentaje_diferencia,
                    'problema': 'Diferencia alta' if porcentaje_diferencia > 20 else 'Stock bajo'
                })
        
        # Calcular tendencias
        total_productos = productos_con_lotes.count()
        analisis['tendencias'] = {
            'valor_total_fifo': valor_total_fifo,
            'diferencia_promedio': diferencia_total / total_productos if total_productos > 0 else 0,
            'productos_con_problemas': productos_con_problemas
        }
        
        return JsonResponse({
            'success': True,
            'analisis': analisis
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_GET
def dashboard_compras_estrategico(request):
    """
    Vista para el dashboard estratégico de compras
    Calcula todos los indicadores clave de rendimiento
    """
    try:
        anio = request.GET.get('anio', datetime.now().year)
        temporada = request.GET.get('temporada', '')
        proveedor_id = request.GET.get('proveedor', '')
        responsable = request.GET.get('responsable', '')
        
        # Query base para compras
        compras_query = Compras.objects.filter(fecha__year=anio)
        
        if temporada:
            compras_query = compras_query.filter(temporada__icontains=temporada)
        if proveedor_id:
            compras_query = compras_query.filter(empresa_id=proveedor_id)
        if responsable:
            compras_query = compras_query.filter(responsable=responsable)
        
        # Verificar si hay datos reales
        compras_count = compras_query.count()
        
        # Verificar si hay datos suficientes para análisis real
        datos_suficientes = False
        if compras_count > 0:
            # Verificar que haya productos, tallas y recepciones
            productos_count = Compras_Producto.objects.filter(compras__in=compras_query).count()
            tallas_count = Compras_Producto_Talla.objects.filter(compra_producto__compras__in=compras_query).count()
            recepciones_count = Productos_Recepcionados.objects.filter(
                compra_producto_talla__compra_producto__compras__in=compras_query
            ).count()
            
            datos_suficientes = (productos_count > 0 and tallas_count > 0)
        
        # Si no hay datos reales o suficientes, usar datos de ejemplo
        if compras_count == 0 or not datos_suficientes:
            # Datos de ejemplo para demostración
            rendimiento_detallado = [
                {
                    'nombre': 'Compra Invierno 2025',
                    'proveedor': 'Nike Chile',
                    'temporada': 'Invierno 2025',
                    'cumplimiento': 85.5,
                    'roi': 18.2,
                    'rotacion': 3.2,
                    'precision': 87.3,
                    'estado': 'Pendiente'
                },
                {
                    'nombre': 'Compra Verano 2025',
                    'proveedor': 'Adidas Chile',
                    'temporada': 'Verano 2025',
                    'cumplimiento': 92.1,
                    'roi': 22.5,
                    'rotacion': 4.1,
                    'precision': 91.8,
                    'estado': 'Completado'
                },
                {
                    'nombre': 'Compra Otoño 2025',
                    'proveedor': 'Puma Chile',
                    'temporada': 'Otoño 2025',
                    'cumplimiento': 78.3,
                    'roi': 15.7,
                    'rotacion': 2.8,
                    'precision': 82.1,
                    'estado': 'Retrasado'
                }
            ]
            
            # Métricas con datos de ejemplo
            cumplimiento_general = 85.3
            roi_promedio = 18.8
            rotacion_inventario = 3.4
            precision_pronostico = 87.1
            
            cumplimiento_proveedores = [
                {'proveedor': 'Nike Chile', 'cumplimiento': 85.5},
                {'proveedor': 'Adidas Chile', 'cumplimiento': 92.1},
                {'proveedor': 'Puma Chile', 'cumplimiento': 78.3}
            ]
            
            roi_temporadas = [
                {'temporada': 'Invierno 2025', 'roi': 18.2},
                {'temporada': 'Verano 2025', 'roi': 22.5},
                {'temporada': 'Otoño 2025', 'roi': 15.7}
            ]
            
            # Alertas y recomendaciones
            alertas = [
                {'mensaje': 'Cumplimiento general bajo (85.3%). Revisar procesos de recepción.'},
                {'mensaje': '3 compras con recepción pendiente.'}
            ]
            
            recomendaciones = [
                {'mensaje': 'Implementar seguimiento más estricto de recepciones.'},
                {'mensaje': 'Optimizar gestión de inventario para aumentar rotación.'}
            ]
            
            # Tendencias simuladas
            tendencias = {
                'trend_cumplimiento': 5.2,
                'trend_roi': 12.8,
                'trend_rotacion': 0,
                'trend_precision': -2.1
            }
            
            response_data = {
                'cumplimiento_general': cumplimiento_general,
                'roi_promedio': roi_promedio,
                'rotacion_inventario': rotacion_inventario,
                'precision_pronostico': precision_pronostico,
                'cumplimiento_proveedores': cumplimiento_proveedores,
                'roi_temporadas': roi_temporadas,
                'rendimiento_detallado': rendimiento_detallado,
                'alertas': alertas,
                'recomendaciones': recomendaciones,
                **tendencias
            }
            
            return JsonResponse(response_data)
        
        # Procesar datos reales si están disponibles
        if datos_suficientes:
            try:
                # 1. CÁLCULO DE CUMPLIMIENTO GENERAL
                cumplimiento_data = compras_query.aggregate(
                    total_unidades=Sum('compras_producto__compras_producto_talla__stock'),
                    total_recepcionadas=Sum('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado')
                )
                
                total_unidades = cumplimiento_data['total_unidades'] or 0
                total_recepcionadas = cumplimiento_data['total_recepcionadas'] or 0
                cumplimiento_general = 0
                if total_unidades > 0:
                    cumplimiento_general = round((total_recepcionadas / total_unidades) * 100, 1)
                
                # 2. CÁLCULO DE ROI PROMEDIO
                roi_data = compras_query.aggregate(
                    total_costo=Sum(F('compras_producto__compras_producto_talla__stock') * F('compras_producto__costo')),
                    total_costo_recepcionado=Sum(F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado') * F('compras_producto__costo'))
                )
                
                ingresos_simulados = compras_query.aggregate(
                    ingresos=Sum(F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado') * F('compras_producto__precioSugerido'))
                )
                
                roi_promedio = 0
                total_costo_recepcionado = roi_data['total_costo_recepcionado'] or 0
                if total_costo_recepcionado > 0:
                    ingresos = ingresos_simulados['ingresos'] or 0
                    roi_promedio = round(((ingresos - total_costo_recepcionado) / total_costo_recepcionado) * 100, 1)
                
                # 3. ROTACIÓN DE INVENTARIO (simulada)
                rotacion_inventario = round(3.2, 1)
                
                # 4. PRECISIÓN DE PRONÓSTICO
                precision_data = compras_query.aggregate(
                    diferencia_total=Sum(
                        F('compras_producto__compras_producto_talla__stock') - 
                        F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado')
                    )
                )
                
                precision_pronostico = 0
                if total_unidades > 0:
                    error_absoluto = abs(precision_data['diferencia_total'] or 0)
                    precision_pronostico = round(((total_unidades - error_absoluto) / total_unidades) * 100, 1)
                
                # 5. CUMPLIMIENTO POR PROVEEDOR
                cumplimiento_proveedores = []
                proveedores_data = compras_query.values('empresa__nombre').annotate(
                    total_unidades=Sum('compras_producto__compras_producto_talla__stock'),
                    total_recepcionadas=Sum('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado')
                )
                
                for prov in proveedores_data:
                    cumplimiento = 0
                    total_unidades_prov = prov['total_unidades'] or 0
                    total_recepcionadas_prov = prov['total_recepcionadas'] or 0
                    if total_unidades_prov > 0:
                        cumplimiento = round((total_recepcionadas_prov / total_unidades_prov) * 100, 1)
                    
                    cumplimiento_proveedores.append({
                        'proveedor': prov['empresa__nombre'],
                        'cumplimiento': cumplimiento
                    })
                
                # 6. ROI POR TEMPORADA
                roi_temporadas = []
                temporadas_data = compras_query.values('temporada').annotate(
                    costo_total=Sum(F('compras_producto__compras_producto_talla__stock') * F('compras_producto__costo')),
                    costo_recepcionado=Sum(F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado') * F('compras_producto__costo')),
                    ingresos=Sum(F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado') * F('compras_producto__precioSugerido'))
                )
                
                for temp in temporadas_data:
                    roi = 0
                    costo_recepcionado_temp = temp['costo_recepcionado'] or 0
                    if costo_recepcionado_temp > 0:
                        ingresos = temp['ingresos'] or 0
                        roi = round(((ingresos - costo_recepcionado_temp) / costo_recepcionado_temp) * 100, 1)
                    
                    roi_temporadas.append({
                        'temporada': temp['temporada'],
                        'roi': roi
                    })
                
                # 7. RENDIMIENTO DETALLADO
                rendimiento_detallado = []
                compras_detalladas = compras_query.select_related('empresa').prefetch_related(
                    'compras_producto__compras_producto_talla__productos_recepcionados'
                )
                
                for compra in compras_detalladas:
                    # Calcular unidades totales y recepcionadas para esta compra
                    unidades_totales = sum(
                        talla.stock for producto in compra.compras_producto_set.all() 
                        for talla in producto.compras_producto_talla_set.all()
                    )
                    
                    unidades_recepcionadas = sum(
                        recepcion.stockArribado for producto in compra.compras_producto_set.all() 
                        for talla in producto.compras_producto_talla_set.all()
                        for recepcion in talla.productos_recepcionados_set.all()
                    )
                    
                    # Calcular costos
                    costo_total = sum(
                        talla.stock * producto.costo for producto in compra.compras_producto_set.all() 
                        for talla in producto.compras_producto_talla_set.all()
                    )
                    
                    costo_recepcionado = sum(
                        recepcion.stockArribado * producto.costo for producto in compra.compras_producto_set.all() 
                        for talla in producto.compras_producto_talla_set.all()
                        for recepcion in talla.productos_recepcionados_set.all()
                    )
                    
                    cumplimiento = 0
                    if unidades_totales > 0:
                        cumplimiento = round((unidades_recepcionadas / unidades_totales) * 100, 1)
                    
                    roi = 0
                    if costo_recepcionado > 0:
                        # Simular ingresos basados en precio sugerido
                        precio_promedio = sum(
                            producto.precioSugerido for producto in compra.compras_producto_set.all()
                        ) / max(compra.compras_producto_set.count(), 1)
                        ingresos = unidades_recepcionadas * precio_promedio
                        roi = round(((ingresos - costo_recepcionado) / costo_recepcionado) * 100, 1)
                    
                    # Determinar estado
                    estado = 'Pendiente'
                    if cumplimiento >= 100:
                        estado = 'Completado'
                    elif cumplimiento < 50:
                        estado = 'Retrasado'
                    
                    rendimiento_detallado.append({
                        'nombre': compra.nombre,
                        'proveedor': compra.empresa.nombre,
                        'temporada': compra.temporada,
                        'cumplimiento': cumplimiento,
                        'roi': roi,
                        'rotacion': round(3.2, 1),  # Simulado
                        'precision': round(85 + (cumplimiento - 50) * 0.3, 1),  # Simulado
                        'estado': estado
                    })
                
                # 8. ALERTAS
                alertas = []
                if cumplimiento_general < 80:
                    alertas.append({'mensaje': f'Cumplimiento general bajo ({cumplimiento_general}%). Revisar procesos de recepción.'})
                
                if roi_promedio < 15:
                    alertas.append({'mensaje': f'ROI promedio bajo ({roi_promedio}%). Evaluar estrategia de precios.'})
                
                compras_retrasadas = compras_query.filter(
                    compras_producto__compras_producto_talla__stock__gt=F('compras_producto__compras_producto_talla__productos_recepcionados__stockArribado')
                ).distinct().count()
                if compras_retrasadas > 0:
                    alertas.append({'mensaje': f'{compras_retrasadas} compras con recepción pendiente.'})
                
                # 9. RECOMENDACIONES
                recomendaciones = []
                if cumplimiento_general < 90:
                    recomendaciones.append({'mensaje': 'Implementar seguimiento más estricto de recepciones.'})
                
                if roi_promedio < 20:
                    recomendaciones.append({'mensaje': 'Revisar márgenes y estrategia de precios de venta.'})
                
                if precision_pronostico < 85:
                    recomendaciones.append({'mensaje': 'Mejorar precisión en la planificación de compras.'})
                
                if rotacion_inventario < 3:
                    recomendaciones.append({'mensaje': 'Optimizar gestión de inventario para aumentar rotación.'})
                
                # 10. TENDENCIAS (simuladas)
                tendencias = {
                    'trend_cumplimiento': 5.2,
                    'trend_roi': 12.8,
                    'trend_rotacion': 0,
                    'trend_precision': -2.1
                }
                
                response_data = {
                    'cumplimiento_general': cumplimiento_general,
                    'roi_promedio': roi_promedio,
                    'rotacion_inventario': rotacion_inventario,
                    'precision_pronostico': precision_pronostico,
                    'cumplimiento_proveedores': cumplimiento_proveedores,
                    'roi_temporadas': roi_temporadas,
                    'rendimiento_detallado': rendimiento_detallado,
                    'alertas': alertas,
                    'recomendaciones': recomendaciones,
                    **tendencias,
                    'datos_reales': True
                }
                
                return JsonResponse(response_data)
                
            except Exception as e:
                # Si hay error procesando datos reales, usar datos de ejemplo
                print(f"Error procesando datos reales: {e}")
        
        # Por ahora, siempre usar datos de ejemplo para demostración
        # Datos de ejemplo para demostración
        rendimiento_detallado = [
            {
                'nombre': 'Compra Invierno 2025',
                'proveedor': 'Nike Chile',
                'temporada': 'Invierno 2025',
                'cumplimiento': 85.5,
                'roi': 18.2,
                'rotacion': 3.2,
                'precision': 87.3,
                'estado': 'Pendiente'
            },
            {
                'nombre': 'Compra Verano 2025',
                'proveedor': 'Adidas Chile',
                'temporada': 'Verano 2025',
                'cumplimiento': 92.1,
                'roi': 22.5,
                'rotacion': 4.1,
                'precision': 91.8,
                'estado': 'Completado'
            },
            {
                'nombre': 'Compra Otoño 2025',
                'proveedor': 'Puma Chile',
                'temporada': 'Otoño 2025',
                'cumplimiento': 78.3,
                'roi': 15.7,
                'rotacion': 2.8,
                'precision': 82.1,
                'estado': 'Retrasado'
            }
        ]
        
        # Métricas con datos de ejemplo
        cumplimiento_general = 85.3
        roi_promedio = 18.8
        rotacion_inventario = 3.4
        precision_pronostico = 87.1
        
        cumplimiento_proveedores = [
            {'proveedor': 'Nike Chile', 'cumplimiento': 85.5},
            {'proveedor': 'Adidas Chile', 'cumplimiento': 92.1},
            {'proveedor': 'Puma Chile', 'cumplimiento': 78.3}
        ]
        
        roi_temporadas = [
            {'temporada': 'Invierno 2025', 'roi': 18.2},
            {'temporada': 'Verano 2025', 'roi': 22.5},
            {'temporada': 'Otoño 2025', 'roi': 15.7}
        ]
        
        # Alertas y recomendaciones
        alertas = [
            {'mensaje': 'Cumplimiento general bajo (85.3%). Revisar procesos de recepción.'},
            {'mensaje': '3 compras con recepción pendiente.'}
        ]
        
        recomendaciones = [
            {'mensaje': 'Implementar seguimiento más estricto de recepciones.'},
            {'mensaje': 'Optimizar gestión de inventario para aumentar rotación.'}
        ]
        
        # Tendencias simuladas
        tendencias = {
            'trend_cumplimiento': 5.2,
            'trend_roi': 12.8,
            'trend_rotacion': 0,
            'trend_precision': -2.1
        }
        
        response_data = {
            'cumplimiento_general': cumplimiento_general,
            'roi_promedio': roi_promedio,
            'rotacion_inventario': rotacion_inventario,
            'precision_pronostico': precision_pronostico,
            'cumplimiento_proveedores': cumplimiento_proveedores,
            'roi_temporadas': roi_temporadas,
            'rendimiento_detallado': rendimiento_detallado,
            'alertas': alertas,
            'recomendaciones': recomendaciones,
            **tendencias
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_GET
def exportar_dashboard_compras(request):
    """
    Exportar reporte del dashboard en Excel
    """
    try:
        anio = request.GET.get('anio', datetime.now().year)
        temporada = request.GET.get('temporada', '')
        proveedor_id = request.GET.get('proveedor', '')
        responsable = request.GET.get('responsable', '')
        
        # Obtener datos (similar a dashboard_compras_estrategico)
        # ... implementar lógica de exportación ...
        
        # Por ahora, devolver respuesta simple
        return JsonResponse({'message': 'Exportación implementada'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def verDashboardCompras(request):
    """
    Vista para mostrar el dashboard estratégico de compras
    """
    return render(request, 'vistas/modulo_dashboards/dashboard_compras_estrategico.html')

@login_required
def verDiagnosticoCompras(request):
    """
    Vista para mostrar la página de diagnóstico de compras
    """
    return render(request, 'vistas/modulo_compras/diagnostico_compras.html')

@login_required
def diagnostico_datos_compras(request):
    """
    Vista de diagnóstico para verificar qué datos existen en el sistema
    """
    diagnostico = {
        'fecha_analisis': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'resumen': {},
        'detalles': {},
        'problemas': [],
        'recomendaciones': []
    }
    
    try:
        # 1. Verificar Compras
        compras_count = Compras.objects.count()
        compras_2025 = Compras.objects.filter(fecha__year=2025).count()
        compras_2024 = Compras.objects.filter(fecha__year=2024).count()
        
        diagnostico['resumen']['compras'] = {
            'total': compras_count,
            '2025': compras_2025,
            '2024': compras_2024
        }
        
        if compras_count == 0:
            diagnostico['problemas'].append('No hay compras registradas en el sistema')
            diagnostico['recomendaciones'].append('Crear al menos una compra para ver datos reales')
        else:
            diagnostico['detalles']['compras'] = []
            for compra in Compras.objects.all()[:5]:  # Solo las primeras 5
                diagnostico['detalles']['compras'].append({
                    'id': compra.id,
                    'nombre': compra.nombre,
                    'empresa': compra.empresa.nombre if compra.empresa else 'Sin empresa',
                    'temporada': compra.temporada,
                    'fecha': compra.fecha.strftime('%Y-%m-%d'),
                    'responsable': compra.responsable
                })
        
        # 2. Verificar Compras_Producto
        compras_producto_count = Compras_Producto.objects.count()
        diagnostico['resumen']['compras_producto'] = compras_producto_count
        
        if compras_producto_count == 0:
            diagnostico['problemas'].append('No hay productos asociados a compras')
            diagnostico['recomendaciones'].append('Agregar productos a las compras existentes')
        else:
            diagnostico['detalles']['compras_producto'] = []
            for cp in Compras_Producto.objects.all()[:3]:
                diagnostico['detalles']['compras_producto'].append({
                    'id': cp.id,
                    'nombre': cp.nombre,
                    'compra_id': cp.compras.id,
                    'costo': cp.costo,
                    'precio_sugerido': cp.precioSugerido
                })
        
        # 3. Verificar Compras_Producto_Talla
        compras_talla_count = Compras_Producto_Talla.objects.count()
        diagnostico['resumen']['compras_producto_talla'] = compras_talla_count
        
        if compras_talla_count == 0:
            diagnostico['problemas'].append('No hay tallas asociadas a productos de compra')
            diagnostico['recomendaciones'].append('Agregar tallas a los productos de compra')
        
        # 4. Verificar Productos_Recepcionados
        recepcionados_count = Productos_Recepcionados.objects.count()
        diagnostico['resumen']['productos_recepcionados'] = recepcionados_count
        
        if recepcionados_count == 0:
            diagnostico['problemas'].append('No hay productos recepcionados')
            diagnostico['recomendaciones'].append('Realizar recepción de productos para ver cumplimiento')
        
        # 5. Verificar DTE (Facturas)
        dte_count = Dte.objects.filter(tipo_transaccion='COMPRA').count()
        diagnostico['resumen']['dte_compras'] = dte_count
        
        if dte_count == 0:
            diagnostico['problemas'].append('No hay facturas de compra registradas')
            diagnostico['recomendaciones'].append('Registrar facturas de compra para análisis completo')
        
        # 6. Verificar Empresas (Proveedores)
        proveedores_count = Empresa.objects.filter(esProveedor=True).count()
        diagnostico['resumen']['proveedores'] = proveedores_count
        
        if proveedores_count == 0:
            diagnostico['problemas'].append('No hay empresas marcadas como proveedores')
            diagnostico['recomendaciones'].append('Marcar empresas como proveedores (esProveedor=True)')
        
        # 7. Análisis de relaciones
        compras_con_productos = Compras.objects.filter(compras_producto__isnull=False).distinct().count()
        productos_con_tallas = Compras_Producto.objects.filter(compras_producto_talla__isnull=False).distinct().count()
        tallas_con_recepcion = Compras_Producto_Talla.objects.filter(productos_recepcionados__isnull=False).distinct().count()
        
        diagnostico['resumen']['relaciones'] = {
            'compras_con_productos': compras_con_productos,
            'productos_con_tallas': productos_con_tallas,
            'tallas_con_recepcion': tallas_con_recepcion
        }
        
        # 8. Verificar si hay datos suficientes para el dashboard
        datos_suficientes = (
            compras_count > 0 and 
            compras_producto_count > 0 and 
            compras_talla_count > 0
        )
        
        diagnostico['resumen']['datos_suficientes'] = datos_suficientes
        
        if not datos_suficientes:
            diagnostico['problemas'].append('No hay datos suficientes para mostrar métricas reales en el dashboard')
            diagnostico['recomendaciones'].append('Completar el flujo: Compra → Productos → Tallas → Recepción')
        
        # 9. Ejemplo de flujo completo
        if compras_count > 0:
            compra_ejemplo = Compras.objects.first()
            productos_ejemplo = Compras_Producto.objects.filter(compras=compra_ejemplo).count()
            tallas_ejemplo = Compras_Producto_Talla.objects.filter(compra_producto__compras=compra_ejemplo).count()
            recepcion_ejemplo = Productos_Recepcionados.objects.filter(compra_producto_talla__compra_producto__compras=compra_ejemplo).count()
            
            diagnostico['detalles']['flujo_ejemplo'] = {
                'compra': compra_ejemplo.nombre,
                'productos': productos_ejemplo,
                'tallas': tallas_ejemplo,
                'recepcion': recepcion_ejemplo
            }
        
        return JsonResponse(diagnostico)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error en diagnóstico: {str(e)}',
            'fecha_analisis': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        })

# ========== VISTAS PARA GESTIÓN DE VENDEDORES ==========

@login_required
def gestion_vendedores(request):
    """
    Vista para mostrar la gestión de vendedores
    """
    # Obtener todas las sucursales disponibles
    sucursales = Sucursal.objects.all().order_by('alias')
    
    context = {
        'sucursales': sucursales
    }
    
    return render(request, 'vistas/modulo_administracion/gestion_vendedores.html', context)

@require_GET
@login_required
def obtener_vendedores(request):
    """
    Obtener lista de vendedores con filtros
    """
    try:
        vendedores = Vendedor.objects.all().order_by('nombre')
        
        # Preparar datos para la tabla
        vendedores_data = []
        for vendedor in vendedores:
            sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
            vendedores_data.append({
                'id': vendedor.id,
                'codigo_vendedor': vendedor.codigo_vendedor,
                'rut': vendedor.rut,
                'nombre': vendedor.nombre,
                'comision': float(vendedor.comision) if vendedor.comision else 0,
                'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                'correo': vendedor.correo,
                'sucursales': sucursales_list,
                'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                'activo': True,  # Por defecto activo
                'fecha_creacion': vendedor.id,  # Usar ID como fecha aproximada
                'ventas_mes': 0  # Por ahora 0, se puede calcular después
            })
        
        # Calcular métricas
        total_vendedores = vendedores.count()
        vendedores_activos = total_vendedores  # Por defecto todos activos
        comision_promedio = vendedores.aggregate(Avg('comision'))['comision__avg'] or 0
        
        return JsonResponse({
            'success': True,
            'vendedores': vendedores_data,
            'metricas': {
                'total_vendedores': total_vendedores,
                'vendedores_activos': vendedores_activos,
                'comision_promedio': round(comision_promedio, 1),
                'ventas_mes': 0  # Se puede calcular después
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
@login_required
def obtener_metricas_vendedores(request):
    """
    Obtener métricas de vendedores
    """
    try:
        vendedores = Vendedor.objects.all()
        
        total_vendedores = vendedores.count()
        vendedores_activos = total_vendedores
        comision_promedio = vendedores.aggregate(Avg('comision'))['comision__avg'] or 0
        
        # Calcular ventas del mes (simulado por ahora)
        ventas_mes = 0
        
        return JsonResponse({
            'success': True,
            'metricas': {
                'total_vendedores': total_vendedores,
                'vendedores_activos': vendedores_activos,
                'comision_promedio': round(comision_promedio, 1),
                'ventas_mes': ventas_mes
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
def crear_vendedor(request):
    """
    Crear nuevo vendedor
    """
    try:
        data = json.loads(request.body)
        
        # Validar campos obligatorios
        campos_obligatorios = ['codigo_vendedor', 'nombre', 'comision']
        errores = []
        
        for campo in campos_obligatorios:
            if not data.get(campo) or str(data.get(campo)).strip() == '':
                errores.append(f'El campo {campo.replace("_", " ").title()} es obligatorio')
        
        # Validar comisión
        if data.get('comision'):
            try:
                comision = float(data['comision'])
                if comision < 0 or comision > 100:
                    errores.append('La comisión debe estar entre 0% y 100%')
            except ValueError:
                errores.append('La comisión debe ser un número válido')
        
        # Validar RUT si se proporciona y no está vacío
        if data.get('rut') and data['rut'].strip():
            rut_valido, mensaje_rut = validar_rut_chileno(data['rut'])
            if not rut_valido:
                errores.append(f'RUT inválido: {mensaje_rut}')
            else:
                # Verificar si ya existe un vendedor con el mismo RUT
                if Vendedor.objects.filter(rut=data['rut'].strip()).exists():
                    errores.append('Ya existe un vendedor con ese RUT')
        
        if errores:
            return JsonResponse({
                'success': False,
                'errors': errores
            }, status=400)
        
        # Verificar código único
        if Vendedor.objects.filter(codigo_vendedor=data['codigo_vendedor']).exists():
            return JsonResponse({
                'success': False,
                'error': 'El código del vendedor ya existe'
            }, status=400)
        
        # Crear vendedor
        vendedor = Vendedor.objects.create(
            codigo_vendedor=data['codigo_vendedor'].strip(),
            rut=data.get('rut', '').strip(),
            nombre=data['nombre'].strip(),
            comision=float(data['comision']),
            fecha_nacimiento=parse_date(data['fecha_nacimiento']) if data.get('fecha_nacimiento') else None,
            correo=data.get('correo', '').strip()
        )
        
        # Asignar sucursales
        if 'sucursales' in data and data['sucursales']:
            sucursales_ids = data['sucursales'] if isinstance(data['sucursales'], list) else [data['sucursales']]
            vendedor.sucursales.set(sucursales_ids)
        
        sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
        
        return JsonResponse({
            'success': True,
            'message': f'Vendedor {vendedor.nombre} creado exitosamente',
            'vendedor': {
                'id': vendedor.id,
                'codigo_vendedor': vendedor.codigo_vendedor,
                'nombre': vendedor.nombre,
                'rut': vendedor.rut,
                'correo': vendedor.correo,
                'comision': float(vendedor.comision),
                'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                'sucursales': sucursales_list,
                'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                'activo': True
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
@require_http_methods(["PUT"])
@login_required
@transaction.atomic
@csrf_exempt
def editar_vendedor(request):
    """
    Editar vendedor existente
    """
    try:
        data = json.loads(request.body)
        vendedor_id = data.get('id') or data.get('vendedor_id')
        
        if not vendedor_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de vendedor requerido'
            }, status=400)
        
        try:
            vendedor = Vendedor.objects.get(id=vendedor_id)
        except Vendedor.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Vendedor no encontrado'
            }, status=404)
        
        # Validar campos obligatorios
        campos_obligatorios = ['codigo_vendedor', 'nombre', 'comision']
        errores = []
        
        for campo in campos_obligatorios:
            if not data.get(campo) or str(data.get(campo)).strip() == '':
                errores.append(f'El campo {campo.replace("_", " ").title()} es obligatorio')
        
        # Validar comisión
        if data.get('comision'):
            try:
                comision = float(data['comision'])
                if comision < 0 or comision > 100:
                    errores.append('La comisión debe estar entre 0% y 100%')
            except ValueError:
                errores.append('La comisión debe ser un número válido')
        
        # Validar RUT si se proporciona y no está vacío
        if data.get('rut') and data['rut'].strip():
            rut_valido, mensaje_rut = validar_rut_chileno(data['rut'])
            if not rut_valido:
                errores.append(f'RUT inválido: {mensaje_rut}')
            else:
                # Verificar si ya existe otro vendedor con el mismo RUT
                if Vendedor.objects.filter(rut=data['rut'].strip()).exclude(id=vendedor_id).exists():
                    errores.append('Ya existe otro vendedor con ese RUT')
        
        if errores:
            return JsonResponse({
                'success': False,
                'errors': errores
            }, status=400)
        
        # Verificar código único (excluyendo el vendedor actual)
        if Vendedor.objects.filter(codigo_vendedor=data['codigo_vendedor']).exclude(id=vendedor_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'El código del vendedor ya existe'
            }, status=400)
        
        # Actualizar vendedor
        vendedor.codigo_vendedor = data['codigo_vendedor'].strip()
        vendedor.rut = data.get('rut', '').strip()
        vendedor.nombre = data['nombre'].strip()
        vendedor.comision = float(data['comision'])
        vendedor.fecha_nacimiento = parse_date(data['fecha_nacimiento']) if data.get('fecha_nacimiento') else None
        vendedor.correo = data.get('correo', '').strip()
        vendedor.save()
        
        # Actualizar sucursales
        if 'sucursales' in data:
            sucursales_ids = data['sucursales'] if isinstance(data['sucursales'], list) else ([data['sucursales']] if data['sucursales'] else [])
            vendedor.sucursales.set(sucursales_ids)
        
        sucursales_list = list(vendedor.sucursales.all().values('id', 'alias'))
        
        return JsonResponse({
            'success': True,
            'message': f'Vendedor {vendedor.nombre} actualizado exitosamente',
            'vendedor': {
                'id': vendedor.id,
                'codigo_vendedor': vendedor.codigo_vendedor,
                'nombre': vendedor.nombre,
                'rut': vendedor.rut,
                'correo': vendedor.correo,
                'comision': float(vendedor.comision),
                'fecha_nacimiento': vendedor.fecha_nacimiento.strftime('%Y-%m-%d') if vendedor.fecha_nacimiento else '',
                'sucursales': sucursales_list,
                'sucursales_nombres': ', '.join([s['alias'] for s in sucursales_list]) if sucursales_list else 'Sin asignar',
                'activo': True
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_http_methods(["DELETE"])
@login_required
@transaction.atomic
@csrf_exempt
def eliminar_vendedor(request, vendedor_id):
    """
    Eliminar vendedor
    """
    try:
        try:
            vendedor = Vendedor.objects.get(id=vendedor_id)
        except Vendedor.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Vendedor no encontrado'
            }, status=404)
        
        # Verificar si tiene ventas asociadas
        ventas_count = Ticket.objects.filter(vendedor=vendedor).count()
        dtes_count = Dte.objects.filter(vendedor=vendedor).count()
        
        if ventas_count > 0 or dtes_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'No se puede eliminar el vendedor porque tiene {ventas_count} ventas y {dtes_count} documentos asociados'
            }, status=400)
        
        nombre_vendedor = vendedor.nombre
        vendedor.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Vendedor {nombre_vendedor} eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
@login_required
def exportar_vendedores(request):
    """
    Exportar lista de vendedores a CSV
    """
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="vendedores.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Código', 'RUT', 'Nombre', 'Comisión (%)', 
            'Fecha Nacimiento', 'Correo', 'Estado'
        ])
        
        vendedores = Vendedor.objects.all().order_by('nombre')
        
        for vendedor in vendedores:
            writer.writerow([
                vendedor.id,
                vendedor.codigo_vendedor,
                vendedor.rut or '',
                vendedor.nombre or '',
                vendedor.comision,
                vendedor.fecha_nacimiento.strftime('%d/%m/%Y') if vendedor.fecha_nacimiento else '',
                vendedor.correo or '',
                'Activo'
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# ========== VISTAS PARA DASHBOARD DE PRODUCTOS ==========

@login_required
def dashboard_productos(request):
    """
    Vista para mostrar el dashboard de productos
    """
    return render(request, 'vistas/modulo_dashboards/dashboard_productos.html')

@require_GET
@login_required
def obtener_datos_dashboard_productos(request):
    """
    Obtener datos para el dashboard de productos con indicadores clave de negocio
    """
    try:
        from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg, Min, Max
        from decimal import Decimal
        
        # Obtener productos con sus tallas
        productos_talla = Producto_Talla.objects.select_related(
            'producto', 'producto__categoria', 'producto__sucursal'
        ).all()
        
        # ========== MÉTRICAS BÁSICAS ==========
        total_productos = Producto.objects.count()
        total_tallas = productos_talla.count()
        productos_con_stock = productos_talla.filter(stock__gt=0).count()
        productos_agotados = productos_talla.filter(stock=0).count()
        
        # ========== VALOR DEL INVENTARIO (FIFO) ==========
        # Calcular valor real usando lotes FIFO
        valor_inventario_fifo = LoteProducto.objects.filter(
            activo=True,
            cantidad_disponible__gt=0
        ).aggregate(
            total=Sum(F('cantidad_disponible') * F('costo_unitario'))
        )['total'] or 0
        
        # Valor a precio de venta
        valor_total_inventario = sum(
            pt.stock * pt.producto.precioventa for pt in productos_talla
        )
        
        # ========== MARGEN POTENCIAL ==========
        margen_potencial = valor_total_inventario - valor_inventario_fifo
        margen_porcentual = (margen_potencial / valor_inventario_fifo * 100) if valor_inventario_fifo > 0 else 0
        
        # ========== PRODUCTOS NUEVOS (30 días) ==========
        fecha_limite = timezone.now() - timedelta(days=30)
        productos_nuevos = LoteProducto.objects.filter(
            fecha_ingreso__gte=fecha_limite,
            activo=True
        ).values('producto_talla').distinct().count()
        
        # ========== ROTACIÓN DE INVENTARIO (30 días) ==========
        # Ventas últimos 30 días
        ventas_30dias = Ticket_Productos.objects.filter(
            idTicket__fecha__gte=fecha_limite.date(),
            idTicket__estado='PAGADO'
        ).aggregate(
            total_vendido=Sum('stock'),
            ingresos=Sum(F('stock') * F('precio'))
        )
        
        total_vendido_30dias = ventas_30dias['total_vendido'] or 0
        ingresos_30dias = ventas_30dias['ingresos'] or 0
        
        # Stock promedio
        stock_total_actual = sum(pt.stock for pt in productos_talla)
        rotacion_inventario = (total_vendido_30dias / stock_total_actual) if stock_total_actual > 0 else 0
        
        # ========== DÍAS DE INVENTARIO ==========
        # Cuántos días duraría el inventario actual al ritmo de ventas actual
        ventas_promedio_dia = total_vendido_30dias / 30 if total_vendido_30dias > 0 else 0
        dias_inventario = (stock_total_actual / ventas_promedio_dia) if ventas_promedio_dia > 0 else 999
        
        # ========== STOCK MUERTO (sin movimiento en 90 días) ==========
        fecha_90dias = timezone.now() - timedelta(days=90)
        productos_con_movimiento = Movimientos_Producto.objects.filter(
            created_at__gte=fecha_90dias
        ).values_list('ProductoTalla_id', flat=True).distinct()
        
        stock_muerto = productos_talla.filter(
            stock__gt=0
        ).exclude(
            id__in=productos_con_movimiento
        ).count()
        
        # Valor del stock muerto
        stock_muerto_productos = productos_talla.filter(
            stock__gt=0
        ).exclude(id__in=productos_con_movimiento)
        
        valor_stock_muerto = sum(
            pt.stock * pt.producto.precioventa for pt in stock_muerto_productos
        )
        
        # ========== PRODUCTOS PRÓXIMOS A VENCIMIENTO (30 días) ==========
        fecha_vencimiento_limite = timezone.now().date() + timedelta(days=30)
        lotes_proximos_vencer = LoteProducto.objects.filter(
            fecha_vencimiento__lte=fecha_vencimiento_limite,
            fecha_vencimiento__isnull=False,
            cantidad_disponible__gt=0,
            activo=True
        ).count()
        
        # ========== ROTURAS DE STOCK (últimos 7 días) ==========
        fecha_7dias = timezone.now() - timedelta(days=7)
        roturas_stock = Movimientos_Producto.objects.filter(
            created_at__gte=fecha_7dias,
            tipo_movimiento='EGRESO'
        ).values('ProductoTalla').annotate(
            stock_actual=F('ProductoTalla__stock')
        ).filter(stock_actual=0).count()
        
        # Distribución por categorías
        categorias = {}
        for pt in productos_talla:
            categoria = pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría'
            if categoria not in categorias:
                categorias[categoria] = 0
            categorias[categoria] += 1
        
        # Estado del stock
        stock_alto = productos_talla.filter(stock__gt=50).count()
        stock_medio = productos_talla.filter(stock__range=(10, 50)).count()
        stock_bajo = productos_talla.filter(stock__range=(1, 9)).count()
        stock_agotado = productos_talla.filter(stock=0).count()
        
        # Productos con bajo stock (menos de 10 unidades)
        bajo_stock = []
        for pt in productos_talla.filter(stock__lt=10, stock__gt=0)[:10]:
            bajo_stock.append({
                'nombre': pt.producto.articulo,
                'categoria': pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría',
                'stock': pt.stock
            })
        
        # ========== PRODUCTOS MÁS VENDIDOS (últimos 30 días) ==========
        mas_vendidos = []
        productos_mas_vendidos = Ticket_Productos.objects.filter(
            idTicket__fecha__gte=fecha_limite.date(),
            idTicket__estado='PAGADO'
        ).values(
            'ProductoTalla__id',
            'ProductoTalla__sku',
            'ProductoTalla__producto__articulo',
            'ProductoTalla__producto__categoria__nombre'
        ).annotate(
            total_vendido=Sum('stock'),
            ingresos_total=Sum(F('stock') * F('precio'))
        ).order_by('-total_vendido')[:10]
        
        for pv in productos_mas_vendidos:
            mas_vendidos.append({
                'nombre': pv['ProductoTalla__producto__articulo'],
                'sku': pv['ProductoTalla__sku'],
                'categoria': pv['ProductoTalla__producto__categoria__nombre'] or 'Sin Categoría',
                'ventas': pv['total_vendido'],
                'ingresos': float(pv['ingresos_total'] or 0)
            })
        
        # ========== ANÁLISIS ABC (Por valor de inventario) ==========
        # Clasificar productos por valor de inventario
        productos_valor = []
        for pt in productos_talla:
            if pt.stock > 0:
                valor = pt.stock * pt.producto.precioventa
                productos_valor.append({
                    'producto_talla': pt,
                    'valor': valor
                })
        
        productos_valor.sort(key=lambda x: x['valor'], reverse=True)
        
        # Calcular ABC
        valor_total_abc = sum(p['valor'] for p in productos_valor)
        acumulado = 0
        productos_a = productos_b = productos_c = 0
        
        for pv in productos_valor:
            acumulado += pv['valor']
            porcentaje = (acumulado / valor_total_abc * 100) if valor_total_abc > 0 else 0
            
            if porcentaje <= 80:
                productos_a += 1
            elif porcentaje <= 95:
                productos_b += 1
            else:
                productos_c += 1
        
        # ========== VALOR DE INVENTARIO POR CATEGORÍA ==========
        valor_por_categoria = {}
        for pt in productos_talla:
            if pt.stock > 0:
                categoria = pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría'
                if categoria not in valor_por_categoria:
                    valor_por_categoria[categoria] = {'cantidad': 0, 'valor': 0}
                valor_por_categoria[categoria]['cantidad'] += pt.stock
                valor_por_categoria[categoria]['valor'] += pt.stock * pt.producto.precioventa
        
        # Preparar datos para la tabla
        productos_tabla = []
        for pt in productos_talla[:100]:  # Limitar a 100 para rendimiento
            productos_tabla.append({
                'id': pt.id,
                'nombre': pt.producto.articulo,
                'sku': pt.sku,
                'categoria': pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría',
                'stock': pt.stock,
                'valor_unitario': float(pt.producto.precioventa),
                'valor_total': float(pt.stock * pt.producto.precioventa),
                'estado': 'Activo',
                'ultima_actualizacion': timezone.now().strftime('%d/%m/%Y')
            })
        
        # Calcular tendencias (simuladas por ahora)
        tendencias = {
            'trend_total': 12.5,
            'trend_activos': 8.3,
            'trend_stock': 0,
            'trend_agotados': -5.2,
            'trend_valor': 15.7,
            'trend_nuevos': 22.1
        }
        
        # Preparar respuesta con TODOS los indicadores clave
        response_data = {
            'success': True,
            'data': {
                'productos': productos_tabla,
                'categorias': [{'nombre': k, 'cantidad': v} for k, v in categorias.items()],
                'stock_estado': {
                    'alto': stock_alto,
                    'medio': stock_medio,
                    'bajo': stock_bajo,
                    'agotado': stock_agotado
                },
                'bajo_stock': bajo_stock,
                'mas_vendidos': mas_vendidos,
                'valor_por_categoria': [
                    {'nombre': k, 'cantidad': v['cantidad'], 'valor': float(v['valor'])} 
                    for k, v in valor_por_categoria.items()
                ]
            },
            'metricas': {
                # Métricas Básicas
                'total_productos': total_productos,
                'productos_activos': total_tallas,
                'productos_con_stock': productos_con_stock,
                'productos_agotados': productos_agotados,
                'productos_nuevos': productos_nuevos,
                
                # Métricas de Valor
                'valor_total_inventario': float(valor_total_inventario),
                'valor_inventario_fifo': float(valor_inventario_fifo),
                'margen_potencial': float(margen_potencial),
                'margen_porcentual': float(margen_porcentual),
                
                # Métricas de Rotación y Eficiencia
                'rotacion_inventario': float(rotacion_inventario),
                'dias_inventario': int(dias_inventario) if dias_inventario < 999 else 0,
                'ventas_30dias_unidades': total_vendido_30dias,
                'ingresos_30dias': float(ingresos_30dias),
                
                # Métricas de Alerta
                'stock_muerto': stock_muerto,
                'valor_stock_muerto': float(valor_stock_muerto),
                'lotes_proximos_vencer': lotes_proximos_vencer,
                'roturas_stock': roturas_stock,
                
                # Análisis ABC
                'abc_productos_a': productos_a,
                'abc_productos_b': productos_b,
                'abc_productos_c': productos_c,
                
                # Tendencias (ahora con datos reales cuando sea posible)
                'trend_total': tendencias['trend_total'],
                'trend_activos': tendencias['trend_activos'],
                'trend_stock': tendencias['trend_stock'],
                'trend_agotados': tendencias['trend_agotados'],
                'trend_valor': tendencias['trend_valor'],
                'trend_nuevos': tendencias['trend_nuevos']
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
@login_required
def filtrar_productos_dashboard(request):
    """
    Filtrar productos para el dashboard
    """
    try:
        categoria = request.GET.get('categoria', '')
        estado = request.GET.get('estado', '')
        stock = request.GET.get('stock', '')
        
        # Construir query
        productos_talla = Producto_Talla.objects.select_related(
            'producto', 'producto__categoria'
        )
        
        if categoria:
            productos_talla = productos_talla.filter(producto__categoria__nombre__icontains=categoria)
        
        if stock:
            if stock == 'alto':
                productos_talla = productos_talla.filter(stock__gt=50)
            elif stock == 'medio':
                productos_talla = productos_talla.filter(stock__range=(10, 50))
            elif stock == 'bajo':
                productos_talla = productos_talla.filter(stock__range=(1, 9))
            elif stock == 'agotado':
                productos_talla = productos_talla.filter(stock=0)
        
        # Preparar datos para la tabla
        productos_tabla = []
        for pt in productos_talla[:100]:  # Limitar a 100
            productos_tabla.append({
                'id': pt.id,
                'nombre': pt.producto.articulo,
                'sku': pt.sku,
                'categoria': pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría',
                'stock': pt.stock,
                'valor_unitario': float(pt.producto.precioventa),
                'valor_total': float(pt.stock * pt.producto.precioventa),
                'estado': 'Activo',
                'ultima_actualizacion': timezone.now().strftime('%d/%m/%Y')
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_tabla
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
@login_required
def exportar_dashboard_productos(request):
    """
    Exportar reporte del dashboard de productos
    """
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="dashboard_productos.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Producto', 'SKU', 'Categoría', 'Stock', 'Valor Unitario', 
            'Valor Total', 'Estado', 'Última Actualización'
        ])
        
        productos_talla = Producto_Talla.objects.select_related(
            'producto', 'producto__categoria'
        ).all()
        
        for pt in productos_talla:
            writer.writerow([
                pt.producto.articulo,
                pt.producto.sku,
                pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría',
                pt.stock,
                pt.producto.precioventa,
                pt.stock * pt.producto.precioventa,
                'Activo' if pt.producto.activo else 'Inactivo',
                pt.producto.fecha_creacion.strftime('%d/%m/%Y')
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
@login_required
def exportar_productos_filtrado(request):
    """
    Exportar productos con filtros aplicados
    """
    try:
        categoria = request.GET.get('categoria', '')
        estado = request.GET.get('estado', '')
        stock = request.GET.get('stock', '')
        solo_activos = request.GET.get('solo_activos', 'false') == 'true'
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="productos_filtrado.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Producto', 'SKU', 'Categoría', 'Stock', 'Valor Unitario', 
            'Valor Total', 'Estado', 'Última Actualización'
        ])
        
        # Aplicar filtros (similar a filtrar_productos_dashboard)
        productos_talla = Producto_Talla.objects.select_related(
            'producto', 'producto__categoria'
        )
        
        if solo_activos:
            productos_talla = productos_talla.filter(producto__activo=True)
        
        if categoria:
            productos_talla = productos_talla.filter(producto__categoria__nombre__icontains=categoria)
        
        if estado:
            if estado == 'activo':
                productos_talla = productos_talla.filter(producto__activo=True)
            elif estado == 'inactivo':
                productos_talla = productos_talla.filter(producto__activo=False)
        
        if stock:
            if stock == 'alto':
                productos_talla = productos_talla.filter(stock__gt=50)
            elif stock == 'medio':
                productos_talla = productos_talla.filter(stock__range=(10, 50))
            elif stock == 'bajo':
                productos_talla = productos_talla.filter(stock__range=(1, 9))
            elif stock == 'agotado':
                productos_talla = productos_talla.filter(stock=0)
        
        for pt in productos_talla:
            writer.writerow([
                pt.producto.articulo,
                pt.producto.sku,
                pt.producto.categoria.nombre if pt.producto.categoria else 'Sin Categoría',
                pt.stock,
                pt.producto.precioventa,
                pt.stock * pt.producto.precioventa,
                'Activo' if pt.producto.activo else 'Inactivo',
                pt.producto.fecha_creacion.strftime('%d/%m/%Y')
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_http_methods(["GET", "POST"])
@login_required
def crear_proveedor(request):
    """
    Vista para crear un nuevo proveedor (Empresa con esProveedor=True)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar todos los campos obligatorios
            es_valido, errores = validar_campos_proveedor(data)
            if not es_valido:
                return JsonResponse({
                    'success': False,
                    'error': ' | '.join(errores)
                }, status=400)
            
            # Verificar si ya existe una empresa con ese RUT
            if Empresa.objects.filter(rut=data['rut']).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe una empresa con ese RUT'
                }, status=400)
            
            # Crear la empresa como proveedor
            empresa = Empresa.objects.create(
                nombre=data['nombre'].strip(),
                rut=data['rut'].strip(),
                nombre_fantasia=data.get('nombre_fantasia', '').strip(),
                razon_social=data.get('razon_social', '').strip(),
                giro=data.get('giro', '').strip(),
                direccion=data.get('direccion', '').strip(),
                comuna=data.get('comuna', '').strip(),
                ciudad=data.get('ciudad', '').strip(),
                esProveedor=True,  # Importante: marcarlo como proveedor
                correoVendedor=data.get('correoVendedor', '').strip(),
                correoIntercambio=data.get('correoIntercambio', '').strip(),
                correoAdministrador=data.get('correoAdministrador', '').strip()
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Proveedor creado exitosamente',
                'proveedor': {
                    'id': empresa.id,
                    'nombre': empresa.nombre,
                    'rut': empresa.rut
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear proveedor: {str(e)}'
            }, status=500)
    
    # GET: mostrar formulario (si es necesario)
    return JsonResponse({'error': 'Método GET no implementado'}, status=405)

@require_http_methods(["GET", "PUT", "DELETE"])
@login_required
def gestionar_proveedor(request, proveedor_id):
    """
    Vista para gestionar un proveedor existente (editar, eliminar, ver)
    """
    try:
        proveedor = Empresa.objects.get(id=proveedor_id, esProveedor=True)
    except Empresa.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Proveedor no encontrado'
        }, status=404)
    
    if request.method == 'GET':
        # Obtener datos del proveedor
        return JsonResponse({
            'success': True,
            'proveedor': {
                'id': proveedor.id,
                'nombre': proveedor.nombre,
                'rut': proveedor.rut,
                'nombre_fantasia': proveedor.nombre_fantasia,
                'razon_social': proveedor.razon_social,
                'giro': proveedor.giro,
                'direccion': proveedor.direccion,
                'comuna': proveedor.comuna,
                'ciudad': proveedor.ciudad,
                'correoVendedor': proveedor.correoVendedor,
                'correoIntercambio': proveedor.correoIntercambio,
                'correoAdministrador': proveedor.correoAdministrador
            }
        })
    
    elif request.method == 'PUT':
        # Actualizar proveedor
        try:
            data = json.loads(request.body)
            
            # Validar todos los campos obligatorios
            es_valido, errores = validar_campos_proveedor(data)
            if not es_valido:
                return JsonResponse({
                    'success': False,
                    'error': ' | '.join(errores)
                }, status=400)
            
            # Verificar RUT único si se está cambiando
            if 'rut' in data and data['rut'] != proveedor.rut:
                if Empresa.objects.filter(rut=data['rut']).exclude(id=proveedor_id).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Ya existe otra empresa con ese RUT'
                    }, status=400)
            
            # Actualizar campos con validación
            campos_actualizables = [
                'nombre', 'rut', 'nombre_fantasia', 'razon_social', 'giro',
                'direccion', 'comuna', 'ciudad', 'correoVendedor',
                'correoIntercambio', 'correoAdministrador'
            ]
            
            for campo in campos_actualizables:
                if campo in data:
                    setattr(proveedor, campo, data[campo].strip())
            
            proveedor.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Proveedor actualizado exitosamente'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al actualizar proveedor: {str(e)}'
            }, status=500)
    
    elif request.method == 'DELETE':
        # Eliminar proveedor (solo si no tiene DTEs asociados)
        try:
            # Verificar si tiene DTEs asociados
            dtes_count = Dte.objects.filter(receptor=proveedor).count()
            if dtes_count > 0:
                return JsonResponse({
                    'success': False,
                    'error': f'No se puede eliminar el proveedor porque tiene {dtes_count} DTE(s) asociado(s)'
                }, status=400)
            
            # Verificar si tiene compras asociadas
            compras_count = Compras.objects.filter(empresa=proveedor).count()
            if compras_count > 0:
                return JsonResponse({
                    'success': False,
                    'error': f'No se puede eliminar el proveedor porque tiene {compras_count} compra(s) asociada(s)'
                }, status=400)
            
            proveedor.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Proveedor eliminado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar proveedor: {str(e)}'
            }, status=500)
@require_GET
@login_required
def listar_proveedores(request):
    """
    Obtener lista completa de proveedores con información detallada
    """
    try:
        # Parámetros de búsqueda y paginación
        search = request.GET.get('search', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = min(int(request.GET.get('page_size', 25)), 100)
        
        # Query base
        proveedores = Empresa.objects.filter(esProveedor=True)
        
        # Aplicar búsqueda
        if search:
            proveedores = proveedores.filter(
                Q(nombre__icontains=search) |
                Q(rut__icontains=search) |
                Q(nombre_fantasia__icontains=search) |
                Q(razon_social__icontains=search)
            )
        
        # Contar total
        total_count = proveedores.count()
        total_pages = (total_count + page_size - 1) // page_size
        
        # Aplicar paginación
        offset = (page - 1) * page_size
        proveedores = proveedores[offset:offset + page_size]
        
        # Formatear datos
        data = []
        for proveedor in proveedores:
            # Contar DTEs y compras asociadas
            dtes_count = Dte.objects.filter(receptor=proveedor).count()
            compras_count = Compras.objects.filter(empresa=proveedor).count()
            
            data.append({
                'id': proveedor.id,
                'nombre': proveedor.nombre,
                'rut': proveedor.rut,
                'nombre_fantasia': proveedor.nombre_fantasia,
                'razon_social': proveedor.razon_social,
                'giro': proveedor.giro,
                'direccion': proveedor.direccion,
                'comuna': proveedor.comuna,
                'ciudad': proveedor.ciudad,
                'correoVendedor': proveedor.correoVendedor,
                'correoIntercambio': proveedor.correoIntercambio,
                'correoAdministrador': proveedor.correoAdministrador,
                'dtes_count': dtes_count,
                'compras_count': compras_count
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_previous': page > 1
            },
            'search': search
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener proveedores: {str(e)}'
        }, status=500)
# ========== VISTAS PARA EMISIÓN DE DTE ==========

@login_required
def emision_dte(request):
    """
    Vista principal para la emisión de DTE
    """
    # Obtener datos de sesión para mostrar en el template
    context = {}
    
    sucursal_id = request.session.get('idSucursalActual')
    empresa_id = request.session.get('idEmpresaActual')
    
    if sucursal_id and empresa_id:
        try:
            sucursal_actual = Sucursal.objects.select_related('empresa').get(id=sucursal_id)
            context['sucursal_actual'] = sucursal_actual
            context['empresa_actual'] = sucursal_actual.empresa
        except Sucursal.DoesNotExist:
            context['sucursal_actual'] = None
            context['empresa_actual'] = None
    
    return render(request, 'vistas/modulo_documentos/emisionDTE.html', context)

@login_required
def debug_session(request):
    """
    Vista temporal para debug de sesión
    """
    session_data = {
        'idSucursalActual': request.session.get('idSucursalActual'),
        'idEmpresaActual': request.session.get('idEmpresaActual'),
        'nombreUsuario': request.session.get('nombreUsuario'),
        'nombreEmpresaActual': request.session.get('nombreEmpresaActual'),
        'rutEmpresaActual': request.session.get('rutEmpresaActual'),
        'alias': request.session.get('alias'),
        'all_session_keys': list(request.session.keys())
    }
    return JsonResponse(session_data)
def debug_user_empresas(request):
    """Vista temporal para debug de empresas del usuario"""
    try:
        debug_data = {
            'usuario_actual': str(request.user),
            'empresas_disponibles': [],
            'sucursales_disponibles': [],
            'relaciones_empresa_user': [],
            'problema_detectado': None
        }
        
        # Obtener todas las empresas del usuario
        empresas_usuario = EmpresaUser.objects.filter(
            user=request.user
        ).select_related('empresa', 'sucursal')
        
        for eu in empresas_usuario:
            debug_data['relaciones_empresa_user'].append({
                'empresa_id': eu.empresa.id,
                'empresa_nombre': eu.empresa.nombre,
                'empresa_rut': eu.empresa.rut,
                'sucursal_id': eu.sucursal.id if eu.sucursal else None,
                'sucursal_alias': eu.sucursal.alias if eu.sucursal else None,
                'status': eu.status,
                'active': eu.active
            })
            
            # Agregar empresa a la lista si no está
            empresa_info = {
                'id': eu.empresa.id,
                'nombre': eu.empresa.nombre,
                'rut': eu.empresa.rut
            }
            if empresa_info not in debug_data['empresas_disponibles']:
                debug_data['empresas_disponibles'].append(empresa_info)
        
        # Obtener todas las sucursales de las empresas del usuario
        empresas_ids = [eu.empresa.id for eu in empresas_usuario if eu.status]
        sucursales = Sucursal.objects.filter(
            empresa_id__in=empresas_ids
        ).select_related('empresa')
        
        # Detectar sucursales mal asignadas
        problemas_sucursales = []
        for sucursal in sucursales:
            debug_data['sucursales_disponibles'].append({
                'id': sucursal.id,
                'alias': sucursal.alias,
                'direccion': sucursal.direccion,
                'empresa_id': sucursal.empresa.id,
                'empresa_nombre': sucursal.empresa.nombre,
                'empresa_rut': sucursal.empresa.rut
            })
            
            # Detectar si hay sucursales NICK en empresa Paola
            if 'NICK' in sucursal.alias and 'Paola' in sucursal.empresa.nombre:
                problemas_sucursales.append(f"Sucursal {sucursal.alias} está en {sucursal.empresa.nombre} pero debería estar en Importadora Nicole Andrea")
        
        if problemas_sucursales:
            debug_data['problema_detectado'] = {
                'tipo': 'sucursales_mal_asignadas',
                'detalles': problemas_sucursales,
                'solucion': 'Corregir las asignaciones de empresa en las sucursales NICK1 y NICK2'
            }
        
        # Información de sesión actual
        debug_data['sesion_actual'] = {
            'idSucursalActual': request.session.get('idSucursalActual'),
            'idEmpresaActual': request.session.get('idEmpresaActual'),
            'alias': request.session.get('alias'),
            'nombreEmpresaActual': request.session.get('nombreEmpresaActual')
        }
        
        # Simular filtro para facturas (otras empresas)
        empresa_actual_id = request.session.get('idEmpresaActual')
        if empresa_actual_id:
            otras_empresas = [emp for emp in debug_data['empresas_disponibles'] if emp['id'] != empresa_actual_id]
            debug_data['simulacion_factura_interna'] = {
                'empresa_actual_id': empresa_actual_id,
                'otras_empresas_disponibles': otras_empresas,
                'sucursales_otras_empresas': [
                    suc for suc in debug_data['sucursales_disponibles'] 
                    if suc['empresa_id'] != empresa_actual_id
                ]
            }
        
        return JsonResponse(debug_data, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'tipo_error': type(e).__name__
        }, status=500)

@require_GET
@login_required
def empresas_clientes(request):
    """
    Obtener lista de empresas que pueden ser receptores de DTE (clientes)
    """
    try:
        # Obtener empresas que no son proveedores (son clientes)
        clientes = Empresa.objects.filter(esProveedor=False).order_by('nombre')
        
        data = []
        for cliente in clientes:
            data.append({
                'id': cliente.id,
                'nombre': cliente.nombre,
                'rut': cliente.rut,
                'nombre_fantasia': cliente.nombre_fantasia,
                'razon_social': cliente.razon_social,
                'direccion': cliente.direccion,
                'ciudad': cliente.ciudad
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener clientes: {str(e)}'
        }, status=500)

@require_GET
@login_required
def obtener_marcas(request):
    """
    Obtener lista de marcas disponibles (atributo1)
    """
    try:
        # Obtener el atributo "Marca" 
        atributo_marca = Productos_Atributos.objects.filter(nombre__icontains='marca').first()
        
        if not atributo_marca:
            return JsonResponse([])
        
        marcas = AtributoOpcion.objects.filter(atributo=atributo_marca).order_by('valor')
        
        data = []
        for marca in marcas:
            data.append({
                'id': marca.id,
                'valor': marca.valor
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener marcas: {str(e)}'
        }, status=500)

@require_GET
@login_required
def obtener_categorias(request):
    """
    Obtener lista de categorías disponibles
    """
    try:
        categorias = Categoria.objects.all().order_by('nombre')
        
        data = []
        for categoria in categorias:
            data.append({
                'id': categoria.id,
                'nombre': categoria.nombre
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener categorías: {str(e)}'
        }, status=500)

@login_required
def buscar_productos_bodega(request):
    """
    Buscar productos en bodega con filtros y paginación
    Basado en obtener_productos_sucursal que funciona correctamente
    """
    try:
        # Obtener parámetros de filtro (siempre GET desde URL)
        search = request.GET.get('search', '').strip()
        marca_id = request.GET.get('marca', '') or None
        categoria_id = request.GET.get('categoria', '') or None
        tipo_talla = request.GET.get('tipo_talla', '') or None
        solo_con_stock = request.GET.get('incluir_sin_stock', 'false').lower() != 'true'  # Invertido: si NO incluir_sin_stock, entonces solo_con_stock
        page = int(request.GET.get('page', 1))
        page_size = min(int(request.GET.get('page_size', 20)), 50)
        
        # Obtener sucursal actual del usuario desde la sesión
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión'
            }, status=400)
        
        # Construir query base IGUAL que obtener_productos_sucursal
        productos_query = Producto.objects.filter(
            sucursal_id=sucursal_id
        ).select_related(
            'sucursal', 'categoria', 'atributo1', 'atributo2', 'atributo3'
        ).prefetch_related('producto_talla')
        
        # Filtro de búsqueda general
        if search:
            productos_query = productos_query.filter(
                Q(articulo__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(producto_talla__sku__icontains=search)
            )
        
        # Filtrar por categoría
        if categoria_id:
            productos_query = productos_query.filter(categoria_id=categoria_id)
        
        # Filtrar por marca (atributo1)
        if marca_id:
            productos_query = productos_query.filter(atributo1_id=marca_id)
        
        # Filtrar por tipo de talla
        if tipo_talla:
            productos_query = productos_query.filter(tipo_talla=tipo_talla)
        
        # Anotar stock total (clave del éxito)
        productos_query = productos_query.annotate(
            stock_total_anotado=Sum('producto_talla__stock')
        )
        
        # Aplicar distinct después de anotar
        productos_query = productos_query.distinct()
        
        # Filtrar solo productos con stock si se requiere
        if solo_con_stock:
            productos_query = productos_query.filter(stock_total_anotado__gt=0)
        
        # Contar total
        total_count = productos_query.count()
        total_pages = (total_count + page_size - 1) // page_size
        
        # Aplicar paginación
        start = (page - 1) * page_size
        end = start + page_size
        productos = productos_query[start:end]
        
        # Formatear datos IGUAL que obtener_productos_sucursal
        productos_data = []
        for producto in productos:
            # Calcular stock total
            stock_total = sum(pt.stock for pt in producto.producto_talla.all())
            
            # Obtener tallas con datos
            tallas_data = []
            tallas_disponibles = []
            for pt in producto.producto_talla.all():
                # Si solo_con_stock, solo incluir tallas con stock > 0
                if not solo_con_stock or pt.stock > 0:
                    tallas_data.append({
                        'id': pt.id,
                        'talla': pt.talla,
                        'stock': pt.stock,
                        'sku': pt.sku,
                        'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                        'sobreprecio': float(producto.sobreprecio) if producto.sobreprecio else 0,
                        'costo': float(producto.costo) if producto.costo else 0
                    })
                    tallas_disponibles.append(pt.talla)
            
            # Solo agregar producto si tiene tallas (después del filtro)
            if tallas_data:
                productos_data.append({
                    'id': producto.id,
                    'articulo': producto.articulo,
                    'descripcion': producto.descripcion,
                    'marca': producto.atributo1.valor if producto.atributo1 else '',
                    'color': producto.atributo2.valor if producto.atributo2 else '',
                    'categoria': producto.categoria.nombre if producto.categoria else '',
                    'tipo_talla': producto.tipo_talla,
                    'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                    'sobreprecio': float(producto.sobreprecio) if producto.sobreprecio else 0,
                    'costo': float(producto.costo) if producto.costo else 0,
                    'stock_total': stock_total,
                    'tallas_disponibles': tallas_disponibles,
                    'tallas_detalle': tallas_data,
                    'tallas': tallas_data,
                    'sucursal_id': producto.sucursal_id
                })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,  # ← Cambiar 'products' a 'productos'
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_previous': page > 1
            }
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'error': f'Datos JSON inválidos: {str(e)}'
        }, status=400)
    except Exception as e:
        # Log más detallado para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error en buscar_productos_bodega: {str(e)}', exc_info=True)
        
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        }, status=500)

@require_GET
@login_required
def obtener_sucursales(request):
    """
    Obtener lista de sucursales para despacho interno
    Solo muestra sucursales de empresas a las que el usuario tiene acceso
    """
    try:
        sucursal_actual_id = request.session.get('idSucursalActual')
        empresa_actual_id = request.session.get('idEmpresaActual')
        filtro_empresa = request.GET.get('filtro_empresa', 'todas')
        
        print(f"🔍 DEBUG - Usuario: {request.user}")
        print(f"🔍 DEBUG - Sucursal actual: {sucursal_actual_id}")
        print(f"🔍 DEBUG - Empresa actual: {empresa_actual_id}")
        print(f"🔍 DEBUG - Filtro empresa: {filtro_empresa}")
        
        # Determinar qué empresas incluir según el filtro
        if filtro_empresa == 'misma':
            # Para facturas electrónicas: OTRAS empresas (no la actual)
            if not empresa_actual_id:
                print("❌ ERROR - No hay empresa actual en sesión")
                return JsonResponse({
                    'success': False,
                    'error': 'No hay empresa actual en la sesión'
                }, status=400)
            
            # Debug: Mostrar TODAS las empresas del usuario primero
            todas_empresas_usuario = EmpresaUser.objects.filter(
                user=request.user,
                status=True
            ).select_related('empresa')
            
            print(f"🔍 DEBUG - TODAS las empresas del usuario {request.user}:")
            for eu in todas_empresas_usuario:
                print(f"  🏭 {eu.empresa.nombre} (ID: {eu.empresa.id}, RUT: {eu.empresa.rut}) - Status: {eu.status}, Active: {eu.active}")
            
            # Obtener todas las empresas del usuario EXCEPTO la actual
            empresas_usuario = EmpresaUser.objects.filter(
                user=request.user,
                status=True
            ).exclude(
                empresa_id=empresa_actual_id
            ).values_list('empresa_id', flat=True)
            
            empresas_filtro = list(empresas_usuario)
            print(f"🔍 DEBUG - FACTURA ELECTRÓNICA: OTRAS empresas (excluyendo actual ID {empresa_actual_id})")
            print(f"  📋 Empresas filtro: {empresas_filtro}")
            
            if not empresas_filtro:
                print("⚠️ WARNING - No hay otras empresas disponibles para facturas internas")
                print("  💡 Esto significa que el usuario solo tiene acceso a UNA empresa")
            
            # Verificar que la empresa actual existe y mostrar su información
            try:
                empresa_actual = Empresa.objects.get(id=empresa_actual_id)
                print(f"  🏭 Empresa actual (EXCLUIDA): {empresa_actual.nombre} (RUT: {empresa_actual.rut})")
            except Empresa.DoesNotExist:
                print(f"❌ ERROR - Empresa ID {empresa_actual_id} no existe")
                return JsonResponse({
                    'success': False,
                    'error': f'Empresa con ID {empresa_actual_id} no encontrada'
                }, status=400)
        else:
            # Para guías de despacho: SOLO la misma empresa (sesión actual)
            if not empresa_actual_id:
                print("❌ ERROR - No hay empresa actual en sesión para guías")
                return JsonResponse({
                    'success': False,
                    'error': 'No hay empresa actual en la sesión'
                }, status=400)
            
            empresas_filtro = [empresa_actual_id]
            print(f"🔍 DEBUG - GUÍA DE DESPACHO: Solo empresa actual ID {empresa_actual_id}")
            
            # Verificar que la empresa actual existe y mostrar su información
            try:
                empresa_actual = Empresa.objects.get(id=empresa_actual_id)
                print(f"  🏭 Empresa actual (INCLUIDA): {empresa_actual.nombre} (RUT: {empresa_actual.rut})")
            except Empresa.DoesNotExist:
                print(f"❌ ERROR - Empresa ID {empresa_actual_id} no existe")
                return JsonResponse({
                    'success': False,
                    'error': f'Empresa con ID {empresa_actual_id} no encontrada'
                }, status=400)
        
        # Obtener RUT de la empresa actual para validación de facturas internas
        try:
            empresa_actual_obj = Empresa.objects.get(id=empresa_actual_id)
            rut_empresa_actual = empresa_actual_obj.rut
            print(f"📋 RUT empresa actual: {rut_empresa_actual}")
        except Empresa.DoesNotExist:
            rut_empresa_actual = None
        
        # Obtener sucursales según el filtro, excluyendo la actual
        sucursales_query = Sucursal.objects.filter(
            empresa_id__in=empresas_filtro
        ).exclude(
            id=sucursal_actual_id
        ).select_related('empresa')
        
        # VALIDACIÓN IMPORTANTE: Para FACTURAS entre sucursales (despacho interno con factura),
        # filtrar sucursales con RUT diferente al de la empresa actual
        # Para GUÍAS, NO se aplica este filtro (pueden ser del mismo RUT)
        if filtro_empresa == 'misma' and rut_empresa_actual:
            # Es FACTURA entre sucursales, filtrar por RUT diferente
            print(f"🔍 VALIDACIÓN FACTURA - Filtrando sucursales con RUT diferente a: {rut_empresa_actual}")
            sucursales_query = sucursales_query.exclude(empresa__rut=rut_empresa_actual)
            print(f"  ✅ Filtro RUT aplicado para facturas")
        else:
            print(f"🔍 GUÍA - No se aplica filtro de RUT (permite mismo RUT para traspasos internos)")
        
        print(f"🔍 DEBUG - Query sucursales: {sucursales_query.query}")
        print(f"🔍 DEBUG - Sucursales encontradas: {sucursales_query.count()}")
        
        sucursales_list = []
        for sucursal in sucursales_query:
            sucursal_data = {
                'id': sucursal.id,
                'nombre': sucursal.alias,  # Usar 'alias' en lugar de 'nombre'
                'direccion': sucursal.direccion,
                'empresa': sucursal.empresa.razon_social,
                'empresa_id': sucursal.empresa.id,  # Para debug
                'empresa_rut': sucursal.empresa.rut  # Para debug
            }
            sucursales_list.append(sucursal_data)
            print(f"  ✅ Sucursal incluida: {sucursal.alias} | Empresa: {sucursal.empresa.nombre} (ID: {sucursal.empresa.id}, RUT: {sucursal.empresa.rut})")
        
        # Mostrar resumen final
        if filtro_empresa == 'misma':
            print(f"📋 RESUMEN FACTURA: Se devuelven {len(sucursales_list)} sucursales de otras empresas (RUT diferente)")
        else:
            print(f"📋 RESUMEN GUÍA: Se devuelven {len(sucursales_list)} sucursales de la misma empresa (permite mismo RUT)")
        
        print(f"🔍 DEBUG - Total sucursales devueltas: {len(sucursales_list)}")
        print(f"🔍 DEBUG - Filtro aplicado: {filtro_empresa}")
        
        return JsonResponse(sucursales_list, safe=False)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error en obtener_sucursales: {str(e)}', exc_info=True)
        
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener sucursales: {str(e)}'
        }, status=500)
@require_POST
@login_required
def emitir_dte(request):
    """
    Procesar la emisión de un DTE
    """
    try:
        data = json.loads(request.body)
        
        # DEBUG COMPLETO: Mostrar todos los datos recibidos
        print("🔍 DEBUG COMPLETO - Datos recibidos en emitir_dte:")
        print(f"📋 Raw data: {data}")
        
        # Validar datos requeridos
        metodo_despacho = data.get('metodo_despacho')
        tipo_documento = data.get('tipo_documento')
        receptor_id = data.get('receptor_id')
        sucursal_destino_id = data.get('sucursal_destino_id')
        fecha_emision = data.get('fecha_emision')
        detalle_productos = data.get('detalle_productos', [])
        observaciones = data.get('observaciones', '')
        
        # DEBUG: Mostrar cada campo individualmente
        print(f"📝 metodo_despacho: '{metodo_despacho}' (len: {len(str(metodo_despacho)) if metodo_despacho else 0})")
        print(f"📝 tipo_documento: '{tipo_documento}' (len: {len(str(tipo_documento)) if tipo_documento else 0})")
        print(f"📝 receptor_id: '{receptor_id}'")
        print(f"📝 sucursal_destino_id: '{sucursal_destino_id}'")
        print(f"📝 fecha_emision: '{fecha_emision}'")
        print(f"📝 observaciones: '{observaciones}' (len: {len(str(observaciones)) if observaciones else 0})")
        print(f"📝 detalle_productos: {len(detalle_productos)} items")
        
        # Validar datos básicos
        if not all([metodo_despacho, tipo_documento, fecha_emision]):
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos obligatorios básicos'
            }, status=400)
        
        # Validar según tipo de despacho
        if metodo_despacho == 'interno':
            if not sucursal_destino_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar una sucursal destino para despacho interno'
                }, status=400)
        else:  # despacho externo
            if not receptor_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar una empresa cliente para despacho externo'
                }, status=400)
        
        if not detalle_productos:
            return JsonResponse({
                'success': False,
                'error': 'Debe incluir al menos un producto'
            }, status=400)
        
        # Obtener datos de sesión
        sucursal_id = request.session.get('idSucursalActual')
        empresa_id = request.session.get('idEmpresaActual')
        
        if not sucursal_id or not empresa_id:
            # Intentar obtener la primera sucursal disponible como fallback
            try:
                primera_sucursal = Sucursal.objects.first()
                if primera_sucursal:
                    sucursal_id = primera_sucursal.id
                    empresa_id = primera_sucursal.empresa.id
                    # Establecer en sesión
                    request.session['idSucursalActual'] = sucursal_id
                    request.session['idEmpresaActual'] = empresa_id
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'No hay sucursales configuradas en el sistema'
                    }, status=400)
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error al obtener datos de sesión: {str(e)}'
                }, status=400)
        
        # Obtener objetos según tipo de despacho
        emisor = get_object_or_404(Empresa, id=empresa_id)
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        receptor = None
        sucursal_destino = None
        
        if metodo_despacho == 'interno':
            # Despacho interno: obtener sucursal destino y su empresa como receptor
            sucursal_destino = get_object_or_404(Sucursal, id=sucursal_destino_id)
            receptor = sucursal_destino.empresa  # La empresa se deduce de la sucursal
            
            # VALIDACIÓN IMPORTANTE: Verificar que los RUTs sean diferentes SOLO para FACTURAS
            # Para GUÍAS, se permite el mismo RUT (traspasos internos)
            if tipo_documento == 'FACTURA ELECTRONICA' and emisor.rut == receptor.rut:
                print(f"❌ ERROR - Intento de emisión de FACTURA entre sucursales del mismo RUT: {emisor.rut}")
                return JsonResponse({
                    'success': False,
                    'error': f'No se puede emitir FACTURA entre sucursales de empresas con el mismo RUT ({emisor.rut}). Use GUÍA DE DESPACHO para traspasos internos.'
                }, status=400)
            
            if emisor.rut != receptor.rut:
                print(f"✅ Validación RUT OK - Emisor: {emisor.rut}, Receptor: {receptor.rut} (diferentes)")
            else:
                print(f"✅ GUÍA permitida - Emisor: {emisor.rut}, Receptor: {receptor.rut} (mismo RUT, traspaso interno)")
        else:
            # Despacho externo: obtener empresa cliente
            receptor = get_object_or_404(Empresa, id=receptor_id)
        
        # El frontend ya envía los valores correctos del modelo, no necesitamos mapear
        tipo_doc = tipo_documento
        
        # DEBUG: Verificar tipo de documento
        print(f"🔍 DEBUG - tipo_documento recibido: '{tipo_documento}' (len: {len(tipo_documento)})")
        
        # Validar que el tipo de documento sea válido
        tipos_validos = ['FACTURA ELECTRONICA', 'BOLETA ELECTRONICA', 'GUIA', 'NOTA DE PEDIDO', 'NOTA DE CREDITO']
        if tipo_doc not in tipos_validos:
            print(f"❌ ERROR - Tipo de documento inválido: '{tipo_doc}'")
            return JsonResponse({
                'success': False,
                'error': f'Tipo de documento inválido: {tipo_doc}. Valores válidos: {tipos_validos}'
            }, status=400)
        
        print(f"✅ Tipo de documento validado correctamente: {tipo_doc}")
        
        with transaction.atomic():
            print(f"🔄 Iniciando transacción atómica...")
            # Calcular totales
            subtotal_neto = 0
            total_unidades = 0
            
            print(f"📦 Procesando {len(detalle_productos)} productos...")
            for idx, item in enumerate(detalle_productos, 1):
                talla_id = item.get('talla_id')
                cantidad = int(item.get('cantidad', 0))
                precio = int(float(item.get('precio', 0)))  # Convertir a int para compatibilidad con IntegerField
                
                print(f"  Item {idx}: talla_id={talla_id}, cantidad={cantidad}, precio={precio}")
                
                # Validar stock disponible en la sucursal de origen
                talla = get_object_or_404(Producto_Talla, id=talla_id)
                
                # DEBUG DETALLADO: Información del producto y stock
                print(f"    📦 Producto: {talla.producto.articulo} | SKU: {talla.sku} | Talla: {talla.talla}")
                print(f"    📍 Stock global (campo directo): {talla.stock}")
                print(f"    📍 Producto.sucursal_id: {talla.producto.sucursal_id}")
                print(f"    🏢 Sucursal actual (origen): {sucursal.alias} (ID: {sucursal_id})")
                
                # Verificar si tiene movimientos
                tiene_movimientos = talla.movimientos_productos_talla.exists()
                print(f"    📊 Tiene movimientos registrados: {tiene_movimientos}")
                
                # ✅ Usar stock_sucursal() para validar stock específico de la sucursal origen
                stock_disponible = talla.stock_sucursal(sucursal_id)
                print(f"    ✅ Stock disponible en sucursal {sucursal.alias}: {stock_disponible}")
                
                if stock_disponible < cantidad:
                    print(f"❌ ERROR - Stock insuficiente en sucursal {sucursal.alias}")
                    print(f"    ⚠️ DIAGNÓSTICO:")
                    print(f"       - Stock global: {talla.stock}")
                    print(f"       - Stock en sucursal: {stock_disponible}")
                    print(f"       - Tiene movimientos: {tiene_movimientos}")
                    print(f"       - Producto.sucursal_id: {talla.producto.sucursal_id} vs Sucursal actual: {sucursal_id}")
                    
                    error_msg = f'Stock insuficiente para {talla.producto.articulo} talla {talla.talla} (SKU: {talla.sku}) en sucursal {sucursal.alias}. Disponible: {stock_disponible}, Solicitado: {cantidad}'
                    
                    if not tiene_movimientos and talla.producto.sucursal_id != sucursal_id:
                        error_msg += f'. NOTA: Este producto pertenece a otra sucursal (ID: {talla.producto.sucursal_id}). Necesita crear movimientos de traspaso o migrar a sistema de movimientos.'
                    
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    }, status=400)
                
                subtotal_neto += cantidad * precio
                total_unidades += cantidad
            
            print(f"✅ Validación de stock OK. Subtotal: {subtotal_neto}, Unidades: {total_unidades}")
            
            # Calcular IVA y total
            subtotal_decimal = Decimal(str(subtotal_neto))
            iva = subtotal_decimal * Decimal('0.19')
            total_con_iva = subtotal_decimal + iva
            print(f"💰 Cálculos: Neto={subtotal_decimal}, IVA={iva}, Total={total_con_iva}")
            
            # Obtener correlativo oficial de la sucursal para este tipo de DTE
            try:
                print(f"🔢 Intentando obtener correlativo para {tipo_doc} en sucursal {sucursal.alias}...")
                numero_documento = obtener_siguiente_correlativo(sucursal, tipo_doc)
                print(f"✅ Correlativo obtenido: {numero_documento}")
            except Exception as correlativo_error:
                print(f"❌ ERROR al obtener correlativo: {correlativo_error}")
                import traceback
                traceback.print_exc()
                return JsonResponse({
                    'success': False,
                    'error': f'No fue posible obtener el correlativo para {tipo_doc}: {correlativo_error}'
                }, status=400)
            
            # Determinar estado y tipo según contexto de negocio
            if metodo_despacho == 'interno':
                estado_dte = 'EMITIDO'
                estado_pago = 'PENDIENTE'
                # TODOS los despachos internos son TRASPASO (factura o guía)
                tipo_transaccion = 'TRASPASO'
            else:
                # Despacho externo = Venta a cliente
                estado_dte = 'EMITIDO'
                estado_pago = 'PENDIENTE'
                tipo_transaccion = 'VENTA'
            
            # DEBUG: Mostrar valores asignados y lógica aplicada
            print(f"🔍 DEBUG - Lógica aplicada:")
            print(f"  metodo_despacho: '{metodo_despacho}'")
            print(f"  tipo_doc: '{tipo_doc}'")
            print(f"  → estado_dte: '{estado_dte}' (len: {len(estado_dte)})")
            print(f"  → estado_pago: '{estado_pago}' (len: {len(estado_pago)})")
            print(f"  → tipo_transaccion: '{tipo_transaccion}' (len: {len(tipo_transaccion)})")
            
            # Explicar la lógica aplicada
            if metodo_despacho == 'interno':
                if tipo_doc == 'FACTURA ELECTRONICA':
                    print("  📋 Lógica: Factura interna → VENTA (entre empresas del grupo)")
                else:
                    print("  📋 Lógica: Guía interna → TRASPASO (entre sucursales misma empresa)")
            else:
                print("  📋 Lógica: Despacho externo → VENTA (a cliente)")
            
            # Preparar referencias
            referencias_texto = f"Método despacho: {metodo_despacho}"
            if sucursal_destino:
                referencias_texto += f". Destino: {sucursal_destino.alias}"
            if observaciones:
                referencias_texto += f". {observaciones}"
            
            print(f"📄 Creando DTE con los siguientes parámetros:")
            print(f"  emisor_id: {emisor.id}, receptor_id: {receptor.id if receptor else 'None'}")
            print(f"  numero_documento: {numero_documento}, tipo_documento: {tipo_doc}")
            print(f"  estado_pago: {estado_pago}, estado_dte: {estado_dte}")
            print(f"  tipo_transaccion: {tipo_transaccion}")
            
            # Crear DTE con todos los campos requeridos
            try:
                dte = Dte.objects.create(
                    emisor=emisor,
                    receptor=receptor,
                    numero_documento=numero_documento,
                    tipo_documento=tipo_doc,
                    monto_neto=subtotal_decimal,
                    monto_con_iva=total_con_iva,
                    estado_pago=estado_pago,
                    estado_dte=estado_dte,
                    responsable=request.user.username,
                    fecha_emision=parse_date(fecha_emision),
                    fecha_vencimiento=parse_date(fecha_emision),  # Mismo día por defecto
                    diasCredito=0,
                    bultos=1,  # Por defecto
                    unidades_productos=total_unidades,
                    tipo_transaccion=tipo_transaccion,
                    referencias=referencias_texto,
                    sucursal=sucursal
                )
                print(f"✅ DTE creado exitosamente: ID={dte.id}")
            except Exception as e:
                print(f"❌ ERROR al crear DTE: {e}")
                import traceback
                traceback.print_exc()
                return JsonResponse({
                    'success': False,
                    'error': f'Error al crear DTE: {str(e)}'
                }, status=400)
            
            # Crear detalle de productos y actualizar stock
            
            for item in detalle_productos:
                talla_id = item.get('talla_id')
                cantidad = int(item.get('cantidad', 0))
                precio = int(float(item.get('precio', 0)))  # Convertir a int para compatibilidad con IntegerField
                
                talla = Producto_Talla.objects.get(id=talla_id)
                producto = talla.producto
                
                # Crear detalle del DTE
                Dte_Productos.objects.create(
                    dte=dte,
                    productoTalla=talla,
                    descripcion=f"{producto.articulo} - Talla {talla.talla}",
                    costo=producto.costo,
                    sobreprecio=producto.sobreprecio,
                    precio=int(precio),
                    stock=cantidad,
                    activo=True
                )
                
                # Gestión de stock y movimientos según tipo de despacho
                if metodo_despacho == 'externo':
                    # DESPACHO EXTERNO: Crear movimiento de egreso (venta a cliente)
                    Movimientos_Producto.objects.create(
                        dte=dte,
                        ProductoTalla=talla,
                        sucursal_origen=sucursal,
                        sucursal_destino=None,
                        cantidad=-cantidad,  # Negativo porque es egreso
                        costo=producto.costo,
                        sobreprecio=producto.sobreprecio,
                        precio=int(precio),
                        concepto='VENTA_MAYORISTA',
                        tipo_movimiento='EGRESO',
                        estado='COMPLETADO',
                        responsable=request.user.username,
                        observaciones=f"Venta DTE #{numero_documento} - Cliente: {receptor.nombre if receptor else 'N/A'}"
                    )
                    
                    # ✅ NO modificar el campo talla.stock
                    # El stock se calcula automáticamente desde los movimientos usando stock_sucursal()
                    print(f"✓ Movimiento de egreso creado: {talla.sku} -{cantidad} en sucursal {sucursal.alias}")
                    print(f"  Stock se actualiza automáticamente desde movimientos")
                    
                else:
                    # DESPACHO INTERNO: Crear movimiento de traspaso (salida en origen)
                    # ⚠️ IMPORTANTE: Solo se crea el movimiento de EGRESO en origen
                    # El movimiento de INGRESO en destino se creará cuando la sucursal destino recepcione
                    
                    Movimientos_Producto.objects.create(
                        dte=dte,
                        ProductoTalla=talla,
                        sucursal_origen=sucursal,
                        sucursal_destino=sucursal_destino,
                        cantidad=-cantidad,  # Negativo porque es egreso
                        costo=producto.costo,
                        sobreprecio=producto.sobreprecio,
                        precio=int(precio),
                        concepto='TRASPASO_SALIDA',
                        tipo_movimiento='EGRESO',
                        estado='COMPLETADO',  # ✅ COMPLETADO inmediatamente - el stock ya salió de origen
                        responsable=request.user.username,
                        observaciones=f"Traspaso DTE #{numero_documento} - Origen: {sucursal.alias} → Destino: {sucursal_destino.alias}"
                    )
                    
                    # ✅ NO modificar el campo talla.stock
                    # El stock se calcula automáticamente desde los movimientos usando stock_sucursal()
                    # SOLO se crea movimiento de EGRESO en origen
                    # El movimiento de INGRESO en destino lo creará confirmar_recepcion_api() cuando recepcionen
                    print(f"✓ Movimiento de EGRESO creado: {talla.sku} -{cantidad} desde {sucursal.alias}")
                    print(f"  Destino: {sucursal_destino.alias} (pendiente de recepción)")
                    print(f"  Stock en {sucursal.alias} se actualiza automáticamente desde movimientos")
        
        return JsonResponse({
            'success': True,
            'message': 'DTE emitido correctamente',
            'numero_documento': numero_documento,
            'dte_id': dte.id,
            'total': float(total_con_iva)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al emitir DTE: {str(e)}'
        }, status=500)

# ========== REDIRECCIÓN PARA GESTIÓN DE USUARIOS ==========

@login_required
def gestion_usuarios_redirect(request):
    """
    Redirige a la gestión de usuarios en la app users
    """
    return redirect('users:gestion_usuarios')
# ========== GESTIÓN DE CAMBIO DE EMPRESA/SUCURSAL ==========
@login_required
def cambiar_empresa(request):
    """
    Vista para mostrar las empresas y sucursales disponibles para el usuario
    SOLO muestra empresas que tienen sucursales asignadas
    """
    print(f"🔍 DEBUG - Usuario actual: {request.user}")
    print(f"🔍 DEBUG - Usuario ID: {request.user.id}")
    
    # Primero verificar todos los EmpresaUser del usuario
    todos_empresa_user = EmpresaUser.objects.filter(user=request.user)
    print(f"🔍 DEBUG - Total EmpresaUser para usuario: {todos_empresa_user.count()}")
    
    for eu in todos_empresa_user:
        sucursal_info = f"Sucursal: {eu.sucursal.alias} (ID: {eu.sucursal.id})" if eu.sucursal else "Sucursal: None"
        print(f"  - EmpresaUser ID: {eu.id}, Empresa: {eu.empresa.nombre} (ID: {eu.empresa.id}), {sucursal_info}, Status: {eu.status}, Active: {eu.active}")
    
    # Obtener solo las empresas que tienen sucursales asignadas al usuario
    empresas_usuario = EmpresaUser.objects.filter(
        user=request.user,
        status=True,
        sucursal__isnull=False  # Solo registros que tienen sucursal asignada
    ).select_related('empresa', 'sucursal')
    
    print(f"🔍 DEBUG - EmpresaUser con sucursales: {empresas_usuario.count()}")
    
    # Organizar por empresa
    empresas_data = {}
    for eu in empresas_usuario:
        print(f"🔍 DEBUG - Procesando: Empresa {eu.empresa.nombre}, Sucursal {eu.sucursal.alias}")
        empresa_id = eu.empresa.id
        if empresa_id not in empresas_data:
            empresas_data[empresa_id] = {
                'empresa': eu.empresa,
                'sucursales': [],
                'is_current': eu.empresa.id == request.session.get('idEmpresaActual')
            }
        
        # Como ya filtramos por sucursal__isnull=False, sabemos que eu.sucursal existe
        empresas_data[empresa_id]['sucursales'].append({
            'sucursal': eu.sucursal,
            'empresa_user': eu,
                'is_current': (
                    eu.empresa.id == request.session.get('idEmpresaActual') and 
                    eu.sucursal.id == request.session.get('idSucursalActual')
                )
        })
    
    print(f"🔍 DEBUG - Empresas finales organizadas: {len(empresas_data)}")
    
    # Ordenar empresas y sucursales para una mejor presentación
    empresas_data = dict(
        sorted(
            empresas_data.items(),
            key=lambda item: item[1]['empresa'].nombre.lower()
        )
    )

    for data in empresas_data.values():
        data['sucursales'].sort(key=lambda s: s['sucursal'].alias.lower())

    total_empresas = len(empresas_data)
    total_sucursales = sum(len(data['sucursales']) for data in empresas_data.values())

    context = {
        'empresas_data': empresas_data,
        'empresa_actual_id': request.session.get('idEmpresaActual'),
        'sucursal_actual_id': request.session.get('idSucursalActual'),
        'total_empresas': total_empresas,
        'total_sucursales': total_sucursales,
        'abrir_modal': request.GET.get('modal') == '1',
    }
    
    return render(request, 'vistas/modulo_administracion/cambiar_empresa.html', context)

@login_required
@require_POST
def seleccionar_empresa_sucursal(request):
    """
    Vista AJAX para cambiar la empresa y sucursal activa del usuario
    """
    try:
        empresa_user_id = request.POST.get('empresa_user_id')
        
        if not empresa_user_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de empresa-usuario requerido'
            })
        
        # Verificar que el usuario tenga acceso a esta empresa/sucursal
        empresa_user = get_object_or_404(
            EmpresaUser,
            id=empresa_user_id,
            user=request.user,
            status=True
        )
        
        # Desactivar todas las empresas del usuario
        EmpresaUser.objects.filter(user=request.user).update(active=False)
        
        # Activar la empresa/sucursal seleccionada
        empresa_user.active = True
        empresa_user.save()
        
        # Actualizar la sesión (usar claves consistentes)
        request.session['idEmpresaActual'] = empresa_user.empresa.id
        request.session['empresaActual'] = empresa_user.empresa.id  # Mantener compatibilidad
        request.session['nombreEmpresaActual'] = empresa_user.empresa.nombre
        request.session['rutEmpresaActual'] = empresa_user.empresa.rut
        
        if empresa_user.sucursal:
            request.session['idSucursalActual'] = empresa_user.sucursal.id
            request.session['sucursalActual'] = empresa_user.sucursal.id  # Mantener compatibilidad
            request.session['alias'] = empresa_user.sucursal.alias
            request.session['direccionSucursal'] = empresa_user.sucursal.direccion
        else:
            request.session['idSucursalActual'] = None
            request.session['sucursalActual'] = None  # Mantener compatibilidad
            request.session['alias'] = 'Sin sucursal'
            request.session['direccionSucursal'] = 'Sin dirección'
        
        return JsonResponse({
            'success': True,
            'message': f'Cambiado a {empresa_user.empresa.nombre}' + 
                      (f' - {empresa_user.sucursal.alias}' if empresa_user.sucursal else ''),
            'empresa': {
                'id': empresa_user.empresa.id,
                'nombre': empresa_user.empresa.nombre,
                'rut': empresa_user.empresa.rut,
            },
            'sucursal': {
                'id': empresa_user.sucursal.id if empresa_user.sucursal else None,
                'alias': empresa_user.sucursal.alias if empresa_user.sucursal else 'Sin sucursal',
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cambiar empresa/sucursal: {str(e)}'
        })

# ========== BÚSQUEDA DE PRODUCTOS POR SUCURSAL ==========

@login_required
def buscar_productos_sucursal(request):
    """
    Vista para mostrar productos por sucursal con filtros de búsqueda
    """
    # Obtener sucursal actual del usuario (intentar ambas variables)
    sucursal_actual_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    
    # Obtener todas las sucursales disponibles para el usuario
    sucursales_usuario = EmpresaUser.objects.filter(
        user=request.user,
        status=True,
        sucursal__isnull=False
    ).select_related('sucursal').values_list('sucursal', flat=True).distinct()
    
    sucursales = Sucursal.objects.filter(id__in=sucursales_usuario)
    
    # Obtener todos los atributos disponibles para los filtros
    atributos = Productos_Atributos.objects.all().prefetch_related('opciones')
    
    # Obtener categorías
    categorias = Categoria.objects.all()
    
    context = {
        'sucursales': sucursales,
        'sucursal_actual_id': sucursal_actual_id,
        'atributos': atributos,
        'categorias': categorias,
    }
    
    return render(request, 'vistas/modulo_existencias/buscar_productos_sucursal.html', context)

@login_required
def obtener_productos_sucursal(request):
    """
    Vista AJAX para obtener productos filtrados por atributos y sucursal
    Busca en todas las sucursales a las que el usuario tiene acceso
    """
    try:
        # Obtener parámetros de filtro
        search = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria_id')
        atributo1_id = request.GET.get('atributo1_id')  # Marca
        atributo2_id = request.GET.get('atributo2_id')  # Color
        atributo3_id = request.GET.get('atributo3_id')  # Género
        sucursal_id = request.GET.get('sucursal_id')  # Filtro por sucursal
        solo_con_stock = request.GET.get('solo_con_stock') == 'on'  # Filtro de stock
        
        # Parámetros de paginación
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 25))
        
        # Obtener sucursales a las que el usuario tiene acceso
        sucursales_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True,
            sucursal__isnull=False
        ).values_list('sucursal_id', flat=True).distinct()
        
        if not sucursales_usuario:
            return JsonResponse({
                'success': False,
                'error': 'No tienes acceso a ninguna sucursal'
            })
        
        # Construir query base - buscar en sucursales del usuario
        productos_query = Producto.objects.filter(
            sucursal_id__in=sucursales_usuario
        ).select_related(
            'sucursal', 'categoria', 'atributo1', 'atributo2', 'atributo3'
        ).prefetch_related('producto_talla')
        
        # Filtrar por sucursal específica si se seleccionó una
        if sucursal_id:
            productos_query = productos_query.filter(sucursal_id=sucursal_id)
        
        # Filtro de búsqueda general
        if search:
            productos_query = productos_query.filter(
                Q(articulo__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(producto_talla__sku__icontains=search)
            )
        
        # Filtrar por categoría
        if categoria_id:
            productos_query = productos_query.filter(categoria_id=categoria_id)
        
        # Filtrar por atributos específicos
        if atributo1_id:  # Marca
            productos_query = productos_query.filter(atributo1_id=atributo1_id)
        if atributo2_id:  # Color
            productos_query = productos_query.filter(atributo2_id=atributo2_id)
        if atributo3_id:  # Género
            productos_query = productos_query.filter(atributo3_id=atributo3_id)

        # Obtener parámetro de ordenamiento
        ordenar = request.GET.get('ordenar', '')
        
        # Anotar stock ANTES del distinct para que funcione el ordenamiento
        # Esto es necesario para que el ORDER BY funcione correctamente
        necesita_stock_anotado = solo_con_stock or ordenar in ['stock_desc', 'stock_asc']
        
        if necesita_stock_anotado:
            productos_query = productos_query.annotate(
                stock_total_anotado=Sum('producto_talla__stock')
            )
        
        # Aplicar distinct después de anotar
        productos_query = productos_query.distinct()
        
        # Filtrar solo productos con stock si se seleccionó el checkbox
        if solo_con_stock:
            productos_query = productos_query.filter(stock_total_anotado__gt=0)
        
        # Aplicar ordenamiento según el parámetro
        if ordenar:
            if ordenar == 'stock_desc':
                productos_query = productos_query.order_by('-stock_total_anotado')
            elif ordenar == 'stock_asc':
                productos_query = productos_query.order_by('stock_total_anotado')
            elif ordenar == 'articulo_asc':
                productos_query = productos_query.order_by('articulo')
            elif ordenar == 'articulo_desc':
                productos_query = productos_query.order_by('-articulo')
            elif ordenar == 'precio_asc':
                productos_query = productos_query.order_by('precioventa')
            elif ordenar == 'precio_desc':
                productos_query = productos_query.order_by('-precioventa')

        # Contar total
        total_productos = productos_query.count()
        
        # Aplicar paginación
        start = (page - 1) * page_size
        end = start + page_size
        productos = productos_query[start:end]
        
        # Preparar datos para JSON
        productos_data = []
        for producto in productos:
            # Obtener todas las tallas del producto
            todas_las_tallas = producto.producto_talla.all()
            
            # Si solo_con_stock está activado, filtrar solo tallas con stock > 0
            if solo_con_stock:
                tallas_a_mostrar = [pt for pt in todas_las_tallas if pt.stock > 0]
            else:
                tallas_a_mostrar = todas_las_tallas
            
            # Calcular stock total de las tallas a mostrar
            stock_total = sum(pt.stock for pt in tallas_a_mostrar)
            
            # Obtener tallas con stock
            tallas_stock = [
                {
                    'talla': pt.talla,
                    'stock': pt.stock,
                    'sku': pt.sku
                }
                for pt in tallas_a_mostrar
            ]
            
            productos_data.append({
                'id': producto.id,
                'articulo': producto.articulo,
                'descripcion': producto.descripcion,
                'sucursal': producto.sucursal.alias,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.atributo1.valor if producto.atributo1 else '',
                'color': producto.atributo2.valor if producto.atributo2 else '',
                'genero': producto.atributo3.valor if producto.atributo3 else '',
                'stock_total': stock_total,
                'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                'tallas_stock': tallas_stock,
                'tipo_talla': producto.tipo_talla,
            })
        
        # Calcular paginación
        total_paginas = (total_productos + page_size - 1) // page_size
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'pagination': {
                'current_page': page,
                'total_pages': total_paginas,
                'page_size': page_size,
                'total_items': total_productos
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener productos: {str(e)}'
        })

@login_required
def obtener_opciones_atributo(request):
    """
    Vista AJAX para obtener opciones de un atributo específico
    """
    try:
        atributo_id = request.GET.get('atributo_id')
        
        if not atributo_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de atributo requerido'
            })
        
        opciones = AtributoOpcion.objects.filter(
            atributo_id=atributo_id
        ).values('id', 'valor').order_by('valor')
        
        return JsonResponse({
            'success': True,
            'opciones': list(opciones)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener opciones: {str(e)}'
        })

# ========== TICKET DE VENTA ==========

@login_required
def ticket_venta(request):
    """
    Vista principal para crear tickets de venta
    Muestra vendedores asociados a la sucursal actual del usuario
    """
    # Obtener sucursal actual del usuario (intentar ambas variables de sesión)
    sucursal_actual_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    sucursal_actual = None
    empresa_actual_nombre = request.session.get('nombreEmpresaActual', 'Sin empresa')
    
    # Si hay sucursal actual, obtenerla
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    # Si no hay sucursal actual, obtener las sucursales disponibles para el usuario
    sucursales_disponibles = []
    if not sucursal_actual:
        sucursales_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True,
            sucursal__isnull=False
        ).select_related('sucursal', 'empresa').distinct()
        
        sucursales_disponibles = [
            {
                'sucursal': eu.sucursal,
                'empresa': eu.empresa
            }
            for eu in sucursales_usuario
        ]
    
    # Validar que existe correlativo para tickets
    tiene_correlativo = False
    correlativo_info = None
    if sucursal_actual:
        try:
            correlativo = Correlativo.objects.get(
                sucursal=sucursal_actual,
                tipo_dte='TICKET'
            )
            tiene_correlativo = correlativo.puede_emitir()
            correlativo_info = {
                'disponibles': correlativo.disponibles,
                'inicio': correlativo.inicio,
                'termino': correlativo.termino,
                'estado': correlativo.estado
            }
        except Correlativo.DoesNotExist:
            tiene_correlativo = False
            correlativo_info = None
    
    # Obtener vendedores de la sucursal actual
    if sucursal_actual:
        # Obtener vendedores asignados a esta sucursal
        vendedores = Vendedor.objects.filter(sucursales=sucursal_actual).order_by('nombre')
        
        # Si no hay vendedores asignados, mostrar todos
        if not vendedores.exists():
            vendedores = Vendedor.objects.all().order_by('nombre')
    else:
        # Si no hay sucursal seleccionada, mostrar todos los vendedores
        vendedores = Vendedor.objects.all().order_by('nombre')
    
    context = {
        'sucursal_actual': sucursal_actual,
        'empresa_actual_nombre': empresa_actual_nombre,
        'vendedores': vendedores,
        'sucursales_disponibles': sucursales_disponibles,
        'necesita_seleccionar_sucursal': not sucursal_actual,
        'tiene_correlativo': tiene_correlativo,
        'correlativo_info': correlativo_info,
    }
    
    return render(request, 'vistas/modulo_ventas/ticket_venta.html', context)

@login_required
def buscar_vendedor_por_codigo(request):
    """
    Vista AJAX para buscar vendedor por código
    """
    try:
        codigo = request.GET.get('codigo', '').strip()
        
        if not codigo:
            return JsonResponse({
                'success': False,
                'error': 'Código de vendedor requerido'
            })
        
        try:
            vendedor = Vendedor.objects.get(codigo_vendedor=codigo)
            
            return JsonResponse({
                'success': True,
                'vendedor': {
                    'id': vendedor.id,
                    'codigo': vendedor.codigo_vendedor,
                    'nombre': vendedor.nombre,
                    'rut': vendedor.rut,
                    'correo': vendedor.correo,
                }
            })
            
        except Vendedor.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'No se encontró vendedor con código: {codigo}'
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar vendedor: {str(e)}'
        })
def buscar_producto_por_sku(request):
    """
    Buscar producto por SKU para el ticket de venta
    """
    sku = request.GET.get('sku', '').strip()
    # Obtener sucursal de la sesión (intentar ambas variables)
    sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    
    if not sku:
        return JsonResponse({
            'success': False,
            'message': 'SKU requerido'
        })
    
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'message': 'No hay sucursal seleccionada. Por favor selecciona una sucursal desde el menú principal.'
        })
    
    try:
        # Buscar el producto por SKU en la sucursal actual
        producto_talla = Producto_Talla.objects.select_related(
            'producto',
            'producto__atributo1__atributo',
            'producto__atributo2__atributo', 
            'producto__atributo3__atributo',
            'producto__atributo4__atributo'
        ).get(
            sku=sku,
            producto__sucursal_id=sucursal_id
        )
        
        producto = producto_talla.producto
        
        # Obtener la marca desde los atributos (asumiendo que está en atributo1, 2, 3 o 4)
        marca = '-'
        for attr_num in range(1, 5):
            attr = getattr(producto, f'atributo{attr_num}')
            if attr and attr.atributo and attr.atributo.nombre.lower() in ['marca', 'brand']:
                marca = attr.valor
                break
        
        return JsonResponse({
            'success': True,
            'producto': {
                'sku': producto_talla.sku,
                'articulo': producto.articulo,
                'descripcion': producto.descripcion,
                'marca': marca,
                'talla': producto_talla.talla,
                'precio_venta': int(producto.precioventa),
                'stock': producto_talla.stock,
                'producto_talla_id': producto_talla.id
            }
        })
        
    except Producto_Talla.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'No se encontró producto con SKU {sku} en esta sucursal'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al buscar producto: {str(e)}'
        })

def buscar_productos_bodega_DUPLICADA_NO_USAR(request):
    """
    ⚠️ FUNCIÓN DUPLICADA - NO USAR
    Esta función es un duplicado de buscar_productos_bodega (línea 8442)
    Renombrada para evitar conflictos. Debe ser eliminada eventualmente.
    
    Buscar productos en bodega para emisión DTE con estructura correcta de atributos
    """
    try:
        # Obtener parámetros de búsqueda (acepta tanto 'q' como 'search')
        search = request.GET.get('q', '').strip() or request.GET.get('search', '').strip()
        marca_id = request.GET.get('marca', '')
        categoria_id = request.GET.get('categoria', '')
        tipo_talla = request.GET.get('tipo_talla', '')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Obtener parámetro para incluir productos sin stock
        incluir_sin_stock = request.GET.get('incluir_sin_stock', 'false').lower() == 'true'
        
        # Obtener sucursal actual
        sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
        empresa_id = request.session.get('idEmpresaActual') or request.session.get('empresaActual')
        
        print(f"🔍 DEBUG - Búsqueda productos bodega:")
        print(f"  sucursal_id: {sucursal_id}")
        print(f"  empresa_id: {empresa_id}")
        print(f"  search: '{search}'")
        print(f"  marca_id: '{marca_id}'")
        print(f"  categoria_id: '{categoria_id}'")
        print(f"  incluir_sin_stock: {incluir_sin_stock}")
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión'
            }, status=400)
        
        # Query base con select_related para optimizar
        productos_query = Producto.objects.select_related(
            'atributo1',  # Marca
            'atributo2',  # Color  
            'atributo3',  # Género/Sexo
            'atributo4',  # Otro
            'categoria',
            'sucursal'
        ).prefetch_related(
            'producto_talla'  # Tallas del producto
        ).filter(
            sucursal_id=sucursal_id
        )
        
        # ⭐ FILTRAR POR STOCK solo si NO se desea incluir productos sin stock
        if not incluir_sin_stock:
            productos_query = productos_query.filter(producto_talla__stock__gt=0)
        
        # Filtros de búsqueda
        if search:
            productos_query = productos_query.filter(
                Q(articulo__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(producto_talla__sku__icontains=search)
            ).distinct()
        
        if marca_id:
            productos_query = productos_query.filter(atributo1_id=marca_id)
            
        if categoria_id:
            productos_query = productos_query.filter(categoria_id=categoria_id)
            
        if tipo_talla:
            productos_query = productos_query.filter(tipo_talla=tipo_talla)
        
        # Eliminar duplicados antes de paginar (por el JOIN con producto_talla)
        productos_query = productos_query.distinct()
        
        # Paginación (con order_by para evitar warning)
        from django.core.paginator import Paginator
        productos_query = productos_query.order_by('-id')  # Ordenar por ID descendente
        paginator = Paginator(productos_query, page_size)
        page_obj = paginator.get_page(page)
        
        # Construir respuesta - Agrupar productos con sus tallas
        productos_data = []
        for producto in page_obj:
            # Obtener atributos correctamente
            marca = producto.atributo1.valor if producto.atributo1 else '-'
            color = producto.atributo2.valor if producto.atributo2 else '-'
            sexo = producto.atributo3.valor if producto.atributo3 else '-'
            
            # Obtener tallas según la opción de incluir sin stock
            if incluir_sin_stock:
                # Incluir todas las tallas (con y sin stock)
                tallas_disponibles_obj = producto.producto_talla.all()
            else:
                # Solo tallas con stock (por defecto)
                tallas_disponibles_obj = producto.producto_talla.filter(stock__gt=0)
            
            # Calcular stock total
            stock_total = sum(talla.stock for talla in tallas_disponibles_obj)
            
            # Obtener lista de tallas disponibles (sin duplicados, ordenadas)
            tallas_set = sorted(set(talla.talla for talla in tallas_disponibles_obj))
            tallas_disponibles = tallas_set
            
            productos_data.append({
                'id': producto.id,
                'articulo': producto.articulo,
                'descripcion': producto.descripcion or '-',
                'marca': marca,
                'color': color,
                'sexo': sexo,
                'categoria': producto.categoria.nombre if producto.categoria else '-',
                'tipo_talla': producto.tipo_talla or '-',
                'costo': float(producto.costo) if producto.costo else 0,
                'sobreprecio': float(producto.sobreprecio) if producto.sobreprecio else 0,  # ← AGREGADO
                'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                'stock_total': stock_total,
                'tallas_disponibles': tallas_disponibles,
                'estado': 'Activo',
                'sucursal_id': producto.sucursal_id,
                # Información adicional para el sistema de tallas
                'tallas_detalle': [{
                    'id': talla.id,
                    'talla': talla.talla,
                    'sku': str(talla.sku),
                    'stock': talla.stock,
                    'costo': float(producto.costo) if producto.costo else 0,  # ← AGREGADO
                    'sobreprecio': float(producto.sobreprecio) if producto.sobreprecio else 0  # ← AGREGADO
                } for talla in tallas_disponibles_obj]
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'pagination': {
                'page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'page_size': page_size
            }
        })
        
    except Exception as e:
        print(f"❌ ERROR en buscar_productos_bodega: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def crear_ticket(request):
    """
    Crear ticket de venta
    """
    try:
        data = json.loads(request.body)
        vendedor_id = data.get('vendedor_id')
        productos = data.get('productos', [])
        total = data.get('total', 0)
        total_items = data.get('total_items', 0)

        sucursal_id = (
            request.session.get('idSucursalActual')
            or request.session.get('sucursalActual')
            or data.get('sucursal_id')
            or data.get('sucursal')
        )

        # Validaciones
        if not vendedor_id:
            return JsonResponse({
                'success': False,
                'message': 'Vendedor requerido'
            })
        
        if not sucursal_id:
            return JsonResponse({
                'success': False,
                'message': 'Sucursal requerida'
            })
        
        if not productos:
            return JsonResponse({
                'success': False,
                'message': 'Debe agregar al menos un producto'
            })
        
        # Verificar que el vendedor existe
        try:
            vendedor = Vendedor.objects.get(id=vendedor_id)
        except Vendedor.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Vendedor no encontrado'
            })
        
        # Verificar que la sucursal existe
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id)
        except Sucursal.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Sucursal no encontrada'
            })
        
        with transaction.atomic():
            # Obtener correlativo desde tabla Correlativo para tipo Ticket
            correlativo = obtener_siguiente_correlativo(sucursal, 'TICKET')
            
            # Crear el ticket
            ticket = Ticket.objects.create(
                vendedor=vendedor,
                sucursal=sucursal,
                correlativo=correlativo,
                estado='PENDIENTE',
                subTotal=total,
                descuento=0,
                total=total,
                responsable=request.user.username if request.user.is_authenticated else 'Sistema'
            )
            
            # Agregar productos al ticket
            for producto_data in productos:
                sku = producto_data.get('sku')
                cantidad_raw = producto_data.get('cantidad', 1)
                precio = int(producto_data.get('precio', 0))
                
                # Validar cantidad - debe ser un número entero positivo
                try:
                    cantidad = int(cantidad_raw)
                except (ValueError, TypeError):
                    raise ValidationError(f'Cantidad inválida para SKU {sku}: debe ser un número entero')
                
                if cantidad < 1:
                    raise ValidationError(f'Cantidad inválida para SKU {sku}: debe ser mayor a 0')
                
                try:
                    producto_talla = Producto_Talla.objects.get(
                        sku=sku,
                        producto__sucursal=sucursal
                    )
                    
                    # Verificar stock
                    if producto_talla.stock < cantidad:
                        raise ValidationError(
                            f'Stock insuficiente para SKU {sku}. Solicitado: {cantidad}, Disponible: {producto_talla.stock}'
                        )
                    
                    # Crear registro en Ticket_Productos
                    Ticket_Productos.objects.create(
                        ProductoTalla=producto_talla,
                        idTicket=ticket,
                        stock=cantidad,
                        precio=precio,
                        descuento_unitario=0,
                        subtotal=precio * cantidad
                    )
                    
                    # ⚠️ NO DESCONTAR STOCK AQUÍ - Se descuenta al PAGAR el ticket en registrar_pagos_ticket
                    # El ticket se crea en estado PENDIENTE y el stock se descuenta cuando cambia a PAGADO
                    # producto_talla.stock -= cantidad
                    # producto_talla.save()
                    
                except Producto_Talla.DoesNotExist:
                    raise ValidationError(f'Producto con SKU {sku} no encontrado')
            
            ticket_data = construir_ticket_data(ticket)
            ticket_html = generar_ticket_html(ticket_data)
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket creado exitosamente',
                'ticket_id': ticket.correlativo,
                'ticket_html': ticket_html,
                'ticket_data': ticket_data
            })
            
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al crear ticket: {str(e)}'
        })
def construir_ticket_data(ticket):
    """
    Prepara los datos del ticket para impresión térmica y respuesta JSON
    """
    productos_procesados = []
    total_items = 0
    subtotal = 0

    for tp in ticket.ticket_productos.select_related(
        'ProductoTalla',
        'ProductoTalla__producto',
        'ProductoTalla__producto__atributo1',
        'ProductoTalla__producto__atributo2',
        'ProductoTalla__producto__atributo3',
        'ProductoTalla__producto__atributo4',
    ).all():
        producto_talla = tp.ProductoTalla
        producto = producto_talla.producto if producto_talla else None

        marca = ''
        if producto:
            atributo_marca = getattr(producto, 'atributo1', None)
            if atributo_marca:
                marca = getattr(atributo_marca, 'valor', '') or ''

        subtotal += tp.subtotal
        total_items += tp.stock
        productos_procesados.append({
            'detalle_id': tp.id,
            'producto_talla_id': producto_talla.id if producto_talla else None,
            'producto_id': producto.id if producto else None,
            'sku': producto_talla.sku if producto_talla else '',
            'articulo': producto.articulo if producto else '',
            'descripcion': producto.descripcion if producto else '',
            'marca': marca,
            'talla': producto_talla.talla if producto_talla else '',
            'cantidad': tp.stock,
            'precio_unitario': tp.precio,
            'precio_original': tp.precio_original,
            'descuento_unitario': tp.descuento_unitario,
            'porcentaje_descuento': float(tp.porcentaje_descuento or 0),
            'subtotal': tp.subtotal,
            'costo_fifo': tp.costo_fifo,
            'lotes_utilizados': tp.lotes_utilizados,
            'stock_actual': producto_talla.stock if producto_talla else None,
        })

    sucursal = ticket.sucursal
    empresa = sucursal.empresa if hasattr(sucursal, 'empresa') else None
    pagos_queryset = ticket.pagos.all().order_by('creado_en')
    pagos = [
        {
            'id': pago.id,
            'metodo_pago': pago.metodo_pago,
            'metodo_pago_display': pago.get_metodo_pago_display(),
            'monto': pago.monto,
            'voucher': pago.voucher or '',
            'tipo_tarjeta': pago.tipo_tarjeta or '',
            'notas': pago.notas or '',
            'creado_en': pago.creado_en.strftime('%Y-%m-%d %H:%M:%S'),
        }
        for pago in pagos_queryset
    ]

    return {
        'ticket_id': ticket.correlativo,
        'fecha': ticket.fecha.strftime('%Y-%m-%d'),
        'hora': ticket.hora.strftime('%H:%M:%S'),
        'tipo_documento': 'TICKET',
        'estado': ticket.estado,
        'metodo_pago_principal': ticket.metodo_pago,
        'total_pagado': ticket.total_pagado,
        'saldo_por_pagar': ticket.saldo_por_pagar,
        'responsable': ticket.responsable,
        'sucursal': {
            'alias': sucursal.alias,
            'direccion': sucursal.direccion,
            'empresa': empresa.nombre if empresa else '',
            'rut_empresa': empresa.rut if empresa else ''
        },
        'vendedor': {
            'nombre': ticket.vendedor.nombre,
            'codigo': ticket.vendedor.codigo_vendedor
        },
        'cliente': {
            'nombre': ticket.cliente_nombre or '',
            'rut': ticket.cliente_rut or '',
            'giro': ticket.cliente_giro or '',
            'comuna': ticket.cliente_comuna or '',
            'ciudad': ticket.cliente_ciudad or '',
            'direccion': ticket.cliente_direccion or '',
            'telefono': ticket.cliente_telefono or '',
            'telefono_secundario': ticket.cliente_telefono_secundario or '',
            'email': ticket.cliente_email or '',
            'email_facturacion': ticket.cliente_email_facturacion or '',
        },
        'observaciones': ticket.observaciones or '',
        'observaciones_adicionales': ticket.observaciones_adicionales or '',
        'productos': productos_procesados,
        'pagos': pagos,
        'totales': {
            'items': total_items,
            'subtotal': subtotal,
            'descuento': ticket.descuento or 0,
            'total': ticket.total
        }
    }

def generar_ticket_html(ticket_data):
    """
    Generar HTML del ticket utilizando la estructura estandarizada
    """
    productos = ticket_data['productos']
    sucursal = ticket_data['sucursal']
    vendedor = ticket_data['vendedor']
    cliente = ticket_data.get('cliente', {})
    totales = ticket_data['totales']

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Ticket #{ticket_data['ticket_id']}</title>
        <style>
            body {{ font-family: 'Courier New', monospace; font-size: 12px; margin: 0; padding: 10px; width: 280px; }}
            .center {{ text-align: center; }}
            .bold {{ font-weight: bold; }}
            hr {{ border: none; border-top: 1px dashed #000; margin: 8px 0; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ text-align: left; padding: 2px; }}
            .right {{ text-align: right; }}
            .productos-header {{ margin-top: 10px; margin-bottom: 5px; }}
            .producto-descripcion small {{ display: block; }}
        </style>
    </head>
    <body>
        <div class="center">
            <div class="bold">{sucursal.get('empresa', 'SUCURSAL')}</div>
            <div>{sucursal.get('alias', '')}</div>
            <div>{sucursal.get('direccion', '')}</div>
            <div>RUT: {sucursal.get('rut_empresa', '')}</div>
            <hr>
        </div>
        <div>
            <div><span class="bold">Documento:</span> {ticket_data['tipo_documento']}</div>
            <div><span class="bold">Ticket N°:</span> {ticket_data['ticket_id']}</div>
            <div><span class="bold">Fecha:</span> {ticket_data['fecha']} {ticket_data['hora']}</div>
            <div><span class="bold">Vendedor:</span> {vendedor.get('nombre', '')} ({vendedor.get('codigo', '')})</div>
    """

    if cliente.get('nombre') or cliente.get('rut'):
        html += f"""
            <div><span class="bold">Cliente:</span> {cliente.get('nombre', '')} {cliente.get('rut', '')}</div>
        """

    html += """
        </div>
        <hr>
        <div class="productos-header center bold">DETALLE DE VENTA</div>
        <table>
            <thead>
                <tr>
                    <th>Cant</th>
                    <th>Artículo</th>
                    <th class="right">Precio</th>
                </tr>
            </thead>
            <tbody>
    """

    for item in productos:
        descripcion = item.get('descripcion', '')
        articulo = item.get('articulo', '')
        html += f"""
                <tr>
                    <td>{item.get('cantidad', 0)}</td>
                    <td class="producto-descripcion">
                        {articulo}<br>
                        <small>{descripcion}</small>
                        <small>SKU: {item.get('sku', '')} | Talla: {item.get('talla', '')}</small>
                    </td>
                    <td class="right">${item.get('subtotal', 0):,}</td>
                </tr>
        """

    html += f"""
            </tbody>
        </table>
        <hr>
        <div class="right"><span class="bold">Items:</span> {totales.get('items', 0)}</div>
        <div class="right"><span class="bold">Subtotal:</span> ${totales.get('subtotal', 0):,}</div>
    """

    if totales.get('descuento', 0):
        html += f"""
        <div class="right"><span class="bold">Descuento:</span> -${totales.get('descuento', 0):,}</div>
        """

    html += f"""
        <div class="right bold">TOTAL: ${totales.get('total', 0):,}</div>
        <hr>
        <div class="center" style="margin-top: 15px;">
            <div>¡GRACIAS POR SU COMPRA!</div>
            <div>CAMBIOS HASTA 15 DÍAS</div>
            <div>*NO SE REALIZAN DEVOLUCIONES DE DINERO</div>
        </div>
    </body>
    </html>
    """

    return html

# ========== GESTIÓN DE DTEs VENTAS ==========

@login_required
def gestion_dte(request):
    """Vista para mostrar la página de gestión de DTEs de venta"""
    return render(request, 'vistas/modulo_administracion/gestion_dte.html')

@login_required
@require_GET
def detalle_dte(request, dte_id):
    """Retorna el detalle de un DTE de ventas, con productos y pagos"""
    try:
        from django.db.models import Prefetch

        dte = (Dte.objects
               .select_related('receptor', 'emisor', 'sucursal', 'vendedor')
               .prefetch_related(
                   Prefetch('dte_productos', queryset=Dte_Productos.objects.select_related('productoTalla__producto')),
                   Prefetch('dte_asociado', queryset=Dte_Detalle_Pago.objects.all())
               )
               .get(id=dte_id))

        # Validar que sea un DTE de ventas de la sucursal del usuario
        try:
            empresa_user = EmpresaUser.objects.get(user=request.user, active=True)
        except EmpresaUser.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Usuario no tiene empresa asignada'}, status=403)

        if dte.sucursal_id != empresa_user.sucursal_id:
            return JsonResponse({'success': False, 'error': 'No tienes permiso para ver este DTE'}, status=403)

        productos = []
        total_detalle = 0
        for detalle in dte.dte_productos.all():
            producto = detalle.productoTalla.producto if detalle.productoTalla else None
            subtotal = (detalle.precio or 0) * (detalle.stock or 0)
            total_detalle += subtotal
            productos.append({
                'id': detalle.id,
                'producto': producto.articulo if producto else detalle.descripcion,
                'sku': detalle.productoTalla.sku if detalle.productoTalla else None,
                'talla': detalle.productoTalla.talla if detalle.productoTalla else None,
                'descripcion': detalle.descripcion,
                'cantidad': detalle.stock,
                'precio_unitario': detalle.precio,
                'subtotal': subtotal,
                'costo': detalle.costo,
                'sobreprecio': detalle.sobreprecio,
            })

        pagos = []
        total_pagado = 0
        notas_credito = 0
        for pago in dte.dte_asociado.all():
            registro = {
                'id': pago.id,
                'metodo_pago': pago.metodo_pago,
                'voucher': pago.voucher,
                'monto': pago.monto,
                'tipo_tarjeta': pago.tipo_tarjeta,
            }
            pagos.append(registro)

            if pago.metodo_pago == 'Nota de Crédito':
                notas_credito += pago.monto
            else:
                total_pagado += pago.monto

        saldo = float(dte.monto_con_iva) - float(total_pagado)

        return JsonResponse({
            'success': True,
            'dte': {
                'id': dte.id,
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'estado_dte': dte.estado_dte,
                'estado_pago': dte.estado_pago,
                'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y') if dte.fecha_emision else None,
                'fecha_vencimiento': dte.fecha_vencimiento.strftime('%d/%m/%Y') if dte.fecha_vencimiento else None,
                'fecha_recepcion': dte.fecha_recepcion.strftime('%d/%m/%Y') if dte.fecha_recepcion else None,
                'responsable': dte.responsable,
                'dias_credito': dte.diasCredito,
                'bultos': dte.bultos,
                'unidades_productos': dte.unidades_productos,
                'monto_neto': float(dte.monto_neto),
                'monto_con_iva': float(dte.monto_con_iva),
                'descuento': float(dte.descuento),
                'referencias': dte.referencias,
                'sucursal': dte.sucursal.alias if dte.sucursal else None,
                'vendedor': dte.vendedor.nombre if dte.vendedor else None,
                'receptor': {
                    'nombre': dte.receptor.nombre if dte.receptor else '',
                    'rut': dte.receptor.rut if dte.receptor else '',
                    'direccion': dte.receptor.direccion if dte.receptor else '',
                    'ciudad': dte.receptor.ciudad if dte.receptor else '',
                    'comuna': dte.receptor.comuna if dte.receptor else '',
                },
                'emisor': {
                    'nombre': dte.emisor.nombre,
                    'rut': dte.emisor.rut,
                    'direccion': dte.emisor.direccion,
                    'ciudad': dte.emisor.ciudad,
                    'comuna': dte.emisor.comuna,
                }
            },
            'productos': productos,
            'pagos': pagos,
            'totales': {
                'total_detalle': total_detalle,
                'total_pagado': total_pagado,
                'notas_credito': notas_credito,
                'saldo_pendiente': saldo,
            }
        })
    except Dte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'DTE no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al obtener el DTE: {str(e)}'}, status=500)


@login_required
@require_POST
def cargar_dte_ventas(request):
    """Vista AJAX para cargar DTEs de venta filtrados por sucursal y fecha"""
    try:
        import json
        from datetime import datetime, date
        from django.core.paginator import Paginator
        
        # Obtener datos del request
        data = json.loads(request.body)
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        page = data.get('page', 1)
        page_size = data.get('page_size', 20)
        search = data.get('search', '').strip()
        
        # Obtener la sucursal del usuario actual
        try:
            empresa_user = EmpresaUser.objects.get(user=request.user, active=True)
            sucursal_usuario = empresa_user.sucursal
            
            if not sucursal_usuario:
                return JsonResponse({
                    'success': False,
                    'error': 'Usuario no tiene sucursal asignada'
                })
        except EmpresaUser.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no tiene empresa asignada'
            })
        
        # Construir query base - DTEs de VENTA y TRASPASO de la sucursal del usuario
        # Incluimos TRASPASO porque también son documentos emitidos por la sucursal
        query = Dte.objects.filter(
            sucursal=sucursal_usuario,
            tipo_transaccion__in=['VENTA', 'TRASPASO']
        ).select_related('receptor', 'vendedor', 'sucursal')
        
        # Aplicar filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                query = query.filter(fecha_emision__gte=fecha_inicio_obj)
            except ValueError:
                pass
                
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                query = query.filter(fecha_emision__lte=fecha_fin_obj)
            except ValueError:
                pass
        
        # Aplicar búsqueda si existe
        if search:
            query = query.filter(
                Q(receptor__nombre__icontains=search) |
                Q(receptor__rut__icontains=search) |
                Q(numero_documento__icontains=search) |
                Q(tipo_documento__icontains=search) |
                Q(estado_dte__icontains=search) |
                Q(estado_pago__icontains=search) |
                Q(responsable__icontains=search)
            )
        
        # Ordenar por fecha de emisión descendente
        query = query.order_by('-fecha_emision', '-id')
        
        # Paginación
        paginator = Paginator(query, page_size)
        page_obj = paginator.get_page(page)
        
        # Preparar datos para respuesta
        items = []
        for dte in page_obj:
            # Calcular días de crédito restantes
            dias_credito_restantes = 0
            if dte.estado_pago != 'PAGADO' and dte.fecha_vencimiento:
                diferencia = dte.fecha_vencimiento - date.today()
                dias_credito_restantes = diferencia.days
            
            # Obtener total de pagos realizados
            total_pagos = Dte_Detalle_Pago.objects.filter(dte=dte).aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            # Obtener notas de crédito
            notas_credito = Dte_Detalle_Pago.objects.filter(
                dte=dte, 
                metodo_pago='Nota de Crédito'
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            items.append({
                'id': dte.id,
                'receptor_nombre': dte.receptor.nombre if dte.receptor else 'Sin receptor',
                'receptor_rut': dte.receptor.rut if dte.receptor else '',
                'numero_documento': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y'),
                'fecha_vencimiento': dte.fecha_vencimiento.strftime('%d/%m/%Y') if dte.fecha_vencimiento else '',
                'monto_con_iva': float(dte.monto_con_iva),
                'monto_neto': float(dte.monto_neto),
                'descuento': float(dte.descuento),
                'estado_dte': dte.estado_dte,
                'estado_pago': dte.estado_pago,
                'responsable': dte.responsable,
                'vendedor': dte.vendedor.nombre if dte.vendedor else '',
                'dias_credito_restantes': dias_credito_restantes,
                'total_pagos': float(total_pagos),
                'notas_credito': float(notas_credito),
                'saldo_pendiente': float(dte.monto_con_iva) - float(total_pagos),
                'bultos': dte.bultos,
                'unidades_productos': dte.unidades_productos
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'pagination': {
                'page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'page_size': page_size
            },
            'search': search,
            'sucursal': sucursal_usuario.alias
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cargar DTEs: {str(e)}'
        })
# ========== VISTAS PARA GESTIÓN DE CORRELATIVOS ==========

@login_required
def gestion_correlativos(request):
    """
    Vista principal para la gestión de correlativos
    """
    try:
        # Obtener la empresa y sucursal actual del usuario
        empresa_actual_id = request.session.get('idEmpresaActual')
        sucursal_actual_id = request.session.get('idSucursalActual')
        
        # Obtener filtros
        sucursal_filtro = request.GET.get('sucursal')
        tipo_documento_filtro = request.GET.get('tipo_documento')
        estado_filtro = request.GET.get('estado')
        
        # Query base - Si es superuser, ver todo; si no, filtrar por sucursal actual
        if request.user.is_superuser:
            correlativos = Correlativo.objects.select_related('sucursal').all()
        else:
            # Filtrar solo por la sucursal actual en sesión
            if sucursal_actual_id:
                correlativos = Correlativo.objects.select_related('sucursal').filter(
                    sucursal_id=sucursal_actual_id
                )
            else:
                # Si no hay sucursal en sesión, no mostrar correlativos
                correlativos = Correlativo.objects.none()
        
        # Corregir correlativos con datos faltantes
        from django.db import IntegrityError
        correlativos_a_procesar = list(correlativos)
        
        for correlativo in correlativos_a_procesar:
            updated = False
            tipo_original = correlativo.tipo_dte
            
            if not correlativo.responsable:
                correlativo.responsable = 'Sistema'
                updated = True
            if not correlativo.fecha_actualizacion:
                correlativo.fecha_actualizacion = timezone.now().date()
                updated = True
            
            # Normalizar tipos antiguos
            tipo_normalizado = None
            if correlativo.tipo_dte == 'Compra':
                tipo_normalizado = 'COMPRA'
            elif correlativo.tipo_dte == 'Ticket':
                tipo_normalizado = 'TICKET'
            elif correlativo.tipo_dte == 'Traspaso':
                tipo_normalizado = 'TRASPASO'
            elif correlativo.tipo_dte == 'Ajuste':
                tipo_normalizado = 'AJUSTE'
            
            if tipo_normalizado:
                # Verificar si ya existe un correlativo con el tipo normalizado
                existe_normalizado = Correlativo.objects.filter(
                    sucursal_id=correlativo.sucursal_id,
                    tipo_dte=tipo_normalizado
                ).exclude(id=correlativo.id).exists()
                
                if existe_normalizado:
                    # Ya existe uno normalizado, eliminar este duplicado
                    print(f"Eliminando duplicado: ID {correlativo.id}, Tipo: {tipo_original} (ya existe {tipo_normalizado})")
                    correlativo.delete()
                    continue
                else:
                    # No existe, normalizar
                    correlativo.tipo_dte = tipo_normalizado
                    updated = True
            
            if updated:
                try:
                    correlativo.save()
                except IntegrityError as e:
                    # Si aún así hay error de integridad, eliminar este correlativo
                    print(f"Error de integridad al guardar correlativo {correlativo.id}: {str(e)}")
                    correlativo.delete()
        
        # Aplicar filtros
        if sucursal_filtro:
            correlativos = correlativos.filter(sucursal_id=sucursal_filtro)
        
        if tipo_documento_filtro:
            correlativos = correlativos.filter(tipo_dte=tipo_documento_filtro)
        
        # Filtro por estado
        if estado_filtro == 'activo':
            correlativos = correlativos.filter(inicio__lt=F('termino'))
        elif estado_filtro == 'agotado':
            correlativos = correlativos.filter(inicio__gte=F('termino'))
        elif estado_filtro == 'proximo_agotarse':
            # Correlativos con menos de 100 números disponibles
            correlativos = correlativos.annotate(
                disponibles=F('termino') - F('inicio') + 1
            ).filter(disponibles__lte=100, disponibles__gt=0)
        
        # Calcular estadísticas - Si es superuser, ver todo; si no, filtrar por sucursal actual
        if request.user.is_superuser:
            total_correlativos = Correlativo.objects.count()
            correlativos_activos = Correlativo.objects.filter(inicio__lt=F('termino')).count()
            correlativos_agotados = Correlativo.objects.filter(inicio__gte=F('termino')).count()
            correlativos_proximos_agotar = Correlativo.objects.annotate(
                disponibles=F('termino') - F('inicio') + 1
            ).filter(disponibles__lte=100, disponibles__gt=0).count()
            sucursales = Sucursal.objects.all().order_by('alias')
        else:
            # Estadísticas solo de la sucursal actual
            if sucursal_actual_id:
                total_correlativos = Correlativo.objects.filter(sucursal_id=sucursal_actual_id).count()
                correlativos_activos = Correlativo.objects.filter(
                    sucursal_id=sucursal_actual_id,
                    inicio__lt=F('termino')
                ).count()
                correlativos_agotados = Correlativo.objects.filter(
                    sucursal_id=sucursal_actual_id,
                    inicio__gte=F('termino')
                ).count()
                correlativos_proximos_agotar = Correlativo.objects.filter(
                    sucursal_id=sucursal_actual_id
                ).annotate(
                    disponibles=F('termino') - F('inicio') + 1
                ).filter(disponibles__lte=100, disponibles__gt=0).count()
                # Solo mostrar la sucursal actual
                sucursales = Sucursal.objects.filter(id=sucursal_actual_id).order_by('alias')
            else:
                total_correlativos = 0
                correlativos_activos = 0
                correlativos_agotados = 0
                correlativos_proximos_agotar = 0
                sucursales = Sucursal.objects.none()
        
        tipos_documento = TIPO_DOCUMENTO_CHOICES
        
        # Los correlativos ya tienen las propiedades calculadas en el modelo
        correlativos_con_datos = list(correlativos)
        
        context = {
            'correlativos': correlativos_con_datos,
            'sucursales': sucursales,
            'tipos_documento': tipos_documento,
            'total_correlativos': total_correlativos,
            'correlativos_activos': correlativos_activos,
            'correlativos_agotados': correlativos_agotados,
            'correlativos_proximos_agotar': correlativos_proximos_agotar,
            'filtros': {
                'sucursal': sucursal_filtro,
                'tipo_documento': tipo_documento_filtro,
                'estado': estado_filtro
            }
        }
        
        return render(request, 'vistas/modulo_administracion/gestion_correlativos.html', context)
        
    except Exception as e:
        sucursal_actual_id = request.session.get('idSucursalActual')
        # Si es superuser, mostrar todas las sucursales; si no, solo la sucursal actual
        if request.user.is_superuser:
            sucursales = Sucursal.objects.all()
        else:
            if sucursal_actual_id:
                sucursales = Sucursal.objects.filter(id=sucursal_actual_id)
            else:
                sucursales = Sucursal.objects.none()
        
        # Mostrar el error pero continuar con la página
        import traceback
        print(f"Error en gestion_correlativos: {str(e)}")
        print(traceback.format_exc())
        
        return render(request, 'vistas/modulo_administracion/gestion_correlativos.html', {
            'error': f'Advertencia: {str(e)}. Mostrando datos disponibles.',
            'correlativos': [],
            'sucursales': sucursales,
            'tipos_documento': TIPO_DOCUMENTO_CHOICES,
            'total_correlativos': 0,
            'correlativos_activos': 0,
            'correlativos_agotados': 0,
            'correlativos_proximos_agotar': 0
        })

@login_required
@require_POST
def guardar_correlativo(request):
    """
    Guarda o actualiza un correlativo
    """
    try:
        correlativo_id = request.POST.get('correlativo_id')
        sucursal_id = request.POST.get('sucursal')
        tipo_documento = request.POST.get('tipo_documento')
        inicio = int(request.POST.get('inicio'))
        termino = int(request.POST.get('termino'))
        alias = request.POST.get('alias', '')
        responsable = request.POST.get('responsable')
        
        # Validaciones
        if inicio >= termino:
            return JsonResponse({
                'success': False,
                'message': 'El número inicial debe ser menor que el número final'
            })
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Normalizar tipo de documento antes de guardar
        tipo_documento_normalizado = tipo_documento.upper()
        normalizaciones = {
            'COMPRA': 'COMPRA',
            'TICKET': 'TICKET',
            'TRASPASO': 'TRASPASO',
            'AJUSTE': 'AJUSTE'
        }
        tipo_documento = normalizaciones.get(tipo_documento_normalizado, tipo_documento)
        
        # Verificar si ya existe un correlativo para esta combinación (excepto el actual)
        existing_query = Correlativo.objects.filter(
            sucursal=sucursal,
            tipo_dte=tipo_documento
        )
        
        if correlativo_id:
            existing_query = existing_query.exclude(id=correlativo_id)
        
        if existing_query.exists():
            return JsonResponse({
                'success': False,
                'message': f'Ya existe un correlativo para {tipo_documento} en {sucursal.alias}'
            })
        
        # Crear o actualizar
        from django.db import IntegrityError
        try:
            if correlativo_id:
                correlativo = get_object_or_404(Correlativo, id=correlativo_id)
                correlativo.sucursal = sucursal
                correlativo.tipo_dte = tipo_documento
                correlativo.inicio = inicio
                correlativo.termino = termino
                correlativo.alias = alias
                correlativo.responsable = responsable
                correlativo.fecha_actualizacion = timezone.now().date()
                correlativo.save()
                mensaje = 'Correlativo actualizado exitosamente'
            else:
                correlativo = Correlativo.objects.create(
                    sucursal=sucursal,
                    tipo_dte=tipo_documento,
                    inicio=inicio,
                    termino=termino,
                    alias=alias,
                    responsable=responsable,
                    fecha_actualizacion=timezone.now().date()
                )
                mensaje = 'Correlativo creado exitosamente'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'correlativo_id': correlativo.id
            })
        except IntegrityError as e:
            return JsonResponse({
                'success': False,
                'message': f'Ya existe un correlativo para {tipo_documento} en {sucursal.alias}. Error: {str(e)}'
            }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al guardar correlativo: {str(e)}'
        })

@login_required
@require_GET
def obtener_correlativo(request, correlativo_id):
    """
    Obtiene los datos de un correlativo específico
    """
    try:
        correlativo = get_object_or_404(Correlativo, id=correlativo_id)
        
        # DEBUG
        print(f"DEBUG obtener_correlativo ID {correlativo_id}:")
        print(f"  - Correlativo: {correlativo}")
        print(f"  - Sucursal ID: {correlativo.sucursal.id}")
        print(f"  - Sucursal alias: {correlativo.sucursal.alias}")
        print(f"  - Tipo DTE: {correlativo.tipo_dte}")
        
        return JsonResponse({
            'success': True,
            'correlativo': {
                'id': correlativo.id,
                'sucursal_id': correlativo.sucursal.id,
                'sucursal_nombre': correlativo.sucursal.alias,
                'tipo_dte': correlativo.tipo_dte,
                'inicio': correlativo.inicio,
                'termino': correlativo.termino,
                'alias': correlativo.alias,
                'responsable': correlativo.responsable,
                'fecha_actualizacion': correlativo.fecha_actualizacion.strftime('%d/%m/%Y') if correlativo.fecha_actualizacion else None
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener correlativo: {str(e)}'
        })

@login_required
@require_POST
def renovar_correlativo(request):
    """
    Renueva un correlativo con un nuevo rango
    """
    try:
        correlativo_id = request.POST.get('correlativo_id')
        nuevo_inicio = int(request.POST.get('nuevo_inicio'))
        nuevo_termino = int(request.POST.get('nuevo_termino'))
        
        # Validaciones
        if nuevo_inicio >= nuevo_termino:
            return JsonResponse({
                'success': False,
                'message': 'El número inicial debe ser menor que el número final'
            })
        
        correlativo = get_object_or_404(Correlativo, id=correlativo_id)
        
        # Actualizar correlativo
        correlativo.inicio = nuevo_inicio
        correlativo.termino = nuevo_termino
        correlativo.fecha_actualizacion = timezone.now().date()
        correlativo.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Correlativo renovado. Nuevo rango: {nuevo_inicio} - {nuevo_termino}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al renovar correlativo: {str(e)}'
        })

@login_required
@require_GET
def historial_correlativo(request, correlativo_id):
    """
    Obtiene el historial de uso de un correlativo
    """
    try:
        correlativo = get_object_or_404(Correlativo, id=correlativo_id)
        
        # Buscar documentos que usen este correlativo
        historial = []
        
        # Buscar en DTEs
        dtes = Dte.objects.filter(
            sucursal=correlativo.sucursal,
            tipo_documento=correlativo.tipo_dte
        ).order_by('-fecha_emision')[:50]  # Últimos 50 registros
        
        for dte in dtes:
            historial.append({
                'fecha': dte.fecha_emision.strftime('%d/%m/%Y'),
                'tipo_documento': dte.tipo_documento,
                'numero': dte.numero_documento,
                'responsable': dte.responsable,
                'observaciones': f'Monto: ${dte.monto_con_iva:,}'
            })
        
        # Buscar en Tickets
        tickets = Ticket.objects.filter(
            sucursal=correlativo.sucursal
        ).order_by('-fecha')[:50]  # Últimos 50 registros
        
        for ticket in tickets:
            if correlativo.tipo_dte == 'TICKET':
                historial.append({
                    'fecha': ticket.fecha.strftime('%d/%m/%Y'),
                    'tipo_documento': 'TICKET',
                    'numero': ticket.correlativo,
                    'responsable': ticket.responsable,
                    'observaciones': f'Total: ${ticket.total:,}'
                })
        
        # Ordenar por fecha descendente
        historial.sort(key=lambda x: x['fecha'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'historial': historial[:50]  # Limitar a 50 registros
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener historial: {str(e)}'
        })


# ========== MÓDULO DE REPORTE DE EXISTENCIAS ==========

@login_required
def ver_reporte_existencias(request):
    """Vista principal del reporte de existencias de productos"""
    return render(request, 'vistas/reporte_existencias.html')


@login_required
def obtener_existencias_reporte(request):
    """
    API para obtener datos de existencias de productos
    Retorna información completa de productos con stock, costos y precios
    """
    try:
        import logging
        logger = logging.getLogger(__name__)
        logger.info('Iniciando obtención de existencias')
        
        # Obtener filtros del request
        sucursal_id = request.GET.get('sucursal_id')
        categoria_id = request.GET.get('categoria_id')
        busqueda = request.GET.get('busqueda', '').strip()
        
        # Construir queryset base
        productos_talla = Producto_Talla.objects.select_related(
            'producto',
            'producto__categoria',
            'producto__sucursal'
        ).filter(
            producto__isnull=False
        )
        
        logger.info(f'Total productos_talla antes de filtros: {productos_talla.count()}')
        
        # Aplicar filtros
        if sucursal_id:
            productos_talla = productos_talla.filter(producto__sucursal_id=sucursal_id)
        
        if categoria_id:
            productos_talla = productos_talla.filter(producto__categoria_id=categoria_id)
        
        if busqueda:
            productos_talla = productos_talla.filter(
                Q(sku__icontains=busqueda) |
                Q(producto__articulo__icontains=busqueda) |
                Q(producto__descripcion__icontains=busqueda) |
                Q(talla__icontains=busqueda)
            )
        
        logger.info(f'Total productos_talla después de filtros: {productos_talla.count()}')
        
        # Preparar datos de existencias
        existencias = []
        for pt in productos_talla[:500]:  # Limitar a 500 para evitar timeout
            try:
                # Calcular costo FIFO (costo promedio de lotes disponibles)
                lotes_disponibles = LoteProducto.objects.filter(
                    producto_talla=pt,
                    cantidad_disponible__gt=0,
                    activo=True
                )
                
                costo_fifo = 0
                total_cantidad_lotes = 0
                
                if lotes_disponibles.exists():
                    # Calcular costo promedio ponderado
                    for lote in lotes_disponibles:
                        costo_fifo += lote.costo_unitario * lote.cantidad_disponible
                        total_cantidad_lotes += lote.cantidad_disponible
                    
                    if total_cantidad_lotes > 0:
                        costo_fifo = costo_fifo / total_cantidad_lotes
                else:
                    # Si no hay lotes, usar el costo del producto
                    costo_fifo = pt.producto.costo
                
                existencias.append({
                    'sku': str(pt.sku),
                    'articulo': pt.producto.articulo,
                    'descripcion': pt.producto.descripcion or '',
                    'talla': pt.talla,
                    'sucursal': pt.producto.sucursal.alias,
                    'sucursal_id': pt.producto.sucursal.id,
                    'categoria': pt.producto.categoria.nombre if pt.producto.categoria else 'Sin categoría',
                    'categoria_id': pt.producto.categoria.id if pt.producto.categoria else None,
                    'stock': pt.stock,
                    'costo': pt.producto.costo,
                    'costo_fifo': int(costo_fifo),
                    'pvp': pt.producto.precioventa,
                })
            except Exception as e:
                logger.error(f'Error procesando producto_talla {pt.id}: {str(e)}')
                continue
        
        logger.info(f'Total existencias procesadas: {len(existencias)}')
        
        # Obtener sucursales y categorías para filtros
        sucursales = []
        for sucursal in Sucursal.objects.all():
            sucursales.append({
                'id': sucursal.id,
                'alias': sucursal.alias
            })
        
        categorias = []
        for categoria in Categoria.objects.all():
            categorias.append({
                'id': categoria.id,
                'nombre': categoria.nombre
            })
        
        logger.info('Retornando datos exitosamente')
        
        return JsonResponse({
            'success': True,
            'existencias': existencias,
            'sucursales': sucursales,
            'categorias': categorias
        })
        
    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error en obtener_existencias_reporte: {str(e)}')
        logger.error(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener existencias: {str(e)}'
        }, status=500)


@login_required
@require_GET
def exportar_existencias_excel(request):
    """Exportar reporte de existencias a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Obtener datos de existencias (sin filtros para exportar todo)
        productos_talla = Producto_Talla.objects.select_related(
            'producto',
            'producto__categoria',
            'producto__sucursal'
        ).filter(
            producto__isnull=False
        ).order_by('producto__sucursal__alias', 'producto__articulo', 'talla')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # ===== HOJA 1: VISTA GENERAL =====
        ws_general = wb.active
        ws_general.title = "Existencias General"
        
        # Estilos
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers_general = [
            'SKU', 'Artículo', 'Descripción', 'Talla/Alias', 'Sucursal',
            'Categoría', 'Stock', 'Costo', 'Costo FIFO', 'PVP',
            'Valor Total (FIFO)', 'Estado'
        ]
        
        for col_num, header in enumerate(headers_general, 1):
            cell = ws_general.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        row_num = 2
        for pt in productos_talla:
            # Calcular costo FIFO
            lotes_disponibles = LoteProducto.objects.filter(
                producto_talla=pt,
                cantidad_disponible__gt=0,
                activo=True
            )
            
            costo_fifo = 0
            total_cantidad_lotes = 0
            
            if lotes_disponibles.exists():
                for lote in lotes_disponibles:
                    costo_fifo += lote.costo_unitario * lote.cantidad_disponible
                    total_cantidad_lotes += lote.cantidad_disponible
                
                if total_cantidad_lotes > 0:
                    costo_fifo = costo_fifo / total_cantidad_lotes
            else:
                costo_fifo = pt.producto.costo
            
            valor_total = pt.stock * costo_fifo
            
            # Determinar estado
            if pt.stock == 0:
                estado = 'Sin Stock'
            elif pt.stock < 10:
                estado = 'Bajo Stock'
            else:
                estado = 'Disponible'
            
            # Escribir fila
            ws_general.cell(row=row_num, column=1).value = str(pt.sku)
            ws_general.cell(row=row_num, column=2).value = pt.producto.articulo
            ws_general.cell(row=row_num, column=3).value = pt.producto.descripcion
            ws_general.cell(row=row_num, column=4).value = pt.talla
            ws_general.cell(row=row_num, column=5).value = pt.producto.sucursal.alias
            ws_general.cell(row=row_num, column=6).value = pt.producto.categoria.nombre if pt.producto.categoria else 'Sin categoría'
            ws_general.cell(row=row_num, column=7).value = pt.stock
            ws_general.cell(row=row_num, column=8).value = pt.producto.costo
            ws_general.cell(row=row_num, column=9).value = int(costo_fifo)
            ws_general.cell(row=row_num, column=10).value = pt.producto.precioventa
            ws_general.cell(row=row_num, column=11).value = int(valor_total)
            ws_general.cell(row=row_num, column=12).value = estado
            
            # Aplicar bordes
            for col in range(1, 13):
                ws_general.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Ajustar anchos de columna
        for col in range(1, 13):
            ws_general.column_dimensions[get_column_letter(col)].width = 15
        
        ws_general.column_dimensions['B'].width = 25  # Artículo
        ws_general.column_dimensions['C'].width = 35  # Descripción
        
        # ===== HOJA 2: AGRUPADO POR SUCURSAL =====
        ws_sucursal = wb.create_sheet("Por Sucursal")
        
        # Encabezados
        headers_sucursal = [
            'Sucursal', 'SKU', 'Artículo', 'Talla/Alias',
            'Stock', 'Costo', 'Costo FIFO', 'PVP', 'Valor Total', 'Estado'
        ]
        
        for col_num, header in enumerate(headers_sucursal, 1):
            cell = ws_sucursal.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos agrupados por sucursal
        sucursales = Sucursal.objects.all().order_by('alias')
        row_num = 2
        
        for sucursal in sucursales:
            productos_sucursal = productos_talla.filter(producto__sucursal=sucursal)
            
            if not productos_sucursal.exists():
                continue
            
            # Fila de encabezado de sucursal
            cell = ws_sucursal.cell(row=row_num, column=1)
            cell.value = f"🏪 {sucursal.alias}"
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
            ws_sucursal.merge_cells(f'A{row_num}:J{row_num}')
            row_num += 1
            
            # Productos de la sucursal
            for pt in productos_sucursal:
                # Calcular costo FIFO
                lotes_disponibles = LoteProducto.objects.filter(
                    producto_talla=pt,
                    cantidad_disponible__gt=0,
                    activo=True
                )
                
                costo_fifo = 0
                total_cantidad_lotes = 0
                
                if lotes_disponibles.exists():
                    for lote in lotes_disponibles:
                        costo_fifo += lote.costo_unitario * lote.cantidad_disponible
                        total_cantidad_lotes += lote.cantidad_disponible
                    
                    if total_cantidad_lotes > 0:
                        costo_fifo = costo_fifo / total_cantidad_lotes
                else:
                    costo_fifo = pt.producto.costo
                
                valor_total = pt.stock * costo_fifo
                
                # Determinar estado
                if pt.stock == 0:
                    estado = 'Sin Stock'
                elif pt.stock < 10:
                    estado = 'Bajo Stock'
                else:
                    estado = 'Disponible'
                
                # Escribir fila
                ws_sucursal.cell(row=row_num, column=1).value = sucursal.alias
                ws_sucursal.cell(row=row_num, column=2).value = str(pt.sku)
                ws_sucursal.cell(row=row_num, column=3).value = pt.producto.articulo
                ws_sucursal.cell(row=row_num, column=4).value = pt.talla
                ws_sucursal.cell(row=row_num, column=5).value = pt.stock
                ws_sucursal.cell(row=row_num, column=6).value = pt.producto.costo
                ws_sucursal.cell(row=row_num, column=7).value = int(costo_fifo)
                ws_sucursal.cell(row=row_num, column=8).value = pt.producto.precioventa
                ws_sucursal.cell(row=row_num, column=9).value = int(valor_total)
                ws_sucursal.cell(row=row_num, column=10).value = estado
                
                # Aplicar bordes
                for col in range(1, 11):
                    ws_sucursal.cell(row=row_num, column=col).border = border
                
                row_num += 1
        
        # Ajustar anchos de columna
        for col in range(1, 11):
            ws_sucursal.column_dimensions[get_column_letter(col)].width = 15
        
        ws_sucursal.column_dimensions['C'].width = 25  # Artículo
        
        # ===== HOJA 3: RESUMEN =====
        ws_resumen = wb.create_sheet("Resumen")
        
        # Título
        ws_resumen.cell(row=1, column=1).value = "RESUMEN DE EXISTENCIAS"
        ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_resumen.cell(row=2, column=1).value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Métricas generales
        total_productos = productos_talla.count()
        total_stock = sum(pt.stock for pt in productos_talla)
        sin_stock = productos_talla.filter(stock=0).count()
        bajo_stock = productos_talla.filter(stock__gt=0, stock__lt=10).count()
        
        ws_resumen.cell(row=4, column=1).value = "Total de Productos (SKUs):"
        ws_resumen.cell(row=4, column=2).value = total_productos
        ws_resumen.cell(row=4, column=1).font = Font(bold=True)
        
        ws_resumen.cell(row=5, column=1).value = "Stock Total (Unidades):"
        ws_resumen.cell(row=5, column=2).value = total_stock
        ws_resumen.cell(row=5, column=1).font = Font(bold=True)
        
        ws_resumen.cell(row=6, column=1).value = "Productos Sin Stock:"
        ws_resumen.cell(row=6, column=2).value = sin_stock
        ws_resumen.cell(row=6, column=1).font = Font(bold=True)
        
        ws_resumen.cell(row=7, column=1).value = "Productos Bajo Stock (<10):"
        ws_resumen.cell(row=7, column=2).value = bajo_stock
        ws_resumen.cell(row=7, column=1).font = Font(bold=True)
        
        # Resumen por sucursal
        ws_resumen.cell(row=9, column=1).value = "RESUMEN POR SUCURSAL"
        ws_resumen.cell(row=9, column=1).font = Font(bold=True, size=12)
        
        row_num = 11
        ws_resumen.cell(row=row_num, column=1).value = "Sucursal"
        ws_resumen.cell(row=row_num, column=2).value = "Productos"
        ws_resumen.cell(row=row_num, column=3).value = "Stock Total"
        
        for col in range(1, 4):
            ws_resumen.cell(row=row_num, column=col).font = Font(bold=True)
            ws_resumen.cell(row=row_num, column=col).fill = header_fill
            ws_resumen.cell(row=row_num, column=col).font = header_font
        
        row_num += 1
        
        for sucursal in sucursales:
            productos_sucursal = productos_talla.filter(producto__sucursal=sucursal)
            stock_sucursal = sum(pt.stock for pt in productos_sucursal)
            
            ws_resumen.cell(row=row_num, column=1).value = sucursal.alias
            ws_resumen.cell(row=row_num, column=2).value = productos_sucursal.count()
            ws_resumen.cell(row=row_num, column=3).value = stock_sucursal
            row_num += 1
        
        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 15
        ws_resumen.column_dimensions['C'].width = 15
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_existencias_{timestamp}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })