"""
Módulo de Gestión de Precios - RetailMind
Sistema avanzado de gestión de precios con recomendaciones inteligentes
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Sum, F, Q, Avg, Count, ExpressionWrapper, DecimalField, Min, Max
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
import json
import logging
from datetime import datetime, timedelta

from .models import (
    Producto, Producto_Talla, LoteProducto, Categoria, AtributoOpcion,
    Sucursal, Movimientos_Producto, Ticket_Productos, Ticket,
    CambioPrecioPendiente, NotificacionCambioPrecio, HistorialCambioPrecio,
    ParametroGlobal, EmpresaUser
)

logger = logging.getLogger('app')


# ========== VISTAS PRINCIPALES ==========

@login_required
def gestion_precios_view(request):
    """Vista principal del módulo de gestión de precios"""
    return render(request, 'vistas/modulo_existencias/gestion_precios.html')


@login_required
def revisar_cambios_precios_view(request):
    """Vista para revisar y aprobar cambios de precios pendientes"""
    return render(request, 'vistas/modulo_existencias/revisar_cambios_precios.html')


@login_required
def edicion_rapida_precios_view(request):
    """Vista de edición rápida con navegación por Tab"""
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import EmpresaUser
    
    # 🔍 VERIFICAR SESIÓN DE SUCURSAL
    sucursal_id = request.session.get('idSucursalActual')
    
    if not sucursal_id:
        # Intentar obtener sucursal del usuario
        try:
            empresa_user = EmpresaUser.objects.filter(
                user=request.user,
                status=True
            ).select_related('sucursal', 'empresa').first()
            
            if empresa_user and empresa_user.sucursal:
                # Establecer en sesión
                request.session['idSucursalActual'] = empresa_user.sucursal.id
                request.session['idEmpresaActual'] = empresa_user.empresa.id
                request.session['alias'] = empresa_user.sucursal.alias
                request.session['nombreEmpresaActual'] = empresa_user.empresa.razon_social
                sucursal_id = empresa_user.sucursal.id
                logger.info("Sesion de precios inicializada: usuario_id=%s, sucursal_id=%s", request.user.id, sucursal_id)
            else:
                # Redirigir a selección de sucursal
                messages.error(request, 'Por favor selecciona una sucursal para continuar')
                return redirect('seleccionar_empresa_sucursal')
        except Exception as e:
            logger.exception("Error al obtener sucursal para edicion rapida de precios usuario_id=%s", request.user.id)
            messages.error(request, f'Error al obtener sucursal: {str(e)}')
            return redirect('verHome')
    
    context = {
        'sucursal_actual': sucursal_id,
        'alias_sucursal': request.session.get('alias', ''),
    }
    
    return render(request, 'vistas/modulo_existencias/edicion_rapida_precios.html', context)


@login_required
def debug_session_precios(request):
    """Endpoint temporal para verificar sesión (SOLO PARA DEBUG)"""
    from .models import EmpresaUser
    
    # Información de sesión
    session_data = {
        'idSucursalActual': request.session.get('idSucursalActual'),
        'idEmpresaActual': request.session.get('idEmpresaActual'),
        'alias': request.session.get('alias'),
        'nombreEmpresaActual': request.session.get('nombreEmpresaActual'),
        'session_keys': list(request.session.keys()),
        'user': request.user.username,
    }
    
    # Información de EmpresaUser
    try:
        empresa_user = EmpresaUser.objects.filter(
            user=request.user,
            status=True
        ).select_related('sucursal', 'empresa').first()
        
        if empresa_user:
            session_data['empresa_user'] = {
                'empresa_id': empresa_user.empresa.id if empresa_user.empresa else None,
                'empresa_nombre': empresa_user.empresa.razon_social if empresa_user.empresa else None,
                'sucursal_id': empresa_user.sucursal.id if empresa_user.sucursal else None,
                'sucursal_alias': empresa_user.sucursal.alias if empresa_user.sucursal else None,
                'status': empresa_user.status,
            }
        else:
            session_data['empresa_user'] = None
    except Exception as e:
        session_data['empresa_user_error'] = str(e)
    
    return JsonResponse(session_data)


# ========== ESTADÍSTICAS GENERALES ==========

@require_GET
@login_required
def obtener_estadisticas(request):
    """Obtener estadísticas generales para el dashboard"""
    try:
        sucursal_id = request.session.get('idSucursalActual')
        
        # Filtrar por sucursal si está seleccionada
        queryset_productos = Producto.objects.all()
        queryset_tallas = Producto_Talla.objects.all()
        
        if sucursal_id:
            queryset_productos = queryset_productos.filter(sucursal_id=sucursal_id)
            queryset_tallas = queryset_tallas.filter(producto__sucursal_id=sucursal_id)
        
        # Total de productos activos
        total_productos = queryset_tallas.count()
        
        # Valor total del inventario
        valor_inventario = 0
        suma_margenes = 0
        count_margenes = 0
        
        for pt in queryset_tallas:
            # Calcular valor del inventario basado en lotes FIFO
            lotes = LoteProducto.objects.filter(
                producto_talla=pt,
                cantidad_disponible__gt=0,
                activo=True
            )
            
            for lote in lotes:
                valor_inventario += lote.cantidad_disponible * lote.costo_unitario
                
                # Calcular margen
                if lote.precio_venta_unitario > 0:
                    margen = ((lote.precio_venta_unitario - lote.costo_unitario) / lote.precio_venta_unitario) * 100
                    suma_margenes += margen
                    count_margenes += 1
        
        margen_promedio = suma_margenes / count_margenes if count_margenes > 0 else 0
        
        # Inventario antiguo (más de 365 días)
        fecha_limite = timezone.now() - timedelta(days=365)
        inventario_antiguo = LoteProducto.objects.filter(
            fecha_ingreso__lt=fecha_limite,
            cantidad_disponible__gt=0,
            activo=True
        )
        
        if sucursal_id:
            inventario_antiguo = inventario_antiguo.filter(
                producto_talla__producto__sucursal_id=sucursal_id
            )
        
        inventario_antiguo_count = inventario_antiguo.count()
        
        return JsonResponse({
            'success': True,
            'total_productos': total_productos,
            'valor_inventario': float(valor_inventario),
            'margen_promedio': float(margen_promedio),
            'inventario_antiguo': inventario_antiguo_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estadísticas: {str(e)}'
        })


# ========== BÚSQUEDA Y FILTRADO DE PRODUCTOS ==========

@require_GET
@login_required
def buscar_productos(request):
    """Buscar productos con filtros avanzados (agrupados por producto, no por talla)"""
    try:
        # Parámetros de búsqueda
        search = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria')
        marca_id = request.GET.get('marca')
        sucursal_id = request.GET.get('sucursal') or request.session.get('idSucursalActual')
        precio_min = request.GET.get('precio_min')
        precio_max = request.GET.get('precio_max')
        margen_min = request.GET.get('margen_min')
        stock_min = request.GET.get('stock_min')
        antiguedad = request.GET.get('antiguedad')
        anio = request.GET.get('anio')
        
        logger.debug(
            "Busqueda productos precios: search=%s, sucursal_get=%s, sucursal_session=%s, "
            "sucursal_final=%s, usuario=%s",
            search,
            request.GET.get('sucursal'),
            request.session.get('idSucursalActual'),
            sucursal_id,
            request.user.username,
        )
        
        # 🚨 VALIDAR QUE HAYA SUCURSAL (OBLIGATORIO)
        if not sucursal_id:
            logger.warning("Busqueda productos precios sin sucursal activa usuario=%s", request.user.username)
            return JsonResponse({
                'success': False,
                'error': 'No hay sucursal activa en la sesión. Por favor, selecciona una sucursal.',
                'productos': [],
                'total': 0,
                'debug': {
                    'session_keys': list(request.session.keys()),
                    'user': request.user.username,
                }
            }, status=400)
        
        # Paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        
        # Construir queryset base - AGRUPADO POR PRODUCTO
        queryset = Producto.objects.select_related(
            'categoria',
            'atributo1',
            'atributo2',
            'atributo3',
            'atributo4',
            'sucursal'
        ).prefetch_related('producto_talla').all()
        
        # Filtro de búsqueda por texto
        if search:
            queryset = queryset.filter(
                Q(articulo__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(producto_talla__sku__icontains=search)
            ).distinct()
        
        # Filtro por categoría
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)
        
        # Filtro por marca (atributo1)
        if marca_id:
            queryset = queryset.filter(atributo1_id=marca_id)
        
        # Filtro por sucursal (OBLIGATORIO - ya validado arriba)
        queryset = queryset.filter(sucursal_id=sucursal_id)
        total_queryset_inicial = queryset.count()
        logger.debug("Busqueda productos precios filtrada por sucursal_id=%s total_inicial=%s", sucursal_id, total_queryset_inicial)
        
        # Preparar datos para respuesta
        productos_data = []
        productos_excluidos = {
            'sin_tallas': 0,
            'sin_stock_ni_lotes': 0,
            'stock_minimo': 0,
            'precio': 0,
            'margen': 0,
            'antiguedad': 0,
            'anio': 0
        }
        
        for producto in queryset:
            # Obtener todas las tallas del producto
            tallas = producto.producto_talla.all()
            
            if not tallas.exists():
                logger.debug("Producto excluido de precios sin tallas: producto_id=%s articulo=%s", producto.id, producto.articulo)
                productos_excluidos['sin_tallas'] += 1
                continue
            
            # Calcular totales de todas las tallas
            stock_total = 0
            costo_total_ponderado = 0
            cantidad_total = 0
            fecha_ingreso_mas_antiguo = None
            tallas_list = []
            
            for pt in tallas:
                stock_total += pt.stock
                tallas_list.append(pt.talla)
                
                # Calcular datos de lotes para cada talla
                lotes = LoteProducto.objects.filter(
                    producto_talla=pt,
                    cantidad_disponible__gt=0,
                    activo=True
                )
                
                for lote in lotes:
                    costo_total_ponderado += lote.cantidad_disponible * lote.costo_unitario
                    cantidad_total += lote.cantidad_disponible
                    
                    if fecha_ingreso_mas_antiguo is None or lote.fecha_ingreso < fecha_ingreso_mas_antiguo:
                        fecha_ingreso_mas_antiguo = lote.fecha_ingreso
            
            # ✅ CAMBIO: No excluir productos sin lotes si tienen stock
            # Si no hay lotes pero hay stock, usar el costo del producto
            if cantidad_total == 0 and stock_total == 0:
                # Solo excluir si NO tiene stock ni lotes
                logger.debug("Producto excluido de precios sin stock ni lotes: producto_id=%s articulo=%s", producto.id, producto.articulo)
                productos_excluidos['sin_stock_ni_lotes'] += 1
                continue
            
            # Filtro por stock mínimo
            if stock_min and stock_total < int(stock_min):
                logger.debug(
                    "Producto excluido de precios por stock minimo: producto_id=%s stock=%s stock_min=%s",
                    producto.id,
                    stock_total,
                    stock_min,
                )
                productos_excluidos['stock_minimo'] += 1
                continue
            
            # Calcular costo promedio
            if cantidad_total > 0:
                # Tiene lotes: usar costo ponderado de lotes
                costo_promedio = costo_total_ponderado / cantidad_total
            else:
                # No tiene lotes pero tiene stock: usar costo del producto
                costo_promedio = float(producto.costo) if producto.costo else 0
            
            precio_venta = producto.precioventa
            
            # Aplicar filtros de precio
            if precio_min and precio_venta < float(precio_min):
                logger.debug(
                    "Producto excluido de precios por precio minimo: producto_id=%s precio=%s precio_min=%s",
                    producto.id,
                    precio_venta,
                    precio_min,
                )
                productos_excluidos['precio'] += 1
                continue
            if precio_max and precio_venta > float(precio_max):
                logger.debug(
                    "Producto excluido de precios por precio maximo: producto_id=%s precio=%s precio_max=%s",
                    producto.id,
                    precio_venta,
                    precio_max,
                )
                productos_excluidos['precio'] += 1
                continue
            
            # Calcular margen
            margen = ((precio_venta - costo_promedio) / precio_venta * 100) if precio_venta > 0 else 0
            
            # Aplicar filtro de margen
            if margen_min and margen < float(margen_min):
                logger.debug(
                    "Producto excluido de precios por margen minimo: producto_id=%s margen=%.2f margen_min=%s",
                    producto.id,
                    margen,
                    margen_min,
                )
                productos_excluidos['margen'] += 1
                continue
            
            # Calcular antigüedad
            dias_inventario = 0
            if fecha_ingreso_mas_antiguo:
                dias_inventario = (timezone.now() - fecha_ingreso_mas_antiguo).days
            
            # Aplicar filtro de antigüedad
            if antiguedad:
                if antiguedad == 'nuevo' and dias_inventario >= 180:
                    logger.debug(
                        "Producto excluido de precios por antiguedad: producto_id=%s dias=%s filtro=%s",
                        producto.id,
                        dias_inventario,
                        antiguedad,
                    )
                    productos_excluidos['antiguedad'] += 1
                    continue
                elif antiguedad == 'medio' and (dias_inventario < 180 or dias_inventario >= 365):
                    logger.debug(
                        "Producto excluido de precios por antiguedad: producto_id=%s dias=%s filtro=%s",
                        producto.id,
                        dias_inventario,
                        antiguedad,
                    )
                    productos_excluidos['antiguedad'] += 1
                    continue
                elif antiguedad == 'antiguo' and dias_inventario < 365:
                    logger.debug(
                        "Producto excluido de precios por antiguedad: producto_id=%s dias=%s filtro=%s",
                        producto.id,
                        dias_inventario,
                        antiguedad,
                    )
                    productos_excluidos['antiguedad'] += 1
                    continue
            
            # Aplicar filtro de año
            if anio and fecha_ingreso_mas_antiguo:
                if fecha_ingreso_mas_antiguo.year != int(anio):
                    logger.debug(
                        "Producto excluido de precios por anio: producto_id=%s anio_producto=%s anio_filtro=%s",
                        producto.id,
                        fecha_ingreso_mas_antiguo.year,
                        anio,
                    )
                    productos_excluidos['anio'] += 1
                    continue
            
            # Buscar último cambio de precio
            ultimo_cambio = HistorialCambioPrecio.objects.filter(
                producto=producto
            ).select_related('usuario').first()

            # Buscar productos similares en otras sucursales (con stock agregado por sucursal)
            productos_similares = Producto.objects.filter(
                articulo=producto.articulo,
                atributo1=producto.atributo1,
                atributo2=producto.atributo2
            ).exclude(sucursal=producto.sucursal).select_related('sucursal').annotate(
                stock_sucursal=Sum('producto_talla__stock')
            )

            # ===== FECHAS DE TRAZABILIDAD =====
            # 1) fecha_creacion del producto en ESTA sucursal (campo directo).
            # 2) fecha_ultimo_despacho: último INGRESO de stock considerando TODAS
            #    las bodegas (sucursal actual + todas las similares con mismo
            #    articulo y atributos). Útil para decisiones de descuento/precio.
            productos_red_ids = list(productos_similares.values_list('id', flat=True)) + [producto.id]
            agg_ingreso = Movimientos_Producto.objects.filter(
                ProductoTalla__producto_id__in=productos_red_ids,
                tipo_movimiento='INGRESO',
            ).aggregate(ultima_fecha=Max('fecha'))
            fecha_ultimo_despacho_red = agg_ingreso['ultima_fecha']
            dias_desde_ultimo_despacho = (
                (timezone.localdate() - fecha_ultimo_despacho_red).days
                if fecha_ultimo_despacho_red else None
            )
            fecha_creacion_local = producto.fecha_creacion.date() if producto.fecha_creacion else None

            # Detalle de stock por sucursal: incluye la sucursal actual + las similares
            sucursales_detalle = []
            if producto.sucursal:
                sucursales_detalle.append({
                    'alias': producto.sucursal.alias,
                    'stock': int(stock_total or 0),
                    'es_actual': True,
                })
            for p in productos_similares:
                if p.sucursal:
                    sucursales_detalle.append({
                        'alias': p.sucursal.alias,
                        'stock': int(p.stock_sucursal or 0),
                        'es_actual': False,
                    })

            # Stock total sumado en toda la red de sucursales (actual + similares)
            stock_total_red = sum(s['stock'] for s in sucursales_detalle)

            # Mantener compatibilidad: lista/cantidad de sucursales "similares" (sin la actual)
            sucursales_lista = [s['alias'] for s in sucursales_detalle if not s['es_actual']]
            sucursales_count = len(sucursales_lista)
            
            # Agregar a resultados
            logger.debug(
                "Producto incluido en busqueda precios: producto_id=%s articulo=%s stock=%s lotes=%s margen=%.2f",
                producto.id,
                producto.articulo,
                stock_total,
                cantidad_total,
                margen,
            )
            
            productos_data.append({
                'id': producto.id,  # ID del producto (no de la talla)
                'sku': ', '.join([str(t.sku) for t in tallas[:3]]),  # Primeros 3 SKUs
                'nombre': producto.articulo,
                'descripcion': producto.descripcion or '',
                'talla': f"{len(tallas_list)} tallas: {', '.join(str(t) for t in tallas_list[:5])}",  # Mostrar tallas
                'categoria': producto.categoria.nombre if producto.categoria else None,
                'marca': producto.atributo1.valor if producto.atributo1 else None,
                'color': producto.atributo2.valor if producto.atributo2 else None,
                'genero': producto.atributo3.valor if producto.atributo3 else None,
                'otro': producto.atributo4.valor if producto.atributo4 else None,
                'temporada': producto.temporada or None,
                'anio_temporada': producto.anio_temporada or None,
                'rango_precio': producto.rango_precio or None,
                'sucursal': producto.sucursal.alias,
                'costo': float(costo_promedio),
                'precio_venta': float(precio_venta),
                'stock': stock_total,
                'dias_inventario': dias_inventario,
                'margen': float(margen),
                'cantidad_tallas': len(tallas_list),
                'ultimo_cambio': {
                    'usuario': ultimo_cambio.usuario.username if ultimo_cambio and ultimo_cambio.usuario else None,
                    'fecha': ultimo_cambio.fecha_cambio.strftime('%d/%m/%Y %H:%M') if ultimo_cambio else None,
                    'hace_cuanto': ultimo_cambio.hace_cuanto if ultimo_cambio else None
                } if ultimo_cambio else None,
                'sucursales_similares': sucursales_count,
                'sucursales_lista': sucursales_lista,
                'sucursales_detalle': sucursales_detalle,
                'stock_total_red': stock_total_red,
                # Fechas de trazabilidad
                'fecha_creacion': fecha_creacion_local.strftime('%d/%m/%Y') if fecha_creacion_local else None,
                'fecha_ultimo_despacho': fecha_ultimo_despacho_red.strftime('%d/%m/%Y') if fecha_ultimo_despacho_red else None,
                'dias_desde_ultimo_despacho': dias_desde_ultimo_despacho,
            })
        
        # Paginación manual
        paginator = Paginator(productos_data, per_page)
        page_obj = paginator.get_page(page)
        
        # Resumen de logging
        total_excluidos = sum(productos_excluidos.values())
        logger.info(
            "Busqueda productos precios completada: sucursal_id=%s incluidos=%s excluidos=%s razones=%s pagina=%s total_paginas=%s",
            sucursal_id,
            len(productos_data),
            total_excluidos,
            {razon: cantidad for razon, cantidad in productos_excluidos.items() if cantidad > 0},
            page_obj.number,
            paginator.num_pages,
        )
        
        return JsonResponse({
            'success': True,
            'productos': list(page_obj),
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar productos: {str(e)}'
        })


# ========== SISTEMA DE RECOMENDACIONES INTELIGENTES ==========

@require_GET
@login_required
def obtener_recomendaciones(request, producto_id):
    """
    Sistema de recomendaciones inteligentes de precio
    Analiza múltiples factores para sugerir el precio óptimo
    Trabaja a nivel de producto (todas las tallas)
    """
    try:
        producto = Producto.objects.select_related('categoria', 'atributo1', 'sucursal').get(id=producto_id)
        tallas = producto.producto_talla.all()
        
        if not tallas.exists():
            return JsonResponse({
                'success': False,
                'error': 'No hay tallas para este producto'
            })
        
        # Obtener todos los lotes de todas las tallas
        lotes = LoteProducto.objects.filter(
            producto_talla__producto=producto,
            activo=True
        ).order_by('fecha_ingreso')
        
        lotes_disponibles = lotes.filter(cantidad_disponible__gt=0)
        
        if not lotes_disponibles.exists():
            return JsonResponse({
                'success': False,
                'error': 'No hay lotes disponibles para este producto'
            })
        
        # === ANÁLISIS ACTUAL ===
        costo_total = 0
        cantidad_total = 0
        precio_venta_promedio = 0
        dias_total = 0
        
        for lote in lotes_disponibles:
            costo_total += lote.cantidad_disponible * lote.costo_unitario
            cantidad_total += lote.cantidad_disponible
            precio_venta_promedio += lote.precio_venta_unitario
            dias_inventario = (timezone.now() - lote.fecha_ingreso).days
            dias_total += dias_inventario * lote.cantidad_disponible
        
        costo_promedio = costo_total / cantidad_total if cantidad_total > 0 else 0
        precio_actual = precio_venta_promedio / lotes_disponibles.count() if lotes_disponibles.count() > 0 else 0
        margen_actual = ((precio_actual - costo_promedio) / precio_actual * 100) if precio_actual > 0 else 0
        dias_promedio = dias_total / cantidad_total if cantidad_total > 0 else 0
        
        # === ANÁLISIS DE VENTAS (ÚLTIMOS 30 DÍAS) ===
        fecha_inicio = timezone.now() - timedelta(days=30)
        
        # Sumar ventas de todas las tallas del producto
        ventas_recientes = Ticket_Productos.objects.filter(
            ProductoTalla__producto=producto,
            idTicket__fecha__gte=fecha_inicio.date(),
            idTicket__estado='PAGADO'
        ).aggregate(
            total_vendido=Sum('stock')
        )
        
        ventas_30_dias = ventas_recientes['total_vendido'] or 0
        
        # Ventas de los 30 días anteriores para comparar tendencia
        fecha_inicio_anterior = fecha_inicio - timedelta(days=30)
        ventas_anteriores = Ticket_Productos.objects.filter(
            ProductoTalla__producto=producto,
            idTicket__fecha__gte=fecha_inicio_anterior.date(),
            idTicket__fecha__lt=fecha_inicio.date(),
            idTicket__estado='PAGADO'
        ).aggregate(
            total_vendido=Sum('stock')
        )
        
        ventas_periodo_anterior = ventas_anteriores['total_vendido'] or 0
        
        # === CALCULAR FACTORES DE AJUSTE ===
        
        # Factor 1: Antigüedad del Inventario
        factor_antiguedad = 0
        if dias_promedio > 365:  # Más de 1 año
            factor_antiguedad = -0.20  # Descuento del 20%
        elif dias_promedio > 180:  # 6 meses a 1 año
            factor_antiguedad = -0.10  # Descuento del 10%
        elif dias_promedio > 90:  # 3 a 6 meses
            factor_antiguedad = -0.05  # Descuento del 5%
        else:  # Menos de 3 meses
            factor_antiguedad = 0  # Sin ajuste
        
        # Factor 2: Rotación de Inventario
        if cantidad_total > 0:
            dias_para_agotar = (cantidad_total / (ventas_30_dias / 30)) if ventas_30_dias > 0 else 999
        else:
            dias_para_agotar = 0
        
        factor_rotacion = 0
        velocidad_rotacion = "Sin ventas"
        
        if ventas_30_dias == 0:
            factor_rotacion = -0.15  # Sin ventas, bajar precio
            velocidad_rotacion = "Sin ventas (descuento necesario)"
        elif dias_para_agotar < 30:
            factor_rotacion = 0.05  # Se vende rápido, puede subir precio
            velocidad_rotacion = "Muy rápida (alta demanda)"
        elif dias_para_agotar < 90:
            factor_rotacion = 0  # Rotación normal
            velocidad_rotacion = "Normal"
        elif dias_para_agotar < 180:
            factor_rotacion = -0.05  # Rotación lenta
            velocidad_rotacion = "Lenta"
        else:
            factor_rotacion = -0.10  # Muy lenta
            velocidad_rotacion = "Muy lenta"
        
        # Factor 3: Tendencia de Ventas
        tendencia_ventas = "Estable"
        if ventas_periodo_anterior > 0:
            variacion = ((ventas_30_dias - ventas_periodo_anterior) / ventas_periodo_anterior) * 100
            if variacion > 20:
                tendencia_ventas = "Creciente (+{:.0f}%)".format(variacion)
            elif variacion < -20:
                tendencia_ventas = "Decreciente ({:.0f}%)".format(variacion)
                factor_rotacion -= 0.05  # Ajuste adicional si ventas están cayendo
            else:
                tendencia_ventas = "Estable"
        elif ventas_30_dias > 0:
            tendencia_ventas = "Nuevas ventas"
        
        # Factor 4: Nivel de Stock
        factor_stock = 0
        if cantidad_total < 3:
            factor_stock = 0.10  # Poco stock, puede subir precio
        elif cantidad_total < 10:
            factor_stock = 0  # Stock normal
        elif cantidad_total < 50:
            factor_stock = -0.05  # Mucho stock
        else:
            factor_stock = -0.10  # Stock excesivo
        
        # === CALCULAR PRECIO RECOMENDADO ===
        
        # Factor combinado
        factor_total = factor_antiguedad + factor_rotacion + factor_stock
        
        # Aplicar factor al precio actual
        precio_recomendado = precio_actual * (1 + factor_total)
        
        # Asegurar margen mínimo del 10%
        precio_minimo_margen = costo_promedio / 0.90  # 10% margen mínimo
        if precio_recomendado < precio_minimo_margen:
            precio_recomendado = precio_minimo_margen
        
        # Redondeo psicológico (terminar en 90 o 490)
        precio_recomendado = int(precio_recomendado)
        ultimo_digito = precio_recomendado % 1000
        
        if ultimo_digito < 490:
            precio_recomendado = (precio_recomendado // 1000) * 1000 + 490
        else:
            precio_recomendado = (precio_recomendado // 1000) * 1000 + 990
        
        # Calcular margen recomendado
        margen_recomendado = ((precio_recomendado - costo_promedio) / precio_recomendado * 100) if precio_recomendado > 0 else 0
        
        # === JUSTIFICACIÓN ===
        justificacion_partes = []
        
        if factor_antiguedad < 0:
            justificacion_partes.append(f"Inventario antiguo ({int(dias_promedio)} días)")
        
        if factor_rotacion > 0:
            justificacion_partes.append("Alta rotación")
        elif factor_rotacion < -0.05:
            justificacion_partes.append("Baja rotación")
        
        if factor_stock < 0:
            justificacion_partes.append("Stock elevado")
        elif factor_stock > 0:
            justificacion_partes.append("Stock limitado")
        
        if ventas_30_dias == 0:
            justificacion_partes.append("Sin ventas recientes")
        
        if not justificacion_partes:
            justificacion = "Precio óptimo según análisis de mercado"
        else:
            justificacion = "Ajuste por: " + ", ".join(justificacion_partes)
        
        # === RESPUESTA ===
        skus_list = [str(t.sku) for t in tallas]
        return JsonResponse({
            'success': True,
            'recomendaciones': {
                'producto_nombre': producto.articulo,
                'sku': ', '.join(skus_list[:3]) + (f' (+{len(skus_list)-3} más)' if len(skus_list) > 3 else ''),
                
                # Análisis actual
                'precio_actual': float(precio_actual),
                'costo_promedio': float(costo_promedio),
                'margen_actual': float(margen_actual),
                'stock_actual': int(cantidad_total),
                'dias_promedio_inventario': int(dias_promedio),
                
                # Análisis de ventas
                'ventas_30_dias': int(ventas_30_dias),
                'velocidad_rotacion': velocidad_rotacion,
                'tendencia_ventas': tendencia_ventas,
                
                # Factores de ajuste
                'factor_antiguedad': float(factor_antiguedad),
                'factor_rotacion': float(factor_rotacion),
                'factor_stock': float(factor_stock),
                'factor_total': float(factor_total),
                
                # Recomendación
                'precio_recomendado': float(precio_recomendado),
                'margen_recomendado': float(margen_recomendado),
                'justificacion': justificacion,
                'variacion_porcentual': float(((precio_recomendado - precio_actual) / precio_actual * 100) if precio_actual > 0 else 0)
            }
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al generar recomendaciones: {str(e)}'
        })


# ========== ACTUALIZACIÓN DE PRECIOS ==========

@require_POST
@login_required
@transaction.atomic
def actualizar_precio(request):
    """Actualizar precio de un producto (todas las tallas) y registrar en historial"""
    try:
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        nuevo_precio = data.get('nuevo_precio')
        motivo = data.get('motivo', 'Cambio manual de precio')
        tipo_cambio = data.get('tipo_cambio', 'MANUAL')
        sincronizar_sucursales = data.get('sincronizar_sucursales', True)  # Por defecto sincroniza
        
        if not producto_id or not nuevo_precio:
            return JsonResponse({
                'success': False,
                'error': 'Parámetros incompletos'
            })
        
        # Convertir a entero
        nuevo_precio = int(nuevo_precio)
        
        producto = Producto.objects.select_related('sucursal').get(id=producto_id)
        precio_anterior = producto.precioventa
        sucursal_origen = producto.sucursal
        
        # Solo registrar si el precio realmente cambió
        if precio_anterior == nuevo_precio:
            return JsonResponse({
                'success': True,
                'message': 'Sin cambios (precio igual)',
                'sin_cambios': True
            })
        
        # Actualizar precio base del producto
        producto.precioventa = nuevo_precio
        producto.save()
        
        # Actualizar precio en lotes activos de TODAS las tallas
        lotes_actualizados = LoteProducto.objects.filter(
            producto_talla__producto=producto,
            cantidad_disponible__gt=0,
            activo=True
        ).update(precio_venta_unitario=nuevo_precio)
        
        # Contar tallas actualizadas
        tallas_actualizadas = producto.producto_talla.count()
        
        # === REGISTRAR EN HISTORIAL ===
        diferencia = nuevo_precio - precio_anterior
        porcentaje = (diferencia / precio_anterior * 100) if precio_anterior > 0 else 0
        
        # Obtener IP del usuario
        ip_address = request.META.get('REMOTE_ADDR')
        
        HistorialCambioPrecio.objects.create(
            producto=producto,
            precio_anterior=precio_anterior,
            precio_nuevo=nuevo_precio,
            diferencia=diferencia,
            porcentaje_cambio=porcentaje,
            motivo=motivo,
            tipo_cambio=tipo_cambio,
            usuario=request.user,
            ip_address=ip_address,
            tallas_afectadas=tallas_actualizadas,
            lotes_afectados=lotes_actualizados
        )
        
        # === SINCRONIZAR PRECIOS Y CREAR ALERTAS EN OTRAS SUCURSALES ===
        # Igual que en creación de producto: sincroniza Y crea alerta
        sucursales_notificadas = 0
        productos_sincronizados = 0
        notificaciones_creadas = 0
        
        if sincronizar_sucursales:
            from django.db.models import Sum
            
            # Buscar productos similares en TODAS las demás sucursales (con o sin stock)
            productos_otras_sucursales = Producto.objects.filter(
                articulo=producto.articulo,
                atributo1=producto.atributo1,
                atributo2=producto.atributo2
            ).exclude(
                sucursal=sucursal_origen  # Excluir sucursal donde se edita
            ).annotate(
                stock_total=Sum('producto_talla__stock')
            ).select_related('sucursal')

            logger.debug(
                "Edicion rapida precios buscando similares en otras sucursales: producto_id=%s encontrados=%s",
                producto.id,
                productos_otras_sucursales.count(),
            )

            for prod_similar in productos_otras_sucursales:
                precio_anterior_sync = int(prod_similar.precioventa or 0)

                # Obtener primera talla (sin requerir stock > 0)
                primera_talla = Producto_Talla.objects.filter(
                    producto=prod_similar
                ).first()

                if not primera_talla:
                    logger.debug(
                        "Precio similar omitido sin tallas: producto_id=%s sucursal=%s",
                        prod_similar.id,
                        prod_similar.sucursal.alias,
                    )
                    continue

                stock_display = prod_similar.stock_total or 0
                logger.debug(
                    "Precio similar evaluado: producto_id=%s sucursal=%s stock=%s precio_actual=%s",
                    prod_similar.id,
                    prod_similar.sucursal.alias,
                    stock_display,
                    precio_anterior_sync,
                )
                
                # Solo procesar si hay diferencia de precio
                if precio_anterior_sync == nuevo_precio:
                    logger.debug(
                        "Precio similar omitido sin diferencia: producto_id=%s sucursal=%s precio=%s",
                        prod_similar.id,
                        prod_similar.sucursal.alias,
                        precio_anterior_sync,
                    )
                    continue
                
                # === 1. SINCRONIZAR PRECIO ===
                prod_similar.precioventa = nuevo_precio
                prod_similar.save()
                
                # Actualizar lotes activos
                LoteProducto.objects.filter(
                    producto_talla__producto=prod_similar,
                    cantidad_disponible__gt=0,
                    activo=True
                ).update(precio_venta_unitario=nuevo_precio)
                
                # Calcular diferencia para esta sucursal
                diferencia_sync = nuevo_precio - precio_anterior_sync
                porcentaje_sync = round((diferencia_sync / precio_anterior_sync * 100), 2) if precio_anterior_sync else 0
                
                # Registrar en historial de la otra sucursal
                HistorialCambioPrecio.objects.create(
                    producto=prod_similar,
                    precio_anterior=precio_anterior_sync,
                    precio_nuevo=nuevo_precio,
                    diferencia=diferencia_sync,
                    porcentaje_cambio=porcentaje_sync,
                    tipo_cambio='SINCRONIZACION',
                    motivo=f'Sincronización automática desde edición rápida en {sucursal_origen.alias}',
                    usuario=request.user,
                    ip_address=ip_address
                )
                
                # === 2. CREAR ALERTA INFORMATIVA (ya aplicado) ===
                # Determinar prioridad según la diferencia
                prioridad = 'MEDIA'
                if abs(porcentaje_sync) > 20:
                    prioridad = 'ALTA'
                if abs(porcentaje_sync) > 50:
                    prioridad = 'URGENTE'
                
                # === UMBRAL DE DIVERGENCIA: Si supera el umbral, crear PENDIENTE (requiere revisión) ===
                # en lugar de APLICADO (informativo). El parámetro UMBRAL_DIVERGENCIA_PRECIO_PCT
                # en ParametroGlobal define el % a partir del cual se requiere aprobación manual.
                # Valor 0 = umbral desactivado (siempre APLICADO automático).
                try:
                    umbral_param = ParametroGlobal.objects.filter(
                        nombre='UMBRAL_DIVERGENCIA_PRECIO_PCT'
                    ).first()
                    umbral_divergencia = umbral_param.valor_entero if umbral_param else 0
                except Exception:
                    umbral_divergencia = 0
                
                # Si el umbral está activo (> 0) y el cambio supera el umbral → PENDIENTE sin auto-aplicar
                supera_umbral = umbral_divergencia > 0 and abs(porcentaje_sync) >= umbral_divergencia
                
                if supera_umbral:
                    # No aplicar precio automáticamente — revertir el cambio hecho en prod_similar
                    prod_similar.precioventa = precio_anterior_sync
                    prod_similar.save()
                    LoteProducto.objects.filter(
                        producto_talla__producto=prod_similar,
                        cantidad_disponible__gt=0,
                        activo=True
                    ).update(precio_venta_unitario=precio_anterior_sync)
                    estado_cambio = 'PENDIENTE'
                    motivo_cambio = (
                        f'Cambio de precio desde {sucursal_origen.alias} supera el umbral de divergencia '
                        f'({umbral_divergencia}%). Requiere aprobación.'
                    )
                    productos_sincronizados -= 1  # No se sincronizó, solo queda pendiente
                else:
                    estado_cambio = 'APLICADO'
                    motivo_cambio = f'Precio sincronizado automáticamente desde {sucursal_origen.alias}'
                
                cambio = CambioPrecioPendiente.objects.create(
                    producto_talla=primera_talla,
                    sucursal=prod_similar.sucursal,
                    precio_anterior=precio_anterior_sync,
                    precio_nuevo=nuevo_precio,
                    diferencia=diferencia_sync,
                    porcentaje_cambio=porcentaje_sync,
                    tipo_cambio='SINCRONIZACION',
                    estado=estado_cambio,
                    motivo=motivo_cambio,
                    creado_por=request.user,
                    prioridad=prioridad,
                    fecha_vencimiento=timezone.now() + timedelta(days=7),
                    notificado=True
                )
                
                # Crear notificaciones para usuarios de esa sucursal
                usuarios_sucursal = EmpresaUser.objects.filter(
                    sucursal=prod_similar.sucursal,
                    status=True
                ).select_related('user')
                
                if supera_umbral:
                    mensaje_notif = (
                        f"⚠️ Cambio de precio pendiente de aprobación en {producto.articulo}: "
                        f"${precio_anterior_sync:,} → ${nuevo_precio:,} "
                        f"({abs(porcentaje_sync):.1f}% — supera umbral {umbral_divergencia}%). "
                        f"Enviado desde {sucursal_origen.alias}. Requiere revisión."
                    )
                else:
                    mensaje_notif = (
                        f"💰 Precio actualizado en {producto.articulo}: "
                        f"${precio_anterior_sync:,} → ${nuevo_precio:,} "
                        f"(desde {sucursal_origen.alias})"
                    )
                
                for empresa_user in usuarios_sucursal:
                    NotificacionCambioPrecio.objects.create(
                        cambio_precio=cambio,
                        usuario=empresa_user.user,
                        tipo='NUEVA',
                        mensaje=mensaje_notif
                    )
                    notificaciones_creadas += 1
                
                if not supera_umbral:
                    productos_sincronizados += 1
                sucursales_notificadas += 1
                logger.info(
                    "Precio %s en sucursal similar: producto_id=%s sucursal=%s precio_anterior=%s "
                    "precio_nuevo=%s notificaciones=%s",
                    'pendiente_revision' if supera_umbral else 'sincronizado',
                    prod_similar.id,
                    prod_similar.sucursal.alias,
                    precio_anterior_sync,
                    nuevo_precio,
                    usuarios_sucursal.count(),
                )
            
            if productos_sincronizados > 0:
                logger.info(
                    "Sincronizacion de precios completada: producto_origen_id=%s sincronizados=%s notificaciones=%s",
                    producto.id,
                    productos_sincronizados,
                    notificaciones_creadas,
                )
        
        return JsonResponse({
            'success': True,
            'message': f'Precio actualizado para {tallas_actualizadas} tallas',
            'lotes_actualizados': lotes_actualizados,
            'tallas_actualizadas': tallas_actualizadas,
            'historial_registrado': True,
            'sucursales_notificadas': sucursales_notificadas,
            'productos_sincronizados': productos_sincronizados
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar precio: {str(e)}'
        })


# ========== MODIFICACIÓN MASIVA ==========

@require_POST
@login_required
@transaction.atomic
def modificacion_masiva(request):
    """Modificar precios de múltiples productos de forma masiva (actualiza todas las tallas)"""
    try:
        data = json.loads(request.body)
        productos_ids = data.get('productos', [])
        tipo_modificacion = data.get('tipo_modificacion')
        valor = Decimal(str(data.get('valor', 0)))
        
        if not productos_ids or not tipo_modificacion:
            return JsonResponse({
                'success': False,
                'error': 'Parámetros incompletos'
            })
        
        productos_actualizados = 0
        tallas_actualizadas_total = 0
        
        for producto_id in productos_ids:
            try:
                producto = Producto.objects.get(id=producto_id)
                
                # Obtener precio actual del producto
                precio_actual = Decimal(str(producto.precioventa))
                
                # Obtener costo promedio de todos los lotes
                lotes = LoteProducto.objects.filter(
                    producto_talla__producto=producto,
                    cantidad_disponible__gt=0,
                    activo=True
                )
                
                if not lotes.exists():
                    continue
                
                costo_promedio = lotes.aggregate(
                    promedio=Avg('costo_unitario')
                )['promedio'] or 0
                costo_promedio = Decimal(str(costo_promedio))
                
                # Calcular nuevo precio según tipo
                if tipo_modificacion == 'fixed':
                    nuevo_precio = valor
                elif tipo_modificacion == 'percentage':
                    nuevo_precio = precio_actual * (Decimal('1') + (valor / Decimal('100')))
                elif tipo_modificacion == 'amount':
                    nuevo_precio = precio_actual + valor
                elif tipo_modificacion == 'margin':
                    # Precio = Costo / (1 - Margen/100)
                    nuevo_precio = costo_promedio / (Decimal('1') - (valor / Decimal('100')))
                else:
                    continue
                
                # Validar precio mínimo (costo + 10%)
                precio_minimo = costo_promedio * Decimal('1.1')
                
                if nuevo_precio < precio_minimo:
                    nuevo_precio = precio_minimo
                
                # Convertir a entero (los precios son IntegerField)
                nuevo_precio_int = int(nuevo_precio)
                
                # Actualizar producto principal
                producto.precioventa = nuevo_precio_int
                producto.save()
                
                # Actualizar TODOS los lotes de TODAS las tallas
                lotes_actualizados = LoteProducto.objects.filter(
                    producto_talla__producto=producto,
                    cantidad_disponible__gt=0,
                    activo=True
                ).update(precio_venta_unitario=nuevo_precio_int)
                
                # Contar tallas
                tallas_count = producto.producto_talla.count()
                tallas_actualizadas_total += tallas_count
                
                productos_actualizados += 1
                
            except Producto.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'{productos_actualizados} productos actualizados ({tallas_actualizadas_total} tallas)',
            'productos_actualizados': productos_actualizados,
            'tallas_actualizadas': tallas_actualizadas_total
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en modificación masiva: {str(e)}'
        })


# ========== SINCRONIZACIÓN MULTI-SUCURSAL ==========

@require_POST
@login_required
@transaction.atomic
def sincronizar_sucursales(request):
    """Sincronizar precios de productos similares en múltiples sucursales (todas las tallas)"""
    try:
        data = json.loads(request.body)
        productos_ids = data.get('productos', [])
        sucursales_destino = data.get('sucursales_destino', [])
        ajuste_porcentual = Decimal(str(data.get('ajuste_porcentual', 0)))
        
        if not productos_ids or not sucursales_destino:
            return JsonResponse({
                'success': False,
                'error': 'Parámetros incompletos'
            })
        
        productos_sincronizados = 0
        sucursales_afectadas = set()
        
        for producto_id in productos_ids:
            try:
                producto_origen = Producto.objects.get(id=producto_id)
                
                # Obtener precio del producto origen
                precio_origen = Decimal(str(producto_origen.precioventa))
                
                # Buscar productos similares en otras sucursales
                for sucursal_id in sucursales_destino:
                    # Buscar productos con mismo nombre y atributos
                    productos_similares = Producto.objects.filter(
                        articulo=producto_origen.articulo,
                        atributo1=producto_origen.atributo1,
                        atributo2=producto_origen.atributo2,
                        sucursal_id=sucursal_id
                    )
                    
                    for prod_similar in productos_similares:
                        # Calcular precio ajustado
                        nuevo_precio = precio_origen * (Decimal('1') + (ajuste_porcentual / Decimal('100')))
                        nuevo_precio_int = int(nuevo_precio)
                        
                        # Actualizar producto
                        prod_similar.precioventa = nuevo_precio_int
                        prod_similar.save()
                        
                        # Actualizar lotes de TODAS las tallas
                        LoteProducto.objects.filter(
                            producto_talla__producto=prod_similar,
                            cantidad_disponible__gt=0,
                            activo=True
                        ).update(precio_venta_unitario=nuevo_precio_int)
                        
                        productos_sincronizados += 1
                        sucursales_afectadas.add(sucursal_id)
                        
            except Producto.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'{productos_sincronizados} productos sincronizados',
            'productos_sincronizados': productos_sincronizados,
            'sucursales_afectadas': len(sucursales_afectadas)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en sincronización: {str(e)}'
        })


# ========== ANÁLISIS DE INVENTARIO ANTIGUO ==========

@require_GET
@login_required
def analisis_inventario_antiguo(request):
    """Análisis detallado de inventario antiguo con recomendaciones"""
    try:
        sucursal_id = request.session.get('idSucursalActual')
        
        # Fechas de corte
        fecha_6_meses = timezone.now() - timedelta(days=180)
        fecha_12_meses = timezone.now() - timedelta(days=365)
        
        queryset = LoteProducto.objects.filter(
            cantidad_disponible__gt=0,
            activo=True
        ).select_related('producto_talla__producto')
        
        if sucursal_id:
            queryset = queryset.filter(producto_talla__producto__sucursal_id=sucursal_id)
        
        # Categorizar por antigüedad
        inventario_antiguo = []
        
        for lote in queryset:
            dias = (timezone.now() - lote.fecha_ingreso).days
            
            if dias >= 180:  # Más de 6 meses
                valor_inventario = lote.cantidad_disponible * lote.costo_unitario
                
                # Sugerencia de descuento basado en antigüedad
                if dias >= 730:  # 2 años
                    descuento_sugerido = 40
                elif dias >= 365:  # 1 año
                    descuento_sugerido = 25
                else:  # 6 meses
                    descuento_sugerido = 15
                
                precio_sugerido = lote.precio_venta_unitario * (1 - descuento_sugerido / 100)
                
                inventario_antiguo.append({
                    'sku': lote.producto_talla.sku,
                    'producto': lote.producto_talla.producto.articulo,
                    'lote': lote.numero_lote,
                    'cantidad': lote.cantidad_disponible,
                    'dias_inventario': dias,
                    'precio_actual': float(lote.precio_venta_unitario),
                    'costo': float(lote.costo_unitario),
                    'valor_inventario': float(valor_inventario),
                    'descuento_sugerido': descuento_sugerido,
                    'precio_sugerido': float(precio_sugerido),
                    'categoria_edad': 'Crítico' if dias >= 365 else 'Atención'
                })
        
        # Ordenar por antigüedad descendente
        inventario_antiguo.sort(key=lambda x: x['dias_inventario'], reverse=True)
        
        # Calcular totales
        total_valor = sum(item['valor_inventario'] for item in inventario_antiguo)
        total_items = len(inventario_antiguo)
        
        return JsonResponse({
            'success': True,
            'inventario_antiguo': inventario_antiguo[:100],  # Top 100
            'resumen': {
                'total_items': total_items,
                'valor_total': float(total_valor),
                'promedio_dias': sum(item['dias_inventario'] for item in inventario_antiguo) / total_items if total_items > 0 else 0
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al analizar inventario antiguo: {str(e)}'
        })


# ========== ENDPOINTS AUXILIARES ==========

@require_GET
@login_required
def listar_categorias(request):
    """Listar todas las categorías activas"""
    try:
        categorias = Categoria.objects.all().order_by('nombre')
        
        categorias_data = []
        for cat in categorias:
            categorias_data.append({
                'id': cat.id,
                'nombre': cat.nombre,
                'padre_id': cat.padre_id if cat.padre else None
            })
        
        return JsonResponse({
            'success': True,
            'categorias': categorias_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al listar categorías: {str(e)}'
        })


@require_GET
@login_required
def listar_atributos(request):
    """Listar opciones de atributos (marcas, colores, etc.)"""
    try:
        tipo = request.GET.get('tipo', 'marca')
        
        # Mapear tipo a nombre de atributo
        tipo_map = {
            'marca': 'Marca',
            'color': 'Color',
            'genero': 'Género',
            'otro': 'Otro'
        }
        
        atributo_nombre = tipo_map.get(tipo, 'Marca')
        
        # Buscar el atributo
        from .models import Productos_Atributos
        try:
            atributo = Productos_Atributos.objects.get(nombre__iexact=atributo_nombre)
            opciones = AtributoOpcion.objects.filter(atributo=atributo).order_by('valor')
            
            opciones_data = []
            for opcion in opciones:
                opciones_data.append({
                    'id': opcion.id,
                    'valor': opcion.valor
                })
            
            return JsonResponse({
                'success': True,
                'opciones': opciones_data
            })
            
        except Productos_Atributos.DoesNotExist:
            return JsonResponse({
                'success': True,
                'opciones': []
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al listar atributos: {str(e)}'
        })


@require_GET
@login_required
def listar_sucursales(request):
    """Listar todas las sucursales"""
    try:
        sucursales = Sucursal.objects.all().order_by('alias')
        
        sucursales_data = []
        for suc in sucursales:
            sucursales_data.append({
                'id': suc.id,
                'alias': suc.alias,
                'direccion': suc.direccion
            })
        
        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al listar sucursales: {str(e)}'
        })


@require_GET
@login_required
def obtener_historial_precio(request, producto_id):
    """Obtener historial de cambios de precio de un producto, enriquecido con
    sucursales afectadas y referencia a la compra/recepción de origen."""
    try:
        producto = Producto.objects.select_related('sucursal').get(id=producto_id)
        
        historial = HistorialCambioPrecio.objects.filter(
            producto=producto
        ).select_related('usuario').order_by('-fecha_cambio')[:10]

        # Prefetch sucursales afectadas por SINCRONIZACION del mismo artículo
        from datetime import timedelta as _td
        articulo = producto.articulo
        atrib1 = producto.atributo1_id if hasattr(producto, 'atributo1_id') else None

        historial_data = []
        for cambio in historial:
            # --- Sucursales afectadas ---
            ventana_inicio = cambio.fecha_cambio - _td(seconds=10)
            ventana_fin = cambio.fecha_cambio + _td(seconds=10)
            sync_items = HistorialCambioPrecio.objects.filter(
                tipo_cambio='SINCRONIZACION',
                precio_nuevo=cambio.precio_nuevo,
                fecha_cambio__range=(ventana_inicio, ventana_fin),
                producto__articulo=articulo,
            ).exclude(producto=producto).select_related('producto__sucursal').values(
                'producto__sucursal__alias', 'producto__sucursal__id', 'precio_anterior'
            )
            sucursales_afectadas = [
                {
                    'alias': s['producto__sucursal__alias'],
                    'precio_anterior': s['precio_anterior'],
                }
                for s in sync_items
            ]

            # --- Origen compra (link a compras/DTE) ---
            origen_compra = None
            tipo_raw = cambio.tipo_cambio
            if tipo_raw in ('ACTUALIZACION_RECEPCION', 'ACTUALIZACION_MANUAL'):
                from .models import Productos_Recepcionados, Movimientos_Producto as _Mov
                # Buscar primer movimiento de ingreso para este producto alrededor de la fecha
                mov = _Mov.objects.filter(
                    ProductoTalla__producto=producto,
                    concepto__in=['INGRESO_INICIAL', 'INGRESO_MANUAL'],
                ).order_by('fecha', 'id').first()
                if mov and mov.dte_id:
                    origen_compra = {
                        'tipo': 'DTE',
                        'id': mov.dte_id,
                        'url': f'/app/verGestionCompras/',
                        'label': f'Ver Compra (DTE #{mov.dte_id})',
                    }
                elif mov:
                    # Buscar a través de Productos_Recepcionados
                    recep = Productos_Recepcionados.objects.filter(
                        movimiento_ingreso=mov
                    ).select_related(
                        'compra_producto_talla__compra_producto__compras'
                    ).first()
                    if recep and recep.compra_producto_talla and recep.compra_producto_talla.compra_producto.compras:
                        compra = recep.compra_producto_talla.compra_producto.compras
                        origen_compra = {
                            'tipo': 'COMPRA',
                            'id': compra.id,
                            'url': f'/app/verGestionCompras/',
                            'label': f'Compra: {compra.nombre}',
                        }

            # --- Margen en el momento del cambio ---
            costo_producto = producto.costo or 0
            margen = round((cambio.precio_nuevo - costo_producto) / cambio.precio_nuevo * 100, 1) if cambio.precio_nuevo > 0 else None

            historial_data.append({
                'id': cambio.id,
                'precio_anterior': cambio.precio_anterior,
                'precio_nuevo': cambio.precio_nuevo,
                'diferencia': cambio.diferencia,
                'porcentaje_cambio': float(cambio.porcentaje_cambio),
                'motivo': cambio.motivo or 'Sin motivo',
                'tipo_cambio': cambio.tipo_cambio,
                'tipo_cambio_label': cambio.get_tipo_cambio_display(),
                'usuario': cambio.usuario.username if cambio.usuario else 'Sistema',
                'fecha_cambio': cambio.fecha_cambio.strftime('%d/%m/%Y %H:%M'),
                'hace_cuanto': cambio.hace_cuanto,
                'tallas_afectadas': cambio.tallas_afectadas,
                'costo_ref': costo_producto,
                'margen_estimado': margen,
                'sucursales_afectadas': sucursales_afectadas,
                'origen_compra': origen_compra,
            })
        
        # Último cambio
        ultimo_cambio = historial.first()
        
        return JsonResponse({
            'success': True,
            'historial': historial_data,
            'producto_articulo': producto.articulo,
            'producto_sucursal': producto.sucursal.alias if producto.sucursal else None,
            'ultimo_cambio': {
                'usuario': ultimo_cambio.usuario.username if ultimo_cambio and ultimo_cambio.usuario else None,
                'fecha': ultimo_cambio.fecha_cambio.strftime('%d/%m/%Y %H:%M') if ultimo_cambio else None,
                'hace_cuanto': ultimo_cambio.hace_cuanto if ultimo_cambio else None,
                'precio': ultimo_cambio.precio_nuevo if ultimo_cambio else producto.precioventa
            } if ultimo_cambio else None
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener historial: {str(e)}'
        })


@require_GET
@login_required
def obtener_historial_ediciones_recientes(request):
    """Obtener historial general de ediciones de precios con filtros de auditoría

    Parámetros GET opcionales:
    - search: Término de búsqueda (artículo, descripción o usuario)
    - usuario: ID del usuario que realizó el cambio
    - tipo_cambio: código del tipo (MANUAL, MASIVO, SINCRONIZACION, ...)
    - fecha_desde / fecha_hasta: rango de fechas YYYY-MM-DD
    - alcance: 'actual' (sucursal de la sesión, default) o 'todas'
    - page / per_page: modo paginado (usado por el modal de auditoría)
    - limit: modo sin paginación (default: 20, máximo: 100)
    """
    try:
        # Obtener sucursal activa del usuario (clave correcta)
        sucursal_id = request.session.get('idSucursalActual')

        # Parámetros de búsqueda
        search_term = request.GET.get('search', '').strip()
        usuario_id = request.GET.get('usuario', '').strip()
        tipo_cambio = request.GET.get('tipo_cambio', '').strip()
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()
        alcance = request.GET.get('alcance', 'actual')
        paginado = 'page' in request.GET

        logger.debug(
            "Historial precios solicitado: sucursal_id=%s search=%s usuario=%s tipo=%s desde=%s hasta=%s alcance=%s",
            sucursal_id, search_term, usuario_id, tipo_cambio, fecha_desde, fecha_hasta, alcance,
        )

        # Obtener cambios de precio
        query = HistorialCambioPrecio.objects.select_related(
            'producto', 'producto__sucursal', 'producto__atributo1', 'usuario'
        ).order_by('-fecha_cambio')

        # Alcance: por defecto solo la sucursal activa; 'todas' muestra toda la red
        if alcance != 'todas' and sucursal_id:
            query = query.filter(producto__sucursal_id=sucursal_id)

        # Filtrar por término de búsqueda si existe
        if search_term:
            query = query.filter(
                Q(producto__articulo__icontains=search_term) |
                Q(producto__descripcion__icontains=search_term) |
                Q(usuario__username__icontains=search_term)
            )

        if usuario_id:
            query = query.filter(usuario_id=usuario_id)

        if tipo_cambio:
            query = query.filter(tipo_cambio=tipo_cambio)

        if fecha_desde:
            query = query.filter(
                fecha_cambio__date__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            )

        if fecha_hasta:
            query = query.filter(
                fecha_cambio__date__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            )

        pagination = None
        if paginado:
            page = int(request.GET.get('page', 1))
            per_page = min(int(request.GET.get('per_page', 25)), 100)
            paginator = Paginator(query, per_page)
            page_obj = paginator.get_page(page)
            historial = list(page_obj)
            pagination = {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            }
        else:
            limit = min(int(request.GET.get('limit', 20)), 100)  # Máximo 100 resultados
            if search_term:
                # Si hay búsqueda, aumentar el límite para mostrar más resultados relevantes
                limit = min(50, limit * 2)
            historial = list(query[:limit])

        logger.debug("Historial precios registros encontrados: total=%s", len(historial))

        historial_data = []
        for cambio in historial:
            fecha_local = timezone.localtime(cambio.fecha_cambio)
            fecha_alta = cambio.producto.fecha_creacion
            historial_data.append({
                'id': cambio.id,
                'producto_id': cambio.producto.id,
                'producto_nombre': cambio.producto.articulo,
                'producto_talla': cambio.producto.atributo1.valor if cambio.producto.atributo1 else '',
                'marca': cambio.producto.atributo1.valor if cambio.producto.atributo1 else '',
                'producto_fecha_creacion': fecha_alta.strftime('%d/%m/%Y') if fecha_alta else None,
                'sucursal': cambio.producto.sucursal.alias if cambio.producto.sucursal else 'N/A',
                'precio_anterior': cambio.precio_anterior,
                'precio_nuevo': cambio.precio_nuevo,
                'diferencia': cambio.diferencia,
                'porcentaje_cambio': float(cambio.porcentaje_cambio),
                'tipo_cambio': cambio.get_tipo_cambio_display(),
                'tipo_cambio_codigo': cambio.tipo_cambio,
                'motivo': cambio.motivo or '',
                'usuario': cambio.usuario.username if cambio.usuario else 'Sistema',
                'usuario_id': cambio.usuario_id,
                'fecha_cambio': fecha_local.strftime('%d/%m/%Y %H:%M'),
                'hace_cuanto': cambio.hace_cuanto,
            })

        return JsonResponse({
            'success': True,
            'historial': historial_data,
            'total': len(historial_data),
            'pagination': pagination,
            'search_term': search_term if search_term else None
        })

    except Exception as e:
        logger.exception("Error al obtener historial de precios")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener historial: {str(e)}'
        })


def _serializar_cambio_alerta(cambio):
    """Serializa un HistorialCambioPrecio para las listas de alertas del panel de KPIs"""
    return {
        'producto_id': cambio.producto_id,
        'producto': cambio.producto.articulo,
        'sucursal': cambio.producto.sucursal.alias if cambio.producto.sucursal else 'N/A',
        'usuario': cambio.usuario.username if cambio.usuario else 'Sistema',
        'fecha': timezone.localtime(cambio.fecha_cambio).strftime('%d/%m/%Y %H:%M'),
        'precio_anterior': cambio.precio_anterior,
        'precio_nuevo': cambio.precio_nuevo,
        'porcentaje_cambio': float(cambio.porcentaje_cambio),
        'costo': cambio.producto.costo or 0,
        'tipo_cambio': cambio.tipo_cambio,
    }


@require_GET
@login_required
def obtener_kpis_cambios_precios(request):
    """KPIs de modificaciones de precios por usuario + alertas de auditoría

    Parámetros GET opcionales:
    - fecha_desde / fecha_hasta: rango YYYY-MM-DD (default: últimos 30 días)
    - alcance: 'todas' (default) o 'actual' (sucursal de la sesión)
    - incluir_sincronizacion: 'true' para contar también los ecos automáticos
      de sincronización multi-sucursal (default: false, solo acciones directas)
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        alcance = request.GET.get('alcance', 'todas')
        incluir_sync = request.GET.get('incluir_sincronizacion', 'false') == 'true'
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()

        hoy = timezone.localdate()
        desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else hoy - timedelta(days=30)
        hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else hoy

        qs = HistorialCambioPrecio.objects.filter(
            fecha_cambio__date__gte=desde,
            fecha_cambio__date__lte=hasta,
        )

        if alcance == 'actual' and sucursal_id:
            qs = qs.filter(producto__sucursal_id=sucursal_id)

        if not incluir_sync:
            # Las sincronizaciones son ecos automáticos de un cambio manual:
            # contarlas inflaría los KPIs por usuario
            qs = qs.exclude(tipo_cambio='SINCRONIZACION')

        # Precio nuevo por debajo del costo ACTUAL del producto (aproximación:
        # el historial no guarda el costo al momento del cambio)
        bajo_costo_q = Q(precio_nuevo__lt=F('producto__costo'), producto__costo__gt=0)
        # El lookup __hour respeta TIME_ZONE (America/Santiago) con USE_TZ=True
        fuera_horario_q = Q(fecha_cambio__hour__lt=8) | Q(fecha_cambio__hour__gte=21)
        UMBRAL_REBAJA_FUERTE = 30  # % de rebaja que se considera "llama la atención"

        # === KPIs GLOBALES ===
        globales = qs.aggregate(
            total=Count('id'),
            alzas=Count('id', filter=Q(diferencia__gt=0)),
            rebajas=Count('id', filter=Q(diferencia__lt=0)),
            pct_promedio=Avg('porcentaje_cambio'),
            rebaja_max=Min('porcentaje_cambio'),
            productos=Count('producto', distinct=True),
            usuarios=Count('usuario', distinct=True),
            bajo_costo=Count('id', filter=bajo_costo_q),
            fuera_horario=Count('id', filter=fuera_horario_q),
            rebajas_fuertes=Count('id', filter=Q(porcentaje_cambio__lte=-UMBRAL_REBAJA_FUERTE)),
        )

        # === KPIs POR USUARIO ===
        por_usuario_raw = qs.values('usuario__id', 'usuario__username').annotate(
            total=Count('id'),
            alzas=Count('id', filter=Q(diferencia__gt=0)),
            rebajas=Count('id', filter=Q(diferencia__lt=0)),
            pct_promedio=Avg('porcentaje_cambio'),
            rebaja_max=Min('porcentaje_cambio'),
            alza_max=Max('porcentaje_cambio'),
            productos=Count('producto', distinct=True),
            bajo_costo=Count('id', filter=bajo_costo_q),
            fuera_horario=Count('id', filter=fuera_horario_q),
            rebajas_fuertes=Count('id', filter=Q(porcentaje_cambio__lte=-UMBRAL_REBAJA_FUERTE)),
            ultima_fecha=Max('fecha_cambio'),
        ).order_by('-total')

        por_usuario = []
        for u in por_usuario_raw:
            por_usuario.append({
                'usuario_id': u['usuario__id'],
                'usuario': u['usuario__username'] or 'Sistema',
                'total': u['total'],
                'alzas': u['alzas'],
                'rebajas': u['rebajas'],
                'pct_promedio': round(float(u['pct_promedio'] or 0), 1),
                'rebaja_max': round(float(u['rebaja_max'] or 0), 1),
                'alza_max': round(float(u['alza_max'] or 0), 1),
                'productos': u['productos'],
                'bajo_costo': u['bajo_costo'],
                'fuera_horario': u['fuera_horario'],
                'rebajas_fuertes': u['rebajas_fuertes'],
                'ultima_fecha': timezone.localtime(u['ultima_fecha']).strftime('%d/%m/%Y %H:%M') if u['ultima_fecha'] else None,
            })

        # === ALERTAS: cosas que llaman la atención ===
        base_alertas = qs.select_related('producto', 'producto__sucursal', 'usuario')

        alertas_rebajas_fuertes = [
            _serializar_cambio_alerta(c)
            for c in base_alertas.filter(porcentaje_cambio__lte=-UMBRAL_REBAJA_FUERTE).order_by('porcentaje_cambio')[:15]
        ]

        alertas_bajo_costo = [
            _serializar_cambio_alerta(c)
            for c in base_alertas.filter(bajo_costo_q).order_by('-fecha_cambio')[:15]
        ]

        alertas_fuera_horario = [
            _serializar_cambio_alerta(c)
            for c in base_alertas.filter(fuera_horario_q).order_by('-fecha_cambio')[:15]
        ]

        # Ping-pong: mismo producto cambiado 3+ veces dentro del período
        alertas_ping_pong = list(
            qs.values('producto_id', 'producto__articulo', 'producto__sucursal__alias').annotate(
                cambios=Count('id'),
                usuarios_distintos=Count('usuario', distinct=True),
                pct_min=Min('porcentaje_cambio'),
                pct_max=Max('porcentaje_cambio'),
            ).filter(cambios__gte=3).order_by('-cambios')[:10]
        )
        for p in alertas_ping_pong:
            p['producto'] = p.pop('producto__articulo')
            p['sucursal'] = p.pop('producto__sucursal__alias') or 'N/A'
            p['pct_min'] = round(float(p['pct_min'] or 0), 1)
            p['pct_max'] = round(float(p['pct_max'] or 0), 1)

        return JsonResponse({
            'success': True,
            'periodo': {
                'desde': desde.strftime('%Y-%m-%d'),
                'hasta': hasta.strftime('%Y-%m-%d'),
                'alcance': alcance,
                'incluye_sincronizacion': incluir_sync,
            },
            'globales': {
                'total': globales['total'],
                'alzas': globales['alzas'],
                'rebajas': globales['rebajas'],
                'pct_promedio': round(float(globales['pct_promedio'] or 0), 1),
                'rebaja_max': round(float(globales['rebaja_max'] or 0), 1),
                'productos': globales['productos'],
                'usuarios': globales['usuarios'],
                'bajo_costo': globales['bajo_costo'],
                'fuera_horario': globales['fuera_horario'],
                'rebajas_fuertes': globales['rebajas_fuertes'],
            },
            'por_usuario': por_usuario,
            'alertas': {
                'umbral_rebaja_fuerte': UMBRAL_REBAJA_FUERTE,
                'rebajas_fuertes': alertas_rebajas_fuertes,
                'bajo_costo': alertas_bajo_costo,
                'fuera_horario': alertas_fuera_horario,
                'ping_pong': alertas_ping_pong,
            },
        })

    except Exception as e:
        logger.exception("Error al obtener KPIs de cambios de precios")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener KPIs: {str(e)}'
        })


@require_GET
@login_required
def buscar_productos_similares_sucursales(request, producto_id):
    """Buscar productos similares en otras sucursales"""
    try:
        producto = Producto.objects.select_related('atributo1', 'atributo2', 'sucursal').get(id=producto_id)
        
        # Buscar productos con mismo nombre y atributos en OTRAS sucursales
        productos_similares = Producto.objects.filter(
            articulo=producto.articulo,
            atributo1=producto.atributo1,
            atributo2=producto.atributo2
        ).exclude(
            sucursal=producto.sucursal
        ).select_related('sucursal').distinct()
        
        sucursales_data = []
        for prod in productos_similares:
            # Calcular stock total
            stock_total = sum(pt.stock for pt in prod.producto_talla.all())
            
            # Obtener último cambio
            ultimo_cambio = HistorialCambioPrecio.objects.filter(
                producto=prod
            ).select_related('usuario').first()
            
            sucursales_data.append({
                'sucursal_id': prod.sucursal.id,
                'sucursal': prod.sucursal.alias,
                'precio_actual': prod.precioventa,
                'stock_total': stock_total,
                'tallas_count': prod.producto_talla.count(),
                'ultimo_cambio': {
                    'usuario': ultimo_cambio.usuario.username if ultimo_cambio and ultimo_cambio.usuario else None,
                    'fecha': ultimo_cambio.fecha_cambio.strftime('%d/%m/%Y') if ultimo_cambio else None,
                    'hace_cuanto': ultimo_cambio.hace_cuanto if ultimo_cambio else None
                } if ultimo_cambio else None
            })
        
        return JsonResponse({
            'success': True,
            'sucursal_actual': {
                'id': producto.sucursal.id,
                'nombre': producto.sucursal.alias,
                'precio': producto.precioventa
            },
            'otras_sucursales': sucursales_data,
            'total_sucursales': len(sucursales_data)
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al buscar sucursales: {str(e)}'
        })


# ========== SISTEMA DE APROBACIÓN DE CAMBIOS DE PRECIOS ==========

@require_POST
@login_required
@transaction.atomic
def proponer_cambio_precio(request):
    """
    Proponer un cambio de precio (crea registro pendiente en lugar de aplicar directamente)
    Trabaja a nivel de producto (todas las tallas)
    """
    try:
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        nuevo_precio = int(data.get('nuevo_precio'))
        motivo = data.get('motivo', '')
        tipo_cambio = data.get('tipo_cambio', 'INDIVIDUAL')
        prioridad = data.get('prioridad', 'MEDIA')
        dias_vencimiento = int(data.get('dias_vencimiento', 7))
        
        if not producto_id or not nuevo_precio:
            return JsonResponse({
                'success': False,
                'error': 'Parámetros incompletos'
            })
        
        producto = Producto.objects.select_related('sucursal').get(id=producto_id)
        
        # Obtener primera talla del producto (para el registro)
        primera_talla = producto.producto_talla.first()
        if not primera_talla:
            return JsonResponse({
                'success': False,
                'error': 'Producto sin tallas definidas'
            })
        
        # Obtener precio actual
        precio_actual = producto.precioventa
        
        # Calcular diferencia y porcentaje
        diferencia = nuevo_precio - precio_actual
        porcentaje_cambio = (diferencia / precio_actual * 100) if precio_actual > 0 else 0
        
        # Crear cambio pendiente (usamos primera talla como referencia, pero afecta a todas)
        cambio = CambioPrecioPendiente.objects.create(
            producto_talla=primera_talla,
            sucursal=producto.sucursal,
            precio_anterior=precio_actual,
            precio_nuevo=nuevo_precio,
            diferencia=diferencia,
            porcentaje_cambio=porcentaje_cambio,
            tipo_cambio=tipo_cambio,
            estado='PENDIENTE',
            motivo=motivo,
            creado_por=request.user,
            prioridad=prioridad,
            fecha_vencimiento=timezone.now() + timedelta(days=dias_vencimiento)
        )
        
        # Crear notificación para usuarios de la sucursal
        from .models import EmpresaUser
        usuarios_sucursal = EmpresaUser.objects.filter(
            sucursal=producto.sucursal,
            status=True
        ).select_related('user')
        
        tallas_count = producto.producto_talla.count()
        mensaje = f"Nuevo cambio de precio propuesto para {producto.articulo} ({tallas_count} tallas). " \
                  f"Precio actual: ${precio_actual:,} → Nuevo: ${nuevo_precio:,} ({porcentaje_cambio:+.1f}%)"
        
        notificaciones_creadas = 0
        for empresa_user in usuarios_sucursal:
            if empresa_user.user != request.user:  # No notificar al creador
                # Evitar duplicados
                existe = NotificacionCambioPrecio.objects.filter(
                    cambio_precio=cambio,
                    usuario=empresa_user.user,
                    tipo='NUEVA'
                ).exists()
                
                if not existe:
                    NotificacionCambioPrecio.objects.create(
                        cambio_precio=cambio,
                        usuario=empresa_user.user,
                        tipo='NUEVA',
                        mensaje=mensaje
                    )
                    notificaciones_creadas += 1
        
        cambio.notificado = True
        cambio.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Cambio de precio propuesto correctamente',
            'cambio_id': cambio.id,
            'notificaciones_enviadas': notificaciones_creadas
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al proponer cambio: {str(e)}'
        })


@require_GET
@login_required
def obtener_indicadores_precios_pendientes(request):
    """
    Obtener indicadores de precios pendientes para el dashboard de ventas
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')
        
        # Filtrar por sucursal si está definida
        queryset = CambioPrecioPendiente.objects.all()
        
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
        
        # Contar por estado
        total_pendientes = queryset.filter(estado='PENDIENTE').count()
        total_aplicados = queryset.filter(estado='APLICADO').count()
        total_revisados = queryset.filter(estado='REVISADO').count()
        total_aprobados = queryset.filter(estado='APROBADO').count()
        total_rechazados = queryset.filter(estado='RECHAZADO').count()
        
        # Cambios urgentes (prioridad alta o vencidos)
        cambios_urgentes = queryset.filter(
            Q(prioridad__in=['ALTA', 'URGENTE']) | 
            Q(fecha_vencimiento__lt=timezone.now(), estado='PENDIENTE')
        ).count()
        
        # Cambios sin revisar (más de 3 días)
        fecha_limite = timezone.now() - timedelta(days=3)
        sin_revisar_antiguos = queryset.filter(
            estado='PENDIENTE',
            fecha_creacion__lt=fecha_limite
        ).count()
        
        # Últimos cambios pendientes (top 5)
        ultimos_cambios = queryset.filter(
            estado='PENDIENTE'
        ).select_related(
            'producto_talla__producto',
            'creado_por',
            'sucursal'
        ).order_by('-fecha_creacion')[:5]
        
        cambios_data = []
        for cambio in ultimos_cambios:
            cambios_data.append({
                'id': cambio.id,
                'sku': cambio.producto_talla.sku,
                'producto': cambio.producto_talla.producto.articulo,
                'precio_anterior': float(cambio.precio_anterior),
                'precio_nuevo': float(cambio.precio_nuevo),
                'porcentaje_cambio': float(cambio.porcentaje_cambio),
                'dias_pendiente': cambio.dias_pendiente,
                'prioridad': cambio.prioridad,
                'creado_por': cambio.creado_por.username if cambio.creado_por else 'Sistema',
                'fecha_creacion': cambio.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'requiere_atencion': cambio.requiere_atencion
            })
        
        return JsonResponse({
            'success': True,
            'indicadores': {
                'total_pendientes': total_pendientes,
                'total_aplicados': total_aplicados,
                'total_revisados': total_revisados,
                'total_aprobados': total_aprobados,
                'total_rechazados': total_rechazados,
                'cambios_urgentes': cambios_urgentes,
                'sin_revisar_antiguos': sin_revisar_antiguos,
                'requiere_atencion': cambios_urgentes + sin_revisar_antiguos
            },
            'ultimos_cambios': cambios_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener indicadores: {str(e)}'
        })


def _filtrar_cambios_precios(request):
    """
    Construye el queryset de CambioPrecioPendiente según los filtros GET.
    Compartido por listar_cambios_pendientes y exportar_cambios_precios_excel.
    Retorna (queryset ordenado, sucursal_id).
    """
    sucursal_id = request.GET.get('sucursal_id') or request.session.get('idSucursalActual')
    estado = request.GET.get('estado')
    prioridad = request.GET.get('prioridad')
    tipo_cambio = request.GET.get('tipo_cambio')
    marca_id = request.GET.get('marca')
    busqueda = request.GET.get('busqueda', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    mostrar_descartados = request.GET.get('mostrar_descartados', 'false') == 'true'
    solo_descartados = request.GET.get('solo_descartados', 'false') == 'true'

    queryset = CambioPrecioPendiente.objects.select_related(
        'producto_talla__producto__atributo1',
        'sucursal',
        'creado_por',
        'revisado_por',
        'aprobado_por'
    ).prefetch_related('producto_talla__producto__producto_talla')

    # Filtrar por descartados según parámetros
    if solo_descartados:
        # Mostrar SOLO los descartados
        queryset = queryset.filter(descartado=True)
    elif not mostrar_descartados:
        # Por defecto NO mostrar descartados
        queryset = queryset.filter(descartado=False)

    # Filtros
    if sucursal_id:
        queryset = queryset.filter(sucursal_id=sucursal_id)

    if estado:
        queryset = queryset.filter(estado=estado)

    if prioridad:
        queryset = queryset.filter(prioridad=prioridad)

    if tipo_cambio:
        queryset = queryset.filter(tipo_cambio=tipo_cambio)

    if marca_id:
        queryset = queryset.filter(producto_talla__producto__atributo1_id=marca_id)

    if busqueda:
        queryset = queryset.filter(
            Q(producto_talla__producto__articulo__icontains=busqueda) |
            Q(producto_talla__producto__descripcion__icontains=busqueda) |
            Q(producto_talla__producto__atributo1__valor__icontains=busqueda) |
            Q(producto_talla__sku__icontains=busqueda) |
            Q(motivo__icontains=busqueda)
        )

    # Filtro de fechas
    if fecha_desde:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d')
        queryset = queryset.filter(fecha_creacion__date__gte=fecha_inicio.date())

    if fecha_hasta:
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        queryset = queryset.filter(fecha_creacion__date__lte=fecha_fin.date())

    return queryset.order_by('-fecha_creacion'), sucursal_id


def _stock_info_producto(producto):
    """
    Stock del producto en su sucursal (cada Producto pertenece a una sucursal,
    por lo que la suma de sus tallas ES el stock de esa sucursal).
    Retorna (stock_total, detalle "talla: stock, ...").
    """
    tallas = list(producto.producto_talla.all())
    stock_total = sum(max(0, t.stock or 0) for t in tallas)
    detalle = ', '.join(f"{t.talla}: {max(0, t.stock or 0)}" for t in tallas[:10])
    if len(tallas) > 10:
        detalle += f" (+{len(tallas) - 10} más)"
    return stock_total, detalle


@require_GET
@login_required
def listar_cambios_pendientes(request):
    """
    Listar todos los cambios de precio pendientes con filtros
    """
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        queryset, sucursal_id = _filtrar_cambios_precios(request)

        # Obtener resumen de contadores
        base_queryset = CambioPrecioPendiente.objects.all()
        if sucursal_id:
            base_queryset = base_queryset.filter(sucursal_id=sucursal_id)
        
        # Contadores de activos (no descartados)
        activos = base_queryset.filter(descartado=False)
        resumen = {
            'pendientes': activos.filter(estado='PENDIENTE').count(),
            'revisados': activos.filter(estado='REVISADO').count(),
            'aprobados': activos.filter(estado='APROBADO').count(),
            'rechazados': activos.filter(estado='RECHAZADO').count(),
            'aplicados': activos.filter(estado='APLICADO').count(),
            'descartados': base_queryset.filter(descartado=True).count(),  # Total descartados
        }

        # Paginación
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        cambios_data = []
        for cambio in page_obj:
            producto = cambio.producto_talla.producto
            tallas = list(producto.producto_talla.all())
            tallas_count = len(tallas)
            tallas_list = [t.talla for t in tallas[:5]]
            stock_producto, stock_detalle = _stock_info_producto(producto)

            cambios_data.append({
                'id': cambio.id,
                'sku': cambio.producto_talla.sku,
                'producto': producto.articulo,
                'marca': producto.atributo1.valor if producto.atributo1 else '',
                'talla': cambio.producto_talla.talla,
                'stock_talla': max(0, cambio.producto_talla.stock or 0),
                'stock_producto': stock_producto,
                'stock_detalle': stock_detalle,
                'tallas_count': tallas_count,
                'tallas_preview': ', '.join(str(t) for t in tallas_list) + (f' (+{tallas_count-5} más)' if tallas_count > 5 else ''),
                'sucursal': cambio.sucursal.alias,
                'precio_anterior': float(cambio.precio_anterior),
                'precio_nuevo': float(cambio.precio_nuevo),
                'diferencia': float(cambio.diferencia),
                'porcentaje_cambio': float(cambio.porcentaje_cambio),
                'tipo_cambio': cambio.get_tipo_cambio_display(),
                'estado': cambio.estado,
                'estado_display': cambio.get_estado_display(),
                'prioridad': cambio.prioridad,
                'motivo': cambio.motivo or '',
                'creado_por': cambio.creado_por.username if cambio.creado_por else 'Sistema',
                'revisado_por': cambio.revisado_por.username if cambio.revisado_por else None,
                'aprobado_por': cambio.aprobado_por.username if cambio.aprobado_por else None,
                'fecha_creacion': cambio.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'fecha_revision': cambio.fecha_revision.strftime('%d/%m/%Y %H:%M') if cambio.fecha_revision else None,
                'fecha_aprobacion': cambio.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if cambio.fecha_aprobacion else None,
                'dias_pendiente': cambio.dias_pendiente,
                'esta_vencido': cambio.esta_vencido,
                'requiere_atencion': cambio.requiere_atencion,
                # Campos de descarte (para historial)
                'descartado': cambio.descartado,
                'fecha_descarte': cambio.fecha_descarte.strftime('%d/%m/%Y %H:%M') if cambio.fecha_descarte else None,
                'descartado_por': cambio.descartado_por.username if cambio.descartado_por else None
            })
        
        return JsonResponse({
            'success': True,
            'cambios': cambios_data,
            'resumen': resumen,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })
        
    except Exception as e:
        logger.exception("Error en listar_cambios_pendientes")
        return JsonResponse({
            'success': False,
            'error': f'Error al listar cambios: {str(e)}'
        })


@require_GET
@login_required
def exportar_cambios_precios_excel(request):
    """
    Exportar a Excel los cambios de precio según los filtros actuales
    (mismos filtros que listar_cambios_pendientes, sin paginación).
    Incluye columnas vacías 'Ejecutado' y 'Observaciones' para trabajo manual;
    la columna ID permite ubicar el registro y descartarlo luego en la vista.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        queryset, sucursal_id = _filtrar_cambios_precios(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cambios de precio"

        headers = [
            'ID', 'SKU', 'Artículo', 'Descripción', 'Marca', 'Talla',
            'Stock talla', 'Stock producto (sucursal)', 'Sucursal',
            'Precio anterior', 'Precio nuevo', 'Diferencia', 'Variación %',
            'Tipo', 'Estado', 'Prioridad', 'Motivo',
            'Creado por', 'Fecha creación', 'Días pendiente',
            'Ejecutado (SÍ/NO)', 'Observaciones',
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color='405189', end_color='405189', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_num in range(1, len(headers) + 1):
            celda = ws.cell(row=1, column=col_num)
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center', vertical='center')

        for cambio in queryset:
            producto = cambio.producto_talla.producto
            stock_producto, _ = _stock_info_producto(producto)
            ws.append([
                cambio.id,
                cambio.producto_talla.sku,
                producto.articulo,
                producto.descripcion,
                producto.atributo1.valor if producto.atributo1 else '',
                cambio.producto_talla.talla,
                max(0, cambio.producto_talla.stock or 0),
                stock_producto,
                cambio.sucursal.alias,
                float(cambio.precio_anterior),
                float(cambio.precio_nuevo),
                float(cambio.diferencia),
                float(cambio.porcentaje_cambio),
                cambio.get_tipo_cambio_display(),
                cambio.get_estado_display() + (' (descartado)' if cambio.descartado else ''),
                cambio.prioridad,
                cambio.motivo or '',
                cambio.creado_por.username if cambio.creado_por else 'Sistema',
                cambio.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                cambio.dias_pendiente,
                '',
                '',
            ])

        anchos = [8, 14, 26, 32, 16, 8, 10, 12, 12, 14, 14, 12, 11, 18, 20, 11, 30, 14, 16, 12, 14, 30]
        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

        fecha_archivo = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cambios_precios_{fecha_archivo}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        logger.exception("Error en exportar_cambios_precios_excel")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def eliminar_cambios_aplicados(request):
    """
    Descartar (archivar) registros de cambios - NO los elimina, solo los marca como descartados
    para mantener el historial completo.
    """
    try:
        data = json.loads(request.body)
        cambio_ids = data.get('cambio_ids', [])
        
        if not cambio_ids:
            return JsonResponse({
                'success': False,
                'error': 'No se especificaron cambios a descartar'
            })
        
        # Marcar como descartados (NO eliminar)
        descartados = CambioPrecioPendiente.objects.filter(
            id__in=cambio_ids
        ).update(
            descartado=True,
            fecha_descarte=timezone.now(),
            descartado_por=request.user
        )
        
        # Marcar las notificaciones como leídas
        NotificacionCambioPrecio.objects.filter(
            cambio_precio_id__in=cambio_ids,
            leida=False
        ).update(leida=True, fecha_lectura=timezone.now())
        
        return JsonResponse({
            'success': True,
            'eliminados': descartados,  # Mantener nombre para compatibilidad JS
            'message': f'{descartados} registro(s) descartados correctamente'
        })
        
    except Exception as e:
        logger.exception("Error en eliminar_cambios_aplicados")
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def revisar_cambio_precio(request):
    """
    Marcar un cambio de precio como revisado
    """
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        observaciones = data.get('observaciones', '')
        
        cambio = CambioPrecioPendiente.objects.get(id=cambio_id)
        
        if cambio.estado != 'PENDIENTE':
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden revisar cambios pendientes'
            })
        
        cambio.estado = 'REVISADO'
        cambio.revisado_por = request.user
        cambio.fecha_revision = timezone.now()
        cambio.observaciones_revision = observaciones
        cambio.save()
        
        # Notificar al creador (evitando duplicados)
        if cambio.creado_por and cambio.creado_por != request.user:
            existe = NotificacionCambioPrecio.objects.filter(
                cambio_precio=cambio,
                usuario=cambio.creado_por,
                tipo='REVISION'
            ).exists()
            
            if not existe:
                NotificacionCambioPrecio.objects.create(
                    cambio_precio=cambio,
                    usuario=cambio.creado_por,
                    tipo='REVISION',
                    mensaje=f"Tu cambio de precio para {cambio.producto_talla.producto.articulo} ha sido revisado por {request.user.username}"
                )
        
        return JsonResponse({
            'success': True,
            'message': 'Cambio marcado como revisado'
        })
        
    except CambioPrecioPendiente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al revisar cambio: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def aprobar_cambio_precio(request):
    """
    Aprobar un cambio de precio y aplicarlo a todas las tallas del producto
    """
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        observaciones = data.get('observaciones', '')
        
        cambio = CambioPrecioPendiente.objects.select_related('producto_talla__producto').get(id=cambio_id)
        
        if cambio.estado not in ['PENDIENTE', 'REVISADO']:
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden aprobar cambios pendientes o revisados'
            })
        
        # Aprobar
        cambio.estado = 'APROBADO'
        cambio.aprobado_por = request.user
        cambio.fecha_aprobacion = timezone.now()
        cambio.observaciones_aprobacion = observaciones
        
        # Aplicar el cambio al producto principal
        producto = cambio.producto_talla.producto
        producto.precioventa = cambio.precio_nuevo
        producto.save()
        
        # Actualizar lotes activos de TODAS las tallas del producto
        lotes_actualizados = LoteProducto.objects.filter(
            producto_talla__producto=producto,
            cantidad_disponible__gt=0,
            activo=True
        ).update(precio_venta_unitario=cambio.precio_nuevo)
        
        # Contar tallas afectadas
        tallas_afectadas = producto.producto_talla.count()
        
        cambio.estado = 'APLICADO'
        cambio.fecha_aplicacion = timezone.now()
        cambio.save()
        
        # Notificar al creador (evitando duplicados)
        if cambio.creado_por and cambio.creado_por != request.user:
            existe = NotificacionCambioPrecio.objects.filter(
                cambio_precio=cambio,
                usuario=cambio.creado_por,
                tipo='APROBACION'
            ).exists()
            
            if not existe:
                NotificacionCambioPrecio.objects.create(
                    cambio_precio=cambio,
                    usuario=cambio.creado_por,
                    tipo='APROBACION',
                    mensaje=f"Tu cambio de precio para {producto.articulo} ha sido aprobado y aplicado a {tallas_afectadas} tallas"
                )
        
        return JsonResponse({
            'success': True,
            'message': f'Cambio aprobado y aplicado a {tallas_afectadas} tallas',
            'tallas_afectadas': tallas_afectadas,
            'lotes_actualizados': lotes_actualizados
        })
        
    except CambioPrecioPendiente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al aprobar cambio: {str(e)}'
        })


@require_POST
@login_required
@transaction.atomic
def rechazar_cambio_precio(request):
    """
    Rechazar un cambio de precio
    """
    try:
        data = json.loads(request.body)
        cambio_id = data.get('cambio_id')
        observaciones = data.get('observaciones', 'Cambio rechazado')
        
        cambio = CambioPrecioPendiente.objects.select_related('producto_talla__producto').get(id=cambio_id)
        
        if cambio.estado not in ['PENDIENTE', 'REVISADO']:
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden rechazar cambios pendientes o revisados'
            })
        
        cambio.estado = 'RECHAZADO'
        cambio.aprobado_por = request.user
        cambio.fecha_aprobacion = timezone.now()
        cambio.observaciones_aprobacion = observaciones
        cambio.save()
        
        # Notificar al creador (evitando duplicados)
        if cambio.creado_por and cambio.creado_por != request.user:
            existe = NotificacionCambioPrecio.objects.filter(
                cambio_precio=cambio,
                usuario=cambio.creado_por,
                tipo='RECHAZO'
            ).exists()
            
            if not existe:
                NotificacionCambioPrecio.objects.create(
                    cambio_precio=cambio,
                    usuario=cambio.creado_por,
                    tipo='RECHAZO',
                    mensaje=f"Tu cambio de precio para {cambio.producto_talla.producto.articulo} ha sido rechazado. Motivo: {observaciones}"
                )
        
        return JsonResponse({
            'success': True,
            'message': 'Cambio rechazado'
        })
        
    except CambioPrecioPendiente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al rechazar cambio: {str(e)}'
        })


@require_GET
@login_required
def obtener_notificaciones_precio(request):
    """
    Obtener notificaciones de cambios de precio para el usuario actual.
    Muestra notificaciones de cambios PENDIENTES y APLICADOS (informativas).
    IMPORTANTE: Solo muestra alertas de la sucursal actual del usuario y NO descartadas.
    """
    try:
        solo_no_leidas = request.GET.get('solo_no_leidas', 'false') == 'true'
        limit = int(request.GET.get('limit', 10))
        
        # Obtener la sucursal actual del usuario
        sucursal_actual_id = request.session.get('idSucursalActual')
        
        # Mostrar notificaciones de cambios PENDIENTES y APLICADOS (no descartados)
        # APLICADO = sincronización automática (informativa)
        # PENDIENTE = requiere acción manual
        queryset = NotificacionCambioPrecio.objects.filter(
            usuario=request.user,
            cambio_precio__estado__in=['PENDIENTE', 'APLICADO'],  # Ambos tipos
            cambio_precio__descartado=False     # No mostrar descartados
        ).select_related('cambio_precio__producto_talla__producto', 'cambio_precio__sucursal')
        
        # Filtrar por sucursal actual si está definida
        if sucursal_actual_id:
            queryset = queryset.filter(cambio_precio__sucursal_id=sucursal_actual_id)
        
        if solo_no_leidas:
            queryset = queryset.filter(leida=False)
        
        queryset = queryset.order_by('-fecha_creacion')[:limit]
        
        notificaciones_data = []
        for notif in queryset:
            notificaciones_data.append({
                'id': notif.id,
                'cambio_id': notif.cambio_precio.id,
                'tipo': notif.get_tipo_display(),
                'mensaje': notif.mensaje,
                'leida': notif.leida,
                'fecha_creacion': notif.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'producto': notif.cambio_precio.producto_talla.producto.articulo,
                'estado': notif.cambio_precio.estado,  # PENDIENTE o APLICADO
                'sucursal': notif.cambio_precio.sucursal.alias if notif.cambio_precio.sucursal else 'N/A'
            })
        
        # Contar notificaciones no leídas (PENDIENTES y APLICADOS) de la sucursal actual
        total_no_leidas_qs = NotificacionCambioPrecio.objects.filter(
            usuario=request.user,
            leida=False,
            cambio_precio__estado__in=['PENDIENTE', 'APLICADO'],  # Ambos tipos
            cambio_precio__descartado=False     # No contar descartados
        )
        if sucursal_actual_id:
            total_no_leidas_qs = total_no_leidas_qs.filter(cambio_precio__sucursal_id=sucursal_actual_id)
        total_no_leidas = total_no_leidas_qs.count()
        
        return JsonResponse({
            'success': True,
            'notificaciones': notificaciones_data,
            'total_no_leidas': total_no_leidas
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener notificaciones: {str(e)}'
        })


@require_POST
@login_required
def marcar_notificacion_leida(request):
    """
    Marcar una o todas las notificaciones como leídas
    """
    try:
        data = json.loads(request.body)
        notificacion_id = data.get('notificacion_id')
        marcar_todas = data.get('marcar_todas', False)
        
        if marcar_todas:
            # Marcar TODAS las notificaciones del usuario como leídas
            actualizadas = NotificacionCambioPrecio.objects.filter(
                usuario=request.user,
                leida=False
            ).update(leida=True, fecha_lectura=timezone.now())
            
            return JsonResponse({
                'success': True,
                'message': f'{actualizadas} notificaciones marcadas como leídas',
                'actualizadas': actualizadas
            })
        else:
            # Marcar una sola notificación
            notificacion = NotificacionCambioPrecio.objects.get(
                id=notificacion_id,
                usuario=request.user
            )
            
            notificacion.marcar_leida()
            
            # Contar cuántas quedan sin leer
            restantes = NotificacionCambioPrecio.objects.filter(
                usuario=request.user,
                leida=False
            ).count()
            
            return JsonResponse({
                'success': True,
                'message': 'Notificación marcada como leída',
                'restantes': restantes
            })
        
    except NotificacionCambioPrecio.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notificación no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al marcar notificación: {str(e)}'
        })


@require_POST
@login_required
def marcar_notificacion_leida_por_cambio(request, cambio_id):
    """
    Marca como leída la NotificacionCambioPrecio asociada a un CambioPrecioPendiente.
    Se usa al llegar a revisar-pendientes desde una notificación (?cambio=ID).
    """
    try:
        actualizadas = NotificacionCambioPrecio.objects.filter(
            cambio_precio_id=cambio_id,
            usuario=request.user,
            leida=False
        ).update(leida=True, fecha_lectura=timezone.now())

        restantes = NotificacionCambioPrecio.objects.filter(
            usuario=request.user,
            leida=False
        ).count()

        return JsonResponse({
            'success': True,
            'actualizadas': actualizadas,
            'restantes': restantes
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@require_POST
@login_required
def eliminar_notificaciones_precio(request):
    """
    Eliminar notificaciones de precios del usuario
    """
    try:
        data = json.loads(request.body)
        notificacion_id = data.get('notificacion_id')
        eliminar_todas = data.get('eliminar_todas', False)
        eliminar_leidas = data.get('eliminar_leidas', False)
        
        if eliminar_todas:
            # Eliminar TODAS las notificaciones del usuario
            eliminadas, _ = NotificacionCambioPrecio.objects.filter(
                usuario=request.user
            ).delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{eliminadas} notificaciones eliminadas',
                'eliminadas': eliminadas
            })
        elif eliminar_leidas:
            # Eliminar solo las notificaciones leídas
            eliminadas, _ = NotificacionCambioPrecio.objects.filter(
                usuario=request.user,
                leida=True
            ).delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{eliminadas} notificaciones leídas eliminadas',
                'eliminadas': eliminadas
            })
        elif notificacion_id:
            # Eliminar una sola notificación
            notificacion = NotificacionCambioPrecio.objects.get(
                id=notificacion_id,
                usuario=request.user
            )
            notificacion.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Notificación eliminada'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Debe especificar qué eliminar'
            })
        
    except NotificacionCambioPrecio.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notificación no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar notificación: {str(e)}'
        })


# ========== REGULARIZACIÓN DE PRECIOS ENTRE SUCURSALES ==========

@require_GET
@login_required
def detectar_discrepancias_precios(request):
    """
    Detecta productos con precios inconsistentes entre sucursales.
    Retorna un resumen y listado de productos que necesitan regularización.
    """
    try:
        # Obtener todas las sucursales activas
        sucursales = Sucursal.objects.filter(estado=True)
        
        if sucursales.count() < 2:
            return JsonResponse({
                'success': True,
                'tiene_discrepancias': False,
                'mensaje': 'Solo hay una sucursal activa',
                'total_discrepancias': 0,
                'productos': []
            })
        
        # Buscar productos que existen en múltiples sucursales
        # Agrupamos por artículo + atributo1 + atributo2 (nombre + marca + color)
        productos_agrupados = Producto.objects.values(
            'articulo', 'atributo1', 'atributo2'
        ).annotate(
            count_sucursales=Count('sucursal', distinct=True),
            precio_min=Min('precioventa'),
            precio_max=Max('precioventa'),
            costo_min=Min('costo'),
            costo_max=Max('costo')
        ).filter(
            count_sucursales__gt=1
        )
        
        discrepancias = []
        
        for grupo in productos_agrupados:
            # Calcular diferencia de precio
            diferencia_precio = grupo['precio_max'] - grupo['precio_min']
            diferencia_costo = grupo['costo_max'] - grupo['costo_min']
            
            # Si hay diferencia de precio o costo, es una discrepancia
            if diferencia_precio > 0 or diferencia_costo > 0:
                # Obtener los productos de este grupo
                productos_grupo = Producto.objects.filter(
                    articulo=grupo['articulo'],
                    atributo1_id=grupo['atributo1'],
                    atributo2_id=grupo['atributo2']
                ).select_related('sucursal', 'atributo1', 'atributo2', 'atributo3')
                
                # Obtener detalles por sucursal
                sucursales_detalle = []
                for prod in productos_grupo:
                    stock_total = sum(pt.stock for pt in prod.producto_talla.all())
                    sucursales_detalle.append({
                        'sucursal_id': prod.sucursal.id,
                        'sucursal': prod.sucursal.alias,
                        'producto_id': prod.id,
                        'precio_venta': float(prod.precioventa),
                        'costo': float(prod.costo),
                        'stock': stock_total
                    })
                
                # Calcular porcentaje de variación
                if grupo['precio_min'] > 0:
                    variacion_porcentual = (diferencia_precio / grupo['precio_min']) * 100
                else:
                    variacion_porcentual = 0
                
                # Obtener marca y color
                primer_prod = productos_grupo.first()
                
                discrepancias.append({
                    'articulo': grupo['articulo'],
                    'marca': primer_prod.atributo1.valor if primer_prod.atributo1 else None,
                    'color': primer_prod.atributo2.valor if primer_prod.atributo2 else None,
                    'genero': primer_prod.atributo3.valor if primer_prod.atributo3 else None,
                    'precio_min': float(grupo['precio_min']),
                    'precio_max': float(grupo['precio_max']),
                    'diferencia_precio': float(diferencia_precio),
                    'costo_min': float(grupo['costo_min']),
                    'costo_max': float(grupo['costo_max']),
                    'diferencia_costo': float(diferencia_costo),
                    'variacion_porcentual': round(variacion_porcentual, 2),
                    'cantidad_sucursales': grupo['count_sucursales'],
                    'sucursales': sucursales_detalle,
                    'es_critico': variacion_porcentual > 10  # Más de 10% de diferencia es crítico
                })
        
        # Ordenar por variación porcentual descendente
        discrepancias.sort(key=lambda x: x['variacion_porcentual'], reverse=True)
        
        # Calcular resumen
        total_discrepancias = len(discrepancias)
        criticos = sum(1 for d in discrepancias if d['es_critico'])
        variacion_maxima = max(d['variacion_porcentual'] for d in discrepancias) if discrepancias else 0
        
        return JsonResponse({
            'success': True,
            'tiene_discrepancias': total_discrepancias > 0,
            'total_discrepancias': total_discrepancias,
            'criticos': criticos,
            'variacion_maxima': variacion_maxima,
            'productos': discrepancias[:50]  # Limitar a 50 para el dashboard
        })
        
    except Exception as e:
        logger.exception("Error en detectar_discrepancias_precios")
        return JsonResponse({
            'success': False,
            'error': f'Error al detectar discrepancias: {str(e)}'
        })


@require_POST
@login_required
def regularizar_precio_sucursales(request):
    """
    Regulariza el precio de un producto igualándolo en todas las sucursales.
    """
    try:
        data = json.loads(request.body)
        articulo = data.get('articulo')
        atributo1_id = data.get('atributo1_id')
        atributo2_id = data.get('atributo2_id')
        precio_nuevo = data.get('precio_nuevo')
        costo_nuevo = data.get('costo_nuevo')
        motivo = data.get('motivo', 'Regularización de precio entre sucursales')
        
        if not all([articulo, atributo1_id, atributo2_id, precio_nuevo]):
            return JsonResponse({
                'success': False,
                'error': 'Parámetros incompletos'
            })
        
        # Obtener todos los productos que coinciden
        productos = Producto.objects.filter(
            articulo=articulo,
            atributo1_id=atributo1_id,
            atributo2_id=atributo2_id
        ).select_related('sucursal')
        
        if not productos.exists():
            return JsonResponse({
                'success': False,
                'error': 'No se encontraron productos para regularizar'
            })
        
        productos_actualizados = 0
        sucursales_afectadas = set()
        
        with transaction.atomic():
            for producto in productos:
                precio_anterior = producto.precioventa
                costo_anterior = producto.costo
                
                # Actualizar precios
                producto.precioventa = int(precio_nuevo)
                if costo_nuevo:
                    producto.costo = int(costo_nuevo)
                producto.save()
                
                # Actualizar lotes activos
                LoteProducto.objects.filter(
                    producto_talla__producto=producto,
                    cantidad_disponible__gt=0,
                    activo=True
                ).update(precio_venta_unitario=int(precio_nuevo))
                
                # Registrar en historial
                if precio_anterior != int(precio_nuevo):
                    HistorialCambioPrecio.objects.create(
                        producto=producto,
                        precio_anterior=precio_anterior,
                        precio_nuevo=int(precio_nuevo),
                        tipo_cambio='REGULARIZACION',
                        motivo=motivo,
                        usuario=request.user,
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                
                productos_actualizados += 1
                sucursales_afectadas.add(producto.sucursal.alias)
        
        return JsonResponse({
            'success': True,
            'message': f'Precio regularizado en {productos_actualizados} productos',
            'productos_actualizados': productos_actualizados,
            'sucursales_afectadas': list(sucursales_afectadas)
        })
        
    except Exception as e:
        logger.exception("Error en regularizar_precio_sucursales")
        return JsonResponse({
            'success': False,
            'error': f'Error al regularizar: {str(e)}'
        })


@require_GET
@login_required
def resumen_discrepancias_precios(request):
    """
    Retorna un resumen rápido de discrepancias para el dashboard.
    Optimizado para carga rápida.
    """
    try:
        # Contar productos con discrepancias de precio
        productos_con_discrepancia = Producto.objects.values(
            'articulo', 'atributo1', 'atributo2'
        ).annotate(
            count_sucursales=Count('sucursal', distinct=True),
            precio_min=Min('precioventa'),
            precio_max=Max('precioventa')
        ).filter(
            count_sucursales__gt=1
        )
        
        # Filtrar los que tienen diferencia real
        total_discrepancias = 0
        criticos = 0
        variacion_maxima = 0
        
        for grupo in productos_con_discrepancia:
            diferencia = grupo['precio_max'] - grupo['precio_min']
            if diferencia > 0:
                total_discrepancias += 1
                
                if grupo['precio_min'] > 0:
                    variacion = (diferencia / grupo['precio_min']) * 100
                    if variacion > variacion_maxima:
                        variacion_maxima = variacion
                    if variacion > 10:
                        criticos += 1
        
        return JsonResponse({
            'success': True,
            'total_discrepancias': total_discrepancias,
            'criticos': criticos,
            'variacion_maxima': round(variacion_maxima, 2),
            'requiere_atencion': total_discrepancias > 0
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'total_discrepancias': 0,
            'criticos': 0,
            'variacion_maxima': 0,
            'requiere_atencion': False
        })
