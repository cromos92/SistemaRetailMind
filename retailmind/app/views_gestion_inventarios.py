"""
Módulo de Gestión de Inventarios - RetailMind
============================================

Sistema completo de toma de inventario físico con:
- Inventario por segmentos (marca, categoría, atributo)
- Fecha de corte para congelamiento de datos
- Análisis previo antes de aplicar ajustes
- Procesamiento en lotes para grandes volúmenes
- Optimización de queries para evitar N+1

Mejores Prácticas de Logística Implementadas:
- Conteo cíclico y ABC
- Reconteo automático para diferencias significativas
- Trazabilidad completa de ajustes
- Reportes de análisis antes de aprobar
"""

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.db.models import Sum, F, Q, Count, Case, When, Prefetch, Value, CharField, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction, connection
from django.core.exceptions import ValidationError, PermissionDenied
import threading
from decimal import Decimal
import csv
import io
import json
import logging

from .models import (
    Producto, Producto_Talla, Productos_Atributos, AtributoOpcion, Categoria,
    LoteProducto, Movimientos_Producto, Sucursal, Empresa, EmpresaUser,
    TomaInventario, TomaInventarioDetalle, TomaInventarioLog, TareaAplicacionAjustes
)
from .utils_permisos import (
    puede_ver_sucursal, obtener_empresas_usuario, obtener_sucursales_usuario
)

logger = logging.getLogger('app')

# ==============================================================================
# CONSTANTES Y CONFIGURACIÓN
# ==============================================================================

BATCH_SIZE = 500  # Tamaño de lote para operaciones masivas
DIFERENCIA_RECONTEO_PORCENTAJE = 10  # % de diferencia para requerir reconteo
DIFERENCIA_RECONTEO_UNIDADES = 5  # Unidades de diferencia para requerir reconteo

# Estados en los que el inventario todavía se está contando
ESTADOS_EN_PROCESO = ['BORRADOR', 'EN_CONTEO', 'CONTEO_FINALIZADO', 'EN_REVISION']


def _inventario_del_usuario(request, inventario_id):
    """
    Recupera la toma verificando que pertenezca a una empresa/sucursal del usuario.

    Sin esto cualquier usuario autenticado podía abrir (y aprobar) el inventario de
    otra empresa simplemente cambiando el id en la URL.
    """
    inventario = get_object_or_404(
        TomaInventario.objects.select_related('sucursal', 'empresa'), id=inventario_id
    )
    empresas_ids = set(obtener_empresas_usuario(request.user).values_list('id', flat=True))
    if inventario.empresa_id not in empresas_ids:
        return None
    if not puede_ver_sucursal(request.user, inventario.sucursal_id):
        return None
    return inventario


def _error_sin_acceso():
    return JsonResponse(
        {'success': False, 'error': 'No tiene acceso a este inventario'}, status=403
    )


# ==============================================================================
# VISTAS PRINCIPALES
# ==============================================================================

@login_required
def gestion_inventarios(request):
    """Vista principal del módulo de Gestión de Inventarios"""
    return render(request, 'vistas/modulo_existencias/gestion_inventarios.html')


@login_required
def detalle_inventario(request, inventario_id):
    """Vista de detalle de un inventario específico"""
    inventario = _inventario_del_usuario(request, inventario_id)
    if inventario is None:
        raise PermissionDenied('No tiene acceso a este inventario')
    return render(request, 'vistas/modulo_existencias/detalle_inventario.html', {
        'inventario': inventario,
        'puede_aplicar_ajustes': inventario.estado in ('APROBADO', 'APLICANDO'),
    })


# ==============================================================================
# API: LISTADO Y FILTROS
# ==============================================================================

@require_GET
@login_required
def obtener_inventarios(request):
    """
    Obtener lista de inventarios con filtros y paginación.
    Optimizado para evitar N+1 queries.
    """
    try:
        sucursal_id = request.session.get('idSucursalActual')

        # Antes se filtraba por `EmpresaUser...first().empresa`: con multi-empresa eso
        # tomaba UNA empresa al azar del usuario (un administrador tiene 1.699) y podía
        # dejar fuera inventarios que sí le corresponden.
        empresas_ids = list(obtener_empresas_usuario(request.user).values_list('id', flat=True))
        if not empresas_ids:
            return JsonResponse({'success': False, 'error': 'Usuario sin empresa asignada'})

        # Parámetros de paginación
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Parámetros de filtro
        estado = request.GET.get('estado')
        tipo = request.GET.get('tipo')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        search = request.GET.get('search', '').strip()
        grupo = request.GET.get('grupo', '').strip()  # EN_PROCESO | CON_DIFERENCIAS

        # Construir queryset optimizado
        queryset = TomaInventario.objects.select_related(
            'sucursal', 'empresa', 'creado_por', 'aprobado_por'
        ).filter(empresa_id__in=empresas_ids)

        # Filtrar por sucursal actual (respetando el acceso del usuario)
        if sucursal_id and puede_ver_sucursal(request.user, sucursal_id):
            queryset = queryset.filter(sucursal_id=sucursal_id)
        else:
            sucursales_ids = list(
                obtener_sucursales_usuario(request.user).values_list('id', flat=True)
            )
            queryset = queryset.filter(sucursal_id__in=sucursales_ids)

        # Aplicar filtros
        if estado:
            queryset = queryset.filter(estado=estado)
        
        if tipo:
            queryset = queryset.filter(tipo_inventario=tipo)
        
        if fecha_desde:
            queryset = queryset.filter(fecha_corte__date__gte=fecha_desde)
        
        if fecha_hasta:
            queryset = queryset.filter(fecha_corte__date__lte=fecha_hasta)
        
        if search:
            queryset = queryset.filter(
                Q(numero_inventario__icontains=search) |
                Q(nombre__icontains=search)
            )

        # Los tabs "En Proceso" y "Con Diferencias" antes solo cambiaban una variable
        # de JS que nunca se enviaba: filtraban nada.
        if grupo == 'EN_PROCESO':
            queryset = queryset.filter(estado__in=ESTADOS_EN_PROCESO)
        elif grupo == 'CON_DIFERENCIAS':
            queryset = queryset.exclude(estado__in=['COMPLETADO', 'CANCELADO']).filter(
                Q(total_diferencias_positivas__gt=0) | Q(total_diferencias_negativas__gt=0)
            )

        # Resumen sobre TODO el conjunto filtrado (antes los KPIs contaban solo la
        # página visible: con 20 por página el total nunca podía pasar de 20).
        base_resumen = queryset
        resumen = {
            'total': base_resumen.count(),
            'en_proceso': base_resumen.filter(estado__in=ESTADOS_EN_PROCESO).count(),
            'pendientes_aprobacion': base_resumen.filter(estado='PENDIENTE_APROBACION').count(),
            'completados': base_resumen.filter(estado='COMPLETADO').count(),
            'con_diferencias': base_resumen.exclude(estado__in=['COMPLETADO', 'CANCELADO']).filter(
                Q(total_diferencias_positivas__gt=0) | Q(total_diferencias_negativas__gt=0)
            ).count(),
        }

        # Ordenar y paginar
        queryset = queryset.order_by('-created_at')
        paginator = Paginator(queryset, per_page)
        inventarios_page = paginator.get_page(page)

        # SKUs realmente contados (el campo total_productos_contados del modelo guarda
        # UNIDADES, no líneas: mostrarlo contra total_productos_esperados comparaba
        # peras con manzanas — en INV-6-20260115-001 decía "9.490 / 37.389 productos"
        # cuando lo contado eran 4.301 SKUs).
        skus_map = {
            row['toma_inventario_id']: row
            for row in TomaInventarioDetalle.objects
            .filter(toma_inventario__in=list(inventarios_page.object_list), excluir_de_analisis=False)
            .values('toma_inventario_id')
            .annotate(lineas=Count('id'), contados=Count('id', filter=Q(contado=True)))
        }

        # Serializar datos
        inventarios_data = []
        for inv in inventarios_page:
            conteo = skus_map.get(inv.id) or {}
            inventarios_data.append({
                'skus_contados': conteo.get('contados', 0),
                'skus_esperados': conteo.get('lineas', inv.total_productos_esperados),
                'unidades_contadas': inv.total_productos_contados,
                'id': inv.id,
                'numero_inventario': inv.numero_inventario,
                'nombre': inv.nombre,
                'sucursal': inv.sucursal.alias,
                'tipo_inventario': inv.tipo_inventario,
                'tipo_inventario_display': inv.get_tipo_inventario_display(),
                'estado': inv.estado,
                'estado_display': inv.get_estado_display(),
                'fecha_corte': inv.fecha_corte.strftime('%d/%m/%Y %H:%M'),
                'progreso_conteo': float(inv.progreso_conteo),
                'total_productos_esperados': inv.total_productos_esperados,
                'total_productos_contados': inv.total_productos_contados,
                'total_diferencias_positivas': inv.total_diferencias_positivas,
                'total_diferencias_negativas': inv.total_diferencias_negativas,
                'valor_diferencias': float(inv.valor_diferencias_positivas - inv.valor_diferencias_negativas),
                'creado_por': inv.creado_por.get_full_name() if inv.creado_por else '',
                'created_at': inv.created_at.strftime('%d/%m/%Y %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'inventarios': inventarios_data,
            'resumen': resumen,
            'pagination': {
                'current_page': inventarios_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': inventarios_page.has_next(),
                'has_previous': inventarios_page.has_previous(),
            }
        })

    except Exception as e:
        logger.error(f"Error al obtener inventarios: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def obtener_filtros_disponibles(request):
    """
    Obtener opciones de filtros disponibles para crear inventario.
    Devuelve marcas, categorías y atributos activos.
    """
    try:
        empresa_user = EmpresaUser.objects.filter(
            user=request.user, 
            active=True
        ).select_related('empresa').first()
        
        if not empresa_user:
            return JsonResponse({'success': False, 'error': 'Usuario sin empresa asignada'})
        
        sucursal_id = request.session.get('idSucursalActual')

        # Marcas y categorías ACOTADAS a la sucursal activa y con stock: antes se
        # ofrecían las de todo el holding, así que era fácil segmentar por una marca
        # que no existe en la tienda y crear una toma vacía.
        productos_suc = Producto.objects.all()
        if sucursal_id:
            productos_suc = productos_suc.filter(sucursal_id=sucursal_id)
        productos_con_stock = productos_suc.filter(producto_talla__stock__gt=0)

        marcas = AtributoOpcion.objects.filter(
            atributo__nombre__icontains='marca',
            productos_marca__in=productos_con_stock
        ).distinct().values('id', 'valor').order_by('valor')

        categorias = Categoria.objects.filter(
            categoria_productos__in=productos_con_stock
        ).distinct().values('id', 'nombre').order_by('nombre')

        # Obtener atributos disponibles (color, género, etc.)
        atributos = Productos_Atributos.objects.filter(
            opciones__isnull=False
        ).distinct().prefetch_related('opciones')
        
        atributos_data = []
        for attr in atributos:
            if attr.nombre.lower() != 'marca':  # Excluir marca que ya tiene su filtro
                atributos_data.append({
                    'id': attr.id,
                    'nombre': attr.nombre,
                    'opciones': list(attr.opciones.values('id', 'valor').order_by('valor'))
                })
        
        return JsonResponse({
            'success': True,
            'filtros': {
                'marcas': list(marcas),
                'categorias': list(categorias),
                'atributos': atributos_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error al obtener filtros: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================================================================
# API: CREACIÓN DE INVENTARIO
# ==============================================================================

@require_POST
@login_required
@transaction.atomic
def crear_inventario(request):
    """
    Crear una nueva toma de inventario.
    
    Proceso:
    1. Validar datos de entrada
    2. Crear encabezado de inventario
    3. Generar snapshot de productos según filtros
    4. Calcular stock del sistema en fecha de corte
    """
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        nombre = data.get('nombre')
        tipo_inventario = data.get('tipo_inventario', 'COMPLETO')
        fecha_corte_str = data.get('fecha_corte')
        filtros = data.get('filtros', {})
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es requerido'})

        # Tipos declarados en el modelo pero sin lógica de segmentación implementada:
        # si se aceptan, generan igual un inventario COMPLETO (335.000 líneas en EDEL)
        # haciendo creer al usuario que contará solo una muestra.
        if tipo_inventario in ('SELECTIVO', 'CICLICO', 'ALEATORIO'):
            return JsonResponse({
                'success': False,
                'error': f'El tipo "{tipo_inventario}" todavía no está implementado. '
                         f'Usa Por Marca, Por Categoría o Por Atributo para una toma parcial.'
            })

        segmentado = {
            'POR_MARCA': 'marcas',
            'POR_CATEGORIA': 'categorias',
        }.get(tipo_inventario)
        if segmentado and not filtros.get(segmentado):
            return JsonResponse({
                'success': False,
                'error': 'Debe seleccionar al menos un valor de segmentación para este tipo de inventario'
            })
        if tipo_inventario == 'POR_ATRIBUTO' and not any((filtros.get('atributos') or {}).values()):
            return JsonResponse({
                'success': False,
                'error': 'Debe seleccionar al menos un atributo para este tipo de inventario'
            })

        # Obtener empresa y sucursal
        empresa_user = EmpresaUser.objects.filter(
            user=request.user,
            active=True
        ).select_related('empresa', 'sucursal').first()

        if not empresa_user:
            return JsonResponse({'success': False, 'error': 'Usuario sin empresa asignada'})

        sucursal_id = request.session.get('idSucursalActual')
        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una sucursal'})

        if not puede_ver_sucursal(request.user, sucursal_id):
            return JsonResponse({'success': False, 'error': 'No tiene acceso a la sucursal activa'})

        sucursal = get_object_or_404(Sucursal, id=sucursal_id)

        # Procesar fecha de corte
        if fecha_corte_str:
            from datetime import datetime
            fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%dT%H:%M')
            fecha_corte = timezone.make_aware(fecha_corte) if timezone.is_naive(fecha_corte) else fecha_corte
        else:
            fecha_corte = timezone.now()
        
        # Crear inventario
        numero_inventario = TomaInventario.generar_numero_inventario(sucursal)
        
        inventario = TomaInventario.objects.create(
            numero_inventario=numero_inventario,
            nombre=nombre,
            sucursal=sucursal,
            empresa=empresa_user.empresa,
            tipo_inventario=tipo_inventario,
            filtros_aplicados=filtros,
            fecha_corte=fecha_corte,
            estado='BORRADOR',
            observaciones=(data.get('observaciones') or '').strip() or None,
            creado_por=request.user
        )

        # Generar detalles de productos a inventariar
        total_productos = _generar_detalles_inventario(inventario, filtros, sucursal_id)

        if total_productos == 0:
            # Sin líneas la toma es inútil y queda basura en el listado.
            raise ValidationError(
                'Los filtros seleccionados no arrojaron productos con stock en esta sucursal.'
            )

        # Actualizar total esperado
        inventario.total_productos_esperados = total_productos
        inventario.save()

        # Registrar log
        _registrar_log(
            inventario=inventario,
            tipo_accion='CREACION',
            descripcion=f'Inventario creado con {total_productos} productos a contar',
            usuario=request.user,
            datos={'filtros': filtros, 'total_productos': total_productos}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Inventario {numero_inventario} creado exitosamente',
            'inventario_id': inventario.id,
            'numero_inventario': numero_inventario,
            'total_productos': total_productos
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': '; '.join(e.messages)})
    except Exception as e:
        logger.error(f"Error al crear inventario: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


def _generar_detalles_inventario(inventario, filtros, sucursal_id):
    """
    Genera los detalles del inventario según los filtros.
    Optimizado para grandes volúmenes usando bulk_create.
    """
    # Construir queryset de productos según filtros
    queryset = Producto_Talla.objects.select_related(
        'producto', 
        'producto__atributo1',  # marca
        'producto__categoria'
    ).filter(
        producto__sucursal_id=sucursal_id
    )
    
    # Aplicar filtros
    marcas = filtros.get('marcas', [])
    categorias = filtros.get('categorias', [])
    atributos = filtros.get('atributos', {})
    productos_ids = filtros.get('productos', [])

    if marcas:
        queryset = queryset.filter(producto__atributo1_id__in=marcas)

    if categorias:
        queryset = queryset.filter(producto__categoria_id__in=categorias)

    if productos_ids:
        queryset = queryset.filter(producto_id__in=productos_ids)

    # Filtros de atributos específicos
    for attr_id, opciones in atributos.items():
        if opciones:
            # Atributo2 = color, Atributo3 = género, etc.
            if attr_id == 'color':
                queryset = queryset.filter(producto__atributo2_id__in=opciones)
            elif attr_id == 'genero':
                queryset = queryset.filter(producto__atributo3_id__in=opciones)

    # "Solo productos con stock" evita generar decenas de miles de líneas en cero.
    # En bodega EDEL hay 341.945 tallas y solo 275 con stock: sin este filtro la
    # toma nace con 335.000 líneas imposibles de contar.
    if filtros.get('solo_con_stock', True):
        queryset = queryset.filter(stock__gt=0)

    # Procesar productos en lotes para reducir consultas N+1
    detalles = []
    batch = []

    def _procesar_batch(batch_items):
        if not batch_items:
            return

        ids = [pt.id for pt in batch_items]
        # Movimientos ocurridos DESPUÉS del corte: el stock al corte se reconstruye
        # restándolos del stock actual (ver _obtener_movimientos_desde_corte_batch).
        posteriores_map = _obtener_movimientos_desde_corte_batch(ids, inventario.fecha_corte, sucursal_id)
        costo_map = _obtener_costo_promedio_batch(ids)

        for pt in batch_items:
            # Stock del sistema en la fecha de corte.
            # OJO: la base SIEMPRE parte del stock plano (Producto_Talla.stock),
            # que es el número que usa el resto del ERP (POS, reportes, ecommerce).
            # Antes se usaba la suma del kardex y, como kardex y stock plano no
            # cuadran en 126k SKUs, el ajuste dejaba el stock distinto de lo contado.
            stock_sistema = (pt.stock or 0) - posteriores_map.get(pt.id, 0)

            # Calcular costo promedio FIFO
            costo_promedio = costo_map.get(pt.id)
            if costo_promedio is None:
                costo_promedio = Decimal(pt.producto.costo or 0)

            # Obtener nombres desnormalizados
            marca_nombre = pt.producto.atributo1.valor if pt.producto.atributo1 else ''
            categoria_nombre = pt.producto.categoria.nombre if pt.producto.categoria else ''

            detalles.append(TomaInventarioDetalle(
                toma_inventario=inventario,
                producto_talla=pt,
                sku=str(pt.sku),
                producto_nombre=pt.producto.articulo,
                talla_nombre=pt.talla if pt.talla else '',
                marca_nombre=marca_nombre,
                categoria_nombre=categoria_nombre,
                stock_sistema=stock_sistema,
                stock_movimientos_post_corte=0,
                stock_sistema_ajustado=stock_sistema,
                costo_unitario_sistema=costo_promedio,
                precio_venta_sistema=Decimal(pt.producto.precioventa or 0)
            ))

        if len(detalles) >= BATCH_SIZE:
            TomaInventarioDetalle.objects.bulk_create(detalles, ignore_conflicts=True)
            detalles.clear()

    for pt in queryset.iterator(chunk_size=BATCH_SIZE):
        batch.append(pt)
        if len(batch) >= BATCH_SIZE:
            _procesar_batch(batch)
            batch = []

    if batch:
        _procesar_batch(batch)

    if detalles:
        TomaInventarioDetalle.objects.bulk_create(detalles, ignore_conflicts=True)

    return inventario.detalles.count()


def _fecha_hora_local(momento):
    """
    Movimientos_Producto guarda `fecha`/`hora` en horario local (timezone.localdate()
    y timezone.localtime()). Los DateTimeField llegan en UTC, así que comparar
    `momento.date()` contra `Movimientos_Producto.fecha` corría el corte 3-4 horas
    (todo lo vendido después de las 20:00 caía en el día siguiente).
    """
    local = timezone.localtime(momento) if timezone.is_aware(momento) else momento
    return local.date(), local.time()


def _obtener_movimientos_desde_corte_batch(producto_talla_ids, fecha_corte, sucursal_id):
    """
    Suma neta de movimientos ocurridos DESPUÉS de la fecha de corte y hasta ahora.

    Sirve para reconstruir el stock al corte: stock_al_corte = stock_actual - esta suma.
    Así la base de comparación queda anclada al stock plano (el que ve el POS) y no
    a la suma del kardex, que en producción difiere en 126.455 SKUs.
    """
    corte_date, corte_time = _fecha_hora_local(fecha_corte)

    movimientos = Movimientos_Producto.objects.filter(
        ProductoTalla_id__in=producto_talla_ids
    ).filter(
        Q(sucursal_destino_id=sucursal_id) | Q(sucursal_origen_id=sucursal_id)
    ).filter(
        Q(fecha__gt=corte_date) |
        Q(fecha=corte_date, hora__gt=corte_time)
    ).values('ProductoTalla_id').annotate(
        total_cantidad=Coalesce(Sum('cantidad'), 0)
    )

    return {m['ProductoTalla_id']: m['total_cantidad'] for m in movimientos}


def _obtener_costo_promedio_batch(producto_talla_ids):
    """
    Calcula costo promedio ponderado FIFO por lote de productos.
    Evita N+1 consultando lotes agregados.
    """
    lotes = LoteProducto.objects.filter(
        producto_talla_id__in=producto_talla_ids,
        cantidad_disponible__gt=0,
        activo=True
    ).values('producto_talla_id').annotate(
        total_valor=Coalesce(
            Sum(ExpressionWrapper(
                F('cantidad_disponible') * F('costo_unitario'),
                output_field=DecimalField(max_digits=18, decimal_places=6)
            )),
            Value(0),
            output_field=DecimalField(max_digits=18, decimal_places=6)
        ),
        total_cantidad=Coalesce(Sum('cantidad_disponible'), 0)
    )

    costo_map = {}
    for lote in lotes:
        if lote['total_cantidad'] > 0:
            costo_map[lote['producto_talla_id']] = lote['total_valor'] / lote['total_cantidad']

    return costo_map


def _obtener_movimientos_post_corte_batch(producto_talla_ids, fecha_corte, fecha_conteo, sucursal_id):
    """
    Obtiene la suma neta de movimientos entre fecha de corte y fecha de conteo.
    Considera sucursal origen/destino y respeta fecha/hora.
    """
    fecha_corte_date, fecha_corte_time = _fecha_hora_local(fecha_corte)
    fecha_conteo_date, fecha_conteo_time = _fecha_hora_local(fecha_conteo)

    movimientos = Movimientos_Producto.objects.filter(
        ProductoTalla_id__in=producto_talla_ids
    ).filter(
        Q(sucursal_destino_id=sucursal_id) | Q(sucursal_origen_id=sucursal_id)
    ).filter(
        Q(fecha__gt=fecha_corte_date) |
        Q(fecha=fecha_corte_date, hora__gt=fecha_corte_time)
    ).filter(
        Q(fecha__lt=fecha_conteo_date) |
        Q(fecha=fecha_conteo_date, hora__lte=fecha_conteo_time)
    ).values('ProductoTalla_id').annotate(
        total_cantidad=Coalesce(Sum('cantidad'), 0)
    )

    movimientos_map = {}
    for mov in movimientos:
        movimientos_map[mov['ProductoTalla_id']] = mov['total_cantidad']

    return movimientos_map


# NOTA: se eliminaron _calcular_stock_fecha_corte() y _calcular_costo_promedio_fifo().
# Estaban muertas (nadie las llamaba) y la primera reconstruía el stock desde
# tipo_movimiento='INGRESO'/'EGRESO', clasificación que en este proyecto no es
# confiable (el default del modelo es 'INGRESO'). El stock al corte ahora se
# calcula en _procesar_batch a partir del stock plano.


# ==============================================================================
# API: CONTEO DE PRODUCTOS
# ==============================================================================

@require_GET
@login_required
def obtener_productos_conteo(request, inventario_id):
    """
    Obtener productos para conteo con paginación y filtros.
    Optimizado para grandes volúmenes.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        # Parámetros
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        estado_conteo = request.GET.get('estado_conteo')  # contado, pendiente, reconteo
        search = request.GET.get('search', '').strip()
        solo_diferencias = request.GET.get('solo_diferencias') == 'true'
        signo = request.GET.get('signo', '')  # positiva | negativa | cero
        marca = request.GET.get('marca')
        categoria = request.GET.get('categoria')
        
        # Construir queryset
        queryset = inventario.detalles.all()
        
        if estado_conteo == 'contado':
            queryset = queryset.filter(contado=True)
        elif estado_conteo == 'pendiente':
            queryset = queryset.filter(contado=False)
        elif estado_conteo == 'reconteo':
            queryset = queryset.filter(reconteo_requerido=True, stock_reconteo__isnull=True)
        
        if search:
            queryset = queryset.filter(
                Q(sku__icontains=search) |
                Q(producto_nombre__icontains=search)
            )
        
        if solo_diferencias:
            queryset = queryset.filter(contado=True).exclude(diferencia=0)

        # Permite que las tarjetas "Sobrantes"/"Faltantes"/"Coinciden" del detalle
        # filtren la tabla en vez de ser sólo decorativas
        if signo == 'positiva':
            queryset = queryset.filter(contado=True, diferencia__gt=0)
        elif signo == 'negativa':
            queryset = queryset.filter(contado=True, diferencia__lt=0)
        elif signo == 'cero':
            queryset = queryset.filter(contado=True, diferencia=0)

        if marca:
            queryset = queryset.filter(marca_nombre__icontains=marca)

        if categoria:
            queryset = queryset.filter(categoria_nombre__icontains=categoria)

        # Ordenar
        queryset = queryset.order_by('producto_nombre', 'talla_nombre')
        
        # Paginar
        paginator = Paginator(queryset, per_page)
        productos_page = paginator.get_page(page)
        
        # Serializar
        productos_data = []
        for det in productos_page:
            productos_data.append({
                'id': det.id,
                'sku': det.sku,
                'producto_nombre': det.producto_nombre,
                'talla_nombre': det.talla_nombre,
                'marca_nombre': det.marca_nombre,
                'categoria_nombre': det.categoria_nombre,
                'stock_sistema': det.stock_sistema,
                'stock_movimientos_post_corte': det.stock_movimientos_post_corte,
                'stock_sistema_ajustado': det.stock_sistema_ajustado,
                'stock_fisico': det.stock_fisico,
                'diferencia': det.diferencia,
                'porcentaje_diferencia': round(det.porcentaje_diferencia, 2),
                'valor_diferencia': float(det.valor_diferencia),
                'contado': det.contado,
                'excluido': det.excluir_de_analisis,
                'fecha_conteo': det.fecha_conteo.strftime('%d/%m/%Y %H:%M') if det.fecha_conteo else None,
                'reconteo_requerido': det.reconteo_requerido,
                'stock_reconteo': det.stock_reconteo,
                'ubicacion': det.ubicacion,
                'observaciones': det.observaciones,
                'costo_unitario': float(det.costo_unitario_sistema),
                'precio_venta': float(det.precio_venta_sistema)
            })
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'pagination': {
                'current_page': productos_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': productos_page.has_next(),
                'has_previous': productos_page.has_previous(),
            }
        })
        
    except Exception as e:
        logger.error(f"Error al obtener productos para conteo: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
@transaction.atomic
def registrar_conteo(request, inventario_id):
    """
    Registrar conteo físico de uno o más productos.
    Soporta conteo individual y masivo.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        # Verificar estado
        if inventario.estado not in ['BORRADOR', 'EN_CONTEO']:
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está en estado de conteo'
            })
        
        data = json.loads(request.body)
        conteos = data.get('conteos', [])
        
        if not conteos:
            return JsonResponse({'success': False, 'error': 'No hay conteos para registrar'})
        
        # Actualizar estado si es el primer conteo
        if inventario.estado == 'BORRADOR':
            inventario.estado = 'EN_CONTEO'
            inventario.fecha_inicio_conteo = timezone.now()
            inventario.save()
        
        # Procesar conteos
        conteos_realizados = 0
        errores = []
        fecha_conteo = timezone.now()

        conteos_map = {
            c.get('detalle_id'): c for c in conteos if c.get('detalle_id') is not None
        }
        detalle_ids = list(conteos_map.keys())
        detalles = inventario.detalles.filter(id__in=detalle_ids).select_related('producto_talla')

        producto_talla_ids = [d.producto_talla_id for d in detalles]
        movimientos_map = _obtener_movimientos_post_corte_batch(
            producto_talla_ids, inventario.fecha_corte, fecha_conteo, inventario.sucursal_id
        )

        detalles_por_id = {d.id: d for d in detalles}

        for detalle_id, conteo in conteos_map.items():
            detalle = detalles_por_id.get(detalle_id)
            if not detalle:
                errores.append(f"Detalle {detalle_id} no encontrado")
                continue

            stock_fisico = conteo.get('stock_fisico')
            ubicacion = conteo.get('ubicacion', '')
            observaciones = conteo.get('observaciones', '')

            try:
                movimientos_post_corte = movimientos_map.get(detalle.producto_talla_id, 0)
                detalle.stock_movimientos_post_corte = movimientos_post_corte
                detalle.stock_sistema_ajustado = detalle.stock_sistema + movimientos_post_corte
                detalle.stock_fisico = int(stock_fisico)
                detalle.contado = True
                detalle.fecha_conteo = fecha_conteo
                detalle.usuario_conteo = request.user
                detalle.ubicacion = ubicacion
                detalle.observaciones = observaciones
                detalle.save()  # El save() calcula diferencia automáticamente
                conteos_realizados += 1
            except Exception as e:
                errores.append(f"Error en detalle {detalle_id}: {str(e)}")
        
        # Recalcular métricas del inventario
        inventario.calcular_metricas()
        
        # Registrar log
        _registrar_log(
            inventario=inventario,
            tipo_accion='REGISTRO_CONTEO',
            descripcion=f'{conteos_realizados} productos contados',
            usuario=request.user,
            datos={'conteos_realizados': conteos_realizados, 'errores': errores}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{conteos_realizados} conteos registrados',
            'conteos_realizados': conteos_realizados,
            'errores': errores if errores else None,
            'progreso': float(inventario.progreso_conteo)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        logger.error(f"Error al registrar conteo: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


def _leer_archivo_conteo(archivo, nombre_hoja='', max_rows=None, read_only=False):
    filas = []
    total_leidas = 0

    if archivo.name.lower().endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            return None, 'openpyxl no está instalado'

        wb = openpyxl.load_workbook(archivo, data_only=True, read_only=read_only)
        hoja = wb.active
        if nombre_hoja:
            if nombre_hoja not in wb.sheetnames:
                return None, f'No existe la hoja "{nombre_hoja}"'
            hoja = wb[nombre_hoja]

        for row in hoja.iter_rows(values_only=True):
            fila = [str(c).strip() if c is not None else '' for c in row]
            if any(fila):
                filas.append(fila)
                total_leidas += 1
                if max_rows and total_leidas >= max_rows:
                    break
    else:
        sample = archivo.read(4096)
        archivo.seek(0)
        try:
            sample_text = sample.decode('utf-8-sig')
        except UnicodeDecodeError:
            sample_text = sample.decode('latin-1', errors='replace')

        if not sample_text.strip():
            return None, 'El archivo está vacío'

        try:
            dialect = csv.Sniffer().sniff(sample_text, delimiters=";,|\t,")
        except csv.Error:
            dialect = csv.excel

        text_stream = io.TextIOWrapper(archivo, encoding='utf-8-sig', errors='replace')
        reader = csv.reader(text_stream, dialect)
        for fila in reader:
            if any(c.strip() for c in fila):
                filas.append(fila)
                total_leidas += 1
                if max_rows and total_leidas >= max_rows:
                    break

    if not filas:
        return None, 'No se encontraron filas válidas'

    return filas, None


def _detectar_indices_conteo(filas):
    encabezado = [str(c).strip().lower() for c in filas[0]]
    tiene_encabezado = any('sku' in c or 'codigo' in c or 'cantidad' in c or 'stock' in c for c in encabezado)

    sku_idx = 0
    cantidad_idx = 1
    if tiene_encabezado:
        mapa = {c: i for i, c in enumerate(encabezado)}
        for key in ['sku', 'codigo', 'codigo_barra', 'codigo_barras', 'barra', 'barcode']:
            if key in mapa:
                sku_idx = mapa[key]
                break
        for key in ['cantidad', 'conteo', 'stock', 'stock_fisico', 'cant', 'qty']:
            if key in mapa:
                cantidad_idx = mapa[key]
                break

    return sku_idx, cantidad_idx, tiene_encabezado


def _extraer_preview_conteo(filas, sku_idx, cantidad_idx, limite=20):
    preview = []
    errores = []

    for fila in filas:
        if len(preview) >= limite:
            break
        if len(fila) <= max(sku_idx, cantidad_idx):
            continue
        sku = str(fila[sku_idx]).strip()
        cantidad_raw = str(fila[cantidad_idx]).strip()
        if not sku:
            continue
        try:
            cantidad = int(float(cantidad_raw.replace(',', '.')))
            preview.append({'sku': sku, 'cantidad': cantidad, 'valido': True})
        except ValueError:
            preview.append({'sku': sku, 'cantidad': cantidad_raw, 'valido': False})
            errores.append(f"Cantidad inválida para SKU {sku}: {cantidad_raw}")

    return preview, errores


@require_POST
@login_required
@transaction.atomic
def importar_conteo_pistola(request, inventario_id):
    """
    Importa conteos desde archivo de pistola (CSV/TXT/XLSX).
    Formato esperado: sku,cantidad (con o sin encabezados).
    """
    inventario = _inventario_del_usuario(request, inventario_id)
    if inventario is None:
        return _error_sin_acceso()

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Debe adjuntar un archivo'})

    # Sin esta guarda se podían importar conteos sobre un inventario ya APROBADO o
    # COMPLETADO, cambiando las diferencias después de la aprobación.
    if inventario.estado not in ['BORRADOR', 'EN_CONTEO']:
        return JsonResponse({
            'success': False,
            'error': f'El inventario está en estado {inventario.get_estado_display()} y ya no admite conteos'
        })

    try:
        # Actualizar estado si es el primer conteo
        if inventario.estado == 'BORRADOR':
            inventario.estado = 'EN_CONTEO'
            inventario.fecha_inicio_conteo = inventario.fecha_inicio_conteo or timezone.now()
            inventario.save()

        nombre_hoja = request.POST.get('nombre_hoja', '').strip()

        filas, error = _leer_archivo_conteo(archivo, nombre_hoja)
        if error:
            return JsonResponse({'success': False, 'error': error})

        sku_idx, cantidad_idx, tiene_encabezado = _detectar_indices_conteo(filas)
        if tiene_encabezado:
            filas = filas[1:]

        conteos_por_sku = {}
        errores = []
        for fila in filas:
            if len(fila) <= max(sku_idx, cantidad_idx):
                continue
            sku = str(fila[sku_idx]).strip()
            cantidad_raw = str(fila[cantidad_idx]).strip()
            if not sku:
                continue
            try:
                cantidad = int(float(cantidad_raw.replace(',', '.')))
            except ValueError:
                errores.append(f"Cantidad inválida para SKU {sku}: {cantidad_raw}")
                continue
            conteos_por_sku[sku] = conteos_por_sku.get(sku, 0) + cantidad

        if not conteos_por_sku:
            return JsonResponse({'success': False, 'error': 'No se encontraron conteos válidos'})

        detalles = inventario.detalles.filter(sku__in=conteos_por_sku.keys()).select_related('producto_talla')
        detalles_por_sku = {d.sku: d for d in detalles}

        fecha_conteo = timezone.now()
        movimientos_map = _obtener_movimientos_post_corte_batch(
            [d.producto_talla_id for d in detalles],
            inventario.fecha_corte,
            fecha_conteo,
            inventario.sucursal_id
        )

        actualizados = 0
        no_encontrados = []
        for sku, cantidad in conteos_por_sku.items():
            detalle = detalles_por_sku.get(sku)
            if not detalle:
                no_encontrados.append(sku)
                continue

            movimientos_post_corte = movimientos_map.get(detalle.producto_talla_id, 0)
            detalle.stock_movimientos_post_corte = movimientos_post_corte
            detalle.stock_sistema_ajustado = detalle.stock_sistema + movimientos_post_corte
            detalle.stock_fisico = cantidad
            detalle.contado = True
            detalle.fecha_conteo = fecha_conteo
            detalle.usuario_conteo = request.user
            detalle.save()
            actualizados += 1

        inventario.calcular_metricas()

        _registrar_log(
            inventario=inventario,
            tipo_accion='REGISTRO_CONTEO',
            descripcion=f'Importación de pistola: {actualizados} productos actualizados',
            usuario=request.user,
            datos={'actualizados': actualizados, 'no_encontrados': no_encontrados, 'errores': errores}
        )

        return JsonResponse({
            'success': True,
            'actualizados': actualizados,
            'no_encontrados': no_encontrados,
            'errores': errores if errores else None,
            'progreso': float(inventario.progreso_conteo)
        })
    except Exception as e:
        logger.error(f"Error al importar conteo: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
@transaction.atomic
def actualizar_exclusion_detalle(request, inventario_id, detalle_id):
    """
    Marca un detalle como excluido/incluido del análisis.
    """
    inventario = _inventario_del_usuario(request, inventario_id)
    if inventario is None:
        return _error_sin_acceso()
    # Excluir/incluir después de aprobar cambia lo que se ajusta al stock
    if inventario.estado not in ['BORRADOR', 'EN_CONTEO', 'CONTEO_FINALIZADO', 'EN_REVISION']:
        return JsonResponse({
            'success': False,
            'error': f'El inventario está en estado {inventario.get_estado_display()}: '
                     f'ya no se pueden cambiar exclusiones'
        })

    try:
        data = json.loads(request.body)
        excluir = bool(data.get('excluir'))
        detalle = inventario.detalles.get(id=detalle_id)
        detalle.excluir_de_analisis = excluir
        detalle.save()
        inventario.calcular_metricas()

        _registrar_log(
            inventario=inventario,
            tipo_accion='MODIFICACION',
            descripcion=f'Detalle {detalle.sku} {"excluido" if excluir else "incluido"} del análisis',
            usuario=request.user,
            datos={'detalle_id': detalle_id, 'excluir': excluir}
        )

        return JsonResponse({
            'success': True,
            'excluido': detalle.excluir_de_analisis,
            'progreso': float(inventario.progreso_conteo)
        })
    except TomaInventarioDetalle.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Detalle no encontrado'})
    except Exception as e:
        logger.error(f"Error al excluir detalle: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
def preview_conteo_pistola(request, inventario_id):
    """
    Previsualiza los primeros registros del archivo de conteo.
    """
    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Debe adjuntar un archivo'})

    nombre_hoja = request.POST.get('nombre_hoja', '').strip()
    filas, error = _leer_archivo_conteo(archivo, nombre_hoja, max_rows=101, read_only=True)
    if error:
        return JsonResponse({'success': False, 'error': error})

    sku_idx, cantidad_idx, tiene_encabezado = _detectar_indices_conteo(filas)
    filas_data = filas[1:] if tiene_encabezado else filas
    preview, errores = _extraer_preview_conteo(filas_data, sku_idx, cantidad_idx, limite=100)

    return JsonResponse({
        'success': True,
        'preview': preview,
        'errores': errores[:10],
        'total_filas': len(filas_data)
    })


@require_POST
@login_required
@transaction.atomic
def registrar_reconteo(request, inventario_id):
    """
    Registrar reconteo de productos con diferencias significativas.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado not in ['EN_CONTEO', 'CONTEO_FINALIZADO', 'EN_REVISION']:
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está en un estado válido para reconteo'
            })
        
        data = json.loads(request.body)
        reconteos = data.get('reconteos', [])
        
        reconteos_realizados = 0
        
        for rec in reconteos:
            detalle_id = rec.get('detalle_id')
            stock_reconteo = rec.get('stock_reconteo')
            observaciones = rec.get('observaciones', '')
            
            try:
                detalle = inventario.detalles.get(id=detalle_id, reconteo_requerido=True)
                
                detalle.stock_reconteo = int(stock_reconteo)
                detalle.fecha_reconteo = timezone.now()
                detalle.usuario_reconteo = request.user
                
                # Si el reconteo confirma el conteo original, usar ese valor
                # Si es diferente, usar el reconteo
                if detalle.stock_reconteo != detalle.stock_fisico:
                    detalle.stock_fisico = detalle.stock_reconteo
                    base_stock = detalle.stock_sistema_ajustado if detalle.stock_sistema_ajustado is not None else detalle.stock_sistema
                    detalle.diferencia = detalle.stock_fisico - base_stock
                    if observaciones:
                        detalle.observaciones = f"{detalle.observaciones or ''}\nReconteo: {observaciones}".strip()
                
                detalle.reconteo_requerido = False
                detalle.save()
                
                reconteos_realizados += 1
                
            except TomaInventarioDetalle.DoesNotExist:
                pass
        
        # Recalcular métricas
        inventario.calcular_metricas()
        
        # Registrar log
        _registrar_log(
            inventario=inventario,
            tipo_accion='RECONTEO',
            descripcion=f'{reconteos_realizados} productos recontados',
            usuario=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{reconteos_realizados} reconteos registrados',
            'reconteos_realizados': reconteos_realizados
        })
        
    except Exception as e:
        logger.error(f"Error al registrar reconteo: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================================================================
# API: ANÁLISIS Y REPORTES
# ==============================================================================

@require_GET
@login_required
def obtener_analisis_inventario(request, inventario_id):
    """
    Obtener análisis completo del inventario antes de aprobar.
    Incluye métricas, tendencias y alertas.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        # Obtener detalles contados (solo los considerados en análisis)
        detalles = inventario.detalles.filter(contado=True, excluir_de_analisis=False)
        
        # === ANÁLISIS DE DIFERENCIAS ===
        diferencias_positivas = detalles.filter(diferencia__gt=0)
        diferencias_negativas = detalles.filter(diferencia__lt=0)
        sin_diferencia = detalles.filter(diferencia=0)
        
        # Top 10 mayores faltantes
        top_faltantes = diferencias_negativas.order_by('diferencia')[:10]
        top_faltantes_data = [
            {
                'sku': d.sku,
                'producto': d.producto_nombre,
                'talla': d.talla_nombre,
                'diferencia': d.diferencia,
                'valor': float(d.valor_diferencia),
                'porcentaje': round(d.porcentaje_diferencia, 2)
            }
            for d in top_faltantes
        ]
        
        # Top 10 mayores sobrantes
        top_sobrantes = diferencias_positivas.order_by('-diferencia')[:10]
        top_sobrantes_data = [
            {
                'sku': d.sku,
                'producto': d.producto_nombre,
                'talla': d.talla_nombre,
                'diferencia': d.diferencia,
                'valor': float(d.valor_diferencia),
                'porcentaje': round(d.porcentaje_diferencia, 2)
            }
            for d in top_sobrantes
        ]
        
        # === ANÁLISIS POR CATEGORÍA ===
        analisis_categorias = detalles.values('categoria_nombre').annotate(
            total_productos=Count('id'),
            productos_con_diferencia=Count('id', filter=~Q(diferencia=0)),
            suma_diferencias=Sum('diferencia'),
            valor_diferencias=Sum(F('diferencia') * F('costo_unitario_sistema'))
        ).order_by('-valor_diferencias')
        
        # === ANÁLISIS POR MARCA ===
        analisis_marcas = detalles.values('marca_nombre').annotate(
            total_productos=Count('id'),
            productos_con_diferencia=Count('id', filter=~Q(diferencia=0)),
            suma_diferencias=Sum('diferencia'),
            valor_diferencias=Sum(F('diferencia') * F('costo_unitario_sistema'))
        ).order_by('-valor_diferencias')
        
        # === PRODUCTOS QUE REQUIEREN RECONTEO ===
        requieren_reconteo = inventario.detalles.filter(
            reconteo_requerido=True,
            stock_reconteo__isnull=True,
            excluir_de_analisis=False
        ).count()
        
        # === INDICADORES DE PRECISIÓN ===
        total_contados = detalles.count()
        precision_inventario = (sin_diferencia.count() / total_contados * 100) if total_contados > 0 else 0

        # === UNIDADES (panel de comparación físico vs sistema) ===
        detalles_analisis = inventario.detalles.filter(excluir_de_analisis=False)
        unidades = detalles.aggregate(
            fisicas=Coalesce(Sum('stock_fisico'), 0),
            sistema=Coalesce(Sum('stock_sistema_ajustado'), 0),
        )
        total_lineas = detalles_analisis.count()
        pendientes_contar = detalles_analisis.filter(contado=False).count()

        # === SEGMENTOS (para los filtros del detalle) ===
        segmentos_marcas = list(
            detalles_analisis.values('marca_nombre').annotate(n=Count('id')).order_by('-n')[:40]
        )
        segmentos_categorias = list(
            detalles_analisis.values('categoria_nombre').annotate(n=Count('id')).order_by('-n')[:40]
        )

        # === RESUMEN FINANCIERO ===
        resumen_financiero = {
            'valor_inventario_sistema': float(inventario.valor_inventario_sistema),
            'valor_inventario_fisico': float(inventario.valor_inventario_fisico),
            'diferencia_total': float(inventario.valor_inventario_fisico - inventario.valor_inventario_sistema),
            'valor_faltantes': float(inventario.valor_diferencias_negativas),
            'valor_sobrantes': float(inventario.valor_diferencias_positivas),
            'impacto_neto': float(inventario.valor_diferencias_positivas - inventario.valor_diferencias_negativas)
        }
        
        # === ALERTAS ===
        alertas = []
        
        if requieren_reconteo > 0:
            alertas.append({
                'tipo': 'warning',
                'mensaje': f'{requieren_reconteo} productos requieren reconteo por diferencias significativas'
            })
        
        if precision_inventario < 90:
            alertas.append({
                'tipo': 'error',
                'mensaje': f'Precisión del inventario ({precision_inventario:.1f}%) está por debajo del 90% recomendado'
            })
        
        if abs(resumen_financiero['impacto_neto']) > 1000000:  # > 1 millón
            alertas.append({
                'tipo': 'warning',
                'mensaje': f'Impacto financiero significativo: ${abs(resumen_financiero["impacto_neto"]):,.0f}'
            })
        
        if inventario.total_productos_contados < inventario.total_productos_esperados:
            faltantes = inventario.total_productos_esperados - inventario.total_productos_contados
            alertas.append({
                'tipo': 'info',
                'mensaje': f'{faltantes} productos aún no han sido contados'
            })
        
        analisis = {
            'resumen': {
                # OJO: total_contados son LÍNEAS/SKUs contados. Las unidades van aparte.
                'total_esperados': total_lineas,
                'total_contados': total_contados,
                'pendientes_contar': pendientes_contar,
                'unidades_fisicas': unidades['fisicas'] or 0,
                'unidades_sistema': unidades['sistema'] or 0,
                'progreso': float(inventario.progreso_conteo),
                'excluidos': inventario.detalles.filter(excluir_de_analisis=True).count(),
                'con_diferencia': diferencias_positivas.count() + diferencias_negativas.count(),
                'sin_diferencia': sin_diferencia.count(),
                'sobrantes': diferencias_positivas.count(),
                'faltantes': diferencias_negativas.count(),
                'requieren_reconteo': requieren_reconteo,
                'precision_inventario': round(precision_inventario, 2)
            },
            'resumen_financiero': resumen_financiero,
            'top_faltantes': top_faltantes_data,
            'top_sobrantes': top_sobrantes_data,
            'analisis_categorias': [
                {
                    'categoria': a['categoria_nombre'] or 'Sin categoría',
                    'total_productos': a['total_productos'],
                    'productos_con_diferencia': a['productos_con_diferencia'],
                    'suma_diferencias': a['suma_diferencias'] or 0,
                    'valor_diferencias': float(a['valor_diferencias'] or 0)
                }
                for a in analisis_categorias
            ],
            'analisis_marcas': [
                {
                    'marca': a['marca_nombre'] or 'Sin marca',
                    'total_productos': a['total_productos'],
                    'productos_con_diferencia': a['productos_con_diferencia'],
                    'suma_diferencias': a['suma_diferencias'] or 0,
                    'valor_diferencias': float(a['valor_diferencias'] or 0)
                }
                for a in analisis_marcas
            ],
            'alertas': alertas,
            'segmentos': {
                'marcas': [
                    {'valor': s['marca_nombre'] or 'Sin marca', 'total': s['n']}
                    for s in segmentos_marcas if s['marca_nombre']
                ],
                'categorias': [
                    {'valor': s['categoria_nombre'] or 'Sin categoría', 'total': s['n']}
                    for s in segmentos_categorias if s['categoria_nombre']
                ],
            },
            'estado': inventario.estado,
            'estado_display': inventario.get_estado_display(),
            # puede_aprobar() del modelo compara unidades contra líneas; se sustituye por
            # la condición real: nada pendiente de contar ni de recontar.
            'puede_aprobar': (
                inventario.estado == 'PENDIENTE_APROBACION' and
                requieren_reconteo == 0 and
                pendientes_contar == 0
            ),
            'puede_enviar_aprobacion': (
                inventario.estado in ('CONTEO_FINALIZADO', 'EN_REVISION') and
                requieren_reconteo == 0
            ),
            'puede_aplicar_ajustes': inventario.estado == 'APROBADO',
        }
        
        return JsonResponse({
            'success': True,
            'analisis': analisis
        })
        
    except Exception as e:
        logger.error(f"Error al obtener análisis: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def exportar_inventario(request, inventario_id):
    """
    Exportar inventario a Excel.

    Se genera en modo `write_only`: openpyxl va escribiendo las filas al vuelo
    en vez de mantener todas las celdas en memoria. Las tomas 4 y 5 de
    producción tienen 335.216 líneas cada una; con el Workbook normal esa
    exportación mataba al worker antes de responder.
    """
    try:
        import openpyxl
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()

        # Crear workbook en modo streaming (sin hoja activa por defecto)
        wb = openpyxl.Workbook(write_only=True)

        # === HOJA DE RESUMEN ===
        ws_resumen = wb.create_sheet("Resumen")

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        
        # Información del inventario
        ws_resumen.append(['TOMA DE INVENTARIO'])
        ws_resumen.append(['Número:', inventario.numero_inventario])
        ws_resumen.append(['Nombre:', inventario.nombre])
        ws_resumen.append(['Sucursal:', inventario.sucursal.alias])
        ws_resumen.append(['Fecha Corte:', inventario.fecha_corte.strftime('%d/%m/%Y %H:%M')])
        ws_resumen.append(['Estado:', inventario.get_estado_display()])
        ws_resumen.append([])
        ws_resumen.append(['MÉTRICAS'])
        ws_resumen.append(['Total Productos:', inventario.total_productos_esperados])
        ws_resumen.append(['Productos Contados:', inventario.total_productos_contados])
        ws_resumen.append(['Progreso:', f'{inventario.progreso_conteo}%'])
        ws_resumen.append(['Diferencias Positivas:', inventario.total_diferencias_positivas])
        ws_resumen.append(['Diferencias Negativas:', inventario.total_diferencias_negativas])
        ws_resumen.append(['Valor Sistema:', f'${inventario.valor_inventario_sistema:,.0f}'])
        ws_resumen.append(['Valor Físico:', f'${inventario.valor_inventario_fisico:,.0f}'])
        
        # === HOJA DE DETALLE ===
        ws_detalle = wb.create_sheet("Detalle")
        
        headers = [
            'SKU', 'Producto', 'Talla', 'Marca', 'Categoría',
            'Stock Sistema', 'Mov. Post Corte', 'Stock Ajustado', 'Stock Físico', 'Diferencia', '% Diferencia',
            'Costo Unit.', 'Valor Diferencia', 'Contado', 'Reconteo Req.',
            'Ubicación', 'Observaciones', 'Excluido'
        ]
        
        # Anchos de columna: en modo write_only hay que fijarlos ANTES de
        # escribir filas, porque después la hoja ya no es direccionable.
        for col_num in range(1, len(headers) + 1):
            ws_detalle.column_dimensions[get_column_letter(col_num)].width = 15

        # Encabezados con estilo: en write_only no se puede volver a la celda
        # con ws.cell(row=1, ...), así que el estilo va en la propia celda.
        celdas_encabezado = []
        for header in headers:
            celda = WriteOnlyCell(ws_detalle, value=header)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')
            celdas_encabezado.append(celda)
        ws_detalle.append(celdas_encabezado)

        # Agregar datos. `.iterator(chunk_size=2000)` evita materializar los
        # 335.216 detalles en una lista de Python.
        detalles_qs = (
            inventario.detalles.all()
            .order_by('producto_nombre', 'talla_nombre', 'id')
            .iterator(chunk_size=2000)
        )
        for det in detalles_qs:
            ws_detalle.append([
                det.sku,
                det.producto_nombre,
                det.talla_nombre or '',
                det.marca_nombre or '',
                det.categoria_nombre or '',
                det.stock_sistema,
                det.stock_movimientos_post_corte,
                det.stock_sistema_ajustado,
                det.stock_fisico,
                det.diferencia,
                round(det.porcentaje_diferencia, 2),
                float(det.costo_unitario_sistema),
                float(det.valor_diferencia),
                'Sí' if det.contado else 'No',
                'Sí' if det.reconteo_requerido else 'No',
                det.ubicacion or '',
                det.observaciones or '',
                'Sí' if det.excluir_de_analisis else 'No'
            ])

        # (los anchos de columna ya se fijaron antes de escribir las filas:
        #  en modo write_only no se pueden tocar después)

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventario_{inventario.numero_inventario}.xlsx"'

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error al exportar inventario: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def exportar_diferencias_inventario(request, inventario_id):
    """
    Exporta solo productos con diferencias (excluidos fuera).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        detalles = inventario.detalles.filter(
            contado=True,
            excluir_de_analisis=False
        ).exclude(diferencia=0).order_by('producto_nombre', 'talla_nombre')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diferencias"

        headers = [
            'SKU', 'Producto', 'Talla', 'Marca', 'Categoría',
            'Stock Sistema', 'Mov. Post Corte', 'Stock Ajustado', 'Stock Físico', 'Diferencia', '% Diferencia',
            'Costo Unit.', 'Valor Diferencia'
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for det in detalles:
            ws.append([
                det.sku,
                det.producto_nombre,
                det.talla_nombre or '',
                det.marca_nombre or '',
                det.categoria_nombre or '',
                det.stock_sistema,
                det.stock_movimientos_post_corte,
                det.stock_sistema_ajustado,
                det.stock_fisico,
                det.diferencia,
                round(det.porcentaje_diferencia, 2),
                float(det.costo_unitario_sistema),
                float(det.valor_diferencia),
            ])

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 16

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventario_{inventario.numero_inventario}_diferencias.xlsx"'

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error al exportar diferencias: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================================================================
# API: FLUJO DE APROBACIÓN
# ==============================================================================

@require_POST
@login_required
@transaction.atomic
def finalizar_conteo(request, inventario_id):
    """
    Finalizar el conteo y pasar a revisión.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado not in ['EN_CONTEO']:
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está en estado de conteo'
            })
        
        # Verificar que se hayan contado todas las LÍNEAS (no unidades: el campo
        # total_productos_contados del modelo acumula stock_fisico, así que una tienda
        # con más unidades que SKUs podía cerrar un conteo a medias).
        detalles_analisis = inventario.detalles.filter(excluir_de_analisis=False)
        pendientes = detalles_analisis.filter(contado=False).count()
        if pendientes > 0:
            return JsonResponse({
                'success': False,
                'error': f'Faltan {pendientes} productos por contar'
            })

        # Verificar si hay reconteos pendientes
        reconteos_pendientes = inventario.detalles.filter(
            reconteo_requerido=True,
            stock_reconteo__isnull=True
        ).count()
        
        if reconteos_pendientes > 0:
            inventario.estado = 'EN_REVISION'
            mensaje = f'Inventario en revisión. {reconteos_pendientes} productos requieren reconteo.'
        else:
            inventario.estado = 'CONTEO_FINALIZADO'
            mensaje = 'Conteo finalizado exitosamente.'
        
        inventario.fecha_fin_conteo = timezone.now()
        inventario.save()
        
        _registrar_log(
            inventario=inventario,
            tipo_accion='CAMBIO_ESTADO',
            descripcion=f'Estado cambiado a {inventario.get_estado_display()}',
            usuario=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'estado': inventario.estado,
            'reconteos_pendientes': reconteos_pendientes
        })
        
    except Exception as e:
        logger.error(f"Error al finalizar conteo: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
@transaction.atomic
def enviar_aprobacion(request, inventario_id):
    """
    Enviar inventario para aprobación.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado not in ['CONTEO_FINALIZADO', 'EN_REVISION']:
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está en un estado válido para enviar a aprobación'
            })
        
        # Verificar que no haya reconteos pendientes
        reconteos_pendientes = inventario.detalles.filter(
            reconteo_requerido=True,
            stock_reconteo__isnull=True
        ).count()
        
        if reconteos_pendientes > 0:
            return JsonResponse({
                'success': False,
                'error': f'Hay {reconteos_pendientes} productos pendientes de reconteo'
            })
        
        inventario.estado = 'PENDIENTE_APROBACION'
        inventario.save()
        
        _registrar_log(
            inventario=inventario,
            tipo_accion='ENVIO_APROBACION',
            descripcion='Inventario enviado para aprobación',
            usuario=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Inventario enviado para aprobación'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar a aprobación: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
@transaction.atomic
def aprobar_inventario(request, inventario_id):
    """
    Aprobar inventario. Solo actualiza el estado, no aplica ajustes.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado != 'PENDIENTE_APROBACION':
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está pendiente de aprobación'
            })
        
        # No se aprueba un conteo incompleto: si quedan líneas sin contar el ajuste
        # posterior no toca ese stock y el inventario se cierra en falso.
        pendientes = inventario.detalles.filter(excluir_de_analisis=False, contado=False).count()
        if pendientes > 0:
            return JsonResponse({
                'success': False,
                'error': f'No se puede aprobar: quedan {pendientes} productos sin contar. '
                         f'Cuéntelos o exclúyalos del análisis.'
            })

        reconteos_pendientes = inventario.detalles.filter(
            reconteo_requerido=True, stock_reconteo__isnull=True, excluir_de_analisis=False
        ).count()
        if reconteos_pendientes > 0:
            return JsonResponse({
                'success': False,
                'error': f'No se puede aprobar: {reconteos_pendientes} productos esperan reconteo'
            })

        data = json.loads(request.body) if request.body else {}
        observaciones = data.get('observaciones', '')

        inventario.estado = 'APROBADO'
        inventario.aprobado_por = request.user
        inventario.fecha_aprobacion = timezone.now()
        if observaciones:
            inventario.observaciones = f"{inventario.observaciones or ''}\nAprobación: {observaciones}".strip()
        inventario.save()
        
        _registrar_log(
            inventario=inventario,
            tipo_accion='APROBACION',
            descripcion='Inventario aprobado',
            usuario=request.user,
            datos={'observaciones': observaciones}
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Inventario aprobado. Puede proceder a aplicar los ajustes.'
        })
        
    except Exception as e:
        logger.error(f"Error al aprobar inventario: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
@login_required
@transaction.atomic
def rechazar_inventario(request, inventario_id):
    """
    Rechazar inventario y devolverlo a conteo.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado != 'PENDIENTE_APROBACION':
            return JsonResponse({
                'success': False, 
                'error': 'El inventario no está pendiente de aprobación'
            })
        
        data = json.loads(request.body)
        motivo = data.get('motivo', '')
        
        if not motivo:
            return JsonResponse({
                'success': False,
                'error': 'Debe indicar el motivo del rechazo'
            })
        
        inventario.estado = 'EN_CONTEO'
        inventario.observaciones = f"{inventario.observaciones or ''}\nRechazo: {motivo}".strip()
        inventario.save()
        
        _registrar_log(
            inventario=inventario,
            tipo_accion='RECHAZO',
            descripcion=f'Inventario rechazado: {motivo}',
            usuario=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Inventario rechazado y devuelto a conteo'
        })
        
    except Exception as e:
        logger.error(f"Error al rechazar inventario: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================================================================
# API: APLICACIÓN DE AJUSTES (BACKGROUND THREAD + PROGRESS TRACKING)
# ==============================================================================

def _ejecutar_ajustes_background(inventario_id, usuario_id):
    """
    Worker que corre en un thread separado para aplicar ajustes de inventario.
    Actualiza TareaAplicacionAjustes cada PROGRESS_UPDATE_INTERVAL SKUs para
    que el frontend pueda hacer polling del progreso.
    La conexión de BD se cierra al finalizar para evitar leaks.
    """
    PROGRESS_UPDATE_INTERVAL = 25

    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        usuario = User.objects.get(pk=usuario_id)
        inventario = TomaInventario.objects.get(pk=inventario_id)
        tarea = TareaAplicacionAjustes.objects.get(inventario=inventario)

        detalles_pendientes = list(
            inventario.detalles.filter(
                contado=True,
                ajuste_aplicado=False,
                excluir_de_analisis=False  # lo excluido del análisis NO debe ajustar stock
            ).exclude(diferencia=0)
        )

        tarea.total = len(detalles_pendientes)
        tarea.procesados = 0
        tarea.save(update_fields=['total', 'procesados'])

        if not detalles_pendientes:
            inventario.estado = 'COMPLETADO'
            inventario.save()
            tarea.estado = 'COMPLETADO'
            tarea.finalizada_en = timezone.now()
            tarea.save(update_fields=['estado', 'finalizada_en'])
            return

        ajustes_aplicados = 0
        errores = []

        for i, detalle in enumerate(detalles_pendientes):
            try:
                _aplicar_ajuste_individual(detalle, inventario, usuario)
                ajustes_aplicados += 1
            except Exception as e:
                errores.append({'sku': detalle.sku, 'error': str(e)})
                logger.error(f"Error al aplicar ajuste para {detalle.sku}: {str(e)}")

            # Actualizar progreso periódicamente
            if (i + 1) % PROGRESS_UPDATE_INTERVAL == 0:
                tarea.procesados = i + 1
                tarea.save(update_fields=['procesados'])

        # El inventario solo se cierra si TODOS los ajustes entraron. Si alguno falló,
        # queda en APROBADO para reintentar (el filtro ajuste_aplicado=False hace que
        # el reintento sea idempotente) en vez de darse por completado a medias.
        inventario.estado = 'COMPLETADO' if not errores else 'APROBADO'
        inventario.save()

        # Registrar log
        _registrar_log(
            inventario=inventario,
            tipo_accion='APLICACION_AJUSTES',
            descripcion=(
                f'{ajustes_aplicados} ajustes aplicados'
                + (f', {len(errores)} con error (inventario queda en Aprobado para reintentar)'
                   if errores else '')
            ),
            usuario=usuario,
            datos={'ajustes_aplicados': ajustes_aplicados, 'errores': errores}
        )

        tarea.procesados = ajustes_aplicados
        tarea.errores = errores
        tarea.estado = 'COMPLETADO' if not errores else 'ERROR'
        tarea.finalizada_en = timezone.now()
        tarea.save(update_fields=['procesados', 'errores', 'estado', 'finalizada_en'])

    except Exception as e:
        logger.error(f"Error crítico en background task de ajustes: {str(e)}")
        try:
            tarea = TareaAplicacionAjustes.objects.get(inventario_id=inventario_id)
            tarea.estado = 'ERROR'
            tarea.errores = [{'error': str(e)}]
            tarea.finalizada_en = timezone.now()
            tarea.save(update_fields=['estado', 'errores', 'finalizada_en'])
            inventario = TomaInventario.objects.get(pk=inventario_id)
            if inventario.estado == 'APLICANDO':
                inventario.estado = 'APROBADO'
                inventario.save()
        except Exception:
            pass
    finally:
        connection.close()


@require_POST
@login_required
def aplicar_ajustes_inventario(request, inventario_id):
    """
    Inicia la aplicación de ajustes de inventario en un thread background.
    Retorna inmediatamente con el task_id para que el frontend haga polling
    al endpoint estado_tarea_ajustes.

    IMPORTANTE: Esta función modifica el stock real del sistema.
    Solo debe ejecutarse después de la aprobación.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()

        if inventario.estado not in ('APROBADO', 'APLICANDO'):
            return JsonResponse({
                'success': False,
                'error': 'El inventario debe estar aprobado para aplicar ajustes'
            })

        # Verificar si ya hay un proceso en curso
        tarea, created = TareaAplicacionAjustes.objects.get_or_create(
            inventario=inventario,
            defaults={'creada_por': request.user}
        )

        if tarea.estado == 'EN_PROCESO':
            return JsonResponse({
                'success': True,
                'task_id': tarea.id,
                'already_running': True,
                'message': 'El proceso ya está en ejecución'
            })

        # Reiniciar tarea si es reintento
        tarea.estado = 'EN_PROCESO'
        tarea.procesados = 0
        tarea.total = 0
        tarea.errores = []
        tarea.iniciada_en = timezone.now()
        tarea.finalizada_en = None
        tarea.creada_por = request.user
        tarea.save()

        inventario.estado = 'APLICANDO'
        inventario.save()

        thread = threading.Thread(
            target=_ejecutar_ajustes_background,
            args=(inventario_id, request.user.id),
            daemon=True
        )
        thread.start()

        return JsonResponse({
            'success': True,
            'task_id': tarea.id,
            'message': 'Proceso de ajustes iniciado. Usa el endpoint de estado para monitorear el progreso.'
        })

    except Exception as e:
        logger.error(f"Error al iniciar aplicación de ajustes: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def estado_tarea_ajustes(request, inventario_id):
    """
    Endpoint de polling para monitorear el progreso de aplicación de ajustes.
    El frontend consulta este endpoint cada ~2s para actualizar la barra de progreso.
    """
    try:
        tarea = TareaAplicacionAjustes.objects.filter(inventario_id=inventario_id).first()
        if not tarea:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró tarea para este inventario'
            })

        return JsonResponse({
            'success': True,
            'estado': tarea.estado,
            'total': tarea.total,
            'procesados': tarea.procesados,
            'porcentaje': tarea.porcentaje,
            'errores': tarea.errores,
            'iniciada_en': tarea.iniciada_en.isoformat() if tarea.iniciada_en else None,
            'finalizada_en': tarea.finalizada_en.isoformat() if tarea.finalizada_en else None,
        })
    except Exception as e:
        logger.error(f"Error al obtener estado de tarea: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

def _aplicar_ajuste_individual(detalle, inventario, usuario):
    """
    Aplica el ajuste de un producto individual: kardex + stock plano + lotes FIFO.

    SOBRANTE (diferencia > 0): lote FIFO nuevo al costo del corte + movimiento
    AJUSTE_INVENTARIO_ENTRADA.
    FALTANTE (diferencia < 0): movimiento AJUSTE_INVENTARIO_SALIDA; registrar_movimiento_producto
    descuenta el stock plano y consume los lotes FIFO best-effort.

    Cambios respecto de la versión anterior (todos por bugs reales):
    - Antes el faltante iba por consumir_stock_fifo(), que LEVANTA ValidationError si no
      hay lotes suficientes. La excepción se tragaba con un logger.warning y el detalle
      igual quedaba `ajuste_aplicado=True`: el stock NO se corregía y la pantalla decía
      "ajustes aplicados exitosamente". En la toma INV-6-20260115-001, 19 de 61 faltantes
      (31%) caen en ese caso (16 sin lotes, 3 con lotes insuficientes).
    - Además ese camino grababa concepto AJUSTE_NEGATIVO y sin referencia_externa, con lo
      que el movimiento no se podía rastrear hasta la toma que lo originó.
    - Se toma lock de fila sobre el Producto_Talla: el stock se actualiza con un
      read-modify-write y el proceso corre en un thread en paralelo con las ventas del POS.
    """
    from .views import registrar_movimiento_producto
    from .views_modulo_productos import crear_lote_producto

    diferencia = detalle.diferencia
    if diferencia == 0:
        return

    referencia = inventario.numero_inventario
    observaciones = f'Ajuste inventario {referencia}'

    with transaction.atomic():
        producto_talla = (
            Producto_Talla.objects.select_for_update()
            .select_related('producto')
            .get(pk=detalle.producto_talla_id)
        )

        if diferencia > 0:
            # El lote se crea primero: si falla, no se toca el stock y el detalle
            # queda pendiente para reintentar.
            lote = crear_lote_producto(
                producto_talla=producto_talla,
                cantidad=diferencia,
                costo_unitario=detalle.costo_unitario_sistema,
                sobreprecio_unitario=0,
                precio_venta_unitario=detalle.precio_venta_sistema,
                observaciones=f'{observaciones} - Sobrante'
            )
            movimiento = registrar_movimiento_producto(
                producto_talla=producto_talla,
                concepto='AJUSTE_INVENTARIO_ENTRADA',
                cantidad=diferencia,
                responsable=usuario,
                sucursal_destino=inventario.sucursal,
                observaciones=f'{observaciones} - Sobrante',
                referencia_externa=referencia,
                crear_lote_fifo=False,  # el lote ya se creó arriba
            )
            if lote is not None and movimiento is not None:
                lote.movimiento = movimiento
                lote.save(update_fields=['movimiento'])
        else:
            # Un ajuste no puede dejar el stock en negativo: si eso pasa la diferencia
            # se calculó contra una base que ya no existe (o hubo ventas después del
            # conteo). Se rechaza y queda listado como error para revisión manual en
            # vez de dejar stock negativo circulando por el POS.
            stock_resultante = (producto_talla.stock or 0) + diferencia
            if stock_resultante < 0:
                raise ValidationError(
                    f'El ajuste dejaría el stock en {stock_resultante} '
                    f'(actual {producto_talla.stock}, diferencia {diferencia}). '
                    f'Recontar el SKU antes de aplicar.'
                )

            # Salida: actualiza stock plano SIEMPRE y consume lotes best-effort.
            registrar_movimiento_producto(
                producto_talla=producto_talla,
                concepto='AJUSTE_INVENTARIO_SALIDA',
                cantidad=-abs(diferencia),
                responsable=usuario,
                sucursal_origen=inventario.sucursal,
                observaciones=f'{observaciones} - Faltante',
                referencia_externa=referencia,
                consumir_lotes=True,
            )

        # Marcar como aplicado solo si el ajuste efectivamente se registró
        detalle.ajuste_aplicado = True
        detalle.fecha_ajuste = timezone.now()
        detalle.save(update_fields=['ajuste_aplicado', 'fecha_ajuste'])


# ==============================================================================
# API: CANCELACIÓN
# ==============================================================================

@require_POST
@login_required
@transaction.atomic
def cancelar_inventario(request, inventario_id):
    """
    Cancelar un inventario.
    Solo se puede cancelar si no se han aplicado ajustes.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        if inventario.estado == 'COMPLETADO':
            return JsonResponse({
                'success': False, 
                'error': 'No se puede cancelar un inventario completado'
            })
        
        if inventario.estado == 'APLICANDO':
            return JsonResponse({
                'success': False, 
                'error': 'No se puede cancelar mientras se aplican ajustes'
            })
        
        data = json.loads(request.body)
        motivo = data.get('motivo', '')
        
        if not motivo:
            return JsonResponse({
                'success': False,
                'error': 'Debe indicar el motivo de cancelación'
            })
        
        inventario.estado = 'CANCELADO'
        inventario.motivo_cancelacion = motivo
        inventario.save()
        
        _registrar_log(
            inventario=inventario,
            tipo_accion='CANCELACION',
            descripcion=f'Inventario cancelado: {motivo}',
            usuario=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Inventario cancelado'
        })
        
    except Exception as e:
        logger.error(f"Error al cancelar inventario: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================================================================
# UTILIDADES
# ==============================================================================

def _registrar_log(inventario, tipo_accion, descripcion, usuario, datos=None):
    """
    Registra una entrada en el log de auditoría.
    """
    TomaInventarioLog.objects.create(
        toma_inventario=inventario,
        tipo_accion=tipo_accion,
        descripcion=descripcion,
        usuario=usuario,
        datos_adicionales=datos or {}
    )


@require_GET
@login_required
def obtener_historial_inventario(request, inventario_id):
    """
    Obtener historial de cambios del inventario.
    """
    try:
        inventario = _inventario_del_usuario(request, inventario_id)
        if inventario is None:
            return _error_sin_acceso()
        
        logs = inventario.logs.select_related('usuario').order_by('-created_at')
        
        logs_data = [
            {
                'tipo_accion': log.tipo_accion,
                'tipo_accion_display': log.get_tipo_accion_display(),
                'descripcion': log.descripcion,
                'usuario': log.usuario.get_full_name() if log.usuario else 'Sistema',
                'fecha': log.created_at.strftime('%d/%m/%Y %H:%M:%S'),
                'datos': log.datos_adicionales
            }
            for log in logs
        ]
        
        return JsonResponse({
            'success': True,
            'historial': logs_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def registrar_movimiento_producto(producto_talla, concepto, cantidad, responsable, 
                                  sucursal_origen=None, sucursal_destino=None,
                                  ticket=None, observaciones=None, referencia_externa=None):
    """
    Registra un movimiento de producto.
    Wrapper para mantener compatibilidad con el sistema existente.
    """
    tipo_movimiento = 'INGRESO' if cantidad > 0 else 'EGRESO'
    
    movimiento = Movimientos_Producto.objects.create(
        ProductoTalla=producto_talla,
        concepto=concepto,
        tipo_movimiento=tipo_movimiento,
        cantidad=cantidad,
        responsable=str(responsable) if responsable else '',
        sucursal_origen=sucursal_origen,
        sucursal_destino=sucursal_destino,
        ticket=ticket,
        observaciones=observaciones or '',
        referencia_externa=referencia_externa or '',
        estado='COMPLETADO',
        fecha=timezone.localdate(),
        hora=timezone.localtime().time()
    )
    
    return movimiento
