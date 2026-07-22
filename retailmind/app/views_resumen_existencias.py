"""
Módulo de Resumen de Existencias - RetailMind
Vista para mostrar resumen de existencias agrupado por sucursal
Con soporte para inventario histórico y agrupación por categoría
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from decimal import Decimal
from datetime import datetime, date
from django.db.models import Sum, Q, F, Value, IntegerField
from django.db.models.functions import Coalesce
from django.utils import timezone
import json
import logging

from .models import Sucursal, Producto, Producto_Talla, Movimientos_Producto, Categoria, EmpresaUser

logger = logging.getLogger('app')


@login_required
def ver_resumen_existencias(request):
    """Vista principal del reporte de resumen de existencias"""
    return render(request, 'vistas/modulo_reportes/resumen_existencias.html')


def calcular_stock_historico(talla, sucursal_id, fecha_corte):
    """
    Calcula el stock de una talla en una sucursal a una fecha específica.

    Fórmula: stock_en_fecha = stock_actual - ingresos_después + abs(egresos_después)

    Los movimientos posteriores a fecha_corte se revierten:
    - Ingresos ocurridos después → se restan (esos pares aún no habían llegado)
    - Egresos ocurridos después → se suman (esos pares todavía estaban en esa fecha)
      Las cantidades de egresos se almacenan como negativas, de ahí el abs().

    Args:
        talla: Instancia de Producto_Talla
        sucursal_id: ID de la sucursal
        fecha_corte: Fecha hasta la cual calcular (date object)

    Returns:
        int: Stock calculado a esa fecha
    """
    # Stock actual
    stock_actual = talla.stock
    
    # Si es la fecha de hoy, retornar stock actual
    if fecha_corte >= timezone.localdate():
        return max(0, stock_actual)
    
    # Calcular movimientos DESPUÉS de la fecha de corte
    # Ingresos después de la fecha (los restamos del stock actual)
    ingresos_posteriores = Movimientos_Producto.objects.filter(
        ProductoTalla=talla,
        fecha__gt=fecha_corte,
        estado='COMPLETADO'
    ).filter(
        Q(sucursal_destino_id=sucursal_id) & 
        (Q(tipo_movimiento='INGRESO') | Q(concepto='TRASPASO_ENTRADA'))
    ).aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Egresos después de la fecha (los sumamos al stock actual, porque egresos son negativos)
    egresos_posteriores = Movimientos_Producto.objects.filter(
        ProductoTalla=talla,
        fecha__gt=fecha_corte,
        estado='COMPLETADO'
    ).filter(
        Q(sucursal_origen_id=sucursal_id) &
        (Q(tipo_movimiento='EGRESO') | Q(concepto='TRASPASO_SALIDA'))
    ).aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Stock histórico = Stock actual - ingresos posteriores + egresos posteriores
    # Los ingresos después de la fecha aumentaron el stock actual → los restamos
    # Los egresos después de la fecha disminuyeron el stock actual → los sumamos de vuelta
    # cantidad de egresos se almacena como negativa, por eso usamos abs()
    stock_historico = stock_actual - ingresos_posteriores + abs(egresos_posteriores)
    
    return max(0, stock_historico)


def _aggregar_por_sucursal_actual(empresas_usuario, marca_id=None, categoria_id=None, excluir_ids=None):
    """
    Agregación SQL optimizada del inventario ACTUAL agrupado por sucursal.
    Ejecuta UNA sola query y devuelve un dict {sucursal_id: {pares, costo, p_interno, p_venta}}.
    """
    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto__excluir_de_analitica=False,
        stock__gt=0,
    )
    if marca_id:
        talla_qs = talla_qs.filter(producto__atributo1_id=marca_id)
    if categoria_id:
        talla_qs = talla_qs.filter(producto__categoria_id=categoria_id)
    if excluir_ids:
        talla_qs = talla_qs.exclude(producto_id__in=excluir_ids)

    zero_int = Value(0, output_field=IntegerField())
    agg = (
        talla_qs
        .values('producto__sucursal_id')
        .annotate(
            total_pares=Coalesce(Sum('stock'), zero_int),
            total_costo=Coalesce(Sum(F('producto__costo') * F('stock')), zero_int),
            total_p_interno=Coalesce(Sum((F('producto__costo') + F('producto__sobreprecio')) * F('stock')), zero_int),
            total_p_venta=Coalesce(Sum(F('producto__precioventa') * F('stock')), zero_int),
        )
    )

    return {
        row['producto__sucursal_id']: {
            'pares': int(row['total_pares'] or 0),
            'costo': int(row['total_costo'] or 0),
            'p_interno': int(row['total_p_interno'] or 0),
            'p_venta': int(row['total_p_venta'] or 0),
        }
        for row in agg
    }


def _aggregar_por_sucursal_historico(empresas_usuario, fecha_corte, marca_id=None, categoria_id=None, excluir_ids=None):
    """
    Agregación optimizada del inventario HISTÓRICO agrupado por sucursal.
    Ejecuta 3 queries (tallas + ingresos + egresos) y arma todo en Python con dicts.
    """
    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto__excluir_de_analitica=False,
    )
    if marca_id:
        talla_qs = talla_qs.filter(producto__atributo1_id=marca_id)
    if categoria_id:
        talla_qs = talla_qs.filter(producto__categoria_id=categoria_id)
    if excluir_ids:
        talla_qs = talla_qs.exclude(producto_id__in=excluir_ids)

    # 1) Tallas con datos del producto necesarios
    tallas_data = list(
        talla_qs.values(
            'id', 'stock',
            'producto__sucursal_id',
            'producto__costo', 'producto__sobreprecio', 'producto__precioventa',
        )
    )
    if not tallas_data:
        return {}

    talla_ids = [t['id'] for t in tallas_data]

    # 2) Ingresos posteriores agrupados por (talla, sucursal_destino)
    ingresos = (
        Movimientos_Producto.objects
        .filter(
            ProductoTalla_id__in=talla_ids,
            fecha__gt=fecha_corte,
            estado='COMPLETADO',
        )
        .filter(Q(tipo_movimiento='INGRESO') | Q(concepto='TRASPASO_ENTRADA'))
        .values('ProductoTalla_id', 'sucursal_destino_id')
        .annotate(total=Sum('cantidad'))
    )
    ingresos_map = {(r['ProductoTalla_id'], r['sucursal_destino_id']): int(r['total'] or 0) for r in ingresos}

    # 3) Egresos posteriores agrupados por (talla, sucursal_origen)
    egresos = (
        Movimientos_Producto.objects
        .filter(
            ProductoTalla_id__in=talla_ids,
            fecha__gt=fecha_corte,
            estado='COMPLETADO',
        )
        .filter(Q(tipo_movimiento='EGRESO') | Q(concepto='TRASPASO_SALIDA'))
        .values('ProductoTalla_id', 'sucursal_origen_id')
        .annotate(total=Sum('cantidad'))
    )
    egresos_map = {(r['ProductoTalla_id'], r['sucursal_origen_id']): int(r['total'] or 0) for r in egresos}

    # 4) Agregar en Python por sucursal aplicando reversión de movimientos
    acum = {}
    for t in tallas_data:
        sid = t['producto__sucursal_id']
        ingresos_post = ingresos_map.get((t['id'], sid), 0)
        egresos_post = egresos_map.get((t['id'], sid), 0)
        stock_actual = t['stock'] or 0
        # stock_actual - ingresos_posteriores + |egresos_posteriores| (egresos guardan cantidad negativa)
        stock_hist = stock_actual - ingresos_post + abs(egresos_post)
        if stock_hist <= 0:
            continue

        costo = t['producto__costo'] or 0
        sobre = t['producto__sobreprecio'] or 0
        venta = t['producto__precioventa'] or 0

        bucket = acum.get(sid)
        if bucket is None:
            bucket = {'pares': 0, 'costo': 0, 'p_interno': 0, 'p_venta': 0}
            acum[sid] = bucket
        bucket['pares'] += stock_hist
        bucket['costo'] += costo * stock_hist
        bucket['p_interno'] += (costo + sobre) * stock_hist
        bucket['p_venta'] += venta * stock_hist

    return acum


def _construir_mapa_categoria_raiz():
    """
    Devuelve un dict {categoria_id: categoria_raiz_id} recorriendo padres
    hasta dos niveles (raíz, sub, sub-sub) en una sola query.
    """
    cats = list(Categoria.objects.all().values('id', 'padre_id'))
    padre_map = {c['id']: c['padre_id'] for c in cats}

    cat_to_root = {}
    for cat_id in padre_map:
        actual = cat_id
        # Limitar profundidad por seguridad
        for _ in range(5):
            padre = padre_map.get(actual)
            if padre is None:
                break
            actual = padre
        cat_to_root[cat_id] = actual
    return cat_to_root


def _aggregar_por_categoria_actual(empresas_usuario, sucursales_ids, marca_id=None, excluir_ids=None):
    """
    Agregación SQL del inventario ACTUAL por (categoria_producto, sucursal).
    El mapeo a categoría raíz se hace en Python.
    """
    # OJO: no se exige categoria__isnull=False — el stock sin categoría se
    # agrega como bucket "Sin categoría" en vez de quedar invisible.
    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto__sucursal_id__in=sucursales_ids,
        producto__excluir_de_analitica=False,
        stock__gt=0,
    )
    if marca_id:
        talla_qs = talla_qs.filter(producto__atributo1_id=marca_id)
    if excluir_ids:
        talla_qs = talla_qs.exclude(producto_id__in=excluir_ids)

    zero_int = Value(0, output_field=IntegerField())
    agg = (
        talla_qs
        .values('producto__categoria_id', 'producto__sucursal_id')
        .annotate(
            total_pares=Coalesce(Sum('stock'), zero_int),
            total_costo=Coalesce(Sum(F('producto__costo') * F('stock')), zero_int),
            total_p_interno=Coalesce(Sum((F('producto__costo') + F('producto__sobreprecio')) * F('stock')), zero_int),
            total_p_venta=Coalesce(Sum(F('producto__precioventa') * F('stock')), zero_int),
        )
    )
    return list(agg)


def _aggregar_por_categoria_historico(empresas_usuario, sucursales_ids, fecha_corte, marca_id=None, excluir_ids=None):
    """
    Variante histórica: 3 queries (tallas con categoría/sucursal, ingresos, egresos)
    y agregación en Python por (categoria, sucursal).
    Devuelve la misma estructura que _aggregar_por_categoria_actual
    (lista de dicts con producto__categoria_id, producto__sucursal_id, total_*).
    """
    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto__sucursal_id__in=sucursales_ids,
        producto__excluir_de_analitica=False,
    )
    if marca_id:
        talla_qs = talla_qs.filter(producto__atributo1_id=marca_id)
    if excluir_ids:
        talla_qs = talla_qs.exclude(producto_id__in=excluir_ids)

    tallas_data = list(
        talla_qs.values(
            'id', 'stock',
            'producto__categoria_id',
            'producto__sucursal_id',
            'producto__costo', 'producto__sobreprecio', 'producto__precioventa',
        )
    )
    if not tallas_data:
        return []

    talla_ids = [t['id'] for t in tallas_data]

    ingresos = (
        Movimientos_Producto.objects
        .filter(
            ProductoTalla_id__in=talla_ids,
            fecha__gt=fecha_corte,
            estado='COMPLETADO',
        )
        .filter(Q(tipo_movimiento='INGRESO') | Q(concepto='TRASPASO_ENTRADA'))
        .values('ProductoTalla_id', 'sucursal_destino_id')
        .annotate(total=Sum('cantidad'))
    )
    ingresos_map = {(r['ProductoTalla_id'], r['sucursal_destino_id']): int(r['total'] or 0) for r in ingresos}

    egresos = (
        Movimientos_Producto.objects
        .filter(
            ProductoTalla_id__in=talla_ids,
            fecha__gt=fecha_corte,
            estado='COMPLETADO',
        )
        .filter(Q(tipo_movimiento='EGRESO') | Q(concepto='TRASPASO_SALIDA'))
        .values('ProductoTalla_id', 'sucursal_origen_id')
        .annotate(total=Sum('cantidad'))
    )
    egresos_map = {(r['ProductoTalla_id'], r['sucursal_origen_id']): int(r['total'] or 0) for r in egresos}

    acum = {}  # (cat_id, suc_id) -> dict
    for t in tallas_data:
        sid = t['producto__sucursal_id']
        cid = t['producto__categoria_id']  # None => bucket "Sin categoría"
        ingresos_post = ingresos_map.get((t['id'], sid), 0)
        egresos_post = egresos_map.get((t['id'], sid), 0)
        stock_hist = (t['stock'] or 0) - ingresos_post + abs(egresos_post)
        if stock_hist <= 0:
            continue

        costo = t['producto__costo'] or 0
        sobre = t['producto__sobreprecio'] or 0
        venta = t['producto__precioventa'] or 0

        key = (cid, sid)
        bucket = acum.get(key)
        if bucket is None:
            bucket = {
                'producto__categoria_id': cid,
                'producto__sucursal_id': sid,
                'total_pares': 0,
                'total_costo': 0,
                'total_p_interno': 0,
                'total_p_venta': 0,
            }
            acum[key] = bucket
        bucket['total_pares'] += stock_hist
        bucket['total_costo'] += costo * stock_hist
        bucket['total_p_interno'] += (costo + sobre) * stock_hist
        bucket['total_p_venta'] += venta * stock_hist

    return list(acum.values())


def _parse_excluir_articulos(request):
    """Parsea el parámetro `excluir_articulos` (CSV de IDs de Producto) en una lista de ints."""
    raw = request.GET.get('excluir_articulos', '') or ''
    ids = []
    for token in raw.split(','):
        token = token.strip()
        if token.isdigit():
            ids.append(int(token))
    return ids


def _calcular_totales_excluidos(empresas_usuario, excluir_ids, fecha_corte=None, es_historico=False):
    """
    Calcula cuántas unidades (y su valor) representan los artículos excluidos,
    para que el usuario vea de forma explícita qué está dejando fuera del análisis.

    Respeta la fecha de corte histórica: si es histórico, reconstruye el stock a esa
    fecha con la misma lógica de reversión de movimientos que el resto del reporte.

    Devuelve un dict {pares, costo, precio_interno, precio_venta, articulos}.
    """
    vacio = {'pares': 0, 'costo': 0, 'precio_interno': 0, 'precio_venta': 0, 'articulos': 0}
    if not excluir_ids:
        return vacio

    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto_id__in=excluir_ids,
    )

    if es_historico and fecha_corte:
        # Reconstrucción histórica (misma fórmula que _aggregar_por_sucursal_historico)
        tallas_data = list(
            talla_qs.values(
                'id', 'stock', 'producto_id',
                'producto__sucursal_id',
                'producto__costo', 'producto__sobreprecio', 'producto__precioventa',
            )
        )
        if not tallas_data:
            return vacio

        talla_ids = [t['id'] for t in tallas_data]

        ingresos = (
            Movimientos_Producto.objects
            .filter(ProductoTalla_id__in=talla_ids, fecha__gt=fecha_corte, estado='COMPLETADO')
            .filter(Q(tipo_movimiento='INGRESO') | Q(concepto='TRASPASO_ENTRADA'))
            .values('ProductoTalla_id', 'sucursal_destino_id')
            .annotate(total=Sum('cantidad'))
        )
        ingresos_map = {(r['ProductoTalla_id'], r['sucursal_destino_id']): int(r['total'] or 0) for r in ingresos}

        egresos = (
            Movimientos_Producto.objects
            .filter(ProductoTalla_id__in=talla_ids, fecha__gt=fecha_corte, estado='COMPLETADO')
            .filter(Q(tipo_movimiento='EGRESO') | Q(concepto='TRASPASO_SALIDA'))
            .values('ProductoTalla_id', 'sucursal_origen_id')
            .annotate(total=Sum('cantidad'))
        )
        egresos_map = {(r['ProductoTalla_id'], r['sucursal_origen_id']): int(r['total'] or 0) for r in egresos}

        pares = costo = p_interno = p_venta = 0
        productos_con_stock = set()
        for t in tallas_data:
            sid = t['producto__sucursal_id']
            ingresos_post = ingresos_map.get((t['id'], sid), 0)
            egresos_post = egresos_map.get((t['id'], sid), 0)
            stock_hist = (t['stock'] or 0) - ingresos_post + abs(egresos_post)
            if stock_hist <= 0:
                continue
            c = t['producto__costo'] or 0
            s = t['producto__sobreprecio'] or 0
            v = t['producto__precioventa'] or 0
            pares += stock_hist
            costo += c * stock_hist
            p_interno += (c + s) * stock_hist
            p_venta += v * stock_hist
            productos_con_stock.add(t['producto_id'])

        return {
            'pares': int(pares),
            'costo': int(costo),
            'precio_interno': int(p_interno),
            'precio_venta': int(p_venta),
            'articulos': len(productos_con_stock),
        }

    # Actual — una sola query agregada
    zero_int = Value(0, output_field=IntegerField())
    agg = (
        talla_qs.filter(stock__gt=0)
        .aggregate(
            total_pares=Coalesce(Sum('stock'), zero_int),
            total_costo=Coalesce(Sum(F('producto__costo') * F('stock')), zero_int),
            total_p_interno=Coalesce(Sum((F('producto__costo') + F('producto__sobreprecio')) * F('stock')), zero_int),
            total_p_venta=Coalesce(Sum(F('producto__precioventa') * F('stock')), zero_int),
        )
    )
    articulos = (
        talla_qs.filter(stock__gt=0)
        .values('producto_id').distinct().count()
    )
    return {
        'pares': int(agg['total_pares'] or 0),
        'costo': int(agg['total_costo'] or 0),
        'precio_interno': int(agg['total_p_interno'] or 0),
        'precio_venta': int(agg['total_p_venta'] or 0),
        'articulos': articulos,
    }


@require_GET
@login_required
def obtener_resumen_existencias(request):
    """API para obtener resumen de existencias agrupado por sucursal"""
    try:
        # Parámetros de filtro
        marca_id = request.GET.get('marca_id')
        categoria_id = request.GET.get('categoria_id')
        fecha_corte_str = request.GET.get('fecha_corte')  # Formato: YYYY-MM-DD
        agrupar_por = request.GET.get('agrupar_por', 'sucursal')  # sucursal | categoria
        excluir_ids = _parse_excluir_articulos(request)

        # Parsear fecha de corte
        fecha_corte = None
        es_historico = False
        if fecha_corte_str:
            try:
                fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
                es_historico = fecha_corte < timezone.localdate()
            except ValueError:
                pass

        # Filtrar por empresas del usuario (seguridad multi-tenant)
        empresas_usuario = list(EmpresaUser.objects.filter(
            user=request.user,
            status=True
        ).values_list('empresa_id', flat=True))

        # Si es agrupación por categoría
        if agrupar_por == 'categoria':
            return _resumen_por_categoria(
                request, marca_id, fecha_corte, es_historico,
                empresas_usuario, excluir_ids,
            )

        # ===== Agrupación por sucursal (default) — agregación SQL optimizada =====
        if es_historico and fecha_corte:
            agg_por_suc = _aggregar_por_sucursal_historico(
                empresas_usuario, fecha_corte,
                marca_id=marca_id, categoria_id=categoria_id, excluir_ids=excluir_ids,
            )
        else:
            agg_por_suc = _aggregar_por_sucursal_actual(
                empresas_usuario,
                marca_id=marca_id, categoria_id=categoria_id, excluir_ids=excluir_ids,
            )

        # Cargar metadata de las sucursales que tienen agregados
        sucursales = (
            Sucursal.objects
            .filter(empresa_id__in=empresas_usuario)
            .select_related('empresa')
            .order_by('alias')
        )

        resumen_sucursales = []
        empresas_acum = {}

        for sucursal in sucursales:
            bucket = agg_por_suc.get(sucursal.id)
            if not bucket or bucket['pares'] <= 0:
                continue

            total_pares = int(bucket['pares'])
            total_costo = int(bucket['costo'])
            total_p_interno = int(bucket['p_interno'])
            total_p_venta = int(bucket['p_venta'])

            resumen_sucursales.append({
                'sucursal_id': sucursal.id,
                'sucursal': sucursal.alias,
                'direccion': sucursal.direccion or '-',
                'empresa_id': sucursal.empresa_id,
                'empresa': sucursal.empresa.nombre if sucursal.empresa else 'Sin Empresa',
                'total_pares': total_pares,
                'total_costo': float(total_costo),
                'total_precio_interno': float(total_p_interno),
                'total_precio_venta': float(total_p_venta),
            })

            emp_id = sucursal.empresa_id
            emp_acum = empresas_acum.get(emp_id)
            if emp_acum is None:
                emp_acum = {
                    'empresa_id': emp_id,
                    'empresa': sucursal.empresa.nombre if sucursal.empresa else 'Sin Empresa',
                    'num_sucursales': 0,
                    'total_pares': 0,
                    'total_costo': 0,
                    'total_precio_interno': 0,
                    'total_precio_venta': 0,
                }
                empresas_acum[emp_id] = emp_acum
            emp_acum['num_sucursales'] += 1
            emp_acum['total_pares'] += total_pares
            emp_acum['total_costo'] += total_costo
            emp_acum['total_precio_interno'] += total_p_interno
            emp_acum['total_precio_venta'] += total_p_venta

        # Calcular totales generales
        total_general = {
            'pares': sum(s['total_pares'] for s in resumen_sucursales),
            'costo': sum(s['total_costo'] for s in resumen_sucursales),
            'precio_interno': sum(s['total_precio_interno'] for s in resumen_sucursales),
            'precio_venta': sum(s['total_precio_venta'] for s in resumen_sucursales),
        }

        # Serializar resumen por empresa ordenado por nombre
        resumen_empresas = []
        for emp in sorted(empresas_acum.values(), key=lambda e: (e['empresa'] or '').lower()):
            resumen_empresas.append({
                'empresa_id': emp['empresa_id'],
                'empresa': emp['empresa'],
                'num_sucursales': emp['num_sucursales'],
                'total_pares': emp['total_pares'],
                'total_costo': float(emp['total_costo']),
                'total_precio_interno': float(emp['total_precio_interno']),
                'total_precio_venta': float(emp['total_precio_venta']),
            })

        # Totales de lo excluido (para que el usuario vea qué deja fuera)
        totales_excluidos = _calcular_totales_excluidos(
            empresas_usuario, excluir_ids, fecha_corte=fecha_corte, es_historico=es_historico,
        )

        fecha_info = f"al {fecha_corte.strftime('%d/%m/%Y')}" if es_historico else "actual"
        logger.info(
            "Resumen existencias generado: sucursales=%s, fecha=%s, total_pares=%s, excluidos=%s (pares_excluidos=%s)",
            len(resumen_sucursales),
            fecha_info,
            total_general['pares'],
            len(excluir_ids),
            totales_excluidos['pares'],
        )

        return JsonResponse({
            'success': True,
            'datos': resumen_sucursales,
            'total_general': total_general,
            'resumen_empresas': resumen_empresas,
            'es_historico': es_historico,
            'fecha_corte': fecha_corte_str if es_historico else None,
            'agrupar_por': 'sucursal',
            'excluir_articulos_count': len(excluir_ids),
            'totales_excluidos': totales_excluidos,
        })
        
    except Exception as e:
        logger.exception("Error en resumen de existencias")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener resumen: {str(e)}'
        })


def _resumen_por_categoria(request, marca_id, fecha_corte, es_historico, empresas_usuario, excluir_ids=None):
    """Genera resumen agrupado por categoría/departamento — versión optimizada con SQL aggregation."""
    try:
        excluir_ids = excluir_ids or []

        # Sucursales del usuario (con empresa para el resumen final)
        sucursales_qs = (
            Sucursal.objects
            .filter(empresa_id__in=empresas_usuario)
            .select_related('empresa')
        )
        sucursales_map = {s.id: s for s in sucursales_qs}
        sucursales_ids = list(sucursales_map.keys())

        # Mapeo categoría -> categoría raíz (1 query)
        cat_to_root = _construir_mapa_categoria_raiz()

        # Categorías raíz para nombres y orden estable
        categorias_raiz = list(
            Categoria.objects.filter(padre__isnull=True).order_by('nombre')
        )
        nombre_raiz = {c.id: c.nombre for c in categorias_raiz}

        # Árbol v1.2: solo las raíces CON hijas son padres reales (Calzado /
        # Ropa / Accesorios...). Las raíces sin hijas son categorías planas
        # antiguas aún vivas: se agrupan en un bucket único "Sin recategorizar"
        # en vez de aparecer mezcladas como raíces del árbol.
        padres_v12 = set(
            Categoria.objects.filter(
                padre__isnull=True, subcategorias__isnull=False,
            ).values_list('id', flat=True)
        )
        LEGACY_KEY = '_legacy'
        SINCAT_KEY = '_sincat'
        legacy_nombres = set()

        # Datos agregados por (categoria_id, sucursal_id)
        if es_historico and fecha_corte:
            agg_rows = _aggregar_por_categoria_historico(
                empresas_usuario, sucursales_ids, fecha_corte,
                marca_id=marca_id, excluir_ids=excluir_ids,
            )
        else:
            agg_rows = _aggregar_por_categoria_actual(
                empresas_usuario, sucursales_ids,
                marca_id=marca_id, excluir_ids=excluir_ids,
            )

        # Acumular por categoría raíz
        cat_acum = {}  # root_cat_id -> bucket
        empresas_acum = {}  # empresa_id -> bucket
        sucursales_por_empresa = {}  # empresa_id -> set(sucursal_id)

        for row in agg_rows:
            cat_id = row['producto__categoria_id']
            suc_id = row['producto__sucursal_id']
            if cat_id is None:
                root_id = SINCAT_KEY
            else:
                root_id = cat_to_root.get(cat_id)
                if root_id is None:
                    continue
                if root_id not in padres_v12:
                    # Plana antigua viva (o residuo _ZZ_): bucket único.
                    legacy_nombres.add(nombre_raiz.get(root_id, f'ID {root_id}'))
                    root_id = LEGACY_KEY

            pares = int(row['total_pares'] or 0)
            if pares <= 0:
                continue

            costo = int(row['total_costo'] or 0)
            p_interno = int(row['total_p_interno'] or 0)
            p_venta = int(row['total_p_venta'] or 0)

            # Por categoría raíz
            cat_b = cat_acum.get(root_id)
            if cat_b is None:
                cat_b = {
                    'pares': 0, 'costo': 0, 'p_interno': 0, 'p_venta': 0,
                    'sucursales': set(),
                }
                cat_acum[root_id] = cat_b
            cat_b['pares'] += pares
            cat_b['costo'] += costo
            cat_b['p_interno'] += p_interno
            cat_b['p_venta'] += p_venta
            cat_b['sucursales'].add(suc_id)

            # Por empresa (resumen final)
            sucursal = sucursales_map.get(suc_id)
            emp_id = sucursal.empresa_id if sucursal else None
            if emp_id is not None:
                emp_b = empresas_acum.get(emp_id)
                if emp_b is None:
                    emp_b = {
                        'empresa_id': emp_id,
                        'empresa': sucursal.empresa.nombre if sucursal and sucursal.empresa else 'Sin Empresa',
                        'pares': 0, 'costo': 0, 'p_interno': 0, 'p_venta': 0,
                    }
                    empresas_acum[emp_id] = emp_b
                    sucursales_por_empresa[emp_id] = set()
                emp_b['pares'] += pares
                emp_b['costo'] += costo
                emp_b['p_interno'] += p_interno
                emp_b['p_venta'] += p_venta
                sucursales_por_empresa[emp_id].add(suc_id)

        # Construir resumen: padres v1.2 en orden alfabético, luego el bucket
        # "Sin recategorizar" (planas antiguas) y "Sin categoría" al final.
        resumen_categorias = []
        for cat in categorias_raiz:
            if cat.id not in padres_v12:
                continue
            b = cat_acum.get(cat.id)
            if not b or b['pares'] <= 0:
                continue
            resumen_categorias.append({
                'categoria_id': cat.id,
                'categoria': cat.nombre,
                'num_sucursales': len(b['sucursales']),
                'total_pares': b['pares'],
                'total_costo': float(b['costo']),
                'total_precio_interno': float(b['p_interno']),
                'total_precio_venta': float(b['p_venta']),
            })
        for key, etiqueta in (
            (LEGACY_KEY, 'Sin recategorizar ({n} categorías antiguas)'),
            (SINCAT_KEY, 'Sin categoría'),
        ):
            b = cat_acum.get(key)
            if not b or b['pares'] <= 0:
                continue
            resumen_categorias.append({
                'categoria_id': None,
                'categoria': etiqueta.format(n=len(legacy_nombres)),
                'detalle_categorias': sorted(legacy_nombres) if key == LEGACY_KEY else [],
                'num_sucursales': len(b['sucursales']),
                'total_pares': b['pares'],
                'total_costo': float(b['costo']),
                'total_precio_interno': float(b['p_interno']),
                'total_precio_venta': float(b['p_venta']),
            })

        # Total general
        total_general = {
            'pares': sum(c['total_pares'] for c in resumen_categorias),
            'costo': sum(c['total_costo'] for c in resumen_categorias),
            'precio_interno': sum(c['total_precio_interno'] for c in resumen_categorias),
            'precio_venta': sum(c['total_precio_venta'] for c in resumen_categorias),
        }

        # Resumen por empresa
        resumen_empresas = []
        for emp in sorted(empresas_acum.values(), key=lambda e: (e['empresa'] or '').lower()):
            resumen_empresas.append({
                'empresa_id': emp['empresa_id'],
                'empresa': emp['empresa'],
                'num_sucursales': len(sucursales_por_empresa.get(emp['empresa_id'], set())),
                'total_pares': emp['pares'],
                'total_costo': float(emp['costo']),
                'total_precio_interno': float(emp['p_interno']),
                'total_precio_venta': float(emp['p_venta']),
            })

        totales_excluidos = _calcular_totales_excluidos(
            empresas_usuario, excluir_ids, fecha_corte=fecha_corte, es_historico=es_historico,
        )

        return JsonResponse({
            'success': True,
            'datos': resumen_categorias,
            'total_general': total_general,
            'resumen_empresas': resumen_empresas,
            'es_historico': es_historico,
            'fecha_corte': fecha_corte.strftime('%Y-%m-%d') if es_historico and fecha_corte else None,
            'agrupar_por': 'categoria',
            'excluir_articulos_count': len(excluir_ids),
            'totales_excluidos': totales_excluidos,
        })

    except Exception as e:
        logger.exception("Error en resumen de existencias por categoria")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener resumen por categoría: {str(e)}'
        })


@require_GET
@login_required
def exportar_resumen_existencias_excel(request):
    """Exportar resumen de existencias a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Obtener datos del reporte
        marca_id = request.GET.get('marca_id')
        categoria_id = request.GET.get('categoria_id')
        fecha_corte = request.GET.get('fecha_corte')
        agrupar_por = request.GET.get('agrupar_por', 'sucursal')
        
        # Reutilizar la función de obtención de datos
        response_data = obtener_resumen_existencias(request)
        datos = json.loads(response_data.content)
        
        if not datos.get('success'):
            return JsonResponse({
                'success': False,
                'error': 'No se pudieron obtener los datos'
            })
        
        datos_resumen = datos.get('datos', [])
        total_general = datos.get('total_general', {})
        es_historico = datos.get('es_historico', False)
        resumen_empresas = datos.get('resumen_empresas', []) or []
        excluir_count = datos.get('excluir_articulos_count', 0)
        totales_excluidos = datos.get('totales_excluidos', {}) or {}

        if not datos_resumen:
            return JsonResponse({
                'success': False,
                'error': 'No hay datos para exportar'
            })
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen Existencias"
        
        # Estilos NEXO
        header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
        total_font = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        titulo = "RESUMEN DE EXISTENCIAS"
        if agrupar_por == 'categoria':
            titulo += " POR CATEGORÍA"
        else:
            titulo += " POR SUCURSAL"
        
        if es_historico and fecha_corte:
            titulo += f" - Cierre del {fecha_corte}"

        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = titulo
        cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Encabezados dinámicos según agrupación
        if agrupar_por == 'categoria':
            headers = ['Categoría', 'Sucursales', 'Total Pares', 'Total Costo', 'Total Precio Interno', 'Total Precio Venta']
        else:
            headers = ['Sucursal', 'Dirección', 'Total Pares', 'Total Costo', 'Total Precio Interno', 'Total Precio Venta']
        
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Datos
        fila = 3
        for item in datos_resumen:
            if agrupar_por == 'categoria':
                ws.cell(row=fila, column=1, value=item['categoria']).border = border
                ws.cell(row=fila, column=2, value=item['num_sucursales']).border = border
            else:
                ws.cell(row=fila, column=1, value=item['sucursal']).border = border
                ws.cell(row=fila, column=2, value=item['direccion']).border = border
            
            cell = ws.cell(row=fila, column=3, value=item['total_pares'])
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            cell = ws.cell(row=fila, column=4, value=item['total_costo'])
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            cell = ws.cell(row=fila, column=5, value=item['total_precio_interno'])
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            cell = ws.cell(row=fila, column=6, value=item['total_precio_venta'])
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            fila += 1
        
        # Fila de TOTALES
        ws.cell(row=fila, column=1, value="TOTALES:").font = total_font
        ws.cell(row=fila, column=1).fill = total_fill
        ws.cell(row=fila, column=1).border = border
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        ws.cell(row=fila, column=2, value="").fill = total_fill
        ws.cell(row=fila, column=2).border = border
        
        cell = ws.cell(row=fila, column=3, value=total_general.get('pares', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        cell = ws.cell(row=fila, column=4, value=total_general.get('costo', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        cell = ws.cell(row=fila, column=5, value=total_general.get('precio_interno', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        cell = ws.cell(row=fila, column=6, value=total_general.get('precio_venta', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        ws.row_dimensions[fila].height = 25

        # ===== Resumen por Empresa =====
        if resumen_empresas:
            fila += 2  # espacio en blanco

            # Título de la sección
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
            cell = ws.cell(row=fila, column=1, value="RESUMEN POR EMPRESA")
            cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[fila].height = 24
            fila += 1

            # Encabezados
            empresa_headers = ['Empresa', 'Sucursales', 'Total Pares', 'Total Costo', 'Total Precio Interno', 'Total Precio Venta']
            for idx, header in enumerate(empresa_headers, start=1):
                c = ws.cell(row=fila, column=idx, value=header)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border
            ws.row_dimensions[fila].height = 22
            fila += 1

            # Filas
            for emp in resumen_empresas:
                ws.cell(row=fila, column=1, value=emp.get('empresa', '')).border = border

                c = ws.cell(row=fila, column=2, value=emp.get('num_sucursales', 0))
                c.alignment = Alignment(horizontal='right')
                c.border = border

                c = ws.cell(row=fila, column=3, value=emp.get('total_pares', 0))
                c.alignment = Alignment(horizontal='right')
                c.border = border

                c = ws.cell(row=fila, column=4, value=emp.get('total_costo', 0))
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right')
                c.border = border

                c = ws.cell(row=fila, column=5, value=emp.get('total_precio_interno', 0))
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right')
                c.border = border

                c = ws.cell(row=fila, column=6, value=emp.get('total_precio_venta', 0))
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right')
                c.border = border

                fila += 1

        # Nota de exclusiones (si aplica) — incluye unidades y valores realmente excluidos
        if excluir_count:
            fila += 1
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
            pares_exc = int(totales_excluidos.get('pares', 0) or 0)
            venta_exc = int(totales_excluidos.get('precio_venta', 0) or 0)
            costo_exc = int(totales_excluidos.get('costo', 0) or 0)
            detalle_exc = ''
            if pares_exc > 0:
                detalle_exc = (
                    f" — {pares_exc:,} unidad(es) por ${venta_exc:,} a precio venta "
                    f"(${costo_exc:,} costo)"
                ).replace(',', '.')
            nota = ws.cell(
                row=fila, column=1,
                value=(f"Nota: {excluir_count} artículo(s) excluidos del análisis "
                       f"(filtro temporal de la sesión){detalle_exc}.")
            )
            nota.font = Font(italic=True, color="8a6914")
            nota.alignment = Alignment(horizontal='left', vertical='center')

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 20

        # ===== Hoja: Top 30 por Sucursal (solo agrupación por sucursal) =====
        if agrupar_por == 'sucursal':
            _agregar_hoja_top_por_sucursal_excel(
                wb, request, datos_resumen, es_historico, fecha_corte,
                header_fill, header_font, total_fill, total_font, border,
            )

        # Nombre del archivo
        filename = "resumen_existencias"
        if agrupar_por == 'categoria':
            filename += "_por_categoria"
        if es_historico and fecha_corte:
            filename += f"_{fecha_corte}"
        filename += ".xlsx"
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Error al exportar resumen de existencias a Excel")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


def _obtener_filtros_top_export(request):
    """Extrae empresas del usuario + filtros (marca, exclusiones, fecha) para los Top-N de export."""
    marca_id = request.GET.get('marca_id')
    excluir_ids = _parse_excluir_articulos(request)

    fecha_corte = None
    es_historico = False
    fecha_corte_str = request.GET.get('fecha_corte')
    if fecha_corte_str:
        try:
            fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
            es_historico = fecha_corte < timezone.localdate()
        except ValueError:
            pass

    empresas_usuario = list(EmpresaUser.objects.filter(
        user=request.user, status=True,
    ).values_list('empresa_id', flat=True))

    return empresas_usuario, marca_id, excluir_ids, fecha_corte, es_historico


def _agregar_hoja_top_por_sucursal_excel(wb, request, datos_resumen, es_historico_reporte,
                                          fecha_corte_reporte, header_fill, header_font,
                                          total_fill, total_font, border):
    """Añade al workbook una hoja con el Top 30 de productos (por unidades) de cada sucursal."""
    from openpyxl.styles import Font, PatternFill, Alignment

    empresas_usuario, marca_id, excluir_ids, fecha_corte, es_historico = _obtener_filtros_top_export(request)

    ws = wb.create_sheet(title="Top 30 por Sucursal")
    seccion_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    seccion_font = Font(bold=True, color="FFFFFF", size=12)
    right = Alignment(horizontal='right')

    fila = 1
    hubo_datos = False

    for suc in datos_resumen:
        sucursal_id = suc.get('sucursal_id')
        if not sucursal_id:
            continue

        detalle = _calcular_detalle_top_por_sucursal(
            empresas_usuario, sucursal_id, limite=30,
            marca_id=marca_id, excluir_ids=excluir_ids,
            fecha_corte=fecha_corte, es_historico=es_historico,
        )
        items = detalle['items']
        if not items:
            continue
        hubo_datos = True

        # Título de sección (nombre sucursal)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        titulo = f"🏪 {suc.get('sucursal', 'Sucursal')}  ·  {detalle['total_productos']} productos con stock"
        c = ws.cell(row=fila, column=1, value=titulo)
        c.fill = seccion_fill
        c.font = seccion_font
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[fila].height = 22
        fila += 1

        # Encabezados
        headers = ['#', 'Artículo', 'Unidades', 'Costo', 'Precio Venta']
        for idx, h in enumerate(headers, start=1):
            hc = ws.cell(row=fila, column=idx, value=h)
            hc.fill = header_fill
            hc.font = header_font
            hc.alignment = Alignment(horizontal='center', vertical='center')
            hc.border = border
        fila += 1

        # Filas del top
        for pos, it in enumerate(items, start=1):
            ws.cell(row=fila, column=1, value=pos).border = border
            art = f"{it['articulo']}"
            if it.get('descripcion'):
                art += f" — {it['descripcion']}"
            ws.cell(row=fila, column=2, value=art).border = border

            u = ws.cell(row=fila, column=3, value=it['pares'])
            u.alignment = right
            u.border = border

            co = ws.cell(row=fila, column=4, value=it['costo'])
            co.number_format = '#,##0'
            co.alignment = right
            co.border = border

            v = ws.cell(row=fila, column=5, value=it['precio_venta'])
            v.number_format = '#,##0'
            v.alignment = right
            v.border = border
            fila += 1

        fila += 1  # espacio entre sucursales

    if not hubo_datos:
        ws.cell(row=1, column=1, value="No hay productos con stock para los filtros actuales.")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16


@require_GET
@login_required
def exportar_resumen_existencias_pdf(request):
    """Exportar resumen de existencias a PDF con estética NEXO usando ReportLab."""
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            KeepTogether, PageBreak,
        )

        # Reutiliza la lógica del JSON (respeta excluir_articulos, fecha_corte, etc.)
        agrupar_por = request.GET.get('agrupar_por', 'sucursal')
        fecha_corte = request.GET.get('fecha_corte')

        response_data = obtener_resumen_existencias(request)
        try:
            datos = json.loads(response_data.content)
        except Exception:
            return JsonResponse({'success': False, 'error': 'No se pudieron obtener los datos'})

        if not datos.get('success'):
            return JsonResponse({
                'success': False,
                'error': datos.get('error', 'No se pudieron obtener los datos'),
            })

        datos_resumen = datos.get('datos', []) or []
        total_general = datos.get('total_general', {}) or {}
        resumen_empresas = datos.get('resumen_empresas', []) or []
        es_historico = datos.get('es_historico', False)
        excluir_count = datos.get('excluir_articulos_count', 0)
        totales_excluidos = datos.get('totales_excluidos', {}) or {}

        if not datos_resumen:
            return JsonResponse({'success': False, 'error': 'No hay datos para exportar'})

        # ===== Paleta NEXO =====
        nexo_secondary = colors.HexColor('#1A1A2E')
        nexo_primary = colors.HexColor('#0066FF')
        nexo_primary_light = colors.HexColor('#E6F0FF')
        nexo_success = colors.HexColor('#00D4AA')
        nexo_warning = colors.HexColor('#FFB020')
        nexo_error = colors.HexColor('#FF4D4D')
        nexo_gray_700 = colors.HexColor('#4A4A5A')
        nexo_gray_500 = colors.HexColor('#8A8A9A')
        nexo_gray_300 = colors.HexColor('#D1D1D9')
        nexo_gray_100 = colors.HexColor('#F5F5F7')

        # ===== Documento =====
        buffer = BytesIO()
        page_w, page_h = A4
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=1.4 * cm,
            rightMargin=1.4 * cm,
            topMargin=1.4 * cm,
            bottomMargin=1.6 * cm,
            title='Resumen de Existencias - NEXO',
            author='NEXO RetailMind',
        )
        content_w = page_w - doc.leftMargin - doc.rightMargin

        styles = getSampleStyleSheet()
        st_title = ParagraphStyle(
            'TitleNexo', parent=styles['Title'],
            fontSize=18, leading=22, textColor=colors.white,
            alignment=TA_LEFT, spaceAfter=2, fontName='Helvetica-Bold',
        )
        st_subtitle = ParagraphStyle(
            'SubNexo', parent=styles['Normal'],
            fontSize=9, leading=12, textColor=colors.HexColor('#CFD3DC'),
            alignment=TA_LEFT, fontName='Helvetica',
        )
        st_section = ParagraphStyle(
            'Section', parent=styles['Heading2'],
            fontSize=11, leading=14, textColor=nexo_secondary,
            spaceBefore=8, spaceAfter=6, fontName='Helvetica-Bold',
        )
        st_card_label = ParagraphStyle(
            'CardLabel', parent=styles['Normal'],
            fontSize=7, leading=9, textColor=nexo_gray_500,
            alignment=TA_CENTER, fontName='Helvetica-Bold',
        )
        st_card_value = ParagraphStyle(
            'CardValue', parent=styles['Normal'],
            fontSize=14, leading=17, textColor=nexo_primary,
            alignment=TA_CENTER, fontName='Helvetica-Bold',
        )
        st_meta = ParagraphStyle(
            'Meta', parent=styles['Normal'],
            fontSize=8, leading=11, textColor=nexo_gray_700,
            alignment=TA_LEFT, fontName='Helvetica',
        )
        st_note = ParagraphStyle(
            'Note', parent=styles['Normal'],
            fontSize=8, leading=11, textColor=colors.HexColor('#8a6914'),
            alignment=TA_LEFT, fontName='Helvetica-Oblique',
        )

        def fmt_int(n):
            try:
                return f"{int(round(float(n))):,}".replace(',', '.')
            except (TypeError, ValueError):
                return '0'

        def fmt_money(n):
            return '$' + fmt_int(n)

        elements = []

        # ===== HEADER (banda oscura tipo NEXO) =====
        titulo_principal = 'RESUMEN DE EXISTENCIAS'
        if agrupar_por == 'categoria':
            titulo_principal += ' POR CATEGORÍA'
        else:
            titulo_principal += ' POR SUCURSAL'

        sub_partes = []
        if es_historico and fecha_corte:
            sub_partes.append(f"Stock al cierre del <b>{fecha_corte}</b>")
        else:
            sub_partes.append(f"Inventario actual al {timezone.localdate().strftime('%d/%m/%Y')}")

        if request.GET.get('marca_id'):
            sub_partes.append('Filtro de marca aplicado')
        if request.GET.get('categoria_id') and agrupar_por != 'categoria':
            sub_partes.append('Filtro de categoría aplicado')
        if excluir_count:
            sub_partes.append(f"<b>{excluir_count}</b> artículo(s) excluidos del análisis")

        header_inner = Table(
            [[Paragraph(titulo_principal, st_title)],
             [Paragraph(' &nbsp;·&nbsp; '.join(sub_partes), st_subtitle)]],
            colWidths=[content_w - 1.0 * cm],
        )
        header_inner.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))

        header = Table([[header_inner]], colWidths=[content_w])
        header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), nexo_secondary),
            ('LEFTPADDING', (0, 0), (-1, -1), 16),
            ('RIGHTPADDING', (0, 0), (-1, -1), 16),
            ('TOPPADDING', (0, 0), (-1, -1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header)
        elements.append(Spacer(1, 0.4 * cm))

        # ===== TARJETAS DE TOTALES =====
        cards = [
            ('TOTAL UNIDADES', fmt_int(total_general.get('pares', 0)), nexo_primary),
            ('VALOR COSTO', fmt_money(total_general.get('costo', 0)), nexo_success),
            ('VALOR P. INTERNO', fmt_money(total_general.get('precio_interno', 0)), colors.HexColor('#c99200')),
            ('VALOR P. VENTA', fmt_money(total_general.get('precio_venta', 0)), nexo_error),
        ]

        card_cells = []
        for label, value, color_top in cards:
            inner = Table(
                [
                    [Paragraph(' ', st_card_label)],  # franja superior
                    [Paragraph(label, st_card_label)],
                    [Paragraph(value, ParagraphStyle(
                        'CardValueCustom', parent=st_card_value, textColor=color_top))],
                ],
                colWidths=[(content_w - 0.3 * cm * 3) / 4],
                rowHeights=[0.18 * cm, 0.6 * cm, 1.0 * cm],
            )
            inner.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), color_top),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('BOX', (0, 1), (-1, -1), 0.5, nexo_gray_300),
                ('LEFTPADDING', (0, 1), (-1, -1), 6),
                ('RIGHTPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, 1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, 1), 0),
                ('TOPPADDING', (0, 2), (-1, 2), 0),
                ('BOTTOMPADDING', (0, 2), (-1, 2), 6),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ]))
            card_cells.append(inner)

        cards_row = Table(
            [card_cells],
            colWidths=[(content_w - 0.3 * cm * 3) / 4] * 4,
        )
        cards_row.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(cards_row)
        elements.append(Spacer(1, 0.5 * cm))

        # ===== TABLA PRINCIPAL =====
        if agrupar_por == 'categoria':
            elements.append(Paragraph('Detalle por Categoría', st_section))
            headers = ['Categoría', 'Sucursales', 'Unidades', 'Costo', 'P. Interno', 'P. Venta']
            col_widths = [
                content_w * 0.32,  # Categoría
                content_w * 0.10,  # Sucursales
                content_w * 0.12,  # Unidades
                content_w * 0.15,  # Costo
                content_w * 0.15,  # P. Interno
                content_w * 0.16,  # P. Venta
            ]
            table_data = [headers]
            for item in datos_resumen:
                table_data.append([
                    str(item.get('categoria', '') or ''),
                    fmt_int(item.get('num_sucursales', 0)),
                    fmt_int(item.get('total_pares', 0)),
                    fmt_money(item.get('total_costo', 0)),
                    fmt_money(item.get('total_precio_interno', 0)),
                    fmt_money(item.get('total_precio_venta', 0)),
                ])
        else:
            elements.append(Paragraph('Detalle por Sucursal', st_section))
            headers = ['Sucursal', 'Empresa', 'Unidades', 'Costo', 'P. Interno', 'P. Venta']
            col_widths = [
                content_w * 0.18,
                content_w * 0.24,
                content_w * 0.11,
                content_w * 0.15,
                content_w * 0.15,
                content_w * 0.17,
            ]
            table_data = [headers]
            for item in datos_resumen:
                table_data.append([
                    str(item.get('sucursal', '') or ''),
                    str(item.get('empresa', '') or '-'),
                    fmt_int(item.get('total_pares', 0)),
                    fmt_money(item.get('total_costo', 0)),
                    fmt_money(item.get('total_precio_interno', 0)),
                    fmt_money(item.get('total_precio_venta', 0)),
                ])

        # Fila TOTAL
        total_row = ['TOTAL GENERAL', '',
                     fmt_int(total_general.get('pares', 0)),
                     fmt_money(total_general.get('costo', 0)),
                     fmt_money(total_general.get('precio_interno', 0)),
                     fmt_money(total_general.get('precio_venta', 0))]
        table_data.append(total_row)

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        last_row = len(table_data) - 1
        tbl_style = [
            ('BACKGROUND', (0, 0), (-1, 0), nexo_secondary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, nexo_gray_300),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, nexo_gray_100]),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            # Fila TOTAL GENERAL
            ('BACKGROUND', (0, last_row), (-1, last_row), nexo_success),
            ('TEXTCOLOR', (0, last_row), (-1, last_row), colors.white),
            ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
            ('FONTSIZE', (0, last_row), (-1, last_row), 9),
            ('SPAN', (0, last_row), (1, last_row)),
            ('ALIGN', (0, last_row), (1, last_row), 'RIGHT'),
        ]
        tbl.setStyle(TableStyle(tbl_style))
        elements.append(tbl)

        # ===== RESUMEN POR EMPRESA =====
        if resumen_empresas and len(resumen_empresas) > 1:
            elements.append(Spacer(1, 0.6 * cm))
            elements.append(Paragraph('Resumen por Empresa', st_section))

            emp_headers = ['Empresa', 'Sucursales', 'Unidades', 'Costo', 'P. Interno', 'P. Venta']
            emp_widths = [
                content_w * 0.32,
                content_w * 0.10,
                content_w * 0.12,
                content_w * 0.15,
                content_w * 0.15,
                content_w * 0.16,
            ]
            emp_data = [emp_headers]
            for emp in resumen_empresas:
                emp_data.append([
                    str(emp.get('empresa', '') or ''),
                    fmt_int(emp.get('num_sucursales', 0)),
                    fmt_int(emp.get('total_pares', 0)),
                    fmt_money(emp.get('total_costo', 0)),
                    fmt_money(emp.get('total_precio_interno', 0)),
                    fmt_money(emp.get('total_precio_venta', 0)),
                ])

            tbl_emp = Table(emp_data, colWidths=emp_widths, repeatRows=1)
            tbl_emp.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), nexo_primary),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.4, nexo_gray_300),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, nexo_primary_light]),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(KeepTogether(tbl_emp))

        # ===== NOTAS =====
        if excluir_count:
            elements.append(Spacer(1, 0.4 * cm))
            pares_exc = int(totales_excluidos.get('pares', 0) or 0)
            detalle_exc = ''
            if pares_exc > 0:
                detalle_exc = (
                    f" Representan <b>{fmt_int(pares_exc)}</b> unidad(es) por "
                    f"<b>{fmt_money(totales_excluidos.get('precio_venta', 0))}</b> a precio venta "
                    f"(<b>{fmt_money(totales_excluidos.get('costo', 0))}</b> a costo)."
                )
            elements.append(Paragraph(
                f"Nota: {excluir_count} artículo(s) excluidos del análisis "
                f"(filtro temporal del navegador del usuario; no afecta otros reportes)."
                + detalle_exc,
                st_note,
            ))

        # ===== TOP 30 POR SUCURSAL (solo agrupación por sucursal) =====
        if agrupar_por == 'sucursal':
            empresas_usuario, marca_id_f, excluir_ids_f, fecha_corte_f, es_hist_f = \
                _obtener_filtros_top_export(request)

            secciones_top = []
            for suc in datos_resumen:
                sucursal_id = suc.get('sucursal_id')
                if not sucursal_id:
                    continue
                detalle = _calcular_detalle_top_por_sucursal(
                    empresas_usuario, sucursal_id, limite=30,
                    marca_id=marca_id_f, excluir_ids=excluir_ids_f,
                    fecha_corte=fecha_corte_f, es_historico=es_hist_f,
                )
                items = detalle['items']
                if not items:
                    continue

                bloque = []
                bloque.append(Paragraph(
                    f"Top 30 · {suc.get('sucursal', 'Sucursal')} "
                    f"<font size=7 color='#8A8A9A'>({detalle['total_productos']} productos con stock)</font>",
                    st_section,
                ))

                top_headers = ['#', 'Artículo', 'Unidades', 'Costo', 'P. Venta']
                top_widths = [
                    content_w * 0.06,
                    content_w * 0.50,
                    content_w * 0.14,
                    content_w * 0.15,
                    content_w * 0.15,
                ]
                top_data = [top_headers]
                for pos, it in enumerate(items, start=1):
                    art = str(it.get('articulo', '') or '')
                    if it.get('descripcion'):
                        art += f" — {it['descripcion']}"
                    if len(art) > 70:
                        art = art[:67] + '…'
                    top_data.append([
                        str(pos),
                        art,
                        fmt_int(it.get('pares', 0)),
                        fmt_money(it.get('costo', 0)),
                        fmt_money(it.get('precio_venta', 0)),
                    ])

                tbl_top = Table(top_data, colWidths=top_widths, repeatRows=1)
                tbl_top.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), nexo_secondary),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('GRID', (0, 0), (-1, -1), 0.4, nexo_gray_300),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, nexo_gray_100]),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                bloque.append(tbl_top)
                bloque.append(Spacer(1, 0.4 * cm))
                secciones_top.append(KeepTogether(bloque))

            if secciones_top:
                elements.append(PageBreak())
                titulo_top = 'DETALLE TOP 30 POR SUCURSAL'
                if es_hist_f and fecha_corte_f:
                    titulo_top += f" — Al {fecha_corte_f.strftime('%d/%m/%Y')}"
                header_top = Table([[Paragraph(titulo_top, st_title)]], colWidths=[content_w])
                header_top.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), nexo_secondary),
                    ('LEFTPADDING', (0, 0), (-1, -1), 16),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 16),
                    ('TOPPADDING', (0, 0), (-1, -1), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(header_top)
                elements.append(Spacer(1, 0.4 * cm))
                elements.extend(secciones_top)

        # ===== FOOTER =====
        usuario_nombre = (
            request.user.get_full_name()
            if request.user.is_authenticated and request.user.get_full_name()
            else getattr(request.user, 'username', 'Sistema')
        )
        fecha_gen = timezone.localtime().strftime('%d/%m/%Y %H:%M')

        def _footer(canvas, doc_):
            canvas.saveState()
            canvas.setStrokeColor(nexo_gray_300)
            canvas.setLineWidth(0.4)
            canvas.line(
                doc_.leftMargin, 1.2 * cm,
                page_w - doc_.rightMargin, 1.2 * cm,
            )
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(nexo_gray_500)
            canvas.drawString(
                doc_.leftMargin, 0.7 * cm,
                f"NEXO · RetailMind  ·  Generado {fecha_gen}  ·  Por {usuario_nombre}",
            )
            canvas.drawRightString(
                page_w - doc_.rightMargin, 0.7 * cm,
                f"Página {doc_.page}",
            )
            canvas.restoreState()

        doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
        buffer.seek(0)

        # Nombre de archivo
        filename = 'resumen_existencias'
        if agrupar_por == 'categoria':
            filename += '_por_categoria'
        if es_historico and fecha_corte:
            filename += f'_{fecha_corte}'
        filename += '.pdf'

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.read())
        return response

    except Exception as e:
        logger.exception("Error al exportar resumen de existencias a PDF")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar PDF: {str(e)}',
        })


@require_GET
@login_required
def listar_sucursales_resumen(request):
    """
    Devuelve las sucursales pertenecientes a las empresas del usuario actual.
    Usado por el modal de exclusiones del reporte de resumen de existencias.
    """
    try:
        empresas_usuario = list(EmpresaUser.objects.filter(
            user=request.user,
            status=True,
        ).values_list('empresa_id', flat=True))

        sucursales = (
            Sucursal.objects
            .filter(empresa_id__in=empresas_usuario)
            .select_related('empresa')
            .order_by('empresa__nombre', 'alias')
        )

        items = []
        for s in sucursales:
            items.append({
                'id': s.id,
                'alias': s.alias,
                'empresa': s.empresa.nombre if s.empresa else '-',
            })

        return JsonResponse({'success': True, 'items': items})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'items': []})


def _calcular_detalle_top_por_sucursal(empresas_usuario, sucursal_id, limite=30,
                                       marca_id=None, excluir_ids=None,
                                       fecha_corte=None, es_historico=False):
    """
    Núcleo de cálculo del Top-N de productos con stock en UNA sucursal, ordenado por
    unidades (pares) descendente. Agrega a nivel de PRODUCTO (suma de todas sus tallas).

    Reutilizado por el endpoint del modal y por los exports (Excel/PDF).

    Devuelve un dict con:
      items (top-N), total_productos, total_pares, total_costo, total_precio_venta.
    """
    excluir_ids = excluir_ids or []
    limite = max(1, min(int(limite or 30), 200))

    talla_qs = Producto_Talla.objects.filter(
        producto__sucursal_id=sucursal_id,
        producto__sucursal__empresa_id__in=empresas_usuario,
        producto__excluir_de_analitica=False,
    )
    if marca_id:
        talla_qs = talla_qs.filter(producto__atributo1_id=marca_id)
    if excluir_ids:
        talla_qs = talla_qs.exclude(producto_id__in=excluir_ids)

    productos = {}  # producto_id -> bucket

    if es_historico and fecha_corte:
        tallas_data = list(
            talla_qs.values(
                'id', 'stock', 'producto_id',
                'producto__articulo', 'producto__descripcion',
                'producto__costo', 'producto__sobreprecio', 'producto__precioventa',
            )
        )
        talla_ids = [t['id'] for t in tallas_data]

        ingresos = (
            Movimientos_Producto.objects
            .filter(ProductoTalla_id__in=talla_ids, fecha__gt=fecha_corte, estado='COMPLETADO')
            .filter(Q(tipo_movimiento='INGRESO') | Q(concepto='TRASPASO_ENTRADA'))
            .filter(sucursal_destino_id=sucursal_id)
            .values('ProductoTalla_id')
            .annotate(total=Sum('cantidad'))
        )
        ingresos_map = {r['ProductoTalla_id']: int(r['total'] or 0) for r in ingresos}

        egresos = (
            Movimientos_Producto.objects
            .filter(ProductoTalla_id__in=talla_ids, fecha__gt=fecha_corte, estado='COMPLETADO')
            .filter(Q(tipo_movimiento='EGRESO') | Q(concepto='TRASPASO_SALIDA'))
            .filter(sucursal_origen_id=sucursal_id)
            .values('ProductoTalla_id')
            .annotate(total=Sum('cantidad'))
        )
        egresos_map = {r['ProductoTalla_id']: int(r['total'] or 0) for r in egresos}

        for t in tallas_data:
            stock_hist = (t['stock'] or 0) - ingresos_map.get(t['id'], 0) + abs(egresos_map.get(t['id'], 0))
            if stock_hist <= 0:
                continue
            _acumular_producto_detalle(productos, t, stock_hist)
    else:
        tallas_data = talla_qs.filter(stock__gt=0).values(
            'stock', 'producto_id',
            'producto__articulo', 'producto__descripcion',
            'producto__costo', 'producto__sobreprecio', 'producto__precioventa',
        )
        for t in tallas_data:
            _acumular_producto_detalle(productos, t, t['stock'] or 0)

    items_ordenados = sorted(productos.values(), key=lambda p: p['pares'], reverse=True)
    top = items_ordenados[:limite]

    return {
        'items': [{
            'producto_id': p['producto_id'],
            'articulo': p['articulo'],
            'descripcion': p['descripcion'],
            'pares': p['pares'],
            'costo': p['costo'],
            'precio_venta': p['p_venta'],
        } for p in top],
        'top_n': limite,
        'total_productos': len(items_ordenados),
        'total_pares': sum(p['pares'] for p in items_ordenados),
        'total_costo': sum(p['costo'] for p in items_ordenados),
        'total_precio_venta': sum(p['p_venta'] for p in items_ordenados),
    }


@require_GET
@login_required
def detalle_stock_sucursal(request):
    """
    Devuelve el detalle de productos con stock de UNA sucursal, ordenado por
    unidades (pares) descendente, para el modal "Ver lo que queda" del reporte.

    Respeta:
      - las empresas del usuario (multi-tenant / seguridad),
      - el filtro de marca (marca_id),
      - las exclusiones activas (excluir_articulos),
      - la fecha de corte histórica (fecha_corte) reconstruyendo el stock a esa fecha.

    Parámetros GET:
      - sucursal_id (obligatorio)
      - marca_id, fecha_corte, excluir_articulos (opcionales, mismos que el reporte)
      - limite (opcional, default 30, máx 200)
    """
    try:
        sucursal_id = request.GET.get('sucursal_id')
        if not sucursal_id or not str(sucursal_id).isdigit():
            return JsonResponse({'success': False, 'error': 'sucursal_id inválido'})
        sucursal_id = int(sucursal_id)

        marca_id = request.GET.get('marca_id')
        excluir_ids = _parse_excluir_articulos(request)

        try:
            limite = int(request.GET.get('limite', 30))
        except (TypeError, ValueError):
            limite = 30

        # Fecha de corte histórica
        fecha_corte = None
        es_historico = False
        fecha_corte_str = request.GET.get('fecha_corte')
        if fecha_corte_str:
            try:
                fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
                es_historico = fecha_corte < timezone.localdate()
            except ValueError:
                pass

        # Multi-tenant + validar que la sucursal pertenece al usuario
        empresas_usuario = list(EmpresaUser.objects.filter(
            user=request.user, status=True,
        ).values_list('empresa_id', flat=True))

        sucursal = (
            Sucursal.objects
            .filter(id=sucursal_id, empresa_id__in=empresas_usuario)
            .select_related('empresa')
            .first()
        )
        if not sucursal:
            return JsonResponse({'success': False, 'error': 'Sucursal no encontrada o sin permisos'})

        detalle = _calcular_detalle_top_por_sucursal(
            empresas_usuario, sucursal_id, limite=limite,
            marca_id=marca_id, excluir_ids=excluir_ids,
            fecha_corte=fecha_corte, es_historico=es_historico,
        )

        return JsonResponse({
            'success': True,
            'sucursal_id': sucursal.id,
            'sucursal': sucursal.alias,
            'empresa': sucursal.empresa.nombre if sucursal.empresa else '-',
            'items': detalle['items'],
            'top_n': detalle['top_n'],
            'total_productos': detalle['total_productos'],
            'total_pares': detalle['total_pares'],
            'total_costo': detalle['total_costo'],
            'total_precio_venta': detalle['total_precio_venta'],
            'es_historico': es_historico,
            'fecha_corte': fecha_corte_str if es_historico else None,
        })
    except Exception as e:
        logger.exception("Error en detalle_stock_sucursal")
        return JsonResponse({'success': False, 'error': str(e), 'items': []})


def _acumular_producto_detalle(productos, talla_row, stock):
    """Acumula una talla dentro del bucket de su producto para el detalle por sucursal."""
    pid = talla_row['producto_id']
    b = productos.get(pid)
    if b is None:
        b = {
            'producto_id': pid,
            'articulo': talla_row['producto__articulo'] or '',
            'descripcion': talla_row['producto__descripcion'] or '',
            'pares': 0, 'costo': 0, 'p_venta': 0,
        }
        productos[pid] = b
    costo = talla_row['producto__costo'] or 0
    venta = talla_row['producto__precioventa'] or 0
    b['pares'] += stock
    b['costo'] += costo * stock
    b['p_venta'] += venta * stock


@require_GET
@login_required
def listar_articulos_para_excluir(request):
    """
    Devuelve hasta 50 productos del usuario para alimentar el modal de
    exclusiones del reporte de resumen de existencias.

    Filtra por las empresas del usuario (multi-tenant).
    Acepta los parámetros:
      - sucursal_id (opcional): filtra por una sucursal específica
      - q (opcional): texto a buscar en articulo, descripcion o sku
      - ids (opcional): CSV de IDs a recuperar siempre (para mostrar
        los ya excluidos aunque no calcen con la búsqueda)
    """
    try:
        sucursal_id = request.GET.get('sucursal_id') or None
        q = (request.GET.get('q') or '').strip()
        ids_raw = request.GET.get('ids') or ''
        ids_solicitados = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]

        empresas_usuario = list(EmpresaUser.objects.filter(
            user=request.user,
            status=True,
        ).values_list('empresa_id', flat=True))

        base_qs = Producto.objects.filter(
            sucursal__empresa_id__in=empresas_usuario,
        ).select_related('sucursal', 'sucursal__empresa')

        if sucursal_id:
            try:
                base_qs = base_qs.filter(sucursal_id=int(sucursal_id))
            except (TypeError, ValueError):
                pass

        # Búsqueda combinada
        productos = []
        seen_ids = set()

        if q:
            buscar_qs = base_qs.filter(
                Q(articulo__icontains=q)
                | Q(descripcion__icontains=q)
                | Q(producto_talla__sku__icontains=q)
            ).distinct().order_by('articulo')[:50]
            for p in buscar_qs:
                if p.id in seen_ids:
                    continue
                seen_ids.add(p.id)
                productos.append(p)
        elif not ids_solicitados:
            # Sin búsqueda y sin ids puntuales: devolvemos los primeros 50
            for p in base_qs.order_by('articulo')[:50]:
                if p.id in seen_ids:
                    continue
                seen_ids.add(p.id)
                productos.append(p)

        # Anexar IDs explícitamente solicitados (los excluidos actuales)
        if ids_solicitados:
            faltantes = [i for i in ids_solicitados if i not in seen_ids]
            if faltantes:
                for p in base_qs.filter(id__in=faltantes):
                    if p.id in seen_ids:
                        continue
                    seen_ids.add(p.id)
                    productos.append(p)

        # Calcular stock total de cada producto en su sucursal
        productos_ids = [p.id for p in productos]
        stock_por_producto = {}
        if productos_ids:
            agg = (
                Producto_Talla.objects
                .filter(producto_id__in=productos_ids)
                .values('producto_id')
                .annotate(total=Sum('stock'))
            )
            for row in agg:
                stock_por_producto[row['producto_id']] = max(0, row['total'] or 0)

        items = []
        for p in productos:
            items.append({
                'id': p.id,
                'articulo': p.articulo,
                'descripcion': p.descripcion,
                'sucursal_id': p.sucursal_id,
                'sucursal': p.sucursal.alias if p.sucursal else '-',
                'empresa': p.sucursal.empresa.nombre if p.sucursal and p.sucursal.empresa else '-',
                'stock_total': stock_por_producto.get(p.id, 0),
            })

        return JsonResponse({
            'success': True,
            'items': items,
            'total': len(items),
            'limite': 50,
        })
    except Exception as e:
        logger.exception("Error en listar_articulos_para_excluir")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'items': [],
        })


@require_GET
@login_required
def verificar_disponibilidad_historico(request):
    """
    Verifica si hay suficientes movimientos registrados para calcular inventario histórico.
    Útil para mostrar alertas al usuario.
    """
    try:
        # Contar movimientos por tipo
        from django.db.models.functions import TruncMonth
        
        total_movimientos = Movimientos_Producto.objects.count()
        
        # Rango de fechas disponible
        primer_movimiento = Movimientos_Producto.objects.order_by('fecha').first()
        ultimo_movimiento = Movimientos_Producto.objects.order_by('-fecha').first()
        
        if total_movimientos == 0:
            return JsonResponse({
                'success': True,
                'disponible': False,
                'mensaje': 'No hay movimientos registrados. El inventario histórico no está disponible.',
                'total_movimientos': 0,
                'fecha_inicio': None,
                'fecha_fin': None
            })
        
        return JsonResponse({
            'success': True,
            'disponible': True,
            'total_movimientos': total_movimientos,
            'fecha_inicio': primer_movimiento.fecha.strftime('%Y-%m-%d') if primer_movimiento else None,
            'fecha_fin': ultimo_movimiento.fecha.strftime('%Y-%m-%d') if ultimo_movimiento else None,
            'mensaje': f'Historial disponible desde {primer_movimiento.fecha.strftime("%d/%m/%Y")} hasta {ultimo_movimiento.fecha.strftime("%d/%m/%Y")}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
