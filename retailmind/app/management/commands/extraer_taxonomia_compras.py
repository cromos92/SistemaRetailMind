# -*- coding: utf-8 -*-
"""
Cruza el catálogo del ERP contra las planillas de Compras y extrae, por artículo:
proveedor, año, planilla de origen, flag escolar (carpeta/planilla BTS/ESCOLAR/
SCHOOL/COLEGIAL) y la taxonomía que traiga la planilla (CATEGORIA/FAMILIA/GENDER…).

Mejoras sobre el match de preview_recategorizacion_v12 (auditoría 2026-07-15):
  - Recorre TODOS los años por defecto (2013-2027 + "2025 ACCESORIOS"), no solo 2025-27.
  - Lee TODAS las hojas de cada planilla, no solo la primera.
  - Reconoce más cabeceras de código (STYLE, REFERENCIA, SKU, EAN…) y arma el
    código completo STYLE-COLOR cuando vienen en columnas separadas (SKECHERS).
  - Match multi-vía: exacto → estilo sin color → prefijo largo (≥5), y reporta
    si la marca del ERP coincide con la carpeta del proveedor (filtra falsos).

100% de solo lectura (BD y planillas). Salida: Excel.

Uso:
    python manage.py extraer_taxonomia_compras                          # todo (tarda)
    python manage.py extraer_taxonomia_compras --proveedor SKECHERS     # prueba rápida
    python manage.py extraer_taxonomia_compras --anios 2024,2025,2026,2027
"""
import glob
import logging
import os
import re
import unicodedata
from collections import Counter, defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from app.models import Producto

logger = logging.getLogger('app')

COMPRAS_DEFAULT = r'C:\Users\cromo\Documents\Compras'

# Cabeceras aceptadas (foldeadas). 'STYLE' se combina con 'COLOR' si existe.
HDR_CODIGO = {'PRODUCTO', 'CODIGO', 'CÓDIGO', 'COD', 'ARTICULO', 'ARTÍCULO', 'STYLE',
              'REFERENCIA', 'REF', 'SKU', 'MATERIAL', 'ITEM', 'ESTILO'}
HDR_COLOR = {'COLOR', 'COD COLOR', 'COLOR CODE'}
HDR_DESC = {'DESCRIPCION', 'DESCRIPCIÓN', 'DETALLE NOMBRE', 'MODEL DESCRIPTION',
            'DESCRIPTION', 'MODELO', 'NOMBRE', 'DETALLE', 'OUTSOLE'}
HDR_CAT = {'CATEGORIA', 'CATEGORÍA', 'FAMILIA', 'SUPER FAMILIA', 'MUNDO', 'SHOE TYPE',
           'DIVISION', 'DIVISIÓN', 'LINEA', 'LÍNEA', 'AREA', 'SILUETA', 'GRUPO'}
HDR_GEN = {'GENERO', 'GÉNERO', 'GENDER', 'GENDER CLASS', 'SEXO', 'SUB-GENDER'}

RE_ESCOLAR = re.compile(r'BTS|ESCOLAR|SCHOOL|COLEGIAL', re.I)


def fold(s):
    s = unicodedata.normalize('NFKD', str(s or ''))
    return ''.join(c for c in s if not unicodedata.combining(c)).upper().strip()


def detectar_columnas(ws):
    """Busca la fila de cabecera en las primeras 30 filas. Devuelve (start, cols) o None.
    cols = dict con índices 1-based: codigo, color, desc, cat, gen (los que haya)."""
    for fila in ws.iter_rows(min_row=1, max_row=30):
        cols = {}
        for c in fila:
            v = fold(c.value)
            if not v:
                continue
            if v in HDR_CODIGO and 'codigo' not in cols:
                cols['codigo'] = c.column
            elif v in HDR_COLOR and 'color' not in cols:
                cols['color'] = c.column
            elif v in HDR_DESC and 'desc' not in cols:
                cols['desc'] = c.column
            elif v in HDR_CAT and 'cat' not in cols:
                cols['cat'] = c.column
            elif v in HDR_GEN and 'gen' not in cols:
                cols['gen'] = c.column
        if 'codigo' in cols:
            return fila[0].row + 1, cols
    return None


def valido(cod):
    return 3 <= len(cod) <= 40 and not cod.isspace() and cod not in ('0', 'X', '-')


class Command(BaseCommand):
    help = "Extrae proveedor/planilla/taxonomía de Compras por artículo del ERP (solo lectura)"

    def add_arguments(self, parser):
        parser.add_argument('--compras-dir', default=COMPRAS_DEFAULT)
        parser.add_argument('--anios', default='',
                            help='Años separados por coma (default: todas las carpetas)')
        parser.add_argument('--proveedor', default='',
                            help='Solo carpetas de proveedor que contengan este texto (prueba rápida)')
        parser.add_argument('--salida', default='',
                            help='Ruta del xlsx (default: taxonomia_compras_<fecha>.xlsx)')

    # ------------------------------------------------------------------
    def _indexar(self, base, anios, filtro_prov):
        """{codigo: meta} con meta del PRIMER hallazgo (prefiere años recientes)."""
        if anios:
            carpetas = [os.path.join(base, a) for a in anios]
        else:
            carpetas = sorted((d.path for d in os.scandir(base) if d.is_dir()), reverse=True)
        idx = {}
        por_estilo = defaultdict(set)   # STYLE (o código sin -COLOR) -> códigos completos
        n_files = n_ok = 0
        from openpyxl import load_workbook
        for carpeta in carpetas:
            anio = os.path.basename(carpeta)
            paths = [p for p in glob.glob(os.path.join(carpeta, '**', '*.xlsx'), recursive=True)
                     if '~$' not in p]
            paths += [p for p in glob.glob(os.path.join(carpeta, '**', '*.xlsm'), recursive=True)
                      if '~$' not in p]
            for path in sorted(paths):
                rel = os.path.relpath(path, base)
                partes = rel.split(os.sep)
                proveedor = partes[1] if len(partes) > 2 else os.path.splitext(partes[-1])[0]
                if filtro_prov and filtro_prov.upper() not in fold(proveedor):
                    continue
                n_files += 1
                escolar = bool(RE_ESCOLAR.search(rel))
                ok = False
                try:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    for ws in wb.worksheets:
                        det = detectar_columnas(ws)
                        if not det:
                            continue
                        start, cols = det
                        ok = True
                        idxs = sorted(cols.values())
                        lo, hi = idxs[0], idxs[-1]
                        for fila in ws.iter_rows(min_row=start, min_col=lo, max_col=hi,
                                                 values_only=True):
                            def cell(k):
                                return str(fila[cols[k] - lo] or '').strip() if k in cols else ''
                            cod = cell('codigo').upper()
                            if not valido(cod):
                                continue
                            color = cell('color').upper()
                            # SKECHERS y similares: STYLE + COLOR en columnas separadas
                            if color and valido(color) and len(color) <= 6 and '-' not in cod:
                                completo = f'{cod}-{color}'
                            else:
                                completo = cod
                            if completo not in idx:
                                idx[completo] = {
                                    'proveedor': proveedor, 'anio': anio,
                                    'planilla': partes[-1], 'escolar': escolar,
                                    'cat': cell('cat'), 'gen': cell('gen'),
                                    'desc': cell('desc'),
                                }
                            pref = completo.rsplit('-', 1)[0] if '-' in completo else completo
                            por_estilo[pref].add(completo)
                    wb.close()
                except Exception:
                    pass
                if ok:
                    n_ok += 1
                if n_files % 200 == 0:
                    self.stdout.write(f'    {n_files} planillas... ({len(idx):,} códigos)')
        self.stdout.write(f'  {n_ok}/{n_files} planillas legibles, {len(idx):,} códigos únicos.')
        return idx, por_estilo

    # ------------------------------------------------------------------
    def handle(self, *args, **opts):
        base = opts['compras_dir']
        if not os.path.isdir(base):
            raise CommandError(f'No existe la carpeta: {base}')
        anios = [a.strip() for a in opts['anios'].split(',') if a.strip()]

        self.stdout.write('1/3 Indexando planillas de Compras (todas las hojas)...')
        idx, por_estilo = self._indexar(base, anios, opts['proveedor'])

        self.stdout.write('2/3 Leyendo artículos del ERP (solo lectura)...')
        arts = {}
        qs = (Producto.objects.values('articulo', 'descripcion', 'atributo1__valor')
              .distinct())
        for p in qs.iterator(chunk_size=5000):
            a = (p['articulo'] or '').strip().upper()
            if a and a not in arts:
                arts[a] = {'desc': p['descripcion'] or '', 'marca': p['atributo1__valor'] or ''}
        self.stdout.write(f'  {len(arts):,} artículos.')

        self.stdout.write('3/3 Matcheando y escribiendo Excel...')
        from openpyxl import Workbook
        stamp = timezone.now().strftime('%Y%m%d_%H%M')
        salida = opts['salida'] or f'taxonomia_compras_{stamp}.xlsx'
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Match')
        ws.append(['articulo', 'marca_erp', 'descripcion_erp', 'match', 'codigo_planilla',
                   'proveedor', 'anio', 'planilla', 'marca_coincide', 'escolar',
                   'categoria_planilla', 'genero_planilla', 'desc_planilla'])

        stats = Counter()
        por_marca = defaultdict(Counter)
        por_prov = defaultdict(Counter)
        for a, info in sorted(arts.items()):
            meta = None
            tipo = ''
            cod = ''
            if a in idx:
                tipo, cod, meta = 'exacto', a, idx[a]
            else:
                pref = a.rsplit('-', 1)[0] if '-' in a else a
                cands = por_estilo.get(a) or (por_estilo.get(pref) if pref != a else None)
                if cands:
                    cod = sorted(cands)[0]
                    tipo, meta = ('estilo_otro_color' if pref != a else 'sin_sufijo'), idx[cod]
                elif len(a) >= 5:
                    cands = [c for p, cs in por_estilo.items() if p.startswith(a) for c in cs]
                    if cands:
                        cod = sorted(cands)[0]
                        tipo, meta = 'prefijo?', idx[cod]
            marca = fold(info['marca'])
            stats[tipo or 'no_encontrado'] += 1
            por_marca[info['marca'] or '(sin marca)'][tipo or 'no_encontrado'] += 1
            if meta:
                coincide = 'Sí' if marca and marca.split()[0] in fold(meta['proveedor']) else ''
                por_prov[meta['proveedor']][tipo] += 1
                ws.append([a, info['marca'], info['desc'], tipo, cod,
                           meta['proveedor'], meta['anio'], meta['planilla'], coincide,
                           'Sí' if meta['escolar'] else '', meta['cat'], meta['gen'],
                           meta['desc']])

        ws2 = wb.create_sheet('Resumen_marca')
        ws2.append(['marca_erp', 'exacto', 'estilo_otro_color', 'sin_sufijo', 'prefijo?',
                    'no_encontrado', 'total'])
        for m, c in sorted(por_marca.items(), key=lambda x: -sum(x[1].values())):
            ws2.append([m, c.get('exacto', 0), c.get('estilo_otro_color', 0),
                        c.get('sin_sufijo', 0), c.get('prefijo?', 0),
                        c.get('no_encontrado', 0), sum(c.values())])
        ws3 = wb.create_sheet('Resumen_proveedor')
        ws3.append(['proveedor', 'exacto', 'estilo_otro_color', 'sin_sufijo', 'prefijo?', 'total'])
        for pv, c in sorted(por_prov.items(), key=lambda x: -sum(x[1].values())):
            ws3.append([pv, c.get('exacto', 0), c.get('estilo_otro_color', 0),
                        c.get('sin_sufijo', 0), c.get('prefijo?', 0), sum(c.values())])
        wb.save(salida)

        total = sum(stats.values())
        self.stdout.write(self.style.SUCCESS(f'Listo: {salida}'))
        for k in ('exacto', 'estilo_otro_color', 'sin_sufijo', 'prefijo?', 'no_encontrado'):
            if stats.get(k):
                self.stdout.write(f'  {k:18}: {stats[k]:6,} ({100.0*stats[k]/total:.1f}%)')
        logger.info('extraer_taxonomia_compras: %d articulos, %s', total, dict(stats))
