"""
Genera un Excel con TODOS los DTE de un mes, agrupados por sucursal.

Comando de SOLO LECTURA: no escribe nada en la base de datos ni en `Dte`,
solo lee y arma un .xlsx nuevo. Combina dos fuentes porque son
complementarias, no intercambiables:

- Para BOLETA ELECTRONICA, el estado SII real (aceptado/rechazado/con
  reparos) NO vive en `Dte.estado_dte` (ese campo se queda en 'EMITIDO'
  aunque el SII haya rechazado el folio). El estado real solo está en el
  export de Acepta (--acepta-xls, columnas folio/estado_sii/monto_total).
  Ese archivo además solo cubre el RUT de la cuenta Acepta consultada, no
  necesariamente todas las empresas del sistema.
- Para el resto de tipos de documento (factura, NC, ND, guía) no hay
  export de Acepta equivalente en este flujo, así que se usa `Dte`
  directamente con su estado interno.

La base de "qué documentos existen y de qué sucursal son" es siempre
`Dte` (es la única fuente que cubre el 100% del mes); Acepta solo enriquece
el estado real de las boletas que encuentra por folio.

Uso:

    python manage.py reporte_dte_mensual_por_sucursal --mes 2026-06

    python manage.py reporte_dte_mensual_por_sucursal --mes 2026-06 \
        --acepta-xls "acepta mes.xls" \
        --mapeo-xlsx "C:\\ruta\\mapeo_junio_PAO4.xlsx" \
        --salida "reporte_dte_2026-06_por_sucursal.xlsx"
"""

import datetime
import logging
import os
from calendar import monthrange

from django.core.management.base import BaseCommand, CommandError

from app.models import Dte

logger = logging.getLogger('app')

HEADER_FILL = '405189'
HEADER_FONT_COLOR = 'FFFFFF'
WARN_FILL = 'FFF3CD'


class Command(BaseCommand):
    help = (
        'Genera un Excel con todos los DTE de un mes agrupados por sucursal '
        '(boletas enriquecidas con el estado SII real desde un export de '
        'Acepta). Solo lectura: no modifica la base de datos.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--mes', type=str, default='2026-06',
            help='Mes a reportar en formato AAAA-MM (default: 2026-06).',
        )
        parser.add_argument(
            '--acepta-xls', type=str, default='acepta mes.xls',
            help='Ruta al export de Acepta (.xls/.xlsx) con columnas '
                 'folio/estado_sii/monto_total/informacion_sii. '
                 'Si no existe, las boletas quedan sin estado SII real '
                 '(solo con el estado interno de Dte).',
        )
        parser.add_argument(
            '--mapeo-xlsx', type=str, default=None,
            help='Ruta opcional a la planilla de refacturación '
                 '(hoja "Refoliado", columnas Folio Viejo/Folio Nuevo/...). '
                 'Si se pasa, agrega la hoja "Refacturación - Control" y '
                 'anota el folio nuevo en la hoja de boletas.',
        )
        parser.add_argument(
            '--salida', type=str, default=None,
            help='Ruta del .xlsx de salida. Default: '
                 'reporte_dte_<mes>_por_sucursal.xlsx en el directorio actual.',
        )

    # ------------------------------------------------------------------ #
    def handle(self, *args, **options):
        desde, hasta = self._rango_mes(options['mes'])
        salida = options['salida'] or f'reporte_dte_{options["mes"]}_por_sucursal.xlsx'

        acepta_por_folio = {}
        acepta_ruta = options['acepta_xls']
        if acepta_ruta and os.path.exists(acepta_ruta):
            acepta_por_folio = self._cargar_acepta_boletas(acepta_ruta)
            self.stdout.write(
                f'  Acepta: {len(acepta_por_folio)} folios de boleta cargados '
                f'desde {os.path.basename(acepta_ruta)}'
            )
        else:
            self.stdout.write(self.style.WARNING(
                f'  No se encontró el export de Acepta ("{acepta_ruta}"). '
                'Las boletas van sin estado SII real (solo estado interno).'
            ))

        mapeo_folio_nuevo = {}
        mapeo_df = None
        if options['mapeo_xlsx']:
            mapeo_folio_nuevo, mapeo_df = self._cargar_mapeo_refoliado(options['mapeo_xlsx'])
            self.stdout.write(
                f'  Mapeo de refacturación: {len(mapeo_folio_nuevo)} folios '
                f'cargados desde {os.path.basename(options["mapeo_xlsx"])}'
            )

        qs = (
            Dte.objects
            .filter(fecha_emision__gte=desde, fecha_emision__lte=hasta)
            .select_related('sucursal')
            .order_by('sucursal__alias', 'tipo_documento', 'numero_documento')
        )
        total = qs.count()
        if total == 0:
            self.stdout.write(self.style.WARNING(
                f'No se encontraron DTE con fecha_emision entre {desde} y {hasta}.'
            ))
            return

        boletas = [d for d in qs if d.tipo_documento == 'BOLETA ELECTRONICA']
        otros = [d for d in qs if d.tipo_documento != 'BOLETA ELECTRONICA']

        self.stdout.write('=' * 70)
        self.stdout.write(f'[REPORTE DTE MENSUAL POR SUCURSAL]  mes: {options["mes"]}')
        self.stdout.write(f'  Total DTE: {total}  (boletas: {len(boletas)}, otros: {len(otros)})')
        self.stdout.write('=' * 70)

        wb = self._construir_workbook(
            boletas, otros, acepta_por_folio, mapeo_folio_nuevo, mapeo_df, options['mes'],
        )
        wb.save(salida)

        self.stdout.write(self.style.SUCCESS(f'  Archivo generado: {os.path.abspath(salida)}'))
        con_estado_sii = sum(1 for d in boletas if d.numero_documento in acepta_por_folio)
        self.stdout.write(
            f'  Boletas con estado SII real (encontradas en Acepta): '
            f'{con_estado_sii} / {len(boletas)}'
        )
        if con_estado_sii < len(boletas):
            self.stdout.write(self.style.WARNING(
                f'  {len(boletas) - con_estado_sii} boletas sin dato de Acepta '
                '(otra empresa/RUT no cubierto por ese export, o fuera de su rango de fechas).'
            ))

    # ------------------------------------------------------------------ #
    def _construir_workbook(self, boletas, otros, acepta_por_folio,
                             mapeo_folio_nuevo, mapeo_df, mes):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        self._hoja_resumen(wb.active, boletas, otros, acepta_por_folio, mes)
        self._hoja_boletas(wb.create_sheet('Boletas'), boletas, acepta_por_folio, mapeo_folio_nuevo)
        self._hoja_otros(wb.create_sheet('Otros Documentos'), otros)
        if mapeo_df is not None:
            self._hoja_refacturacion(wb.create_sheet('Refacturación - Control'), mapeo_df)

        return wb

    # ------------------------------------------------------------------ #
    def _estilo_header(self, ws, encabezados, fila=1):
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill('solid', fgColor=HEADER_FILL)
        font = Font(bold=True, color=HEADER_FONT_COLOR)
        for col, texto in enumerate(encabezados, start=1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Coordenada como string: usar ws.cell(...) aquí materializaría una
        # celda vacía en esa fila y pandas la leería de vuelta como una fila
        # NaN espuria entre el encabezado y los datos.
        ws.freeze_panes = f'A{fila + 1}'

    def _autoancho(self, ws, anchos):
        for i, w in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl_letter(i)].width = w

    # ------------------------------------------------------------------ #
    def _hoja_resumen(self, ws, boletas, otros, acepta_por_folio, mes):
        ws.title = 'Resumen por Sucursal'
        from openpyxl.styles import Font

        ws.append([f'Resumen DTE {mes} por sucursal'])
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws.append([])

        # --- Tabla A: boletas por sucursal x estado SII ---
        ws.append(['Boletas electrónicas — estado SII (fuente: Acepta)'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        fila_header_a = ws.max_row + 1
        self._estilo_header(ws, ['Sucursal', 'Estado SII', 'Cantidad', 'Monto'], fila=fila_header_a)

        agregados = {}
        for d in boletas:
            sucursal = d.sucursal.alias if d.sucursal_id else '(Sin sucursal)'
            info = acepta_por_folio.get(d.numero_documento)
            estado = info['estado_sii'] if info else 'Sin dato Acepta'
            clave = (sucursal, estado)
            cant, monto = agregados.get(clave, (0, 0))
            agregados[clave] = (cant + 1, monto + int(d.monto_con_iva or 0))

        for (sucursal, estado), (cant, monto) in sorted(agregados.items()):
            ws.append([sucursal, estado, cant, monto])
            ws.cell(row=ws.max_row, column=4).number_format = '#,##0'

        ws.append([])

        # --- Tabla B: otros documentos por sucursal x tipo ---
        ws.append(['Otros documentos (factura, NC, ND, guía, etc.) — estado interno del sistema'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        fila_header_b = ws.max_row + 1
        self._estilo_header(
            ws, ['Sucursal', 'Tipo Documento', 'Estado (interno)', 'Cantidad', 'Monto'],
            fila=fila_header_b,
        )

        agregados_b = {}
        for d in otros:
            sucursal = d.sucursal.alias if d.sucursal_id else '(Sin sucursal)'
            clave = (sucursal, d.tipo_documento, d.estado_dte)
            cant, monto = agregados_b.get(clave, (0, 0))
            agregados_b[clave] = (cant + 1, monto + int(d.monto_con_iva or 0))

        for (sucursal, tipo, estado), (cant, monto) in sorted(agregados_b.items()):
            ws.append([sucursal, tipo, estado, cant, monto])
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0'

        ws.append([])
        ws.append(['Notas:'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for nota in [
            '- "Sin dato Acepta": el folio no aparece en el archivo de Acepta pasado por '
            '--acepta-xls (puede ser de otra empresa/RUT, o el archivo no cubre ese rango).',
            '- El estado interno de Dte (columna "Estado (interno)") NO refleja el resultado '
            'real del SII para boletas; para boletas usar siempre la tabla de arriba (Acepta).',
            '- Ver hoja "Refacturación - Control" para el detalle de folios PAO4 refacturados '
            'y su estado pendiente de confirmar en un reporte de julio.',
        ]:
            ws.append([nota])

        self._autoancho(ws, [22, 26, 22, 12, 16])

    # ------------------------------------------------------------------ #
    def _hoja_boletas(self, ws, boletas, acepta_por_folio, mapeo_folio_nuevo):
        encabezados = [
            'Sucursal', 'Folio', 'Fecha Emisión', 'Monto', 'Estado (interno Dte)',
            'Estado SII (Acepta)', 'Info SII', 'Folio Nuevo (refacturación)',
        ]
        self._estilo_header(ws, encabezados)
        fill_rechazo = None
        from openpyxl.styles import PatternFill
        fill_rechazo = PatternFill('solid', fgColor=WARN_FILL)

        for d in boletas:
            info = acepta_por_folio.get(d.numero_documento)
            estado_sii = info['estado_sii'] if info else ''
            info_sii = info['informacion_sii'] if info else ''
            folio_nuevo = mapeo_folio_nuevo.get(d.numero_documento, '')
            ws.append([
                d.sucursal.alias if d.sucursal_id else '(Sin sucursal)',
                d.numero_documento,
                d.fecha_emision.strftime('%Y-%m-%d') if d.fecha_emision else '',
                int(d.monto_con_iva or 0),
                d.estado_dte,
                estado_sii,
                info_sii,
                folio_nuevo,
            ])
            ws.cell(row=ws.max_row, column=4).number_format = '#,##0'
            if estado_sii and 'RECHAZ' in str(estado_sii).upper():
                for col in range(1, len(encabezados) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = fill_rechazo

        self._autoancho(ws, [14, 10, 14, 14, 20, 26, 45, 14])

    # ------------------------------------------------------------------ #
    def _hoja_otros(self, ws, otros):
        encabezados = ['Sucursal', 'Tipo Documento', 'Folio', 'Fecha Emisión',
                       'Monto', 'Estado (interno)', 'Estado Pago']
        self._estilo_header(ws, encabezados)
        for d in otros:
            ws.append([
                d.sucursal.alias if d.sucursal_id else '(Sin sucursal)',
                d.tipo_documento,
                d.numero_documento,
                d.fecha_emision.strftime('%Y-%m-%d') if d.fecha_emision else '',
                int(d.monto_con_iva or 0),
                d.estado_dte,
                d.estado_pago,
            ])
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0'
        self._autoancho(ws, [14, 20, 10, 14, 14, 20, 14])

    # ------------------------------------------------------------------ #
    def _hoja_refacturacion(self, ws, mapeo_df):
        encabezados = list(mapeo_df.columns) + ['Confirmación julio']
        self._estilo_header(ws, encabezados)
        for _, row in mapeo_df.iterrows():
            valores = [row[c] for c in mapeo_df.columns]
            valores.append('PENDIENTE — falta reporte de Acepta que cubra julio')
            ws.append(valores)
        self._autoancho(ws, [6] + [16] * (len(encabezados) - 2) + [40])

    # ------------------------------------------------------------------ #
    @staticmethod
    def _rango_mes(texto):
        try:
            anio, mes = texto.strip().split('-')
            anio, mes = int(anio), int(mes)
            primer_dia = datetime.date(anio, mes, 1)
            ultimo_dia = datetime.date(anio, mes, monthrange(anio, mes)[1])
            return primer_dia, ultimo_dia
        except (ValueError, TypeError):
            raise CommandError(f'Mes inválido "{texto}". Formato: AAAA-MM.')

    def _cargar_acepta_boletas(self, ruta):
        """
        Lee el export de Acepta (.xls/.xlsx) y devuelve
        {folio:int -> {'estado_sii', 'monto_total', 'informacion_sii'}}.
        Si un folio se repite, gana la fila con estado RECHAZADO (mismo
        criterio que generar_txt_acepta_batch). Solo lectura.
        """
        try:
            import pandas as pd
        except ImportError:
            raise CommandError(
                'Falta pandas para leer el archivo de Acepta '
                '(pip install pandas; para .xls también xlrd).'
            )
        engine = 'xlrd' if ruta.lower().endswith('.xls') else None
        try:
            df = pd.read_excel(ruta, engine=engine)
        except ImportError:
            raise CommandError('Para leer .xls falta xlrd (ver requirements-dev.txt).')

        cols = {c.lower().strip(): c for c in df.columns}
        requeridas = ('folio', 'estado_sii')
        if any(r not in cols for r in requeridas):
            raise CommandError(
                f'El archivo de Acepta debe tener columnas {requeridas}. '
                f'Encontradas: {list(df.columns)}'
            )
        col_folio = cols['folio']
        col_estado = cols['estado_sii']
        col_monto = cols.get('monto_total')
        col_info = cols.get('informacion_sii')

        mapa = {}
        for _, row in df.iterrows():
            try:
                folio = int(row[col_folio])
            except (ValueError, TypeError):
                continue
            estado = str(row[col_estado]).strip()
            previo = mapa.get(folio)
            if previo is not None and 'RECHAZ' not in estado.upper():
                continue
            mapa[folio] = {
                'estado_sii': estado,
                'monto_total': row[col_monto] if col_monto else None,
                'informacion_sii': str(row[col_info]).strip() if col_info and pd.notna(row[col_info]) else '',
            }
        return mapa

    def _cargar_mapeo_refoliado(self, ruta):
        """
        Lee la hoja "Refoliado" del mapeo de refacturación y devuelve
        ({folio_viejo:int -> folio_nuevo:int}, DataFrame crudo sin la fila
        TOTAL, para volcar tal cual en la hoja de control). Solo lectura.
        """
        if not os.path.exists(ruta):
            raise CommandError(f'No existe el archivo de mapeo: {ruta}')
        try:
            import pandas as pd
        except ImportError:
            raise CommandError('Falta pandas para leer el archivo de mapeo.')

        df = pd.read_excel(ruta, sheet_name='Refoliado')
        df.columns = [str(c).strip() for c in df.columns]
        if 'Folio Viejo' not in df.columns or 'Folio Nuevo' not in df.columns:
            raise CommandError(
                'El mapeo debe tener columnas "Folio Viejo" y "Folio Nuevo". '
                f'Encontradas: {list(df.columns)}'
            )
        df = df.dropna(subset=['Folio Nuevo']).copy()
        df['Folio Viejo'] = df['Folio Viejo'].astype(int)
        df['Folio Nuevo'] = df['Folio Nuevo'].astype(int)

        mapa = dict(zip(df['Folio Viejo'], df['Folio Nuevo']))
        return mapa, df


def openpyxl_letter(i):
    from openpyxl.utils import get_column_letter
    return get_column_letter(i)
