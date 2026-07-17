# -*- coding: utf-8 -*-
"""Asigna género BEBÉ (AtributoOpcion id 779, atributo Sexo) a los artículos
confirmados en docs/candidatos_bebe_v12.xlsx (columna confirmar_bebe = SI).

Solo cambia atributo3 en TODAS las filas/bodegas del artículo. No toca categoría
ni nada más. DRY-RUN por defecto; --apply escribe. Requiere haber corrido
sembrar_taxonomia_v12 --apply (que crea la opción BEBÉ).

Uso:
    python manage.py asignar_genero_bebe                 # dry-run
    python manage.py asignar_genero_bebe --apply         # escribe
    python manage.py asignar_genero_bebe --excel ..\\docs\\candidatos_bebe_v12.xlsx
"""
import logging
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from app.models import Producto, AtributoOpcion

logger = logging.getLogger('app')

EXCEL_DEFAULT = os.path.join('..', 'docs', 'candidatos_bebe_v12.xlsx')


class Command(BaseCommand):
    help = "Asigna género BEBÉ a los artículos confirmados en el Excel (dry-run por defecto)"

    def add_arguments(self, parser):
        parser.add_argument('--excel', default=EXCEL_DEFAULT)
        parser.add_argument('--apply', action='store_true', help='Escribe en la BD')

    def handle(self, *args, **opts):
        from openpyxl import load_workbook
        apply_ = opts['apply']
        modo = 'APPLY' if apply_ else 'DRY-RUN'

        bebe = AtributoOpcion.objects.filter(valor__iexact='BEBÉ').first()
        if bebe is None:
            raise CommandError("No existe la opción de género BEBÉ — corre sembrar_taxonomia_v12 --apply")

        ruta = opts['excel']
        if not os.path.exists(ruta):
            raise CommandError(f"No existe el Excel: {ruta}")
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb['Candidatos_BEBE']
        it = ws.iter_rows(values_only=True)
        head = [str(h) if h is not None else '' for h in next(it)]
        idx = {h: i for i, h in enumerate(head)}
        if 'articulo' not in idx or 'confirmar_bebe' not in idx:
            raise CommandError("Faltan columnas 'articulo'/'confirmar_bebe'")

        articulos = []
        for row in it:
            a = str(row[idx['articulo']] or '').strip()
            conf = str(row[idx['confirmar_bebe']] or '').strip().upper()
            if a and conf in ('SI', 'SÍ', 'S', 'X', 'TRUE', '1'):
                articulos.append(a)

        n_art = n_prod = 0
        with transaction.atomic():
            for a in articulos:
                qs = Producto.objects.filter(articulo__iexact=a).exclude(atributo3=bebe)
                afectados = qs.count()
                if afectados:
                    n_art += 1
                    n_prod += afectados
                    if apply_:
                        qs.update(atributo3=bebe)
            if not apply_:
                transaction.set_rollback(True)

        self.stdout.write(self.style.MIGRATE_HEADING(f"[{modo}] asignar_genero_bebe"))
        self.stdout.write(f"  confirmados en Excel : {len(articulos)}")
        self.stdout.write(f"  artículos afectados  : {n_art}")
        self.stdout.write(f"  productos (filas)    : {n_prod}")
        self.stdout.write(self.style.SUCCESS(
            f"  {'ESCRITO' if apply_ else 'sin escribir (usa --apply)'}"))
        logger.info("asignar_genero_bebe %s: %d articulos, %d productos", modo, n_art, n_prod)
