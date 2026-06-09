"""
Re-categorización del catálogo (Paola / RealSport) — v2.

Propone una categoría para cada producto LÓGICO (deduplicado por `articulo`),
combinando cuatro señales:
  1) Categoría actual  -> mapeo determinista (capa 1)
  2) Palabra de tipo en `descripcion` + marca (capa 2)
  3) `Producto_Talla.talla` -> forma física (CALZADO / ROPA / BALON_ACC) para
     detectar MAL-CATEGORIZADOS (ej. camiseta talla "L" en categoría de calzado)
  4) Diccionario marca+modelo -> sub-tipo, resuelto por conocimiento + búsqueda
     web (lo produce el workflow de clasificación; se ingiere con --modelos)

FLUJO:
  - --export-modelos <json> : exporta los modelos únicos a clasificar (top 80%).
  - (sin args)              : dry-run; genera Excel de previsualización. NO escribe BD.
  - --modelos <json>        : ingiere las clasificaciones del workflow al pipeline.
  - --commit                : aplica (propaga por `articulo` a todas las sucursales).
  - --from-excel <xlsx>     : aplica la columna categoria_propuesta editada a mano.

Uso típico:
    python manage.py recategorizar_catalogo --export-modelos modelos.json
    # (clasificar modelos.json con el workflow -> modelos_clasificados.json)
    python manage.py recategorizar_catalogo --modelos modelos_clasificados.json
    python manage.py recategorizar_catalogo --modelos modelos_clasificados.json --commit
"""

import json
import logging
import re
import unicodedata
from collections import Counter, defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Sum
from django.utils import timezone

from app.models import Producto, Producto_Talla, Categoria
from ._data_recategorizacion import (
    TAXONOMIA, MAPEO_DIRECTO, REGLAS_KEYWORD, KEYWORDS_AMBIGUOS, REGLAS_MARCA,
    NO_PRODUCTO_EXACTO, NO_PRODUCTO_PREFIJOS, IDS_DEFAULT_ZAPATILLAS,
    GENERIC_DESC, SUBTIPO_A_NODO, FORMA_ESPERADA, CALZADO, OTROS,
)

logger = logging.getLogger('app')

EXCEL_HEADERS = [
    'articulo', 'descripcion', 'marca', 'sexo', 'n_filas', 'n_sucursales',
    'sucursales', 'stock_total', 'tallas_muestra', 'forma_talla',
    'categoria_original', 'inconsistente_bodegas', 'categoria_propuesta',
    'departamento_destino', 'capa_regla', 'fuente', 'confianza',
    'mal_categorizado', 'motivo', 'revisar',
]

# Categorías "residuo" cuyos productos pasan a la heurística / clasificación por modelo.
RESIDUAL_CATS = [1, 63, 73, 75, 78, 83, 79, 80, 81, 82]


# ───────────────────────── helpers de texto ─────────────────────────
def _fold(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s.upper().strip()


_KW_FOLDED = [([_fold(t) for t in toks], dep, tipo) for (toks, dep, tipo) in REGLAS_KEYWORD]
_KW_AMBIG_FOLDED = {_fold(t) for t in KEYWORDS_AMBIGUOS}
_NOPROD_EXACT_FOLDED = {_fold(t) for t in NO_PRODUCTO_EXACTO}
_NOPROD_PREFIX_FOLDED = tuple(_fold(t) for t in NO_PRODUCTO_PREFIJOS)
_MARCA_FOLDED = {_fold(k): v for k, v in REGLAS_MARCA.items()}
_GENERIC_FOLDED = {_fold(t) for t in GENERIC_DESC}


# ───────────────────────── helpers de talla ─────────────────────────
_RE_SHOE_DEC = re.compile(r'^\d{1,2}[,.]\d$')      # 8,0  9,5  10,5
_RE_SHOE_RANGE = re.compile(r'^\d{2}/\d{2}$')      # 39/40
_RE_INT = re.compile(r'^\d{1,3}$')
_ALPHA_ROPA = {'XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL'}
_TALLA_NULA = {'', '0', '00', '000', 'UNICA', 'U', 'S/T', 'ST', 'N/A', 'SIN TALLA'}


def forma_fisica(tallas):
    """Clasifica la forma física del producto según sus tallas reales."""
    shoe = ropa = other = 0
    for t in tallas:
        tu = (t or '').strip().upper()
        if tu in _TALLA_NULA:
            continue
        if tu in _ALPHA_ROPA:
            ropa += 1
        elif _RE_SHOE_DEC.match(tu) or _RE_SHOE_RANGE.match(tu):
            shoe += 1
        elif _RE_INT.match(tu):
            v = int(tu)
            if 18 <= v <= 47:
                shoe += 1            # talla de calzado (infantil 18 a adulto 47)
            else:
                other += 1           # 1-17 (años?) o 100/700/850 (ruido/balón)
        else:
            other += 1
    if shoe and shoe >= ropa:
        return 'CALZADO'
    if ropa and ropa > shoe:
        return 'ROPA'
    return 'INDEFINIDO'


def clasificar(cat_id, descripcion, marca):
    """Capa 1 + Capa 2. Devuelve (dep, tipo, regla, revisar, excluir)."""
    if cat_id in MAPEO_DIRECTO:
        dep, tipo, rev = MAPEO_DIRECTO[cat_id]
        return dep, tipo, 'mapeo_directo', rev, False

    desc_f = _fold(descripcion)
    marca_f = _fold(marca)

    if desc_f in _NOPROD_EXACT_FOLDED or desc_f.startswith(_NOPROD_PREFIX_FOLDED):
        return OTROS, 'No-Producto', 'no_producto', False, True

    for toks, dep, tipo in _KW_FOLDED:
        for tok in toks:
            if tok and tok in desc_f:
                return dep, tipo, 'keyword', (tok in _KW_AMBIG_FOLDED), False

    if marca_f in _MARCA_FOLDED:
        dep, tipo = _MARCA_FOLDED[marca_f]
        return dep, tipo, 'marca', True, False

    if cat_id in IDS_DEFAULT_ZAPATILLAS:
        return CALZADO, 'Zapatillas Urbanas', 'default_zapatillas', True, False

    return None, None, 'sin_resolver', True, False


class Command(BaseCommand):
    help = 'Re-categoriza el catálogo (talla + categoría + keyword + modelo web). Dry-run genera Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--commit', action='store_true')
        parser.add_argument('--fase', choices=['taxonomia', 'capa1', 'capa2', 'limpieza'], default=None)
        parser.add_argument('--from-excel', dest='from_excel', default=None)
        parser.add_argument('--modelos', default=None,
                            help='JSON de clasificaciones marca+modelo del workflow.')
        parser.add_argument('--export-modelos', dest='export_modelos', default=None,
                            help='Exporta los modelos únicos a clasificar (top 80%) y sale.')
        parser.add_argument('--cobertura', type=float, default=0.80,
                            help='Cobertura objetivo para --export-modelos (0-1).')
        parser.add_argument('--out', default=None)
        parser.add_argument('--limit', type=int, default=0)

    # ──────────────────────────────────────────────────────────────────
    def handle(self, *args, **opts):
        self.stdout.write(self.style.SUCCESS('=' * 78))
        self.stdout.write(self.style.SUCCESS('  RE-CATEGORIZACIÓN DEL CATÁLOGO (v2: talla + modelo web)'))
        self.stdout.write(self.style.SUCCESS('=' * 78))

        if opts['export_modelos']:
            return self._exportar_modelos(opts['export_modelos'], opts['cobertura'])

        if opts['from_excel']:
            if not opts['commit']:
                raise CommandError('--from-excel requiere --commit.')
            return self._aplicar_desde_excel(opts['from_excel'], opts['fase'])

        model_dict = self._cargar_modelos(opts['modelos']) if opts['modelos'] else {}
        if model_dict:
            self.stdout.write(f'  Modelos clasificados cargados: {len(model_dict):,}')
        if not opts['commit']:
            self.stdout.write(self.style.WARNING('  DRY-RUN (no escribe BD). Se generará el Excel.'))

        self.stdout.write('\n[1/3] Agregando productos por `articulo` + tallas...')
        articulos = self._construir_dataset(limit=opts['limit'])
        self.stdout.write(f'      {len(articulos):,} artículos lógicos.')

        self.stdout.write('[2/3] Clasificando (capa1 + capa2 + talla + modelo)...')
        filas, resumen = self._clasificar_todos(articulos, model_dict)

        if opts['commit']:
            self.stdout.write('[3/3] Aplicando en la BD...')
            self._commit(filas, opts['fase'])
        else:
            self.stdout.write('[3/3] Escribiendo Excel...')
            ruta = self._exportar_excel(filas, opts['out'])
            self.stdout.write(self.style.SUCCESS(f'      Excel: {ruta}'))

        self._imprimir_resumen(resumen, len(filas))

    # ──────────────────────────────────────────────────────────────────
    def _construir_dataset(self, limit=0):
        agg = {}
        qs = Producto.objects.values_list(
            'articulo', 'descripcion', 'categoria_id', 'categoria__nombre',
            'atributo1__valor', 'atributo3__valor', 'sucursal__alias',
        ).iterator(chunk_size=5000)
        for articulo, desc, cat_id, cat_nom, marca, sexo, suc in qs:
            a = agg.get(articulo)
            if a is None:
                a = agg[articulo] = {
                    'desc': Counter(), 'marca': Counter(), 'sexo': Counter(),
                    'cat_nom': Counter(), 'cat_id': Counter(), 'suc': set(),
                    'n_filas': 0, 'tallas': set(),
                }
            a['n_filas'] += 1
            if desc:
                a['desc'][desc] += 1
            if marca:
                a['marca'][marca] += 1
            if sexo:
                a['sexo'][sexo] += 1
            a['cat_nom'][cat_nom or '(sin categoría)'] += 1
            a['cat_id'][cat_id] += 1
            if suc:
                a['suc'].add(suc)

        # Tallas por artículo.
        for art, talla in Producto_Talla.objects.values_list('producto__articulo', 'talla').iterator(chunk_size=10000):
            a = agg.get(art)
            if a is not None and talla:
                a['tallas'].add(talla)

        # Stock total por artículo.
        stock_map = defaultdict(int)
        for r in Producto_Talla.objects.values('producto__articulo').annotate(s=Sum('stock')):
            stock_map[r['producto__articulo']] = r['s'] or 0
        for art, a in agg.items():
            a['stock_total'] = stock_map.get(art, 0)

        if limit:
            return dict(list(agg.items())[:limit])
        return agg

    def _rep(self, counter):
        return counter.most_common(1)[0][0] if counter else ''

    def _clasificar_todos(self, articulos, model_dict):
        filas = []
        resumen = {
            'por_departamento': Counter(), 'por_propuesta': Counter(),
            'por_regla': Counter(), 'por_forma': Counter(),
            'revisar': 0, 'sin_resolver': 0, 'inconsistentes': 0, 'mal_categorizados': 0,
        }
        for articulo, a in articulos.items():
            desc = self._rep(a['desc'])
            marca = self._rep(a['marca'])
            sexo = self._rep(a['sexo'])
            cat_nom = self._rep(a['cat_nom'])
            cat_id = self._rep(a['cat_id'])
            ids_distintos = {k for k in a['cat_id'] if k is not None}
            inconsistente = len(ids_distintos) > 1
            forma = forma_fisica(a['tallas'])

            dep, tipo, regla, revisar, excluir = clasificar(cat_id, desc, marca)
            fuente = 'reglas'
            confianza = 'alta' if regla in ('mapeo_directo', 'keyword', 'no_producto') else 'media'

            # Refinar el residuo con el diccionario de modelos (conocimiento/web).
            if regla in ('default_zapatillas', 'sin_resolver') and desc:
                key = (_fold(marca), _fold(desc))
                if _fold(desc) not in _GENERIC_FOLDED and key in model_dict:
                    m = model_dict[key]
                    dep, tipo = m['dep'], m['tipo']
                    fuente = m.get('fuente', 'modelo')
                    confianza = m.get('confianza', 'media')
                    regla = 'modelo'
                    revisar = confianza != 'alta'

            # Detección de mal-categorizado por talla (CALZADO vs ROPA).
            mal, motivo = '', ''
            forma_esp = FORMA_ESPERADA.get(dep)
            if forma_esp and forma in ('CALZADO', 'ROPA') and forma != forma_esp:
                mal, motivo = 'Sí', f'talla indica {forma} pero propuesta es {dep}'
                revisar = True

            propuesta = f'{dep} / {tipo}' if dep and tipo else '(SIN RESOLVER)'
            tallas_muestra = ', '.join(sorted(t for t in a['tallas'] if (t or '').strip() not in _TALLA_NULA)[:8])

            filas.append({
                'articulo': articulo, 'descripcion': desc, 'marca': marca, 'sexo': sexo,
                'n_filas': a['n_filas'], 'n_sucursales': len(a['suc']),
                'sucursales': ', '.join(sorted(a['suc'])), 'stock_total': a['stock_total'],
                'tallas_muestra': tallas_muestra, 'forma_talla': forma,
                'categoria_original': cat_nom,
                'inconsistente_bodegas': 'Sí' if inconsistente else '',
                'categoria_propuesta': propuesta, 'departamento_destino': dep or '',
                'capa_regla': regla, 'fuente': fuente, 'confianza': confianza,
                'mal_categorizado': mal, 'motivo': motivo,
                'revisar': 'Sí' if revisar else '',
                '_dep': dep, '_tipo': tipo, '_excluir': excluir,
            })
            resumen['por_departamento'][dep or '(sin resolver)'] += 1
            resumen['por_propuesta'][propuesta] += 1
            resumen['por_regla'][regla] += 1
            resumen['por_forma'][forma] += 1
            resumen['revisar'] += 1 if revisar else 0
            resumen['sin_resolver'] += 1 if regla == 'sin_resolver' else 0
            resumen['inconsistentes'] += 1 if inconsistente else 0
            resumen['mal_categorizados'] += 1 if mal else 0
        return filas, resumen

    # ──────────────────────── export de modelos ────────────────────────
    def _exportar_modelos(self, ruta, cobertura):
        from django.db.models import Count
        base = (Producto.objects.filter(categoria_id__in=[1, 63, 83])
                .exclude(descripcion=''))
        total = base.count()
        modelos = list(base.values('atributo1__valor', 'descripcion')
                       .annotate(n=Count('id')).order_by('-n'))
        seleccion, acc = [], 0
        for m in modelos:
            desc = (m['descripcion'] or '').strip()
            if _fold(desc) in _GENERIC_FOLDED:
                continue   # genéricos: no se buscan, quedan Zapatillas Urbanas
            seleccion.append({'marca': m['atributo1__valor'] or '', 'modelo': desc, 'n': m['n']})
            acc += m['n']
            if acc >= total * cobertura:
                break
        with open(ruta, 'w', encoding='utf-8') as fh:
            json.dump(seleccion, fh, ensure_ascii=False, indent=1)
        self.stdout.write(self.style.SUCCESS(
            f'  Exportados {len(seleccion):,} modelos (cubren ~{cobertura:.0%} de {total:,} filas) -> {ruta}'))

    def _cargar_modelos(self, ruta):
        with open(ruta, encoding='utf-8') as fh:
            data = json.load(fh)
        d = {}
        for m in data:
            dep = m.get('dep') or m.get('departamento')
            tipo = m.get('tipo')
            if not dep or not tipo:
                # Resolver desde sub-tipo libre.
                sub = (m.get('subtipo') or m.get('sub') or '').strip().lower()
                dep, tipo = SUBTIPO_A_NODO.get(sub, (CALZADO, 'Zapatillas Urbanas'))
            d[(_fold(m.get('marca')), _fold(m.get('modelo')))] = {
                'dep': dep, 'tipo': tipo,
                'fuente': m.get('fuente', 'modelo'),
                'confianza': m.get('confianza', 'media'),
            }
        return d

    # ──────────────────────────── salida ───────────────────────────────
    def _exportar_excel(self, filas, out):
        try:
            from openpyxl import Workbook
        except ImportError:
            raise CommandError('openpyxl no está instalado (pip install openpyxl).')
        if not out:
            from django.conf import settings
            fecha = timezone.now().strftime('%Y%m%d_%H%M')
            out = str(settings.BASE_DIR / f'recategorizacion_preview_{fecha}.xlsx')
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Recategorizacion')
        ws.append(EXCEL_HEADERS)
        for f in sorted(filas, key=lambda x: (
                x['mal_categorizado'] != 'Sí', x['revisar'] != 'Sí',
                x['departamento_destino'], x['categoria_propuesta'], -x['stock_total'])):
            ws.append([f[h] for h in EXCEL_HEADERS])
        wb.save(out)
        return out

    def _imprimir_resumen(self, resumen, total):
        s = self.stdout
        s.write('\n' + '=' * 78)
        s.write(self.style.SUCCESS(f'  RESUMEN  ({total:,} artículos lógicos)'))
        s.write('=' * 78)
        s.write('\n  Por departamento destino:')
        for dep, n in resumen['por_departamento'].most_common():
            s.write(f'    {dep:<20} {n:>8,}')
        s.write('\n  Por regla / fuente:')
        for regla, n in resumen['por_regla'].most_common():
            s.write(f'    {regla:<20} {n:>8,}')
        s.write('\n  Por forma física (talla):')
        for forma, n in resumen['por_forma'].most_common():
            s.write(f'    {forma:<20} {n:>8,}')
        s.write('\n  Top 12 categorías propuestas:')
        for prop, n in resumen['por_propuesta'].most_common(12):
            s.write(f'    {prop:<32} {n:>8,}')
        s.write('')
        s.write(self.style.WARNING(f'  A revisar manualmente : {resumen["revisar"]:,}'))
        s.write(self.style.WARNING(f'  Sin resolver          : {resumen["sin_resolver"]:,}'))
        s.write(self.style.WARNING(f'  Mal-categorizados     : {resumen["mal_categorizados"]:,}'))
        s.write(self.style.WARNING(f'  Inconsistentes/bodega : {resumen["inconsistentes"]:,}'))
        s.write('=' * 78)

    # ──────────────────────────── commit ───────────────────────────────
    def _ensure_taxonomia(self):
        nodos = {}
        for dep, tipos in TAXONOMIA.items():
            raiz, _ = Categoria.objects.get_or_create(nombre=dep, padre=None)
            nodos[(dep, None)] = raiz.id
            for tipo in tipos:
                hijo, _ = Categoria.objects.get_or_create(nombre=tipo, padre=raiz)
                nodos[(dep, tipo)] = hijo.id
        return nodos

    def _commit(self, filas, fase):
        with transaction.atomic():
            nodos = self._ensure_taxonomia()
            if fase == 'taxonomia':
                self.stdout.write(f'      Taxonomía: {len(nodos)} nodos.')
                return
            if fase in (None, 'capa1', 'capa2'):
                reglas_capa1 = {'mapeo_directo'}
                reglas_capa2 = {'keyword', 'marca', 'default_zapatillas', 'no_producto', 'modelo'}
                aplicar = (reglas_capa1 | reglas_capa2) if fase is None else (
                    reglas_capa1 if fase == 'capa1' else reglas_capa2)
                por_destino = defaultdict(list)
                for f in filas:
                    if f['capa_regla'] in aplicar and f['_dep'] and f['_tipo']:
                        por_destino[(f['_dep'], f['_tipo'], f['_excluir'])].append(f['articulo'])
                total = 0
                for (dep, tipo, excluir), arts in por_destino.items():
                    cat_id = nodos[(dep, tipo)]
                    for i in range(0, len(arts), 1000):
                        upd = {'categoria_id': cat_id}
                        if excluir:
                            upd['excluir_de_analitica'] = True
                        total += Producto.objects.filter(articulo__in=arts[i:i + 1000]).update(**upd)
                self.stdout.write(self.style.SUCCESS(f'      Filas actualizadas: {total:,}'))
            if fase in (None, 'limpieza'):
                self._deprecar_categorias_vacias()

    def _deprecar_categorias_vacias(self):
        protegidas = set()
        for dep in TAXONOMIA:
            try:
                protegidas.add(Categoria.objects.get(nombre=dep, padre=None).id)
            except Categoria.DoesNotExist:
                pass
        n = 0
        for cat in Categoria.objects.filter(padre__isnull=True).exclude(id__in=protegidas):
            if cat.nombre.startswith('_ZZ_'):
                continue
            if (not Producto.objects.filter(categoria_id=cat.id).exists()
                    and not Categoria.objects.filter(padre_id=cat.id).exists()):
                cat.nombre = f'_ZZ_{cat.nombre}'[:100]
                cat.save(update_fields=['nombre'])
                n += 1
        self.stdout.write(self.style.SUCCESS(f'      Categorías viejas deprecadas: {n}'))

    def _aplicar_desde_excel(self, ruta, fase):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl no está instalado.')
        wb = load_workbook(ruta, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else '' for h in next(rows)]
        try:
            i_art = header.index('articulo')
            i_prop = header.index('categoria_propuesta')
        except ValueError:
            raise CommandError("El Excel debe tener columnas 'articulo' y 'categoria_propuesta'.")
        with transaction.atomic():
            nodos = self._ensure_taxonomia()
            por_destino = defaultdict(list)
            ignoradas = 0
            for row in rows:
                art = row[i_art]
                prop = row[i_prop] if isinstance(row[i_prop], str) else ''
                if not art or '/' not in prop:
                    ignoradas += 1
                    continue
                dep, _, tipo = prop.partition('/')
                key = (dep.strip(), tipo.strip())
                if key not in nodos:
                    ignoradas += 1
                    continue
                por_destino[key].append(str(art))
            total = 0
            for (dep, tipo), arts in por_destino.items():
                cat_id = nodos[(dep, tipo)]
                for i in range(0, len(arts), 1000):
                    total += Producto.objects.filter(articulo__in=arts[i:i + 1000]).update(categoria_id=cat_id)
            self.stdout.write(self.style.SUCCESS(f'      Actualizadas desde Excel: {total:,} (ignoradas: {ignoradas:,})'))
            if fase in (None, 'limpieza'):
                self._deprecar_categorias_vacias()
