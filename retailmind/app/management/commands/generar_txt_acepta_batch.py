"""
Genera archivos TXT de Acepta en LOTE (por mes, rango de fechas, rango de
folios o total), replicando exactamente el TXT que descarga el botón
"Acciones" de la vista de DTE (endpoint generar_txt_desde_dte_existente).

Es un comando de SOLO LECTURA: no escribe nada en la base de datos, solo lee
DTEs y escribe archivos .txt (o un .zip) a disco. Reutiliza las mismas
funciones "fuente única de verdad" que usa la vista web, de modo que cada TXT
es byte-idéntico al del botón:

    construir_datos_txt_desde_dte(dte)  -> dict
    generar_txt_dte_acepta(datos)       -> str (contenido TXT)
    validar_dte_antes_de_guardar(datos) -> dict con 'warnings'

Uso (ejemplos):

    # Un solo folio (para comparar contra el botón web)
    python manage.py generar_txt_acepta_batch \
        --sucursal PAO4 --tipo "BOLETA ELECTRONICA" \
        --folio-desde 355001 --folio-hasta 355001

    # Un bloque de folios acotado, empaquetado en un ZIP
    python manage.py generar_txt_acepta_batch \
        --sucursal PAO4 --folio-desde 355001 --folio-hasta 355368 --zip

    # Todas las boletas electrónicas de PAO4 de junio 2026
    python manage.py generar_txt_acepta_batch \
        --sucursal PAO4 --tipo "BOLETA ELECTRONICA" --mes 2026-06

    # Ver qué se generaría sin escribir archivos
    python manage.py generar_txt_acepta_batch --mes 2026-06 --dry-run
"""

import datetime
import logging
import os
import zipfile
from calendar import monthrange

from django.core.management.base import BaseCommand, CommandError

from app.models import Dte, Sucursal

logger = logging.getLogger('app')

# Tipos que NO generan TXT de Acepta (mismo criterio que la vista:
# generar_txt_desde_dte_existente rechaza 'BOLETA PAPEL').
TIPOS_SIN_TXT = {'BOLETA PAPEL'}

# Rango de folios afectados por el incidente de reemisión conocido de esta
# instalación: dos sucursales que comparten RUT emitieron el mismo bloque de
# boletas, porque cada ficha de `Empresa` corre su propio correlativo y el SII
# solo ve el RUT. Se usa para DESTACARLO en el resumen; no filtra nada.
# Configurable por si aparece otro bloque.
BLOQUE_FOLIOS_REEMITIDOS = (355001, 355368)


class Command(BaseCommand):
    help = (
        'Genera TXT de Acepta en lote (por mes, rango de fechas, rango de '
        'folios o total). Solo lectura: no modifica la base de datos.'
    )

    def add_arguments(self, parser):
        # --- Filtros de fecha (mutuamente relacionados) ---
        parser.add_argument(
            '--mes', type=str, default=None,
            help='Mes a exportar en formato AAAA-MM (ej. 2026-06).',
        )
        parser.add_argument(
            '--desde', type=str, default=None,
            help='Fecha inicial AAAA-MM-DD (alternativa a --mes).',
        )
        parser.add_argument(
            '--hasta', type=str, default=None,
            help='Fecha final AAAA-MM-DD (alternativa a --mes).',
        )
        parser.add_argument(
            '--total', action='store_true',
            help='Exportar TODOS los DTEs sin filtro de fecha (mucho volumen).',
        )

        # --- Filtros adicionales ---
        parser.add_argument(
            '--sucursal', type=str, default=None,
            help='Alias de la sucursal (ej. PAO4). Opcional.',
        )
        parser.add_argument(
            '--tipo', type=str, default=None,
            help='tipo_documento exacto (ej. "BOLETA ELECTRONICA"). '
                 'Por defecto todos los tipos que generan TXT.',
        )
        parser.add_argument(
            '--folio-desde', type=int, default=None,
            help='numero_documento mínimo (inclusive).',
        )
        parser.add_argument(
            '--folio-hasta', type=int, default=None,
            help='numero_documento máximo (inclusive).',
        )

        # --- Salida ---
        parser.add_argument(
            '--salida', type=str, default=None,
            help='Carpeta destino. Por defecto ./exports_txt/<filtro>/.',
        )
        parser.add_argument(
            '--zip', action='store_true',
            help='Empaquetar todos los TXT en un solo archivo .zip.',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Solo contar y listar; no escribe archivos.',
        )

        # --- Refoliado / correcciones "por fuera del sistema" ---
        # NO tocan la BD: solo cambian folio y fecha en el TXT/Excel de salida.
        parser.add_argument(
            '--folio-nuevo-desde', type=int, default=None,
            help='Renumera las boletas seleccionadas con folios NUEVOS '
                 'secuenciales a partir de este número, en orden de folio '
                 'viejo (ej. 3000 -> 3000, 3001, 3002...). No modifica la BD.',
        )
        parser.add_argument(
            '--fecha-nueva', type=str, default=None,
            help='Fuerza esta fecha (AAAA-MM-DD) en todos los documentos de '
                 'salida (ej. 2026-06-30). No modifica la BD.',
        )
        parser.add_argument(
            '--excel', nargs='?', const='__DEFAULT__', default=None,
            help='Genera un Excel con el mapeo folio_viejo -> folio_nuevo, '
                 'fecha nueva y monto. Si se pasa solo --excel escribe en la '
                 'carpeta de salida; opcionalmente indica una ruta .xlsx. '
                 'En esta etapa NO genera los TXT.',
        )
        parser.add_argument(
            '--cruzar-sii', type=str, default=None,
            help='Ruta a un .xls/.xlsx exportado del SII (con columnas folio y '
                 'estado_sii). En el Excel de mapeo agrega columnas "Estado '
                 'SII" y "En archivo SII" marcando qué folios viejos aparecen '
                 'rechazados. Requiere --excel. Solo lectura.',
        )

    # ------------------------------------------------------------------ #
    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.usar_zip = options['zip']

        # Parámetros de refoliado (por fuera del sistema).
        self.folio_nuevo_desde = options['folio_nuevo_desde']
        self.fecha_nueva = (
            self._parse_fecha(options['fecha_nueva'])
            if options['fecha_nueva'] else None
        )
        self.solo_excel = options['excel'] is not None

        # Cruce opcional contra el archivo de estados del SII.
        self.sii_por_folio = None
        if options['cruzar_sii']:
            if not self.solo_excel:
                raise CommandError('--cruzar-sii requiere --excel.')
            self.sii_por_folio = self._cargar_estados_sii(options['cruzar_sii'])

        # Import diferido de las funciones de la vista: viven en
        # views_modulo_documentos y arrastran dependencias de vista; se
        # importan aquí para no cargarlas si el command falla antes.
        from app.views_modulo_documentos import (
            construir_datos_txt_desde_dte,
            generar_txt_dte_acepta,
            validar_dte_antes_de_guardar,
        )
        self._construir = construir_datos_txt_desde_dte
        self._generar = generar_txt_dte_acepta
        self._validar = validar_dte_antes_de_guardar

        queryset, etiqueta = self._construir_queryset(options)

        total = queryset.count()
        if total == 0:
            self.stdout.write(self.style.WARNING(
                'No se encontraron DTEs con los filtros indicados.'
            ))
            return

        # Carpeta / zip de salida
        salida_dir = options['salida'] or os.path.join('exports_txt', etiqueta)
        self.stdout.write('=' * 70)
        self.stdout.write(f'[GENERAR TXT ACEPTA BATCH]  filtro: {etiqueta}')
        self.stdout.write(f'  DTEs seleccionados: {total}')
        if self.folio_nuevo_desde is not None:
            self.stdout.write(self.style.WARNING(
                f'  Refoliado: folio nuevo desde {self.folio_nuevo_desde} '
                f'(secuencial por folio viejo). NO modifica la BD.'
            ))
        if self.fecha_nueva is not None:
            self.stdout.write(self.style.WARNING(
                f'  Fecha forzada: {self.fecha_nueva.strftime("%Y-%m-%d")} '
                f'en todos los documentos. NO modifica la BD.'
            ))
        self.stdout.write('=' * 70)

        # Modo Excel de mapeo: solo genera el Excel, NO los TXT (etapa previa
        # para revisar folio viejo -> nuevo, fecha y monto antes de emitir).
        if self.solo_excel:
            self._generar_excel_mapeo(queryset, salida_dir, etiqueta, options['excel'])
            return

        if self.dry_run:
            self.stdout.write(self.style.WARNING(
                '  [DRY-RUN] No se escribirá ningún archivo.'
            ))
        else:
            self.stdout.write(f'  Salida: {os.path.abspath(salida_dir)}'
                              f'{".zip" if self.usar_zip else "/"}')

        self._procesar(queryset, salida_dir, etiqueta)

    # ------------------------------------------------------------------ #
    def _construir_queryset(self, options):
        """Arma el queryset de Dte según filtros y devuelve (qs, etiqueta)."""
        qs = Dte.objects.select_related('emisor', 'receptor', 'sucursal', 'vendedor')
        partes_etiqueta = []

        # --- Sucursal ---
        if options['sucursal']:
            alias = options['sucursal'].strip()
            sucs = list(Sucursal.objects.filter(alias=alias))
            if not sucs:
                raise CommandError(f'No existe la sucursal con alias "{alias}".')
            qs = qs.filter(sucursal__in=sucs)
            partes_etiqueta.append(alias)

        # --- Tipo de documento ---
        if options['tipo']:
            tipo = options['tipo'].strip()
            qs = qs.filter(tipo_documento=tipo)
            partes_etiqueta.append(tipo.replace(' ', '_'))
        # Siempre excluir los tipos que no generan TXT.
        qs = qs.exclude(tipo_documento__in=TIPOS_SIN_TXT)

        # --- Rango de folios ---
        if options['folio_desde'] is not None:
            qs = qs.filter(numero_documento__gte=options['folio_desde'])
        if options['folio_hasta'] is not None:
            qs = qs.filter(numero_documento__lte=options['folio_hasta'])
        if options['folio_desde'] is not None or options['folio_hasta'] is not None:
            partes_etiqueta.append(
                f'folios_{options["folio_desde"] or "min"}-{options["folio_hasta"] or "max"}'
            )

        # --- Fecha: --mes | --desde/--hasta | --total ---
        modos_fecha = [
            bool(options['mes']),
            bool(options['desde'] or options['hasta']),
            bool(options['total']),
        ]
        if sum(modos_fecha) > 1:
            raise CommandError(
                'Usa solo uno de: --mes, --desde/--hasta, o --total.'
            )

        if options['mes']:
            desde, hasta = self._rango_mes(options['mes'])
            qs = qs.filter(fecha_emision__gte=desde, fecha_emision__lte=hasta)
            partes_etiqueta.append(options['mes'])
        elif options['desde'] or options['hasta']:
            desde = self._parse_fecha(options['desde']) if options['desde'] else None
            hasta = self._parse_fecha(options['hasta']) if options['hasta'] else None
            if desde:
                qs = qs.filter(fecha_emision__gte=desde)
            if hasta:
                qs = qs.filter(fecha_emision__lte=hasta)
            partes_etiqueta.append(
                f'{options["desde"] or "min"}_a_{options["hasta"] or "max"}'
            )
        elif options['total']:
            partes_etiqueta.append('total')
        else:
            # Sin filtro de fecha explícito y sin --total: exige acotar,
            # salvo que ya haya un filtro de folios que limite el volumen.
            if options['folio_desde'] is None and options['folio_hasta'] is None:
                raise CommandError(
                    'Debes acotar la exportación: usa --mes AAAA-MM, '
                    '--desde/--hasta, un rango de --folio-desde/--folio-hasta, '
                    'o --total para exportar todo.'
                )

        etiqueta = '__'.join(partes_etiqueta) if partes_etiqueta else 'export'
        qs = qs.order_by('sucursal_id', 'tipo_documento', 'numero_documento')
        return qs, etiqueta

    # ------------------------------------------------------------------ #
    def _generar_excel_mapeo(self, queryset, salida_dir, etiqueta, opcion_excel):
        """
        Escribe un Excel con el mapeo de refoliado:
            folio_viejo -> folio_nuevo | fecha_nueva | monto | ...

        El folio nuevo se asigna secuencial en el mismo orden del queryset
        (sucursal, tipo, folio viejo ascendente). SOLO LECTURA de la BD.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Refoliado'

        cruzar = self.sii_por_folio is not None
        encabezados = [
            'N°', 'Sucursal', 'Tipo Documento', 'Folio Viejo', 'Folio Nuevo',
            'Fecha Original', 'Fecha Nueva', 'Monto', 'Estado', 'Receptor RUT',
            'DTE ID',
        ]
        if cruzar:
            encabezados += ['En archivo SII', 'Estado SII']
        ws.append(encabezados)
        cabecera_fill = PatternFill('solid', fgColor='405189')
        cabecera_font = Font(bold=True, color='FFFFFF')
        for col, _ in enumerate(encabezados, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = cabecera_fill
            c.font = cabecera_font
            c.alignment = Alignment(horizontal='center')

        # Resaltado para filas cuyo folio NO aparece en el archivo del SII
        # (útil para ver de un vistazo las que quedaron fuera del rechazo).
        fill_no_sii = PatternFill('solid', fgColor='FFF3CD')  # amarillo suave

        fecha_nueva_str = (
            self.fecha_nueva.strftime('%Y-%m-%d') if self.fecha_nueva else ''
        )
        folio_nuevo = self.folio_nuevo_desde
        total_monto = 0
        fila = 0
        en_sii = 0
        rechazadas = 0

        for dte in queryset.iterator(chunk_size=500):
            fila += 1
            fnuevo = folio_nuevo + (fila - 1) if folio_nuevo is not None else ''
            monto = int(dte.monto_con_iva) if dte.monto_con_iva is not None else 0
            total_monto += monto
            receptor_rut = getattr(dte.receptor, 'rut', '') if dte.receptor_id else ''
            valores = [
                fila,
                dte.sucursal.alias if dte.sucursal_id else '',
                dte.tipo_documento,
                dte.numero_documento,
                fnuevo,
                dte.fecha_emision.strftime('%Y-%m-%d') if dte.fecha_emision else '',
                fecha_nueva_str,
                monto,
                dte.estado_dte,
                receptor_rut,
                dte.id,
            ]
            estado_sii = None
            if cruzar:
                estado_sii = self.sii_por_folio.get(dte.numero_documento)
                esta = estado_sii is not None
                valores += ['SÍ' if esta else 'NO', estado_sii or '']
                if esta:
                    en_sii += 1
                    if 'RECHAZ' in str(estado_sii).upper():
                        rechazadas += 1
            ws.append(valores)
            # Resaltar en amarillo las filas que NO están en el archivo SII.
            if cruzar and estado_sii is None:
                for col in range(1, len(valores) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = fill_no_sii

        # Fila de totales
        ws.append([])
        fila_total = ws.max_row + 1
        ws.cell(row=fila_total, column=3, value='TOTAL').font = Font(bold=True)
        ws.cell(row=fila_total, column=4, value=fila)  # cantidad de docs
        c_total = ws.cell(row=fila_total, column=8, value=total_monto)
        c_total.font = Font(bold=True)

        # Formato de miles en columna Monto (H) y ancho de columnas
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=8).number_format = '#,##0'
        anchos = [5, 10, 20, 12, 12, 14, 14, 14, 14, 16, 10]
        if cruzar:
            anchos += [14, 22]
        for i, w in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        # Ruta destino
        if opcion_excel and opcion_excel != '__DEFAULT__':
            ruta = opcion_excel
            if not ruta.lower().endswith('.xlsx'):
                ruta += '.xlsx'
        else:
            ruta = os.path.abspath(salida_dir) + f'__mapeo_refoliado.xlsx'

        os.makedirs(os.path.dirname(os.path.abspath(ruta)) or '.', exist_ok=True)
        wb.save(ruta)

        self.stdout.write('\n' + '-' * 70)
        self.stdout.write('EXCEL DE MAPEO GENERADO')
        self.stdout.write('-' * 70)
        self.stdout.write(self.style.SUCCESS(f'  Documentos: {fila}'))
        self.stdout.write(self.style.SUCCESS(f'  Monto total: {total_monto:,}'))
        if folio_nuevo is not None:
            self.stdout.write(
                f'  Folios nuevos: {folio_nuevo} .. {folio_nuevo + fila - 1}'
            )
        else:
            self.stdout.write(self.style.WARNING(
                '  (Sin --folio-nuevo-desde: la columna Folio Nuevo va vacía.)'
            ))
        if not fecha_nueva_str:
            self.stdout.write(self.style.WARNING(
                '  (Sin --fecha-nueva: la columna Fecha Nueva va vacía.)'
            ))
        if cruzar:
            fuera = fila - en_sii
            self.stdout.write('')
            self.stdout.write('  Cruce con archivo SII:')
            self.stdout.write(f'     En el archivo SII: {en_sii}')
            self.stdout.write(self.style.WARNING(
                f'     RECHAZADAS por el SII: {rechazadas}'
            ))
            self.stdout.write(
                f'     NO están en el archivo SII (fila amarilla): {fuera}'
            )
        self.stdout.write(f'  Archivo: {os.path.abspath(ruta)}')
        self.stdout.write(self.style.WARNING(
            '  Revisa este Excel ANTES de generar los TXT. No se tocó la BD.'
        ))

    # ------------------------------------------------------------------ #
    def _procesar(self, queryset, salida_dir, etiqueta):
        generados = 0
        con_warnings = 0
        con_error = 0
        errores = []
        en_bloque_error = 0
        lo, hi = BLOQUE_FOLIOS_REEMITIDOS

        zipf = None
        if not self.dry_run:
            if self.usar_zip:
                os.makedirs(os.path.dirname(os.path.abspath(salida_dir)), exist_ok=True)
                zip_path = os.path.abspath(salida_dir) + '.zip'
                zipf = zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED)
            else:
                os.makedirs(salida_dir, exist_ok=True)

        # Las funciones de la vista emiten muchos logger.warning de debug por
        # cada DTE. En un batch eso ensucia los logs; se silencia el logger de
        # ese módulo durante el proceso y se restaura al final.
        vista_logger = logging.getLogger('app.views_modulo_documentos')
        nivel_previo = vista_logger.level
        vista_logger.setLevel(logging.ERROR)

        idx = 0
        try:
            for dte in queryset.iterator(chunk_size=500):
                try:
                    datos = self._construir(dte)

                    # Refoliado / fecha forzada "por fuera del sistema": se
                    # sobrescribe SOLO el folio (y opcionalmente la fecha) en el
                    # dict antes de generar el TXT. Todo lo demás (montos,
                    # detalle, receptor, referencias) queda intacto -> el TXT es
                    # idéntico al original salvo el folio/fecha.
                    folio_usado = dte.numero_documento
                    if self.folio_nuevo_desde is not None:
                        folio_usado = self.folio_nuevo_desde + idx
                        datos['documento']['folio'] = folio_usado
                    if self.fecha_nueva is not None:
                        datos['documento']['fecha_emision'] = (
                            self.fecha_nueva.strftime('%Y-%m-%d')
                        )
                    idx += 1

                    contenido = self._generar(datos)
                    validacion = self._validar(datos) or {}
                    warnings = validacion.get('warnings') or []

                    tipo_num = datos['documento']['tipo_documento']
                    # El nombre del archivo usa el folio realmente escrito en el
                    # TXT (el nuevo si se refolió) y la fecha del documento.
                    fecha_doc = (
                        self.fecha_nueva if self.fecha_nueva is not None
                        else dte.fecha_emision
                    )
                    fecha_str = fecha_doc.strftime('%Y%m%d') if fecha_doc else 'sinfecha'
                    nombre = f'dte_{tipo_num}_{folio_usado}_{fecha_str}.txt'

                    if not self.dry_run:
                        if zipf is not None:
                            zipf.writestr(nombre, contenido)
                        else:
                            # newline='' evita que Windows convierta '\n' en
                            # '\r\n': así el .txt en disco es byte-idéntico al
                            # que entrega el botón web (HttpResponse sirve '\n'
                            # puro). Crítico para que Acepta reciba lo mismo.
                            with open(os.path.join(salida_dir, nombre), 'w',
                                      encoding='utf-8', newline='') as f:
                                f.write(contenido)

                    generados += 1
                    if warnings:
                        con_warnings += 1
                        logger.warning(
                            '[TXT-BATCH] DTE %s (folio %s) con warnings: %s',
                            dte.id, dte.numero_documento, warnings,
                        )
                    if lo <= dte.numero_documento <= hi:
                        en_bloque_error += 1

                    if generados % 200 == 0:
                        self.stdout.write(f'  ... {generados} generados')

                except Exception as e:  # noqa: BLE001 - el batch no debe caerse
                    con_error += 1
                    msg = f'DTE id={dte.id} folio={dte.numero_documento}: {e}'
                    errores.append(msg)
                    logger.error('[TXT-BATCH] Error en %s', msg)
        finally:
            vista_logger.setLevel(nivel_previo)
            if zipf is not None:
                zipf.close()

        self._imprimir_resumen(
            generados, con_warnings, con_error, errores,
            en_bloque_error, salida_dir,
        )

    # ------------------------------------------------------------------ #
    def _imprimir_resumen(self, generados, con_warnings, con_error, errores,
                          en_bloque_error, salida_dir):
        self.stdout.write('\n' + '-' * 70)
        self.stdout.write('RESUMEN')
        self.stdout.write('-' * 70)
        verbo = 'se generarían' if self.dry_run else 'generados'
        self.stdout.write(self.style.SUCCESS(f'  TXT {verbo}: {generados}'))
        if con_warnings:
            self.stdout.write(self.style.WARNING(
                f'  Con warnings (largos SII truncados, etc.): {con_warnings}'
            ))
        if con_error:
            self.stdout.write(self.style.ERROR(f'  Con error (omitidos): {con_error}'))
            for m in errores[:20]:
                self.stdout.write(self.style.ERROR(f'     - {m}'))
            if len(errores) > 20:
                self.stdout.write(self.style.ERROR(
                    f'     ... y {len(errores) - 20} más (ver logs).'
                ))

        if en_bloque_error:
            lo, hi = BLOQUE_FOLIOS_REEMITIDOS
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(
                f'  [!] {en_bloque_error} de estos DTEs caen en el bloque de '
                f'folios reemitidos {lo}-{hi} (dos sucursales del mismo RUT '
                f'consumieron el mismo bloque).'
            ))

        if not self.dry_run and generados:
            destino = os.path.abspath(salida_dir) + ('.zip' if self.usar_zip else '/')
            self.stdout.write('')
            self.stdout.write(f'  Archivos en: {destino}')

    # ------------------------------------------------------------------ #
    @staticmethod
    def _parse_fecha(texto):
        try:
            return datetime.datetime.strptime(texto.strip(), '%Y-%m-%d').date()
        except ValueError:
            raise CommandError(f'Fecha inválida "{texto}". Formato: AAAA-MM-DD.')

    @staticmethod
    def _rango_mes(texto):
        try:
            anio, mes = texto.strip().split('-')
            anio, mes = int(anio), int(mes)
            primer_dia = datetime.date(anio, mes, 1)
            ultimo_dia = datetime.date(anio, mes, monthrange(anio, mes)[1])
            return primer_dia, ultimo_dia
        except (ValueError, TypeError):
            raise CommandError(f'Mes inválido "{texto}". Formato: AAAA-MM.')

    def _cargar_estados_sii(self, ruta):
        """
        Lee el export de estados del SII (.xls/.xlsx) y devuelve
        {folio:int -> estado_sii:str}. Requiere columnas 'folio' y
        'estado_sii'. Si un folio se repite en el archivo, gana el que tenga
        estado 'RECHAZADO' (para no perder el rechazo). Solo lectura.
        """
        if not os.path.exists(ruta):
            raise CommandError(f'No existe el archivo SII: {ruta}')
        try:
            import pandas as pd
        except ImportError:
            raise CommandError(
                'Falta pandas para leer el archivo SII '
                '(pip install pandas; para .xls también xlrd).'
            )
        engine = 'xlrd' if ruta.lower().endswith('.xls') else None
        try:
            df = pd.read_excel(ruta, engine=engine)
        except ImportError:
            raise CommandError(
                'Para leer .xls falta xlrd: pip install xlrd '
                '(ver requirements-dev.txt).'
            )
        cols = {c.lower().strip(): c for c in df.columns}
        if 'folio' not in cols or 'estado_sii' not in cols:
            raise CommandError(
                'El archivo SII debe tener columnas "folio" y "estado_sii". '
                f'Encontradas: {list(df.columns)}'
            )
        col_folio, col_estado = cols['folio'], cols['estado_sii']
        mapa = {}
        for _, row in df.iterrows():
            try:
                folio = int(row[col_folio])
            except (ValueError, TypeError):
                continue
            estado = str(row[col_estado]).strip()
            previo = mapa.get(folio)
            # Un folio repetido: conservar el estado RECHAZADO si aparece.
            if previo is None or 'RECHAZ' in estado.upper():
                mapa[folio] = estado
        self.stdout.write(
            f'  Archivo SII: {len(df)} filas, {len(mapa)} folios únicos '
            f'cargados de {os.path.basename(ruta)}'
        )
        return mapa
