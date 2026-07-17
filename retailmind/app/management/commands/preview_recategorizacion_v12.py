# -*- coding: utf-8 -*-
"""
Genera un Excel con la recategorización v1.2 COMO QUEDARÍA, sin tocar la BD.

Cruza tres fuentes:
  1. La BD de productos (agrupados por artículo, todas las bodegas).
  2. El mapeo v1.2 (categoría vieja -> árbol físico + especialidades múltiples),
     con refuerzo por keywords de descripción, forma de talla y marca.
  3. (Opcional) Las planillas de Compras por proveedor, para proponer el código
     completo de artículos cortados o sin sufijo -COLOR.

100% de solo lectura. Salida: recategorizacion_v12_preview_<fecha>.xlsx

Uso:
    python manage.py preview_recategorizacion_v12
    python manage.py preview_recategorizacion_v12 --compras-dir "C:\\Users\\cromo\\Documents\\Compras" --anios 2025,2026,2027
    python manage.py preview_recategorizacion_v12 --limite 500     # prueba rápida
"""
import glob
import logging
import os
import re
import unicodedata
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.utils import timezone

from app.models import Producto, Producto_Talla

from ._data_recategorizacion import (
    MAPEO_DIRECTO, REGLAS_KEYWORD, KEYWORDS_AMBIGUOS, REGLAS_MARCA,
    NO_PRODUCTO_EXACTO, NO_PRODUCTO_PREFIJOS, IDS_DEFAULT_ZAPATILLAS,
)

logger = logging.getLogger('app')


def fold(s):
    s = unicodedata.normalize('NFKD', str(s or ''))
    return ''.join(c for c in s if not unicodedata.combining(c)).upper().strip()


GEN_MAP = {"DAMA": "Mujer", "MUJER": "Mujer", "HOMBRE": "Hombre", "NINO": "Niño",
           "NINA": "Niña", "JUVENIL": "Juvenil", "UNISEX": "Unisex", "BEBE": "Bebé"}

ESP_DEP = {"Fútbol": "pasto", "Running": "running", "Training y Fitness": "training",
           "Basketball": "basket", "Tenis": "tenis", "Boxeo": "boxeo",
           "Artes Marciales": "artesm", "Natación": "natacion", "Voleyball": "voley",
           "Atletismo": "atletismo", "Ciclismo": "ciclismo",
           "Outdoor y Camping": "trekking", "Patinaje y Skate": "skate",
           "Pesca y Caza": "pesca", "Gimnasia": "gim", "Handball": "handball",
           "Rugby": "rugby", "Hockey": "hockey", "Golf": "golf", "Béisbol": "beisbol",
           "Ping Pong": "pingpong", "Squash": "tenis", "Pool y Billar": "pool",
           "Bádminton": "badminton"}

CAL = {"Zapatillas Urbanas": ("zapatillas", ["urbano"]),
       "Zapatillas Lona": ("zapatillas", ["urbano"]),
       "Botas": ("botas", []), "Botines": ("botines", ["urbano"]),
       "Mocasines": ("mocasin", ["urbano"]),
       "Sandalias y Chalas": ("sandalias", ["urbano"]),
       "Ballerinas y Bajas": ("ballerinas", ["urbano"]),
       "Plataformas": ("plataformas", ["urbano"]),
       "Pantuflas": ("pantuflas", ["descanso"]),
       "Zuecos": ("danza", ["cueca"]),
       "Zapatos de Vestir": ("zvestir", ["vestir"]),
       "Alpargatas": ("alpargata", ["urbano"]),
       "Cueca": ("zvestir", ["cueca"]),
       "Bebé y Gateadores": ("gateadores", []),
       "Confort y Ortopédico": ("zvestir", ["descanso"])}

ROPA_MAP = {"Poleras y Camisetas": ("vestuario", "poleras"),
            "Polerones y Chaquetas": ("vestuario", "chaquetas"),
            "Pantalones y Buzos": ("vestuario", "buzo"),
            "Shorts": ("vestuario", "shorts"),
            "Calcetines y Medias": ("accesorios", "medias")}

ACC_MAP = {"Bolsos y Mochilas": "mochilas", "Gorros": "gorros", "Correas": "accvest",
           "Pulseras y Joyas": "accvest", "Accesorios de Pelo": "accvest",
           "Protecciones y Canilleras": "protecciones"}


def sub_acc_por_desc(d):
    if re.search(r'BALON|PELOTA', d): return "balones"
    if "GUANTE" in d: return "guantes"
    if re.search(r'CANILLERA|ESPINILLERA|RODILLERA|MUNEQUERA|PROTEC|CODERA|VENDA|GUARD', d): return "protecciones"
    if re.search(r'MOCHILA|BOLSO|MORRAL|CARTERA|BALONERO', d): return "mochilas"
    if re.search(r'GORRO|JOCKEY|VISERA', d): return "gorros"
    if re.search(r'CALCET|SOQUETE|MEDIA', d): return "medias"
    if re.search(r'PESA|DISCO|MANCUERNA|BANDA ELAST|CUERDA|BARRA', d): return "equipamiento"
    if re.search(r'TROFEO|MEDALLA', d): return "trofeos"
    return "accdep"


def sub_ves_por_desc(d):
    if re.search(r'POLERON|PARKA|CHAQUETA|CHALECO|HOODIE|CANGURO', d): return "chaquetas"
    if re.search(r'SHORT|CALZA|BERMUDA|PATA', d): return "shorts"
    if re.search(r'PANTALON|JOGGER|LEGGING|BUZO', d): return "buzo"
    if re.search(r'BIKINI|TRAJE DE BA|TRIKINI|SUNGA', d): return "tbano"
    if "MALLA" in d: return "mallas"
    return "poleras"


def forma_por_tallas(tallas):
    """CALZADO si predominan tallas numéricas 16-49; ROPA si S/M/L/XL; '' si mezcla."""
    num = alpha = 0
    for t in tallas:
        tf = fold(t).replace(',', '.')
        try:
            v = float(tf)
            if 15 <= v <= 50:
                num += 1
                continue
        except ValueError:
            pass
        if tf in {"XS", "S", "M", "L", "XL", "XXL", "XXXL", "2XL", "3XL"}:
            alpha += 1
    if num > alpha and num: return "CALZADO"
    if alpha > num and alpha: return "ROPA"
    return ""


def derivar_v12(dep, tipo, desc, forma, gen):
    """(dep, tipo) del mapeo viejo + descripción + forma -> (cat, sub, [esp], notas)."""
    d = fold(desc)
    notas = []
    cat = sub = ""
    esp = []
    if dep == "Calzado":
        cat = "calzado"
        if tipo == "Escolar":
            sub = "ballerinas" if re.search(r'MAFALDA|GUILLERMINA|REINITA', d) else "zapatillas"
            if "BLANC" in d: esp = ["blanco"]
            elif "NEGR" in d: esp = ["negro"]
            else: esp = ["oficial"]; notas.append("escolar: definir oficial/blanco/negro")
            if sub == "zapatillas": esp.append("urbano")
        elif tipo in CAL:
            sub, esp = CAL[tipo][0], list(CAL[tipo][1])
            if tipo == "Bebé y Gateadores" and gen != "Bebé":
                notas.append("sugerir gen=Bebé")
            if tipo in ("Zuecos", "Confort y Ortopédico"):
                notas.append(f"{tipo}: confirmar destino")
    elif dep == "Deportes":
        e = ESP_DEP.get(tipo)
        if tipo == "Fútbol":
            if re.search(r'\bTF\b|SINTETIC|BABY', d): e = "baby"
            elif re.search(r'SALA|FUTSAL|INDOOR', d): e = "sala"
        if tipo == "Outdoor y Camping" and re.search(r'CARPA|CAMPING|SACO', d): e = "camping"
        esp = [e] if e else []
        if forma == "CALZADO":
            cat = "calzado"
            sub = "botines" if re.search(r'BOTIN|\bBTN\b', d) else "zapatillas"
        elif forma == "ROPA":
            cat, sub = "vestuario", sub_ves_por_desc(d)
        else:
            cat, sub = "accesorios", sub_acc_por_desc(d)
            if sub == "accdep": notas.append("accesorio genérico: confirmar sub")
    elif dep == "Ropa":
        if tipo == "Trajes de Baño y Mallas":
            cat = "vestuario"
            if "MALLA" in d: sub, esp = "mallas", ["gim"]
            else: sub, esp = "tbano", ["natacion"]
        elif tipo in ROPA_MAP:
            cat, sub = ROPA_MAP[tipo]
    elif dep == "Accesorios":
        cat, sub = "accesorios", ACC_MAP.get(tipo, "accdep")
    elif dep == "Otros":
        if tipo == "Trofeos y Medallas": cat, sub = "accesorios", "trofeos"
        elif tipo == "Juegos de Salón": cat, sub, esp = "accesorios", "accdep", ["recre"]
        elif tipo == "Implementos y Repuestos": cat, sub = "accesorios", sub_acc_por_desc(d)
        elif tipo == "Seguridad":
            cat = "calzado"
            sub = "botas" if ("BOTA" in d or "BOTIN" in d) else "zapatillas"
            esp = ["seguridad"]
        elif tipo == "No-Producto":
            return "", "", [], ["NO-PRODUCTO: excluir de analítica"]
    return cat, sub, esp, notas


def clasificar_dep_tipo(cat_id, cat_nombre, desc, marca):
    """Capas del pipeline viejo: mapeo directo -> keywords -> marca. -> (dep,tipo,rev,fuente)."""
    d = ' ' + fold(desc) + ' '
    dstrip = fold(desc)
    if dstrip in NO_PRODUCTO_EXACTO or dstrip.startswith(NO_PRODUCTO_PREFIJOS):
        return "Otros", "No-Producto", False, "no-producto"
    if cat_id in MAPEO_DIRECTO:
        dep, tipo, rev = MAPEO_DIRECTO[cat_id]
        return dep, tipo, rev, "mapeo_directo"
    for toks, dep, tipo in REGLAS_KEYWORD:
        for t in toks:
            if t in d:
                return dep, tipo, (t in KEYWORDS_AMBIGUOS), "keyword"
    m = fold(marca)
    if m in REGLAS_MARCA:
        dep, tipo = REGLAS_MARCA[m]
        return dep, tipo, True, "marca"
    if cat_id in IDS_DEFAULT_ZAPATILLAS:
        return "Calzado", "Zapatillas Urbanas", True, "default_casual"
    return "", "", True, "sin_resolver"


# ──────────────────────────── Compras ────────────────────────────
def indexar_compras(base_dir, anios, stdout):
    """Recorre las planillas y devuelve {prefijo_sin_color: [codigos_completos]}."""
    codigos = set()
    patrones = []
    for a in anios:
        patrones += glob.glob(os.path.join(base_dir, a, '**', '*.xls*'), recursive=True)
    stdout.write(f"  Compras: {len(patrones)} planillas en {anios} — indexando (puede tardar)...")
    from openpyxl import load_workbook
    ok = 0
    for i, path in enumerate(patrones, 1):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            col = None
            for fila in ws.iter_rows(min_row=1, max_row=30):
                for c in fila:
                    if fold(c.value) in ("PRODUCTO", "CODIGO", "COD", "ARTICULO", "ARTÍCULO"):
                        col = c.column; start = c.row + 1; break
                if col: break
            if col:
                for fila in ws.iter_rows(min_row=start, min_col=col, max_col=col, values_only=True):
                    v = str(fila[0] or '').strip()
                    if 4 <= len(v) <= 40 and not v.isspace():
                        codigos.add(v.upper())
                ok += 1
            wb.close()
        except Exception:
            pass
        if i % 100 == 0:
            stdout.write(f"    {i}/{len(patrones)} archivos...")
    idx = defaultdict(list)
    for c in codigos:
        pref = c.rsplit('-', 1)[0] if '-' in c else c
        idx[pref].append(c)
    stdout.write(f"  Compras: {ok} planillas legibles, {len(codigos):,} códigos únicos.")
    return idx


def match_compras(articulo, idx):
    a = str(articulo or '').strip().upper()
    if not a or not idx:
        return "", ""
    pref = a.rsplit('-', 1)[0] if '-' in a else a
    if a in idx.get(pref, []):
        return "exacto", ""
    cands = idx.get(a, [])          # ERP sin sufijo, compras con -COLOR
    if cands:
        return "sin_sufijo", " | ".join(sorted(cands)[:3])
    cands = [c for p, lst in idx.items() if p.startswith(a) for c in lst][:3]
    if cands:
        return "truncado?", " | ".join(sorted(cands))
    return "no_encontrado", ""


class Command(BaseCommand):
    help = "Genera Excel de recategorización v1.2 (solo lectura) desde la BD + mapeo + Compras"

    def add_arguments(self, parser):
        parser.add_argument('--salida', default='',
                            help='Ruta del xlsx de salida (default: recategorizacion_v12_preview_<fecha>.xlsx)')
        parser.add_argument('--compras-dir', default='',
                            help=r'Carpeta base de Compras (ej: C:\Users\cromo\Documents\Compras)')
        parser.add_argument('--anios', default='2025,2026,2027',
                            help='Años de Compras a indexar, separados por coma')
        parser.add_argument('--limite', type=int, default=0,
                            help='Procesar solo N artículos (prueba rápida)')

    def handle(self, *args, **opts):
        stamp = timezone.now().strftime('%Y%m%d_%H%M')
        salida = opts['salida'] or f'recategorizacion_v12_preview_{stamp}.xlsx'

        # 1) Productos agrupados por artículo
        self.stdout.write("1/4 Leyendo productos de la BD (solo lectura)...")
        arts = {}
        qs = (Producto.objects
              .select_related('categoria', 'atributo1', 'atributo3', 'sucursal')
              .values('id', 'articulo', 'descripcion', 'categoria_id',
                      'categoria__nombre', 'atributo1__valor', 'atributo3__valor',
                      'sucursal__nombre'))
        for p in qs.iterator(chunk_size=5000):
            a = (p['articulo'] or '').strip()
            r = arts.setdefault(a, {'desc': p['descripcion'] or '', 'marca': p['atributo1__valor'] or '',
                                    'sexo': p['atributo3__valor'] or '', 'cat_id': p['categoria_id'],
                                    'cat_nombre': p['categoria__nombre'] or '', 'n': 0,
                                    'sucs': set(), 'ids': []})
            r['n'] += 1
            r['sucs'].add(p['sucursal__nombre'] or '')
            r['ids'].append(p['id'])
        self.stdout.write(f"  {len(arts):,} artículos ({sum(r['n'] for r in arts.values()):,} filas).")

        # 2) Tallas y stock por artículo
        self.stdout.write("2/4 Leyendo tallas y stock...")
        tallas = defaultdict(list)
        stock = defaultdict(int)
        for t in (Producto_Talla.objects
                  .values('producto__articulo', 'talla', 'stock').iterator(chunk_size=10000)):
            a = (t['producto__articulo'] or '').strip()
            if len(tallas[a]) < 12:
                tallas[a].append(t['talla'])
            stock[a] += max(0, t['stock'] or 0)

        # 3) Compras (opcional)
        idx_compras = {}
        if opts['compras_dir']:
            self.stdout.write("3/4 Indexando planillas de Compras...")
            anios = [a.strip() for a in opts['anios'].split(',') if a.strip()]
            idx_compras = indexar_compras(opts['compras_dir'], anios, self.stdout)
        else:
            self.stdout.write("3/4 Compras: omitido (usa --compras-dir para activarlo).")

        # 4) Derivar y escribir Excel
        self.stdout.write("4/4 Derivando v1.2 y escribiendo Excel...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Articulos_v12")
        head = ['articulo', 'descripcion', 'marca', 'gen_v12', 'categoria_original',
                'cat_v12', 'sub_v12', 'esp_v12', 'confianza', 'revisar', 'nota',
                'fuente', 'n_filas', 'sucursales', 'stock_total',
                'match_compras', 'codigo_sugerido']
        ws.append(head)

        resumen = defaultdict(int)
        resumen_esp = defaultdict(int)
        n = n_rev = 0
        for a, r in sorted(arts.items()):
            if opts['limite'] and n >= opts['limite']:
                break
            n += 1
            gen = GEN_MAP.get(fold(r['sexo']), 'Unisex')
            dep, tipo, rev, fuente = clasificar_dep_tipo(
                r['cat_id'], r['cat_nombre'], r['desc'], r['marca'])
            forma = forma_por_tallas(tallas.get(a, []))
            if dep:
                cat, sub, esp, notas = derivar_v12(dep, tipo, r['desc'], forma, gen)
            else:
                cat = sub = ''; esp = []; notas = ['sin resolver — manual']
            conf = 'alta' if (fuente == 'mapeo_directo' and not rev and not notas) else \
                   ('media' if cat else 'baja')
            revisar = 'Sí' if (rev or notas or not cat) else ''
            if revisar: n_rev += 1
            mc, sug = match_compras(a, idx_compras)
            ws.append([a, r['desc'], r['marca'], gen, r['cat_nombre'],
                       cat, sub, ', '.join(dict.fromkeys(esp)), conf, revisar,
                       '; '.join(notas), fuente, r['n'],
                       ', '.join(sorted(s for s in r['sucs'] if s)), stock[a],
                       mc, sug])
            if cat:
                resumen[(cat, sub)] += 1
                for e in esp:
                    resumen_esp[e] += 1

        ws2 = wb.create_sheet("Resumen_cat_sub")
        ws2.append(['cat_v12', 'sub_v12', 'articulos'])
        for (c, s), v in sorted(resumen.items(), key=lambda x: -x[1]):
            ws2.append([c, s, v])
        ws3 = wb.create_sheet("Resumen_especialidades")
        ws3.append(['especialidad', 'articulos'])
        for e, v in sorted(resumen_esp.items(), key=lambda x: -x[1]):
            ws3.append([e, v])
        wb.save(salida)

        self.stdout.write(self.style.SUCCESS(
            f"Listo: {salida} — {n:,} artículos, {n_rev:,} con flag revisar."))
        logger.info("preview_recategorizacion_v12: %d articulos -> %s", n, salida)
