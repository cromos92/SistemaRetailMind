# -*- coding: utf-8 -*-
"""
Aplica la recategorización v1.2 por artículo, leyendo
docs/recategorizacion_v12_por_articulo.xlsx (hoja Articulos_v12).

Por cada artículo del Excel:
  - Producto.categoria -> Categoria hija v1.2 (según sub_v12), en TODAS las
    filas/bodegas de ese artículo.
  - Especialidades (esp_v12, separadas por coma) -> ProductoAtributoValor
    con el atributo "Especialidad" (multi-etiqueta; no duplica).
  - Género: si el producto tiene DAMA en atributo3, se migra a MUJER.

Seguridad:
  - DRY-RUN por defecto; --apply escribe.
  - --lote N procesa solo los primeros N artículos (para probar).
  - --filtro-cat "TEXTO" limita a artículos cuya categoria_original contenga TEXTO.
  - --solo-confiables omite filas con revisar_final = Sí.
  - Siempre escribe un CSV de log con antes/después junto al Excel.
  - No borra categorías viejas ni productos. No toca filas sin cat_v12
    (no-producto / sin resolver) — solo las reporta.

Requiere haber corrido antes:  python manage.py sembrar_taxonomia_v12 --apply

Uso típico:
    python manage.py aplicar_recategorizacion_v12 --lote 200            # dry-run chico
    python manage.py aplicar_recategorizacion_v12 --lote 200 --apply    # primer lote real
    python manage.py aplicar_recategorizacion_v12 --solo-confiables --apply
    python manage.py aplicar_recategorizacion_v12 --apply               # todo
"""
import csv
import logging
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from app.models import (
    Producto, Categoria, Productos_Atributos, AtributoOpcion, ProductoAtributoValor,
)

from ._data_recategorizacion_v12 import (
    PADRES, SUBS, ESPECIALIDADES, NOMBRE_ATRIBUTO_ESPECIALIDAD, GENERO_MIGRAR,
)

logger = logging.getLogger('app')

EXCEL_DEFAULT = os.path.join('..', 'docs', 'recategorizacion_v12_por_articulo.xlsx')


class Command(BaseCommand):
    help = "Aplica la recategorización v1.2 por artículo desde el Excel (dry-run por defecto)"

    def add_arguments(self, parser):
        parser.add_argument('--excel', default=EXCEL_DEFAULT,
                            help='Ruta al xlsx (hoja Articulos_v12)')
        parser.add_argument('--apply', action='store_true', help='Escribe en la BD')
        parser.add_argument('--lote', type=int, default=0,
                            help='Procesar solo los primeros N artículos')
        parser.add_argument('--filtro-cat', default='',
                            help='Solo artículos cuya categoria_original contenga este texto')
        parser.add_argument('--solo-confiables', action='store_true',
                            help='Omitir filas con revisar_final = Sí')

    # ------------------------------------------------------------------
    def _cargar_excel(self, ruta):
        from openpyxl import load_workbook
        if not os.path.exists(ruta):
            raise CommandError(f"No existe el Excel: {ruta}")
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb['Articulos_v12']
        rows = ws.iter_rows(values_only=True)
        head = [str(h) if h is not None else '' for h in next(rows)]
        idx = {h: i for i, h in enumerate(head)}
        req = ['articulo', 'cat_v12', 'sub_v12', 'esp_v12', 'revisar_final', 'categoria_original']
        for r in req:
            if r not in idx:
                raise CommandError(f"Falta la columna '{r}' en Articulos_v12")
        for row in rows:
            yield {h: (row[i] if i < len(row) else None) for h, i in idx.items()}

    def _mapa_categorias(self):
        """slug de sub -> objeto Categoria (valida que la siembra esté hecha)."""
        mapa = {}
        for slug, (padre_slug, nombre) in SUBS.items():
            obj = (Categoria.objects
                   .filter(nombre__iexact=nombre, padre__nombre__iexact=PADRES[padre_slug])
                   .first())
            if obj is None:
                raise CommandError(
                    f"Categoria '{nombre}' (slug {slug}) no existe — "
                    f"corre primero: python manage.py sembrar_taxonomia_v12 --apply")
            mapa[slug] = obj
        return mapa

    def _mapa_especialidades(self):
        attr = Productos_Atributos.objects.filter(
            nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD).first()
        if attr is None:
            raise CommandError("Atributo 'Especialidad' no existe — corre sembrar_taxonomia_v12 --apply")
        opciones = {o.valor: o for o in AtributoOpcion.objects.filter(atributo=attr)}
        faltan = set(ESPECIALIDADES) - set(opciones)
        if faltan:
            raise CommandError(f"Faltan opciones de Especialidad: {sorted(faltan)}")
        return attr, opciones

    # ------------------------------------------------------------------
    def handle(self, *args, **opts):
        apply_ = opts['apply']
        modo = 'APPLY' if apply_ else 'DRY-RUN'
        cat_map = self._mapa_categorias()
        esp_attr, esp_ops = self._mapa_especialidades()

        # Género DAMA -> MUJER (ids resueltos por valor, con verificación)
        gen_map = {}
        for origen, destino in GENERO_MIGRAR.items():
            o = AtributoOpcion.objects.filter(valor__iexact=origen).first()
            d = AtributoOpcion.objects.filter(valor__iexact=destino).first()
            if o and d and o.atributo_id == d.atributo_id:
                gen_map[o.id] = d

        stamp = timezone.now().strftime('%Y%m%d_%H%M')
        log_path = os.path.join(os.path.dirname(os.path.abspath(opts['excel'])),
                                f'recategorizacion_v12_log_{stamp}_{modo}.csv')

        n_art = n_prod = n_esp = n_gen = n_skip = n_sin = 0
        with open(log_path, 'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(['articulo', 'productos_afectados', 'cat_antes(ids)', 'cat_despues',
                        'especialidades', 'genero_migrado', 'accion'])

            with transaction.atomic():
                for fila in self._cargar_excel(opts['excel']):
                    if opts['lote'] and n_art >= opts['lote']:
                        break
                    art = str(fila['articulo'] or '').strip()
                    if not art:
                        continue
                    if opts['filtro_cat'] and opts['filtro_cat'].upper() not in str(
                            fila['categoria_original'] or '').upper():
                        continue
                    if opts['solo_confiables'] and str(fila['revisar_final'] or '') == 'Sí':
                        n_skip += 1
                        continue
                    cat_slug = str(fila['cat_v12'] or '').strip()
                    sub_slug = str(fila['sub_v12'] or '').strip()
                    if not cat_slug or sub_slug not in cat_map:
                        n_sin += 1
                        w.writerow([art, 0, '', '', '', '', 'SIN DESTINO (no-producto/sin resolver)'])
                        continue

                    n_art += 1
                    destino = cat_map[sub_slug]
                    esps = [e.strip() for e in str(fila['esp_v12'] or '').split(',')
                            if e.strip() and e.strip() in esp_ops]

                    productos = list(Producto.objects.filter(articulo=art))
                    if not productos:
                        w.writerow([art, 0, '', destino.nombre, ','.join(esps), '', 'ARTÍCULO NO ENCONTRADO EN BD'])
                        continue

                    antes = sorted({p.categoria_id for p in productos if p.categoria_id})
                    gen_migrados = 0
                    for p in productos:
                        if apply_:
                            cambios = []
                            if p.categoria_id != destino.id:
                                p.categoria = destino
                                cambios.append('categoria')
                            if p.atributo3_id in gen_map:
                                p.atributo3 = gen_map[p.atributo3_id]
                                cambios.append('atributo3')
                                gen_migrados += 1
                            if cambios:
                                p.save(update_fields=cambios + ['fecha_actualizacion'])
                            for slug in esps:
                                _, creado = ProductoAtributoValor.objects.get_or_create(
                                    producto=p, atributo=esp_attr, opcion=esp_ops[slug])
                                if creado:
                                    n_esp += 1
                        else:
                            if p.atributo3_id in gen_map:
                                gen_migrados += 1
                            n_esp += len(esps)   # estimación dry-run
                    n_prod += len(productos)
                    n_gen += gen_migrados
                    w.writerow([art, len(productos), ';'.join(map(str, antes)), destino.nombre,
                                ','.join(esps), gen_migrados, modo])

                if not apply_:
                    transaction.set_rollback(True)

        self.stdout.write(self.style.MIGRATE_HEADING(f"[{modo}] aplicar_recategorizacion_v12"))
        self.stdout.write(f"  artículos procesados : {n_art}")
        self.stdout.write(f"  productos (filas)    : {n_prod}")
        self.stdout.write(f"  etiquetas esp.       : {n_esp}{' (estimado)' if not apply_ else ''}")
        self.stdout.write(f"  géneros DAMA→MUJER   : {n_gen}")
        self.stdout.write(f"  omitidos (revisar)   : {n_skip}")
        self.stdout.write(f"  sin destino          : {n_sin}")
        self.stdout.write(self.style.SUCCESS(f"  log: {log_path}"))
        logger.info("aplicar_recategorizacion_v12 %s: %d articulos, %d productos, log=%s",
                    modo, n_art, n_prod, log_path)
