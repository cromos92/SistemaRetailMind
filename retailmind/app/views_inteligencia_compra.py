"""
Inteligencia de Compra — análisis histórico + pronóstico de una marca para
decidir cuánto/qué comprar. 100% SOLO LECTURA sobre el kardex.

Vista nueva del módulo de Reportes (FBV, patrón del proyecto):
  - ver_inteligencia_compra:      página con selector de marca + gráficos ApexCharts
  - obtener_inteligencia_compra:  API JSON con todo el análisis de una marca

Reglas de dominio:
  - Marca = Producto.atributo1 (AtributoOpcion). Excluye excluir_de_analitica=True.
  - Demanda de público = ventas en TIENDAS vendedoras (es_centro_distribucion=False).
    Las BODEGAS/CD (EDEL, IMP, PA00, PAO0, GILD) son distribución/mayorista.
  - Conceptos SIEMPRE desde constants_kardex (no listas inline).
  - Unidades = Σ|cantidad| (robusto ante bugs de signo de la migración).
  - Pronóstico: seasonal-naive con tendencia acotada.
"""
import logging
import re
from datetime import timedelta

from django.db.models import (
    BigIntegerField, Case, CharField, Count, F, Max, OuterRef, Q, Subquery,
    Sum, Value, When,
)
from django.db.models.functions import Abs, Coalesce, ExtractMonth, ExtractYear
from django.db.utils import ProgrammingError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from .constants_kardex import (
    CONCEPTOS_ABASTECIMIENTO, CONCEPTOS_REINGRESO, CONCEPTOS_VENTA,
    REF_SALDO_INICIAL_SINTETICO,
)
from .decorators import requiere_permiso
from .models import (
    AtributoOpcion, CampanaLiquidacionProducto, Categoria, Compras_Producto,
    EmpresaUser, LoteProducto, Movimientos_Producto, Producto, Producto_Talla,
    ProductoAtributoValor, Productos_Atributos, Sucursal,
)

logger = logging.getLogger('app')

BI = BigIntegerField()
MESES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']


# ------------------------------------------------------------------ helpers
def _sucursales_usuario(request):
    empresas = EmpresaUser.objects.filter(
        user=request.user, status=True
    ).values_list('empresa_id', flat=True)
    return Sucursal.objects.filter(empresa_id__in=empresas).order_by('alias')


def _talla_norm(t):
    """Normaliza talla a talla EU de calzado. Fuera de escala -> None ('Otros')."""
    if t is None:
        return None
    s = str(t).strip().replace(',', '.')
    try:
        v = float(s)
    except (ValueError, TypeError):
        return None
    if 19 <= v <= 48:
        if abs(v - round(v)) < 0.01:
            return str(int(round(v)))
        return f"{v:.1f}"
    return None


def _talla_key(t):
    try:
        return (0, float(str(t).replace(',', '.')))
    except (ValueError, TypeError):
        return (1, str(t))


def _u(qs):
    return qs.aggregate(u=Sum(Abs('cantidad'), output_field=BI))['u'] or 0


def _um(qs, campo):
    return qs.aggregate(m=Sum(Abs(F('cantidad')) * F(campo), output_field=BI))['m'] or 0


def _mes_siguiente(y, m):
    return (y, m + 1) if m < 12 else (y + 1, 1)


def _mes_anterior(y, m):
    return (y, m - 1) if m > 1 else (y - 1, 12)


def _forecast(serie, hoy):
    """serie: dict {(y,m): u}. Devuelve dict con ttm, yoy, forward y forecast 12m."""
    # índices estacionales: promedio de participación mensual en años completos recientes
    years = sorted({y for (y, _m) in serie})
    shares = {m: [] for m in range(1, 13)}
    for y in years:
        tot = sum(serie.get((y, m), 0) for m in range(1, 13))
        # solo años con los 12 meses con algo de dato y no el año en curso
        if tot > 0 and all((y, m) in serie for m in range(1, 13)) and y < hoy.year:
            for m in range(1, 13):
                shares[m].append(serie.get((y, m), 0) / tot)
    seas = {m: (sum(v) / len(v) if v else 1 / 12) for m, v in shares.items()}
    ssum = sum(seas.values()) or 1
    seas = {m: seas[m] / ssum for m in seas}

    last_y, last_m = _mes_anterior(hoy.year, hoy.month)  # último mes completo

    def ttm_end(y, m):
        tot, yy, mm = 0, y, m
        for _ in range(12):
            tot += serie.get((yy, mm), 0)
            yy, mm = _mes_anterior(yy, mm)
        return tot

    ttm = ttm_end(last_y, last_m)
    py, pm = last_y - 1, last_m
    ttm_prev = ttm_end(py, pm)
    yoy = (ttm / ttm_prev - 1) if ttm_prev else 0.0
    g = max(-0.35, min(0.10, yoy))
    forward = ttm * (1 + g)

    fc = []
    yy, mm = _mes_siguiente(last_y, last_m)
    for _ in range(12):
        base = forward * seas[mm]
        fc.append({'label': f"{MESES[mm]} {yy}", 'y': yy, 'm': mm,
                   'base': round(base), 'low': round(base * 0.85), 'high': round(base * 1.15)})
        yy, mm = _mes_siguiente(yy, mm)

    return {
        'ttm': ttm, 'ttm_prev': ttm_prev, 'yoy': round(yoy * 100, 1),
        'g': round(g * 100, 1), 'forward_annual': round(forward),
        'forecast': fc, 'forecast_total': round(sum(f['base'] for f in fc)),
        'seasonal': [{'mes': MESES[m], 'indice': round(seas[m] * 100, 1)} for m in range(1, 13)],
    }


# ------------------------------------------------------------------ vista
# El código 'inteligencia_compra' lo crea `inicializar_permisos` (Fase B):
# hasta correr el command en prod, el decorador es fail-closed (403) —
# esperado, mismo despliegue que diferencias/tránsito (incidente 05-ago).
@requiere_permiso('inteligencia_compra', 'puede_ver')
def ver_inteligencia_compra(request):
    """Página 'Inteligencia de Compra' con selector de marca."""
    sucursales = list(_sucursales_usuario(request))
    tiendas = [s for s in sucursales if not s.es_centro_distribucion]

    atributo_marca = Productos_Atributos.objects.filter(nombre__icontains='marca').first()
    marcas = []
    if atributo_marca:
        marcas = list(AtributoOpcion.objects.filter(atributo=atributo_marca)
                      .order_by('valor').values('id', 'valor'))

    context = {
        'marcas': marcas,
        'tiendas': tiendas,
        'sucursal_actual_id': request.session.get('idSucursalActual'),
    }
    return render(request, 'vistas/modulo_reportes/inteligencia_compra.html', context)


@require_GET
@requiere_permiso('inteligencia_compra', 'puede_ver')
def obtener_inteligencia_compra(request):
    """API: análisis + pronóstico de compra para una marca."""
    try:
        marca_id = request.GET.get('marca_id')
        sucursal_id = request.GET.get('sucursal_id')  # 'todas' o id de tienda
        if not marca_id:
            return JsonResponse({'success': False, 'error': 'Selecciona una marca.'})

        marca = AtributoOpcion.objects.filter(id=marca_id).first()
        if not marca:
            return JsonResponse({'success': False, 'error': 'Marca no encontrada.'})

        sucursales = list(_sucursales_usuario(request))
        all_ids = [s.id for s in sucursales]
        tiendas_ids = [s.id for s in sucursales if not s.es_centro_distribucion]
        bodegas_ids = [s.id for s in sucursales if s.es_centro_distribucion]

        tienda_especifica = None
        if sucursal_id and str(sucursal_id).lower() != 'todas':
            try:
                sid = int(sucursal_id)
                if sid in tiendas_ids:
                    tienda_especifica = sid
            except (ValueError, TypeError):
                pass
        tienda_scope = [tienda_especifica] if tienda_especifica else tiendas_ids

        hoy = timezone.localdate()

        # -------- base querysets (joins, sin materializar IN gigante) --------
        prod_f = {
            'ProductoTalla__producto__atributo1_id': marca_id,
            'ProductoTalla__producto__excluir_de_analitica': False,
            'ProductoTalla__producto__sucursal_id__in': all_ids,
            'estado': 'COMPLETADO',
        }
        movs = Movimientos_Producto.objects.filter(**prod_f)
        ventas = movs.filter(concepto__in=CONCEPTOS_VENTA)
        # Venta NETA: los reingresos post-venta (NC, devoluciones, cambios —
        # CONCEPTOS_REINGRESO) se RESTAN de la demanda en el mismo universo
        # (marca/analítica/sucursal/período). Con demanda bruta SKECHERS 90d
        # sumaba 1.410 ventas con 169 reingresos (+12%): velocidad, TTM,
        # pronóstico y compra sugerida inflados (auditoría 2026-08, P1-6a).
        # Sell-through: la apertura sintética de la migración Laravel entró
        # como INGRESO_INICIAL pero NO es abastecimiento (es la foto inicial);
        # contarla duplica "ingresado" y hunde el STR (26,8% vs 60,2% real
        # en SKECHERS 2026 — auditoría 2026-08, P1-6b).
        abast = movs.filter(concepto__in=CONCEPTOS_ABASTECIMIENTO).exclude(
            concepto='INGRESO_INICIAL',
            referencia_externa=REF_SALDO_INICIAL_SINTETICO,
        )
        en_tienda = {
            'ProductoTalla__producto__sucursal__es_centro_distribucion': False,
            'ProductoTalla__producto__sucursal_id__in': tienda_scope,
        }
        ventas_tienda = ventas.filter(**en_tienda)

        # Fase C (perf): ventas y reingresos se agregan JUNTOS en una pasada
        # por (grupo, año, mes, tienda) y las series (anual / por tienda /
        # mensual) se derivan sumando parciales enteros — mismo resultado
        # campo a campo que las 8 agregaciones separadas de antes, con 2
        # scans del kardex en vez de 8 (auditoría 2026-08, sección 10).
        GRUPO = Case(When(concepto__in=CONCEPTOS_VENTA, then=Value('v')),
                     default=Value('r'), output_field=CharField())
        vr = movs.filter(concepto__in=CONCEPTOS_VENTA + CONCEPTOS_REINGRESO)
        vr_tienda = vr.filter(**en_tienda)

        # -------- 1+2+4. una pasada tiendas: año/mes/tienda (NETA) --------
        ALIAS_K = 'ProductoTalla__producto__sucursal__alias'
        va_bruta, ra = {}, {}   # año  -> {'u','monto'} (ventas / reingresos)
        vt_bruta, rt = {}, {}   # alias-> {'u','monto'}
        serie = {}              # (año, mes) -> unidades netas
        for r in (vr_tienda
                  .annotate(g=GRUPO, a=ExtractYear('fecha'), m=ExtractMonth('fecha'))
                  .values('g', 'a', 'm', ALIAS_K)
                  .annotate(u=Sum(Abs('cantidad'), output_field=BI),
                            monto=Sum(Abs(F('cantidad')) * F('precio'), output_field=BI))):
            u, monto = r['u'] or 0, r['monto'] or 0
            es_venta = r['g'] == 'v'
            acc = (va_bruta if es_venta else ra).setdefault(r['a'], {'u': 0, 'monto': 0})
            acc['u'] += u
            acc['monto'] += monto
            acc = (vt_bruta if es_venta else rt).setdefault(r[ALIAS_K], {'u': 0, 'monto': 0})
            acc['u'] += u
            acc['monto'] += monto
            k = (r['a'], r['m'])
            serie[k] = serie.get(k, 0) + (u if es_venta else -u)

        va = {}
        for a in sorted(set(va_bruta) | set(ra)):
            va[a] = {
                'u': ((va_bruta.get(a) or {}).get('u') or 0) - ((ra.get(a) or {}).get('u') or 0),
                'monto': ((va_bruta.get(a) or {}).get('monto') or 0) - ((ra.get(a) or {}).get('monto') or 0),
            }
        ventas_anual = [{'anio': a, 'unidades': va[a]['u'], 'monto': va[a]['monto']}
                        for a in sorted(va)]
        aa = {r['a']: (r['u'] or 0) for r in abast.annotate(a=ExtractYear('fecha')).values('a')
              .annotate(u=Sum(Abs('cantidad'), output_field=BI)).order_by('a')}

        ventas_por_tienda = sorted(
            [{'alias': al,
              'unidades': ((vt_bruta.get(al) or {}).get('u') or 0) - ((rt.get(al) or {}).get('u') or 0),
              'monto': ((vt_bruta.get(al) or {}).get('monto') or 0) - ((rt.get(al) or {}).get('monto') or 0)}
             for al in set(vt_bruta) | set(rt)],
            key=lambda x: -x['unidades'])

        # bodegas: una pasada (solo unidades — es lo único que se publica)
        vb, rb = {}, {}
        for r in (vr.filter(ProductoTalla__producto__sucursal__es_centro_distribucion=True)
                  .annotate(g=GRUPO).values('g', ALIAS_K)
                  .annotate(u=Sum(Abs('cantidad'), output_field=BI))):
            (vb if r['g'] == 'v' else rb)[r[ALIAS_K]] = r['u'] or 0
        distribucion_bodega = sorted(
            [{'alias': al, 'unidades': vb.get(al, 0) - rb.get(al, 0)}
             for al in set(vb) | set(rb)],
            key=lambda x: -x['unidades'])

        # -------- 3. stock actual --------
        pt = Producto_Talla.objects.filter(
            producto__atributo1_id=marca_id, producto__excluir_de_analitica=False,
            producto__sucursal_id__in=all_ids, stock__gt=0)
        stock_suc = [
            {'alias': r['producto__sucursal__alias'],
             'tipo': 'Bodega' if r['producto__sucursal__es_centro_distribucion'] else 'Tienda',
             'stock': (r['u'] or 0), 'skus': r['skus']}
            for r in pt.values('producto__sucursal__alias', 'producto__sucursal__es_centro_distribucion')
            .annotate(u=Sum('stock', output_field=BI), skus=Count('id')).order_by('-u')]
        stock_tiendas = pt.filter(producto__sucursal__es_centro_distribucion=False,
                                  producto__sucursal_id__in=tienda_scope
                                  ).aggregate(s=Sum('stock', output_field=BI))['s'] or 0

        # -------- 4. serie mensual venta pública NETA (gráfico + pronóstico) --------
        # (`serie` ya viene de la pasada consolidada 1+2+4)
        # serie continua últimos 36 meses para el chart
        serie_chart = []
        yy, mm = hoy.year, hoy.month
        tmp = []
        for _ in range(36):
            tmp.append({'label': f"{MESES[mm]} {str(yy)[2:]}", 'u': serie.get((yy, mm), 0)})
            yy, mm = _mes_anterior(yy, mm)
        serie_chart = list(reversed(tmp))

        fc = _forecast(serie, hoy)

        # -------- 5. sell-through anual --------
        sellthrough = []
        for a in sorted(set(va) | set(aa)):
            v = (va.get(a, {}) or {}).get('u') or 0
            i = aa.get(a, 0)
            sellthrough.append({'anio': a, 'vendido': v, 'ingresado': i,
                                'str': round(100.0 * v / i, 1) if i else None})

        # -------- 6. velocidad 90d + montos TTM (una pasada 455d) --------
        # Las 4 ventanas (90d, LY 455→365d y los montos TTM del bloque 9) son
        # agregados condicionales sobre el MISMO scan de 455 días: idéntico a
        # los 8 aggregates separados de antes, con 1 query.
        d90 = hoy - timedelta(days=90)
        d12 = hoy - timedelta(days=365)
        d455 = hoy - timedelta(days=455)
        win = {r['g']: r for r in (
            vr_tienda.filter(fecha__gte=d455).annotate(g=GRUPO).values('g')
            .annotate(
                u90=Sum(Abs('cantidad'),
                        filter=Q(fecha__gte=d90, fecha__lt=hoy), output_field=BI),
                u_ly=Sum(Abs('cantidad'),
                         filter=Q(fecha__lt=d12), output_field=BI),
                mp=Sum(Abs(F('cantidad')) * F('precio'),
                       filter=Q(fecha__gte=d12), output_field=BI),
                mc=Sum(Abs(F('cantidad')) * F('costo'),
                       filter=Q(fecha__gte=d12), output_field=BI)))}
        _wv, _wr = win.get('v') or {}, win.get('r') or {}
        v90 = (_wv.get('u90') or 0) - (_wr.get('u90') or 0)
        v90_ly = (_wv.get('u_ly') or 0) - (_wr.get('u_ly') or 0)
        ttm_venta_m = (_wv.get('mp') or 0) - (_wr.get('mp') or 0)
        ttm_costo_m = (_wv.get('mc') or 0) - (_wr.get('mc') or 0)

        # -------- 7. curva de tallas (EU normalizada, demanda NETA 24m) --------
        venta_talla, stock_talla = {}, {}
        TALLA_K = 'ProductoTalla__talla'
        rows_curva = list(
            vr_tienda.filter(fecha__gte=hoy - timedelta(days=730))
            .annotate(g=GRUPO).values('g', TALLA_K)
            .annotate(u=Sum(Abs('cantidad'), output_field=BI)))
        if not any(r['g'] == 'v' for r in rows_curva):
            # fallback: marca sin venta reciente → all-time (mismo criterio
            # que antes: manda la existencia de VENTAS en 24m)
            rows_curva = list(
                vr_tienda.annotate(g=GRUPO).values('g', TALLA_K)
                .annotate(u=Sum(Abs('cantidad'), output_field=BI)))
        for r in rows_curva:
            t = _talla_norm(r[TALLA_K])
            if t:
                delta = (r['u'] or 0) if r['g'] == 'v' else -(r['u'] or 0)
                venta_talla[t] = venta_talla.get(t, 0) + delta
        # Piso 0: una talla con más reingresos que ventas no puede pesar negativo.
        venta_talla = {t: max(0, v) for t, v in venta_talla.items()}
        for r in (pt.filter(producto__sucursal__es_centro_distribucion=False,
                            producto__sucursal_id__in=tienda_scope)
                  .values('talla').annotate(u=Sum('stock', output_field=BI))):
            t = _talla_norm(r['talla'])
            if t:
                stock_talla[t] = stock_talla.get(t, 0) + (r['u'] or 0)
        tot_v = sum(venta_talla.values()) or 1
        tot_s = sum(stock_talla.values()) or 1
        curva = []
        for t in sorted(set(venta_talla) | set(stock_talla), key=_talla_key):
            vp = 100.0 * venta_talla.get(t, 0) / tot_v
            sp = 100.0 * stock_talla.get(t, 0) / tot_s
            gap = vp - sp
            flag = 'quebrada' if (vp >= 3 and stock_talla.get(t, 0) == 0) else (
                'sobre_stock' if gap <= -1.0 else ('gap' if gap >= 1.0 else 'ok'))
            curva.append({'talla': t, 'venta_u': venta_talla.get(t, 0), 'venta_pct': round(vp, 1),
                          'stock_u': stock_talla.get(t, 0), 'stock_pct': round(sp, 1),
                          'gap': round(gap, 1), 'flag': flag})

        # -------- 8. recomendación --------
        forward = fc['forward_annual']
        mensual = forward / 12 if forward else 0
        cobertura = round(stock_tiendas / mensual, 1) if mensual else None
        season = forward / 2  # ~6 meses
        gross_need = max(0, season - stock_tiendas)
        newness = round(season * 0.25)
        comprar = int(round(max(gross_need, newness)))

        if cobertura is None:
            veredicto = 'Sin datos de venta suficientes para proyectar.'
        elif cobertura > 12:
            veredicto = (f'Sobre-stock ({cobertura} meses de cobertura). NO reposición amplia: '
                         f'comprar solo frescura + relleno de tallas núcleo.')
        elif cobertura > 6:
            veredicto = f'Cobertura holgada ({cobertura} meses). Compra moderada, muy enfocada en curva.'
        elif cobertura >= 3:
            veredicto = f'Cobertura sana ({cobertura} meses). Reponer según curva de demanda.'
        else:
            veredicto = f'Riesgo de quiebre ({cobertura} meses). Reponer con prioridad.'
            comprar = int(round(max(gross_need, forward * 0.75 - stock_tiendas, newness)))

        # curva de compra: reparte por demanda, excluye tallas sobre-stockeadas
        pesos = {c['talla']: c['venta_u'] for c in curva if c['flag'] != 'sobre_stock' and c['venta_u'] > 0}
        suma_pesos = sum(pesos.values()) or 1
        curva_compra = sorted(
            [{'talla': t, 'pct': round(100.0 * w / suma_pesos, 1), 'pares': int(round(comprar * w / suma_pesos))}
             for t, w in pesos.items()],
            key=lambda x: _talla_key(x['talla']))
        evitar = [c['talla'] for c in curva if c['flag'] == 'sobre_stock']

        # asignación por tienda (por participación de venta pública en scope)
        base_alloc = ventas_por_tienda if ventas_por_tienda else []
        suma_alloc = sum(t['unidades'] for t in base_alloc) or 1
        asignacion = [{'alias': t['alias'],
                       'pct': round(100.0 * t['unidades'] / suma_alloc, 1),
                       'pares': int(round(comprar * t['unidades'] / suma_alloc))}
                      for t in base_alloc]

        tot_pub = sum(x['unidades'] for x in ventas_anual)
        tot_bod = sum(x['unidades'] for x in distribucion_bodega)

        # -------- 9. finanzas: valor inventario, margen, rotación, WOS, GMROI --------
        pt_tienda = pt.filter(producto__sucursal__es_centro_distribucion=False,
                              producto__sucursal_id__in=tienda_scope)
        fin = pt_tienda.aggregate(
            c=Sum(F('stock') * F('producto__costo'), output_field=BI),
            p=Sum(F('stock') * F('producto__precioventa'), output_field=BI),
            n=Count('id'))
        inv_costo = fin['c'] or 0
        inv_precio = fin['p'] or 0

        # ttm_venta_m / ttm_costo_m ya vienen de la pasada 455d del bloque 6.
        # margen: realizado si el costo viene en las ventas; si no, margen de lista del inventario
        if ttm_venta_m > 0 and 0 < ttm_costo_m < ttm_venta_m:
            margen_pct = round(100.0 * (ttm_venta_m - ttm_costo_m) / ttm_venta_m, 1)
            margen_src = 'realizado'
            margen_anual = ttm_venta_m - ttm_costo_m
        elif inv_precio > 0:
            margen_pct = round(100.0 * (inv_precio - inv_costo) / inv_precio, 1)
            margen_src = 'lista'
            margen_anual = ttm_venta_m * margen_pct / 100.0
        else:
            margen_pct, margen_src, margen_anual = None, None, 0
        rotacion = round(12.0 / cobertura, 2) if cobertura else None
        wos = round(cobertura * 4.345, 1) if cobertura else None
        gmroi = round(margen_anual / inv_costo, 2) if inv_costo else None

        # -------- 10. dead stock / antigüedad (SKUs con stock sin venta reciente) --------
        skus_total = fin['n'] or 0
        vend90 = ventas_tienda.filter(fecha__gte=hoy - timedelta(days=90)).values('ProductoTalla')
        vend180 = ventas_tienda.filter(fecha__gte=hoy - timedelta(days=180)).values('ProductoTalla')
        dead90_n = pt_tienda.exclude(id__in=vend90).count()
        d180 = pt_tienda.exclude(id__in=vend180).aggregate(
            n=Count('id'), s=Sum('stock', output_field=BI),
            c=Sum(F('stock') * F('producto__costo'), output_field=BI))
        dead180_n = d180['n'] or 0
        dead180_u = d180['s'] or 0
        dead180_costo = d180['c'] or 0

        # -------- 11. clasificación ABC (motor de predicción, snapshot batch) --------
        abc = None
        try:
            from .models.predicciones import ClasificacionABC
            abc_qs = ClasificacionABC.objects.filter(articulo__atributo1_id=marca_id)
            max_anio = abc_qs.aggregate(x=Max('anio'))['x']
            if max_anio:
                abc_qs = abc_qs.filter(anio=max_anio)
                dist = {r['clasificacion_abc']: r['n'] for r in
                        abc_qs.values('clasificacion_abc').annotate(n=Count('articulo', distinct=True))}
                fecha_abc = abc_qs.aggregate(x=Max('fecha_calculo'))['x']
                abc = {'A': dist.get('A', 0), 'B': dist.get('B', 0), 'C': dist.get('C', 0),
                       'anio': max_anio,
                       'fecha': fecha_abc.strftime('%Y-%m-%d') if fecha_abc else None}
        except Exception:
            logger.warning('ABC no disponible para marca %s', marca_id)

        # -------- 12. lead time (proveedor dominante de la marca) --------
        lead = {'dias': 21, 'proveedor': None, 'fuente': 'default'}
        try:
            articulos = list(Producto.objects.filter(atributo1_id=marca_id)
                             .values_list('articulo', flat=True).distinct()[:500])
            if articulos:
                top_prov = (Compras_Producto.objects.filter(nombre__in=articulos)
                            .values('compras__empresa__nombre', 'compras__empresa__lead_time_dias')
                            .annotate(n=Count('id')).order_by('-n').first())
                if top_prov and top_prov['compras__empresa__lead_time_dias']:
                    lead = {'dias': top_prov['compras__empresa__lead_time_dias'],
                            'proveedor': top_prov['compras__empresa__nombre'], 'fuente': 'proveedor'}
        except Exception:
            logger.warning('Lead time no disponible para marca %s', marca_id)

        # -------- 13. liquidación (dado el sobre-stock, cuánto/qué liberar) --------
        tallas_sobre = [c['talla'] for c in curva if c['flag'] == 'sobre_stock']
        if cobertura and cobertura > 12:
            liq_texto = (f'Liberar capital: {dead180_n:,} SKUs sin venta en 180 días '
                         f'({dead180_u:,} pares, ${dead180_costo:,} a costo) son candidatos a liquidación/markdown.')
        elif dead180_u > 0:
            liq_texto = (f'{dead180_n:,} SKUs sin venta 180d ({dead180_u:,} pares) para depurar; '
                         f'el resto rota bien.')
        else:
            liq_texto = 'Inventario sano: sin dead-stock relevante.'
        liquidacion = {'skus': dead180_n, 'unidades': dead180_u, 'valor_costo': dead180_costo,
                       'tallas_sobre': tallas_sobre, 'texto': liq_texto}

        data = {
            'marca': marca.valor,
            'scope': 'Todas las tiendas' if not tienda_especifica else
                     next((s.alias for s in sucursales if s.id == tienda_especifica), ''),
            'kpis': {
                'venta_publica_hist': tot_pub,
                'distribucion_bodega': tot_bod,
                'stock_tiendas': stock_tiendas,
                'vel_dia': round(v90 / 90, 1),
                'vel_yoy': round((v90 / v90_ly - 1) * 100, 1) if v90_ly else None,
                'pronostico_12m': fc['forecast_total'],
                'cobertura_meses': cobertura,
            },
            'finanzas': {
                'inv_costo': inv_costo, 'inv_precio': inv_precio,
                'margen_pct': margen_pct, 'margen_src': margen_src,
                'rotacion': rotacion, 'wos': wos, 'gmroi': gmroi,
            },
            'salud': {
                'skus_total': skus_total, 'dead90_n': dead90_n, 'dead180_n': dead180_n,
                'dead180_u': dead180_u, 'dead180_costo': dead180_costo,
                'pct_dead180': round(100.0 * dead180_n / skus_total, 1) if skus_total else None,
                'abc': abc, 'lead_time': lead,
            },
            'liquidacion': liquidacion,
            'ventas_anual': ventas_anual,
            'ventas_por_tienda': ventas_por_tienda,
            'distribucion_bodega': distribucion_bodega,
            'stock_suc': stock_suc,
            'serie_chart': serie_chart,
            'forecast': fc,
            'sellthrough': sellthrough,
            'curva': curva,
            'recomendacion': {
                'veredicto': veredicto, 'comprar': comprar, 'cobertura_meses': cobertura,
                'ttm': fc['ttm'], 'yoy': fc['yoy'], 'lead_dias': lead['dias'],
                'curva_compra': curva_compra, 'evitar_tallas': evitar, 'asignacion': asignacion,
            },
        }
        return JsonResponse({'success': True, 'data': data})

    except Exception as e:
        logger.exception('Error en inteligencia de compra')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ══════════════════════════════════════════════════════════════
#  PLAN DE LIQUIDACIÓN — ranking consolidado de TODAS las marcas
# ══════════════════════════════════════════════════════════════
@requiere_permiso('plan_liquidacion', 'puede_ver')
def ver_plan_liquidacion(request):
    return render(request, 'vistas/modulo_reportes/plan_liquidacion.html', {})


# ------------------------------------------ plan de liquidación (v2)
NOMBRE_ATRIBUTO_ESPECIALIDAD = 'Especialidad'
ES_TIENDA_PT = Q(producto__sucursal__es_centro_distribucion=False)
ES_CD_PT = Q(producto__sucursal__es_centro_distribucion=True)
MAX_EXPORT_FILAS = 20000
BUCKETS_ANTIGUEDAD = {
    '0-90': (0, 90), '90-180': (90, 180), '180-365': (180, 365), '365+': (365, None),
}
# Descuento de liquidación sugerido por antigüedad (mismo criterio que
# analisis_inventario_antiguo: 6 meses / 1 año / 2 años -> 15 / 25 / 40 %).
ESCALA_DESCUENTO_LIQUIDACION = [(730, 40), (365, 25), (180, 15)]
PISO_COSTO_FACTOR = 1.1

# Tramos de antigüedad por DÍAS EXACTOS desde hoy — NO por año calendario.
# Los cortes son los mismos umbrales de ESCALA_DESCUENTO_LIQUIDACION, así el
# % del tramo es EXACTAMENTE el % que lleva cada fila del detalle y el mismo
# producto ya no puede salir con 0% en su fila y 25% en el resumen (bucket
# por año calendario sobrestimaba ≥$220,7M el "1 año a liquidar" — auditoría
# 2026-08, P1-7b). (clave, label, días desde, días hasta exclusivo).
TRAMOS_ANTIGUEDAD = (
    ('t0', '<6 meses', 0, 180),
    ('t1', '6-12 meses', 180, 365),
    ('t2', '1-2 años', 365, 730),
    ('t3', '2+ años', 730, None),
)


def _descuento_sugerido(dias):
    """% de descuento de liquidación sugerido según días de antigüedad."""
    if dias is None:
        return 0
    for umbral, pct in ESCALA_DESCUENTO_LIQUIDACION:
        if dias >= umbral:
            return pct
    return 0


def _tramo_de_dias(dias):
    """Clave de TRAMOS_ANTIGUEDAD para una antigüedad en días exactos.
    Espeja el orden del Case de `_tramo_case` (None => sin dato)."""
    if dias is None:
        return None
    if dias >= 730:
        return 't3'
    if dias >= 365:
        return 't2'
    if dias >= 180:
        return 't1'
    return 't0'


def _precio_liq_sugerido(precioventa, costo, dias):
    """Precio de liquidación sugerido (precioventa - descuento sugerido), con
    piso en costo*1.1. Devuelve (precio, descuento_pct)."""
    pct = _descuento_sugerido(dias)
    if not pct:
        return (precioventa or 0), 0
    nuevo = int(round((precioventa or 0) * (1 - pct / 100.0)))
    piso = int((costo or 0) * PISO_COSTO_FACTOR)
    return max(nuevo, piso), pct


def _scope_plan(request):
    """Alcance y filtros GET comunes del plan de liquidación.

    Devuelve (base_pt, mov_base, ctx):
      - base_pt: Producto_Talla con stock>0 en el alcance. Con incluir_cd=1
        entran también las bodegas/CD del usuario (su stock se reporta
        aparte, nunca en los denominadores de rotación/GMROI).
      - mov_base: ventas del kardex SIEMPRE solo en tiendas vendedoras.
      - ctx: filtros parseados + sucursales del usuario.
    """
    sucursales = list(_sucursales_usuario(request))
    tiendas_ids = [s.id for s in sucursales if not s.es_centro_distribucion]
    cd_ids = [s.id for s in sucursales if s.es_centro_distribucion]

    incluir_cd = request.GET.get('incluir_cd') == '1'
    sucursal_id = request.GET.get('sucursal_id') or None
    marca_id = request.GET.get('marca_id') or None
    especialidad_id = request.GET.get('especialidad_id') or None
    categoria_id = request.GET.get('categoria_id') or None

    alcance_ids = tiendas_ids + (cd_ids if incluir_cd else [])
    if sucursal_id:
        # Elegir una bodega explícitamente la incluye aunque incluir_cd=0.
        sid = int(sucursal_id)
        alcance_ids = [sid] if sid in tiendas_ids + cd_ids else []

    base_pt = Producto_Talla.objects.filter(
        stock__gt=0, producto__excluir_de_analitica=False,
        producto__sucursal_id__in=alcance_ids,
    )
    mov_scope = [i for i in alcance_ids if i in tiendas_ids]
    # Mismo universo que base_pt: sin excluir_de_analitica aquí, las ventas
    # de productos excluidos (bolsas IMP/PA00, etc.) inflaban rotación/GMROI
    # de la marca y el plan recomendaba "Reponer" lo que había que liquidar
    # (PAOLA: rotación 147 con exclusiones vs 0,56 real — auditoría 20-ago).
    mov_base = Movimientos_Producto.objects.filter(
        estado='COMPLETADO', concepto__in=CONCEPTOS_VENTA,
        ProductoTalla__producto__sucursal_id__in=mov_scope,
        ProductoTalla__producto__excluir_de_analitica=False,
    )

    if categoria_id:
        from app.views_modulo_reportes import _expandir_categoria_ids
        cat_scope = _expandir_categoria_ids(categoria_id)
        base_pt = base_pt.filter(producto__categoria_id__in=cat_scope)
        mov_base = mov_base.filter(ProductoTalla__producto__categoria_id__in=cat_scope)
    if marca_id:
        base_pt = base_pt.filter(producto__atributo1_id=marca_id)
        mov_base = mov_base.filter(ProductoTalla__producto__atributo1_id=marca_id)
    if especialidad_id:
        # Subconsulta IN (no join): multi-etiqueta sin duplicar filas.
        prods_esp = ProductoAtributoValor.objects.filter(
            atributo__nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD,
            opcion_id=especialidad_id,
        ).values('producto_id')
        base_pt = base_pt.filter(producto_id__in=prods_esp)
        mov_base = mov_base.filter(ProductoTalla__producto_id__in=prods_esp)

    ctx = {
        'sucursales': sucursales, 'tiendas_ids': tiendas_ids, 'cd_ids': cd_ids,
        'incluir_cd': incluir_cd, 'sucursal_id': sucursal_id,
        'marca_id': marca_id, 'especialidad_id': especialidad_id,
        'categoria_id': categoria_id,
    }
    return base_pt, mov_base, ctx


def _opciones_filtros_plan(ctx):
    """Catálogos para poblar los selects de filtros (solo primera carga)."""
    suc_ids = ctx['tiendas_ids'] + ctx['cd_ids']
    marcas = list(
        AtributoOpcion.objects.filter(
            productos_marca__sucursal_id__in=suc_ids,
            productos_marca__producto_talla__stock__gt=0,
        ).values('id', 'valor').distinct().order_by('valor'))
    categorias = list(
        Categoria.objects.select_related('padre')
        .values('id', 'nombre', 'padre_id', 'padre__nombre')
        .order_by('padre__nombre', 'nombre'))
    especialidades = list(
        AtributoOpcion.objects.filter(
            atributo__nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD,
        ).values('id', 'valor').order_by('valor'))
    sucursales = [{'id': s.id, 'alias': s.alias, 'es_cd': s.es_centro_distribucion}
                  for s in ctx['sucursales']]
    return {'marcas': marcas, 'categorias': categorias,
            'especialidades': especialidades, 'sucursales': sucursales}


def _fila_liquidacion(ir, t, dd):
    """Métricas de una fila del ranking (dimensión agnóstica).

    ir: agregados de inventario; t: ventas TTM 365d; dd: dead-stock 180d.
    Ratios (rotación/cobertura/GMROI) usan SOLO stock/ventas de tiendas; el
    stock CD se exhibe aparte. Fila solo-bodega => acción 'Traspasar'.
    """
    stock = ir.get('stock_u') or 0
    stock_cd = ir.get('stock_cd') or 0
    if stock + stock_cd <= 0:
        return None
    valor_costo = ir.get('valor_costo') or 0
    valor_cd = ir.get('valor_cd') or 0
    skus = ir.get('skus') or 0
    u = t.get('u') or 0
    venta = t.get('venta') or 0
    costo = t.get('costo') or 0
    dead_u = dd.get('dead_u') or 0
    dead_costo = dd.get('dead_costo') or 0
    dead_skus = dd.get('dead_skus') or 0

    rotacion = round(u / stock, 2) if stock else None
    cobertura = round(stock / (u / 12.0), 1) if (u and stock) else None
    margen = (venta - costo) if (venta > 0 and 0 < costo < venta) else None
    gmroi = round(margen / valor_costo, 2) if (margen and valor_costo) else None
    pct_dead = round(100.0 * dead_skus / skus, 1) if skus else 0

    if stock == 0 and stock_cd > 0:
        accion = 'Traspasar'
    elif gmroi is not None and gmroi < 1 and (cobertura is None or cobertura > 12):
        accion = 'Liquidar'
    elif gmroi is not None and gmroi >= 3 and cobertura is not None and cobertura < 6:
        accion = 'Reponer'
    elif pct_dead >= 70:
        accion = 'Depurar'
    else:
        accion = 'Monitorear'

    return {
        'stock': stock, 'stock_cd': stock_cd, 'skus': skus,
        'valor_costo': valor_costo, 'valor_cd': valor_cd, 'ttm_u': u,
        'rotacion': rotacion, 'cobertura': cobertura, 'gmroi': gmroi,
        'dead_u': dead_u, 'dead_costo': dead_costo, 'dead_skus': dead_skus,
        'pct_dead': pct_dead, 'accion': accion,
    }


@require_GET
@requiere_permiso('plan_liquidacion', 'puede_ver')
def obtener_plan_liquidacion(request):
    """Ranking por capital inmovilizado (dead-stock 180d) + GMROI/rotación.

    v2 — dimensiones: marca, categoría hija v1.2 ("Padre › Hijo"),
    especialidad (multi-etiqueta, vista de atribución) y sucursal. Filtros
    GET combinables: categoria_id (un padre incluye su rama), marca_id,
    especialidad_id, sucursal_id, incluir_cd. Ventas/rotación/GMROI se
    calculan SIEMPRE solo con tiendas vendedoras; el stock de bodegas/CD
    (incluir_cd=1) se reporta aparte (stock_cd/valor_cd, acción Traspasar).
    `con_opciones=1` agrega los catálogos para poblar los selects.
    """
    try:
        hoy = timezone.localdate()
        base_pt, mov_base, ctx = _scope_plan(request)

        vend180 = (mov_base.filter(fecha__gte=hoy - timedelta(days=180))
                   .values('ProductoTalla'))

        # ---- Fase C (perf): marca/categoría/sucursal comparten base. Se
        # agrega UNA vez por la terna (marca, categoría, sucursal) y cada
        # dimensión (y los totales) se deriva sumando parciales enteros:
        # mismo resultado que las 9 queries anteriores con 3. Especialidad
        # sigue aparte (multi-etiqueta: su join duplica filas).
        K_MAR, K_CAT, K_SUC = ('producto__atributo1_id', 'producto__categoria_id',
                               'producto__sucursal_id')
        MK_MAR, MK_CAT, MK_SUC = ('ProductoTalla__producto__atributo1_id',
                                  'ProductoTalla__producto__categoria_id',
                                  'ProductoTalla__producto__sucursal_id')
        inv_rows = list(
            base_pt.values(K_MAR, 'producto__atributo1__valor',
                           K_CAT, 'producto__categoria__nombre',
                           'producto__categoria__padre__nombre',
                           K_SUC, 'producto__sucursal__alias',
                           'producto__sucursal__es_centro_distribucion')
            .annotate(stock_u=Sum('stock', filter=ES_TIENDA_PT, output_field=BI),
                      stock_cd=Sum('stock', filter=ES_CD_PT, output_field=BI),
                      valor_costo=Sum(F('stock') * F('producto__costo'),
                                      filter=ES_TIENDA_PT, output_field=BI),
                      valor_cd=Sum(F('stock') * F('producto__costo'),
                                   filter=ES_CD_PT, output_field=BI),
                      skus=Count('id', filter=ES_TIENDA_PT)))
        mv_rows = list(
            mov_base.filter(fecha__gte=hoy - timedelta(days=365))
            .values(MK_MAR, MK_CAT, MK_SUC)
            .annotate(u=Sum(Abs('cantidad'), output_field=BI),
                      venta=Sum(Abs(F('cantidad')) * F('precio'), output_field=BI),
                      costo=Sum(Abs(F('cantidad')) * F('costo'), output_field=BI)))
        dead_rows = list(
            base_pt.filter(ES_TIENDA_PT).exclude(id__in=vend180)
            .values(K_MAR, K_CAT, K_SUC)
            .annotate(dead_u=Sum('stock', output_field=BI),
                      dead_costo=Sum(F('stock') * F('producto__costo'), output_field=BI),
                      dead_skus=Count('id')))

        CAMPOS_INV = ('stock_u', 'stock_cd', 'valor_costo', 'valor_cd', 'skus')
        CAMPOS_MV = ('u', 'venta', 'costo')
        CAMPOS_DEAD = ('dead_u', 'dead_costo', 'dead_skus')

        def _acum(rows, key, campos, etiquetas=()):
            """Suma parciales agrupando por rows[key]. Los labels dependen
            funcionalmente del id, así que basta la primera aparición."""
            out = {}
            for r in rows:
                acc = out.get(r[key])
                if acc is None:
                    acc = {c: 0 for c in campos}
                    for e in etiquetas:
                        acc[e] = r.get(e)
                    out[r[key]] = acc
                for c in campos:
                    acc[c] += r.get(c) or 0
            return out

        def _filas_dim(key, mkey, etiquetas, nombre_key, etiquetar,
                       excluir_none=True, extra_fila=None):
            """Ranking de una dimensión, derivado de las pasadas comunes.
            `excluir_none` espeja el filtro isnull=False que la versión por
            queries aplicaba a marca/categoría."""
            inv = _acum([r for r in inv_rows
                         if not (excluir_none and r[key] is None)],
                        key, CAMPOS_INV, etiquetas)
            ttm = _acum(mv_rows, mkey, CAMPOS_MV)
            dead = _acum(dead_rows, key, CAMPOS_DEAD)
            filas = []
            for did, ir in inv.items():
                fila = _fila_liquidacion(ir, ttm.get(did, {}), dead.get(did, {}))
                if fila is None:
                    continue
                fila[nombre_key] = etiquetar(ir)
                fila['id'] = did
                if extra_fila:
                    fila.update(extra_fila(ir))
                filas.append(fila)
            filas.sort(key=lambda x: (-(x['dead_costo'] or 0), -(x['valor_cd'] or 0)))
            return filas

        filas = _filas_dim(
            K_MAR, MK_MAR, ('producto__atributo1__valor',), 'marca',
            lambda ir: ir['producto__atributo1__valor'],
        )

        def _label_cat(ir):
            nombre = ir.get('producto__categoria__nombre') or 'Sin categoría'
            padre = ir.get('producto__categoria__padre__nombre') or ''
            return f'{padre} › {nombre}' if padre else nombre

        filas_cat = _filas_dim(
            K_CAT, MK_CAT,
            ('producto__categoria__nombre', 'producto__categoria__padre__nombre'),
            'categoria', _label_cat,
        )

        filas_suc = _filas_dim(
            K_SUC, MK_SUC,
            ('producto__sucursal__alias', 'producto__sucursal__es_centro_distribucion'),
            'sucursal', lambda ir: ir['producto__sucursal__alias'],
            excluir_none=False,
            extra_fila=lambda ir: {
                'es_cd': bool(ir['producto__sucursal__es_centro_distribucion'])},
        )
        filas_suc.sort(key=lambda x: (x.get('es_cd', False),
                                      -(x['dead_costo'] or 0), -(x['valor_cd'] or 0)))

        def _filas_por_especialidad():
            """Dimensión especialidad v1.2 — multi-etiqueta => ATRIBUCIÓN:
            un producto suma en cada especialidad que tiene, así que el
            total de la tabla puede exceder el 100% del inventario."""
            ESP = Q(producto__atributos__atributo__nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD)
            k = 'producto__atributos__opcion_id'
            inv = {r[k]: r for r in
                   base_pt.filter(ESP)
                   .values(k, 'producto__atributos__opcion__valor')
                   .annotate(stock_u=Sum('stock', filter=ES_TIENDA_PT, output_field=BI),
                             stock_cd=Sum('stock', filter=ES_CD_PT, output_field=BI),
                             valor_costo=Sum(F('stock') * F('producto__costo'),
                                             filter=ES_TIENDA_PT, output_field=BI),
                             valor_cd=Sum(F('stock') * F('producto__costo'),
                                          filter=ES_CD_PT, output_field=BI),
                             skus=Count('id', filter=ES_TIENDA_PT))}
            km = 'ProductoTalla__producto__atributos__opcion_id'
            mv = (mov_base.filter(
                      fecha__gte=hoy - timedelta(days=365),
                      ProductoTalla__producto__atributos__atributo__nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD)
                  .values(km)
                  .annotate(u=Sum(Abs('cantidad'), output_field=BI),
                            venta=Sum(Abs(F('cantidad')) * F('precio'), output_field=BI),
                            costo=Sum(Abs(F('cantidad')) * F('costo'), output_field=BI)))
            ttm = {r[km]: r for r in mv}
            dead = {r[k]: r for r in
                    base_pt.filter(ESP & ES_TIENDA_PT).exclude(id__in=vend180)
                    .values(k)
                    .annotate(dead_u=Sum('stock', output_field=BI),
                              dead_costo=Sum(F('stock') * F('producto__costo'), output_field=BI),
                              dead_skus=Count('id'))}
            filas_e = []
            for did, ir in inv.items():
                fila = _fila_liquidacion(ir, ttm.get(did, {}), dead.get(did, {}))
                if fila is None:
                    continue
                fila['especialidad'] = ir['producto__atributos__opcion__valor']
                fila['id'] = did
                filas_e.append(fila)
            filas_e.sort(key=lambda x: (-(x['dead_costo'] or 0), -(x['valor_cd'] or 0)))
            return filas_e

        filas_esp = _filas_por_especialidad()

        # Totales sobre el universo filtrado completo (no solo filas con
        # marca): TODAS las filas de la pasada común, ids nulos incluidos.
        tot = {'stock_u': 0, 'valor_costo': 0, 'stock_cd': 0, 'valor_cd': 0}
        for r in inv_rows:
            for c in tot:
                tot[c] += r.get(c) or 0
        dead_tot = {'dead_u': 0, 'dead_costo': 0}
        for r in dead_rows:
            for c in dead_tot:
                dead_tot[c] += r.get(c) or 0
        tot_valor = tot['valor_costo'] or 0
        tot_dead_costo = dead_tot['dead_costo'] or 0

        data = {
            'totales': {
                'marcas': len(filas),
                'stock_u': tot['stock_u'] or 0, 'valor_costo': tot_valor,
                'stock_cd': tot['stock_cd'] or 0, 'valor_cd': tot['valor_cd'] or 0,
                'dead_u': dead_tot['dead_u'] or 0, 'dead_costo': tot_dead_costo,
                'dead_pct_valor': round(100.0 * tot_dead_costo / tot_valor, 1) if tot_valor else 0,
                'n_liquidar': sum(1 for f in filas if f['accion'] == 'Liquidar'),
                'n_reponer': sum(1 for f in filas if f['accion'] == 'Reponer'),
            },
            'marcas': filas,
            'categorias': filas_cat,
            'especialidades': filas_esp,
            'sucursales': filas_suc,
            'filtros_aplicados': {
                'categoria_id': ctx['categoria_id'], 'marca_id': ctx['marca_id'],
                'especialidad_id': ctx['especialidad_id'],
                'sucursal_id': ctx['sucursal_id'], 'incluir_cd': ctx['incluir_cd'],
            },
        }
        if request.GET.get('con_opciones') == '1':
            data['opciones'] = _opciones_filtros_plan(ctx)
        return JsonResponse({'success': True, 'data': data})

    except Exception as e:
        logger.exception('Error en plan de liquidación')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Orden GET -> expresión de order_by en el queryset de Producto anotado.
# 'antiguedad' ordena por la fecha FIFO (más viejo = más días); 'sin_venta'
# por la última venta (venta más vieja / nunca = más días sin vender).
ORDENES_DETALLE = {
    'valor_costo': F('valor_ord'), 'stock_u': F('stock_u'),
    'precioventa': F('precioventa'), 'articulo': F('articulo'),
    'dias_antiguedad': F('aging_dt'), 'dias_sin_venta': F('ultima_venta'),
    'u365': F('u365'),
}


def _detalle_query(base_pt, mov_base, hoy, q=None, con_ventas=True):
    """Queryset de Producto anotado para el drill-down, paginable en el DB.

    Todo el cálculo pesado (agregados + antigüedad FIFO + última venta) se
    resuelve con anotaciones/Subquery para que Postgres ordene y pagine
    (LIMIT/OFFSET) sin materializar los ~15k productos del catálogo en Python.
    Antigüedad = lote vivo más antiguo; si el producto no tiene lotes (stock
    migrado), `fecha_creacion` (corregida en prod desde el kardex el
    2026-05-20). Ventas siempre solo tiendas (mov_base ya viene acotado).

    `con_ventas=False` (Fase C perf): omite los Subquery correlacionados de
    última venta y u365 — el EXPORT los resuelve con 2 pasadas agrupadas en
    `_filas_export` (con ~15k filas los 2 subqueries por fila eran el grueso
    de los 33s del Excel); el detalle paginado los sigue usando porque
    ordena/pagina por ellos en el DB.
    """
    prod = Producto.objects.filter(id__in=base_pt.values('producto_id'))
    if q:
        prod = prod.filter(Q(articulo__icontains=q) |
                           Q(descripcion__icontains=q) |
                           Q(atributo1__valor__icontains=q))

    lote_sq = (LoteProducto.objects.filter(
        producto_talla__producto=OuterRef('pk'),
        activo=True, agotado=False, cantidad_disponible__gt=0,
    ).order_by('fecha_ingreso').values('fecha_ingreso')[:1])

    stock_f = Q(producto_talla__stock__gt=0)
    annots = dict(
        stock_u=Coalesce(Sum('producto_talla__stock', filter=stock_f, output_field=BI), 0),
        tallas=Count('producto_talla', filter=stock_f),
        fecha_lote=Subquery(lote_sq),
    )
    if con_ventas:
        venta_sq = (mov_base.filter(ProductoTalla__producto=OuterRef('pk'))
                    .order_by('-fecha').values('fecha')[:1])
        u365_sq = (mov_base.filter(ProductoTalla__producto=OuterRef('pk'),
                                   fecha__gte=hoy - timedelta(days=365))
                   .values('ProductoTalla__producto')
                   .annotate(s=Sum(Abs('cantidad'), output_field=BI)).values('s')[:1])
        annots['ultima_venta'] = Subquery(venta_sq)
        annots['u365'] = Coalesce(Subquery(u365_sq, output_field=BI), 0)

    return prod.annotate(**annots).annotate(
        aging_dt=Coalesce('fecha_lote', 'fecha_creacion'),
        valor_ord=F('costo') * F('stock_u'),
    )


def _bucket_filter(qs, bucket, hoy):
    """Aplica el rango de antigüedad al queryset anotado (sobre aging_dt)."""
    if bucket == 'sin-dato':
        return qs.filter(aging_dt__isnull=True)
    rango = BUCKETS_ANTIGUEDAD.get(bucket)
    if not rango:
        return qs
    lo, hi = rango
    # dias in [lo, hi)  <=>  aging_date in (hoy-hi, hoy-lo]
    qs = qs.filter(aging_dt__date__lte=hoy - timedelta(days=lo))
    if hi is not None:
        qs = qs.filter(aging_dt__date__gt=hoy - timedelta(days=hi))
    return qs


def _parse_tramos(raw):
    """'t1,t2' -> tuplas de TRAMOS_ANTIGUEDAD (ignora claves desconocidas)."""
    claves = {p.strip() for p in (raw or '').split(',') if p.strip()}
    return [t for t in TRAMOS_ANTIGUEDAD if t[0] in claves]


def _tramo_filter(qs, tramos, hoy):
    """Filtra el queryset del detalle por TRAMO(s) de antigüedad en días
    exactos (sobre aging_dt) — mismo criterio que el gráfico y el Excel."""
    if not tramos:
        return qs
    cond = Q()
    for _clave, _label, d0, d1 in tramos:
        # dias en [d0, d1)  <=>  aging_date in (hoy-d1, hoy-d0]
        if d0 > 0:
            c = Q(aging_dt__date__lte=hoy - timedelta(days=d0))
        else:
            c = Q(aging_dt__isnull=False)  # incluye fechas futuras (días<0)
        if d1 is not None:
            c &= Q(aging_dt__date__gt=hoy - timedelta(days=d1))
        cond |= c
    return qs.filter(cond)


def _serializar_detalle(qs, hoy, ventas_bulk=None):
    """Convierte una página del queryset (ya sliceada) en dicts JSON.

    `ventas_bulk=(venta_map, u365_map)` (Fase C perf): última venta y u365
    resueltos por mapas {producto_id: valor} de 2 queries agrupadas, para el
    queryset liviano de `_detalle_query(..., con_ventas=False)`. Sin mapas,
    se leen de las anotaciones (comportamiento original del detalle paginado).
    """
    campos = [
        'id', 'articulo', 'descripcion', 'atributo1__valor', 'atributo2__valor',
        'categoria__nombre', 'categoria__padre__nombre', 'sucursal_id',
        'sucursal__alias', 'sucursal__es_centro_distribucion', 'precioventa',
        'costo', 'stock_u', 'tallas', 'fecha_lote', 'fecha_creacion']
    if ventas_bulk is None:
        campos += ['ultima_venta', 'u365']
    filas = list(qs.values(*campos))
    out = []
    for r in filas:
        if ventas_bulk is not None:
            r['ultima_venta'] = ventas_bulk[0].get(r['id'])
            r['u365'] = ventas_bulk[1].get(r['id'], 0)
        f = r['fecha_lote'] or r['fecha_creacion']
        fuente = 'lote' if r['fecha_lote'] else ('creacion' if r['fecha_creacion'] else None)
        fecha_fifo = timezone.localtime(f).date() if f else None
        ult = r['ultima_venta']
        cat_n = r['categoria__nombre'] or 'Sin categoría'
        cat_p = r['categoria__padre__nombre'] or ''
        dias = (hoy - fecha_fifo).days if fecha_fifo else None
        precio_liq, pct = _precio_liq_sugerido(r['precioventa'], r['costo'], dias)
        out.append({
            'producto_id': r['id'], 'articulo': r['articulo'],
            'descripcion': r['descripcion'], 'marca': r['atributo1__valor'],
            'color': r['atributo2__valor'],
            'categoria': f'{cat_p} › {cat_n}' if cat_p else cat_n,
            'sucursal_id': r['sucursal_id'], 'sucursal': r['sucursal__alias'],
            'es_cd': bool(r['sucursal__es_centro_distribucion']),
            'tallas': r['tallas'], 'stock_u': r['stock_u'] or 0,
            'valor_costo': (r['stock_u'] or 0) * (r['costo'] or 0),
            'precioventa': r['precioventa'], 'costo': r['costo'],
            'fecha_fifo': fecha_fifo, 'antiguedad_fuente': fuente,
            'anio': fecha_fifo.year if fecha_fifo else None,
            'ultima_venta': ult,
            'dias_sin_venta': (hoy - ult).days if ult else None,
            'u365': r['u365'] or 0,
            'dias_antiguedad': dias,
            'descuento_sugerido': pct,
            'precio_liquidacion': precio_liq,
        })
    return out


@require_GET
@requiere_permiso('plan_liquidacion', 'puede_ver')
def obtener_plan_liquidacion_detalle(request):
    """Drill-down del plan de liquidación a nivel Producto, paginado en el DB.

    GET: filtros de _scope_plan + antiguedad (0-90|90-180|180-365|365+|
    sin-dato), q (artículo/descripción/marca), orden (dias_antiguedad|
    dias_sin_venta|valor_costo|stock_u|u365|precioventa|articulo, prefijo
    '-' = descendente), page, page_size (máx 200).
    """
    try:
        hoy = timezone.localdate()
        base_pt, mov_base, ctx = _scope_plan(request)
        qs = _detalle_query(base_pt, mov_base, hoy, q=request.GET.get('q') or None)

        bucket = request.GET.get('antiguedad') or None
        if bucket:
            qs = _bucket_filter(qs, bucket, hoy)
        qs = _tramo_filter(qs, _parse_tramos(request.GET.get('tramo')), hoy)

        orden = request.GET.get('orden') or '-valor_costo'
        rev = orden.startswith('-')
        key = orden.lstrip('-')
        expr = ORDENES_DETALLE.get(key, F('valor_ord'))
        # nulls_last para no llenar la primera página de productos sin dato,
        # salvo dias_sin_venta desc (nunca vendido = más días => va primero).
        if key == 'dias_sin_venta':
            ordering = expr.asc(nulls_first=True) if rev else expr.desc(nulls_last=True)
        else:
            ordering = expr.desc(nulls_last=True) if rev else expr.asc(nulls_last=True)
        qs = qs.order_by(ordering, 'id')

        total = qs.count()
        try:
            page = max(int(request.GET.get('page') or 1), 1)
            page_size = min(max(int(request.GET.get('page_size') or 50), 10), 200)
        except (TypeError, ValueError):
            page, page_size = 1, 50
        ini = (page - 1) * page_size
        pagina = _serializar_detalle(qs[ini:ini + page_size], hoy)

        # Badge "en campaña activa" solo para los productos de la página.
        # Tolerante a que la migración 0188 aún no esté aplicada (si el código
        # se despliega antes de migrar, el reporte sigue funcionando sin badge).
        try:
            en_campana = set(CampanaLiquidacionProducto.objects.filter(
                activo=True, producto_id__in=[f['producto_id'] for f in pagina],
            ).values_list('producto_id', flat=True))
        except ProgrammingError:
            logger.warning('CampanaLiquidacionProducto no existe aún (migración 0188 pendiente)')
            en_campana = set()

        for f in pagina:
            f['fecha_fifo'] = f['fecha_fifo'].strftime('%Y-%m-%d') if f['fecha_fifo'] else None
            f['ultima_venta'] = f['ultima_venta'].strftime('%Y-%m-%d') if f['ultima_venta'] else None
            f['en_campana'] = f['producto_id'] in en_campana

        return JsonResponse({'success': True, 'total': total, 'page': page,
                             'page_size': page_size, 'filas': pagina})
    except Exception as e:
        logger.exception('Error en detalle de plan de liquidación')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _aging_dt_expr():
    """Antigüedad FIFO como fecha (lote vivo más antiguo o creación)."""
    lote_sq = (LoteProducto.objects.filter(
        producto_talla__producto=OuterRef('pk'),
        activo=True, agotado=False, cantidad_disponible__gt=0,
    ).order_by('fecha_ingreso').values('fecha_ingreso')[:1])
    return Coalesce(Subquery(lote_sq), F('fecha_creacion'))


def _tramo_case(hoy):
    """Case SQL: aging_dt (anotado) -> clave de TRAMOS_ANTIGUEDAD por días
    exactos desde hoy. Sin fecha => NULL (queda fuera, igual que antes)."""
    return Case(
        When(aging_dt__date__lte=hoy - timedelta(days=730), then=Value('t3')),
        When(aging_dt__date__lte=hoy - timedelta(days=365), then=Value('t2')),
        When(aging_dt__date__lte=hoy - timedelta(days=180), then=Value('t1')),
        When(aging_dt__isnull=False, then=Value('t0')),
        default=Value(None),
        output_field=CharField(),
    )


def _tramos_liquidacion(base_pt, hoy):
    """Capital + pares por TRAMO de antigüedad (días exactos desde hoy).

    Reemplaza el bucket por año calendario: "todo 2025 = 1 año = 25%"
    metía en '≥1 año' stock con lote de <365 días. Los cortes son los de
    ESCALA_DESCUENTO_LIQUIDACION, así el % del tramo coincide 1:1 con el
    % de cada fila del detalle y del Excel."""
    stock_f = Q(producto_talla__stock__gt=0)
    rows = (Producto.objects.filter(id__in=base_pt.values('producto_id'))
            .annotate(aging_dt=_aging_dt_expr())
            .annotate(tramo=_tramo_case(hoy))
            .values('tramo')
            .annotate(
                valor=Coalesce(Sum(F('producto_talla__stock') * F('costo'),
                                   filter=stock_f, output_field=BI), 0),
                pares=Coalesce(Sum('producto_talla__stock', filter=stock_f, output_field=BI), 0),
                productos=Count('id', distinct=True)))
    por_tramo = {r['tramo']: r for r in rows}
    tramos = []
    for clave, label, d0, d1 in TRAMOS_ANTIGUEDAD:
        r = por_tramo.get(clave)
        if not r:
            continue
        tramos.append({
            'tramo': clave, 'label': label,
            'dias_desde': d0, 'dias_hasta': d1,
            'antiguedad_anios': d0 // 365,
            'valor': r['valor'] or 0, 'pares': r['pares'] or 0,
            'productos': r['productos'] or 0,
            'descuento_sugerido': _descuento_sugerido(d0)})
    return tramos


def _dimension_liquidacion(base_pt, hoy, dim='marca', top_n=10, con_productos=False):
    """Top items de una dimensión (marca|especialidad) con su capital dividido
    en buckets de urgencia por DÍAS EXACTOS del lote (<1 año / 1-2 años /
    2+ años — mismo criterio que los tramos y el detalle). Especialidad es
    multi-etiqueta (atribución): un producto suma en cada una que tenga.

    `con_productos=True` (Fase C perf): agrega Count(distinct) de productos
    por grupo y devuelve (top, rows) para que `obtener_plan_liquidacion_por_anio`
    derive los tramos de ESTA misma pasada (el costo está en el subquery de
    aging por producto) en vez de re-escanear con `_tramos_liquidacion`.
    """
    stock_f = Q(producto_talla__stock__gt=0)
    qs = (Producto.objects.filter(id__in=base_pt.values('producto_id'))
          .annotate(aging_dt=_aging_dt_expr())
          .annotate(tramo=_tramo_case(hoy)))
    if dim == 'especialidad':
        qs = qs.filter(atributos__atributo__nombre__iexact=NOMBRE_ATRIBUTO_ESPECIALIDAD)
        campo, etiqueta = 'atributos__opcion__valor', 'especialidad'
    else:
        campo, etiqueta = 'atributo1__valor', 'marca'
    annots = dict(
        valor=Coalesce(Sum(F('producto_talla__stock') * F('costo'),
                           filter=stock_f, output_field=BI), 0),
        pares=Coalesce(Sum('producto_talla__stock', filter=stock_f, output_field=BI), 0))
    if con_productos:
        annots['productos'] = Count('id', distinct=True)
    rows = list(qs.values(campo, 'tramo').annotate(**annots))
    acc = {}
    for r in rows:
        nombre = r[campo] or f'Sin {etiqueta}'
        tramo = r['tramo']
        valor, pares = r['valor'] or 0, r['pares'] or 0
        it = acc.setdefault(nombre, {'nombre': nombre, 'val_reciente': 0,
                                     'val_1anio': 0, 'val_2mas': 0, 'valor': 0, 'pares': 0})
        it['valor'] += valor
        it['pares'] += pares
        if tramo in (None, 't0', 't1'):   # <1 año (o sin dato)
            it['val_reciente'] += valor
        elif tramo == 't2':               # 1-2 años
            it['val_1anio'] += valor
        else:                             # 't3' => 2+ años
            it['val_2mas'] += valor
    items = sorted(acc.values(), key=lambda x: -x['valor'])
    top = items[:top_n]
    if len(items) > top_n:
        otras = {'nombre': 'Otras', 'val_reciente': 0, 'val_1anio': 0,
                 'val_2mas': 0, 'valor': 0, 'pares': 0}
        for m in items[top_n:]:
            for k in ('val_reciente', 'val_1anio', 'val_2mas', 'valor', 'pares'):
                otras[k] += m[k]
        top.append(otras)
    if con_productos:
        return top, rows
    return top


def _tramos_desde_rows(rows):
    """Tramos de antigüedad derivados de las filas (grupo, tramo) de
    `_dimension_liquidacion(..., con_productos=True)`.

    Cada producto vive en exactamente UN (grupo, tramo) — atributo1 es una FK
    simple, no multi-etiqueta — así que sumar los parciales (valor, pares y
    Count distinct de productos) reproduce 1:1 la query dedicada de
    `_tramos_liquidacion` sin re-escanear el aging por producto."""
    por_tramo = {}
    for r in rows:
        acc = por_tramo.setdefault(r['tramo'],
                                   {'valor': 0, 'pares': 0, 'productos': 0})
        acc['valor'] += r['valor'] or 0
        acc['pares'] += r['pares'] or 0
        acc['productos'] += r.get('productos') or 0
    tramos = []
    for clave, label, d0, d1 in TRAMOS_ANTIGUEDAD:
        r = por_tramo.get(clave)
        if not r:
            continue
        tramos.append({
            'tramo': clave, 'label': label,
            'dias_desde': d0, 'dias_hasta': d1,
            'antiguedad_anios': d0 // 365,
            'valor': r['valor'] or 0, 'pares': r['pares'] or 0,
            'productos': r['productos'] or 0,
            'descuento_sugerido': _descuento_sugerido(d0)})
    return tramos


@require_GET
@requiere_permiso('plan_liquidacion', 'puede_ver')
def obtener_plan_liquidacion_por_anio(request):
    """Análisis de antigüedad para los gráficos: capital/pares por TRAMO de
    días exactos y por dimensión (marca|especialidad, param `dim`).
    `solo=dim` omite los tramos (para el toggle de dimensión, que no
    recalcula el gráfico de antigüedad). Era la única ruta del plan sin
    @requiere_permiso (auditoría 2026-08, P1-7a)."""
    try:
        hoy = timezone.localdate()
        base_pt, mov_base, ctx = _scope_plan(request)
        dim = request.GET.get('dim') or 'marca'
        solo = request.GET.get('solo')
        if solo == 'dim':
            por_dimension = _dimension_liquidacion(base_pt, hoy, dim=dim)
            return JsonResponse({'success': True, 'dim': dim, 'por_dimension': por_dimension})
        if dim == 'especialidad':
            # multi-etiqueta: sus filas duplican productos — tramos aparte.
            por_dimension = _dimension_liquidacion(base_pt, hoy, dim=dim)
            tramos = _tramos_liquidacion(base_pt, hoy)
        else:
            # Fase C (perf): dimensión y tramos comparten la MISMA pasada de
            # aging por producto (1 query en vez de 2, resultado idéntico).
            por_dimension, rows_dim = _dimension_liquidacion(
                base_pt, hoy, dim=dim, con_productos=True)
            tramos = _tramos_desde_rows(rows_dim)
        total_valor = sum(t['valor'] for t in tramos)
        total_pares = sum(t['pares'] for t in tramos)
        # "A liquidar" = tramos con descuento sugerido (≥6 meses)…
        valor_liquidar = sum(t['valor'] for t in tramos if (t['descuento_sugerido'] or 0) > 0)
        pares_liquidar = sum(t['pares'] for t in tramos if (t['descuento_sugerido'] or 0) > 0)
        # …y aparte el corte clásico "≥1 año" (ahora por días reales).
        valor_1anio_mas = sum(t['valor'] for t in tramos if t['dias_desde'] >= 365)
        pares_1anio_mas = sum(t['pares'] for t in tramos if t['dias_desde'] >= 365)
        return JsonResponse({'success': True, 'tramos': tramos, 'dim': dim,
                             'por_dimension': por_dimension,
                             'tramos_disponibles': [{'tramo': t['tramo'], 'label': t['label']}
                                                    for t in tramos],
                             'totales': {'valor': total_valor, 'pares': total_pares,
                                         'valor_liquidar': valor_liquidar,
                                         'pares_liquidar': pares_liquidar,
                                         'valor_1anio_mas': valor_1anio_mas,
                                         'pares_1anio_mas': pares_1anio_mas}})
    except Exception as e:
        logger.exception('Error en plan de liquidación por antigüedad')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _filas_export(request, hoy):
    """Filas serializadas del detalle filtrado (sin paginar, cap
    MAX_EXPORT_FILAS). Compartido por el Excel y el formulario de impresión.

    Fase C (perf): el export NO usa los Subquery por fila de última venta /
    u365 (con ~15k filas eran el grueso de los 33s del Excel): las mismas
    cifras salen de 2 pasadas agrupadas por producto sobre `mov_base`
    (Max(fecha) ≡ subquery order_by -fecha [:1]; Sum(|cantidad|) 365d ≡
    subquery agrupado) y se cruzan por id en la serialización.
    """
    base_pt, mov_base, ctx = _scope_plan(request)
    qs = _detalle_query(base_pt, mov_base, hoy, q=request.GET.get('q') or None,
                        con_ventas=False)
    bucket = request.GET.get('antiguedad') or None
    if bucket:
        qs = _bucket_filter(qs, bucket, hoy)
    qs = _tramo_filter(qs, _parse_tramos(request.GET.get('tramo')), hoy)
    # valor desc SOLO para que el cap priorice lo más valioso; el orden final
    # (marca → año → artículo asc) se aplica en Python por sucursal.
    qs = qs.order_by(F('valor_ord').desc(nulls_last=True), 'id')
    venta_map = {r['ProductoTalla__producto_id']: r['f'] for r in
                 mov_base.values('ProductoTalla__producto_id')
                 .annotate(f=Max('fecha'))}
    u365_map = {r['ProductoTalla__producto_id']: (r['s'] or 0) for r in
                mov_base.filter(fecha__gte=hoy - timedelta(days=365))
                .values('ProductoTalla__producto_id')
                .annotate(s=Sum(Abs('cantidad'), output_field=BI))}
    # truncado sin COUNT aparte: se pide UNA fila extra y se descarta.
    filas = _serializar_detalle(qs[:MAX_EXPORT_FILAS + 1], hoy,
                                ventas_bulk=(venta_map, u365_map))
    truncado = len(filas) > MAX_EXPORT_FILAS
    if truncado:
        filas = filas[:MAX_EXPORT_FILAS]
    return filas, ctx, bucket, truncado


def _agrupar_por_sucursal(filas):
    """{(es_cd, alias): [filas orden marca→año→artículo asc]} — tiendas
    alfabético primero, bodegas/CD al final."""
    grupos = {}
    for f in filas:
        grupos.setdefault((bool(f['es_cd']), f['sucursal'] or '—'), []).append(f)
    for g in grupos.values():
        g.sort(key=lambda f: ((f['marca'] or 'zzz').lower(),
                              f['anio'] or 9999,
                              (f['articulo'] or '').lower()))
    return dict(sorted(grupos.items()))


def _titulo_hoja(alias, es_cd):
    """Título de hoja Excel válido (≤31 chars, sin caracteres prohibidos)."""
    t = re.sub(r'[\[\]:*?/\\]', '-', alias or 'Sucursal')
    if es_cd:
        t += ' CD'
    return t[:31]


# Cabeceras del formato de verificación por tienda (Excel e impresión).
# ID SIEMPRE en la columna A: el re-import y la extracción por IA lo usan
# como clave. Las 3 últimas columnas van en blanco (se llenan a mano).
HEADERS_VERIFICACION = [
    'ID', 'N°', 'Marca', 'Año', 'Artículo', 'Descripción', 'Color', 'Tallas',
    'Pares', 'F. ingreso', 'Últ. venta', 'Precio lista', 'Desc. %',
    'Precio liq.', 'Precio caja $', '¿Coincide?', 'Observación',
]
ANCHOS_VERIFICACION = [9, 5, 14, 6, 16, 22, 12, 7, 7, 11, 11, 11, 8, 11, 12, 10, 24]


@require_GET
@requiere_permiso('plan_liquidacion', 'puede_exportar')
def exportar_plan_liquidacion_excel(request):
    """Excel de verificación: UNA hoja por sucursal (orden marca → año →
    artículo asc), con columnas en blanco para la verificación física
    (Precio caja / ¿Coincide? / Observación) y SIN costo en las hojas de
    tienda. Hoja Resumen (con costo, para el analista) + hoja Filtros.
    Mismos GET params que el detalle; cap MAX_EXPORT_FILAS."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties

        hoy = timezone.localdate()
        filas, ctx, bucket, truncado = _filas_export(request, hoy)
        tramos_sel = _parse_tramos(request.GET.get('tramo'))
        anios_txt = ', '.join(t[1] for t in tramos_sel) or 'Todos'
        grupos = _agrupar_por_sucursal(filas)

        wb = Workbook()
        wb.remove(wb.active)

        bold = Font(bold=True)
        head_fill = PatternFill('solid', fgColor='E8EAF0')
        thin = Side(style='thin', color='808080')
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)
        centrado = Alignment(horizontal='center')

        resumen_suc = {}    # (alias, anio) -> {pares, costo, lista, liq}
        resumen_tramo = {}  # clave de tramo (días exactos) -> {pares, costo, liq}
        for (es_cd, alias), items in grupos.items():
            ws = wb.create_sheet(_titulo_hoja(alias, es_cd))
            for i, ancho in enumerate(ANCHOS_VERIFICACION, start=1):
                ws.column_dimensions[get_column_letter(i)].width = ancho
            # Page setup imprimible: horizontal, ajustado al ancho, headers
            # repetidos en cada página.
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.page_margins.left = ws.page_margins.right = 0.3
            ws.page_margins.top = ws.page_margins.bottom = 0.5
            ws.print_title_rows = '1:2'
            ws.freeze_panes = 'A3'

            titulo = (f'Plan de Liquidación · {alias}{" (CD)" if es_cd else ""}'
                      f' · {hoy.strftime("%d-%m-%Y")} · Antigüedad: {anios_txt}')
            c0 = ws.cell(row=1, column=1, value=titulo)
            c0.font = Font(bold=True, size=13)
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=len(HEADERS_VERIFICACION))
            for c, h in enumerate(HEADERS_VERIFICACION, start=1):
                cell = ws.cell(row=2, column=c, value=h)
                cell.font = bold
                cell.fill = head_fill
                cell.border = borde
                cell.alignment = centrado

            tot_pares = tot_lista = tot_liq = 0
            fila_n = 3
            for n, f in enumerate(items, start=1):
                pares = f['stock_u'] or 0
                lista = f['precioventa'] or 0
                liq = f['precio_liquidacion'] or 0
                valores = [
                    f['producto_id'], n, f['marca'] or '—', f['anio'] or 's/d',
                    f['articulo'], f['descripcion'], f['color'] or '',
                    f['tallas'], pares,
                    f['fecha_fifo'].strftime('%d-%m-%Y') if f['fecha_fifo'] else 's/d',
                    f['ultima_venta'].strftime('%d-%m-%Y') if f['ultima_venta'] else 'Nunca',
                    lista, f['descuento_sugerido'] or 0, liq,
                    '', '', '',  # Precio caja / ¿Coincide? / Observación (a mano)
                ]
                for c, v in enumerate(valores, start=1):
                    ws.cell(row=fila_n, column=c, value=v).border = borde
                tot_pares += pares
                tot_lista += lista * pares
                tot_liq += liq * pares
                r = resumen_suc.setdefault((alias, f['anio'] or 0), {
                    'pares': 0, 'costo': 0, 'lista': 0, 'liq': 0})
                r['pares'] += pares
                r['costo'] += f['valor_costo'] or 0
                r['lista'] += lista * pares
                r['liq'] += liq * pares
                # Resumen por tramo con los MISMOS días exactos de la fila:
                # el % del tramo es el % que lleva la fila (antes el bloque
                # "por año" del Resumen usaba años×365 y contradecía al detalle).
                rt = resumen_tramo.setdefault(_tramo_de_dias(f['dias_antiguedad']), {
                    'pares': 0, 'costo': 0, 'liq': 0})
                rt['pares'] += pares
                rt['costo'] += f['valor_costo'] or 0
                rt['liq'] += liq * pares
                fila_n += 1

            for col, val in ((2, 'TOTAL'), (9, tot_pares), (12, tot_lista), (14, tot_liq)):
                cell = ws.cell(row=fila_n, column=col, value=val)
                cell.font = bold
                cell.border = borde

        # ---- Hoja Resumen (analista: aquí SÍ va el costo) ----
        ws_r = wb.create_sheet('Resumen')
        ws_r.append(['Sucursal', 'Año', 'Pares', 'Valor a costo',
                     'Valor precio lista', 'Valor liquidación est.'])
        for c in range(1, 7):
            ws_r.cell(row=1, column=c).font = bold
        for (alias, anio), r in sorted(resumen_suc.items()):
            ws_r.append([alias, anio or 's/d', r['pares'], r['costo'],
                         r['lista'], r['liq']])
        ws_r.append([])
        ws_r.append(['Tramo antigüedad', 'Días', 'Descuento sugerido %',
                     'Pares', 'Valor a costo', 'Valor liquidación est.'])
        # Tramos por días EXACTOS (mismo criterio que cada fila del detalle);
        # antes este bloque usaba años calendario ×365 y el mismo producto
        # podía llevar 0% en su fila y 25% aquí.
        meta_tramo = {clave: (label, d0, d1) for clave, label, d0, d1 in TRAMOS_ANTIGUEDAD}
        orden_tramo = {clave: i for i, (clave, _l, _d0, _d1) in enumerate(TRAMOS_ANTIGUEDAD)}
        for clave in sorted(resumen_tramo, key=lambda c: orden_tramo.get(c, 99)):
            a = resumen_tramo[clave]
            if clave is None:
                ws_r.append(['Sin dato', '—', 0, a['pares'], a['costo'], a['liq']])
                continue
            label, d0, d1 = meta_tramo[clave]
            rango = f'{d0}–{d1 - 1}' if d1 is not None else f'≥{d0}'
            ws_r.append([label, rango, _descuento_sugerido(d0),
                         a['pares'], a['costo'], a['liq']])

        # ---- Hoja Filtros ----
        ws2 = wb.create_sheet('Filtros')
        ws2.append(['Filtro', 'Valor'])
        ws2.append(['Generado', hoy.strftime('%Y-%m-%d')])
        for nombre, valor in _etiquetas_filtros(ctx):
            ws2.append([nombre, valor])
        ws2.append(['Tramos antigüedad', anios_txt])
        ws2.append(['Bucket antigüedad', bucket or 'Todos'])
        ws2.append(['Búsqueda', request.GET.get('q') or '—'])
        if truncado:
            ws2.append(['ADVERTENCIA', f'Export truncado a {MAX_EXPORT_FILAS} filas'])

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = (
            f'attachment; filename=plan_liquidacion_{hoy.strftime("%Y%m%d")}.xlsx')
        wb.save(resp)
        return resp
    except Exception as e:
        logger.exception('Error exportando plan de liquidación')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _fmt_clp(v):
    return f'{int(v or 0):,}'.replace(',', '.')


@require_GET
@requiere_permiso('plan_liquidacion', 'puede_exportar')
def imprimir_plan_liquidacion(request):
    """Formulario de verificación IMPRIMIBLE: una sección por tienda con
    layout fijo pensado para extracción por IA desde la foto del papel —
    IDs impresos a máquina, casillas ☐ SÍ ☐ NO, barcode CODE128 por tienda
    y encabezado repetido en cada página. Mismos GET params que el export."""
    hoy = timezone.localdate()
    filas, ctx, bucket, truncado = _filas_export(request, hoy)
    grupos = _agrupar_por_sucursal(filas)
    tramos_sel = _parse_tramos(request.GET.get('tramo'))
    anios_txt = ', '.join(t[1] for t in tramos_sel) or 'Todos'

    secciones = []
    for (es_cd, alias), items in grupos.items():
        for f in items:
            f['lista_fmt'] = _fmt_clp(f['precioventa'])
            f['liq_fmt'] = _fmt_clp(f['precio_liquidacion'])
            f['fecha_txt'] = f['fecha_fifo'].strftime('%d-%m-%Y') if f['fecha_fifo'] else 's/d'
        sid = items[0]['sucursal_id'] if items else 0
        secciones.append({
            'alias': alias, 'es_cd': es_cd,
            'codigo': f'LIQ-{sid}-{hoy.strftime("%Y%m%d")}',
            'items': items,
            'total_pares': sum(i['stock_u'] or 0 for i in items),
            'total_lista_fmt': _fmt_clp(sum((i['precioventa'] or 0) * (i['stock_u'] or 0) for i in items)),
            'total_liq_fmt': _fmt_clp(sum((i['precio_liquidacion'] or 0) * (i['stock_u'] or 0) for i in items)),
        })
    return render(request, 'vistas/modulo_reportes/plan_liquidacion_imprimir.html', {
        'secciones': secciones,
        'fecha': hoy,
        'anios': anios_txt,
        'truncado': truncado,
        'max_filas': MAX_EXPORT_FILAS,
    })


def _leer_archivo_verificacion(archivo):
    """Lee un .xlsx/.csv de verificación. Devuelve (ids, verificaciones):
    verificaciones = {producto_id: {'precio_caja', 'coincide', 'observacion'}}.

    XLSX: solo hojas cuyo header (fila 1 o 2) empiece con 'ID' — las hojas
    por tienda del export; ignora Resumen/Filtros (evita leer años como IDs).
    CSV (p.ej. producido por la IA desde la foto del papel): columnas por
    nombre — ID obligatoria; PRECIO_CAJA / COINCIDE / OBSERVACION opcionales.
    """
    def _norm(v):
        return re.sub(r'[^A-Z_]', '', str(v or '').upper().replace(' ', '_'))

    def _mapear(headers):
        """{campo: índice} desde una fila de headers."""
        m = {}
        for i, h in enumerate(headers):
            h = _norm(h)
            if h == 'ID':
                m.setdefault('id', i)
            elif 'PRECIO_CAJA' in h or h == 'PRECIOCAJA':
                m.setdefault('precio_caja', i)
            elif 'COINCIDE' in h:
                m.setdefault('coincide', i)
            elif 'OBSERV' in h:
                m.setdefault('observacion', i)
        return m if 'id' in m else None

    def _consumir(filas_iter, mapa, ids, verifs):
        for row in filas_iter:
            if not row:
                continue
            raw = row[mapa['id']] if mapa['id'] < len(row) else None
            s = str(raw).strip() if raw is not None else ''
            if not s.replace('.0', '').isdigit():
                continue
            pid = int(float(s))
            ids.append(pid)
            v = {}
            for campo in ('precio_caja', 'coincide', 'observacion'):
                idx = mapa.get(campo)
                if idx is not None and idx < len(row) and row[idx] not in (None, ''):
                    v[campo] = str(row[idx]).strip()
            if v:
                verifs[pid] = v

    nombre = (archivo.name or '').lower()
    ids, verifs = [], {}
    if nombre.endswith('.csv'):
        import csv
        import io
        txt = archivo.read().decode('utf-8-sig', errors='replace')
        rows = list(csv.reader(io.StringIO(txt)))
        if not rows:
            return [], {}
        mapa = _mapear(rows[0])
        if mapa:
            _consumir(rows[1:], mapa, ids, verifs)
        else:
            # Sin header: primera columna = IDs (formato viejo).
            _consumir(rows, {'id': 0}, ids, verifs)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(archivo, read_only=True, data_only=True)
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            mapa = None
            for _ in range(2):  # header en fila 1 (formato viejo) o 2 (título+header)
                fila = next(it, None)
                if fila is None:
                    break
                mapa = _mapear(fila)
                if mapa:
                    break
            if mapa:
                _consumir(it, mapa, ids, verifs)
    return ids, verifs


@require_POST
@requiere_permiso('plan_liquidacion', 'puede_ver')
def importar_seleccion_liquidacion(request):
    """Importa un .xlsx (multi-tab del export) o .csv (p.ej. extraído por IA
    del formulario impreso: ID, PRECIO_CAJA, COINCIDE, OBSERVACION). Devuelve
    los producto_ids válidos para seleccionar y armar campaña y, si vienen
    columnas de verificación, un resumen de discrepancias (no persiste)."""
    try:
        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'success': False, 'error': 'Sube un archivo .xlsx o .csv'}, status=400)
        ids, verifs = _leer_archivo_verificacion(archivo)
        if not ids:
            return JsonResponse({'success': False,
                                 'error': 'No se encontraron IDs (usá el Excel exportado o un CSV con columna ID).'}, status=400)
        sucursales = [s.id for s in _sucursales_usuario(request)]
        productos = {p['id']: p for p in Producto.objects.filter(
            id__in=ids, sucursal_id__in=sucursales,
        ).values('id', 'articulo', 'precioventa')}
        validos = list(productos.keys())

        resultado = {'success': True, 'producto_ids': validos,
                     'importados': len(validos),
                     'no_encontrados': len(set(ids)) - len(validos)}

        if verifs:
            coincide_si = coincide_no = precio_distinto = 0
            detalles = []
            for pid, v in verifs.items():
                p = productos.get(pid)
                if not p:
                    continue
                resp = re.sub(r'[^A-Z]', '', str(v.get('coincide', '')).upper())
                if resp.startswith('S'):
                    coincide_si += 1
                elif resp.startswith('N'):
                    coincide_no += 1
                pc_raw = re.sub(r'[^\d]', '', str(v.get('precio_caja', '')))
                if pc_raw:
                    pc = int(pc_raw)
                    if pc != (p['precioventa'] or 0):
                        precio_distinto += 1
                        if len(detalles) < 20:
                            detalles.append({'id': pid, 'articulo': p['articulo'],
                                             'precio_lista': p['precioventa'],
                                             'precio_caja': pc,
                                             'observacion': v.get('observacion', '')})
            resultado['verificacion'] = {
                'con_datos': len(verifs), 'coincide_si': coincide_si,
                'coincide_no': coincide_no, 'precio_distinto': precio_distinto,
                'detalles': detalles,
            }
        return JsonResponse(resultado)
    except Exception as e:
        logger.exception('Error importando selección de liquidación')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _etiquetas_filtros(ctx):
    """Pares (etiqueta, valor legible) de los filtros aplicados, para la
    hoja Filtros del export."""
    pares = []
    if ctx['sucursal_id']:
        alias = next((s.alias for s in ctx['sucursales']
                      if s.id == int(ctx['sucursal_id'])), ctx['sucursal_id'])
        pares.append(['Sucursal', alias])
    if ctx['marca_id']:
        op = AtributoOpcion.objects.filter(id=ctx['marca_id']).first()
        pares.append(['Marca', op.valor if op else ctx['marca_id']])
    if ctx['categoria_id']:
        cat = Categoria.objects.select_related('padre').filter(id=ctx['categoria_id']).first()
        if cat:
            label = f'{cat.padre.nombre} › {cat.nombre}' if cat.padre else cat.nombre
        else:
            label = ctx['categoria_id']
        pares.append(['Categoría', label])
    if ctx['especialidad_id']:
        op = AtributoOpcion.objects.filter(id=ctx['especialidad_id']).first()
        pares.append(['Especialidad', op.valor if op else ctx['especialidad_id']])
    pares.append(['Incluye bodegas/CD', 'Sí' if ctx['incluir_cd'] else 'No'])
    return pares
