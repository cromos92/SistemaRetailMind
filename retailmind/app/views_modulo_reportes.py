"""
Módulo de Reportes - RetailMind
Contiene todas las vistas relacionadas con reportes, dashboards estratégicos y análisis de datos
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, Http404, HttpResponseBadRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count, Q, Avg, Max, OuterRef, Subquery, IntegerField
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
import json
import re
import logging
from collections import defaultdict
from decimal import Decimal
from datetime import datetime, timedelta

from .models import (
    Compras, Compras_Producto, Compras_Producto_Talla, Productos_Recepcionados,
    Dte, Dte_Productos, Dte_Detalle_Pago, Ticket, Ticket_Productos,
    Producto, Producto_Talla, Movimientos_Producto, Sucursal, EmpresaUser,
    Empresa, Vendedor, LoteProducto, Traspaso, AjusteInventario,
    TicketDetallePago, METODO_PAGO_TICKET_CHOICES, TIPO_DOCUMENTO_CHOICES,
    Categoria, AtributoOpcion, Productos_Atributos,
    PermisoRol, PedidoEcommerce, CANAL_ECOMMERCE_CHOICES,
    CambioDevolucionDetalle,
)
from .utils_permisos import (
    obtener_sucursales_usuario,
    ids_empresas_alcance,
    ids_sucursales_alcance,
    puede_ver_sucursal,
    filtrar_queryset_por_sucursal,
    usuario_puede_ver_todas_sucursales,
    obtener_contexto_sucursales,
)

logger = logging.getLogger('app')


# ═══════════════════════════════════════════════════════════════════════════
# ALCANCE POR EMPRESA (scoping) — helpers compartidos del módulo de reportes
# ═══════════════════════════════════════════════════════════════════════════
# El middleware de permisos es FAIL-OPEN (una URL que no está en
# `URL_PERMISO_MAP` pasa) y además sólo mira `puede_ver`, así que la frontera
# entre las empresas del holding tiene que estar DENTRO de la vista. Sin esto,
# cualquier usuario autenticado leía inversión, ventas, márgenes y proveedores
# de otra empresa cambiando `?sucursal_id=` a mano.
#
# Convención (la misma de `_rp_filtros`, más abajo): una lista de ids en
# `None` significa "sin restricción" — rol administrador o flag
# `puede_ver_todas_sucursales`. Materializar los ~1.700 ids de un admin dentro
# del SQL no acotaría nada.

def _sin_acceso_reporte(mensaje='No tienes acceso a esa sucursal: pertenece a otra empresa.'):
    """403 uniforme, mismo contrato JSON que el resto del módulo de reportes."""
    return JsonResponse({'success': False, 'error': mensaje}, status=403)


def _alcance_sucursales(request, sucursal_id=None):
    """
    Universo de sucursales del usuario + validación del filtro recibido por GET.

    Devuelve ``(sucursales_ids, error)``:
      · ``sucursales_ids`` — lista de ids permitidos, o ``None`` si ve todo.
      · ``error`` — ``JsonResponse`` 403 listo para devolver cuando la sucursal
        pedida por querystring no pertenece a ninguna empresa del usuario.
        Nunca se confía en el id que llega del cliente.
    """
    permitidas = ids_sucursales_alcance(request.user)

    if permitidas is None or sucursal_id in (None, ''):
        return permitidas, None

    try:
        pedida = int(sucursal_id)
    except (TypeError, ValueError):
        return permitidas, _sin_acceso_reporte('Sucursal no válida.')

    if pedida not in permitidas:
        logger.warning(
            'scoping reportes: usuario %s pidió la sucursal %s fuera de su alcance',
            getattr(request.user, 'username', request.user), sucursal_id,
        )
        return permitidas, _sin_acceso_reporte()

    return permitidas, None


def _filtro_alcance(campo, sucursal_pedida, permitidas):
    """
    kwargs de filtro para un campo de sucursal ya validado por `_alcance_sucursales`.

    · ``{campo: valor}``     si el usuario pidió una sucursal concreta.
    · ``{campo__in: [...]}`` si no pidió ninguna y su alcance está acotado.
    · ``{}``                 si ve todo el holding.
    """
    if sucursal_pedida:
        return {campo: sucursal_pedida}
    if permitidas is not None:
        return {f'{campo}__in': permitidas}
    return {}


def _q_alcance_dte(usuario):
    """
    Q de alcance para `Dte`, con el mismo criterio de 3 puntas que usa
    `views_modulo_reportes_diferencias._queryset_recepciones`.

    `Dte.sucursal` sería el anclaje natural, pero en producción los DTE de
    COMPRA vienen con `sucursal=NULL` (ver auditoría de sucursales), así que
    filtrar sólo por sucursal borraría del reporte toda la compra histórica de
    un usuario acotado. Un documento entra si lo emitió o lo recibió una
    empresa del usuario, o si quedó anclado a una sucursal suya.
    """
    empresas_ids = ids_empresas_alcance(usuario)
    if empresas_ids is None:
        return Q()
    return (
        Q(emisor_id__in=empresas_ids)
        | Q(receptor_id__in=empresas_ids)
        | Q(sucursal__empresa_id__in=empresas_ids)
    )


def _q_alcance_recepcion(usuario):
    """
    Q de alcance para `Productos_Recepcionados` (mismas 3 puntas que el DTE,
    más la sucursal donde aterrizó la mercadería).
    """
    empresas_ids = ids_empresas_alcance(usuario)
    if empresas_ids is None:
        return Q()
    return (
        Q(dte__emisor_id__in=empresas_ids)
        | Q(dte__receptor_id__in=empresas_ids)
        | Q(sucursal_destino__empresa_id__in=empresas_ids)
    )


def _expandir_categoria_ids(categoria_id):
    """[id] + ids de sus hijas directas.

    Árbol v1.2 (2 niveles): filtrar por un PADRE debe incluir toda su rama;
    filtrar por una hija se comporta igual que antes (sin hijas propias).
    """
    ids = [categoria_id]
    ids += list(
        Categoria.objects.filter(padre_id=categoria_id).values_list('id', flat=True)
    )
    return ids


# ========== NOTAS DE CRÉDITO DEL LADO VENTA (helper único) ==========

# Una NC de venta NO nace con `tipo_transaccion='VENTA'`: los flujos que la
# emiten (`emisionNotaCredito`, `anular_factura_dte`, devolución garantía,
# `pos_service`) la graban como 'DEVOLUCION' (mueve plata) o 'ANULACION'
# (solo contable). 'VENTA'/'VENTA_PUBLICO' quedan en la lista únicamente
# porque `emitir_dte_concepto` sí crea NC/ND heredando el tipo de la venta
# (bug conocido, ver AUDITORIA_DOCUMENTOS_VENTAS_2026-07-30).
TIPOS_TRANSACCION_NC_VENTA = ('VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION')

# Estados que representan facturación válida (mismo criterio que
# `_queryset_dtes_ventas_vendedor` / `obtener_ventas_por_sucursal_reporte`).
ESTADOS_DTE_FACTURACION = ('EMITIDO', 'ACEPTADO', 'ANULADO')


def _queryset_ncs_venta(fecha_inicio, fecha_fin, user=None, request=None,
                        estados=ESTADOS_DTE_FACTURACION,
                        excluir_internas=False, base_queryset=None):
    """Notas de crédito REALES del lado venta en el rango [fecha_inicio, fecha_fin].

    POR QUÉ EXISTE
    --------------
    Varios reportes armaban el universo con
    ``tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO']`` y recién después
    hacían ``.filter(tipo_documento='NOTA DE CREDITO')``. Ese orden no
    devuelve casi nada: las NC de venta se emiten con
    ``tipo_transaccion`` en ('DEVOLUCION', 'ANULACION'), así que el primer
    filtro ya las había descartado. Consecuencia visible: la "Tasa de
    Devolución" del comparativo salía ~0% siempre y los netos quedaban
    inflados por el monto completo de las devoluciones.

    Parámetros
    ----------
    fecha_inicio / fecha_fin : date
        Rango sobre ``fecha_emision`` (inclusivo en ambos extremos).
    user / request :
        Si se pasan, aplica ``filtrar_queryset_por_sucursal`` (permisos).
    estados : tuple
        Estados de ``estado_dte`` admitidos. Por defecto los de
        facturación histórica; los reportes de caja pasan
        ``('EMITIDO', 'ACEPTADO')`` para dejar fuera anulados.
    excluir_internas : bool
        Excluye documentos donde receptor == emisor (facturación entre
        sucursales de la misma empresa).
    base_queryset : QuerySet | None
        Permite reusar un queryset ya filtrado (sucursal, tipo de
        documento, etc.) en vez de partir de ``Dte.objects``.

    Nota: siempre filtra ``descartado=False``. Las NC de modalidad OCULTA
    se marcan descartadas a propósito para quedar fuera de caja y de los
    listados; contarlas acá las haría restar dos veces (el DTE original
    ya queda en estado ANULADO).
    """
    qs = base_queryset if base_queryset is not None else Dte.objects.all()
    qs = qs.filter(
        fecha_emision__gte=fecha_inicio,
        fecha_emision__lte=fecha_fin,
        tipo_documento='NOTA DE CREDITO',
        tipo_transaccion__in=TIPOS_TRANSACCION_NC_VENTA,
        descartado=False,
    )
    if estados:
        qs = qs.filter(estado_dte__in=list(estados))
    if excluir_internas:
        qs = qs.exclude(receptor__isnull=False, receptor_id=F('emisor_id'))
    if user is not None:
        qs = filtrar_queryset_por_sucursal(qs, user, request)
    return qs


# ========== REPORTE DE VENTAS POR SUCURSAL ==========

@login_required
def ver_reporte_ventas_sucursal(request):
    """Vista principal del reporte de ventas mensual por vendedor y sucursal.

    Adicionalmente expone los flags `puede_ver_reporte_comisiones` y
    `puede_exportar_reporte_comisiones` para que el template muestre /
    oculte el botón "Comisiones" y los controles del modal según el
    permiso `reporte_comisiones_vendedor`.
    """
    context = obtener_contexto_sucursales(request.user, request)
    sucursal_id_sesion = _sucursal_id_sesion(request)
    context['puede_ver_reporte_comisiones'] = _puede_ver_reporte_comisiones(
        request.user, sucursal_id=sucursal_id_sesion,
    )
    context['puede_exportar_reporte_comisiones'] = _puede_exportar_reporte_comisiones(
        request.user, sucursal_id=sucursal_id_sesion,
    )
    return render(request, 'vistas/modulo_reportes/reporte_ventas_sucursal.html', context)


def _queryset_dtes_ventas_vendedor(request, fi, ff, vendedor_id=None):
    """Universo canónico de DTE del reporte de ventas por vendedor.

    Lo comparten el agregado (`obtener_ventas_por_vendedor_reporte`) y su
    drill-down (`obtener_documentos_vendedor_reporte`). Si cada uno arma su
    propio filtro, el detalle deja de explicar la fila desde la que se abrió
    (era el caso: el detalle pedía `tipo_transaccion='VENTA_PUBLICO'` y
    `estado_dte in (EMITIDO, PAGADO)`, y no restaba notas de crédito).

    Criterio (idéntico a `obtener_ventas_por_sucursal_reporte`): facturación
    histórica, no caja — entra ANULADO y las NC de ambas modalidades
    (DEVOLUCION / ANULACION) restan.
    """
    queryset = Dte.objects.filter(
        fecha_emision__gte=fi,
        fecha_emision__lte=ff,
        tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
        estado_dte__in=['EMITIDO', 'ACEPTADO', 'ANULADO'],
        descartado=False,
    ).select_related('vendedor', 'sucursal')

    queryset = filtrar_queryset_por_sucursal(queryset, request.user, request)

    if vendedor_id:
        queryset = queryset.filter(vendedor_id=vendedor_id)

    return queryset


@require_GET
@login_required
def obtener_ventas_por_vendedor_reporte(request):
    """API para obtener datos de ventas por vendedor.

    Mismos filtros que ``obtener_ventas_por_sucursal_reporte``: cuenta
    toda venta emitida (incluso ANULADO; el reporte es facturación
    histórica, no caja). Las NC restan independiente de modalidad.
    """
    try:
        vendedor_id = request.GET.get('vendedor_id')

        # Rango de fechas (mes / fecha puntual / rango; default mes actual TZ Chile).
        fi, ff = _parse_rango_fechas_reporte(request)

        # ========== DTEs (facturación histórica) ==========
        # Universo compartido con el drill-down de documentos por vendedor.
        queryset_dtes = _queryset_dtes_ventas_vendedor(request, fi, ff, vendedor_id)

        queryset_ventas = queryset_dtes.exclude(tipo_documento='NOTA DE CREDITO')
        # TODAS las NC restan (ambas modalidades anulan facturación).
        queryset_ncs = queryset_dtes.filter(tipo_documento='NOTA DE CREDITO')
        queryset_ncs_anulacion = queryset_ncs.filter(tipo_transaccion='ANULACION')

        # Ventas por vendedor
        ventas_por_vend = queryset_ventas.values(
            'vendedor__id',
            'vendedor__nombre',
            'vendedor__codigo_vendedor',
        ).annotate(
            total_ventas=Sum('monto_con_iva'),
            total_descuentos=Sum('descuento'),
            total_documentos=Count('id'),
        )

        # Devoluciones (NC DEVOLUCION) por vendedor — restan del total.
        # Incluimos también las NC sin vendedor asignado (vendedor__id=None)
        # para que el row "Sin vendedor" refleje el neto correcto.
        ncs_por_vend = {
            r['vendedor__id']: {
                'total': int(r['total'] or 0),
                'cantidad': int(r['cant'] or 0),
            }
            for r in queryset_ncs.values('vendedor__id').annotate(
                total=Sum('monto_con_iva'),
                cant=Count('id'),
            )
        }

        # NC ANULACION (informativas) por vendedor — solo conteo, no restan.
        ncs_anul_por_vend = {
            r['vendedor__id']: {
                'total': int(r['total'] or 0),
                'cantidad': int(r['cant'] or 0),
            }
            for r in queryset_ncs_anulacion.values('vendedor__id').annotate(
                total=Sum('monto_con_iva'),
                cant=Count('id'),
            )
        }

        # Unidades por vendedor (para UPT = unidades por documento).
        unidades_por_vend = {
            r['dte__vendedor__id']: int(r['u'] or 0)
            for r in Dte_Productos.objects.filter(dte__in=queryset_ventas)
            .values('dte__vendedor__id').annotate(u=Sum('stock'))
        }

        # Consolidar.
        # Usamos `None` como key para los DTEs sin vendedor asignado para
        # que el total de la tabla calce con el KPI "Total Ventas" del
        # reporte. Antes eso se descartaba con `if not vid: continue`,
        # creando un gap que confundía al usuario (veía $X arriba y la
        # tabla sumaba $X − sin_vendedor).
        ventas_acumuladas = {}
        for item in ventas_por_vend:
            vid = item['vendedor__id']  # puede ser None
            nc = ncs_por_vend.get(vid, {'total': 0, 'cantidad': 0})
            nc_anul = ncs_anul_por_vend.get(vid, {'total': 0, 'cantidad': 0})
            es_sin_vendedor = (not vid)
            ventas_acumuladas[vid] = {
                'nombre': item['vendedor__nombre'] if not es_sin_vendedor else 'Sin vendedor asignado',
                'codigo': item['vendedor__codigo_vendedor'] if not es_sin_vendedor else '',
                'ventas_brutas': int(item['total_ventas'] or 0),
                'descuentos': int(item['total_descuentos'] or 0),
                'documentos': int(item['total_documentos'] or 0),
                'devoluciones': nc['total'],
                'cantidad_devoluciones': nc['cantidad'],
                'nc_anulacion_total': nc_anul['total'],
                'cantidad_nc_anulacion': nc_anul['cantidad'],
                'unidades': unidades_por_vend.get(vid, 0),
                'sin_vendedor': es_sin_vendedor,
            }

        # Si hay NCs sin vendedor pero ningún DTE de venta sin vendedor,
        # igual mostramos la fila para que el total cierre correctamente.
        if None in ncs_por_vend and None not in ventas_acumuladas:
            nc = ncs_por_vend[None]
            nc_anul = ncs_anul_por_vend.get(None, {'total': 0, 'cantidad': 0})
            ventas_acumuladas[None] = {
                'nombre': 'Sin vendedor asignado',
                'codigo': '',
                'ventas_brutas': 0,
                'descuentos': 0,
                'documentos': 0,
                'devoluciones': nc['total'],
                'cantidad_devoluciones': nc['cantidad'],
                'nc_anulacion_total': nc_anul['total'],
                'cantidad_nc_anulacion': nc_anul['cantidad'],
                'sin_vendedor': True,
            }

        # Total general (ventas netas = brutas − devoluciones).
        # NCs sin vendedor se excluyen: existen en SII pero no pertenecen
        # a ningún vendedor y no deben restar del total del reporte.
        total_general = sum(
            v['ventas_brutas'] - v['devoluciones']
            for v in ventas_acumuladas.values()
            if not v.get('sin_vendedor')
        )

        # Armar salida
        vendedores_data = []
        for vid, data in sorted(
            ventas_acumuladas.items(),
            key=lambda x: x[1]['ventas_brutas'] - x[1]['devoluciones'],
            reverse=True,
        ):
            ventas_netas = data['ventas_brutas'] - data['devoluciones']
            participacion = (ventas_netas / total_general * 100) if total_general > 0 else 0
            vendedores_data.append({
                'id': vid,
                'nombre': data['nombre'],
                'codigo': data['codigo'],
                'ventas': ventas_netas,
                'ventas_brutas': data['ventas_brutas'],
                'descuentos': data['descuentos'],
                'devoluciones': data['devoluciones'],
                'cantidad_devoluciones': data['cantidad_devoluciones'],
                # NC informativas (ANULACION) — no afectan ventas, solo conteo.
                'nc_anulacion_total': int(data.get('nc_anulacion_total', 0)),
                'cantidad_nc_anulacion': int(data.get('cantidad_nc_anulacion', 0)),
                'documentos': data['documentos'],
                # Calidad de venta: % de descuento sobre precio lista y
                # unidades por documento (venta cruzada).
                'unidades': int(data.get('unidades', 0)),
                'upt': round(data.get('unidades', 0) / data['documentos'], 2)
                       if data['documentos'] else 0,
                'pct_descuento': round(
                    100 * data['descuentos'] / (data['ventas_brutas'] + data['descuentos']), 1
                ) if (data['ventas_brutas'] + data['descuentos']) > 0 else 0,
                'participacion': round(participacion, 1),
                # Marca para que el frontend pinte la fila distinto y NO
                # ofrezca el botón "ver documentos" (no hay vendedor_id).
                'sin_vendedor': bool(data.get('sin_vendedor', False)),
            })

        # KPIs.
        # `top_vendedor` ignora la fila virtual "Sin vendedor" para que
        # nunca aparezca como ranking #1.
        total_documentos = sum(v['documentos'] for v in ventas_acumuladas.values() if not v.get('sin_vendedor'))
        ticket_promedio = total_general / total_documentos if total_documentos > 0 else 0
        top_vendedor = next(
            (v['nombre'] for v in vendedores_data if not v.get('sin_vendedor')),
            '-',
        )
        total_devoluciones_general = sum(
            v['devoluciones'] for v in ventas_acumuladas.values() if not v.get('sin_vendedor')
        )

        return JsonResponse({
            'success': True,
            'vendedores': vendedores_data,
            'kpis': {
                'total_ventas': int(total_general),
                'total_ventas_brutas': int(total_general + total_devoluciones_general),
                'total_devoluciones': int(total_devoluciones_general),
                'total_documentos': total_documentos,
                'ticket_promedio': int(ticket_promedio),
                'top_vendedor': top_vendedor,
            }
        })
        
    except Exception as e:
        logger.exception("Error al obtener ventas por vendedor")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener ventas por vendedor: {str(e)}'
        })


# ========== HELPERS COMUNES PARA REPORTES DE COMISIONES ==========

# Código del `OpcionMenu` que protege el reporte de comisiones por
# vendedor. Configurable por rol/sucursal desde la pantalla de gestión
# de permisos. Lo crea la migración 0150_permiso_reporte_comisiones_vendedor.
CODIGO_PERMISO_COMISIONES = 'reporte_comisiones_vendedor'


def _puede_ver_reporte_comisiones(user, sucursal_id=None) -> bool:
    """Atajo para chequear el permiso del reporte de comisiones.

    El permiso se verifica con `puede_ver`, que es la convención del
    resto de reportes del módulo (`reporte_ventas_sucursal`,
    `reporte_existencias`, etc.). Si el usuario no está autenticado
    devuelve False sin tocar la BD.
    """
    if user is None or not getattr(user, 'is_authenticated', False):
        return False
    return PermisoRol.tiene_permiso(
        user, CODIGO_PERMISO_COMISIONES, 'puede_ver',
        sucursal_id=sucursal_id,
    )


def _puede_exportar_reporte_comisiones(user, sucursal_id=None) -> bool:
    """Chequea `puede_exportar` para el endpoint de Excel.

    Si el rol no tiene `puede_exportar=True` pero sí `puede_ver`, podrá
    consultar el reporte en pantalla pero no descargarlo. La migración
    inicial otorga `puede_ver=puede_exportar=True` al rol administrador.
    """
    if user is None or not getattr(user, 'is_authenticated', False):
        return False
    return PermisoRol.tiene_permiso(
        user, CODIGO_PERMISO_COMISIONES, 'puede_exportar',
        sucursal_id=sucursal_id,
    )


def _sucursal_id_sesion(request):
    """Sucursal activa en la sesión (None si no hay).

    Se usa para que `PermisoRol.tiene_permiso` evalúe también el override
    por sucursal (`PermisoSucursal`). Replica la convención del resto
    del proyecto (idSucursalActual / sucursalActual).
    """
    return (
        request.session.get('idSucursalActual')
        or request.session.get('sucursalActual')
    )


def _parse_rango_fechas_reporte(request):
    """Calcula `(fecha_inicio, fecha_fin)` (date) según los parámetros GET.

    Soporta los mismos modos que el resto del reporte de ventas-sucursal:
      * `fecha=YYYY-MM-DD` → un día puntual.
      * `fecha_inicio=...&fecha_fin=...` → rango.
      * `mes=YYYY-MM` → mes completo (default: mes actual en TZ Chile).

    Devuelve un tuple `(fecha_inicio, fecha_fin)` con `datetime.date`.
    """
    mes = request.GET.get('mes')
    fecha = request.GET.get('fecha')
    fecha_inicio_param = request.GET.get('fecha_inicio')
    fecha_fin_param = request.GET.get('fecha_fin')

    if fecha:
        fi = datetime.strptime(fecha, '%Y-%m-%d').date()
        ff = fi
    elif fecha_inicio_param and fecha_fin_param:
        fi = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').date()
        ff = datetime.strptime(fecha_fin_param, '%Y-%m-%d').date()
    else:
        if not mes:
            # Mes actual en zona horaria Chile (regla `timezone-chile`).
            mes = timezone.localdate().strftime('%Y-%m')
        primer_dia = datetime.strptime(mes, '%Y-%m').replace(day=1).date()
        if primer_dia.month == 12:
            ultimo_dia = primer_dia.replace(
                year=primer_dia.year + 1, month=1, day=1
            ) - timedelta(days=1)
        else:
            ultimo_dia = primer_dia.replace(
                month=primer_dia.month + 1, day=1
            ) - timedelta(days=1)
        fi = primer_dia
        ff = ultimo_dia
    return fi, ff


def _calcular_comisiones_vendedor(request):
    """Calcula la matriz de comisiones por vendedor para los filtros pedidos.

    La "venta neta" se define como la suma de `monto_neto` (sin IVA) de los
    DTEs de venta menos la suma de `monto_neto` de las Notas de Crédito
    imputadas al mismo vendedor, empresa y sucursal en el período.

    La comisión se calcula como `venta_neta_sin_iva * (Vendedor.comision / 100)`.

    El resultado se agrupa por **(vendedor, empresa emisora, sucursal)**: si
    un vendedor operó para más de una empresa o en más de una sucursal en el
    período aparece una fila por cada combinación, para que la liquidación
    contable pueda hacerse por razón social y por local.

    Las NC que salen en una sucursal distinta a la de la venta original se
    imputan a la sucursal donde efectivamente se emitió la NC (la cardinalidad
    natural del agregado SQL).

    DINERO REAL — dos correcciones respecto de la versión anterior
    -------------------------------------------------------------
    1. **Se excluye ``estado_dte='ANULADO'``.** Antes entraba "porque el
       reporte de facturación histórica lo incluye". Pero un DTE queda en
       ANULADO solo en dos casos, y en los dos la venta dejó de existir:
       anulación manual (`anular_documento`) o NC de modalidad OCULTA. En
       ese segundo caso la NC se graba con ``descartado=True``, así que
       nunca entraba al queryset y NO compensaba nada: la venta anulada
       pagaba comisión completa. Excluir ANULADO no genera doble resta
       justamente porque su NC está descartada.

       Efecto lateral consciente: comisiones deja de calzar exactamente con
       `obtener_ventas_por_vendedor_reporte`, que sí es facturación
       histórica y mantiene los ANULADO.

    2. **Las NC sin vendedor ya no se descartan.** `emisionNotaCredito` no
       copia el vendedor a la NC, así que ``vendedor_id`` viene NULL en
       prácticamente todas. El filtro anterior (``if r['vendedor__id'] and
       r['emisor__id']``) las botaba todas: ninguna devolución descontaba
       comisión de nadie. La imputación ahora es, en este orden:

         a) ``vendedor`` de la propia NC, si lo tiene;
         b) ``documento_afectado.vendedor`` — el vendedor de la venta
            original. Es un dato real, no un prorrateo estimado;
         c) si tampoco hay (NC huérfana), la NC va a una fila explícita
            **"Sin vendedor asignado"** dentro de su empresa/sucursal, que
            SÍ resta de los subtotales y del total global pero no descuenta
            comisión a ninguna persona en particular.

       Se eligió (b)+(c) sobre el prorrateo por sucursal porque prorratear
       inventa una atribución: le bajaría el sueldo a vendedores que no
       hicieron esa venta. El caso (c) queda visible en pantalla y en el
       Excel para que la liquidación lo resuelva a mano.

    Cuando una NC se imputa a un vendedor que no registró ventas en esa
    empresa/sucursal en el período (vendió el mes pasado, la devolución
    llegó este mes), se agrega igual su fila con ventas brutas en 0: la
    venta neta y la comisión quedan NEGATIVAS. Es un descuento real de
    liquidación, no un error de cálculo.

    Devuelve un dict con::

        {
          'empresas': [
            {
              'id': 1,
              'nombre': 'Empresa A',
              'rut': '76.123.456-7',
              'sucursales': [
                {
                  'id': 10,
                  'alias': 'Centro',
                  'direccion': 'Av. ...',
                  'vendedores': [<dict por vendedor en esta sucursal>],
                  'subtotales': {<mismas keys que totales>},
                },
                ...
              ],
              'subtotales': {<suma de las sucursales + cantidad_sucursales>},
            },
            ...
          ],
          'vendedores': [...],   # flat (1 por trío vendedor-empresa-sucursal)
          'fecha_inicio': 'YYYY-MM-DD',
          'fecha_fin': 'YYYY-MM-DD',
          'sucursal_id': str | None,
          'vendedor_id': str | None,
          'totales': {<global, incluye cantidad_sucursales>},
        }
    """
    fi, ff = _parse_rango_fechas_reporte(request)
    sucursal_id = request.GET.get('sucursal_id')
    vendedor_id = request.GET.get('vendedor_id')

    # `ANULADO` fuera: son ventas que dejaron de existir y su NC (cuando la
    # hay) está descartada, así que nunca las compensaba. Ver docstring.
    queryset_dtes = Dte.objects.filter(
        fecha_emision__gte=fi,
        fecha_emision__lte=ff,
        tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
        estado_dte__in=['EMITIDO', 'ACEPTADO'],
        descartado=False,
    ).select_related('vendedor', 'sucursal', 'emisor')

    # Respetar permisos de sucursal del usuario y filtro explícito si vino.
    queryset_dtes = filtrar_queryset_por_sucursal(queryset_dtes, request.user, request)
    if sucursal_id:
        queryset_dtes = queryset_dtes.filter(sucursal_id=sucursal_id)

    queryset_ventas = queryset_dtes.exclude(tipo_documento='NOTA DE CREDITO')
    queryset_ncs = queryset_dtes.filter(tipo_documento='NOTA DE CREDITO')

    # El filtro por vendedor se aplica DESPUÉS de separar: si se aplicara al
    # queryset común, las NC (que traen vendedor_id NULL) desaparecerían y el
    # vendedor filtrado aparecería sin ninguna devolución descontada.
    if vendedor_id:
        queryset_ventas = queryset_ventas.filter(vendedor_id=vendedor_id)
        queryset_ncs = queryset_ncs.filter(
            Q(vendedor_id=vendedor_id)
            | Q(vendedor__isnull=True, documento_afectado__vendedor_id=vendedor_id)
        )

    # Agrupación triple (vendedor, emisor=empresa, sucursal). Un mismo
    # vendedor que vendió para varias empresas o en varias sucursales
    # produce una fila por cada combinación.
    ventas_agg = queryset_ventas.values(
        'vendedor__id',
        'vendedor__nombre',
        'vendedor__codigo_vendedor',
        'vendedor__comision',
        'emisor__id',
        'emisor__nombre',
        'emisor__rut',
        'sucursal__id',
        'sucursal__alias',
        'sucursal__direccion',
    ).annotate(
        total_ventas=Sum('monto_con_iva'),
        total_neto=Sum('monto_neto'),
        total_documentos=Count('id'),
    )

    # NC con la misma clave triple. Si una NC se emite en otra sucursal
    # respecto de la venta original, se imputa donde se emitió.
    #
    # El vendedor sale de la NC o, si viene NULL (el caso normal: la NC no
    # hereda vendedor), del documento afectado. Las que no tienen ninguno de
    # los dos van a `ncs_sin_vendedor` y terminan en una fila explícita.
    ncs_agg: dict[tuple, dict] = {}
    ncs_sin_vendedor: dict[tuple, dict] = {}
    for r in queryset_ncs.values(
        'vendedor__id', 'documento_afectado__vendedor__id',
        'emisor__id', 'sucursal__id',
    ).annotate(
        total=Sum('monto_con_iva'),
        neto=Sum('monto_neto'),
        cant=Count('id'),
    ):
        eid = r['emisor__id']
        if not eid:
            # Sin empresa emisora no hay liquidación posible; se registra
            # para no perder la plata en silencio.
            logger.warning(
                "Comisiones: NC sin empresa emisora en %s..%s (monto=%s)",
                fi, ff, r['total'],
            )
            continue
        vid = r['vendedor__id'] or r['documento_afectado__vendedor__id']
        sid = r['sucursal__id']
        destino = ncs_agg if vid else ncs_sin_vendedor
        clave = (vid, eid, sid) if vid else (eid, sid)
        acum = destino.setdefault(clave, {'total': 0, 'neto': 0, 'cantidad': 0})
        acum['total'] += int(r['total'] or 0)
        acum['neto'] += int(r['neto'] or 0)
        acum['cantidad'] += int(r['cant'] or 0)

    # Claves de NC efectivamente consumidas por una fila de ventas; el resto
    # se materializa después como fila propia (comisión negativa).
    ncs_consumidas: set[tuple] = set()

    def _nuevo_subtotal() -> dict:
        return {
            'total_ventas_brutas_con_iva': 0,
            'total_ventas_brutas_neto': 0,
            'total_devoluciones_con_iva': 0,
            'total_devoluciones_neto': 0,
            'total_ventas_netas_sin_iva': 0,
            'total_ventas_netas_con_iva': 0,
            'total_comisiones': 0,
            'total_documentos': 0,
            'cantidad_vendedores': 0,
        }

    # Construcción anidada Empresa -> Sucursal -> Vendedor + lista flat.
    empresas_map: dict[int, dict] = {}
    vendedores_data: list[dict] = []
    glob = _nuevo_subtotal()
    sucursales_distintas: set[int] = set()

    def _registrar_fila(fila, cuenta_como_vendedor=True):
        """Agrega `fila` a la lista flat y propaga a los 3 niveles de subtotal.

        `cuenta_como_vendedor=False` para la fila sintética "Sin vendedor
        asignado": suma su plata a los subtotales (para eso existe) pero no
        infla el conteo de vendedores de la liquidación.
        """
        eid = fila['empresa_id']
        sid = fila['sucursal_id']
        vendedores_data.append(fila)

        emp = empresas_map.setdefault(eid, {
            'id': eid,
            'nombre': fila['empresa_nombre'],
            'rut': fila['empresa_rut'],
            'sucursales_map': {},  # interno, se elimina antes de devolver
            'subtotales': {**_nuevo_subtotal(), 'cantidad_sucursales': 0},
        })

        # Clave robusta para sucursales sin id (DTEs huérfanos): usa el
        # alias o un placeholder negativo derivado del id de empresa.
        suc_key = sid if sid is not None else f"_no_suc_{eid}"
        suc = emp['sucursales_map'].setdefault(suc_key, {
            'id': sid,
            'alias': fila['sucursal_alias'],
            'direccion': fila['sucursal_direccion'],
            'vendedores': [],
            'subtotales': _nuevo_subtotal(),
        })
        suc['vendedores'].append(fila)

        for sub in (suc['subtotales'], emp['subtotales'], glob):
            sub['total_ventas_brutas_con_iva'] += fila['ventas_brutas']
            sub['total_ventas_brutas_neto'] += fila['ventas_brutas_neto']
            sub['total_devoluciones_con_iva'] += fila['devoluciones']
            sub['total_devoluciones_neto'] += fila['devoluciones_neto']
            sub['total_ventas_netas_sin_iva'] += fila['ventas_netas_sin_iva']
            sub['total_ventas_netas_con_iva'] += fila['ventas_netas_con_iva']
            sub['total_comisiones'] += fila['comision_monto']
            sub['total_documentos'] += fila['documentos']
            if cuenta_como_vendedor:
                sub['cantidad_vendedores'] += 1

        if sid is not None:
            sucursales_distintas.add(sid)

    for item in ventas_agg:
        vid = item['vendedor__id']
        eid = item['emisor__id']
        sid = item['sucursal__id']
        if not vid or not eid:
            continue
        ventas_brutas_iva = int(item['total_ventas'] or 0)
        ventas_brutas_neto = int(item['total_neto'] or 0)
        nc = ncs_agg.get((vid, eid, sid), {'total': 0, 'neto': 0})
        ncs_consumidas.add((vid, eid, sid))
        ventas_netas_iva = ventas_brutas_iva - nc['total']
        ventas_netas_neto = ventas_brutas_neto - nc['neto']
        try:
            comision_pct = float(item['vendedor__comision'] or 0)
        except (TypeError, ValueError):
            comision_pct = 0.0
        comision_monto = int(round(ventas_netas_neto * comision_pct / 100.0))

        _registrar_fila({
            'id': vid,
            'nombre': item['vendedor__nombre'] or '(sin nombre)',
            'codigo': item['vendedor__codigo_vendedor'] or '',
            'empresa_id': eid,
            'empresa_nombre': item['emisor__nombre'] or '(sin empresa)',
            'empresa_rut': item['emisor__rut'] or '',
            'sucursal_id': sid,
            'sucursal_alias': item['sucursal__alias'] or '',
            'sucursal_direccion': item['sucursal__direccion'] or '',
            'ventas_brutas': ventas_brutas_iva,
            'ventas_brutas_neto': ventas_brutas_neto,
            'devoluciones': nc['total'],
            'devoluciones_neto': nc['neto'],
            'ventas_netas_con_iva': ventas_netas_iva,
            'ventas_netas_sin_iva': ventas_netas_neto,
            'documentos': int(item['total_documentos'] or 0),
            'comision_pct': comision_pct,
            'comision_monto': comision_monto,
            'sin_vendedor': False,
        })

    # --- NC que no encontraron una fila de ventas donde restarse ---
    # Caso típico: el vendedor vendió en un período anterior y la devolución
    # cae en éste. Antes desaparecían y la comisión salía sin descontar.
    ncs_huerfanas = {
        k: v for k, v in ncs_agg.items() if k not in ncs_consumidas
    }
    if ncs_huerfanas:
        vendedores_meta = {
            v.id: v for v in Vendedor.objects.filter(
                id__in={k[0] for k in ncs_huerfanas}
            ).only('id', 'nombre', 'codigo_vendedor', 'comision')
        }
        empresas_meta = {
            e.id: e for e in Empresa.objects.filter(
                id__in={k[1] for k in ncs_huerfanas}
            ).only('id', 'nombre', 'rut')
        }
        sucursales_meta = {
            s.id: s for s in Sucursal.objects.filter(
                id__in={k[2] for k in ncs_huerfanas if k[2] is not None}
            ).only('id', 'alias', 'direccion')
        }
        for (vid, eid, sid), nc in ncs_huerfanas.items():
            vend = vendedores_meta.get(vid)
            emp_obj = empresas_meta.get(eid)
            suc_obj = sucursales_meta.get(sid)
            try:
                comision_pct = float(getattr(vend, 'comision', 0) or 0)
            except (TypeError, ValueError):
                comision_pct = 0.0
            _registrar_fila({
                'id': vid,
                'nombre': getattr(vend, 'nombre', None) or '(sin nombre)',
                'codigo': getattr(vend, 'codigo_vendedor', None) or '',
                'empresa_id': eid,
                'empresa_nombre': getattr(emp_obj, 'nombre', None) or '(sin empresa)',
                'empresa_rut': getattr(emp_obj, 'rut', None) or '',
                'sucursal_id': sid,
                'sucursal_alias': getattr(suc_obj, 'alias', None) or '',
                'sucursal_direccion': getattr(suc_obj, 'direccion', None) or '',
                'ventas_brutas': 0,
                'ventas_brutas_neto': 0,
                'devoluciones': nc['total'],
                'devoluciones_neto': nc['neto'],
                'ventas_netas_con_iva': -nc['total'],
                'ventas_netas_sin_iva': -nc['neto'],
                'documentos': 0,
                'comision_pct': comision_pct,
                'comision_monto': int(round(-nc['neto'] * comision_pct / 100.0)),
                'sin_vendedor': False,
                'solo_devoluciones': True,
            })

    # --- NC realmente huérfanas (sin vendedor propio ni en el documento
    # afectado): fila explícita que SÍ resta de los subtotales, pero sin
    # descontarle comisión a ninguna persona concreta. ---
    if ncs_sin_vendedor:
        empresas_meta_sv = {
            e.id: e for e in Empresa.objects.filter(
                id__in={k[0] for k in ncs_sin_vendedor}
            ).only('id', 'nombre', 'rut')
        }
        sucursales_meta_sv = {
            s.id: s for s in Sucursal.objects.filter(
                id__in={k[1] for k in ncs_sin_vendedor if k[1] is not None}
            ).only('id', 'alias', 'direccion')
        }
        for (eid, sid), nc in ncs_sin_vendedor.items():
            emp_obj = empresas_meta_sv.get(eid)
            suc_obj = sucursales_meta_sv.get(sid)
            _registrar_fila({
                'id': None,
                'nombre': 'Sin vendedor asignado',
                'codigo': '',
                'empresa_id': eid,
                'empresa_nombre': getattr(emp_obj, 'nombre', None) or '(sin empresa)',
                'empresa_rut': getattr(emp_obj, 'rut', None) or '',
                'sucursal_id': sid,
                'sucursal_alias': getattr(suc_obj, 'alias', None) or '',
                'sucursal_direccion': getattr(suc_obj, 'direccion', None) or '',
                'ventas_brutas': 0,
                'ventas_brutas_neto': 0,
                'devoluciones': nc['total'],
                'devoluciones_neto': nc['neto'],
                'ventas_netas_con_iva': -nc['total'],
                'ventas_netas_sin_iva': -nc['neto'],
                'documentos': 0,
                'comision_pct': 0.0,
                # 0 a propósito: no se le descuenta a nadie en particular.
                'comision_monto': 0,
                'sin_vendedor': True,
                'solo_devoluciones': True,
            }, cuenta_como_vendedor=False)

    total_comisiones = glob['total_comisiones']
    total_ventas_brutas_con_iva = glob['total_ventas_brutas_con_iva']
    total_ventas_brutas_sin_iva = glob['total_ventas_brutas_neto']
    total_ventas_netas_sin_iva = glob['total_ventas_netas_sin_iva']
    total_ventas_netas_con_iva = glob['total_ventas_netas_con_iva']
    total_documentos = glob['total_documentos']
    total_devoluciones = glob['total_devoluciones_con_iva']
    total_devoluciones_neto = glob['total_devoluciones_neto']

    # Ordenamientos: empresas alfabéticamente; dentro de cada empresa las
    # sucursales por alias alfabético; dentro de cada sucursal los
    # vendedores por ventas netas descendentes.
    for emp in empresas_map.values():
        sucursales_lista = list(emp['sucursales_map'].values())
        for suc in sucursales_lista:
            suc['vendedores'].sort(
                key=lambda v: v['ventas_netas_sin_iva'], reverse=True,
            )
        sucursales_lista.sort(
            key=lambda s: ((s['alias'] or '').lower(), (s['direccion'] or '').lower()),
        )
        emp['sucursales'] = sucursales_lista
        emp['subtotales']['cantidad_sucursales'] = len(sucursales_lista)
        del emp['sucursales_map']

    empresas_data = sorted(
        empresas_map.values(), key=lambda e: (e['nombre'] or '').lower(),
    )
    vendedores_data.sort(
        key=lambda v: (
            v['empresa_nombre'].lower(),
            (v['sucursal_alias'] or '').lower(),
            -v['ventas_netas_sin_iva'],
        ),
    )

    return {
        'empresas': empresas_data,
        'vendedores': vendedores_data,
        'fecha_inicio': fi.strftime('%Y-%m-%d'),
        'fecha_fin': ff.strftime('%Y-%m-%d'),
        'sucursal_id': sucursal_id or None,
        'vendedor_id': vendedor_id or None,
        'totales': {
            # "Total original" — la cifra que coincide con el KPI principal
            # del reporte (ventas brutas con IVA, antes de devoluciones).
            'total_ventas_brutas_con_iva': total_ventas_brutas_con_iva,
            'total_ventas_brutas_sin_iva': total_ventas_brutas_sin_iva,
            'total_ventas_netas_sin_iva': total_ventas_netas_sin_iva,
            'total_ventas_netas_con_iva': total_ventas_netas_con_iva,
            'total_comisiones': total_comisiones,
            'total_documentos': total_documentos,
            'total_devoluciones': total_devoluciones,
            'total_devoluciones_neto': total_devoluciones_neto,
            # Excluye las filas sintéticas "Sin vendedor asignado".
            'cantidad_vendedores': glob['cantidad_vendedores'],
            'cantidad_empresas': len(empresas_data),
            'cantidad_sucursales': len(sucursales_distintas),
            # Devoluciones que no se pudieron imputar a ninguna persona:
            # restan del neto pero NO descuentan comisión de nadie. Si esto
            # es > 0, la liquidación necesita revisión manual.
            'devoluciones_sin_vendedor': sum(
                v['total'] for v in ncs_sin_vendedor.values()
            ),
            'cantidad_ncs_sin_vendedor': sum(
                v['cantidad'] for v in ncs_sin_vendedor.values()
            ),
        },
    }


@require_GET
@login_required
def obtener_comisiones_por_vendedor(request):
    """API JSON: comisiones por vendedor en el período/sucursal indicados.

    Reusa los filtros del reporte de ventas-sucursal (mes/fecha/rango y
    sucursal/vendedor). La comisión se calcula como
    `venta_neta_sin_iva × Vendedor.comision / 100`.

    Requiere el permiso `reporte_comisiones_vendedor` (acción `puede_ver`)
    configurable por rol/sucursal desde la pantalla de gestión de permisos.
    """
    if not _puede_ver_reporte_comisiones(
        request.user, sucursal_id=_sucursal_id_sesion(request)
    ):
        return JsonResponse(
            {
                'success': False,
                'error': 'No tienes permiso para ver el reporte de comisiones por vendedor',
            },
            status=403,
        )
    try:
        data = _calcular_comisiones_vendedor(request)
        return JsonResponse({'success': True, **data})
    except ValueError as e:
        return JsonResponse(
            {'success': False, 'error': f'Parámetros inválidos: {e}'},
            status=400,
        )
    except Exception as e:
        logger.exception("Error al calcular comisiones de vendedor")
        return JsonResponse(
            {'success': False, 'error': f'Error al calcular comisiones: {e}'},
            status=500,
        )


@require_GET
@login_required
def exportar_comisiones_vendedor_excel(request):
    """Exporta el reporte de comisiones por vendedor a Excel con formato.

    Encabezado con color institucional, totales en negrita, formato moneda
    y porcentaje, columnas auto-dimensionadas y bordes en toda la tabla.
    Reusa `_calcular_comisiones_vendedor` para garantizar que el Excel
    refleje exactamente los mismos números que el modal en pantalla.

    Requiere el permiso `reporte_comisiones_vendedor` con la acción
    `puede_exportar` (puede ser distinta de `puede_ver`: un rol puede
    consultar el reporte en pantalla pero no descargarlo).
    """
    sucursal_id_sesion = _sucursal_id_sesion(request)
    if not _puede_ver_reporte_comisiones(request.user, sucursal_id_sesion):
        return JsonResponse(
            {
                'success': False,
                'error': 'No tienes permiso para ver el reporte de comisiones',
            },
            status=403,
        )
    if not _puede_exportar_reporte_comisiones(request.user, sucursal_id_sesion):
        return JsonResponse(
            {
                'success': False,
                'error': 'No tienes permiso para exportar el reporte de comisiones',
            },
            status=403,
        )
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        data = _calcular_comisiones_vendedor(request)

        # Resolver alias de sucursal y vendedor para mostrar en el header.
        sucursal_label = 'Todas'
        if data['sucursal_id']:
            try:
                suc = Sucursal.objects.get(id=data['sucursal_id'])
                sucursal_label = suc.alias or suc.direccion or f"#{suc.id}"
            except Sucursal.DoesNotExist:
                sucursal_label = f"#{data['sucursal_id']}"

        vendedor_label = 'Todos'
        if data['vendedor_id']:
            try:
                v = Vendedor.objects.get(id=data['vendedor_id'])
                vendedor_label = f"{v.codigo_vendedor} - {v.nombre}"
            except Vendedor.DoesNotExist:
                vendedor_label = f"#{data['vendedor_id']}"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Comisiones por Vendedor'

        # Estilos institucionales NEXO (azul corporativo).
        header_fill = PatternFill(start_color='0066FF', end_color='0066FF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        sub_fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')
        total_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        total_font = Font(bold=True, color='FFFFFF', size=11)
        empresa_fill = PatternFill(start_color='0052CC', end_color='0052CC', fill_type='solid')
        empresa_font = Font(bold=True, color='FFFFFF', size=12)
        # Banner de sucursal: tono azul intermedio (más suave que empresa,
        # más fuerte que subtotal). Permite separar visualmente los 3
        # niveles de jerarquía: TOTAL > EMPRESA > SUCURSAL > vendedor.
        sucursal_fill = PatternFill(start_color='4D7BC9', end_color='4D7BC9', fill_type='solid')
        sucursal_font = Font(bold=True, color='FFFFFF', size=11)
        # Subtotal sucursal: tono más claro que el subtotal de empresa para
        # que la jerarquía siga reflejándose en las filas de totales.
        subtotal_suc_fill = PatternFill(start_color='EAF1FB', end_color='EAF1FB', fill_type='solid')
        subtotal_suc_font = Font(bold=True, color='1A1A2E', size=10)
        # Subtotal empresa (más oscuro que el de sucursal).
        subtotal_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        subtotal_font = Font(bold=True, color='1A1A2E', size=10)
        thin = Side(style='thin', color='B0B0B0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        left_indent = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

        # 8 columnas: #, Vendedor, Código, Ventas Brutas, Devoluciones,
        # Ventas Netas, % Comisión, Comisión $. La sucursal y dirección
        # se muestran como banner agrupador, no como columnas.
        N_COLS = 8
        last_col_letter = get_column_letter(N_COLS)

        # ===== Banner / título =====
        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = 'REPORTE DE COMISIONES POR VENDEDOR'
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
        ws['A1'].fill = total_fill
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 26

        # ===== Sub-banner: contexto de filtros =====
        ws.merge_cells(f'A2:{last_col_letter}2')
        rango_str = (
            f"Período: {data['fecha_inicio']} → {data['fecha_fin']}  ·  "
            f"Sucursal: {sucursal_label}  ·  Vendedor: {vendedor_label}"
        )
        ws['A2'] = rango_str
        ws['A2'].font = Font(italic=True, size=10, color='4A4A5A')
        ws['A2'].alignment = center
        ws['A2'].fill = sub_fill
        ws.row_dimensions[2].height = 20

        ws.merge_cells(f'A3:{last_col_letter}3')
        ws['A3'] = (
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}  ·  "
            f"Empresas: {data['totales'].get('cantidad_empresas', 0)}  ·  "
            f"Sucursales: {data['totales'].get('cantidad_sucursales', 0)}  ·  "
            f"Vendedores: {data['totales']['cantidad_vendedores']}"
        )
        ws['A3'].font = Font(italic=True, size=9, color='8A8A9A')
        ws['A3'].alignment = center

        # ===== Headers de tabla =====
        # "Ventas Brutas (c/IVA)" para que sea distinto a "Ventas Netas (s/IVA)"
        # cuando no hay devoluciones; concilia con el KPI "Total Ventas c/IVA".
        headers = [
            '#', 'Vendedor', 'Código',
            'Ventas Brutas (c/IVA)', 'Devoluciones (s/IVA)',
            'Ventas Netas (s/IVA)', '% Comisión', 'Comisión $',
        ]
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[row].height = 32

        # ===== Filas de datos: Empresa -> Sucursal -> Vendedor =====
        money_fmt = '"$"#,##0'
        pct_fmt = '0.00"%"'
        row = 6
        idx_global = 0
        empresas = data.get('empresas') or []

        def _label_sucursal_compacto(suc: dict) -> str:
            """Texto del banner de sucursal: alias — dirección (compacto)."""
            alias = (suc.get('alias') or '').strip()
            direc = (suc.get('direccion') or '').strip()
            if alias and direc and alias.lower() != direc.lower():
                return f"{alias} — {direc}"
            return alias or direc or '(sin sucursal)'

        for emp in empresas:
            sucursales_emp = emp.get('sucursales') or []
            # Backcompat: si por algún motivo el backend mandara solo
            # `vendedores` (formato anterior), agruparlos en una
            # "sucursal sintética" sin alias para no romper el Excel.
            if not sucursales_emp and emp.get('vendedores'):
                sucursales_emp = [{
                    'id': None,
                    'alias': '',
                    'direccion': '',
                    'vendedores': emp.get('vendedores') or [],
                    'subtotales': emp.get('subtotales') or {},
                }]
            total_vend_emp = sum(
                len(s.get('vendedores') or []) for s in sucursales_emp
            )
            if not total_vend_emp:
                continue

            sub_emp = emp.get('subtotales') or {}
            cant_suc = sub_emp.get('cantidad_sucursales') or len(sucursales_emp)

            # Banner empresa.
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
            cabecera_emp = ws.cell(
                row=row, column=1,
                value=(
                    f"EMPRESA: {emp.get('nombre') or '(sin nombre)'}"
                    + (f"  ·  RUT {emp.get('rut')}" if emp.get('rut') else '')
                    + f"  ·  {cant_suc} sucursal"
                    + ('' if cant_suc == 1 else 'es')
                    + f"  ·  {total_vend_emp} vendedor"
                    + ('' if total_vend_emp == 1 else 'es')
                ),
            )
            cabecera_emp.fill = empresa_fill
            cabecera_emp.font = empresa_font
            cabecera_emp.alignment = left
            ws.row_dimensions[row].height = 22
            for col in range(1, N_COLS + 1):
                ws.cell(row=row, column=col).border = border
            row += 1

            for suc in sucursales_emp:
                vendedores_suc = suc.get('vendedores') or []
                if not vendedores_suc:
                    continue
                ssub = suc.get('subtotales') or {}
                label_suc = _label_sucursal_compacto(suc)

                # Banner sucursal (nivel intermedio entre empresa y
                # subtotal). Compacto: alias y dirección en una sola línea.
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
                cabecera_suc = ws.cell(
                    row=row, column=1,
                    value=(
                        f"↳ SUCURSAL: {label_suc}"
                        f"  ·  {len(vendedores_suc)} vendedor"
                        + ('' if len(vendedores_suc) == 1 else 'es')
                    ),
                )
                cabecera_suc.fill = sucursal_fill
                cabecera_suc.font = sucursal_font
                cabecera_suc.alignment = left_indent
                ws.row_dimensions[row].height = 20
                for col in range(1, N_COLS + 1):
                    ws.cell(row=row, column=col).border = border
                row += 1

                for v in vendedores_suc:
                    idx_global += 1
                    ws.cell(row=row, column=1, value=idx_global).alignment = center
                    ws.cell(row=row, column=2, value=v.get('nombre', '')).alignment = left
                    ws.cell(row=row, column=3, value=v.get('codigo', '')).alignment = center

                    c4 = ws.cell(row=row, column=4, value=v.get('ventas_brutas', 0))
                    c4.number_format = money_fmt
                    c4.alignment = right

                    c5 = ws.cell(row=row, column=5, value=v.get('devoluciones_neto', 0))
                    c5.number_format = money_fmt
                    c5.alignment = right

                    c6 = ws.cell(row=row, column=6, value=v.get('ventas_netas_sin_iva', 0))
                    c6.number_format = money_fmt
                    c6.alignment = right
                    c6.font = Font(bold=True, color='1A1A2E')

                    c7 = ws.cell(row=row, column=7, value=v.get('comision_pct', 0))
                    c7.number_format = pct_fmt
                    c7.alignment = center

                    c8 = ws.cell(row=row, column=8, value=v.get('comision_monto', 0))
                    c8.number_format = money_fmt
                    c8.alignment = right
                    c8.font = Font(bold=True, color='00B38A')

                    for col in range(1, N_COLS + 1):
                        ws.cell(row=row, column=col).border = border
                    row += 1

                # Subtotal por sucursal.
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                cell_lbl = ws.cell(
                    row=row, column=1,
                    value=f"↳ Subtotal {label_suc}",
                )
                cell_lbl.fill = subtotal_suc_fill
                cell_lbl.font = subtotal_suc_font
                cell_lbl.alignment = right

                ss4 = ws.cell(row=row, column=4, value=ssub.get('total_ventas_brutas_con_iva', 0))
                ss4.number_format = money_fmt
                ss4.fill = subtotal_suc_fill
                ss4.font = subtotal_suc_font
                ss4.alignment = right

                ss5 = ws.cell(row=row, column=5, value=ssub.get('total_devoluciones_neto', 0))
                ss5.number_format = money_fmt
                ss5.fill = subtotal_suc_fill
                ss5.font = subtotal_suc_font
                ss5.alignment = right

                ss6 = ws.cell(row=row, column=6, value=ssub.get('total_ventas_netas_sin_iva', 0))
                ss6.number_format = money_fmt
                ss6.fill = subtotal_suc_fill
                ss6.font = subtotal_suc_font
                ss6.alignment = right

                ws.cell(row=row, column=7, value='').fill = subtotal_suc_fill

                ss8 = ws.cell(row=row, column=8, value=ssub.get('total_comisiones', 0))
                ss8.number_format = money_fmt
                ss8.fill = subtotal_suc_fill
                ss8.font = Font(bold=True, color='00B38A', size=10)
                ss8.alignment = right

                for col in range(1, N_COLS + 1):
                    ws.cell(row=row, column=col).border = border
                ws.row_dimensions[row].height = 18
                row += 1

            # Subtotal por empresa (suma de sus sucursales).
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell_lbl = ws.cell(
                row=row, column=1,
                value=f"Subtotal {emp.get('nombre') or ''}",
            )
            cell_lbl.fill = subtotal_fill
            cell_lbl.font = subtotal_font
            cell_lbl.alignment = right

            cs4 = ws.cell(row=row, column=4, value=sub_emp.get('total_ventas_brutas_con_iva', 0))
            cs4.number_format = money_fmt
            cs4.fill = subtotal_fill
            cs4.font = subtotal_font
            cs4.alignment = right

            cs5 = ws.cell(row=row, column=5, value=sub_emp.get('total_devoluciones_neto', 0))
            cs5.number_format = money_fmt
            cs5.fill = subtotal_fill
            cs5.font = subtotal_font
            cs5.alignment = right

            cs6 = ws.cell(row=row, column=6, value=sub_emp.get('total_ventas_netas_sin_iva', 0))
            cs6.number_format = money_fmt
            cs6.fill = subtotal_fill
            cs6.font = subtotal_font
            cs6.alignment = right

            ws.cell(row=row, column=7, value='').fill = subtotal_fill

            cs8 = ws.cell(row=row, column=8, value=sub_emp.get('total_comisiones', 0))
            cs8.number_format = money_fmt
            cs8.fill = subtotal_fill
            cs8.font = Font(bold=True, color='00B38A', size=10)
            cs8.alignment = right

            for col in range(1, N_COLS + 1):
                ws.cell(row=row, column=col).border = border
            ws.row_dimensions[row].height = 20
            row += 1

            # Fila vacía como separador entre empresas.
            row += 1

        # ===== Fila de TOTAL GENERAL =====
        if data['vendedores']:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell_label = ws.cell(row=row, column=1, value='TOTAL GENERAL')
            cell_label.fill = total_fill
            cell_label.font = total_font
            cell_label.alignment = right

            t_brutas = sum(v['ventas_brutas'] for v in data['vendedores'])
            t_dev = sum(v['devoluciones_neto'] for v in data['vendedores'])

            c4 = ws.cell(row=row, column=4, value=t_brutas)
            c4.number_format = money_fmt
            c4.fill = total_fill
            c4.font = total_font
            c4.alignment = right

            c5 = ws.cell(row=row, column=5, value=t_dev)
            c5.number_format = money_fmt
            c5.fill = total_fill
            c5.font = total_font
            c5.alignment = right

            c6 = ws.cell(
                row=row, column=6,
                value=data['totales']['total_ventas_netas_sin_iva'],
            )
            c6.number_format = money_fmt
            c6.fill = total_fill
            c6.font = total_font
            c6.alignment = right

            ws.cell(row=row, column=7, value='').fill = total_fill

            c8 = ws.cell(
                row=row, column=8,
                value=data['totales']['total_comisiones'],
            )
            c8.number_format = money_fmt
            c8.fill = total_fill
            c8.font = total_font
            c8.alignment = right

            for col in range(1, N_COLS + 1):
                ws.cell(row=row, column=col).border = border
            ws.row_dimensions[row].height = 22

        # ===== Anchos =====
        # #, Vendedor, Código, Brutas, Dev, Netas, %, Comisión.
        anchos = [5, 32, 14, 20, 20, 22, 12, 20]
        for col, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Congelar encabezado de tabla.
        ws.freeze_panes = 'A6'

        # Render binario.
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        nombre = (
            f"comisiones_{data['fecha_inicio']}_{data['fecha_fin']}.xlsx"
        )
        response = HttpResponse(
            bio.read(),
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{nombre}"'
        )
        return response
    except Exception as e:
        logger.exception("Error al exportar comisiones de vendedor")
        return JsonResponse(
            {'success': False, 'error': f'Error al exportar comisiones: {e}'},
            status=500,
        )


@require_GET
@login_required
def obtener_ventas_por_sucursal_reporte(request):
    """API para obtener datos de ventas por sucursal.

    Reporte de **facturación histórica** (no de caja):
      * Cuenta toda venta emitida, INCLUSO si después se anuló
        (``estado_dte=ANULADO`` también entra). Si vendiste $X el día Y,
        ese día Y vendiste $X — independiente de que más tarde se anule
        con NC.
      * Las NC restan del día en que se emitieron (sea ``DEVOLUCION`` o
        ``ANULACION``). Así, si emitiste la NC en otro mes, el reporte
        del mes original sigue mostrando la venta y el mes de la NC
        muestra la devolución.
      * No es lo mismo que cuadratura-caja: cuadratura excluye los
        ANULADOS (no hay dinero) y trata las NC ANULACION como
        informativas. Para conciliar las dos vistas usar el botón
        "Diagnóstico vs cuadratura" del tab sucursales.

    Filtros que SÍ aplica:
      * ``tipo_transaccion in {VENTA, VENTA_PUBLICO, DEVOLUCION, ANULACION}``
        (las dos primeras para ventas; las dos últimas para que las NC
        del día entren al queryset)
      * ``estado_dte in {EMITIDO, ACEPTADO, ANULADO}`` — excluye
        ``CANCELADO``/``RECHAZADO``/``PENDIENTE`` (rechazos del SII y
        documentos sin emitir).
      * ``descartado=False`` — NC ocultas y eliminaciones lógicas no
        cuentan (operación interna).
      * Incluye FACTURA EXENTA, autoventas, CDs y facturas sin receptor
        (todo lo que se haya emitido oficialmente).
    """
    try:
        # Parámetros de filtro
        mes = request.GET.get('mes')  # Formato: YYYY-MM
        fecha = request.GET.get('fecha')  # Formato: YYYY-MM-DD (fecha específica)
        fecha_inicio_param = request.GET.get('fecha_inicio')  # Formato: YYYY-MM-DD (rango)
        fecha_fin_param = request.GET.get('fecha_fin')  # Formato: YYYY-MM-DD (rango)
        sucursal_id = request.GET.get('sucursal_id')

        # Determinar rango de fechas según el tipo de filtro.
        # Default: mes actual en zona horaria Chile (regla `timezone-chile`).
        if fecha:
            fecha_inicio = datetime.strptime(fecha, '%Y-%m-%d')
            fecha_fin = fecha_inicio
        elif fecha_inicio_param and fecha_fin_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d')
        else:
            if not mes:
                mes = timezone.localdate().strftime('%Y-%m')
            fecha_inicio = datetime.strptime(mes, '%Y-%m').replace(day=1)
            if fecha_inicio.month == 12:
                fecha_fin = fecha_inicio.replace(year=fecha_inicio.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                fecha_fin = fecha_inicio.replace(month=fecha_inicio.month + 1, day=1) - timedelta(days=1)

        fi = fecha_inicio.date() if hasattr(fecha_inicio, 'date') else fecha_inicio
        ff = fecha_fin.date() if hasattr(fecha_fin, 'date') else fecha_fin

        # ========== DTEs (facturación histórica) ==========
        # Acepta ANULADO porque ese DTE fue una venta real el día que se
        # emitió; la NC asociada se contabiliza el día en que se emitió.
        queryset_dtes = Dte.objects.filter(
            fecha_emision__gte=fi,
            fecha_emision__lte=ff,
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'],
            estado_dte__in=['EMITIDO', 'ACEPTADO', 'ANULADO'],
            descartado=False,
        ).select_related('sucursal', 'vendedor')

        queryset_dtes = filtrar_queryset_por_sucursal(queryset_dtes, request.user, request)

        # Separar ventas (no-NC) de notas de crédito.
        queryset_ventas = queryset_dtes.exclude(tipo_documento='NOTA DE CREDITO')
        # Restar TODAS las NC (DEVOLUCION + ANULACION). En este reporte
        # ambas representan ventas que se deshicieron oficialmente, así
        # que ambas afectan el total del mes. La distinción DEVOLUCION/
        # ANULACION es relevante para cuadratura-caja, no para un reporte
        # de facturación histórica.
        queryset_ncs = queryset_dtes.filter(tipo_documento='NOTA DE CREDITO')
        queryset_ncs_devolucion = queryset_ncs.filter(tipo_transaccion='DEVOLUCION')
        queryset_ncs_anulacion = queryset_ncs.filter(tipo_transaccion='ANULACION')

        # Agregar ventas por sucursal
        ventas_por_suc = queryset_ventas.values(
            'sucursal__id',
            'sucursal__alias',
            'sucursal__direccion',
        ).annotate(
            total_ventas=Sum('monto_con_iva'),
            total_neto=Sum('monto_neto'),
            total_descuentos=Sum('descuento'),
            total_documentos=Count('id'),
        )

        # NCs por sucursal: AMBAS modalidades (DEVOLUCION y ANULACION)
        # restan del total. Es un reporte de facturación, así que cualquier
        # NC válida cancela una venta previa.
        ncs_por_suc = {
            r['sucursal__id']: {
                'total': int(r['total'] or 0),
                'neto': int(r['neto'] or 0),
                'cantidad': int(r['cant'] or 0),
            }
            for r in queryset_ncs.values('sucursal__id').annotate(
                total=Sum('monto_con_iva'),
                neto=Sum('monto_neto'),
                cant=Count('id'),
            )
        }

        # Desglose informativo (no afecta totales — solo se muestra como
        # conteo en cada fila para que el operador sepa cuántas
        # ANULACIONes hay vs DEVOLUCIONes).
        ncs_anul_por_suc = {
            r['sucursal__id']: {
                'total': int(r['total'] or 0),
                'cantidad': int(r['cant'] or 0),
            }
            for r in queryset_ncs_anulacion.values('sucursal__id').annotate(
                total=Sum('monto_con_iva'),
                cant=Count('id'),
            )
        }

        # Vendedores distintos por sucursal
        vendedores_por_sucursal = {}
        for r in queryset_dtes.values('sucursal_id', 'vendedor_id').distinct():
            sid = r['sucursal_id']
            vid = r['vendedor_id']
            if sid and vid:
                vendedores_por_sucursal.setdefault(sid, set()).add(vid)

        # Consolidar ventas + devoluciones por sucursal
        ventas_acumuladas = {}
        for item in ventas_por_suc:
            sid = item['sucursal__id']
            if not sid:
                continue
            total_brutas = int(item['total_ventas'] or 0)
            neto_brutas = int(item['total_neto'] or 0)
            iva_brutas = total_brutas - neto_brutas
            nc = ncs_por_suc.get(sid, {'total': 0, 'neto': 0, 'cantidad': 0})
            nc_anul = ncs_anul_por_suc.get(sid, {'total': 0, 'cantidad': 0})
            ventas_acumuladas[sid] = {
                'alias': item['sucursal__alias'],
                'direccion': item['sucursal__direccion'],
                'ventas_brutas': total_brutas,
                'neto_brutas': neto_brutas,
                'iva_brutas': iva_brutas,
                'descuentos': int(item['total_descuentos'] or 0),
                'documentos': int(item['total_documentos'] or 0),
                'devoluciones': nc['total'],
                'devoluciones_neto': nc['neto'],
                'devoluciones_iva': nc['total'] - nc['neto'],
                'cantidad_devoluciones': nc['cantidad'],
                'nc_anulacion_total': nc_anul['total'],
                'cantidad_nc_anulacion': nc_anul['cantidad'],
            }

        # Sucursales con NC pero sin ventas en el período
        for sid, nc in ncs_por_suc.items():
            if not sid or sid in ventas_acumuladas:
                continue
            try:
                from .models import Sucursal as _Sucursal
                suc = _Sucursal.objects.get(id=sid)
                nc_anul = ncs_anul_por_suc.get(sid, {'total': 0, 'cantidad': 0})
                ventas_acumuladas[sid] = {
                    'alias': suc.alias,
                    'direccion': suc.direccion,
                    'ventas_brutas': 0,
                    'neto_brutas': 0,
                    'iva_brutas': 0,
                    'descuentos': 0,
                    'documentos': 0,
                    'devoluciones': nc['total'],
                    'devoluciones_neto': nc['neto'],
                    'devoluciones_iva': nc['total'] - nc['neto'],
                    'cantidad_devoluciones': nc['cantidad'],
                    'nc_anulacion_total': nc_anul['total'],
                    'cantidad_nc_anulacion': nc_anul['cantidad'],
                }
            except Exception:
                pass

        # Sucursales con NC ANULACION pero sin ventas ni NC devolución
        for sid, nc_anul in ncs_anul_por_suc.items():
            if not sid or sid in ventas_acumuladas:
                continue
            try:
                from .models import Sucursal as _Sucursal
                suc = _Sucursal.objects.get(id=sid)
                ventas_acumuladas[sid] = {
                    'alias': suc.alias,
                    'direccion': suc.direccion,
                    'ventas_brutas': 0,
                    'neto_brutas': 0,
                    'iva_brutas': 0,
                    'descuentos': 0,
                    'documentos': 0,
                    'devoluciones': 0,
                    'devoluciones_neto': 0,
                    'devoluciones_iva': 0,
                    'cantidad_devoluciones': 0,
                    'nc_anulacion_total': nc_anul['total'],
                    'cantidad_nc_anulacion': nc_anul['cantidad'],
                }
            except Exception:
                pass

        # Ventas netas (descontando devoluciones) para totales y ordenamiento
        total_general = sum(
            d['ventas_brutas'] - d['devoluciones']
            for d in ventas_acumuladas.values()
        )

        sucursales_data = []
        for sid, data in sorted(
            ventas_acumuladas.items(),
            key=lambda x: x[1]['ventas_brutas'] - x[1]['devoluciones'],
            reverse=True,
        ):
            ventas_netas = data['ventas_brutas'] - data['devoluciones']
            neto_netas = data['neto_brutas'] - data['devoluciones_neto']
            iva_netas = data['iva_brutas'] - data['devoluciones_iva']
            total_docs = data['documentos']
            participacion = (ventas_netas / total_general * 100) if total_general > 0 else 0

            sucursales_data.append({
                'id': sid,
                'nombre': data['alias'],
                'direccion': data['direccion'],
                'neto': int(neto_netas),
                'iva': int(iva_netas),
                'descuentos': int(data['descuentos']),
                'ventas': int(ventas_netas),
                'ventas_brutas': int(data['ventas_brutas']),
                'devoluciones': int(data['devoluciones']),
                'cantidad_devoluciones': int(data['cantidad_devoluciones']),
                # NC informativas (ANULACION) — no restan, solo se cuentan
                # para que el operador vea el dato sin descuadrar el total.
                'nc_anulacion_total': int(data.get('nc_anulacion_total', 0)),
                'cantidad_nc_anulacion': int(data.get('cantidad_nc_anulacion', 0)),
                'documentos': total_docs,
                'vendedores': len(vendedores_por_sucursal.get(sid, set())),
                'participacion': round(participacion, 1),
            })

        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener ventas por sucursal: {str(e)}'
        })


# ========== DIAGNÓSTICO: cuadratura-caja vs reporte ventas-sucursal ==========

# Filtros que aplica el reporte de ventas-sucursal (para detectar exclusiones).
# Mantener sincronizado con `obtener_ventas_por_sucursal_reporte`.
_REPORTE_TIPO_TRANSACCION_PERMITIDA = {
    'VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION',
}
# El reporte acepta ANULADO (es facturación histórica). Cuadratura no lo
# acepta (es caja del día). Por eso el diagnóstico mostrará el DTE ANULADO
# como "solo en reporte" — diferencia esperada y deseable.
_REPORTE_ESTADO_DTE_PERMITIDO = {'EMITIDO', 'ACEPTADO', 'ANULADO'}


def _diagnosticar_dte_vs_reporte(dte):
    """Devuelve la lista de motivos por los que un DTE NO entra al reporte
    de ventas-sucursal según los filtros que apliquen.

    Lista vacía => el DTE entraría al reporte. La función refleja los
    filtros REALES vigentes en `obtener_ventas_por_sucursal_reporte`
    tras alinear con cuadratura (sin exclusiones por CENTRO_DISTRIBUCION,
    autoventa, FACTURA_EXENTA o factura-sin-receptor). Si en el futuro
    se reintroduce alguna de esas exclusiones, sumar el motivo aquí
    para que el diagnóstico siga siendo fiel.
    """
    motivos = []
    if dte.tipo_transaccion not in _REPORTE_TIPO_TRANSACCION_PERMITIDA:
        motivos.append(f'TIPO_TRANSACCION={dte.tipo_transaccion}')
    if dte.estado_dte not in _REPORTE_ESTADO_DTE_PERMITIDO:
        motivos.append(f'ESTADO={dte.estado_dte}')
    if getattr(dte, 'descartado', False):
        motivos.append('DESCARTADO')
    return motivos


def _diagnosticar_dte_vs_cuadratura(dte):
    """Devuelve True si el DTE pasa los filtros que aplica
    `_calcular_cuadratura_data` al iterar DTEs del día.

    Filtros replicados (lectura — no se invoca):
      - estado_dte in {EMITIDO, ACEPTADO}
      - tipo_transaccion in {VENTA, VENTA_PUBLICO, DEVOLUCION, ANULACION}
      - descartado=False
    """
    if dte.estado_dte not in {'EMITIDO', 'ACEPTADO'}:
        return False
    if dte.tipo_transaccion not in {
        'VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION'
    }:
        return False
    if getattr(dte, 'descartado', False):
        return False
    return True


@require_GET
@login_required
def api_diagnostico_cuadratura_vs_reporte(request):
    """Compara, para una sucursal y fecha puntual, los totales de
    cuadratura-caja contra los del reporte ventas-sucursal y lista los
    DTEs en disputa con su motivo de exclusión.

    Parámetros:
      - `fecha=YYYY-MM-DD` (requerido)
      - `sucursal_id=<int>` (requerido)

    Devuelve JSON con::

        {
          "success": True,
          "fecha": "...",
          "sucursal": {"id": ..., "alias": "..."},
          "cuadratura_total": <int>,
          "reporte_total": <int>,
          "diferencia": <cuadratura - reporte>,
          "solo_en_cuadratura": [
              {id, folio, tipo_documento, tipo_transaccion, estado, monto,
               descartado, receptor_id, motivos_exclusion_reporte: [...]}
          ],
          "solo_en_reporte": [...],
          "tickets_sin_dte": [
              {id, correlativo, total, folio_dte}
          ],
          "resumen_cuadratura": {  # extracto del dict de cuadratura
              "venta_total", "total_tickets", "total_boletas_electronicas",
              ...
          }
        }

    NO modifica datos. Es una herramienta de auditoría que llama al
    helper `_calcular_cuadratura_data` (lectura) y reconstruye los
    filtros del reporte para comparar dte por dte.
    """
    from .views_modulo_ventas import _calcular_cuadratura_data

    fecha_str = request.GET.get('fecha')
    sucursal_id = request.GET.get('sucursal_id')

    if not fecha_str:
        return JsonResponse({
            'success': False,
            'error': 'Parametro `fecha` requerido (YYYY-MM-DD).'
        }, status=400)
    if not sucursal_id:
        return JsonResponse({
            'success': False,
            'error': 'Parametro `sucursal_id` requerido.'
        }, status=400)

    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Formato de fecha invalido. Usar YYYY-MM-DD.'
        }, status=400)

    try:
        sucursal = Sucursal.objects.get(id=int(sucursal_id))
    except (Sucursal.DoesNotExist, ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'error': 'Sucursal no encontrada.'
        }, status=404)

    # Cuadratura: fuente de verdad. NO se toca.
    cuadratura = _calcular_cuadratura_data(sucursal, fecha_str)

    # Set de DTEs "puro" del día y sucursal (sin exclusiones), para listar
    # caso por caso lo que entra/no entra a cada reporte.
    dtes_puros = list(
        Dte.objects.filter(
            sucursal=sucursal,
            fecha_emision=fecha_obj,
        ).select_related('receptor', 'emisor', 'sucursal').order_by(
            'tipo_documento', 'numero_documento'
        )
    )

    # Recalcular los totales del reporte para esta fecha/sucursal usando
    # la misma logica del endpoint del reporte (tras los fixes), sin
    # llamar al endpoint para no depender de la request.
    reporte_total = 0
    solo_en_cuadratura = []
    solo_en_reporte = []
    en_ambos = 0

    for dte in dtes_puros:
        motivos_excl_reporte = _diagnosticar_dte_vs_reporte(dte)
        entra_reporte = (len(motivos_excl_reporte) == 0)
        entra_cuadratura = _diagnosticar_dte_vs_cuadratura(dte)

        # Calcular el aporte al "total ventas" del reporte:
        # el reporte hace `ventas_brutas - devoluciones`. Como es un
        # reporte de facturación histórica, restamos TODAS las NC
        # (DEVOLUCION y ANULACION). Cuadratura solo resta DEVOLUCION;
        # esa diferencia se puede observar en el listado de DTE en
        # disputa que devuelve este endpoint.
        es_nc = dte.tipo_documento == 'NOTA DE CREDITO'
        aporte_reporte = 0
        if entra_reporte:
            monto = int(dte.monto_con_iva or 0)
            if es_nc:
                aporte_reporte = -monto
            else:
                aporte_reporte = monto
        reporte_total += aporte_reporte

        info = {
            'id': dte.id,
            'folio': dte.numero_documento,
            'tipo_documento': dte.tipo_documento,
            'tipo_transaccion': dte.tipo_transaccion,
            'estado': dte.estado_dte,
            'monto': int(dte.monto_con_iva or 0),
            'descartado': bool(getattr(dte, 'descartado', False)),
            'receptor_id': dte.receptor_id,
            'receptor_nombre': dte.receptor.nombre if dte.receptor else None,
            'motivos_exclusion_reporte': motivos_excl_reporte,
            'entra_cuadratura': entra_cuadratura,
            'entra_reporte': entra_reporte,
        }

        if entra_cuadratura and not entra_reporte:
            solo_en_cuadratura.append(info)
        elif entra_reporte and not entra_cuadratura:
            solo_en_reporte.append(info)
        elif entra_cuadratura and entra_reporte:
            en_ambos += 1

    # Tickets pagados sin DTE (cuadratura los suma desde Ticket; el reporte
    # solo mira DTE, así que para alinear hay que conocerlos).
    tickets_pagados = Ticket.objects.filter(
        sucursal=sucursal,
        fecha=fecha_obj,
        estado='PAGADO',
    ).only('id', 'correlativo', 'total', 'folio_dte')

    folios_emitidos = {
        d.numero_documento for d in dtes_puros if d.numero_documento
    }
    tickets_sin_dte = []
    for t in tickets_pagados:
        if t.folio_dte and t.folio_dte in folios_emitidos:
            continue
        tickets_sin_dte.append({
            'id': t.id,
            'correlativo': t.correlativo,
            'total': int(t.total or 0),
            'folio_dte': t.folio_dte,
        })

    cuadratura_total = int(cuadratura.get('venta_total', 0) or 0)
    diferencia = cuadratura_total - reporte_total

    resumen_cuadratura = {
        k: int(cuadratura.get(k, 0) or 0)
        for k in (
            'venta_total',
            'total_tickets',
            'total_boletas_electronicas',
            'total_boletas_papel',
            'total_facturas',
            'total_facturas_exentas',
            'total_notas_credito',
            'cantidad_tickets',
            'cantidad_boletas_electronicas',
            'cantidad_boletas_papel',
            'cantidad_facturas',
            'cantidad_facturas_exentas',
            'cantidad_notas_credito',
        )
    }

    return JsonResponse({
        'success': True,
        'fecha': fecha_str,
        'sucursal': {'id': sucursal.id, 'alias': sucursal.alias},
        'cuadratura_total': cuadratura_total,
        'reporte_total': reporte_total,
        'diferencia': diferencia,
        'solo_en_cuadratura': solo_en_cuadratura,
        'solo_en_reporte': solo_en_reporte,
        'en_ambos': en_ambos,
        'tickets_sin_dte': tickets_sin_dte,
        'tickets_sin_dte_total': sum(t['total'] for t in tickets_sin_dte),
        'resumen_cuadratura': resumen_cuadratura,
    })


@require_GET
@login_required
def obtener_vendedores_reporte(request):
    """API para obtener lista de vendedores que tienen ventas en la sucursal"""
    try:
        sucursal_id = request.GET.get('sucursal_id')

        if sucursal_id:
            # Obtener vendedores que tienen ventas EN esta sucursal (Tickets o DTEs)
            from django.db.models import Q
            
            # IDs de vendedores con Tickets en esta sucursal
            vendedores_tickets = set(Ticket.objects.filter(
                sucursal_id=sucursal_id
            ).values_list('vendedor_id', flat=True).distinct())
            
            # IDs de vendedores con DTEs de venta al público en esta sucursal
            vendedores_dtes = set(Dte.objects.filter(
                sucursal_id=sucursal_id,
                tipo_transaccion='VENTA_PUBLICO',
            ).exclude(
                receptor__isnull=False,
                receptor_id=F('emisor_id')
            ).values_list('vendedor_id', flat=True).distinct())
            
            # Unir ambos conjuntos
            vendedores_ids = vendedores_tickets | vendedores_dtes
            vendedores_ids.discard(None)  # Remover None si existe
            
            vendedores = Vendedor.objects.filter(
                id__in=vendedores_ids,
                nombre__isnull=False
            ).order_by('nombre')
        else:
            # Sin filtro de sucursal, mostrar todos
            vendedores = Vendedor.objects.filter(
                nombre__isnull=False
            ).order_by('nombre')

        vendedores_data = []
        for vendedor in vendedores:
            vendedores_data.append({
                'id': vendedor.id,
                'nombre': vendedor.nombre,
                'codigo': vendedor.codigo_vendedor
            })

        return JsonResponse({
            'success': True,
            'vendedores': vendedores_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener vendedores: {str(e)}'
        })


@require_GET
@login_required
def obtener_sucursales_reporte(request):
    """API para obtener lista de sucursales según permisos del usuario"""
    try:
        sucursales = obtener_sucursales_usuario(request.user)

        sucursales_data = []
        for sucursal in sucursales:
            sucursales_data.append({
                'id': sucursal.id,
                'nombre': sucursal.alias,
                'direccion': sucursal.direccion
            })

        return JsonResponse({
            'success': True,
            'sucursales': sucursales_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener sucursales: {str(e)}'
        })


@require_GET
@login_required
def obtener_documentos_vendedor_reporte(request):
    """Detalle (drill-down) de la fila de un vendedor del reporte ventas-sucursal.

    Usa el MISMO universo que `obtener_ventas_por_vendedor_reporte` vía
    `_queryset_dtes_ventas_vendedor`, así que la suma de la columna monto es
    exactamente la venta neta de la fila: documentos de venta en positivo y
    notas de crédito en negativo (así restan, igual que en el agregado).

    Los tickets POS ya no se suman al total:
      * el ticket que generó boleta está representado por su DTE — sumar ambos
        contaba dos veces la misma venta;
      * el ticket sin DTE no forma parte del agregado (que es 100% DTE).
    Los tickets sin DTE se devuelven aparte, sólo como dato informativo.
    """
    try:
        vendedor_id = request.GET.get('vendedor_id')
        if not vendedor_id:
            return JsonResponse({'success': False, 'error': 'Se requiere vendedor_id'})

        fi, ff = _parse_rango_fechas_reporte(request)

        qs_dtes = _queryset_dtes_ventas_vendedor(
            request, fi, ff, vendedor_id,
        ).select_related('receptor')

        documentos = []
        for d in qs_dtes.order_by('-fecha_emision', '-id'):
            es_nc = (d.tipo_documento == 'NOTA DE CREDITO')
            monto = int(d.monto_con_iva or 0)
            documentos.append({
                'origen': 'DTE',
                'tipo': d.tipo_documento or '',
                'numero': d.numero_documento or '',
                'fecha': d.fecha_emision.strftime('%d/%m/%Y') if d.fecha_emision else '',
                'cliente': d.receptor.razon_social if d.receptor else 'Público General',
                'sucursal': d.sucursal.alias if d.sucursal else '-',
                # NC con signo negativo: la fila del vendedor es venta NETA.
                'monto': -monto if es_nc else monto,
                'estado': d.estado_dte or '',
                'es_nota_credito': es_nc,
            })

        # ---- Tickets POS sin DTE (informativos, NO suman al total) ----
        qs_tickets = Ticket.objects.filter(
            vendedor_id=vendedor_id,
            created_at__date__gte=fi,
            created_at__date__lte=ff,
            estado='PAGADO',
            dte_generado=False,
        ).select_related('sucursal')
        qs_tickets = filtrar_queryset_por_sucursal(qs_tickets, request.user, request)

        tickets_sin_dte = [{
            'origen': 'TICKET',
            'tipo': 'Ticket POS sin DTE',
            'numero': str(t.id),
            'fecha': t.created_at.strftime('%d/%m/%Y') if t.created_at else '',
            'cliente': 'Venta Directa',
            'sucursal': t.sucursal.alias if t.sucursal else '-',
            'monto': int(t.total or 0),
            'estado': 'PAGADO',
            'computa_en_total': False,
        } for t in qs_tickets.order_by('-created_at')]

        return JsonResponse({
            'success': True,
            'periodo': {'inicio': fi.strftime('%Y-%m-%d'), 'fin': ff.strftime('%Y-%m-%d')},
            'documentos': documentos,
            'total': sum(d['monto'] for d in documentos),
            'cantidad': len(documentos),
            'tickets_sin_dte': tickets_sin_dte,
            'tickets_sin_dte_total': sum(t['monto'] for t in tickets_sin_dte),
            'tickets_sin_dte_cantidad': len(tickets_sin_dte),
        })

    except Exception as e:
        logger.exception("Error al obtener documentos para reporte")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def obtener_comparativa_mensual(request):
    """API para obtener datos de comparativa mensual de sucursales (últimos 6 meses)"""
    try:
        fecha_fin = timezone.now()
        fecha_inicio = fecha_fin - timedelta(days=180)

        sucursales_dict = {}
        meses_set = set()

        # ========== TICKETS (POS nuevo) ==========
        # Solo tickets SIN DTE: un ticket con boleta ya está representado en
        # la suma de DTEs de abajo (misma regla anti-doble-conteo que
        # ventas-global y ventas-comparativo). Sin este filtro la comparativa
        # contaba ~2x las ventas POS (medido jun-2026: $180M duplicados/mes).
        queryset_tickets = Ticket.objects.filter(
            created_at__gte=fecha_inicio,
            estado='PAGADO',
            dte_generado=False,
            modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
        ).select_related('sucursal')

        queryset_tickets = filtrar_queryset_por_sucursal(queryset_tickets, request.user, request)

        tickets_mensuales = queryset_tickets.annotate(
            mes=TruncMonth('created_at')
        ).values(
            'mes',
            'sucursal__alias'
        ).annotate(
            total_ventas=Sum('total')
        ).order_by('mes', 'sucursal__alias')

        for item in tickets_mensuales:
            sucursal_nombre = item['sucursal__alias']
            mes_str = item['mes'].strftime('%Y-%m')
            meses_set.add(mes_str)
            if sucursal_nombre not in sucursales_dict:
                sucursales_dict[sucursal_nombre] = {}
            sucursales_dict[sucursal_nombre][mes_str] = \
                sucursales_dict[sucursal_nombre].get(mes_str, 0) + int(item['total_ventas'] or 0)

        # ========== DTEs (ventas al público, sin facturas entre sucursales) ==========
        queryset_dtes = Dte.objects.filter(
            fecha_emision__gte=fecha_inicio.date(),
            tipo_transaccion='VENTA_PUBLICO',
        ).exclude(
            estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
        ).exclude(
            receptor__isnull=False,
            receptor_id=F('emisor_id')
        ).select_related('sucursal')

        queryset_dtes = filtrar_queryset_por_sucursal(queryset_dtes, request.user, request)

        # Ventas DTE (excluir Notas de Crédito)
        dtes_mensuales = queryset_dtes.exclude(
            tipo_documento='NOTA DE CREDITO'
        ).annotate(
            mes=TruncMonth('fecha_emision')
        ).values(
            'mes',
            'sucursal__alias'
        ).annotate(
            total_ventas=Sum('monto_con_iva')
        ).order_by('mes', 'sucursal__alias')

        for item in dtes_mensuales:
            sucursal_nombre = item['sucursal__alias']
            mes_str = item['mes'].strftime('%Y-%m')
            meses_set.add(mes_str)
            if sucursal_nombre not in sucursales_dict:
                sucursales_dict[sucursal_nombre] = {}
            sucursales_dict[sucursal_nombre][mes_str] = \
                sucursales_dict[sucursal_nombre].get(mes_str, 0) + int(item['total_ventas'] or 0)

        # Restar Notas de Crédito (devoluciones).
        # `queryset_dtes` ya filtró `tipo_transaccion='VENTA_PUBLICO'`, y las
        # NC nunca llevan ese tipo: pedirlas ahí devolvía ~0 y la comparativa
        # mostraba las sucursales sin ninguna devolución descontada. Se
        # rearman desde el helper compartido.
        nc_mensuales = _queryset_ncs_venta(
            fecha_inicio.date(), fecha_fin.date(),
            user=request.user, request=request,
            estados=('EMITIDO', 'ACEPTADO'),
            excluir_internas=True,
        ).select_related('sucursal').annotate(
            mes=TruncMonth('fecha_emision')
        ).values(
            'mes',
            'sucursal__alias'
        ).annotate(
            total_nc=Sum('monto_con_iva')
        ).order_by('mes', 'sucursal__alias')

        for item in nc_mensuales:
            sucursal_nombre = item['sucursal__alias']
            mes_str = item['mes'].strftime('%Y-%m')
            meses_set.add(mes_str)
            if sucursal_nombre not in sucursales_dict:
                sucursales_dict[sucursal_nombre] = {}
            sucursales_dict[sucursal_nombre][mes_str] = \
                sucursales_dict[sucursal_nombre].get(mes_str, 0) - int(item['total_nc'] or 0)

        # Ordenar meses
        meses_ordenados = sorted(list(meses_set))

        # Formatear datos para el gráfico
        series_data = []
        for sucursal, ventas_por_mes in sucursales_dict.items():
            data = [ventas_por_mes.get(mes, 0) for mes in meses_ordenados]
            series_data.append({
                'name': sucursal,
                'data': data
            })

        meses_labels = [datetime.strptime(m, '%Y-%m').strftime('%b %Y') for m in meses_ordenados]

        return JsonResponse({
            'success': True,
            'series': series_data,
            'categories': meses_labels
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener comparativa mensual: {str(e)}'
        })


# ========== REPORTE DE DOCUMENTOS EMITIDOS ==========

@login_required
def ver_documentos_emitidos(request):
    """Vista principal del reporte de documentos emitidos"""
    sucursal_activa_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    sucursal_actual = None
    if sucursal_activa_id:
        sucursal_actual = Sucursal.objects.filter(id=sucursal_activa_id).first()
    context = obtener_contexto_sucursales(request.user, request)
    context['sucursal_actual'] = sucursal_actual
    return render(request, 'vistas/modulo_reportes/documentos_emitidos.html', context)


@require_GET
@login_required
def obtener_documentos_emitidos(request):
    """API para obtener documentos emitidos con filtros"""
    try:
        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_documento = request.GET.get('tipo_documento')
        metodo_pago_filtro = request.GET.get('metodo_pago')
        sucursal_param = (request.GET.get('sucursal_id') or '').strip()

        # Si no se proporcionan fechas, usar el día actual
        if not fecha_desde or not fecha_hasta:
            fecha_fin = timezone.localdate()
            fecha_desde = fecha_fin
            fecha_hasta = fecha_fin
        else:
            fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

        # Consultar DTEs (Boletas y Facturas Electrónicas)
        # NOTA: Incluimos facturas entre sucursales (receptor=emisor) para coincidir
        # con el comportamiento de /app/ventas/documentos/.
        #
        # Se excluyen los estados que NO representan un documento vigente
        # (RECHAZADO por el SII, CANCELADO, ANULADO) y los `descartado=True`
        # (borrado lógico desde gestión de documentos). Antes entraban todos:
        # un documento rechazado o eliminado seguía sumando a los totales por
        # método de pago y al KPI de ventas del reporte.
        queryset = Dte.objects.select_related(
            'vendedor',
            'receptor',
            'sucursal'
        ).prefetch_related(
            'dte_asociado'  # Prefetch de los detalles de pago
        ).filter(
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
            fecha_emision__gte=fecha_desde,
            fecha_emision__lte=fecha_hasta,
            descartado=False,
        ).exclude(
            estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
        )

        # Notas de crédito reales del período (helper compartido). El filtro
        # de arriba jamás las traía: una NC no lleva tipo_transaccion 'VENTA'.
        # Se filtran en paralelo con los mismos criterios de sucursal/tipo/
        # método de pago y se unen a la lista antes de paginar.
        queryset_ncs = _queryset_ncs_venta(
            fecha_desde, fecha_hasta,
            estados=('EMITIDO', 'ACEPTADO'),
        ).select_related(
            'vendedor', 'receptor', 'sucursal'
        ).prefetch_related('dte_asociado')

        # --- Filtro de sucursal con semántica explícita ---
        # Reglas:
        #   - sucursal_param == 'all'  → sin filtro (traer todas las sucursales accesibles)
        #   - sucursal_param == '<id>' → filtrar por esa sucursal (si tiene permiso)
        #   - sucursal_param vacío     → usar sucursal activa de la sesión (default seguro
        #                                que coincide con /app/ventas/documentos/)
        sucursal_sesion_id = (
            request.session.get('idSucursalActual')
            or request.session.get('sucursalActual')
        )
        filtro_sucursal_desc = ''
        sucursal_ids_permitidas = list(
            obtener_sucursales_usuario(request.user).values_list('id', flat=True)
        )

        # La decisión de sucursal se resuelve UNA vez y se aplica igual al
        # queryset de ventas y al de notas de crédito (si divergieran, las NC
        # de otra sucursal restarían de un total que no las contiene).
        if sucursal_param == 'all':
            if usuario_puede_ver_todas_sucursales(request.user):
                filtro_sucursal_desc = 'Todas las sucursales accesibles'
            else:
                # Usuario no puede ver todas → se restringe a las asignadas
                filtro_sucursal_desc = 'Sucursales asignadas al usuario'

            def _aplicar_sucursal(qs):
                return qs.filter(sucursal_id__in=sucursal_ids_permitidas) \
                    if sucursal_ids_permitidas else qs.none()
        elif sucursal_param:
            # Validar permiso sobre esa sucursal específica
            if puede_ver_sucursal(request.user, sucursal_param):
                filtro_sucursal_desc = f'Sucursal específica (id={sucursal_param})'

                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_param)
            elif sucursal_sesion_id:
                # Fallback: sucursal de sesión si la tiene
                filtro_sucursal_desc = (
                    f'Sin permiso para id={sucursal_param}, se usó sucursal de sesión'
                )

                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_sesion_id)
            else:
                filtro_sucursal_desc = 'Sin permiso y sin sucursal de sesión'

                def _aplicar_sucursal(qs):
                    return qs.none()
        else:
            # Default: sucursal activa de la sesión
            if sucursal_sesion_id:
                filtro_sucursal_desc = f'Sucursal activa de sesión (id={sucursal_sesion_id})'

                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_sesion_id)
            else:
                # Si no hay sesión, restringir a las sucursales del usuario
                filtro_sucursal_desc = 'Sin sucursal de sesión, se usaron las accesibles'

                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id__in=sucursal_ids_permitidas) \
                        if sucursal_ids_permitidas else qs.none()

        queryset = _aplicar_sucursal(queryset)
        queryset_ncs = _aplicar_sucursal(queryset_ncs)

        # Aplicar filtros
        MAPA_TIPO_DOCUMENTO = {
            'BOLETA_ELECTRONICA': 'BOLETA ELECTRONICA',
            'BOLETA_PAPEL': 'BOLETA PAPEL',
            'FACTURA_ELECTRONICA': 'FACTURA ELECTRONICA',
            'FACTURA_EXENTA': 'FACTURA EXENTA',
            'BOLETA': 'BOLETA',
        }
        if tipo_documento:
            tipo_doc_db = MAPA_TIPO_DOCUMENTO.get(tipo_documento)
            if tipo_doc_db:
                queryset = queryset.filter(tipo_documento=tipo_doc_db)
                # Al pedir un tipo específico de venta, las NC no corresponden
                # (el usuario está mirando un subconjunto, no un neto).
                queryset_ncs = queryset_ncs.none()
            elif tipo_documento in ('NOTA_DE_CREDITO', 'NOTA DE CREDITO'):
                queryset = queryset.none()

        # ✅ FILTRO POR MÉTODO DE PAGO - Aplicado ANTES de limitar registros
        if metodo_pago_filtro:
            from django.db.models import Q
            # Filtrar por método de pago en Dte_Detalle_Pago
            # Usamos distinct() porque un DTE puede tener múltiples detalles de pago
            queryset = queryset.filter(
                Q(dte_asociado__metodo_pago__icontains=metodo_pago_filtro)
            ).distinct()
            queryset_ncs = queryset_ncs.filter(
                Q(dte_asociado__metodo_pago__icontains=metodo_pago_filtro)
            ).distinct()

            # Si el filtro es por método de pago específico del sistema de tickets,
            # también necesitamos incluir DTEs sin detalles de pago pero con ticket relacionado
            # (Esta es una optimización adicional que puede implementarse después)

        # Una sola materialización para la lista y el resumen, con los pagos
        # prefetched (antes: 2 recorridos del queryset + 2-4 queries por DTE).
        # Ventas y NC se ordenan juntas para que la paginación no separe las
        # devoluciones al final del listado.
        dtes_todos = sorted(
            list(queryset) + list(queryset_ncs),
            key=lambda d: (d.fecha_emision, int(d.numero_documento or 0)),
            reverse=True,
        )

        # Fallback masivo de tickets para los DTEs sin detalle de pago
        # (reemplaza la búsqueda de Ticket/TicketDetallePago fila a fila).
        #
        # El vínculo Ticket↔Dte es `Ticket.folio_dte == Dte.numero_documento`
        # (mismo criterio que views_modulo_documentos y la cuadratura). Antes
        # se cruzaba contra `Ticket.correlativo`, que es la secuencia interna
        # del ticket y NO tiene relación con el folio del CAF: el "match"
        # devolvía el ticket equivocado o ninguno, y el método de pago caía al
        # default 'Efectivo'. Las NC se excluyen del cruce a propósito: su
        # folio pertenece a otra serie y podría colisionar con el de una boleta.
        claves_sin_pago = {
            (int(d.numero_documento), d.sucursal_id)
            for d in dtes_todos
            if not d.dte_asociado.all()
            and d.numero_documento is not None
            and (d.tipo_documento or '').upper() != 'NOTA DE CREDITO'
        }
        tickets_map = {}
        pagos_ticket_map = {}
        if claves_sin_pago:
            folios = {f for f, _ in claves_sin_pago}
            sucursales_cruce = {s for _, s in claves_sin_pago if s is not None}
            # Acotado por sucursal y por fecha (antes: `correlativo__in=...` a
            # secas, que barría la tabla Ticket completa de todas las
            # sucursales). La banda es holgada a propósito: `Ticket.fecha` es
            # `auto_now=True`, o sea se mueve con cada save posterior, así que
            # un margen justo dejaría fuera tickets reeditados.
            qs_tickets_cruce = Ticket.objects.filter(
                folio_dte__in=folios,
                fecha__gte=fecha_desde - timedelta(days=365),
                fecha__lte=fecha_hasta + timedelta(days=365),
            )
            if sucursales_cruce:
                qs_tickets_cruce = qs_tickets_cruce.filter(
                    sucursal_id__in=sucursales_cruce
                )
            for t in qs_tickets_cruce:
                tickets_map.setdefault((t.folio_dte, t.sucursal_id), t)
            ids_tickets = [t.id for t in tickets_map.values()]
            if ids_tickets:
                for pt in TicketDetallePago.objects.filter(
                        ticket_id__in=ids_tickets).order_by('id'):
                    pagos_ticket_map.setdefault(pt.ticket_id, []).append(pt)

        # Preparar datos de documentos
        documentos_data = []
        total_real = len(dtes_todos)

        # Paginación real (antes: truncado silencioso a 100 filas que dejaba
        # documentos invisibles sin aviso cuando el rango superaba los 100).
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (TypeError, ValueError):
            page = 1
        try:
            per_page = int(request.GET.get('per_page', 100))
        except (TypeError, ValueError):
            per_page = 100
        per_page = max(10, min(per_page, 500))
        offset = (page - 1) * per_page

        for dte in dtes_todos[offset:offset + per_page]:
            # Obtener información del cliente
            cliente_info = 'N.N'
            if dte.receptor:
                cliente_info = dte.receptor.nombre
                if dte.receptor.rut:
                    cliente_info += f' ({dte.receptor.rut})'
            
            # Obtener método de pago (desde detalles de pago)
            metodo_pago_display = 'No especificado'
            voucher = ''
            tipo_tarjeta = ''
            pagos_detalle = {}
            
            # Buscar en detalles de pago del DTE (prefetched, sin query extra)
            detalles_pago = list(dte.dte_asociado.all())

            if detalles_pago:
                # Si hay múltiples pagos, concatenar
                metodos = []
                vouchers = []
                for detalle in detalles_pago:
                    metodos.append(detalle.metodo_pago)
                    if detalle.voucher:
                        vouchers.append(str(detalle.voucher))
                    if detalle.tipo_tarjeta and not tipo_tarjeta:
                        tipo_tarjeta = detalle.tipo_tarjeta
                    pagos_detalle[detalle.metodo_pago] = pagos_detalle.get(detalle.metodo_pago, 0) + int(detalle.monto or 0)
                
                metodo_pago_display = ', '.join(metodos) if metodos else 'No especificado'
                voucher = ', '.join(vouchers) if vouchers else ''
            else:
                # Sin detalles de pago: fallback al ticket relacionado desde el
                # mapa precalculado (sin query por fila). Cruce por folio_dte.
                ticket_relacionado = tickets_map.get(
                    (int(dte.numero_documento), dte.sucursal_id)
                ) if dte.numero_documento is not None else None

                if ticket_relacionado:
                    metodo_pago_display = dict(METODO_PAGO_TICKET_CHOICES).get(
                        ticket_relacionado.metodo_pago, 
                        ticket_relacionado.metodo_pago
                    )
                    
                    # Buscar voucher en detalles de pago del ticket (mapa)
                    pagos_ticket = pagos_ticket_map.get(ticket_relacionado.id, [])
                    if pagos_ticket:
                        vouchers = []
                        for pago_ticket in pagos_ticket:
                            if pago_ticket.voucher:
                                vouchers.append(str(pago_ticket.voucher))
                            if pago_ticket.tipo_tarjeta and not tipo_tarjeta:
                                tipo_tarjeta = pago_ticket.tipo_tarjeta
                            metodo_ticket = dict(METODO_PAGO_TICKET_CHOICES).get(
                                pago_ticket.metodo_pago,
                                pago_ticket.metodo_pago
                            )
                            pagos_detalle[metodo_ticket] = pagos_detalle.get(metodo_ticket, 0) + int(pago_ticket.monto or 0)
                        voucher = ', '.join(vouchers) if vouchers else ''
                else:
                    metodo_pago_display = 'Efectivo'  # Default si no hay info
                    pagos_detalle['Efectivo'] = int(dte.monto_con_iva or 0)
            
            pagos_detalle_str = ', '.join(
                [f"{metodo} ${int(monto):,}".replace(',', '.') for metodo, monto in pagos_detalle.items()]
            ) if pagos_detalle else 'No especificado'

            # Monto efectivamente pagado = suma de pagos registrados.
            # `monto_con_iva` YA viene neto del descuento (ver el cálculo de
            # `total_dte` en views_modulo_ventas.generar_dte_desde_ticket), así
            # que el fallback NO debe volver a restarlo.
            pagado = sum(pagos_detalle.values())
            if not pagado:
                pagado = int(dte.monto_con_iva or 0)

            # Descuento real = el que quedó grabado en el DTE, punto.
            # Antes, si `dte.descuento` era 0 y los pagos sumaban menos que el
            # total, la diferencia se declaraba "descuento". Pero esa
            # diferencia es un SALDO NO PAGADO: una venta a crédito, un
            # convenio o un pago parcial. Se expone aparte con su nombre real.
            descuento_dte = int(dte.descuento) if dte.descuento else 0
            saldo_no_pagado_doc = max(int(dte.monto_con_iva or 0) - pagado, 0)

            documentos_data.append({
                'id': dte.id,
                'tipo_documento': dte.tipo_documento,
                'tipo_documento_display': dte.tipo_documento,
                'metodo_pago': metodo_pago_display,
                'metodo_pago_display': metodo_pago_display,
                'pagos_detalle': pagos_detalle,
                'pagos_detalle_str': pagos_detalle_str,
                'correlativo': dte.numero_documento,
                'cliente_info': cliente_info,
                'total': int(dte.monto_con_iva),
                'descuento': descuento_dte,
                'saldo_no_pagado': saldo_no_pagado_doc,
                'pagado': pagado,
                'es_nota_credito': (dte.tipo_documento or '').upper() == 'NOTA DE CREDITO',
                'vendedor': dte.vendedor.nombre if dte.vendedor else 'N/A',
                'vendedor_codigo': dte.vendedor.codigo_vendedor if dte.vendedor else '',
                'fecha': dte.fecha_emision.strftime('%d/%m/%Y'),
                'voucher': voucher,
                'tipo_tarjeta': tipo_tarjeta
            })
        
        # Calcular resúmenes por método de pago
        resumen = {
            'efectivo': 0,
            'tbk_credito': 0,
            'tbk_debito': 0,
            'tarjeta_comercial': 0,
            'convenio': 0,
            'descuentos': 0,
            'venta_internet': 0,
            'transferencia': 0,
            'credito_trabajador': 0,
            'otros': 0,
            'notas_credito': 0,   # total de NC (se resta del global)
            # `monto_con_iva` YA viene neto del descuento de línea, así que
            # `ventas_brutas` es en realidad "vendido después de descuento".
            'ventas_brutas': 0,
            # Informativo: descuentos ya aplicados dentro de ventas_brutas.
            # NO se vuelven a restar (ver total_global).
            'descuentos': 0,
            # Diferencia entre lo facturado y lo efectivamente pagado
            # (crédito, convenio, pago parcial). Antes esto se reportaba
            # como "descuento" y se restaba del total.
            'saldo_no_pagado': 0,
            'total_global': 0     # ventas_brutas - notas_credito
        }
        
        # --- Helper: clasifica un método de pago en la categoría del resumen ---
        # Importante: el orden de verificación evita colisiones (ej. CREDITO_TRABAJADOR
        # debe detectarse ANTES que CREDITO genérico de tarjeta).
        def clasificar_metodo(metodo_raw):
            m = (metodo_raw or '').upper()
            if not m:
                return 'otros'
            # Específicos primero
            if 'TRABAJADOR' in m:
                return 'credito_trabajador'
            if 'INTERNET' in m or 'WEB' in m:
                return 'venta_internet'
            if 'TRANSFERENCIA' in m:
                return 'transferencia'
            if 'CONVENIO' in m:
                return 'convenio'
            if 'COMERCIAL' in m or 'PARIS' in m or 'RIPLEY' in m or 'FALABELLA' in m:
                return 'tarjeta_comercial'
            if 'EFECTIVO' in m:
                return 'efectivo'
            # Tarjetas TBK/bancarias
            if 'DEBITO' in m or 'REDCOMPRA' in m:
                return 'tbk_debito'
            if ('CREDITO' in m or 'VISA' in m or 'MASTERCARD' in m
                    or 'AMEX' in m or 'DINER' in m):
                return 'tbk_credito'
            # Otros (CHEQUE, OTRO, ORDEN_COMPRA, TBK_MANUAL, TBK_PREPAGO_POS,
            # CREDITO_EXTERNO sin TRABAJADOR, etc.)
            return 'otros'

        # --- Diagnóstico: contadores para cuadrar con /app/ventas/documentos/ ---
        diag_conteo_por_tipo = {}
        diag_sucursales = {}          # {sucursal_id: {'alias': ..., 'cantidad': N, 'monto': $}}
        diag_cant_dtes = 0            # DTEs "normales" (no NC) que aportan a ventas brutas
        diag_cant_nc = 0              # Notas de crédito
        diag_sin_pagos = 0            # DTEs sin registros en dte_asociado
        # Equivalente al 'total_ventas' de /app/ventas/documentos/: suma de
        # pagos cuando existen, monto_con_iva cuando no (misma regla que
        # `listar_documentos_ventas`).
        total_equivalente_ventas_doc = 0

        # Calcular totales por método de pago desde Dte_Detalle_Pago
        # (misma lista materializada; sin segunda evaluación del queryset)
        for dte in dtes_todos:
            total = int(dte.monto_con_iva or 0)
            descuento_db = int(dte.descuento) if dte.descuento else 0
            tipo_doc_upper = dte.tipo_documento.upper() if dte.tipo_documento else ''

            # --- Diagnóstico: conteos por tipo y sucursal ---
            tipo_key = dte.tipo_documento or 'SIN_TIPO'
            if tipo_key not in diag_conteo_por_tipo:
                diag_conteo_por_tipo[tipo_key] = {'cantidad': 0, 'monto': 0}
            diag_conteo_por_tipo[tipo_key]['cantidad'] += 1
            diag_conteo_por_tipo[tipo_key]['monto'] += total

            suc_id = dte.sucursal_id
            if suc_id not in diag_sucursales:
                suc_alias = getattr(dte.sucursal, 'alias', None) or getattr(dte.sucursal, 'nombreSucursal', '') or f'ID {suc_id}'
                diag_sucursales[suc_id] = {
                    'id': suc_id,
                    'alias': suc_alias,
                    'cantidad': 0,
                    'monto': 0,
                }
            diag_sucursales[suc_id]['cantidad'] += 1
            diag_sucursales[suc_id]['monto'] += total

            # Las Notas de Crédito se acumulan aparte y se restan del total
            if 'NOTA' in tipo_doc_upper and 'CREDITO' in tipo_doc_upper:
                resumen['notas_credito'] += total
                diag_cant_nc += 1
                continue

            diag_cant_dtes += 1

            # Obtener métodos de pago del DTE (prefetched)
            detalles_pago = list(dte.dte_asociado.all())
            pagado_sum = 0

            if detalles_pago:
                for detalle in detalles_pago:
                    monto = int(detalle.monto or 0)
                    pagado_sum += monto
                    categoria = clasificar_metodo(detalle.metodo_pago)
                    resumen[categoria] = resumen.get(categoria, 0) + monto
            else:
                diag_sin_pagos += 1

            # Descuento = el grabado en el DTE. Es SOLO informativo: ya está
            # descontado dentro de `monto_con_iva`, así que no se vuelve a
            # restar del total. La diferencia entre lo facturado y lo pagado
            # NO es descuento (es crédito / convenio / pago parcial) y va a
            # su propio acumulador.
            resumen['descuentos'] += descuento_db
            if pagado_sum > 0 and pagado_sum < total:
                resumen['saldo_no_pagado'] += total - pagado_sum

            resumen['ventas_brutas'] += total
            total_equivalente_ventas_doc += pagado_sum if pagado_sum > 0 else total

            # Si no se procesó método de pago (no había registros en dte_asociado),
            # buscar en el ticket relacionado como fallback (mapa precalculado).
            if not detalles_pago:
                ticket_relacionado = tickets_map.get(
                    (int(dte.numero_documento), dte.sucursal_id)
                ) if dte.numero_documento is not None else None

                if ticket_relacionado:
                    metodo_ticket = ticket_relacionado.metodo_pago

                    if metodo_ticket == 'TBK_POS_INTEGRADO':
                        # Verificar tipo de tarjeta en detalles (mapa)
                        pagos_tr = pagos_ticket_map.get(ticket_relacionado.id, [])
                        pago_ticket = pagos_tr[0] if pagos_tr else None
                        if pago_ticket and pago_ticket.tipo_tarjeta and \
                                'DEBITO' in pago_ticket.tipo_tarjeta.upper():
                            resumen['tbk_debito'] += total
                        elif pago_ticket and pago_ticket.tipo_tarjeta:
                            resumen['tbk_credito'] += total
                        else:
                            resumen['tbk_debito'] += total  # default
                    else:
                        categoria = clasificar_metodo(metodo_ticket)
                        resumen[categoria] = resumen.get(categoria, 0) + total
                else:
                    # Si no hay ticket ni detalle de pago, asumir efectivo
                    resumen['efectivo'] += total

        # Total neto = ventas brutas - notas de crédito.
        # Los descuentos NO se restan acá: `monto_con_iva` se graba ya neto
        # del descuento de línea (`total_dte = suma_items - descuentos` en
        # generar_dte_desde_ticket), así que restarlos de nuevo bajaba el
        # total del reporte por el monto completo de los descuentos del período.
        resumen['total_global'] = resumen['ventas_brutas'] - resumen['notas_credito']

        # --- Bloque de diagnóstico (para cuadrar con /app/ventas/documentos/) ---
        # Permite identificar rápidamente discrepancias: universo de sucursales,
        # tipos de documento incluidos, DTEs sin pagos registrados, etc.
        sucursales_incluidas = sorted(
            diag_sucursales.values(),
            key=lambda s: s.get('monto', 0),
            reverse=True,
        )
        conteo_por_tipo = [
            {'tipo': t, 'cantidad': v['cantidad'], 'monto': v['monto']}
            for t, v in sorted(
                diag_conteo_por_tipo.items(),
                key=lambda kv: kv[1]['monto'],
                reverse=True,
            )
        ]
        diagnostico = {
            'filtro_sucursal_aplicado': filtro_sucursal_desc,
            'sucursal_param_recibido': sucursal_param or '(vacío)',
            'sucursal_sesion_id': sucursal_sesion_id,
            'fecha_desde': str(fecha_desde),
            'fecha_hasta': str(fecha_hasta),
            'sucursales_incluidas': sucursales_incluidas,
            'cantidad_sucursales_incluidas': len(sucursales_incluidas),
            'conteo_por_tipo': conteo_por_tipo,
            'cantidad_dtes': diag_cant_dtes,
            'cantidad_notas_credito': diag_cant_nc,
            'dtes_sin_pagos_registrados': diag_sin_pagos,
            # Totales útiles para cuadrar entre vistas
            'total_bruto_sin_descuento': resumen['ventas_brutas'],
            'total_equivalente_ventas_documentos': total_equivalente_ventas_doc,
            'total_neto': resumen['total_global'],
        }

        total_pages = max(1, (total_real + per_page - 1) // per_page)
        return JsonResponse({
            'success': True,
            'documentos': documentos_data,
            'resumen': resumen,
            'diagnostico': diagnostico,
            'total_registros': len(documentos_data),
            'total_real': total_real,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total_real,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_previous': page > 1,
            },
        })

    except Exception as e:
        logger.exception("Error al obtener documentos emitidos")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener documentos: {str(e)}'
        })


@require_GET
@login_required
def exportar_documentos_emitidos_excel(request):
    """Exportar documentos emitidos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Parámetros de filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_documento = request.GET.get('tipo_documento')
        metodo_pago_filtro = request.GET.get('metodo_pago')
        sucursal_param = (request.GET.get('sucursal_id') or '').strip()

        # Si no se proporcionan fechas, usar el día actual
        if not fecha_desde or not fecha_hasta:
            fecha_fin = timezone.localdate()
            fecha_desde = fecha_fin
            fecha_hasta = fecha_fin
        else:
            fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

        # Consultar DTEs (Boletas y Facturas Electrónicas)
        # NOTA: Sin exclusión de facturas internas (receptor=emisor) para que el
        # Excel coincida con /app/ventas/documentos/ y con la vista en pantalla.
        # Mismos filtros de vigencia que `obtener_documentos_emitidos`: fuera
        # los rechazados/cancelados/anulados y los descartados.
        queryset = Dte.objects.select_related(
            'vendedor',
            'receptor',
            'sucursal'
        ).prefetch_related(
            'dte_asociado'
        ).filter(
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
            fecha_emision__gte=fecha_desde,
            fecha_emision__lte=fecha_hasta,
            descartado=False,
        ).exclude(
            estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
        )

        # Las NC nunca entraban al Excel (no llevan tipo_transaccion 'VENTA'),
        # así que el archivo listaba solo ventas y quien sumaba la columna
        # "Total" obtenía el bruto sin devoluciones.
        queryset_ncs = _queryset_ncs_venta(
            fecha_desde, fecha_hasta,
            estados=('EMITIDO', 'ACEPTADO'),
        ).select_related(
            'vendedor', 'receptor', 'sucursal'
        ).prefetch_related('dte_asociado')

        # Filtro de sucursal con la misma semántica que obtener_documentos_emitidos:
        #   'all' = todas, vacío = sucursal de sesión, id = específica.
        sucursal_sesion_id = (
            request.session.get('idSucursalActual')
            or request.session.get('sucursalActual')
        )
        sucursal_ids_permitidas = list(
            obtener_sucursales_usuario(request.user).values_list('id', flat=True)
        )

        if sucursal_param == 'all':
            def _aplicar_sucursal(qs):
                return qs.filter(sucursal_id__in=sucursal_ids_permitidas) \
                    if sucursal_ids_permitidas else qs.none()
        elif sucursal_param:
            if puede_ver_sucursal(request.user, sucursal_param):
                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_param)
            elif sucursal_sesion_id:
                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_sesion_id)
            else:
                def _aplicar_sucursal(qs):
                    return qs.none()
        else:
            if sucursal_sesion_id:
                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id=sucursal_sesion_id)
            else:
                def _aplicar_sucursal(qs):
                    return qs.filter(sucursal_id__in=sucursal_ids_permitidas) \
                        if sucursal_ids_permitidas else qs.none()

        queryset = _aplicar_sucursal(queryset)
        queryset_ncs = _aplicar_sucursal(queryset_ncs)

        # Aplicar filtros
        MAPA_TIPO_DOCUMENTO = {
            'BOLETA_ELECTRONICA': 'BOLETA ELECTRONICA',
            'BOLETA_PAPEL': 'BOLETA PAPEL',
            'FACTURA_ELECTRONICA': 'FACTURA ELECTRONICA',
            'FACTURA_EXENTA': 'FACTURA EXENTA',
            'BOLETA': 'BOLETA',
        }
        if tipo_documento:
            tipo_doc_db = MAPA_TIPO_DOCUMENTO.get(tipo_documento)
            if tipo_doc_db:
                queryset = queryset.filter(tipo_documento=tipo_doc_db)
                queryset_ncs = queryset_ncs.none()
            elif tipo_documento in ('NOTA_DE_CREDITO', 'NOTA DE CREDITO'):
                queryset = queryset.none()

        if metodo_pago_filtro:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(dte_asociado__metodo_pago__icontains=metodo_pago_filtro)
            ).distinct()
            queryset_ncs = queryset_ncs.filter(
                Q(dte_asociado__metodo_pago__icontains=metodo_pago_filtro)
            ).distinct()

        documentos_excel = sorted(
            list(queryset) + list(queryset_ncs),
            key=lambda d: (d.fecha_emision, int(d.numero_documento or 0)),
            reverse=True,
        )

        # Cruce masivo Ticket↔Dte por `folio_dte` (la clave real; antes se
        # consultaba `correlativo=dte.numero_documento` fila a fila, que además
        # de ser N+1 comparaba dos secuencias distintas y nunca acertaba).
        claves_cruce = {
            (int(d.numero_documento), d.sucursal_id)
            for d in documentos_excel
            if not d.dte_asociado.all()
            and d.numero_documento is not None
            and (d.tipo_documento or '').upper() != 'NOTA DE CREDITO'
        }
        tickets_map = {}
        primer_pago_ticket = {}
        if claves_cruce:
            folios = {f for f, _ in claves_cruce}
            sucursales_cruce = {s for _, s in claves_cruce if s is not None}
            qs_tickets_cruce = Ticket.objects.filter(
                folio_dte__in=folios,
                fecha__gte=fecha_desde - timedelta(days=365),
                fecha__lte=fecha_hasta + timedelta(days=365),
            )
            if sucursales_cruce:
                qs_tickets_cruce = qs_tickets_cruce.filter(
                    sucursal_id__in=sucursales_cruce
                )
            for t in qs_tickets_cruce:
                tickets_map.setdefault((t.folio_dte, t.sucursal_id), t)
            ids_tickets = [t.id for t in tickets_map.values()]
            if ids_tickets:
                for pt in TicketDetallePago.objects.filter(
                        ticket_id__in=ids_tickets).order_by('id'):
                    primer_pago_ticket.setdefault(pt.ticket_id, pt)

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documentos Emitidos"

        # Encabezados
        headers = [
            'ID', 'Tipo Documento', 'Método Pago', 'N° Documento', 'Cliente/RUT',
            'Total', 'Descuento', 'Pagado', 'Vendedor', 'Fecha', 'Voucher'
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4154F1", end_color="4154F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for dte in documentos_excel:
            es_nc = (dte.tipo_documento or '').upper() == 'NOTA DE CREDITO'

            # Cliente
            cliente_info = 'N.N'
            if dte.receptor:
                cliente_info = dte.receptor.nombre
                if dte.receptor.rut:
                    cliente_info += f' ({dte.receptor.rut})'

            # Método de pago y voucher
            metodo_pago_display = 'No especificado'
            voucher = ''
            tipo_tarjeta = ''

            detalles_pago = list(dte.dte_asociado.all())
            if detalles_pago:
                metodos = []
                vouchers = []
                for detalle in detalles_pago:
                    metodos.append(detalle.metodo_pago)
                    if detalle.voucher:
                        vouchers.append(str(detalle.voucher))
                    if detalle.tipo_tarjeta and not tipo_tarjeta:
                        tipo_tarjeta = detalle.tipo_tarjeta

                metodo_pago_display = ', '.join(metodos) if metodos else 'No especificado'
                voucher = ', '.join(vouchers) if vouchers else ''
            else:
                ticket_relacionado = tickets_map.get(
                    (int(dte.numero_documento), dte.sucursal_id)
                ) if (dte.numero_documento is not None and not es_nc) else None

                if ticket_relacionado:
                    metodo_pago_display = dict(METODO_PAGO_TICKET_CHOICES).get(
                        ticket_relacionado.metodo_pago,
                        ticket_relacionado.metodo_pago
                    )
                    pago_ticket = primer_pago_ticket.get(ticket_relacionado.id)
                    if pago_ticket:
                        voucher = pago_ticket.voucher or ''
                        tipo_tarjeta = pago_ticket.tipo_tarjeta or ''
                elif not es_nc:
                    metodo_pago_display = 'Efectivo'

            vendedor_nombre = dte.vendedor.nombre if dte.vendedor else 'N/A'
            vendedor_codigo = dte.vendedor.codigo_vendedor if dte.vendedor else ''
            vendedor_display = (
                f'{vendedor_nombre} ({vendedor_codigo})' if vendedor_codigo else vendedor_nombre
            )

            # `monto_con_iva` ya viene neto del descuento (ver
            # generar_dte_desde_ticket), así que "Pagado" NO puede volver a
            # restarlo: el Excel mostraba cada venta con descuento por debajo
            # de lo realmente cobrado. Las NC van con signo negativo para que
            # la suma de la columna "Total" dé el neto del período.
            signo = -1 if es_nc else 1
            total = signo * int(dte.monto_con_iva or 0)
            descuento = int(dte.descuento) if dte.descuento else 0
            pagado = total

            ws.append([
                dte.id,
                dte.tipo_documento,
                metodo_pago_display,
                dte.numero_documento,
                cliente_info,
                total,
                descuento,
                pagado,
                vendedor_display,
                dte.fecha_emision.strftime('%d/%m/%Y'),
                voucher
            ])

        # Ajustar ancho de columnas
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"documentos_emitidos_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar documentos emitidos: {str(e)}'
        })


# ========== REPORTE DE EXISTENCIAS POR MARCA ==========

@login_required
def ver_reporte_existencias_marca(request):
    """Vista principal del reporte de existencias por marca"""
    # Obtener sucursal actual para pasarla al template
    sucursal_actual_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    sucursal_actual = None
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            pass
    
    context = {
        'sucursal_actual': sucursal_actual,
        'sucursal_actual_id': sucursal_actual_id,
    }
    return render(request, 'vistas/modulo_reportes/reporte_existencias_marca.html', context)


@require_GET
@login_required
def obtener_reporte_existencias_marca(request):
    """
    API OPTIMIZADA para reporte de existencias por marca.
    
    CAMBIOS DE OPTIMIZACIÓN:
    1. Requiere al menos un filtro obligatorio (marca, departamento, búsqueda o sucursal)
    2. Usa la sucursal actual del usuario por defecto
    3. Limita resultados a 500 productos máximo
    4. Usa agregación en BD en lugar de Python
    5. Solo carga los campos necesarios con .only()
    """
    try:
        # ========== PARÁMETROS DE FILTRO ==========
        marca_id = request.GET.get('marca_id')
        departamento_id = request.GET.get('departamento_id')
        busqueda = request.GET.get('busqueda', '').strip()
        sucursal_id = request.GET.get('sucursal_id')
        solo_con_stock = request.GET.get('solo_con_stock', 'true') == 'true'
        limite = int(request.GET.get('limite', 500))
        sin_filtro = request.GET.get('sin_filtro', 'false') == 'true'
        
        # 'todas' indica explícitamente todas las sucursales del usuario (no aplicar fallback a sesión)
        todas_sucursales = (str(sucursal_id).lower() == 'todas') if sucursal_id else False
        if todas_sucursales:
            sucursal_id = None

        # Obtener sucursal actual del usuario si no se especifica (y no se pidió 'todas')
        if not sucursal_id and not todas_sucursales:
            sucursal_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')

        # ========== VALIDACIÓN: REQUIERE AL MENOS UN FILTRO (A MENOS QUE FUERCE SIN FILTRO) ==========
        tiene_filtro = any([marca_id, departamento_id, busqueda])

        if not tiene_filtro and not sin_filtro:
            return JsonResponse({
                'success': False,
                'requiere_filtro': True,
                'error': 'Por favor selecciona al menos un filtro: Marca, Departamento o usa el buscador de artículos.',
                'sugerencia': 'Esto optimiza la consulta y evita cargar los 50,000+ productos.'
            })

        # Si no hay filtro, limitar más estrictamente
        if not tiene_filtro:
            limite = min(limite, 500)  # Máximo 500 sin filtro

        # ========== OBTENER SUCURSALES ==========
        # El universo permitido se resuelve SIEMPRE desde las empresas del
        # usuario; una sucursal pedida por querystring se intersecta con ese
        # universo. Antes se confiaba en el sucursal_id recibido y bastaba
        # pasar el id de otra empresa para ver su stock, costo y PVP.
        empresas_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True
        ).values_list('empresa_id', flat=True)
        permitidas = Sucursal.objects.filter(empresa_id__in=empresas_usuario)

        if sucursal_id:
            sucursales = permitidas.filter(id=sucursal_id)
            if not sucursales.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes acceso a esa sucursal',
                }, status=403)
        else:
            sucursales = permitidas.order_by('alias')


        sucursales_list = list(sucursales)
        sucursales_ids = [s.id for s in sucursales_list]
        
        # ========== CONSULTA OPTIMIZADA ==========
        # Base: productos de las sucursales del usuario
        queryset = Producto.objects.filter(
            sucursal_id__in=sucursales_ids,
            excluir_de_analitica=False
        ).select_related(
            'atributo1', 'atributo2', 'categoria', 'sucursal'
        ).only(
            'id', 'articulo', 'costo', 'precioventa', 'sucursal_id',
            'atributo1__id', 'atributo1__valor',
            'atributo2__id', 'atributo2__valor',
            'categoria__id', 'categoria__nombre',
            'sucursal__id', 'sucursal__alias'
        )

        # ========== APLICAR FILTROS ==========
        if marca_id:
            queryset = queryset.filter(atributo1_id=marca_id)

        if departamento_id:
            # Árbol v1.2: un padre incluye todas sus hijas
            queryset = queryset.filter(categoria_id__in=_expandir_categoria_ids(departamento_id))

        if busqueda:
            # Búsqueda por artículo, marca o color
            queryset = queryset.filter(
                Q(articulo__icontains=busqueda) |
                Q(atributo1__valor__icontains=busqueda) |
                Q(atributo2__valor__icontains=busqueda)
            )

        # ========== AGREGAR STOCK CON ANOTACIÓN ==========
        # `base_filtrado` conserva una fila Producto por sucursal; se necesita
        # entera más abajo para armar la columna de cada sucursal.
        base_filtrado = queryset
        queryset = queryset.annotate(
            total_stock_anotado=Sum('producto_talla__stock')
        )

        # Filtrar solo productos con stock si se requiere
        if solo_con_stock:
            queryset = queryset.filter(total_stock_anotado__gt=0)

        # ========== LIMITAR POR ARTÍCULO, NO POR FILA ==========
        # El límite se aplicaba sobre filas Producto (una por sucursal), así que
        # con N sucursales se mostraban ~limite/N artículos y, peor, cada
        # artículo quedaba con solo algunas de sus sucursales: los totales de la
        # tabla eran incompletos sin avisarlo. Ahora se eligen primero los
        # artículos y después se traen TODAS sus filas.
        articulos_visibles = list(
            queryset.order_by('atributo1__valor', 'articulo')
            .values_list('articulo', flat=True)
            .distinct()[:limite]
        )
        total_articulos_disponibles = queryset.values('articulo').distinct().count()
        truncado = total_articulos_disponibles > len(articulos_visibles)

        # ========== PROCESAR DATOS (PIVOT POR SKU) ==========
        # Cada producto en BD pertenece a 1 sucursal. Aquí agrupamos por
        # (articulo, marca, color, departamento) y armamos un dict
        # stock_por_sucursal: {sucursal_id: stock} con TODAS las sucursales del usuario.
        productos_lista = list(
            base_filtrado.filter(articulo__in=articulos_visibles)
            .order_by('atributo1__valor', 'articulo')
        ) if articulos_visibles else []
        productos_procesados = len(productos_lista)

        # Prefetch de tallas para los productos filtrados (stock real por producto)
        productos_ids = [p.id for p in productos_lista]
        tallas_por_producto = {}
        if productos_ids:
            from django.db.models import Sum as DjangoSum
            tallas = Producto_Talla.objects.filter(
                producto_id__in=productos_ids
            ).values('producto_id').annotate(
                stock_total=DjangoSum('stock')
            )
            tallas_por_producto = {t['producto_id']: t['stock_total'] or 0 for t in tallas}

        # Agrupar productos por (articulo, marca_id, atributo2_id, categoria_id)
        # para fusionar variantes que existen en distintas sucursales en una sola fila.
        agrupados = {}
        for producto in productos_lista:
            stock_total_producto = tallas_por_producto.get(producto.id, 0)
            clave = (
                producto.articulo,
                producto.atributo1_id,
                producto.atributo2_id,
                producto.categoria_id,
            )
            if clave not in agrupados:
                agrupados[clave] = {
                    'articulo': producto.articulo,
                    'marca': producto.atributo1.valor if producto.atributo1 else 'Sin Marca',
                    'marca_id': producto.atributo1.id if producto.atributo1 else None,
                    'color': producto.atributo2.valor if producto.atributo2 else '-',
                    'departamento': producto.categoria.nombre if producto.categoria else '-',
                    'costo': float(producto.costo) if producto.costo else 0,
                    'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                    'stock_por_sucursal': {},  # {sucursal_id: stock}
                    'sucursales': {},          # {alias: {stock, sucursal_id}} — retrocompat
                    'total_stock': 0,
                }

            fila = agrupados[clave]
            if producto.sucursal_id:
                # Si el SKU aparece varias veces en la misma sucursal (no debería),
                # sumamos para no perder stock.
                key_id = str(producto.sucursal_id)
                fila['stock_por_sucursal'][key_id] = (
                    fila['stock_por_sucursal'].get(key_id, 0) + stock_total_producto
                )
                if producto.sucursal:
                    fila['sucursales'][producto.sucursal.alias] = {
                        'stock': fila['sucursales'].get(producto.sucursal.alias, {}).get('stock', 0) + stock_total_producto,
                        'sucursal_id': producto.sucursal_id,
                    }
            fila['total_stock'] += stock_total_producto

        # Aplicar filtro solo_con_stock sobre el TOTAL agrupado
        datos_reporte = [
            fila for fila in agrupados.values()
            if (not solo_con_stock) or fila['total_stock'] > 0
        ]

        # Ordenar por marca, luego artículo
        datos_reporte.sort(key=lambda f: (f.get('marca') or '', f.get('articulo') or ''))
        
        # ========== SALUD DEL UNIVERSO FILTRADO (cobertura / stock viejo) ==========
        salud = None
        if productos_ids:
            from datetime import timedelta as _td
            from app.constants_kardex import CONCEPTOS_VENTA as _CV
            from app.models import LoteProducto as _Lote
            _hoy = timezone.localdate()
            _stock_univ = sum(tallas_por_producto.values())
            _vend30 = abs(
                Movimientos_Producto.objects.filter(
                    concepto__in=_CV, estado='COMPLETADO',
                    fecha__gte=_hoy - _td(days=30),
                    ProductoTalla__producto_id__in=productos_ids,
                ).aggregate(s=Sum('cantidad'))['s'] or 0
            )
            _viejo = _Lote.objects.filter(
                activo=True, cantidad_disponible__gt=0,
                producto_talla__producto_id__in=productos_ids,
                fecha_ingreso__date__lte=_hoy - _td(days=181),
            ).aggregate(s=Sum('cantidad_disponible'))['s'] or 0
            _vel = _vend30 / 30 if _vend30 else 0
            salud = {
                'cobertura_dias': int(_stock_univ / _vel) if _vel else None,
                'pct_stock_viejo': round(100 * _viejo / _stock_univ, 1) if _stock_univ else 0,
                'vendidas_30': _vend30,
                'stock_total': _stock_univ,
            }

        # ========== RESPUESTA ==========
        sucursales_data = [{'id': s.id, 'alias': s.alias} for s in sucursales_list]

        return JsonResponse({
            'success': True,
            'datos': datos_reporte,
            'sucursales': sucursales_data,
            'salud': salud,
            'filtros_aplicados': {
                'marca_id': marca_id,
                'departamento_id': departamento_id,
                'busqueda': busqueda,
                'sucursal_id': sucursal_id,
                'solo_con_stock': solo_con_stock,
            },
            # Visible para la UI: si la lista viene cortada hay que decirlo,
            # o el usuario lee un inventario parcial como si fuera el total.
            'truncado': truncado,
            'articulos_mostrados': len(articulos_visibles),
            'articulos_disponibles': total_articulos_disponibles,
            'debug': {
                'productos_procesados': productos_procesados,
                'registros_reporte': len(datos_reporte),
                'total_sucursales': len(sucursales_data),
                'limite_aplicado': limite
            }
        })
        
    except Exception as e:
        logger.exception("Error en reporte de existencias por marca")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener reporte: {str(e)}'
        })


@require_GET
@login_required
def exportar_existencias_marca_excel(request):
    """Exportar reporte de existencias por marca a Excel (por sucursales)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Obtener datos del reporte
        marca_id = request.GET.get('marca_id')
        departamento_id = request.GET.get('departamento_id')
        
        # Reutilizar la función de obtención de datos
        temp_request = request
        response_data = obtener_reporte_existencias_marca(temp_request)
        datos = json.loads(response_data.content)
        
        if not datos.get('success'):
            return JsonResponse({
                'success': False,
                'error': 'No se pudieron obtener los datos'
            })
        
        datos_reporte = datos.get('datos', [])
        sucursales_lista = datos.get('sucursales', [])
        
        if not datos_reporte:
            return JsonResponse({
                'success': False,
                'error': 'No hay datos para exportar'
            })
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Existencias por Marca"
        
        # Estilos
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sucursal_fill = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
        sucursal_font = Font(bold=True, color="004085", size=10)
        total_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        total_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Agrupar por marca
        por_marca = {}
        for item in datos_reporte:
            marca = item['marca']
            if marca not in por_marca:
                por_marca[marca] = []
            por_marca[marca].append(item)
        
        fila_actual = 1
        
        # Procesar cada marca
        for marca, productos in sorted(por_marca.items()):
            # Filtrar solo sucursales con stock en esta marca
            sucursales_con_stock = []
            for sucursal in sucursales_lista:
                tiene_stock = any(
                    sucursal['alias'] in p['sucursales'] and 
                    p['sucursales'][sucursal['alias']]['stock'] > 0 
                    for p in productos
                )
                if tiene_stock:
                    sucursales_con_stock.append(sucursal)
            
            # Título de la marca
            ws.merge_cells(f'A{fila_actual}:E{fila_actual}')
            cell = ws[f'A{fila_actual}']
            cell.value = f"MARCA: {marca}"
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            fila_actual += 1
            
            # Encabezados principales
            headers_row1 = ['Artículo', 'Color', 'Depart', 'Costo', 'PrecioV.']
            
            # Agregar solo columnas de sucursales con stock
            for sucursal in sucursales_con_stock:
                headers_row1.append(sucursal['alias'])
                headers_row1.append('')  # Para el stock
                
            headers_row1.append('TT (TOTAL)')
            headers_row1.append('')  # Para el stock total
            
            # Escribir primera fila de encabezados
            for idx, header in enumerate(headers_row1, start=1):
                cell = ws.cell(row=fila_actual, column=idx, value=header if header else None)
                if header and header not in ['Artículo', 'Color', 'Depart', 'Costo', 'PrecioV.']:
                    cell.fill = sucursal_fill
                    cell.font = sucursal_font
                else:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            fila_actual += 1
            
            # Subencabezados (Stock)
            headers_row2 = ['', '', '', '', '']  # Columnas fijas sin subencabezado
            for sucursal in sucursales_con_stock:
                headers_row2.extend(['Stock', 'Stock'])
            headers_row2.extend(['Stock', 'Stock'])
            
            for idx, header in enumerate(headers_row2, start=1):
                if header:
                    cell = ws.cell(row=fila_actual, column=idx, value=header)
                    if idx > 5:  # Solo las columnas de sucursales
                        cell.fill = sucursal_fill
                        cell.font = sucursal_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
            
            fila_actual += 1
            
            # Inicializar totales por columna
            totales_por_sucursal = {suc['alias']: {'stock': 0} for suc in sucursales_con_stock}
            gran_total_stock = 0

            # Datos de productos
            for producto in productos:
                row_data = [
                    producto['articulo'],
                    producto['color'],
                    producto['departamento'],
                    producto['costo'],
                    producto['precio_venta']
                ]

                # Datos por sucursal (solo las que tienen stock)
                for sucursal in sucursales_con_stock:
                    stock_suc = producto['sucursales'].get(sucursal['alias'])
                    if stock_suc:
                        stock_val = stock_suc.get('stock', 0)
                        row_data.append(stock_val)
                        row_data.append(stock_val)
                        # Acumular totales
                        totales_por_sucursal[sucursal['alias']]['stock'] += stock_val
                    else:
                        row_data.append('-')
                        row_data.append('-')

                # Totales del producto
                total_stock_prod = producto.get('total_stock', 0)
                row_data.append(total_stock_prod)
                row_data.append(total_stock_prod)
                gran_total_stock += total_stock_prod
                
                # Escribir fila
                for idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=fila_actual, column=idx, value=value)
                    
                    # Aplicar formato
                    if idx > 5:  # Columnas numéricas
                        if idx in [len(row_data) - 1, len(row_data)]:  # Columnas de total
                            cell.fill = total_fill
                            cell.font = total_font
                        
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.alignment = Alignment(horizontal='left' if idx <= 2 else 'center')
                    
                    cell.border = border
                
                fila_actual += 1
            
            # Fila de TOTALES
            total_fill_yellow = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            ws.merge_cells(f'A{fila_actual}:E{fila_actual}')
            cell = ws[f'A{fila_actual}']
            cell.value = "TOTALES:"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = total_fill
            cell.border = border
            
            col_idx = 6
            for sucursal in sucursales_con_stock:
                total_suc = totales_por_sucursal[sucursal['alias']]
                stock_val = total_suc['stock']

                cell = ws.cell(row=fila_actual, column=col_idx, value=stock_val)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                col_idx += 1

                cell = ws.cell(row=fila_actual, column=col_idx, value=stock_val)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                col_idx += 1

            # Gran total
            cell = ws.cell(row=fila_actual, column=col_idx, value=gran_total_stock)
            cell.font = Font(bold=True, size=12)
            cell.fill = total_fill_yellow
            cell.alignment = Alignment(horizontal='right')
            cell.border = border

            cell = ws.cell(row=fila_actual, column=col_idx + 1, value=gran_total_stock)
            cell.font = Font(bold=True, size=12)
            cell.fill = total_fill_yellow
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            fila_actual += 1
            
            # Espacio entre marcas
            fila_actual += 2
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        for i in range(6, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_existencias_marca.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Error al exportar reporte de existencias por marca a Excel")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        })


# ========== REPORTE DE EXISTENCIAS POR SUCURSAL ==========

@login_required
def ver_reporte_existencias_sucursal(request):
    """Vista principal del reporte de existencias por sucursal"""
    return render(request, 'vistas/modulo_reportes/reporte_existencias_sucursal.html')


@require_GET
@login_required
def obtener_reporte_existencias_sucursal(request):
    """API para obtener datos del reporte de existencias por sucursal"""
    try:
        sucursal_id = request.GET.get('sucursal_id')
        marca_id = request.GET.get('marca_id')
        incluir_sin_stock = request.GET.get('incluir_sin_stock', 'false').lower() == 'true'

        if not sucursal_id:
            return JsonResponse({'success': False, 'error': 'Sucursal es requerida'})

        # Validar que el usuario tenga acceso a esta sucursal
        empresas_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True
        ).values_list('empresa_id', flat=True)
        sucursal = get_object_or_404(Sucursal, id=sucursal_id, empresa_id__in=empresas_usuario)

        # Obtener productos de esta sucursal con sus tallas en un solo queryset
        queryset = Producto.objects.filter(
            sucursal_id=sucursal_id,
            excluir_de_analitica=False
        ).select_related(
            'atributo1', 'atributo2', 'atributo3', 'categoria'
        ).prefetch_related('producto_talla')

        if marca_id:
            queryset = queryset.filter(atributo1_id=marca_id)

        # Pre-cargar total recibido desde Movimientos_Producto (todos los ingresos reales)
        # La FK en Movimientos_Producto se llama "ProductoTalla" (mayúsculas)
        talla_ids = list(
            Producto_Talla.objects.filter(
                producto__sucursal_id=sucursal_id,
                **({'producto__atributo1_id': marca_id} if marca_id else {})
            ).values_list('id', flat=True)
        )

        from django.db.models import Sum as _Sum
        # "Recibido hist." = TODO lo que alguna vez SUMÓ stock a este SKU.
        #
        # Antes se usaba una lista blanca de conceptos y se sumaba `cantidad`
        # sin exigir signo, lo que producía el absurdo "recibido histórico <
        # stock actual":
        #   - dejaba fuera los traspasos legacy de una pierna
        #     (TRASPASO_SUCURSAL/BODEGA/VITRINA), que son la vía de entrada de
        #     buena parte del catálogo migrado de Laravel;
        #   - RECEPCION_COMPRA con cantidad negativa RESTABA del recibido.
        # Medición contra prod (jul-2026): SKECHERS tenía 1.063 SKUs con stock
        # vivo y "recibido" por debajo (1.534 unidades; 976 de esos SKUs en 0),
        # todos explicados por TRASPASO_SUCURSAL.
        #
        # Clasificar por SIGNO en vez de por concepto no hay que mantenerlo
        # cuando se agrega un concepto nuevo, y hace cumplir el invariante por
        # construcción: como stock = SUM(entradas) + SUM(salidas) y las salidas
        # son <= 0, sumar solo cantidad > 0 garantiza recibido >= stock siempre
        # que el kardex esté completo.
        stocks_iniciales_qs = (
            Movimientos_Producto.objects.filter(
                ProductoTalla_id__in=talla_ids,
                estado='COMPLETADO',
                cantidad__gt=0,
            )
            .values('ProductoTalla_id')
            .annotate(total=_Sum('cantidad'))
        )
        stocks_iniciales_map = {
            row['ProductoTalla_id']: row['total']
            for row in stocks_iniciales_qs
        }

        datos_reporte = []
        total_stock = 0
        valor_inventario = 0
        valor_venta_potencial = 0
        skus_con_stock = 0
        productos_sin_stock = 0

        for producto in queryset:
            # Costo de adquisicion: para vendedoras incluye sobreprecio (precio interno)
            costo_unitario = (producto.costo or 0) + (producto.sobreprecio or 0)

            for talla in producto.producto_talla.all():
                stock = talla.stock or 0

                if stock <= 0:
                    productos_sin_stock += 1
                    if not incluir_sin_stock:
                        continue
                else:
                    skus_con_stock += 1
                    total_stock += stock
                    valor_inventario += costo_unitario * stock
                    valor_venta_potencial += (producto.precioventa or 0) * stock

                stock_inicial = stocks_iniciales_map.get(talla.id, 0) or 0

                datos_reporte.append({
                    'articulo': producto.articulo,
                    'descripcion': producto.descripcion or '-',
                    'categoria': producto.categoria.nombre if producto.categoria else '-',
                    'marca': producto.atributo1.valor if producto.atributo1 else '-',
                    'color': producto.atributo2.valor if producto.atributo2 else '-',
                    'genero': producto.atributo3.valor if producto.atributo3 else '-',
                    'talla': talla.talla if talla.talla else '-',
                    'sku': str(talla.sku),
                    'stock_inicial': stock_inicial,
                    'stock': stock,
                    'costo': float(producto.costo or 0),
                    'sobreprecio': float(producto.sobreprecio) if producto.sobreprecio else 0,
                    'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                    'sin_stock': stock <= 0,
                })

        resumen = {
            'total_productos': skus_con_stock,
            'stock_total': total_stock,
            'valor_inventario': valor_inventario,
            'sin_stock': productos_sin_stock,
            'sucursal': sucursal.alias or sucursal.nombre or f'Sucursal {sucursal.id}',
        }

        # KPIs de salud de inventario de la sucursal: cobertura en días,
        # % de stock envejecido (lotes >180d) y valor venta potencial.
        from datetime import timedelta as _td
        from app.constants_kardex import CONCEPTOS_VENTA as _CV
        from app.models import LoteProducto as _Lote
        _hoy = timezone.localdate()
        vendidas_30 = abs(
            Movimientos_Producto.objects.filter(
                concepto__in=_CV, estado='COMPLETADO',
                fecha__gte=_hoy - _td(days=30),
                sucursal_origen_id=sucursal_id,
            ).aggregate(s=_Sum('cantidad'))['s'] or 0
        )
        stock_viejo = _Lote.objects.filter(
            activo=True, cantidad_disponible__gt=0,
            producto_talla__producto__sucursal_id=sucursal_id,
            fecha_ingreso__date__lte=_hoy - _td(days=181),
        ).aggregate(s=_Sum('cantidad_disponible'))['s'] or 0
        _velocidad = vendidas_30 / 30 if vendidas_30 else 0
        resumen.update({
            'valor_venta_potencial': valor_venta_potencial,
            'cobertura_dias': int(total_stock / _velocidad) if _velocidad else None,
            'pct_stock_viejo': round(100 * stock_viejo / total_stock, 1) if total_stock else 0,
            'vendidas_30': vendidas_30,
        })

        return JsonResponse({'success': True, 'datos': datos_reporte, 'resumen': resumen})

    except Exception as e:
        logger.exception("Error en reporte de existencias por sucursal")
        return JsonResponse({'success': False, 'error': f'Error al obtener reporte: {str(e)}'})


def _get_existencias_datos(request):
    """Helper que devuelve (datos_reporte, resumen, error) para exportaciones."""
    response_data = obtener_reporte_existencias_sucursal(request)
    datos = json.loads(response_data.content)
    if not datos.get('success'):
        return None, None, datos.get('error', 'Error desconocido')
    return datos.get('datos', []), datos.get('resumen', {}), None


@require_GET
@login_required
def exportar_existencias_sucursal_excel(request):
    """Exportar reporte de existencias por sucursal a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not request.GET.get('sucursal_id'):
            return JsonResponse({'success': False, 'error': 'Sucursal es requerida'})

        datos_reporte, resumen, error = _get_existencias_datos(request)
        if error:
            return JsonResponse({'success': False, 'error': error})
        if not datos_reporte:
            return JsonResponse({'success': False, 'error': 'No hay datos para exportar'})

        nombre_sucursal = resumen.get('sucursal', 'Sucursal')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Exist. {nombre_sucursal}"[:31]

        verde = "1E7E34"
        verde_claro = "28A745"
        rojo_claro = "FFF0F0"
        header_fill = PatternFill(start_color=verde, end_color=verde, fill_type="solid")
        subheader_fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Fila 1: Título
        ws.merge_cells('A1:M1')
        cell = ws['A1']
        cell.value = f"REPORTE DE EXISTENCIAS — {nombre_sucursal.upper()}"
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Fila 2: Fecha
        ws.merge_cells('A2:M2')
        from datetime import date as _date
        ws['A2'].value = f"Generado: {timezone.localdate().strftime('%d/%m/%Y')}   |   Total registros: {resumen.get('total_productos', 0)}   |   Stock total: {resumen.get('stock_total', 0):,}"
        ws['A2'].font = Font(italic=True, size=10, color="444444")
        ws['A2'].alignment = Alignment(horizontal='center')

        # Fila 3: Encabezados
        headers = [
            'SKU', 'Artículo', 'Descripción', 'Categoría', 'Marca', 'Color', 'Género',
            'Talla', 'Stock Inicial', 'Stock Actual', 'Costo', 'Sobreprecio', 'Precio Venta'
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.fill = subheader_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[3].height = 20

        # Datos
        for fila_idx, item in enumerate(datos_reporte, start=4):
            sin_stock = item.get('sin_stock', False)
            row_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid") if sin_stock else None

            def _cell(col, val, align='left'):
                c = ws.cell(row=fila_idx, column=col, value=val)
                c.border = border
                c.alignment = Alignment(horizontal=align)
                if row_fill:
                    c.fill = row_fill
                return c

            _cell(1, item['sku'])
            _cell(2, item['articulo'])
            _cell(3, item['descripcion'])
            _cell(4, item['categoria'])
            _cell(5, item['marca'])
            _cell(6, item['color'])
            _cell(7, item['genero'])
            _cell(8, item['talla'], 'center')
            c = _cell(9, item['stock_inicial'], 'right')
            c.number_format = '#,##0'
            c = _cell(10, item['stock'], 'right')
            c.number_format = '#,##0'
            if sin_stock:
                c.font = Font(color="CC0000", bold=True)
            c = _cell(11, item['costo'], 'right')
            c.number_format = '$#,##0'
            c = _cell(12, item['sobreprecio'], 'right')
            c.number_format = '$#,##0'
            c = _cell(13, item['precio_venta'], 'right')
            c.number_format = '$#,##0'

        # Anchos
        col_widths = [14, 22, 32, 16, 16, 14, 12, 8, 12, 12, 12, 12, 14]
        for i, w in enumerate(col_widths, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A4'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_name = nombre_sucursal.replace(' ', '_').replace('/', '-')
        response['Content-Disposition'] = f'attachment; filename="existencias_{safe_name}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        logger.exception("Error al exportar existencias por sucursal a Excel")
        return JsonResponse({'success': False, 'error': f'Error al exportar: {str(e)}'})


@require_GET
@login_required
def exportar_existencias_sucursal_pdf(request):
    """Exportar reporte de existencias por sucursal a PDF usando ReportLab"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from io import BytesIO

        if not request.GET.get('sucursal_id'):
            return JsonResponse({'success': False, 'error': 'Sucursal es requerida'})

        datos_reporte, resumen, error = _get_existencias_datos(request)
        if error:
            return JsonResponse({'success': False, 'error': error})
        if not datos_reporte:
            return JsonResponse({'success': False, 'error': 'No hay datos para exportar'})

        nombre_sucursal = resumen.get('sucursal', 'Sucursal')
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        verde = colors.HexColor('#1E7E34')
        verde_claro = colors.HexColor('#28A745')
        gris_claro = colors.HexColor('#F8F9FA')
        rojo_claro = colors.HexColor('#FFE0E0')

        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, textColor=verde, spaceAfter=4)
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey, spaceAfter=12)

        elements = []
        from datetime import date as _date
        elements.append(Paragraph(f"Reporte de Existencias — {nombre_sucursal}", title_style))
        elements.append(Paragraph(
            f"Generado: {timezone.localdate().strftime('%d/%m/%Y')}  |  "
            f"Total SKUs: {resumen.get('total_productos', 0):,}  |  "
            f"Stock total: {resumen.get('stock_total', 0):,}  |  "
            f"Valor inventario: ${resumen.get('valor_inventario', 0):,.0f}",
            sub_style
        ))

        # Tabla
        col_labels = ['Artículo', 'Descripción', 'Marca', 'Color', 'Género', 'Talla',
                      'Stock Ini.', 'Stock Act.', 'Costo', 'Precio V.']
        col_widths_pdf = [3.5*cm, 6*cm, 3*cm, 2.8*cm, 2.5*cm, 1.8*cm,
                          2*cm, 2.2*cm, 2.5*cm, 2.5*cm]

        table_data = [col_labels]
        for item in datos_reporte:
            row = [
                item['articulo'],
                item['descripcion'][:45] + ('…' if len(item['descripcion']) > 45 else ''),
                item['marca'],
                item['color'],
                item['genero'],
                item['talla'],
                f"{item['stock_inicial']:,}",
                f"{item['stock']:,}",
                f"${item['costo']:,.0f}",
                f"${item['precio_venta']:,.0f}",
            ]
            table_data.append(row)

        table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

        row_count = len(table_data)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), verde),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DEE2E6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, gris_claro]),
            ('ALIGN', (6, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        # Resaltar filas sin stock
        for row_idx, item in enumerate(datos_reporte, start=1):
            if item.get('sin_stock'):
                style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), rojo_claro))
                style_cmds.append(('TEXTCOLOR', (7, row_idx), (7, row_idx), colors.red))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.grey)
            canvas.drawString(1.5*cm, 1*cm, f"RetailMind — {nombre_sucursal} — {timezone.localdate().strftime('%d/%m/%Y')}")
            canvas.drawRightString(landscape(A4)[0] - 1.5*cm, 1*cm, f"Página {doc.page}")
            canvas.restoreState()

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        buffer.seek(0)

        response = HttpResponse(content_type='application/pdf')
        safe_name = nombre_sucursal.replace(' ', '_').replace('/', '-')
        response['Content-Disposition'] = f'attachment; filename="existencias_{safe_name}.pdf"'
        response.write(buffer.read())
        return response

    except Exception as e:
        logger.exception("Error al exportar existencias por sucursal a PDF")
        return JsonResponse({'success': False, 'error': f'Error al exportar PDF: {str(e)}'})


# ========== REPORTE DE COMPRAS INTEGRAL ==========

@login_required
def ver_reporte_compras(request):
    """Vista principal del reporte de compras"""
    return render(request, 'vistas/modulo_reportes/reporte_compras.html')


# ========== REPORTE: PRODUCTOS CREADOS POR ORIGEN (gap H4) ==========
# El ORIGEN de un SKU = el concepto de su PRIMER movimiento de INGRESO.
# Permite ver, de forma agregada, cuántos SKU nacieron por compra vs alta
# manual vs traspaso vs ajuste, en lugar de auditar SKU por SKU.

ORIGEN_LABELS = {
    'RECEPCION_COMPRA': 'Recepción de compra',
    'INGRESO_INICIAL': 'Alta inicial / saldo',
    'INGRESO_MANUAL': 'Alta manual',
    # En datos migrados de Laravel el primer ingreso del SKU a la tienda suele ser
    # el despacho interno bodega→tienda, registrado como TRASPASO_SUCURSAL (INGRESO).
    'TRASPASO_SUCURSAL': 'Despacho interno (traspaso)',
    'TRASPASO_ENTRADA': 'Traspaso desde otra sucursal',
    'REPOSICION_STOCK': 'Reposición',
    'AJUSTE_POSITIVO': 'Ajuste de inventario',
    'AJUSTE_INVENTARIO': 'Ajuste de inventario',
    'DEVOLUCION_CLIENTE': 'Devolución de cliente',
}


@login_required
def ver_reporte_productos_origen(request):
    """Vista principal: productos creados en el período clasificados por origen."""
    return render(request, 'vistas/modulo_reportes/productos_por_origen.html')


@login_required
@require_GET
def api_productos_por_origen(request):
    """
    Altas de catálogo del período, clasificadas por el ORIGEN del SKU = el
    concepto de su PRIMER movimiento de INGRESO. Responde el gap H4 de la
    auditoría de trazabilidad.

    GET params:
        anio      — año a analizar (default = actual)
        sucursal  — id de sucursal (opcional)

    SCOPING: el universo de SKU se acota a las sucursales de las empresas del
    usuario. Antes no había ningún filtro por empresa, así que sin pasar
    `sucursal` se veían las altas de catálogo, unidades y costo de TODO el
    holding, y pasando el id de otra empresa se veían las de esa empresa.
    """
    from django.db.models import Min
    from .utils_analitica import productos_activos_qs

    try:
        anio = int(request.GET.get('anio', timezone.localdate().year))
    except (TypeError, ValueError):
        anio = timezone.localdate().year
    sucursal_id = request.GET.get('sucursal') or None

    permitidas, sin_acceso = _alcance_sucursales(request, sucursal_id)
    if sin_acceso:
        return sin_acceso

    # `Producto` tiene FK a `Sucursal`, así que la sucursal del producto es el
    # anclaje directo del SKU a la empresa dueña del catálogo.
    pt_ids = (
        productos_activos_qs()
        .filter(**_filtro_alcance('producto__sucursal_id', sucursal_id, permitidas))
        .values_list('id', flat=True)
    )

    base = Movimientos_Producto.objects.filter(
        tipo_movimiento='INGRESO', estado='COMPLETADO',
        ProductoTalla_id__in=pt_ids,
    )

    # 1) Primer ingreso (fecha mínima) por SKU
    primeras = {
        r['ProductoTalla_id']: r['primera']
        for r in base.values('ProductoTalla_id').annotate(primera=Min('fecha'))
    }
    # SKU cuyo primer ingreso cae en el año pedido = "creados" ese año
    pt_anio = [pt for pt, f in primeras.items() if f and f.year == anio]

    # 2) Concepto / costo / cantidad del PRIMER movimiento de esos SKU.
    #    Recorremos ordenado por (SKU, fecha, hora, id) y tomamos el primero por SKU.
    por_origen = defaultdict(lambda: {'skus': 0, 'unidades': 0, 'costo': 0.0})
    por_mes = defaultdict(lambda: {'skus': 0, 'unidades': 0})
    vistos = set()
    total_skus = total_unidades = 0
    total_costo = 0.0

    if pt_anio:
        movs = (
            base.filter(ProductoTalla_id__in=pt_anio)
            .values('ProductoTalla_id', 'fecha', 'hora', 'concepto', 'cantidad', 'costo')
            .order_by('ProductoTalla_id', 'fecha', 'hora', 'id')
        )
        for m in movs.iterator(chunk_size=5000):
            pt = m['ProductoTalla_id']
            if pt in vistos:
                continue   # ya tomamos su primer movimiento
            vistos.add(pt)
            uds = abs(int(m['cantidad'] or 0))
            costo = float(m['costo'] or 0) * uds
            o = por_origen[m['concepto'] or 'OTRO']
            o['skus'] += 1
            o['unidades'] += uds
            o['costo'] += costo
            pm = por_mes[m['fecha'].month if m['fecha'] else 0]
            pm['skus'] += 1
            pm['unidades'] += uds
            total_skus += 1
            total_unidades += uds
            total_costo += costo

    _CONCEPTOS_IRREGULARES = {'AJUSTE_POSITIVO', 'AJUSTE_INVENTARIO', 'DEVOLUCION_CLIENTE', 'INGRESO_MANUAL'}

    por_origen_data = sorted((
        {
            'concepto': concepto,
            'origen': ORIGEN_LABELS.get(concepto, 'Otro'),
            'skus': v['skus'],
            'unidades': v['unidades'],
            'costo': round(v['costo'], 0),
            'pct': round(v['skus'] / total_skus * 100, 1) if total_skus else 0,
            'pct_costo': round(v['costo'] / total_costo * 100, 1) if total_costo else 0,
            'costo_por_sku': round(v['costo'] / v['skus']) if v['skus'] else 0,
            'alerta': concepto in _CONCEPTOS_IRREGULARES,
        }
        for concepto, v in por_origen.items()
    ), key=lambda x: x['skus'], reverse=True)

    por_mes_data = [
        {'mes': mm, 'skus': por_mes.get(mm, {}).get('skus', 0),
         'unidades': por_mes.get(mm, {}).get('unidades', 0)}
        for mm in range(1, 13)
    ]

    mes_pico = max(por_mes_data, key=lambda m: m['skus']) if por_mes_data else {'mes': 0, 'skus': 0}

    anios = sorted({f.year for f in primeras.values() if f}, reverse=True)
    anio_actual = timezone.localdate().year
    if anio_actual not in anios:
        anios.insert(0, anio_actual)

    return JsonResponse({
        'success': True,
        'anio': anio,
        'resumen': {
            'total_skus': total_skus,
            'total_unidades': total_unidades,
            'total_costo': round(total_costo, 0),
            'costo_por_sku': round(total_costo / total_skus) if total_skus else 0,
            'mes_pico': mes_pico['mes'],
            'mes_pico_skus': mes_pico['skus'],
        },
        'por_origen': por_origen_data,
        'por_mes': por_mes_data,
        'anios_disponibles': anios,
    })


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE DE COMPRAS INTEGRAL
# ══════════════════════════════════════════════════════════════════════════════
# Reglas transversales del reporte (antes cada función aplicaba las suyas):
#   · SCOPING: el universo de DTE se acota a las empresas del usuario
#     (EmpresaUser status=True) en el rol de RECEPTOR. Sin esto cualquier
#     usuario veía inversión, deuda y proveedores de TODO el holding.
#   · Un DTE descartado o ANULADO/CANCELADO/RECHAZADO no es una compra viva.
#   · Las NOTA DE CREDITO de compra RESTAN (devolución al proveedor). Antes
#     sumaban a la inversión igual que una factura.
#   · La DEUDA se calcula por saldo real (monto_con_iva − pagos registrados).
#     El campo estado_pago no sirve: en producción convive 'PAGADO'/'Pagado'/
#     'Pendiente'/'Abonado' (fuera de choices) y nunca descuenta los abonos.
#   · CUMPLIMIENTO del proveedor = unidades recepcionadas / unidades pedidas.
#     El ratio de facturas pagadas es un indicador de tesorería y vive en la
#     pestaña de Pagos.
ESTADOS_DTE_ANULADOS_COMPRAS = ['ANULADO', 'CANCELADO', 'RECHAZADO']
TIPO_DOC_NC_COMPRAS = 'NOTA DE CREDITO'
ESTADOS_COMPRA_VIGENTES = ['ACTIVA', 'COMPLETADA']

# Familia normalizada de temporada -> (valor de Compras.temporada_familia,
# fragmentos de texto libre usados antes de que existiera ese campo).
TEMPORADAS_COMPRAS = {
    'VERANO': ('VERANO', ['verano']),
    'INVIERNO': ('INVIERNO', ['invierno']),
    'OTONO': ('OTONO', ['otoño', 'otono']),
    'PRIMAVERA': ('PRIMAVERA', ['primavera']),
}


def _normalizar_texto_temporada(valor):
    """'Otoño' -> 'OTONO'. El <select> manda la ñ y la BD guarda ambas formas."""
    import unicodedata
    if not valor:
        return ''
    descompuesto = unicodedata.normalize('NFKD', str(valor))
    return ''.join(c for c in descompuesto if not unicodedata.combining(c)).strip().upper()


def _rango_periodo_compras(anio, periodo, hoy):
    """
    Traduce el filtro 'Período' a un rango de fechas real dentro del año.
    Antes se calculaba en la vista y no se usaba en ninguna query: el selector
    se veía en pantalla pero no filtraba nada.
    """
    from datetime import date
    inicio_anio = date(anio, 1, 1)
    fin_anio = date(anio, 12, 31)
    # Para un año que no es el actual, "últimos N días" se ancla al cierre de
    # ese año; si no, cualquier período relativo devolvería siempre vacío.
    referencia = hoy if anio == hoy.year else fin_anio
    dias = {'semana': 7, 'mes': 30, 'trimestre': 90}.get(periodo)
    if not dias:
        return inicio_anio, fin_anio
    return max(inicio_anio, referencia - timedelta(days=dias)), min(fin_anio, referencia)


def _construir_filtros_compras(request):
    """
    Resuelve UNA vez los filtros del request y el universo de empresas que el
    usuario puede ver. Todas las funciones del reporte reciben este dict.
    """
    from .models import Dte, Sucursal
    from .utils_permisos import obtener_empresas_usuario

    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get('anio') or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year
    periodo = (request.GET.get('periodo') or 'anual').strip().lower()
    temporada = (request.GET.get('temporada') or '').strip()
    proveedor_id = (request.GET.get('proveedor') or '').strip()
    if not proveedor_id.isdigit():
        proveedor_id = ''

    empresas_permitidas = list(
        obtener_empresas_usuario(request.user).values_list('id', flat=True)
    )

    sucursal_activa_id = request.session.get('idSucursalActual')
    sucursal_activa = None
    if sucursal_activa_id:
        sucursal_activa = Sucursal.objects.filter(id=sucursal_activa_id).first()
    es_vendedora = bool(sucursal_activa and sucursal_activa.es_solo_vendedora)

    # La empresa de la sesión ACOTA, nunca amplía: si no está entre las
    # permitidas se descarta (viene de sesión, no es un dato confiable).
    empresa_activa_id = (getattr(sucursal_activa, 'empresa_id', None)
                         or request.session.get('idEmpresaActual'))
    try:
        # La sesión puede traerlo como string; sin castear no matchea nunca
        # contra los ids permitidos y el reporte se abría a todas las empresas.
        empresa_activa_id = int(empresa_activa_id) if empresa_activa_id else None
    except (TypeError, ValueError):
        empresa_activa_id = None
    if empresa_activa_id and empresa_activa_id in empresas_permitidas:
        empresas_ids = [empresa_activa_id]
    else:
        empresa_activa_id = None
        empresas_ids = list(empresas_permitidas)

    empresa_receptora_id = empresa_activa_id if es_vendedora else None
    if es_vendedora and not empresa_receptora_id:
        # Sucursal vendedora fuera del alcance del usuario: universo vacío.
        empresas_ids = []

    fecha_inicio, fecha_fin = _rango_periodo_compras(anio, periodo, hoy)

    # El modelo Compras (órdenes) NO tiene empresa compradora: `empresa` es el
    # PROVEEDOR, así que las órdenes no se pueden scopear por sí solas. Se
    # atribuyen únicamente a quien opera un centro de distribución Y además
    # recibe facturas de compra en el período; de lo contrario una tienda vería
    # las órdenes (y el cumplimiento) del CD del holding.
    incluye_cd = (
        not es_vendedora
        and bool(empresas_ids)
        and Sucursal.objects.filter(
            empresa_id__in=empresas_ids, es_centro_distribucion=True
        ).exists()
        and Dte.objects.filter(
            tipo_transaccion='COMPRA',
            receptor_id__in=empresas_ids,
            fecha_emision__range=(fecha_inicio, fecha_fin),
        ).exists()
    )

    filtros = {
        'anio': anio,
        'periodo': periodo,
        'temporada': temporada,
        'proveedor_id': proveedor_id,
        'empresas_ids': empresas_ids,
        'empresa_activa_id': empresa_activa_id,
        'es_vendedora': es_vendedora,
        'empresa_receptora_id': empresa_receptora_id,
        'sucursal_activa_id': sucursal_activa_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'hoy': hoy,
        'incluye_cd': incluye_cd,
    }
    filtros['compras_ids'] = _compras_ids_por_temporada(filtros)
    filtros['temporada_aplicada'] = filtros['compras_ids'] is not None
    return filtros


def _compras_ids_por_temporada(filtros):
    """
    Ids de Compras de la temporada pedida, o None si no hay filtro.
    `temporada_familia` es un campo reciente: el histórico solo tiene texto
    libre ('Verano 2026', 'verano'), por eso se cruzan ambos.
    """
    from .models import Compras
    clave = _normalizar_texto_temporada(filtros['temporada'])
    if clave not in TEMPORADAS_COMPRAS:
        return None
    if filtros['es_vendedora'] or not filtros['incluye_cd']:
        # La temporada solo existe en las órdenes de compra del centro de
        # distribución: sin ellas el filtro no es aplicable y se informa como
        # tal (temporada_aplicada=False) en vez de fingir que filtró.
        return None
    familia, textos = TEMPORADAS_COMPRAS[clave]
    # La orden se emite hasta un año antes de que llegue la mercadería (la
    # temporada de verano se compra el año anterior), así que la ventana de
    # órdenes se abre 365 días hacia atrás. Los DTE siguen acotados al período:
    # el resultado es "lo recepcionado en el período de esa temporada".
    qs = Compras.objects.filter(
        estado__in=ESTADOS_COMPRA_VIGENTES,
        fecha__range=(filtros['fecha_inicio'] - timedelta(days=365), filtros['fecha_fin']),
    )
    if filtros['proveedor_id']:
        qs = qs.filter(empresa_id=filtros['proveedor_id'])
    condicion = Q(temporada_familia=familia)
    for texto in textos:
        condicion |= Q(temporada__icontains=texto)
    return list(qs.filter(condicion).values_list('id', flat=True))


def _get_dtes_base_compras(filtros, anio=None):
    """
    Queryset base de DTEs del reporte, ya scopeado y saneado:

    - CENTRO_DISTRIBUCION: DTEs tipo_transaccion='COMPRA' cuyo RECEPTOR es una
      de las empresas del usuario (emisor = proveedor externo).
    - VENDEDORA: DTEs tipo_transaccion='TRASPASO' recibidos desde el CD.

    `anio` permite reusar el mismo scoping para el año anterior (comparativas).
    """
    from .models import Dte, Productos_Recepcionados

    if anio is None:
        inicio, fin = filtros['fecha_inicio'], filtros['fecha_fin']
    else:
        inicio, fin = _rango_periodo_compras(anio, filtros['periodo'], filtros['hoy'])

    if filtros['es_vendedora'] and filtros['empresa_receptora_id']:
        qs = Dte.objects.filter(
            tipo_transaccion='TRASPASO',
            receptor_id=filtros['empresa_receptora_id'],
            fecha_emision__range=(inicio, fin),
        )
    else:
        qs = Dte.objects.filter(
            tipo_transaccion='COMPRA',
            receptor_id__in=filtros['empresas_ids'],
            fecha_emision__range=(inicio, fin),
        )

    qs = qs.exclude(descartado=True).exclude(estado_dte__in=ESTADOS_DTE_ANULADOS_COMPRAS)

    if filtros['proveedor_id']:
        qs = qs.filter(emisor_id=filtros['proveedor_id'])

    if filtros['compras_ids'] is not None:
        # Filtro de temporada: solo los DTE que recepcionaron alguna línea de
        # las órdenes de compra de esa temporada.
        qs = qs.filter(id__in=Productos_Recepcionados.objects.filter(
            compra_producto_talla__compra_producto__compras_id__in=filtros['compras_ids'],
            dte__isnull=False,
        ).values('dte_id'))

    return qs.select_related('emisor')


def _get_compras_base(filtros, anio=None):
    """Órdenes de compra (modelo Compras) vigentes del período."""
    from .models import Compras
    if not filtros['incluye_cd']:
        return Compras.objects.none()
    qs = Compras.objects.filter(estado__in=ESTADOS_COMPRA_VIGENTES)
    if filtros['compras_ids'] is not None:
        # Con filtro de temporada los ids ya vienen resueltos (ventana propia:
        # la orden puede ser del año anterior a la recepción).
        qs = qs.filter(id__in=filtros['compras_ids'])
    else:
        if anio is None:
            inicio, fin = filtros['fecha_inicio'], filtros['fecha_fin']
        else:
            inicio, fin = _rango_periodo_compras(anio, filtros['periodo'], filtros['hoy'])
        qs = qs.filter(fecha__range=(inicio, fin))
    if filtros['proveedor_id']:
        qs = qs.filter(empresa_id=filtros['proveedor_id'])
    return qs


def _anotar_saldo_compras(qs):
    """
    Anota `pagado` (suma de Dte_Detalle_Pago, incluye NC y compensaciones como
    medio de pago) y `saldo` = monto_con_iva − pagado.

    Se usa Subquery y no Sum() sobre el join para que `saldo` siga siendo una
    expresión agregable (Django no deja agregar sobre otro agregado) y para no
    multiplicar filas cuando un DTE tiene varios pagos.
    """
    from django.db.models import Value
    from .models import Dte_Detalle_Pago
    dinero = DecimalField(max_digits=16, decimal_places=2)
    pagos = (Dte_Detalle_Pago.objects
             .filter(dte=OuterRef('pk'))
             .values('dte')
             .annotate(total=Sum('monto'))
             .values('total')[:1])
    return qs.annotate(
        pagado=Coalesce(Subquery(pagos, output_field=dinero),
                        Value(Decimal('0'), output_field=dinero)),
    ).annotate(
        saldo=ExpressionWrapper(F('monto_con_iva') - F('pagado'), output_field=dinero),
    )


def _datos_compras_por_proveedor(filtros):
    """
    Agregados por proveedor desde las órdenes de compra, en 3 queries.
    Antes eran ~4 queries POR PROVEEDOR (N+1) más un Empresa.objects.get().
    """
    from .models import Compras_Producto_Talla, Productos_Recepcionados

    compras_qs = _get_compras_base(filtros)
    datos = {}

    def _fila(pid):
        return datos.setdefault(pid, {
            'total_compras': 0, 'pares_comprados': 0, 'pares_recepcionados': 0,
            'inversion': 0.0, 'venta_lista': 0.0, 'nombre': '',
        })

    for row in compras_qs.values('empresa_id', 'empresa__nombre').annotate(n=Count('id')):
        if not row['empresa_id']:
            continue
        fila = _fila(row['empresa_id'])
        fila['total_compras'] = row['n']
        fila['nombre'] = row['empresa__nombre'] or ''

    pedidos = (Compras_Producto_Talla.objects
               .filter(compra_producto__compras__in=compras_qs)
               .values('compra_producto__compras__empresa_id')
               .annotate(unidades=Sum('stock'),
                         costo=Sum(F('compra_producto__costo') * F('stock')),
                         venta=Sum(F('compra_producto__precioSugerido') * F('stock'))))
    for row in pedidos:
        pid = row['compra_producto__compras__empresa_id']
        if not pid:
            continue
        fila = _fila(pid)
        fila['pares_comprados'] = int(row['unidades'] or 0)
        fila['inversion'] = float(row['costo'] or 0)
        fila['venta_lista'] = float(row['venta'] or 0)

    recibidos = (Productos_Recepcionados.objects
                 .filter(compra_producto_talla__compra_producto__compras__in=compras_qs)
                 .values('compra_producto_talla__compra_producto__compras__empresa_id')
                 .annotate(unidades=Sum('stockArribado')))
    for row in recibidos:
        pid = row['compra_producto_talla__compra_producto__compras__empresa_id']
        if not pid:
            continue
        _fila(pid)['pares_recepcionados'] = int(row['unidades'] or 0)

    return datos


@login_required
@require_GET
def api_reporte_compras(request):
    """
    API completa del reporte de compras.

    Modo según la sucursal activa:
    - CENTRO_DISTRIBUCION (EDEL): DTEs tipo_transaccion='COMPRA' a proveedores.
    - VENDEDORA: DTEs tipo_transaccion='TRASPASO' recibidos desde el CD.

    Todo el reporte comparte un único dict de filtros ya scopeado y un único
    ranking de proveedores (ver comentario de cabecera del módulo de compras).
    """
    try:
        filtros = _construir_filtros_compras(request)

        # El ranking se calcula UNA vez y se reparte: top / pareto / cumplimiento.
        ranking = calcular_ranking_proveedores_compras(filtros)

        estado_pagos = calcular_estado_pagos_compras(filtros)
        estado_recepciones = calcular_estado_recepciones_compras(filtros)
        metricas_recepcion = calcular_metricas_recepcion_compras(filtros, estado_recepciones)
        metricas = calcular_metricas_compras(filtros, ranking, estado_pagos, metricas_recepcion)

        evolucion_mensual = calcular_evolucion_mensual_compras(filtros)
        top_proveedores = calcular_top_proveedores_compras(ranking)
        pareto_proveedores = calcular_pareto_proveedores_compras(ranking)
        cumplimiento_proveedores = calcular_cumplimiento_proveedores_compras(ranking)
        roi_temporadas = calcular_roi_temporadas_compras(evolucion_mensual)
        comparativa_anual = calcular_comparativa_anual_compras(filtros)
        recepciones_pendientes = obtener_recepciones_pendientes_compras(filtros)
        vencimientos = calcular_vencimientos_compras(filtros)
        pagos_pendientes = obtener_pagos_pendientes_compras(filtros)

        insights = generar_insights_compras(metricas, top_proveedores, comparativa_anual)
        alertas = generar_alertas_compras(metricas, estado_pagos, cumplimiento_proveedores)

        return JsonResponse({
            'success': True,
            'metricas': metricas,
            'evolucion_mensual': evolucion_mensual,
            'top_proveedores': top_proveedores,
            'pareto_proveedores': pareto_proveedores,
            'cumplimiento_proveedores': cumplimiento_proveedores,
            'roi_temporadas': roi_temporadas,
            'comparativa_anual': comparativa_anual,
            'estado_recepciones': estado_recepciones,
            'metricas_recepcion': metricas_recepcion,
            'recepciones_pendientes': recepciones_pendientes,
            'estado_pagos': estado_pagos,
            'vencimientos': vencimientos,
            'pagos_pendientes': pagos_pendientes,
            'insights': insights,
            'alertas': alertas,
            'filtros_aplicados': {
                'anio': filtros['anio'],
                'periodo': filtros['periodo'],
                'proveedor_id': filtros['proveedor_id'],
                'temporada': filtros['temporada'],
                'temporada_aplicada': filtros['temporada_aplicada'],
                'desde': filtros['fecha_inicio'].isoformat(),
                'hasta': filtros['fecha_fin'].isoformat(),
                'empresas_ids': filtros['empresas_ids'],
                'modo': 'vendedora' if filtros['es_vendedora'] else 'centro_distribucion',
            }
        })

    except Exception as e:
        logger.exception("Error al generar reporte de compras")
        return JsonResponse({
            'success': False,
            'error': f'Error al generar reporte: {str(e)}',
        }, status=500)


def calcular_metricas_compras(filtros, ranking, estado_pagos, metricas_recepcion):
    """
    Métricas de cabecera. Recibe ya calculados el ranking, el estado de pagos y
    las métricas de recepción para no repetir sus queries.
    """
    from .models import Dte_Productos, Productos_Recepcionados

    dtes_qs = _get_dtes_base_compras(filtros)
    es_nc = Q(tipo_documento=TIPO_DOC_NC_COMPRAS)

    agg = dtes_qs.aggregate(
        documentos=Count('id', filter=~es_nc),
        notas_credito=Count('id', filter=es_nc),
        neto=Sum('monto_neto', filter=~es_nc),
        neto_nc=Sum('monto_neto', filter=es_nc),
        con_iva=Sum('monto_con_iva', filter=~es_nc),
        con_iva_nc=Sum('monto_con_iva', filter=es_nc),
        proveedores=Count('emisor_id', distinct=True, filter=~es_nc),
    )
    total_compras = agg['documentos'] or 0
    total_nc = agg['notas_credito'] or 0
    devoluciones_nc = float(agg['neto_nc'] or 0)
    inversion_bruta = float(agg['neto'] or 0)
    inversion_total = inversion_bruta - devoluciones_nc
    monto_con_iva_total = float(agg['con_iva'] or 0) - float(agg['con_iva_nc'] or 0)
    proveedores_activos = agg['proveedores'] or 0

    unidades_compradas = metricas_recepcion['unidades_esperadas']
    unidades_recibidas = metricas_recepcion['unidades_recibidas']
    tasa_recepcion = metricas_recepcion['porcentaje_cumplimiento']

    # Complemento con órdenes de compra (Compras) cuando el flujo no dejó DTE.
    # No aplica con filtro de temporada: ahí la cabecera debe seguir siendo lo
    # FACTURADO del período (si no, elegir una temporada disparaba la inversión
    # al costo de todas las órdenes de esa temporada, incluidas las de otro año).
    if filtros['compras_ids'] is None and (
            not total_compras or not unidades_compradas or not inversion_total):
        pedidas = sum(p['unidades_pedidas'] for p in ranking)
        recibidas = sum(p['unidades_recibidas'] for p in ranking)
        ordenes = sum(p['ordenes_compra'] for p in ranking)
        costo_ordenes = sum(p['costo_ordenes'] for p in ranking)
        if not unidades_compradas and pedidas:
            # Al cambiar la fuente hay que recalcular el cumplimiento con ella:
            # cruzar unidades de las órdenes con el % de los DTE daba 0%.
            unidades_compradas = pedidas
            unidades_recibidas = recibidas
            tasa_recepcion = round(recibidas / pedidas * 100, 1)
        if not total_compras:
            total_compras = ordenes
        if not inversion_total:
            inversion_total = costo_ordenes
            monto_con_iva_total = costo_ordenes
        if not proveedores_activos:
            proveedores_activos = len([p for p in ranking if p['ordenes_compra']])
    costo_promedio = round(inversion_total / unidades_compradas) if unidades_compradas else 0

    # ===== TENDENCIAS vs AÑO ANTERIOR (mismo scoping y mismo período) =====
    dtes_anterior = _get_dtes_base_compras(filtros, anio=filtros['anio'] - 1)
    agg_ant = dtes_anterior.aggregate(
        documentos=Count('id', filter=~es_nc),
        neto=Sum('monto_neto', filter=~es_nc),
        neto_nc=Sum('monto_neto', filter=es_nc),
        proveedores=Count('emisor_id', distinct=True, filter=~es_nc),
    )
    total_anterior = agg_ant['documentos'] or 0
    inversion_anterior = float(agg_ant['neto'] or 0) - float(agg_ant['neto_nc'] or 0)
    proveedores_anterior = agg_ant['proveedores'] or 0
    # Unidades del año anterior por la misma fuente que las actuales
    # (Productos_Recepcionados), si no la tendencia compara peras con manzanas.
    unidades_anterior = int(Productos_Recepcionados.objects.filter(
        dte__in=dtes_anterior.values('id')
    ).exclude(
        producto_talla__producto__excluir_de_analitica=True
    ).aggregate(total=Sum('stockArribado'))['total'] or 0)
    if not unidades_anterior:
        unidades_anterior = int(Dte_Productos.objects.filter(
            dte__in=dtes_anterior.values('id'),
            productoTalla__producto__excluir_de_analitica=False,
        ).aggregate(total=Sum('stock'))['total'] or 0)

    tendencia_compras = calcular_porcentaje_cambio(total_compras, total_anterior)
    tendencia_inversion = calcular_porcentaje_cambio(inversion_total, inversion_anterior)
    tendencia_unidades = calcular_porcentaje_cambio(unidades_compradas, unidades_anterior)
    tendencia_proveedores = calcular_porcentaje_cambio(proveedores_activos, proveedores_anterior)

    # ── Markup teórico de lista ─────────────────────────────────────────────
    # Antes el "valor de venta esperado" era Sum(monto_con_iva) de la FACTURA DE
    # COMPRA y el "margen"/"ROI" su diferencia con el neto: o sea, el IVA
    # soportado (siempre 19,0%). Ahora se calcula contra el precio sugerido de
    # la orden de compra y se rotula TEÓRICO: aún no cruza con la venta real.
    costo_lista = sum(p['costo_ordenes'] for p in ranking)
    venta_lista = sum(p['venta_ordenes'] for p in ranking)
    margen_teorico = venta_lista - costo_lista
    markup_teorico = round((margen_teorico / costo_lista) * 100, 1) if costo_lista > 0 else 0

    return {
        'total_compras': total_compras,
        'total_notas_credito': total_nc,
        'devoluciones_nc': devoluciones_nc,
        'inversion_bruta': inversion_bruta,
        'proveedores_activos': proveedores_activos,
        'unidades_esperadas': unidades_compradas,
        'unidades_recepcionadas': unidades_recibidas,
        'cumplimiento_general': tasa_recepcion,
        'inversion_total': inversion_total,
        'iva_soportado': monto_con_iva_total - inversion_total,
        # El trío teórico se calcula sobre las ÓRDENES de compra del período,
        # no sobre los DTE: se expone su costo base para comparar peras con
        # peras (inversion_total viene de los DTE recibidos).
        'costo_ordenes_compra': costo_lista,
        'valor_venta_esperado': venta_lista,
        'margen_bruto_esperado': margen_teorico,
        'roi_promedio': markup_teorico,
        'markup_teorico_calculable': costo_lista > 0,
        'costo_promedio_unidad': float(costo_promedio),
        'pendiente_pago': estado_pagos['pendientes'] + estado_pagos['vencidos'],
        'pct_facturas_pagadas': estado_pagos['pct_facturas_pagadas'],
        'tendencia_compras': tendencia_compras,
        'tendencia_inversion': tendencia_inversion,
        'tendencia_unidades': tendencia_unidades,
        'tendencia_proveedores': tendencia_proveedores,
        'tendencia_cumplimiento': 0,
        'tendencia_roi': 0
    }


def calcular_porcentaje_cambio(actual, anterior):
    """Calcula el porcentaje de cambio entre dos valores"""
    if anterior == 0:
        return 100 if actual > 0 else 0
    return round(((actual - anterior) / anterior) * 100, 1)


def calcular_evolucion_mensual_compras(filtros):
    """
    Evolución mensual en 4 queries agrupadas por mes.
    Antes: 12 iteraciones × (count + sum + lista de ids + unidades + 2 queries
    de Compras) sobre el mismo queryset.
    """
    from django.db.models.functions import ExtractMonth
    from .models import Compras_Producto_Talla, Dte_Productos, Productos_Recepcionados

    dtes_qs = _get_dtes_base_compras(filtros)
    es_nc = Q(tipo_documento=TIPO_DOC_NC_COMPRAS)

    por_mes = {m: {'mes': m, 'total_compras': 0, 'inversion': 0.0, 'unidades': 0}
               for m in range(1, 13)}

    for row in (dtes_qs.annotate(m=ExtractMonth('fecha_emision'))
                .values('m')
                .annotate(documentos=Count('id', filter=~es_nc),
                          neto=Sum('monto_neto', filter=~es_nc),
                          neto_nc=Sum('monto_neto', filter=es_nc))):
        fila = por_mes.get(row['m'])
        if not fila:
            continue
        fila['total_compras'] = row['documentos'] or 0
        fila['inversion'] = float(row['neto'] or 0) - float(row['neto_nc'] or 0)

    for row in (Dte_Productos.objects
                .filter(dte__in=dtes_qs.values('id'), productoTalla__producto__excluir_de_analitica=False)
                .annotate(m=ExtractMonth('dte__fecha_emision'))
                .values('m').annotate(u=Sum('stock'))):
        fila = por_mes.get(row['m'])
        if fila:
            fila['unidades'] = int(row['u'] or 0)

    # El flujo de compras no escribe Dte_Productos: sus unidades viven en
    # Productos_Recepcionados.
    for row in (Productos_Recepcionados.objects
                .filter(dte__in=dtes_qs.values('id'))
                .exclude(producto_talla__producto__excluir_de_analitica=True)
                .annotate(m=ExtractMonth('dte__fecha_emision'))
                .values('m').annotate(u=Sum('stockArribado'))):
        fila = por_mes.get(row['m'])
        if fila and not fila['unidades']:
            fila['unidades'] = int(row['u'] or 0)

    # Complemento con órdenes de compra SOLO si el período no tiene ningún DTE:
    # mes a mes mezclaba órdenes con facturas y la suma de la serie ya no
    # cuadraba con la inversión de la cabecera.
    if filtros['compras_ids'] is None and not any(
            fila['total_compras'] for fila in por_mes.values()):
        for row in (Compras_Producto_Talla.objects
                    .filter(compra_producto__compras__in=_get_compras_base(filtros))
                    .annotate(m=ExtractMonth('compra_producto__compras__fecha'))
                    .values('m')
                    .annotate(u=Sum('stock'),
                              costo=Sum(F('compra_producto__costo') * F('stock')),
                              ordenes=Count('compra_producto__compras', distinct=True))):
            fila = por_mes.get(row['m'])
            if not fila:
                continue
            fila['total_compras'] = row['ordenes'] or 0
            fila['inversion'] = float(row['costo'] or 0)
            fila['unidades'] = int(row['u'] or 0)

    return [por_mes[m] for m in range(1, 13)]


def calcular_ranking_proveedores_compras(filtros):
    """
    Ranking ÚNICO de proveedores (emisores), en ~8 queries agrupadas.

    Antes: un loop N+1 con ~6 queries + un Empresa.objects.get() por proveedor,
    y encima pareto y cumplimiento volvían a invocar el cálculo completo. Con
    70 proveedores eso eran miles de queries por request.

    Métricas por proveedor:
      · inversion  = neto de facturas − neto de notas de crédito
      · cumplimiento = unidades recepcionadas / unidades pedidas (ENTREGAS)
      · pct_facturas_pagadas = facturas con saldo 0 / facturas (tesorería)
      · markup = precio sugerido vs costo de la orden (teórico)
    """
    from .models import Dte_Productos, Productos_Recepcionados

    dtes_qs = _get_dtes_base_compras(filtros)
    es_nc = Q(tipo_documento=TIPO_DOC_NC_COMPRAS)

    proveedores = {}

    def _fila(pid, nombre='', rut=''):
        if pid not in proveedores:
            proveedores[pid] = {
                'id': pid, 'nombre': nombre or 'Sin nombre', 'rut': rut or '',
                'total_compras': 0, 'notas_credito': 0, 'inversion': 0.0,
                'devoluciones': 0.0, 'unidades': 0,
                'unidades_pedidas': 0, 'unidades_recibidas': 0,
                'ordenes_compra': 0, 'costo_ordenes': 0.0, 'venta_ordenes': 0.0,
                'facturas_pagadas': 0, 'saldo_pendiente': 0.0,
            }
        elif nombre and proveedores[pid]['nombre'] == 'Sin nombre':
            proveedores[pid]['nombre'] = nombre
        return proveedores[pid]

    for row in (dtes_qs.values('emisor_id', 'emisor__nombre', 'emisor__rut')
                .annotate(documentos=Count('id', filter=~es_nc),
                          notas=Count('id', filter=es_nc),
                          neto=Sum('monto_neto', filter=~es_nc),
                          neto_nc=Sum('monto_neto', filter=es_nc))):
        if not row['emisor_id']:
            continue
        fila = _fila(row['emisor_id'], row['emisor__nombre'], row['emisor__rut'])
        fila['total_compras'] = row['documentos'] or 0
        fila['notas_credito'] = row['notas'] or 0
        fila['devoluciones'] = float(row['neto_nc'] or 0)
        fila['inversion'] = float(row['neto'] or 0) - fila['devoluciones']

    # Facturas totalmente pagadas y deuda viva, por saldo real (no estado_pago).
    saldos = _anotar_saldo_compras(dtes_qs.exclude(tipo_documento=TIPO_DOC_NC_COMPRAS))
    for row in (saldos.values('emisor_id')
                .annotate(pagadas=Count('id', filter=Q(saldo__lte=0)),
                          deuda=Sum('saldo', filter=Q(saldo__gt=0)))):
        if not row['emisor_id']:
            continue
        fila = _fila(row['emisor_id'])
        fila['facturas_pagadas'] = row['pagadas'] or 0
        fila['saldo_pendiente'] = float(row['deuda'] or 0)

    for row in (Dte_Productos.objects
                .filter(dte__in=dtes_qs.values('id'), productoTalla__producto__excluir_de_analitica=False)
                .values('dte__emisor_id').annotate(u=Sum('stock'))):
        if row['dte__emisor_id']:
            _fila(row['dte__emisor_id'])['unidades'] = int(row['u'] or 0)

    for row in (Productos_Recepcionados.objects
                .filter(dte__in=dtes_qs.values('id'))
                .exclude(producto_talla__producto__excluir_de_analitica=True)
                .values('dte__emisor_id').annotate(u=Sum('stockArribado'))):
        pid = row['dte__emisor_id']
        if pid and not _fila(pid)['unidades']:
            _fila(pid)['unidades'] = int(row['u'] or 0)

    # Órdenes de compra: entregas (pedidas vs recepcionadas) y markup teórico.
    for pid, datos in _datos_compras_por_proveedor(filtros).items():
        fila = _fila(pid, datos['nombre'])
        fila['ordenes_compra'] = datos['total_compras']
        fila['unidades_pedidas'] = datos['pares_comprados']
        fila['unidades_recibidas'] = datos['pares_recepcionados']
        fila['costo_ordenes'] = datos['inversion']
        fila['venta_ordenes'] = datos['venta_lista']

    resultado = []
    for fila in proveedores.values():
        if not fila['total_compras'] and not fila['ordenes_compra']:
            continue
        if not fila['total_compras'] and fila['ordenes_compra']:
            # Proveedor con orden de compra pero sin DTE recibido todavía.
            fila['total_compras'] = fila['ordenes_compra']
            fila['inversion'] = fila['inversion'] or fila['costo_ordenes']
            fila['unidades'] = fila['unidades'] or fila['unidades_pedidas']

        pedidas = fila['unidades_pedidas']
        fila['cumplimiento'] = round(fila['unidades_recibidas'] / pedidas * 100, 1) if pedidas else 0
        fila['cumplimiento_calculable'] = pedidas > 0
        docs = fila['total_compras']
        fila['pct_facturas_pagadas'] = round(fila['facturas_pagadas'] / docs * 100, 1) if docs else 0
        costo = fila['costo_ordenes']
        fila['markup'] = round((fila['venta_ordenes'] - costo) / costo * 100, 1) if costo > 0 else 0
        resultado.append(fila)

    inversion_general = sum(f['inversion'] for f in resultado)
    for fila in resultado:
        fila['participacion'] = round(fila['inversion'] / inversion_general * 100, 1) if inversion_general else 0

    resultado.sort(key=lambda x: x['inversion'], reverse=True)
    return resultado


def calcular_top_proveedores_compras(ranking):
    """Top 15 por inversión (el ranking ya viene ordenado y calculado)."""
    return ranking[:15]


def calcular_pareto_proveedores_compras(ranking):
    """Pareto: top 10 con el acumulado de participación sobre el total."""
    total = sum(p['inversion'] for p in ranking)
    acumulado = 0.0
    resultado = []
    for proveedor in ranking[:10]:
        acumulado += proveedor['inversion']
        fila = dict(proveedor)
        fila['acumulado_pct'] = round(acumulado / total * 100, 1) if total else 0
        resultado.append(fila)
    return resultado


def calcular_cumplimiento_proveedores_compras(ranking):
    """
    Ranking de cumplimiento de ENTREGAS (recepcionado/pedido).
    Antes ordenaba por % de facturas pagadas, que es tesorería, no logística.
    Solo entran proveedores con órdenes de compra: sin unidades pedidas no hay
    cumplimiento que medir (antes salían todos con 0% o 100% inventado).
    """
    medibles = [p for p in ranking if p['cumplimiento_calculable']]
    # Peores primero: el gráfico es para detectar entregas incompletas, no para
    # premiar. Con orden descendente los proveedores problemáticos quedaban
    # fuera del top 10 y la alerta de "cumplimiento bajo" nunca se disparaba.
    medibles.sort(key=lambda x: (x['cumplimiento'], -x['unidades_pedidas']))
    return medibles[:10]


def calcular_roi_temporadas_compras(evolucion_mensual):
    """
    Inversión neta por trimestre. Se deriva de la evolución mensual ya
    calculada: antes lanzaba 8 queries propias (2 por trimestre) sobre
    exactamente los mismos datos.
    """
    por_mes = {fila['mes']: fila['inversion'] for fila in evolucion_mensual}

    trimestres = [
        ('Q1 (Ene-Mar)', [1, 2, 3]),
        ('Q2 (Abr-Jun)', [4, 5, 6]),
        ('Q3 (Jul-Sep)', [7, 8, 9]),
        ('Q4 (Oct-Dic)', [10, 11, 12]),
    ]
    resultado = []
    for nombre, meses in trimestres:
        inversion = sum(por_mes.get(m, 0.0) for m in meses)
        if inversion:
            resultado.append({'temporada': nombre, 'inversion': inversion})
    return resultado


def calcular_comparativa_anual_compras(filtros):
    """
    Inversión mensual de este año vs el anterior: 2 queries agrupadas por mes
    (antes 24 queries, una por mes y año).
    """
    from django.db.models.functions import ExtractMonth

    es_nc = Q(tipo_documento=TIPO_DOC_NC_COMPRAS)

    def _serie(anio):
        qs = _get_dtes_base_compras(filtros, anio=anio)
        por_mes = {}
        for row in (qs.annotate(m=ExtractMonth('fecha_emision'))
                    .values('m')
                    .annotate(neto=Sum('monto_neto', filter=~es_nc),
                              neto_nc=Sum('monto_neto', filter=es_nc))):
            por_mes[row['m']] = float(row['neto'] or 0) - float(row['neto_nc'] or 0)
        return [por_mes.get(m, 0.0) for m in range(1, 13)]

    return {
        'actual': _serie(filtros['anio']),
        'anterior': _serie(filtros['anio'] - 1),
    }


def calcular_estado_recepciones_compras(filtros):
    """Estado de recepción de los DTE del período (1 query agrupada)."""
    dtes_qs = _get_dtes_base_compras(filtros).exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)

    conteo = {row['estado_dte']: row['n']
              for row in dtes_qs.values('estado_dte').annotate(n=Count('id'))}

    return {
        'recibido_ok': conteo.get('ACEPTADO', 0) + conteo.get('RECEPCIONADO_COMPLETO', 0)
                       + conteo.get('RECEPCIONADO_SOBRANTE', 0),
        'parcial': conteo.get('RECEPCIONADO_PARCIAL', 0),
        'pendiente': conteo.get('EMITIDO', 0),
        'con_problemas': conteo.get('EN_REGULARIZACION', 0),
    }


def calcular_metricas_recepcion_compras(filtros, estado_recepciones=None):
    """
    Unidades esperadas vs recibidas del período. Fuente real:
    Productos_Recepcionados (el flujo de compras no escribe Dte_Productos).
    """
    from django.db.models import Case, When
    from .models import Dte_Productos, Productos_Recepcionados

    dtes_qs = _get_dtes_base_compras(filtros).exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)

    # Esperadas y recibidas se leen de la MISMA fila de recepción: mezclar el
    # total de Dte_Productos (que solo existe en algunos DTE) con el arribo de
    # todos daba cumplimientos de 0% o de más de 100% sin sentido. Cuando la
    # línea no declaró cantidad esperada se asume la recibida (no se inventan
    # faltantes).
    esperada_fila = Case(
        When(cantidad_esperada__gt=0, then=F('cantidad_esperada')),
        default=F('stockArribado'),
        output_field=IntegerField(),
    )
    recepciones = (Productos_Recepcionados.objects
                   .filter(dte__in=dtes_qs.values('id'))
                   .exclude(producto_talla__producto__excluir_de_analitica=True)
                   .aggregate(recibidas=Sum('stockArribado'),
                              esperadas=Sum(esperada_fila),
                              faltantes=Sum('cantidad_faltante')))
    unidades_recibidas = int(recepciones['recibidas'] or 0)
    unidades_esperadas = int(recepciones['esperadas'] or 0)
    faltantes = int(recepciones['faltantes'] or 0)

    if not unidades_esperadas:
        # DTE con líneas pero sin recepción registrada (traspasos en tránsito).
        unidades_esperadas = int(Dte_Productos.objects.filter(
            dte__in=dtes_qs.values('id'),
            productoTalla__producto__excluir_de_analitica=False,
        ).aggregate(total=Sum('stock'))['total'] or 0)

    if unidades_esperadas > unidades_recibidas and not faltantes:
        faltantes = unidades_esperadas - unidades_recibidas

    if estado_recepciones is not None:
        con_problemas = estado_recepciones['con_problemas'] + estado_recepciones['parcial']
    else:
        con_problemas = dtes_qs.filter(
            estado_dte__in=['EN_REGULARIZACION', 'RECEPCIONADO_PARCIAL']
        ).count()

    porcentaje = round(unidades_recibidas / unidades_esperadas * 100, 1) if unidades_esperadas else 0

    return {
        'unidades_esperadas': unidades_esperadas,
        'unidades_recibidas': unidades_recibidas,
        'con_problemas': con_problemas,
        'faltantes': max(0, faltantes),
        'porcentaje_cumplimiento': porcentaje,
    }


def obtener_recepciones_pendientes_compras(filtros):
    """
    DTEs pendientes de recepción o con problemas, con las unidades REALES.
    Antes las parciales usaban `int(unidades * 0.5)  # Estimación`: un número
    inventado que además alimentaba la tabla de la pestaña Recepciones.
    """
    from .models import Dte_Productos, Productos_Recepcionados

    dtes = list(_get_dtes_base_compras(filtros).filter(
        estado_dte__in=['EMITIDO', 'RECEPCIONADO_PARCIAL', 'EN_REGULARIZACION']
    ).order_by('-fecha_emision')[:10])
    if not dtes:
        return []

    ids = [d.id for d in dtes]
    esperadas = {row['dte_id']: int(row['u'] or 0) for row in
                 Dte_Productos.objects.filter(dte_id__in=ids)
                 .values('dte_id').annotate(u=Sum('stock'))}
    recibidas = {}
    esperadas_recepcion = {}
    for row in (Productos_Recepcionados.objects.filter(dte_id__in=ids)
                .values('dte_id')
                .annotate(u=Sum('stockArribado'), e=Sum('cantidad_esperada'))):
        recibidas[row['dte_id']] = int(row['u'] or 0)
        esperadas_recepcion[row['dte_id']] = int(row['e'] or 0)

    estados = {
        'EMITIDO': 'PENDIENTE',
        'RECEPCIONADO_PARCIAL': 'PARCIAL',
        'EN_REGULARIZACION': 'PROBLEMA',
    }

    resultado = []
    for dte in dtes:
        recibido = recibidas.get(dte.id, 0)
        esperado = esperadas.get(dte.id) or esperadas_recepcion.get(dte.id) or dte.unidades_productos or recibido
        resultado.append({
            'id': dte.id,
            'numero': dte.numero_documento,
            'proveedor': dte.emisor.nombre if dte.emisor else 'N/A',
            'fecha': dte.fecha_emision.strftime('%d/%m/%Y'),
            'esperadas': esperado,
            'recibidas': recibido,
            'estado': estados.get(dte.estado_dte, 'PROBLEMA'),
        })
    return resultado


def calcular_estado_pagos_compras(filtros):
    """
    Estado de pagos por SALDO REAL (monto_con_iva − pagos registrados).

    El campo estado_pago no es utilizable: `_recalcular_estado_pago` escribe
    'PAGADO'/'Parcial'/'Pendiente' y la migración dejó 'Pagado'/'Abonado', así
    que una factura con abono parcial desaparecía del reporte y a las
    pendientes se les imputaba el monto COMPLETO sin restar los abonos.
    Las notas de crédito no son deuda: se excluyen (ya restaron vía pago).
    """
    hoy = filtros['hoy']
    saldos = _anotar_saldo_compras(
        _get_dtes_base_compras(filtros).exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)
    )

    abierto = Q(saldo__gt=0)
    totales = saldos.aggregate(
        documentos=Count('id'),
        pagadas=Count('id', filter=Q(saldo__lte=0)),
        con_iva=Sum('monto_con_iva'),
        pendientes=Sum('saldo', filter=abierto & (
            Q(fecha_vencimiento__gte=hoy) | Q(fecha_vencimiento__isnull=True))),
        vencidos=Sum('saldo', filter=abierto & Q(fecha_vencimiento__lt=hoy)),
        vencen_semana=Sum('saldo', filter=abierto & Q(
            fecha_vencimiento__range=(hoy, hoy + timedelta(days=7)))),
    )
    pendientes = totales['pendientes'] or 0
    vencidos = totales['vencidos'] or 0
    vencen_semana = totales['vencen_semana'] or 0

    con_iva = float(totales['con_iva'] or 0)
    documentos = totales['documentos'] or 0
    pagados = con_iva - float(pendientes) - float(vencidos)

    return {
        'pagados': max(0.0, pagados),
        'pendientes': float(pendientes),
        'vencidos': float(vencidos),
        'vencen_semana': float(vencen_semana),
        'documentos': documentos,
        'facturas_pagadas': totales['pagadas'] or 0,
        'pct_facturas_pagadas': round((totales['pagadas'] or 0) / documentos * 100, 1) if documentos else 0,
    }


def calcular_vencimientos_compras(filtros):
    """
    Vencimientos por tramo, en 1 query y sobre el saldo pendiente real.
    Antes eran 6 queries SIN scoping por empresa y sin filtro de año: sumaban
    el monto completo de toda factura marcada 'Pendiente' del holding.
    """
    hoy = filtros['hoy']
    saldos = _anotar_saldo_compras(
        _get_dtes_base_compras(filtros).exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)
    ).filter(saldo__gt=0)

    tramos = [
        {'periodo': 'Vencidos', 'monto': 0.0, 'dias_restantes': -1},
        {'periodo': 'Hoy', 'monto': 0.0, 'dias_restantes': 0},
        {'periodo': '1-7 días', 'monto': 0.0, 'dias_restantes': 3},
        {'periodo': '8-15 días', 'monto': 0.0, 'dias_restantes': 11},
        {'periodo': '16-30 días', 'monto': 0.0, 'dias_restantes': 23},
        {'periodo': '+30 días', 'monto': 0.0, 'dias_restantes': 45},
    ]

    for fila in saldos.values('fecha_vencimiento', 'saldo'):
        monto = float(fila['saldo'] or 0)
        vence = fila['fecha_vencimiento']
        if not vence:
            continue
        dias = (vence - hoy).days
        if dias < 0:
            tramos[0]['monto'] += monto
        elif dias == 0:
            tramos[1]['monto'] += monto
        elif dias <= 7:
            tramos[2]['monto'] += monto
        elif dias <= 15:
            tramos[3]['monto'] += monto
        elif dias <= 30:
            tramos[4]['monto'] += monto
        else:
            tramos[5]['monto'] += monto

    return tramos


def obtener_pagos_pendientes_compras(filtros):
    """
    Facturas con saldo pendiente (scopeadas). El monto mostrado es el SALDO,
    no el total del documento: antes una factura abonada figuraba por su valor
    completo y las 'Parcial'/'Abonado' ni siquiera aparecían.
    """
    hoy = filtros['hoy']
    saldos = _anotar_saldo_compras(
        _get_dtes_base_compras(filtros).exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)
    ).filter(saldo__gt=0).order_by('fecha_vencimiento')[:20]

    resultado = []
    for dte in saldos:
        dias = (dte.fecha_vencimiento - hoy).days if dte.fecha_vencimiento else 0
        resultado.append({
            'id': dte.id,
            'numero_dte': dte.numero_documento,
            'proveedor': dte.emisor.nombre if dte.emisor else 'N/A',
            'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y'),
            'fecha_vencimiento': dte.fecha_vencimiento.strftime('%d/%m/%Y') if dte.fecha_vencimiento else 'N/A',
            'monto': float(dte.monto_con_iva),
            'pagado': float(dte.pagado),
            'saldo': float(dte.saldo),
            'dias_restantes': dias,
        })
    return resultado


def generar_insights_compras(metricas, top_proveedores, comparativa):
    """Genera insights estratégicos basados en los datos"""
    insights = []

    # Insight sobre markup teórico (precio sugerido vs costo de la OC).
    # Solo se opina si el dato existe: antes se opinaba sobre el IVA.
    if metricas.get('markup_teorico_calculable'):
        if metricas['roi_promedio'] >= 100:
            insights.append({
                'tipo': 'success',
                'titulo': 'Markup teórico saludable',
                'descripcion': (
                    f"El precio sugerido de las órdenes duplica el costo "
                    f"({metricas['roi_promedio']}% sobre costo). Falta contrastarlo "
                    f"con la venta real y el descuento aplicado."
                )
            })
        elif metricas['roi_promedio'] < 60:
            insights.append({
                'tipo': 'warning',
                'titulo': 'Markup teórico bajo',
                'descripcion': (
                    f"El precio sugerido deja solo {metricas['roi_promedio']}% sobre el costo. "
                    f"Con descuentos de temporada el margen real puede desaparecer."
                )
            })

    # Insight sobre cumplimiento de ENTREGAS (unidades recibidas vs pedidas).
    # Antes este bloque hablaba de "cumplimiento de proveedores" con el % de
    # facturas pagadas: opinaba sobre tesorería creyendo hablar de logística.
    if metricas['unidades_esperadas']:
        cumplimiento = metricas['cumplimiento_general']
        if cumplimiento >= 95:
            insights.append({
                'tipo': 'success',
                'titulo': 'Entregas completas',
                'descripcion': (
                    f"Se recepcionó el {cumplimiento}% de las unidades esperadas "
                    f"({metricas['unidades_recepcionadas']:,} de {metricas['unidades_esperadas']:,})."
                )
            })
        elif cumplimiento < 80:
            insights.append({
                'tipo': 'warning',
                'titulo': 'Entregas incompletas',
                'descripcion': (
                    f"Solo llegó el {cumplimiento}% de las unidades esperadas. "
                    f"Revise faltantes y regularizaciones con los proveedores."
                )
            })

    # Insight sobre concentración de proveedores
    if top_proveedores and len(top_proveedores) >= 3:
        top3_participacion = sum(p['participacion'] for p in top_proveedores[:3])
        if top3_participacion > 70:
            insights.append({
                'tipo': 'info',
                'titulo': 'Alta concentración de proveedores',
                'descripcion': f"El {top3_participacion:.0f}% de las compras se concentra en 3 proveedores. Considere diversificar."
            })

    # Insight sobre devoluciones al proveedor (notas de crédito de compra)
    if metricas.get('devoluciones_nc') and metricas.get('inversion_bruta'):
        peso_nc = metricas['devoluciones_nc'] / metricas['inversion_bruta'] * 100
        if peso_nc >= 5:
            insights.append({
                'tipo': 'warning',
                'titulo': 'Devoluciones a proveedor relevantes',
                'descripcion': (
                    f"Las notas de crédito de compra equivalen al {peso_nc:.1f}% de lo facturado "
                    f"(${metricas['devoluciones_nc']:,.0f}). Revise calidad y diferencias de facturación."
                )
            })

    # Insight sobre tendencia
    if metricas['tendencia_inversion'] > 20:
        insights.append({
            'tipo': 'info',
            'titulo': 'Inversión en crecimiento',
            'descripcion': f"La inversión ha aumentado un {metricas['tendencia_inversion']}% respecto al año anterior."
        })
    elif metricas['tendencia_inversion'] < -20:
        insights.append({
            'tipo': 'warning',
            'titulo': 'Reducción de inversión',
            'descripcion': f"La inversión ha disminuido un {abs(metricas['tendencia_inversion'])}% respecto al año anterior."
        })

    return insights


def generar_alertas_compras(metricas, estado_pagos, cumplimiento_proveedores):
    """Genera alertas importantes basadas en los datos"""
    alertas = []

    # Alerta de pagos vencidos
    if estado_pagos['vencidos'] > 0:
        alertas.append({
            'nivel': 'error',
            'titulo': 'Pagos vencidos',
            'descripcion': f"Tiene ${estado_pagos['vencidos']:,.0f} en saldo vencido. Gestione con urgencia."
        })

    # Alerta de vencimientos próximos
    if estado_pagos['vencen_semana'] > 0:
        alertas.append({
            'nivel': 'warning',
            'titulo': 'Vencimientos próximos',
            'descripcion': f"${estado_pagos['vencen_semana']:,.0f} vencen esta semana. Planifique los pagos."
        })

    # Alerta de cumplimiento bajo (entregas, solo proveedores medibles)
    if cumplimiento_proveedores:
        proveedores_bajos = [p for p in cumplimiento_proveedores if p['cumplimiento'] < 70]
        if proveedores_bajos:
            alertas.append({
                'nivel': 'warning',
                'titulo': f'{len(proveedores_bajos)} proveedores con entregas incompletas',
                'descripcion': 'Recepcionaron menos del 70% de las unidades pedidas. Revise las condiciones comerciales.'
            })

    # Alerta de pendiente de pago alto
    if metricas['inversion_total'] and metricas['pendiente_pago'] > metricas['inversion_total'] * 0.5:
        alertas.append({
            'nivel': 'warning',
            'titulo': 'Alto nivel de deuda',
            'descripcion': 'El saldo pendiente de pago representa más del 50% de la inversión del período.'
        })

    return alertas


@login_required
@require_GET
def exportar_reporte_compras_excel(request):
    """Exporta el reporte de compras a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Mismos filtros y mismo scoping que la vista web: el Excel no puede
        # exportar un universo distinto al que el usuario ve en pantalla.
        filtros = _construir_filtros_compras(request)
        anio = filtros['anio']

        ranking = calcular_ranking_proveedores_compras(filtros)
        estado_pagos = calcular_estado_pagos_compras(filtros)
        metricas_recepcion = calcular_metricas_recepcion_compras(filtros)
        metricas = calcular_metricas_compras(filtros, ranking, estado_pagos, metricas_recepcion)
        top_proveedores = calcular_top_proveedores_compras(ranking)
        evolucion = calcular_evolucion_mensual_compras(filtros)

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte Compras {anio}"

        # Estilos
        header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = f"REPORTE DE COMPRAS - AÑO {anio}"
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sección KPIs
        ws['A3'] = "INDICADORES PRINCIPALES"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = subheader_fill
        ws.merge_cells('A3:H3')
        
        kpis = [
            ('Total Compras', metricas['total_compras']),
            ('Inversión Total', f"${metricas['inversion_total']:,.0f}"),
            ('Unidades Compradas', metricas['unidades_esperadas']),
            ('Cumplimiento entregas', f"{metricas['cumplimiento_general']}%"),
            ('Markup teórico', f"{metricas['roi_promedio']}%"),
            ('Proveedores Activos', metricas['proveedores_activos']),
            ('Pendiente Pago (saldo)', f"${metricas['pendiente_pago']:,.0f}"),
            ('Devoluciones (NC)', f"${metricas['devoluciones_nc']:,.0f}"),
            ('% facturas pagadas', f"{metricas['pct_facturas_pagadas']}%"),
        ]
        
        fila = 4
        for idx, (nombre, valor) in enumerate(kpis):
            col = (idx % 4) * 2 + 1
            if idx > 0 and idx % 4 == 0:
                fila += 1
            ws.cell(row=fila, column=col, value=nombre).font = Font(bold=True)
            ws.cell(row=fila, column=col + 1, value=valor)
        
        fila += 3
        
        # Sección Proveedores
        ws.cell(row=fila, column=1, value="TOP PROVEEDORES")
        ws.cell(row=fila, column=1).font = Font(bold=True, size=12)
        ws.cell(row=fila, column=1).fill = subheader_fill
        ws.merge_cells(f'A{fila}:H{fila}')
        
        fila += 1
        headers_prov = ['Proveedor', 'Compras', 'Inversión', 'Unidades',
                        'Entregas (recep/ped)', 'Markup teórico', 'Participación']
        for idx, header in enumerate(headers_prov, start=1):
            cell = ws.cell(row=fila, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        fila += 1
        for proveedor in top_proveedores:
            ws.cell(row=fila, column=1, value=proveedor['nombre']).border = border
            ws.cell(row=fila, column=2, value=proveedor['total_compras']).border = border
            ws.cell(row=fila, column=3, value=f"${proveedor['inversion']:,.0f}").border = border
            ws.cell(row=fila, column=4, value=proveedor['unidades']).border = border
            ws.cell(row=fila, column=5,
                    value=f"{proveedor['cumplimiento']}%" if proveedor['cumplimiento_calculable'] else 's/d').border = border
            ws.cell(row=fila, column=6,
                    value=f"{proveedor['markup']}%" if proveedor['markup'] else 's/d').border = border
            ws.cell(row=fila, column=7, value=f"{proveedor['participacion']}%").border = border
            fila += 1
        
        fila += 2
        
        # Sección Evolución Mensual
        ws.cell(row=fila, column=1, value="EVOLUCIÓN MENSUAL")
        ws.cell(row=fila, column=1).font = Font(bold=True, size=12)
        ws.cell(row=fila, column=1).fill = subheader_fill
        ws.merge_cells(f'A{fila}:H{fila}')
        
        fila += 1
        headers_evol = ['Mes', 'Compras', 'Inversión', 'Unidades']
        for idx, header in enumerate(headers_evol, start=1):
            cell = ws.cell(row=fila, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
        fila += 1
        for item in evolucion:
            ws.cell(row=fila, column=1, value=meses_nombres[item['mes'] - 1]).border = border
            ws.cell(row=fila, column=2, value=item['total_compras']).border = border
            ws.cell(row=fila, column=3, value=f"${item['inversion']:,.0f}").border = border
            ws.cell(row=fila, column=4, value=item['unidades']).border = border
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_compras_{anio}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Error al exportar reporte de compras")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}',
        }, status=500)


# ========== API RENDIMIENTO DE COMPRAS: ROTACIÓN REAL + TRAZABILIDAD ==========
# Estrategia: Reconstruir ciclo de vida COMPLETO desde Movimientos_Producto
#
# ENTRADA (comprado):   concepto RECEPCION_COMPRA + INGRESO_INICIAL + INGRESO_MANUAL
# DESPACHO (sucursales): concepto TRASPASO_SALIDA (saliente real) + DTEs TRASPASO
# VENTA (al cliente):   concepto VENTA_PUBLICO/MAYORISTA + Ticket_Productos
# OTROS EGRESOS:        ajustes a la baja, pérdidas, devoluciones a proveedor, cambios
#
# Esto cubre TANTO datos migrados de Laravel como datos nuevos del sistema.
# Laravel no tenía módulo de Compras, todo se registraba como movimientos.

# INGRESO_MANUAL cuenta como entrada: el alta manual de catálogo deja kardex
# (igual que FIFO/trazabilidad, que ya lo tratan como ingreso). Antes quedaba
# fuera del embudo y distorsionaba rotación/inversión de los SKU creados a mano.
CONCEPTOS_ENTRADA = ['RECEPCION_COMPRA', 'INGRESO_INICIAL', 'INGRESO_MANUAL']
# Despacho SALIENTE a sucursales = solo TRASPASO_SALIDA (egreso, cantidad
# negativa → se normaliza con abs() al consumir). TRASPASO_SUCURSAL NO va aquí:
# en datos migrados es la pata de ENTRADA a la tienda (tipo INGRESO), así que
# sumarlo mezclaba signos y el "despachado" salía negativo.
CONCEPTOS_DESPACHO = ['TRASPASO_SALIDA']
CONCEPTOS_VENTA = ['VENTA_PUBLICO', 'VENTA_MAYORISTA']
# Egresos que NO son venta ni traspaso interno: mermas, ajustes a la baja,
# devoluciones a proveedor y salida por cambio. Se cuentan aparte para cerrar la
# reconciliación entrada → (venta + despacho + otros egresos + remanente), en vez
# de que ese stock que salió aparezca como "no vendido" en góndola.
CONCEPTOS_OTROS_EGRESOS = [
    'AJUSTE_NEGATIVO', 'PERDIDA_ROBO', 'PERDIDA_DETERIORO',
    'DONACION_ENTREGADA', 'DEVOLUCION_PROVEEDOR', 'CAMBIO_PRODUCTO_SALIDA',
]


@login_required
@require_GET
def api_rendimiento_compras(request):
    """
    Análisis de rendimiento: Entrada → Despacho → Venta por año.
    DUAL-SOURCE:
      - Si el año tiene >= 500 movimientos en Movimientos_Producto → usa movimientos (sistema actual/2025-2026)
      - Si tiene < 500 movimientos → usa DTEs migrados de Laravel (histórico 2018-2024)
    Soporta comparativa interanual (param comparar=1).

    SCOPING: todas las puntas del embudo (movimientos, tickets y DTE) se acotan
    al alcance del usuario. Antes esta vista no tenía ni un filtro por empresa:
    bastaba llamarla sin `sucursal` para leer inversión, ventas, margen y ROI
    de TODO el holding, o pasar el id de una sucursal ajena para leer los de
    esa empresa.
    """
    from .models import Movimientos_Producto
    from .utils_analitica import productos_activos_qs

    UMBRAL_MOVIMIENTOS = 500

    sucursal_id = request.GET.get('sucursal', '') or None
    permitidas, sin_acceso = _alcance_sucursales(request, sucursal_id)
    if sin_acceso:
        return sin_acceso

    def _dtes(**filtros):
        """DTE dentro del alcance del usuario (ver `_q_alcance_dte`)."""
        base = Dte.objects.filter(_q_alcance_dte(request.user), **filtros)
        return base.filter(sucursal_id=sucursal_id) if sucursal_id else base

    def _datos_desde_movimientos(anio, sucursal_id, pt_ids):
        """Agrega datos de Movimientos_Producto para el año dado."""
        bf = {'fecha__year': anio, 'estado': 'COMPLETADO', 'ProductoTalla_id__in': pt_ids}
        bf.update(_filtro_alcance('sucursal_origen_id', sucursal_id, permitidas))

        ent = Movimientos_Producto.objects.filter(concepto__in=CONCEPTOS_ENTRADA, **bf).aggregate(
            uds=Sum('cantidad'),
            costo=Sum(F('costo') * F('cantidad')),
        )
        desp = Movimientos_Producto.objects.filter(concepto__in=CONCEPTOS_DESPACHO, **bf).aggregate(
            uds=Sum('cantidad'),
        )
        vtas = Movimientos_Producto.objects.filter(concepto__in=CONCEPTOS_VENTA, **bf).aggregate(
            uds=Sum('cantidad'),
            ingreso=Sum(F('precio') * F('cantidad')),
        )
        # Otros egresos (mermas/ajustes a la baja/devoluciones a proveedor/cambios).
        # Son EGRESO → cantidad negativa; se normaliza a positivo con abs() abajo.
        otros = Movimientos_Producto.objects.filter(concepto__in=CONCEPTOS_OTROS_EGRESOS, **bf).aggregate(
            uds=Sum('cantidad'),
        )
        q_tk = {'idTicket__fecha__year': anio, 'idTicket__estado__in': ['PAGADO', 'PENDIENTE'],
                'ProductoTalla_id__in': pt_ids}
        q_tk.update(_filtro_alcance('idTicket__sucursal_id', sucursal_id, permitidas))
        vtas_tk = Ticket_Productos.objects.filter(**q_tk).aggregate(
            uds=Sum('stock'), ingreso=Sum('subtotal'),
        )
        return {
            'entrada': int(ent['uds'] or 0),
            'inversion': float(ent['costo'] or 0),
            'despacho': abs(int(desp['uds'] or 0)),
            'venta_mov': int(vtas['uds'] or 0),
            'venta_tk': int(vtas_tk['uds'] or 0),
            'ingreso_mov': float(vtas['ingreso'] or 0),
            'ingreso_tk': float(vtas_tk['ingreso'] or 0),
            'otros_egresos': abs(int(otros['uds'] or 0)),
            'fuente': 'movimientos',
            '_bf': bf,
        }

    def _datos_desde_dte(anio, sucursal_id):
        """Agrega datos de DTEs migrados de Laravel para el año dado."""
        # El filtro por sucursal (pedida o alcance) lo aplica `_dtes`.
        dte_f = {'fecha_emision__year': anio}

        # Entrada: DTEs tipo COMPRA → Dte_Productos (stock=unidades, costo=costo unit)
        dte_compra_ids = _dtes(tipo_transaccion='COMPRA', **dte_f).values_list('id', flat=True)
        ent = Dte_Productos.objects.filter(
            dte_id__in=dte_compra_ids,
            productoTalla__producto__excluir_de_analitica=False,
        ).aggregate(
            uds=Sum('stock'),
            costo=Sum(F('costo') * F('stock')),
        )

        # Despacho: DTEs tipo TRASPASO
        dte_traspaso_ids = _dtes(tipo_transaccion='TRASPASO', **dte_f).values_list('id', flat=True)
        desp = Dte_Productos.objects.filter(
            dte_id__in=dte_traspaso_ids,
            productoTalla__producto__excluir_de_analitica=False,
        ).aggregate(
            uds=Sum('stock'),
        )

        # Ventas: DTEs tipo VENTA_PUBLICO + VENTA (ingreso = precio * stock)
        dte_venta_ids = _dtes(
            tipo_transaccion__in=['VENTA_PUBLICO', 'VENTA'], **dte_f
        ).values_list('id', flat=True)
        vtas = Dte_Productos.objects.filter(
            dte_id__in=dte_venta_ids,
            productoTalla__producto__excluir_de_analitica=False,
        ).aggregate(
            uds=Sum('stock'),
            ingreso=Sum(F('precio') * F('stock')),
        )

        return {
            'entrada': int(ent['uds'] or 0),
            'inversion': float(ent['costo'] or 0),
            'despacho': abs(int(desp['uds'] or 0)),
            'venta_mov': int(vtas['uds'] or 0),
            'venta_tk': 0,
            'ingreso_mov': float(vtas['ingreso'] or 0),
            'ingreso_tk': 0,
            'otros_egresos': 0,  # DTEs migrados no registran mermas/ajustes
            'fuente': 'dte',
            '_bf': dte_f,
        }

    def _evolucion_desde_movimientos(anio, sucursal_id, pt_ids, bf):
        ent_m = {e['mes']: e for e in Movimientos_Producto.objects.filter(
            concepto__in=CONCEPTOS_ENTRADA, **bf
        ).values(mes=F('fecha__month')).annotate(uds=Sum('cantidad'))}

        vtas_m = {v['mes']: v for v in Movimientos_Producto.objects.filter(
            concepto__in=CONCEPTOS_VENTA, **bf
        ).values(mes=F('fecha__month')).annotate(uds=Sum('cantidad'), ingreso=Sum(F('precio') * F('cantidad')))}

        q_tk = {'idTicket__fecha__year': anio, 'idTicket__estado__in': ['PAGADO', 'PENDIENTE']}
        q_tk.update(_filtro_alcance('idTicket__sucursal_id', sucursal_id, permitidas))
        vtas_tk = {v['mes']: v for v in Ticket_Productos.objects.filter(**q_tk).values(
            mes=F('idTicket__fecha__month')
        ).annotate(uds=Sum('stock'), ingreso=Sum('subtotal'))}

        evol = []
        for m in range(1, 13):
            e_m = ent_m.get(m, {})
            vm = vtas_m.get(m, {})
            vt = vtas_tk.get(m, {})
            evol.append({
                'mes': m,
                'uds_compradas': e_m.get('uds', 0) or 0,
                'uds_vendidas': max(vm.get('uds', 0) or 0, vt.get('uds', 0) or 0),
                'ingreso': max(float(vm.get('ingreso', 0) or 0), float(vt.get('ingreso', 0) or 0)),
            })
        return evol

    def _evolucion_desde_dte(anio, sucursal_id):
        dte_f = {'fecha_emision__year': anio}

        dte_compra_ids = list(_dtes(tipo_transaccion='COMPRA', **dte_f).values_list('id', flat=True))
        dte_venta_ids = list(_dtes(
            tipo_transaccion__in=['VENTA_PUBLICO', 'VENTA'], **dte_f
        ).values_list('id', flat=True))

        ent_m = {e['mes']: e for e in Dte_Productos.objects.filter(
            dte_id__in=dte_compra_ids,
            productoTalla__producto__excluir_de_analitica=False,
        ).annotate(
            mes=F('dte__fecha_emision__month')
        ).values('mes').annotate(uds=Sum('stock'))} if dte_compra_ids else {}

        vtas_m = {v['mes']: v for v in Dte_Productos.objects.filter(
            dte_id__in=dte_venta_ids,
            productoTalla__producto__excluir_de_analitica=False,
        ).annotate(
            mes=F('dte__fecha_emision__month')
        ).values('mes').annotate(uds=Sum('stock'), ingreso=Sum(F('precio') * F('stock')))} if dte_venta_ids else {}

        evol = []
        for m in range(1, 13):
            e_m = ent_m.get(m, {})
            vm = vtas_m.get(m, {})
            evol.append({
                'mes': m,
                'uds_compradas': e_m.get('uds', 0) or 0,
                'uds_vendidas': vm.get('uds', 0) or 0,
                'ingreso': float(vm.get('ingreso', 0) or 0),
            })
        return evol

    def _sucursales_desde_dte(anio, sucursal_id):
        """Desglose por sucursal usando DTEs (por sucursal del DTE)."""
        dte_f = {'fecha_emision__year': anio}

        # Ventas agrupadas por sucursal del DTE
        vtas_suc = {}
        for v in _dtes(
            tipo_transaccion__in=['VENTA_PUBLICO', 'VENTA'], **dte_f
        ).annotate(
            uds=Sum('dte_productos__stock',
                    filter=Q(dte_productos__productoTalla__producto__excluir_de_analitica=False)),
            ingreso=Sum(F('dte_productos__precio') * F('dte_productos__stock'),
                        filter=Q(dte_productos__productoTalla__producto__excluir_de_analitica=False)),
        ).values('sucursal_id', 'sucursal__alias', 'uds', 'ingreso'):
            sid = v['sucursal_id']
            if sid not in vtas_suc:
                vtas_suc[sid] = {'alias': v['sucursal__alias'], 'uds': 0, 'ingreso': 0.0}
            vtas_suc[sid]['uds'] += v['uds'] or 0
            vtas_suc[sid]['ingreso'] += float(v['ingreso'] or 0)

        result = []
        for sid, v in vtas_suc.items():
            result.append({
                'sucursal': v['alias'] or 'Sin sucursal',
                'uds_entrada': 0,
                'uds_despachadas': 0,
                'uds_vendidas': v['uds'],
                'ingreso': v['ingreso'],
                'costo_despacho': 0,
                'rotacion_pct': 0,
                'stock_restante': 0,
                'roi': 0,
                'explicacion': 'Datos desde DTEs migrados de Laravel.',
            })
        result.sort(key=lambda x: x['uds_vendidas'], reverse=True)
        return result

    def _calcular_resumen(d):
        """Calcula KPIs a partir del dict de datos brutos."""
        entrada = d['entrada']
        inversion = d['inversion']
        despacho = d['despacho']
        otros_egresos = d.get('otros_egresos', 0)
        vendido = max(d['venta_mov'], d['venta_tk'])
        ingreso = max(d['ingreso_mov'], d['ingreso_tk'])

        costo_vendido = inversion * (vendido / entrada) if entrada > 0 else 0
        margen = ingreso - costo_vendido
        rotacion_raw = (vendido / entrada) * 100 if entrada > 0 else 0
        rotacion = min(999.0, round(rotacion_raw, 1))
        roi = round((margen / costo_vendido) * 100, 1) if costo_vendido > 0 else 0

        return {
            'total_comprado': entrada,
            'total_recibido': entrada,
            'total_despachado': despacho,
            'total_vendido': vendido,
            'total_inversion': inversion,
            'total_ingreso_ventas': ingreso,
            'margen_real': margen,
            'rotacion_global': rotacion,
            'rotacion_raw_pct': round(rotacion_raw, 1),
            'ventas_superan_entrada': vendido > entrada,
            'roi_real': roi,
            'tasa_despacho': round((despacho / entrada) * 100, 1) if entrada > 0 else 0,
            'total_otros_egresos': otros_egresos,
            # Remanente real = lo que entró menos lo que salió por CUALQUIER vía
            # (venta + otros egresos). Antes solo restaba ventas, así que las
            # mermas/devoluciones inflaban el "no vendido".
            'stock_sin_vender': entrada - vendido - otros_egresos,
            'productos_analizados': entrada,
        }

    def _delta(actual, anterior, campo):
        """Calcula delta % entre actual y año anterior."""
        a = actual.get(campo, 0) or 0
        b = anterior.get(campo, 0) or 0
        if b == 0:
            return None
        return round((a - b) / abs(b) * 100, 1)

    try:
        anio = int(request.GET.get('anio', timezone.localdate().year))
        comparar = request.GET.get('comparar', '0') == '1'

        pt_ids = (
            productos_activos_qs()
            .filter(**_filtro_alcance('producto__sucursal_id', sucursal_id, permitidas))
            .values_list('id', flat=True)
        )

        # ── Determinar fuente para el año solicitado ──
        # `count_movs` es un HEURÍSTICO de disponibilidad de datos del sistema
        # (¿este año está en el sistema nuevo o sólo en los DTE migrados?), no
        # un dato de negocio de una empresa; se deja global a propósito para no
        # cambiar la fuente —y con ella los números— según quién mire.
        count_movs = Movimientos_Producto.objects.filter(
            estado='COMPLETADO', fecha__year=anio
        ).count()
        usar_dte = count_movs < UMBRAL_MOVIMIENTOS

        if usar_dte:
            datos = _datos_desde_dte(anio, sucursal_id)
        else:
            datos = _datos_desde_movimientos(anio, sucursal_id, pt_ids)
            # Fallback: si hay movimientos en general pero sin entradas de stock,
            # complementar con DTEs para tener al menos los KPIs de ventas históricas
            if datos['entrada'] == 0:
                datos_dte = _datos_desde_dte(anio, sucursal_id)
                if datos_dte['entrada'] > 0 or datos_dte['venta_mov'] > 0:
                    # DTE tiene más datos; usarlo en modo mixto
                    datos = datos_dte
                    usar_dte = True
                else:
                    # Mantener movimientos con ventas aunque entrada=0
                    # (puede haber ventas de stock de años anteriores)
                    datos['fuente'] = 'movimientos'

        # ── Años disponibles (basados en DTEs que tienen datos) ──
        anios_dte = sorted(set(
            _dtes(
                tipo_transaccion__in=['COMPRA', 'VENTA_PUBLICO', 'VENTA']
            ).dates('fecha_emision', 'year')
        ), key=lambda d: d.year, reverse=True)
        anios_disponibles = [d.year for d in anios_dte]
        # Asegurar que el año actual también aparezca
        anio_actual = timezone.localdate().year
        if anio_actual not in anios_disponibles:
            anios_disponibles.insert(0, anio_actual)

        # ── KPIs del año actual ──
        resumen = _calcular_resumen(datos)

        # ── Nota de metodología ──
        if usar_dte:
            nota_metodologia = (
                f'Datos del año {anio} obtenidos desde DTEs migrados de Laravel '
                f'(el año tiene solo {count_movs} movimientos registrados en el nuevo sistema). '
                'Entrada = DTEs tipo COMPRA, Ventas = DTEs tipo VENTA_PUBLICO/VENTA, '
                'Despacho = DTEs tipo TRASPASO.'
            )
        else:
            if resumen['total_comprado'] == 0:
                nota_metodologia = (
                    f'El año {anio} tiene {count_movs} movimientos pero sin ingresos de stock '
                    '(RECEPCION_COMPRA / INGRESO_INICIAL). Las ventas corresponden a stock ingresado en períodos anteriores. '
                    'Las métricas de rotación e inversión no están disponibles para este año.'
                )
            else:
                nota_metodologia = (
                    'Datos desde Movimientos_Producto del sistema actual. '
                    'Embudo coherente: mismo año y sucursal para entrada, despacho saliente y ventas.'
                )
        if resumen['ventas_superan_entrada']:
            nota_metodologia += ' Rotación >100% indica ventas con stock de ciclos anteriores.'

        # ── Comparativa interanual ──
        comparacion = None
        if comparar:
            anio_ant = anio - 1
            count_movs_ant = Movimientos_Producto.objects.filter(
                estado='COMPLETADO', fecha__year=anio_ant
            ).count()
            usar_dte_ant = count_movs_ant < UMBRAL_MOVIMIENTOS
            if usar_dte_ant:
                datos_ant = _datos_desde_dte(anio_ant, sucursal_id)
            else:
                datos_ant = _datos_desde_movimientos(anio_ant, sucursal_id, pt_ids)
            resumen_ant = _calcular_resumen(datos_ant)

            comparacion = {
                'anio_anterior': anio_ant,
                'fuente_anterior': datos_ant['fuente'],
                'delta_entrada': _delta(resumen, resumen_ant, 'total_comprado'),
                'delta_vendido': _delta(resumen, resumen_ant, 'total_vendido'),
                'delta_inversion': _delta(resumen, resumen_ant, 'total_inversion'),
                'delta_ingreso': _delta(resumen, resumen_ant, 'total_ingreso_ventas'),
                'delta_rotacion': _delta(resumen, resumen_ant, 'rotacion_global'),
                'delta_roi': _delta(resumen, resumen_ant, 'roi_real'),
                'resumen_anterior': resumen_ant,
            }

        # ── Evolución mensual ──
        bf = datos.get('_bf', {})
        if usar_dte:
            evolucion = _evolucion_desde_dte(anio, sucursal_id)
        else:
            evolucion = _evolucion_desde_movimientos(anio, sucursal_id, pt_ids, bf)

        # ── Por sucursal ──
        if usar_dte:
            por_sucursal = _sucursales_desde_dte(anio, sucursal_id)
            por_compra = []
        else:
            # Lógica existente de movimientos
            entradas_suc = {e['sucursal_origen_id']: e for e in Movimientos_Producto.objects.filter(
                concepto__in=CONCEPTOS_ENTRADA, **bf
            ).values('sucursal_origen_id', 'sucursal_origen__alias').annotate(
                uds=Sum('cantidad'), costo=Sum(F('costo') * F('cantidad')),
            )}
            despachos_suc = {d['sucursal_origen_id']: d for d in Movimientos_Producto.objects.filter(
                concepto__in=CONCEPTOS_DESPACHO, **bf
            ).values('sucursal_origen_id', 'sucursal_origen__alias').annotate(
                uds=Sum('cantidad'), costo=Sum(F('costo') * F('cantidad')),
            )}
            ventas_suc_mov = {v['sucursal_origen_id']: v for v in Movimientos_Producto.objects.filter(
                concepto__in=CONCEPTOS_VENTA, **bf
            ).values('sucursal_origen_id', 'sucursal_origen__alias').annotate(
                uds=Sum('cantidad'), ingreso=Sum(F('precio') * F('cantidad')),
            )}
            q_tk_suc = {'idTicket__fecha__year': anio, 'idTicket__estado__in': ['PAGADO', 'PENDIENTE']}
            q_tk_suc.update(_filtro_alcance('idTicket__sucursal_id', sucursal_id, permitidas))
            ventas_suc_ticket = {v['idTicket__sucursal_id']: v for v in Ticket_Productos.objects.filter(
                **q_tk_suc
            ).values('idTicket__sucursal_id', 'idTicket__sucursal__alias').annotate(
                uds=Sum('stock'), ingreso=Sum('subtotal'),
            )}

            all_suc_ids = set(list(entradas_suc.keys()) + list(despachos_suc.keys()) +
                             list(ventas_suc_mov.keys()) + list(ventas_suc_ticket.keys()))
            por_sucursal = []
            por_compra = []
            for sid in all_suc_ids:
                if not sid:
                    continue
                e = entradas_suc.get(sid, {})
                d = despachos_suc.get(sid, {})
                vm = ventas_suc_mov.get(sid, {})
                vt = ventas_suc_ticket.get(sid, {})

                uds_ent = e.get('uds', 0) or 0
                uds_desp = abs(d.get('uds', 0) or 0)
                uds_vend = max(vm.get('uds', 0) or 0, vt.get('uds', 0) or 0)
                ingreso = max(float(vm.get('ingreso', 0) or 0), float(vt.get('ingreso', 0) or 0))
                costo_e = float(e.get('costo', 0) or 0)
                costo_d = abs(float(d.get('costo', 0) or 0))
                suc_nombre = (e.get('sucursal_origen__alias') or d.get('sucursal_origen__alias') or
                              vm.get('sucursal_origen__alias') or vt.get('idTicket__sucursal__alias') or 'Sin sucursal')

                rot = round((uds_vend / uds_ent) * 100, 1) if uds_ent > 0 else (
                    round((uds_vend / uds_desp) * 100, 1) if uds_desp > 0 else 0)
                roi = round(((ingreso - costo_e) / costo_e) * 100, 1) if costo_e > 0 else 0

                explicacion = ''
                if uds_ent < uds_vend and uds_ent > 0:
                    explicacion = 'Ventas mayores que entradas del año: el stock vendido puede venir de años anteriores o ingresos por traspaso desde otra bodega.'
                elif uds_ent == 0 and uds_vend > 0:
                    explicacion = 'Sin ingreso inicial en esta sucursal; ventas usan stock histórico o traspasos registrados con origen en central.'

                por_sucursal.append({
                    'sucursal': suc_nombre,
                    'uds_entrada': uds_ent,
                    'uds_despachadas': uds_desp,
                    'uds_vendidas': uds_vend,
                    'ingreso': ingreso,
                    'costo_despacho': costo_d,
                    'rotacion_pct': rot,
                    'stock_restante': uds_ent - uds_vend if uds_ent > 0 else uds_desp - uds_vend,
                    'roi': roi,
                    'explicacion': explicacion,
                })
                if uds_ent > 0:
                    por_compra.append({
                        'compra': f"{suc_nombre} (Ingreso)",
                        'fuente': 'Movimiento',
                        'productos': 0,
                        'uds_compradas': uds_ent,
                        'uds_despachadas': uds_desp,
                        'uds_vendidas': uds_vend,
                        'rotacion_pct': rot,
                        'inversion': costo_e,
                        'ingreso_venta': ingreso,
                        'roi_real': roi,
                    })

            por_sucursal.sort(key=lambda x: x['rotacion_pct'], reverse=True)
            por_compra.sort(key=lambda x: x['rotacion_pct'], reverse=True)

        # ── Top productos ──
        por_producto = []
        if not usar_dte:
            q_top_tk = {'idTicket__fecha__year': anio, 'idTicket__estado__in': ['PAGADO', 'PENDIENTE']}
            q_top_tk.update(_filtro_alcance('idTicket__sucursal_id', sucursal_id, permitidas))
            top_vendidos = list(Ticket_Productos.objects.filter(**q_top_tk).values(
                'ProductoTalla_id', 'ProductoTalla__producto__articulo', 'ProductoTalla__talla',
            ).annotate(uds_vendidas=Sum('stock'), ingreso=Sum('subtotal')).order_by('-uds_vendidas')[:150])

            if not top_vendidos:
                top_vendidos = list(Movimientos_Producto.objects.filter(
                    concepto__in=CONCEPTOS_VENTA, **bf
                ).values(
                    'ProductoTalla_id', 'ProductoTalla__producto__articulo', 'ProductoTalla__talla',
                ).annotate(uds_vendidas=Sum('cantidad'), ingreso=Sum(F('precio') * F('cantidad'))
                ).order_by('-uds_vendidas')[:150])

            top_pt_ids = [t['ProductoTalla_id'] for t in top_vendidos]
            entradas_top = {}
            if top_pt_ids:
                for e in Movimientos_Producto.objects.filter(
                    ProductoTalla_id__in=top_pt_ids,
                    concepto__in=CONCEPTOS_ENTRADA, fecha__year=anio, estado='COMPLETADO',
                ).values('ProductoTalla_id').annotate(uds=Sum('cantidad'), costo=Sum(F('costo') * F('cantidad'))):
                    entradas_top[e['ProductoTalla_id']] = e
            despachos_top = {}
            if top_pt_ids:
                for d in Movimientos_Producto.objects.filter(
                    ProductoTalla_id__in=top_pt_ids,
                    concepto__in=CONCEPTOS_DESPACHO, fecha__year=anio, estado='COMPLETADO',
                ).values('ProductoTalla_id').annotate(uds=Sum('cantidad')):
                    despachos_top[d['ProductoTalla_id']] = abs(d['uds'] or 0)

            for tv in top_vendidos:
                pt_id = tv['ProductoTalla_id']
                uds_v = tv['uds_vendidas'] or 0
                ing = float(tv['ingreso'] or 0)
                e = entradas_top.get(pt_id, {})
                uds_e = e.get('uds', 0) or 0
                costo_t = float(e.get('costo', 0) or 0)
                uds_d = despachos_top.get(pt_id, 0)
                cu = round(costo_t / uds_e) if uds_e > 0 else 0
                margen = ing - (cu * uds_v)
                rot = round((uds_v / uds_e) * 100, 1) if uds_e > 0 else 0
                roi_p = round((margen / costo_t) * 100, 1) if costo_t > 0 else 0
                por_producto.append({
                    'articulo': tv['ProductoTalla__producto__articulo'] or '-',
                    'talla': tv['ProductoTalla__talla'] or '-',
                    'origen': '-', 'tipo_entrada': 'Movimiento', 'fuente': 'Movimiento',
                    'uds_compradas': uds_e, 'uds_recibidas': uds_e, 'uds_despachadas': uds_d,
                    'uds_vendidas': uds_v, 'costo_unitario': cu, 'ingreso_venta': ing,
                    'margen_real': margen, 'rotacion_pct': rot, 'roi_real': roi_p,
                    'stock_restante': uds_e - uds_v if uds_e > 0 else 0,
                })
            por_producto.sort(key=lambda x: x['rotacion_pct'], reverse=True)
        else:
            # Para años DTE: top ventas desde DTEs
            dte_venta_ids = list(_dtes(
                tipo_transaccion__in=['VENTA_PUBLICO', 'VENTA'], fecha_emision__year=anio
            ).values_list('id', flat=True))
            if dte_venta_ids:
                for v in Dte_Productos.objects.filter(
                    dte_id__in=dte_venta_ids,
                    productoTalla__producto__excluir_de_analitica=False,
                ).values(
                    'descripcion'
                ).annotate(uds_vendidas=Sum('stock'), ingreso=Sum(F('precio') * F('stock'))
                ).order_by('-uds_vendidas')[:100]:
                    por_producto.append({
                        'articulo': v['descripcion'] or '-',
                        'talla': '-', 'origen': '-', 'tipo_entrada': 'DTE', 'fuente': 'DTE',
                        'uds_compradas': 0, 'uds_recibidas': 0, 'uds_despachadas': 0,
                        'uds_vendidas': v['uds_vendidas'] or 0,
                        'costo_unitario': 0, 'ingreso_venta': float(v['ingreso'] or 0),
                        'margen_real': 0, 'rotacion_pct': 0, 'roi_real': 0, 'stock_restante': 0,
                    })

        return JsonResponse({
            'success': True,
            'resumen': resumen,
            'nota_metodologia': nota_metodologia,
            'fuente_datos': datos['fuente'],
            'anios_disponibles': anios_disponibles,
            'comparacion': comparacion,
            'por_compra': por_compra[:50],
            'por_producto': por_producto[:100],
            'por_sucursal': por_sucursal,
            'evolucion_mensual': evolucion,
        })

    except Exception as e:
        logger.exception("Error al generar rendimiento de compras")
        return JsonResponse({
            'success': False,
            'error': str(e),
        }, status=500)


def _resumen_vacio():
    return {
        'total_comprado': 0, 'total_recibido': 0, 'total_despachado': 0,
        'total_vendido': 0, 'total_inversion': 0, 'total_ingreso_ventas': 0,
        'margen_real': 0, 'rotacion_global': 0, 'roi_real': 0,
        'tasa_despacho': 0, 'stock_sin_vender': 0, 'productos_analizados': 0,
    }


# ========== REPORTE DE MOVIMIENTOS POR SUCURSAL (INICIAL VS RESTANTE) ==========

# NOTA: helper interno, NO es una vista. No lleva @login_required: se invoca como
# _mapas_movimientos_sucursal(productos_ids, ...) y el wrapper de login_required
# tomaría productos_ids (una lista) como `request` -> 'list' object has no
# attribute 'user'. El control de acceso ya está en las vistas que lo llaman.
def _mapas_movimientos_sucursal(productos_ids, sucursales_ids, filtro_fecha):
    """Mapas {producto_id: {sucursal_id: unidades}} para el reporte de
    movimientos por sucursal, separando compras/abastecimiento, traspasos
    recibidos, traspasos enviados y ventas. Los sets de conceptos viven en
    app.constants_kardex para que todos los reportes cuenten lo mismo.
    Nota: los traspasos legacy de una sola pierna (TRASPASO_SUCURSAL/BODEGA/
    VITRINA) se clasifican por signo."""
    from app.constants_kardex import (
        CONCEPTOS_ABASTECIMIENTO,
        CONCEPTOS_TRASPASO_ENTRADA,
        CONCEPTOS_TRASPASO_LEGACY,
        CONCEPTOS_TRASPASO_SALIDA,
        CONCEPTOS_VENTA,
    )

    conceptos_compras = CONCEPTOS_ABASTECIMIENTO + (
        'CAMBIO_PRODUCTO_ENTRADA', 'AJUSTE_POSITIVO',
        'AJUSTE_INVENTARIO_ENTRADA', 'SOBRANTE_INGRESO', 'DONACION_RECIBIDA',
    )

    base = Movimientos_Producto.objects.filter(
        ProductoTalla__producto_id__in=productos_ids,
        estado='COMPLETADO',
        **filtro_fecha,
    )

    def _mapa(queryset, campo_sucursal):
        mapa = {}
        for item in queryset:
            prod = item['ProductoTalla__producto_id']
            mapa.setdefault(prod, {})[item[campo_sucursal]] = abs(item['total'] or 0)
        return mapa

    def _mapa_entrada(qs_base):
        """Suma entradas por (producto, sucursal RECEPTORA).

        La sucursal receptora es SIEMPRE la dueña del SKU
        (``ProductoTalla.producto.sucursal_id``). En este modelo el stock vive
        por sucursal en la propia fila ``Producto``, así que un movimiento
        sobre un SKU solo puede afectar el stock de esa sucursal:
        ``sucursal_origen``/``sucursal_destino`` describen el traslado, no
        dónde quedó el saldo.

        La versión anterior atribuía por ``COALESCE(destino, origen)``. Eso es
        correcto para INGRESO_INICIAL/RECEPCION_COMPRA (la migración de Laravel
        dejó ``sucursal_destino=NULL`` y la sucursal en ``sucursal_origen``),
        pero es al revés para los traspasos legacy de una pierna: ahí
        ``sucursal_origen`` es la sucursal EMISORA, así que la entrada se
        sumaba a quien despachó y no a quien recibió. Medición contra prod
        (jul-2026): en SKECHERS ~13.000 unidades quedaban contadas en
        PAO4/NICK1/PAO3/PAO1 siendo de EDEL, dejando la fila de EDEL con
        Recib. por debajo de Rest.; en PANAMA JACK eran 215 de 4.724 entradas
        (4,5%). Contar por la sucursal dueña elimina el caso y no puede
        duplicar (cada SKU pertenece a una sola sucursal)."""
        mapa = {}
        filas = (
            qs_base.filter(ProductoTalla__producto__sucursal_id__in=sucursales_ids)
            .values('ProductoTalla__producto_id',
                    'ProductoTalla__producto__sucursal_id')
            .annotate(total=Sum('cantidad'))
        )
        for item in filas:
            prod = item['ProductoTalla__producto_id']
            suc = item['ProductoTalla__producto__sucursal_id']
            d = mapa.setdefault(prod, {})
            d[suc] = d.get(suc, 0) + abs(item['total'] or 0)
        return mapa

    # Entradas (abastecimiento y traspasos recibidos): solo movimientos que
    # SUMAN stock (cantidad > 0), contados por la sucursal receptora.
    compras = _mapa_entrada(
        base.filter(concepto__in=conceptos_compras, cantidad__gt=0)
    )
    traspasos_in = _mapa_entrada(
        base.filter(cantidad__gt=0).filter(
            Q(concepto__in=CONCEPTOS_TRASPASO_ENTRADA)
            | Q(concepto__in=CONCEPTOS_TRASPASO_LEGACY)
        )
    )
    traspasos_out = _mapa(
        base.filter(sucursal_origen_id__in=sucursales_ids, cantidad__lt=0)
        .filter(Q(concepto__in=CONCEPTOS_TRASPASO_SALIDA)
                | Q(concepto__in=CONCEPTOS_TRASPASO_LEGACY))
        .values('ProductoTalla__producto_id', 'sucursal_origen_id')
        .annotate(total=Sum('cantidad')),
        'sucursal_origen_id',
    )
    ventas = _mapa(
        base.filter(sucursal_origen_id__in=sucursales_ids,
                    concepto__in=CONCEPTOS_VENTA)
        .values('ProductoTalla__producto_id', 'sucursal_origen_id')
        .annotate(total=Sum('cantidad')),
        'sucursal_origen_id',
    )
    return compras, traspasos_in, traspasos_out, ventas


# NOTA: helper interno, NO es una vista (mismo motivo que el de arriba: no
# lleva @login_required porque se invoca con args posicionales).
def _saldos_periodo(productos_ids, sucursales_ids, filtro_fecha):
    """Mapas para reconstruir el saldo de un período (kardex real).

    Devuelve tres mapas ``{producto_id: {sucursal_id: unidades}}``:

    - ``entradas``  — SUM(cantidad) de los movimientos con ``cantidad > 0``
      DENTRO de la ventana de fechas.
    - ``salidas``   — idem con ``cantidad < 0``, en valor absoluto.
    - ``posterior`` — SUM(cantidad) CON SIGNO de los movimientos POSTERIORES a
      ``fecha_hasta``. Vacío si no se filtró por ``fecha_hasta``.

    El sistema no guarda snapshots históricos de stock en ninguna tabla: el
    único saldo conocido es ``Producto_Talla.stock`` de HOY. Así que el saldo
    del período se obtiene rebobinando ese stock::

        saldo_final   = stock_hoy   - posterior
        saldo_inicial = saldo_final - (entradas - salidas)

    Con esa definición se cumple por construcción
    ``saldo_inicial + entradas - salidas = saldo_final``.

    Sin esto el reporte comparaba el RECIBIDO DE LA VENTANA contra el STOCK DE
    HOY —dos magnitudes que no se relacionan—, así que todo lo que hubiera
    llegado antes de ``fecha_desde`` aparecía como "Restante sin Recibido".
    Medición contra prod (jul-2026): PANAMA JACK filtrado desde 2026-07-01
    daba Recib. 0 vs Rest. 368, o sea el 100% del stock como fantasma.

    SALDO DE APERTURA DE LA MIGRACIÓN: se excluye (ver
    ``REF_SALDO_INICIAL_SINTETICO``). Esa carga entró como ``INGRESO_INICIAL``
    fechada ≈2026-01-22, pero el kardex legacy ANTERIOR también se migró, así
    que contarla como entrada suma el mismo stock dos veces. Al excluirla queda
    absorbida en el saldo inicial, que es lo que realmente es: la foto de
    arranque. Medición contra prod (ago-2026): 1066-1-CA en PAO2, período desde
    2026-01-01 → la apertura (+7) caía DENTRO de la ventana y el rebobinado
    daba Original 0 con Actual 6, cuando el kardex legacy cerraba 2025 en 7.
    Sin el arreglo, cualquier ventana que cruce la fecha de corte da originales
    en 0 o negativos.

    Todo se atribuye a la sucursal DUEÑA del SKU, igual que ``_mapa_entrada``.
    """
    from app.constants_kardex import REF_SALDO_INICIAL_SINTETICO

    base = Movimientos_Producto.objects.filter(
        ProductoTalla__producto_id__in=productos_ids,
        ProductoTalla__producto__sucursal_id__in=sucursales_ids,
        estado='COMPLETADO',
    ).exclude(
        concepto='INGRESO_INICIAL',
        referencia_externa=REF_SALDO_INICIAL_SINTETICO,
    )

    def _agrupar(queryset, absoluto):
        mapa = {}
        filas = (
            queryset.values('ProductoTalla__producto_id',
                            'ProductoTalla__producto__sucursal_id')
            .annotate(total=Sum('cantidad'))
        )
        for item in filas:
            prod = item['ProductoTalla__producto_id']
            suc = item['ProductoTalla__producto__sucursal_id']
            valor = item['total'] or 0
            d = mapa.setdefault(prod, {})
            d[suc] = d.get(suc, 0) + (abs(valor) if absoluto else valor)
        return mapa

    ventana = base.filter(**filtro_fecha)
    entradas = _agrupar(ventana.filter(cantidad__gt=0), True)
    salidas = _agrupar(ventana.filter(cantidad__lt=0), True)

    posterior = {}
    if filtro_fecha.get('fecha__lte'):
        posterior = _agrupar(
            base.filter(fecha__gt=filtro_fecha['fecha__lte']), False
        )
    return entradas, salidas, posterior


# NOTA: helper interno, NO es una vista (se invoca con args posicionales, ver
# el comentario de _mapas_movimientos_sucursal).
def _sucursales_reporte_movimientos(user, solo_tiendas):
    """Sucursales visibles en el reporte de movimientos por sucursal.

    Con ``solo_tiendas`` se dejan fuera las bodegas proveedoras / centros de
    distribución (EDEL, GILD y compañía). Se replica la property
    ``Sucursal.es_compradora`` —que es Python puro, no una columna— para poder
    decidirlo en memoria sobre las sucursales ya cargadas.

    Excluir una sucursal la saca del reporte COMPLETO, no solo de sus columnas:
    cada ``Producto`` pertenece a una sola sucursal, así que sus filas, sus
    totales y sus KPI desaparecen juntos. Es justo lo que se quiere —el total
    debe cuadrar con lo que se ve— pero significa que el stock que duerme en
    bodega no está escondido en ninguna columna: no está.

    Devuelve ``(visibles, cantidad_ocultas)``.
    """
    empresas_usuario = EmpresaUser.objects.filter(
        user=user,
        status=True
    ).values_list('empresa_id', flat=True)
    todas = list(Sucursal.objects.filter(empresa_id__in=empresas_usuario).order_by('alias'))
    if not solo_tiendas:
        return todas, 0
    visibles = [s for s in todas if not s.es_compradora]
    return visibles, len(todas) - len(visibles)


@login_required
def ver_reporte_movimientos_sucursal(request):
    """
    Vista principal del reporte de movimientos por sucursal (inicial vs restante).

    `@login_required` es obligatorio: el middleware de permisos devuelve None
    para usuarios anónimos (no los bloquea), así que sin el decorador la vista
    corría con `AnonymousUser`, el filtro `EmpresaUser.objects.filter(user=...)`
    reventaba y la página respondía 500 en vez de mandar al login.
    """
    sucursal_actual_id = request.session.get('idSucursalActual') or request.session.get('sucursalActual')
    sucursal_actual = None
    if sucursal_actual_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_actual_id)
        except Sucursal.DoesNotExist:
            pass
    
    # Obtener todas las sucursales del usuario
    empresas_usuario = EmpresaUser.objects.filter(
        user=request.user,
        status=True
    ).values_list('empresa_id', flat=True)
    sucursales = Sucursal.objects.filter(empresa_id__in=empresas_usuario).order_by('alias')
    
    context = {
        'sucursal_actual': sucursal_actual,
        'sucursal_actual_id': sucursal_actual_id,
        'sucursales': sucursales,
    }
    return render(request, 'vistas/modulo_reportes/reporte_movimientos_sucursal.html', context)


@require_GET
@login_required
def obtener_reporte_movimientos_sucursal(request):
    """
    API OPTIMIZADA para reporte de stock inicial (recibido) vs restante por sucursal.
    
    OPTIMIZACIÓN: Usa agregaciones en BD en lugar de queries individuales.
    Reduce de N*M queries a solo 3 queries totales.
    """
    try:
        # ========== PARÁMETROS ==========
        marca_id = request.GET.get('marca_id')
        departamento_id = request.GET.get('departamento_id')
        busqueda = request.GET.get('busqueda', '').strip()
        limite_param = request.GET.get('limite', '')
        limite = int(limite_param) if limite_param else None
        sin_filtro = request.GET.get('sin_filtro', 'false') == 'true'
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()
        # Por defecto el reporte es de TIENDAS: las bodegas proveedoras/CD
        # abastecen, no venden, y sus columnas solo ensuciaban la comparación.
        solo_tiendas = request.GET.get('solo_tiendas', 'true') != 'false'

        tiene_filtro = any([marca_id, departamento_id, busqueda, fecha_desde, fecha_hasta])

        if not tiene_filtro and not sin_filtro:
            return JsonResponse({
                'success': False,
                'requiere_filtro': True,
                'error': 'Por favor selecciona al menos un filtro: Marca, Departamento o busca un artículo.',
                'sugerencia': 'Usa los filtros para optimizar la consulta.'
            })

        # ========== OBTENER SUCURSALES DEL USUARIO ==========
        sucursales_list, bodegas_ocultas = _sucursales_reporte_movimientos(
            request.user, solo_tiendas
        )
        sucursales_ids = [s.id for s in sucursales_list]
        sucursales_map = {s.id: s.alias for s in sucursales_list}

        if not sucursales_ids:
            return JsonResponse({
                'success': True,
                'datos': [],
                'sucursales': [],
                'periodo': {'activo': bool(fecha_desde or fecha_hasta),
                            'desde': fecha_desde or None, 'hasta': fecha_hasta or None},
                'bodegas_ocultas': bodegas_ocultas,
                'error_visible': (
                    'Todas tus sucursales son bodegas proveedoras. Desmarca '
                    '"Solo tiendas" para verlas.' if bodegas_ocultas else None
                ),
                'debug': {'total_productos': 0, 'total_sucursales': 0},
            })

        # ========== QUERY 1: PRODUCTOS BASE (con filtros) ==========
        queryset = Producto.objects.filter(
            sucursal_id__in=sucursales_ids,
            excluir_de_analitica=False
        ).select_related(
            'atributo1', 'atributo2', 'categoria', 'sucursal'
        ).only(
            'id', 'articulo', 'costo', 'precioventa', 'sucursal_id',
            'atributo1__id', 'atributo1__valor',
            'atributo2__id', 'atributo2__valor',
            'categoria__id', 'categoria__nombre',
            'sucursal__id', 'sucursal__alias'
        )

        if marca_id:
            queryset = queryset.filter(atributo1_id=marca_id)

        if departamento_id:
            # Árbol v1.2: un padre incluye todas sus hijas
            queryset = queryset.filter(categoria_id__in=_expandir_categoria_ids(departamento_id))

        if busqueda:
            queryset = queryset.filter(
                Q(articulo__icontains=busqueda) |
                Q(atributo1__valor__icontains=busqueda) |
                Q(atributo2__valor__icontains=busqueda)
            )

        queryset = queryset.order_by('atributo1__valor', 'articulo')
        
        if limite:
            queryset = queryset[:limite]
        
        # Obtener IDs de productos para las siguientes queries
        productos_list = list(queryset)
        productos_ids = [p.id for p in productos_list]
        
        if not productos_ids:
            return JsonResponse({
                'success': True,
                'datos': [],
                'sucursales': [{'id': s.id, 'alias': s.alias} for s in sucursales_list],
                'periodo': {'activo': bool(fecha_desde or fecha_hasta),
                            'desde': fecha_desde or None, 'hasta': fecha_hasta or None},
                'bodegas_ocultas': bodegas_ocultas,
                'debug': {'total_productos': 0, 'total_sucursales': len(sucursales_list)}
            })
        
        # ========== PARSEAR FILTRO DE FECHAS ==========
        filtro_fecha = {}
        if fecha_desde:
            try:
                filtro_fecha['fecha__gte'] = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            except ValueError:
                pass
        if fecha_hasta:
            try:
                filtro_fecha['fecha__lte'] = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            except ValueError:
                pass

        # ========== QUERIES 2/4: MAPAS AGREGADOS POR TIPO DE FLUJO ==========
        # Compras/abastecimiento y traspasos recibidos van SEPARADOS (antes se
        # mezclaban en "recibido"), se agregan los traspasos enviados, y las
        # ventas usan el set canónico de constants_kardex (incluye
        # VENTA_DIRECTA, invisible hasta la auditoría jul-2026).
        compras_map, traspasos_in_map, traspasos_out_map, ventas_map = (
            _mapas_movimientos_sucursal(productos_ids, sucursales_ids, filtro_fecha)
        )

        # ========== SALDOS DEL PERÍODO (kardex real) ==========
        # Con filtro de fechas el reporte deja de ser "recibido histórico vs
        # stock de hoy" y pasa a ser un kardex: saldo inicial + entradas -
        # salidas = saldo final. Ver _saldos_periodo.
        periodo_activo = bool(filtro_fecha)
        entradas_map, salidas_map, posterior_map = _saldos_periodo(
            productos_ids, sucursales_ids, filtro_fecha
        )

        # ========== QUERY 3: STOCK ACTUAL AGREGADO POR PRODUCTO Y SUCURSAL ==========
        stock_query = Producto_Talla.objects.filter(
            producto_id__in=productos_ids
        ).values(
            'producto_id', 'producto__sucursal_id'
        ).annotate(
            total_stock=Sum('stock')
        )

        # Crear mapa de stock: {producto_id: {sucursal_id: stock}}
        stock_map = {}
        for item in stock_query:
            prod_id = item['producto_id']
            suc_id = item['producto__sucursal_id']
            stock = item['total_stock'] or 0

            if prod_id not in stock_map:
                stock_map[prod_id] = {}
            stock_map[prod_id][suc_id] = stock

        # (Las ventas vienen del mismo helper que compras/traspasos — ver
        # _mapas_movimientos_sucursal.)

        # ========== PROCESAR RESULTADOS (sin queries adicionales) ==========
        datos_reporte = []

        for producto in productos_list:
            prod_id = producto.id

            # Obtener datos de este producto desde los mapas
            compras_producto = compras_map.get(prod_id, {})
            tin_producto = traspasos_in_map.get(prod_id, {})
            tout_producto = traspasos_out_map.get(prod_id, {})
            stock_producto = stock_map.get(prod_id, {})
            ventas_producto = ventas_map.get(prod_id, {})
            ent_producto = entradas_map.get(prod_id, {})
            sal_producto = salidas_map.get(prod_id, {})
            post_producto = posterior_map.get(prod_id, {})

            # Construir datos por sucursal
            stock_por_sucursal = {}
            total_compras = total_tin = total_tout = 0
            total_restante = 0
            total_vendido = 0
            total_entradas = total_salidas = total_saldo_inicial = 0
            total_stock_hoy = 0
            total_stock_original = 0

            sucursales_con_datos = (
                set(compras_producto) | set(tin_producto) | set(tout_producto)
                | set(stock_producto) | set(ventas_producto)
                | set(ent_producto) | set(sal_producto) | set(post_producto)
            )

            for suc_id in sucursales_con_datos:
                if suc_id not in sucursales_map:
                    continue

                compras = compras_producto.get(suc_id, 0)
                traspasos_in = tin_producto.get(suc_id, 0)
                traspasos_out = tout_producto.get(suc_id, 0)
                vendido = ventas_producto.get(suc_id, 0)
                inicial = compras + traspasos_in  # compat con la UI actual

                # Kardex del período. Sin filtro de fechas `posterior` es 0 y
                # `restante` sigue siendo el stock de hoy (comportamiento
                # histórico intacto). Con filtro, `restante` pasa a ser el saldo
                # al cierre de la ventana y aparece el saldo inicial.
                stock_hoy = stock_producto.get(suc_id, 0)
                entradas = ent_producto.get(suc_id, 0)
                salidas = sal_producto.get(suc_id, 0)
                restante = stock_hoy - post_producto.get(suc_id, 0)
                saldo_inicial = restante - (entradas - salidas)

                # Las dos columnas que muestra la UI. "Original" cambia de
                # definición según haya período: sin fechas es todo lo que
                # entró alguna vez; con fechas, el saldo del día "Desde".
                # "Actual" es el stock de hoy, o el saldo al cierre si se
                # filtró por fechas.
                stock_original = saldo_inicial if periodo_activo else inicial
                stock_actual = restante

                if (inicial or restante or vendido or traspasos_out
                        or entradas or salidas or saldo_inicial):
                    suc_alias = sucursales_map[suc_id]
                    stock_por_sucursal[suc_alias] = {
                        'sucursal_id': suc_id,
                        'stock_original': stock_original,
                        'stock_actual': stock_actual,
                        'inicial': inicial,
                        'compras': compras,
                        'traspasos_in': traspasos_in,
                        'traspasos_out': traspasos_out,
                        'restante': restante,
                        'vendido': vendido,
                        # Kardex del período (solo con sentido si periodo_activo)
                        'saldo_inicial': saldo_inicial,
                        'entradas': entradas,
                        'salidas': salidas,
                        'stock_hoy': stock_hoy,
                        # Lo que no cae en compras/traspasos/ventas: ajustes,
                        # correcciones, cambios, devoluciones, anulaciones.
                        'otros_entradas': max(entradas - compras - traspasos_in, 0),
                        'otros_salidas': max(salidas - traspasos_out - vendido, 0),
                    }
                    total_compras += compras
                    total_tin += traspasos_in
                    total_tout += traspasos_out
                    total_restante += restante
                    total_vendido += vendido
                    total_entradas += entradas
                    total_salidas += salidas
                    total_saldo_inicial += saldo_inicial
                    total_stock_hoy += stock_hoy
                    total_stock_original += stock_original

            total_inicial = total_compras + total_tin

            # Solo incluir productos con datos
            if stock_por_sucursal or total_inicial > 0 or total_restante > 0:
                datos_reporte.append({
                    'articulo': producto.articulo,
                    'marca': producto.atributo1.valor if producto.atributo1 else 'Sin Marca',
                    'color': producto.atributo2.valor if producto.atributo2 else '-',
                    'departamento': producto.categoria.nombre if producto.categoria else '-',
                    'costo': float(producto.costo) if producto.costo else 0,
                    'precio_venta': float(producto.precioventa) if producto.precioventa else 0,
                    'sucursales': stock_por_sucursal,
                    'total_stock_original': total_stock_original,
                    'total_stock_actual': total_restante,
                    'total_inicial': total_inicial,
                    'total_compras': total_compras,
                    'total_traspasos_in': total_tin,
                    'total_traspasos_out': total_tout,
                    'total_restante': total_restante,
                    'total_vendido': total_vendido,
                    'total_saldo_inicial': total_saldo_inicial,
                    'total_entradas': total_entradas,
                    'total_salidas': total_salidas,
                    'total_stock_hoy': total_stock_hoy,
                })

        sucursales_data = [{'id': s.id, 'alias': s.alias} for s in sucursales_list]

        return JsonResponse({
            'success': True,
            'datos': datos_reporte,
            'sucursales': sucursales_data,
            'bodegas_ocultas': bodegas_ocultas,
            # La UI muestra siempre dos columnas (Original / Actual); esto solo
            # cambia QUÉ significa "Original": sin período es todo lo recibido
            # alguna vez, con período es el saldo del día "Desde".
            'periodo': {
                'activo': periodo_activo,
                'desde': fecha_desde or None,
                'hasta': fecha_hasta or None,
            },
            'debug': {
                'total_productos': len(datos_reporte),
                'total_sucursales': len(sucursales_data),
                'optimizado': True
            }
        })
        
    except Exception as e:
        logger.exception("Error en reporte movimientos por sucursal")
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener reporte: {str(e)}'
        })


@require_GET
@login_required
def exportar_movimientos_sucursal_excel(request):
    """
    Exportar reporte de movimientos por sucursal a Excel.
    OPTIMIZADO: Usa agregaciones en BD (3 queries totales).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Parámetros
        marca_id = request.GET.get('marca_id')
        departamento_id = request.GET.get('departamento_id')
        busqueda = request.GET.get('busqueda', '').strip()
        limite = request.GET.get('limite')
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()
        solo_tiendas = request.GET.get('solo_tiendas', 'true') != 'false'

        # Parsear filtro de fechas
        filtro_fecha = {}
        if fecha_desde:
            try:
                filtro_fecha['fecha__gte'] = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            except ValueError:
                pass
        if fecha_hasta:
            try:
                filtro_fecha['fecha__lte'] = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Obtener sucursales del usuario (mismo criterio que la vista web: con
        # `solo_tiendas` las bodegas proveedoras quedan fuera del export).
        sucursales_list, bodegas_ocultas = _sucursales_reporte_movimientos(
            request.user, solo_tiendas
        )
        sucursales_ids = [s.id for s in sucursales_list]
        sucursales_map = {s.id: s.alias for s in sucursales_list}

        # QUERY 1: Productos base
        queryset = Producto.objects.filter(
            sucursal_id__in=sucursales_ids,
            excluir_de_analitica=False
        ).select_related(
            'atributo1', 'atributo2', 'categoria', 'sucursal'
        ).only(
            'id', 'articulo', 'costo', 'precioventa', 'sucursal_id',
            'atributo1__id', 'atributo1__valor',
            'atributo2__id', 'atributo2__valor',
            'categoria__id', 'categoria__nombre',
            'sucursal__id', 'sucursal__alias'
        )

        if marca_id:
            queryset = queryset.filter(atributo1_id=marca_id)

        if departamento_id:
            # Árbol v1.2: un padre incluye sus hijas, igual que la vista web.
            # Sin esto el Excel traía menos filas que la pantalla al elegir un
            # departamento padre.
            queryset = queryset.filter(categoria_id__in=_expandir_categoria_ids(departamento_id))

        if busqueda:
            queryset = queryset.filter(
                Q(articulo__icontains=busqueda) |
                Q(atributo1__valor__icontains=busqueda) |
                Q(atributo2__valor__icontains=busqueda)
            )
        
        queryset = queryset.order_by('atributo1__valor', 'articulo')
        
        if limite:
            queryset = queryset[:int(limite)]
        
        productos_list = list(queryset)
        productos_ids = [p.id for p in productos_list]
        
        # QUERY 2/4: mapas por tipo de flujo (mismo helper que la vista API)
        compras_map, traspasos_in_map, traspasos_out_map, ventas_map = (
            _mapas_movimientos_sucursal(productos_ids, sucursales_ids, filtro_fecha)
            if productos_ids else ({}, {}, {}, {})
        )

        # Saldos del período: el Excel tiene que decir lo mismo que la web, o
        # el usuario exporta un "Rest." de hoy junto a un "Recib." de la ventana.
        periodo_activo = bool(filtro_fecha)
        entradas_map, salidas_map, posterior_map = (
            _saldos_periodo(productos_ids, sucursales_ids, filtro_fecha)
            if productos_ids else ({}, {}, {})
        )

        # QUERY 3: Stock actual agregado
        stock_map = {}
        if productos_ids:
            stock_query = Producto_Talla.objects.filter(
                producto_id__in=productos_ids
            ).values(
                'producto_id', 'producto__sucursal_id'
            ).annotate(
                total_stock=Sum('stock')
            )

            for item in stock_query:
                prod_id = item['producto_id']
                suc_id = item['producto__sucursal_id']
                if prod_id not in stock_map:
                    stock_map[prod_id] = {}
                stock_map[prod_id][suc_id] = item['total_stock'] or 0

        # (Ventas incluidas en el helper de mapas.)

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos Sucursal"
        
        # Estilos
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        inicial_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        restante_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        total_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "STOCK ORIGINAL VS STOCK ACTUAL POR SUCURSAL"
        ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        ws['A1'].fill = header_fill

        # Info
        leyenda_periodo = (
            f" | Período {fecha_desde or 'inicio'} → {fecha_hasta or 'hoy'}"
            " | Original = saldo del día Desde; Actual = saldo al cierre"
            if periodo_activo else
            " | Original = todo lo recibido alguna vez; Actual = stock de HOY"
        )
        leyenda_bodegas = (f" | {bodegas_ocultas} bodega(s) proveedora(s) excluida(s)"
                           if bodegas_ocultas else "")
        ws['A2'] = (f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | "
                    f"{len(productos_list)} productos{leyenda_periodo}{leyenda_bodegas}")
        ws['A2'].font = Font(italic=True, size=9)

        # Headers fila 4. Por sucursal solo van las dos columnas que muestra la
        # pantalla; el desglose de movimientos (entradas/salidas/traspasos) se
        # conserva a nivel de TOTAL para no perder la trazabilidad del período.
        row = 4
        headers = ['Artículo', 'Marca', 'Color', 'Departamento', 'Costo', 'Precio Venta']

        for suc in [s.alias for s in sucursales_list]:
            headers.extend([f'{suc} Original', f'{suc} Actual'])
        headers.extend(['TOTAL Original', 'TOTAL Actual', 'TOTAL Entradas',
                        'TOTAL Salidas', 'VENDIDO'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Datos (sin queries adicionales)
        row = 5
        for producto in productos_list:
            prod_id = producto.id
            compras_prod = compras_map.get(prod_id, {})
            tin_prod = traspasos_in_map.get(prod_id, {})
            tout_prod = traspasos_out_map.get(prod_id, {})
            stock_prod = stock_map.get(prod_id, {})
            ventas_prod = ventas_map.get(prod_id, {})

            # Calcular totales
            total_compras = sum(compras_prod.values())
            total_tin = sum(tin_prod.values())
            total_tout = sum(tout_prod.values())
            total_vendido = sum(ventas_prod.values())

            ent_prod = entradas_map.get(prod_id, {})
            sal_prod = salidas_map.get(prod_id, {})
            post_prod = posterior_map.get(prod_id, {})
            # Saldo al cierre = stock de hoy rebobinado por lo posterior al corte.
            saldo_fin_prod = {
                suc.id: stock_prod.get(suc.id, 0) - post_prod.get(suc.id, 0)
                for suc in sucursales_list
            }
            saldo_ini_prod = {
                suc.id: (saldo_fin_prod[suc.id]
                         - (ent_prod.get(suc.id, 0) - sal_prod.get(suc.id, 0)))
                for suc in sucursales_list
            }
            total_restante = sum(saldo_fin_prod.values())
            total_entradas = sum(ent_prod.values())
            total_salidas = sum(sal_prod.values())

            # Misma definición que la API: con período "original" es el saldo
            # del día Desde; sin período, todo lo que entró alguna vez.
            def _original(suc_id):
                if periodo_activo:
                    return saldo_ini_prod[suc_id]
                return compras_prod.get(suc_id, 0) + tin_prod.get(suc_id, 0)

            total_original = sum(_original(s.id) for s in sucursales_list)

            if (total_compras > 0 or total_tin > 0 or total_tout > 0
                    or total_restante > 0 or total_vendido > 0
                    or total_entradas > 0 or total_salidas > 0):
                col = 1

                ws.cell(row=row, column=col, value=producto.articulo).border = border
                col += 1
                ws.cell(row=row, column=col, value=producto.atributo1.valor if producto.atributo1 else '-').border = border
                col += 1
                ws.cell(row=row, column=col, value=producto.atributo2.valor if producto.atributo2 else '-').border = border
                col += 1
                ws.cell(row=row, column=col, value=producto.categoria.nombre if producto.categoria else '-').border = border
                col += 1

                costo_val = float(producto.costo) if producto.costo else 0
                cell_costo = ws.cell(row=row, column=col, value=costo_val)
                cell_costo.border = border
                cell_costo.number_format = '"$"#,##0'
                cell_costo.alignment = Alignment(horizontal='right')
                col += 1

                precio_val = float(producto.precioventa) if producto.precioventa else 0
                cell_precio = ws.cell(row=row, column=col, value=precio_val)
                cell_precio.border = border
                cell_precio.number_format = '"$"#,##0'
                cell_precio.alignment = Alignment(horizontal='right')
                cell_precio.font = Font(bold=True, color="00875A")
                col += 1

                # Datos por sucursal (desde los mapas, sin queries)
                for suc in sucursales_list:
                    for valor, relleno in (
                        (_original(suc.id), inicial_fill),
                        (saldo_fin_prod[suc.id], restante_fill),
                    ):
                        celda = ws.cell(row=row, column=col, value=valor)
                        celda.border = border
                        celda.fill = relleno
                        celda.alignment = Alignment(horizontal='center')
                        col += 1

                # Totales
                for valor, color_fuente in (
                    (total_original, None),
                    (total_restante, None),
                    (total_entradas, None),
                    (total_salidas, None),
                    (total_vendido, 'CC0000'),
                ):
                    celda = ws.cell(row=row, column=col, value=valor)
                    celda.border = border
                    celda.fill = total_fill
                    celda.font = (Font(bold=True, color=color_fuente)
                                  if color_fuente else Font(bold=True))
                    celda.alignment = Alignment(horizontal='center')
                    col += 1

                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13

        for col in range(7, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_str = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="stock_inicial_restante_{fecha_str}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Error exportando stock inicial restante a Excel")
        return JsonResponse({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        }, status=500)


# ========== REPORTES MEJORADOS: RECEPCIONES Y DESPACHOS ==========

@require_GET
@login_required
def api_reporte_recepciones_detallado(request):
    """
    API mejorada para reporte de recepciones con:
    - Datos reales de Productos_Recepcionados (no solo DTEs)
    - Desglose por sucursal destino
    - Split reposicion vs compra nueva
    - Diferencias de precio detectadas

    HUÉRFANO (jul-2026): ningún template ni JS llama a esta URL. Se le aplica
    igual el alcance por empresa porque la ruta sigue publicada y el middleware
    de permisos es fail-open; queda pendiente decidir si se elimina.

    SCOPING: el universo se acota con las mismas 3 puntas del reporte de
    diferencias (emisor / receptor del DTE, o sucursal donde aterrizó la
    mercadería). Antes no había ningún filtro por empresa.
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        proveedor_id = request.GET.get('proveedor_id')
        sucursal_id = request.GET.get('sucursal_id')

        _permitidas, sin_acceso = _alcance_sucursales(request, sucursal_id)
        if sin_acceso:
            return sin_acceso

        if not fecha_inicio or not fecha_fin:
            fecha_fin_dt = timezone.localdate()
            fecha_inicio_dt = fecha_fin_dt - timedelta(days=30)
        else:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

        # Para compras históricas vinculadas retroactivamente, fecha_recepcion
        # se setea con compra.fecha (fecha real). Filtramos por ese campo
        # cuando existe, y caemos a `fecha` (auto_now) para registros legados.
        recepciones_qs = Productos_Recepcionados.objects.filter(
            _q_alcance_recepcion(request.user),
            Q(fecha_recepcion__date__range=[fecha_inicio_dt, fecha_fin_dt]) |
            (Q(fecha_recepcion__isnull=True) & Q(fecha__range=[fecha_inicio_dt, fecha_fin_dt]))
        ).select_related(
            'compra_producto_talla__compra_producto__compras__empresa',
            'producto_talla__producto',
            'dte__emisor',
            'sucursal_destino',
        )

        if proveedor_id:
            # Match por proveedor en DTE o en la compra vinculada (cubre históricas sin DTE).
            recepciones_qs = recepciones_qs.filter(
                Q(dte__emisor_id=proveedor_id) |
                Q(compra_producto_talla__compra_producto__compras__empresa_id=proveedor_id)
            )
        if sucursal_id:
            recepciones_qs = recepciones_qs.filter(sucursal_destino_id=sucursal_id)

        # --- Totals ---
        total_items = recepciones_qs.count()
        total_unidades = recepciones_qs.aggregate(t=Sum('stockArribado'))['t'] or 0
        total_reposicion = recepciones_qs.filter(es_reposicion=True).count()
        total_nuevo = recepciones_qs.filter(es_reposicion=False).count()
        total_historicas = recepciones_qs.filter(es_historica=True).count()
        total_con_cambio_precio = recepciones_qs.exclude(
            precio_anterior__isnull=True,
        ).exclude(
            precio_nuevo__isnull=True,
        ).exclude(
            precio_anterior=F('precio_nuevo'),
        ).count()

        # --- By sucursal ---
        por_sucursal = (
            recepciones_qs
            .values('sucursal_destino__alias', 'sucursal_destino_id')
            .annotate(
                items=Count('id'),
                unidades=Sum('stockArribado'),
                reposiciones=Count('id', filter=Q(es_reposicion=True)),
                nuevos=Count('id', filter=Q(es_reposicion=False)),
            )
            .order_by('-unidades')
        )
        por_sucursal_data = [
            {
                'sucursal': row['sucursal_destino__alias'] or 'Sin asignar',
                'sucursal_id': row['sucursal_destino_id'],
                'items': row['items'],
                'unidades': row['unidades'] or 0,
                'reposiciones': row['reposiciones'],
                'nuevos': row['nuevos'],
            }
            for row in por_sucursal
        ]

        # --- By proveedor ---
        # Rama 1: recepciones con DTE → proveedor es dte.emisor
        por_proveedor_dte = (
            recepciones_qs
            .filter(dte__isnull=False)
            .values('dte__emisor__nombre', 'dte__emisor__rut')
            .annotate(
                items=Count('id'),
                unidades=Sum('stockArribado'),
                reposiciones=Count('id', filter=Q(es_reposicion=True)),
                nuevos=Count('id', filter=Q(es_reposicion=False)),
            )
        )
        # Rama 2: recepciones sin DTE (típico de compras históricas vinculadas)
        # → proveedor es la empresa de la compra.
        por_proveedor_compra = (
            recepciones_qs
            .filter(dte__isnull=True, compra_producto_talla__isnull=False)
            .values(
                'compra_producto_talla__compra_producto__compras__empresa__nombre',
                'compra_producto_talla__compra_producto__compras__empresa__rut',
            )
            .annotate(
                items=Count('id'),
                unidades=Sum('stockArribado'),
                reposiciones=Count('id', filter=Q(es_reposicion=True)),
                nuevos=Count('id', filter=Q(es_reposicion=False)),
            )
        )

        # Merge por nombre+rut acumulando contadores.
        merged_proveedor = {}
        for row in por_proveedor_dte:
            key = (row['dte__emisor__nombre'], row['dte__emisor__rut'])
            if not key[0]:
                continue
            merged_proveedor[key] = {
                'proveedor': key[0],
                'rut': key[1],
                'items': row['items'],
                'unidades': row['unidades'] or 0,
                'reposiciones': row['reposiciones'],
                'nuevos': row['nuevos'],
            }
        for row in por_proveedor_compra:
            nombre = row['compra_producto_talla__compra_producto__compras__empresa__nombre']
            rut = row['compra_producto_talla__compra_producto__compras__empresa__rut']
            if not nombre:
                continue
            key = (nombre, rut)
            if key in merged_proveedor:
                bucket = merged_proveedor[key]
                bucket['items'] += row['items']
                bucket['unidades'] += row['unidades'] or 0
                bucket['reposiciones'] += row['reposiciones']
                bucket['nuevos'] += row['nuevos']
            else:
                merged_proveedor[key] = {
                    'proveedor': nombre,
                    'rut': rut,
                    'items': row['items'],
                    'unidades': row['unidades'] or 0,
                    'reposiciones': row['reposiciones'],
                    'nuevos': row['nuevos'],
                }
        por_proveedor_data = sorted(
            merged_proveedor.values(),
            key=lambda d: d['unidades'],
            reverse=True,
        )

        # --- Cambios de precio detectados ---
        cambios_precio = []
        for rec in recepciones_qs.exclude(
            precio_anterior__isnull=True,
        ).exclude(
            precio_nuevo__isnull=True,
        ).exclude(
            precio_anterior=F('precio_nuevo'),
        ).order_by('-fecha')[:50]:
            cambios_precio.append({
                'fecha': rec.fecha.strftime('%d/%m/%Y') if rec.fecha else '',
                'sku': rec.producto_talla.sku if rec.producto_talla else '',
                'producto': rec.producto_talla.producto.articulo if rec.producto_talla and rec.producto_talla.producto else '',
                'precio_anterior': rec.precio_anterior,
                'precio_nuevo': rec.precio_nuevo,
                'diferencia': (rec.precio_nuevo or 0) - (rec.precio_anterior or 0),
                'sucursal': rec.sucursal_destino.alias if rec.sucursal_destino else '',
                'proveedor': rec.dte.emisor.nombre if rec.dte and rec.dte.emisor else '',
            })

        return JsonResponse({
            'success': True,
            'resumen': {
                'total_items': total_items,
                'total_unidades': total_unidades,
                'total_reposicion': total_reposicion,
                'total_nuevo': total_nuevo,
                'total_historicas': total_historicas,
                'total_con_cambio_precio': total_con_cambio_precio,
            },
            'por_sucursal': por_sucursal_data,
            'por_proveedor': por_proveedor_data,
            'cambios_precio': cambios_precio,
            'parametros': {
                'fecha_inicio': fecha_inicio_dt.strftime('%d/%m/%Y'),
                'fecha_fin': fecha_fin_dt.strftime('%d/%m/%Y'),
            },
        })

    except Exception as e:
        logger.exception("Error al generar reporte de recepciones detallado")
        return JsonResponse({
            'success': False,
            'error': str(e),
        }, status=500)


@require_GET
@login_required
def api_reporte_despachos_detallado(request):
    """
    API mejorada para despachos con datos reales de recepción,
    desglose por sucursal destino y split reposición/nuevo.

    HUÉRFANO (jul-2026): ningún template ni JS llama a esta URL. Se le aplica
    igual el alcance por empresa porque la ruta sigue publicada; queda
    pendiente decidir si se elimina.

    SCOPING: los DTE se acotan al alcance del usuario (`_q_alcance_dte`) y las
    recepciones de cada documento a las sucursales de sus empresas. `proveedor_id`
    filtra por EMISOR (el proveedor externo), así que no se valida contra el
    alcance: quien acota es el universo de DTE, no el parámetro.
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        proveedor_id = request.GET.get('proveedor_id')

        permitidas = ids_sucursales_alcance(request.user)

        if not fecha_inicio or not fecha_fin:
            fecha_fin_dt = timezone.localdate()
            fecha_inicio_dt = fecha_fin_dt - timedelta(days=30)
        else:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

        dtes_compra = Dte.objects.filter(
            _q_alcance_dte(request.user),
            tipo_transaccion='COMPRA',
            fecha_emision__range=[fecha_inicio_dt, fecha_fin_dt],
        ).select_related('emisor')

        if proveedor_id:
            dtes_compra = dtes_compra.filter(emisor_id=proveedor_id)

        despachos = []
        for dte in dtes_compra.order_by('-fecha_emision'):
            # OJO nombres reales: Dte_Productos.productoTalla (camelCase) y la
            # cantidad de línea es `stock` — con 'producto_talla'/'cantidad'
            # este reporte devolvía 500 siempre (detectado por la suite
            # _test_reportes_readonly, jul-2026).
            productos_dte = Dte_Productos.objects.filter(dte=dte).select_related(
                'productoTalla__producto',
            )
            recepciones = Productos_Recepcionados.objects.filter(
                dte=dte,
            ).select_related('sucursal_destino', 'producto_talla__producto')
            if permitidas is not None:
                # Un DTE del alcance puede tener recepciones repartidas en
                # sucursales de otra empresa; el desglose sólo muestra las propias.
                recepciones = recepciones.filter(sucursal_destino_id__in=permitidas)

            total_esperado = sum(p.stock or 0 for p in productos_dte)
            total_recibido = sum(r.stockArribado or 0 for r in recepciones)

            sucursales_destino = {}
            for rec in recepciones:
                suc_name = rec.sucursal_destino.alias if rec.sucursal_destino else 'Sin asignar'
                if suc_name not in sucursales_destino:
                    sucursales_destino[suc_name] = {'unidades': 0, 'reposicion': 0, 'nuevo': 0}
                sucursales_destino[suc_name]['unidades'] += rec.stockArribado
                if rec.es_reposicion:
                    sucursales_destino[suc_name]['reposicion'] += rec.stockArribado
                else:
                    sucursales_destino[suc_name]['nuevo'] += rec.stockArribado

            despachos.append({
                'dte_id': dte.id,
                'numero_dte': dte.numero_documento,
                'tipo_documento': dte.tipo_documento,
                'fecha_emision': dte.fecha_emision.strftime('%d/%m/%Y') if dte.fecha_emision else '',
                'proveedor': dte.emisor.nombre if dte.emisor else '',
                'rut_proveedor': dte.emisor.rut if dte.emisor else '',
                'total': float(dte.monto_con_iva or 0),
                'estado': dte.estado_dte,
                'total_esperado': total_esperado,
                'total_recibido': total_recibido,
                'porcentaje_recepcion': round(total_recibido / total_esperado * 100, 1) if total_esperado > 0 else 0,
                'sucursales_destino': [
                    {'sucursal': k, **v} for k, v in sucursales_destino.items()
                ],
            })

        resumen = {
            'total_documentos': len(despachos),
            'total_esperado': sum(d['total_esperado'] for d in despachos),
            'total_recibido': sum(d['total_recibido'] for d in despachos),
            'monto_total': sum(d['total'] for d in despachos),
        }
        if resumen['total_esperado'] > 0:
            resumen['porcentaje_recepcion_global'] = round(
                resumen['total_recibido'] / resumen['total_esperado'] * 100, 1
            )
        else:
            resumen['porcentaje_recepcion_global'] = 0

        return JsonResponse({
            'success': True,
            'despachos': despachos,
            'resumen': resumen,
            'parametros': {
                'fecha_inicio': fecha_inicio_dt.strftime('%d/%m/%Y'),
                'fecha_fin': fecha_fin_dt.strftime('%d/%m/%Y'),
            },
        })

    except Exception as e:
        logger.exception("Error al generar reporte de despachos detallado")
        return JsonResponse({
            'success': False,
            'error': str(e),
        }, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE: RENDIMIENTO POR PROVEEDOR  (Compra → Recepción → Venta a público)
# ══════════════════════════════════════════════════════════════════════════════
#
# ATRIBUCIÓN DE LA VENTA — REGLA CANÓNICA (documentada a propósito)
# ----------------------------------------------------------------
# Antes la venta se atribuía por COINCIDENCIA DE TEXTO entre
# `Compras_Producto.nombre` (lo que trae el formato de importación de la OC) y
# `Producto.articulo` (el código del catálogo). Eso producía dos errores
# opuestos y ambos malos: ceros falsos, y —cuando dos proveedores compartían un
# nombre— la MISMA venta contada en los dos, con lo que los totales de venta y
# margen no eran sumables y el "% venta global" podía superar el 100%.
#
# Ahora el origen de un SKU se resuelve por la relación REAL del sistema, la
# misma que usa el resto del módulo de compras:
#
#     Compras → Compras_Producto → Compras_Producto_Talla
#            → Productos_Recepcionados.producto_talla   (enlace preferido)
#       o     Compras_Producto_Talla.producto_talla     (enlace de respaldo)
#
# Un detalle del modelo obliga a un paso más: `Producto` tiene FK a `Sucursal`,
# así que el mismo artículo físico existe como un `Producto_Talla` DISTINTO en
# cada sucursal. La recepción enlaza el SKU del centro de distribución, pero la
# venta a público ocurre sobre el SKU de la tienda (después del traspaso): medido
# contra producción, atribuir por `producto_talla_id` exacto da CERO ventas para
# todos los proveedores. Por eso la atribución se hace a nivel de IDENTIDAD DE
# ARTÍCULO —`Producto.articulo` normalizado + marca (`Producto.atributo1`)—, que
# es la clave de identidad que ya usa la deduplicación de productos.
#
# DESEMPATE (un artículo comprado a más de un proveedor): gana el proveedor de la
# RECEPCIÓN MÁS RECIENTE (`fecha_recepcion`, con fallback a `fecha`); si el
# artículo no tiene ninguna recepción en el universo consultado, gana el de la
# ORDEN DE COMPRA más reciente; a igualdad de fecha, el id de proveedor mayor
# (criterio arbitrario pero estable, para que el reporte no cambie entre cargas).
# Cada artículo pertenece así a UN solo proveedor: las filas son disjuntas y los
# totales SÍ son sumables.
#
# Lo que no se puede atribuir NO se adivina: se informa aparte como
# `unidades_venta_sin_atribuir` / `cobertura_atribucion_pct`.
#
# OTRAS REGLAS
# ------------
# · Órdenes: solo estados vigentes (ACTIVA/COMPLETADA) → ELIMINADA y CANCELADA
#   quedan fuera, igual que en el resto del módulo.
# · Scoping: el universo se acota a las empresas del usuario (EmpresaUser
#   status=True vía `obtener_empresas_usuario`). Los DTE de compra por su
#   RECEPTOR; las órdenes por la empresa de la sucursal donde se recepcionó,
#   porque `Compras.empresa` es el PROVEEDOR y el modelo no tiene empresa
#   compradora.
# · "No inventariable" excluye notas de crédito, descartados y anulados: antes
#   una NC de proveedor sumaba en positivo a esa columna.
# · Ventas: solo tickets PAGADO (un PENDIENTE aún no es venta), por
#   `created_at` (Ticket.fecha es auto_now) y sin productos marcados
#   `excluir_de_analitica`.

# Módulos de origen que cuentan como "venta a público" en este reporte.
RP_MODULOS_VENTA_PUBLICO = ['VENTA_PUBLICO', 'POS', 'ECOMMERCE']
# Sentinela para ordenar registros sin fecha sin tener que comparar None.
RP_FECHA_MINIMA = datetime(1900, 1, 1).date()


@login_required
def ver_reporte_rendimiento_proveedor(request):
    return render(request, 'vistas/modulo_reportes/reporte_rendimiento_proveedor.html')


def _rp_filtros(request):
    """Resuelve UNA vez los filtros del request y el universo de empresas."""
    from .utils_permisos import obtener_empresas_usuario

    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get('anio') or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year

    def _fecha(clave):
        valor = (request.GET.get(clave) or '').strip()
        if not valor:
            return None
        try:
            return datetime.strptime(valor, '%Y-%m-%d').date()
        except ValueError:
            return None

    proveedor_id = (request.GET.get('proveedor_id') or '').strip()
    sucursal_id = (request.GET.get('sucursal_id') or '').strip()

    # `empresas_ids = None` significa "sin restricción": para un administrador
    # materializar la lista serían ~1.700 ids viajando dentro del SQL sin acotar
    # nada, porque `obtener_empresas_usuario` ya le devuelve todas.
    ve_todas = usuario_puede_ver_todas_sucursales(request.user)
    empresas_ids = None if ve_todas else list(
        obtener_empresas_usuario(request.user).values_list('id', flat=True)
    )

    return {
        'anio': anio,
        'fecha_inicio': _fecha('fecha_inicio'),
        'fecha_fin': _fecha('fecha_fin'),
        'proveedor_id': int(proveedor_id) if proveedor_id.isdigit() else None,
        'sucursal_id': int(sucursal_id) if sucursal_id.isdigit() else None,
        'empresas_ids': empresas_ids,
        've_todas': ve_todas,
    }


def _rp_q_periodo_recepcion(filtros):
    """
    Q sobre Productos_Recepcionados para el período pedido.

    Las vinculaciones retroactivas llevan la fecha real en `fecha_recepcion`;
    las recepciones antiguas pueden tenerla NULL y hay que caer a `fecha`.
    Con rango explícito manda el rango (antes el universo cruzaba año Y rango,
    así que un rango de otro año devolvía siempre vacío).
    """
    inicio, fin = filtros['fecha_inicio'], filtros['fecha_fin']
    if not inicio and not fin:
        return (Q(fecha_recepcion__year=filtros['anio']) |
                (Q(fecha_recepcion__isnull=True) & Q(fecha__year=filtros['anio'])))
    condicion = Q()
    if inicio:
        condicion &= (Q(fecha_recepcion__date__gte=inicio) |
                      (Q(fecha_recepcion__isnull=True) & Q(fecha__gte=inicio)))
    if fin:
        condicion &= (Q(fecha_recepcion__date__lte=fin) |
                      (Q(fecha_recepcion__isnull=True) & Q(fecha__lte=fin)))
    return condicion


def _rp_compras_scope(filtros):
    """
    Órdenes de compra vigentes del período, ya scopeadas por empresa.

    Universo = órdenes del año pedido ∪ órdenes con alguna recepción dentro del
    período (una OC del año pasado que llega este año debe medirse aquí).
    Los ids viajan como subconsulta, no como lista Python: antes se
    materializaban dos sets y se mandaban como IN gigante.
    """
    qs = Compras.objects.filter(estado__in=ESTADOS_COMPRA_VIGENTES)
    if filtros['proveedor_id']:
        qs = qs.filter(empresa_id=filtros['proveedor_id'])

    con_recepcion = (Productos_Recepcionados.objects
                     .filter(_rp_q_periodo_recepcion(filtros),
                             compra_producto_talla__isnull=False)
                     .values('compra_producto_talla__compra_producto__compras_id'))
    qs = qs.filter(Q(fecha__year=filtros['anio']) | Q(id__in=con_recepcion))

    if not filtros['ve_todas']:
        # `Compras.empresa` es el PROVEEDOR: la orden no tiene empresa
        # compradora, así que el único anclaje real al holding del usuario es la
        # sucursal donde se recepcionó la mercadería.
        propias = (Productos_Recepcionados.objects
                   .filter(compra_producto_talla__isnull=False,
                           sucursal_destino__empresa_id__in=filtros['empresas_ids'])
                   .values('compra_producto_talla__compra_producto__compras_id'))
        qs = qs.filter(id__in=propias)
    return qs


def _rp_clave_articulo(articulo, marca_id):
    """Identidad de artículo usada para atribuir: código normalizado + marca."""
    if not articulo:
        return None
    return ((articulo or '').strip().upper(), marca_id)


def _rp_mapa_articulo_proveedor(compras_qs):
    """
    Resuelve el proveedor de origen de cada artículo, en 3 queries acotadas.

    Devuelve `(mapa_enlace, mapa_codigo, multiproveedor)`:

    · `mapa_enlace` — {(articulo, marca_id): proveedor_id} construido con el
      enlace REAL del sistema (recepción, y en su defecto la línea de la OC).
      Es la fuente fuerte y la que manda siempre.
    · `mapa_codigo` — {articulo: proveedor_id} de respaldo, construido con el
      nombre de la línea de OC cuando ese nombre ES un código de catálogo. Solo
      se puebla para códigos que NO tienen enlace real: medido contra
      producción, 3.537 de 6.085 líneas de OC del período todavía no tienen
      `producto_talla` (mercadería pedida y aún no recepcionada), y sin este
      respaldo el reporte perdería ~1.200 unidades vendidas que sí son de esos
      proveedores. A diferencia del match por texto anterior, este respaldo es
      DISJUNTO: cada código va a un único proveedor.
    · `multiproveedor` — cuántos artículos tenían más de un proveedor candidato,
      para poder decirle al usuario que ahí se aplicó una regla y no una suma.

    Desempate en ambos mapas: gana la recepción/orden más reciente y, a igualdad
    de fecha, el id de proveedor mayor (arbitrario pero estable).
    """
    mapa_enlace = {}
    ganador_enlace = {}
    mapa_codigo = {}
    ganador_codigo = {}
    candidatos = defaultdict(set)

    def _postular(destino, ganadores, clave, proveedor_id, prioridad, fecha):
        if not clave or not proveedor_id:
            return
        candidatos[clave].add(proveedor_id)
        marca = (prioridad, fecha or RP_FECHA_MINIMA, proveedor_id)
        if clave not in ganadores or marca > ganadores[clave]:
            ganadores[clave] = marca
            destino[clave] = proveedor_id

    # Prioridad 1: recepción real (es el hecho físico, y la más reciente manda).
    recepciones = (Productos_Recepcionados.objects
                   .filter(compra_producto_talla__compra_producto__compras__in=compras_qs,
                           producto_talla__isnull=False)
                   .values('producto_talla__producto__articulo',
                           'producto_talla__producto__atributo1_id',
                           'compra_producto_talla__compra_producto__compras__empresa_id',
                           'fecha_recepcion', 'fecha'))
    for fila in recepciones:
        fecha = fila['fecha_recepcion']
        fecha = timezone.localtime(fecha).date() if fecha else fila['fecha']
        _postular(
            mapa_enlace, ganador_enlace,
            _rp_clave_articulo(fila['producto_talla__producto__articulo'],
                               fila['producto_talla__producto__atributo1_id']),
            fila['compra_producto_talla__compra_producto__compras__empresa_id'],
            1, fecha,
        )

    # Prioridad 0: enlace de la orden, sin recepción registrada todavía.
    lineas_oc = (Compras_Producto_Talla.objects
                 .filter(compra_producto__compras__in=compras_qs,
                         producto_talla__isnull=False)
                 .values('producto_talla__producto__articulo',
                         'producto_talla__producto__atributo1_id',
                         'compra_producto__compras__empresa_id',
                         'compra_producto__compras__fecha'))
    for fila in lineas_oc:
        _postular(
            mapa_enlace, ganador_enlace,
            _rp_clave_articulo(fila['producto_talla__producto__articulo'],
                               fila['producto_talla__producto__atributo1_id']),
            fila['compra_producto__compras__empresa_id'],
            0, fila['compra_producto__compras__fecha'],
        )

    # Respaldo: el nombre de la línea de OC como código de artículo.
    for fila in (Compras_Producto.objects
                 .filter(compras__in=compras_qs)
                 .values('nombre', 'compras__empresa_id', 'compras__fecha')):
        codigo = (fila['nombre'] or '').strip().upper()
        if not codigo:
            continue
        _postular(mapa_codigo, ganador_codigo, codigo,
                  fila['compras__empresa_id'], 0, fila['compras__fecha'])

    # El enlace real manda: si un código ya tiene enlace, el respaldo se descarta.
    con_enlace = {clave[0] for clave in mapa_enlace}
    mapa_codigo = {codigo: proveedor_id for codigo, proveedor_id in mapa_codigo.items()
                   if codigo not in con_enlace}

    multiproveedor = sum(1 for provs in candidatos.values() if len(provs) > 1)
    return mapa_enlace, mapa_codigo, multiproveedor


def _rp_ventas_por_articulo(filtros):
    """
    Venta a público del período agrupada por identidad de artículo, en 1 query.

    Se agrupa por (artículo, marca) y NO se filtra por una lista de SKU en
    Python: eso era el `IN` de decenas de miles de ids que hacía caer el reporte.
    """
    condiciones = {
        'idTicket__estado': 'PAGADO',
        'idTicket__modulo_origen__in': RP_MODULOS_VENTA_PUBLICO,
        'ProductoTalla__isnull': False,
    }
    inicio, fin = filtros['fecha_inicio'], filtros['fecha_fin']
    if inicio and fin:
        condiciones['idTicket__created_at__date__range'] = [inicio, fin]
    elif inicio:
        condiciones['idTicket__created_at__date__gte'] = inicio
    elif fin:
        condiciones['idTicket__created_at__date__lte'] = fin
    else:
        condiciones['idTicket__created_at__year'] = filtros['anio']
    if filtros['sucursal_id']:
        condiciones['idTicket__sucursal_id'] = filtros['sucursal_id']
    if filtros['empresas_ids'] is not None:
        # Un usuario acotado no puede ver la facturación de las tiendas de otra
        # empresa del holding. Antes el reporte no scopeaba la venta en absoluto.
        condiciones['idTicket__sucursal__empresa_id__in'] = filtros['empresas_ids']

    return (Ticket_Productos.objects
            .filter(**condiciones)
            .exclude(ProductoTalla__producto__excluir_de_analitica=True)
            .values('ProductoTalla__producto__articulo',
                    'ProductoTalla__producto__atributo1_id')
            .annotate(unidades=Sum('stock'), venta=Sum('subtotal')))


def _rp_no_inventariables(filtros):
    """
    DTE de compra sin ninguna línea con Producto_Talla (compra por concepto),
    agrupados por emisor en 1 query.

    Se excluyen notas de crédito, descartados y anulados: una NC no tiene líneas
    inventariables y antes caía aquí SUMANDO, inflando lo que el usuario lee como
    "gasto por concepto de este proveedor".
    """
    qs = Dte.objects.filter(tipo_transaccion='COMPRA')
    if filtros['empresas_ids'] is not None:
        qs = qs.filter(receptor_id__in=filtros['empresas_ids'])
    inicio, fin = filtros['fecha_inicio'], filtros['fecha_fin']
    if inicio:
        qs = qs.filter(fecha_emision__gte=inicio)
    if fin:
        qs = qs.filter(fecha_emision__lte=fin)
    if not inicio and not fin:
        qs = qs.filter(fecha_emision__year=filtros['anio'])
    if filtros['proveedor_id']:
        qs = qs.filter(emisor_id=filtros['proveedor_id'])

    qs = (qs.exclude(descartado=True)
            .exclude(estado_dte__in=ESTADOS_DTE_ANULADOS_COMPRAS)
            .exclude(es_nota_credito=True)
            .exclude(tipo_documento=TIPO_DOC_NC_COMPRAS)
            .exclude(dte_productos__productoTalla__isnull=False))

    return qs.values('emisor_id').annotate(documentos=Count('id'),
                                           monto=Sum('monto_con_iva'))


@require_GET
@login_required
def api_reporte_rendimiento_proveedor(request):
    """
    Métricas por proveedor: comprados (OC) → recepcionados → vendidos a público.

    Toda la agregación es por `values(...).annotate(...)`: no hay un loop por
    proveedor. La atribución de la venta y su regla de desempate están
    documentadas en la cabecera de este bloque.
    """
    try:
        filtros = _rp_filtros(request)
        compras_qs = _rp_compras_scope(filtros)

        proveedores = {}

        def _fila(proveedor_id):
            return proveedores.setdefault(proveedor_id, {
                'proveedor_id': proveedor_id,
                'proveedor_nombre': '', 'proveedor_rut': '',
                'pares_comprados': 0, 'pares_recepcionados': 0, 'pares_vendidos': 0,
                'inversion_total': 0.0, 'venta_total': 0.0,
                'articulos_atribuidos': 0,
                'documentos_no_inventariables': 0, 'monto_no_inventariable': 0.0,
            })

        # 1) Comprado e inversión, agrupado por proveedor.
        compras_por_proveedor = (
            Compras_Producto_Talla.objects
            .filter(compra_producto__compras__in=compras_qs)
            .values('compra_producto__compras__empresa_id')
            .annotate(pares=Sum('stock'),
                      inversion=Sum(F('compra_producto__costo') * F('stock')))
        )
        for row in compras_por_proveedor:
            proveedor_id = row['compra_producto__compras__empresa_id']
            if not proveedor_id:
                continue
            fila = _fila(proveedor_id)
            fila['pares_comprados'] = row['pares'] or 0
            fila['inversion_total'] = float(row['inversion'] or 0)

        # 2) Recepcionado en el período, agrupado por proveedor.
        recepciones_por_proveedor = (
            Productos_Recepcionados.objects
            .filter(_rp_q_periodo_recepcion(filtros),
                    compra_producto_talla__compra_producto__compras__in=compras_qs)
            .values('compra_producto_talla__compra_producto__compras__empresa_id')
            .annotate(unidades=Sum('stockArribado'))
        )
        for row in recepciones_por_proveedor:
            proveedor_id = row['compra_producto_talla__compra_producto__compras__empresa_id']
            if not proveedor_id:
                continue
            _fila(proveedor_id)['pares_recepcionados'] = row['unidades'] or 0

        # 3) Venta a público: se agrupa por identidad de artículo y se mapea al
        #    proveedor de origen. Cada artículo va a UN proveedor (ver cabecera),
        #    así que las filas son disjuntas y los totales son sumables.
        mapa_enlace, mapa_codigo, articulos_multiproveedor = _rp_mapa_articulo_proveedor(compras_qs)
        articulos_por_proveedor = defaultdict(int)
        for proveedor_id in mapa_enlace.values():
            articulos_por_proveedor[proveedor_id] += 1
        for proveedor_id in mapa_codigo.values():
            articulos_por_proveedor[proveedor_id] += 1

        unidades_sin_atribuir = 0
        venta_sin_atribuir = 0.0
        unidades_por_respaldo = 0
        for row in _rp_ventas_por_articulo(filtros):
            clave = _rp_clave_articulo(row['ProductoTalla__producto__articulo'],
                                       row['ProductoTalla__producto__atributo1_id'])
            unidades = row['unidades'] or 0
            venta = float(row['venta'] or 0)
            proveedor_id = mapa_enlace.get(clave)
            if not proveedor_id and clave:
                proveedor_id = mapa_codigo.get(clave[0])
                if proveedor_id:
                    unidades_por_respaldo += unidades
            if not proveedor_id:
                # Artículo vendido que no proviene de ninguna OC del universo
                # consultado (o sin enlace al catálogo): se informa, no se reparte.
                unidades_sin_atribuir += unidades
                venta_sin_atribuir += venta
                continue
            fila = _fila(proveedor_id)
            fila['pares_vendidos'] += unidades
            fila['venta_total'] += venta

        # 4) Compras por concepto / sin detalle inventariable.
        for row in _rp_no_inventariables(filtros):
            proveedor_id = row['emisor_id']
            if not proveedor_id:
                continue
            fila = _fila(proveedor_id)
            fila['documentos_no_inventariables'] = row['documentos'] or 0
            fila['monto_no_inventariable'] = float(row['monto'] or 0)

        # 5) Nombres, en una sola query (antes era un Empresa.get() por proveedor).
        for empresa in Empresa.objects.filter(id__in=list(proveedores.keys())).values('id', 'nombre', 'rut'):
            fila = proveedores.get(empresa['id'])
            if fila:
                fila['proveedor_nombre'] = empresa['nombre'] or 'Sin nombre'
                fila['proveedor_rut'] = empresa['rut'] or ''

        resultados = []
        for fila in proveedores.values():
            comprados = fila['pares_comprados']
            recepcionados = fila['pares_recepcionados']
            vendidos = fila['pares_vendidos']
            if not comprados and not recepcionados and not vendidos \
                    and not fila['documentos_no_inventariables']:
                continue
            fila['proveedor_nombre'] = fila['proveedor_nombre'] or 'Sin nombre'
            fila['articulos_atribuidos'] = articulos_por_proveedor.get(fila['proveedor_id'], 0)
            fila['pct_recepcion'] = round((recepcionados / comprados) * 100, 1) if comprados else 0
            fila['pct_venta'] = round((vendidos / recepcionados) * 100, 1) if recepcionados else 0
            # El costo de lo vendido se prorratea sobre la inversión de la OC:
            # Ticket_Productos.costo_fifo está en 0 en todo el histórico, así que
            # no hay costo real por línea del que tirar.
            costo_vendido = fila['inversion_total'] * (vendidos / comprados) if comprados else 0
            fila['margen'] = round(fila['venta_total'] - costo_vendido, 0)
            fila['stock_disponible'] = recepcionados - vendidos
            fila['pendiente_recepcion'] = max(comprados - recepcionados, 0)
            fila['metodologia'] = (
                'Incluye compras por concepto/sin detalle inventariable; se informan como monto no inventariable.'
                if fila['documentos_no_inventariables'] else ''
            )
            resultados.append(fila)

        resultados.sort(key=lambda r: r['inversion_total'], reverse=True)

        total_comprados = sum(r['pares_comprados'] for r in resultados)
        total_recepcionados = sum(r['pares_recepcionados'] for r in resultados)
        total_vendidos = sum(r['pares_vendidos'] for r in resultados)
        total_venta_atribuida = sum(r['venta_total'] for r in resultados)
        cobertura = (total_vendidos / (total_vendidos + unidades_sin_atribuir) * 100
                     if (total_vendidos + unidades_sin_atribuir) else 0)

        return JsonResponse({'success': True, 'kpis': {
            'total_proveedores': len(resultados),
            'total_comprados': total_comprados,
            'total_recepcionados': total_recepcionados,
            'total_vendidos': total_vendidos,
            'pct_recepcion_global': round((total_recepcionados / total_comprados) * 100, 1) if total_comprados else 0,
            'pct_venta_global': round((total_vendidos / total_recepcionados) * 100, 1) if total_recepcionados else 0,
            'total_inversion': sum(r['inversion_total'] for r in resultados),
            'total_monto_no_inventariable': sum(r['monto_no_inventariable'] for r in resultados),
            'total_documentos_no_inventariables': sum(r['documentos_no_inventariables'] for r in resultados),
            'total_pendiente_recepcion': sum(r['pendiente_recepcion'] for r in resultados),
            'total_venta': total_venta_atribuida,
            'total_margen': sum(r['margen'] for r in resultados),
            # Calidad del dato: lo que el reporte NO pudo atribuir a un proveedor.
            'unidades_venta_sin_atribuir': unidades_sin_atribuir,
            'venta_sin_atribuir': round(venta_sin_atribuir, 0),
            'cobertura_atribucion_pct': round(cobertura, 1),
            'articulos_multiproveedor': articulos_multiproveedor,
            'unidades_venta_por_nombre_oc': unidades_por_respaldo,
        }, 'proveedores': resultados, 'metodologia': {
            'comprados': 'Unidades desde Compras_Producto_Talla de ordenes ACTIVA/COMPLETADA (ELIMINADA y CANCELADA quedan fuera).',
            'recepcionados': 'Unidades desde Productos_Recepcionados en el periodo consultado.',
            'vendidos': (
                'La venta se atribuye por el enlace real OC -> Compras_Producto_Talla -> '
                'Productos_Recepcionados -> Producto_Talla, a nivel de identidad de articulo '
                '(codigo + marca) porque cada sucursal tiene su propio Producto_Talla. '
                'Para las lineas de OC todavia sin recepcionar se usa como respaldo el nombre '
                'de la linea cuando ES un codigo de catalogo (ver unidades_venta_por_nombre_oc). '
                'Si un articulo se compro a mas de un proveedor NO se duplica: se asigna al '
                'proveedor de la recepcion (u orden) mas reciente. Solo tickets PAGADO, por '
                'created_at, sin productos excluidos de analitica.'
            ),
            'sin_atribuir': (
                'Las unidades vendidas que no provienen de ninguna orden del universo '
                'consultado se informan aparte (unidades_venta_sin_atribuir) en vez de repartirse.'
            ),
            'margen': 'Venta atribuida menos costo prorrateado de la OC (costo_fifo no esta poblado en el historico).',
            'no_inventariable': 'DTE de compra sin lineas con Producto_Talla; excluye notas de credito, descartados y anulados. No suma pares ni recepcion.',
            'sucursal': 'El filtro de sucursal aplica SOLO a la venta: comprados y recepcionados siguen siendo de toda la cadena.',
            'scoping': (
                'Un usuario acotado a ciertas empresas ve solo las ordenes recepcionadas en '
                'sucursales de esas empresas, los DTE de compra que ellas recibieron y la venta '
                'de sus propias tiendas; comprados y recepcionados de una orden compartida siguen '
                'siendo los de la orden completa, porque el modelo Compras no tiene empresa compradora.'
            ),
        }})

    except Exception as e:
        logger.exception("Error al generar rendimiento por proveedor")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
@login_required
def exportar_rendimiento_proveedor_excel(request):
    """Exporta rendimiento por proveedor a Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        resp_api = api_reporte_rendimiento_proveedor(request)
        data = json.loads(resp_api.content)
        if not data.get('success'):
            return JsonResponse({'error': 'No se pudo generar'}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rendimiento Proveedor'
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='405189', end_color='405189', fill_type='solid')
        brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Proveedor', 'RUT', 'Comprados', 'Recepcionados', 'Vendidos',
                    '% Recep', '% Venta', 'Inversion inventariable', 'Monto no inventariable',
                    'Docs no inventariables', 'Venta Total', 'Margen', 'Pend. recepcion', 'Stock Disp.']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, Alignment(horizontal='center'), brd

        for i, p in enumerate(data['proveedores'], 2):
            vals = [p['proveedor_nombre'], p['proveedor_rut'], p['pares_comprados'], p['pares_recepcionados'],
                    p['pares_vendidos'], p['pct_recepcion'], p['pct_venta'], p['inversion_total'],
                    p.get('monto_no_inventariable', 0), p.get('documentos_no_inventariables', 0),
                    p['venta_total'], p['margen'], p.get('pendiente_recepcion', 0), p['stock_disponible']]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=v)
                c.border = brd
                if col >= 3:
                    c.alignment = Alignment(horizontal='right')

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col) + 3, 25)

        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename=rendimiento_proveedor_{request.GET.get("anio", timezone.localdate().year)}.xlsx'
        wb.save(resp)
        return resp
    except Exception as e:
        logger.exception("Error al exportar rendimiento por proveedor")
        return JsonResponse({'error': str(e)}, status=500)


# ========== REPORTE DE VENTAS POR INTERNET ==========


def _decimal_seguro(valor, default=Decimal('0')):
    """Convierte valores de snapshots ecommerce sin romper el reporte."""
    try:
        return Decimal(str(valor if valor not in (None, '') else default))
    except (TypeError, ValueError, ArithmeticError):
        return default


def _entero_seguro(valor, default=0):
    try:
        return int(_decimal_seguro(valor, Decimal(default)))
    except (TypeError, ValueError, ArithmeticError):
        return default


def _rango_ventas_internet(request):
    hoy = timezone.localdate()
    if request.GET.get('todo_historico') == '1':
        return None, None, True

    inicio_default = hoy.replace(month=1, day=1)
    try:
        fecha_inicio = datetime.strptime(
            request.GET.get('fecha_inicio') or inicio_default.isoformat(),
            '%Y-%m-%d',
        ).date()
        fecha_fin = datetime.strptime(
            request.GET.get('fecha_fin') or hoy.isoformat(),
            '%Y-%m-%d',
        ).date()
    except ValueError as exc:
        raise ValidationError('Las fechas deben usar el formato YYYY-MM-DD.') from exc

    if fecha_inicio > fecha_fin:
        raise ValidationError('La fecha de inicio no puede ser posterior a la fecha de fin.')
    return fecha_inicio, fecha_fin, False


def _nombre_usuario(usuario):
    if not usuario:
        return 'Sin operador'
    return usuario.get_full_name() or usuario.username


def _fecha_hora_ticket(ticket):
    if ticket.created_at:
        return timezone.localtime(ticket.created_at)
    if ticket.fecha:
        hora = ticket.hora or datetime.min.time()
        naive = datetime.combine(ticket.fecha, hora)
        return timezone.make_aware(naive) if timezone.is_naive(naive) else naive
    return None


def _decimal_desde_int(valor):
    return Decimal(str(valor or 0))


def _nombre_plataforma(valor):
    valor = (valor or '').strip()
    return valor or 'Sin plataforma'


def _codigos_canal_ecommerce(valor):
    normalizado = (valor or '').strip().upper()
    if not normalizado:
        return []
    codigos = {normalizado}
    for codigo, nombre in CANAL_ECOMMERCE_CHOICES:
        if normalizado in {codigo.upper(), nombre.upper()}:
            codigos.add(codigo)
    return sorted(codigos)


def _normalizar_tipo_boleta(tipo_documento):
    tipo = (tipo_documento or '').strip().upper().replace('_', ' ')
    if 'BOLETA' not in tipo:
        return ''
    if 'EXENTA' in tipo:
        return 'Boleta exenta'
    if 'ELECTRONICA' in tipo:
        return 'Boleta electronica'
    if 'PAPEL' in tipo or tipo == 'BOLETA':
        return 'Boleta papel/manual'
    return 'Boleta'


def _tipo_boleta_emitida(ticket, dte=None):
    if dte and getattr(dte, 'tipo_documento', None):
        return _normalizar_tipo_boleta(dte.tipo_documento)
    if ticket.folio_dte:
        return _normalizar_tipo_boleta(ticket.tipo_dte)
    return ''


def _linea_producto_nombre(linea):
    if linea.ProductoTalla and linea.ProductoTalla.producto:
        producto = linea.ProductoTalla.producto
        return producto.descripcion or producto.articulo or 'Producto sin nombre'
    return linea.descripcion_linea or 'Linea manual'


def _linea_producto_sku(linea):
    if linea.ProductoTalla:
        return str(linea.ProductoTalla.sku or 'Sin SKU')
    return 'MANUAL'


def _construir_reporte_ventas_internet(request, paginar=True):
    fecha_inicio, fecha_fin, todo_historico = _rango_ventas_internet(request)
    sucursales_permitidas = obtener_sucursales_usuario(request.user).select_related('empresa')
    sucursal_ids = list(sucursales_permitidas.values_list('id', flat=True))

    queryset = Ticket.objects.filter(
        estado='PAGADO',
        sucursal_id__in=sucursal_ids,
    ).filter(
        Q(modulo_origen='ECOMMERCE') | Q(pagos__metodo_pago='VENTA_INTERNET')
    ).select_related(
        'sucursal__empresa', 'vendedor',
    ).prefetch_related(
        'pagos',
        'ticket_productos__ProductoTalla__producto',
        'pedidos_ecommerce__dte',
        'pedidos_ecommerce__facturado_por',
    ).distinct().order_by('-created_at', '-id')

    if not todo_historico:
        # created_at es la fecha real de la venta; Ticket.fecha es auto_now
        # (se reescribe en cada save) y filtraba por última modificación.
        queryset = queryset.filter(
            created_at__date__gte=fecha_inicio, created_at__date__lte=fecha_fin
        )

    empresa_id = request.GET.get('empresa_id', '').strip()
    sucursal_id = request.GET.get('sucursal_id', '').strip()
    plataforma = (request.GET.get('plataforma') or request.GET.get('canal') or '').strip()
    origen = request.GET.get('origen', '').strip().upper()
    vendedor_id = request.GET.get('vendedor_id', '').strip()
    busqueda = request.GET.get('q', '').strip()

    if empresa_id:
        queryset = queryset.filter(sucursal__empresa_id=empresa_id)
    if sucursal_id:
        if not str(sucursal_id).isdigit() or int(sucursal_id) not in sucursal_ids:
            queryset = queryset.none()
        else:
            queryset = queryset.filter(sucursal_id=sucursal_id)
    if plataforma:
        if plataforma == '__SIN_PLATAFORMA__':
            queryset = queryset.filter(
                Q(pagos__metodo_pago='VENTA_INTERNET', pagos__tipo_tarjeta__isnull=True)
                | Q(pagos__metodo_pago='VENTA_INTERNET', pagos__tipo_tarjeta='')
            )
        else:
            queryset = queryset.filter(
                Q(pagos__metodo_pago='VENTA_INTERNET', pagos__tipo_tarjeta__iexact=plataforma)
                | Q(pedidos_ecommerce__canal_origen__in=_codigos_canal_ecommerce(plataforma))
            ).distinct()
    if origen == 'ECOMMERCE':
        queryset = queryset.filter(modulo_origen='ECOMMERCE')
    elif origen == 'POS':
        queryset = queryset.exclude(modulo_origen='ECOMMERCE').filter(pagos__metodo_pago='VENTA_INTERNET')
    if vendedor_id:
        queryset = queryset.filter(vendedor_id=vendedor_id)
    if busqueda:
        queryset = queryset.filter(
            Q(correlativo__icontains=busqueda)
            | Q(cliente_nombre__icontains=busqueda)
            | Q(cliente_rut__icontains=busqueda)
            | Q(pagos__voucher__icontains=busqueda)
            | Q(pagos__tipo_tarjeta__icontains=busqueda)
            | Q(pedidos_ecommerce__numero_ticket_rm__icontains=busqueda)
            | Q(pedidos_ecommerce__numero_pedido_canal__icontains=busqueda)
            | Q(pedidos_ecommerce__correlativo__icontains=busqueda)
            | Q(pedidos_ecommerce__cliente_nombre__icontains=busqueda)
            | Q(pedidos_ecommerce__cliente_documento__icontains=busqueda)
        ).distinct()

    empresas = {}
    plataformas = defaultdict(lambda: {'tickets': set(), 'ventas': Decimal('0'), 'unidades': 0})
    origenes = defaultdict(lambda: {'tickets': 0, 'ventas': Decimal('0'), 'unidades': 0})
    vendedores = defaultdict(lambda: {'tickets': 0, 'ventas': Decimal('0'), 'unidades': 0})
    boletas_empresas = defaultdict(lambda: {
        'id': None,
        'empresa': '',
        'rut': '',
        'boletas_total': 0,
        'boleta_electronica': 0,
        'boleta_papel': 0,
        'boleta_exenta': 0,
        'otras_boletas': 0,
        'tickets': set(),
        'monto_internet': Decimal('0'),
    })
    productos = defaultdict(lambda: {
        'sku': '', 'producto': '', 'unidades': 0, 'venta_internet_asignada': Decimal('0'),
        'venta_ticket_total': Decimal('0'), 'tickets': set(), 'plataformas': set(), 'origenes': set(),
    })
    ventas_diarias = defaultdict(lambda: {'tickets': 0, 'ventas': Decimal('0'), 'unidades': 0})
    detalle = []
    alertas = {
        'ecommerce_sin_pago_internet': {'count': 0, 'monto': 0, 'tickets': []},
        'tickets_mixtos': {'count': 0, 'monto_internet': 0, 'tickets': []},
        'sin_plataforma': {'count': 0, 'monto': 0, 'tickets': []},
    }
    total_ventas = Decimal('0')
    total_unidades = 0
    total_ecommerce = 0
    total_pos = 0
    total_boletas = 0
    monto_fallback_ecommerce = Decimal('0')

    for ticket in queryset:
        pagos_internet = [pago for pago in ticket.pagos.all() if pago.metodo_pago == 'VENTA_INTERNET']
        pedido = next(iter(ticket.pedidos_ecommerce.all()), None)
        es_ecommerce = ticket.modulo_origen == 'ECOMMERCE' or pedido is not None
        monto_pagos_internet = sum((pago.monto or 0) for pago in pagos_internet)

        if monto_pagos_internet <= 0 and not es_ecommerce:
            continue

        usa_fallback = False
        if monto_pagos_internet > 0:
            monto_internet = _decimal_desde_int(monto_pagos_internet)
        else:
            monto_internet = _decimal_desde_int(ticket.total)
            usa_fallback = True
            monto_fallback_ecommerce += monto_internet

        total_ticket = _decimal_desde_int(ticket.total)
        proporcion_internet = (monto_internet / total_ticket) if total_ticket > 0 else Decimal('0')
        proporcion_internet = min(proporcion_internet, Decimal('1'))
        plataformas_ticket = sorted({
            _nombre_plataforma(pago.tipo_tarjeta)
            for pago in pagos_internet
        })
        if not plataformas_ticket and pedido:
            plataformas_ticket = [_nombre_plataforma(pedido.get_canal_origen_display())]
        if not plataformas_ticket:
            plataformas_ticket = ['Sin plataforma']

        origen_codigo = 'ECOMMERCE' if es_ecommerce else 'POS'
        origen_nombre = 'Ecommerce / AllConnected' if es_ecommerce else 'POS Venta Internet'
        fecha_dt = _fecha_hora_ticket(ticket)
        fecha_local = fecha_dt.date() if fecha_dt else ticket.fecha
        fecha_key = fecha_local.isoformat() if fecha_local else ''
        lineas = list(ticket.ticket_productos.all())
        unidades_ticket = sum(max(linea.stock or 0, 0) for linea in lineas)

        total_ventas += monto_internet
        total_unidades += unidades_ticket
        if es_ecommerce:
            total_ecommerce += 1
        else:
            total_pos += 1

        if usa_fallback:
            alertas['ecommerce_sin_pago_internet']['count'] += 1
            alertas['ecommerce_sin_pago_internet']['monto'] += int(monto_internet)
            if len(alertas['ecommerce_sin_pago_internet']['tickets']) < 20:
                alertas['ecommerce_sin_pago_internet']['tickets'].append(ticket.correlativo)
        if monto_pagos_internet > 0 and total_ticket > monto_internet:
            alertas['tickets_mixtos']['count'] += 1
            alertas['tickets_mixtos']['monto_internet'] += int(monto_internet)
            if len(alertas['tickets_mixtos']['tickets']) < 20:
                alertas['tickets_mixtos']['tickets'].append(ticket.correlativo)
        if 'Sin plataforma' in plataformas_ticket:
            alertas['sin_plataforma']['count'] += 1
            alertas['sin_plataforma']['monto'] += int(monto_internet)
            if len(alertas['sin_plataforma']['tickets']) < 20:
                alertas['sin_plataforma']['tickets'].append(ticket.correlativo)

        empresa = ticket.sucursal.empresa
        empresa_data = empresas.setdefault(empresa.id, {
            'id': empresa.id,
            'nombre': empresa.nombre,
            'rut': empresa.rut,
            'tickets': 0,
            'ventas': Decimal('0'),
            'unidades': 0,
            'sucursales': {},
        })
        sucursal_data = empresa_data['sucursales'].setdefault(ticket.sucursal_id, {
            'id': ticket.sucursal_id,
            'nombre': ticket.sucursal.alias,
            'tickets': 0,
            'ventas': Decimal('0'),
            'unidades': 0,
        })
        empresa_data['tickets'] += 1
        empresa_data['ventas'] += monto_internet
        empresa_data['unidades'] += unidades_ticket
        sucursal_data['tickets'] += 1
        sucursal_data['ventas'] += monto_internet
        sucursal_data['unidades'] += unidades_ticket

        for plataforma_item in plataformas_ticket:
            plataformas[plataforma_item]['tickets'].add(ticket.id)
            if pagos_internet:
                ventas_plataforma = sum(
                    _decimal_desde_int(pago.monto)
                    for pago in pagos_internet
                    if _nombre_plataforma(pago.tipo_tarjeta) == plataforma_item
                )
            else:
                ventas_plataforma = monto_internet
            plataformas[plataforma_item]['ventas'] += ventas_plataforma
            plataformas[plataforma_item]['unidades'] += unidades_ticket

        origenes[origen_codigo]['codigo'] = origen_codigo
        origenes[origen_codigo]['nombre'] = origen_nombre
        origenes[origen_codigo]['tickets'] += 1
        origenes[origen_codigo]['ventas'] += monto_internet
        origenes[origen_codigo]['unidades'] += unidades_ticket

        vendedor = ticket.vendedor
        vendedor_key = str(vendedor.id) if vendedor else 'sin-vendedor'
        vendedor_nombre = str(vendedor) if vendedor else 'Sin vendedor asociado'
        vendedores[vendedor_key]['id'] = vendedor.id if vendedor else None
        vendedores[vendedor_key]['nombre'] = vendedor_nombre
        vendedores[vendedor_key]['codigo'] = vendedor.codigo_vendedor if vendedor else '-'
        vendedores[vendedor_key]['tickets'] += 1
        vendedores[vendedor_key]['ventas'] += monto_internet
        vendedores[vendedor_key]['unidades'] += unidades_ticket

        if fecha_key:
            ventas_diarias[fecha_key]['tickets'] += 1
            ventas_diarias[fecha_key]['ventas'] += monto_internet
            ventas_diarias[fecha_key]['unidades'] += unidades_ticket

        for linea in lineas:
            cantidad = max(linea.stock or 0, 0)
            subtotal_linea = _decimal_desde_int(linea.subtotal)
            venta_asignada = subtotal_linea * proporcion_internet
            sku = _linea_producto_sku(linea)
            nombre = _linea_producto_nombre(linea)
            producto_key = f'{sku}|{nombre}'
            producto = productos[producto_key]
            producto['sku'] = sku
            producto['producto'] = nombre
            producto['unidades'] += cantidad
            producto['venta_internet_asignada'] += venta_asignada
            producto['venta_ticket_total'] += subtotal_linea
            producto['tickets'].add(ticket.id)
            producto['plataformas'].update(plataformas_ticket)
            producto['origenes'].add(origen_nombre)

        dte = pedido.dte if pedido and pedido.dte_id else None
        dte_label = ''
        if dte:
            dte_label = f'{dte.tipo_documento} #{dte.numero_documento}'
        elif ticket.folio_dte:
            dte_label = f'{ticket.tipo_dte or "DTE"} #{ticket.folio_dte}'
        tipo_boleta = _tipo_boleta_emitida(ticket, dte)
        if tipo_boleta:
            total_boletas += 1
            boleta_empresa = boletas_empresas[empresa.id]
            boleta_empresa['id'] = empresa.id
            boleta_empresa['empresa'] = empresa.nombre
            boleta_empresa['rut'] = empresa.rut
            boleta_empresa['boletas_total'] += 1
            boleta_empresa['tickets'].add(ticket.id)
            boleta_empresa['monto_internet'] += monto_internet
            if tipo_boleta == 'Boleta electronica':
                boleta_empresa['boleta_electronica'] += 1
            elif tipo_boleta == 'Boleta papel/manual':
                boleta_empresa['boleta_papel'] += 1
            elif tipo_boleta == 'Boleta exenta':
                boleta_empresa['boleta_exenta'] += 1
            else:
                boleta_empresa['otras_boletas'] += 1
        fecha_venta_str = fecha_dt.strftime('%d/%m/%Y %H:%M') if fecha_dt else ''
        fecha_venta_iso = fecha_dt.isoformat() if fecha_dt else ''
        fecha_recepcion = timezone.localtime(pedido.fecha_recepcion).strftime('%d/%m/%Y %H:%M') if pedido and pedido.fecha_recepcion else ''
        fecha_facturacion = timezone.localtime(pedido.fecha_facturacion).strftime('%d/%m/%Y %H:%M') if pedido and pedido.fecha_facturacion else ''
        operador = _nombre_usuario(pedido.facturado_por) if pedido else ticket.responsable or 'POS'

        detalle.append({
            'id': ticket.id,
            'ticket_id': ticket.id,
            'correlativo': ticket.correlativo,
            'fecha': fecha_venta_str,
            'fecha_venta': fecha_venta_str,
            'fecha_iso': fecha_venta_iso,
            'fecha_recepcion': fecha_recepcion,
            'fecha_facturacion': fecha_facturacion,
            'origen': origen_nombre,
            'origen_codigo': origen_codigo,
            'plataforma': ', '.join(plataformas_ticket),
            'voucher': ', '.join([pago.voucher for pago in pagos_internet if pago.voucher]) or '-',
            'pedido_ecommerce_id': pedido.id if pedido else None,
            'ticket_rm': pedido.numero_ticket_rm if pedido else f'Ticket #{ticket.correlativo}',
            'numero_ticket_rm': pedido.numero_ticket_rm if pedido else '',
            'pedido_canal': pedido.numero_pedido_canal if pedido else '',
            'numero_pedido_canal': pedido.numero_pedido_canal if pedido else '',
            'dte': dte_label or '-',
            'tipo_boleta': tipo_boleta,
            'empresa': empresa.nombre,
            'sucursal': ticket.sucursal.alias,
            'cliente': ticket.cliente_nombre or (pedido.cliente_nombre if pedido else ''),
            'documento_cliente': ticket.cliente_rut or (pedido.cliente_documento if pedido else '') or '-',
            'vendedor': vendedor_nombre,
            'codigo_vendedor': vendedor.codigo_vendedor if vendedor else '-',
            'operador': operador,
            'unidades': unidades_ticket,
            'monto_internet': float(monto_internet),
            'total': float(monto_internet),
            'total_ticket': float(total_ticket),
            'es_mixto': monto_pagos_internet > 0 and total_ticket > monto_internet,
            'usa_fallback': usa_fallback,
            'pos_url': f'/app/pos-dashboard/?ticket={ticket.correlativo}',
            'ecommerce_url': f'/app/ecommerce/pedidos/{pedido.id}/' if pedido else '',
        })

    total_pedidos = len(detalle)
    for empresa in empresas.values():
        empresa['pedidos'] = empresa['tickets']
        empresa['ticket_promedio'] = float(empresa['ventas'] / empresa['tickets']) if empresa['tickets'] else 0
        empresa['participacion'] = float(empresa['ventas'] / total_ventas * 100) if total_ventas else 0
        empresa['ventas'] = float(empresa['ventas'])
        empresa['sucursales'] = sorted(
            [
                {
                    **sucursal,
                    'pedidos': sucursal['tickets'],
                    'ventas': float(sucursal['ventas']),
                    'ticket_promedio': float(sucursal['ventas'] / sucursal['tickets']) if sucursal['tickets'] else 0,
                }
                for sucursal in empresa['sucursales'].values()
            ],
            key=lambda item: item['ventas'],
            reverse=True,
        )

    productos_data = sorted([
        {
            'sku': item['sku'],
            'producto': item['producto'],
            'unidades': item['unidades'],
            'venta_internet_asignada': float(item['venta_internet_asignada']),
            'venta_ticket_total': float(item['venta_ticket_total']),
            'ventas': float(item['venta_internet_asignada']),
            'pedidos': len(item['tickets']),
            'tickets': len(item['tickets']),
            'canales': ', '.join(sorted(item['plataformas'])),
            'plataformas': ', '.join(sorted(item['plataformas'])),
            'origenes': ', '.join(sorted(item['origenes'])),
            'precio_promedio': float(item['venta_internet_asignada'] / item['unidades']) if item['unidades'] else 0,
        }
        for item in productos.values()
    ], key=lambda item: (item['venta_internet_asignada'], item['unidades']), reverse=True)

    vendedores_data = sorted([
        {
            **item,
            'pedidos': item['tickets'],
            'ventas': float(item['ventas']),
            'ticket_promedio': float(item['ventas'] / item['tickets']) if item['tickets'] else 0,
        }
        for item in vendedores.values()
    ], key=lambda item: item['ventas'], reverse=True)
    vendedores_reales = [item for item in vendedores_data if item['id'] is not None]
    pedidos_sin_vendedor = vendedores.get('sin-vendedor', {}).get('tickets', 0)

    boletas_empresas_data = sorted([
        {
            'id': item['id'],
            'empresa': item['empresa'],
            'nombre': item['empresa'],
            'rut': item['rut'],
            'boletas_total': item['boletas_total'],
            'boleta_electronica': item['boleta_electronica'],
            'boleta_papel': item['boleta_papel'],
            'boleta_exenta': item['boleta_exenta'],
            'otras_boletas': item['otras_boletas'],
            'tickets': len(item['tickets']),
            'monto_internet': float(item['monto_internet']),
            'participacion': (item['boletas_total'] / total_boletas * 100) if total_boletas else 0,
        }
        for item in boletas_empresas.values()
    ], key=lambda item: (item['boletas_total'], item['monto_internet']), reverse=True)

    paginator = Paginator(detalle, min(max(_entero_seguro(request.GET.get('page_size'), 25), 10), 100))
    pagina = paginator.get_page(request.GET.get('page', 1)) if paginar else None

    empresas_filtro = {}
    for sucursal in sucursales_permitidas:
        empresas_filtro[sucursal.empresa_id] = sucursal.empresa
    plataformas_filtro = set(
        TicketDetallePago.objects.filter(
            ticket__sucursal_id__in=sucursal_ids,
            metodo_pago='VENTA_INTERNET',
        )
        .exclude(tipo_tarjeta__isnull=True)
        .exclude(tipo_tarjeta='')
        .values_list('tipo_tarjeta', flat=True)
    )
    canal_display = dict(CANAL_ECOMMERCE_CHOICES)
    canales_ecommerce = PedidoEcommerce.objects.filter(
        ticket__sucursal_id__in=sucursal_ids,
    ).exclude(canal_origen__isnull=True).exclude(canal_origen='').values_list('canal_origen', flat=True).distinct()
    for canal in canales_ecommerce:
        plataformas_filtro.add(canal_display.get(canal, canal))
    vendedores_ticket_ids = Ticket.objects.filter(
        sucursal_id__in=sucursal_ids,
        vendedor__isnull=False,
    ).values_list('vendedor_id', flat=True).distinct()
    vendedores_filtro = Vendedor.objects.filter(
        Q(sucursales__id__in=sucursal_ids) | Q(id__in=vendedores_ticket_ids)
    ).distinct().order_by('nombre')

    return {
        'success': True,
        'periodo': {
            'inicio': fecha_inicio.isoformat() if fecha_inicio else '',
            'fin': fecha_fin.isoformat() if fecha_fin else '',
            'todo_historico': todo_historico,
            'label': 'Todo historico' if todo_historico else f'{fecha_inicio.isoformat()} al {fecha_fin.isoformat()}',
        },
        'resumen': {
            'ventas': float(total_ventas),
            'venta_internet': float(total_ventas),
            'pedidos': total_pedidos,
            'tickets': total_pedidos,
            'unidades': total_unidades,
            'ticket_promedio': float(total_ventas / total_pedidos) if total_pedidos else 0,
            'empresas': len(empresas),
            'sucursales': sum(len(empresa['sucursales']) for empresa in empresas.values()),
            'productos': len(productos_data),
            'plataformas': len(plataformas),
            'ecommerce_count': total_ecommerce,
            'pos_count': total_pos,
            'boletas': total_boletas,
            'boletas_empresas': len(boletas_empresas_data),
            'monto_fallback_ecommerce': float(monto_fallback_ecommerce),
            'vendedores': len(vendedores_reales),
            'vendedor_unico': len(vendedores_reales) == 1 and pedidos_sin_vendedor == 0,
            'pedidos_sin_vendedor': pedidos_sin_vendedor,
        },
        'empresas': sorted(empresas.values(), key=lambda item: item['ventas'], reverse=True),
        'canales': sorted([
            {
                'codigo': nombre,
                'nombre': nombre,
                'pedidos': len(item['tickets']),
                'tickets': len(item['tickets']),
                'ventas': float(item['ventas']),
                'unidades': item['unidades'],
                'participacion': float(item['ventas'] / total_ventas * 100) if total_ventas else 0,
            }
            for nombre, item in plataformas.items()
        ], key=lambda item: item['ventas'], reverse=True),
        'plataformas': sorted([
            {
                'codigo': nombre,
                'nombre': nombre,
                'tickets': len(item['tickets']),
                'ventas': float(item['ventas']),
                'unidades': item['unidades'],
                'participacion': float(item['ventas'] / total_ventas * 100) if total_ventas else 0,
            }
            for nombre, item in plataformas.items()
        ], key=lambda item: item['ventas'], reverse=True),
        'origenes': sorted([
            {
                'codigo': item['codigo'],
                'nombre': item['nombre'],
                'tickets': item['tickets'],
                'ventas': float(item['ventas']),
                'unidades': item['unidades'],
                'participacion': float(item['ventas'] / total_ventas * 100) if total_ventas else 0,
            }
            for item in origenes.values()
        ], key=lambda item: item['ventas'], reverse=True),
        'vendedores': vendedores_data,
        'boletas_empresas': boletas_empresas_data,
        'productos': productos_data,
        'ventas_diarias': [
            {'fecha': fecha, 'pedidos': item['tickets'], 'tickets': item['tickets'], 'ventas': float(item['ventas']), 'unidades': item['unidades']}
            for fecha, item in sorted(ventas_diarias.items())
        ],
        'detalle': list(pagina.object_list) if pagina else detalle,
        'alertas': alertas,
        'paginacion': {
            'pagina': pagina.number if pagina else 1,
            'paginas': paginator.num_pages if paginar else 1,
            'total': paginator.count,
            'page_size': paginator.per_page,
            'tiene_anterior': pagina.has_previous() if pagina else False,
            'tiene_siguiente': pagina.has_next() if pagina else False,
        },
        'filtros': {
            'empresas': [
                {'id': empresa.id, 'nombre': empresa.nombre}
                for empresa in sorted(empresas_filtro.values(), key=lambda item: item.nombre.lower())
            ],
            'sucursales': [
                {'id': sucursal.id, 'nombre': sucursal.alias, 'empresa_id': sucursal.empresa_id}
                for sucursal in sucursales_permitidas
            ],
            'canales': [],
            'plataformas': [
                {'codigo': item['nombre'], 'nombre': item['nombre']}
                for item in sorted(
                    [{'nombre': nombre} for nombre in plataformas_filtro],
                    key=lambda item: item['nombre'].lower(),
                )
            ],
            'origenes': [
                {'codigo': 'ECOMMERCE', 'nombre': 'Ecommerce / AllConnected'},
                {'codigo': 'POS', 'nombre': 'POS Venta Internet'},
            ],
            'vendedores': [
                {'id': vendedor.id, 'nombre': str(vendedor), 'codigo': vendedor.codigo_vendedor or '-'}
                for vendedor in vendedores_filtro
            ],
        },
    }


@login_required
def ver_reporte_ventas_internet(request):
    context = obtener_contexto_sucursales(request.user, request)
    return render(request, 'vistas/modulo_reportes/reporte_ventas_internet.html', context)


@require_GET
@login_required
def obtener_reporte_ventas_internet(request):
    try:
        return JsonResponse(_construir_reporte_ventas_internet(request))
    except ValidationError as exc:
        mensaje = exc.messages[0] if exc.messages else str(exc)
        return JsonResponse({'success': False, 'error': mensaje}, status=400)


@require_GET
@login_required
def exportar_reporte_ventas_internet(request):
    try:
        data = _construir_reporte_ventas_internet(request, paginar=False)
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = openpyxl.Workbook()
        resumen = workbook.active
        resumen.title = 'Resumen'
        encabezado = PatternFill('solid', fgColor='405189')
        fuente_encabezado = Font(color='FFFFFF', bold=True)

        resumen.append(['INFORME VENTAS INTERNET', f"{data['periodo']['inicio']} al {data['periodo']['fin']}"])
        resumen.append([])
        resumen.append(['Indicador', 'Valor'])
        indicadores = [
            ('Venta internet', data['resumen']['venta_internet']),
            ('Tickets', data['resumen']['tickets']),
            ('Unidades vendidas', data['resumen']['unidades']),
            ('Ticket promedio', data['resumen']['ticket_promedio']),
            ('Ecommerce / AllConnected', data['resumen']['ecommerce_count']),
            ('POS Venta Internet', data['resumen']['pos_count']),
            ('Boletas emitidas', data['resumen']['boletas']),
            ('Plataformas', data['resumen']['plataformas']),
            ('Fallback ecommerce sin pago internet', data['resumen']['monto_fallback_ecommerce']),
            ('Empresas', data['resumen']['empresas']),
            ('Sucursales', data['resumen']['sucursales']),
            ('Vendedores detectados', data['resumen']['vendedores']),
        ]
        for indicador in indicadores:
            resumen.append(indicador)

        productos_ws = workbook.create_sheet('Productos')
        productos_ws.append([
            'SKU', 'Producto', 'Unidades', 'Venta internet asignada', 'Venta ticket total',
            'Tickets', 'Precio promedio internet', 'Plataformas', 'Origenes',
        ])
        for producto in data['productos']:
            productos_ws.append([
                producto['sku'], producto['producto'], producto['unidades'],
                producto['venta_internet_asignada'], producto['venta_ticket_total'],
                producto['tickets'], producto['precio_promedio'], producto['plataformas'],
                producto['origenes'],
            ])

        plataformas_ws = workbook.create_sheet('Plataformas')
        plataformas_ws.append(['Plataforma', 'Tickets', 'Unidades', 'Venta', 'Participacion %'])
        for plataforma in data['plataformas']:
            plataformas_ws.append([
                plataforma['nombre'], plataforma['tickets'], plataforma['unidades'],
                plataforma['ventas'], plataforma['participacion'],
            ])

        boletas_ws = workbook.create_sheet('Boletas empresa')
        boletas_ws.append([
            'Empresa', 'RUT', 'Boletas total', 'Boleta electronica',
            'Boleta papel/manual', 'Boleta exenta', 'Otras boletas',
            'Tickets', 'Monto internet', 'Participacion %',
        ])
        for item in data['boletas_empresas']:
            boletas_ws.append([
                item['empresa'], item['rut'], item['boletas_total'],
                item['boleta_electronica'], item['boleta_papel'],
                item['boleta_exenta'], item['otras_boletas'], item['tickets'],
                item['monto_internet'], item['participacion'],
            ])

        detalle_ws = workbook.create_sheet('Detalle pedidos')
        detalle_ws.append([
            'Fecha venta', 'Origen', 'Plataforma', 'Ticket', 'Pedido RM', 'Pedido canal',
            'DTE/Folio', 'Tipo boleta', 'Empresa', 'Sucursal', 'Cliente', 'RUT/Doc', 'Vendedor',
            'Codigo vendedor', 'Operador', 'Fecha recepcion', 'Fecha facturacion',
            'Voucher', 'Unidades', 'Monto internet', 'Total ticket', 'Mixto', 'Fallback',
        ])
        for pedido in data['detalle']:
            detalle_ws.append([
                pedido['fecha_venta'], pedido['origen'], pedido['plataforma'], pedido['correlativo'],
                pedido['numero_ticket_rm'], pedido['numero_pedido_canal'], pedido['dte'],
                pedido['tipo_boleta'], pedido['empresa'], pedido['sucursal'], pedido['cliente'],
                pedido['documento_cliente'], pedido['vendedor'], pedido['codigo_vendedor'], pedido['operador'],
                pedido['fecha_recepcion'], pedido['fecha_facturacion'], pedido['voucher'],
                pedido['unidades'], pedido['monto_internet'], pedido['total_ticket'],
                'Si' if pedido['es_mixto'] else 'No',
                'Si' if pedido['usa_fallback'] else 'No',
            ])

        alertas_ws = workbook.create_sheet('Alertas')
        alertas_ws.append(['Alerta', 'Cantidad', 'Monto', 'Tickets muestra'])
        alertas_ws.append([
            'Ecommerce sin pago VENTA_INTERNET',
            data['alertas']['ecommerce_sin_pago_internet']['count'],
            data['alertas']['ecommerce_sin_pago_internet']['monto'],
            ', '.join(map(str, data['alertas']['ecommerce_sin_pago_internet']['tickets'])),
        ])
        alertas_ws.append([
            'Tickets mixtos',
            data['alertas']['tickets_mixtos']['count'],
            data['alertas']['tickets_mixtos']['monto_internet'],
            ', '.join(map(str, data['alertas']['tickets_mixtos']['tickets'])),
        ])
        alertas_ws.append([
            'Sin plataforma',
            data['alertas']['sin_plataforma']['count'],
            data['alertas']['sin_plataforma']['monto'],
            ', '.join(map(str, data['alertas']['sin_plataforma']['tickets'])),
        ])

        for sheet in workbook.worksheets:
            header_row = 3 if sheet.title == 'Resumen' else 1
            for cell in sheet[header_row]:
                cell.fill = encabezado
                cell.font = fuente_encabezado
                cell.alignment = Alignment(horizontal='center')
            sheet.freeze_panes = f'A{header_row + 1}'
            for column_cells in sheet.columns:
                width = min(max(len(str(cell.value or '')) for cell in column_cells) + 2, 45)
                sheet.column_dimensions[column_cells[0].column_letter].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f"attachment; filename=ventas_internet_{data['periodo']['inicio']}_{data['periodo']['fin']}.xlsx"
        )
        workbook.save(response)
        return response
    except ValidationError as exc:
        mensaje = exc.messages[0] if exc.messages else str(exc)
        return JsonResponse({'success': False, 'error': mensaje}, status=400)


# ========== REPORTE VENTAS GLOBALES POR EMPRESA ==========

@login_required
def ver_reporte_ventas_global(request):
    """Vista del reporte de ventas globales agrupadas por empresa"""
    context = obtener_contexto_sucursales(request.user, request)
    return render(request, 'vistas/modulo_reportes/reporte_ventas_global.html', context)


@require_GET
@login_required
def obtener_ventas_global_por_empresa(request):
    """API: ventas globales de todas las sucursales agrupadas por empresa con comparativo"""
    try:
        mes = request.GET.get('mes')
        fecha_inicio_param = request.GET.get('fecha_inicio')
        fecha_fin_param = request.GET.get('fecha_fin')

        if fecha_inicio_param and fecha_fin_param:
            fecha_inicio = datetime.strptime(fecha_inicio_param, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_param, '%Y-%m-%d')
        else:
            if not mes:
                mes = timezone.localdate().strftime('%Y-%m')
            fecha_inicio = datetime.strptime(mes, '%Y-%m').replace(day=1)
            if fecha_inicio.month == 12:
                fecha_fin = fecha_inicio.replace(year=fecha_inicio.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                fecha_fin = fecha_inicio.replace(month=fecha_inicio.month + 1, day=1) - timedelta(days=1)

        # Periodo anterior (mismo largo) para comparativo
        delta_dias = (fecha_fin - fecha_inicio).days + 1
        fecha_inicio_ant = fecha_inicio - timedelta(days=delta_dias)
        fecha_fin_ant = fecha_inicio - timedelta(days=1)

        # Scoping multi-empresa: sin permiso "ver todas", el reporte global
        # solo agrega las sucursales asignadas al usuario (EmpresaUser).
        _scope = {}
        if not usuario_puede_ver_todas_sucursales(request.user):
            _scope = {'sucursal_id__in': list(
                obtener_sucursales_usuario(request.user).values_list('id', flat=True)
            )}

        def _sumar_periodo(fi, ff):
            """Suma ventas de Tickets + DTEs por sucursal para un rango"""
            fi_date = fi.date() if hasattr(fi, 'date') else fi
            ff_date = ff.date() if hasattr(ff, 'date') else ff

            data = {}  # {sucursal_id: {ventas, descuentos, documentos, devoluciones}}

            # Tickets — SOLO los que no generaron DTE: un ticket con boleta ya
            # está representado en la suma de DTEs de abajo (misma regla
            # anti-doble-conteo del reporte comparativo).
            for r in Ticket.objects.filter(
                created_at__date__gte=fi_date, created_at__date__lte=ff_date,
                estado='PAGADO',
                dte_generado=False,
                modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
                **_scope,
            ).values('sucursal_id', 'sucursal__alias', 'sucursal__empresa_id', 'sucursal__empresa__nombre').annotate(
                total=Sum('total'), dcto=Sum('descuento'), docs=Count('id'),
            ):
                sid = r['sucursal_id']
                if not sid:
                    continue
                if sid not in data:
                    data[sid] = {
                        'alias': r['sucursal__alias'],
                        'empresa_id': r['sucursal__empresa_id'],
                        'empresa_nombre': r['sucursal__empresa__nombre'],
                        'ventas': 0, 'descuentos': 0, 'documentos': 0, 'devoluciones': 0,
                    }
                data[sid]['ventas'] += int(r['total'] or 0)
                data[sid]['descuentos'] += int(r['dcto'] or 0)
                data[sid]['documentos'] += int(r['docs'] or 0)

            # DTEs ventas
            for r in Dte.objects.filter(
                fecha_emision__gte=fi_date, fecha_emision__lte=ff_date,
                tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
                **_scope,
            ).exclude(
                estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
            ).exclude(
                tipo_documento='NOTA DE CREDITO'
            ).values('sucursal_id', 'sucursal__alias', 'sucursal__empresa_id', 'sucursal__empresa__nombre').annotate(
                total=Sum('monto_con_iva'), dcto=Sum('descuento'), docs=Count('id'),
            ):
                sid = r['sucursal_id']
                if not sid:
                    continue
                if sid not in data:
                    data[sid] = {
                        'alias': r['sucursal__alias'],
                        'empresa_id': r['sucursal__empresa_id'],
                        'empresa_nombre': r['sucursal__empresa__nombre'],
                        'ventas': 0, 'descuentos': 0, 'documentos': 0, 'devoluciones': 0,
                    }
                data[sid]['ventas'] += int(r['total'] or 0)
                data[sid]['descuentos'] += int(r['dcto'] or 0)
                data[sid]['documentos'] += int(r['docs'] or 0)

            # DTEs notas de credito.
            # Las NC de venta se emiten con tipo_transaccion DEVOLUCION o
            # ANULACION, no con VENTA: exigir ambas condiciones dejaba el
            # filtro vacío y las ventas "netas" salían iguales a las brutas
            # (junio-2026: $8,7M sin restar). Se excluye COMPRA para no
            # mezclar notas de crédito de proveedores.
            for r in Dte.objects.filter(
                fecha_emision__gte=fi_date, fecha_emision__lte=ff_date,
                tipo_transaccion__in=[
                    'VENTA', 'VENTA_PUBLICO', 'DEVOLUCION', 'ANULACION',
                ],
                tipo_documento='NOTA DE CREDITO',
                **_scope,
            ).exclude(
                estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
            ).values('sucursal_id').annotate(total=Sum('monto_con_iva')):
                sid = r['sucursal_id']
                if sid and sid in data:
                    data[sid]['devoluciones'] += int(r['total'] or 0)

            return data

        actual = _sumar_periodo(fecha_inicio, fecha_fin)
        anterior = _sumar_periodo(fecha_inicio_ant, fecha_fin_ant)

        # Agrupar por empresa
        empresas = {}  # {empresa_id: {nombre, sucursales: [...], totales}}
        for sid, d in actual.items():
            eid = d['empresa_id']
            if not eid:
                continue
            if eid not in empresas:
                empresas[eid] = {
                    'nombre': d['empresa_nombre'],
                    'sucursales': [],
                    'ventas': 0, 'descuentos': 0, 'documentos': 0, 'devoluciones': 0,
                }
            netas = d['ventas'] - d['devoluciones']
            empresas[eid]['ventas'] += netas
            empresas[eid]['descuentos'] += d['descuentos']
            empresas[eid]['documentos'] += d['documentos']
            empresas[eid]['devoluciones'] += d['devoluciones']

            # Comparativo sucursal
            ant = anterior.get(sid, {})
            ventas_ant = (ant.get('ventas', 0) - ant.get('devoluciones', 0))
            variacion = ((netas - ventas_ant) / ventas_ant * 100) if ventas_ant > 0 else (100 if netas > 0 else 0)

            empresas[eid]['sucursales'].append({
                'id': sid,
                'nombre': d['alias'],
                'ventas': netas,
                'ventas_brutas': d['ventas'],
                'descuentos': d['descuentos'],
                'documentos': d['documentos'],
                'devoluciones': d['devoluciones'],
                'ticket_promedio': netas // d['documentos'] if d['documentos'] > 0 else 0,
                'ventas_anterior': ventas_ant,
                'variacion': round(variacion, 1),
            })

        # Totales globales y variaciones por empresa
        total_global = sum(e['ventas'] for e in empresas.values())

        # Anterior por empresa para comparativo
        empresas_ant = {}
        for sid, d in anterior.items():
            eid = d.get('empresa_id')
            if eid:
                if eid not in empresas_ant:
                    empresas_ant[eid] = 0
                empresas_ant[eid] += d['ventas'] - d.get('devoluciones', 0)

        resultado = []
        for eid, e in sorted(empresas.items(), key=lambda x: x[1]['ventas'], reverse=True):
            ventas_ant_emp = empresas_ant.get(eid, 0)
            variacion_emp = ((e['ventas'] - ventas_ant_emp) / ventas_ant_emp * 100) if ventas_ant_emp > 0 else (100 if e['ventas'] > 0 else 0)
            participacion = (e['ventas'] / total_global * 100) if total_global > 0 else 0

            # Ordenar sucursales por ventas
            e['sucursales'].sort(key=lambda s: s['ventas'], reverse=True)

            resultado.append({
                'id': eid,
                'nombre': e['nombre'],
                'ventas': e['ventas'],
                'descuentos': e['descuentos'],
                'documentos': e['documentos'],
                'devoluciones': e['devoluciones'],
                'ticket_promedio': e['ventas'] // e['documentos'] if e['documentos'] > 0 else 0,
                'participacion': round(participacion, 1),
                'ventas_anterior': ventas_ant_emp,
                'variacion': round(variacion_emp, 1),
                'sucursales': e['sucursales'],
            })

        return JsonResponse({
            'success': True,
            'empresas': resultado,
            'kpis': {
                'total_ventas': total_global,
                'total_descuentos': sum(e['descuentos'] for e in empresas.values()),
                'total_documentos': sum(e['documentos'] for e in empresas.values()),
                'total_devoluciones': sum(e['devoluciones'] for e in empresas.values()),
                'total_empresas': len(resultado),
                'total_sucursales': sum(len(e['sucursales']) for e in empresas.values()),
            },
            'periodo': {
                'inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fin': fecha_fin.strftime('%Y-%m-%d'),
                'inicio_anterior': fecha_inicio_ant.strftime('%Y-%m-%d'),
                'fin_anterior': fecha_fin_ant.strftime('%Y-%m-%d'),
            }
        })

    except Exception as e:
        logger.exception("Error al obtener reporte de ventas sucursal")
        return JsonResponse({'success': False, 'error': str(e)})


# ========== REPORTE COMPARATIVO DE VENTAS ==========

def _calcular_rangos_comparativo(tipo_flujo, fecha_inicio_param=None, fecha_fin_param=None):
    """
    Calcula el rango (fi_act, ff_act) del periodo actual y (fi_ant, ff_ant) del
    periodo anterior equivalente, según el tipo de flujo temporal.

    Devuelve 4 objetos datetime.date.
    Flujos soportados:
      - 'hoy'           : hoy vs ayer
      - 'semana'        : lunes-hoy vs lunes-domingo semana anterior
      - 'mes_mtd'       : mes-a-la-fecha vs mismo rango mes anterior
      - 'mes_full'      : mes actual completo vs mes anterior completo
      - 'ultimos_7'     : últimos 7 días vs 7 días previos
      - 'ultimos_30'    : últimos 30 días vs 30 días previos
      - 'yoy'           : mes actual vs mismo mes año anterior (elimina estacionalidad)
      - 'trimestre_yoy' : trimestre actual vs mismo trimestre año anterior
      - 'anio_ytd'      : año a la fecha (ene-hoy) vs mismo rango año anterior
      - 'anio_full'     : año calendario anterior completo vs año antepasado completo
      - 'custom'        : usa fecha_inicio_param/fecha_fin_param y desplaza el mismo largo
    """
    hoy = timezone.localdate()

    def _restar_meses(d, meses):
        y = d.year
        m = d.month - meses
        while m <= 0:
            m += 12
            y -= 1
        from calendar import monthrange
        dia = min(d.day, monthrange(y, m)[1])
        return d.replace(year=y, month=m, day=dia)

    if tipo_flujo == 'hoy':
        fi_act = ff_act = hoy
        fi_ant = ff_ant = hoy - timedelta(days=1)

    elif tipo_flujo == 'semana':
        dia_semana = hoy.weekday()
        fi_act = hoy - timedelta(days=dia_semana)
        ff_act = hoy
        fi_ant = fi_act - timedelta(days=7)
        ff_ant = fi_ant + timedelta(days=dia_semana)

    elif tipo_flujo == 'mes_mtd':
        fi_act = hoy.replace(day=1)
        ff_act = hoy
        fi_ant = _restar_meses(fi_act, 1)
        dia_actual = hoy.day
        from calendar import monthrange
        ultimo_dia_mes_ant = monthrange(fi_ant.year, fi_ant.month)[1]
        ff_ant = fi_ant.replace(day=min(dia_actual, ultimo_dia_mes_ant))

    elif tipo_flujo == 'mes_full':
        fi_act = hoy.replace(day=1)
        from calendar import monthrange
        ultimo_dia = monthrange(hoy.year, hoy.month)[1]
        ff_act = hoy.replace(day=ultimo_dia)
        fi_ant = _restar_meses(fi_act, 1)
        ultimo_dia_ant = monthrange(fi_ant.year, fi_ant.month)[1]
        ff_ant = fi_ant.replace(day=ultimo_dia_ant)

    elif tipo_flujo == 'ultimos_7':
        ff_act = hoy
        fi_act = hoy - timedelta(days=6)
        ff_ant = fi_act - timedelta(days=1)
        fi_ant = ff_ant - timedelta(days=6)

    elif tipo_flujo == 'ultimos_30':
        ff_act = hoy
        fi_act = hoy - timedelta(days=29)
        ff_ant = fi_act - timedelta(days=1)
        fi_ant = ff_ant - timedelta(days=29)

    elif tipo_flujo == 'yoy':
        fi_act = hoy.replace(day=1)
        from calendar import monthrange
        ultimo_dia = monthrange(hoy.year, hoy.month)[1]
        ff_act = hoy.replace(day=ultimo_dia)
        try:
            fi_ant = fi_act.replace(year=fi_act.year - 1)
        except ValueError:
            fi_ant = fi_act.replace(year=fi_act.year - 1, day=28)
        ultimo_dia_ant = monthrange(fi_ant.year, fi_ant.month)[1]
        ff_ant = fi_ant.replace(day=ultimo_dia_ant)

    elif tipo_flujo == 'trimestre_yoy':
        from calendar import monthrange
        trimestre = (hoy.month - 1) // 3 + 1
        mes_inicio_q = (trimestre - 1) * 3 + 1
        mes_fin_q = mes_inicio_q + 2
        fi_act = hoy.replace(month=mes_inicio_q, day=1)
        ff_act = hoy.replace(
            month=mes_fin_q,
            day=monthrange(hoy.year, mes_fin_q)[1],
        )
        fi_ant = fi_act.replace(year=fi_act.year - 1)
        ff_ant = fi_ant.replace(
            day=monthrange(fi_ant.year, fi_ant.month)[1],
        )
        # Ajustar día final por si cambia último día del mes entre años
        ff_ant = ff_ant.replace(
            month=mes_fin_q,
            day=monthrange(ff_ant.year, mes_fin_q)[1],
        )

    elif tipo_flujo == 'anio_ytd':
        fi_act = hoy.replace(month=1, day=1)
        ff_act = hoy
        fi_ant = fi_act.replace(year=fi_act.year - 1)
        try:
            ff_ant = hoy.replace(year=hoy.year - 1)
        except ValueError:
            ff_ant = hoy.replace(year=hoy.year - 1, day=28)

    elif tipo_flujo == 'anio_full':
        from calendar import monthrange
        anio_act = hoy.year - 1
        anio_ant = hoy.year - 2
        fi_act = hoy.replace(year=anio_act, month=1, day=1)
        ff_act = hoy.replace(
            year=anio_act, month=12,
            day=monthrange(anio_act, 12)[1],
        )
        fi_ant = hoy.replace(year=anio_ant, month=1, day=1)
        ff_ant = hoy.replace(
            year=anio_ant, month=12,
            day=monthrange(anio_ant, 12)[1],
        )

    else:  # 'custom' o default
        if fecha_inicio_param and fecha_fin_param:
            fi_act = datetime.strptime(fecha_inicio_param, '%Y-%m-%d').date()
            ff_act = datetime.strptime(fecha_fin_param, '%Y-%m-%d').date()
        else:
            fi_act = hoy.replace(day=1)
            ff_act = hoy
        delta_dias = (ff_act - fi_act).days + 1
        ff_ant = fi_act - timedelta(days=1)
        fi_ant = ff_ant - timedelta(days=delta_dias - 1)

    return fi_act, ff_act, fi_ant, ff_ant


def _sumar_ventas_comparativo(fi, ff, user, request):
    """
    Suma ventas de Tickets + DTEs para el rango (fi, ff), con breakdown por
    sucursal, vendedor y canal, **evitando el doble contado** entre Ticket y Dte.

    Regla anti-doble contado (clave):
      - Un Ticket con `dte_generado=True` ya está representado en la tabla Dte
        como su boleta/factura electrónica. Para NO contarlo dos veces:
          * Total ventas brutas = sum(DTEs) + sum(Tickets con dte_generado=False)
      - ECOMMERCE se calcula INDEPENDIENTE desde Ticket.modulo_origen='ECOMMERCE'
        (da igual si tiene DTE o no; el monto es el del ticket).
      - "POS presencial" (canal 'POS') = Tickets POS/VENTA_PUBLICO sin DTE.
      - Canal 'DTE' = DTEs MENOS los tickets ECOMMERCE que ya generaron DTE
        (para no duplicar ventas de internet dentro de 'presencial').

    Devuelve dict:
    {
      'total_ventas_netas', 'total_ventas_brutas', 'total_documentos',
      'total_descuentos', 'total_devoluciones', 'cantidad_devoluciones',
      'total_unidades',
      'canal': {
          'ECOMMERCE':  {'ventas', 'documentos', 'unidades'},
          'POS':        {'ventas', 'documentos', 'unidades'},
          'DTE':        {'ventas', 'documentos', 'unidades'},
      },
      'sucursales': { sid: {...} },
      'vendedores': { vid: {...} },
    }
    """
    resultado = {
        'total_ventas_netas': 0,
        'total_ventas_brutas': 0,
        'total_documentos': 0,
        'total_descuentos': 0,
        'total_devoluciones': 0,
        'cantidad_devoluciones': 0,
        'total_unidades': 0,
        'canal': {
            'ECOMMERCE': {'ventas': 0, 'documentos': 0, 'unidades': 0},
            'POS':       {'ventas': 0, 'documentos': 0, 'unidades': 0},
            'DTE':       {'ventas': 0, 'documentos': 0, 'unidades': 0},
        },
        'sucursales': {},
        'vendedores': {},
    }

    def _suc_row(sid, nombre):
        return resultado['sucursales'].setdefault(sid, {
            'nombre': nombre or '-',
            'ventas_brutas': 0, 'devoluciones': 0, 'documentos': 0,
            'descuentos': 0, 'unidades': 0,
            'ventas_ecommerce': 0, 'ventas_presencial': 0,
        })

    def _vend_row(vid, nombre, codigo, sucursal):
        return resultado['vendedores'].setdefault(vid, {
            'nombre': nombre or '-',
            'codigo': codigo or '',
            'sucursal': sucursal or '-',
            'ventas_brutas': 0, 'devoluciones': 0, 'documentos': 0,
        })

    # ========== TICKETS (base común) ==========
    qs_tickets = Ticket.objects.filter(
        created_at__date__gte=fi,
        created_at__date__lte=ff,
        estado='PAGADO',
        modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
    ).select_related('sucursal', 'vendedor')
    qs_tickets = filtrar_queryset_por_sucursal(qs_tickets, user, request)

    # --- Canal ECOMMERCE: TODOS los tickets ECOMMERCE (aunque tengan DTE) ---
    # El monto real de la venta está siempre en el Ticket. Se suma aquí y
    # luego se RESTA del canal DTE para no contarlo dos veces.
    qs_ecom = qs_tickets.filter(modulo_origen='ECOMMERCE')
    for r in qs_ecom.values('sucursal_id', 'sucursal__alias').annotate(
        total=Sum('total'), dcto=Sum('descuento'), docs=Count('id'),
    ):
        sid = r['sucursal_id']
        total = int(r['total'] or 0)
        docs = int(r['docs'] or 0)
        dcto = int(r['dcto'] or 0)
        resultado['canal']['ECOMMERCE']['ventas'] += total
        resultado['canal']['ECOMMERCE']['documentos'] += docs
        # OJO: los tickets ECOMMERCE con dte_generado=True YA están en DTEs.
        # NO sumamos al total_ventas_brutas aquí para esos. Solo sumamos los
        # ECOMMERCE SIN DTE (caso raro, pero existe).
        if sid:
            s = _suc_row(sid, r['sucursal__alias'])
            s['ventas_ecommerce'] += total

    # Los tickets ECOMMERCE sin DTE sí se suman al total bruto (no están en DTE)
    for r in qs_ecom.filter(dte_generado=False).values(
        'sucursal_id', 'sucursal__alias'
    ).annotate(total=Sum('total'), dcto=Sum('descuento'), docs=Count('id')):
        sid = r['sucursal_id']
        total = int(r['total'] or 0)
        docs = int(r['docs'] or 0)
        dcto = int(r['dcto'] or 0)
        resultado['total_ventas_brutas'] += total
        resultado['total_documentos'] += docs
        resultado['total_descuentos'] += dcto
        if sid:
            s = _suc_row(sid, r['sucursal__alias'])
            s['ventas_brutas'] += total
            s['documentos'] += docs
            s['descuentos'] += dcto

    # --- Canal POS presencial: tickets POS/VENTA_PUBLICO sin DTE ---
    qs_pos_solo = qs_tickets.filter(
        modulo_origen__in=['POS', 'VENTA_PUBLICO'],
        dte_generado=False,
    )
    for r in qs_pos_solo.values('sucursal_id', 'sucursal__alias').annotate(
        total=Sum('total'), dcto=Sum('descuento'), docs=Count('id'),
    ):
        sid = r['sucursal_id']
        total = int(r['total'] or 0)
        docs = int(r['docs'] or 0)
        dcto = int(r['dcto'] or 0)
        resultado['canal']['POS']['ventas'] += total
        resultado['canal']['POS']['documentos'] += docs
        resultado['total_ventas_brutas'] += total
        resultado['total_documentos'] += docs
        resultado['total_descuentos'] += dcto
        if sid:
            s = _suc_row(sid, r['sucursal__alias'])
            s['ventas_brutas'] += total
            s['documentos'] += docs
            s['descuentos'] += dcto
            s['ventas_presencial'] += total

    # --- Unidades de tickets ECOMMERCE + POS-sin-DTE ---
    tickets_unidades = Ticket_Productos.objects.filter(
        idTicket__in=qs_ecom
    ).values(
        'idTicket__sucursal_id'
    ).annotate(u=Sum('stock'))
    for r in tickets_unidades:
        u = int(r['u'] or 0)
        resultado['canal']['ECOMMERCE']['unidades'] += u
        sid = r['idTicket__sucursal_id']
        # No sumamos al total_unidades aquí; se suma desde DTEs + tickets sin DTE
        if sid and sid in resultado['sucursales']:
            # sólo contar para la sucursal si el ticket NO tiene DTE (evitar doble)
            pass

    # Unidades solo de tickets sin DTE (se suman al total)
    unid_tick_sin_dte = Ticket_Productos.objects.filter(
        idTicket__in=qs_tickets.filter(dte_generado=False)
    ).values(
        'idTicket__sucursal_id', 'idTicket__modulo_origen'
    ).annotate(u=Sum('stock'))
    for r in unid_tick_sin_dte:
        u = int(r['u'] or 0)
        resultado['total_unidades'] += u
        sid = r['idTicket__sucursal_id']
        canal = 'ECOMMERCE' if r['idTicket__modulo_origen'] == 'ECOMMERCE' else 'POS'
        # ECOMMERCE ya sumó unidades arriba; solo sumamos POS sin DTE al canal POS
        if canal == 'POS':
            resultado['canal']['POS']['unidades'] += u
        if sid and sid in resultado['sucursales']:
            resultado['sucursales'][sid]['unidades'] += u

    # --- Tickets por vendedor (solo los sin DTE, para no duplicar con DTE) ---
    for r in qs_tickets.filter(dte_generado=False).values(
        'vendedor_id', 'vendedor__nombre', 'vendedor__codigo_vendedor',
        'sucursal__alias',
    ).annotate(total=Sum('total'), docs=Count('id')):
        vid = r['vendedor_id']
        if not vid:
            continue
        v = _vend_row(vid, r['vendedor__nombre'],
                      r['vendedor__codigo_vendedor'], r['sucursal__alias'])
        v['ventas_brutas'] += int(r['total'] or 0)
        v['documentos'] += int(r['docs'] or 0)

    # Identificar ID de tickets ECOMMERCE con DTE (para restar luego del canal DTE)
    ecom_con_dte_por_suc = {}  # {sid: monto_a_restar}
    for r in qs_ecom.filter(dte_generado=True).values(
        'sucursal_id'
    ).annotate(total=Sum('total'), docs=Count('id')):
        sid = r['sucursal_id']
        ecom_con_dte_por_suc[sid] = {
            'total': int(r['total'] or 0),
            'docs': int(r['docs'] or 0),
        }

    # ========== DTEs ==========
    qs_dtes_base = Dte.objects.filter(
        fecha_emision__gte=fi,
        fecha_emision__lte=ff,
        tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
    ).exclude(
        estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
    ).exclude(
        receptor__isnull=False,
        receptor_id=F('emisor_id')
    ).select_related('sucursal', 'vendedor')
    qs_dtes_base = filtrar_queryset_por_sucursal(qs_dtes_base, user, request)

    # DTEs de venta (excluyendo NC)
    qs_dte_ventas = qs_dtes_base.exclude(tipo_documento='NOTA DE CREDITO')
    for r in qs_dte_ventas.values('sucursal_id', 'sucursal__alias').annotate(
        total=Sum('monto_con_iva'),
        dcto=Sum('descuento'),
        docs=Count('id'),
        unidades=Sum('unidades_productos'),
    ):
        sid = r['sucursal_id']
        total = int(r['total'] or 0)
        docs = int(r['docs'] or 0)
        dcto = int(r['dcto'] or 0)
        unidades = int(r['unidades'] or 0)

        # Restar los tickets ECOMMERCE que YA fueron contados arriba (viven
        # en la tabla Dte como boletas electrónicas)
        ecom_info = ecom_con_dte_por_suc.get(sid, {'total': 0, 'docs': 0})
        dte_neto = total - ecom_info['total']
        dte_docs_neto = max(docs - ecom_info['docs'], 0)

        resultado['canal']['DTE']['ventas'] += dte_neto
        resultado['canal']['DTE']['documentos'] += dte_docs_neto
        resultado['canal']['DTE']['unidades'] += unidades

        # Al total general sí sumamos TODOS los DTEs (incluyen a los ECOMMERCE
        # que se convirtieron en boleta). Por eso NO sumamos arriba los tickets
        # ECOMMERCE con DTE al total: ya están representados acá.
        resultado['total_ventas_brutas'] += total
        resultado['total_documentos'] += docs
        resultado['total_descuentos'] += dcto
        resultado['total_unidades'] += unidades

        if sid:
            s = _suc_row(sid, r['sucursal__alias'])
            s['ventas_brutas'] += total
            s['documentos'] += docs
            s['descuentos'] += dcto
            s['unidades'] += unidades
            # Presencial a nivel sucursal = DTEs totales - parte ECOMMERCE
            s['ventas_presencial'] += dte_neto

    # DTEs por vendedor
    for r in qs_dte_ventas.values(
        'vendedor_id', 'vendedor__nombre', 'vendedor__codigo_vendedor',
        'sucursal__alias'
    ).annotate(total=Sum('monto_con_iva'), docs=Count('id')):
        vid = r['vendedor_id']
        if not vid:
            continue
        v = _vend_row(vid, r['vendedor__nombre'],
                      r['vendedor__codigo_vendedor'], r['sucursal__alias'])
        v['ventas_brutas'] += int(r['total'] or 0)
        v['documentos'] += int(r['docs'] or 0)

    # NC (devoluciones) por sucursal.
    # `qs_dtes_base` está restringido a tipo_transaccion VENTA/VENTA_PUBLICO y
    # ninguna NC lleva esos tipos, así que filtrarla por tipo_documento sobre
    # esa base devolvía casi cero: el KPI "Tasa Devolución" salía ~0% en todos
    # los períodos y `total_ventas_netas` era igual a las brutas.
    qs_nc = _queryset_ncs_venta(
        fi, ff, user=user, request=request,
        estados=('EMITIDO', 'ACEPTADO'),
        excluir_internas=True,
    )
    for r in qs_nc.values('sucursal_id').annotate(
        total=Sum('monto_con_iva'), cant=Count('id')
    ):
        sid = r['sucursal_id']
        total_nc = int(r['total'] or 0)
        cant_nc = int(r['cant'] or 0)
        resultado['total_devoluciones'] += total_nc
        resultado['cantidad_devoluciones'] += cant_nc
        if sid and sid in resultado['sucursales']:
            resultado['sucursales'][sid]['devoluciones'] += total_nc

    # NC por vendedor.
    # `emisionNotaCredito` no copia el vendedor a la NC, así que `vendedor_id`
    # viene NULL en casi todas: sin el fallback al vendedor del documento
    # afectado, la columna "devoluciones" del ranking quedaba en cero.
    for r in qs_nc.values(
        'vendedor_id', 'documento_afectado__vendedor_id'
    ).annotate(total=Sum('monto_con_iva')):
        vid = r['vendedor_id'] or r['documento_afectado__vendedor_id']
        if vid and vid in resultado['vendedores']:
            resultado['vendedores'][vid]['devoluciones'] += int(r['total'] or 0)

    resultado['total_ventas_netas'] = (
        resultado['total_ventas_brutas'] - resultado['total_devoluciones']
    )
    return resultado


@login_required
def ver_reporte_ventas_comparativo(request):
    """Vista principal del reporte comparativo de ventas (actual vs periodo anterior)."""
    context = obtener_contexto_sucursales(request.user, request)
    return render(request, 'vistas/modulo_reportes/reporte_ventas_comparativo.html', context)


@require_GET
@login_required
def obtener_ventas_comparativo(request):
    """API: datos comparativos de ventas entre periodo actual y periodo anterior equivalente.

    Parámetros:
      - tipo_flujo: hoy | semana | mes_mtd | mes_full | ultimos_7 | ultimos_30 | yoy | custom
      - fecha_inicio, fecha_fin: requeridos si tipo_flujo=custom
      - sucursal_id: opcional (filtra por una sucursal; respeta permisos)
    """
    try:
        # Default `mes_mtd` (mes a la fecha vs mismo tramo del mes anterior).
        # Con `mes_full` se comparaba el mes en curso INCOMPLETO contra el mes
        # anterior COMPLETO, así que al abrir el reporte siempre aparecía una
        # caída fantasma que se iba corrigiendo sola a fin de mes.
        tipo_flujo = request.GET.get('tipo_flujo', 'mes_mtd')
        fi_param = request.GET.get('fecha_inicio')
        ff_param = request.GET.get('fecha_fin')

        fi_act, ff_act, fi_ant, ff_ant = _calcular_rangos_comparativo(
            tipo_flujo, fi_param, ff_param
        )

        actual = _sumar_ventas_comparativo(fi_act, ff_act, request.user, request)
        anterior = _sumar_ventas_comparativo(fi_ant, ff_ant, request.user, request)

        def _var_pct(act, ant):
            if ant > 0:
                return round((act - ant) / ant * 100, 1)
            return 100.0 if act > 0 else 0.0

        def _pct(num, den):
            return round(num / den * 100, 1) if den > 0 else 0.0

        # ---------- KPIs globales ----------
        ventas_act = actual['total_ventas_netas']
        ventas_ant = anterior['total_ventas_netas']
        docs_act = actual['total_documentos']
        docs_ant = anterior['total_documentos']
        ticket_act = (ventas_act / docs_act) if docs_act > 0 else 0
        ticket_ant = (ventas_ant / docs_ant) if docs_ant > 0 else 0

        pct_internet_act = _pct(
            actual['canal']['ECOMMERCE']['ventas'], actual['total_ventas_brutas']
        )
        pct_internet_ant = _pct(
            anterior['canal']['ECOMMERCE']['ventas'], anterior['total_ventas_brutas']
        )

        tasa_dev_act = _pct(actual['total_devoluciones'], actual['total_ventas_brutas'])
        tasa_dev_ant = _pct(anterior['total_devoluciones'], anterior['total_ventas_brutas'])

        kpis = {
            'ventas_actual': ventas_act,
            'ventas_anterior': ventas_ant,
            'variacion_pct': _var_pct(ventas_act, ventas_ant),
            'variacion_abs': ventas_act - ventas_ant,

            'pct_internet_actual': pct_internet_act,
            'pct_internet_anterior': pct_internet_ant,
            'delta_internet_pp': round(pct_internet_act - pct_internet_ant, 1),
            'ventas_internet_actual': actual['canal']['ECOMMERCE']['ventas'],
            'ventas_internet_anterior': anterior['canal']['ECOMMERCE']['ventas'],

            'ticket_promedio_actual': int(ticket_act),
            'ticket_promedio_anterior': int(ticket_ant),
            'variacion_ticket_pct': _var_pct(ticket_act, ticket_ant),

            'documentos_actual': docs_act,
            'documentos_anterior': docs_ant,
            'variacion_documentos_pct': _var_pct(docs_act, docs_ant),

            'tasa_devolucion_actual': tasa_dev_act,
            'tasa_devolucion_anterior': tasa_dev_ant,
            'delta_tasa_devolucion_pp': round(tasa_dev_act - tasa_dev_ant, 1),

            'unidades_actual': actual['total_unidades'],
            'unidades_anterior': anterior['total_unidades'],
            'variacion_unidades_pct': _var_pct(actual['total_unidades'], anterior['total_unidades']),

            'descuentos_actual': actual['total_descuentos'],
            'descuentos_anterior': anterior['total_descuentos'],
        }

        # ---------- Por sucursal ----------
        sucursales_rows = []
        todas_sids = set(actual['sucursales'].keys()) | set(anterior['sucursales'].keys())
        for sid in todas_sids:
            a = actual['sucursales'].get(sid, {})
            b = anterior['sucursales'].get(sid, {})
            nombre = a.get('nombre') or b.get('nombre') or '-'
            v_brutas_a = a.get('ventas_brutas', 0)
            v_brutas_b = b.get('ventas_brutas', 0)
            dev_a = a.get('devoluciones', 0)
            dev_b = b.get('devoluciones', 0)
            netas_a = v_brutas_a - dev_a
            netas_b = v_brutas_b - dev_b
            docs_a = a.get('documentos', 0)
            docs_b = b.get('documentos', 0)
            tp_a = int(netas_a / docs_a) if docs_a > 0 else 0
            ecom_a = a.get('ventas_ecommerce', 0)
            ecom_b = b.get('ventas_ecommerce', 0)
            pct_int_a = _pct(ecom_a, v_brutas_a)
            pct_int_b = _pct(ecom_b, v_brutas_b)

            sucursales_rows.append({
                'id': sid,
                'nombre': nombre,
                'ventas_actual': netas_a,
                'ventas_anterior': netas_b,
                'variacion_pct': _var_pct(netas_a, netas_b),
                'variacion_abs': netas_a - netas_b,
                'pct_internet_actual': pct_int_a,
                'pct_internet_anterior': pct_int_b,
                'ventas_internet_actual': ecom_a,
                'documentos_actual': docs_a,
                'documentos_anterior': docs_b,
                'ticket_promedio_actual': tp_a,
                'devoluciones_actual': dev_a,
                'unidades_actual': a.get('unidades', 0),
                'participacion': _pct(netas_a, ventas_act),
            })
        sucursales_rows.sort(key=lambda x: x['ventas_actual'], reverse=True)

        # ---------- Por vendedor ----------
        vendedores_rows = []
        todas_vids = set(actual['vendedores'].keys()) | set(anterior['vendedores'].keys())
        for vid in todas_vids:
            a = actual['vendedores'].get(vid, {})
            b = anterior['vendedores'].get(vid, {})
            nombre = a.get('nombre') or b.get('nombre') or '-'
            codigo = a.get('codigo') or b.get('codigo') or ''
            sucursal = a.get('sucursal') or b.get('sucursal') or '-'
            v_brutas_a = a.get('ventas_brutas', 0)
            v_brutas_b = b.get('ventas_brutas', 0)
            dev_a = a.get('devoluciones', 0)
            dev_b = b.get('devoluciones', 0)
            netas_a = v_brutas_a - dev_a
            netas_b = v_brutas_b - dev_b
            docs_a = a.get('documentos', 0)
            docs_b = b.get('documentos', 0)
            tp_a = int(netas_a / docs_a) if docs_a > 0 else 0

            vendedores_rows.append({
                'id': vid,
                'nombre': nombre,
                'codigo': codigo,
                'sucursal': sucursal,
                'ventas_actual': netas_a,
                'ventas_anterior': netas_b,
                'variacion_pct': _var_pct(netas_a, netas_b),
                'variacion_abs': netas_a - netas_b,
                'documentos_actual': docs_a,
                'documentos_anterior': docs_b,
                'ticket_promedio_actual': tp_a,
                'participacion': _pct(netas_a, ventas_act),
            })
        vendedores_rows.sort(key=lambda x: x['ventas_actual'], reverse=True)

        # ---------- Por canal ----------
        canales_rows = []
        etiquetas_canal = {
            'ECOMMERCE': 'Internet (E-commerce)',
            'POS': 'Presencial (POS/Tickets)',
            'DTE': 'Documentos Tributarios (DTE)',
        }
        for key in ['ECOMMERCE', 'POS', 'DTE']:
            a = actual['canal'][key]
            b = anterior['canal'][key]
            canales_rows.append({
                'canal': key,
                'nombre': etiquetas_canal[key],
                'ventas_actual': a['ventas'],
                'ventas_anterior': b['ventas'],
                'variacion_pct': _var_pct(a['ventas'], b['ventas']),
                'documentos_actual': a['documentos'],
                'documentos_anterior': b['documentos'],
                'unidades_actual': a['unidades'],
                'unidades_anterior': b['unidades'],
                'participacion_actual': _pct(a['ventas'], actual['total_ventas_brutas']),
                'participacion_anterior': _pct(b['ventas'], anterior['total_ventas_brutas']),
            })

        # ---------- Top crecimiento ----------
        top_suc = max(sucursales_rows, key=lambda x: x['variacion_pct'], default=None)
        top_vend = max(vendedores_rows, key=lambda x: x['variacion_pct'], default=None)
        kpis['top_sucursal_crecimiento'] = (
            {'nombre': top_suc['nombre'], 'variacion': top_suc['variacion_pct']}
            if top_suc and top_suc['ventas_actual'] > 0 else None
        )
        kpis['top_vendedor_crecimiento'] = (
            {'nombre': top_vend['nombre'], 'variacion': top_vend['variacion_pct']}
            if top_vend and top_vend['ventas_actual'] > 0 else None
        )

        return JsonResponse({
            'success': True,
            'tipo_flujo': tipo_flujo,
            'periodo': {
                'actual': {
                    'inicio': fi_act.strftime('%Y-%m-%d'),
                    'fin': ff_act.strftime('%Y-%m-%d'),
                },
                'anterior': {
                    'inicio': fi_ant.strftime('%Y-%m-%d'),
                    'fin': ff_ant.strftime('%Y-%m-%d'),
                },
            },
            'kpis': kpis,
            'sucursales': sucursales_rows,
            'vendedores': vendedores_rows,
            'canales': canales_rows,
        })

    except Exception as e:
        logger.exception("Error al obtener comparativo de ventas")
        return JsonResponse({'success': False, 'error': str(e)})


# ========== REPORTE DE PRODUCTOS VENDIDOS ==========

def _aplicar_filtros_producto(qs, prefix, filtros):
    """Aplica filtros de atributo/categoría/temporada a un queryset de líneas de venta.
    'prefix' es el prefijo relacional al Producto, por ejemplo:
      - 'ProductoTalla__producto' para Ticket_Productos
      - 'productoTalla__producto' para Dte_Productos

    Por defecto excluye productos con Producto.excluir_de_analitica=True
    (productos marcados desde Gestión de Productos como no-analíticos:
    exhibición, consignación, catálogo retirado, etc.).
    Pasar filtros['incluir_excluidos']=True para incluirlos (auditoría).
    """
    # Exclusión por defecto de productos marcados como no-analíticos
    if not filtros.get('incluir_excluidos'):
        qs = qs.filter(**{f'{prefix}__excluir_de_analitica': False})

    if filtros.get('marca_id'):
        qs = qs.filter(**{f'{prefix}__atributo1_id': filtros['marca_id']})
    if filtros.get('color_id'):
        qs = qs.filter(**{f'{prefix}__atributo2_id': filtros['color_id']})
    if filtros.get('sexo_id'):
        qs = qs.filter(**{f'{prefix}__atributo3_id': filtros['sexo_id']})
    if filtros.get('genero_id'):
        qs = qs.filter(**{f'{prefix}__atributo4_id': filtros['genero_id']})
    if filtros.get('categoria_id'):
        # Árbol v1.2: filtrar por un PADRE debe incluir todas sus hijas
        # (patrón obtener_ventas_por_especialidad). El resultado expandido se
        # cachea en filtros para las 3 llamadas (tickets, dtes, devoluciones).
        if '_categoria_ids' not in filtros:
            cid = filtros['categoria_id']
            ids = [cid] + [
                str(x) for x in Categoria.objects.filter(padre_id=cid)
                .values_list('id', flat=True)
            ]
            filtros['_categoria_ids'] = ids
        qs = qs.filter(**{f'{prefix}__categoria_id__in': filtros['_categoria_ids']})
    if filtros.get('especialidad_id'):
        # Especialidad v1.2 (multi-etiqueta vía ProductoAtributoValor). Filtrar
        # por UNA opción no duplica líneas (una fila PAV por producto+opción).
        qs = qs.filter(**{f'{prefix}__atributos__opcion_id': filtros['especialidad_id']})
    if filtros.get('temporada'):
        qs = qs.filter(**{f'{prefix}__temporada': filtros['temporada']})
    if filtros.get('anio_temporada'):
        qs = qs.filter(**{f'{prefix}__anio_temporada': filtros['anio_temporada']})
    if filtros.get('rango_precio'):
        qs = qs.filter(**{f'{prefix}__rango_precio': filtros['rango_precio']})
    if filtros.get('busqueda'):
        q = filtros['busqueda']
        qs = qs.filter(
            Q(**{f'{prefix}__articulo__icontains': q}) |
            Q(**{f'{prefix}__descripcion__icontains': q})
        )
    return qs


def _sucursales_efectivas_reporte(user, request):
    """IDs de sucursal a los que quedó acotado el reporte, o None si no hay tope.

    Espeja la decisión de `filtrar_queryset_por_sucursal` (que sólo sabe
    filtrar querysets con FK a sucursal) para poder acotar también agregados
    que no cuelgan de la venta — el stock, por ejemplo. Lista vacía = el
    usuario no puede ver nada (equivalente a `queryset.none()`).
    """
    def _como_lista(valor):
        try:
            return [int(valor)]
        except (TypeError, ValueError):
            return []

    sucursal_id = request.GET.get('sucursal_id')
    if sucursal_id:
        if puede_ver_sucursal(user, sucursal_id):
            return _como_lista(sucursal_id)
        return _como_lista(request.session.get('idSucursalActual'))

    if usuario_puede_ver_todas_sucursales(user):
        return None

    return _como_lista(request.session.get('idSucursalActual'))


def _expr_monto_linea_dte():
    """Monto de una línea de DTE: `monto_item` y, si viene en 0, `precio * stock`.

    El fallback va por línea (no por producto) para que el mismo criterio
    sirva tanto en la agregación como en la corrección de IVA de abajo.
    """
    from django.db.models import Case, When
    return Case(
        When(monto_item__gt=0, then=ExpressionWrapper(
            F('monto_item'), output_field=DecimalField())),
        default=ExpressionWrapper(
            F('precio') * F('stock'), output_field=DecimalField()),
        output_field=DecimalField(),
    )


def _factores_iva_dte_lineas_netas(qs_lineas_base):
    """`{dte_id: factor}` de los DTE cuyas líneas están guardadas en NETO.

    El reporte de productos vendidos expresa TODO con IVA incluido (criterio
    único: es como se lee la venta al público y es lo que ya traen tickets y
    boletas). El problema es que `Dte_Productos` no tiene un criterio único:
    la mayoría de los documentos guarda la línea bruta, pero las guías y
    algunas facturas la guardan neta, así que decidir por `tipo_documento`
    daría un resultado incorrecto.

    Se decide documento por documento comparando la suma de sus líneas contra
    la cabecera: si empata con `monto_neto` y el documento es afecto
    (`monto_con_iva > monto_neto`), sus líneas están netas y se corrigen por
    `monto_con_iva / monto_neto`. Lo que no empata con ninguno de los dos
    (líneas editadas a mano, documentos parciales) se deja como está.

    Recibe el queryset de líneas SIN los filtros de producto: la comparación
    es contra la cabecera, así que necesita todas las líneas del documento
    (y de paso evita arrastrar los joins a Producto al subquery, que cuadruplican
    el costo de esta consulta).
    """
    tolerancia = Decimal('0.01')
    factores = {}
    filas = Dte_Productos.objects.filter(
        dte_id__in=qs_lineas_base.values('dte_id'),
        dte__monto_neto__gt=0,
        dte__monto_con_iva__gt=F('dte__monto_neto'),
    ).values(
        'dte_id',
        neto=F('dte__monto_neto'),
        bruto=F('dte__monto_con_iva'),
    ).annotate(total_lineas=Sum(_expr_monto_linea_dte()))

    for fila in filas:
        neto = Decimal(fila['neto'] or 0)
        bruto = Decimal(fila['bruto'] or 0)
        base = Decimal(fila['total_lineas'] or 0)
        if base <= 0 or neto <= 0:
            continue
        if abs(base - neto) <= neto * tolerancia:
            factores[fila['dte_id']] = bruto / neto
    return factores


def _agregar_productos_vendidos(fi, ff, filtros, user, request):
    """Agrega ventas a nivel producto desde Ticket_Productos + Dte_Productos
    aplicando la regla anti-doble-contado:
      - Tickets: solo los que NO generaron DTE (`dte_generado=False`).
      - DTEs: todos los de venta no-anulados/no-NC (incluyen boletas de tickets).

    CRITERIO DE MONTO: todo se expresa CON IVA. Tickets y boletas ya vienen
    brutos; las líneas que están guardadas en neto (guías, algunas facturas)
    se llevan a bruto con el factor de su propia cabecera
    (ver `_factores_iva_dte_lineas_netas`).

    Métricas calculadas por producto:
      - unidades (NETAS de devoluciones de cliente), unidades_brutas,
        unidades_devueltas, monto, monto_devuelto, costo, margen, margen_pct
      - docs (cantidad de ventas), sucursales_count
    Las devoluciones se restan desde CambioDevolucionDetalle (no desde la Nota
    de Crédito): las ventas por ticket no generan NC.

    Devuelve dict con:
      {
        'productos': [{producto_id, articulo, descripcion, marca, categoria,
                       sexo, genero, unidades, monto, costo, margen, margen_pct, docs}],
        'por_marca': [...], 'por_categoria': [...], 'por_sexo': [...],
        'por_genero': [...], 'por_sucursal_categoria': [{sucursal, categoria, unidades, monto}],
        'kpis': {...}
      }
    """
    # ---------- TICKET_PRODUCTOS (solo tickets sin DTE) ----------
    qs_tp = Ticket_Productos.objects.filter(
        idTicket__created_at__date__gte=fi,
        idTicket__created_at__date__lte=ff,
        idTicket__estado='PAGADO',
        idTicket__modulo_origen__in=['VENTA_PUBLICO', 'POS', 'ECOMMERCE'],
        idTicket__dte_generado=False,
        ProductoTalla__isnull=False,
    )
    # Permisos / sucursal (aplica al ticket padre)
    qs_tp = filtrar_queryset_por_sucursal(
        qs_tp, user, request, campo_sucursal='idTicket__sucursal_id'
    )
    qs_tp = _aplicar_filtros_producto(qs_tp, 'ProductoTalla__producto', filtros)

    # ---------- DTE_PRODUCTOS ----------
    qs_dp = Dte_Productos.objects.filter(
        dte__fecha_emision__gte=fi,
        dte__fecha_emision__lte=ff,
        dte__tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
        productoTalla__isnull=False,
    ).exclude(
        dte__estado_dte__in=['ANULADO', 'CANCELADO', 'RECHAZADO']
    ).exclude(
        dte__tipo_documento='NOTA DE CREDITO'
    ).exclude(
        dte__receptor__isnull=False,
        dte__receptor_id=F('dte__emisor_id'),
    )
    qs_dp_base = filtrar_queryset_por_sucursal(
        qs_dp, user, request, campo_sucursal='dte__sucursal_id'
    )
    qs_dp = _aplicar_filtros_producto(qs_dp_base, 'productoTalla__producto', filtros)

    # ---------- Agregación por producto ----------
    productos_acum = {}  # {producto_id: dict}
    sucursales_por_producto = {}  # {producto_id: set(sucursal_id)}

    def _get_prod(pid, data_row):
        nombre_cat = data_row.get('prod_categoria') or '-'
        padre_cat = data_row.get('prod_categoria_padre') or ''
        return productos_acum.setdefault(pid, {
            'producto_id': pid,
            'articulo': data_row.get('prod_articulo') or '-',
            'descripcion': data_row.get('prod_descripcion') or '',
            'marca': data_row.get('prod_marca') or '-',
            'marca_id': data_row.get('prod_marca_id'),
            'categoria': nombre_cat,
            # Árbol v1.2: label "Padre › Hijo" para desambiguar hijas homónimas
            # (patrón obtener_ventas_por_categoria).
            'categoria_label': f'{padre_cat} › {nombre_cat}' if padre_cat else nombre_cat,
            'categoria_id': data_row.get('prod_categoria_id'),
            'sexo': data_row.get('prod_sexo') or '-',
            'sexo_id': data_row.get('prod_sexo_id'),
            'genero': data_row.get('prod_genero') or '-',
            'genero_id': data_row.get('prod_genero_id'),
            'unidades': 0, 'monto': 0, 'costo': 0, 'docs': 0,
            'unidades_devueltas': 0, 'monto_devuelto': 0,
        })

    # --- Tickets ---
    agg_tp = qs_tp.values(
        'ProductoTalla__producto_id',
        prod_articulo=F('ProductoTalla__producto__articulo'),
        prod_descripcion=F('ProductoTalla__producto__descripcion'),
        prod_marca=F('ProductoTalla__producto__atributo1__valor'),
        prod_marca_id=F('ProductoTalla__producto__atributo1_id'),
        prod_categoria=F('ProductoTalla__producto__categoria__nombre'),
        prod_categoria_padre=F('ProductoTalla__producto__categoria__padre__nombre'),
        prod_categoria_id=F('ProductoTalla__producto__categoria_id'),
        prod_sexo=F('ProductoTalla__producto__atributo3__valor'),
        prod_sexo_id=F('ProductoTalla__producto__atributo3_id'),
        prod_genero=F('ProductoTalla__producto__atributo4__valor'),
        prod_genero_id=F('ProductoTalla__producto__atributo4_id'),
    ).annotate(
        unid=Sum('stock'),
        monto=Sum('subtotal'),
        # costo FIFO total = costo_fifo * cantidad
        costo_total=Sum(ExpressionWrapper(
            F('costo_fifo') * F('stock'), output_field=DecimalField()
        )),
        docs=Count('idTicket', distinct=True),
    )
    for r in agg_tp:
        pid = r['ProductoTalla__producto_id']
        if not pid:
            continue
        p = _get_prod(pid, r)
        p['unidades'] += int(r['unid'] or 0)
        p['monto'] += int(r['monto'] or 0)
        p['costo'] += int(r['costo_total'] or 0)
        p['docs'] += int(r['docs'] or 0)

    # Sucursales por producto (Tickets)
    for r in qs_tp.values('ProductoTalla__producto_id', 'idTicket__sucursal_id').distinct():
        pid = r['ProductoTalla__producto_id']
        sid = r['idTicket__sucursal_id']
        if pid and sid:
            sucursales_por_producto.setdefault(pid, set()).add(sid)

    # --- DTEs ---
    # monto por línea: `monto_item` y, si viene en 0, `precio * stock`.
    agg_dp = qs_dp.values(
        'productoTalla__producto_id',
        prod_articulo=F('productoTalla__producto__articulo'),
        prod_descripcion=F('productoTalla__producto__descripcion'),
        prod_marca=F('productoTalla__producto__atributo1__valor'),
        prod_marca_id=F('productoTalla__producto__atributo1_id'),
        prod_categoria=F('productoTalla__producto__categoria__nombre'),
        prod_categoria_padre=F('productoTalla__producto__categoria__padre__nombre'),
        prod_categoria_id=F('productoTalla__producto__categoria_id'),
        prod_sexo=F('productoTalla__producto__atributo3__valor'),
        prod_sexo_id=F('productoTalla__producto__atributo3_id'),
        prod_genero=F('productoTalla__producto__atributo4__valor'),
        prod_genero_id=F('productoTalla__producto__atributo4_id'),
    ).annotate(
        unid=Sum('stock'),
        monto_total=Sum(_expr_monto_linea_dte()),
        costo_total=Sum(ExpressionWrapper(
            F('costo') * F('stock'), output_field=DecimalField()
        )),
        docs=Count('dte', distinct=True),
    )
    for r in agg_dp:
        pid = r['productoTalla__producto_id']
        if not pid:
            continue
        p = _get_prod(pid, r)
        p['unidades'] += int(r['unid'] or 0)
        p['monto'] += int(r['monto_total'] or 0)
        p['costo'] += int(r['costo_total'] or 0)
        p['docs'] += int(r['docs'] or 0)

    # Llevar a BRUTO las líneas de los documentos guardados en neto: se suma
    # sólo el diferencial, sobre el mismo monto ya agregado arriba.
    factores_iva = _factores_iva_dte_lineas_netas(qs_dp_base)
    if factores_iva:
        for r in qs_dp.filter(dte_id__in=list(factores_iva)).values(
            'productoTalla__producto_id', 'dte_id',
        ).annotate(monto_total=Sum(_expr_monto_linea_dte())):
            p = productos_acum.get(r['productoTalla__producto_id'])
            if not p:
                continue
            base = Decimal(r['monto_total'] or 0)
            p['monto'] += int(base * factores_iva[r['dte_id']]) - int(base)

    # Sucursales por producto (DTEs)
    for r in qs_dp.values('productoTalla__producto_id', 'dte__sucursal_id').distinct():
        pid = r['productoTalla__producto_id']
        sid = r['dte__sucursal_id']
        if pid and sid:
            sucursales_por_producto.setdefault(pid, set()).add(sid)

    # ---------- DEVOLUCIONES DE CLIENTE (netear unidades) ----------
    # Las unidades vendidas deben ser NETAS de lo devuelto. La fuente es
    # CambioDevolucionDetalle (no la Nota de Crédito): las ventas por ticket
    # no generan NC, y el kardex pierde las devoluciones "no aptas" (cant=0).
    # Aquí capturamos toda devolución/cambio de mercadería una sola vez, sin
    # solape (el reporte no cuenta las NC como venta).
    qs_dev = CambioDevolucionDetalle.objects.filter(
        producto_original__isnull=False,
        cantidad_original__gt=0,
        cambio_devolucion__fecha_ejecucion__date__gte=fi,
        cambio_devolucion__fecha_ejecucion__date__lte=ff,
    )
    qs_dev = filtrar_queryset_por_sucursal(
        qs_dev, user, request, campo_sucursal='cambio_devolucion__sucursal_id'
    )
    qs_dev = _aplicar_filtros_producto(
        qs_dev, 'producto_original__ProductoTalla__producto', filtros
    )
    agg_dev = qs_dev.values('producto_original__ProductoTalla__producto_id').annotate(
        dev_u=Sum('cantidad_original'),
        dev_m=Sum(ExpressionWrapper(
            F('precio_original_unitario') * F('cantidad_original'),
            output_field=DecimalField(),
        )),
    )
    dev_omitidas = 0
    for r in agg_dev:
        pid = r['producto_original__ProductoTalla__producto_id']
        p = productos_acum.get(pid)
        if not p:
            # Devolución de un SKU sin ventas en el período: no restamos para
            # no generar filas negativas fantasma.
            dev_omitidas += int(r['dev_u'] or 0)
            continue
        p['unidades_devueltas'] += int(r['dev_u'] or 0)
        p['monto_devuelto'] += int(r['dev_m'] or 0)
        p['unidades'] -= int(r['dev_u'] or 0)   # unidades = NETAS de devoluciones
        p['monto'] -= int(r['dev_m'] or 0)
    if dev_omitidas:
        logger.info(
            'productos_vendidos: %s unidades devueltas omitidas del neteo '
            '(SKUs sin ventas en el período %s..%s)', dev_omitidas, fi, ff
        )

    # ---------- Stock actual por producto (sell-through / cobertura) ----------
    # El denominador tiene que vivir en el mismo universo que el numerador: si
    # el reporte está acotado a una sucursal, el stock también. Antes se
    # comparaban ventas de una sucursal contra el stock de toda la red.
    sucursales_stock = _sucursales_efectivas_reporte(user, request)
    stock_por_producto = {}
    if productos_acum:
        qs_stock = Producto_Talla.objects.filter(
            producto_id__in=productos_acum.keys(), stock__gt=0,
        )
        if sucursales_stock is not None:
            qs_stock = qs_stock.filter(producto__sucursal_id__in=sucursales_stock)
        stock_por_producto = {
            r['producto_id']: int(r['s'] or 0)
            for r in qs_stock.values('producto_id').annotate(s=Sum('stock'))
        }
    dias_periodo = max((ff - fi).days + 1, 1)

    # ---------- Calcular margen por producto + totales ----------
    productos = []
    tot_unid = 0
    tot_monto = 0
    tot_costo = 0
    tot_stock = 0
    tot_dev = 0
    for pid, p in productos_acum.items():
        margen = p['monto'] - p['costo']
        margen_pct = (margen / p['monto'] * 100) if p['monto'] > 0 else 0
        sucs = len(sucursales_por_producto.get(pid, set()))
        stock_actual = stock_por_producto.get(pid, 0)
        disponible = p['unidades'] + stock_actual
        velocidad_dia = p['unidades'] / dias_periodo if p['unidades'] else 0
        productos.append({
            **p,
            'unidades_brutas': p['unidades'] + p['unidades_devueltas'],
            'margen': margen,
            'margen_pct': round(margen_pct, 1),
            'sucursales_count': sucs,
            'stock_actual': stock_actual,
            'sell_through': round(100 * p['unidades'] / disponible, 1) if disponible else 0,
            'cobertura_dias': int(stock_actual / velocidad_dia) if velocidad_dia else None,
        })
        tot_unid += p['unidades']
        tot_monto += p['monto']
        tot_costo += p['costo']
        tot_stock += stock_actual
        tot_dev += p['unidades_devueltas']

    tot_margen = tot_monto - tot_costo
    tot_margen_pct = (tot_margen / tot_monto * 100) if tot_monto > 0 else 0
    tot_disponible = tot_unid + tot_stock
    tot_sell_through = round(100 * tot_unid / tot_disponible, 1) if tot_disponible else 0
    tot_cobertura = (int(tot_stock / (tot_unid / dias_periodo))
                     if tot_unid else None)

    # Orden según parámetro
    orden = filtros.get('orden', 'unidades_desc')
    key_map = {
        'unidades_desc': lambda x: x['unidades'],
        'monto_desc':    lambda x: x['monto'],
        'margen_desc':   lambda x: x['margen'],
        'margen_pct_desc': lambda x: x['margen_pct'],
    }
    productos.sort(key=key_map.get(orden, key_map['unidades_desc']), reverse=True)

    # Participación % (en unidades)
    for p in productos:
        p['participacion'] = round(p['unidades'] / tot_unid * 100, 1) if tot_unid > 0 else 0

    # ---------- Agregaciones por dimensión ----------
    def _agrupar_por(campo_key, campo_nombre):
        acc = {}
        for p in productos:
            key = p.get(campo_key) or None
            nombre = p.get(campo_nombre) or 'Sin clasificar'
            slot = acc.setdefault(key, {
                'id': key,
                'nombre': nombre,
                'unidades': 0, 'monto': 0, 'costo': 0,
                'skus': 0,
            })
            slot['unidades'] += p['unidades']
            slot['monto'] += p['monto']
            slot['costo'] += p['costo']
            slot['skus'] += 1
        filas = []
        for s in acc.values():
            m = s['monto'] - s['costo']
            filas.append({
                **s,
                'margen': m,
                'margen_pct': round((m / s['monto'] * 100), 1) if s['monto'] > 0 else 0,
                'participacion': round(s['unidades'] / tot_unid * 100, 1) if tot_unid > 0 else 0,
            })
        filas.sort(key=lambda x: x['unidades'], reverse=True)
        return filas

    por_marca = _agrupar_por('marca_id', 'marca')
    # Árbol v1.2: agrupar por id pero mostrar "Padre › Hijo" (desambigua
    # hijas homónimas de padres distintos).
    por_categoria = _agrupar_por('categoria_id', 'categoria_label')
    por_sexo = _agrupar_por('sexo_id', 'sexo')
    por_genero = _agrupar_por('genero_id', 'genero')

    # ---------- Agregación por ESPECIALIDAD v1.2 (multi-etiqueta) ----------
    # Vista de ATRIBUCIÓN, no partición: un producto [running, urbano] suma a
    # ambas etiquetas, así que la suma puede exceder el total (mismo criterio
    # que obtener_ventas_por_especialidad del dashboard).
    por_especialidad = []
    if productos_acum:
        from app.models import ProductoAtributoValor
        esp_por_prod = {}
        for pav in ProductoAtributoValor.objects.filter(
            producto_id__in=productos_acum.keys(),
            atributo__nombre__iexact='Especialidad',
        ).values('producto_id', 'opcion_id', 'opcion__valor'):
            esp_por_prod.setdefault(pav['producto_id'], []).append(
                (pav['opcion_id'], pav['opcion__valor'])
            )
        esp_acc = {}
        for pid, p in productos_acum.items():
            for eid, valor in esp_por_prod.get(pid, []):
                slot = esp_acc.setdefault(eid, {
                    'id': eid, 'nombre': valor,
                    'unidades': 0, 'monto': 0, 'skus': 0,
                })
                slot['unidades'] += p['unidades']
                slot['monto'] += p['monto']
                slot['skus'] += 1
        por_especialidad = sorted(
            esp_acc.values(), key=lambda x: x['unidades'], reverse=True)
        for e in por_especialidad:
            e['participacion'] = round(e['unidades'] / tot_unid * 100, 1) if tot_unid > 0 else 0

    # ---------- Heatmap Sucursal × Categoría ----------
    # Recalcular directo en DB para mejor precisión (evitar perder unidades por producto sin categoría)
    heat = {}  # {(sid, cat_nombre): {unidades, monto}}

    def _cat_label(row):
        nombre = row.get('cat_nombre') or 'Sin clasificar'
        padre = row.get('cat_padre') or ''
        return f'{padre} › {nombre}' if padre else nombre

    for r in qs_tp.values(
        'idTicket__sucursal_id',
        suc_nombre=F('idTicket__sucursal__alias'),
        cat_nombre=F('ProductoTalla__producto__categoria__nombre'),
        cat_padre=F('ProductoTalla__producto__categoria__padre__nombre'),
    ).annotate(unid=Sum('stock'), monto=Sum('subtotal')):
        sid = r['idTicket__sucursal_id']
        suc = r['suc_nombre'] or '-'
        cat = _cat_label(r)
        key = (sid, cat)
        h = heat.setdefault(key, {'sucursal_id': sid, 'sucursal': suc,
                                   'categoria': cat, 'unidades': 0, 'monto': 0})
        h['unidades'] += int(r['unid'] or 0)
        h['monto'] += int(r['monto'] or 0)

    for r in qs_dp.values(
        'dte__sucursal_id',
        suc_nombre=F('dte__sucursal__alias'),
        cat_nombre=F('productoTalla__producto__categoria__nombre'),
        cat_padre=F('productoTalla__producto__categoria__padre__nombre'),
    ).annotate(unid=Sum('stock'), monto_total=Sum(_expr_monto_linea_dte())):
        sid = r['dte__sucursal_id']
        suc = r['suc_nombre'] or '-'
        cat = _cat_label(r)
        key = (sid, cat)
        h = heat.setdefault(key, {'sucursal_id': sid, 'sucursal': suc,
                                   'categoria': cat, 'unidades': 0, 'monto': 0})
        h['unidades'] += int(r['unid'] or 0)
        h['monto'] += int(r['monto_total'] or 0)

    # Mismo diferencial neto→bruto que en la agregación por producto.
    if factores_iva:
        for r in qs_dp.filter(dte_id__in=list(factores_iva)).values(
            'dte__sucursal_id', 'dte_id',
            cat_nombre=F('productoTalla__producto__categoria__nombre'),
            cat_padre=F('productoTalla__producto__categoria__padre__nombre'),
        ).annotate(monto_total=Sum(_expr_monto_linea_dte())):
            h = heat.get((r['dte__sucursal_id'], _cat_label(r)))
            if not h:
                continue
            base = Decimal(r['monto_total'] or 0)
            h['monto'] += int(base * factores_iva[r['dte_id']]) - int(base)

    # Restar devoluciones del heatmap (solo celdas ya existentes: evita crear
    # celdas negativas para categorías/sucursales sin ventas en el período).
    for r in qs_dev.values(
        'cambio_devolucion__sucursal_id',
        cat_nombre=F('producto_original__ProductoTalla__producto__categoria__nombre'),
        cat_padre=F('producto_original__ProductoTalla__producto__categoria__padre__nombre'),
    ).annotate(unid=Sum('cantidad_original'),
               monto=Sum(ExpressionWrapper(
                   F('precio_original_unitario') * F('cantidad_original'),
                   output_field=DecimalField()))):
        sid = r['cambio_devolucion__sucursal_id']
        cat = _cat_label(r)
        h = heat.get((sid, cat))
        if h:
            h['unidades'] -= int(r['unid'] or 0)
            h['monto'] -= int(r['monto'] or 0)

    heatmap = sorted(
        heat.values(),
        key=lambda x: (x['sucursal'], -x['unidades']),
    )

    # ---------- KPIs ----------
    top_marca = por_marca[0]['nombre'] if por_marca else '-'
    top_categoria = por_categoria[0]['nombre'] if por_categoria else '-'
    top_producto = productos[0]['articulo'] if productos else '-'

    kpis = {
        'total_unidades': tot_unid,
        'total_devoluciones_unid': tot_dev,
        'total_monto': tot_monto,
        'total_costo': tot_costo,
        'total_margen': tot_margen,
        'total_margen_pct': round(tot_margen_pct, 1),
        'total_skus': len(productos),
        'ticket_promedio_unid': round(tot_monto / tot_unid, 0) if tot_unid > 0 else 0,
        'top_marca': top_marca,
        'top_categoria': top_categoria,
        'top_producto': top_producto,
        # Salud de inventario del universo vendido en el período:
        'stock_actual_total': tot_stock,
        'sell_through_periodo': tot_sell_through,
        'cobertura_dias': tot_cobertura,
        'dias_periodo': dias_periodo,
        # Alcance del denominador de sell-through/cobertura y criterio de monto,
        # para que el consumidor sepa qué está comparando.
        'stock_alcance': 'red' if sucursales_stock is None else 'sucursal',
        'criterio_monto': 'CON_IVA',
    }

    return {
        'productos': productos,
        'por_marca': por_marca,
        'por_categoria': por_categoria,
        'por_sexo': por_sexo,
        'por_genero': por_genero,
        'por_especialidad': por_especialidad,
        'heatmap': heatmap,
        'kpis': kpis,
    }


@login_required
def ver_reporte_productos_vendidos(request):
    """Vista principal del reporte de productos vendidos."""
    context = obtener_contexto_sucursales(request.user, request)
    return render(request, 'vistas/modulo_reportes/reporte_productos_vendidos.html', context)


@require_GET
@login_required
def obtener_productos_vendidos(request):
    """API: productos vendidos con agregaciones por marca/categoría/sexo/género y heatmap.

    Parámetros (todos opcionales):
      - tipo_flujo / fecha_inicio / fecha_fin
      - sucursal_id
      - marca_id, color_id, sexo_id, genero_id, categoria_id
      - temporada, anio_temporada, rango_precio
      - busqueda (texto libre)
      - orden: unidades_desc | monto_desc | margen_desc | margen_pct_desc
      - top_n (default 50) — limita sólo la tabla de productos, no las agregaciones
    """
    try:
        # Mismo motivo que en `obtener_ventas_comparativo`: `mes_full`
        # proyecta el mes en curso como si estuviera cerrado y deja al
        # reporte mostrando una caída que no existe.
        tipo_flujo = request.GET.get('tipo_flujo', 'mes_mtd')
        fi_param = request.GET.get('fecha_inicio')
        ff_param = request.GET.get('fecha_fin')

        # Reusar helper de rangos del reporte comparativo
        fi, ff, _fi_ant, _ff_ant = _calcular_rangos_comparativo(
            tipo_flujo, fi_param, ff_param
        )

        filtros = {
            'marca_id': request.GET.get('marca_id') or None,
            'color_id': request.GET.get('color_id') or None,
            'sexo_id': request.GET.get('sexo_id') or None,
            'genero_id': request.GET.get('genero_id') or None,
            'categoria_id': request.GET.get('categoria_id') or None,
            'especialidad_id': request.GET.get('especialidad_id') or None,
            'temporada': request.GET.get('temporada') or None,
            'anio_temporada': request.GET.get('anio_temporada') or None,
            'rango_precio': request.GET.get('rango_precio') or None,
            'busqueda': (request.GET.get('busqueda') or '').strip() or None,
            'orden': request.GET.get('orden', 'unidades_desc'),
            # Por defecto excluye productos marcados como no-analíticos
            # (exhibición, consignación, etc.) — misma lógica que verGestionProducto
            'incluir_excluidos': request.GET.get('incluir_excluidos') in ('1', 'true', 'True'),
        }

        try:
            top_n = int(request.GET.get('top_n', 50))
        except (TypeError, ValueError):
            top_n = 50
        top_n = max(1, min(top_n, 500))

        data = _agregar_productos_vendidos(fi, ff, filtros, request.user, request)

        return JsonResponse({
            'success': True,
            'tipo_flujo': tipo_flujo,
            'periodo': {
                'inicio': fi.strftime('%Y-%m-%d'),
                'fin': ff.strftime('%Y-%m-%d'),
            },
            'kpis': data['kpis'],
            'productos': data['productos'][:top_n],
            'productos_total': len(data['productos']),
            'por_marca': data['por_marca'][:100],
            'por_categoria': data['por_categoria'][:100],
            'por_sexo': data['por_sexo'],
            'por_genero': data['por_genero'],
            'por_especialidad': data['por_especialidad'],
            'heatmap': data['heatmap'],
        })

    except Exception as e:
        logger.exception("Error al obtener reporte productos vendidos")
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
@login_required
def obtener_atributo_opciones(request):
    """API: lista de opciones de un atributo (marca, color, sexo, género) o categorías.

    Parámetro:
      - tipo: 'marca' | 'color' | 'sexo' | 'genero' | 'categoria'
    """
    try:
        tipo = (request.GET.get('tipo') or '').lower()
        mapping = {
            'marca': 'Marca',
            'color': 'Color',
            'sexo': 'Sexo',
            'genero': 'Género',
        }

        if tipo == 'categoria':
            # Árbol v1.2: padres primero (filtran su rama completa) y luego
            # sus hijas con label "Padre › Hija". Se ocultan las _ZZ_
            # (categorías planas deprecadas por la recategorización).
            padres = list(
                Categoria.objects.filter(padre__isnull=True)
                .exclude(nombre__startswith='_ZZ_')
                .order_by('nombre').values('id', 'nombre')
            )
            hijas_por_padre = {}
            for h in Categoria.objects.filter(padre__isnull=False) \
                    .exclude(nombre__startswith='_ZZ_') \
                    .order_by('nombre').values('id', 'nombre', 'padre_id'):
                hijas_por_padre.setdefault(h['padre_id'], []).append(h)
            opciones = []
            for p in padres:
                hijas = hijas_por_padre.get(p['id'], [])
                if hijas:
                    opciones.append({'id': p['id'], 'nombre': f"{p['nombre']} (toda la rama)"})
                    for h in hijas:
                        opciones.append({'id': h['id'], 'nombre': f"{p['nombre']} › {h['nombre']}"})
                else:
                    # Plana vieja aún con productos: visible pero marcada
                    opciones.append({'id': p['id'], 'nombre': f"{p['nombre']} (sin recategorizar)"})
            return JsonResponse({'success': True, 'opciones': opciones})

        if tipo == 'especialidad':
            qs = AtributoOpcion.objects.filter(
                atributo__nombre__iexact='Especialidad'
            ).order_by('valor').values('id', 'valor')
            return JsonResponse({
                'success': True,
                'opciones': [{'id': x['id'], 'nombre': x['valor']} for x in qs],
            })

        nombre_atributo = mapping.get(tipo)
        if not nombre_atributo:
            return JsonResponse({
                'success': False,
                'error': "Parametro 'tipo' invalido. Use: marca | color | sexo | genero | categoria | especialidad",
            })

        qs = AtributoOpcion.objects.filter(
            atributo__nombre=nombre_atributo
        ).order_by('valor').values('id', 'valor')

        return JsonResponse({
            'success': True,
            'opciones': [{'id': x['id'], 'nombre': x['valor']} for x in qs],
        })
    except Exception as e:
        logger.exception("Error al obtener opciones de atributo para reportes")
        return JsonResponse({'success': False, 'error': str(e)})


# ========== REPORTE DE DESPACHOS A TIENDAS (TRASPASO_SALIDA) ==========

@login_required
def ver_reporte_despachos_tiendas(request):
    """
    Vista principal del reporte de despachos a tiendas (traspasos de salida).

    Muestra lo que las sucursales del usuario despacharon hacia las tiendas en
    un rango de fechas (por defecto, la semana actual), con cantidades por
    artículo, agrupado por tienda destino y filtrable por proveedor.
    """
    empresas_usuario = EmpresaUser.objects.filter(
        user=request.user,
        status=True,
    ).values_list('empresa_id', flat=True)
    sucursales = Sucursal.objects.filter(empresa_id__in=empresas_usuario).order_by('alias')

    hoy = timezone.localdate()
    lunes = hoy - timedelta(days=hoy.weekday())

    context = {
        'sucursales': sucursales,
        'fecha_desde_default': lunes.strftime('%Y-%m-%d'),
        'fecha_hasta_default': hoy.strftime('%Y-%m-%d'),
    }
    return render(request, 'vistas/modulo_reportes/reporte_despachos_tiendas.html', context)


@require_GET
@login_required
def obtener_reporte_despachos_tiendas(request):
    """
    API del reporte de despachos a tiendas.

    Devuelve, para el rango de fechas indicado (por defecto la semana actual),
    los despachos (Movimientos_Producto con concepto TRASPASO_SALIDA, estado
    COMPLETADO) agrupados por artículo y tienda destino, con cantidades.

    El proveedor se deriva de la última recepción de compra (DTE COMPRA) de cada
    artículo: es una aproximación ("último proveedor del artículo"), no un dato
    propio del traspaso.
    """
    try:
        # ===== PARÁMETROS =====
        fecha_desde_str = request.GET.get('fecha_desde', '').strip()
        fecha_hasta_str = request.GET.get('fecha_hasta', '').strip()
        sucursal_destino_id = request.GET.get('sucursal_destino_id') or None
        proveedor_id = request.GET.get('proveedor_id') or None

        hoy = timezone.localdate()
        try:
            fecha_desde = (datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
                           if fecha_desde_str else hoy - timedelta(days=hoy.weekday()))
        except ValueError:
            fecha_desde = hoy - timedelta(days=hoy.weekday())
        try:
            fecha_hasta = (datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
                           if fecha_hasta_str else hoy)
        except ValueError:
            fecha_hasta = hoy

        if proveedor_id:
            try:
                proveedor_id = int(proveedor_id)
            except (TypeError, ValueError):
                proveedor_id = None

        # ===== SUCURSALES DEL USUARIO (despachadoras = origen) =====
        empresas_usuario = EmpresaUser.objects.filter(
            user=request.user,
            status=True,
        ).values_list('empresa_id', flat=True)
        sucursales_ids = list(
            Sucursal.objects.filter(empresa_id__in=empresas_usuario).values_list('id', flat=True)
        )

        # ===== MOVIMIENTOS DE DESPACHO (TRASPASO_SALIDA) =====
        movimientos = Movimientos_Producto.objects.filter(
            concepto='TRASPASO_SALIDA',
            estado='COMPLETADO',
            fecha__gte=fecha_desde,
            fecha__lte=fecha_hasta,
            sucursal_origen_id__in=sucursales_ids,
        )
        if sucursal_destino_id:
            movimientos = movimientos.filter(sucursal_destino_id=sucursal_destino_id)

        # Se agrupa por CÓDIGO de artículo (no por producto_id): en este catálogo
        # el mismo `articulo` existe como varias filas Producto (una por sucursal),
        # y el usuario razona en códigos de artículo.
        # `cantidad` de TRASPASO_SALIDA es negativa (egreso). En un traspaso interno:
        #   costo       = costo origen (lo que costó al origen)
        #   sobreprecio = recargo de traspaso
        #   precio      = costo destino = costo + sobreprecio (lo que se carga al destino)
        # NO es precio de venta al público. Valorizamos los tres por separado.
        agregado = list(movimientos.values(
            'ProductoTalla__producto__articulo',
            'ProductoTalla__producto__descripcion',
            'sucursal_destino__alias',
        ).annotate(
            uds=Sum('cantidad'),
            valor_costo=Sum(F('cantidad') * F('costo')),
            valor_sobreprecio=Sum(F('cantidad') * F('sobreprecio')),
            valor_destino=Sum(F('cantidad') * F('precio')),
            lineas=Count('id'),
        ))

        # ===== CONTRAPARTE RECIBIDA EN DESTINO (TRASPASO_ENTRADA) =====
        # Se atribuye a cada proveedor vía el artículo (igual que el despacho).
        entradas = Movimientos_Producto.objects.filter(
            concepto='TRASPASO_ENTRADA',
            estado='COMPLETADO',
            fecha__gte=fecha_desde,
            fecha__lte=fecha_hasta,
            sucursal_origen_id__in=sucursales_ids,
        )
        if sucursal_destino_id:
            entradas = entradas.filter(sucursal_destino_id=sucursal_destino_id)
        entradas_por_articulo = list(entradas.values(
            'ProductoTalla__producto__articulo',
        ).annotate(uds=Sum('cantidad')))

        # ===== DERIVAR PROVEEDOR POR ARTÍCULO (última recepción de compra) =====
        articulos = {
            row['ProductoTalla__producto__articulo']
            for row in agregado if row['ProductoTalla__producto__articulo']
        } | {
            row['ProductoTalla__producto__articulo']
            for row in entradas_por_articulo if row['ProductoTalla__producto__articulo']
        }
        proveedor_por_articulo = {}
        if articulos:
            recepciones = Productos_Recepcionados.objects.filter(
                producto_talla__producto__articulo__in=articulos,
                dte__tipo_transaccion='COMPRA',
                dte__emisor__isnull=False,
            ).select_related(
                'dte', 'dte__emisor', 'producto_talla__producto'
            ).order_by('dte__fecha_emision')
            # Orden ascendente: la última recepción (más reciente) sobrescribe.
            for rec in recepciones:
                art = rec.producto_talla.producto.articulo
                proveedor_por_articulo[art] = {
                    'id': rec.dte.emisor_id,
                    'nombre': rec.dte.emisor.nombre,
                }

        SIN_PROV = 'Sin recepción'

        def _coincide_filtro(prov):
            return not proveedor_id or (prov and prov['id'] == proveedor_id)

        # ===== PIVOTE ARTÍCULO × TIENDA (+ valorización) + ACUMULADO POR PROVEEDOR =====
        productos = {}
        prov_acc = {}

        for row in agregado:
            art = row['ProductoTalla__producto__articulo']
            uds = abs(row['uds'] or 0)
            if uds == 0:
                continue

            prov = proveedor_por_articulo.get(art)
            if not _coincide_filtro(prov):  # filtro por proveedor (derivado del artículo)
                continue

            v_costo = abs(row['valor_costo'] or 0)
            v_sobre = abs(row['valor_sobreprecio'] or 0)
            v_destino = abs(row['valor_destino'] or 0)
            suc_alias = row['sucursal_destino__alias'] or 'Sin tienda'
            prov_nombre = prov['nombre'] if prov else SIN_PROV

            # --- pivote por artículo ---
            if art not in productos:
                productos[art] = {
                    'articulo': art,
                    'descripcion': row['ProductoTalla__producto__descripcion'] or '',
                    'proveedor': prov_nombre,
                    'tiendas': {},
                    'total': 0,
                    'valor_costo': 0,
                    'valor_sobreprecio': 0,
                    'valor_destino': 0,
                }
            productos[art]['tiendas'][suc_alias] = (
                productos[art]['tiendas'].get(suc_alias, 0) + uds
            )
            productos[art]['total'] += uds
            productos[art]['valor_costo'] += v_costo
            productos[art]['valor_sobreprecio'] += v_sobre
            productos[art]['valor_destino'] += v_destino

            # --- acumulado por proveedor ---
            acc = prov_acc.setdefault(prov_nombre, {
                'proveedor': prov_nombre, 'despachado': 0, 'recibido': 0,
                'valor_costo': 0, 'valor_sobreprecio': 0, 'valor_destino': 0,
                '_articulos': set(), '_tiendas': set(),
            })
            acc['despachado'] += uds
            acc['valor_costo'] += v_costo
            acc['valor_sobreprecio'] += v_sobre
            acc['valor_destino'] += v_destino
            acc['_articulos'].add(art)
            acc['_tiendas'].add(suc_alias)

        datos = sorted(productos.values(), key=lambda p: p['total'], reverse=True)

        # Tiendas presentes (columnas del pivote), tras filtros
        tiendas_presentes = set()
        for p in datos:
            tiendas_presentes.update(p['tiendas'].keys())
        tiendas_list = sorted(tiendas_presentes)

        # Atribuir lo RECIBIDO a cada proveedor (vía artículo)
        for row in entradas_por_articulo:
            art = row['ProductoTalla__producto__articulo']
            recibido = abs(row['uds'] or 0)
            if recibido == 0:
                continue
            prov = proveedor_por_articulo.get(art)
            if not _coincide_filtro(prov):
                continue
            prov_nombre = prov['nombre'] if prov else SIN_PROV
            acc = prov_acc.setdefault(prov_nombre, {
                'proveedor': prov_nombre, 'despachado': 0, 'recibido': 0,
                'valor_costo': 0, 'valor_sobreprecio': 0, 'valor_destino': 0,
                '_articulos': set(), '_tiendas': set(),
            })
            acc['recibido'] += recibido

        # ===== TOTALES =====
        total_despachado = sum(p['total'] for p in datos)
        total_valor_costo = sum(p['valor_costo'] for p in datos)
        total_valor_sobreprecio = sum(p['valor_sobreprecio'] for p in datos)
        total_valor_destino = sum(p['valor_destino'] for p in datos)
        total_recibido = sum(a['recibido'] for a in prov_acc.values())

        # ===== POR PROVEEDOR (indicadores finales) =====
        por_proveedor = []
        for acc in prov_acc.values():
            desp = acc['despachado']
            rec = acc['recibido']
            por_proveedor.append({
                'proveedor': acc['proveedor'],
                'despachado': desp,
                'recibido': rec,
                'pendiente': max(desp - rec, 0),
                'tasa_recepcion': round(rec / desp * 100, 1) if desp > 0 else 0,
                'pct_unidades': round(desp / total_despachado * 100, 1) if total_despachado > 0 else 0,
                'valor_costo': acc['valor_costo'],
                'valor_sobreprecio': acc['valor_sobreprecio'],
                'valor_destino': acc['valor_destino'],
                'articulos': len(acc['_articulos']),
                'tiendas': len(acc['_tiendas']),
            })
        por_proveedor.sort(key=lambda x: x['despachado'], reverse=True)

        # ===== RESUMEN GLOBAL =====
        resumen = {
            'total_articulos': len(datos),
            'total_proveedores': len([p for p in por_proveedor if p['despachado'] > 0]),
            'total_despachado': total_despachado,
            'total_recibido': total_recibido,
            'total_pendiente': max(total_despachado - total_recibido, 0),
            'tasa_recepcion': round(total_recibido / total_despachado * 100, 1) if total_despachado > 0 else 0,
            'total_tiendas': len(tiendas_list),
            'valor_costo': total_valor_costo,
            'valor_sobreprecio': total_valor_sobreprecio,
            'valor_destino': total_valor_destino,
        }

        return JsonResponse({
            'success': True,
            'datos': datos,
            'por_proveedor': por_proveedor,
            'tiendas': tiendas_list,
            'resumen': resumen,
            'parametros': {
                'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
                'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
            },
        })

    except Exception as e:
        logger.exception("Error en reporte de despachos a tiendas")
        return JsonResponse({
            'success': False,
            'error': f'Error al generar reporte: {str(e)}',
        }, status=500)
