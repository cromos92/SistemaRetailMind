from datetime import datetime, time, timedelta
from io import BytesIO

from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from app.models import PedidoEcommerce, Ticket, TicketDetallePago, Ticket_Productos
from .factories import (
    crear_empresa,
    crear_empresa_user,
    crear_producto_con_talla,
    crear_sucursal,
    crear_usuario,
    crear_vendedor,
)


STATICFILES_STORAGE_TEST = 'django.contrib.staticfiles.storage.StaticFilesStorage'


@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class ReporteVentasInternetTest(TestCase):
    def setUp(self):
        self.empresa = crear_empresa(nombre='Empresa Web', rut='76.111.222-3')
        self.sucursal = crear_sucursal(empresa=self.empresa, alias='WEB-1')
        self.vendedor = crear_vendedor(
            nombre='Vendedor Internet', empresa=self.empresa, codigo_vendedor='1000'
        )
        self.vendedor.sucursales.add(self.sucursal)
        self.usuario = crear_usuario(username='reportes', rol='administrador')
        self.producto, self.producto_talla = crear_producto_con_talla(
            self.sucursal, articulo='Zapatilla Web', sku=990001, stock=10,
            precioventa=30000,
        )
        self.client = Client()
        self.client.login(username='reportes', password='TestPass123!')
        self.url = reverse('obtener_reporte_ventas_internet')

    def _set_fecha_ticket(self, ticket, fecha):
        dt = timezone.make_aware(datetime.combine(fecha, time(12, 0)))
        Ticket.objects.filter(pk=ticket.pk).update(fecha=fecha, hora=time(12, 0), created_at=dt)
        ticket.refresh_from_db()
        return ticket

    def _ticket(self, correlativo, total=50000, sucursal=None, vendedor=None,
                modulo_origen='ECOMMERCE', fecha=None):
        ticket_sucursal = sucursal or self.sucursal
        ticket = Ticket.objects.create(
            vendedor=vendedor or self.vendedor,
            sucursal=ticket_sucursal,
            correlativo=correlativo,
            estado='PAGADO',
            subTotal=total,
            descuento=0,
            total=total,
            responsable='Ecommerce',
            metodo_pago='VENTA_INTERNET',
            modulo_origen=modulo_origen,
        )
        if fecha:
            self._set_fecha_ticket(ticket, fecha)
        producto_talla = self.producto_talla
        if ticket_sucursal != self.sucursal:
            _, producto_talla = crear_producto_con_talla(
                ticket_sucursal,
                articulo=f'Producto Web {correlativo}',
                sku=990001 + int(correlativo),
                stock=10,
                precioventa=total // 2,
            )
        Ticket_Productos.objects.create(
            ProductoTalla=producto_talla,
            idTicket=ticket,
            stock=2,
            precio=total // 2,
            subtotal=total,
            precio_original=total // 2,
        )
        return ticket

    def _pago(self, ticket, monto=None, metodo='VENTA_INTERNET', plataforma='Shopify'):
        return TicketDetallePago.objects.create(
            ticket=ticket,
            metodo_pago=metodo,
            tipo_tarjeta=plataforma,
            voucher=f'V-{ticket.correlativo}-{metodo}',
            monto=monto if monto is not None else ticket.total,
        )

    def _pedido(self, numero, ticket, estado='FACTURADO', total=50000, fecha=None):
        fecha = fecha or timezone.now()
        return PedidoEcommerce.objects.create(
            numero_ticket_rm=f'RM-{numero}',
            numero_pedido_canal=f'CANAL-{numero}',
            canal_origen='SHOPIFY',
            sucursal=ticket.sucursal,
            rut_empresa=ticket.sucursal.empresa.rut,
            cliente_nombre='Cliente Web',
            cliente_documento='11111111-1',
            subtotal=total,
            total=total,
            items=[
                {'sku': '990001', 'nombre': 'Zapatilla Web', 'cantidad': 2, 'precio_unitario': total / 2}
            ],
            estado=estado,
            sub_estado='FACTURADO_OK' if estado == 'FACTURADO' else 'CANCELADO_CLIENTE',
            ticket=ticket,
            fecha_facturacion=fecha if estado == 'FACTURADO' else None,
            facturado_por=self.usuario,
        )

    def _get(self, **params):
        params.setdefault('fecha_inicio', timezone.localdate().replace(month=1, day=1).isoformat())
        params.setdefault('fecha_fin', timezone.localdate().isoformat())
        return self.client.get(self.url, params)

    def test_incluye_ecommerce_facturado_con_pago_internet(self):
        ticket = self._ticket(1, total=50000)
        self._pago(ticket, 50000, plataforma='Shopify')
        self._pedido('1', total=50000, ticket=ticket)

        response = self._get()
        self.assertEqual(response.status_code, 200, response.content)
        data = response.json()

        self.assertEqual(data['resumen']['venta_internet'], 50000.0)
        self.assertEqual(data['resumen']['tickets'], 1)
        self.assertEqual(data['resumen']['ecommerce_count'], 1)
        self.assertEqual(data['resumen']['pos_count'], 0)
        self.assertEqual(data['detalle'][0]['origen_codigo'], 'ECOMMERCE')
        self.assertEqual(data['detalle'][0]['numero_ticket_rm'], 'RM-1')

    def test_incluye_pos_con_pago_venta_internet(self):
        ticket = self._ticket(2, total=40000, modulo_origen='VENTA_PUBLICO')
        self._pago(ticket, 40000, plataforma='Paris')

        data = self._get().json()

        self.assertEqual(data['resumen']['venta_internet'], 40000.0)
        self.assertEqual(data['resumen']['pos_count'], 1)
        self.assertEqual(data['detalle'][0]['origen_codigo'], 'POS')
        self.assertEqual(data['plataformas'][0]['nombre'], 'Paris')

    def test_resume_boletas_emitidas_por_empresa(self):
        ticket = self._ticket(8, total=45000, modulo_origen='VENTA_PUBLICO')
        self._pago(ticket, 45000, plataforma='Paris')
        Ticket.objects.filter(pk=ticket.pk).update(
            tipo_dte='BOLETA_ELECTRONICA',
            folio_dte=1234,
        )

        data = self._get().json()

        self.assertEqual(data['resumen']['boletas'], 1)
        self.assertEqual(data['resumen']['boletas_empresas'], 1)
        self.assertEqual(data['boletas_empresas'][0]['empresa'], 'Empresa Web')
        self.assertEqual(data['boletas_empresas'][0]['boletas_total'], 1)
        self.assertEqual(data['boletas_empresas'][0]['boleta_electronica'], 1)
        self.assertEqual(data['detalle'][0]['tipo_boleta'], 'Boleta electronica')

    def test_ticket_mixto_prorratea_productos(self):
        ticket = self._ticket(3, total=100000, modulo_origen='VENTA_PUBLICO')
        self._pago(ticket, 40000, plataforma='Mercado Pago')
        self._pago(ticket, 60000, metodo='EFECTIVO', plataforma='')

        data = self._get().json()

        self.assertEqual(data['resumen']['venta_internet'], 40000.0)
        self.assertEqual(data['alertas']['tickets_mixtos']['count'], 1)
        self.assertEqual(data['productos'][0]['venta_internet_asignada'], 40000.0)
        self.assertEqual(data['productos'][0]['venta_ticket_total'], 100000.0)

    def test_ecommerce_sin_pago_internet_entra_como_fallback(self):
        ticket = self._ticket(4, total=30000)
        self._pedido('4', total=30000, ticket=ticket)

        data = self._get().json()

        self.assertEqual(data['resumen']['venta_internet'], 30000.0)
        self.assertEqual(data['resumen']['monto_fallback_ecommerce'], 30000.0)
        self.assertEqual(data['alertas']['ecommerce_sin_pago_internet']['count'], 1)
        self.assertTrue(data['detalle'][0]['usa_fallback'])

    def test_filtra_fallback_ecommerce_por_canal(self):
        ticket = self._ticket(40, total=30000)
        self._pedido('40', total=30000, ticket=ticket)

        data = self._get(plataforma='Shopify').json()
        self.assertEqual(data['resumen']['tickets'], 1)
        self.assertEqual(data['detalle'][0]['plataforma'], 'Shopify')

        data_sin_match = self._get(plataforma='Paris').json()
        self.assertEqual(data_sin_match['resumen']['tickets'], 0)

    def test_filtros_incluyen_vendedores_y_canales_ecommerce(self):
        ticket = self._ticket(41, total=30000)
        self._pedido('41', total=30000, ticket=ticket)

        data = self._get().json()

        plataformas = {item['nombre'] for item in data['filtros']['plataformas']}
        vendedores = {item['id'] for item in data['filtros']['vendedores']}
        self.assertIn('Shopify', plataformas)
        self.assertIn(self.vendedor.id, vendedores)

    def test_default_anio_actual_incluye_enero(self):
        fecha_enero = timezone.localdate().replace(month=1, day=26)
        ticket = self._ticket(5, total=22000, modulo_origen='VENTA_PUBLICO', fecha=fecha_enero)
        self._pago(ticket, 22000, plataforma='Ripley')

        data = self.client.get(self.url).json()

        self.assertEqual(data['resumen']['tickets'], 1)
        self.assertEqual(data['periodo']['inicio'], timezone.localdate().replace(month=1, day=1).isoformat())

    def test_restringe_sucursales_segun_usuario(self):
        otra_empresa = crear_empresa(nombre='Otra Empresa', rut='76.222.333-4')
        otra_sucursal = crear_sucursal(empresa=otra_empresa, alias='WEB-2')
        otro_vendedor = crear_vendedor(
            nombre='Otro Vendedor', empresa=otra_empresa, codigo_vendedor='2000'
        )
        otro_vendedor.sucursales.add(otra_sucursal)

        ticket_ok = self._ticket(6, modulo_origen='VENTA_PUBLICO')
        self._pago(ticket_ok, 50000, plataforma='Paris')

        ticket_oculto = self._ticket(
            1, sucursal=otra_sucursal, vendedor=otro_vendedor,
            modulo_origen='VENTA_PUBLICO',
        )
        TicketDetallePago.objects.create(
            ticket=ticket_oculto,
            metodo_pago='VENTA_INTERNET',
            tipo_tarjeta='Paris',
            monto=ticket_oculto.total,
        )

        limitado = crear_usuario(username='limitado', rol='vendedor')
        crear_empresa_user(limitado, self.empresa, self.sucursal)
        self.client.logout()
        self.client.login(username='limitado', password='TestPass123!')

        data = self._get().json()
        self.assertEqual(data['resumen']['tickets'], 1)
        self.assertEqual(data['empresas'][0]['nombre'], 'Empresa Web')

    def test_exporta_excel_con_hojas_esperadas_sin_duplicar_ticket(self):
        ticket = self._ticket(7, total=70000, modulo_origen='VENTA_PUBLICO')
        self._pago(ticket, 30000, plataforma='Paris')
        TicketDetallePago.objects.create(
            ticket=ticket,
            metodo_pago='VENTA_INTERNET',
            tipo_tarjeta='Ripley',
            monto=40000,
        )
        Ticket.objects.filter(pk=ticket.pk).update(
            tipo_dte='BOLETA_ELECTRONICA',
            folio_dte=7001,
        )

        response = self.client.get(
            reverse('exportar_reporte_ventas_internet'),
            {
                'fecha_inicio': timezone.localdate().replace(month=1, day=1).isoformat(),
                'fecha_fin': timezone.localdate().isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertIn('Resumen', workbook.sheetnames)
        self.assertIn('Detalle pedidos', workbook.sheetnames)
        self.assertIn('Productos', workbook.sheetnames)
        self.assertIn('Plataformas', workbook.sheetnames)
        self.assertIn('Boletas empresa', workbook.sheetnames)
        detalle = workbook['Detalle pedidos']
        self.assertEqual(detalle.max_row, 2)
        boletas = workbook['Boletas empresa']
        self.assertEqual(boletas.max_row, 2)
