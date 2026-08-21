# -*- coding: utf-8 -*-
"""
Fase B auditoría de Reportes 2026-08 — inteligencia-compra, plan-liquidación
y resumen-existencias (hallazgos P1-6, P1-7 y P1-3 de
docs/AUDITORIA_REPORTES_2026-08.md):

1. Demanda NETA: la inteligencia de compra resta CONCEPTOS_REINGRESO
   (devoluciones/NC/cambios) de las ventas en el mismo universo — antes la
   demanda bruta inflaba velocidad, TTM, pronóstico y compra sugerida (+12%
   medido en SKECHERS).
2. Sell-through sin la apertura sintética de la migración Laravel
   (INGRESO_INICIAL ref MIGRACION_LARAVEL): contarla hundía el STR
   (26,8% vs ~60% real).
3. Permisos: 'inteligencia_compra' (página + API) y
   obtener_plan_liquidacion_por_anio ('plan_liquidacion') son fail-closed.
4. Plan de liquidación: buckets de antigüedad por TRAMOS de días exactos
   (<6m / 6-12m / 1-2a / 2+a), no por año calendario — el mismo criterio en
   JSON, detalle y Excel (antes "todo 2025 = 1 año = 25%" metía en ≥1 año
   stock con lote de <365 días, +$220,7M).
5. resumen-existencias: filtrar por una categoría PADRE v1.2 expande la
   rama (antes devolvía 0 en silencio).

Aislado en SQLite:
    $env:DATABASE_URL="sqlite:///C:/temp/tb2.sqlite3"; python manage.py test app.tests.test_fase_b_inteligencia
"""
import io
import json
from datetime import timedelta

from django.test import RequestFactory, TestCase
from django.utils import timezone

from app.models import (
    AtributoOpcion, Categoria, LoteProducto, ModuloSistema, Movimientos_Producto,
    OpcionMenu, PermisoRol, Producto, Producto_Talla, Productos_Atributos,
)
from app.tests.factories import (
    crear_empresa, crear_empresa_user, crear_lote_fifo,
    crear_producto_con_talla, crear_sucursal, crear_usuario,
)
from app.views_inteligencia_compra import (
    exportar_plan_liquidacion_excel, obtener_inteligencia_compra,
    obtener_plan_liquidacion_detalle, obtener_plan_liquidacion_por_anio,
    ver_inteligencia_compra,
)
from app.views_resumen_existencias import obtener_resumen_existencias

HOY = None  # se fija en setUp (timezone.localdate())


class BaseFaseB(TestCase):
    """Empresa con una tienda vendedora + usuario con permisos reales
    (OpcionMenu/PermisoRol, sin mocks) y un cajero SIN fila de permiso."""

    def setUp(self):
        global HOY
        HOY = timezone.localdate()
        self.empresa = crear_empresa()
        self.tienda = crear_sucursal(self.empresa, alias='TIENDA-1')
        self.analista = crear_usuario(username='analista', rol='jefe_local')
        crear_empresa_user(self.analista, self.empresa, self.tienda)
        self.cajero = crear_usuario(username='cajero1', rol='cajero')
        crear_empresa_user(self.cajero, self.empresa, self.tienda)

        # Las migraciones pueden traer seeds de módulos/opciones/permisos:
        # get_or_create + limpiar la fila del cajero deja el estado exacto.
        modulo, _ = ModuloSistema.objects.get_or_create(
            codigo='reportes', defaults={'nombre': 'Reportes'})
        for codigo in ('inteligencia_compra', 'plan_liquidacion', 'resumen_existencias'):
            opcion, _ = OpcionMenu.objects.get_or_create(
                codigo=codigo, defaults={'modulo': modulo, 'nombre': codigo})
            PermisoRol.objects.update_or_create(
                rol='jefe_local', opcion_menu=opcion,
                defaults={'puede_ver': True, 'puede_exportar': True},
            )
            # cajero: SIN fila => PermisoRol.tiene_permiso -> False (fail-closed)
            PermisoRol.objects.filter(rol='cajero', opcion_menu=opcion).delete()

        # atributo Marca + una opción para los fixtures de inteligencia
        self.attr_marca = Productos_Atributos.objects.create(
            nombre='Marca', descripcion='marca')
        self.marca = AtributoOpcion.objects.create(
            atributo=self.attr_marca, valor='TESTBRAND')

    # ------------------------------------------------------------ helpers
    def _req(self, user=None, ajax=False, **params):
        extra = {'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'} if ajax else {}
        req = RequestFactory().get('/x', data=params, **extra)
        req.user = user or self.analista
        req.session = {'idSucursalActual': self.tienda.id,
                       'idEmpresaActual': self.empresa.id}
        return req

    def _producto_marca(self, articulo, sku, stock=10, **kw):
        return crear_producto_con_talla(
            self.tienda, articulo=articulo, sku=sku, stock=stock,
            atributo1=self.marca, **kw)

    def _mov(self, talla, concepto, cantidad, dias_atras=5, precio=2000, costo=1000):
        return Movimientos_Producto.objects.create(
            ProductoTalla=talla,
            sucursal_origen=self.tienda,
            cantidad=cantidad,
            costo=costo, precio=precio,
            concepto=concepto,
            estado='COMPLETADO',
            fecha=HOY - timedelta(days=dias_atras),
        )

    def _fijar_lote(self, talla, dias_atras, **kw):
        lote = crear_lote_fifo(talla, **kw)
        fecha = timezone.now() - timedelta(days=dias_atras)
        LoteProducto.objects.filter(pk=lote.pk).update(fecha_ingreso=fecha)
        return lote

    def _json(self, view, req):
        resp = view(req)
        return resp.status_code, json.loads(resp.content)


# ═══════════════════════════════ 1. demanda NETA ═══════════════════════════
class VentaNetaTest(BaseFaseB):
    """1 venta de 10 + 1 devolución de 2 => la demanda en TODAS las series
    debe ser 8, no 10 (los reingresos se restan, no se suman como |cant|)."""

    def setUp(self):
        super().setUp()
        _p, self.talla = self._producto_marca('ZAP-NETA', sku=8000001)
        self._mov(self.talla, 'VENTA_PUBLICO', -10, dias_atras=5)
        self._mov(self.talla, 'DEVOLUCION_CLIENTE', +2, dias_atras=4)

    def _data(self):
        status, body = self._json(
            obtener_inteligencia_compra,
            self._req(marca_id=self.marca.id, sucursal_id='todas'))
        self.assertEqual(status, 200)
        self.assertTrue(body['success'], body.get('error'))
        return body['data']

    def test_venta_anual_neta(self):
        d = self._data()
        fila = next(x for x in d['ventas_anual'] if x['anio'] == HOY.year)
        self.assertEqual(fila['unidades'], 8)          # 10 - 2, no 12 ni 10
        self.assertEqual(fila['monto'], 8 * 2000)      # el monto también neto
        self.assertEqual(d['kpis']['venta_publica_hist'], 8)

    def test_velocidad_90d_neta(self):
        d = self._data()
        self.assertEqual(d['kpis']['vel_dia'], round(8 / 90, 1))

    def test_venta_por_tienda_neta(self):
        d = self._data()
        fila = next(x for x in d['ventas_por_tienda'] if x['alias'] == 'TIENDA-1')
        self.assertEqual(fila['unidades'], 8)


# ══════════════════════ 2. sell-through sin apertura ═══════════════════════
class SellThroughSinAperturaTest(BaseFaseB):
    """La apertura sintética (INGRESO_INICIAL ref MIGRACION_LARAVEL) no es
    abastecimiento: 'ingresado' debe excluirla; el INGRESO_INICIAL real y la
    RECEPCION_COMPRA sí cuentan. Vendido = neto de reingresos."""

    def test_ingresado_excluye_apertura_sintetica(self):
        _p, talla = self._producto_marca('ZAP-STR', sku=8000002)
        m = self._mov(talla, 'INGRESO_INICIAL', +10, dias_atras=30)
        Movimientos_Producto.objects.filter(pk=m.pk).update(
            referencia_externa='MIGRACION_LARAVEL')     # apertura sintética
        self._mov(talla, 'INGRESO_INICIAL', +4, dias_atras=30)   # real, sin ref
        self._mov(talla, 'RECEPCION_COMPRA', +10, dias_atras=20)
        self._mov(talla, 'VENTA_PUBLICO', -5, dias_atras=5)
        self._mov(talla, 'DEVOLUCION_CLIENTE', +1, dias_atras=4)

        status, body = self._json(
            obtener_inteligencia_compra,
            self._req(marca_id=self.marca.id, sucursal_id='todas'))
        self.assertEqual(status, 200)
        st = next(s for s in body['data']['sellthrough'] if s['anio'] == HOY.year)
        self.assertEqual(st['ingresado'], 14)   # 4 + 10; NO 24 (sin apertura)
        self.assertEqual(st['vendido'], 4)      # 5 - 1 (neto)
        self.assertEqual(st['str'], round(100.0 * 4 / 14, 1))


# ═══════════════════════════════ 3. permisos ═══════════════════════════════
class PermisosFaseBTest(BaseFaseB):
    """inteligencia_compra (página+API) y por-anio son fail-closed: sin fila
    de PermisoRol => 403 (patrón PermisoRol real, sin mocks)."""

    def test_por_anio_403_sin_permiso(self):
        status, body = self._json(
            obtener_plan_liquidacion_por_anio, self._req(user=self.cajero, ajax=True))
        self.assertEqual(status, 403)
        self.assertTrue(body.get('error'))

    def test_por_anio_200_con_permiso(self):
        status, body = self._json(
            obtener_plan_liquidacion_por_anio, self._req(ajax=True))
        self.assertEqual(status, 200)
        self.assertTrue(body['success'])

    def test_api_inteligencia_403_sin_permiso(self):
        status, _ = self._json(
            obtener_inteligencia_compra,
            self._req(user=self.cajero, ajax=True, marca_id=self.marca.id))
        self.assertEqual(status, 403)

    def test_pagina_inteligencia_403_sin_permiso(self):
        resp = ver_inteligencia_compra(self._req(user=self.cajero, ajax=True))
        self.assertEqual(resp.status_code, 403)

    def test_api_inteligencia_200_con_permiso(self):
        status, body = self._json(
            obtener_inteligencia_compra,
            self._req(ajax=True, marca_id=self.marca.id))
        self.assertEqual(status, 200)
        self.assertTrue(body['success'])


# ═══════════════ 4. tramos de antigüedad por días exactos ══════════════════
class TramosAntiguedadTest(BaseFaseB):
    """3 productos con lote de 300 / 400 / 800 días: con año calendario los
    dos primeros podían caer juntos en '1 año'; por días exactos van a
    6-12 meses (15%), 1-2 años (25%) y 2+ años (40%) respectivamente, y el
    corte '≥1 año' solo suma los dos últimos."""

    def setUp(self):
        super().setUp()
        _p, t_a = self._producto_marca('A-300D', sku=8100001, stock=5, costo=1000)
        _p, t_b = self._producto_marca('B-400D', sku=8100002, stock=3, costo=2000)
        _p, t_c = self._producto_marca('C-800D', sku=8100003, stock=2, costo=500)
        self._fijar_lote(t_a, dias_atras=300)
        self._fijar_lote(t_b, dias_atras=400)
        self._fijar_lote(t_c, dias_atras=800)

    def test_por_anio_bucketiza_por_dias_exactos(self):
        status, body = self._json(obtener_plan_liquidacion_por_anio, self._req())
        self.assertEqual(status, 200)
        tramos = {t['tramo']: t for t in body['tramos']}
        self.assertEqual(set(tramos), {'t1', 't2', 't3'})
        self.assertEqual(tramos['t1']['valor'], 5 * 1000)   # 300d -> 6-12 meses
        self.assertEqual(tramos['t1']['descuento_sugerido'], 15)
        self.assertEqual(tramos['t2']['valor'], 3 * 2000)   # 400d -> 1-2 años
        self.assertEqual(tramos['t2']['descuento_sugerido'], 25)
        self.assertEqual(tramos['t3']['valor'], 2 * 500)    # 800d -> 2+ años
        self.assertEqual(tramos['t3']['descuento_sugerido'], 40)
        # ≥1 año por días REALES: el producto de 300 días queda fuera.
        self.assertEqual(body['totales']['valor_1anio_mas'], 3 * 2000 + 2 * 500)
        self.assertEqual(body['totales']['valor'], 5000 + 6000 + 1000)

    def test_detalle_filtra_por_tramo(self):
        status, body = self._json(
            obtener_plan_liquidacion_detalle, self._req(tramo='t2'))
        self.assertEqual(status, 200)
        arts = {f['articulo'] for f in body['filas']}
        self.assertEqual(arts, {'B-400D'})
        fila = body['filas'][0]
        self.assertEqual(fila['descuento_sugerido'], 25)   # días exactos
        self.assertTrue(365 <= fila['dias_antiguedad'] < 730)

    def test_dimension_por_dias_exactos(self):
        """El stacked por marca reparte por días: 300d es <1 año, no '1 año'."""
        status, body = self._json(obtener_plan_liquidacion_por_anio, self._req())
        marca = next(m for m in body['por_dimension'] if m['nombre'] == 'TESTBRAND')
        self.assertEqual(marca['val_reciente'], 5 * 1000)          # 300d
        self.assertEqual(marca['val_1anio'], 3 * 2000)             # 400d
        self.assertEqual(marca['val_2mas'], 2 * 500)               # 800d

    def test_excel_resumen_usa_tramos(self):
        from openpyxl import load_workbook
        resp = exportar_plan_liquidacion_excel(self._req(tramo='t2'))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
        celdas = [c.value for row in wb['Resumen'].iter_rows() for c in row]
        self.assertIn('Tramo antigüedad', celdas)
        self.assertIn('1-2 años', celdas)          # el tramo filtrado
        self.assertNotIn('6-12 meses', celdas)     # los otros no entran
        # y el filtro quedó declarado en la hoja Filtros
        celdas_f = [c.value for row in wb['Filtros'].iter_rows() for c in row]
        self.assertIn('1-2 años', celdas_f)


# ═══════════════ 5. resumen-existencias: categoría PADRE ═══════════════════
class ResumenCategoriaPadreTest(BaseFaseB):
    """El dropdown ofrece padres v1.2: filtrar por el padre debe traer el
    stock de las hijas (antes: 0 en silencio). Vale para el corte actual y
    para el histórico."""

    def setUp(self):
        super().setUp()
        self.padre = Categoria.objects.create(nombre='CALZADO-T')
        self.hija = Categoria.objects.create(nombre='ZAPATILLAS-T', padre=self.padre)
        prod = Producto.objects.create(
            articulo='HIJA-ART', descripcion='en categoría hija',
            sucursal=self.tienda, costo=1000, sobreprecio=0, precioventa=2000,
            categoria=self.hija)
        Producto_Talla.objects.create(producto=prod, sku=8200001, stock=7, talla='40')

    def test_padre_encuentra_stock_de_la_hija(self):
        status, body = self._json(
            obtener_resumen_existencias, self._req(categoria_id=self.padre.id))
        self.assertEqual(status, 200)
        self.assertTrue(body['success'], body.get('error'))
        self.assertEqual(body['total_general']['pares'], 7)
        fila = next(f for f in body['datos'] if f['sucursal'] == 'TIENDA-1')
        self.assertEqual(fila['total_pares'], 7)

    def test_hija_directa_sigue_funcionando(self):
        status, body = self._json(
            obtener_resumen_existencias, self._req(categoria_id=self.hija.id))
        self.assertEqual(status, 200)
        self.assertEqual(body['total_general']['pares'], 7)

    def test_padre_en_corte_historico(self):
        ayer = (HOY - timedelta(days=1)).isoformat()
        status, body = self._json(
            obtener_resumen_existencias,
            self._req(categoria_id=self.padre.id, fecha_corte=ayer))
        self.assertEqual(status, 200)
        self.assertTrue(body['es_historico'])
        self.assertEqual(body['total_general']['pares'], 7)
