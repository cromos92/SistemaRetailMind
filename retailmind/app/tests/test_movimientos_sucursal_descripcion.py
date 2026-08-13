"""Verificación de la columna Descripción OPCIONAL (SQLite desechable).

Corre las DOS vistas (API JSON y export Excel) con y sin
`incluir_descripcion` y comprueba la alineación de columnas del xlsx.
"""
import json
from io import BytesIO

from django.test import TestCase
from django.test.client import RequestFactory
from openpyxl import load_workbook

from app.models import AtributoOpcion, Productos_Atributos
from app.tests.factories import (
    crear_empresa, crear_empresa_user, crear_producto_con_talla,
    crear_sucursal, crear_usuario,
)
from app.views_modulo_reportes import (
    exportar_movimientos_sucursal_excel,
    obtener_reporte_movimientos_sucursal,
)

DESC = 'ZAPATILLA RUNNING 10" PRO <X>'


class DescripcionOpcionalTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa = crear_empresa()
        cls.suc = crear_sucursal(empresa=cls.empresa, alias='TIENDA1')
        cls.user = crear_usuario()
        crear_empresa_user(cls.user, cls.empresa, cls.suc)

        attr = Productos_Atributos.objects.create(nombre='Marca', descripcion='Marca')
        marca = AtributoOpcion.objects.create(atributo=attr, valor='NIKE')

        cls.prod, _ = crear_producto_con_talla(
            cls.suc, articulo='ABC-123', atributo1=marca)
        cls.prod.descripcion = DESC
        cls.prod.save(update_fields=['descripcion'])

    def _get(self, view, **params):
        req = RequestFactory().get('/x/', params)
        req.user = self.user
        req.session = {}
        return view(req)

    # ---------- API ----------
    def test_api_sin_flag_no_trae_descripcion(self):
        data = json.loads(self._get(
            obtener_reporte_movimientos_sucursal, sin_filtro='true', mostrar='todo'
        ).content)
        self.assertTrue(data['success'], data)
        self.assertTrue(data['datos'], 'no devolvio filas')
        self.assertFalse(data['incluir_descripcion'])
        self.assertNotIn('descripcion', data['datos'][0])

    def test_api_con_flag_trae_descripcion(self):
        data = json.loads(self._get(
            obtener_reporte_movimientos_sucursal, sin_filtro='true',
            mostrar='todo', incluir_descripcion='true'
        ).content)
        self.assertTrue(data['datos'], 'no devolvio filas')
        self.assertTrue(data['incluir_descripcion'])
        self.assertEqual(data['datos'][0]['descripcion'], DESC)

    # ---------- EXCEL ----------
    def _hoja(self, **params):
        resp = self._get(exportar_movimientos_sucursal_excel,
                         sin_filtro='true', mostrar='todo', **params)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        ws = load_workbook(BytesIO(resp.content)).active
        return ws, [c.value for c in ws[4]], [c.value for c in ws[5]]

    def test_excel_sin_flag(self):
        ws, headers, fila = self._hoja()
        self.assertEqual(headers[:6], ['Artículo', 'Marca', 'Color',
                                       'Departamento', 'Costo', 'Precio Venta'])
        self.assertEqual(fila[0], 'ABC-123')
        self.assertEqual(fila[1], 'NIKE')     # Marca pegada al Artículo
        self.assertEqual(fila[4], 15000)      # Costo en su columna
        self.assertEqual(fila[5], 20000)      # Precio en la suya
        self.assertIn('A1:F1', {str(r) for r in ws.merged_cells.ranges})

    def test_excel_con_flag(self):
        ws, headers, fila = self._hoja(incluir_descripcion='true')
        self.assertEqual(headers[:7], ['Artículo', 'Descripción', 'Marca', 'Color',
                                       'Departamento', 'Costo', 'Precio Venta'])
        self.assertEqual(fila[0], 'ABC-123')
        self.assertEqual(fila[1], DESC)
        self.assertEqual(fila[2], 'NIKE')     # todo corrido una columna
        self.assertEqual(fila[5], 15000)
        self.assertEqual(fila[6], 20000)
        self.assertIn('A1:G1', {str(r) for r in ws.merged_cells.ranges})

    def test_totales_bajo_las_mismas_columnas(self):
        """La fila TOTALES debe caer bajo las MISMAS columnas que los datos."""
        for extra, n_fijas in (({}, 6), ({'incluir_descripcion': 'true'}, 7)):
            with self.subTest(extra=extra):
                ws, headers, _ = self._hoja(**extra)
                ultima = ws.max_row
                self.assertEqual(ws.cell(row=ultima, column=1).value, 'TOTALES')
                for c in range(2, n_fijas + 1):
                    self.assertIn(ws.cell(row=ultima, column=c).value, ('', None),
                                  f'col {c} deberia ir vacia en TOTALES')
                # El primer número cae justo donde empieza el bloque por sucursal.
                self.assertIsInstance(
                    ws.cell(row=ultima, column=n_fijas + 1).value, int)
                self.assertTrue(headers[n_fijas].endswith(('Original', 'Actual')),
                                f'header {n_fijas}: {headers[n_fijas]}')

    def test_sin_flag_no_hay_query_extra_por_producto(self):
        """El .only() sin 'descripcion' no debe dejar el campo diferido en uso."""
        data = json.loads(self._get(
            obtener_reporte_movimientos_sucursal, sin_filtro='true', mostrar='todo'
        ).content)
        self.assertNotIn('descripcion', json.dumps(data['datos'][0]))
