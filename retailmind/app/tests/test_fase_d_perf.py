# -*- coding: utf-8 -*-
"""
Fase D auditoría de Reportes 2026-08 — cierre de la deuda de performance del
Plan de Liquidación SIN cambiar ni un número
(docs/AUDITORIA_REPORTES_2026-08.md, secciones 4 y 11 "deuda restante").

Los 3 residuos que la Fase C dejó documentados como "no-hacer" y que acá se
resuelven, cada uno con su invariante bajo test:

1. Excel del plan (openpyxl escribiendo ~253k celdas con estilo): el cuerpo se
   escribe con `ws.append` + una StyleArray de borde COMPARTIDA. El archivo
   tiene que salir idéntico: mismas hojas, mismas filas/columnas, mismos
   valores y los MISMOS estilos (borde en todo el cuerpo, sin heredar la
   negrita ni el relleno del encabezado).
2. Detalle ordenado por última venta (~14s): el orden y la paginación ya no
   pasan por los Subquery correlacionados de venta. `_ventas_bulk` (Max +
   Σ365d en una pasada agrupada) tiene que dar EXACTAMENTE lo mismo que las
   anotaciones, y `_ordenar_ids_por_ventas` tiene que espejar el ORDER BY de
   Postgres (NULLS FIRST/LAST y desempate por id incluidos).
3. Dimensión especialidad de `por-anio` (join multi-etiqueta que duplica
   filas): dimensión y tramos se derivan de UNA pasada por producto. Un
   producto con 2 especialidades suma en las 2 de la tabla pero UNA sola vez
   en los tramos, y el que no tiene ninguna queda fuera de la tabla y dentro
   de los tramos.

Más las equivalencias de los mapas en bloque que habilitan todo lo anterior:
`_lote_map` (Min del lote vivo) == el Subquery de aging, y
`_rows_producto_planas` == el queryset anotado del export (filas y orden).

Aislado en SQLite:
    $env:DATABASE_URL="sqlite:///C:/temp/td6.sqlite3"; python manage.py test app.tests.test_fase_d_perf
"""
import io
import json
from datetime import timedelta

from django.utils import timezone

from app.models import (
    AtributoOpcion, LoteProducto, ProductoAtributoValor, Productos_Atributos,
)
from app.tests.factories import crear_producto_con_talla, crear_sucursal
from app.tests.test_fase_b_inteligencia import BaseFaseB
from app.views_inteligencia_compra import (
    MAX_EXPORT_FILAS, _detalle_query, _lote_map, _ordenar_ids_por_ventas,
    _rows_producto_planas, _scope_plan, _ventas_bulk,
    exportar_plan_liquidacion_excel, obtener_plan_liquidacion_detalle,
    obtener_plan_liquidacion_por_anio,
)


class BaseFaseD(BaseFaseB):
    """Helpers comunes: acceso al scope real de la vista y al detalle JSON."""

    def _scope(self, **params):
        return _scope_plan(self._req(**params))

    def _detalle(self, **params):
        status, body = self._json(obtener_plan_liquidacion_detalle,
                                  self._req(**params))
        self.assertEqual(status, 200)
        self.assertTrue(body.get('success'), body.get('error'))
        return body

    def _ids(self, body):
        return [f['producto_id'] for f in body['filas']]


# ═══════════ 1. mapas en bloque == Subquery correlacionado ═══════════
class VentasBulkEquivalenteTest(BaseFaseD):
    """Última venta y u365 de la pasada agrupada == las anotaciones por fila,
    en los 3 casos que importan: vendido hace poco, vendido solo fuera de la
    ventana de 365 días, y nunca vendido."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.p_rec, t_rec = self._producto_marca('D-RECIENTE', sku=9100001, stock=5)
        self.p_vie, t_vie = self._producto_marca('D-VIEJO', sku=9100002, stock=5)
        self.p_nun, _t = self._producto_marca('D-NUNCA', sku=9100003, stock=5)
        self._mov(t_rec, 'VENTA_PUBLICO', -3, dias_atras=20)
        self._mov(t_rec, 'VENTA_PUBLICO', -2, dias_atras=100)
        self._mov(t_vie, 'VENTA_PUBLICO', -7, dias_atras=400)   # fuera de 365d

    def _anotado(self):
        base_pt, mov_base, _ctx = self._scope()
        qs = _detalle_query(base_pt, mov_base, self.hoy)
        return {r['id']: (r['ultima_venta'], r['u365'])
                for r in qs.values('id', 'ultima_venta', 'u365')}

    def test_maps_iguales_a_las_anotaciones(self):
        base_pt, mov_base, _ctx = self._scope()
        venta_map, u365_map = _ventas_bulk(mov_base, self.hoy)
        anotado = self._anotado()
        for pid, (ult, u365) in anotado.items():
            self.assertEqual(venta_map.get(pid), ult, f'ultima_venta pid={pid}')
            self.assertEqual(u365_map.get(pid, 0), u365, f'u365 pid={pid}')
        # y los valores concretos esperados
        self.assertEqual(venta_map[self.p_rec.id], self.hoy - timedelta(days=20))
        self.assertEqual(u365_map[self.p_rec.id], 5)
        self.assertEqual(venta_map[self.p_vie.id], self.hoy - timedelta(days=400))
        self.assertNotIn(self.p_vie.id, u365_map)   # => 0 al leer con .get(,0)
        self.assertNotIn(self.p_nun.id, venta_map)  # nunca vendido => None

    def test_solo_u365_no_trae_ultima_venta(self):
        """`con_ultima=False` recorre solo los últimos 365 días: mismo u365,
        venta_map vacío (el llamador no lo usa para ordenar por u365)."""
        base_pt, mov_base, _ctx = self._scope()
        venta_map, u365_map = _ventas_bulk(mov_base, self.hoy, con_ultima=False)
        self.assertEqual(venta_map, {})
        self.assertEqual(u365_map, {self.p_rec.id: 5})

    def test_acotado_a_una_pagina(self):
        base_pt, mov_base, _ctx = self._scope()
        venta_map, u365_map = _ventas_bulk(mov_base, self.hoy,
                                           producto_ids=[self.p_vie.id])
        self.assertEqual(set(venta_map), {self.p_vie.id})
        self.assertEqual(u365_map, {})


class LoteMapEquivalenteTest(BaseFaseD):
    """El Min agrupado del lote vivo == el Subquery `fecha_lote`: se queda con
    el más antiguo VIVO, ignora agotados/inactivos/sin disponible, y el
    producto sin lotes no entra al mapa (cae a fecha_creacion)."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.p_ok, t_ok = self._producto_marca('D-LOTES', sku=9110001, stock=9)
        self.p_sin, _t = self._producto_marca('D-SINLOTE', sku=9110002, stock=4)
        self.viejo_vivo = self._fijar_lote(t_ok, dias_atras=300, cantidad=4)
        self._fijar_lote(t_ok, dias_atras=100, cantidad=5)
        self._fijar_lote(t_ok, dias_atras=900, cantidad=0, agotado=True)
        self._fijar_lote(t_ok, dias_atras=800, cantidad=3, activo=False)

    def test_min_lote_vivo(self):
        base_pt, mov_base, _ctx = self._scope()
        mapa = _lote_map(base_pt)
        esperado = LoteProducto.objects.get(pk=self.viejo_vivo.pk).fecha_ingreso
        self.assertEqual(mapa[self.p_ok.id], esperado)
        self.assertNotIn(self.p_sin.id, mapa)
        # idéntico al Subquery de la versión anotada
        qs = _detalle_query(base_pt, mov_base, self.hoy)
        anotado = {r['id']: r['fecha_lote'] for r in qs.values('id', 'fecha_lote')}
        self.assertEqual(mapa.get(self.p_ok.id), anotado[self.p_ok.id])
        self.assertEqual(mapa.get(self.p_sin.id), anotado[self.p_sin.id])


class RowsPlanasEquivalenteTest(BaseFaseD):
    """El armado plano del export (2 queries) devuelve las MISMAS filas, en el
    MISMO orden (valor a costo desc, id) que el queryset anotado."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        # valores distintos y un empate exacto para probar el desempate por id
        crear_producto_con_talla(self.tienda, articulo='D-P1', sku=9120001,
                                 stock=3, costo=1000, atributo1=self.marca)
        crear_producto_con_talla(self.tienda, articulo='D-P2', sku=9120002,
                                 stock=10, costo=1000, atributo1=self.marca)
        crear_producto_con_talla(self.tienda, articulo='D-P3', sku=9120003,
                                 stock=1, costo=3000, atributo1=self.marca)
        crear_producto_con_talla(self.tienda, articulo='D-P4', sku=9120004,
                                 stock=3, costo=1000, atributo1=self.marca)

    def test_mismas_filas_y_orden(self):
        from django.db.models import F
        base_pt, mov_base, _ctx = self._scope()
        anotado = list(
            _detalle_query(base_pt, mov_base, self.hoy, con_ventas=False,
                           con_aging=False)
            .order_by(F('valor_ord').desc(nulls_last=True), 'id')
            .values('id', 'stock_u', 'tallas', 'costo'))
        planas = _rows_producto_planas(base_pt)
        self.assertEqual([r['id'] for r in planas], [r['id'] for r in anotado])
        for a, b in zip(anotado, planas):
            self.assertEqual((a['stock_u'], a['tallas'], a['costo']),
                             (b['stock_u'], b['tallas'], b['costo']))

    def test_filtro_q_se_respeta(self):
        base_pt, _mov, _ctx = self._scope()
        planas = _rows_producto_planas(base_pt, q='D-P3')
        self.assertEqual([r['articulo'] for r in planas], ['D-P3'])


# ═══════════ 2. detalle: orden y paginación por columnas del kardex ═══════
class OrdenPorVentasTest(BaseFaseD):
    """El orden en Python espeja el ORDER BY: nunca-vendido primero con
    '-dias_sin_venta' (NULLS FIRST) y último con 'dias_sin_venta' (NULLS
    LAST), y los empates se rompen por id ascendente."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.p10, t10 = self._producto_marca('D-V10', sku=9200001, stock=5)
        self.p50, t50 = self._producto_marca('D-V50', sku=9200002, stock=5)
        self.pA, tA = self._producto_marca('D-EMP-A', sku=9200003, stock=5)
        self.pB, tB = self._producto_marca('D-EMP-B', sku=9200004, stock=5)
        self.pN, _tN = self._producto_marca('D-NUNCA', sku=9200005, stock=5)
        self._mov(t10, 'VENTA_PUBLICO', -1, dias_atras=10)
        self._mov(t50, 'VENTA_PUBLICO', -9, dias_atras=50)
        self._mov(tA, 'VENTA_PUBLICO', -2, dias_atras=30)
        self._mov(tB, 'VENTA_PUBLICO', -2, dias_atras=30)   # empate con A

    def test_desc_nunca_primero(self):
        ids = self._ids(self._detalle(orden='-dias_sin_venta'))
        # más días sin vender primero: nunca -> 50d -> 30d (A,B por id) -> 10d
        self.assertEqual(ids, [self.pN.id, self.p50.id, min(self.pA.id, self.pB.id),
                               max(self.pA.id, self.pB.id), self.p10.id])

    def test_asc_nunca_ultimo(self):
        ids = self._ids(self._detalle(orden='dias_sin_venta'))
        self.assertEqual(ids, [self.p10.id, min(self.pA.id, self.pB.id),
                               max(self.pA.id, self.pB.id), self.p50.id,
                               self.pN.id])

    def test_orden_igual_al_helper(self):
        """El helper de orden y la vista coinciden fila por fila."""
        base_pt, mov_base, _ctx = self._scope()
        ids = list(_detalle_query(base_pt, mov_base, self.hoy, con_ventas=False,
                                  con_aging=False).values_list('id', flat=True))
        bulk = _ventas_bulk(mov_base, self.hoy)
        for orden in ('-dias_sin_venta', 'dias_sin_venta'):
            esperado = _ordenar_ids_por_ventas(
                ids, orden.lstrip('-'), orden.startswith('-'), bulk)
            self.assertEqual(self._ids(self._detalle(orden=orden)), esperado)

    def test_paginacion_sin_saltos_ni_repetidos(self):
        p1 = self._detalle(orden='-dias_sin_venta', page='1', page_size='10')
        p2 = self._detalle(orden='-dias_sin_venta', page='2', page_size='10')
        self.assertEqual(p1['total'], 5)
        self.assertEqual(len(p1['filas']), 5)
        self.assertEqual(p2['filas'], [])
        chico = self._detalle(orden='-dias_sin_venta', page='1', page_size='10')
        self.assertEqual(self._ids(chico)[:2],
                         self._ids(p1)[:2])

    def test_pagina_2_continua(self):
        p1 = self._detalle(orden='-dias_sin_venta', page='1', page_size='2')
        p2 = self._detalle(orden='-dias_sin_venta', page='2', page_size='2')
        p3 = self._detalle(orden='-dias_sin_venta', page='3', page_size='2')
        todos = self._ids(p1) + self._ids(p2) + self._ids(p3)
        self.assertEqual(todos, self._ids(self._detalle(orden='-dias_sin_venta')))
        self.assertEqual(len(set(todos)), 5)


class OrdenPorU365Test(BaseFaseD):
    """u365 es Coalesce(...,0): nunca NULL, así que el orden es numérico puro
    con desempate por id (la venta de hace 400 días cuenta como 0)."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.p9, t9 = self._producto_marca('D-U9', sku=9210001, stock=5)
        self.p2, t2 = self._producto_marca('D-U2', sku=9210002, stock=5)
        self.p0, t0 = self._producto_marca('D-U0', sku=9210003, stock=5)
        self._mov(t9, 'VENTA_PUBLICO', -9, dias_atras=30)
        self._mov(t2, 'VENTA_PUBLICO', -2, dias_atras=30)
        self._mov(t0, 'VENTA_PUBLICO', -6, dias_atras=400)

    def test_desc_y_asc(self):
        self.assertEqual(self._ids(self._detalle(orden='-u365')),
                         [self.p9.id, self.p2.id, self.p0.id])
        self.assertEqual(self._ids(self._detalle(orden='u365')),
                         [self.p0.id, self.p2.id, self.p9.id])

    def test_u365_y_ultima_venta_de_la_pagina(self):
        """Ordenando por u365 la última venta de la página sale de la pasada
        acotada a sus ids — tiene que seguir siendo la fecha real."""
        filas = {f['producto_id']: f for f in self._detalle(orden='-u365')['filas']}
        self.assertEqual(filas[self.p9.id]['u365'], 9)
        self.assertEqual(filas[self.p0.id]['u365'], 0)
        self.assertEqual(filas[self.p0.id]['ultima_venta'],
                         (self.hoy - timedelta(days=400)).strftime('%Y-%m-%d'))
        self.assertEqual(filas[self.p0.id]['dias_sin_venta'], 400)


class OrdenFichaYFiltrosTest(BaseFaseD):
    """Los órdenes por columnas de la ficha siguen resolviéndose en el DB, y
    la página trae última venta / u365 correctos aunque el orden no dependa
    de ellos. Con filtro de antigüedad el aging vuelve al queryset."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.caro, t_caro = self._producto_marca('D-CARO', sku=9220001,
                                                 stock=10, costo=5000)
        self.barato, t_barato = self._producto_marca('D-BARATO', sku=9220002,
                                                     stock=1, costo=100)
        self._fijar_lote(t_caro, dias_atras=400)     # tramo t2 (1-2 años)
        self._fijar_lote(t_barato, dias_atras=10)    # tramo t0
        self._mov(t_caro, 'VENTA_PUBLICO', -4, dias_atras=15)

    def test_default_valor_costo(self):
        body = self._detalle()
        self.assertEqual(self._ids(body), [self.caro.id, self.barato.id])
        fila = body['filas'][0]
        self.assertEqual(fila['valor_costo'], 50000)
        self.assertEqual(fila['u365'], 4)
        self.assertEqual(fila['ultima_venta'],
                         (self.hoy - timedelta(days=15)).strftime('%Y-%m-%d'))
        self.assertIsNone(body['filas'][1]['ultima_venta'])
        self.assertEqual(body['filas'][1]['u365'], 0)

    def test_orden_por_antiguedad_usa_aging(self):
        # El orden es sobre aging_dt (la FECHA), no sobre los días: '-' =
        # fecha descendente = lote más nuevo primero. Semántica preexistente,
        # intacta — acá solo se comprueba que el aging vuelve al queryset.
        self.assertEqual(self._ids(self._detalle(orden='-dias_antiguedad')),
                         [self.barato.id, self.caro.id])
        self.assertEqual(self._ids(self._detalle(orden='dias_antiguedad')),
                         [self.caro.id, self.barato.id])

    def test_tramo_con_orden_de_ventas(self):
        body = self._detalle(tramo='t2', orden='-dias_sin_venta')
        self.assertEqual(body['total'], 1)
        self.assertEqual(self._ids(body), [self.caro.id])

    def test_bucket_con_orden_de_ventas(self):
        body = self._detalle(antiguedad='365+', orden='dias_sin_venta')
        self.assertEqual(self._ids(body), [self.caro.id])


# ═══════════ 3. por-anio: dimensión y tramos de UNA pasada ═══════════
class PorAnioEspecialidadTest(BaseFaseD):
    """Multi-etiqueta: el producto con 2 especialidades suma en las 2 filas de
    la tabla (atribución) pero UNA sola vez en los tramos; el producto sin
    especialidad no genera fila y sí cuenta en los tramos."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        attr = Productos_Atributos.objects.create(nombre='Especialidad',
                                                  descripcion='esp')
        self.running = AtributoOpcion.objects.create(atributo=attr, valor='running')
        self.urbano = AtributoOpcion.objects.create(atributo=attr, valor='urbano')
        self.dos, t_dos = self._producto_marca('D-DOS-ESP', sku=9300001,
                                               stock=4, costo=1000)
        self.una, t_una = self._producto_marca('D-UNA-ESP', sku=9300002,
                                               stock=2, costo=2000)
        self.nada, t_nada = self._producto_marca('D-SIN-ESP', sku=9300003,
                                                 stock=5, costo=3000)
        for opcion in (self.running, self.urbano):
            ProductoAtributoValor.objects.create(producto=self.dos,
                                                 atributo=attr, opcion=opcion)
        ProductoAtributoValor.objects.create(producto=self.una, atributo=attr,
                                             opcion=self.running)
        for t in (t_dos, t_una, t_nada):
            self._fijar_lote(t, dias_atras=400)   # todos en el tramo t2

    def _body(self, **params):
        status, body = self._json(obtener_plan_liquidacion_por_anio,
                                  self._req(**params))
        self.assertEqual(status, 200)
        self.assertTrue(body.get('success'), body.get('error'))
        return body

    def test_atribucion_sin_duplicar_los_tramos(self):
        body = self._body(dim='especialidad')
        dims = {d['nombre']: d for d in body['por_dimension']}
        self.assertEqual(set(dims), {'running', 'urbano'})
        # running = D-DOS (4x1000) + D-UNA (2x2000); urbano = solo D-DOS
        self.assertEqual(dims['running']['valor'], 4000 + 4000)
        self.assertEqual(dims['running']['pares'], 4 + 2)
        self.assertEqual(dims['urbano']['valor'], 4000)
        # los tramos cuentan cada producto UNA vez (incluido el sin especialidad)
        tramos = {t['tramo']: t for t in body['tramos']}
        self.assertEqual(set(tramos), {'t2'})
        self.assertEqual(tramos['t2']['productos'], 3)
        self.assertEqual(tramos['t2']['valor'], 4000 + 4000 + 15000)
        self.assertEqual(tramos['t2']['pares'], 4 + 2 + 5)

    def test_tramos_iguales_en_ambas_dimensiones(self):
        self.assertEqual(self._body()['tramos'],
                         self._body(dim='especialidad')['tramos'])
        self.assertEqual(self._body()['totales'],
                         self._body(dim='especialidad')['totales'])

    def test_solo_dim_no_calcula_tramos(self):
        body = self._body(dim='especialidad', solo='dim')
        self.assertNotIn('tramos', body)
        self.assertEqual({d['nombre'] for d in body['por_dimension']},
                         {'running', 'urbano'})

    def test_dim_marca_incluye_sin_marca(self):
        crear_producto_con_talla(self.tienda, articulo='D-NOMARCA',
                                 sku=9300004, stock=1, costo=700)
        dims = {d['nombre']: d for d in self._body()['por_dimension']}
        self.assertEqual(set(dims), {'TESTBRAND', 'Sin marca'})
        self.assertEqual(dims['Sin marca']['valor'], 700)


class TramosLimitesTest(BaseFaseD):
    """Los cortes por días exactos del Case SQL, reproducidos en Python:
    179->t0, 180->t1, 364->t1, 365->t2, 729->t2, 730->t3."""

    def test_limites_exactos(self):
        casos = {179: 't0', 180: 't1', 364: 't1', 365: 't2', 729: 't2', 730: 't3'}
        for i, (dias, _tramo) in enumerate(sorted(casos.items())):
            _p, t = self._producto_marca(f'D-L{dias}', sku=9400000 + i,
                                         stock=1, costo=100)
            self._fijar_lote(t, dias_atras=dias)
        status, body = self._json(obtener_plan_liquidacion_por_anio, self._req())
        self.assertEqual(status, 200)
        conteo = {t['tramo']: t['productos'] for t in body['tramos']}
        self.assertEqual(conteo, {'t0': 1, 't1': 2, 't2': 2, 't3': 1})

    def test_producto_sin_fecha_queda_fuera_de_los_tramos(self):
        """Sin lote y sin fecha_creacion no hay tramo: no suma en ningún
        bucket (igual que el `default=Value(None)` del Case)."""
        from app.models import Producto
        p, _t = self._producto_marca('D-SINFECHA', sku=9410001, stock=3, costo=100)
        Producto.objects.filter(pk=p.pk).update(fecha_creacion=None)
        _p2, t2 = self._producto_marca('D-CONFECHA', sku=9410002, stock=1, costo=100)
        self._fijar_lote(t2, dias_atras=400)
        status, body = self._json(obtener_plan_liquidacion_por_anio, self._req())
        self.assertEqual(status, 200)
        self.assertEqual({t['tramo'] for t in body['tramos']}, {'t2'})
        self.assertEqual(body['totales']['pares'], 1)
        # pero sí aparece en la dimensión, en el bucket "reciente" (sin dato)
        dims = {d['nombre']: d for d in body['por_dimension']}
        self.assertEqual(dims['TESTBRAND']['pares'], 4)
        self.assertEqual(dims['TESTBRAND']['val_reciente'], 300)


# ═══════════ 4. Excel: mismo archivo, estilos incluidos ═══════════
class ExcelEstilosTest(BaseFaseD):
    """El cuerpo se escribe con una StyleArray compartida: todas sus celdas
    llevan borde y NINGUNA hereda la negrita/relleno del encabezado."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.cd = crear_sucursal(self.empresa, alias='CD-1',
                                 es_centro_distribucion=True)
        for i in range(4):
            _p, t = self._producto_marca(f'D-X{i}', sku=9500001 + i,
                                         stock=2 + i, costo=1000 * (i + 1))
            self._fijar_lote(t, dias_atras=100 + 200 * i)
            if i == 0:
                self._mov(t, 'VENTA_PUBLICO', -3, dias_atras=25)

    def _wb(self, **params):
        from openpyxl import load_workbook
        resp = exportar_plan_liquidacion_excel(self._req(**params))
        self.assertEqual(resp.status_code, 200)
        return load_workbook(io.BytesIO(resp.content))

    def test_estructura_y_estilos(self):
        wb = self._wb()
        self.assertEqual(wb.sheetnames, ['TIENDA-1', 'Resumen', 'Filtros'])
        ws = wb['TIENDA-1']
        self.assertEqual(ws.max_column, 17)
        self.assertEqual(ws.max_row, 3 + 4)      # título + header + 4 filas + TOTAL
        self.assertEqual(str(ws.merged_cells), 'A1:Q1')
        self.assertEqual(ws.freeze_panes, 'A3')
        self.assertEqual(ws.print_title_rows, '$1:$2')
        self.assertEqual(ws.page_setup.orientation, 'landscape')
        # encabezado: negrita + relleno + centrado + borde
        h = ws.cell(row=2, column=3)
        self.assertTrue(h.font.b)
        self.assertEqual(h.fill.fgColor.rgb[-6:], 'E8EAF0')
        self.assertEqual(h.alignment.horizontal, 'center')
        self.assertEqual(h.border.left.style, 'thin')
        # cuerpo: SOLO borde
        for fila in range(3, 7):
            for col in range(1, 18):
                c = ws.cell(row=fila, column=col)
                self.assertEqual(c.border.left.style, 'thin',
                                 f'sin borde en {c.coordinate}')
                self.assertEqual(c.border.bottom.style, 'thin')
                self.assertFalse(c.font.b, f'negrita colada en {c.coordinate}')
                self.assertIn(c.fill.patternType, (None, 'none'),
                              f'relleno colado en {c.coordinate}')
                self.assertIsNone(c.alignment.horizontal,
                                  f'alineación colada en {c.coordinate}')
        # fila TOTAL: negrita + borde
        total = ws.cell(row=7, column=2)
        self.assertEqual(total.value, 'TOTAL')
        self.assertTrue(total.font.b)
        self.assertEqual(total.border.left.style, 'thin')

    def test_valores_iguales_al_detalle(self):
        """Las columnas ID/Pares/Últ. venta/Precio liq. del Excel son las
        mismas cifras que el JSON del detalle."""
        wb = self._wb()
        ws = wb['TIENDA-1']
        filas = {r[0]: r for r in ws.iter_rows(min_row=3, max_row=6,
                                               values_only=True)}
        detalle = {f['producto_id']: f
                   for f in self._detalle(page_size='200')['filas']}
        self.assertEqual(set(filas), set(detalle))
        for pid, r in filas.items():
            d = detalle[pid]
            self.assertEqual(r[8], d['stock_u'])
            self.assertEqual(r[11], d['precioventa'])
            self.assertEqual(r[12], d['descuento_sugerido'])
            self.assertEqual(r[13], d['precio_liquidacion'])
            esperada = ('Nunca' if not d['ultima_venta'] else
                        '-'.join(reversed(d['ultima_venta'].split('-'))))
            self.assertEqual(r[10], esperada)

    def test_hoja_por_sucursal_con_cd(self):
        crear_producto_con_talla(self.cd, articulo='D-XCD', sku=9500900,
                                 stock=3, costo=1000, atributo1=self.marca)
        wb = self._wb(incluir_cd='1')
        self.assertIn('CD-1 CD', wb.sheetnames)

    def test_filtro_antiguedad_del_export_igual_al_queryset(self):
        """El export filtra antigüedad en Python (`_filtrar_antiguedad`): el
        set de productos tiene que ser el MISMO que el de los filtros SQL
        (`_bucket_filter` / `_tramo_filter`) para tramos, buckets y
        'sin-dato'."""
        from app.models import Producto
        from app.views_inteligencia_compra import (
            _bucket_filter, _filas_export, _filtrar_antiguedad, _parse_tramos,
            _tramo_filter,
        )
        p_sd, _t = self._producto_marca('D-SINFECHA', sku=9600001, stock=1,
                                        costo=100)
        Producto.objects.filter(pk=p_sd.pk).update(fecha_creacion=None)
        base_pt, mov_base, _ctx = self._scope()
        casos = [('t2', None), ('t2,t3', None), ('t0', None),
                 (None, '365+'), (None, '0-90'), (None, 'sin-dato'),
                 ('t3', '365+')]
        for tramo, bucket in casos:
            params = {}
            if tramo:
                params['tramo'] = tramo
            if bucket:
                params['antiguedad'] = bucket
            qs = _detalle_query(base_pt, mov_base, self.hoy, con_ventas=False)
            if bucket:
                qs = _bucket_filter(qs, bucket, self.hoy)
            qs = _tramo_filter(qs, _parse_tramos(tramo), self.hoy)
            esperado = set(qs.values_list('id', flat=True))
            planas = _filtrar_antiguedad(
                _rows_producto_planas(base_pt), _lote_map(base_pt), self.hoy,
                bucket, _parse_tramos(tramo))
            self.assertEqual({r['id'] for r in planas}, esperado,
                             f'tramo={tramo} bucket={bucket}')
            filas, _c, _b, _tr = _filas_export(self._req(**params), self.hoy)
            self.assertEqual({f['producto_id'] for f in filas}, esperado,
                             f'export tramo={tramo} bucket={bucket}')

    def test_cap_de_filas_declarado(self):
        """El cap sigue siendo el mismo y la hoja Filtros no miente cuando no
        hubo truncado."""
        wb = self._wb()
        filtros = {r[0]: r[1] for r in wb['Filtros'].iter_rows(values_only=True)}
        self.assertNotIn('ADVERTENCIA', filtros)
        self.assertEqual(MAX_EXPORT_FILAS, 20000)
