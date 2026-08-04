"""
Tests de regresión para las correcciones del módulo de reportes de ventas
(`app/views_modulo_reportes.py`).

Cubre los cinco hallazgos corregidos:

1. Las NC de venta se emiten con `tipo_transaccion` en ('DEVOLUCION',
   'ANULACION'), así que filtrar primero por ('VENTA', 'VENTA_PUBLICO') y
   recién después por `tipo_documento='NOTA DE CREDITO'` no devolvía nada.
   → helper `_queryset_ncs_venta` usado en los 4 sitios.
2. `documentos_emitidos` cruzaba Ticket con Dte por `correlativo` en vez de
   por `folio_dte` (dos secuencias distintas).
3. Doble resta de descuentos: `monto_con_iva` YA viene neto, y además se
   inventaban descuentos cuando los pagos sumaban menos que el total.
4. `documentos_emitidos` no filtraba estado ni `descartado`.
5. Comisiones: incluía DTEs ANULADO y botaba toda NC sin `vendedor_id`.
"""
from datetime import date, timedelta
from decimal import Decimal

from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from app.models import (
    Dte, Dte_Detalle_Pago, ModuloSistema, OpcionMenu, PermisoRol, Ticket,
)
from app.views_modulo_reportes import (
    _calcular_comisiones_vendedor,
    _queryset_ncs_venta,
)

from .factories import (
    crear_empresa,
    crear_empresa_user,
    crear_sucursal,
    crear_usuario,
    crear_vendedor,
)

STATICFILES_STORAGE_TEST = 'django.contrib.staticfiles.storage.StaticFilesStorage'


class _BaseReportesTest(TestCase):
    """Entorno mínimo: empresa + sucursal + vendedor + usuario administrador."""

    def setUp(self):
        self.hoy = timezone.localdate()
        self.empresa = crear_empresa(nombre='Empresa Reportes', rut='76.555.444-3')
        self.sucursal = crear_sucursal(empresa=self.empresa, alias='REP-1')
        self.vendedor = crear_vendedor(
            nombre='Vendedora Uno', empresa=self.empresa,
            codigo_vendedor='V-01', comision=Decimal('5.00'),
        )
        self.vendedor.sucursales.add(self.sucursal)
        self.usuario = crear_usuario(username='reportes_fix', rol='administrador')
        crear_empresa_user(self.usuario, self.empresa, self.sucursal)
        self._otorgar_permisos_reportes()
        self.client = Client()
        self.client.login(username='reportes_fix', password='TestPass123!')
        self._folio = 1000

    def _otorgar_permisos_reportes(self):
        """`middleware_permisos` exige un PermisoRol explícito por endpoint.

        Sin esto los tests reciben 302 (redirect a bienvenida) en vez de la
        respuesta del reporte.
        """
        modulo, _ = ModuloSistema.objects.get_or_create(
            codigo='reportes', defaults={'nombre': 'Reportes'},
        )
        codigos = (
            'reporte_ventas_comparativo',
            'reporte_productos_vendidos',
            'reporte_documentos_emitidos',
            'reporte_comisiones_vendedor',
        )
        for codigo in codigos:
            opcion, _ = OpcionMenu.objects.get_or_create(
                codigo=codigo,
                defaults={'modulo': modulo, 'nombre': codigo, 'activo': True},
            )
            PermisoRol.objects.update_or_create(
                rol='administrador', opcion_menu=opcion,
                defaults={'puede_ver': True, 'puede_exportar': True},
            )

    # ---------- helpers de creación ----------

    def _siguiente_folio(self):
        self._folio += 1
        return self._folio

    def _dte(self, monto_con_iva, *, tipo_documento='BOLETA ELECTRONICA',
             tipo_transaccion='VENTA_PUBLICO', estado_dte='EMITIDO',
             descuento=0, descartado=False, fecha=None, vendedor=-1,
             sucursal=None, numero=None, documento_afectado=None,
             es_nota_credito=False, receptor=None):
        monto_con_iva = Decimal(monto_con_iva)
        return Dte.objects.create(
            emisor=self.empresa,
            receptor=receptor,
            numero_documento=numero if numero is not None else self._siguiente_folio(),
            tipo_documento=tipo_documento,
            monto_con_iva=monto_con_iva,
            monto_neto=(monto_con_iva / Decimal('1.19')).quantize(Decimal('1')),
            descuento=Decimal(descuento),
            estado_pago='PAGADO',
            estado_dte=estado_dte,
            responsable=self.usuario.username,
            fecha_emision=fecha or self.hoy,
            fecha_vencimiento=fecha or self.hoy,
            diasCredito=0,
            bultos=0,
            unidades_productos=1,
            vendedor=self.vendedor if vendedor == -1 else vendedor,
            sucursal=sucursal or self.sucursal,
            tipo_transaccion=tipo_transaccion,
            descartado=descartado,
            es_nota_credito=es_nota_credito,
            documento_afectado=documento_afectado,
            hora=timezone.localtime().time(),
        )

    def _nc(self, monto_con_iva, *, tipo_transaccion='DEVOLUCION', **kwargs):
        kwargs.setdefault('vendedor', None)
        return self._dte(
            monto_con_iva,
            tipo_documento='NOTA DE CREDITO',
            tipo_transaccion=tipo_transaccion,
            es_nota_credito=True,
            **kwargs,
        )

    def _pago(self, dte, monto, metodo='EFECTIVO'):
        return Dte_Detalle_Pago.objects.create(
            dte=dte, metodo_pago=metodo, monto=int(monto),
        )


# ==========================================================================
# 1. Helper compartido de notas de crédito
# ==========================================================================

@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class QuerysetNCsVentaTest(_BaseReportesTest):

    def test_trae_ncs_de_ambas_modalidades(self):
        """DEVOLUCION y ANULACION son NC de venta reales; ambas deben entrar."""
        self._nc(10000, tipo_transaccion='DEVOLUCION')
        self._nc(5000, tipo_transaccion='ANULACION')

        qs = _queryset_ncs_venta(self.hoy, self.hoy)

        self.assertEqual(qs.count(), 2)

    def test_el_filtro_viejo_no_encontraba_nada(self):
        """Reproduce el bug: filtrar VENTA/VENTA_PUBLICO primero da 0 NC."""
        self._nc(10000, tipo_transaccion='DEVOLUCION')
        self._nc(5000, tipo_transaccion='ANULACION')

        viejo = Dte.objects.filter(
            fecha_emision=self.hoy,
            tipo_transaccion__in=['VENTA', 'VENTA_PUBLICO'],
        ).filter(tipo_documento='NOTA DE CREDITO')

        self.assertEqual(viejo.count(), 0)
        self.assertEqual(_queryset_ncs_venta(self.hoy, self.hoy).count(), 2)

    def test_excluye_descartadas_y_fuera_de_rango(self):
        self._nc(10000)
        self._nc(9000, descartado=True)            # NC modalidad OCULTA
        self._nc(8000, fecha=self.hoy - timedelta(days=5))

        qs = _queryset_ncs_venta(self.hoy, self.hoy)

        self.assertEqual(qs.count(), 1)
        self.assertEqual(int(qs.first().monto_con_iva), 10000)

    def test_no_trae_ncs_de_compra(self):
        self._nc(7000, tipo_transaccion='COMPRA')

        self.assertEqual(_queryset_ncs_venta(self.hoy, self.hoy).count(), 0)

    def test_respeta_filtro_de_estados(self):
        self._nc(4000, estado_dte='ANULADO')

        self.assertEqual(_queryset_ncs_venta(self.hoy, self.hoy).count(), 1)
        self.assertEqual(
            _queryset_ncs_venta(
                self.hoy, self.hoy, estados=('EMITIDO', 'ACEPTADO')
            ).count(),
            0,
        )


# ==========================================================================
# 2 / 3 / 4. Reporte de documentos emitidos
# ==========================================================================

@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class DocumentosEmitidosTest(_BaseReportesTest):

    def _get(self, **extra):
        params = {
            'fecha_desde': self.hoy.strftime('%Y-%m-%d'),
            'fecha_hasta': self.hoy.strftime('%Y-%m-%d'),
            'sucursal_id': str(self.sucursal.id),
        }
        params.update(extra)
        resp = self.client.get(reverse('obtener_documentos_emitidos'), params)
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get('success'), data.get('error'))
        return data

    def test_las_ncs_reales_se_restan_del_total(self):
        venta = self._dte(100000)
        self._pago(venta, 100000)
        self._nc(30000, tipo_transaccion='DEVOLUCION', documento_afectado=venta)

        data = self._get()
        resumen = data['resumen']

        self.assertEqual(resumen['ventas_brutas'], 100000)
        self.assertEqual(resumen['notas_credito'], 30000)
        self.assertEqual(resumen['total_global'], 70000)
        self.assertEqual(data['diagnostico']['cantidad_notas_credito'], 1)
        # La NC debe aparecer también en el listado, no solo en el resumen.
        self.assertTrue(any(d['es_nota_credito'] for d in data['documentos']))

    def test_no_hay_doble_resta_de_descuentos(self):
        """`monto_con_iva` ya viene neto del descuento: no se resta de nuevo."""
        venta = self._dte(100000, descuento=20000)
        self._pago(venta, 100000)
        self._nc(10000, documento_afectado=venta)

        resumen = self._get()['resumen']

        self.assertEqual(resumen['descuentos'], 20000)   # informativo
        self.assertEqual(resumen['ventas_brutas'], 100000)
        # Antes: 100.000 - 10.000 - 20.000 = 70.000 (el descuento salía 2 veces)
        self.assertEqual(resumen['total_global'], 90000)

    def test_no_inventa_descuentos_en_ventas_a_credito(self):
        """Pagar menos que el total es saldo pendiente, no un descuento."""
        venta = self._dte(100000, descuento=0)
        self._pago(venta, 40000, metodo='CREDITO_TRABAJADOR')

        data = self._get()
        resumen = data['resumen']

        self.assertEqual(resumen['descuentos'], 0)
        self.assertEqual(resumen['saldo_no_pagado'], 60000)
        self.assertEqual(resumen['total_global'], 100000)

        fila = data['documentos'][0]
        self.assertEqual(fila['descuento'], 0)
        self.assertEqual(fila['saldo_no_pagado'], 60000)

    def test_cruce_con_ticket_por_folio_dte(self):
        """El vínculo es Ticket.folio_dte == Dte.numero_documento."""
        venta = self._dte(50000, numero=5000)  # sin Dte_Detalle_Pago

        # Ticket correcto: su correlativo (77) NO coincide con el folio.
        Ticket.objects.create(
            vendedor=self.vendedor, sucursal=self.sucursal,
            correlativo=77, estado='PAGADO', subTotal=50000, descuento=0,
            total=50000, responsable='POS', metodo_pago='TRANSFERENCIA',
            folio_dte=5000,
        )
        # Señuelo: un ticket cuyo correlativo sí es 5000 pero de otra venta.
        Ticket.objects.create(
            vendedor=self.vendedor, sucursal=self.sucursal,
            correlativo=5000, estado='PAGADO', subTotal=999, descuento=0,
            total=999, responsable='POS', metodo_pago='CHEQUE',
            folio_dte=None,
        )

        data = self._get()
        fila = next(d for d in data['documentos'] if d['correlativo'] == 5000)

        self.assertEqual(fila['metodo_pago'], 'Transferencia')
        self.assertNotEqual(fila['metodo_pago'], 'Cheque')
        # El resumen debe clasificar el monto en transferencia, no en efectivo.
        self.assertEqual(data['resumen']['transferencia'], 50000)
        self.assertEqual(data['resumen']['efectivo'], 0)
        self.assertEqual(venta.numero_documento, 5000)

    def test_excluye_rechazados_cancelados_anulados_y_descartados(self):
        ok = self._dte(10000)
        self._pago(ok, 10000)
        for estado in ('RECHAZADO', 'CANCELADO', 'ANULADO'):
            self._dte(777000, estado_dte=estado)
        self._dte(555000, descartado=True)

        data = self._get()

        self.assertEqual(data['resumen']['ventas_brutas'], 10000)
        self.assertEqual(data['total_real'], 1)


@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class DocumentosEmitidosExcelTest(_BaseReportesTest):

    def test_excel_incluye_las_ncs_con_signo_negativo(self):
        from io import BytesIO
        from openpyxl import load_workbook

        venta = self._dte(100000, descuento=20000)
        self._pago(venta, 100000)
        self._nc(30000, documento_afectado=venta)

        resp = self.client.get(
            reverse('exportar_documentos_emitidos_excel'),
            {
                'fecha_desde': self.hoy.strftime('%Y-%m-%d'),
                'fecha_hasta': self.hoy.strftime('%Y-%m-%d'),
                'sucursal_id': str(self.sucursal.id),
            },
        )
        self.assertEqual(resp.status_code, 200)

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        filas = list(ws.iter_rows(min_row=2, values_only=True))

        self.assertEqual(len(filas), 2)
        totales = {f[1]: f[5] for f in filas}
        self.assertEqual(totales['BOLETA ELECTRONICA'], 100000)
        self.assertEqual(totales['NOTA DE CREDITO'], -30000)
        # La suma de la columna Total da el neto del período.
        self.assertEqual(sum(f[5] for f in filas), 70000)

        # "Pagado" no vuelve a restar el descuento ya incluido en el total.
        pagados = {f[1]: f[7] for f in filas}
        self.assertEqual(pagados['BOLETA ELECTRONICA'], 100000)


# ==========================================================================
# 1 (bis). Comparativo de ventas
# ==========================================================================

@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class VentasComparativoTest(_BaseReportesTest):

    def test_tasa_devolucion_deja_de_ser_cero(self):
        self._dte(200000)
        self._nc(50000, tipo_transaccion='DEVOLUCION')

        resp = self.client.get(reverse('obtener_ventas_comparativo'), {
            'tipo_flujo': 'custom',
            'fecha_inicio': self.hoy.strftime('%Y-%m-%d'),
            'fecha_fin': self.hoy.strftime('%Y-%m-%d'),
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get('success'), data.get('error'))

        kpis = data['kpis']
        self.assertEqual(kpis['ventas_actual'], 150000)          # 200k - 50k
        self.assertEqual(kpis['tasa_devolucion_actual'], 25.0)

    def test_default_tipo_flujo_es_mes_mtd(self):
        """Sin `tipo_flujo` el período actual termina HOY, no a fin de mes."""
        resp = self.client.get(reverse('obtener_ventas_comparativo'))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get('success'), data.get('error'))

        self.assertEqual(data['tipo_flujo'], 'mes_mtd')
        self.assertEqual(
            data['periodo']['actual']['fin'], self.hoy.strftime('%Y-%m-%d'),
        )

    def test_default_productos_vendidos_es_mes_mtd(self):
        resp = self.client.get(reverse('obtener_productos_vendidos'))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get('success'), data.get('error'))

        self.assertEqual(data['tipo_flujo'], 'mes_mtd')
        self.assertEqual(data['periodo']['fin'], self.hoy.strftime('%Y-%m-%d'))


# ==========================================================================
# 5. Comisiones por vendedor (dinero real)
# ==========================================================================

@override_settings(STATICFILES_STORAGE=STATICFILES_STORAGE_TEST)
class ComisionesVendedorTest(_BaseReportesTest):

    def _calcular(self, **params):
        rf = RequestFactory()
        base = {
            'fecha_inicio': self.hoy.strftime('%Y-%m-%d'),
            'fecha_fin': self.hoy.strftime('%Y-%m-%d'),
        }
        base.update(params)
        request = rf.get('/app/api/reportes/comisiones-vendedor/', base)
        request.user = self.usuario
        request.session = {}
        return _calcular_comisiones_vendedor(request)

    def _fila(self, data, vendedor_id):
        return next(
            (v for v in data['vendedores'] if v['id'] == vendedor_id), None,
        )

    def test_excluye_dtes_anulados(self):
        """Una venta anulada no paga comisión (su NC va descartada)."""
        self._dte(119000)                      # EMITIDO → sí comisiona
        self._dte(238000, estado_dte='ANULADO')

        data = self._calcular()
        fila = self._fila(data, self.vendedor.id)

        self.assertIsNotNone(fila)
        self.assertEqual(fila['ventas_brutas'], 119000)
        neto = int((Decimal('119000') / Decimal('1.19')).quantize(Decimal('1')))
        self.assertEqual(fila['ventas_netas_sin_iva'], neto)
        self.assertEqual(fila['comision_monto'], int(round(neto * 5 / 100)))

    def test_nc_sin_vendedor_se_imputa_por_documento_afectado(self):
        """La NC no hereda vendedor; se toma el de la venta original."""
        venta = self._dte(119000)
        self._nc(59500, documento_afectado=venta)   # vendedor=None

        data = self._calcular()
        fila = self._fila(data, self.vendedor.id)

        self.assertEqual(fila['devoluciones'], 59500)
        neto_venta = int((Decimal('119000') / Decimal('1.19')).quantize(Decimal('1')))
        neto_nc = int((Decimal('59500') / Decimal('1.19')).quantize(Decimal('1')))
        self.assertEqual(fila['ventas_netas_sin_iva'], neto_venta - neto_nc)
        self.assertEqual(
            fila['comision_monto'],
            int(round((neto_venta - neto_nc) * 5 / 100)),
        )
        self.assertEqual(data['totales']['devoluciones_sin_vendedor'], 0)

    def test_el_filtro_viejo_botaba_la_nc_completa(self):
        """Sanity check del bug: sin fallback, la comisión iba sobre el bruto."""
        venta = self._dte(119000)
        self._nc(59500, documento_afectado=venta)

        ncs_con_vendedor = Dte.objects.filter(
            tipo_documento='NOTA DE CREDITO', vendedor__isnull=False,
        )
        self.assertEqual(ncs_con_vendedor.count(), 0)

        data = self._calcular()
        self.assertEqual(data['totales']['total_devoluciones'], 59500)

    def test_nc_huerfana_va_a_fila_sin_vendedor_y_afecta_el_total(self):
        self._dte(119000)
        self._nc(23800)   # sin vendedor y sin documento_afectado

        data = self._calcular()

        sin_vend = [v for v in data['vendedores'] if v.get('sin_vendedor')]
        self.assertEqual(len(sin_vend), 1)
        self.assertEqual(sin_vend[0]['nombre'], 'Sin vendedor asignado')
        self.assertEqual(sin_vend[0]['devoluciones'], 23800)
        # No se le descuenta comisión a ninguna persona concreta...
        self.assertEqual(sin_vend[0]['comision_monto'], 0)
        # ...pero la plata SÍ afecta el total del reporte.
        self.assertEqual(data['totales']['total_devoluciones'], 23800)
        self.assertEqual(data['totales']['devoluciones_sin_vendedor'], 23800)
        self.assertEqual(data['totales']['cantidad_ncs_sin_vendedor'], 1)
        # La fila sintética no infla el conteo de vendedores.
        self.assertEqual(data['totales']['cantidad_vendedores'], 1)

    def test_nc_de_periodo_anterior_genera_fila_con_comision_negativa(self):
        """Vendió el mes pasado, devolvieron este mes: es un descuento real."""
        venta_previa = self._dte(119000, fecha=self.hoy - timedelta(days=40))
        self._nc(119000, documento_afectado=venta_previa)

        data = self._calcular()
        fila = self._fila(data, self.vendedor.id)

        self.assertIsNotNone(fila)
        self.assertEqual(fila['ventas_brutas'], 0)
        self.assertEqual(fila['devoluciones'], 119000)
        self.assertLess(fila['comision_monto'], 0)
        self.assertLess(data['totales']['total_comisiones'], 0)

    def test_filtro_por_vendedor_conserva_sus_devoluciones(self):
        otro = crear_vendedor(
            nombre='Vendedor Dos', empresa=self.empresa,
            codigo_vendedor='V-02', rut='9.876.543-2', comision=Decimal('3.00'),
        )
        otro.sucursales.add(self.sucursal)

        venta = self._dte(119000)
        self._nc(59500, documento_afectado=venta)
        self._dte(238000, vendedor=otro)

        data = self._calcular(vendedor_id=str(self.vendedor.id))

        self.assertEqual(len(data['vendedores']), 1)
        fila = data['vendedores'][0]
        self.assertEqual(fila['id'], self.vendedor.id)
        self.assertEqual(fila['devoluciones'], 59500)
