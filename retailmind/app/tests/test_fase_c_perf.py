# -*- coding: utf-8 -*-
"""
Fase C auditoría de Reportes 2026-08 — performance SIN cambiar ni un número
(docs/AUDITORIA_REPORTES_2026-08.md, secciones 4 y 10).

Qué garantiza esta suite (sobre las versiones CONSOLIDADAS):

1. inteligencia-compra: las series anual/por tienda/mensual, la velocidad 90d
   (con sus bordes de ventana) y las finanzas TTM salen de pasadas únicas por
   (grupo, período) y siguen siendo NETAS de reingresos, con los mismos
   valores que las agregaciones separadas de la Fase B.
2. curva de tallas: el fallback all-time (marca sin ventas en 24m) sigue
   funcionando con la pasada combinada ventas+reingresos.
3. plan-liquidación: las dimensiones marca/categoría/sucursal derivadas de la
   pasada común reproducen ranking, acciones y totales — incluido el stock
   SIN marca que cuenta en los totales pero no genera fila de marca.
4. por-anio: los tramos derivados de la misma pasada que la dimensión
   cuentan productos DISTINTOS por tramo aunque haya varias marcas.
5. export Excel: última venta viene ahora de una pasada agrupada (Max fecha)
   y muestra lo mismo que el Subquery por fila ('Nunca' incluido).
6. resumen-existencias HISTÓRICO: la reconstrucción en SQL respeta la fórmula
   de reversión exacta (ingresos propios restan, egresos propios suman en
   valor absoluto, destino NULL/ajeno no cuenta, clamp a 0, tallas con stock
   0 reviven, estados no COMPLETADO y movimientos pre-corte se ignoran).

Aislado en SQLite:
    $env:DATABASE_URL="sqlite:///C:/temp/tc3.sqlite3"; python manage.py test app.tests.test_fase_c_perf
"""
import io
import json
from datetime import timedelta

from django.utils import timezone

from app.models import AtributoOpcion, Movimientos_Producto
from app.tests.factories import crear_producto_con_talla, crear_sucursal
from app.tests.test_fase_b_inteligencia import BaseFaseB
from app.views_inteligencia_compra import (
    exportar_plan_liquidacion_excel, obtener_inteligencia_compra,
    obtener_plan_liquidacion, obtener_plan_liquidacion_por_anio,
)
from app.views_resumen_existencias import obtener_resumen_existencias


# ══════════════ 1. inteligencia: series consolidadas y bordes ══════════════
class InteligenciaConsolidadaTest(BaseFaseB):
    """2 tiendas + 1 bodega. Ventas en los bordes de las ventanas: la venta
    de HOY entra a la historia y al TTM pero NO a la velocidad 90d
    (fecha__lt=hoy), igual que con los aggregates separados."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.tienda2 = crear_sucursal(self.empresa, alias='TIENDA-2')
        self.bodega = crear_sucursal(self.empresa, alias='BOD-CD',
                                     es_centro_distribucion=True)
        _p, self.t1 = self._producto_marca('ZAP-C1', sku=8300001, talla='42')
        p2, self.t2 = crear_producto_con_talla(
            self.tienda2, articulo='ZAP-C2', sku=8300002, talla='40',
            atributo1=self.marca)
        p3, self.t3 = crear_producto_con_talla(
            self.bodega, articulo='ZAP-C3', sku=8300003, talla='41',
            atributo1=self.marca)
        # tienda 1: venta HOY (borde), venta 50d, devolución 30d
        self._mov(self.t1, 'VENTA_PUBLICO', -4, dias_atras=0)
        self._mov(self.t1, 'VENTA_PUBLICO', -10, dias_atras=50)
        self._mov(self.t1, 'DEVOLUCION_CLIENTE', +2, dias_atras=30)
        # tienda 2: venta 20d
        Movimientos_Producto.objects.create(
            ProductoTalla=self.t2, sucursal_origen=self.tienda2,
            cantidad=-3, costo=1000, precio=2000, concepto='VENTA_PUBLICO',
            estado='COMPLETADO', fecha=self.hoy - timedelta(days=20))
        # bodega: venta mayorista 15d + reingreso 10d
        for cant, conc, dias in ((-5, 'VENTA_MAYORISTA', 15),
                                 (+1, 'DEVOLUCION_CLIENTE', 10)):
            Movimientos_Producto.objects.create(
                ProductoTalla=self.t3, sucursal_origen=self.bodega,
                cantidad=cant, costo=1000, precio=2000, concepto=conc,
                estado='COMPLETADO', fecha=self.hoy - timedelta(days=dias))

    def _data(self, **params):
        params.setdefault('marca_id', self.marca.id)
        params.setdefault('sucursal_id', 'todas')
        resp = obtener_inteligencia_compra(self._req(ajax=True, **params))
        body = json.loads(resp.content)
        self.assertTrue(body['success'], body.get('error'))
        return body['data']

    def test_series_netas_consolidadas(self):
        d = self._data()
        # historia total neta: (4+10-2) + 3 = 15 (sin bodega)
        self.assertEqual(d['kpis']['venta_publica_hist'], 15)
        self.assertEqual(sum(x['unidades'] for x in d['ventas_anual']), 15)
        self.assertEqual(sum(x['monto'] for x in d['ventas_anual']), 15 * 2000)
        por_tienda = {x['alias']: x for x in d['ventas_por_tienda']}
        self.assertEqual(por_tienda['TIENDA-1']['unidades'], 12)
        self.assertEqual(por_tienda['TIENDA-1']['monto'], 12 * 2000)
        self.assertEqual(por_tienda['TIENDA-2']['unidades'], 3)
        # serie mensual (chart 36m) suma lo mismo que la historia
        self.assertEqual(sum(p['u'] for p in d['serie_chart']), 15)

    def test_distribucion_bodega_neta(self):
        d = self._data()
        self.assertEqual(d['distribucion_bodega'],
                         [{'alias': 'BOD-CD', 'unidades': 4}])  # 5 - 1
        self.assertEqual(d['kpis']['distribucion_bodega'], 4)

    def test_velocidad_90d_excluye_hoy(self):
        d = self._data()
        # v90: [hoy-90, hoy) => la venta de HOY queda fuera: 10-2+3 = 11
        self.assertEqual(d['kpis']['vel_dia'], round(11 / 90, 1))

    def test_ttm_incluye_hoy_margen_realizado(self):
        d = self._data()
        # TTM (fecha >= hoy-365, sin tope): 17 ventas - 2 reingresos
        # venta $ = 15*2000 = 30000; costo $ = 15*1000 = 15000 -> 50% realizado
        self.assertEqual(d['finanzas']['margen_pct'], 50.0)
        self.assertEqual(d['finanzas']['margen_src'], 'realizado')

    def test_curva_tallas_neta(self):
        d = self._data()
        curva = {c['talla']: c for c in d['curva']}
        self.assertEqual(curva['42']['venta_u'], 12)  # 14 - 2
        self.assertEqual(curva['40']['venta_u'], 3)


class CurvaFallbackTest(BaseFaseB):
    """Marca sin ventas en 24m: la curva cae a all-time también con la
    pasada combinada (el criterio sigue siendo 'hay VENTAS en 24m')."""

    def test_fallback_all_time(self):
        _p, talla = self._producto_marca('ZAP-VIEJA', sku=8300010, talla='40')
        self._mov(talla, 'VENTA_PUBLICO', -6, dias_atras=800)
        resp = obtener_inteligencia_compra(self._req(
            ajax=True, marca_id=self.marca.id, sucursal_id='todas'))
        body = json.loads(resp.content)
        self.assertTrue(body['success'], body.get('error'))
        curva = {c['talla']: c for c in body['data']['curva']}
        self.assertEqual(curva['40']['venta_u'], 6)


# ══════════════ 3. plan liquidación: dimensiones derivadas ══════════════
class PlanConsolidadoTest(BaseFaseB):
    """Marca A (vendida, sobre-stock), marca B (dead) y un producto SIN
    marca que solo cuenta en totales. CD aparte con incluir_cd."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.marca_b = AtributoOpcion.objects.create(
            atributo=self.attr_marca, valor='MARCA-B')
        _pa, self.ta = self._producto_marca('ART-A', sku=8400001, stock=5,
                                            costo=1000)
        crear_producto_con_talla(self.tienda, articulo='ART-B', sku=8400002,
                                 stock=3, costo=2000, atributo1=self.marca_b)
        crear_producto_con_talla(self.tienda, articulo='ART-N', sku=8400003,
                                 stock=2, costo=500)  # sin marca
        self.cd = crear_sucursal(self.empresa, alias='CD-1',
                                 es_centro_distribucion=True)
        crear_producto_con_talla(self.cd, articulo='ART-C', sku=8400004,
                                 stock=7, costo=1000, atributo1=self.marca)
        # venta TTM de A (60d): 4 u a $2000 (costo $1000)
        self._mov(self.ta, 'VENTA_PUBLICO', -4, dias_atras=60)

    def _data(self, **params):
        resp = obtener_plan_liquidacion(self._req(ajax=True, **params))
        body = json.loads(resp.content)
        self.assertTrue(body['success'], body.get('error'))
        return body['data']

    def test_totales_incluyen_sin_marca(self):
        d = self._data()
        self.assertEqual(d['totales']['stock_u'], 10)          # 5+3+2
        self.assertEqual(d['totales']['valor_costo'], 12000)   # 5k+6k+1k
        self.assertEqual(d['totales']['dead_u'], 5)            # B(3) + N(2)
        self.assertEqual(d['totales']['dead_costo'], 7000)
        # pero solo 2 filas de marca (la sin-marca no genera fila)
        self.assertEqual({f['marca'] for f in d['marcas']},
                         {'TESTBRAND', 'MARCA-B'})

    def test_metricas_por_marca(self):
        d = self._data()
        filas = {f['marca']: f for f in d['marcas']}
        a, b = filas['TESTBRAND'], filas['MARCA-B']
        self.assertEqual((a['stock'], a['ttm_u'], a['rotacion']), (5, 4, 0.8))
        self.assertEqual(a['cobertura'], 15.0)
        self.assertEqual(a['gmroi'], 0.8)          # margen 4000 / costo 5000
        self.assertEqual(a['accion'], 'Liquidar')  # gmroi<1 + cobertura>12
        self.assertEqual((b['dead_u'], b['dead_costo'], b['pct_dead']),
                         (3, 6000, 100.0))
        self.assertEqual(b['accion'], 'Depurar')

    def test_dimensiones_categoria_y_sucursal(self):
        d = self._data()
        # factory: todos en la misma categoría => 1 fila con el total tiendas
        self.assertEqual(len(d['categorias']), 1)
        self.assertEqual(d['categorias'][0]['stock'], 10)
        suc = {f['sucursal']: f for f in d['sucursales']}
        self.assertEqual(suc['TIENDA-1']['stock'], 10)

    def test_incluir_cd_separa_stock(self):
        d = self._data(incluir_cd='1')
        self.assertEqual(d['totales']['stock_cd'], 7)
        self.assertEqual(d['totales']['valor_cd'], 7000)
        a = next(f for f in d['marcas'] if f['marca'] == 'TESTBRAND')
        self.assertEqual((a['stock'], a['stock_cd']), (5, 7))


# ══════════════ 4. por-anio: tramos derivados de la dimensión ══════════════
class PorAnioProductosPorTramoTest(BaseFaseB):
    """2 marcas distintas con lote en el MISMO tramo: el conteo de productos
    del tramo debe ser la suma de distintos por grupo (2), y valor/pares la
    suma de ambos — igual que la query dedicada anterior."""

    def test_tramo_suma_marcas(self):
        marca2 = AtributoOpcion.objects.create(
            atributo=self.attr_marca, valor='OTRA-MARCA')
        _p1, t1 = self._producto_marca('A-400D', sku=8500001, stock=5, costo=1000)
        _p2, t2 = crear_producto_con_talla(
            self.tienda, articulo='B-500D', sku=8500002, stock=2, costo=3000,
            atributo1=marca2)
        self._fijar_lote(t1, dias_atras=400)
        self._fijar_lote(t2, dias_atras=500)
        status, body = self._json(obtener_plan_liquidacion_por_anio, self._req())
        self.assertEqual(status, 200)
        tramos = {t['tramo']: t for t in body['tramos']}
        self.assertEqual(set(tramos), {'t2'})
        self.assertEqual(tramos['t2']['productos'], 2)
        self.assertEqual(tramos['t2']['valor'], 5 * 1000 + 2 * 3000)
        self.assertEqual(tramos['t2']['pares'], 7)
        nombres = {m['nombre'] for m in body['por_dimension']}
        self.assertEqual(nombres, {'TESTBRAND', 'OTRA-MARCA'})


# ══════════════ 5. export: última venta desde la pasada agrupada ══════════
class ExportUltimaVentaTest(BaseFaseB):
    def test_ultima_venta_y_nunca(self):
        from openpyxl import load_workbook
        pv, tv = self._producto_marca('VENDIDO', sku=8600001, stock=4)
        pn, _tn = self._producto_marca('NUNCA-V', sku=8600002, stock=2)
        self._fijar_lote(tv, dias_atras=200)
        self._mov(tv, 'VENTA_PUBLICO', -5, dias_atras=20)
        self._mov(tv, 'VENTA_PUBLICO', -1, dias_atras=40)  # Max = la de 20d
        resp = exportar_plan_liquidacion_excel(self._req())
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb['TIENDA-1']
        hoy = timezone.localdate()
        celdas = {row[0]: row for row in ws.iter_rows(min_row=3, values_only=True)}
        fila_v = celdas[pv.id]
        fila_n = celdas[pn.id]
        # columna 11 = 'Últ. venta'
        self.assertEqual(fila_v[10],
                         (hoy - timedelta(days=20)).strftime('%d-%m-%Y'))
        self.assertEqual(fila_n[10], 'Nunca')


# ══════════════ 6. resumen-existencias histórico en SQL ══════════════
class ResumenHistoricoSQLTest(BaseFaseB):
    """Cada caso de la fórmula de reversión, uno por producto. Corte hace 10
    días; movimientos post a 5 días, pre a 15."""

    def setUp(self):
        super().setUp()
        self.hoy = timezone.localdate()
        self.corte = self.hoy - timedelta(days=10)
        self.otra = crear_sucursal(self.empresa, alias='TIENDA-2')

        def prod(articulo, sku, stock):
            _p, t = crear_producto_con_talla(
                self.tienda, articulo=articulo, sku=sku, stock=stock,
                costo=100, sobreprecio=0, precioventa=300)
            return t

        def mov(talla, cantidad, dias, destino=None, origen=None,
                concepto='REPOSICION_STOCK', estado='COMPLETADO'):
            Movimientos_Producto.objects.create(
                ProductoTalla=talla, cantidad=cantidad,
                sucursal_destino=destino, sucursal_origen=origen,
                costo=100, precio=300, concepto=concepto, estado=estado,
                fecha=self.hoy - timedelta(days=dias))

        t1 = prod('P1-ING', 8700001, 10)      # ingreso post propio: 10-3 = 7
        mov(t1, +3, 5, destino=self.tienda)
        t2 = prod('P2-EGR', 8700002, 5)       # egreso post propio: 5+|−2| = 7
        mov(t2, -2, 5, origen=self.tienda, concepto='VENTA_PUBLICO')
        t3 = prod('P3-CLAMP', 8700003, 4)     # 4-9 = -5 -> fuera
        mov(t3, +9, 5, destino=self.tienda)
        t4 = prod('P4-REVIVE', 8700004, 0)    # 0+|−6| = 6 (stock 0 revive)
        mov(t4, -6, 5, origen=self.tienda, concepto='VENTA_PUBLICO')
        t5 = prod('P5-AJENA', 8700005, 8)     # destino OTRA sucursal: no toca
        mov(t5, +5, 5, destino=self.otra)
        t6 = prod('P6-NULL', 8700006, 3)      # destino NULL: no toca
        mov(t6, +2, 5, destino=None)
        t7 = prod('P7-PRE', 8700007, 2)       # anterior al corte: no toca
        mov(t7, -9, 15, origen=self.tienda, concepto='VENTA_PUBLICO')
        t8 = prod('P8-PEND', 8700008, 6)      # no COMPLETADO: no toca
        mov(t8, -4, 5, origen=self.tienda, concepto='VENTA_PUBLICO',
            estado='PENDIENTE')
        # esperado histórico: 7+7+0+6+8+3+2+6 = 39 · actual: 38 (P4 sin stock)

    def _body(self, **params):
        resp = obtener_resumen_existencias(self._req(ajax=True, **params))
        body = json.loads(resp.content)
        self.assertTrue(body['success'], body.get('error'))
        return body

    def test_reconstruccion_por_sucursal(self):
        body = self._body(fecha_corte=self.corte.isoformat())
        self.assertTrue(body['es_historico'])
        self.assertEqual(body['total_general']['pares'], 39)
        fila = next(f for f in body['datos'] if f['sucursal'] == 'TIENDA-1')
        self.assertEqual(fila['total_pares'], 39)
        # valores: costo 100 / venta 300 por par
        self.assertEqual(fila['total_costo'], 39 * 100.0)
        self.assertEqual(fila['total_precio_venta'], 39 * 300.0)

    def test_corte_actual_no_cambia(self):
        body = self._body()
        self.assertFalse(body['es_historico'])
        self.assertEqual(body['total_general']['pares'], 38)

    def test_reconstruccion_por_categoria(self):
        body = self._body(fecha_corte=self.corte.isoformat(),
                          agrupar_por='categoria')
        self.assertTrue(body['es_historico'])
        self.assertEqual(body['total_general']['pares'], 39)
