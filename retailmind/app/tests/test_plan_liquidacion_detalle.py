"""
Tests del drill-down del Plan de Liquidación: antigüedad FIFO (lote y
fallback a fecha_creacion), buckets de antigüedad, inclusión/exclusión de
bodegas/CD y filtro por especialidad sin duplicar. Aislados en SQLite.
"""
import json
from datetime import timedelta

from django.test import RequestFactory, TestCase
from django.utils import timezone

import io

from django.core.files.uploadedfile import SimpleUploadedFile

from app.models import (
    AtributoOpcion, LoteProducto, ModuloSistema, OpcionMenu, PermisoRol,
    Producto, ProductoAtributoValor, Productos_Atributos,
)
from app.views_inteligencia_compra import (
    exportar_plan_liquidacion_excel, importar_seleccion_liquidacion,
    imprimir_plan_liquidacion, obtener_plan_liquidacion,
    obtener_plan_liquidacion_detalle,
)
from app.tests.factories import (
    crear_empresa, crear_empresa_user, crear_sucursal, crear_producto_con_talla,
    crear_usuario, crear_lote_fifo,
)


class PlanLiquidacionDetalleTests(TestCase):
    def setUp(self):
        self.empresa = crear_empresa()
        self.tienda = crear_sucursal(empresa=self.empresa, alias='TIENDA-1')
        self.bodega = crear_sucursal(
            empresa=self.empresa, alias='BODEGA-CD',
            es_centro_distribucion=True, tipo_sucursal='CENTRO_DISTRIBUCION')
        self.user = crear_usuario(username='analista')
        crear_empresa_user(self.user, self.empresa, self.tienda)
        # Las vistas del Plan de Liquidacion exigen el permiso plan_liquidacion.puede_ver.
        modulo, _ = ModuloSistema.objects.get_or_create(
            codigo='liquidacion', defaults={'nombre': 'Liquidacion', 'orden': 11})
        opcion, _ = OpcionMenu.objects.get_or_create(
            codigo='plan_liquidacion',
            defaults={'modulo': modulo, 'nombre': 'Plan de Liquidacion', 'orden': 1})
        PermisoRol.objects.update_or_create(
            rol=self.user.rol, opcion_menu=opcion,
            defaults={'puede_ver': True, 'puede_exportar': True})

    def _req(self, **params):
        req = RequestFactory().get('/x', data=params)
        req.user = self.user
        req.session = {'idSucursalActual': self.tienda.id,
                       'idEmpresaActual': self.empresa.id}
        return req

    def _fijar_lote(self, talla, dias_atras, **kw):
        lote = crear_lote_fifo(talla, **kw)
        fecha = timezone.now() - timedelta(days=dias_atras)
        LoteProducto.objects.filter(pk=lote.pk).update(fecha_ingreso=fecha)
        return lote

    def test_antiguedad_desde_lote_y_bucket(self):
        prod, talla = crear_producto_con_talla(
            self.tienda, articulo='VIEJO', sku=7000001, stock=5)
        self._fijar_lote(talla, dias_atras=400)

        data = json.loads(obtener_plan_liquidacion_detalle(self._req()).content)
        self.assertTrue(data['success'])
        fila = next(f for f in data['filas'] if f['articulo'] == 'VIEJO')
        self.assertEqual(fila['antiguedad_fuente'], 'lote')
        self.assertGreaterEqual(fila['dias_antiguedad'], 399)

        # Bucket +1 año lo incluye; 0-90 lo excluye.
        d365 = json.loads(obtener_plan_liquidacion_detalle(
            self._req(antiguedad='365+')).content)
        self.assertTrue(any(f['articulo'] == 'VIEJO' for f in d365['filas']))
        d090 = json.loads(obtener_plan_liquidacion_detalle(
            self._req(antiguedad='0-90')).content)
        self.assertFalse(any(f['articulo'] == 'VIEJO' for f in d090['filas']))

    def test_fallback_fecha_creacion_sin_lote(self):
        prod, talla = crear_producto_con_talla(
            self.tienda, articulo='SINLOTE', sku=7000002, stock=3)
        # Sin lote → antigüedad por fecha_creacion (auto_now_add=hoy).
        data = json.loads(obtener_plan_liquidacion_detalle(self._req()).content)
        fila = next(f for f in data['filas'] if f['articulo'] == 'SINLOTE')
        self.assertEqual(fila['antiguedad_fuente'], 'creacion')
        self.assertIsNotNone(fila['fecha_fifo'])

    def test_cd_excluida_por_defecto_incluida_con_flag(self):
        crear_producto_con_talla(self.tienda, articulo='EN-TIENDA',
                                 sku=7000003, stock=4)
        crear_producto_con_talla(self.bodega, articulo='EN-BODEGA',
                                 sku=7000004, stock=9)

        base = json.loads(obtener_plan_liquidacion_detalle(self._req()).content)
        arts = {f['articulo'] for f in base['filas']}
        self.assertIn('EN-TIENDA', arts)
        self.assertNotIn('EN-BODEGA', arts)  # CD excluida por defecto

        con_cd = json.loads(obtener_plan_liquidacion_detalle(
            self._req(incluir_cd='1')).content)
        arts_cd = {f['articulo'] for f in con_cd['filas']}
        self.assertIn('EN-BODEGA', arts_cd)
        bodega_fila = next(f for f in con_cd['filas'] if f['articulo'] == 'EN-BODEGA')
        self.assertTrue(bodega_fila['es_cd'])

    def test_cd_stock_separado_en_agregado(self):
        crear_producto_con_talla(self.tienda, articulo='T', sku=7000005, stock=4)
        crear_producto_con_talla(self.bodega, articulo='B', sku=7000006, stock=9)
        data = json.loads(obtener_plan_liquidacion(self._req(incluir_cd='1')).content)
        tot = data['data']['totales']
        self.assertEqual(tot['stock_u'], 4)   # solo tiendas
        self.assertEqual(tot['stock_cd'], 9)  # bodega aparte

    # ---------------- export multi-tab / impresión / import ----------------

    def _post_import(self, contenido, nombre, ctype):
        req = RequestFactory().post('/x')
        req.user = self.user
        req.session = {'idSucursalActual': self.tienda.id,
                       'idEmpresaActual': self.empresa.id}
        req.FILES['archivo'] = SimpleUploadedFile(nombre, contenido, content_type=ctype)
        return json.loads(importar_seleccion_liquidacion(req).content)

    def _setup_marcas_orden(self):
        """3 productos en la tienda con marcas/artículos para probar el orden."""
        attr_m = Productos_Atributos.objects.create(nombre='Marca', descripcion='m')
        nike = AtributoOpcion.objects.create(atributo=attr_m, valor='NIKE')
        adidas = AtributoOpcion.objects.create(atributo=attr_m, valor='ADIDAS')
        crear_producto_con_talla(self.tienda, articulo='B-ART', sku=7100001,
                                 stock=2, atributo1=nike)
        crear_producto_con_talla(self.tienda, articulo='Z-ART', sku=7100002,
                                 stock=3, atributo1=adidas)
        crear_producto_con_talla(self.tienda, articulo='A-ART', sku=7100003,
                                 stock=4, atributo1=adidas)

    def test_export_multitab_por_tienda_y_orden(self):
        from openpyxl import load_workbook
        self._setup_marcas_orden()
        tienda2 = crear_sucursal(empresa=self.empresa, alias='TIENDA-2')
        crear_producto_con_talla(tienda2, articulo='OTRA', sku=7100004, stock=1)

        resp = exportar_plan_liquidacion_excel(self._req())
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
        # 1 hoja por tienda + Resumen + Filtros.
        self.assertIn('TIENDA-1', wb.sheetnames)
        self.assertIn('TIENDA-2', wb.sheetnames)
        self.assertIn('Resumen', wb.sheetnames)
        self.assertIn('Filtros', wb.sheetnames)
        ws = wb['TIENDA-1']
        self.assertEqual(ws['A2'].value, 'ID')  # header para el re-import
        # Orden marca → artículo asc (ADIDAS A-ART, ADIDAS Z-ART, NIKE B-ART).
        filas = [(ws.cell(row=r, column=3).value, ws.cell(row=r, column=5).value)
                 for r in range(3, 6)]
        self.assertEqual(filas, [('ADIDAS', 'A-ART'), ('ADIDAS', 'Z-ART'),
                                 ('NIKE', 'B-ART')])
        # Sin costo en hojas de tienda (headers).
        headers = [ws.cell(row=2, column=c).value for c in range(1, 18)]
        self.assertNotIn('Costo unit.', headers)
        self.assertIn('Precio caja $', headers)
        self.assertIn('¿Coincide?', headers)

    def test_import_multitab_roundtrip_ignora_resumen(self):
        self._setup_marcas_orden()
        resp = exportar_plan_liquidacion_excel(self._req())
        res = self._post_import(resp.content, 'export.xlsx',
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(res['success'], res.get('error'))
        esperados = set(Producto.objects.filter(
            sucursal=self.tienda).values_list('id', flat=True))
        # Exactamente los productos exportados; los años del Resumen NO entran.
        self.assertEqual(set(res['producto_ids']), esperados)

    def test_import_csv_ia_con_discrepancias(self):
        prod, _ = crear_producto_con_talla(
            self.tienda, articulo='VERIF', sku=7100010, stock=2, precioventa=20000)
        csv = (f'ID,PRECIO_CAJA,COINCIDE,OBSERVACION\n'
               f'{prod.id},15990,NO,caja danada\n').encode('utf-8')
        res = self._post_import(csv, 'ia.csv', 'text/csv')
        self.assertTrue(res['success'], res.get('error'))
        self.assertEqual(res['producto_ids'], [prod.id])
        v = res['verificacion']
        self.assertEqual(v['con_datos'], 1)
        self.assertEqual(v['coincide_no'], 1)
        self.assertEqual(v['precio_distinto'], 1)
        self.assertEqual(v['detalles'][0]['precio_caja'], 15990)

    def test_imprimir_formulario_html(self):
        prod, _ = crear_producto_con_talla(
            self.tienda, articulo='IMPRESO', sku=7100020, stock=3)
        resp = imprimir_plan_liquidacion(self._req())
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode('utf-8')
        self.assertIn('TIENDA-1', html)
        self.assertIn(str(prod.id), html)
        self.assertIn('LIQ-', html)          # barcode/código de lote
        self.assertIn('Coincide', html)      # casillas de verificación

    def test_filtro_especialidad_no_duplica(self):
        prod, talla = crear_producto_con_talla(
            self.tienda, articulo='CANCHA', sku=7000007, stock=6)
        # Producto con DOS especialidades → no debe contar doble el stock.
        attr = Productos_Atributos.objects.create(
            nombre='Especialidad', descripcion='esp')
        op1 = AtributoOpcion.objects.create(atributo=attr, valor='futbol')
        op2 = AtributoOpcion.objects.create(atributo=attr, valor='pasto')
        ProductoAtributoValor.objects.create(producto=prod, atributo=attr, opcion=op1)
        ProductoAtributoValor.objects.create(producto=prod, atributo=attr, opcion=op2)

        data = json.loads(obtener_plan_liquidacion_detalle(
            self._req(especialidad_id=op1.id)).content)
        filas = [f for f in data['filas'] if f['articulo'] == 'CANCHA']
        self.assertEqual(len(filas), 1)          # una sola fila
        self.assertEqual(filas[0]['stock_u'], 6)  # stock sin duplicar
